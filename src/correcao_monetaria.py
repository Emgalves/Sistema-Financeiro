# correcao_monetaria.py


import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import json
from pathlib import Path
from openpyxl import load_workbook
# import requests
import threading
from decimal import Decimal, ROUND_HALF_UP

# Importações do seu sistema
try:
    from src.config.logger_config import system_logger, log_action
    logger = system_logger.get_logger()
except ImportError:
    import logging
    logger = logging.getLogger("correcao_monetaria")

class GerenciadorCorrecaoMonetaria:
    def __init__(self, sistema_principal=None):
        self.sistema = sistema_principal
        # Usar o mesmo caminho das configurações principais
        try:
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            self.config_path = GerenciadorConfiguracoes.CONFIG_PATH
        except ImportError:
            # Fallback caso não consiga importar
            from pathlib import Path
            self.config_path = Path("parametros_sistema.json")
        
        self.config = self.carregar_configuracoes()
        
    def carregar_configuracoes(self):
        """Carrega configurações com suporte a índices"""
        if self.config_path.exists():
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                # Garantir que as seções de correção existam
                if 'indices_correcao' not in config:
                    config['indices_correcao'] = {
                        'indice_padrao': 'IGPM',
                        'indices_disponiveis': {
                            'IGPM': {'nome_completo': 'Índice Geral de Preços do Mercado', 'historico': []},
                            'IPCA': {'nome_completo': 'Índice Nacional de Preços ao Consumidor Amplo', 'historico': []},
                            'INPC': {'nome_completo': 'Índice Nacional de Preços ao Consumidor', 'historico': []}
                        }
                    }
                    self.salvar_configuracoes_config(config)
                
                if 'correcao_automatica' not in config:
                    config['correcao_automatica'] = {
                        'ativa': True,
                        'dia_calculo': 15,
                        'meses_aplicacao': [1, 4, 7, 10],
                        'avisar_antes_dias': 7,
                        'ultimo_processamento': None
                    }
                    self.salvar_configuracoes_config(config)
                    
                return config
            except Exception as e:
                print(f"Erro ao carregar configurações: {str(e)}")
                return self._configuracoes_padrao()
        
        return self._configuracoes_padrao()
    
    def _configuracoes_padrao(self):
        """Configurações padrão"""
        return {
            'indices_correcao': {
                'indice_padrao': 'IGPM',
                'indices_disponiveis': {
                    'IGPM': {'nome_completo': 'Índice Geral de Preços do Mercado', 'historico': []},
                    'IPCA': {'nome_completo': 'Índice Nacional de Preços ao Consumidor Amplo', 'historico': []},
                    'INPC': {'nome_completo': 'Índice Nacional de Preços ao Consumidor', 'historico': []}
                }
            },
            'correcao_automatica': {
                'ativa': True,
                'dia_calculo': 15,
                'meses_aplicacao': [1, 4, 7, 10],
                'avisar_antes_dias': 7,
                'ultimo_processamento': None
            }
        }
    
    def salvar_configuracoes_config(self, config):
        """Salva um config específico"""
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Erro ao salvar configurações: {str(e)}")
    
    def salvar_configuracoes(self):
        """Salva as configurações"""
        self.salvar_configuracoes_config(self.config)
    
    def adicionar_indice_periodo(self, indice, periodo, valor):
        """
        Adiciona um índice para um período específico
        
        Args:
            indice: Nome do índice (IGPM, IPCA, etc.)
            periodo: Período no formato 'MM/AAAA'
            valor: Valor do índice (pode ser positivo ou negativo)
        """
        try:
            if indice not in self.config['indices_correcao']['indices_disponiveis']:
                raise ValueError(f"Índice {indice} não configurado")
            
            # Validar formato do período
            mes, ano = periodo.split('/')
            if not (1 <= int(mes) <= 12):
                raise ValueError("Mês deve estar entre 01 e 12")
            
            # Converter valor para decimal para precisão
            valor_decimal = Decimal(str(valor))
            
            # Verificar se já existe registro para este período
            historico = self.config['indices_correcao']['indices_disponiveis'][indice]['historico']
            
            for i, registro in enumerate(historico):
                if registro['periodo'] == periodo:
                    # Atualizar registro existente
                    historico[i]['valor'] = float(valor_decimal)
                    historico[i]['data_atualizacao'] = datetime.now().isoformat()
                    self.salvar_configuracoes()
                    return True
            
            # Adicionar novo registro
            novo_registro = {
                'periodo': periodo,
                'valor': float(valor_decimal),
                'data_inclusao': datetime.now().isoformat(),
                'data_atualizacao': datetime.now().isoformat()
            }
            
            historico.append(novo_registro)
            # Ordenar por período
            historico.sort(key=lambda x: datetime.strptime(x['periodo'], '%m/%Y'))
            
            self.salvar_configuracoes()
            return True
            
        except Exception as e:
            print(f"Erro ao adicionar índice: {str(e)}")
            return False
    
    def calcular_correcao_acumulada(self, indice, data_inicial, data_final):
        """
        Calcula a correção monetária acumulada entre duas datas
        
        Args:
            indice: Nome do índice
            data_inicial: Data inicial (datetime.date)
            data_final: Data final (datetime.date)
            
        Returns:
            dict: {'fator_correcao': float, 'percentual': float, 'detalhes': list}
        """
        try:
            if indice not in self.config['indices_correcao']['indices_disponiveis']:
                raise ValueError(f"Índice {indice} não encontrado")
            
            historico = self.config['indices_correcao']['indices_disponiveis'][indice]['historico']
            
            # Gerar lista de períodos entre as datas
            periodos_aplicacao = []
            data_atual = data_inicial.replace(day=1)  # Primeiro dia do mês
            
            while data_atual <= data_final:
                periodo_str = data_atual.strftime('%m/%Y')
                periodos_aplicacao.append(periodo_str)
                data_atual += relativedelta(months=1)
            
            # Calcular fator acumulado
            fator_acumulado = Decimal('1.0')
            detalhes = []
            
            for periodo in periodos_aplicacao:
                # Buscar índice do período
                indice_periodo = None
                for registro in historico:
                    if registro['periodo'] == periodo:
                        indice_periodo = Decimal(str(registro['valor']))
                        break
                
                if indice_periodo is not None:
                    # Aplicar correção: fator = fator * (1 + indice/100)
                    fator_periodo = 1 + (indice_periodo / 100)
                    fator_acumulado *= fator_periodo
                    
                    detalhes.append({
                        'periodo': periodo,
                        'indice': float(indice_periodo),
                        'fator_periodo': float(fator_periodo),
                        'fator_acumulado': float(fator_acumulado)
                    })
                else:
                    print(f"AVISO: Índice não encontrado para {periodo}")
            
            percentual_total = float((fator_acumulado - 1) * 100)
            
            return {
                'fator_correcao': float(fator_acumulado),
                'percentual': percentual_total,
                'detalhes': detalhes,
                'periodos_aplicados': len(detalhes),
                'periodos_faltantes': len(periodos_aplicacao) - len(detalhes)
            }
            
        except Exception as e:
            print(f"Erro ao calcular correção: {str(e)}")
            return {
                'fator_correcao': 1.0,
                'percentual': 0.0,
                'detalhes': [],
                'erro': str(e)
            }
    
    def aplicar_correcao_contratos(self, cliente=None, data_corte=None):
        """
        Aplica correção monetária aos contratos com valores fixos
        
        Args:
            cliente: Nome do cliente (None para todos)
            data_corte: Data de referência para a correção
        """
        if not data_corte:
            data_corte = date.today()
        
        resultados = []
        
        try:
            # Se cliente específico
            if cliente:
                clientes = [cliente]
            else:
                # Buscar todos os clientes
                try:
                    from src.config.config import PASTA_CLIENTES
                except ImportError:
                    # Fallback se não conseguir importar
                    from pathlib import Path
                    PASTA_CLIENTES = Path("clientes")
                
                clientes = []
                for arquivo in PASTA_CLIENTES.glob('*.xlsx'):
                    if not arquivo.name.startswith('~'):  # Ignorar arquivos temporários
                        clientes.append(arquivo.stem)
            
            for cliente_nome in clientes:
                try:
                    resultado_cliente = self.processar_correcao_cliente(cliente_nome, data_corte)
                    if resultado_cliente['contratos_processados'] > 0:
                        resultados.append(resultado_cliente)
                except Exception as e:
                    print(f"Erro ao processar cliente {cliente_nome}: {str(e)}")
            
            return {
                'sucesso': True,
                'clientes_processados': len(resultados),
                'detalhes': resultados,
                'data_processamento': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {
                'sucesso': False,
                'erro': str(e),
                'data_processamento': datetime.now().isoformat()
            }
    
    def processar_correcao_cliente(self, cliente, data_corte):
        """Processa correção para um cliente específico"""
        try:
            from src.config.config import PASTA_CLIENTES
        except ImportError:
            from pathlib import Path
            PASTA_CLIENTES = Path("clientes")
        
        arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        
        if not arquivo_cliente.exists():
            raise FileNotFoundError(f"Arquivo do cliente {cliente} não encontrado")
        
        wb = load_workbook(arquivo_cliente)
        
        if 'Contratos_ADM' not in wb.sheetnames:
            wb.close()
            return {'cliente': cliente, 'contratos_processados': 0, 'erro': 'Aba Contratos_ADM não encontrada'}
        
        ws = wb['Contratos_ADM']
        
        contratos_processados = 0
        detalhes_contratos = []
        
        # Buscar contratos ativos com valores fixos
        for row_idx in range(3, ws.max_row + 1):
            try:
                # Verificar se é linha de contrato
                num_contrato = ws.cell(row=row_idx, column=1).value
                if not num_contrato:
                    continue
                
                status_contrato = ws.cell(row=row_idx, column=4).value
                if status_contrato != 'ATIVO':
                    continue
                
                # Verificar linha seguinte para administradores
                if row_idx + 1 <= ws.max_row:
                    admin_contrato = ws.cell(row=row_idx + 1, column=7).value  # Coluna G
                    tipo_remuneracao = ws.cell(row=row_idx + 1, column=10).value  # Coluna J
                    
                    if admin_contrato == num_contrato and tipo_remuneracao == 'Fixo':
                        # Este contrato tem valor fixo - aplicar correção
                        resultado_contrato = self.aplicar_correcao_contrato_especifico(
                            ws, num_contrato, row_idx + 1, data_corte
                        )
                        
                        if resultado_contrato['sucesso']:
                            contratos_processados += 1
                            detalhes_contratos.append(resultado_contrato)
                        
            except Exception as e:
                print(f"Erro ao processar linha {row_idx}: {str(e)}")
        
        # Salvar alterações se houve processamento
        if contratos_processados > 0:
            wb.save(arquivo_cliente)
        
        wb.close()
        
        return {
            'cliente': cliente,
            'contratos_processados': contratos_processados,
            'detalhes': detalhes_contratos,
            'data_processamento': datetime.now().isoformat()
        }
    
    def aplicar_correcao_contrato_especifico(self, worksheet, num_contrato, linha_admin, data_corte):
        """Aplica correção a um contrato específico"""
        try:
            # Obter dados atuais do contrato
            valor_atual = worksheet.cell(row=linha_admin, column=12).value  # Coluna L - Valor Total
            data_inicio = worksheet.cell(row=linha_admin - 1, column=2).value  # Data início do contrato
            
            if not valor_atual or not data_inicio:
                return {'sucesso': False, 'erro': 'Dados insuficientes do contrato'}
            
            # Verificar se já houve correção recente
            ultima_correcao = worksheet.cell(row=linha_admin, column=15).value  # Coluna fictícia para controle
            
            if ultima_correcao and isinstance(ultima_correcao, datetime):
                # Se já foi corrigido no mesmo ano, pular
                if ultima_correcao.year == data_corte.year:
                    return {'sucesso': False, 'motivo': 'Já corrigido neste ano'}
            
            # Calcular período para correção (último ano completo)
            if isinstance(data_inicio, datetime):
                data_inicio = data_inicio.date()
            
            # Data inicial da correção: último aniversário do contrato
            anos_decorridos = data_corte.year - data_inicio.year
            if data_corte < data_inicio.replace(year=data_corte.year):
                anos_decorridos -= 1
            
            if anos_decorridos <= 0:
                return {'sucesso': False, 'motivo': 'Contrato com menos de 1 ano'}
            
            data_inicial_correcao = data_inicio.replace(year=data_inicio.year + anos_decorridos - 1)
            data_final_correcao = data_inicio.replace(year=data_inicio.year + anos_decorridos)
            
            # Calcular correção
            indice_padrao = self.config['indices_correcao']['indice_padrao']
            resultado_correcao = self.calcular_correcao_acumulada(
                indice_padrao, data_inicial_correcao, data_final_correcao
            )
            
            if 'erro' in resultado_correcao:
                return {'sucesso': False, 'erro': resultado_correcao['erro']}
            
            # Aplicar correção
            valor_original = Decimal(str(valor_atual))
            fator_correcao = Decimal(str(resultado_correcao['fator_correcao']))
            novo_valor = valor_original * fator_correcao
            
            # Arredondar para 2 casas decimais
            novo_valor = novo_valor.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            # Atualizar na planilha
            worksheet.cell(row=linha_admin, column=12, value=float(novo_valor))
            worksheet.cell(row=linha_admin, column=15, value=datetime.now())  # Controle de última correção
            
            # Atualizar parcelas pendentes com o novo valor
            self.atualizar_parcelas_pendentes(worksheet, num_contrato, float(novo_valor))
            
            return {
                'sucesso': True,
                'contrato': num_contrato,
                'valor_original': float(valor_original),
                'valor_corrigido': float(novo_valor),
                'fator_correcao': float(fator_correcao),
                'percentual_correcao': resultado_correcao['percentual'],
                'periodo_correcao': f"{data_inicial_correcao.strftime('%m/%Y')} a {data_final_correcao.strftime('%m/%Y')}",
                'indice_utilizado': indice_padrao
            }
            
        except Exception as e:
            return {'sucesso': False, 'erro': str(e)}
    
    def atualizar_parcelas_pendentes(self, worksheet, num_contrato, novo_valor_total):
        """Atualiza o valor das parcelas pendentes após correção"""
        try:
            # Buscar parcelas do contrato
            parcelas_pendentes = []
            
            for row_idx in range(3, worksheet.max_row + 1):
                if (worksheet.cell(row=row_idx, column=25).value == num_contrato and  # Mesmo contrato
                    worksheet.cell(row=row_idx, column=31).value != 'PAGO'):  # Parcela não paga
                    parcelas_pendentes.append(row_idx)
            
            if parcelas_pendentes:
                # Calcular novo valor por parcela
                valor_parcela = novo_valor_total / len(parcelas_pendentes)
                
                # Atualizar cada parcela pendente
                for linha in parcelas_pendentes:
                    worksheet.cell(row=linha, column=30, value=valor_parcela)  # Coluna AD - Valor
                    
                print(f"Atualizadas {len(parcelas_pendentes)} parcelas pendentes do contrato {num_contrato}")
                
        except Exception as e:
            print(f"Erro ao atualizar parcelas: {str(e)}")

# ===== INTERFACE PARA GERENCIAR ÍNDICES =====

class InterfaceIndicesCorrecao:
    def __init__(self, parent=None):
        self.root = tk.Toplevel(parent) if parent else tk.Tk()
        self.root.title("Gerenciamento de Índices de Correção Monetária")
        self.root.geometry("900x700")
        
        self.gerenciador = GerenciadorCorrecaoMonetaria()
        self.setup_gui()
        self.carregar_dados()
    
    def setup_gui(self):
        # Notebook para diferentes seções
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Abas
        self.setup_aba_indices()
        self.setup_aba_correcao()
        self.setup_aba_relatorios()
        
        # Botões globais
        frame_botoes = ttk.Frame(self.root)
        frame_botoes.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(frame_botoes, text="Salvar Alterações",
                  command=self.salvar_alteracoes).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Fechar",
                  command=self.root.destroy).pack(side='right', padx=5)
    
    def setup_aba_indices(self):
        """Aba para gerenciar índices mensais"""
        frame_indices = ttk.Frame(self.notebook)
        self.notebook.add(frame_indices, text='Índices Mensais')
        
        # Seleção de índice
        frame_selecao = ttk.Frame(frame_indices)
        frame_selecao.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_selecao, text="Índice:").pack(side='left', padx=5)
        self.combo_indice = ttk.Combobox(frame_selecao, state='readonly')
        self.combo_indice['values'] = list(self.gerenciador.config['indices_correcao']['indices_disponiveis'].keys())
        self.combo_indice.set('IGPM')
        self.combo_indice.pack(side='left', padx=5)
        self.combo_indice.bind('<<ComboboxSelected>>', self.carregar_historico_indice)
        
        # Frame para adicionar novo índice
        frame_novo = ttk.LabelFrame(frame_indices, text="Adicionar/Editar Índice")
        frame_novo.pack(fill='x', padx=5, pady=5)
        
        # Período
        ttk.Label(frame_novo, text="Período (MM/AAAA):").grid(row=0, column=0, padx=5, pady=5)
        self.entry_periodo = ttk.Entry(frame_novo, width=10)
        self.entry_periodo.grid(row=0, column=1, padx=5, pady=5)
        self.entry_periodo.insert(0, datetime.now().strftime('%m/%Y'))
        
        # Valor
        ttk.Label(frame_novo, text="Valor (%):").grid(row=0, column=2, padx=5, pady=5)
        self.entry_valor = ttk.Entry(frame_novo, width=10)
        self.entry_valor.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Button(frame_novo, text="Adicionar/Atualizar",
                  command=self.adicionar_indice).grid(row=0, column=4, padx=10, pady=5)
        
        # Lista de índices
        frame_lista = ttk.LabelFrame(frame_indices, text="Histórico de Índices")
        frame_lista.pack(fill='both', expand=True, padx=5, pady=5)
        
        colunas = ('Período', 'Valor (%)', 'Data Inclusão', 'Última Atualização')
        self.tree_indices = ttk.Treeview(frame_lista, columns=colunas, show='headings')
        for col in colunas:
            self.tree_indices.heading(col, text=col)
            self.tree_indices.column(col, width=120)
        
        scrollbar = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_indices.yview)
        self.tree_indices.configure(yscrollcommand=scrollbar.set)
        
        self.tree_indices.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Botão para remover
        ttk.Button(frame_lista, text="Remover Selecionado",
                  command=self.remover_indice).pack(pady=5)
    
    def setup_aba_correcao(self):
        """Aba para aplicar correção monetária"""
        frame_correcao = ttk.Frame(self.notebook)
        self.notebook.add(frame_correcao, text='Aplicar Correção')
        
        # Configurações de correção
        frame_config = ttk.LabelFrame(frame_correcao, text="Configurações")
        frame_config.pack(fill='x', padx=5, pady=5)
        
        # Cliente específico ou todos
        ttk.Label(frame_config, text="Cliente:").grid(row=0, column=0, padx=5, pady=5)
        self.combo_cliente = ttk.Combobox(frame_config)
        self.combo_cliente.grid(row=0, column=1, padx=5, pady=5)
        
        # Data de corte
        ttk.Label(frame_config, text="Data de Referência:").grid(row=1, column=0, padx=5, pady=5)
        self.data_corte = DateEntry(frame_config, width=12, locale='pt_BR',
                                   background='darkblue', foreground='white',
                                   borderwidth=2, date_pattern='dd/mm/yyyy')
        self.data_corte.grid(row=1, column=1, padx=5, pady=5)
        
        # Índice a utilizar
        ttk.Label(frame_config, text="Índice:").grid(row=2, column=0, padx=5, pady=5)
        self.combo_indice_correcao = ttk.Combobox(frame_config, state='readonly')
        self.combo_indice_correcao['values'] = list(self.gerenciador.config['indices_correcao']['indices_disponiveis'].keys())
        self.combo_indice_correcao.set(self.gerenciador.config['indices_correcao']['indice_padrao'])
        self.combo_indice_correcao.grid(row=2, column=1, padx=5, pady=5)
        
        # Botões de ação
        frame_acoes = ttk.Frame(frame_correcao)
        frame_acoes.pack(fill='x', padx=5, pady=10)
        
        ttk.Button(frame_acoes, text="Simular Correção",
                  command=self.simular_correcao).pack(side='left', padx=5)
        ttk.Button(frame_acoes, text="Aplicar Correção",
                  command=self.aplicar_correcao).pack(side='left', padx=5)
        
        # Área de resultados
        frame_resultados = ttk.LabelFrame(frame_correcao, text="Resultados")
        frame_resultados.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.text_resultados = tk.Text(frame_resultados, height=15, wrap=tk.WORD)
        scrollbar_result = ttk.Scrollbar(frame_resultados, orient='vertical', command=self.text_resultados.yview)
        self.text_resultados.configure(yscrollcommand=scrollbar_result.set)
        
        self.text_resultados.pack(side='left', fill='both', expand=True)
        scrollbar_result.pack(side='right', fill='y')
    
    def setup_aba_relatorios(self):
        """Aba para relatórios de correção"""
        frame_relatorios = ttk.Frame(self.notebook)
        self.notebook.add(frame_relatorios, text='Relatórios')
        
        ttk.Label(frame_relatorios, text="Funcionalidade de relatórios em desenvolvimento").pack(pady=20)
    
    def carregar_dados(self):
        """Carrega dados iniciais"""
        self.carregar_historico_indice()
        self.carregar_clientes()
    
    def carregar_clientes(self):
        """Carrega lista de clientes"""
        try:
            from src.config.config import PASTA_CLIENTES
            clientes = ['TODOS OS CLIENTES']
            
            for arquivo in PASTA_CLIENTES.glob('*.xlsx'):
                if not arquivo.name.startswith('~'):
                    clientes.append(arquivo.stem)
            
            self.combo_cliente['values'] = sorted(clientes)
            self.combo_cliente.set('TODOS OS CLIENTES')
            
        except Exception as e:
            print(f"Erro ao carregar clientes: {str(e)}")
            self.combo_cliente['values'] = ['TODOS OS CLIENTES']
            self.combo_cliente.set('TODOS OS CLIENTES')
    
    def carregar_historico_indice(self, event=None):
        """Carrega histórico do índice selecionado"""
        indice = self.combo_indice.get()
        
        # Limpar tree
        for item in self.tree_indices.get_children():
            self.tree_indices.delete(item)
        
        if indice in self.gerenciador.config['indices_correcao']['indices_disponiveis']:
            historico = self.gerenciador.config['indices_correcao']['indices_disponiveis'][indice]['historico']
            
            for registro in historico:
                data_inclusao = registro.get('data_inclusao', '')
                if data_inclusao:
                    try:
                        data_inc = datetime.fromisoformat(data_inclusao).strftime('%d/%m/%Y %H:%M')
                    except:
                        data_inc = data_inclusao
                else:
                    data_inc = ''
                
                data_atualizacao = registro.get('data_atualizacao', '')
                if data_atualizacao:
                    try:
                        data_atual = datetime.fromisoformat(data_atualizacao).strftime('%d/%m/%Y %H:%M')
                    except:
                        data_atual = data_atualizacao
                else:
                    data_atual = ''
                
                self.tree_indices.insert('', 'end', values=(
                    registro['periodo'],
                    f"{registro['valor']:.4f}",
                    data_inc,
                    data_atual
                ))
    
    def adicionar_indice(self):
        """Adiciona ou atualiza um índice"""
        try:
            indice = self.combo_indice.get()
            periodo = self.entry_periodo.get().strip()
            valor = self.entry_valor.get().strip()
            
            if not periodo or not valor:
                messagebox.showerror("Erro", "Preencha período e valor!")
                return
            
            # Validar formato do período
            try:
                datetime.strptime(periodo, '%m/%Y')
            except ValueError:
                messagebox.showerror("Erro", "Período deve estar no formato MM/AAAA!")
                return
            
            # Validar valor
            try:
                valor_float = float(valor.replace(',', '.'))
            except ValueError:
                messagebox.showerror("Erro", "Valor inválido!")
                return
            
            # Adicionar índice
            if self.gerenciador.adicionar_indice_periodo(indice, periodo, valor_float):
                messagebox.showinfo("Sucesso", f"Índice {indice} para {periodo} adicionado/atualizado!")
                self.carregar_historico_indice()
                self.entry_periodo.delete(0, tk.END)
                self.entry_valor.delete(0, tk.END)
                
                # Sugerir próximo período
                try:
                    mes, ano = periodo.split('/')
                    prox_data = datetime(int(ano), int(mes), 1) + relativedelta(months=1)
                    self.entry_periodo.insert(0, prox_data.strftime('%m/%Y'))
                except:
                    pass
            else:
                messagebox.showerror("Erro", "Erro ao adicionar índice!")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro: {str(e)}")
    
    def remover_indice(self):
        """Remove o índice selecionado"""
        selecionado = self.tree_indices.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um registro para remover!")
            return
        
        valores = self.tree_indices.item(selecionado)['values']
        periodo = valores[0]
        
        if messagebox.askyesno("Confirmar", f"Remover índice de {periodo}?"):
            try:
                indice = self.combo_indice.get()
                historico = self.gerenciador.config['indices_correcao']['indices_disponiveis'][indice]['historico']
                
                # Remover registro
                for i, registro in enumerate(historico):
                    if registro['periodo'] == periodo:
                        del historico[i]
                        break
                
                self.gerenciador.salvar_configuracoes()
                self.carregar_historico_indice()
                messagebox.showinfo("Sucesso", "Registro removido!")
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao remover: {str(e)}")
    
    def simular_correcao(self):
        """Simula a aplicação da correção monetária"""
        self.text_resultados.delete(1.0, tk.END)
        self.text_resultados.insert(tk.END, "SIMULAÇÃO DE CORREÇÃO MONETÁRIA\n")
        self.text_resultados.insert(tk.END, "=" * 50 + "\n\n")
        
        try:
            cliente = self.combo_cliente.get()
            data_corte = self.data_corte.get_date()
            indice = self.combo_indice_correcao.get()
            
            self.text_resultados.insert(tk.END, f"Cliente: {cliente}\n")
            self.text_resultados.insert(tk.END, f"Data de Referência: {data_corte.strftime('%d/%m/%Y')}\n")
            self.text_resultados.insert(tk.END, f"Índice: {indice}\n\n")
            
            # Temporariamente alterar o índice padrão
            indice_original = self.gerenciador.config['indices_correcao']['indice_padrao']
            self.gerenciador.config['indices_correcao']['indice_padrao'] = indice
            
            if cliente == 'TODOS OS CLIENTES':
                resultado = self.gerenciador.aplicar_correcao_contratos(None, data_corte)
            else:
                resultado = self.gerenciador.aplicar_correcao_contratos(cliente, data_corte)
            
            # Restaurar índice original
            self.gerenciador.config['indices_correcao']['indice_padrao'] = indice_original
            
            if resultado['sucesso']:
                self.text_resultados.insert(tk.END, f"SIMULAÇÃO CONCLUÍDA\n\n")
                self.text_resultados.insert(tk.END, f"Clientes processados: {resultado['clientes_processados']}\n\n")
                
                for detalhe in resultado['detalhes']:
                    self.text_resultados.insert(tk.END, f"CLIENTE: {detalhe['cliente']}\n")
                    self.text_resultados.insert(tk.END, f"Contratos processados: {detalhe['contratos_processados']}\n")
                    
                    for contrato in detalhe['detalhes']:
                        if contrato['sucesso']:
                            self.text_resultados.insert(tk.END, f"\n  Contrato: {contrato['contrato']}\n")
                            self.text_resultados.insert(tk.END, f"  Valor original: R$ {contrato['valor_original']:,.2f}\n")
                            self.text_resultados.insert(tk.END, f"  Valor corrigido: R$ {contrato['valor_corrigido']:,.2f}\n")
                            self.text_resultados.insert(tk.END, f"  Correção: {contrato['percentual_correcao']:.4f}%\n")
                            self.text_resultados.insert(tk.END, f"  Período: {contrato['periodo_correcao']}\n")
                    
                    self.text_resultados.insert(tk.END, "\n" + "-" * 30 + "\n")
            else:
                self.text_resultados.insert(tk.END, f"ERRO: {resultado['erro']}\n")
                
        except Exception as e:
            self.text_resultados.insert(tk.END, f"ERRO NA SIMULAÇÃO: {str(e)}\n")
    
    def aplicar_correcao(self):
        """Aplica efetivamente a correção monetária"""
        if not messagebox.askyesno("Confirmar", 
                                  "Deseja realmente aplicar a correção monetária?\n\n"
                                  "Esta ação alterará os valores dos contratos e não pode ser desfeita automaticamente!"):
            return
        
        self.text_resultados.delete(1.0, tk.END)
        self.text_resultados.insert(tk.END, "APLICANDO CORREÇÃO MONETÁRIA...\n")
        self.text_resultados.insert(tk.END, "=" * 50 + "\n\n")
        
        try:
            cliente = self.combo_cliente.get()
            data_corte = self.data_corte.get_date()
            indice = self.combo_indice_correcao.get()
            
            # Alterar índice padrão permanentemente
            self.gerenciador.config['indices_correcao']['indice_padrao'] = indice
            self.gerenciador.salvar_configuracoes()
            
            if cliente == 'TODOS OS CLIENTES':
                resultado = self.gerenciador.aplicar_correcao_contratos(None, data_corte)
            else:
                resultado = self.gerenciador.aplicar_correcao_contratos(cliente, data_corte)
            
            if resultado['sucesso']:
                self.text_resultados.insert(tk.END, "CORREÇÃO APLICADA COM SUCESSO!\n\n")
                self.text_resultados.insert(tk.END, f"Clientes processados: {resultado['clientes_processados']}\n")
                self.text_resultados.insert(tk.END, f"Data de processamento: {resultado['data_processamento']}\n\n")
                
                # Registrar histórico de correção
                self.registrar_historico_correcao(resultado)
                
                messagebox.showinfo("Sucesso", "Correção monetária aplicada com sucesso!")
            else:
                self.text_resultados.insert(tk.END, f"ERRO: {resultado['erro']}\n")
                messagebox.showerror("Erro", f"Erro ao aplicar correção: {resultado['erro']}")
                
        except Exception as e:
            erro_msg = f"Erro ao aplicar correção: {str(e)}"
            self.text_resultados.insert(tk.END, f"ERRO: {erro_msg}\n")
            messagebox.showerror("Erro", erro_msg)
    
    def registrar_historico_correcao(self, resultado):
        """Registra o histórico de correções aplicadas"""
        try:
            if 'historico_correcoes' not in self.gerenciador.config:
                self.gerenciador.config['historico_correcoes'] = []
            
            registro_historico = {
                'data_aplicacao': datetime.now().isoformat(),
                'usuario': 'Sistema',  # Você pode implementar controle de usuário
                'resultado': resultado,
                'resumo': {
                    'clientes_processados': resultado['clientes_processados'],
                    'total_contratos': sum(d['contratos_processados'] for d in resultado['detalhes'])
                }
            }
            
            self.gerenciador.config['historico_correcoes'].append(registro_historico)
            self.gerenciador.salvar_configuracoes()
            
        except Exception as e:
            print(f"Erro ao registrar histórico: {str(e)}")
    
    def salvar_alteracoes(self):
        """Salva todas as alterações"""
        try:
            self.gerenciador.salvar_configuracoes()
            messagebox.showinfo("Sucesso", "Configurações salvas com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar: {str(e)}")

# ===== FUNÇÕES DE INTEGRAÇÃO =====

def adicionar_correcao_monetaria_ao_menu():
    """
    Função para adicionar a opção de correção monetária ao menu principal
    """
    def abrir_indices_correcao():
        try:
            app = InterfaceIndicesCorrecao()
            app.root.mainloop()
        except Exception as e:
            print(f"Erro ao abrir interface de índices: {str(e)}")
            import tkinter.messagebox as msgbox
            msgbox.showerror("Erro", f"Erro ao abrir correção monetária: {str(e)}")
    
    return abrir_indices_correcao