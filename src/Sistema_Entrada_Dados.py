import os
import sys
import traceback

# Configurar diretório de log no desktop do usuário
log_dir = os.path.join(os.path.expanduser("~"), "Desktop")
log_file = os.path.join(log_dir, "sistema_log.txt")

def log_error(message):
    """Grava mensagens de erro em arquivo no desktop"""
    try:
        with open(log_file, "a") as f:
            f.write(f"{message}\n")
    except:
        pass

# Capturar e registrar exceções não tratadas
def handle_exception(exc_type, exc_value, exc_traceback):
    """Manipulador para exceções não capturadas"""
    error_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    log_error(f"\n\n--- ERRO NÃO TRATADO: {error_msg}")
    # Mostrar mensagem simples para o usuário
    import tkinter.messagebox as msgbox
    msgbox.showerror("Erro", f"Ocorreu um erro: {str(exc_value)}\nDetalhes foram registrados em: {log_file}")
    
# Substituir o manipulador de exceções padrão
sys.excepthook = handle_exception

log_error(f"\n\n--- INICIANDO APLICAÇÃO: {sys.argv[0]} ---")

# Imports da biblioteca padrão Python
import os
import sys
from pathlib import Path
import re
import calendar
from datetime import datetime, timedelta
from decimal import Decimal

# Imports relacionados ao Tkinter (MUITO IMPORTANTE)
import tkinter as tk
from tkinter import ttk, messagebox, StringVar
from tkinter import *
from tkcalendar import DateEntry, Calendar
from tkinter import filedialog, messagebox

# Imports para manipulação de dados e Excel
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import openpyxl
import babel
from dateutil.relativedelta import relativedelta

import requests
import json
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
from urllib.parse import urljoin

# Detectar modo PyInstaller e ajustar paths
if getattr(sys, 'frozen', False):
    # Estamos em um executável criado pelo PyInstaller
    base_dir = Path(sys._MEIPASS)
    # Garantir que src e src/config estão no path
    for subdir in ['src', os.path.join('src', 'config')]:
        path = os.path.join(base_dir, subdir)
        if path not in sys.path:
            sys.path.insert(0, path)
            print(f"PyInstaller: Adicionando {path} ao sys.path")

# Configurar caminhos de importação
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    config_dir = current_dir / 'config'
    
    # Adicionar todos os caminhos necessários
    for path in [str(current_dir), str(project_root), str(config_dir)]:
        if path not in sys.path:
            sys.path.insert(0, path)
            print(f"Adicionado ao path: {path}")

add_project_root()

# Configurar logging básico para diagnóstico
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger("sistema")

# Importações centralizadas com tratamento de erro
try:
    # Primeiro, tentar importação com 'src.config'
    from src.config.utils import *
    from src.config.dialogs import custom_messagebox
    from src.config.logger_config import system_logger, log_action
    from src.config.window_config import configurar_janela
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH,
        ARQUIVO_FORNECEDORES
    )
    logger.info("Configurações importadas com sucesso via src.config")
except ImportError:
    # Se falhar, tentar importação direta de 'config'
    try:
        from config.utils import *
        from config.logger_config import system_logger, log_action
        from config.window_config import configurar_janela
        from config.config import (
            ARQUIVO_CLIENTES,
            ARQUIVO_MODELO,
            PASTA_CLIENTES,
            BASE_PATH,
            ARQUIVO_FORNECEDORES
        )
        logger.info("Configurações importadas com sucesso via config")
    except ImportError as e:
        # Último recurso - importações relativas
        try:
            from src.config.utils import *
            from src.config.logger_config import system_logger, log_action
            from src.config.window_config import configurar_janela
            from src.config.config import (
                ARQUIVO_CLIENTES,
                ARQUIVO_MODELO,
                PASTA_CLIENTES,
                BASE_PATH,
                ARQUIVO_FORNECEDORES
            )
            logger.info("Configurações importadas com sucesso via .config")
        except Exception as e:
            logger.error(f"Erro ao importar configurações: {str(e)}")
            # Não raise aqui para permitir definições alternativas

# Gestão de Locações
from src.gestao_locacoes import GerenciadorLocacoes

# from src.nfe.sistema_nfe_unificado import substituir_sistemas_nfe_por_unificado
from src.materiais.gerenciador_materiais import inicializar_sistema_materiais_completo

# Definir funções de compatibilidade, caso a importação das configurações falhe
def get_categorias_fornecedor():
    """Retorna a lista de categorias de fornecedor"""
    try:
        # Tentar importar do arquivo de configurações
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        return GerenciadorConfiguracoes.get_categorias_fornecedor()
    except:
        # Valores padrão como fallback
        return ['MO', 'MAT', 'SERV', 'DIV', 'ADM', 'LOC', 'TP']

@log_action("Carregar configurações")
def carregar_configuracoes():
    """Carrega as configurações do sistema"""
    try:
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        return GerenciadorConfiguracoes.carregar_configuracoes()
    except:
        # Configurações padrão como fallback
        return {
            'cafe': {'valor_atual': 4.0}
        }

def get_bancos():
    """Retorna a lista de bancos"""
    try:
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        return GerenciadorConfiguracoes.get_bancos()
    except:
        # Valores padrão como fallback
        return [
            '104 - CAIXA ECONÔMICA FEDERAL',
            '001 - BANCO DO BRASIL',
            '033 - BANCO SANTANDER',
            '237 - BRADESCO',
            '341 - ITAÚ',
            '077 - BANCO INTER',
            '260 - NUBANK'
        ]

# Importar configurações
try:
    from src.config.utils import *
    from src.configuracoes_sistema import GerenciadorConfiguracoes
    logger.info("Configurações importadas com sucesso")
except Exception as e:
    logger.error(f"Erro ao importar configurações: {str(e)}")
    raise

# Importar configurações do sistema
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH
    )
    logger.info("Configurações do sistema importadas com sucesso")
except Exception as e:
    logger.error(f"Erro ao importar configurações do sistema: {str(e)}")
    raise

from src.config.window_config import configurar_janela
    

# Modificação para usar o método de utils.py
from src.config.utils import buscar_dados_bancarios_fornecedor


class VisualizadorLancamentos:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.janela = tk.Toplevel(sistema_principal.root)
        configurar_janela(self.janela, "Visualização de Lançamentos Pendentes", 1000, 400)

        # Registrar esta janela para os diálogos
        try:
            from src.config.dialogs import set_main_window
            set_main_window(self.janela)
        except ImportError:
            pass

        self.alteracoes = False
        self.dados_para_incluir = []
        self._fechando = False
        self._dialogo_aberto = False
        
        # Configurar comportamento quando a janela é fechada
        self.janela.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # ESTRATÉGIA BALANCEADA
        self.janela.lift()
        self.janela.focus_force()

        # Frame principal
        self.frame_principal = ttk.Frame(self.janela)
        self.frame_principal.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar Treeview para visualização
        self.criar_treeview()
        
        # Frame para resumo
        self.frame_resumo = ttk.LabelFrame(self.frame_principal, text="Resumo")
        self.frame_resumo.pack(fill='x', pady=5)
        
        self.lbl_total_lancamentos = ttk.Label(self.frame_resumo, text="Total de Lançamentos: 0")
        self.lbl_total_lancamentos.pack(side='left', padx=5)
        
        self.lbl_valor_total = ttk.Label(self.frame_resumo, text="Valor Total: R$ 0,00")
        self.lbl_valor_total.pack(side='left', padx=5)
        
        self.lbl_selecionados = ttk.Label(self.frame_resumo, text="Marcados: 0", 
                                         foreground='blue', font=('TkDefaultFont', 9, 'bold'))
        self.lbl_selecionados.pack(side='left', padx=5)
        
        # Frame para botões de seleção
        self.frame_selecao = ttk.Frame(self.frame_principal)
        self.frame_selecao.pack(fill='x', pady=5)
        
        ttk.Label(self.frame_selecao, text="Marcar:").pack(side='left', padx=5)
        ttk.Button(self.frame_selecao, text="✓ Todos", 
                  command=self.selecionar_todos).pack(side='left', padx=2)
        ttk.Button(self.frame_selecao, text="✗ Nenhum", 
                  command=self.desmarcar_todos).pack(side='left', padx=2)
        ttk.Button(self.frame_selecao, text="⇄ Inverter", 
                  command=self.inverter_selecao).pack(side='left', padx=2)
        
        # Frame para botões principais
        self.frame_botoes = ttk.Frame(self.frame_principal)
        self.frame_botoes.pack(fill='x', pady=5)
        
        ttk.Button(self.frame_botoes, text="✏️ Editar", 
                  command=self.editar_lancamento).pack(side='left', padx=5)
        
        ttk.Button(self.frame_botoes, text="📝 Editar em Massa", 
                    command=self.editar_em_massa).pack(side='left', padx=5)
        
        # Botão de exclusão único - funciona para 1 ou vários
        self.btn_excluir = tk.Button(self.frame_botoes, text="🗑️ Excluir Marcados", 
                                     command=self.excluir_marcados,
                                     bg='#dc3545', fg='white', 
                                     font=('TkDefaultFont', 9, 'bold'),
                                     relief='raised', bd=2, cursor='hand2')
        self.btn_excluir.pack(side='left', padx=5)
        
        ttk.Button(self.frame_botoes, text="📂 Carregar Rascunho", 
                  command=self.carregar_rascunho).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="💾 Salvar na Planilha", 
                  command=self.salvar_na_planilha).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="Fechar", 
                  command=self.fechar_janela).pack(side='right', padx=5)
        
        # Monitorar perda de foco
        self.janela.bind('<FocusOut>', self._on_focus_out)
    
    def _on_focus_out(self, event):
        """Quando janela perde foco, verifica se deve trazer de volta"""
        if not self._dialogo_aberto and not self._fechando:
            self.janela.after(200, self._verificar_trazer_frente)
    
    def _verificar_trazer_frente(self):
        """Verifica se deve trazer janela para frente"""
        try:
            if (self.janela and self.janela.winfo_exists() and 
                not self._dialogo_aberto and not self._fechando):
                if not self._tem_dialogo_filho():
                    self.janela.lift()
                    self.janela.focus_force()
        except:
            pass
    
    def _tem_dialogo_filho(self):
        """Verifica se há algum diálogo filho aberto"""
        try:
            widget_com_foco = self.janela.focus_get()
            if widget_com_foco:
                parent = widget_com_foco.winfo_toplevel()
                if parent != self.janela and isinstance(parent, tk.Toplevel):
                    return True
            return False
        except:
            return False
    
    def fechar_janela(self):
        """Fecha a janela de forma segura"""
        self._fechando = True
        if hasattr(self.sistema, 'on_visualizador_close'):
            self.sistema.on_visualizador_close()
        else:
            self.janela.destroy()
    
    def on_close(self):
        """Manipula o fechamento da janela"""
        self.fechar_janela()

    def criar_treeview(self):
        """Cria a TreeView com coluna de checkbox CLICÁVEL"""
        colunas = ('☑', 'Data', 'Tipo', 'CNPJ/CPF', 'Nome', 'Referência', 'NF', 'Vr. Unit.', 
                   'Dias', 'Valor', 'Vencimento', 'Categoria', 'Forma Pagamento', 
                   'Dados Bancários', 'Observação')
        
        self.tree = ttk.Treeview(self.frame_principal, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree.heading(col, text=col)
            
            if col == '☑':
                width = 30
            elif col in ['CNPJ/CPF', 'Nome', 'Referência', 'Dados Bancários', 'Observação']:
                width = 150
            elif col in ['Data', 'Vencimento']:
                width = 100
            elif col in ['Vr. Unit.', 'Valor', 'NF']:
                width = 100
            elif col == 'Forma Pagamento':
                width = 80
            else:
                width = 80
            self.tree.column(col, width=width)

        scrolly = ttk.Scrollbar(self.frame_principal, orient='vertical', command=self.tree.yview)
        scrollx = ttk.Scrollbar(self.frame_principal, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree.pack(fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Clique simples marca/desmarca checkbox
        self.tree.bind('<Button-1>', self.on_tree_click)
    
    def on_tree_click(self, event):
        """Detecta clique na coluna do checkbox e alterna"""
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            if column == '#1':
                item = self.tree.identify_row(event.y)
                if item:
                    self.toggle_checkbox(item)
                    return "break"
    
    def toggle_checkbox(self, item_id):
        """Alterna o estado do checkbox de um item"""
        valores = list(self.tree.item(item_id)['values'])
        if valores[0] in ['☐', '']:
            valores[0] = '☑'
        else:
            valores[0] = '☐'
        self.tree.item(item_id, values=valores)
        self.atualizar_contador_selecionados()
    
    def selecionar_todos(self):
        """Marca todos os checkboxes"""
        for item in self.tree.get_children():
            valores = list(self.tree.item(item)['values'])
            valores[0] = '☑'
            self.tree.item(item, values=valores)
        self.atualizar_contador_selecionados()
    
    def desmarcar_todos(self):
        """Desmarca todos os checkboxes"""
        for item in self.tree.get_children():
            valores = list(self.tree.item(item)['values'])
            valores[0] = '☐'
            self.tree.item(item, values=valores)
        self.atualizar_contador_selecionados()
    
    def inverter_selecao(self):
        """Inverte a seleção dos checkboxes"""
        for item in self.tree.get_children():
            valores = list(self.tree.item(item)['values'])
            valores[0] = '☑' if valores[0] == '☐' else '☐'
            self.tree.item(item, values=valores)
        self.atualizar_contador_selecionados()
    
    def obter_indices_marcados(self):
        """Retorna lista de índices dos itens marcados"""
        indices_marcados = []
        todos_items = self.tree.get_children()
        
        for idx, item in enumerate(todos_items):
            valores = self.tree.item(item)['values']
            if valores[0] == '☑':
                indices_marcados.append(idx)
        
        return indices_marcados
    
    def excluir_marcados(self):
        """
        Exclui todos os lançamentos marcados PERMANENTEMENTE
        Funciona para 1 ou vários itens marcados
        """
        indices_marcados = self.obter_indices_marcados()
        
        if not indices_marcados:
            self._dialogo_aberto = True
            custom_messagebox("warning", "⚠️ Nenhum Item Marcado", 
                            "Para excluir lançamentos:\n\n"
                            "1. Clique no ☑ para marcar os itens desejados\n"
                            "2. Clique em 'Excluir Marcados'\n\n"
                            "💡 Dica: Use os botões 'Todos', 'Nenhum' ou 'Inverter' "
                            "para facilitar a seleção!")
            self._dialogo_aberto = False
            self.janela.lift()
            return
        
        # MARCAR que há diálogo aberto
        self._dialogo_aberto = True
        
        # Mensagem personalizada para singular/plural
        qtd = len(indices_marcados)
        if qtd == 1:
            titulo = "⚠️ Excluir 1 Lançamento"
            mensagem = (f"Você está prestes a EXCLUIR PERMANENTEMENTE "
                       f"1 lançamento!\n\n"
                       f"❌ Esta ação NÃO PODE ser desfeita!\n"
                       f"❌ O dado será removido IMEDIATAMENTE!\n"
                       f"❌ O rascunho será atualizado AGORA!\n\n"
                       f"Confirma a EXCLUSÃO DEFINITIVA?")
        else:
            titulo = f"⚠️ Excluir {qtd} Lançamentos"
            mensagem = (f"Você está prestes a EXCLUIR PERMANENTEMENTE "
                       f"{qtd} lançamentos!\n\n"
                       f"❌ Esta ação NÃO PODE ser desfeita!\n"
                       f"❌ Os dados serão removidos IMEDIATAMENTE!\n"
                       f"❌ O rascunho será atualizado AGORA!\n\n"
                       f"Confirma a EXCLUSÃO DEFINITIVA?")
        
        resposta = custom_messagebox("yesno", titulo, mensagem)
        
        # MARCAR que diálogo foi fechado
        self._dialogo_aberto = False
        
        if resposta:
            try:
                # REMOVER IMEDIATAMENTE
                itens_removidos = self.remover_itens_especificos(indices_marcados)
                
                # ATUALIZAR SISTEMA PRINCIPAL
                self.sistema.dados_para_incluir = self.dados_para_incluir.copy()
                
                # SALVAR RASCUNHO ATUALIZADO
                self.salvar_rascunho_imediatamente()
                
                # Verificar se ainda há dados
                self._dialogo_aberto = True
                if len(self.dados_para_incluir) == 0:
                    if qtd == 1:
                        msg = "1 lançamento EXCLUÍDO PERMANENTEMENTE!"
                    else:
                        msg = f"{itens_removidos} lançamentos EXCLUÍDOS PERMANENTEMENTE!"
                    
                    custom_messagebox("info", "✅ Exclusão Concluída", 
                                    f"{msg}\n\n"
                                    f"✅ Rascunho ELIMINADO!\n"
                                    f"✅ Não há mais lançamentos pendentes.")
                    self._dialogo_aberto = False
                    self.fechar_janela()
                    return
                else:
                    if qtd == 1:
                        msg = "1 lançamento EXCLUÍDO PERMANENTEMENTE!"
                    else:
                        msg = f"{itens_removidos} lançamentos EXCLUÍDOS PERMANENTEMENTE!"
                    
                    custom_messagebox("info", "✅ Exclusão Concluída", 
                                    f"{msg}\n\n✅ Rascunho ATUALIZADO!")
                self._dialogo_aberto = False
                
            except Exception as e:
                self._dialogo_aberto = True
                custom_messagebox("error", "Erro", 
                                f"Erro ao excluir lançamentos:\n{str(e)}")
                self._dialogo_aberto = False
                import traceback
                traceback.print_exc()
        
        # Trazer janela de volta
        self.janela.lift()
        self.janela.focus_force()
    
    def editar_em_massa(self):
        """Abre editor para alteração em massa dos lançamentos marcados"""
        indices_marcados = self.obter_indices_marcados()
        
        if not indices_marcados:
            self._dialogo_aberto = True
            custom_messagebox("warning", "⚠️ Nenhum Item Marcado", 
                            "Para editar em massa:\n\n"
                            "1. Marque os lançamentos desejados com ☑\n"
                            "2. Clique em 'Editar em Massa'\n\n"
                            "💡 Use os botões auxiliares para marcar múltiplos itens!")
            self._dialogo_aberto = False
            self.janela.lift()
            return
        
        if len(indices_marcados) == 1:
            self._dialogo_aberto = True
            resposta = custom_messagebox("yesno", 
                "Edição Individual ou em Massa?",
                "Você marcou apenas 1 lançamento.\n\n"
                "Deseja usar o editor em massa mesmo assim?\n"
                "(Não = abrirá o editor individual normal)")
            self._dialogo_aberto = False
            
            if not resposta:
                # Abrir editor individual
                item = self.tree.get_children()[indices_marcados[0]]
                self.tree.selection_set(item)
                self.editar_lancamento()
                return
        
        # Coletar dados dos lançamentos marcados
        dados_selecionados = [self.dados_para_incluir[idx] for idx in indices_marcados]
        
        # Abrir editor em massa
        self._dialogo_aberto = True
        editor = EditorEmMassa(self.janela, dados_selecionados, indices_marcados, 
                            self.atualizar_lancamento)
        self._dialogo_aberto = False

    def salvar_rascunho_imediatamente(self):
        """SALVA ou REMOVE o rascunho IMEDIATAMENTE"""
        try:
            temp_file = os.path.join(os.path.expanduser("~"), "Desktop", 
                                    "backup_lancamentos.json")
            
            if not self.dados_para_incluir or len(self.dados_para_incluir) == 0:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    print(f"✅ Rascunho DELETADO: {temp_file}")
                return
            
            backup_data = {
                'cliente': self.sistema.cliente_atual if hasattr(self.sistema, 'cliente_atual') else '',
                'total_lancamentos': len(self.dados_para_incluir),
                'data_sessao': datetime.now().isoformat(),
                'lancamentos': self.dados_para_incluir
            }
            
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(backup_data, f, ensure_ascii=False, indent=2)
            
            print(f"✅ Rascunho SALVO: {temp_file} ({len(self.dados_para_incluir)} lançamentos)")
                
        except Exception as e:
            print(f"❌ Erro ao salvar rascunho: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def atualizar_contador_selecionados(self):
        """Atualiza o contador de itens selecionados"""
        qtd_selecionados = len(self.obter_indices_marcados())
        
        if qtd_selecionados == 1:
            texto = "Marcado: 1"
        else:
            texto = f"Marcados: {qtd_selecionados}"
        
        self.lbl_selecionados.config(text=texto)
        
        # Mudar cor do botão baseado na seleção
        if qtd_selecionados > 0:
            self.btn_excluir.config(bg='#ff0000', state='normal')
            if qtd_selecionados == 1:
                self.btn_excluir.config(text='🗑️ Excluir Marcado')
            else:
                self.btn_excluir.config(text='🗑️ Excluir Marcados')
        else:
            self.btn_excluir.config(bg='#dc3545', state='normal', 
                                   text='🗑️ Excluir Marcados')
    
    def atualizar_dados(self, dados):
        """Atualiza os dados na visualização"""
        self.dados_para_incluir = dados.copy() if dados else []
        
        if hasattr(self, 'dados_originais'):
            self.dados_originais = dados.copy() if dados else []
        else:
            self.dados_originais = dados.copy() if dados else []
        
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        if not dados:
            self.lbl_total_lancamentos.config(text="Total de Lançamentos: 0")
            self.lbl_valor_total.config(text="Valor Total: R$ 0,00")
            self.lbl_selecionados.config(text="Marcados: 0")
            if hasattr(self, 'janela'):
                self.janela.title("Visualização de Lançamentos Pendentes - 0 registros")
            return
            
        valor_total = 0
        for lancamento in self.dados_para_incluir:
            valores = (
                '☐',
                lancamento['data'],
                lancamento['tp_desp'],
                lancamento['cnpj_cpf'],
                lancamento['nome'],
                lancamento['referencia'],
                lancamento.get('nf', ''),
                lancamento['vr_unit'],
                lancamento['dias'],
                lancamento['valor'],
                lancamento['dt_vencto'],
                lancamento['categoria'],
                lancamento.get('forma_pagamento', ''),
                lancamento['dados_bancarios'],
                lancamento['observacao']
            )
            self.tree.insert('', 'end', values=valores)
            
            try:
                valor_total += float(str(lancamento['valor']).replace(',', '.'))
            except (ValueError, TypeError):
                pass
        
        self.lbl_total_lancamentos.config(text=f"Total de Lançamentos: {len(dados)}")
        self.lbl_valor_total.config(text=f"Valor Total: R$ {valor_total:,.2f}")
        self.lbl_selecionados.config(text="Marcados: 0")
        
        if hasattr(self, 'janela'):
            self.janela.title(f"Visualização de Lançamentos Pendentes - {len(dados)} registros")

    def editar_lancamento(self):
        """Abre a janela de edição para o lançamento selecionado na TreeView"""
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            self._dialogo_aberto = True
            custom_messagebox("warning", "Aviso", 
                            "Para editar um lançamento:\n\n"
                            "Clique na LINHA desejada para selecioná-la\n"
                            "(não é o checkbox ☑, é a linha inteira)")
            self._dialogo_aberto = False
            self.janela.lift()
            return

        todos_items = self.tree.get_children()
        indice = todos_items.index(item_selecionado[0])
        
        valores = self.tree.item(item_selecionado)['values']
        dados = {
            'data': valores[1],
            'tp_desp': valores[2],
            'cnpj_cpf': valores[3],
            'nome': valores[4],
            'referencia': valores[5],
            'nf': valores[6],
            'vr_unit': valores[7],
            'dias': valores[8],
            'valor': valores[9],
            'dt_vencto': valores[10],
            'categoria': valores[11],
            'forma_pagamento': valores[12],
            'dados_bancarios': valores[13],
            'observacao': valores[14] if len(valores) > 14 else ''
        }
        
        editor = EditorLancamento(self.janela, dados, indice, self.atualizar_lancamento)

    def atualizar_lancamento(self, indice, novos_dados):
        """Atualiza os dados de um lançamento específico"""
        try:
            cnpj_cpf = str(novos_dados['cnpj_cpf']).replace('.', '').replace('-', '').replace('/', '')
            novos_dados['cnpj_cpf'] = formatar_cnpj_cpf(cnpj_cpf)
            novos_dados['observacao'] = novos_dados['observacao'].upper()

            item = self.tree.get_children()[indice]
            valores_atuais = self.tree.item(item)['values']
            
            valores = (
                valores_atuais[0],
                novos_dados['data'],
                novos_dados['tp_desp'],
                novos_dados['cnpj_cpf'],
                novos_dados['nome'],
                novos_dados['referencia'],
                novos_dados['nf'],
                novos_dados['vr_unit'],
                novos_dados['dias'],
                novos_dados['valor'],
                novos_dados['dt_vencto'],
                novos_dados['categoria'],
                novos_dados.get('forma_pagamento', ''),
                novos_dados['dados_bancarios'],
                novos_dados['observacao']
            )
            
            self.dados_para_incluir[indice] = novos_dados.copy()
            self.tree.item(item, values=valores)
            self.atualizar_resumo()
            
            self.sistema.dados_para_incluir = self.dados_para_incluir.copy()
            self.salvar_rascunho_imediatamente()
            
            return True
        except Exception as e:
            print(f"Erro ao atualizar lançamento: {str(e)}")
            return False

    def salvar_na_planilha(self):
        """Salva os dados diretamente na planilha"""
        if not self.dados_para_incluir:
            self._dialogo_aberto = True
            custom_messagebox("warning", "Aviso", "Não há dados para salvar!")
            self._dialogo_aberto = False
            self.janela.lift()
            return

        try:
            self.sistema.dados_para_incluir = self.dados_para_incluir.copy()
            self._fechando = True
            
            if self.sistema:
                self.sistema.enviar_dados()
                
                try:
                    temp_file = os.path.join(os.path.expanduser("~"), "Desktop", 
                                            "backup_lancamentos.json")
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        print("✅ Rascunho DELETADO após salvamento na planilha")
                except:
                    pass
                
                if self.janela and self.janela.winfo_exists():
                    self.janela.destroy()
            else:
                self._fechando = False
                self._dialogo_aberto = True
                custom_messagebox("error", "Erro", 
                                "Referência ao sistema principal não encontrada")
                self._dialogo_aberto = False
                self.janela.lift()

        except Exception as e:
            self._fechando = False
            self._dialogo_aberto = True
            custom_messagebox("error", "Erro", f"Erro ao salvar dados: {str(e)}")
            self._dialogo_aberto = False
            self.janela.lift()
            print(f"Erro detalhado ao salvar: {str(e)}")

    def atualizar_resumo(self):
        """Atualiza os totais e resumo"""
        items = self.tree.get_children()
        total_lancamentos = len(items)
        valor_total = 0
        
        for item in items:
            try:
                valor_total += float(self.tree.item(item)['values'][9])
            except (ValueError, TypeError, IndexError):
                pass
        
        self.lbl_total_lancamentos.config(text=f"Total de Lançamentos: {total_lancamentos}")
        self.lbl_valor_total.config(text=f"Valor Total: R$ {valor_total:,.2f}")
        self.atualizar_contador_selecionados()

    def get_dados_atualizados(self):
        """Retorna todos os dados atualizados"""
        return self.dados_para_incluir.copy()

    def carregar_rascunho(self):
        """Carrega dados do arquivo de backup"""
        try:
            temp_file = os.path.join(os.path.expanduser("~"), "Desktop", 
                                    "backup_lancamentos.json")
            
            if os.path.exists(temp_file):
                with open(temp_file, 'r', encoding='utf-8') as f:
                    backup_data = json.load(f)
                
                data_backup = datetime.fromisoformat(backup_data['data_sessao'])
                
                self._dialogo_aberto = True
                
                resposta = custom_messagebox("yesno", 
                    "📂 Carregar Rascunho",
                    f"Rascunho encontrado:\n\n"
                    f"• Cliente: {backup_data['cliente']}\n"
                    f"• Lançamentos: {backup_data['total_lancamentos']}\n"
                    f"• Salvo em: {data_backup.strftime('%d/%m/%Y às %H:%M')}\n\n"
                    "Carregar estes dados?\n"
                    "(Os dados atuais serão substituídos)")
                
                if resposta:
                    self.sistema.dados_para_incluir = backup_data['lancamentos']
                    self.sistema.cliente_atual = backup_data['cliente']
                    
                    self.dados_para_incluir = backup_data['lancamentos'].copy()
                    self.atualizar_dados(backup_data['lancamentos'])
                    
                    if hasattr(self.sistema, 'cliente_combobox'):
                        self.sistema.cliente_combobox.set(backup_data['cliente'])
                        self.sistema.selecionar_cliente(None)
                    
                    custom_messagebox("info", "✅ Rascunho Carregado", 
                                    f"Dados carregados com sucesso!\n"
                                    f"{len(backup_data['lancamentos'])} lançamentos carregados.")
                
                self._dialogo_aberto = False
                self.janela.lift()
            else:
                self._dialogo_aberto = True
                custom_messagebox("info", "📂 Rascunho", 
                                "Nenhum rascunho encontrado no Desktop.")
                self._dialogo_aberto = False
                self.janela.lift()
                
        except Exception as e:
            self._dialogo_aberto = True
            custom_messagebox("error", "Erro", f"Erro ao carregar rascunho: {str(e)}")
            self._dialogo_aberto = False
            self.janela.lift()
            import traceback
            traceback.print_exc()

    def popular_tree(self, dados_lancamentos):
        """Popula a TreeView com os dados fornecidos"""
        try:
            self.dados_originais = dados_lancamentos.copy() if dados_lancamentos else []
            
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            if not dados_lancamentos:
                self.atualizar_contador()
                return
            
            valor_total = 0
            for lancamento in dados_lancamentos:
                valores = (
                    '☐',
                    lancamento.get('data', ''),
                    lancamento.get('tp_desp', ''),
                    lancamento.get('cnpj_cpf', ''),
                    lancamento.get('nome', ''),
                    lancamento.get('referencia', ''),
                    lancamento.get('nf', ''),
                    lancamento.get('vr_unit', ''),
                    lancamento.get('dias', '1'),
                    lancamento.get('valor', ''),
                    lancamento.get('dt_vencto', ''),
                    lancamento.get('categoria', ''),
                    lancamento.get('forma_pagamento', ''),
                    lancamento.get('dados_bancarios', ''),
                    lancamento.get('observacao', '')
                )
                
                self.tree.insert('', 'end', values=valores)
                
                try:
                    valor_total += float(str(lancamento.get('valor', 0)).replace(',', '.'))
                except (ValueError, TypeError):
                    pass
            
            self.atualizar_contador()
            
        except Exception as e:
            print(f"Erro ao popular tree: {str(e)}")
            import traceback
            traceback.print_exc()

    def remover_itens_especificos(self, indices_para_remover):
        """Remove itens específicos da TreeView e dos dados"""
        try:
            todos_items = self.tree.get_children()
            
            indices_validos = [idx for idx in indices_para_remover 
                            if 0 <= idx < len(todos_items)]
            
            if not indices_validos:
                return 0
            
            items_removidos = []
            for idx in sorted(indices_validos, reverse=True):
                if idx < len(todos_items):
                    item_id = todos_items[idx]
                    self.tree.delete(item_id)
                    items_removidos.append(idx)
            
            novos_dados = []
            for idx, dados in enumerate(self.dados_para_incluir):
                if idx not in indices_validos:
                    novos_dados.append(dados)
            
            self.dados_para_incluir = novos_dados
            
            if hasattr(self, 'dados_originais'):
                novos_dados_originais = []
                for idx, dados in enumerate(self.dados_originais):
                    if idx not in indices_validos:
                        novos_dados_originais.append(dados)
                self.dados_originais = novos_dados_originais
            
            self.atualizar_contador()
            
            return len(items_removidos)
            
        except Exception as e:
            print(f"Erro ao remover itens específicos: {str(e)}")
            import traceback
            traceback.print_exc()
            return 0

    def atualizar_contador(self):
        """Atualiza o contador de registros e valor total"""
        try:
            qtd_registros = len(self.tree.get_children())
            valor_total = 0
            
            for item in self.tree.get_children():
                try:
                    valores = self.tree.item(item)['values']
                    valor_str = str(valores[9]).replace(',', '.')
                    valor_total += float(valor_str)
                except (ValueError, TypeError, IndexError):
                    continue
            
            if hasattr(self, 'lbl_total_lancamentos'):
                self.lbl_total_lancamentos.config(
                    text=f"Total de Lançamentos: {qtd_registros}"
                )
            
            if hasattr(self, 'lbl_valor_total'):
                self.lbl_valor_total.config(
                    text=f"Valor Total: R$ {valor_total:,.2f}"
                )
            
            if hasattr(self, 'janela') and self.janela:
                self.janela.title(f"Visualização de Lançamentos Pendentes - {qtd_registros} registros")
            
            self.atualizar_contador_selecionados()
            
        except Exception as e:
            print(f"Erro ao atualizar contador: {str(e)}")

    def limpar_visualizacao(self):
        """Limpa completamente a visualização"""
        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            self.dados_para_incluir.clear()
            
            if hasattr(self, 'dados_originais'):
                self.dados_originais.clear()
            
            self.atualizar_contador()
            
        except Exception as e:
            print(f"Erro ao limpar visualização: {str(e)}")

    def fechar_se_vazio(self):
        """Fecha o visualizador se não houver mais dados"""
        try:
            if len(self.dados_para_incluir) == 0:
                self.fechar_janela()
                return True
            return False
        except Exception as e:
            print(f"Erro ao verificar fechamento: {str(e)}")
            return False
        
class EditorLancamento:
    def __init__(self, parent, dados, indice, callback_atualizacao):
        self.janela = tk.Toplevel(parent)
        self.janela.title("Editar Lançamento")
        self.janela.geometry("600x550")  
        
        self._fechando = False
        self.janela.protocol("WM_DELETE_WINDOW", self.on_close)
        self.janela.transient(parent)
        self.janela.grab_set()
        
        self.dados = dados
        self.indice = indice
        self.callback_atualizacao = callback_atualizacao
        
        # Frame principal
        frame = ttk.Frame(self.janela, padding="10")
        frame.pack(fill='both', expand=True)
        
        # Frame para dados do fornecedor (não editáveis)
        frame_fornecedor = ttk.LabelFrame(frame, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', pady=5)
        
        # CNPJ/CPF
        ttk.Label(frame_fornecedor, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2)
        self.cnpj_cpf = ttk.Entry(frame_fornecedor, state='readonly')
        self.cnpj_cpf.grid(row=0, column=1, padx=5, pady=2)
        
        # Nome
        ttk.Label(frame_fornecedor, text="Nome:").grid(row=1, column=0, padx=5, pady=2)
        self.nome = ttk.Entry(frame_fornecedor, state='readonly')
        self.nome.grid(row=1, column=1, padx=5, pady=2)
        
        # Frame para dados da despesa
        frame_despesa = ttk.LabelFrame(frame, text="Dados da Despesa")
        frame_despesa.pack(fill='x', pady=5)
        
        # Data de Referência
        ttk.Label(frame_despesa, text="Data do Relatório:").grid(row=0, column=0, padx=5, pady=2)
        self.data_rel = DateEntry(frame_despesa, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_rel.grid(row=0, column=1, padx=5, pady=2)
        
        # Tipo de Despesa
        ttk.Label(frame_despesa, text="Tipo Despesa:").grid(row=1, column=0, padx=5, pady=2)
        self.tp_desp = ttk.Entry(frame_despesa)
        self.tp_desp.grid(row=1, column=1, padx=5, pady=2)
        
        # Referência
        ttk.Label(frame_despesa, text="Referência:").grid(row=2, column=0, padx=5, pady=2)
        self.referencia = ttk.Entry(frame_despesa)
        self.referencia.grid(row=2, column=1, padx=5, pady=2)

        # Etapa da Obra
        ttk.Label(frame_despesa, text="Etapa da Obra:").grid(row=3, column=0, padx=5, pady=2)
        self.etapa_obra = ttk.Entry(frame_despesa)
        self.etapa_obra.grid(row=3, column=1, padx=5, pady=2)
        
        # Insumo
        ttk.Label(frame_despesa, text="Insumo:").grid(row=4, column=0, padx=5, pady=2)
        self.insumo = ttk.Entry(frame_despesa)
        self.insumo.grid(row=4, column=1, padx=5, pady=2)

        # NF
        ttk.Label(frame_despesa, text="NF:").grid(row=5, column=0, padx=5, pady=2)
        self.nf = ttk.Entry(frame_despesa)
        self.nf.grid(row=5, column=1, padx=5, pady=2)
        
        # Valor Unitário
        ttk.Label(frame_despesa, text="Valor Unitário:").grid(row=6, column=0, padx=5, pady=2)
        self.vr_unit = ttk.Entry(frame_despesa)
        self.vr_unit.grid(row=6, column=1, padx=5, pady=2)
        
        # Dias
        ttk.Label(frame_despesa, text="Dias:").grid(row=7, column=0, padx=5, pady=2)
        self.dias = ttk.Entry(frame_despesa)
        self.dias.grid(row=7, column=1, padx=5, pady=2)
        
        # Valor Total
        ttk.Label(frame_despesa, text="Valor Total:").grid(row=8, column=0, padx=5, pady=2)
        self.valor = ttk.Entry(frame_despesa, state='readonly')
        self.valor.grid(row=8, column=1, padx=5, pady=2)
        
        # Data de Vencimento
        ttk.Label(frame_despesa, text="Data Vencimento:").grid(row=9, column=0, padx=5, pady=2)
        self.dt_vencto = DateEntry(frame_despesa, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.dt_vencto.grid(row=9, column=1, padx=5, pady=2)
        
        # Configurar o calendário para permitir navegação
        def configurar_calendario(event=None):
            if hasattr(self.dt_vencto, '_top_cal'):
                cal = self.dt_vencto._top_cal
                if cal:
                    def permitir_navegacao(event):
                        return "break"
                    
                    # Permitir cliques nas setas e mês/ano
                    for widget in cal.winfo_children():
                        if isinstance(widget, tk.Button):
                            widget.unbind('<Button-1>')
                            widget.bind('<Button-1>', permitir_navegacao)
                        
        self.dt_vencto.bind('<<DateEntryPopup>>', configurar_calendario)
        
        # Forma de Pagamento
        ttk.Label(frame_despesa, text="Forma de Pagamento:").grid(row=10, column=0, padx=5, pady=2)
        self.forma_pagamento = ttk.Combobox(frame_despesa, values=['PIX', 'TED', "DINHEIRO"], state='readonly')
        self.forma_pagamento.grid(row=10, column=1, padx=5, pady=2)
        
        # Observação
        ttk.Label(frame_despesa, text="Observação:").grid(row=11, column=0, padx=5, pady=2)
        self.observacao = ttk.Entry(frame_despesa)
        self.observacao.grid(row=11, column=1, padx=5, pady=2)
        
        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Salvar", command=self.salvar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=self.janela.self.on_close).pack(side='left', padx=5)
        
        # Preencher dados existentes
        self.preencher_dados()
        
        # Vincular eventos
        self.vr_unit.bind('<KeyRelease>', self.calcular_valor_total)
        self.dias.bind('<KeyRelease>', self.calcular_valor_total)

    def on_close(self):
        """Fecha a janela de forma segura"""
        self._fechando = True
        self.janela.grab_release()
        self.janela.destroy()
        
    def preencher_dados(self):
        """Preenche os campos com os dados atuais"""
        self.cnpj_cpf.config(state='normal')
        self.cnpj_cpf.insert(0, self.dados['cnpj_cpf'])
        self.cnpj_cpf.config(state='readonly')
        
        self.nome.config(state='normal')
        self.nome.insert(0, self.dados['nome'])
        self.nome.config(state='readonly')
        
        self.data_rel.set_date(datetime.strptime(self.dados['data'], '%d/%m/%Y'))
        self.tp_desp.insert(0, self.dados['tp_desp'])
        self.referencia.insert(0, self.dados['referencia'])
        
        self.etapa_obra.insert(0, self.dados.get('etapa_obra', ''))
        self.insumo.insert(0, self.dados.get('insumo', ''))

        self.nf.insert(0, self.dados.get('nf', ''))
        self.vr_unit.insert(0, self.dados['vr_unit'])
        self.dias.insert(0, str(self.dados['dias']))
        
        self.valor.config(state='normal')
        self.valor.insert(0, self.dados['valor'])
        self.valor.config(state='readonly')
        
        self.dt_vencto.set_date(datetime.strptime(self.dados['dt_vencto'], '%d/%m/%Y'))
        self.observacao.insert(0, self.dados.get('observacao', ''))
        self.forma_pagamento.set(self.dados.get('forma_pagamento', ''))

    def atualizar_dados_bancarios(self, event=None):
        """Atualiza os dados bancários baseado no tipo de despesa e forma de pagamento"""
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get().strip()

        if not cnpj_cpf:  # Se não houver fornecedor selecionado
            return
        
        forma_pagamento = self.forma_pagamento_var.get()
        
        try:
            # Usar a função centralizada em utils
            from src.config.utils import buscar_dados_bancarios_fornecedor
            dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj_cpf, forma_pagamento)
        except ImportError:
            # Implementação alternativa se a função não estiver disponível
            fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
            if not fornecedor_completo:
                return
            
            if forma_pagamento == "DINHEIRO":
                dados_bancarios = "PAGAMENTO EM DINHEIRO"
            elif forma_pagamento == "PIX" and fornecedor_completo['chave_pix']:
                dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
            else:
                # Estrutura para TED
                dados_ted = []
                if fornecedor_completo['banco']: dados_ted.append(str(fornecedor_completo['banco']))
                if fornecedor_completo['op']: dados_ted.append(str(fornecedor_completo['op']))
                if fornecedor_completo['agencia']: dados_ted.append(str(fornecedor_completo['agencia']))
                if fornecedor_completo['conta']: dados_ted.append(str(fornecedor_completo['conta']))
                # SEMPRE adicionar o CNPJ/CPF para TED
                dados_ted.append(str(fornecedor_completo['cnpj_cpf']))
                
                dados_bancarios = ' - '.join(filter(None, dados_ted))

            if dados_bancarios.strip() in ['', ' - ']:
                dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'

        # Atualizar o campo
        self.campos_fornecedor['dados_bancarios'].config(state='normal')
        self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
        self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
        self.campos_fornecedor['dados_bancarios'].config(state='readonly')
        
        
    def calcular_valor_total(self, event=None):
        """Calcula o valor total baseado no valor unitário e dias"""
        try:
            vr_unit = float(self.vr_unit.get().replace(',', '.'))
            dias = float(self.dias.get() or 1)
            valor_total = vr_unit * dias
            
            self.valor.config(state='normal')
            self.valor.delete(0, tk.END)
            self.valor.insert(0, f"{valor_total:.2f}")
            self.valor.config(state='readonly')
            
        except (ValueError, AttributeError):
            self.valor.config(state='normal')
            self.valor.delete(0, tk.END)
            self.valor.config(state='readonly')
            
    def salvar(self):
        """Salva as alterações e fecha a janela"""
        try:
            # Validar campos obrigatórios
            if not all([self.tp_desp.get(), self.referencia.get(), self.vr_unit.get()]):
                custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
                return
            
            # Validar datas
            for data_entry in [self.data_rel, self.dt_vencto]:
                data_str = data_entry.get()
                try:
                    datetime.strptime(data_str, '%d/%m/%Y')
                except ValueError:
                    custom_messagebox("error", "Erro", "Data inválida!")
                    return
            
            # Atualizar dados
            dados_atualizados = {
                'data': self.data_rel.get(),
                'tp_desp': self.tp_desp.get(),
                'cnpj_cpf': self.dados['cnpj_cpf'],
                'nome': self.dados['nome'],
                'forma_pagamento': self.forma_pagamento.get(),
                'referencia': self.referencia.get(),
                'etapa_obra': self.etapa_obra.get(),  # === NOVO CAMPO ===
                'insumo': self.insumo.get(),          # === NOVO CAMPO ===
                'nf': self.nf.get(),
                'vr_unit': self.vr_unit.get(),
                'dias': float(self.dias.get().replace(',', '.') if self.dias.get() else 1),
                'valor': self.valor.get(),
                'dt_vencto': self.dt_vencto.get(),
                'categoria': self.dados['categoria'],
                'dados_bancarios': self.dados['dados_bancarios'],
                'observacao': self.observacao.get()
            }
            
            # Chamar callback de atualização e verificar sucesso
            if self.callback_atualizacao(self.indice, dados_atualizados):
                custom_messagebox("info", "Sucesso", "Alterações salvas com sucesso!")
                self.on_close()
            else:
                custom_messagebox("error", "Erro", "Não foi possível salvar as alterações!")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar alterações: {str(e)}")


class EditorEmMassa:
    """Editor para alteração em massa de lançamentos marcados"""
    
    def __init__(self, parent, dados_selecionados, indices_selecionados, callback_atualizacao):
        self.janela = tk.Toplevel(parent)
        self.janela.title(f"Edição em Massa - {len(indices_selecionados)} lançamentos")
        self.janela.geometry("500x500")
        
        # Gerenciamento de foco
        self._fechando = False
        self.janela.protocol("WM_DELETE_WINDOW", self.on_close)
        self.janela.transient(parent)
        self.janela.grab_set()
        
        self.dados_selecionados = dados_selecionados
        self.indices_selecionados = indices_selecionados
        self.callback_atualizacao = callback_atualizacao
        
        self._criar_interface()
    
    def _criar_interface(self):
        frame = ttk.Frame(self.janela, padding="10")
        frame.pack(fill='both', expand=True)
        
        info_frame = ttk.LabelFrame(frame, text="Informações")
        info_frame.pack(fill='x', pady=5)
        
        ttk.Label(info_frame, 
                 text=f"Você está editando {len(self.indices_selecionados)} lançamentos simultaneamente.",
                 font=('TkDefaultFont', 9, 'bold')).pack(padx=10, pady=5)
        ttk.Label(info_frame, 
                 text="Marque apenas os campos que deseja alterar.",
                 foreground='blue').pack(padx=10, pady=2)
        
        campos_frame = ttk.LabelFrame(frame, text="Campos para Edição em Massa")
        campos_frame.pack(fill='both', expand=True, pady=10)
        
        self.campos_vars = {}
        self.campos_widgets = {}
        
        # TODOS OS 7 CAMPOS
        self._criar_campo_checkbox(campos_frame, 0, "data_rel", "Data do Relatório:", 
                                   DateEntry, date_pattern='dd/mm/yyyy', locale='pt_BR', width=20)
        
        self._criar_campo_checkbox(campos_frame, 1, "tp_desp", "Tipo de Despesa:", 
                                   ttk.Entry, width=30)
        
        self._criar_campo_checkbox(campos_frame, 2, "referencia", "Referência:", 
                                   ttk.Entry, width=30)
        
        self._criar_campo_checkbox(campos_frame, 3, "etapa_obra", "Etapa da Obra:", 
                                   ttk.Entry, width=30)
        
        self._criar_campo_checkbox(campos_frame, 4, "dt_vencto", "Data de Vencimento:", 
                                   DateEntry, date_pattern='dd/mm/yyyy', locale='pt_BR', width=20)
        
        # self._criar_campo_checkbox(campos_frame, 5, "forma_pagamento", "Forma de Pagamento:", 
        #                            ttk.Combobox, values=['PIX', 'TED', 'DINHEIRO'], 
        #                            state='readonly', width=20)
        
        # self._criar_campo_checkbox(campos_frame, 6, "observacao", "Observação:", 
        #                            ttk.Entry, width=30)
        
        # Frame de atalhos
        atalhos_frame = ttk.Frame(campos_frame)
        atalhos_frame.grid(row=7, column=0, columnspan=3, pady=10, sticky='ew')
        
        ttk.Label(atalhos_frame, text="Marcar:").pack(side='left', padx=5)
        ttk.Button(atalhos_frame, text="✓ Todos os Campos", 
                  command=self.marcar_todos).pack(side='left', padx=2)
        ttk.Button(atalhos_frame, text="✗ Nenhum Campo", 
                  command=self.desmarcar_todos).pack(side='left', padx=2)
        
        # Botões de ação
        botoes_frame = ttk.Frame(frame)
        botoes_frame.pack(fill='x', pady=10)
        
        ttk.Button(botoes_frame, text="💾 Aplicar Alterações", 
                  command=self.aplicar_alteracoes).pack(side='left', padx=5)
        ttk.Button(botoes_frame, text="❌ Cancelar", 
                  command=self.on_close).pack(side='left', padx=5)
        
        ttk.Label(botoes_frame, 
                 text="⚠️ As alterações serão aplicadas imediatamente",
                 foreground='red', font=('TkDefaultFont', 8)).pack(side='right', padx=10)
    
    def _criar_campo_checkbox(self, parent, row, nome_campo, label, widget_class, **widget_kwargs):
        """Cria um campo com checkbox de habilitação"""
        var = tk.BooleanVar(value=False)
        self.campos_vars[nome_campo] = var
        
        chk = ttk.Checkbutton(parent, variable=var, 
                             command=lambda: self._toggle_campo(nome_campo))
        chk.grid(row=row, column=0, padx=5, pady=5)
        
        lbl = ttk.Label(parent, text=label, state='disabled')
        lbl.grid(row=row, column=1, sticky='w', padx=5, pady=5)
        
        # Criar o widget
        if widget_class == DateEntry:
            widget = widget_class(parent, **widget_kwargs)
            # ✅ CORREÇÃO: Não usar widget._entry diretamente
        elif widget_class == ttk.Combobox:
            widget = widget_class(parent, **widget_kwargs)
        else:
            widget = widget_class(parent, **widget_kwargs)
        
        widget.grid(row=row, column=2, padx=5, pady=5, sticky='ew')
        widget.config(state='disabled')
        
        self.campos_widgets[nome_campo] = {'label': lbl, 'widget': widget}
        parent.columnconfigure(2, weight=1)
    
    def _toggle_campo(self, nome_campo):
        """Habilita/desabilita um campo baseado no checkbox"""
        habilitado = self.campos_vars[nome_campo].get()
        estado = 'normal' if habilitado else 'disabled'
        
        self.campos_widgets[nome_campo]['label'].config(state=estado)
        
        widget = self.campos_widgets[nome_campo]['widget']
        
        # Tratamento especial para cada tipo
        if isinstance(widget, DateEntry):
            if habilitado:
                widget.config(state='normal')
                # ✅ CORREÇÃO: Acessar entry interno de forma segura
                try:
                    for child in widget.winfo_children():
                        if isinstance(child, tk.Entry):
                            child.config(state='normal')
                            break
                except:
                    pass
            else:
                widget.config(state='disabled')
        elif isinstance(widget, ttk.Combobox):
            if habilitado:
                widget.config(state='readonly')
            else:
                widget.config(state='disabled')
        else:
            widget.config(state=estado)
    
    def marcar_todos(self):
        """Marca todos os checkboxes e habilita todos os campos"""
        for nome_campo in self.campos_vars:
            self.campos_vars[nome_campo].set(True)
            self._toggle_campo(nome_campo)
    
    def desmarcar_todos(self):
        """Desmarca todos os checkboxes e desabilita todos os campos"""
        for nome_campo in self.campos_vars:
            self.campos_vars[nome_campo].set(False)
            self._toggle_campo(nome_campo)
    
    def aplicar_alteracoes(self):
        """Aplica as alterações em todos os lançamentos marcados"""
        try:
            try:
                from src.config.dialogs import custom_messagebox
            except ImportError:
                from config.dialogs import custom_messagebox
            
            campos_marcados = [campo for campo, var in self.campos_vars.items() if var.get()]
            
            if not campos_marcados:
                custom_messagebox("warning", "⚠️ Atenção", 
                                "Você precisa marcar pelo menos um campo para editar!")
                return
            
            qtd = len(self.indices_selecionados)
            
            traducao_campos = {
                'data_rel': 'Data do Relatório',
                'tp_desp': 'Tipo de Despesa',
                'referencia': 'Referência',
                'etapa_obra': 'Etapa da Obra',
                'dt_vencto': 'Data de Vencimento',
                'forma_pagamento': 'Forma de Pagamento',
                'observacao': 'Observação'
            }
            
            campos_texto = "\n• ".join([traducao_campos.get(c, c) for c in campos_marcados])
            
            resposta = custom_messagebox("yesno", 
                "⚠️ Confirmar Edição em Massa",
                f"Você está prestes a alterar {qtd} lançamentos!\n\n"
                f"Campos que serão alterados:\n• {campos_texto}\n\n"
                f"Esta operação não pode ser desfeita.\n\n"
                f"Confirma a alteração?")
            
            if not resposta:
                return
            
            valores_alteracao = {}
            for campo in campos_marcados:
                widget = self.campos_widgets[campo]['widget']
                
                if isinstance(widget, DateEntry):
                    valores_alteracao[campo] = widget.get()
                elif isinstance(widget, (ttk.Entry, ttk.Combobox)):
                    valores_alteracao[campo] = widget.get()
            
            alteracoes_realizadas = 0
            erros = []
            
            for idx, dados_originais in zip(self.indices_selecionados, self.dados_selecionados):
                try:
                    dados_atualizados = dados_originais.copy()
                    
                    for campo, valor in valores_alteracao.items():
                        if campo == 'data_rel':
                            dados_atualizados['data'] = valor
                        else:
                            dados_atualizados[campo] = valor
                    
                    if self.callback_atualizacao(idx, dados_atualizados):
                        alteracoes_realizadas += 1
                    else:
                        erros.append(f"Lançamento {idx+1}")
                        
                except Exception as e:
                    erros.append(f"Lançamento {idx+1}: {str(e)}")
            
            if alteracoes_realizadas > 0:
                mensagem = f"✅ {alteracoes_realizadas} lançamentos alterados com sucesso!"
                
                if erros:
                    mensagem += f"\n\n⚠️ {len(erros)} lançamentos com erro:\n" + "\n".join(erros[:5])
                    if len(erros) > 5:
                        mensagem += f"\n... e mais {len(erros)-5} erros"
                
                custom_messagebox("info", "Resultado da Edição em Massa", mensagem)
                self.on_close()
            else:
                custom_messagebox("error", "Erro", 
                                "Nenhuma alteração foi realizada.\n\n" + 
                                "Erros:\n" + "\n".join(erros[:10]))
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao aplicar alterações: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def on_close(self):
        """Fecha a janela de forma segura"""
        self._fechando = True
        self.janela.grab_release()
        self.janela.destroy()

class GerenciadorCPFsCriados:
    def __init__(self):
        self.arquivo_fornecedores = ARQUIVO_FORNECEDORES
        
    def gerar_cpf_valido(self):
        """Gera um CPF válido seguindo EXATAMENTE o algoritmo oficial"""
        import random
        
        # Gerar os 9 primeiros dígitos (evitar sequências óbvias)
        while True:
            cpf = [random.randint(0, 9) for _ in range(9)]
            
            # Evitar CPFs com todos os dígitos iguais (000.000.000, 111.111.111, etc.)
            if len(set(cpf)) > 1:
                break
        
        # Calcular PRIMEIRO dígito verificador
        soma = 0
        for i in range(9):
            soma += cpf[i] * (10 - i)
        
        resto = soma % 11
        if resto < 2:
            primeiro_digito = 0
        else:
            primeiro_digito = 11 - resto
        
        cpf.append(primeiro_digito)
        
        # Calcular SEGUNDO dígito verificador
        soma = 0
        for i in range(10):
            soma += cpf[i] * (11 - i)
        
        resto = soma % 11
        if resto < 2:
            segundo_digito = 0
        else:
            segundo_digito = 11 - resto
        
        cpf.append(segundo_digito)
        
        return ''.join(map(str, cpf))
    
    def validar_cpf_gerado(self, cpf):
        """Valida se o CPF gerado está correto"""
        if len(cpf) != 11:
            return False
        
        # Verificar se não são todos iguais
        if cpf == cpf[0] * 11:
            return False
        
        # Calcular primeiro dígito
        soma = 0
        for i in range(9):
            soma += int(cpf[i]) * (10 - i)
        resto = soma % 11
        digito1 = 0 if resto < 2 else 11 - resto
        
        if int(cpf[9]) != digito1:
            return False
        
        # Calcular segundo dígito
        soma = 0
        for i in range(10):
            soma += int(cpf[i]) * (11 - i)
        resto = soma % 11
        digito2 = 0 if resto < 2 else 11 - resto
        
        return int(cpf[10]) == digito2
    
    def obter_proximo_cpf_disponivel(self):
        """Busca o próximo CPF disponível na aba CPF"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)
            
            # Verificar se a aba CPF existe
            if 'CPF' not in wb.sheetnames:
                print("Criando aba CPF...")
                # Criar a aba CPF se não existir
                ws_cpf = wb.create_sheet('CPF')
                ws_cpf.cell(row=1, column=1, value='CPF_CRIADO')
                ws_cpf.cell(row=1, column=2, value='STATUS')
                ws_cpf.cell(row=1, column=3, value='USADO_POR')
                ws_cpf.cell(row=1, column=4, value='DATA_USO')
                wb.save(self.arquivo_fornecedores)
            else:
                ws_cpf = wb['CPF']
            
            # Buscar primeiro CPF disponível
            cpf_disponivel = None
            linha_disponivel = None
            
            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                status = ws_cpf.cell(row=row, column=2).value
                
                if cpf_valor and (not status or status == 'DISPONIVEL'):
                    # Validar se o CPF é realmente válido
                    if self.validar_cpf_gerado(str(cpf_valor)):
                        cpf_disponivel = str(cpf_valor)
                        linha_disponivel = row
                        print(f"CPF disponível encontrado: {cpf_disponivel}")
                        break
                    else:
                        print(f"CPF inválido encontrado na planilha: {cpf_valor}, removendo...")
                        # Marcar como inválido
                        ws_cpf.cell(row=row, column=2, value='INVALIDO')
            
            # Se não encontrou nenhum disponível, gerar novos
            if not cpf_disponivel:
                print("Gerando novos CPFs...")
                # Gerar 20 novos CPFs válidos
                cpfs_gerados = 0
                tentativas = 0
                max_tentativas = 100
                
                while cpfs_gerados < 20 and tentativas < max_tentativas:
                    tentativas += 1
                    novo_cpf = self.gerar_cpf_valido()
                    
                    # Validar o CPF gerado
                    if self.validar_cpf_gerado(novo_cpf):
                        # Verificar se já existe
                        if not self.cpf_ja_existe(ws_cpf, novo_cpf):
                            proxima_linha = ws_cpf.max_row + 1
                            ws_cpf.cell(row=proxima_linha, column=1, value=novo_cpf)
                            ws_cpf.cell(row=proxima_linha, column=2, value='DISPONIVEL')
                            cpfs_gerados += 1
                            
                            print(f"CPF válido gerado: {novo_cpf}")
                            
                            if not cpf_disponivel:  # Pegar o primeiro gerado
                                cpf_disponivel = novo_cpf
                                linha_disponivel = proxima_linha
                    else:
                        print(f"CPF inválido gerado (descartado): {novo_cpf}")
                
                if cpfs_gerados > 0:
                    wb.save(self.arquivo_fornecedores)
                    print(f"Total de CPFs válidos gerados: {cpfs_gerados}")
                else:
                    print("ERRO: Não foi possível gerar CPFs válidos")
            
            wb.close()
            
            if cpf_disponivel:
                print(f"Retornando CPF: {cpf_disponivel}")
                # Validar uma última vez antes de retornar
                if self.validar_cpf_gerado(cpf_disponivel):
                    return cpf_disponivel, linha_disponivel
                else:
                    print(f"ERRO: CPF retornado é inválido: {cpf_disponivel}")
                    return None, None
            else:
                return None, None
            
        except Exception as e:
            print(f"Erro ao obter CPF disponível: {str(e)}")
            import traceback
            traceback.print_exc()
            return None, None
    
    def cpf_ja_existe(self, worksheet, cpf):
        """Verifica se o CPF já existe na planilha"""
        for row in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row, column=1).value) == str(cpf):
                return True
        return False
    
    def marcar_cpf_como_usado(self, cpf, nome_fornecedor):
        """Marca um CPF como usado"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)
            ws_cpf = wb['CPF']
            
            for row in range(2, ws_cpf.max_row + 1):
                if str(ws_cpf.cell(row=row, column=1).value) == str(cpf):
                    ws_cpf.cell(row=row, column=2, value='USADO')
                    ws_cpf.cell(row=row, column=3, value=nome_fornecedor)
                    ws_cpf.cell(row=row, column=4, value=datetime.now().strftime('%d/%m/%Y %H:%M'))
                    break
            
            wb.save(self.arquivo_fornecedores)
            wb.close()
            return True
            
        except Exception as e:
            print(f"Erro ao marcar CPF como usado: {str(e)}")
            return False
    
    def listar_cpfs_disponiveis(self):
        """Lista todos os CPFs disponíveis"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)
            
            if 'CPF' not in wb.sheetnames:
                wb.close()
                return []
            
            ws_cpf = wb['CPF']
            cpfs_disponiveis = []
            
            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                status = ws_cpf.cell(row=row, column=2).value
                
                if cpf_valor and (not status or status == 'DISPONIVEL'):
                    # Validar antes de adicionar à lista
                    if self.validar_cpf_gerado(str(cpf_valor)):
                        cpfs_disponiveis.append(str(cpf_valor))
            
            wb.close()
            return cpfs_disponiveis
            
        except Exception as e:
            print(f"Erro ao listar CPFs disponíveis: {str(e)}")
            return []     
        
    def marcar_cpf_como_disponivel(self, cpf):
        """Marca um CPF como disponível novamente na aba CPF"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)
            
            if 'CPF' not in wb.sheetnames:
                wb.close()
                return False
            
            ws_cpf = wb['CPF']
            
            # Procurar o CPF e marcar como disponível
            for row in range(2, ws_cpf.max_row + 1):
                if str(ws_cpf.cell(row=row, column=1).value).strip() == str(cpf):
                    ws_cpf.cell(row=row, column=2, value='DISPONIVEL')  # Status
                    ws_cpf.cell(row=row, column=3, value='')  # Limpar nome do fornecedor
                    ws_cpf.cell(row=row, column=4, value='')  # Limpar data de uso
                    break
            
            wb.save(self.arquivo_fornecedores)
            wb.close()
            return True
            
        except Exception as e:
            print(f"Erro ao marcar CPF como disponível: {str(e)}")
            return False

    def listar_todos_cpfs_criados(self):
        """Lista todos os CPFs criados da aba CPF (disponíveis e usados)"""
        try:
            wb = load_workbook(self.arquivo_fornecedores, data_only=True)
            
            if 'CPF' not in wb.sheetnames:
                wb.close()
                return []
            
            ws_cpf = wb['CPF']
            cpfs = []
            
            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                if cpf_valor:  # Se tem CPF
                    cpfs.append(str(cpf_valor).strip())
            
            wb.close()
            return cpfs
            
        except Exception as e:
            print(f"Erro ao listar todos os CPFs criados: {str(e)}")
            return []

    def listar_cpfs_usados(self):
        """Lista apenas os CPFs que estão marcados como USADO na aba CPF"""
        try:
            wb = load_workbook(self.arquivo_fornecedores, data_only=True)
            
            if 'CPF' not in wb.sheetnames:
                wb.close()
                return []
            
            ws_cpf = wb['CPF']
            cpfs_usados = []
            
            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                status = ws_cpf.cell(row=row, column=2).value
                
                if cpf_valor and status and str(status).strip().upper() == 'USADO':
                    cpfs_usados.append(str(cpf_valor).strip())
            
            wb.close()
            return cpfs_usados
            
        except Exception as e:
            print(f"Erro ao listar CPFs usados: {str(e)}")
            return []

    def obter_detalhes_cpf_usado(self, cpf):
        """Obtém detalhes de um CPF usado (nome do fornecedor e data de uso)"""
        try:
            wb = load_workbook(self.arquivo_fornecedores, data_only=True)
            
            if 'CPF' not in wb.sheetnames:
                wb.close()
                return None
            
            ws_cpf = wb['CPF']
            
            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                
                if str(cpf_valor).strip() == str(cpf):
                    status = ws_cpf.cell(row=row, column=2).value
                    usado_por = ws_cpf.cell(row=row, column=3).value
                    data_uso = ws_cpf.cell(row=row, column=4).value
                    
                    wb.close()
                    return {
                        'status': str(status) if status else '',
                        'usado_por': str(usado_por) if usado_por else '',
                        'data_uso': str(data_uso) if data_uso else ''
                    }
            
            wb.close()
            return None
            
        except Exception as e:
            print(f"Erro ao obter detalhes do CPF: {str(e)}")
            return None

class SistemaEntradaDados:

    def atualizar_combos_configuracoes(self):
        """Atualiza os valores das Comboboxes baseados nas configurações"""
        if 'categoria' in self.campos_fornecedor and isinstance(self.campos_fornecedor['categoria'], ttk.Combobox):
            categorias = get_categorias_fornecedor()
            self.campos_fornecedor['categoria']['values'] = categorias
            # Define o primeiro valor como padrão se houver categorias
            if categorias:
                self.campos_fornecedor['categoria'].set(categorias[0])
            
    def __init__(self, parent=None):
        print("Inicializando SistemaEntradaDados...")
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Sistema de Entrada de Dados")
        self.dados_para_incluir = []
        self.data_rel = None
        self.cliente_atual = None
        self.visualizador = None
        self._gestor_parcelas = None  # Inicializa como None
        self.gerenciador_lancamentos = None
        self.gerenciador_agenda = None 
        self.cache_fornecedores = CacheFornecedores()

        # Inicializar a variável de forma de pagamento
        self.forma_pagamento_var = tk.StringVar(value="")
            
        # Frame temporário para criar os entries
        temp_frame = ttk.Frame(self.root)

        # Criação dos campos_fornecedor e campos_despesa
        self.campos_fornecedor = {
            'cnpj_cpf': tk.Entry(temp_frame),
            'nome': tk.Entry(temp_frame),
            'categoria': tk.Entry(temp_frame),
            'dados_bancarios': tk.Entry(temp_frame)
        }

        self.campos_despesa = {
            'tp_desp': tk.Entry(temp_frame),
            'referencia': tk.Entry(temp_frame),
            'nf': tk.Entry(temp_frame),
            'vr_unit': tk.Entry(temp_frame),
            'dias': tk.Entry(temp_frame),
            'valor': tk.Entry(temp_frame),
            'dt_vencto': tk.Entry(temp_frame),
            'observacao': tk.Entry(temp_frame)
        }

        self.gestao_taxas = GestaoTaxasFixas(self)

        self.atualizar_combos_configuracoes()
            
        # Configurar interface
        self.setup_gui()
        self.configurar_todos_calendarios()

        # Adicionar estas linhas para configurar cada aba explicitamente
        print("Configurando aba de seleção...")
        self.setup_aba_selecao()
        print("Configurando aba de fornecedor...")
        self.setup_aba_fornecedor()
        print("Configurando aba de dados...")
        self.setup_aba_dados()

       # Configurar sistema de backup automático
        print("Configurando sistema de backup...")
        
        # Verificar dados não salvos após 1 segundo (para interface carregar)
        self.root.after(1000, self.verificar_dados_nao_salvos)
        
        # Configurar proteções
        self.configurar_protecoes()

        # Configurar auto-salvamento
        self.configurar_auto_salvamento()
        
        print("Sistema de backup configurado ✅")

        # Inicializar sistema de materiais
        self.integrador_materiais = inicializar_sistema_materiais_completo(self)
        
        # Inicializar sistema NFe:
        try:
            from src.nfe.extensao_sistema_hibrido import inicializar_sistema_nfe_estendido
            inicializar_sistema_nfe_estendido(self)
        except Exception as e:
            print(f"⚠️ Sistema NFe estendido não carregado: {e}")
            # Fallback para sistema original
            try:
                from src.nfe.sistema_hibrido_nfe import inicializar_sistema_nfe_hibrido
                inicializar_sistema_nfe_hibrido(self)
                print("⚠️ Usando sistema NFe original (sem extensão)")
            except Exception as e2:
                print(f"⚠️ Sistema NFe não carregado: {e2}")

        if hasattr(self, 'configurar_certificado_rapido'):
            print("✅ Certificado A1 disponível!")
        else:
            print("❌ Certificado A1 NÃO disponível")

        # Verificação automática do certificado A1
        try:
            from teste_certificado_automatico import verificar_certificado_a1_automatico
            verificar_certificado_a1_automatico(self)
        except Exception as e:
            print(f"Erro na verificação do certificado: {e}")

        # Adicionar métodos de interface
        try:
            from teste_certificado_automatico import (
                criar_botao_teste_certificado,
                configurar_certificado_interface,
                testar_certificado_interface,
                consultar_nfe_interface,
                processar_nfe_consultada
            )
            
            self.criar_botao_teste_certificado = criar_botao_teste_certificado.__get__(self)
            self.configurar_certificado_interface = configurar_certificado_interface.__get__(self)
            self.testar_certificado_interface = testar_certificado_interface.__get__(self)
            self.consultar_nfe_interface = consultar_nfe_interface.__get__(self)
            self.processar_nfe_consultada = processar_nfe_consultada.__get__(self)
            
        except Exception as e:
            print(f"Erro ao adicionar métodos de interface: {e}")
        
        print("Finalizada inicialização do sistema")

    def setup_gui(self):
        print("Iniciando setup_gui...")
        
        # Remover notebook existente se houver
        for widget in self.root.winfo_children():
            if isinstance(widget, ttk.Notebook):
                print("Notebook existente encontrado e será removido")
                widget.destroy()
        
        # Frame principal com abas
        self.notebook = ttk.Notebook(self.root)
        print("Novo Notebook criado")
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar abas
        self.aba_selecao = ttk.Frame(self.notebook)
        self.aba_fornecedor = ttk.Frame(self.notebook)
        self.aba_dados = ttk.Frame(self.notebook)
        
        print("Adicionando abas ao Notebook")
        self.notebook.add(self.aba_selecao, text='Seleção de Cliente')
        self.notebook.add(self.aba_fornecedor, text='Fornecedor')
        self.notebook.add(self.aba_dados, text='Entrada de Dados')

        print("Setup_gui concluído")

    @property
    def gestor_parcelas(self):
        """Getter para gestor_parcelas - cria apenas quando necessário"""
        if self._gestor_parcelas is None:
            print("Criando nova instância do GestorParcelas")  # Debug
            self._gestor_parcelas = GestorParcelas(self)
        return self._gestor_parcelas

    @gestor_parcelas.setter
    def gestor_parcelas(self, valor):
        """Setter para gestor_parcelas"""
        self._gestor_parcelas = valor        

    def voltar_menu(self):
        """
        Versão alternativa que força redesenho completo
        """
        if self.dados_para_incluir and custom_messagebox("yesno", 
            "Confirmação", 
            "Existem dados não salvos. Deseja salvá-los antes de sair?"):
            self.enviar_dados()
        
        menu_ref = self.menu_principal
        
        # Fechar janela atual
        self.root.destroy()
        
        if menu_ref:
            try:
                geometria = getattr(menu_ref, '_geometria_original', "900x700+100+50")
                
                # ESTRATÉGIA DIFERENTE: Ocultar, configurar, mostrar
                menu_ref.withdraw()  # Garantir que está oculto
                
                # Configurar geometria enquanto oculto
                menu_ref.geometry(geometria)
                menu_ref.update_idletasks()
                menu_ref.update()
                
                # Aguardar um ciclo
                menu_ref.after(10, lambda: None)
                menu_ref.update()
                
                # Mostrar janela configurada
                menu_ref.deiconify()
                
                # Forçar redesenho
                menu_ref.update_idletasks()
                menu_ref.update()
                
                # Foco
                menu_ref.lift()
                menu_ref.focus_force()
                
                print(f"✅ Menu restaurado com redesenho completo")
                
            except Exception as e:
                print(f"⚠️ Erro: {e}")


    def sair_sistema(self):
        """Fecha o sistema verificando dados não salvos"""
        try:
            self.finalizar_sistema()
        except Exception as e:
            print(f"Erro ao finalizar sistema: {str(e)}")
        finally:
            # Forçar saída se necessário
            import sys
            sys.exit()  
    
    def configurar_todos_calendarios(self):
        """Configura a navegação para todos os calendários do sistema"""
        # Lista de todos os campos de data que usam DateEntry
        date_entries = []
        
        # Adicionar campos da interface principal
        if hasattr(self, 'data_rel_entry'):
            date_entries.append(self.data_rel_entry)
        
        # Adicionar campos da aba de dados
        if hasattr(self, 'campos_despesa') and 'dt_vencto' in self.campos_despesa:
            date_entries.append(self.campos_despesa['dt_vencto'])
        
        # Configurar cada DateEntry encontrado
        for date_entry in date_entries:
            if isinstance(date_entry, DateEntry):
                configurar_navegacao_calendario(date_entry)
        
        print("Calendários configurados para permitir navegação livre.")

    def setup_aba_selecao(self):
        """Configura a aba de seleção de cliente - VERSÃO FINAL"""
        # Frame principal para organização
        frame_principal = ttk.Frame(self.aba_selecao)
        frame_principal.pack(expand=True, fill='both', padx=20, pady=10)

        # ===== SEÇÃO DE SELEÇÃO DE CLIENTE - MODERNIZADA =====
        frame_selecao = ttk.LabelFrame(
            frame_principal, 
            text="Seleção do Cliente", 
            padding=(15, 10)
        )
        frame_selecao.pack(fill='x', pady=(0, 15))

        # Container para label e área de busca
        frame_busca = ttk.Frame(frame_selecao)
        frame_busca.pack(fill='x', pady=5)

        # Label com instrução mais clara
        ttk.Label(
            frame_busca, 
            text="Digite ou selecione o cliente:", 
            font=('Arial', 11, 'bold')
        ).pack(anchor='w', pady=(0, 5))

        # Frame para combobox e contador
        frame_combo = ttk.Frame(frame_busca)
        frame_combo.pack(fill='x')

        # Combobox com busca melhorada
        self.cliente_combobox = ttk.Combobox(
            frame_combo, 
            font=('Arial', 11),
            state='normal'
        )
        self.cliente_combobox.pack(side='left', fill='x', expand=True, padx=(0, 5))

        # Label contador de resultados
        self.label_contador = ttk.Label(
            frame_combo, 
            text="", 
            font=('Arial', 9),
            foreground='gray'
        )
        self.label_contador.pack(side='left', padx=5)

        # Botão Ver Todos
        self.btn_ver_todos = ttk.Button(
            frame_combo, 
            text="📋 Ver Todos", 
            command=self.mostrar_todos_clientes,
            width=12
        )
        self.btn_ver_todos.pack(side='right', padx=(5, 0))

        # ===== BINDINGS =====
        self.cliente_combobox.bind('<KeyRelease>', self.filtrar_clientes_digitacao)
        self.cliente_combobox.bind('<Button-1>', self.abrir_dropdown_clientes)
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)
        self.cliente_combobox.bind('<Return>', self.selecionar_primeiro_resultado)
        self.cliente_combobox.bind('<Escape>', self.limpar_busca)

        # Dica visual
        frame_dica = ttk.Frame(frame_selecao)
        frame_dica.pack(fill='x', pady=(5, 0))
        
        ttk.Label(
            frame_dica,
            text="💡 Dica: Comece a digitar para buscar. Use ↑↓ para navegar, Enter para selecionar, ESC para limpar.",
            font=('Arial', 9, 'italic'),
            foreground='#666666'
        ).pack(anchor='w')

        # ===== BOTÕES =====
        frame_gerenciar = ttk.Frame(frame_principal)
        frame_gerenciar.pack(pady=10)

        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 11, 'bold'), padding=(20, 12))
        style.configure('Action.TButton', font=('Arial', 11, 'bold'), padding=(25, 15))

        frame_botoes = ttk.Frame(frame_gerenciar)
        frame_botoes.pack(pady=5)

        ttk.Button(
            frame_botoes, 
            text="➕ Novo Cliente", 
            command=self.criar_novo_cliente,
            style='Big.TButton'
        ).pack(side='left', padx=8)
                
        ttk.Button(
            frame_botoes,
            text="✏️ Editar Cliente",
            command=self.editar_cliente,
            style='Big.TButton'
        ).pack(side='left', padx=8)

        ttk.Button(
            frame_botoes, 
            text="📄 Gerir Contratos",
            command=self.abrir_gestao_contratos,
            style='Big.TButton'
        ).pack(side='left', padx=8)

        # Botão Continuar
        frame_continuar = ttk.Frame(frame_principal)
        frame_continuar.pack(fill='x', pady=15)

        ttk.Separator(frame_continuar, orient='horizontal').pack(fill='x', pady=(0, 15))

        self.btn_continuar = ttk.Button(
            frame_continuar,
            text="Continuar →",
            command=self.continuar_para_fornecedor,
            state='disabled',
            style='Action.TButton'
        )
        self.btn_continuar.pack(pady=5)

        # Botões de navegação
        frame_navegacao = ttk.Frame(frame_principal)
        frame_navegacao.pack(fill='x', side='bottom', pady=(20, 0))

        ttk.Separator(frame_navegacao, orient='horizontal').pack(fill='x', pady=(0, 10))

        frame_botoes_nav = ttk.Frame(frame_navegacao)
        frame_botoes_nav.pack()

        ttk.Button(
            frame_botoes_nav, 
            text="⬅️ Voltar ao Menu", 
            command=self.voltar_menu,
            style='Big.TButton'
        ).pack(side='left', padx=10)
        
        ttk.Button(
            frame_botoes_nav, 
            text="❌ Sair", 
            command=self.sair_sistema,
            style='Big.TButton'
        ).pack(side='left', padx=10)

        # Inicialização
        self.clientes_completos = []
        self.atualizar_lista_clientes()
        self.cliente_combobox.focus_set()

    def filtrar_clientes_digitacao(self, event=None):
        """Filtra clientes em tempo real - VERSÃO FINAL"""
        try:
            texto_busca = self.cliente_combobox.get().upper().strip()
            
            if not texto_busca:
                self.cliente_combobox['values'] = self.clientes_completos
                if hasattr(self, 'atualizar_contador'):
                    self.atualizar_contador(len(self.clientes_completos))
                return
            
            clientes_filtrados = []
            for cliente in self.clientes_completos:
                try:
                    cliente_str = str(cliente) if cliente is not None else ""
                    if texto_busca in cliente_str.upper():
                        clientes_filtrados.append(cliente)
                except Exception:
                    continue
            
            self.cliente_combobox['values'] = clientes_filtrados
            
            if hasattr(self, 'atualizar_contador'):
                self.atualizar_contador(len(clientes_filtrados))
            
            if clientes_filtrados and len(texto_busca) >= 2:
                try:
                    self.cliente_combobox.event_generate('<Down>')
                except Exception:
                    pass
                    
        except Exception as e:
            logger = system_logger.get_logger()
            logger.error(f"Erro ao filtrar clientes: {str(e)}")
            try:
                self.cliente_combobox['values'] = self.clientes_completos
                if hasattr(self, 'atualizar_contador'):
                    self.atualizar_contador(len(self.clientes_completos))
            except Exception:
                pass

    def abrir_dropdown_clientes(self, event=None):
        """Abre dropdown automaticamente ao clicar no campo"""
        # Garantir que mostra todos os clientes se não há filtro
        if not self.cliente_combobox.get():
            self.cliente_combobox['values'] = self.clientes_completos
            self.atualizar_contador(len(self.clientes_completos))
        
        # Abrir dropdown
        self.cliente_combobox.event_generate('<Down>')

    def selecionar_primeiro_resultado(self, event=None):
        """Seleciona o primeiro resultado quando Enter é pressionado"""
        valores = self.cliente_combobox['values']
        
        if valores:
            # Definir primeiro valor
            self.cliente_combobox.set(valores[0])
            # Disparar evento de seleção
            self.cliente_combobox.event_generate('<<ComboboxSelected>>')

    def limpar_busca(self, event=None):
        """Limpa o campo de busca e restaura lista completa"""
        self.cliente_combobox.set('')
        self.cliente_combobox['values'] = self.clientes_completos
        self.atualizar_contador(len(self.clientes_completos))
        self.btn_continuar.config(state='disabled')

    def atualizar_contador(self, quantidade):
        """Atualiza contador de resultados"""
        try:
            if not hasattr(self, 'label_contador'):
                return
                
            if not hasattr(self, 'clientes_completos'):
                self.label_contador.config(text=f"({quantidade} clientes)")
                return
                
            if quantidade == len(self.clientes_completos):
                self.label_contador.config(
                    text=f"({quantidade} clientes)",
                    foreground='gray'
                )
            else:
                self.label_contador.config(
                    text=f"({quantidade} de {len(self.clientes_completos)})",
                    foreground='#0066cc'
                )
        except Exception as e:
            logger = system_logger.get_logger()
            logger.debug(f"Erro ao atualizar contador: {str(e)}")

    def criar_arquivo_clientes(self):
        """Cria arquivo base de clientes se não existir"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Clientes'
            
            # Adicionar cabeçalhos - somente campos básicos agora
            headers = ['Nome', 'Endereco', 'Data_Inicial', 'Observacoes']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            caminho_base = ARQUIVO_CLIENTES
            workbook.save(caminho_base)
            custom_messagebox("info", "Informação", "Arquivo de clientes criado com sucesso!")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao criar arquivo de clientes: {str(e)}")
 
    def criar_novo_cliente(self):
        """Abre janela para cadastro de novo cliente"""
        janela_cliente = tk.Toplevel(self.root)
        janela_cliente.title("Novo Cliente")
        janela_cliente.geometry("700x700")  # Aumentado para incluir novo campo

        # Frame principal
        frame = ttk.Frame(janela_cliente, padding="10")
        frame.pack(fill='both', expand=True)

        # Nome
        ttk.Label(frame, text="Nome do Cliente:*").pack(pady=5)
        nome_entry = ttk.Entry(frame, width=80)
        nome_entry.pack(pady=5)

        # Endereço
        ttk.Label(frame, text="Endereço:*").pack(pady=5)
        endereco_entry = ttk.Entry(frame, width=80)
        endereco_entry.pack(pady=5)

        # Data Inicial
        ttk.Label(frame, text="Data Inicial:* (Dia 5 ou 20)").pack(pady=5)
        data_entry = DateEntry(
            frame,
            width=20,
            date_pattern='yyyy-mm-dd',
            locale='pt_BR'
        )
        data_entry.pack(pady=5)

        # NOVO: Tipo de Taxa
        ttk.Label(frame, text="Tipo de Taxa de Administração:*").pack(pady=5)
        tipo_taxa_var = tk.StringVar(value="Percentual")
        frame_tipo = ttk.Frame(frame)
        frame_tipo.pack(pady=5)
        
        ttk.Radiobutton(
            frame_tipo, 
            text="Percentual", 
            variable=tipo_taxa_var, 
            value="Percentual"
        ).pack(side='left', padx=10)
        
        ttk.Radiobutton(
            frame_tipo, 
            text="Fixo", 
            variable=tipo_taxa_var, 
            value="Fixo"
        ).pack(side='left', padx=10)
        
        ttk.Radiobutton(
            frame_tipo, 
            text="Sem Taxa", 
            variable=tipo_taxa_var, 
            value="Sem Taxa"
        ).pack(side='left', padx=10)

        # Observações
        ttk.Label(frame, text="Observações:").pack(pady=5)
        obs_entry = ttk.Entry(frame, width=80)
        obs_entry.pack(pady=5)

        # CPF
        ttk.Label(frame, text="CPF:").pack(pady=5)
        cpf_entry = ttk.Entry(frame, width=80)
        cpf_entry.pack(pady=5)

        # CNO
        ttk.Label(frame, text="CNO:").pack(pady=5)
        cno_entry = ttk.Entry(frame, width=80)
        cno_entry.pack(pady=5)

        # Estado civil
        ttk.Label(frame, text="Estado Civil:").pack(pady=5)
        estado_civil_var = tk.StringVar(value="Casado(a)")
        frame_estado_civil = ttk.Frame(frame)
        frame_estado_civil.pack(pady=5)
        
        ttk.Radiobutton(
            frame_estado_civil, 
            text="Casado(a)", 
            variable=estado_civil_var, 
            value="Casado(a)"
        ).pack(side='left', padx=10)
        
        ttk.Radiobutton(
            frame_estado_civil, 
            text="Solteiro(a)", 
            variable=estado_civil_var, 
            value="Solteiro(a)"
        ).pack(side='left', padx=10)
        
        ttk.Radiobutton(
            frame_estado_civil, 
            text="Divorciado(a)", 
            variable=estado_civil_var, 
            value="Divorciado(a)"
        ).pack(side='left', padx=10)    
        
        ttk.Radiobutton(
            frame_estado_civil, 
            text="Viúvo(a)", 
            variable=estado_civil_var, 
            value="Viúvo(a)"
        ).pack(side='left', padx=10)

        # Cidade
        ttk.Label(frame, text="Cidade:").pack(pady=5)
        cid_entry = ttk.Entry(frame, width=80)
        cid_entry.pack(pady=5)

        def validar_data(*args):
            """Valida se a data selecionada é dia 5 ou 20"""
            data = data_entry.get_date()
            if data.day not in [5, 20]:
                messagebox.showinfo(
                    "Data Inválida",
                    "A data inicial deve ser dia 5 ou 20 do mês.\n"
                    "Por favor, selecione uma data válida."
                )
                # Ajustar para próximo dia válido
                if data.day < 5:
                    data = data.replace(day=5)
                elif data.day < 20:
                    data = data.replace(day=20)
                else:
                    if data.month == 12:
                        data = data.replace(year=data.year + 1, month=1, day=5)
                    else:
                        data = data.replace(month=data.month + 1, day=5)
                data_entry.set_date(data)

        data_entry.bind("<<DateEntrySelected>>", validar_data)

        def salvar_cliente():
            nome = nome_entry.get().strip()
            endereco = endereco_entry.get().strip()
            data = data_entry.get()
            observacoes = obs_entry.get().strip()
            tipo_taxa = tipo_taxa_var.get()
            cpf = cpf_entry.get().strip()
            cno = cno_entry.get().strip()
            estado_civil = estado_civil_var.get()
            cidade = cid_entry.get().strip()
            
            if not nome or not endereco:
                messagebox.showerror("Erro", "Nome e Endereço são obrigatórios!")
                return

            try:
                data = datetime.strptime(data, '%Y-%m-%d').date()
                if data.day not in [5, 20]:
                    messagebox.showerror("Erro", "A data inicial deve ser dia 5 ou 20 do mês!")
                    return
            except ValueError:
                messagebox.showerror("Erro", "Data inválida!")
                return

            try:
                wb = load_workbook(ARQUIVO_CLIENTES)
                ws = wb['Clientes']

                # Verificar se cliente já existe
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[0].upper() == nome.upper():
                        messagebox.showerror("Erro", "Cliente já cadastrado!")
                        return

                # Adicionar novo cliente
                proxima_linha = ws.max_row + 1
                ws.cell(row=proxima_linha, column=1, value=nome.upper())
                ws.cell(row=proxima_linha, column=2, value=endereco.upper())
                ws.cell(row=proxima_linha, column=3, value=data)
                ws.cell(row=proxima_linha, column=4, value=observacoes.upper())
                ws.cell(row=proxima_linha, column=5, value=None)  # Data Final
                ws.cell(row=proxima_linha, column=6, value=tipo_taxa)  # NOVO: Tipo Taxa
                ws.cell(row=proxima_linha, column=7, value=cpf)  # CPF
                ws.cell(row=proxima_linha, column=8, value=cno)  # CNO
                ws.cell(row=proxima_linha, column=9, value=estado_civil)  # Estado Civil
                ws.cell(row=proxima_linha, column=10, value=cidade)  # Cidade

                wb.save(ARQUIVO_CLIENTES)

                # Criar arquivo do cliente
                if self.criar_arquivo_cliente(nome.upper(), endereco.upper()):
                    messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")
                    self.atualizar_lista_clientes()
                    janela_cliente.destroy()

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao cadastrar cliente: {str(e)}")

        ttk.Button(frame, text="Salvar", command=salvar_cliente).pack(pady=10)
        ttk.Button(frame, text="Cancelar", command=janela_cliente.destroy).pack(pady=5)

    def selecionar_cliente(self, event):
        """
        Atualiza seleção de cliente e habilita botão de continuar
        VERSÃO COM PROTEÇÃO contra perda de dados pendentes + FEEDBACK VISUAL MELHORADO
        """
        try:
            # ==========================================
            # PROTEÇÃO: Verificar dados pendentes ANTES de trocar
            # ==========================================
            novo_cliente = self.cliente_combobox.get()
            
            # Se já existe um cliente selecionado E está tentando trocar para outro
            if (hasattr(self, 'cliente_atual') and 
                self.cliente_atual and 
                self.cliente_atual != novo_cliente):
                
                # Verificar se há dados pendentes
                if hasattr(self, 'dados_para_incluir') and self.dados_para_incluir:
                    qtd_pendentes = len(self.dados_para_incluir)
                    
                    logger = system_logger.get_logger()
                    logger.warning(f"Tentativa de trocar cliente com {qtd_pendentes} lançamentos pendentes")
                    
                    resposta = custom_messagebox(
                        "yesno",
                        "Dados Pendentes - Confirmação Necessária",
                        f"⚠️ ATENÇÃO: Existem {qtd_pendentes} lançamento(s) pendente(s) de envio!\n\n"
                        f"Cliente atual: {self.cliente_atual}\n"
                        f"Novo cliente: {novo_cliente}\n\n"
                        f"Ao trocar de cliente, estes dados serão PERDIDOS e não poderão "
                        f"ser recuperados.\n\n"
                        f"O que deseja fazer?\n\n"
                        f"• SIM = Trocar de cliente e DESCARTAR os {qtd_pendentes} lançamentos pendentes\n"
                        f"• NÃO = Cancelar troca e ENVIAR os lançamentos primeiro"
                    )
                    
                    if not resposta:  # Usuário escolheu NÃO
                        logger.info("Troca de cliente CANCELADA pelo usuário - dados pendentes preservados")
                        
                        # Restaurar seleção anterior no combobox
                        self.cliente_combobox.set(self.cliente_atual)
                        
                        # ===== NOVO: Feedback visual de cancelamento =====
                        if hasattr(self, 'label_contador'):
                            self.label_contador.config(
                                text="⚠️ Troca cancelada - dados preservados",
                                foreground='orange'
                            )
                            # Restaurar após 3 segundos
                            self.root.after(3000, lambda: self.atualizar_contador(len(self.clientes_completos)))
                        
                        # Mensagem informativa
                        custom_messagebox(
                            "info", 
                            "Troca Cancelada", 
                            f"Troca de cliente cancelada.\n\n"
                            f"Os {qtd_pendentes} lançamentos pendentes foram preservados.\n\n"
                            f"Por favor:\n"
                            f"1. Clique em 'Enviar' para salvar os lançamentos\n"
                            f"2. Depois selecione o novo cliente"
                        )
                        
                        return  # IMPORTANTE: Interromper aqui sem trocar cliente
                    
                    # Se chegou aqui, usuário confirmou descarte (SIM)
                    logger.warning(f"Usuário CONFIRMOU descarte de {qtd_pendentes} lançamentos pendentes")
                    
                    # Limpar dados pendentes
                    self.limpar_visualizacao_completa()
                    
                    logger.info(f"Dados pendentes descartados. Trocando de '{self.cliente_atual}' para '{novo_cliente}'")
            
            # ==========================================
            # Continuar com seleção normal do cliente
            # ==========================================
            self.cliente_atual = novo_cliente
            
            # Atualiza label na aba de dados
            self.cliente_label.config(text=f"Cliente: {self.cliente_atual}")
            
            # Atualiza também o label na aba de fornecedor
            if hasattr(self, 'lbl_cliente_fornecedor'):
                self.lbl_cliente_fornecedor.config(text=f"Cliente: {self.cliente_atual}")
            
            # Habilita o botão continuar
            self.btn_continuar.config(state='normal')
            
            # ===== NOVO: Feedback visual de sucesso =====
            if hasattr(self, 'label_contador'):
                self.label_contador.config(
                    text="✓ Cliente selecionado",
                    foreground='green'
                )
                # Restaurar contador após 2 segundos
                self.root.after(2000, lambda: self.atualizar_contador(len(self.clientes_completos)))
            
            # Log de sucesso
            logger = system_logger.get_logger()
            logger.info(f"Cliente selecionado com sucesso: {self.cliente_atual}")
            
            # Não muda de aba automaticamente
            
        except Exception as e:
            logger = system_logger.get_logger()
            logger.error(f"Erro ao selecionar cliente: {str(e)}")
            custom_messagebox("error", "Erro", f"Erro ao selecionar cliente: {str(e)}")


    def editar_cliente(self):
        """Edita o cliente selecionado"""
        cliente_selecionado = self.cliente_combobox.get()
        
        if not cliente_selecionado:
            messagebox.showwarning("Aviso", "Selecione um cliente para editar")
            return

        try:
            # Carregar dados do cliente
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            dados_cliente = None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == cliente_selecionado:
                    dados_cliente = {
                        'nome': row[0],
                        'endereco': row[1],
                        'data_inicial': row[2],
                        'observacoes': row[3],
                        'data_final': row[4] if len(row) > 4 else None,
                        'tipo_taxa': row[5] if len(row) > 5 else 'Percentual',
                        'cpf': row[6] if len(row) > 6 else '',
                        'cno': row[7] if len(row) > 7 else '',
                        'estado_civil': row[8] if len(row) > 8 else '',
                        'cidade': row[9] if len(row) > 9 else ''
                    }
                    break
            
            wb.close()
            
            if not dados_cliente:
                messagebox.showerror("Erro", "Cliente não encontrado!")
                return
                
            # Criar janela de edição
            janela_edicao = tk.Toplevel(self.root)
            janela_edicao.title(f"Editar Cliente - {cliente_selecionado}")
            janela_edicao.geometry("600x430")

            frame = ttk.Frame(janela_edicao, padding="10")
            frame.pack(fill='both', expand=True)

            # Nome
            ttk.Label(frame, text="Nome do Cliente:*").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            nome_entry = ttk.Entry(frame, width=70)
            nome_entry.insert(0, dados_cliente['nome'])
            nome_entry.grid(row=0, column=1, padx=5, pady=5)

            # Endereço
            ttk.Label(frame, text="Endereço:*").grid(row=1, column=0, padx=5, pady=5, sticky='w')
            endereco_entry = ttk.Entry(frame, width=70)
            endereco_entry.insert(0, dados_cliente['endereco'])
            endereco_entry.grid(row=1, column=1, padx=5, pady=5)

            # Data Inicial
            ttk.Label(frame, text="Data Inicial:*").grid(row=2, column=0, padx=5, pady=5, sticky='w')
            data_inicial_entry = DateEntry(
                frame,
                width=20,
                date_pattern='yyyy-mm-dd',
                locale='pt_BR'
            )
            if dados_cliente['data_inicial']:
                data_inicial_entry.set_date(dados_cliente['data_inicial'])
            data_inicial_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')

            # Data Final
            ttk.Label(frame, text="Data Final:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
            
            tem_data_final = tk.BooleanVar(value=bool(dados_cliente['data_final']))
            
            check_data_final = ttk.Checkbutton(
                frame, 
                text="Obra finalizada",
                variable=tem_data_final
            )
            check_data_final.grid(row=3, column=1, padx=5, pady=5, sticky='w')
            
            data_final_entry = DateEntry(
                frame,
                width=20,
                date_pattern='yyyy-mm-dd',
                locale='pt_BR'
            )
            
            if dados_cliente['data_final']:
                data_final_entry.set_date(dados_cliente['data_final'])
            else:
                data_final_entry.delete(0, tk.END)
                data_final_entry.config(state='disabled')
            
            data_final_entry.grid(row=4, column=1, padx=5, pady=5, sticky='w')

            # NOVO: Tipo de Taxa
            ttk.Label(frame, text="Tipo de Taxa:*").grid(row=5, column=0, padx=5, pady=5, sticky='w')
            tipo_taxa_var = tk.StringVar(value=dados_cliente['tipo_taxa'])
            
            frame_tipo = ttk.Frame(frame)
            frame_tipo.grid(row=5, column=1, padx=5, pady=5, sticky='w')
            
            ttk.Radiobutton(
                frame_tipo, 
                text="Percentual", 
                variable=tipo_taxa_var, 
                value="Percentual"
            ).grid(row=5, column=1, padx=5, pady=5, sticky='w')
            
            ttk.Radiobutton(
                frame_tipo, 
                text="Fixo", 
                variable=tipo_taxa_var, 
                value="Fixo"
            ).grid(row=5, column=2, padx=5, pady=5, sticky='w')
            
            ttk.Radiobutton(
                frame_tipo, 
                text="Sem Taxa", 
                variable=tipo_taxa_var, 
                value="Sem Taxa"
            ).grid(row=5, column=3, padx=5, pady=5, sticky='w')

            # Observações
            ttk.Label(frame, text="Observações:").grid(row=6, column=0, padx=5, pady=5, sticky='w')
            obs_entry = ttk.Entry(frame, width=70)
            obs_entry.insert(0, dados_cliente['observacoes'] or '')
            obs_entry.grid(row=6, column=1, padx=5, pady=5)

            # CPF
            ttk.Label(frame, text="CPF:").grid(row=7, column=0, padx=5, pady=5, sticky='w')
            cpf_entry = ttk.Entry(frame, width=70)
            cpf_entry.insert(0, dados_cliente['cpf'] or '')
            cpf_entry.grid(row=7, column=1, padx=5, pady=5)

            # CNO
            ttk.Label(frame, text="CNO:").grid(row=8, column=0, padx=5, pady=5, sticky='w')
            cno_entry = ttk.Entry(frame, width=70)
            cno_entry.insert(0, dados_cliente['cno'] or '')
            cno_entry.grid(row=8, column=1, padx=5, pady=5) 

            # Estado civil
            ttk.Label(frame, text="Estado Civil:").grid(row=9, column=0, padx=5, pady=5, sticky='w')
            estado_civil_var = tk.StringVar(value="Casado(a)")
            frame_estado_civil = ttk.Frame(frame)
            frame_estado_civil.grid(row=9, column=1, padx=5, pady=5, sticky='w')
            
            ttk.Radiobutton(
                frame_estado_civil, 
                text="Casado(a)", 
                variable=estado_civil_var, 
                value="Casado(a)"
            ).grid(row=9, column=1, padx=5, pady=5, sticky='w')
            
            ttk.Radiobutton(
                frame_estado_civil, 
                text="Solteiro(a)", 
                variable=estado_civil_var, 
                value="Solteiro(a)"
            ).grid(row=9, column=2, padx=5, pady=5, sticky='w')
            
            ttk.Radiobutton(
                frame_estado_civil, 
                text="Divorciado(a)", 
                variable=estado_civil_var, 
                value="Divorciado(a)"
            ).grid(row=9, column=3, padx=5, pady=5, sticky='w')    
            
            ttk.Radiobutton(
                frame_estado_civil, 
                text="Viúvo(a)", 
                variable=estado_civil_var, 
                value="Viúvo(a)"
            ).grid(row=9, column=4, padx=5, pady=5, sticky='w') 

            # Cidade
            ttk.Label(frame, text="Cidade:").grid(row=10, column=0, padx=5, pady=5, sticky='w')
            cid_entry = ttk.Entry(frame, width=70)
            cid_entry.insert(0, dados_cliente['cidade'] or '')
            cid_entry.grid(row=10, column=1, padx=5, pady=5)


            def toggle_data_final():
                if tem_data_final.get():
                    data_final_entry.config(state='normal')
                    if not data_final_entry.get():
                        data_final_entry.set_date(datetime.now().date())
                else:
                    data_final_entry.delete(0, tk.END)
                    data_final_entry.config(state='disabled')
            
            check_data_final.config(command=toggle_data_final)

            def salvar_alteracoes():
                try:
                    nome = nome_entry.get().strip()
                    endereco = endereco_entry.get().strip()
                    
                    if not nome or not endereco:
                        messagebox.showerror("Erro", "Nome e Endereço são obrigatórios!")
                        return

                    wb = load_workbook(ARQUIVO_CLIENTES)
                    ws = wb['Clientes']

                    # Remover registros antigos
                    linhas_para_remover = []
                    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if row[0].value == cliente_selecionado:
                            linhas_para_remover.append(idx)

                    for linha in reversed(linhas_para_remover):
                        ws.delete_rows(linha)

                    # Adicionar novo registro
                    proxima_linha = ws.max_row + 1
                    ws.cell(row=proxima_linha, column=1, value=nome.upper())
                    ws.cell(row=proxima_linha, column=2, value=endereco.upper())
                    ws.cell(row=proxima_linha, column=3, value=data_inicial_entry.get_date())
                    ws.cell(row=proxima_linha, column=4, value=obs_entry.get().upper())
                    
                    if tem_data_final.get():
                        ws.cell(row=proxima_linha, column=5, value=data_final_entry.get_date())
                    else:
                        ws.cell(row=proxima_linha, column=5, value=None)
                    
                    # NOVO: Salvar Tipo Taxa
                    ws.cell(row=proxima_linha, column=6, value=tipo_taxa_var.get())
                    ws.cell(row=proxima_linha, column=7, value=cpf_entry.get().strip())
                    ws.cell(row=proxima_linha, column=8, value=cno_entry.get().strip())
                    ws.cell(row=proxima_linha, column=9, value=estado_civil_var.get())
                    ws.cell(row=proxima_linha, column=10, value=cid_entry.get().strip())

                    wb.save(ARQUIVO_CLIENTES)
                    
                    # Renomear arquivo se mudou o nome
                    if nome.upper() != cliente_selecionado:
                        caminho_antigo = PASTA_CLIENTES / f"{cliente_selecionado}.xlsx"
                        caminho_novo = PASTA_CLIENTES / f"{nome.upper()}.xlsx"
                        if os.path.exists(caminho_antigo):
                            os.rename(caminho_antigo, caminho_novo)

                    messagebox.showinfo("Sucesso", "Cliente atualizado com sucesso!")
                    self.atualizar_lista_clientes()
                    janela_edicao.destroy()

                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao salvar alterações: {str(e)}")

            # Botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.grid(row=11, column=0, columnspan=2, pady=20)

            ttk.Button(frame_botoes, text="Salvar", command=salvar_alteracoes).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Cancelar", command=janela_edicao.destroy).pack(side='left', padx=5)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir editor: {str(e)}") 

    def atualizar_lista_clientes(self):
        """Atualiza lista de clientes - LÓGICA CORRETA"""
        try:
            if not os.path.exists(ARQUIVO_CLIENTES):
                self.criar_arquivo_clientes()
                return
                
            workbook = load_workbook(ARQUIVO_CLIENTES)
            sheet = workbook['Clientes']
            
            clientes_ativos = []
            
            for row in range(2, sheet.max_row + 1):
                try:
                    nome = sheet.cell(row=row, column=1).value
                    data_final = sheet.cell(row=row, column=5).value  # Coluna E = Data Final
                    
                    if not nome or not isinstance(nome, str):
                        continue
                    
                    nome = nome.strip()
                    if not nome:
                        continue
                    
                    # LÓGICA CORRETA:
                    # Cliente FINALIZADO = tem Data Final preenchida (qualquer valor não-None)
                    # Cliente ATIVO = Data Final vazia (None)
                    is_finalizado = (data_final is not None)
                    
                    if not is_finalizado:
                        clientes_ativos.append(nome)
                        
                except Exception as e:
                    logger = system_logger.get_logger()
                    logger.warning(f"Erro linha {row}: {str(e)}")
                    continue
            
            workbook.close()
            
            clientes_ativos.sort()
            self.clientes_completos = clientes_ativos
            self.cliente_combobox['values'] = clientes_ativos
            
            if hasattr(self, 'label_contador'):
                self.atualizar_contador(len(clientes_ativos))
            
            if not clientes_ativos and hasattr(self, 'label_contador'):
                self.label_contador.config(
                    text="Nenhum cliente cadastrado",
                    foreground='orange'
                )
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar clientes: {str(e)}")
            logger = system_logger.get_logger()
            logger.error(f"Erro detalhado: {str(e)}")
            if not hasattr(self, 'clientes_completos'):
                self.clientes_completos = []


    def criar_arquivo_clientes(self):
        """Cria arquivo base de clientes se não existir"""
        try:
            print(f"Tentando criar arquivo de clientes em: {ARQUIVO_CLIENTES}")
            print(f"Diretório existe? {os.path.exists(os.path.dirname(ARQUIVO_CLIENTES))}")
            
            # Garantir que o diretório existe
            os.makedirs(os.path.dirname(ARQUIVO_CLIENTES), exist_ok=True)
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Clientes'
            
            # Adicionar cabeçalhos - agora incluindo Data_Final
            headers = ['Nome', 'Endereco', 'Data_Inicial', 'Observacoes', 'Data_Final']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            print(f"Tentando salvar arquivo em: {ARQUIVO_CLIENTES}")
            workbook.save(ARQUIVO_CLIENTES)
            custom_messagebox("info", "Informação", "Arquivo de clientes criado com sucesso!")
            
        except Exception as e:
            print(f"Erro detalhado ao criar arquivo de clientes: {str(e)}")
            print(f"Tipo do erro: {type(e)}")
            custom_messagebox("error", "Erro", f"Erro ao criar arquivo de clientes: {str(e)}")

    def mostrar_todos_clientes(self):
        """Mostra todos os clientes - JANELA CORRIGIDA"""
        try:
            if not os.path.exists(ARQUIVO_CLIENTES):
                custom_messagebox("warning", "Aviso", "Arquivo não encontrado")
                return
                
            workbook = load_workbook(ARQUIVO_CLIENTES)
            sheet = workbook['Clientes']
            
            clientes_ativos = []
            clientes_finalizados = []
            
            for row in range(2, sheet.max_row + 1):
                try:
                    nome = sheet.cell(row=row, column=1).value
                    data_final = sheet.cell(row=row, column=5).value  # Coluna E = Data Final
                    
                    if not nome or not isinstance(nome, str):
                        continue
                    
                    nome = nome.strip()
                    if not nome:
                        continue
                    
                    # LÓGICA CORRETA:
                    # Cliente FINALIZADO = tem Data Final preenchida
                    # Cliente ATIVO = Data Final vazia (None)
                    is_finalizado = (data_final is not None)
                    
                    if is_finalizado:
                        clientes_finalizados.append(nome)
                    else:
                        clientes_ativos.append(nome)
                        
                except Exception as e:
                    logger = system_logger.get_logger()
                    logger.warning(f"Erro linha {row}: {str(e)}")
                    continue
            
            workbook.close()
            
            # ===== JANELA COM DIMENSÕES CORRETAS =====
            janela_todos = tk.Toplevel(self.root)
            janela_todos.title("Todos os Clientes")
            janela_todos.geometry("700x600")  # ← AUMENTADO de 600x500
            janela_todos.minsize(600, 500)    # ← TAMANHO MÍNIMO
            janela_todos.transient(self.root)
            janela_todos.grab_set()
            
            # Frame principal com padding adequado
            frame = ttk.Frame(janela_todos, padding="20")  # ← AUMENTADO padding
            frame.pack(fill='both', expand=True)
            
            # Título
            ttk.Label(
                frame, 
                text="Visualização Completa de Clientes",
                font=('Arial', 14, 'bold')  # ← FONTE MAIOR
            ).pack(pady=(0, 20))
            
            # ===== CLIENTES ATIVOS =====
            frame_ativos = ttk.LabelFrame(
                frame, 
                text=f"Clientes Ativos ({len(clientes_ativos)})",
                padding=(10, 10)
            )
            frame_ativos.pack(fill='both', expand=True, pady=(0, 15))
            
            frame_lista_ativos = ttk.Frame(frame_ativos)
            frame_lista_ativos.pack(fill='both', expand=True)
            
            scrollbar_ativos = ttk.Scrollbar(frame_lista_ativos)
            scrollbar_ativos.pack(side='right', fill='y')
            
            lista_ativos = tk.Listbox(
                frame_lista_ativos,
                font=('Arial', 10),
                yscrollcommand=scrollbar_ativos.set,
                height=10  # ← ALTURA FIXA
            )
            lista_ativos.pack(side='left', fill='both', expand=True)
            scrollbar_ativos.config(command=lista_ativos.yview)
            
            for cliente in sorted(clientes_ativos):
                lista_ativos.insert(tk.END, cliente)
            
            # ===== CLIENTES FINALIZADOS =====
            frame_finalizados = ttk.LabelFrame(
                frame,
                text=f"Clientes Finalizados ({len(clientes_finalizados)})",
                padding=(10, 10)
            )
            frame_finalizados.pack(fill='both', expand=True, pady=(0, 15))
            
            frame_lista_fin = ttk.Frame(frame_finalizados)
            frame_lista_fin.pack(fill='both', expand=True)
            
            scrollbar_fin = ttk.Scrollbar(frame_lista_fin)
            scrollbar_fin.pack(side='right', fill='y')
            
            lista_finalizados = tk.Listbox(
                frame_lista_fin,
                font=('Arial', 10),
                foreground='gray',
                yscrollcommand=scrollbar_fin.set,
                height=10  # ← ALTURA FIXA
            )
            lista_finalizados.pack(side='left', fill='both', expand=True)
            scrollbar_fin.config(command=lista_finalizados.yview)
            
            for cliente in sorted(clientes_finalizados):
                lista_finalizados.insert(tk.END, cliente)
            
            # ===== BOTÕES COM LAYOUT CORRETO =====
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=(15, 0))
            
            # Criar style se não existir
            style = ttk.Style()
            style.configure('Botao.TButton', font=('Arial', 10), padding=(10, 8))
            
            def selecionar_ativo():
                try:
                    selected = lista_ativos.curselection()
                    if not selected:
                        custom_messagebox("warning", "Aviso", "Selecione um cliente ativo")
                        return
                    cliente = lista_ativos.get(selected[0])
                    self.cliente_combobox.set(cliente)
                    self.selecionar_cliente(None)
                    janela_todos.destroy()
                except Exception as e:
                    logger = system_logger.get_logger()
                    logger.error(f"Erro: {str(e)}")
            
            def selecionar_finalizado():
                try:
                    selected = lista_finalizados.curselection()
                    if not selected:
                        custom_messagebox("warning", "Aviso", "Selecione um cliente finalizado")
                        return
                    cliente = lista_finalizados.get(selected[0])
                    self.cliente_combobox.set(cliente)
                    self.selecionar_cliente(None)
                    janela_todos.destroy()
                except Exception as e:
                    logger = system_logger.get_logger()
                    logger.error(f"Erro: {str(e)}")
            
            # LAYOUT DE BOTÕES CORRIGIDO - SEM pack(fill='x')
            # Botões alinhados à esquerda com tamanhos fixos
            btn_ativo = ttk.Button(
                frame_botoes,
                text="✓ Selecionar Cliente Ativo",
                command=selecionar_ativo,
                style='Botao.TButton',
                width=25  # ← LARGURA FIXA
            )
            btn_ativo.pack(side='left', padx=(0, 10))
            
            btn_finalizado = ttk.Button(
                frame_botoes,
                text="✓ Selecionar Cliente Finalizado",
                command=selecionar_finalizado,
                style='Botao.TButton',
                width=28  # ← LARGURA FIXA
            )
            btn_finalizado.pack(side='left', padx=(0, 10))
            
            # Botão Fechar à direita com largura fixa
            btn_fechar = ttk.Button(
                frame_botoes,
                text="Fechar",
                command=janela_todos.destroy,
                style='Botao.TButton',
                width=15  # ← LARGURA FIXA
            )
            btn_fechar.pack(side='right')
            
            # Duplo clique
            lista_ativos.bind('<Double-Button-1>', lambda e: selecionar_ativo())
            lista_finalizados.bind('<Double-Button-1>', lambda e: selecionar_finalizado())
            
            # Centralizar janela na tela
            janela_todos.update_idletasks()
            x = (janela_todos.winfo_screenwidth() // 2) - (janela_todos.winfo_width() // 2)
            y = (janela_todos.winfo_screenheight() // 2) - (janela_todos.winfo_height() // 2)
            janela_todos.geometry(f"+{x}+{y}")
                    
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro: {str(e)}")
            logger = system_logger.get_logger()
            logger.error(f"Erro detalhado: {str(e)}")


    def criar_arquivo_cliente(self, nome_cliente, endereco):
        """Cria um novo arquivo Excel para o cliente baseado no MODELO.xlsx"""
        try:
            print(f"\nTentando criar arquivo para cliente: {nome_cliente}")
            print(f"ARQUIVO_MODELO: {ARQUIVO_MODELO}")
            print(f"ARQUIVO_MODELO existe? {os.path.exists(ARQUIVO_MODELO)}")
            print(f"PASTA_CLIENTES: {PASTA_CLIENTES}")
            print(f"PASTA_CLIENTES existe? {os.path.exists(PASTA_CLIENTES)}")
            
            modelo_path = ARQUIVO_MODELO
            novo_arquivo = PASTA_CLIENTES / f"{nome_cliente}.xlsx"
            
            print(f"Novo arquivo será criado em: {novo_arquivo}")
            print(f"Diretório do novo arquivo existe? {os.path.exists(os.path.dirname(novo_arquivo))}")
                
            if os.path.exists(novo_arquivo):
                print(f"Arquivo {novo_arquivo} já existe!")
                raise Exception("Arquivo do cliente já existe!")
                    
            # Garantir que o diretório existe
            os.makedirs(os.path.dirname(novo_arquivo), exist_ok=True)
                
            print(f"Tentando copiar de {modelo_path} para {novo_arquivo}")
            
            # Copiar o arquivo modelo
            from shutil import copy2
            copy2(modelo_path, novo_arquivo)
            
            print("Arquivo copiado com sucesso")
                
            # Buscar data inicial do arquivo clientes.xlsx
            wb_clientes = load_workbook(ARQUIVO_CLIENTES)
            ws_clientes = wb_clientes['Clientes']
            
            data_inicial = None
            # Procurar o cliente e sua data inicial
            for row in range(2, ws_clientes.max_row + 1):
                if ws_clientes.cell(row=row, column=1).value == nome_cliente:
                    data_valor = ws_clientes.cell(row=row, column=3).value  # Coluna C
                    if not data_valor:
                        raise Exception("Data inicial não informada no cadastro do cliente")
                        
                    if isinstance(data_valor, datetime):
                        data_inicial = data_valor.date()
                    else:
                        try:
                            data_inicial = datetime.strptime(str(data_valor), '%Y-%m-%d').date()
                        except ValueError:
                            raise Exception("Data inicial deve estar no formato AAAA-MM-DD")
                    break
            
            if not data_inicial:
                raise Exception("Cliente não encontrado no cadastro")
                
            # Validar se é dia 5 ou 20
            if data_inicial.day not in [5, 20]:
                raise Exception("A data inicial deve ser dia 5 ou 20 do mês")
                
            # Abrir o novo arquivo para edição
            workbook = load_workbook(novo_arquivo)
            
            # Atualizar planilha RESUMO
            resumo_sheet = workbook["RESUMO"]
            
            # Informações básicas
            resumo_sheet["A3"] = nome_cliente
            resumo_sheet["A4"] = endereco
            
            # Descrições das células
            resumo_sheet["K3"] = "Data Inicial"
            
            # Adicionar data inicial
            resumo_sheet["L3"] = data_inicial
            resumo_sheet["L3"].number_format = 'dd/mm/yyyy'
            
            # Gerar as 96 datas quinzenais
            data_atual = data_inicial
            datas_geradas = []
            
            for i in range(96):  # 4 anos = 96 relatórios
                row = i + 9  # Começar na linha 9
                
                # Verificar se a data já foi usada
                if data_atual in datas_geradas:
                    raise Exception(f"Data duplicada detectada: {data_atual.strftime('%d/%m/%Y')}")
                datas_geradas.append(data_atual)
                
                # Adicionar data e número do relatório
                resumo_sheet.cell(row=row, column=1, value=data_atual)
                resumo_sheet.cell(row=row, column=1).number_format = 'dd/mm/yyyy'
                resumo_sheet.cell(row=row, column=2, value=i + 1)
                
                # Próxima data
                if data_atual.day == 5:
                    data_atual = data_atual.replace(day=20)
                else:  # day == 20
                    if data_atual.month == 12:
                        data_atual = data_atual.replace(year=data_atual.year + 1, month=1, day=5)
                    else:
                        data_atual = data_atual.replace(month=data_atual.month + 1, day=5)

            # Criar aba Contratos_ADM
            contratos_sheet = workbook.create_sheet("Contratos_ADM")
            
            # Definir os blocos na linha 1
            blocos = ["CONTRATOS", "", "", "", "", "",
                     "ADMINISTRADORES_CONTRATO", "", "", "", "", "", "",
                     "ADITIVOS", "", "", "",
                     "ADMINISTRADORES_ADITIVO", "", "", "", "", "", "",
                     "PARCELAS", "", "", "", "", "", "", ""]
            
            for col, valor in enumerate(blocos, 1):
                contratos_sheet.cell(row=1, column=col, value=valor)
            
            # Definir cabeçalhos na linha 2
            headers = [
                # CONTRATOS
                "Nº Contrato", "Data Início", "Data Fim", "Status", "Observações", "",
                # ADMINISTRADORES_CONTRATO
                "Nº Contrato", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total", "Nº Parcelas", 
                # ADITIVOS
                "Nº Contrato", "Nº Aditivo", "Data Início", "Data Fim",
                # ADMINISTRADORES_ADITIVO
                "Nº Contrato", "Nº Aditivo", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total",
                # PARCELAS
                "Referência", "Número", "CNPJ/CPF", "Nome", "Data Vencimento", "Valor", "Status", "Data Pagamento", "Eventos/Fases", "Percentual"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = contratos_sheet.cell(row=2, column=col, value=header)
                # Formatação do cabeçalho
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Ajustar largura das colunas
            for col in range(1, len(headers) + 1):
                contratos_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # Salvar alterações
            workbook.save(novo_arquivo)
            wb_clientes.close()
            
            return True
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao criar arquivo do cliente: {str(e)}")
            if 'wb_clientes' in locals():
                wb_clientes.close()
            return False

    
    def abrir_gestao_contratos(self):
        """Abre a gestão de contratos para o cliente atual"""
        if not self.cliente_atual:
            custom_messagebox("warning", "Aviso", "Selecione um cliente primeiro!")
            return
        
        # Ocultar temporariamente a janela principal
        self.root.withdraw()
        
        # Criar e configurar a janela de gestão de contratos diretamente aqui
        # em vez de delegar para outra classe/método
        janela_gestao = tk.Toplevel(self.root)
        janela_gestao.title(f"Gestão de Contratos - {self.cliente_atual}")
        janela_gestao.geometry("800x750")
        
        # Centralizar a janela (sem depender de um método da classe GestaoContratos)
        janela_gestao.update_idletasks()
        width = janela_gestao.winfo_width()
        height = janela_gestao.winfo_height()
        x = (janela_gestao.winfo_screenwidth() // 2) - (width // 2)
        y = (janela_gestao.winfo_screenheight() // 2) - (height // 2)
        janela_gestao.geometry(f'{width}x{height}+{x}+{y}')
        
        # Colocar a janela em primeiro plano
        janela_gestao.attributes('-topmost', True)
        janela_gestao.after(100, lambda: janela_gestao.attributes('-topmost', False))
        
        # Definir comportamento quando a janela for fechada
        def on_close():
            janela_gestao.destroy()
            self.root.deiconify()  # Mostrar a janela principal novamente
            self.root.lift()
            self.root.focus_force()
        
        # Configurar protocolo de fechamento
        janela_gestao.protocol("WM_DELETE_WINDOW", on_close)
        
        # Criar o restante da interface usando o gestor de contratos
        gestor = GestaoContratos(janela_gestao)  # Passamos a janela_gestao como parent
        gestor.cliente_atual = self.cliente_atual
        gestor.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
        
        # Criar e preencher a interface dentro da janela_gestao
        gestor.criar_interface_contratos(janela_gestao, on_close)

    def abrir_controle_pagamentos(self):
        """Abre o módulo de controle de pagamentos"""
        try:
            # Importar módulo
            from controle_pagamentos import ControlePagamentos
            
            # Instanciar e abrir janela de controle
            controle = ControlePagamentos(self.root)
        except ImportError as e:
            custom_messagebox("error", "Erro", f"Não foi possível importar o módulo de Controle de Pagamentos: {str(e)}")
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir controle de pagamentos: {str(e)}")

    
    def configurar_protecoes(self):
        """
        Configura proteções contra perda de dados
        Deve ser chamado no __init__ da classe principal
        """
        try:
            # Vincular evento de fechamento da janela
            if hasattr(self, 'janela_principal'):
                self.janela_principal.protocol("WM_DELETE_WINDOW", self.fechar_aplicacao)
            
            # Vincular evento de mudança no combo de clientes
            if hasattr(self, 'combo_cliente'):
                # Remover binding anterior se existir
                self.combo_cliente.unbind('<<ComboboxSelected>>')
                # Adicionar novo binding com proteção
                self.combo_cliente.bind('<<ComboboxSelected>>', self.selecionar_cliente)
            
            logger = system_logger.get_logger()
            logger.info("Proteções contra perda de dados configuradas")
            
        except Exception as e:
            logger = system_logger.get_logger()
            logger.error(f"Erro ao configurar proteções: {str(e)}")

    def continuar_para_fornecedor(self):
        """Avança para a aba de fornecedor após confirmar seleção"""
        if self.cliente_atual:
            self.notebook.select(1)  # Vai para aba de fornecedor
        else:
            custom_messagebox("warning",  "Aviso", "Selecione um cliente primeiro!")

    
    def setup_aba_fornecedor(self):
        """Configura a aba de fornecedor com layout elegante e botões de tamanho médio"""
        # Criar estilo para botões médios
        style = ttk.Style()
        style.configure('Medium.TButton', font=('Arial', 11), padding=(10, 6))
        
        # Frame para exibir o cliente selecionado
        frame_cliente = ttk.Frame(self.aba_fornecedor)
        frame_cliente.pack(fill='x', padx=10, pady=5)
        
        # Label para mostrar o cliente selecionado
        self.lbl_cliente_fornecedor = ttk.Label(
            frame_cliente, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_fornecedor.pack(anchor='w', padx=5)
        
        # Frame de busca com tamanho reduzido
        frame_busca = ttk.LabelFrame(self.aba_fornecedor, text="Busca de Fornecedor")
        frame_busca.pack(fill='x', padx=10, pady=5)

        # Frame interno para organizar os elementos de busca
        busca_interno = ttk.Frame(frame_busca)
        busca_interno.pack(fill='x', padx=5, pady=5)

        # Campo de busca
        ttk.Label(busca_interno, text="Nome:", font=('Arial', 10)).pack(side='left', padx=5)
        self.busca_entry = ttk.Entry(busca_interno, font=('Arial', 10), width=40)
        self.busca_entry.pack(side='left', padx=5)
        self.busca_entry.bind('<Return>', lambda e: self.buscar_fornecedor())

        # Botão de busca
        ttk.Button(busca_interno, 
            text="Buscar", 
            command=self.buscar_fornecedor,
            style='Medium.TButton').pack(side='left', padx=10)

        ttk.Button(busca_interno, 
                text="📋 Lançamentos", 
                command=self.abrir_visualizador_fornecedor,
                style='Medium.TButton').pack(side='left', padx=5)

        # Frame principal para resultados
        frame_resultados = ttk.Frame(self.aba_fornecedor)
        frame_resultados.pack(fill='both', expand=True, padx=10, pady=5)

        # Lista de resultados com scrollbar
        frame_tree = ttk.Frame(frame_resultados)
        frame_tree.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Scrollbar vertical
        scroll_y = ttk.Scrollbar(frame_tree, orient='vertical')
        scroll_y.pack(side='right', fill='y')
        
        # Treeview para resultados
        self.tree_fornecedores = ttk.Treeview(frame_tree, 
                                            columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                            show='headings',
                                            yscrollcommand=scroll_y.set) 
        
        self.tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_fornecedores.heading('Nome', text='Nome')
        self.tree_fornecedores.heading('Categoria', text='Categoria')
        
        # Configurar larguras das colunas
        self.tree_fornecedores.column('CNPJ/CPF', width=150)
        self.tree_fornecedores.column('Nome', width=300)
        self.tree_fornecedores.column('Categoria', width=100)
        
        self.tree_fornecedores.pack(side='left', fill='both', expand=True)
        scroll_y.config(command=self.tree_fornecedores.yview)
        
        # Adicionar evento de duplo clique para selecionar fornecedor
        self.tree_fornecedores.bind('<Double-1>', lambda e: self.selecionar_fornecedor())

        # Frame para botões de ação do fornecedor
        frame_acoes = ttk.Frame(self.aba_fornecedor)
        frame_acoes.pack(fill='x', padx=10, pady=5)

        ttk.Button(frame_acoes, 
                text="Novo Fornecedor", 
                command=self.novo_fornecedor,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_acoes, 
                text="Editar Fornecedor", 
                command=self.editar_fornecedor,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_acoes, 
                text="Selecionar", 
                command=self.selecionar_fornecedor,
                style='Medium.TButton').pack(side='left', padx=5)
        
        ttk.Button(frame_acoes,
                text="🗑️ Gerenciar CPFs Criados",
                command=self.gerenciar_cpfs_criados,
                style='Medium.TButton').pack(side='left', padx=5)
                    
        # Label explicativo
        ttk.Label(frame_acoes,
                text="(Excluir fornecedores temporários)",
                font=('Arial', 8),
                foreground='gray').pack(side='left', padx=10)

        self.adicionar_botao_gerenciar_lancamentos()

        # Separador para dividir visualmente as seções
        ttk.Separator(self.aba_fornecedor, orient='horizontal').pack(fill='x', padx=10, pady=5)

        # Frame para taxas e processamento
        frame_taxas = ttk.LabelFrame(self.aba_fornecedor, text="Funções Administrativas")
        frame_taxas.pack(fill='x', padx=10, pady=5)

        # Container para botões de taxas
        frame_botoes_taxas = ttk.Frame(frame_taxas)
        frame_botoes_taxas.pack(fill='x', padx=5, pady=8)

        ttk.Button(
            frame_botoes_taxas, 
            text="Controle de Pagamentos de Taxa",
            command=self.abrir_controle_pagamentos,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_taxas, 
            text="Finalização de Quinzena",
            command=self.abrir_finalizacao_quinzena,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_taxas, 
            text="Correção Monetária",
            command=self.abrir_correcao_monetaria,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        frame_botoes_verificacao = ttk.Frame(frame_taxas)
        frame_botoes_verificacao.pack(fill='x', padx=5, pady=(0, 8))

        ttk.Button(
            frame_botoes_verificacao, 
            text="🔍 Verificar Consistência das Taxas",
            command=lambda: self.verificar_e_mostrar_consistencia(),
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        # Separador para dividir visualmente as seções
        ttk.Separator(self.aba_fornecedor, orient='horizontal').pack(fill='x', padx=10, pady=5)

        # Frame de botões gerais na parte inferior
        frame_botoes_fornecedor = ttk.Frame(self.aba_fornecedor)
        frame_botoes_fornecedor.pack(fill='x', padx=10, pady=10, side='bottom')
        
        ttk.Button(
            frame_botoes_fornecedor, 
            text="🚛 Importar Transporte", 
            command=self.importar_transporte_cafe,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_fornecedor, 
            text="Importar Folha RH", 
            command=self.importar_folha_rh,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_fornecedor, 
            text="🔧 Gestão de Locações", 
            command=self.abrir_gestao_locacoes,
            style='Medium.TButton'
        ).pack(side='left', padx=5)
       
        ttk.Button(frame_botoes_fornecedor, 
                text="Voltar ao Menu", 
                command=self.voltar_menu,
                style='Medium.TButton').pack(side='right', padx=5)
        ttk.Button(frame_botoes_fornecedor, 
                text="Sair", 
                command=self.sair_sistema,
                style='Medium.TButton').pack(side='right', padx=5)
        
        # self.adicionar_secao_materiais_fornecedor()
  
    def buscar_fornecedor(self):
        """Busca fornecedores baseado no termo digitado"""
        try:
            termo = self.busca_entry.get().strip()
            
            # Limpar resultados anteriores
            for item in self.tree_fornecedores.get_children():
                self.tree_fornecedores.delete(item)
            
            if not termo:
                return
            
            # Abrir planilha de fornecedores
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
            
            resultados_encontrados = 0
            termo_upper = termo.upper()
            
            # Buscar na planilha
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular linhas vazias
                    continue
                    
                cnpj_cpf = str(row[0]).strip()
                nome = str(row[3] or '').strip().upper()  # Coluna D = Nome
                categoria = str(row[11] or '').strip()    # Coluna L = Categoria
                
                # Verificar se o termo está no nome
                if termo_upper in nome:
                    # Formatar CNPJ/CPF para exibição
                    cnpj_cpf_formatado = formatar_cnpj_cpf(cnpj_cpf)
                    
                    # Inserir resultado na tree
                    self.tree_fornecedores.insert('', 'end', values=(
                        cnpj_cpf_formatado,
                        nome,
                        categoria
                    ))
                    
                    resultados_encontrados += 1
                    
                    # Limitar resultados para evitar travamento
                    if resultados_encontrados >= 100:
                        break
            
            wb.close()
            
            # Ordenar resultados por nome
            if resultados_encontrados > 1:
                # Obter todos os itens
                items = []
                for item in self.tree_fornecedores.get_children():
                    values = self.tree_fornecedores.item(item)['values']
                    items.append(values)
                
                # Limpar tree
                for item in self.tree_fornecedores.get_children():
                    self.tree_fornecedores.delete(item)
                
                # Ordenar por nome (segundo elemento)
                items.sort(key=lambda x: x[1])
                
                # Reinserir ordenado
                for values in items:
                    self.tree_fornecedores.insert('', 'end', values=values)
            
            if resultados_encontrados == 0:
                # Mostrar mensagem quando não encontrar
                self.tree_fornecedores.insert('', 'end', values=(
                    '', 'Nenhum fornecedor encontrado', ''
                ))
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro na busca: {str(e)}")
            print(f"Erro detalhado na busca: {str(e)}")

    def buscar_fornecedores_por_nome_parcial(self, nome_parcial):
        """
        Busca otimizada por nome parcial usando cache
        VERSÃO FINAL OTIMIZADA
        """
        try:
            if not nome_parcial or len(nome_parcial) < 3:
                return []
            
            # Carregar fornecedores do cache
            fornecedores = self.cache_fornecedores.carregar_cache_se_necessario(ARQUIVO_FORNECEDORES)
            
            if not fornecedores:
                print("DEBUG: Cache vazio, fazendo busca direta")
                return self.buscar_fornecedores_por_nome_parcial_direto(nome_parcial)
            
            nome_busca = nome_parcial.strip().upper()
            fornecedores_encontrados = []
            
            print(f"DEBUG: Buscando '{nome_busca}' em {len(fornecedores)} fornecedores do cache")
            
            for fornecedor in fornecedores:
                nome = fornecedor['nome']
                
                if nome_busca in nome:
                    # Calcular relevância (posição no nome)
                    posicao = nome.find(nome_busca)
                    relevancia = 1000 - posicao
                    
                    fornecedor_resultado = fornecedor.copy()
                    fornecedor_resultado['relevancia'] = relevancia
                    
                    fornecedores_encontrados.append(fornecedor_resultado)
                    
                    # Limitar resultados para performance
                    if len(fornecedores_encontrados) >= 15:
                        break
            
            # Ordenar por relevância
            fornecedores_encontrados.sort(key=lambda x: (-x['relevancia'], x['nome']))
            
            print(f"DEBUG: {len(fornecedores_encontrados)} fornecedores encontrados no cache")
            
            return fornecedores_encontrados
            
        except Exception as e:
            print(f"DEBUG: Erro na busca por cache: {str(e)}")
            # Fallback para busca direta
            return self.buscar_fornecedores_por_nome_parcial_direto(nome_parcial)

    def novo_fornecedor(self):
        """Abre janela para cadastro de novo fornecedor - VERSÃO CORRIGIDA"""
        self.janela_fornecedor = tk.Toplevel(self.root)
        self.janela_fornecedor.title("Novo Fornecedor")
        self.janela_fornecedor.geometry("800x700")
        
        self.janela_fornecedor.transient(self.root)  # Definir como janela filha
        self.janela_fornecedor.grab_set()  # Tornar modal
        self.janela_fornecedor.update_idletasks()  # Garantir que o tamanho seja calculado
        
        # Calcular posição central
        largura = 800
        altura = 700
        pos_x = (self.janela_fornecedor.winfo_screenwidth() // 2) - (largura // 2)
        pos_y = (self.janela_fornecedor.winfo_screenheight() // 2) - (altura // 2)
        
        # Aplicar posição
        self.janela_fornecedor.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")

        self.janela_fornecedor.lift()  # Trazer para frente
        self.janela_fornecedor.focus_force()  # Forçar foco
        self.janela_fornecedor.attributes('-topmost', True)  # Manter no topo temporariamente

        self.janela_fornecedor.after(500, lambda: self.janela_fornecedor.attributes('-topmost', False))
        
        # Configurar formulário
        self.setup_formulario_fornecedor()

        self.janela_fornecedor.after(100, lambda: self.janela_fornecedor.focus_force())

    def editar_fornecedor(self):
        """Abre janela para edição de fornecedor existente"""
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um fornecedor para editar")
            return

        # Buscar dados completos do fornecedor
        fornecedor = self.buscar_fornecedor_completo(
            self.tree_fornecedores.item(selecionado)['values'][0]
        )
        if not fornecedor:
            custom_messagebox("error", "Erro", "Fornecedor não encontrado")
            return

        # Criar janela de edição
        self.janela_fornecedor = tk.Toplevel(self.root)
        self.janela_fornecedor.title("Editar Fornecedor")
        self.setup_formulario_fornecedor(modo_edicao=True)

        try:
            # Determinar tipo de pessoa baseado no tamanho do CNPJ/CPF
            cnpj_cpf = str(fornecedor['cnpj_cpf']).strip()
            tipo_pessoa = 'PJ' if len(cnpj_cpf) > 11 else 'PF'

            # Preencher e configurar campos não editáveis
            # CNPJ/CPF
            self.campos_form['cnpj_cpf'].insert(0, cnpj_cpf.zfill(14 if tipo_pessoa == 'PJ' else 11))
            self.campos_form['cnpj_cpf'].config(state='readonly')
            
            # Tipo Pessoa
            self.campos_form['tipo_pessoa'].set(tipo_pessoa)
            self.campos_form['tipo_pessoa'].config(state='disabled')
            
            # Razão Social
            self.campos_form['razao_social'].insert(0, fornecedor['razao_social'] or '')
            self.campos_form['razao_social'].config(state='readonly')
            
            # Preencher campos editáveis
            self.campos_form['nome'].insert(0, fornecedor['nome'] or '')
            self.campos_form['telefone'].insert(0, fornecedor['telefone'] or '')
            self.campos_form['email'].insert(0, fornecedor['email'] or '')
            self.campos_form['banco'].insert(0, fornecedor['banco'] or '')
            self.campos_form['op'].insert(0, fornecedor['op'] or '')
            self.campos_form['agencia'].insert(0, fornecedor['agencia'] or '')
            self.campos_form['conta'].insert(0, fornecedor['conta'] or '')
            self.campos_form['chave_pix'].insert(0, fornecedor['chave_pix'] or '')
            
            # Categoria (pode ser combobox)
            if isinstance(self.campos_form['categoria'], ttk.Combobox):
                self.campos_form['categoria'].set(fornecedor['categoria'] or '')
            else:
                self.campos_form['categoria'].insert(0, fornecedor['categoria'] or '')
                
            self.campos_form['especificacao'].insert(0, fornecedor['especificacao'] or '')
            self.campos_form['vinculo'].insert(0, fornecedor['vinculo'] or '')
            self.campos_form['endereco'].insert(0, fornecedor['endereco'] or '')

            # Centralizar a janela
            self.janela_fornecedor.update_idletasks()
            width = self.janela_fornecedor.winfo_width()
            height = self.janela_fornecedor.winfo_height()
            x = (self.janela_fornecedor.winfo_screenwidth() // 2) - (width // 2)
            y = (self.janela_fornecedor.winfo_screenheight() // 2) - (height // 2)
            self.janela_fornecedor.geometry('{}x{}+{}+{}'.format(width, height, x, y))
            
            # Tornar a janela modal
            self.janela_fornecedor.transient(self.root)
            self.janela_fornecedor.grab_set()

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar dados do fornecedor: {str(e)}")
            self.janela_fornecedor.destroy()

    def selecionar_fornecedor(self):
        """Seleciona o fornecedor e preenche seus dados"""
        fornecedor = selecionar_fornecedor(
            self.tree_fornecedores, 
            self.campos_fornecedor,
            self.campos_despesa,
            self.notebook,
            self.buscar_fornecedor_completo
        )
        if fornecedor:
            # Formatar CNPJ/CPF
            cnpj_cpf = str(fornecedor[0]).strip()
            self.campos_fornecedor['cnpj_cpf'].config(state='normal')
            self.campos_fornecedor['cnpj_cpf'].delete(0, tk.END)
            self.campos_fornecedor['cnpj_cpf'].insert(0, formatar_cnpj_cpf(cnpj_cpf))
            self.campos_fornecedor['cnpj_cpf'].config(state='readonly')
            
            # Carregar dados completos do fornecedor
            fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
            if fornecedor_completo:
                # Substituir o campo de categoria por Combobox
                self.campos_fornecedor['categoria'] = ttk.Combobox(
                    self.frame_fornecedor,  # Usando o atributo da classe
                    values=get_categorias_fornecedor(),
                    state='readonly',
                    width=30
                )
                self.campos_fornecedor['categoria'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        
                    
                # Definir categoria do fornecedor
                self.campos_fornecedor['categoria'].set(fornecedor_completo['categoria'])
                
                self.campos_fornecedor['dados_bancarios'].config(state='normal')
                self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
                
                # Construir dados bancários
                if fornecedor_completo['chave_pix']:
                    dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
                else:
                    dados_bancarios = (f"{fornecedor_completo['banco'] or ''} "
                                    f"{fornecedor_completo['op'] or ''} - "
                                    f"{fornecedor_completo['agencia'] or ''} "
                                    f"{fornecedor_completo['conta'] or ''}").strip()
                    
                if dados_bancarios.strip() in ['', ' - ']:
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
                
                self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
                self.campos_fornecedor['dados_bancarios'].config(state='readonly')
                
                # NOVO: Preencher campo de referência com a especificação do fornecedor, se disponível
                if fornecedor_completo['especificacao'] and hasattr(self, 'campos_despesa') and 'referencia' in self.campos_despesa:
                    if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                        # Para Combobox, verificamos se o valor está nas opções
                        especificacao = fornecedor_completo['especificacao'].strip()
                        valores = self.campos_despesa['referencia']['values']
                        
                        # Deixamos o campo livre para edição quando não for tipo 1
                        self.campos_despesa['referencia'].config(state='normal')
                        self.campos_despesa['referencia'].delete(0, tk.END)
                        self.campos_despesa['referencia'].insert(0, especificacao)
                    else:
                        # Para Entry normal
                        self.campos_despesa['referencia'].delete(0, tk.END)
                        self.campos_despesa['referencia'].insert(0, fornecedor_completo['especificacao'].strip())
                
                self.notebook.select(2)  # Vai para aba de dados
            
    def buscar_dados_bancarios(self, cnpj_cpf):
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
        
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == cnpj_cpf:
                    print(f"CNPJ/CPF encontrado: {cnpj_cpf}")
                    print(f"Dados da linha: {row}")
                    if row[14]:  # coluna O com dados bancários consolidados
                        return row[14]
                    return ""
            return ""
        except Exception as e:
            print(f"Erro ao buscar dados bancários: {e}")
            return ""

    def buscar_fornecedor_completo(self, cnpj_cpf):
        """Busca todos os dados de um fornecedor - VERSÃO OTIMIZADA"""
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
        
            # Normalizar CNPJ/CPF de entrada
            cnpj_cpf_numeros = ''.join(filter(str.isdigit, str(cnpj_cpf)))
            if len(cnpj_cpf_numeros) <= 11:
                cnpj_cpf_normalizado = cnpj_cpf_numeros.zfill(11)
            else:
                cnpj_cpf_normalizado = cnpj_cpf_numeros.zfill(14)
            
            # Buscar na planilha
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                    
                # Normalizar CNPJ/CPF da planilha
                row_cnpj_numeros = ''.join(filter(str.isdigit, str(row[0])))
                if len(row_cnpj_numeros) <= 11:
                    row_cnpj_normalizado = row_cnpj_numeros.zfill(11)
                else:
                    row_cnpj_normalizado = row_cnpj_numeros.zfill(14)
                    
                if row_cnpj_normalizado == cnpj_cpf_normalizado:
                    fornecedor = {
                        'cnpj_cpf': row[0],
                        'tipo_pessoa': row[1],
                        'razao_social': row[2],
                        'nome': row[3],
                        'telefone': row[4],
                        'email': row[5],
                        'banco': row[6],
                        'op': row[7],
                        'agencia': row[8],
                        'conta': row[9],
                        'chave_pix': row[10],
                        'categoria': row[11],
                        'especificacao': row[12],
                        'vinculo': row[13],
                        'endereco': row[15]
                    }
                    wb.close()
                    return fornecedor
            
            wb.close()
            return None
            
        except Exception as e:
            print(f"Erro ao buscar fornecedor: {e}")
            return None
   
    def setup_formulario_fornecedor(self, modo_edicao=False):
        """Configura o formulário de cadastro/edição de fornecedor com suporte a CPFs criados"""
        formulario = ttk.Frame(self.janela_fornecedor)
        formulario.pack(padx=10, pady=5, fill='both', expand=True)

        # Inicializar gerenciador de CPFs se não existir
        if not hasattr(self, 'gerenciador_cpfs'):
            self.gerenciador_cpfs = GerenciadorCPFsCriados()

        # Campos principais
        campos_principais = ttk.LabelFrame(formulario, text="Dados Principais")
        campos_principais.pack(fill='x', pady=5)

        self.campos_form = {}

        # Frame especial para CNPJ/CPF com botões de CPF criado
        frame_cpf_completo = ttk.Frame(campos_principais)
        frame_cpf_completo.grid(row=0, column=0, columnspan=4, sticky='ew', padx=5, pady=5)

        # Label e campo CNPJ/CPF
        tk.Label(frame_cpf_completo, text="CNPJ/CPF:*").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['cnpj_cpf'] = tk.Entry(frame_cpf_completo, width=20)
        self.campos_form['cnpj_cpf'].grid(row=0, column=1, padx=5, pady=2)
        self.campos_form['cnpj_cpf'].bind('<FocusOut>', self.atualizar_tipo_pessoa)
        
        # Tipo de pessoa
        tk.Label(frame_cpf_completo, text="Tipo:*").grid(row=0, column=2, padx=10, pady=2, sticky='w')
        self.campos_form['tipo_pessoa'] = ttk.Combobox(frame_cpf_completo, 
                                                    values=['PF', 'PJ'],
                                                    state='readonly',
                                                    width=5)
        self.campos_form['tipo_pessoa'].grid(row=0, column=3, padx=5, pady=2)

        # Frame para botões de CPF criado
        frame_botoes_cpf = ttk.Frame(frame_cpf_completo)
        frame_botoes_cpf.grid(row=1, column=0, columnspan=4, pady=5, sticky='ew')

        # Botão para usar CPF criado automaticamente
        btn_cpf_auto = ttk.Button(frame_botoes_cpf, 
                                text="🔄 Obter CPF Criado", 
                                command=self.usar_cpf_criado_auto,
                                width=18)
        btn_cpf_auto.pack(side='left', padx=5)

        # Botão para escolher CPF da lista
        btn_cpf_lista = ttk.Button(frame_botoes_cpf, 
                                text="📋 Escolher da Lista", 
                                command=self.mostrar_cpfs_disponiveis,
                                width=18)
        btn_cpf_lista.pack(side='left', padx=5)

        # Label informativo
        lbl_info = tk.Label(frame_botoes_cpf, 
                        text="💡 Use estes botões para prestadores sem CPF próprio",
                        font=('Arial', 8),
                        fg='gray')
        lbl_info.pack(side='left', padx=20)

        # Razão Social e Nome
        tk.Label(campos_principais, text="Razão Social:*").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['razao_social'] = tk.Entry(campos_principais, width=50)
        self.campos_form['razao_social'].grid(row=1, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        self.campos_form['razao_social'].bind('<FocusOut>', self.copiar_para_nome)

        tk.Label(campos_principais, text="Nome Fantasia:*").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['nome'] = tk.Entry(campos_principais, width=50)
        self.campos_form['nome'].grid(row=2, column=1, columnspan=3, padx=5, pady=2, sticky='ew')

        # Contatos - Frame especial para telefone com botão PIX
        campos_contato = ttk.LabelFrame(formulario, text="Contato")
        campos_contato.pack(fill='x', pady=5)

        # Frame para telefone com botão PIX
        frame_telefone = ttk.Frame(campos_contato)
        frame_telefone.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=2)

        tk.Label(frame_telefone, text="Telefone:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['telefone'] = tk.Entry(frame_telefone, width=20)
        self.campos_form['telefone'].grid(row=0, column=1, padx=5, pady=2)

        # Botão para usar telefone como PIX
        btn_tel_pix = ttk.Button(frame_telefone, 
                                text="📱 Usar como PIX", 
                                command=self.usar_telefone_como_pix,
                                width=15)
        btn_tel_pix.grid(row=0, column=2, padx=10, pady=2)

        # Email
        tk.Label(campos_contato, text="Email:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['email'] = tk.Entry(campos_contato, width=50)
        self.campos_form['email'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Dados Bancários
        campos_bancarios = ttk.LabelFrame(formulario, text="Dados Bancários")
        campos_bancarios.pack(fill='x', pady=5)

        # Carregar configurações
        try:
            carregar_configuracoes()  
            lista_bancos = get_bancos()
        except Exception as e:
            print(f"Erro ao carregar bancos: {str(e)}")
            lista_bancos = []

        tk.Label(campos_bancarios, text="Banco:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['banco'] = ttk.Combobox(
            campos_bancarios,
            values=lista_bancos,
            state='readonly'
        )
        self.campos_form['banco'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Operação:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['op'] = tk.Entry(campos_bancarios)
        self.campos_form['op'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Agência:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['agencia'] = tk.Entry(campos_bancarios)
        self.campos_form['agencia'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Conta:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['conta'] = tk.Entry(campos_bancarios)
        self.campos_form['conta'].grid(row=3, column=1, padx=5, pady=2, sticky='ew')

        # PIX
        campos_pix = ttk.LabelFrame(formulario, text="Chave PIX")
        campos_pix.pack(fill='x', pady=5)

        # Tipo de chave PIX
        ttk.Label(campos_pix, text="Tipo de Chave:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.tipo_pix = ttk.Combobox(
            campos_pix, 
            values=['Selecione', 'CNPJ/CPF', 'Telefone', 'Email'],
            state='readonly'
        )
        self.tipo_pix.grid(row=0, column=1, padx=5, pady=2)
        self.tipo_pix.set('Telefone')  # Padrão para prestadores

        ttk.Label(campos_pix, text="Chave:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['chave_pix'] = ttk.Entry(campos_pix, width=40)
        self.campos_form['chave_pix'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Binding para atualização automática
        self.tipo_pix.bind('<<ComboboxSelected>>', self.atualizar_chave_pix)

        # Classificação
        campos_class = ttk.LabelFrame(formulario, text="Classificação")
        campos_class.pack(fill='x', pady=5)

        # Carregar categorias
        try:
            categorias = get_categorias_fornecedor()
        except Exception as e:
            print(f"Erro ao carregar categorias: {str(e)}")
            categorias = ['ADM', 'DIV', 'LOC', 'MAT', 'MO', 'SERV', 'TP']

        tk.Label(campos_class, text="Categoria:*").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['categoria'] = ttk.Combobox(campos_class, 
                                                    values=categorias,
                                                    state='readonly')
        self.campos_form['categoria'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_class, text="Especificação:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['especificacao'] = tk.Entry(campos_class, width=40)
        self.campos_form['especificacao'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_class, text="Vínculo:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['vinculo'] = tk.Entry(campos_class, width=40)
        self.campos_form['vinculo'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_class, text="Endereço:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.campos_form['endereco'] = tk.Entry(campos_class, width=80)
        self.campos_form['endereco'].grid(row=3, column=1, padx=5, pady=2, sticky='ew')

        # Botões de ação
        frame_botoes = ttk.Frame(formulario)
        frame_botoes.pack(fill='x', pady=10)

        ttk.Button(frame_botoes, 
                text="Salvar", 
                command=self.salvar_fornecedor_com_cpf_criado).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                text="Cancelar", 
                command=self.janela_fornecedor.destroy).pack(side='left', padx=5)

    def usar_cpf_criado_auto(self):
        """Busca e usa automaticamente o próximo CPF criado disponível - VERSÃO MELHORADA"""
        try:
            if not hasattr(self, 'gerenciador_cpfs'):
                self.gerenciador_cpfs = GerenciadorCPFsCriados()
            
            print("Iniciando busca por CPF disponível...")
            cpf_disponivel, linha = self.gerenciador_cpfs.obter_proximo_cpf_disponivel()
            
            if cpf_disponivel:
                print(f"CPF obtido: {cpf_disponivel}")
                
                # TESTE: Validar o CPF antes de usar
                if not self.gerenciador_cpfs.validar_cpf_gerado(cpf_disponivel):
                    custom_messagebox("error", "Erro", 
                                    f"❌ CPF gerado é inválido: {cpf_disponivel}\n\n"
                                    f"Verifique o algoritmo de geração")
                    return
                
                # Formatar CPF
                cpf_formatado = f"{cpf_disponivel[:3]}.{cpf_disponivel[3:6]}.{cpf_disponivel[6:9]}-{cpf_disponivel[9:]}"
                print(f"CPF formatado: {cpf_formatado}")
                
                # TESTE: Validar CPF formatado com a função do sistema
                try:
                    if not validar_cnpj_cpf(cpf_formatado):
                        print(f"AVISO: Sistema não reconheceu CPF como válido: {cpf_formatado}")
                        # Mesmo assim, continuar - pode ser problema na função validar_cnpj_cpf
                except Exception as e:
                    print(f"Erro na validação do sistema: {str(e)}")
                
                # Preencher campo CNPJ/CPF
                self.campos_form['cnpj_cpf'].delete(0, tk.END)
                self.campos_form['cnpj_cpf'].insert(0, cpf_formatado)
                
                # Definir como PF
                self.campos_form['tipo_pessoa'].set('PF')
                
                # Configurar PIX automaticamente
                self.tipo_pix.set('CNPJ/CPF')
                self.campos_form['chave_pix'].delete(0, tk.END)
                self.campos_form['chave_pix'].insert(0, cpf_formatado)
                
                # Focar no campo nome para continuar o cadastro
                self.campos_form['razao_social'].focus()
                
                custom_messagebox("info", "CPF Obtido", 
                                f"✅ CPF criado obtido com sucesso!\n\n"
                                f"📋 CPF: {cpf_formatado}\n"
                                f"🔑 Configurado como chave PIX\n"
                                f"🔍 Validação: CPF matematicamente correto\n\n"
                                f"➡️ Continue preenchendo os dados do prestador")
            else:
                custom_messagebox("error", "Erro", 
                                "❌ Não foi possível obter um CPF criado.\n\n"
                                "Possíveis causas:\n"
                                "• Arquivo Base_Fornecedores.xlsx não encontrado\n"
                                "• Erro na geração de CPFs válidos\n"
                                "• Problema de permissão no arquivo")
                
        except Exception as e:
            print(f"Erro detalhado: {str(e)}")
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"❌ Erro ao obter CPF criado:\n{str(e)}")

    def mostrar_cpfs_disponiveis(self):
        """Mostra lista de CPFs disponíveis para seleção manual"""
        try:
            if not hasattr(self, 'gerenciador_cpfs'):
                self.gerenciador_cpfs = GerenciadorCPFsCriados()
                
            cpfs_disponiveis = self.gerenciador_cpfs.listar_cpfs_disponiveis()
            
            if not cpfs_disponiveis:
                custom_messagebox("info", "CPFs Disponíveis", 
                                "❌ Nenhum CPF disponível no momento.\n\n"
                                "Use o botão 'Obter CPF Criado' para gerar novos CPFs automaticamente.")
                return
            
            # Criar janela de seleção
            janela_cpfs = tk.Toplevel(self.janela_fornecedor)
            janela_cpfs.title("📋 CPFs Criados - Escolher da Lista")
            janela_cpfs.geometry("450x600")
            janela_cpfs.transient(self.janela_fornecedor)
            janela_cpfs.grab_set()
            
            # Centralizar janela
            janela_cpfs.update_idletasks()
            x = (janela_cpfs.winfo_screenwidth() // 2) - (225)
            y = (janela_cpfs.winfo_screenheight() // 2) - (300)
            janela_cpfs.geometry(f"450x600+{x}+{y}")
            
            frame = ttk.Frame(janela_cpfs, padding="15")
            frame.pack(fill='both', expand=True)
            
            # Cabeçalho
            header_frame = ttk.Frame(frame)
            header_frame.pack(fill='x', pady=(0, 15))
            
            ttk.Label(header_frame, text="📋 CPFs Criados Disponíveis", 
                    font=('Arial', 14, 'bold')).pack()
            ttk.Label(header_frame, text=f"Total disponível: {len(cpfs_disponiveis)}", 
                    font=('Arial', 10)).pack()
            
            # Lista de CPFs
            lista_frame = ttk.LabelFrame(frame, text="Selecione um CPF:")
            lista_frame.pack(fill='both', expand=True, pady=(0, 15))
            
            # Frame interno para Treeview e scrollbar
            tree_frame = ttk.Frame(lista_frame)
            tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Treeview para CPFs
            tree_cpfs = ttk.Treeview(tree_frame, columns=('CPF', 'Status'), show='headings', height=15)
            tree_cpfs.heading('CPF', text='CPF Disponível')
            tree_cpfs.heading('Status', text='Status')
            tree_cpfs.column('CPF', width=200, anchor='center')
            tree_cpfs.column('Status', width=100, anchor='center')
            
            # Scrollbar
            scroll_cpfs = ttk.Scrollbar(tree_frame, orient='vertical', command=tree_cpfs.yview)
            tree_cpfs.configure(yscrollcommand=scroll_cpfs.set)
            
            tree_cpfs.pack(side='left', fill='both', expand=True)
            scroll_cpfs.pack(side='right', fill='y')
            
            # Preencher lista
            for i, cpf in enumerate(cpfs_disponiveis, 1):
                cpf_formatado = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
                tree_cpfs.insert('', 'end', values=(cpf_formatado, '✅ Disponível'), tags=(cpf,))
            
            def selecionar_cpf():
                selecionado = tree_cpfs.selection()
                if not selecionado:
                    custom_messagebox("warning", "Aviso", "⚠️ Selecione um CPF da lista")
                    return
                
                # Obter CPF selecionado (não formatado)
                cpf_selecionado = tree_cpfs.item(selecionado[0])['tags'][0]
                cpf_formatado = tree_cpfs.item(selecionado[0])['values'][0]
                
                # Preencher no formulário principal
                self.campos_form['cnpj_cpf'].delete(0, tk.END)
                self.campos_form['cnpj_cpf'].insert(0, cpf_formatado)
                
                # Configurar tipo como PF
                self.campos_form['tipo_pessoa'].set('PF')
                
                # Configurar PIX
                self.tipo_pix.set('CNPJ/CPF')
                self.campos_form['chave_pix'].delete(0, tk.END)
                self.campos_form['chave_pix'].insert(0, cpf_formatado)
                
                # Fechar janela
                janela_cpfs.destroy()
                
                # Focar no campo nome
                self.campos_form['razao_social'].focus()
                
                custom_messagebox("info", "CPF Selecionado", 
                                f"✅ CPF selecionado com sucesso!\n\n"
                                f"📋 {cpf_formatado}\n\n"
                                f"➡️ Continue preenchendo os dados")
            
            # Botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x')
            
            ttk.Button(frame_botoes, text="✅ Selecionar", 
                    command=selecionar_cpf).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="❌ Cancelar", 
                    command=janela_cpfs.destroy).pack(side='left', padx=5)
            
            # Info adicional
            info_label = tk.Label(frame_botoes, 
                                text="💡 Dica: Dê duplo clique no CPF para selecioná-lo rapidamente",
                                font=('Arial', 8), fg='gray')
            info_label.pack(side='right', padx=10)
            
            # Double-click para selecionar rapidamente
            tree_cpfs.bind('<Double-1>', lambda e: selecionar_cpf())
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"❌ Erro ao mostrar CPFs:\n{str(e)}")

    def usar_telefone_como_pix(self):
        """Usa o telefone digitado como chave PIX"""
        telefone = self.campos_form['telefone'].get().strip()
        
        if not telefone:
            custom_messagebox("warning", "Aviso", "⚠️ Digite o telefone primeiro!")
            self.campos_form['telefone'].focus()
            return
        
        # Limpar telefone (remover caracteres especiais)
        telefone_limpo = ''.join(filter(str.isdigit, telefone))
        
        # Verificar se tem pelo menos 10 dígitos
        if len(telefone_limpo) < 10:
            custom_messagebox("warning", "Aviso", "⚠️ Telefone deve ter pelo menos 10 dígitos!")
            self.campos_form['telefone'].focus()
            return
        
        # Formatar telefone para PIX
        if len(telefone_limpo) == 10:
            # Fixo: (XX) XXXX-XXXX
            tel_formatado = f"({telefone_limpo[:2]}) {telefone_limpo[2:6]}-{telefone_limpo[6:]}"
        elif len(telefone_limpo) == 11:
            # Celular: (XX) 9XXXX-XXXX
            tel_formatado = f"({telefone_limpo[:2]}) {telefone_limpo[2:7]}-{telefone_limpo[7:]}"
        else:
            # Manter como estava se tiver mais de 11 dígitos
            tel_formatado = telefone_limpo
        
        # Atualizar campo telefone com formatação
        self.campos_form['telefone'].delete(0, tk.END)
        self.campos_form['telefone'].insert(0, tel_formatado)
        
        # Definir tipo PIX como Telefone
        self.tipo_pix.set('Telefone')
        
        # Preencher chave PIX
        self.campos_form['chave_pix'].delete(0, tk.END)
        self.campos_form['chave_pix'].insert(0, tel_formatado)
        
        custom_messagebox("info", "PIX Configurado", 
                        f"✅ Telefone configurado como chave PIX!\n\n"
                        f"📱 {tel_formatado}\n\n"
                        f"➡️ Continue preenchendo os outros dados")

    def atualizar_chave_pix(self, event=None):
        """Atualiza o campo de chave PIX baseado no tipo selecionado"""
        tipo_selecionado = self.tipo_pix.get()
        self.campos_form['chave_pix'].delete(0, tk.END)
        
        if tipo_selecionado == 'CNPJ/CPF':
            self.campos_form['chave_pix'].insert(0, self.campos_form['cnpj_cpf'].get())
        elif tipo_selecionado == 'Telefone':
            self.campos_form['chave_pix'].insert(0, self.campos_form['telefone'].get())
        elif tipo_selecionado == 'Email':
            self.campos_form['chave_pix'].insert(0, self.campos_form['email'].get())

    def atualizar_tipo_pessoa(self, event=None):
        """Determina automaticamente o tipo de pessoa baseado no CNPJ/CPF"""
        cnpj_cpf = self.campos_form['cnpj_cpf'].get().strip()
        # Remove caracteres não numéricos
        cnpj_cpf = ''.join(filter(str.isdigit, cnpj_cpf))
        
        if len(cnpj_cpf) <= 11:
            self.campos_form['tipo_pessoa'].set('PF')
        else:
            self.campos_form['tipo_pessoa'].set('PJ')

    def copiar_para_nome(self, event=None):
        """Copia a razão social para o nome se este estiver vazio"""
        razao_social = self.campos_form['razao_social'].get().strip()
        nome_atual = self.campos_form['nome'].get().strip()
        
        if razao_social and not nome_atual:
            self.campos_form['nome'].insert(0, razao_social)

    def salvar_fornecedor_com_cpf_criado(self):
        """Salva fornecedor e marca CPF criado como usado - VERSÃO CORRIGIDA"""
        # Validar campos obrigatórios
        campos_obrigatorios = ['tipo_pessoa', 'cnpj_cpf', 'razao_social', 'nome', 'categoria']
        for campo in campos_obrigatorios:
            if not self.campos_form[campo].get().strip():
                custom_messagebox("error", "Erro", f"❌ O campo {campo} é obrigatório!")
                return

        # Validar CNPJ/CPF
        tipo_pessoa = self.campos_form['tipo_pessoa'].get()
        cnpj_cpf_original = self.campos_form['cnpj_cpf'].get().strip()
        
        # Limpar CNPJ/CPF mantendo apenas números
        cnpj_cpf_numeros = ''.join(filter(str.isdigit, cnpj_cpf_original))
        
        # Validar com números limpos
        if not self.validar_cnpj_cpf_numeros(cnpj_cpf_numeros):
            custom_messagebox("error", "Erro", f"❌ {'CPF' if tipo_pessoa == 'PF' else 'CNPJ'} inválido!")
            return
        
        # Verificar se é um CPF criado
        eh_cpf_criado = False
        if tipo_pessoa == 'PF' and len(cnpj_cpf_numeros) == 11:
            try:
                if not hasattr(self, 'gerenciador_cpfs'):
                    self.gerenciador_cpfs = GerenciadorCPFsCriados()
                cpfs_disponiveis = self.gerenciador_cpfs.listar_cpfs_disponiveis()
                if cnpj_cpf_numeros in cpfs_disponiveis:
                    eh_cpf_criado = True
            except Exception as e:
                print(f"Erro ao verificar CPF criado: {str(e)}")

        # Montar dados bancários
        if self.campos_form['chave_pix'].get():
            dados_bancarios = f"PIX: {self.campos_form['chave_pix'].get()}"
        else:
            dados_bancarios = (f"{self.campos_form['banco'].get()} "
                            f"{self.campos_form['op'].get()} - "
                            f"{self.campos_form['agencia'].get()} "
                            f"{self.campos_form['conta'].get()}").strip()

        # Preparar dados garantindo que tudo seja string
        dados = {
            'tipo_pessoa': str(tipo_pessoa),
            'cnpj_cpf': str(cnpj_cpf_numeros),  # Salvar apenas números
            'razao_social': str(self.campos_form['razao_social'].get().upper()),
            'nome': str(self.campos_form['nome'].get().upper()),
            'telefone': str(self.campos_form['telefone'].get()),
            'email': str(self.campos_form['email'].get()),
            'banco': str(self.campos_form['banco'].get()),
            'op': str(self.campos_form['op'].get()),
            'agencia': str(self.campos_form['agencia'].get()),
            'conta': str(self.campos_form['conta'].get()),
            'chave_pix': str(self.campos_form['chave_pix'].get()),
            'categoria': str(self.campos_form['categoria'].get().upper()),
            'especificacao': str(self.campos_form['especificacao'].get().upper()),
            'vinculo': str(self.campos_form['vinculo'].get().upper()),
            'dados_bancarios': str(dados_bancarios),
            'endereco': str(self.campos_form['endereco'].get().upper())
        }

        try:
            # Salvar na base de fornecedores
            self.salvar_na_base_fornecedores(dados)
            
            # Se for CPF criado, marcar como usado
            if eh_cpf_criado:
                nome_fornecedor = self.campos_form['nome'].get().upper()
                sucesso_marcacao = self.gerenciador_cpfs.marcar_cpf_como_usado(cnpj_cpf_numeros, nome_fornecedor)
                
                if sucesso_marcacao:
                    cnpj_cpf_formatado = formatar_cnpj_cpf(cnpj_cpf_numeros)
                    mensagem_sucesso = (f"✅ Fornecedor salvo com sucesso!\n\n"
                                    f"🔄 CPF criado marcado como usado:\n"
                                    f"📋 {cnpj_cpf_formatado}\n"
                                    f"👤 {nome_fornecedor}")
                else:
                    mensagem_sucesso = (f"✅ Fornecedor salvo com sucesso!\n\n"
                                    f"⚠️ Aviso: Não foi possível marcar o CPF como usado")
            else:
                mensagem_sucesso = f"✅ Fornecedor salvo com sucesso!"
                
            custom_messagebox("info", "Sucesso", mensagem_sucesso)
            self.janela_fornecedor.destroy()
            self.buscar_fornecedor()  # Atualizar lista
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"❌ Erro ao salvar fornecedor:\n{str(e)}")

    def gerenciar_cpfs_criados(self):
        """Abre interface para gerenciar fornecedores com CPFs criados - NOVA FUNCIONALIDADE"""
        try:
            # Verificar se o arquivo de fornecedores existe
            if not os.path.exists(ARQUIVO_FORNECEDORES):
                custom_messagebox("error", "Erro", "Arquivo de fornecedores não encontrado!")
                return
            
            # Criar janela de gerenciamento
            janela_gerencia = tk.Toplevel(self.root)
            janela_gerencia.title("Gerenciar Fornecedores com CPF Criado")
            janela_gerencia.geometry("900x700")
            janela_gerencia.transient(self.root)
            janela_gerencia.grab_set()
            
            # Centralizar janela
            janela_gerencia.update_idletasks()
            x = (janela_gerencia.winfo_screenwidth() // 2) - (450)
            y = (janela_gerencia.winfo_screenheight() // 2) - (350)
            janela_gerencia.geometry(f"900x800+{x}+{y}")
            
            main_frame = ttk.Frame(janela_gerencia, padding="15")
            main_frame.pack(fill='both', expand=True)
            
            # Cabeçalho
            header_frame = ttk.Frame(main_frame)
            header_frame.pack(fill='x', pady=(0, 15))
            
            ttk.Label(header_frame, text="Gerenciar Fornecedores com CPF Criado", 
                    font=('Arial', 16, 'bold')).pack()
            ttk.Label(header_frame, text="Exclua registros temporários quando obtiver os dados reais", 
                    font=('Arial', 10), foreground='gray').pack()
            
            # Frame de controles
            controls_frame = ttk.Frame(main_frame)
            controls_frame.pack(fill='x', pady=(0, 10))
            
            # Botão para atualizar lista
            ttk.Button(controls_frame, text="Atualizar Lista", 
                    command=lambda: self.carregar_fornecedores_cpf_criado(tree_fornecedores)).pack(side='left', padx=5)
            
            # Botão para excluir selecionados
            ttk.Button(controls_frame, text="Excluir Selecionados", 
                    command=lambda: self.excluir_fornecedores_selecionados(tree_fornecedores)).pack(side='left', padx=5)
            
            # Label para contagem
            self.lbl_contagem = ttk.Label(controls_frame, text="", font=('Arial', 10))
            self.lbl_contagem.pack(side='right', padx=10)
            
            # Frame para lista de fornecedores
            lista_frame = ttk.LabelFrame(main_frame, text="Fornecedores com CPF Criado:")
            lista_frame.pack(fill='both', expand=True, pady=(0, 10))
            
            # Frame interno para Treeview e scrollbars
            tree_frame = ttk.Frame(lista_frame)
            tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Treeview para fornecedores
            tree_fornecedores = ttk.Treeview(tree_frame, 
                                            columns=('CPF', 'Nome', 'Categoria', 'Especificacao', 'Status'), 
                                            show='headings', 
                                            height=20,
                                            selectmode='extended')  # Permite seleção múltipla
            
            # Configurar colunas
            tree_fornecedores.heading('CPF', text='CPF Criado')
            tree_fornecedores.heading('Nome', text='Nome/Razão Social')
            tree_fornecedores.heading('Categoria', text='Categoria')
            tree_fornecedores.heading('Especificacao', text='Especificação')
            tree_fornecedores.heading('Status', text='Status')
            
            tree_fornecedores.column('CPF', width=150, anchor='center')
            tree_fornecedores.column('Nome', width=250)
            tree_fornecedores.column('Categoria', width=100, anchor='center')
            tree_fornecedores.column('Especificacao', width=150)
            tree_fornecedores.column('Status', width=120, anchor='center')
            
            # Scrollbars
            scroll_y = ttk.Scrollbar(tree_frame, orient='vertical', command=tree_fornecedores.yview)
            scroll_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree_fornecedores.xview)
            tree_fornecedores.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            
            tree_fornecedores.pack(side='left', fill='both', expand=True)
            scroll_y.pack(side='right', fill='y')
            scroll_x.pack(side='bottom', fill='x')
            
            # Frame de informações e ações
            info_frame = ttk.LabelFrame(main_frame, text="Informações:")
            info_frame.pack(fill='x', pady=(0, 15))
            
            info_text = tk.Text(info_frame, height=4, wrap='word', bg='#f0f0f0', 
                            font=('Arial', 9), state='normal')
            info_text.pack(fill='x', padx=10, pady=10)
            
            info_text.insert('1.0', 
                "• Selecione um ou múltiplos fornecedores para exclusão (Ctrl+Click para múltipla seleção)\n"
                "• CPFs criados são identificados automaticamente pela base de CPFs gerados\n"
                "• A exclusão é permanente - certifique-se antes de confirmar\n"
                "• Após excluir, o CPF volta a ficar disponível para reutilização")
            info_text.config(state='disabled')
            
            # Botões de ação
            buttons_frame = ttk.Frame(main_frame)
            buttons_frame.pack(fill='x')
            
            ttk.Button(buttons_frame, text="Fechar", 
                    command=janela_gerencia.destroy).pack(side='right', padx=5)
            
            ttk.Button(buttons_frame, text="Exportar Lista", 
                    command=lambda: self.exportar_lista_cpf_criado(tree_fornecedores)).pack(side='left', padx=5)
            
            # Carregar dados iniciais
            self.carregar_fornecedores_cpf_criado(tree_fornecedores)
            
            # Armazenar referência da tree para uso em outras funções
            self.tree_gerencia_cpf = tree_fornecedores
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir gerenciador: {str(e)}")

    def carregar_fornecedores_cpf_criado(self, tree_widget):
        """Carrega fornecedores que usam CPFs criados da aba CPF"""
        try:
            # Limpar tree
            for item in tree_widget.get_children():
                tree_widget.delete(item)
            
            # Inicializar gerenciador de CPFs se não existir
            if not hasattr(self, 'gerenciador_cpfs'):
                self.gerenciador_cpfs = GerenciadorCPFsCriados()
            
            # Obter lista de CPFs criados da aba CPF
            cpfs_criados_total = self.gerenciador_cpfs.listar_todos_cpfs_criados()
            
            if not cpfs_criados_total:
                self.lbl_contagem.config(text="Nenhum CPF criado encontrado na aba CPF")
                tree_widget.insert('', 'end', values=(
                    '', 'Nenhum CPF criado encontrado na aba CPF', '', '', ''
                ))
                return
            
            # Carregar fornecedores da aba Fornecedores
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
            
            fornecedores_encontrados = []
            cpfs_usados = self.gerenciador_cpfs.listar_cpfs_usados()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular linhas vazias
                    continue
                
                cnpj_cpf_row = str(row[0]).strip()
                # Remover formatação para comparação
                cnpj_cpf_numeros = ''.join(filter(str.isdigit, cnpj_cpf_row))
                
                # Verificar se é um CPF criado
                if cnpj_cpf_numeros in cpfs_criados_total:
                    nome = str(row[3] or '').strip()  # Coluna D = Nome
                    categoria = str(row[11] or '').strip()  # Coluna L = Categoria
                    especificacao = str(row[12] or '').strip()  # Coluna M = Especificação
                    
                    # Verificar status na aba CPF
                    status = "Em Uso" if cnpj_cpf_numeros in cpfs_usados else "Disponível"
                    
                    # Obter detalhes do CPF da aba CPF
                    detalhes = self.gerenciador_cpfs.obter_detalhes_cpf_usado(cnpj_cpf_numeros)
                    if detalhes:
                        if detalhes['status'].upper() == 'USADO':
                            status = f"Em Uso ({detalhes['data_uso']})"
                        elif detalhes['status'].upper() == 'DISPONIVEL':
                            status = "Cadastrado (não deveria existir)"
                        elif detalhes['status'].upper() == 'INVALIDO':
                            status = "CPF Inválido"
                    
                    # Formatar CPF para exibição
                    cpf_formatado = f"{cnpj_cpf_numeros[:3]}.{cnpj_cpf_numeros[3:6]}.{cnpj_cpf_numeros[6:9]}-{cnpj_cpf_numeros[9:]}"
                    
                    # Inserir na tree
                    item_id = tree_widget.insert('', 'end', values=(
                        cpf_formatado,
                        nome,
                        categoria,
                        especificacao,
                        status
                    ), tags=(cnpj_cpf_numeros,))  # Armazenar CPF sem formatação nas tags
                    
                    fornecedores_encontrados.append({
                        'cpf': cnpj_cpf_numeros,
                        'nome': nome,
                        'categoria': categoria,
                        'especificacao': especificacao,
                        'status': status
                    })
            
            wb.close()
            
            # Atualizar contador
            total = len(fornecedores_encontrados)
            em_uso = len([f for f in fornecedores_encontrados if 'Em Uso' in f['status']])
            inconsistentes = len([f for f in fornecedores_encontrados if 'não deveria existir' in f['status']])
            
            contador_texto = f"Total: {total} | Em uso: {em_uso}"
            if inconsistentes > 0:
                contador_texto += f" | Inconsistentes: {inconsistentes}"
            
            self.lbl_contagem.config(text=contador_texto)
            
            if total == 0:
                # Inserir mensagem informativa
                tree_widget.insert('', 'end', values=(
                    '', 'Nenhum fornecedor com CPF criado encontrado', '', '', ''
                ))
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar fornecedores: {str(e)}")
            print(f"Erro detalhado: {str(e)}")
            import traceback
            traceback.print_exc()

    def excluir_fornecedores_selecionados(self, tree_widget):
        """Exclui os fornecedores selecionados da base"""
        try:
            selecionados = tree_widget.selection()
            
            if not selecionados:
                custom_messagebox("warning", "Aviso", "Selecione pelo menos um fornecedor para excluir!")
                return
            
            # Coletar dados dos selecionados
            fornecedores_para_excluir = []
            for item in selecionados:
                values = tree_widget.item(item)['values']
                if values[1] == 'Nenhum fornecedor com CPF criado encontrado':
                    continue
                
                cpf_sem_formatacao = tree_widget.item(item)['tags'][0]
                fornecedores_para_excluir.append({
                    'cpf': cpf_sem_formatacao,
                    'cpf_formatado': values[0],
                    'nome': values[1],
                    'categoria': values[2],
                    'especificacao': values[3],
                    'status': values[4]
                })
            
            if not fornecedores_para_excluir:
                custom_messagebox("warning", "Aviso", "Nenhum fornecedor válido selecionado!")
                return
            
            # Confirmar exclusão
            total = len(fornecedores_para_excluir)
            lista_nomes = '\n'.join([f"• {f['cpf_formatado']} - {f['nome']}" for f in fornecedores_para_excluir[:5]])
            if total > 5:
                lista_nomes += f"\n... e mais {total - 5} fornecedores"
            
            resposta = custom_messagebox("question", "Confirmar Exclusão", 
                                    f"Deseja realmente excluir {total} fornecedor(es)?\n\n"
                                    f"{lista_nomes}\n\n"
                                    f"ATENÇÃO: Esta ação é permanente e não pode ser desfeita!")
            
            if resposta != 'yes':
                return
            
            # Executar exclusão
            sucesso, falhas = self.executar_exclusao_fornecedores(fornecedores_para_excluir)
            
            # Mostrar resultado
            if falhas:
                mensagem = f"Exclusão concluída com problemas:\n\n"
                mensagem += f"✓ Excluídos com sucesso: {sucesso}\n"
                mensagem += f"✗ Falhas: {len(falhas)}\n\n"
                mensagem += "Fornecedores com falha:\n"
                for falha in falhas[:3]:
                    mensagem += f"• {falha['nome']} - {falha['erro']}\n"
                if len(falhas) > 3:
                    mensagem += f"... e mais {len(falhas) - 3} falhas"
                
                custom_messagebox("warning", "Exclusão com Problemas", mensagem)
            else:
                custom_messagebox("info", "Sucesso", 
                                f"✓ {sucesso} fornecedor(es) excluído(s) com sucesso!\n\n"
                                f"Os CPFs voltaram a ficar disponíveis para reutilização.")
            
            # Recarregar lista
            self.carregar_fornecedores_cpf_criado(tree_widget)
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro durante exclusão: {str(e)}")

    def executar_exclusao_fornecedores(self, fornecedores_para_excluir):
        """Executa a exclusão dos fornecedores da base e atualiza CPFs criados"""
        sucesso = 0
        falhas = []
        
        try:
            # Carregar planilha de fornecedores
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Coletar todas as linhas que NÃO devem ser excluídas
            linhas_manter = []
            cpfs_excluir = [f['cpf'] for f in fornecedores_para_excluir]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Linha vazia
                    continue
                
                cnpj_cpf_row = str(row[0]).strip()
                cnpj_cpf_numeros = ''.join(filter(str.isdigit, cnpj_cpf_row))
                
                # Se não está na lista de exclusão, manter
                if cnpj_cpf_numeros not in cpfs_excluir:
                    linhas_manter.append(row)
            
            # Limpar planilha (manter apenas cabeçalho)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            
            # Reescrever apenas as linhas que devem ser mantidas
            for i, row_data in enumerate(linhas_manter, start=2):
                for j, value in enumerate(row_data, start=1):
                    if j <= 15:  # Limitar às colunas existentes
                        ws.cell(row=i, column=j, value=value)
            
            # Salvar alterações na base de fornecedores
            wb.save(ARQUIVO_FORNECEDORES)
            
            # Marcar CPFs como disponíveis novamente
            if not hasattr(self, 'gerenciador_cpfs'):
                self.gerenciador_cpfs = GerenciadorCPFsCriados()
            
            for fornecedor in fornecedores_para_excluir:
                try:
                    # Tentar marcar CPF como disponível novamente
                    self.gerenciador_cpfs.marcar_cpf_como_disponivel(fornecedor['cpf'])
                    sucesso += 1
                except Exception as e:
                    falhas.append({
                        'nome': fornecedor['nome'],
                        'cpf': fornecedor['cpf'],
                        'erro': str(e)
                    })
            
        except Exception as e:
            # Se der erro geral, considerar falha para todos
            for fornecedor in fornecedores_para_excluir:
                falhas.append({
                    'nome': fornecedor['nome'],
                    'cpf': fornecedor['cpf'],
                    'erro': f"Erro geral: {str(e)}"
                })
        
        return sucesso, falhas

    def exportar_lista_cpf_criado(self, tree_widget):
        """Exporta a lista de fornecedores com CPF criado para Excel"""
        try:
            from tkinter import filedialog
            import pandas as pd
            from datetime import datetime
            
            # Verificar se há dados
            if not tree_widget.get_children():
                custom_messagebox("warning", "Aviso", "Não há dados para exportar!")
                return
            
            # Coletar dados da tree
            dados = []
            for item in tree_widget.get_children():
                values = tree_widget.item(item)['values']
                if values[1] != 'Nenhum fornecedor com CPF criado encontrado':
                    dados.append({
                        'CPF_Criado': values[0],
                        'Nome_Razao_Social': values[1],
                        'Categoria': values[2],
                        'Especificacao': values[3],
                        'Status': values[4]
                    })
            
            if not dados:
                custom_messagebox("warning", "Aviso", "Não há dados válidos para exportar!")
                return
            
            # Escolher local para salvar
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo = filedialog.asksaveasfilename(
                title="Salvar lista de fornecedores com CPF criado",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialvalue=f"fornecedores_cpf_criado_{timestamp}.xlsx"
            )
            
            if not arquivo:
                return
            
            # Criar DataFrame e salvar
            df = pd.DataFrame(dados)
            
            with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fornecedores_CPF_Criado', index=False)
                
                # Ajustar largura das colunas
                worksheet = writer.sheets['Fornecedores_CPF_Criado']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            custom_messagebox("info", "Sucesso", 
                            f"Lista exportada com sucesso!\n\n"
                            f"Arquivo: {arquivo}\n"
                            f"Total de registros: {len(dados)}")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao exportar: {str(e)}")

    def validar_cnpj_cpf_numeros(self, numeros):
        """Valida CNPJ ou CPF usando apenas números"""
        if not numeros or not numeros.isdigit():
            return False
        
        if len(numeros) == 11:
            return self.validar_cpf_algoritmo(numeros)
        elif len(numeros) == 14:
            return self.validar_cnpj_algoritmo(numeros)
        else:
            return False

    def validar_cpf_algoritmo(self, cpf):
        """Valida CPF usando algoritmo oficial"""
        if cpf == cpf[0] * 11:
            return False
        
        # Calcular primeiro dígito verificador
        soma = 0
        for i in range(9):
            soma += int(cpf[i]) * (10 - i)
        
        resto = soma % 11
        digito1 = 0 if resto < 2 else 11 - resto
        
        if int(cpf[9]) != digito1:
            return False
        
        # Calcular segundo dígito verificador
        soma = 0
        for i in range(10):
            soma += int(cpf[i]) * (11 - i)
        
        resto = soma % 11
        digito2 = 0 if resto < 2 else 11 - resto
        
        return int(cpf[10]) == digito2

    def validar_cnpj_algoritmo(self, cnpj):
        """Valida CNPJ usando algoritmo oficial"""
        if cnpj == cnpj[0] * 14:
            return False
        
        # Calcular primeiro dígito verificador
        peso = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        soma = sum(int(cnpj[i]) * peso[i] for i in range(12))
        resto = soma % 11
        digito1 = 0 if resto < 2 else 11 - resto
        
        if int(cnpj[12]) != digito1:
            return False
        
        # Calcular segundo dígito verificador
        peso = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        soma = sum(int(cnpj[i]) * peso[i] for i in range(13))
        resto = soma % 11
        digito2 = 0 if resto < 2 else 11 - resto
        
        return int(cnpj[13]) == digito2

    def salvar_na_base_fornecedores(self, dados):
        """Salva os dados na planilha de fornecedores - VERSÃO CORRIGIDA"""
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Coletar todos os dados existentes e o novo
            fornecedores = []
            
            # Converter dados existentes mantendo formato original da planilha
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tem CNPJ/CPF
                    fornecedor = {
                        'cnpj_cpf': str(row[0]) if row[0] is not None else '',  # CORREÇÃO: Garantir string
                        'tipo_pessoa': str(row[1]) if row[1] is not None else '',
                        'razao_social': str(row[2]) if row[2] is not None else '',
                        'nome': str(row[3]) if row[3] is not None else '',
                        'telefone': str(row[4]) if row[4] is not None else '',
                        'email': str(row[5]) if row[5] is not None else '',
                        'banco': str(row[6]) if row[6] is not None else '',
                        'op': str(row[7]) if row[7] is not None else '',
                        'agencia': str(row[8]) if row[8] is not None else '',
                        'conta': str(row[9]) if row[9] is not None else '',
                        'chave_pix': str(row[10]) if row[10] is not None else '',
                        'categoria': str(row[11]) if row[11] is not None else '',
                        'especificacao': str(row[12]) if row[12] is not None else '',
                        'vinculo': str(row[13]) if row[13] is not None else '',
                        'dados_bancarios': str(row[14]) if row[14] is not None else '',
                        'endereco': str(row[15]) if row[15] is not None else ''
                    }
                    fornecedores.append(fornecedor)
            
            # CORREÇÃO: Garantir que todos os campos do novo fornecedor sejam strings
            dados_corrigidos = {}
            for key, value in dados.items():
                if value is None:
                    dados_corrigidos[key] = ''
                else:
                    dados_corrigidos[key] = str(value)
            
            # Adicionar novo fornecedor ou atualizar existente
            fornecedor_encontrado = False
            for i, fornecedor in enumerate(fornecedores):
                # CORREÇÃO: Comparar strings com strings
                if str(fornecedor['cnpj_cpf']).strip() == str(dados_corrigidos['cnpj_cpf']).strip():
                    fornecedores[i] = dados_corrigidos.copy()
                    fornecedor_encontrado = True
                    break
            
            if not fornecedor_encontrado:
                fornecedores.append(dados_corrigidos.copy())
            
            # CORREÇÃO: Ordenar de forma segura convertendo tudo para string maiúscula
            try:
                fornecedores_ordenados = sorted(
                    fornecedores,
                    key=lambda x: (str(x.get('nome', '')).upper().strip(), str(x.get('cnpj_cpf', '')).strip())
                )
            except Exception as e:
                print(f"Erro na ordenação, mantendo ordem original: {str(e)}")
                # Se der erro na ordenação, manter ordem original
                fornecedores_ordenados = fornecedores
            
            # Limpar planilha existente
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            
            # Reescrever dados ordenados
            for i, fornecedor in enumerate(fornecedores_ordenados, start=2):
                ws.cell(row=i, column=1, value=fornecedor.get('cnpj_cpf', ''))
                ws.cell(row=i, column=2, value=fornecedor.get('tipo_pessoa', ''))
                ws.cell(row=i, column=3, value=fornecedor.get('razao_social', ''))
                ws.cell(row=i, column=4, value=fornecedor.get('nome', ''))
                ws.cell(row=i, column=5, value=fornecedor.get('telefone', ''))
                ws.cell(row=i, column=6, value=fornecedor.get('email', ''))
                ws.cell(row=i, column=7, value=fornecedor.get('banco', ''))
                ws.cell(row=i, column=8, value=fornecedor.get('op', ''))
                ws.cell(row=i, column=9, value=fornecedor.get('agencia', ''))
                ws.cell(row=i, column=10, value=fornecedor.get('conta', ''))
                ws.cell(row=i, column=11, value=fornecedor.get('chave_pix', ''))
                ws.cell(row=i, column=12, value=fornecedor.get('categoria', ''))
                ws.cell(row=i, column=13, value=fornecedor.get('especificacao', ''))
                ws.cell(row=i, column=14, value=fornecedor.get('vinculo', ''))
                ws.cell(row=i, column=15, value=fornecedor.get('dados_bancarios', ''))
                ws.cell(row=i, column=16, value=fornecedor.get('endereco', ''))
            
            wb.save(ARQUIVO_FORNECEDORES)
            
        except Exception as e:
            raise Exception(f"Erro ao salvar na planilha: {str(e)}")

    def atualizar_linha_fornecedor(self, row, dados):
        """Atualiza uma linha existente com novos dados"""
        row[0].value = dados['cnpj_cpf']
        row[1].value = dados['tipo_pessoa']  # Nova coluna para tipo de pessoa
        row[2].value = dados['razao_social']
        row[3].value = dados['nome']
        row[4].value = dados['telefone']
        row[5].value = dados['email']
        row[6].value = dados['banco']
        row[7].value = dados['op']
        row[8].value = dados['agencia']
        row[9].value = dados['conta']
        row[10].value = dados['chave_pix']
        row[11].value = dados['categoria']
        row[12].value = dados['especificacao']
        row[13].value = dados['vinculo']
        row[14].value = dados['dados_bancarios']
        row[15].value = dados['endereco']

    def adicionar_linha_fornecedor(self, ws, linha, dados):
        """Adiciona uma nova linha com os dados do fornecedor"""
        ws.cell(row=linha, column=1, value=dados['cnpj_cpf'])
        ws.cell(row=linha, column=2, value=dados['tipo_pessoa'])
        ws.cell(row=linha, column=3, value=dados['razao_social'])
        ws.cell(row=linha, column=4, value=dados['nome'])
        ws.cell(row=linha, column=5, value=dados['telefone'])
        ws.cell(row=linha, column=6, value=dados['email'])
        ws.cell(row=linha, column=7, value=dados['banco'])
        ws.cell(row=linha, column=8, value=dados['op'])
        ws.cell(row=linha, column=9, value=dados['agencia'])
        ws.cell(row=linha, column=10, value=dados['conta'])
        ws.cell(row=linha, column=11, value=dados['chave_pix'])
        ws.cell(row=linha, column=12, value=dados['categoria'])
        ws.cell(row=linha, column=13, value=dados['especificacao'])
        ws.cell(row=linha, column=14, value=dados['vinculo'])
        ws.cell(row=linha, column=15, value=dados['dados_bancarios'])
        ws.cell(row=linha, column=16, value=dados['endereco'])
      
    def atualizar_fornecedor(self):
        """Atualiza dados do fornecedor existente"""
        # Validações semelhantes ao salvar_fornecedor
        campos_obrigatorios = ['razao_social', 'nome', 'categoria']
        for campo in campos_obrigatorios:
            if not self.campos_form[campo].get().strip():
                custom_messagebox("error", "Erro", f"O campo {campo} é obrigatório!")
                return

        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            cnpj_cpf = self.campos_form['cnpj_cpf'].get()
            for row in ws.iter_rows(min_row=2):
                if row[0].value == cnpj_cpf:
                    # Atualizar dados na linha existente
                    row[1].value = self.campos_form['tipo_pessoa'].get().upper()
                    row[2].value = self.campos_form['razao_social'].get().upper()
                    row[3].value = self.campos_form['nome'].get().upper()
                    row[4].value = self.campos_form['telefone'].get()
                    row[5].value = self.campos_form['email'].get()
                    row[6].value = self.campos_form['banco'].get()
                    row[7].value = self.campos_form['op'].get()
                    row[8].value = self.campos_form['agencia'].get()
                    row[9].value = self.campos_form['conta'].get()
                    row[10].value = self.campos_form['chave_pix'].get()
                    row[11].value = self.campos_form['categoria'].get()
                    row[12].value = self.campos_form['especificacao'].get().upper()
                    row[13].value = self.campos_form['vinculo'].get().upper()
                    row[14].value = self.campos_form['dados_bancarios'].get().upper()
                    row[15].value = self.campos_form['endereco'].get().upper()
                    break

            wb.save(ARQUIVO_FORNECEDORES)
            custom_messagebox("info", "Sucesso", "Fornecedor atualizado com sucesso!")
            self.janela_fornecedor.destroy()
            self.buscar_fornecedor()  # Atualiza a lista
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao atualizar fornecedor: {str(e)}")

    def preencher_dados_fornecedor(self, dados):
        """Preenche os campos do fornecedor na aba de entrada"""
        self.campos_fornecedor['cnpj_cpf'].delete(0, tk.END)
        self.campos_fornecedor['cnpj_cpf'].insert(0, dados[0])
        
        self.campos_fornecedor['nome'].delete(0, tk.END)
        self.campos_fornecedor['nome'].insert(0, dados[1])
        
        self.campos_fornecedor['categoria'].delete(0, tk.END)
        self.campos_fornecedor['categoria'].insert(0, dados[2])

    def abrir_visualizador_fornecedor(self):
        """Abre o visualizador de lançamentos para o fornecedor selecionado"""
        # Verificar se há fornecedor selecionado
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um fornecedor primeiro!")
            return
        
        # Obter dados do fornecedor
        valores = self.tree_fornecedores.item(selecionado[0])['values']
        cnpj_cpf = valores[0]
        nome = valores[1]
        
        # Verificar se há cliente selecionado
        if not self.cliente_atual:
            custom_messagebox("error", "Erro", "Selecione um cliente primeiro!")
            return
        
        # Abrir visualizador
        visualizador = VisualizadorLancamentosFornecedor(self.root, self)
        visualizador.abrir_visualizador(cnpj_cpf, nome)

    def validar_tipo_despesa(self, P):
        """
        Valida entrada do tipo de despesa
        Args:
            P: valor proposto após a modificação
        """
        if P == "": return True  # Permite campo vazio
        if not P.isdigit(): return False  # Permite apenas dígitos
        return 1 <= int(P) <= 6  # Permite apenas valores entre 1 e 6

    def setup_aba_dados(self):
        """Configura a aba de entrada de dados com layout aprimorado e ordem de campos otimizada"""
        # Verificar se o estilo Medium.TButton já existe
        style = ttk.Style()
        if not style.lookup('Medium.TButton', 'font'):
            style.configure('Medium.TButton', font=('Arial', 11), padding=(10, 6))

        # Frame para cabeçalho com informações do cliente
        frame_cabecalho = ttk.Frame(self.aba_dados)
        frame_cabecalho.pack(fill='x', padx=10, pady=5)
        
        # Label do cliente destacado
        self.cliente_label = ttk.Label(frame_cabecalho, 
                                    text="Cliente: Nenhum selecionado", 
                                    font=('Arial', 12, 'bold'),
                                    foreground='#0056b3')
        self.cliente_label.pack(side='left', padx=5)
        
        # Frame para data de referência
        frame_data = ttk.LabelFrame(self.aba_dados, text="Data de Referência")
        frame_data.pack(fill='x', padx=10, pady=8)
        
        # Container interno para organização da data
        frame_data_interno = ttk.Frame(frame_data)
        frame_data_interno.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_data_interno, text="Data do Relatório:", font=('Arial', 10)).pack(side='left', padx=5)
        
        
        self.data_rel_entry = DateEntry(
            frame_data_interno,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2,
            font=('Arial', 10),
        )
        self.data_rel_entry.pack(side='left', padx=5, pady=5)
        
        # Definir data de referência inicial
        data_rel_inicial = self.calcular_data_rel()
        self.data_rel_entry.set_date(data_rel_inicial)
        
        def validar_entrada_data(event=None):
            data = self.data_rel_entry.get()
            if not validar_data(data):
                custom_messagebox("error", "Erro", "Data inválida! Use o formato dd/mm/aaaa")
                self.data_rel_entry.delete(0, tk.END)
                self.data_rel_entry.insert(0, datetime.now().strftime('%d/%m/%Y'))
                return False
            return True
        
        self.data_rel_entry.bind('<FocusOut>', validar_entrada_data)  # Valida quando perde o foco
        
        # Frame para dados do fornecedor
        frame_fornecedor = ttk.LabelFrame(self.aba_dados, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', padx=10, pady=8)
        
        # Adicione esta linha para tornar frame_fornecedor um atributo da classe
        self.frame_fornecedor = frame_fornecedor
        
        # Grid para organizar os campos de fornecedor de forma mais equilibrada
        self.campos_fornecedor = {}
        campos = [('cnpj_cpf', 'CNPJ/CPF:'), 
                ('nome', 'Nome:'), 
                ('categoria', 'Categoria:')]
        
        for row, (campo, label) in enumerate(campos):
            ttk.Label(frame_fornecedor, text=label, font=('Arial', 10)).grid(row=row, column=0, padx=5, pady=5, sticky='e')
            entry = ttk.Entry(frame_fornecedor, width=40, font=('Arial', 10))
            entry.grid(row=row, column=1, padx=5, pady=5, sticky='ew')
            if campo != 'categoria':
                entry.config(state='readonly')
            self.campos_fornecedor[campo] = entry
        
        # Frame para forma de pagamento
        frame_pagamento = ttk.Frame(frame_fornecedor)
        frame_pagamento.grid(row=len(campos), column=0, columnspan=2, pady=5, sticky='ew')
        
        ttk.Label(frame_pagamento, text="Forma de Pagamento:", font=('Arial', 10)).pack(side='left', padx=5)
        self.forma_pagamento_combo = ttk.Combobox(
            frame_pagamento,
            textvariable=self.forma_pagamento_var,
            values=["PIX", "TED", "DINHEIRO"],
            state="readonly",
            width=10,
            font=('Arial', 10)
        )
        self.forma_pagamento_combo.pack(side='left', padx=5)
        self.forma_pagamento_combo.bind('<<ComboboxSelected>>', self.atualizar_dados_bancarios)
        
        # Dados Bancários (agora após a forma de pagamento)
        ttk.Label(frame_fornecedor, text="Dados Bancários:", font=('Arial', 10)).grid(row=len(campos) + 1, column=0, padx=5, pady=5, sticky='e')
        entry = ttk.Entry(frame_fornecedor, width=40, state='readonly', font=('Arial', 10))
        entry.grid(row=len(campos) + 1, column=1, padx=5, pady=5, sticky='ew')
        self.campos_fornecedor['dados_bancarios'] = entry
        
        # Configure expandability of columns
        frame_fornecedor.columnconfigure(1, weight=1)
        
        # Frame para botões de parcelamento
        frame_parcelamento = ttk.Frame(self.aba_dados)
        frame_parcelamento.pack(fill='x', padx=10, pady=5)
        
        # Inicializar o gestor de parcelas com a janela root
        self.gestor_parcelas = GestorParcelas(self)
        
        ttk.Button(
            frame_parcelamento,
            text="Parcelar Despesa",
            command=self.abrir_parcelamento,
            style='Medium.TButton'
        ).pack(side='left', padx=5)
        
        # Frame para dados da despesa com layout em grid otimizado
        frame_despesa = ttk.LabelFrame(self.aba_dados, text="Dados da Despesa")
        frame_despesa.pack(fill='both', expand=True, padx=10, pady=8)
        
        # Adicionar as opções de referência para tipo 1
        self.opcoes_referencia_tipo1 = [
            'DIÁRIA', 'SALÁRIO', 'TRANSPORTE', 
            'FÉRIAS', '13º SALÁRIO', 'RESCISÃO', 'CAFÉ'
        ]
        
        self.campos_despesa = {}
        
        # Criar grid layout de 4x4 para os campos da despesa
        # Coluna 0 e 1: Labels e campos da esquerda
        # Coluna 2 e 3: Labels e campos da direita
        
        # ===== LADO ESQUERDO (valores numéricos) =====
    
        # Tipo Despesa (row=0)
        ttk.Label(frame_despesa, text="Tipo Despesa (1-6):", font=('Arial', 10)).grid(
            row=0, column=0, padx=5, pady=5, sticky='e')
        vcmd = (frame_despesa.register(self.validar_tipo_despesa), '%P')
        self.campos_despesa['tp_desp'] = ttk.Entry(
            frame_despesa, validate='key', validatecommand=vcmd, font=('Arial', 10), width=10)
        self.campos_despesa['tp_desp'].grid(row=0, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # Valor Unitário (row=1)
        ttk.Label(frame_despesa, text="Valor Unitário:", font=('Arial', 10)).grid(
            row=1, column=0, padx=5, pady=5, sticky='e')
        self.campos_despesa['vr_unit'] = ttk.Entry(frame_despesa, font=('Arial', 10), width=15)
        self.campos_despesa['vr_unit'].grid(row=1, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # Dias (row=2)
        ttk.Label(frame_despesa, text="Dias:", font=('Arial', 10)).grid(
            row=2, column=0, padx=5, pady=5, sticky='e')
        self.campos_despesa['dias'] = ttk.Entry(frame_despesa, font=('Arial', 10), width=8)
        self.campos_despesa['dias'].grid(row=2, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # Valor Total (row=3)
        ttk.Label(frame_despesa, text="Valor Total:", font=('Arial', 10)).grid(
            row=3, column=0, padx=5, pady=5, sticky='e')
        self.campos_despesa['valor'] = ttk.Entry(
            frame_despesa, state='readonly', font=('Arial', 10), width=15)
        self.campos_despesa['valor'].grid(row=3, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # Data Vencimento (row=4) - MOVIDO PARA A ESQUERDA
        ttk.Label(frame_despesa, text="Data Vencimento:", font=('Arial', 10)).grid(
            row=4, column=0, padx=5, pady=5, sticky='e')
        self.campos_despesa['dt_vencto'] = DateEntry(
            frame_despesa,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2,
            font=('Arial', 10),
            width=15
        )
        self.campos_despesa['dt_vencto'].grid(row=4, column=1, padx=(5, 20), pady=5, sticky='w')
        self.campos_despesa['dt_vencto'].delete(0, tk.END)  # Inicializa vazio
        
        # ===== LADO DIREITO (texto) =====
        
        # Referência (row=0)
        ttk.Label(frame_despesa, text="Referência:", font=('Arial', 10)).grid(
            row=0, column=2, padx=5, pady=5, sticky='e')
        self.campos_despesa['referencia'] = ttk.Combobox(
            frame_despesa, font=('Arial', 10), width=40)
        self.campos_despesa['referencia']['values'] = self.opcoes_referencia_tipo1
        self.campos_despesa['referencia'].grid(row=0, column=3, padx=5, pady=5, sticky='ew')
        self.campos_despesa['referencia'].bind(
            '<<ComboboxSelected>>', lambda e: self.calcular_valor_total())
        
        # Etapa da Obra (row=1)
        ttk.Label(frame_despesa, text="Etapa da Obra:", font=('Arial', 10)).grid(
            row=1, column=2, padx=5, pady=5, sticky='e')
        
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        etapas_obra = GerenciadorConfiguracoes.get_etapas_obra()
        
        self.campos_despesa['etapa_obra'] = ComboboxAutocompletar(
            frame_despesa, 
            values=etapas_obra,
            config_key='etapas_obra',
            config_manager=GerenciadorConfiguracoes,
            font=('Arial', 10), 
            width=40, 
            state='normal'  # Mudado de 'readonly' para 'normal'
        )
        self.campos_despesa['etapa_obra'].grid(row=1, column=3, padx=5, pady=5, sticky='ew')

        # Insumo (row=2)
        ttk.Label(frame_despesa, text="Insumo:", font=('Arial', 10)).grid(
            row=2, column=2, padx=5, pady=5, sticky='e')
        
        # Obter lista de insumos das configurações
        insumos = GerenciadorConfiguracoes.get_insumos()
        
        self.campos_despesa['insumo'] = ComboboxAutocompletar(
            frame_despesa, 
            values=insumos,
            config_key='insumos',
            config_manager=GerenciadorConfiguracoes,
            font=('Arial', 10), 
            width=40, 
            state='normal'  # Mudado de 'readonly' para 'normal'
        )
        self.campos_despesa['insumo'].grid(row=2, column=3, padx=5, pady=5, sticky='ew')

        # NF + Checkbox para materiais (row=3)
        ttk.Label(frame_despesa, text="NF:", font=('Arial', 10)).grid(
            row=3, column=2, padx=5, pady=5, sticky='e')

        # Frame para NF e checkbox de materiais
        frame_nf = ttk.Frame(frame_despesa)
        frame_nf.grid(row=3, column=3, padx=5, pady=5, sticky='ew')

        self.campos_despesa['nf'] = ttk.Entry(frame_nf, font=('Arial', 10), width=15)
        self.campos_despesa['nf'].pack(side='left')

        # Checkbox para indicar se há materiais vinculados
        self.tem_materiais_var = tk.BooleanVar()
        self.checkbox_materiais = ttk.Checkbutton(
            frame_nf,
            text="Tem materiais",
            variable=self.tem_materiais_var,
            command=self.handle_checkbox_change
        )
        self.checkbox_materiais.pack(side='left', padx=(10, 0))
                
        # Observação (row=4)
        ttk.Label(frame_despesa, text="Observação:", font=('Arial', 10)).grid(
            row=4, column=2, padx=5, pady=5, sticky='e')
        self.campos_despesa['observacao'] = ttk.Entry(frame_despesa, font=('Arial', 10), width=40)
        self.campos_despesa['observacao'].grid(row=4, column=3, padx=5, pady=5, sticky='ew')
    
        # Configurar peso da coluna para expandir apenas os campos de referência e observação
        frame_despesa.columnconfigure(3, weight=1)  # Apenas a coluna 3 (campos expansíveis) cresce
        
        # Inserir valor padrão para dias
        self.campos_despesa['dias'].insert(0, "1")
        
        # Bindings
        self.campos_despesa['vr_unit'].bind('<KeyRelease>', self.calcular_valor_total)
        self.campos_despesa['dias'].bind('<KeyRelease>', self.calcular_valor_total)
        self.campos_despesa['tp_desp'].bind('<KeyRelease>', self.verificar_tipo_despesa)
        
        # Configurar a ordem de tab para seguir o fluxo de trabalho natural
        self.campos_despesa['tp_desp'].bind('<Return>', lambda e: self.campos_despesa['referencia'].focus())
        self.campos_despesa['referencia'].bind('<Return>', lambda e: self.campos_despesa['vr_unit'].focus())
        self.campos_despesa['vr_unit'].bind('<Return>', lambda e: self.campos_despesa['dias'].focus())
        self.campos_despesa['dias'].bind('<Return>', lambda e: self.campos_despesa['etapa_obra'].focus())
        self.campos_despesa['etapa_obra'].bind('<Return>', lambda e: self.campos_despesa['insumo'].focus())
        self.campos_despesa['insumo'].bind('<Return>', lambda e: self.campos_despesa['nf'].focus())
        self.campos_despesa['nf'].bind('<Return>', lambda e: self.campos_despesa['dt_vencto'].focus())
        self.campos_despesa['dt_vencto'].bind('<Return>', lambda e: self.campos_despesa['observacao'].focus())
        
        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.aba_dados)
        frame_botoes.pack(fill='x', padx=10, pady=10, side='bottom')
        
        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.aba_dados)
        frame_botoes.pack(fill='x', padx=10, pady=10, side='bottom')
        
        # Organizar botões com Adicionar em destaque à direita
        ttk.Button(frame_botoes, text="Cancelar", 
                command=self.cancelar_entrada,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Visualizar Lançamentos", 
                command=self.visualizar_lancamentos,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Enviar", 
                command=self.enviar_dados,
                style='Medium.TButton').pack(side='left', padx=5)
        
        # Botão Adicionar em destaque (lado direito)
        adicionar_btn = ttk.Button(frame_botoes, text="Adicionar", 
                                command=self.adicionar_dados,
                                style='Medium.TButton')
        adicionar_btn.pack(side='right', padx=5)
        
        # Configurar um estilo especial para o botão Adicionar (opcional)
        style = ttk.Style()
        style.configure('Destaque.TButton', 
                    background='#0056b3',  # Esta propriedade pode não ter efeito em todos os temas
                    font=('Arial', 11, 'bold'))
        adicionar_btn.configure(style='Destaque.TButton')

        # ADICIONAR NO FINAL DO MÉTODO:
        print("DEBUG: setup_aba_dados executado completamente!")
        print(f"DEBUG: tem_materiais_var criado: {hasattr(self, 'tem_materiais_var')}")
        print(f"DEBUG: checkbox_materiais criado: {hasattr(self, 'checkbox_materiais')}")
        
        if hasattr(self, 'checkbox_materiais'):
            try:
                comando = self.checkbox_materiais.cget('command')
                print(f"DEBUG: Comando do checkbox: {comando}")
            except Exception as e:
                print(f"DEBUG: Erro ao verificar comando: {e}")

    def atualizar_comboboxes_dinamicamente(self):
        """
        Método para atualizar os comboboxes quando as configurações são alteradas
        Chame este método sempre que as configurações forem modificadas
        """
        try:
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            
            # Atualizar etapas da obra
            if hasattr(self, 'campos_despesa') and 'etapa_obra' in self.campos_despesa:
                etapas_atualizadas = GerenciadorConfiguracoes.get_etapas_obra()
                self.campos_despesa['etapa_obra'].atualizar_valores(etapas_atualizadas)
            
            # Atualizar insumos
            if hasattr(self, 'campos_despesa') and 'insumo' in self.campos_despesa:
                insumos_atualizados = GerenciadorConfiguracoes.get_insumos()
                self.campos_despesa['insumo'].atualizar_valores(insumos_atualizados)
                
        except Exception as e:
            print(f"Erro ao atualizar comboboxes: {e}")

    def calcular_data_rel(self):
        """
        Calcula a data de referência seguindo a regra dos dias 5 e 20
        """
        hoje = datetime.now()
        if 6 <= hoje.day <= 20:
            data_rel = hoje.replace(day=20)
        else:
            if hoje.day > 20:
                data_rel = (hoje + relativedelta(months=1)).replace(day=5)
            else:
                data_rel = hoje.replace(day=5)
        return data_rel

    def visualizar_lancamentos(self):
        """Abre a janela de visualização de lançamentos pendentes"""
        if hasattr(self, 'visualizador') and self.visualizador and hasattr(self.visualizador, 'janela') and self.visualizador.janela.winfo_exists():
            # Se o visualizador já existir, apenas trazê-lo para frente
            self.visualizador.janela.lift()
            self.visualizador.janela.focus_force()
            return
        
        # Criar nova instância do visualizador
        self.visualizador = VisualizadorLancamentos(self)
        
        # Configurar callback para quando a janela for fechada
        self.visualizador.janela.protocol("WM_DELETE_WINDOW", self.on_visualizador_close)
        
        # Atualizar dados
        self.visualizador.dados_para_incluir = self.dados_para_incluir.copy()
        self.visualizador.atualizar_dados(self.dados_para_incluir)
        
        # Garantir que a janela fique na frente
        self.visualizador.janela.lift()
        self.visualizador.janela.focus_force()

    def on_visualizador_close(self):
        """Manipula o fechamento da janela do visualizador"""
        # Atualizar dados_para_incluir com os dados mais recentes do visualizador
        if self.visualizador:
            self.dados_para_incluir = self.visualizador.get_dados_atualizados()
            self.visualizador.janela.destroy()
            self.visualizador = None

    def adicionar_botao_gerenciar_lancamentos(self):
        """Adiciona botão para gerenciar lançamentos na aba de fornecedor"""
        frame_gerenciamento = ttk.LabelFrame(self.aba_fornecedor, text="Gerenciamento de Dados")
        frame_gerenciamento.pack(fill='x', padx=10, pady=5)
        
        frame_botoes_ger = ttk.Frame(frame_gerenciamento)
        frame_botoes_ger.pack(fill='x', padx=5, pady=8)
        
        ttk.Button(
            frame_botoes_ger, 
            text="📅 Agenda",
            command=self.abrir_agenda,
            style='Medium.TButton'
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes_ger, 
            text="Gerenciar Lançamentos",
            command=self.abrir_gerenciador_lancamentos,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_ger, 
            text="📋 Visualizar Lançamentos",
            command=self.visualizar_lancamentos,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

    def abrir_gerenciador_lancamentos(self):
        """Abre o gerenciador de lançamentos"""
        if not hasattr(self, 'gerenciador_lancamentos') or self.gerenciador_lancamentos is None:
            self.gerenciador_lancamentos = GerenciadorLancamentos(self)
        self.gerenciador_lancamentos.abrir_gerenciador()

    def processar_parcelas(self):
        """Processa as parcelas geradas mantendo os dados do fornecedor"""
        print("Iniciando processamento de parcelas...")
        
        # Verificar se há parcelas para processar
        if not hasattr(self, 'gestor_parcelas') or not self.gestor_parcelas.parcelas:
            print("Nenhuma parcela para processar")
            return False
            
        # Validar se há fornecedor selecionado
        if not self.campos_fornecedor['cnpj_cpf'].get():
            custom_messagebox("error", "Erro", "Selecione um fornecedor antes de processar as parcelas!")
            return False
            
        # Guardar dados do fornecedor atual
        dados_fornecedor = {
            'cnpj_cpf': self.campos_fornecedor['cnpj_cpf'].get(),
            'nome': self.campos_fornecedor['nome'].get(),
            'categoria': self.campos_fornecedor['categoria'].get(),
            'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get()
        }
        
        print(f"Dados do fornecedor capturados: {dados_fornecedor}")
        total_parcelas = len(self.gestor_parcelas.parcelas)
        print(f"Total de parcelas a processar: {total_parcelas}")
        
        try:
            processadas = 0
            for i, parcela in enumerate(self.gestor_parcelas.parcelas, 1):
                print(f"\nProcessando parcela {i} de {total_parcelas}")
                
                # Restaurar dados do fornecedor antes de cada parcela
                for campo, valor in dados_fornecedor.items():
                    entry = self.campos_fornecedor[campo]
                    entry.config(state='normal')
                    entry.delete(0, tk.END)
                    entry.insert(0, valor)
                    if campo != 'categoria':
                        entry.config(state='readonly')
                
                print(f"Dados do fornecedor restaurados para parcela {i}")

                
                # Preencher dados da parcela
                self.data_rel_entry.set_date(datetime.strptime(parcela['data_rel'], '%d/%m/%Y'))
                self.campos_despesa['tp_desp'].delete(0, tk.END)
                self.campos_despesa['tp_desp'].insert(0, self.gestor_parcelas.tipo_despesa_valor)
                self.campos_despesa['nf'].delete(0, tk.END)
                self.campos_despesa['nf'].insert(0, parcela['nf'])
                
                if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                    self.campos_despesa['referencia'].set(parcela['referencia'])
                else:
                    self.campos_despesa['referencia'].delete(0, tk.END)
                    self.campos_despesa['referencia'].insert(0, parcela['referencia'])
                
                self.campos_despesa['vr_unit'].delete(0, tk.END)
                self.campos_despesa['vr_unit'].insert(0, f"{parcela['valor']:.2f}")
                
                self.campos_despesa['valor'].config(state='normal')
                self.campos_despesa['valor'].delete(0, tk.END)
                self.campos_despesa['valor'].insert(0, f"{parcela['valor']:.2f}")
                self.campos_despesa['valor'].config(state='readonly')
                
                self.campos_despesa['dt_vencto'].set_date(
                    datetime.strptime(parcela['dt_vencto'], '%d/%m/%Y')
                )

                self.campos_despesa['etapa_obra'].delete(0, tk.END)
                self.campos_despesa['etapa_obra'].insert(0, parcela.get('etapa_obra', ''))

                self.campos_despesa['insumo'].delete(0, tk.END)
                self.campos_despesa['insumo'].insert(0, parcela.get('insumo', ''))
                
                # Adicionar à lista de dados e verificar sucesso
                if self.adicionar_dados(eh_parcelamento=True):
                    processadas += 1
                    print(f"Parcela {i} processada com sucesso")
                else:
                    print(f"Falha ao processar parcela {i}")
            
            # Relatório final
            if processadas == total_parcelas:
                custom_messagebox("info", "Sucesso", 
                                  f"Todas as {total_parcelas} parcelas foram processadas com sucesso!")
            else:
                custom_messagebox("warning",  "Aviso", 
                                     f"Apenas {processadas} de {total_parcelas} parcelas foram processadas.")
            
            return processadas == total_parcelas
            
        except Exception as e:
            erro_msg = f"Erro ao processar parcelas: {str(e)}"
            print(erro_msg)
            custom_messagebox("error", "Erro", erro_msg)
            return False
            
        finally:
            self.limpar_campos_despesa()
            print("Processamento de parcelas finalizado")

    def abrir_parcelamento(self):
        """Abre a janela de parcelamento e processa os dados após o fechamento"""
        print("\nIniciando processo de parcelamento...")
        
        # Verificar se há fornecedor selecionado
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get()
        if not cnpj_cpf:
            print("Erro: Fornecedor não selecionado")
            custom_messagebox("error", "Erro", "Selecione um fornecedor antes de criar parcelas!")
            return

        print("\nCapturando dados do fornecedor...")
        dados_fornecedor = {
            'cnpj_cpf': cnpj_cpf,
            'nome': self.campos_fornecedor['nome'].get(),
            'categoria': self.campos_fornecedor['categoria'].get(),
            'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get()
        }
        print(f"Dados capturados: {dados_fornecedor}")
        
        # Validar se todos os campos do fornecedor estão preenchidos
        if not all(dados_fornecedor.values()):
            print("Erro: Dados do fornecedor incompletos")
            custom_messagebox("error", "Erro", "Dados do fornecedor incompletos!")
            return

        print("Abrindo janela de parcelamento...")
        self.gestor_parcelas.abrir_janela_parcelas()
        self.root.wait_window(self.gestor_parcelas.janela_parcelas)

        if hasattr(self.gestor_parcelas, 'parcelas') and self.gestor_parcelas.parcelas:
            print(f"Processando {len(self.gestor_parcelas.parcelas)} parcelas...")
            
            success = True
            for i, parcela in enumerate(self.gestor_parcelas.parcelas, 1):
                try:
                    print(f"\nProcessando parcela {i}")
                    
                    # Restaurar dados do fornecedor
                    for campo, valor in dados_fornecedor.items():
                        entry = self.campos_fornecedor[campo]
                        entry.config(state='normal')
                        entry.delete(0, tk.END)
                        entry.insert(0, valor)
                        if campo != 'categoria':
                            entry.config(state='readonly')
                    
                    # Preencher dados da parcela
                    self.data_rel_entry.set_date(datetime.strptime(parcela['data_rel'], '%d/%m/%Y'))
                    
                    self.campos_despesa['tp_desp'].delete(0, tk.END)
                    self.campos_despesa['tp_desp'].insert(0, self.gestor_parcelas.tipo_despesa_valor)
                    self.campos_despesa['nf'].delete(0, tk.END)
                    self.campos_despesa['nf'].insert(0, parcela['nf'])
                    
                    if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                        self.campos_despesa['referencia'].set(parcela['referencia'])
                    else:
                        self.campos_despesa['referencia'].delete(0, tk.END)
                        self.campos_despesa['referencia'].insert(0, parcela['referencia'])
                    
                    self.campos_despesa['vr_unit'].delete(0, tk.END)
                    self.campos_despesa['vr_unit'].insert(0, f"{parcela['valor']:.2f}")
                    
                    self.campos_despesa['dias'].delete(0, tk.END)
                    self.campos_despesa['dias'].insert(0, '1')
                    
                    self.campos_despesa['valor'].config(state='normal')
                    self.campos_despesa['valor'].delete(0, tk.END)
                    self.campos_despesa['valor'].insert(0, f"{parcela['valor']:.2f}")
                    self.campos_despesa['valor'].config(state='readonly')
                    
                    self.campos_despesa['dt_vencto'].set_date(
                        datetime.strptime(parcela['dt_vencto'], '%d/%m/%Y')
                    )
                    
                    self.campos_despesa['etapa_obra'].delete(0, tk.END)
                    self.campos_despesa['etapa_obra'].insert(0, parcela.get('etapa_obra', ''))

                    self.campos_despesa['insumo'].delete(0, tk.END)
                    self.campos_despesa['insumo'].insert(0, parcela.get('insumo', ''))

                    # Adicionar à lista de dados
                    if not self.adicionar_dados(eh_parcelamento=True):
                        print(f"Falha ao adicionar parcela {i}")
                        success = False
                        break
                    
                    print(f"Parcela {i} processada com sucesso")
                    
                except Exception as e:
                    success = False
                    print(f"Erro ao processar parcela {i}: {str(e)}")
                    custom_messagebox("error", "Erro", f"Erro ao processar parcela {i}: {str(e)}")
                    break
            
            if success:
                custom_messagebox("info", "Sucesso", 
                                  f"Todas as {len(self.gestor_parcelas.parcelas)} parcelas foram processadas!")
                # Calcular a data de referência padrão
                hoje = datetime.now()
                if 6 <= hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    if hoje.day > 20:
                        data_rel = (hoje + relativedelta(months=1)).replace(day=5)
                    else:
                        data_rel = hoje.replace(day=5)
                
                # Restaurar a data de referência padrão
                self.data_rel_entry.set_date(data_rel)
                
                self.limpar_campos_despesa()
                self.notebook.select(1)  # Volta para aba fornecedor
            else:
                custom_messagebox("error", "Erro", "Houve um erro no processamento das parcelas.")
        else:
            print("Nenhuma parcela para processar")

    def abrir_calendario(self):
        try:
            top = Toplevel(self.root)
            top.title("Selecionar Data")
            top.geometry("300x250")
            top.grab_set()  # Torna a janela modal
        
            cal = Calendar(top, selectmode='day', 
                          date_pattern='dd/mm/yyyy',
                          locale='pt_BR')
            cal.pack(padx=10, pady=10)
        
            def selecionar_data():
                data = cal.get_date()
                self.data_rel_entry.delete(0, tk.END)
                self.data_rel_entry.insert(0, data)
                top.destroy()
        
            ttk.Button(top, text="OK", command=selecionar_data).pack(pady=5)
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir calendário: {str(e)}")

    def atualizar_campo_referencia(self, event=None):
        """Atualiza o campo de referência baseado no tipo de despesa"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
    
        try:
            if tp_desp == '1':
                # Redefine as opções e configura como readonly
                self.campos_despesa['referencia']['values'] = self.opcoes_referencia_tipo1
                self.campos_despesa['referencia'].config(state='readonly')
                # Seleciona o primeiro item como padrão
                if self.opcoes_referencia_tipo1:
                    self.campos_despesa['referencia'].set(self.opcoes_referencia_tipo1[0])
            else:
                # Para outros tipos, limpa a seleção e permite digitação
                self.campos_despesa['referencia'].set('')
                self.campos_despesa['referencia']['values'] = []
                self.campos_despesa['referencia'].config(state='normal')
            
        except Exception as e:
            print(f"Erro ao atualizar campo referência: {str(e)}")

    def atualizar_dados_bancarios(self, event=None):
        """Atualiza os dados bancários baseado no tipo de despesa"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get().strip()
    
        if not cnpj_cpf:  # Se não houver fornecedor selecionado
            return
        
        fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
        if not fornecedor_completo:
            return
        
        self.campos_fornecedor['dados_bancarios'].config(state='normal')
        self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
        
        # Construir dados bancários baseado na forma de pagamento
        forma_pagamento = self.forma_pagamento_var.get()
        
        if forma_pagamento == "DINHEIRO":
            dados_bancarios = "PAGAMENTO EM DINHEIRO"
        elif forma_pagamento == "PIX" and fornecedor_completo['chave_pix']:
            dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
        else:
            # Estrutura para TED
            dados_ted = []
            if fornecedor_completo['banco']: dados_ted.append(str(fornecedor_completo['banco']))
            if fornecedor_completo['op']: dados_ted.append(str(fornecedor_completo['op']))
            if fornecedor_completo['agencia']: dados_ted.append(str(fornecedor_completo['agencia']))
            if fornecedor_completo['conta']: dados_ted.append(str(fornecedor_completo['conta']))
            # SEMPRE adicionar o CNPJ/CPF para TED
            dados_ted.append(str(fornecedor_completo['cnpj_cpf']))
            
            dados_bancarios = ' - '.join(filter(None, dados_ted))

        if dados_bancarios.strip() in ['', ' - ']:
            dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
            
        self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
        self.campos_fornecedor['dados_bancarios'].config(state='readonly')
 
    def cancelar_entrada(self):
        """Cancela a entrada de dados atual e retorna à aba fornecedor"""
        if any(self.campos_despesa[campo].get() for campo in ['tp_desp', 'referencia', 'vr_unit']):
            if custom_messagebox("yesno", "Confirmação", "Deseja descartar os dados atuais?"):
                self.limpar_campos_despesa()
                self.notebook.select(1)  # Volta para aba fornecedor
        else:
            self.notebook.select(1)  # Volta para aba fornecedor
    
    def abrir_gestao_locacoes(self):
        """Abre o módulo de gestão de locações"""
        if not self.cliente_atual:
            custom_messagebox("warning", "Aviso", "Selecione um cliente primeiro!")
            return
        
        try:
            gerenciador = GerenciadorLocacoes(self)
            gerenciador.abrir_gestao_locacoes()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir gestão de locações: {str(e)}")
            import traceback
            traceback.print_exc()

    def abrir_agenda(self):
        """Abre o gerenciador de agenda"""
        try:
            if not self.cliente_atual:
                custom_messagebox("error", "Erro", "Selecione um cliente primeiro!")
                return
            
            # Importar e instanciar o gerenciador de agenda apenas quando necessário
            if not hasattr(self, 'gerenciador_agenda') or self.gerenciador_agenda is None:
                self.gerenciador_agenda = GerenciadorAgenda(self)
            
            self.gerenciador_agenda.abrir_agenda()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir agenda: {str(e)}")
            print(f"DEBUG: Erro ao abrir agenda: {str(e)}")

    def validar_campos_agenda(self):
        """Validação específica para dados vindos da agenda - SIMPLIFICADA"""
        try:
            print("DEBUG: Validação simplificada para agenda")
            
            # Verificar apenas campos essenciais sem mostrar erros
            # (os erros serão tratados pelo fluxo normal se necessário)
            
            # CNPJ/CPF
            cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get().strip()
            if not cnpj_cpf:
                print("DEBUG: CNPJ/CPF vazio")
                return False
            
            # Nome
            nome = self.campos_fornecedor['nome'].get().strip()
            if not nome:
                print("DEBUG: Nome vazio")
                return False
            
            # Valor unitário
            vr_unit_str = self.campos_despesa['vr_unit'].get().strip()
            if not vr_unit_str:
                print("DEBUG: Valor unitário vazio")
                return False
            
            try:
                vr_unit = float(vr_unit_str.replace(',', '.'))
                if vr_unit <= 0:
                    print("DEBUG: Valor unitário inválido")
                    return False
            except ValueError:
                print("DEBUG: Valor unitário não numérico")
                return False
            
            print("DEBUG: Validação simplificada passou")
            return True
            
        except Exception as e:
            print(f"DEBUG: Erro na validação simplificada: {str(e)}")
            return False

    def inserir_lancamento_completo(self, dados_lancamento):
        """
        Método para inserir lançamento completo chamado pela agenda
        VERSÃO CORRIGIDA - SEGUINDO O FLUXO CORRETO DO SISTEMA
        """
        try:
            print("DEBUG: Iniciando inserção de lançamento via agenda")
            
            # 1. Validar dados básicos
            if not self.validar_dados_basicos_agenda(dados_lancamento):
                return False
            
            # 2. Verificar se cliente está selecionado
            if not hasattr(self, 'cliente_atual') or not self.cliente_atual:
                custom_messagebox("error", "Erro", "Nenhum cliente selecionado!")
                return False
            
            # 3. Configurar campos do formulário principal com os dados
            self.preencher_campos_desde_agenda(dados_lancamento)
            
            # 4. FLUXO CORRETO: Usar adicionar_dados que adiciona à lista
            print("DEBUG: Chamando adicionar_dados para adicionar à lista")
            
            # Temporariamente substituir o método de validação
            metodo_validacao_original = self.validar_campos
            self.validar_campos = self.validar_campos_agenda
            
            try:
                sucesso = self.adicionar_dados(eh_parcelamento=False)
            finally:
                # Restaurar método original
                self.validar_campos = metodo_validacao_original
            
            if sucesso:
                print("DEBUG: Dados adicionados à lista com sucesso")
                
                # 5. FLUXO CORRETO: Chamar enviar_dados para salvar na planilha
                print("DEBUG: Chamando enviar_dados para salvar na planilha")
                
                try:
                    # Verificar se há dados para enviar
                    if not hasattr(self, 'dados_para_incluir') or not self.dados_para_incluir:
                        print("DEBUG: Nenhum dado na lista para enviar")
                        return False
                    
                    # Chamar enviar_dados que faz todo o processo de validação e salvamento
                    self.enviar_dados()
                    
                    print("DEBUG: Dados enviados com sucesso")
                    return True
                    
                except Exception as e:
                    print(f"DEBUG: Erro ao enviar dados: {str(e)}")
                    return False
            else:
                print("DEBUG: Erro ao adicionar dados à lista")
                return False
            
        except Exception as e:
            print(f"DEBUG: Erro geral ao inserir lançamento da agenda: {str(e)}")
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao inserir lançamento: {str(e)}")
            return False
        
    def validar_dados_basicos_agenda(self, dados):
        """Valida dados vindos da agenda"""
        campos_obrigatorios = ['nome', 'valor', 'data_rel', 'dt_vencto']
        
        for campo in campos_obrigatorios:
            if not dados.get(campo):
                custom_messagebox("error", "Erro", f"Campo '{campo}' é obrigatório!")
                return False
        
        # Validar valor numérico
        try:
            float(dados['valor'])
        except ValueError:
            custom_messagebox("error", "Erro", "Valor deve ser numérico!")
            return False
        
        return True

    def preencher_campos_desde_agenda(self, dados):
        """Preenche os campos do formulário principal com dados da agenda - NOMES CORRETOS"""
        try:
            print("DEBUG: Iniciando preenchimento de campos desde agenda")
            
            # DATA_REL - usar o nome correto do campo
            if dados.get('data_rel'):
                if isinstance(dados['data_rel'], str):
                    data_obj = datetime.strptime(dados['data_rel'], '%d/%m/%Y').date()
                else:
                    data_obj = dados['data_rel']
                
                # CORREÇÃO: usar data_rel_entry (baseado no código adicionar_dados)
                if hasattr(self, 'data_rel_entry'):
                    self.data_rel_entry.set_date(data_obj)
                    print(f"DEBUG: Data_rel preenchida: {data_obj}")
                else:
                    print("DEBUG: Campo data_rel_entry não encontrado")
            
            # TIPO DE DESPESA - CORREÇÃO: usar delete/insert para Entry
            if dados.get('tp_desp'):
                if hasattr(self, 'campos_despesa') and 'tp_desp' in self.campos_despesa:
                    campo_tp_desp = self.campos_despesa['tp_desp']
                    
                    # Verificar se é Combobox ou Entry
                    if hasattr(campo_tp_desp, 'set'):
                        # É um Combobox
                        campo_tp_desp.set(str(dados['tp_desp']))
                        print(f"DEBUG: Tipo despesa preenchido (Combobox): {dados['tp_desp']}")
                    else:
                        # É um Entry
                        campo_tp_desp.delete(0, tk.END)
                        campo_tp_desp.insert(0, str(dados['tp_desp']))
                        print(f"DEBUG: Tipo despesa preenchido (Entry): {dados['tp_desp']}")
                else:
                    print("DEBUG: Campo tp_desp não encontrado em campos_despesa")
            
            # FORNECEDOR - CNPJ/CPF
            if dados.get('cnpj_cpf'):
                if hasattr(self, 'campos_fornecedor') and 'cnpj_cpf' in self.campos_fornecedor:
                    campo_cnpj = self.campos_fornecedor['cnpj_cpf']
                    campo_cnpj.config(state='normal')  # Habilitar temporariamente
                    campo_cnpj.delete(0, tk.END)
                    campo_cnpj.insert(0, dados['cnpj_cpf'])
                    campo_cnpj.config(state='readonly')  # Voltar ao readonly
                    print(f"DEBUG: CNPJ/CPF preenchido: {dados['cnpj_cpf']}")
                    
                    # Tentar buscar fornecedor existente
                    try:
                        self.buscar_fornecedor_por_cnpj_agenda_manual(dados['cnpj_cpf'])
                    except Exception as e:
                        print(f"DEBUG: Erro ao buscar fornecedor: {str(e)}")
                else:
                    print("DEBUG: Campo cnpj_cpf não encontrado em campos_fornecedor")

            # FORNECEDOR - NOME
            if dados.get('nome'):
                if hasattr(self, 'campos_fornecedor') and 'nome' in self.campos_fornecedor:
                    campo_nome = self.campos_fornecedor['nome']
                    campo_nome.config(state='normal')  # Habilitar temporariamente
                    campo_nome.delete(0, tk.END)
                    campo_nome.insert(0, dados['nome'].upper())
                    campo_nome.config(state='readonly')  # Voltar ao readonly
                    print(f"DEBUG: Nome preenchido: {dados['nome']}")
                else:
                    print("DEBUG: Campo nome não encontrado em campos_fornecedor")
            
            # DESPESA - REFERÊNCIA
            if dados.get('referencia'):
                if hasattr(self, 'campos_despesa') and 'referencia' in self.campos_despesa:
                    self.campos_despesa['referencia'].delete(0, tk.END)
                    self.campos_despesa['referencia'].insert(0, dados['referencia'].upper())
                    print(f"DEBUG: Referência preenchida: {dados['referencia']}")
                else:
                    print("DEBUG: Campo referencia não encontrado em campos_despesa")
            
            # DESPESA - NF
            if dados.get('nf'):
                if hasattr(self, 'campos_despesa') and 'nf' in self.campos_despesa:
                    self.campos_despesa['nf'].delete(0, tk.END)
                    self.campos_despesa['nf'].insert(0, dados['nf'].upper())
                    print(f"DEBUG: NF preenchida: {dados['nf']}")
                else:
                    print("DEBUG: Campo nf não encontrado em campos_despesa")
            
            # DESPESA - VALOR UNITÁRIO E DIAS (baseado no método adicionar_dados)
            if dados.get('valor'):
                valor = float(dados['valor'])
                dias = dados.get('dias', 1)
                if isinstance(dias, str):
                    dias = float(dias) if dias else 1
                
                # Calcular valor unitário
                vr_unit = valor / dias if dias > 0 else valor
                
                # Preencher valor unitário
                if hasattr(self, 'campos_despesa') and 'vr_unit' in self.campos_despesa:
                    valor_unit_formatado = f"{vr_unit:.2f}".replace('.', ',')
                    self.campos_despesa['vr_unit'].delete(0, tk.END)
                    self.campos_despesa['vr_unit'].insert(0, valor_unit_formatado)
                    print(f"DEBUG: Valor unitário preenchido: {valor_unit_formatado}")
                
                # Preencher dias
                if hasattr(self, 'campos_despesa') and 'dias' in self.campos_despesa:
                    self.campos_despesa['dias'].delete(0, tk.END)
                    self.campos_despesa['dias'].insert(0, str(int(dias)))
                    print(f"DEBUG: Dias preenchido: {dias}")
                
                # Preencher valor total
                if hasattr(self, 'campos_despesa') and 'valor' in self.campos_despesa:
                    valor_formatado = f"{valor:.2f}".replace('.', ',')
                    campo_valor = self.campos_despesa['valor']
                    campo_valor.config(state='normal')  # Habilitar temporariamente
                    campo_valor.delete(0, tk.END)
                    campo_valor.insert(0, valor_formatado)
                    campo_valor.config(state='readonly')  # Voltar ao readonly
                    print(f"DEBUG: Valor total preenchido: {valor_formatado}")
            
            # DATA DE VENCIMENTO
            if dados.get('dt_vencto'):
                if isinstance(dados['dt_vencto'], str):
                    data_vencto_obj = datetime.strptime(dados['dt_vencto'], '%d/%m/%Y').date()
                else:
                    data_vencto_obj = dados['dt_vencto']
                
                if hasattr(self, 'campos_despesa') and 'dt_vencto' in self.campos_despesa:
                    self.campos_despesa['dt_vencto'].set_date(data_vencto_obj)
                    print(f"DEBUG: Data vencimento preenchida: {data_vencto_obj}")
                else:
                    print("DEBUG: Campo dt_vencto não encontrado em campos_despesa")
            
            # OBSERVAÇÃO
            if dados.get('observacao'):
                if hasattr(self, 'campos_despesa') and 'observacao' in self.campos_despesa:
                    self.campos_despesa['observacao'].delete(0, tk.END)
                    self.campos_despesa['observacao'].insert(0, dados['observacao'].upper())
                    print(f"DEBUG: Observação preenchida: {dados['observacao']}")
                else:
                    print("DEBUG: Campo observacao não encontrado em campos_despesa")
            
            print(f"DEBUG: Preenchimento concluído para {dados.get('nome', 'N/A')}")
            
        except Exception as e:
            print(f"DEBUG: Erro ao preencher campos desde agenda: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def buscar_fornecedor_por_cnpj(self, cnpj_cpf):
        """Busca e seleciona fornecedor pelo CNPJ/CPF"""
        try:
            fornecedor = self.buscar_fornecedor_completo(cnpj_cpf)
            if fornecedor:
                # Preencher dados do fornecedor encontrado
                self.campos_fornecedor['nome'].delete(0, tk.END)
                self.campos_fornecedor['nome'].insert(0, fornecedor['nome'])
                
                if fornecedor.get('categoria'):
                    self.campos_fornecedor['categoria'].delete(0, tk.END)
                    self.campos_fornecedor['categoria'].insert(0, fornecedor['categoria'])
                
                # Preencher dados bancários baseado na forma de pagamento padrão
                if fornecedor.get('chave_pix'):
                    dados_bancarios = f"PIX: {fornecedor['chave_pix']}"
                else:
                    # Construir dados para TED
                    partes_dados = []
                    if fornecedor.get('banco'): partes_dados.append(fornecedor['banco'])
                    if fornecedor.get('op'): partes_dados.append(fornecedor['op'])
                    if fornecedor.get('agencia'): partes_dados.append(fornecedor['agencia'])
                    if fornecedor.get('conta'): partes_dados.append(fornecedor['conta'])
                    partes_dados.append(fornecedor['cnpj_cpf'])
                    dados_bancarios = ' - '.join(partes_dados)
                
                self.campos_despesa['dados_bancarios'].delete(0, tk.END)
                self.campos_despesa['dados_bancarios'].insert(0, dados_bancarios)
                
                print(f"DEBUG: Fornecedor encontrado e dados preenchidos: {fornecedor['nome']}")
                return True
            else:
                print(f"DEBUG: Fornecedor não encontrado para CNPJ/CPF: {cnpj_cpf}")
                return False
                
        except Exception as e:
            print(f"DEBUG: Erro ao buscar fornecedor por CNPJ: {str(e)}")
            return False

    def buscar_fornecedor_por_nome_agenda(self, nome_fornecedor):
        """Busca fornecedor pelo nome para uso na agenda"""
        try:
            if not nome_fornecedor or not nome_fornecedor.strip():
                return None
            
            # Abrir planilha de fornecedores
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
            
            nome_busca = nome_fornecedor.strip().upper()
            fornecedor_encontrado = None
            melhor_match = 0
            
            # Buscar na planilha
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular linhas vazias
                    continue
                    
                cnpj_cpf = str(row[0]).strip()           # Coluna A
                nome = str(row[3] or '').strip().upper() # Coluna D = Nome
                
                # Verificar correspondência
                if nome_busca in nome:
                    # Calcular qualidade do match
                    match_quality = len(nome_busca) / len(nome) if nome else 0
                    
                    # Se for match exato ou melhor que o anterior
                    if nome_busca == nome or match_quality > melhor_match:
                        melhor_match = match_quality
                        fornecedor_encontrado = {
                            'cnpj_cpf': cnpj_cpf,
                            'nome': nome,
                            'categoria': str(row[11] or '').strip(),      # Coluna L = CATEGORIA
                            'telefone': str(row[4] or '').strip(),        # Coluna E = TELEFONE
                            'email': str(row[5] or '').strip(),           # Coluna F = EMAIL
                            'banco': str(row[6] or '').strip(),           # Coluna G = BANCO
                            'op': str(row[7] or '').strip(),              # Coluna H = OP
                            'agencia': str(row[8] or '').strip(),         # Coluna I = AGENC
                            'conta': str(row[9] or '').strip(),           # Coluna J = CONTA
                            'chave_pix': str(row[10] or '').strip(),      # Coluna K = Chave_PIX ← CORRIGIDO!
                            'dados_bancarios': str(row[14] or '').strip() # Coluna O = DADOS BANCÁRIOS
                        }
                        
                        # Se for match exato, parar a busca
                        if nome_busca == nome:
                            break
            
            wb.close()
            
            if fornecedor_encontrado:
                print(f"DEBUG: Fornecedor encontrado: {fornecedor_encontrado['nome']} - {fornecedor_encontrado['cnpj_cpf']}")
                print(f"DEBUG: Chave PIX: {fornecedor_encontrado['chave_pix']}")
                print(f"DEBUG: Dados bancários: {fornecedor_encontrado['dados_bancarios']}")
            else:
                print(f"DEBUG: Nenhum fornecedor encontrado para: {nome_fornecedor}")
            
            return fornecedor_encontrado
            
        except Exception as e:
            print(f"DEBUG: Erro ao buscar fornecedor por nome: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def buscar_fornecedor_por_cnpj_agenda(self, cnpj_cpf):
        """Busca fornecedor pelo CNPJ/CPF para uso na agenda"""
        try:
            if not cnpj_cpf or not cnpj_cpf.strip():
                return None
            
            # Limpar formatação do CNPJ/CPF
            cnpj_limpo = ''.join(filter(str.isdigit, cnpj_cpf))
            
            # Abrir planilha de fornecedores
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
            
            # Buscar na planilha
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular linhas vazias
                    continue
                    
                cnpj_planilha = ''.join(filter(str.isdigit, str(row[0])))
                
                if cnpj_limpo == cnpj_planilha:
                    fornecedor = {
                        'cnpj_cpf': str(row[0]).strip(),              # Coluna A
                        'nome': str(row[3] or '').strip().upper(),    # Coluna D = NOME
                        'categoria': str(row[11] or '').strip(),      # Coluna L = CATEGORIA
                        'telefone': str(row[4] or '').strip(),        # Coluna E = TELEFONE
                        'email': str(row[5] or '').strip(),           # Coluna F = EMAIL
                        'banco': str(row[6] or '').strip(),           # Coluna G = BANCO
                        'op': str(row[7] or '').strip(),              # Coluna H = OP
                        'agencia': str(row[8] or '').strip(),         # Coluna I = AGENC
                        'conta': str(row[9] or '').strip(),           # Coluna J = CONTA
                        'chave_pix': str(row[10] or '').strip(),      # Coluna K = Chave_PIX ← CORRIGIDO!
                        'dados_bancarios': str(row[14] or '').strip() # Coluna O = DADOS BANCÁRIOS
                    }
                    
                    wb.close()
                    print(f"DEBUG: Fornecedor encontrado por CNPJ: {fornecedor['nome']}")
                    print(f"DEBUG: Chave PIX: {fornecedor['chave_pix']}")
                    print(f"DEBUG: Dados bancários: {fornecedor['dados_bancarios']}")
                    return fornecedor
            
            wb.close()
            print(f"DEBUG: Nenhum fornecedor encontrado para CNPJ: {cnpj_cpf}")
            return None
            
        except Exception as e:
            print(f"DEBUG: Erro ao buscar fornecedor por CNPJ: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def obter_dados_bancarios_fornecedor(self, cnpj_cpf, forma_pagamento_preferida="PIX"):
        """Obtém dados bancários formatados do fornecedor"""
        print(f"DEBUG ===== obter_dados_bancarios_fornecedor =====")
        print(f"DEBUG: cnpj_cpf recebido: '{cnpj_cpf}'")
        print(f"DEBUG: forma_pagamento: '{forma_pagamento_preferida}'")
        
        try:
            fornecedor = self.buscar_fornecedor_por_cnpj_agenda(cnpj_cpf)
            
            print(f"DEBUG: Fornecedor retornado: {fornecedor}")
            
            if not fornecedor:
                print(f"DEBUG: Fornecedor não encontrado - retornando vazio")
                return ""
            
            # PRIORIDADE 1: Usar coluna DADOS BANCÁRIOS se já estiver preenchida
            dados_bancarios_coluna = fornecedor.get('dados_bancarios', '').strip()
            if dados_bancarios_coluna:
                print(f"DEBUG: Usando dados da coluna DADOS BANCÁRIOS: '{dados_bancarios_coluna}'")
                return dados_bancarios_coluna
            
            # PRIORIDADE 2: Construir dados baseado na forma de pagamento PIX
            if forma_pagamento_preferida == "PIX":
                chave_pix = fornecedor.get('chave_pix', '').strip()
                if chave_pix:
                    resultado = f"PIX: {chave_pix}"
                    print(f"DEBUG: Construindo dados PIX: '{resultado}'")
                    return resultado
                else:
                    print(f"DEBUG: Chave PIX não cadastrada - retornando vazio")
                    return ""
            
            # PRIORIDADE 3: Construir dados para TED (apenas se houver dados completos)
            partes_dados = []
            
            if fornecedor.get('banco'): 
                partes_dados.append(fornecedor['banco'])
            if fornecedor.get('op'): 
                partes_dados.append(fornecedor['op'])
            if fornecedor.get('agencia'): 
                partes_dados.append(fornecedor['agencia'])
            if fornecedor.get('conta'): 
                partes_dados.append(fornecedor['conta'])
            
            # Só adicionar CNPJ/CPF se houver pelo menos um dado bancário
            if partes_dados:
                partes_dados.append(fornecedor['cnpj_cpf'])
                resultado = ' - '.join(partes_dados)
                print(f"DEBUG: Construindo dados TED: '{resultado}'")
                return resultado
            else:
                # Se não houver nenhum dado bancário, retornar vazio
                print(f"DEBUG: Nenhum dado bancário cadastrado - retornando vazio")
                return ""
            
        except Exception as e:
            print(f"DEBUG: ERRO em obter_dados_bancarios_fornecedor: {str(e)}")
            import traceback
            traceback.print_exc()
            return ""
    
    def buscar_fornecedor_por_cnpj_agenda_manual(self, cnpj_cpf):
        """Método auxiliar para buscar e preencher dados do fornecedor"""
        try:
            # Usar o método que você já tem para buscar fornecedor
            fornecedor = self.buscar_fornecedor_por_cnpj_agenda(cnpj_cpf)
            
            if fornecedor:
                # Preencher campos do fornecedor se encontrado
                if fornecedor.get('nome'):
                    self.campos_fornecedor['nome'].delete(0, tk.END)
                    self.campos_fornecedor['nome'].insert(0, fornecedor['nome'])
                
                if fornecedor.get('categoria'):
                    # Verificar se campo categoria existe
                    if hasattr(self, 'campos_fornecedor') and 'categoria' in self.campos_fornecedor:
                        self.campos_fornecedor['categoria'].delete(0, tk.END)
                        self.campos_fornecedor['categoria'].insert(0, fornecedor['categoria'])
                
                # Preencher dados bancários se existir o campo
                dados_bancarios = self.obter_dados_bancarios_fornecedor(cnpj_cpf)
                if hasattr(self, 'campos_fornecedor') and 'dados_bancarios' in self.campos_fornecedor:
                    campo_dados_bancarios = self.campos_fornecedor['dados_bancarios']
                    campo_dados_bancarios.config(state='normal')  # Habilitar temporariamente
                    campo_dados_bancarios.delete(0, tk.END)
                    campo_dados_bancarios.insert(0, dados_bancarios)
                    campo_dados_bancarios.config(state='readonly')  # Voltar ao readonly
                
                print(f"DEBUG: Dados do fornecedor preenchidos: {fornecedor['nome']}")
                return True
            else:
                print(f"DEBUG: Fornecedor não encontrado para CNPJ/CPF: {cnpj_cpf}")
                return False
                
        except Exception as e:
            print(f"DEBUG: Erro ao buscar fornecedor para agenda: {str(e)}")
            return False
    
    def calcular_valor_total(self, event=None):
        """Calcula o valor total baseado no tipo de despesa"""
        try:
            # Pegar valor unitário
            vr_unit_str = self.campos_despesa['vr_unit'].get().strip()
            if not vr_unit_str:
                self.campos_despesa['valor'].config(state='normal')
                self.campos_despesa['valor'].delete(0, tk.END)
                self.campos_despesa['valor'].config(state='readonly')
                return
                
            vr_unit = float(vr_unit_str.replace(',', '.'))
            
            # Pegar tipo de despesa
            tp_desp = self.campos_despesa['tp_desp'].get()
            
            # Calcular com base no tipo
            if tp_desp == '1':  # Tipo que usa dias
                dias_str = self.campos_despesa['dias'].get().strip()
                dias = float(dias_str.replace(',', '.')) if dias_str else 1
                valor_total = vr_unit * dias
            else:
                valor_total = vr_unit
                
            # Atualizar campo de valor
            self.campos_despesa['valor'].config(state='normal')
            self.campos_despesa['valor'].delete(0, tk.END)
            self.campos_despesa['valor'].insert(0, f"{valor_total:.2f}")
            self.campos_despesa['valor'].config(state='readonly')
            
        except ValueError:
            # Em caso de erro, limpa o campo valor
            self.campos_despesa['valor'].config(state='normal')
            self.campos_despesa['valor'].delete(0, tk.END)
            self.campos_despesa['valor'].config(state='readonly')

    def verificar_tipo_despesa(self, event=None):
        """Verifica o tipo de despesa e ajusta campos conforme necessário"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        
        # Salvar a referência atual antes de qualquer modificação
        referencia_atual = ""
        if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
            referencia_atual = self.campos_despesa['referencia'].get()
        else:
            referencia_atual = self.campos_despesa['referencia'].get() 

        if not tp_desp.isdigit():
            self.campos_despesa['tp_desp'].delete(0, tk.END)
            return

        tp_desp_num = int(tp_desp)
        if not (1 <= tp_desp_num <= 6):
            self.campos_despesa['tp_desp'].delete(0, tk.END)
            return
            
        # Configura o campo dias
        if tp_desp == '1':
            self.campos_despesa['dias'].config(state='normal')
        else:
            self.campos_despesa['dias'].config(state='disabled')
            self.campos_despesa['dias'].delete(0, tk.END)
            self.campos_despesa['dias'].insert(0, '1')

        # Configura o campo nf
        if tp_desp != '1':
            self.campos_despesa['nf'].config(state='normal')
        else:
            self.campos_despesa['nf'].config(state='disabled')
            self.campos_despesa['nf'].delete(0, tk.END)
            
        # Atualiza o campo referência
        self.atualizar_campo_referencia(event)
        
        # NOVO: Restaurar o valor da referência se for uma especificação personalizada
        # e não uma das opções padrão do tipo 1
        if tp_desp != '1' and referencia_atual and referencia_atual not in self.opcoes_referencia_tipo1:
            if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                self.campos_despesa['referencia'].delete(0, tk.END)
                self.campos_despesa['referencia'].insert(0, referencia_atual)
            else:
                self.campos_despesa['referencia'].delete(0, tk.END)
                self.campos_despesa['referencia'].insert(0, referencia_atual)

        # Move para o campo referência
        self.campos_despesa['referencia'].focus()
        
    def adicionar_dados(self, eh_parcelamento=False):
        """Adiciona dados à lista temporária e retorna à aba fornecedor"""
        logger = system_logger.get_logger()
        logger.info(f"Iniciando adição de dados - Cliente: {self.cliente_atual}")

        if not self.validar_campos():
            logger.warning(f"Falha na validação dos campos")
            return False
        
        try:
            # Coleta do primeiro conjunto de dados
            vr_unit_str = self.campos_despesa['vr_unit'].get().strip()
            if not vr_unit_str:
                custom_messagebox("error", "Erro", "Valor unitário é obrigatório!")
                return False
            vr_unit = float(vr_unit_str.replace(',', '.'))
        
            valor_str = self.campos_despesa['valor'].get().strip()
            if not valor_str:
                custom_messagebox("error", "Erro", "Valor total não foi calculado!")
                return False
            valor = float(valor_str.replace(',', '.'))

            # Coletar dados do lançamento
            dados = {
                'data': self.data_rel_entry.get(),
                'cnpj_cpf': self.campos_fornecedor['cnpj_cpf'].get(),
                'nome': self.campos_fornecedor['nome'].get(),
                'categoria': self.campos_fornecedor['categoria'].get().upper(),
                'tp_desp': self.campos_despesa['tp_desp'].get(),
                'referencia': self.campos_despesa['referencia'].get().upper(),
                'etapa_obra': self.campos_despesa['etapa_obra'].get(),
                'insumo': self.campos_despesa['insumo'].get(), 
                'nf': self.campos_despesa['nf'].get().upper(),
                'vr_unit': f"{vr_unit:.2f}",
                'dias': float(self.campos_despesa['dias'].get().replace(',', '.')) if self.campos_despesa['dias'].get() else 1,
                'valor': f"{valor:.2f}",
                'dt_vencto': self.campos_despesa['dt_vencto'].get(),
                'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get(),
                'observacao': self.campos_despesa['observacao'].get().upper(),
                'forma_pagamento': self.forma_pagamento_var.get()
            }
            self.dados_para_incluir.append(dados)

            # Verificar se é um lançamento de TRANSPORTE e criar lançamento automático de CAFÉ
            if dados['tp_desp'] == '1' and dados['referencia'] == 'TRANSPORTE':
                try:
                    # Buscar valor do café nas configurações
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    config = GerenciadorConfiguracoes.carregar_configuracoes()
                    
                    if config and 'cafe' in config and 'valor_atual' in config['cafe']:
                        vr_unit_cafe = float(config['cafe']['valor_atual'])
                    else:
                        vr_unit_cafe = 4.0  # Valor padrão caso não encontre configuração
                        
                    dias_cafe = int(dados['dias'])
                    valor_cafe = vr_unit_cafe * dias_cafe
                    
                    # Criar dados do lançamento do CAFÉ
                    dados_cafe = dados.copy()
                    dados_cafe.update({
                        'referencia': 'CAFÉ',
                        'vr_unit': f"{vr_unit_cafe:.2f}",
                        'valor': f"{valor_cafe:.2f}"
                    })
                    self.dados_para_incluir.append(dados_cafe)
                    custom_messagebox("info", "Informação", 
                        f"Lançamento de CAFÉ adicionado automaticamente com valor de R$ {vr_unit_cafe:.2f} por dia!")
                except Exception as e:
                    custom_messagebox("warning", "Aviso", 
                        f"Erro ao processar lançamento automático do café: {str(e)}\n"
                        "O lançamento principal foi salvo, mas o café não foi gerado.")

            # Só limpa os campos e mostra mensagem se não for parcelamento
            if not eh_parcelamento:
                self.limpar_campos_despesa()
                
                # Limpar campos do fornecedor
                for campo, entry in self.campos_fornecedor.items():
                    entry.config(state='normal')
                    entry.delete(0, tk.END)
                    if campo != 'categoria':
                        entry.config(state='readonly')
                
                # Desmarcar checkbox de materiais
                self.tem_materiais_var.set(False)
                
                custom_messagebox("info", "Sucesso", "Dados adicionados com sucesso!")
                
                # Voltar para a aba fornecedor
                self.notebook.select(1)
                self.tree_fornecedores.selection_remove(self.tree_fornecedores.selection())
                self.busca_entry.delete(0, tk.END)
            
            logger.info(f"Dados adicionados com sucesso - Cliente: {self.cliente_atual}, Total: {len(self.dados_para_incluir)}")
            return True
            
        except ValueError as e:
            logger.error(f"Erro ao processar valores: {str(e)}")
            custom_messagebox("error", "Erro", f"Erro ao processar valores: {str(e)}")
            return False

    def handle_checkbox_change(self):
        """
        Versão modificada que inclui sugestão de importar NFe quando materiais são marcados
        """
        print(f"DEBUG: handle_checkbox_change chamado. tem_materiais_var: {self.tem_materiais_var.get()}")
        
        if self.tem_materiais_var.get():
            # Checkbox foi marcado
            print("DEBUG: Checkbox marcado - materiais vinculados")
            
            # Perguntar se usuário quer importar NFe
            resposta = messagebox.askyesno(
                "Importar NFe", 
                "🏗️ MATERIAIS VINCULADOS!\n\n"
                "Esta despesa possui materiais associados.\n\n"
                "💡 Dica: Você pode importar os materiais diretamente de uma NFe "
                "para o controle de obra, incluindo:\n"
                "• Descrições detalhadas\n"
                "• Quantidades e valores\n"
                "• Dados do fornecedor\n"
                "• Classificação automática\n\n"
                "Deseja importar materiais de uma NFe agora?"
            )
            
            if resposta:
                self.abrir_importacao_nfe_completa()
        else:
            # Checkbox foi desmarcado
            print("DEBUG: Checkbox desmarcado - sem materiais")

    def abrir_importacao_nfe_completa(self):
        """
        Método para abrir importação NFe completa
        """
        try:
            # Verificar se cliente está selecionado
            if not self.cliente_atual:
                messagebox.showerror("Erro", "Selecione um cliente antes de importar NFe!")
                return
            
            # Abrir seletor de arquivo XML
            from tkinter import filedialog
            arquivo_xml = filedialog.askopenfilename(
                title="Selecionar XML da NFe",
                filetypes=[
                    ("Arquivos XML", "*.xml"),
                    ("Todos os arquivos", "*.*")
                ]
            )
            
            if arquivo_xml:
                # Processar XML usando o sistema híbrido existente
                if hasattr(self, 'processador_nfe'):
                    try:
                        dados_nfe = self.processador_nfe.processar_xml_nfe(arquivo_xml)
                        
                        if dados_nfe:
                            # Abrir integrador completo
                            from src.nfe.integrador_nfe_sistema import IntegradorNFeFinanceiroMateriais
                            integrador = IntegradorNFeFinanceiroMateriais(self)
                            integrador.criar_interface_integracao_nfe(dados_nfe)
                        else:
                            messagebox.showerror("Erro", "Erro ao processar XML da NFe!")
                            
                    except Exception as e:
                        messagebox.showerror("Erro", f"Erro ao processar XML:\n{str(e)}")
                else:
                    messagebox.showerror("Erro", "Sistema NFe não inicializado!")
                    
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar NFe:\n{str(e)}")

    def validar_campos(self):
        """Valida os campos antes de adicionar/enviar dados"""
        # Validar data
        if not self.data_rel_entry.get():
            custom_messagebox("error", "Erro", "Data de referência é obrigatória!")
            return False

        # Validar fornecedor
        if not self.campos_fornecedor['cnpj_cpf'].get():
            custom_messagebox("error", "Erro", "Selecione um fornecedor!")
            return False


        # Validar tipo de despesa
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        if not tp_desp or not tp_desp.isdigit() or not (1 <= int(tp_desp) <= 7):
            custom_messagebox("error", "Erro", "Tipo de despesa deve ser um número entre 1 e 7!")
            return False

        # Validar valor unitário
        vr_unit = self.campos_despesa['vr_unit'].get().strip()
        if not vr_unit:
            custom_messagebox("error", "Erro", "Valor unitário é obrigatório!")
            return False
        try:
            float(vr_unit.replace(',', '.'))
        except ValueError:
            custom_messagebox("error", "Erro", "Valor unitário inválido!")
            return False

        # Validar dias para tipo de despesa 1
        if tp_desp == '1':
            dias = self.campos_despesa['dias'].get().strip()
            if not dias:
                custom_messagebox("error", "Erro", "Quantidade de dias é obrigatória para tipo 1!")
                return False
            try:
                dias_float = float(dias.replace(',', '.'))
                if dias_float <= 0:
                    custom_messagebox("error", "Erro", "Quantidade de dias deve ser maior que zero!")
                    return False
            except ValueError:
                custom_messagebox("error", "Erro", "Quantidade de dias inválida!")
                return False

        # Validar referência
        if not self.campos_despesa['referencia'].get().strip():
            custom_messagebox("error", "Erro", "Referência é obrigatória!")
            return False

        # Validar data de vencimento
        if not self.campos_despesa['dt_vencto'].get():
            custom_messagebox("error", "Erro", "Data de vencimento é obrigatória!")
            return False

        return True

    def importar_folha_rh(self):
        """Inicia o processo de importação de dados da folha de RH"""
        # Verificar se um cliente está selecionado
        if not self.cliente_atual:
            if custom_messagebox("yesno", 
                "Importação RH",
                "Nenhum cliente está selecionado. A importação será feita baseada nos dados da planilha RH.\n\n"
                "Deseja continuar?"
            ):
                importador = ImportadorRH(self)
                importador.importar_dados()
        else:
            importador = ImportadorRH(self)
            importador.importar_dados()

    def importar_transporte_cafe(self):
        """Chama a importação de transporte através do ImportadorRH"""
        try:
            importador = ImportadorRH(self)
            importador.importar_transporte_cafe()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir importador de transporte: {str(e)}")

    def limpar_campos_despesa(self):
        """Limpa todos os campos da despesa"""
        campos_para_limpar = ['tp_desp', 'referencia', 'etapa_obra', 'insumo', 'nf', 'vr_unit', 
                            'dias', 'valor', 'observacao']
        
        for campo in campos_para_limpar:
            if campo in self.campos_despesa:
                widget = self.campos_despesa[campo]
                
                # Tratamento especial para campo 'valor' (Entry readonly)
                if campo == 'valor':
                    widget.config(state='normal')
                    widget.delete(0, tk.END)
                    widget.config(state='readonly')
                
                # Tratamento para Combobox readonly (etapa_obra, insumo)
                elif isinstance(widget, ttk.Combobox):
                    widget.set('')
                
                # Tratamento para Entry normais
                elif hasattr(widget, 'delete'):
                    widget.delete(0, tk.END)
        
        # Reinicializar valor padrão para dias
        if 'dias' in self.campos_despesa:
            self.campos_despesa['dias'].insert(0, "1")
        
        # Limpar data de vencimento
        if 'dt_vencto' in self.campos_despesa:
            self.campos_despesa['dt_vencto'].delete(0, tk.END)

    def verificar_duplicidade_antes_salvar(self, sheet, dados):
        """
        Verifica se um lançamento similar já existe na planilha usando critérios inteligentes
        VERSÃO CORRIGIDA - Considera NFs diferentes como NÃO-DUPLICATAS
        """
        # TESTE: Log bem visível
        print("🔍 EXECUTANDO VERIFICAÇÃO DE DUPLICIDADE!")
        print(f"🔍 Dados recebidos: {dados}")
        
        logger = system_logger.get_logger()
        logger.error("🔍 MÉTODO DE VERIFICAÇÃO FOI CHAMADO!")  # Use ERROR para garantir que apareça
        
        try:
            # Normalizar dados para comparação
            nome_novo = str(dados['nome']).strip().upper()
            referencia_nova = str(dados['referencia']).strip().upper()
            nf_nova = str(dados.get('nf', '')).strip().upper()
            dt_vencto_nova = str(dados['dt_vencto']).strip()
            cnpj_novo = str(dados.get('cnpj_cpf', '')).strip()
            
            try:
                valor_novo = float(str(dados['valor']).replace(',', '.'))
            except (ValueError, TypeError):
                logger.error(f"Erro ao converter valor para verificação: {dados['valor']}")
                return False
            
            logger.info(f"=== INICIANDO VERIFICAÇÃO DE DUPLICIDADE ===")
            logger.info(f"Dados novos: {nome_novo} | {referencia_nova} | NF: {nf_nova} | R$ {valor_novo:.2f} | {dt_vencto_nova}")
            
            duplicatas_encontradas = 0
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Pular linhas vazias
                if not row[0]:
                    continue

                # Verificar status - pular se excluído
                status = row[13] if len(row) > 13 else 'ATIVO'
                if status == 'EXCLUIDO':
                    continue
                    
                # Normalizar dados da planilha para comparação
                nome_planilha = str(row[3] or '').strip().upper()  # NOME
                referencia_planilha = str(row[4] or '').strip().upper()  # REFERÊNCIA
                nf_planilha = str(row[5] or '').strip().upper()  # NF
                cnpj_planilha = str(row[2] or '').strip()  # CNPJ_CPF
                
                # Comparar datas de vencimento
                dt_vencto_planilha = ""
                if row[9]:  # DT_VENCTO
                    if isinstance(row[9], datetime):
                        dt_vencto_planilha = row[9].strftime('%d/%m/%Y')
                    else:
                        dt_vencto_planilha = str(row[9]).strip()
                
                # Comparar valores
                try:
                    valor_planilha = float(str(row[8] or 0).replace(',', '.'))  # VALOR
                    diferenca_valor = abs(valor_planilha - valor_novo)
                except (ValueError, TypeError):
                    continue
                
                # LOG para debug
                logger.debug(f"Linha {row_num}: {nome_planilha} | {referencia_planilha} | NF: {nf_planilha} | R$ {valor_planilha:.2f} | {dt_vencto_planilha}")
                
                # =============================================================
                # CRITÉRIO 1: DUPLICATA EXATA POR NF
                # =============================================================
                if (nf_nova and nf_planilha and 
                    nf_nova == nf_planilha and 
                    nome_planilha == nome_novo and 
                    diferenca_valor < 0.01 and 
                    dt_vencto_planilha == dt_vencto_nova):
                    
                    logger.error(f"🚨 DUPLICATA EXATA DETECTADA (Critério: NF)!")
                    logger.error(f"   NF: {nf_nova}")
                    logger.error(f"   Nome: {nome_novo}")
                    logger.error(f"   Valor: R$ {valor_novo:.2f}")
                    logger.error(f"   Vencimento: {dt_vencto_nova}")
                    logger.error(f"   Linha existente: {row_num}")
                    return True
                
                # =============================================================
                # CRITÉRIO 2: MESMO FORNECEDOR + VALOR + DATA + REFERÊNCIA SIMILAR
                # (mas apenas se NFs não existirem OU forem iguais)
                # =============================================================
                if (nome_planilha == nome_novo and 
                    diferenca_valor < 0.01 and 
                    dt_vencto_planilha == dt_vencto_nova):
                    
                    # SE AMBAS AS NFs EXISTEM E SÃO DIFERENTES, NÃO É DUPLICATA
                    if nf_nova and nf_planilha and nf_nova != nf_planilha:
                        logger.debug(f"   ✅ NFs diferentes detectadas ('{nf_nova}' vs '{nf_planilha}') - NÃO é duplicata")
                        continue
                    
                    # Verificar similaridade da referência
                    similaridade = self._calcular_similaridade_simples(referencia_nova, referencia_planilha)
                    
                    logger.debug(f"   Similaridade entre '{referencia_nova}' e '{referencia_planilha}': {similaridade:.2%}")
                    
                    if similaridade >= 0.7:  # 70% de similaridade
                        logger.error(f"🚨 DUPLICATA PROVÁVEL DETECTADA (Critério: Fornecedor+Valor+Data+Referência Similar)!")
                        logger.error(f"   Nome: {nome_novo}")
                        logger.error(f"   Referência nova: '{referencia_nova}'")
                        logger.error(f"   Referência existente: '{referencia_planilha}'")
                        logger.error(f"   Similaridade: {similaridade:.2%}")
                        logger.error(f"   Valor: R$ {valor_novo:.2f}")
                        logger.error(f"   Vencimento: {dt_vencto_nova}")
                        logger.error(f"   NF nova: '{nf_nova}' | NF existente: '{nf_planilha}'")
                        logger.error(f"   Linha existente: {row_num}")
                        return True
                
                # =============================================================
                # CRITÉRIO 3: MESMO CNPJ + VALOR + DATA (independente da referência)
                # (mas apenas se NFs não existirem OU forem iguais)
                # =============================================================
                if (cnpj_novo and cnpj_planilha and 
                    cnpj_novo == cnpj_planilha and 
                    diferenca_valor < 0.01 and 
                    dt_vencto_planilha == dt_vencto_nova):
                    
                    # SE AMBAS AS NFs EXISTEM E SÃO DIFERENTES, NÃO É DUPLICATA
                    if nf_nova and nf_planilha and nf_nova != nf_planilha:
                        logger.debug(f"   ✅ NFs diferentes detectadas ('{nf_nova}' vs '{nf_planilha}') - NÃO é duplicata")
                        continue
                    
                    logger.error(f"🚨 DUPLICATA SUSPEITA DETECTADA (Critério: CNPJ+Valor+Data)!")
                    logger.error(f"   CNPJ: {cnpj_novo}")
                    logger.error(f"   Nome: {nome_novo}")
                    logger.error(f"   Valor: R$ {valor_novo:.2f}")
                    logger.error(f"   Vencimento: {dt_vencto_nova}")
                    logger.error(f"   Referência nova: '{referencia_nova}'")
                    logger.error(f"   Referência existente: '{referencia_planilha}'")
                    logger.error(f"   NF nova: '{nf_nova}' | NF existente: '{nf_planilha}'")
                    logger.error(f"   Linha existente: {row_num}")
                    return True
            
            logger.info(f"✅ Nenhuma duplicata encontrada para: {nome_novo} - {referencia_nova} - NF: {nf_nova}")
            return False
            
        except Exception as e:
            logger.error(f"❌ ERRO na verificação de duplicidade: {str(e)}", exc_info=True)
            return False
        
        finally:
            if hasattr(self, '_is_saving'):
                self._is_saving = False

    def _calcular_similaridade_simples(self, texto1, texto2):
        """
        Calcula similaridade simples entre dois textos
        """
        if not texto1 or not texto2:
            return 0.0
        
        # Converter para minúsculas e remover espaços extras
        t1 = ' '.join(texto1.lower().split())
        t2 = ' '.join(texto2.lower().split())
        
        # Se um texto está contido no outro, alta similaridade
        if t1 in t2 or t2 in t1:
            return 0.9
        
        # Calcular similaridade por caracteres comuns
        comum = sum(1 for c in t1 if c in t2)
        total = max(len(t1), len(t2))
        
        if total == 0:
            return 0.0
        
        return comum / total
    # Correção para o método enviar_dados() - SistemaEntradaDados.py

    def enviar_dados(self):
        """
        Versão OTIMIZADA com gestão inteligente de duplicatas
        - Remove automaticamente duplicatas da visualização
        - Mantém apenas registros únicos pendentes
        - Protege contra envios para cliente errado
        """
        logger = system_logger.get_logger()
        logger.info(f"Iniciando envio de dados - Cliente: {self.cliente_atual}, Registros: {len(self.dados_para_incluir) if self.dados_para_incluir else 0}")
        
        # Desabilitar botão para evitar múltiplos cliques
        btn_enviar = None
        for aba in [self.aba_dados, self.aba_fornecedor]:
            for child in aba.winfo_children():
                if isinstance(child, ttk.Frame):
                    for widget in child.winfo_children():
                        if isinstance(widget, ttk.Button) and widget['text'] == "Enviar":
                            btn_enviar = widget
                            break
        
        if btn_enviar:
            btn_enviar.config(state='disabled')
        
        datas_recalculadas = []
        
        try:
            # ==========================================
            # VALIDAÇÃO 1: Cliente selecionado
            # ==========================================
            if not self.cliente_atual:
                custom_messagebox("error", "Erro", "Selecione um cliente!")
                return
            
            # ==========================================
            # VALIDAÇÃO 2: Verificar NF com materiais
            # ==========================================
            nf_numero = self.campos_despesa['nf'].get().strip()
            tem_materiais_marcado = self.tem_materiais_var.get()
            
            if nf_numero and tem_materiais_marcado:
                if not self.verificar_nf_ja_processada(nf_numero):
                    resposta = custom_messagebox(
                        "yesno",
                        "NFe Detectada",
                        f"NOTA FISCAL DETECTADA!\n\n"
                        f"NF: {nf_numero}\n"
                        f"Materiais marcados como vinculados\n\n"
                        f"SUGESTÃO: Para melhor controle, importe os materiais "
                        f"diretamente do XML da NFe. Isso permitirá:\n\n"
                        f"• Controle detalhado de cada item\n"
                        f"• Valores unitários e totais precisos\n"
                        f"• Classificação automática por categoria\n"
                        f"• Controle de localização na obra\n"
                        f"• Preenchimento automático dos dados\n\n"
                        f"Deseja importar esta NFe agora?\n"
                        f"(Clique 'Não' para continuar com o lançamento manual)"
                    )
                    
                    if resposta:
                        if self.localizar_e_processar_nfe(nf_numero):
                            return
            
            # ==========================================
            # OBTER DADOS PARA PROCESSAR
            # ==========================================
            dados_para_processar = []
            if hasattr(self, 'visualizador') and self.visualizador and self.visualizador.tree.winfo_exists():
                logger.info("Obtendo dados do visualizador")
                dados_para_processar = self.visualizador.get_dados_atualizados()
            elif self.dados_para_incluir:
                logger.info("Usando dados_para_incluir existentes")
                dados_para_processar = self.dados_para_incluir.copy()
                    
            if not dados_para_processar:
                custom_messagebox("warning", "Aviso", "Não há dados para enviar!")
                return

            logger.info(f"Total de registros a processar: {len(dados_para_processar)}")

            # ==========================================
            # GARANTIR CAMPOS OBRIGATÓRIOS
            # ==========================================
            for lancamento in dados_para_processar:
                if 'etapa_obra' not in lancamento:
                    lancamento['etapa_obra'] = ''
                if 'insumo' not in lancamento:
                    lancamento['insumo'] = ''

            # ==========================================
            # CAPTURAR DATAS AFETADAS
            # ==========================================
            datas_afetadas = set()
            for lancamento in dados_para_processar:
                try:
                    data_rel = datetime.strptime(lancamento['data'], '%d/%m/%Y').date()
                    datas_afetadas.add(data_rel)
                except:
                    continue
            
            logger.info(f"Datas que serão afetadas pela inserção: {[d.strftime('%d/%m/%Y') for d in datas_afetadas]}")

            # ==========================================
            # ADICIONAR IDENTIFICADORES ÚNICOS
            # ==========================================
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            for i, lancamento in enumerate(dados_para_processar):
                lancamento['id'] = f"{timestamp}-{i}"

            arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
            logger.info(f"Salvando em: {arquivo_cliente}")
            
            # ==========================================
            # VERIFICAÇÃO DE DUPLICATAS
            # ==========================================
            try:
                self.verificar_e_corrigir_ids_antes_insercao(arquivo_cliente)
                
                workbook = load_workbook(arquivo_cliente)
                sheet = workbook["Dados"]

                self.verificar_e_adicionar_cabecalho_etapa_obra(sheet)
                self.verificar_e_adicionar_cabecalho_insumo(sheet)
                
                # Separar lançamentos válidos e duplicados
                lancamentos_duplicados = []
                lancamentos_validos = []
                
                for dados in dados_para_processar:
                    if self.verificar_duplicidade_antes_salvar(sheet, dados):
                        lancamentos_duplicados.append(dados)
                        logger.warning(f"Duplicata detectada: {dados['nome']} - {dados['referencia']} - R$ {dados['valor']}")
                    else:
                        lancamentos_validos.append(dados)

                # ==========================================
                # TRATAMENTO INTELIGENTE DE DUPLICATAS
                # ==========================================
                if lancamentos_duplicados:
                    logger.warning(f"Detectados {len(lancamentos_duplicados)} possíveis lançamentos duplicados")
                    
                    # Construir mensagem detalhada
                    msg_duplicados = f"DUPLICATAS DETECTADAS!\n\n"
                    msg_duplicados += f"Foram encontrados {len(lancamentos_duplicados)} possíveis lançamentos duplicados:\n\n"
                    
                    for i, dados in enumerate(lancamentos_duplicados[:5], 1):
                        msg_duplicados += f"{i}. {dados['nome']}\n"
                        msg_duplicados += f"   {dados['referencia']}\n"
                        msg_duplicados += f"   R$ {dados['valor']}\n"
                        msg_duplicados += f"   Vencimento: {dados['dt_vencto']}\n"
                        if dados.get('nf'):
                            msg_duplicados += f"   NF: {dados['nf']}\n"
                        msg_duplicados += "\n"
                    
                    if len(lancamentos_duplicados) > 5:
                        msg_duplicados += f"... e mais {len(lancamentos_duplicados) - 5} duplicatas.\n\n"
                    
                    msg_duplicados += f"Resumo:\n"
                    msg_duplicados += f"• Lançamentos únicos: {len(lancamentos_validos)}\n"
                    msg_duplicados += f"• Possíveis duplicatas: {len(lancamentos_duplicados)}\n\n"
                    
                    # NOVA OPÇÃO: Se todos são duplicatas
                    if len(lancamentos_validos) == 0:
                        msg_duplicados += "TODOS os lançamentos são duplicatas!\n\n"
                        msg_duplicados += "O que deseja fazer?\n\n"
                        msg_duplicados += "• SIM = Cancelar operação e limpar visualização\n"
                        msg_duplicados += "• NÃO = Manter na visualização para revisar"
                        
                        resposta = custom_messagebox("yesno", "Todas Duplicatas", msg_duplicados)
                        
                        workbook.close()
                        
                        if resposta:  # SIM = Limpar tudo
                            logger.info("Usuário optou por limpar todos os lançamentos duplicados")
                            self.limpar_visualizacao_completa()
                            custom_messagebox("info", "Limpeza Concluída", 
                                            "Todos os lançamentos duplicados foram removidos da visualização.\n\n"
                                            "Nenhum dado foi salvo na planilha.")
                        else:  # NÃO = Manter para revisão
                            logger.info("Usuário optou por manter lançamentos para revisão")
                            custom_messagebox("info", "Mantido para Revisão", 
                                            "Os lançamentos foram mantidos na visualização para que você possa:\n\n"
                                            "• Revisar individualmente cada um\n"
                                            "• Remover duplicatas manualmente\n"
                                            "• Editar se necessário\n\n"
                                            "Nenhum dado foi salvo na planilha.")
                        return
                    
                    # Se há MIX de únicos e duplicados
                    msg_duplicados += "Deseja continuar salvando APENAS os lançamentos únicos?\n\n"
                    msg_duplicados += "• SIM = Salvar únicos e remover duplicatas da visualização\n"
                    msg_duplicados += "• NÃO = Cancelar e manter todos para revisar"
                    
                    resposta = custom_messagebox("yesno", "Duplicatas Detectadas", msg_duplicados)
                    
                    if not resposta:  # Usuário cancelou
                        logger.info("Operação cancelada pelo usuário devido a duplicatas")
                        workbook.close()
                        custom_messagebox("info", "Operação Cancelada", 
                                        "Operação cancelada. Os lançamentos permanecem na visualização para revisão.\n\n"
                                        "Você pode:\n"
                                        "• Remover as duplicatas manualmente\n"
                                        "• Editar os registros\n"
                                        "• Tentar enviar novamente")
                        return
                    
                    # Usuário confirmou: salvar únicos e limpar duplicatas
                    dados_para_processar = lancamentos_validos
                    logger.info(f"Usuário optou por continuar. Processando {len(dados_para_processar)} lançamentos únicos")
                    
                    # IMPORTANTE: Já marcar para limpeza posterior
                    self._duplicatas_para_remover = lancamentos_duplicados

                # ==========================================
                # REMOVER TABELAS EXISTENTES
                # ==========================================
                if sheet.tables:
                    table_name = list(sheet.tables.keys())[0]
                    sheet.tables.pop(table_name)
                    
                # ==========================================
                # PROCESSAR REGISTROS VÁLIDOS
                # ==========================================
                registros_salvos = 0
                for dados in dados_para_processar:
                    try:
                        proxima_linha = sheet.max_row + 1
                        
                        # Data de referência
                        data_rel = datetime.strptime(dados['data'], '%d/%m/%Y')
                        data_cell = sheet.cell(row=proxima_linha, column=1, value=data_rel)
                        data_cell.number_format = 'DD/MM/YYYY'

                        # Tipo de despesa
                        tp_desp_cell = sheet.cell(row=proxima_linha, column=2, value=int(dados['tp_desp']))
                        tp_desp_cell.number_format = '0'

                        # Dados gerais
                        sheet.cell(row=proxima_linha, column=3, value=dados['cnpj_cpf'])
                        sheet.cell(row=proxima_linha, column=4, value=dados['nome'])
                        sheet.cell(row=proxima_linha, column=5, value=dados['referencia'])
                        sheet.cell(row=proxima_linha, column=6, value=dados['nf'])

                        # Valores numéricos
                        vr_unit = float(dados['vr_unit'].replace(',', '.'))
                        vr_unit_cell = sheet.cell(row=proxima_linha, column=7, value=vr_unit)
                        aplicar_formatacao_celula(vr_unit_cell)

                        sheet.cell(row=proxima_linha, column=8, value=int(dados.get('dias', 1)))

                        valor = float(dados['valor'].replace(',', '.'))
                        valor_cell = sheet.cell(row=proxima_linha, column=9, value=valor)
                        aplicar_formatacao_celula(valor_cell)

                        dt_vencto = datetime.strptime(dados['dt_vencto'], '%d/%m/%Y')
                        dt_vencto_cell = sheet.cell(row=proxima_linha, column=10, value=dt_vencto)
                        dt_vencto_cell.number_format = 'DD/MM/YYYY'

                        sheet.cell(row=proxima_linha, column=11, value=dados['categoria'])
                        sheet.cell(row=proxima_linha, column=12, value=dados['dados_bancarios'])
                        sheet.cell(row=proxima_linha, column=13, value=dados['observacao'])
                        
                        # ID sequencial
                        novo_id = self.obter_proximo_id_sequencial(sheet)
                        sheet.cell(row=proxima_linha, column=14, value='ATIVO')
                        sheet.cell(row=proxima_linha, column=15, value=novo_id)

                        # Campos adicionais
                        sheet.cell(row=proxima_linha, column=17, value=dados.get('etapa_obra', ''))
                        sheet.cell(row=proxima_linha, column=18, value=dados.get('insumo', ''))

                        logger.info(f"Lançamento inserido com ID {novo_id} na linha {proxima_linha}")
                        
                        registros_salvos += 1
                        
                    except Exception as e:
                        logger.error(f"Erro ao processar registro {dados.get('nome', 'DESCONHECIDO')}: {str(e)}")
                        continue

                # ==========================================
                # SALVAR ARQUIVO E RECALCULAR
                # ==========================================
                try:
                    workbook.save(arquivo_cliente)
                    
                    if registros_salvos > 0:
                        logger.info(f"Iniciando verificação de recálculo para {len(datas_afetadas)} datas")
                        
                        import time
                        time.sleep(0.5)
                        
                        for data_rel in datas_afetadas:
                            try:
                                logger.info(f"Verificando necessidade de recálculo para {data_rel}")
                                
                                resultado = self.chamar_apos_operacao_lancamento(data_rel, "INCLUSAO")
                                
                                if resultado["sucesso"]:
                                    if "taxas recalculadas" in resultado["mensagem"]:
                                        datas_recalculadas.append(data_rel.strftime('%d/%m/%Y'))
                                        logger.info(f"Taxas recalculadas para {data_rel}: {resultado['mensagem']}")
                                    else:
                                        logger.info(f"{data_rel}: {resultado['mensagem']}")
                                else:
                                    logger.warning(f"Erro no recálculo para {data_rel}: {resultado['mensagem']}")
                                    
                            except Exception as e:
                                logger.error(f"Erro ao verificar recálculo para {data_rel}: {str(e)}")
                                continue
                    
                    # ==========================================
                    # MENSAGEM DE SUCESSO E LIMPEZA
                    # ==========================================
                    if lancamentos_duplicados:
                        mensagem_sucesso = f"Dados salvos com sucesso!\n\n"
                        mensagem_sucesso += f"Resumo da operação:\n"
                        mensagem_sucesso += f"• Lançamentos salvos: {registros_salvos}\n"
                        mensagem_sucesso += f"• Duplicatas ignoradas: {len(lancamentos_duplicados)}\n\n"
                        if datas_recalculadas:
                            mensagem_sucesso += f"Taxas recalculadas para: {', '.join(datas_recalculadas)}\n\n"
                        mensagem_sucesso += f"As duplicatas foram automaticamente removidas da visualização."
                        custom_messagebox("info", "Sucesso com Filtro", mensagem_sucesso)
                        
                        # LIMPAR DUPLICATAS DA VISUALIZAÇÃO
                        self.remover_duplicatas_da_visualizacao(lancamentos_duplicados)
                    else:
                        mensagem_sucesso = f"Dados salvos com sucesso! {registros_salvos} lançamentos processados."
                        if datas_recalculadas:
                            mensagem_sucesso += f"\n\nTaxas recalculadas para: {', '.join(datas_recalculadas)}"
                        custom_messagebox("info", "Sucesso", mensagem_sucesso)
                    
                    self.limpar_backup()

                    # ==========================================
                    # LIMPAR APENAS OS DADOS SALVOS
                    # ==========================================
                    if registros_salvos > 0:
                        self.limpar_dados_salvos(dados_para_processar)
                    
                except PermissionError:
                    logger.error("Permissão negada ao salvar arquivo - provável arquivo aberto")
                    custom_messagebox("error", 
                        "Erro", 
                        f"Não foi possível salvar! A planilha '{self.cliente_atual}.xlsx' está aberta.\n\n"
                        "Por favor:\n"
                        "1. Feche a planilha\n"
                        "2. Clique em OK\n"
                        "3. Tente enviar novamente"
                    )
                except Exception as e:
                    logger.error(f"Erro ao salvar arquivo: {str(e)}")
                    custom_messagebox("error", "Erro", f"Erro ao salvar arquivo: {str(e)}")
                            
            except Exception as e:
                logger.error(f"Erro ao processar dados: {str(e)}", exc_info=True)
                custom_messagebox("error", "Erro", f"Erro ao processar dados: {str(e)}")
                
        except Exception as e:
            logger.error(f"Erro geral no método enviar_dados: {str(e)}", exc_info=True)
            custom_messagebox("error", "Erro", f"Erro ao enviar dados: {str(e)}")
            
        finally:
            # Reabilitar botão
            if btn_enviar:
                btn_enviar.config(state='normal')
            
            if hasattr(self, '_is_saving'):
                self._is_saving = False


    # ==========================================
    # MÉTODOS AUXILIARES PARA GESTÃO DE DUPLICATAS
    # ==========================================

    def limpar_visualizacao_completa(self):
        """
        Limpa completamente a visualização de lançamentos pendentes
        Usado quando usuário confirma descarte ao trocar cliente
        """
        try:
            logger = system_logger.get_logger()
            logger.info("Limpando visualização completa de lançamentos pendentes")
            
            # Limpar lista principal
            if hasattr(self, 'dados_para_incluir'):
                qtd_antes = len(self.dados_para_incluir)
                self.dados_para_incluir.clear()
                logger.info(f"Lista dados_para_incluir limpa ({qtd_antes} registros removidos)")
            
            # Fechar e destruir visualizador se existir
            if hasattr(self, 'visualizador') and self.visualizador:
                if hasattr(self.visualizador, 'janela') and self.visualizador.janela:
                    try:
                        self.visualizador.janela.destroy()
                        logger.info("Janela do visualizador destruída")
                    except Exception as e:
                        logger.warning(f"Erro ao destruir janela do visualizador: {str(e)}")
                self.visualizador = None
                logger.info("Referência ao visualizador removida")
            
            logger.info("Visualização limpa com sucesso")
            
        except Exception as e:
            logger = system_logger.get_logger()
            logger.error(f"Erro ao limpar visualização completa: {str(e)}")

    def remover_duplicatas_da_visualizacao(self, lancamentos_duplicados):
        """
        Remove APENAS os lançamentos duplicados da visualização
        Mantém os lançamentos únicos que não foram salvos
        """
        try:
            logger = system_logger.get_logger()
            logger.info(f"Removendo {len(lancamentos_duplicados)} duplicatas da visualização")
            
            # Criar set de IDs das duplicatas para busca rápida
            ids_duplicatas = set()
            for dup in lancamentos_duplicados:
                # Criar chave única baseada nos campos principais
                chave = f"{dup['nome']}|{dup['referencia']}|{dup['valor']}|{dup['dt_vencto']}"
                ids_duplicatas.add(chave)
            
            # Filtrar dados_para_incluir removendo duplicatas
            dados_filtrados = []
            for dados in self.dados_para_incluir:
                chave = f"{dados['nome']}|{dados['referencia']}|{dados['valor']}|{dados['dt_vencto']}"
                if chave not in ids_duplicatas:
                    dados_filtrados.append(dados)
            
            self.dados_para_incluir = dados_filtrados
            
            # Atualizar visualizador se existir
            if hasattr(self, 'visualizador') and self.visualizador:
                if hasattr(self.visualizador, 'tree') and self.visualizador.tree.winfo_exists():
                    # Limpar tree
                    for item in self.visualizador.tree.get_children():
                        self.visualizador.tree.delete(item)
                    
                    # Repopular apenas com dados filtrados
                    if dados_filtrados:
                        self.visualizador.popular_tree(dados_filtrados)
                        logger.info(f"Visualizador atualizado com {len(dados_filtrados)} registros únicos")
                    else:
                        # Se não sobrou nada, fechar visualizador
                        self.visualizador.janela.destroy()
                        self.visualizador = None
                        logger.info("Visualizador fechado - nenhum registro único restante")
            
            logger.info("Duplicatas removidas com sucesso")
            
        except Exception as e:
            logger.error(f"Erro ao remover duplicatas da visualização: {str(e)}")

    def limpar_dados_salvos(self, dados_salvos):
        """
        Remove da visualização APENAS os dados que foram efetivamente salvos
        Mantém dados que ainda não foram processados
        """
        try:
            logger = system_logger.get_logger()
            logger.info(f"Limpando {len(dados_salvos)} registros salvos da visualização")
            
            # Criar set de IDs dos dados salvos
            ids_salvos = set()
            for dados in dados_salvos:
                if 'id' in dados:
                    ids_salvos.add(dados['id'])
            
            # Filtrar dados_para_incluir
            dados_restantes = [d for d in self.dados_para_incluir if d.get('id') not in ids_salvos]
            
            self.dados_para_incluir = dados_restantes
            
            # Atualizar ou fechar visualizador
            if hasattr(self, 'visualizador') and self.visualizador:
                if dados_restantes:
                    # Ainda há dados - atualizar visualização
                    if hasattr(self.visualizador, 'tree') and self.visualizador.tree.winfo_exists():
                        for item in self.visualizador.tree.get_children():
                            self.visualizador.tree.delete(item)
                        self.visualizador.popular_tree(dados_restantes)
                        logger.info(f"Visualizador atualizado com {len(dados_restantes)} registros pendentes")
                else:
                    # Não há mais dados - fechar visualizador
                    self.visualizador.janela.destroy()
                    self.visualizador = None
                    logger.info("Visualizador fechado - todos os registros foram processados")
            
            logger.info("Dados salvos removidos com sucesso da visualização")
            
        except Exception as e:
            logger.error(f"Erro ao limpar dados salvos: {str(e)}")

    def limpar_dados_ao_trocar_cliente(self):
        """
        PROTEÇÃO: Limpa dados pendentes ao trocar de cliente
        Deve ser chamado no método de seleção de cliente
        """
        try:
            if hasattr(self, 'dados_para_incluir') and self.dados_para_incluir:
                qtd_pendentes = len(self.dados_para_incluir)
                
                resposta = custom_messagebox(
                    "yesno",
                    "Dados Pendentes",
                    f"ATENÇÃO: Existem {qtd_pendentes} lançamentos pendentes de envio!\n\n"
                    f"Ao trocar de cliente, estes dados serão perdidos.\n\n"
                    f"Deseja realmente trocar de cliente agora?\n\n"
                    f"• SIM = Trocar e descartar lançamentos pendentes\n"
                    f"• NÃO = Cancelar troca e enviar os lançamentos primeiro"
                )
                
                if not resposta:
                    return False  # Cancelar troca de cliente
                
                # Confirmar descarte
                self.limpar_visualizacao_completa()
                
                logger = system_logger.get_logger()
                logger.warning(f"Dados pendentes descartados ao trocar de cliente: {qtd_pendentes} registros")
            
            return True  # Permitir troca de cliente
            
        except Exception as e:
            logger = system_logger.get_logger()
            logger.error(f"Erro ao limpar dados na troca de cliente: {str(e)}")
            return True  # Em caso de erro, permitir troca
    
    def verificar_e_adicionar_cabecalho_etapa_obra(self, sheet):
        """
        Verifica se o cabeçalho da coluna Etapa da Obra existe e adiciona se necessário
        """
        try:
            # Verificar se existe cabeçalho na linha 1, coluna 17
            cabecalho_atual = sheet.cell(row=1, column=17).value
            
            if cabecalho_atual is None or cabecalho_atual == "":
                # Adicionar cabeçalho para Etapa da Obra
                sheet.cell(row=1, column=17, value="ETAPA_OBRA")
                logger.info("Cabeçalho 'ETAPA_OBRA' adicionado na coluna Q (17)")
                
            return True
            
        except Exception as e:
            logger.error(f"Erro ao verificar/adicionar cabeçalho de etapa da obra: {str(e)}")
            return False

    def verificar_e_adicionar_cabecalho_insumo(self, sheet):
        """
        Verifica se o cabeçalho da coluna Insumo existe e adiciona se necessário
        """
        try:
            # Verificar se existe cabeçalho na linha 1, coluna 18
            cabecalho_atual = sheet.cell(row=1, column=18).value
            
            if cabecalho_atual is None or cabecalho_atual == "":
                # Adicionar cabeçalho para Insumo
                sheet.cell(row=1, column=18, value="INSUMO")
                logger.info("Cabeçalho 'INSUMO' adicionado na coluna R (18)")
                
            return True
            
        except Exception as e:
            logger.error(f"Erro ao verificar/adicionar cabeçalho de insumo: {str(e)}")
            return False
        
    def verificar_e_corrigir_ids_antes_insercao(self, arquivo_cliente):
        """
        Verifica e corrige IDs duplicados ANTES de inserir novos lançamentos
        """
        try:
            logger = system_logger.get_logger()
            logger.info("Verificando integridade dos IDs antes da inserção")
            
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            # Verificar se coluna ID_LANCAMENTO existe
            if ws.cell(row=1, column=15).value != 'ID_LANCAMENTO':
                ws.cell(row=1, column=15, value='ID_LANCAMENTO')
                logger.info("Coluna ID_LANCAMENTO criada")
            
            # Coletar todos os IDs existentes
            ids_existentes = {}
            linhas_com_id_invalido = []
            
            for row in range(2, ws.max_row + 1):
                id_atual = ws.cell(row=row, column=15).value
                
                if id_atual is None or id_atual == '':
                    linhas_com_id_invalido.append(row)
                else:
                    try:
                        id_int = int(float(id_atual))
                        if id_int in ids_existentes:
                            # ID duplicado encontrado
                            logger.warning(f"ID duplicado {id_int} nas linhas {ids_existentes[id_int]} e {row}")
                            linhas_com_id_invalido.append(row)
                        else:
                            ids_existentes[id_int] = row
                    except (ValueError, TypeError):
                        logger.warning(f"ID inválido na linha {row}: {id_atual}")
                        linhas_com_id_invalido.append(row)
            
            # Se há IDs inválidos, corrigir
            if linhas_com_id_invalido:
                logger.info(f"Corrigindo {len(linhas_com_id_invalido)} IDs inválidos/duplicados")
                
                # Encontrar próximo ID disponível
                proximo_id = max(ids_existentes.keys()) + 1 if ids_existentes else 1
                
                for linha in linhas_com_id_invalido:
                    ws.cell(row=linha, column=15, value=proximo_id)
                    logger.info(f"ID {proximo_id} atribuído à linha {linha}")
                    proximo_id += 1
                
                # Salvar correções
                wb.save(arquivo_cliente)
                logger.info("Correções de ID salvas com sucesso")
            
            wb.close()
            
        except Exception as e:
            logger.error(f"Erro ao verificar IDs: {str(e)}")

    # ===== MÉTODO AUXILIAR NOVO: Obter próximo ID sequencial =====
    def obter_proximo_id_sequencial(self, worksheet):
        """
        Obtém o próximo ID sequencial disponível
        """
        try:
            max_id = 0
            
            # Percorrer coluna 15 (ID_LANCAMENTO) para encontrar o maior ID
            for row in range(2, worksheet.max_row + 1):
                id_valor = worksheet.cell(row=row, column=15).value
                if id_valor is not None:
                    try:
                        id_int = int(float(id_valor))
                        if id_int > max_id:
                            max_id = id_int
                    except (ValueError, TypeError):
                        continue
            
            return max_id + 1
            
        except Exception as e:
            print(f"Erro ao obter próximo ID: {str(e)}")
            # Fallback: usar número da linha como ID
            return worksheet.max_row

    # ========== MÉTODOS AUXILIARES NECESSÁRIOS NFe==========
    def verificar_nf_ja_processada(self, numero_nf):
        """
        Verifica se uma NF já foi processada via integrador
        """
        try:
            # Verificar no sistema de materiais se há materiais desta NF
            if hasattr(self, 'gerenciador_materiais'):
                df_materiais = self.gerenciador_materiais.carregar_materiais_cliente(self.cliente_atual)
                if not df_materiais.empty and 'Numero_NF' in df_materiais.columns:
                    nfs_existentes = df_materiais['Numero_NF'].dropna().astype(str).str.strip()
                    return numero_nf in nfs_existentes.values
            return False
        except Exception as e:
            print(f"Erro ao verificar NF processada: {e}")
            return False

    def localizar_e_processar_nfe(self, numero_nf):
        """
        Tenta localizar e processar NFe por número
        """
        try:
            from tkinter import filedialog
            
            # Solicitar localização do XML
            custom_messagebox(
                "info",
                "Localizar XML", 
                f"📁 LOCALIZAR XML DA NFe\n\n"
                f"📋 NF: {numero_nf}\n\n"
                f"Na próxima janela, selecione o arquivo XML "
                f"correspondente a esta nota fiscal."
            )
            
            arquivo_xml = filedialog.askopenfilename(
                title=f"Selecionar XML da NFe {numero_nf}",
                filetypes=[
                    ("Arquivos XML", "*.xml"),
                    ("Todos os arquivos", "*.*")
                ]
            )
            
            if arquivo_xml:
                # Verificar se o XML corresponde à NF informada
                if hasattr(self, 'processador_nfe'):
                    dados_nfe = self.processador_nfe.processar_xml_nfe(arquivo_xml)
                    
                    if dados_nfe and dados_nfe.get('numero_nf') == numero_nf:
                        # XML correto, abrir integrador
                        from src.nfe.integrador_nfe_sistema import IntegradorNFeFinanceiroMateriais
                        integrador = IntegradorNFeFinanceiroMateriais(self)
                        integrador.criar_interface_integracao_nfe(dados_nfe)
                        return True
                    else:
                        nf_encontrada = dados_nfe.get('numero_nf', 'não identificada') if dados_nfe else 'erro ao ler arquivo'
                        custom_messagebox(
                            "error",
                            "Arquivo Incorreto", 
                            f"❌ ARQUIVO INCORRETO!\n\n"
                            f"O XML selecionado não corresponde à NF {numero_nf}.\n\n"
                            f"📋 NF esperada: {numero_nf}\n"
                            f"📄 NF do arquivo: {nf_encontrada}\n\n"
                            f"Selecione o arquivo XML correto ou continue "
                            f"com o lançamento manual."
                        )
                else:
                    custom_messagebox(
                        "error",
                        "Sistema NFe", 
                        "❌ Sistema NFe não inicializado!\n\n"
                        "O processador de NFe não está disponível. "
                        "Continue com o lançamento manual."
                    )
            
            return False
            
        except Exception as e:
            logger = system_logger.get_logger()
            logger.error(f"Erro ao processar NFe: {str(e)}")
            custom_messagebox("error", "Erro", f"Erro ao processar NFe:\n{str(e)}")
            return False
    
    # ===== FUNÇÕES UTILITÁRIAS PARA VERIFICAÇÃO E RECÁLCULO =====
    def abrir_finalizacao_quinzena(self):
        """Abre a finalização de quinzena"""
        try:
            from src.finalizacao_quinzena import FinalizacaoQuinzena
            app = FinalizacaoQuinzena(self.root)
            self.root.withdraw()  # Esconde menu principal
            app.run()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir finalização: {str(e)}")


    def verificar_consistencia_taxas(self, data_referencia=None):
        """
        Método para verificar consistência das taxas de administração
        """
        if not self.cliente_atual:
            return "Nenhum cliente selecionado"
        
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
            
            if not os.path.exists(arquivo_cliente):
                return "Arquivo do cliente não encontrado"
            
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")
            
            # Se não especificou data, usar todas as datas com taxas
            if data_referencia is None:
                df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
                datas_com_taxas = set()
                
                # Identificar datas que têm taxas
                for _, row in df.iterrows():
                    tp_desp = row.get('TP_DESP', 0)
                    if tp_desp == 7:  # Tipo específico para taxas de administração
                        if pd.notna(row['DATA_REL']):
                            datas_com_taxas.add(row['DATA_REL'].date())
                
                if not datas_com_taxas:
                    return "Nenhuma taxa de administração encontrada"
                
                # Verificar cada data
                relatorio = []
                relatorio.append(f"Cliente: {self.cliente_atual}")
                relatorio.append(f"Total de datas com taxas: {len(datas_com_taxas)}")
                relatorio.append("-" * 50)
                
                try:
                    gestor_taxas = GestorTaxasAdministracao(self)
                    for data in sorted(datas_com_taxas):
                        try:
                            precisa, motivo = gestor_taxas.verificar_necessidade_recalculo(data)
                            status = "❌ INCONSISTENTE" if precisa else "✅ OK"
                            relatorio.append(f"{data.strftime('%d/%m/%Y')}: {status} - {motivo}")
                        except Exception as e:
                            relatorio.append(f"{data.strftime('%d/%m/%Y')}: ⚠️ ERRO - {str(e)}")
                except Exception as e:
                    # Se não conseguir usar o gestor, fazer verificação básica
                    for data in sorted(datas_com_taxas):
                        relatorio.append(f"{data.strftime('%d/%m/%Y')}: ℹ️ DETECTADA - Possui taxa de administração")
                    relatorio.append("")
                    relatorio.append(f"⚠️ Erro no gestor de taxas: {str(e)}")
                    relatorio.append("Verificação básica realizada.")
                
                return "\n".join(relatorio)
            
            else:
                # Verificar data específica
                try:
                    gestor_taxas = GestorTaxasAdministracao(self)
                    precisa, motivo = gestor_taxas.verificar_necessidade_recalculo(data_referencia)
                    return f"Data {data_referencia}: {'❌ INCONSISTENTE' if precisa else '✅ OK'} - {motivo}"
                except Exception as e:
                    return f"Data {data_referencia}: ⚠️ ERRO - {str(e)}"
                
        except Exception as e:
            return f"Erro na verificação: {str(e)}"
        
    def verificar_e_mostrar_consistencia(self):
        """
        Verifica e mostra consistência das taxas em interface gráfica
        """
        if not self.cliente_atual:
            custom_messagebox("warning", "Aviso", "Selecione um cliente primeiro!")
            return
        
        try:
            print("DEBUG: Iniciando verificação de consistência das taxas...")
            print(f"DEBUG: Cliente atual: {self.cliente_atual}")
            
            # Agora chama o método da própria classe
            relatorio = self.verificar_consistencia_taxas()
            print(f"DEBUG: Relatório gerado com sucesso")
            
            # Criar janela para mostrar o relatório
            janela = tk.Toplevel(self.root)
            janela.title(f"Verificação de Taxas - {self.cliente_atual}")
            janela.geometry("700x500")
            janela.grab_set()  # Modal
            
            # Centralizar janela
            janela.transient(self.root)
            
            frame = ttk.Frame(janela, padding="15")
            frame.pack(fill='both', expand=True)
            
            ttk.Label(frame, text="📊 Verificação de Consistência das Taxas", 
                    font=('Arial', 14, 'bold')).pack(pady=(0, 15))
            
            # Área de texto com scroll
            text_frame = ttk.Frame(frame)
            text_frame.pack(fill='both', expand=True)
            
            text_widget = tk.Text(text_frame, wrap=tk.WORD, font=('Consolas', 10))
            scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
            text_widget.configure(yscrollcommand=scrollbar.set)
            
            text_widget.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            text_widget.insert(tk.END, relatorio)
            text_widget.config(state='disabled')
            
            # Botões
            botoes_frame = ttk.Frame(frame)
            botoes_frame.pack(fill='x', pady=(15, 0))
            
            ttk.Button(botoes_frame, text="Fechar", 
                    command=janela.destroy).pack(side='right')
                        
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro na verificação: {traceback.format_exc()}")
            custom_messagebox("error", "Erro", f"Erro na verificação: {str(e)}")

    # def chamar_apos_operacao_lancamento(self, data_lancamento, tipo_operacao):
    #     """
    #     Método utilitário para chamar após qualquer operação de lançamento
        
    #     Args:
    #         data_lancamento: Data do lançamento afetado
    #         tipo_operacao: "INCLUSAO", "EXCLUSAO", "ALTERACAO"
        
    #     Use este método após:
    #     - Adicionar novo lançamento
    #     - Excluir lançamento existente  
    #     - Alterar valor de lançamento existente
    #     - Alterar status de lançamento (ATIVO <-> EXCLUIDO)
    #     """
    #     try:
    #         # Só verificar se a data não for None/vazia
    #         if not data_lancamento:
    #             print("DEBUG: Data de lançamento não fornecida")
    #             return {"sucesso": True, "mensagem": "Sem data para verificar"}
            
    #         # Normalizar tipo de operação
    #         operacao_map = {
    #             "INCLUSAO": "INCLUSÃO",
    #             "EXCLUSAO": "EXCLUSÃO", 
    #             "ALTERACAO": "ALTERAÇÃO",
    #             "INCLUSÃO": "INCLUSÃO",
    #             "EXCLUSÃO": "EXCLUSÃO",
    #             "ALTERAÇÃO": "ALTERAÇÃO"
    #         }
            
    #         operacao = operacao_map.get(tipo_operacao.upper(), tipo_operacao)
            
    #         print(f"DEBUG: Operação de lançamento: {operacao} em {data_lancamento}")
            
    #         # Chamar verificação
    #         resultado = self.verificar_necessidade_recalculo_apos_nova_despesa(data_lancamento, operacao)
            
    #         return resultado
            
    #     except Exception as e:
    #         print(f"DEBUG: Erro ao processar operação de lançamento: {str(e)}")
    #         return {"sucesso": False, "mensagem": f"Erro: {str(e)}"}

    # def verificar_necessidade_recalculo_apos_nova_despesa(self, data_lancamento, operacao="INCLUSÃO"):
    #     """
    #     VERSÃO MELHORADA - Distingue entre quinzena atual e histórico
        
    #     Para QUINZENA ATUAL: Mantém comportamento original (exclui e recria)
    #     Para HISTÓRICO: Sugere ajuste compensatório
    #     """
    #     try:
    #         print(f"DEBUG: Verificando necessidade de recálculo após {operacao} em {data_lancamento}")
            
    #         # Converter data se necessário
    #         if isinstance(data_lancamento, str):
    #             data_obj = datetime.strptime(data_lancamento, '%d/%m/%Y').date()
    #         else:
    #             data_obj = data_lancamento
            
    #         # Verificar se existe taxa na data
    #         if not self._existe_taxa_na_data(data_obj):
    #             print(f"DEBUG: Nenhuma taxa encontrada em {data_obj.strftime('%d/%m/%Y')}")
    #             return {"sucesso": True, "mensagem": "Sem taxas na data"}
            
    #         # IMPORTANTE: Verificar se é quinzena atual ou histórico
    #         hoje = datetime.now().date()
    #         quinzena_atual = self._obter_quinzena_atual()
            
    #         # Determinar se a data é da quinzena atual
    #         eh_quinzena_atual = self._eh_mesma_quinzena(data_obj, quinzena_atual)
            
    #         print(f"DEBUG: Data {data_obj} - Quinzena atual: {eh_quinzena_atual}")
            
    #         gestor_taxas = GestorTaxasAdministracao(self)
            
    #         if eh_quinzena_atual:
    #             # QUINZENA ATUAL: Usar lógica original (excluir e recriar)
    #             return self._recalculo_quinzena_atual(data_obj, gestor_taxas, operacao)
    #         else:
    #             # HISTÓRICO: Apenas informar, não alterar
    #             return self._informar_diferenca_historica(data_obj, gestor_taxas, operacao)
                
    #     except Exception as e:
    #         import traceback
    #         print(f"DEBUG: Erro na verificação: {traceback.format_exc()}")
    #         return {"sucesso": False, "mensagem": f"Erro: {str(e)}"}
        
    def configurar_auto_salvamento(self):
        """Configura o auto-salvamento automático - MÉTODO NECESSÁRIO"""
        def executar_auto_salvamento():
            try:
                self.auto_salvar_dados()
                # Reagendar para 2 minutos (120000ms)
                self.root.after(120000, executar_auto_salvamento)
            except Exception as e:
                print(f"❌ Erro no auto-salvamento automático: {str(e)}")
                # Reagendar mesmo em caso de erro
                self.root.after(120000, executar_auto_salvamento)
        
        # Iniciar o timer após 2 minutos
        self.root.after(120000, executar_auto_salvamento)
        print("🔄 Auto-salvamento configurado (a cada 2 minutos)")

    def auto_salvar_dados(self):
        """Salva automaticamente os dados - VERSÃO CORRIGIDA COM NOMES ESPECÍFICOS"""
        try:
            if self.dados_para_incluir and self.cliente_atual:  # CORREÇÃO: Verificar se há cliente
                # Preparar dados do backup
                backup_data = {
                    'cliente': self.cliente_atual,
                    'data_sessao': datetime.now().isoformat(),
                    'lancamentos': self.dados_para_incluir,
                    'total_lancamentos': len(self.dados_para_incluir),
                    'estacao': os.environ.get('COMPUTERNAME', 'Desconhecido'),
                    'usuario': os.environ.get('USERNAME', 'Desconhecido'),
                    'versao_backup': '2.1'  # Nova versão com correções
                }
                
                backup_salvo = False
                
                # PRIORIDADE 1: Google Drive (já estava correto)
                try:
                    pasta_backup = PASTA_CLIENTES / "Backups_Sistema"
                    os.makedirs(pasta_backup, exist_ok=True)
                    
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    nome_arquivo = f"backup_{self.cliente_atual}_{timestamp}.json"
                    arquivo_backup = pasta_backup / nome_arquivo
                    
                    with open(arquivo_backup, 'w', encoding='utf-8') as f:
                        json.dump(backup_data, f, ensure_ascii=False, indent=2)
                    
                    print(f"✅ Backup salvo no Google Drive: {arquivo_backup}")
                    backup_salvo = True
                    
                    self.limpar_backups_antigos(pasta_backup, self.cliente_atual)
                    
                except Exception as e:
                    print(f"⚠️ Erro ao salvar backup no Google Drive: {str(e)}")
                
                # FALLBACK 1: Desktop local - CORREÇÃO: Nome específico do cliente
                if not backup_salvo:
                    try:
                        desktop_local = os.path.join(os.path.expanduser("~"), "Desktop")
                        if os.path.exists(desktop_local) and os.access(desktop_local, os.W_OK):
                            # CORREÇÃO: Usar nome específico do cliente
                            arquivo_local = os.path.join(desktop_local, f"backup_{self.cliente_atual}.json")
                            
                            with open(arquivo_local, 'w', encoding='utf-8') as f:
                                json.dump(backup_data, f, ensure_ascii=False, indent=2)
                            
                            print(f"✅ Backup salvo no Desktop: {arquivo_local}")
                            backup_salvo = True
                            
                    except Exception as e:
                        print(f"⚠️ Erro ao salvar backup no Desktop: {str(e)}")
                
                # FALLBACK 2: Pasta temporária - CORREÇÃO: Nome específico do cliente
                if not backup_salvo:
                    try:
                        import tempfile
                        # CORREÇÃO: Usar nome específico do cliente na pasta temp
                        arquivo_temp = os.path.join(tempfile.gettempdir(), f"backup_{self.cliente_atual}.json")
                        
                        with open(arquivo_temp, 'w', encoding='utf-8') as f:
                            json.dump(backup_data, f, ensure_ascii=False, indent=2)
                        
                        print(f"✅ Backup salvo na pasta temporária: {arquivo_temp}")
                        backup_salvo = True
                        
                    except Exception as e:
                        print(f"⚠️ Erro ao salvar backup na pasta temp: {str(e)}")
                
                if backup_salvo:
                    print(f"🔄 Auto-salvamento realizado para {self.cliente_atual}: {len(self.dados_para_incluir)} itens")
                else:
                    print("❌ ERRO: Não foi possível salvar backup em nenhum local!")
            else:
                if not self.cliente_atual:
                    print("⚠️ Auto-salvamento cancelado: Nenhum cliente selecionado")
                if not self.dados_para_incluir:
                    print("ℹ️ Auto-salvamento cancelado: Nenhum dado para salvar")
                    
        except Exception as e:
            print(f"❌ Erro geral no auto-salvamento: {str(e)}")

    def limpar_backups_antigos(self, pasta_backup, cliente):
        """Remove backups antigos mantendo apenas os últimos 5"""
        try:
            # Buscar todos os backups do cliente
            pattern = f"backup_{cliente}_*.json"
            backups = list(pasta_backup.glob(pattern))
            
            # Ordenar por data de modificação (mais recente primeiro)
            backups.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            
            # Remover backups excedentes (manter apenas os 5 mais recentes)
            for backup_antigo in backups[5:]:
                try:
                    backup_antigo.unlink()
                    print(f"🗑️ Backup antigo removido: {backup_antigo.name}")
                except Exception as e:
                    print(f"⚠️ Erro ao remover backup antigo {backup_antigo}: {str(e)}")
                    
        except Exception as e:
            print(f"⚠️ Erro ao limpar backups antigos: {str(e)}")

    def verificar_dados_nao_salvos(self):
        """Verifica se existem dados não salvos - VERSÃO CORRIGIDA COM VALIDAÇÃO"""
        try:
            backups_encontrados = []
            
            # BUSCA 1: Google Drive (prioridade)
            try:
                pasta_backup = PASTA_CLIENTES / "Backups_Sistema"
                if pasta_backup.exists():
                    agora = datetime.now()
                    limite_tempo = agora - relativedelta(hours=24)
                    
                    for arquivo_backup in pasta_backup.glob("backup_*.json"):
                        try:
                            # Verificar se é recente
                            data_modificacao = datetime.fromtimestamp(arquivo_backup.stat().st_mtime)
                            if data_modificacao >= limite_tempo:
                                
                                with open(arquivo_backup, 'r', encoding='utf-8') as f:
                                    backup_data = json.load(f)
                                
                                # CORREÇÃO: Verificar se tem dados válidos E se é do cliente atual
                                if (backup_data.get('lancamentos') and 
                                    len(backup_data['lancamentos']) > 0):
                                    
                                    backup_info = {
                                        'arquivo': arquivo_backup,
                                        'data': backup_data,
                                        'origem': 'Google Drive'
                                    }
                                    backups_encontrados.append(backup_info)
                                    
                        except Exception as e:
                            print(f"Erro ao processar backup {arquivo_backup}: {str(e)}")
                            continue
            except Exception as e:
                print(f"Erro ao buscar backups no Google Drive: {str(e)}")
            
            # BUSCA 2: Desktop local (fallback) - CORRIGIR NOME DO ARQUIVO
            if not backups_encontrados:
                try:
                    # CORREÇÃO: Usar nome específico do cliente se disponível
                    if self.cliente_atual:
                        temp_file = os.path.join(os.path.expanduser("~"), "Desktop", f"backup_{self.cliente_atual}.json")
                    else:
                        temp_file = os.path.join(os.path.expanduser("~"), "Desktop", "backup_lancamentos.json")
                    
                    if os.path.exists(temp_file):
                        data_modificacao = datetime.fromtimestamp(os.path.getmtime(temp_file))
                        if (datetime.now() - data_modificacao).days < 1:
                            
                            with open(temp_file, 'r', encoding='utf-8') as f:
                                backup_data = json.load(f)
                            
                            if backup_data.get('lancamentos'):
                                backup_info = {
                                    'arquivo': temp_file,
                                    'data': backup_data,
                                    'origem': 'Desktop Local'
                                }
                                backups_encontrados.append(backup_info)
                                
                except Exception as e:
                    print(f"Erro ao buscar backup no Desktop: {str(e)}")
            
            # Processar backups encontrados COM VALIDAÇÃO DE CLIENTE
            if backups_encontrados:
                # Ordenar por data (mais recente primeiro)
                backups_encontrados.sort(
                    key=lambda x: datetime.fromisoformat(x['data']['data_sessao']), 
                    reverse=True
                )
                
                backup_mais_recente = backups_encontrados[0]
                backup_data = backup_mais_recente['data']
                cliente_backup = backup_data.get('cliente', '')
                
                # CORREÇÃO PRINCIPAL: Validar se o cliente do backup é compatível
                if self.cliente_atual and cliente_backup and self.cliente_atual != cliente_backup:
                    # Cliente do backup é diferente do atual - perguntar ao usuário
                    resposta_cliente = custom_messagebox(
                        "yesno",
                        "⚠️ Cliente Diferente Detectado",
                        f"ATENÇÃO: Foi encontrado um backup com dados de outro cliente!\n\n"
                        f"Cliente atual selecionado: {self.cliente_atual}\n"
                        f"Cliente do backup: {cliente_backup}\n\n"
                        f"Lançamentos no backup: {len(backup_data['lancamentos'])}\n"
                        f"Data do backup: {datetime.fromisoformat(backup_data['data_sessao']).strftime('%d/%m/%Y às %H:%M:%S')}\n\n"
                        f"IMPORTANTE: Se você carregar este backup, o sistema mudará\n"
                        f"automaticamente para o cliente '{cliente_backup}'.\n\n"
                        f"Deseja continuar e mudar para o cliente do backup?"
                    )
                    
                    if not resposta_cliente:
                        # Usuário recusou - perguntar se quer remover o backup
                        if custom_messagebox("yesno", "Remover Backup", 
                                            f"Deseja remover este backup de '{cliente_backup}' "
                                            f"para não ser perguntado novamente?"):
                            try:
                                if backup_mais_recente['origem'] == 'Google Drive':
                                    backup_mais_recente['arquivo'].unlink()
                                else:
                                    os.remove(backup_mais_recente['arquivo'])
                                print(f"🗑️ Backup de cliente diferente removido")
                            except Exception as e:
                                print(f"⚠️ Erro ao remover backup: {str(e)}")
                        return False
                
                # Se chegou até aqui, pode continuar com a recuperação
                data_backup = datetime.fromisoformat(backup_data['data_sessao'])
                origem = backup_mais_recente['origem']
                estacao = backup_data.get('estacao', 'Desconhecida')
                usuario = backup_data.get('usuario', 'Desconhecido')
                total_lancamentos = backup_data.get('total_lancamentos', len(backup_data['lancamentos']))
                
                # Mensagem de confirmação melhorada
                mensagem_recuperacao = (
                    f"🔄 RECUPERAÇÃO DE DADOS DISPONÍVEL\n\n"
                    f"📋 Cliente: {cliente_backup}\n"
                    f"📊 Lançamentos: {total_lancamentos}\n"
                    f"📅 Data/Hora: {data_backup.strftime('%d/%m/%Y às %H:%M:%S')}\n"
                    f"💾 Origem: {origem}\n"
                    f"🖥️ Estação: {estacao}\n"
                    f"👤 Usuário: {usuario}\n\n"
                )
                
                # CORREÇÃO: Aviso claro sobre mudança de cliente
                if self.cliente_atual != cliente_backup:
                    mensagem_recuperacao += (
                        f"⚠️ ATENÇÃO: O sistema mudará do cliente '{self.cliente_atual}' "
                        f"para '{cliente_backup}'\n\n"
                    )
                
                mensagem_recuperacao += "Deseja recuperar estes dados?"
                
                if custom_messagebox("yesno", "Recuperação de Dados", mensagem_recuperacao):
                    
                    # Recuperar dados
                    self.dados_para_incluir = backup_data['lancamentos']
                    
                    # CORREÇÃO: Log da mudança de cliente
                    if self.cliente_atual != cliente_backup:
                        print(f"🔄 Cliente alterado: {self.cliente_atual} → {cliente_backup}")
                    
                    self.cliente_atual = cliente_backup
                    
                    # Atualizar interface
                    if self.cliente_atual:
                        self.cliente_combobox.set(self.cliente_atual)
                        self.selecionar_cliente(None)
                    
                    # Mostrar visualizador
                    self.visualizar_lancamentos()
                    
                    # Remover backup após recuperação bem-sucedida
                    try:
                        if origem == 'Google Drive':
                            backup_mais_recente['arquivo'].unlink()
                            print(f"✅ Backup removido após recuperação: {backup_mais_recente['arquivo']}")
                        else:
                            os.remove(backup_mais_recente['arquivo'])
                            print(f"✅ Backup removido após recuperação: {backup_mais_recente['arquivo']}")
                    except Exception as e:
                        print(f"⚠️ Erro ao remover backup: {str(e)}")
                    
                    custom_messagebox("info", "Recuperação Realizada", 
                                    f"✅ Dados recuperados com sucesso!\n\n"
                                    f"📊 {total_lancamentos} lançamentos carregados\n"
                                    f"📋 Cliente: {cliente_backup}\n"
                                    f"💾 Origem: {origem}")
                    
                    return True
                else:
                    # Usuário recusou recuperação - perguntar se quer remover backup
                    if custom_messagebox("yesno", "Remover Backup", 
                                        "Deseja remover este backup para não ser perguntado novamente?"):
                        try:
                            if origem == 'Google Drive':
                                backup_mais_recente['arquivo'].unlink()
                            else:
                                os.remove(backup_mais_recente['arquivo'])
                            print(f"🗑️ Backup removido por solicitação do usuário")
                        except Exception as e:
                            print(f"⚠️ Erro ao remover backup: {str(e)}")
            
            return False
            
        except Exception as e:
            print(f"❌ Erro na verificação de recuperação: {str(e)}")
            return False

    def limpar_backup(self):
        """Remove arquivos de backup - VERSÃO CORRIGIDA PARA TODOS OS FORMATOS"""
        try:
            backups_removidos = 0
            
            # LIMPAR 1: Google Drive (já estava correto)
            try:
                pasta_backup = PASTA_CLIENTES / "Backups_Sistema"
                if pasta_backup.exists() and self.cliente_atual:
                    pattern = f"backup_{self.cliente_atual}_*.json"
                    for backup_file in pasta_backup.glob(pattern):
                        try:
                            backup_file.unlink()
                            backups_removidos += 1
                            print(f"🗑️ Backup removido do Google Drive: {backup_file.name}")
                        except Exception as e:
                            print(f"⚠️ Erro ao remover backup {backup_file}: {str(e)}")
            except Exception as e:
                print(f"⚠️ Erro ao limpar backups do Google Drive: {str(e)}")
            
            # LIMPAR 2: Desktop local - CORREÇÃO: Buscar ambos os formatos
            try:
                desktop_local = os.path.expanduser("~") + "/Desktop"
                
                # Formato antigo (genérico)
                arquivo_antigo = os.path.join(desktop_local, "backup_lancamentos.json")
                if os.path.exists(arquivo_antigo):
                    os.remove(arquivo_antigo)
                    backups_removidos += 1
                    print("🗑️ Backup genérico removido do Desktop")
                
                # Formato novo (específico do cliente)
                if self.cliente_atual:
                    arquivo_novo = os.path.join(desktop_local, f"backup_{self.cliente_atual}.json")
                    if os.path.exists(arquivo_novo):
                        os.remove(arquivo_novo)
                        backups_removidos += 1
                        print(f"🗑️ Backup de {self.cliente_atual} removido do Desktop")
                        
            except Exception as e:
                print(f"⚠️ Erro ao remover backup do Desktop: {str(e)}")
            
            # LIMPAR 3: Pasta temporária - CORREÇÃO: Buscar ambos os formatos
            try:
                import tempfile
                temp_dir = tempfile.gettempdir()
                
                # Formato antigo (genérico)
                arquivo_antigo = os.path.join(temp_dir, "backup_lancamentos.json")
                if os.path.exists(arquivo_antigo):
                    os.remove(arquivo_antigo)
                    backups_removidos += 1
                    print("🗑️ Backup genérico removido da pasta temp")
                
                # Formato novo (específico do cliente)
                if self.cliente_atual:
                    arquivo_novo = os.path.join(temp_dir, f"backup_{self.cliente_atual}.json")
                    if os.path.exists(arquivo_novo):
                        os.remove(arquivo_novo)
                        backups_removidos += 1
                        print(f"🗑️ Backup de {self.cliente_atual} removido da pasta temp")
                        
            except Exception as e:
                print(f"⚠️ Erro ao remover backup da pasta temp: {str(e)}")
            
            if backups_removidos > 0:
                print(f"✅ Total de backups limpos: {backups_removidos}")
            else:
                print("ℹ️ Nenhum backup encontrado para limpeza")
                
        except Exception as e:
            print(f"❌ Erro ao limpar backup: {str(e)}")

    # MÉTODO ADICIONAL: Visualizar backups disponíveis (para debug/administração)
    def listar_backups_disponiveis(self):
        """Lista todos os backups disponíveis (método de administração)"""
        try:
            print("\n📋 LISTAGEM DE BACKUPS DISPONÍVEIS")
            print("=" * 50)
            
            total_backups = 0
            
            # Listar Google Drive
            try:
                pasta_backup = PASTA_CLIENTES / "Backups_Sistema"
                if pasta_backup.exists():
                    print(f"\n💾 Google Drive ({pasta_backup}):")
                    for arquivo_backup in pasta_backup.glob("backup_*.json"):
                        try:
                            data_mod = datetime.fromtimestamp(arquivo_backup.stat().st_mtime)
                            tamanho = arquivo_backup.stat().st_size
                            print(f"  📄 {arquivo_backup.name}")
                            print(f"      📅 {data_mod.strftime('%d/%m/%Y %H:%M:%S')}")
                            print(f"      📏 {tamanho:,} bytes")
                            total_backups += 1
                        except Exception as e:
                            print(f"  ❌ Erro ao ler {arquivo_backup}: {str(e)}")
                else:
                    print("\n💾 Google Drive: Pasta não encontrada")
            except Exception as e:
                print(f"\n💾 Google Drive: Erro ao acessar - {str(e)}")
            
            # Listar Desktop
            try:
                temp_file = os.path.join(os.path.expanduser("~"), "Desktop", "backup_lancamentos.json")
                if os.path.exists(temp_file):
                    print(f"\n🖥️ Desktop Local:")
                    data_mod = datetime.fromtimestamp(os.path.getmtime(temp_file))
                    tamanho = os.path.getsize(temp_file)
                    print(f"  📄 backup_lancamentos.json")
                    print(f"      📅 {data_mod.strftime('%d/%m/%Y %H:%M:%S')}")
                    print(f"      📏 {tamanho:,} bytes")
                    total_backups += 1
                else:
                    print("\n🖥️ Desktop Local: Nenhum backup encontrado")
            except Exception as e:
                print(f"\n🖥️ Desktop Local: Erro ao acessar - {str(e)}")
            
            print(f"\n📊 TOTAL DE BACKUPS: {total_backups}")
            print("=" * 50)
            
            return total_backups
            
        except Exception as e:
            print(f"❌ Erro ao listar backups: {str(e)}")
            return 0

    def limpar_backups_orfaos(self):
        """Limpa backups órfãos de todos os clientes - MÉTODO DE MANUTENÇÃO"""
        try:
            print("\n🧹 LIMPEZA DE BACKUPS ÓRFÃOS")
            print("=" * 50)
            
            backups_removidos = 0
            
            # LIMPAR Desktop: backups genéricos antigos
            try:
                desktop_local = os.path.expanduser("~") + "/Desktop"
                arquivo_generico = os.path.join(desktop_local, "backup_lancamentos.json")
                
                if os.path.exists(arquivo_generico):
                    # Verificar se é antigo (mais de 1 dia)
                    data_modificacao = datetime.fromtimestamp(os.path.getmtime(arquivo_generico))
                    if (datetime.now() - data_modificacao).days >= 1:
                        os.remove(arquivo_generico)
                        backups_removidos += 1
                        print("🗑️ Backup genérico antigo removido do Desktop")
                        
            except Exception as e:
                print(f"⚠️ Erro ao limpar Desktop: {str(e)}")
            
            # LIMPAR Pasta temporária: backups genéricos antigos
            try:
                import tempfile
                temp_dir = tempfile.gettempdir()
                arquivo_generico = os.path.join(temp_dir, "backup_lancamentos.json")
                
                if os.path.exists(arquivo_generico):
                    # Verificar se é antigo (mais de 1 dia)
                    data_modificacao = datetime.fromtimestamp(os.path.getmtime(arquivo_generico))
                    if (datetime.now() - data_modificacao).days >= 1:
                        os.remove(arquivo_generico)
                        backups_removidos += 1
                        print("🗑️ Backup genérico antigo removido da pasta temp")
                        
            except Exception as e:
                print(f"⚠️ Erro ao limpar pasta temp: {str(e)}")
            
            # LIMPAR Google Drive: backups muito antigos (mais de 7 dias)
            try:
                pasta_backup = PASTA_CLIENTES / "Backups_Sistema"
                if pasta_backup.exists():
                    agora = datetime.now()
                    limite_tempo = agora - relativedelta(days=7)
                    
                    for arquivo_backup in pasta_backup.glob("backup_*.json"):
                        try:
                            data_modificacao = datetime.fromtimestamp(arquivo_backup.stat().st_mtime)
                            if data_modificacao < limite_tempo:
                                arquivo_backup.unlink()
                                backups_removidos += 1
                                print(f"🗑️ Backup antigo removido: {arquivo_backup.name}")
                        except Exception as e:
                            print(f"⚠️ Erro ao processar {arquivo_backup}: {str(e)}")
                            
            except Exception as e:
                print(f"⚠️ Erro ao limpar Google Drive: {str(e)}")
            
            print(f"\n✅ Limpeza concluída: {backups_removidos} backups órfãos removidos")
            print("=" * 50)
            
            return backups_removidos
            
        except Exception as e:
            print(f"❌ Erro na limpeza de backups órfãos: {str(e)}")
            return 0


    def verificar_integridade_backups(self):
        """Verifica integridade de todos os backups disponíveis"""
        try:
            print("\n🔍 VERIFICAÇÃO DE INTEGRIDADE DOS BACKUPS")
            print("=" * 60)
            
            backups_validos = 0
            backups_corrompidos = 0
            
            # Verificar Google Drive
            try:
                pasta_backup = PASTA_CLIENTES / "Backups_Sistema"
                if pasta_backup.exists():
                    print(f"\n💾 Verificando Google Drive ({pasta_backup}):")
                    
                    for arquivo_backup in pasta_backup.glob("backup_*.json"):
                        try:
                            with open(arquivo_backup, 'r', encoding='utf-8') as f:
                                backup_data = json.load(f)
                            
                            # Verificar campos obrigatórios
                            campos_obrigatorios = ['cliente', 'data_sessao', 'lancamentos']
                            campos_ausentes = [campo for campo in campos_obrigatorios 
                                            if campo not in backup_data]
                            
                            if campos_ausentes:
                                print(f"  ❌ {arquivo_backup.name} - Campos ausentes: {campos_ausentes}")
                                backups_corrompidos += 1
                            else:
                                cliente = backup_data['cliente']
                                total_lancamentos = len(backup_data['lancamentos'])
                                data_backup = datetime.fromisoformat(backup_data['data_sessao'])
                                
                                print(f"  ✅ {arquivo_backup.name}")
                                print(f"      Cliente: {cliente}")
                                print(f"      Lançamentos: {total_lancamentos}")
                                print(f"      Data: {data_backup.strftime('%d/%m/%Y %H:%M:%S')}")
                                backups_validos += 1
                                
                        except json.JSONDecodeError:
                            print(f"  ❌ {arquivo_backup.name} - JSON corrompido")
                            backups_corrompidos += 1
                        except Exception as e:
                            print(f"  ❌ {arquivo_backup.name} - Erro: {str(e)}")
                            backups_corrompidos += 1
                else:
                    print("\n💾 Google Drive: Pasta não encontrada")
            except Exception as e:
                print(f"\n💾 Google Drive: Erro ao acessar - {str(e)}")
            
            print(f"\n📊 RESUMO DA VERIFICAÇÃO:")
            print(f"✅ Backups válidos: {backups_validos}")
            print(f"❌ Backups corrompidos: {backups_corrompidos}")
            print("=" * 60)
            
            return {'validos': backups_validos, 'corrompidos': backups_corrompidos}
            
        except Exception as e:
            print(f"❌ Erro na verificação de integridade: {str(e)}")
            return {'validos': 0, 'corrompidos': 0}
    
    def abrir_correcao_monetaria(self):
        """Abre o gerenciador de correção monetária"""
        try:
            from src.correcao_monetaria import InterfaceIndicesCorrecao
            interface = InterfaceIndicesCorrecao(self.root)
        except ImportError as e:
            custom_messagebox("error", "Erro", f"Erro ao importar módulo de correção: {str(e)}")
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir correção monetária: {str(e)}")

    def finalizar_sistema(self):
        """Finaliza o sistema de forma segura"""
        try:
            # Salvar dados pendentes se necessário
            if hasattr(self, 'dados_para_incluir') and self.dados_para_incluir:
                if hasattr(self, 'custom_messagebox'):
                    if self.custom_messagebox("yesno", "Confirmação", 
                        "Existem dados não salvos. Deseja salvá-los antes de sair?"):
                        self.enviar_dados()
            
            # Limpar backup temporário
            if hasattr(self, 'limpar_backup'):
                self.limpar_backup()
            
            # Fechar janelas filhas primeiro
            if hasattr(self, 'visualizador') and self.visualizador:
                try:
                    self.visualizador.janela.destroy()
                except:
                    pass
            
            # Fechar janela principal
            if hasattr(self, 'root') and self.root:
                try:
                    self.root.quit()  # Usar quit() ao invés de destroy()
                    self.root.destroy()
                except:
                    pass
                    
        except Exception as e:
            print(f"Aviso durante finalização: {str(e)}")

class EditorCliente:
    def __init__(self, parent): 
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("Editor de Clientes")
        self.root.geometry("800x600")
        
        self.setup_gui()
        self.carregar_clientes()

    def setup_gui(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Lista de clientes
        frame_clientes = ttk.LabelFrame(main_frame, text="Clientes")
        frame_clientes.pack(fill='both', expand=True, pady=5)

        self.tree_clientes = ttk.Treeview(frame_clientes, 
                                        columns=('Nome', 'Endereço', 'Taxa ADM'),
                                        show='headings')
        self.tree_clientes.heading('Nome', text='Nome')
        self.tree_clientes.heading('Endereço', text='Endereço')
        self.tree_clientes.heading('Taxa ADM', text='Taxa ADM (%)')
        self.tree_clientes.pack(fill='both', expand=True, padx=5, pady=5)

        # Frame para edição
        frame_edicao = ttk.LabelFrame(main_frame, text="Edição")
        frame_edicao.pack(fill='x', pady=5)

        ttk.Label(frame_edicao, text="Taxa de Administração (%):").pack(side='left', padx=5)
        self.taxa_entry = ttk.Entry(frame_edicao, width=10)
        self.taxa_entry.pack(side='left', padx=5)

        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(frame_botoes, 
                  text="Atualizar Taxa", 
                  command=self.atualizar_taxa).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Remover Taxa", 
                  command=self.remover_taxa).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Fechar", 
                  command=self.root.destroy).pack(side='right', padx=5)

    def carregar_clientes(self):
        """Carrega a lista de clientes do arquivo Excel"""
        try:
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            # Limpar lista atual
            for item in self.tree_clientes.get_children():
                self.tree_clientes.delete(item)
            
            # Adicionar clientes
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome não vazio
                    self.tree_clientes.insert('', 'end', values=(
                        row[0],  # Nome
                        row[1],  # Endereço
                        row[6] if row[6] else "0.00"  # Taxa ADM
                    ))
            
            wb.close()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar clientes: {str(e)}")

    def atualizar_taxa(self):
        """Atualiza a taxa de administração do cliente selecionado"""
        selecionado = self.tree_clientes.selection()
        if not selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um cliente")
            return

        try:
            taxa = float(self.taxa_entry.get().replace(',', '.'))
            if not (0 <= taxa <= 100):
                custom_messagebox("error", "Erro", "Taxa deve estar entre 0 e 100")
                return
                
            cliente = self.tree_clientes.item(selecionado)['values'][0]
            
            # Atualizar no arquivo
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == cliente:
                    row[6].value = taxa  # Coluna da taxa de administração
                    
            wb.save(ARQUIVO_CLIENTES)
            
            # Atualizar na treeview
            self.tree_clientes.set(selecionado, 'Taxa ADM', f"{taxa:.2f}")
            custom_messagebox("info", "Sucesso", "Taxa atualizada com sucesso!")
            
        except ValueError:
            custom_messagebox("error", "Erro", "Taxa inválida")
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao atualizar taxa: {str(e)}")

    def remover_taxa(self):
        """Remove a taxa de administração do cliente selecionado"""
        selecionado = self.tree_clientes.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um cliente")
            return

        if custom_messagebox("yesno", "Confirmar", "Deseja remover a taxa de administração?"):
            try:
                cliente = self.tree_clientes.item(selecionado)['values'][0]
                
                # Atualizar no arquivo
                wb = load_workbook(ARQUIVO_CLIENTES)
                ws = wb['Clientes']
                
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == cliente:
                        row[6].value = None  # Remover taxa
                        
                wb.save(ARQUIVO_CLIENTES)
                
                # Atualizar na treeview
                self.tree_clientes.set(selecionado, 'Taxa ADM', "0.00")
                custom_messagebox("info", "Sucesso", "Taxa removida com sucesso!")
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao remover taxa: {str(e)}")

class GestaoContratos:
    def __init__(self, parent):
        self.parent = parent
        self.arquivo_cliente = None
        self.cliente_atual = None

    def centralizar_janela(self, janela, largura=800, altura=600, parent=None):
        """
        Centraliza uma janela na tela ou relativa ao parent se fornecido.
        Também define o tamanho padrão da janela.
        """
        # Definir tamanho
        janela.geometry(f"{largura}x{altura}")
        
        # Atualizar a janela para garantir que as dimensões sejam aplicadas
        janela.update_idletasks()
        
        # Se parent for fornecido, centralize em relação a ele
        if parent and parent.winfo_exists():
            # Calcular o centro da janela pai
            x_parent = parent.winfo_x() + parent.winfo_width() // 2
            y_parent = parent.winfo_y() + parent.winfo_height() // 2
            
            # Calcular a posição da nova janela
            x = x_parent - largura // 2
            y = y_parent - altura // 2
        else:
            # Centralizar na tela
            x = (janela.winfo_screenwidth() // 2) - (largura // 2)
            y = (janela.winfo_screenheight() // 2) - (altura // 2)
        
        # Definir posição
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        
        # Tornar a janela modal (quando aplicável)
        if parent and hasattr(janela, 'transient') and hasattr(janela, 'grab_set'):
            janela.transient(parent)
            janela.grab_set()
            
        # Trazer para frente
        janela.lift()
        janela.focus_force()

    def criar_interface_contratos(self, janela, on_close_callback):
        """Cria a interface de gestão de contratos em uma janela já existente"""
        try:
            # Verificar se o arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                custom_messagebox("error", "Erro", f"Arquivo do cliente {self.cliente_atual} não encontrado!")
                on_close_callback()  # Fechar a janela em caso de erro
                return

            # Abrir arquivo e verificar aba
            wb = load_workbook(self.arquivo_cliente)
            if 'Contratos_ADM' not in wb.sheetnames:
                # Se não existir a aba, criar
                print(f"Criando aba Contratos_ADM para {self.cliente_atual}")
                ws = wb.create_sheet("Contratos_ADM")
                
                # Definir os blocos na linha 1
                blocos = ["CONTRATOS", "", "", "", "", "",
                        "ADMINISTRADORES_CONTRATO", "", "", "", "", "", "",
                        "ADITIVOS", "", "", "",
                        "ADMINISTRADORES_ADITIVO", "", "", "", "", "", "",
                        "PARCELAS", "", "", "", "", "", "", "", ""]
                
                for col, valor in enumerate(blocos, 1):
                    ws.cell(row=1, column=col, value=valor)
                
                # Definir cabeçalhos na linha 2
                headers = [
                    # CONTRATOS
                    "Nº Contrato", "Data Início", "Data Fim", "Status", "Observações", "",
                    # ADMINISTRADORES_CONTRATO
                    "Nº Contrato", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total", "Nº Parcelas", 
                    # ADITIVOS
                    "Nº Contrato", "Nº Aditivo", "Data Início", "Data Fim",
                    # ADMINISTRADORES_ADITIVO
                    "Nº Contrato", "Nº Aditivo", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total",
                    # PARCELAS
                    "Referência", "Número", "CNPJ/CPF", "Nome", "Data Vencimento", "Valor", "Status", "Data Pagamento", "Eventos/Fases", "Percentual %"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                # Salvar as alterações
                wb.save(self.arquivo_cliente)

            # Frame principal
            frame_principal = ttk.Frame(janela, padding="10")
            frame_principal.pack(fill='both', expand=True)

            # Frame para lista de contratos existentes
            frame_contratos = ttk.LabelFrame(frame_principal, text="Contratos Existentes")
            frame_contratos.pack(fill='both', expand=True, pady=5)

            # Treeview para contratos
            colunas = ('Nº Contrato', 'Data Início', 'Data Fim', 'Status')
            self.tree_contratos = ttk.Treeview(frame_contratos, columns=colunas, show='headings')
            for col in colunas:
                self.tree_contratos.heading(col, text=col)
                self.tree_contratos.column(col, width=100)
            
            # Adicionar scrollbars
            scroll_y = ttk.Scrollbar(frame_contratos, orient='vertical', command=self.tree_contratos.yview)
            scroll_x = ttk.Scrollbar(frame_contratos, orient='horizontal', command=self.tree_contratos.xview)
            self.tree_contratos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            
            self.tree_contratos.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y.pack(side='right', fill='y')
            scroll_x.pack(side='bottom', fill='x')

            # Frame para lista de administradores do contrato selecionado
            frame_admins = ttk.LabelFrame(frame_principal, text="Administradores do Contrato")
            frame_admins.pack(fill='both', expand=True, pady=5)

            # Treeview para administradores
            colunas_adm = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial Pagamento')
            self.tree_adm_contrato = ttk.Treeview(frame_admins, columns=colunas_adm, show='headings')
            for col in colunas_adm:
                self.tree_adm_contrato.heading(col, text=col)
                self.tree_adm_contrato.column(col, width=100)
            
            # Adicionar scrollbars para administradores
            scroll_y_adm = ttk.Scrollbar(frame_admins, orient='vertical', command=self.tree_adm_contrato.yview)
            scroll_x_adm = ttk.Scrollbar(frame_admins, orient='horizontal', command=self.tree_adm_contrato.xview)
            self.tree_adm_contrato.configure(yscrollcommand=scroll_y_adm.set, xscrollcommand=scroll_x_adm.set)
            
            self.tree_adm_contrato.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y_adm.pack(side='right', fill='y')
            scroll_x_adm.pack(side='bottom', fill='x')

            # Botões de ação
            frame_botoes = ttk.Frame(frame_principal)
            frame_botoes.pack(fill='x', pady=5)

            ttk.Button(frame_botoes, text="Novo Contrato", 
                    command=lambda: self.criar_novo_contrato(janela)).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Editar Contrato", 
                    command=self.editar_contrato).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Excluir Contrato", 
                    command=self.excluir_contrato).pack(side='left', padx=5)
            
            # Botão Fechar com callback personalizado
            ttk.Button(frame_botoes, text="Fechar", 
                    command=on_close_callback).pack(side='right', padx=5)

            # Carregar contratos existentes
            self.carregar_contratos()

            # Binding para atualizar administradores quando selecionar contrato
            self.tree_contratos.bind('<<TreeviewSelect>>', self.mostrar_administradores)

        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao abrir janela de contratos: {str(e)}")
            if 'wb' in locals():
                wb.close()
            # Garantir que a janela principal seja restaurada em caso de erro
            on_close_callback()

    def carregar_contratos(self):
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            for item in self.tree_contratos.get_children():
                self.tree_contratos.delete(item)
            
            contratos_processados = set()
            for row in ws.iter_rows(min_row=3, values_only=True):
                num_contrato = row[0]
                if num_contrato and num_contrato not in contratos_processados:
                    # Processar datas
                    data_inicio = ''
                    if row[1]:
                        try:
                            if isinstance(row[1], datetime):
                                data_inicio = row[1].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[1]), '%Y-%m-%d')
                                data_inicio = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_inicio = str(row[1])

                    data_fim = ''
                    if row[2]:
                        try:
                            if isinstance(row[2], datetime):
                                data_fim = row[2].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[2]), '%Y-%m-%d')
                                data_fim = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_fim = str(row[2])

                    
                    self.tree_contratos.insert('', 'end', values=(
                        num_contrato,
                        data_inicio,
                        data_fim,
                        row[3] or ''
                    ))
                    contratos_processados.add(num_contrato)
            
            wb.close()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar contratos: {str(e)}")

    def mostrar_administradores(self, event=None):
        """Mostra administradores do contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            return
            
        try:
            # Limpar lista atual
            for item in self.tree_adm_contrato.get_children():
                self.tree_adm_contrato.delete(item)
                
            num_contrato = self.tree_contratos.item(selecionado)['values'][0]
            
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[6] == num_contrato:  # Coluna G - Nº Contrato
                    if row[26]:  # Data Inicial de Pagamento
                        data_inicial = row[26].strftime('%d/%m/%Y') if isinstance(row[26], datetime) else str(row[26])
                    else:
                        data_inicial = ''
                        
                    self.tree_adm_contrato.insert('', 'end', values=(
                        row[7],   # CNPJ/CPF
                        row[8],   # Nome
                        row[9],   # Tipo
                        row[10],  # Valor/Percentual
                        row[11],  # Valor Total
                        row[12],  # Nº Parcelas
                        data_inicial  # Data Inicial de Pagamento
                    ))
            
            wb.close()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar administradores: {str(e)}")
  
    def criar_novo_contrato(self, janela_principal):
        """Abre janela para criar novo contrato com suporte a parcelas fixas ou eventos"""
        janela = tk.Toplevel(self.parent)
        janela.title(f"Novo Contrato - {self.cliente_atual}")
        janela.geometry("800x650")
        
        # Frame principal com scrollbar para garantir acesso a todos os campos
        main_frame = ttk.Frame(janela)
        main_frame.pack(fill='both', expand=True)
        
        # Adicionar canvas com scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)
        
        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Frame principal dentro do scrollable frame
        frame = ttk.Frame(scroll_frame, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame para dados do contrato
        frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
        frame_contrato.pack(fill='x', pady=5)

        # Número do Contrato
        ttk.Label(frame_contrato, text="Nº Contrato:*", width=15).grid(row=0, column=0, padx=5, pady=5, sticky='w')
        num_contrato = ttk.Entry(frame_contrato, width=20)
        num_contrato.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Datas 
        ttk.Label(frame_contrato, text="Data Início:*", width=15).grid(row=1, column=0, padx=5, pady=5, sticky='w')
        data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_inicio.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame_contrato, text="Data Fim:*", width=15).grid(row=2, column=0, padx=5, pady=5, sticky='w')
        data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_fim.grid(row=2, column=1, padx=5, pady=5, sticky='w')

        # Observações
        ttk.Label(frame_contrato, text="Observações:", width=15).grid(row=3, column=0, padx=5, pady=5, sticky='nw')
        observacoes = ttk.Entry(frame_contrato, width=25)
        observacoes.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        
        # Adicionar campo para valor global do contrato
        ttk.Label(frame_contrato, text="Valor Global:*", width=15).grid(row=4, column=0, padx=5, pady=5, sticky='w')
        valor_global = ttk.Entry(frame_contrato, width=20)
        valor_global.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        
        # Tipo de pagamento (metodo)
        ttk.Label(frame_contrato, text="Método de Pagamento:*", width=22).grid(row=5, column=0, padx=5, pady=5, sticky='w')
        metodo_pagamento = ttk.Combobox(frame_contrato, values=[
            "Percentual da Quinzena", 
            "Valor Fixo em Parcelas", 
            "Eventos/Fases"
        ], state='readonly', width=20)
        metodo_pagamento.grid(row=5, column=1, padx=5, pady=5, sticky='w')
        metodo_pagamento.current(0)  # Valor padrão
        
        # Frame para Administradores
        frame_adm = ttk.LabelFrame(frame, text="Administradores")
        frame_adm.pack(fill='both', expand=True, pady=5)

        # Lista de Administradores
        colunas = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial')
        self.tree_adm = ttk.Treeview(frame_adm, columns=colunas, show='headings', height=5)
        
        for col in colunas:
            self.tree_adm.heading(col, text=col)
            self.tree_adm.column(col, width=100)
        
        # Adicionar scrollbars
        scroll_y = ttk.Scrollbar(frame_adm, orient='vertical', command=self.tree_adm.yview)
        scroll_x = ttk.Scrollbar(frame_adm, orient='horizontal', command=self.tree_adm.xview)
        self.tree_adm.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_adm.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')
        
        # Botões para administradores
        frame_botoes_adm = ttk.Frame(frame_adm)
        frame_botoes_adm.pack(fill='x', pady=5)

        # Botões para administradores - explicitamente configurados
        ttk.Button(
            frame_botoes_adm, 
            text="Adicionar Administrador",
            command=lambda: self.adicionar_administrador_modificado(self.tree_adm, valor_global, metodo_pagamento)
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_adm, 
            text="Remover Administrador",
            command=lambda: self.remover_administrador(self.tree_adm)
        ).pack(side='left', padx=5)

        def salvar():
            # Validar campos obrigatórios
            if not num_contrato.get() or not data_inicio.get() or not data_fim.get() or not valor_global.get():
                custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios do contrato!")
                return
                
            # Validar valor global
            try:
                valor_global_float = float(valor_global.get().replace(',', '.'))
                if valor_global_float <= 0:
                    custom_messagebox("error", "Erro", "Valor global deve ser maior que zero!")
                    return
            except ValueError:
                custom_messagebox("error", "Erro", "Valor global inválido!")
                return
                
            # Validar administradores
            if not self.tree_adm.get_children():
                custom_messagebox("error", "Erro", "Adicione pelo menos um administrador!")
                return
                
            # Criar contrato
            self.salvar_contrato_com_opcoes(
                num_contrato.get(),
                data_inicio.get_date(),
                data_fim.get_date(),
                observacoes.get(),
                valor_global_float,
                metodo_pagamento.get(),
                {},  # Opções simplificadas pois foram movidas para o administrador
                janela
            )
            
            janela_principal.focus_set()
            self.carregar_contratos()

        ttk.Button(frame, text="Salvar", command=salvar).pack(side='left', padx=5, pady=10)
        ttk.Button(frame, text="Cancelar", command=janela.destroy).pack(side='left', padx=5, pady=10)                          

    def processar_eventos(self, ws, num_contrato, valor_global, eventos):
        """Processa os eventos do contrato e cria parcelas vinculadas"""
        for i, (descricao, percentual, valor_evento) in enumerate(eventos, 1):
            # Para cada administrador, criar um registro de parcela vinculada ao evento
            for item in self.tree_adm.get_children():
                valores_adm = self.tree_adm.item(item)['values']
                cnpj_cpf_adm = str(valores_adm[0]).strip()
                cnpj_cpf_adm = formatar_cnpj_cpf(cnpj_cpf_adm)
                nome_adm = valores_adm[1]
                
                # Calcular valor para este administrador (proporcional ao percentual definido)
                if valores_adm[2] == 'Percentual':
                    perc_adm = float(str(valores_adm[3]).replace('%', '').replace(',', '.'))
                    valor_admin_evento = (perc_adm / 100) * valor_evento
                else:  # Fixo
                    # Distribuir o valor total entre os eventos conforme percentuais
                    valor_total_adm = float(str(valores_adm[4]).replace('.', '').replace(',', '.'))
                    valor_admin_evento = (percentual / 100) * valor_total_adm
                
                # Registrar parcela vinculada ao evento, combinando as informações de evento e parcela
                proxima_linha = ws.max_row + 1
                ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                ws.cell(row=proxima_linha, column=26, value=i)  # Número do evento como número da parcela
                ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                ws.cell(row=proxima_linha, column=29, value=None)  # Data vencimento (vazio)
                ws.cell(row=proxima_linha, column=30, value=valor_admin_evento)  # Valor
                ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                ws.cell(row=proxima_linha, column=32, value=i)  # ID do evento vinculado
                ws.cell(row=proxima_linha, column=33, value=descricao.upper())  # Descrição do evento
                ws.cell(row=proxima_linha, column=34, value=f"{percentual:.2f}%")  # Percentual do evento
       
    def editar_contrato(self):
        """Edita o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um contrato para editar")
            return

        try:
            dados_contrato = self.tree_contratos.item(selecionado)['values']
            
            janela = tk.Toplevel(self.parent)
            janela.title(f"Editar Contrato - {self.cliente_atual}")
            janela.geometry("600x500")

            # Frame principal
            frame = ttk.Frame(janela, padding="10")
            frame.pack(fill='both', expand=True)

            # Dados do Contrato
            frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
            frame_contrato.pack(fill='x', pady=5)

            # Número do Contrato (readonly)
            ttk.Label(frame_contrato, text="Nº Contrato:").grid(row=0, column=0, padx=5, pady=2)
            num_contrato = ttk.Entry(frame_contrato, state='readonly')
            num_contrato.grid(row=0, column=1, padx=5, pady=2)
            num_contrato.insert(0, dados_contrato[0])

            # Datas
            ttk.Label(frame_contrato, text="Data Início:").grid(row=1, column=0, padx=5, pady=2)
            data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_inicio.grid(row=1, column=1, padx=5, pady=2)
            data_inicio.set_date(datetime.strptime(dados_contrato[1], '%d/%m/%Y'))

            ttk.Label(frame_contrato, text="Data Fim:").grid(row=2, column=0, padx=5, pady=2)
            data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_fim.grid(row=2, column=1, padx=5, pady=2)
            data_fim.set_date(datetime.strptime(dados_contrato[2], '%d/%m/%Y'))

            # Status
            ttk.Label(frame_contrato, text="Status:").grid(row=3, column=0, padx=5, pady=2)
            status_combo = ttk.Combobox(frame_contrato, values=['ATIVO', 'INATIVO'], state='readonly')
            status_combo.grid(row=3, column=1, padx=5, pady=2)
            status_combo.set(dados_contrato[3])

            def salvar_alteracoes():
                try:
                    wb = load_workbook(self.arquivo_cliente)
                    ws = wb['Contratos_ADM']
                    
                    # Atualizar dados do contrato
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value == dados_contrato[0]:
                            row[1].value = data_inicio.get_date()
                            row[2].value = data_fim.get_date()
                            row[3].value = status_combo.get()
                    
                    wb.save(self.arquivo_cliente)
                    custom_messagebox("info", "Sucesso", "Contrato atualizado com sucesso!")
                    janela.destroy()
                    self.carregar_contratos()
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao salvar alterações: {str(e)}")

            # Botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)

            ttk.Button(frame_botoes, text="Salvar", command=salvar_alteracoes).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir edição: {str(e)}")

    def adicionar_administrador_modificado(self, tree, valor_global_entry, metodo_pagamento_combo):
        """Versão modificada para incluir os detalhes de parcelas/eventos na tela do administrador"""
        # Verificar se valor global foi informado
        if not valor_global_entry.get():
            custom_messagebox("error", "Erro", "Informe o valor global do contrato primeiro")
            return
            
        try:
            valor_global_float = float(valor_global_entry.get().replace(',', '.'))
            if valor_global_float <= 0:
                custom_messagebox("error", "Erro", "Valor global deve ser maior que zero")
                return
        except ValueError:
            custom_messagebox("error", "Erro", "Valor global inválido")
            return
            
        # Obter o método de pagamento selecionado
        metodo = metodo_pagamento_combo.get()
        
        # Chamar método para abrir janela de administrador
        janela_admin = tk.Toplevel(self.parent)
        janela_admin.title("Adicionar Administrador")
        
        # Ajustar tamanho baseado no método (maior para eventos)
        if metodo == "Eventos/Fases":
            janela_admin.geometry("800x700")
        else:
            janela_admin.geometry("600x650")
        
        # Frame principal com scrollbar para permitir mais conteúdo
        main_frame = ttk.Frame(janela_admin)
        main_frame.pack(fill='both', expand=True)
        
        # Adicionar canvas com scrollbar se for Eventos/Fases
        if metodo == "Eventos/Fases":
            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            frame_admin = ttk.Frame(canvas)
            
            frame_admin.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=frame_admin, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
        else:
            frame_admin = ttk.Frame(main_frame, padding="10")
            frame_admin.pack(fill='both', expand=True)
        
        # Frame de busca
        frame_busca = ttk.LabelFrame(frame_admin, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        busca_entry = ttk.Entry(frame_busca, width=40)
        busca_entry.pack(side='left', padx=5)
        
        # Lista de fornecedores
        frame_lista = ttk.LabelFrame(frame_admin, text="Fornecedores")
        frame_lista.pack(fill='x', padx=5, pady=5)
        
        tree_fornecedores = ttk.Treeview(frame_lista,
                                    columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                    show='headings',
                                    height=3)
        tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        tree_fornecedores.heading('Nome', text='Nome')
        tree_fornecedores.heading('Categoria', text='Categoria')
        tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para dados do administrador
        frame_dados = ttk.LabelFrame(frame_admin, text="Dados do Administrador")
        frame_dados.pack(fill='x', padx=5, pady=5)
        
        # CNPJ/CPF
        ttk.Label(frame_dados, text="CNPJ/CPF:*").grid(row=0, column=0, padx=5, pady=2)
        cnpj_cpf_entry = ttk.Entry(frame_dados, state='readonly')
        cnpj_cpf_entry.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        
        # Nome
        ttk.Label(frame_dados, text="Nome/Razão Social:*").grid(row=1, column=0, padx=5, pady=2)
        nome_entry = ttk.Entry(frame_dados, state='readonly')
        nome_entry.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        
        # Mostrar informações do contrato (somente leitura)
        ttk.Label(frame_dados, text="Valor Global:").grid(row=2, column=0, padx=5, pady=2)
        valor_global_label = ttk.Label(frame_dados, text=f"R$ {valor_global_float:,.2f}")
        valor_global_label.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_dados, text="Método de Pagamento:").grid(row=3, column=0, padx=5, pady=2)
        metodo_label = ttk.Label(frame_dados, text=metodo)
        metodo_label.grid(row=3, column=1, padx=5, pady=2, sticky='w')
        
        # Tipo de remuneração
        ttk.Label(frame_dados, text="Tipo de Remuneração:*").grid(row=4, column=0, padx=5, pady=2)
        
        if metodo == "Percentual da Quinzena":
            tipo_combo = ttk.Combobox(frame_dados, values=['Percentual'], state='readonly')
            tipo_combo.grid(row=4, column=1, padx=5, pady=2, sticky='w')
            tipo_combo.set('Percentual')
            
            # Percentual
            ttk.Label(frame_dados, text="Percentual (%):*").grid(row=5, column=0, padx=5, pady=2)
            percentual_entry = ttk.Entry(frame_dados)
            percentual_entry.grid(row=5, column=1, padx=5, pady=2, sticky='w')
            
        else:  # Valor Fixo em Parcelas ou Eventos/Fases
            tipo_valores = ['Percentual', 'Fixo']
            tipo_combo = ttk.Combobox(frame_dados, values=tipo_valores, state='readonly')
            tipo_combo.grid(row=4, column=1, padx=5, pady=2)
            tipo_combo.set('Fixo')  # Padrão para eventos/parcelas fixas
            
            # Frame para valores percentuais
            frame_percentual_admin = ttk.Frame(frame_dados)
            frame_percentual_admin.grid(row=5, column=0, columnspan=2, pady=5)
            
            # Percentual
            ttk.Label(frame_percentual_admin, text="Percentual do Contrato (%):*").grid(row=0, column=0, padx=5, pady=2)
            percentual_entry = ttk.Entry(frame_percentual_admin)
            percentual_entry.grid(row=0, column=1, padx=5, pady=2)
            
            # Frame para valores fixos
            frame_fixo = ttk.Frame(frame_dados)
            frame_fixo.grid(row=6, column=0, columnspan=2, pady=5)
            
            # Valor Total
            ttk.Label(frame_fixo, text="Valor Total:*").grid(row=0, column=0, padx=5, pady=2)
            valor_total_entry = ttk.Entry(frame_fixo)
            valor_total_entry.grid(row=0, column=1, padx=5, pady=2)
            
            def atualizar_campos_tipo(*args):
                """Atualiza campos baseado no tipo selecionado"""
                if tipo_combo.get() == 'Percentual':
                    frame_percentual_admin.grid()
                    frame_fixo.grid_remove()
                elif tipo_combo.get() == 'Fixo':
                    frame_percentual_admin.grid_remove()
                    frame_fixo.grid()
                
            # Configurar evento
            tipo_combo.bind('<<ComboboxSelected>>', atualizar_campos_tipo)
            
            # Configurar interface inicial
            atualizar_campos_tipo()
        
        # Forma de pagamento para dados bancários
        ttk.Label(frame_dados, text="Forma de Pagamento:").grid(row=7, column=0, padx=5, pady=2)
        forma_pagamento = ttk.Combobox(frame_dados, values=['PIX', 'TED'], state='readonly')
        forma_pagamento.grid(row=7, column=1, padx=5, pady=2)
        forma_pagamento.set('PIX')  # Valor padrão
        
        # Área para configurações específicas do método de pagamento
        # ===============================================================
        # Frame para configurações específicas de método de pagamento
        frame_config_metodo = ttk.LabelFrame(frame_admin, text="Configuração de Pagamento")
        
        if metodo in ["Valor Fixo em Parcelas", "Eventos/Fases"]:
            frame_config_metodo.pack(fill='x', padx=5, pady=5, after=frame_dados)
        
        # 1. Frame para Parcelas Fixas
        if metodo == "Valor Fixo em Parcelas":
            frame_parcelas = ttk.Frame(frame_config_metodo)
            frame_parcelas.pack(fill='x', padx=5, pady=5)
            
            # Número de parcelas
            ttk.Label(frame_parcelas, text="Número de Parcelas:*").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            num_parcelas_entry = ttk.Entry(frame_parcelas, width=10)
            num_parcelas_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            
            # Checkbox para entrada
            var_tem_entrada = tk.BooleanVar(value=False)
            check_entrada = ttk.Checkbutton(frame_parcelas, text="Possui entrada?", variable=var_tem_entrada)
            check_entrada.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky='w')
            
            # Frame para entrada
            frame_entrada = ttk.Frame(frame_parcelas)
            frame_entrada.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky='w')
            
            ttk.Label(frame_entrada, text="Valor da Entrada:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            valor_entrada_entry = ttk.Entry(frame_entrada, width=15)
            valor_entrada_entry.grid(row=0, column=1, padx=5, pady=2, sticky='w')
            
            ttk.Label(frame_entrada, text="Data da Entrada:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
            data_entrada = DateEntry(frame_entrada, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_entrada.grid(row=1, column=1, padx=5, pady=2, sticky='w')
            
            ttk.Label(frame_entrada, text="Descrição da Entrada:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
            descricao_entrada = ttk.Entry(frame_entrada, width=40)
            descricao_entrada.grid(row=2, column=1, padx=5, pady=2, sticky='w')
            descricao_entrada.insert(0, "ENTRADA")  # Valor padrão
            
            # Ocultar frame de entrada inicialmente
            frame_entrada.grid_remove()
            
            # Função para mostrar/ocultar frame de entrada
            def toggle_entrada():
                if var_tem_entrada.get():
                    frame_entrada.grid()
                else:
                    frame_entrada.grid_remove()
            
            # Configurar checkbox para chamar a função
            check_entrada.config(command=toggle_entrada)
            
            # Frame para gerenciar descrições individuais das parcelas
            frame_descricoes = ttk.LabelFrame(frame_parcelas, text="Descrições Individuais das Parcelas")
            frame_descricoes.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky='ew')
            
            ttk.Label(frame_descricoes, text="Para configurar descrições individuais, primeiro defina o número de parcelas e clique em:").grid(
                row=0, column=0, columnspan=2, padx=5, pady=2, sticky='w')
            
            # Lista para armazenar descrições individuais
            descricoes_parcelas = []
            
            def configurar_descricoes_parcelas():
                try:
                    # Validar número de parcelas
                    if not num_parcelas_entry.get():
                        custom_messagebox("error", "Erro", "Informe o número de parcelas primeiro")
                        return
                        
                    num_parcelas = int(num_parcelas_entry.get())
                    if num_parcelas <= 0:
                        custom_messagebox("error", "Erro", "Número de parcelas deve ser maior que zero")
                        return
                    
                    # Criar janela para configurar descrições
                    janela_descricoes = tk.Toplevel(janela_admin)
                    janela_descricoes.title("Descrições Individuais das Parcelas")
                    janela_descricoes.geometry("500x600")
                    
                    # Frame com scrollbar
                    frame_scroll = ttk.Frame(janela_descricoes)
                    frame_scroll.pack(fill='both', expand=True, padx=10, pady=10)
                    
                    canvas = tk.Canvas(frame_scroll)
                    scrollbar = ttk.Scrollbar(frame_scroll, orient="vertical", command=canvas.yview)
                    frame_content = ttk.Frame(canvas)
                    
                    frame_content.bind(
                        "<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                    )
                    
                    canvas.create_window((0, 0), window=frame_content, anchor="nw")
                    canvas.configure(yscrollcommand=scrollbar.set)
                    
                    canvas.pack(side="left", fill="both", expand=True)
                    scrollbar.pack(side="right", fill="y")
                    
                    # Inicializar ou redimensionar a lista de descrições
                    if len(descricoes_parcelas) < num_parcelas:
                        # Adicionar novas entradas para as parcelas adicionais
                        for _ in range(num_parcelas - len(descricoes_parcelas)):
                            descricoes_parcelas.append("")
                    else:
                        # Truncar a lista se o número de parcelas diminuiu
                        del descricoes_parcelas[num_parcelas:]
                    
                    # Criar campos para cada parcela
                    for i in range(num_parcelas):
                        ttk.Label(frame_content, text=f"Parcela {i+1}:").grid(
                            row=i, column=0, padx=5, pady=5, sticky='w')
                        
                        desc_entry = ttk.Entry(frame_content, width=40)
                        desc_entry.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                        
                        # Preencher com valor existente, se houver
                        if i < len(descricoes_parcelas) and descricoes_parcelas[i]:
                            desc_entry.insert(0, descricoes_parcelas[i])
                        else:
                            desc_entry.insert(0, f"PARCELA {i+1}")
                        
                        # Armazenar referência à entrada para recuperar valores depois
                        desc_entry.idx = i
                    
                    def salvar_descricoes():
                        # Coletar todas as descrições dos campos
                        for child in frame_content.winfo_children():
                            if isinstance(child, ttk.Entry):
                                idx = getattr(child, 'idx', -1)
                                if 0 <= idx < len(descricoes_parcelas):
                                    descricoes_parcelas[idx] = child.get().strip()
                        
                        # Confirmar para o usuário
                        custom_messagebox("info", "Sucesso", "Descrições salvas!")
                        janela_descricoes.destroy()
                    
                    # Botões
                    frame_botoes = ttk.Frame(janela_descricoes)
                    frame_botoes.pack(fill='x', pady=10)
                    
                    ttk.Button(frame_botoes, text="Salvar Descrições", 
                            command=salvar_descricoes).pack(side='right', padx=10)
                    
                    ttk.Button(frame_botoes, text="Cancelar", 
                            command=janela_descricoes.destroy).pack(side='right', padx=10)
                    
                    # Centralizar a janela
                    janela_descricoes.update_idletasks()
                    w = janela_descricoes.winfo_width()
                    h = janela_descricoes.winfo_height()
                    x = (janela_descricoes.winfo_screenwidth() // 2) - (w // 2)
                    y = (janela_descricoes.winfo_screenheight() // 2) - (h // 2)
                    janela_descricoes.geometry(f'{w}x{h}+{x}+{y}')
                    
                    # Tornar a janela modal
                    janela_descricoes.transient(janela_admin)
                    janela_descricoes.grab_set()
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao configurar descrições: {str(e)}")
                    
            # Botão para configurar descrições individuais
            btn_config_descricoes = ttk.Button(frame_descricoes, 
                                            text="Configurar Descrições", 
                                            command=configurar_descricoes_parcelas)
            btn_config_descricoes.grid(row=1, column=0, padx=5, pady=5, sticky='w')
        
        # 2. Frame para Eventos/Fases
        elif metodo == "Eventos/Fases":
            frame_eventos = ttk.Frame(frame_config_metodo)
            frame_eventos.pack(fill='x', padx=5, pady=5)
            
            # Lista de eventos
            colunas_evento = ('Nº', 'Descrição', 'Percentual', 'Valor')
            tree_eventos = ttk.Treeview(frame_eventos, columns=colunas_evento, show='headings', height=5)
            tree_eventos.heading('Nº', text='Nº')
            tree_eventos.heading('Descrição', text='Descrição')
            tree_eventos.heading('Percentual', text='Percentual (%)')
            tree_eventos.heading('Valor', text='Valor (R$)')
            
            tree_eventos.column('Nº', width=50, anchor='center')
            tree_eventos.column('Descrição', width=300)
            tree_eventos.column('Percentual', width=100, anchor='e')
            tree_eventos.column('Valor', width=100, anchor='e')
            
            # Adicionar scrollbars para eventos
            scroll_y_eventos = ttk.Scrollbar(frame_eventos, orient='vertical', command=tree_eventos.yview)
            scroll_x_eventos = ttk.Scrollbar(frame_eventos, orient='horizontal', command=tree_eventos.xview)
            tree_eventos.configure(yscrollcommand=scroll_y_eventos.set, xscrollcommand=scroll_x_eventos.set)
            
            tree_eventos.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y_eventos.pack(side='right', fill='y')
            scroll_x_eventos.pack(side='bottom', fill='x')
            
            # Frame para adicionar evento
            frame_add_evento = ttk.Frame(frame_eventos)
            frame_add_evento.pack(fill='x', padx=5, pady=5)
            
            ttk.Label(frame_add_evento, text="Descrição:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            evento_descricao = ttk.Entry(frame_add_evento, width=40)
            evento_descricao.grid(row=0, column=1, padx=5, pady=2, sticky='w')
            
            ttk.Label(frame_add_evento, text="Percentual (%):").grid(row=0, column=2, padx=5, pady=2, sticky='w')
            evento_percentual = ttk.Entry(frame_add_evento, width=10)
            evento_percentual.grid(row=0, column=3, padx=5, pady=2, sticky='w')
            
            # Botões para eventos
            frame_botoes_evento = ttk.Frame(frame_eventos)
            frame_botoes_evento.pack(fill='x', pady=5)
            
            # Variável para rastrear o total de percentuais
            total_percentual_var = tk.StringVar(value="Total: 0%")
            lbl_total_percentual = ttk.Label(frame_botoes_evento, textvariable=total_percentual_var)
            lbl_total_percentual.pack(side='left', padx=5)
            
            # Lista para armazenar eventos
            eventos = []
            
            def calcular_valor_evento(percentual, valor_total_str):
                """Calcula o valor do evento baseado no percentual e valor total"""
                try:
                    percentual_float = float(percentual.replace(',', '.'))
                    valor_float = float(valor_total_str.replace(',', '.'))
                    return (percentual_float / 100) * valor_float
                except (ValueError, AttributeError):
                    return 0
            
            def adicionar_evento():
                """Adiciona um evento à lista"""
                if not valor_global_entry.get():
                    custom_messagebox("error", "Erro", "Informe o valor global do contrato primeiro")
                    return
                    
                descricao = evento_descricao.get().strip()
                percentual_str = evento_percentual.get().strip()
                
                if not descricao:
                    custom_messagebox("error", "Erro", "Informe a descrição do evento")
                    return
                    
                try:
                    percentual = float(percentual_str.replace(',', '.'))
                    if percentual <= 0 or percentual > 100:
                        custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100")
                        return
                except ValueError:
                    custom_messagebox("error", "Erro", "Percentual inválido")
                    return
                    
                # Calcular total atual
                total_atual = sum(float(e[1]) for e in eventos)
                
                # Verificar se ultrapassa 100%
                if total_atual + percentual > 100:
                    custom_messagebox("error", "Erro", "Total de percentual não pode exceder 100%")
                    return
                    
                # Calcular valor baseado no percentual
                valor_total = valor_global_entry.get().replace(',', '.')
                try:
                    valor_total_float = float(valor_total)
                    valor_evento = (percentual / 100) * valor_total_float
                except (ValueError, TypeError):
                    valor_evento = 0
                    
                # Adicionar à lista
                eventos.append((descricao, percentual, valor_evento))
                
                # Adicionar ao treeview
                tree_eventos.insert('', 'end', values=(
                    len(eventos),  # Número sequencial
                    descricao, 
                    f"{percentual:.2f}", 
                    f"R$ {valor_evento:.2f}"
                ))
                
                # Atualizar total
                total_percentual_var.set(f"Total: {total_atual + percentual:.2f}%")
                
                # Limpar campos
                evento_descricao.delete(0, tk.END)
                evento_percentual.delete(0, tk.END)
                
            def remover_evento():
                """Remove o evento selecionado"""
                selecionado = tree_eventos.selection()
                if not selecionado:
                    custom_messagebox("warning",  "Aviso", "Selecione um evento para remover")
                    return
                    
                # Obter valores
                valores = tree_eventos.item(selecionado)['values']
                indice = int(valores[0]) - 1  # Ajusta para índice 0-based
                
                if 0 <= indice < len(eventos):
                    # Remover da lista
                    eventos.pop(indice)
                    
                    # Limpar e recriar treeview para atualizar numeração
                    for item in tree_eventos.get_children():
                        tree_eventos.delete(item)
                        
                    for i, (desc, perc, valor) in enumerate(eventos, 1):
                        tree_eventos.insert('', 'end', values=(i, desc, f"{perc:.2f}", f"R$ {valor:.2f}"))
                    
                    # Atualizar total
                    total_atual = sum(float(e[1]) for e in eventos)
                    total_percentual_var.set(f"Total: {total_atual:.2f}%")
            
            # Configurar botões de eventos
            ttk.Button(frame_botoes_evento, text="Adicionar Evento", command=adicionar_evento).pack(side='right', padx=5)
            ttk.Button(frame_botoes_evento, text="Remover Evento", command=remover_evento).pack(side='right', padx=5)
        
        # Função de busca para fornecedores
        def busca_local():
            """Função de busca"""
            termo = busca_entry.get()
            buscar_fornecedor(tree_fornecedores, termo)
            
        ttk.Button(frame_busca, text="Buscar", command=busca_local).pack(side='left', padx=5)
        busca_entry.bind('<Return>', lambda e: busca_local())
        
        def selecionar_e_preencher(event=None):
            """Seleciona fornecedor e preenche campos"""
            selecionado = tree_fornecedores.selection()
            if not selecionado:
                return
                
            valores = tree_fornecedores.item(selecionado)['values']
            cnpj_cpf_entry.config(state='normal')
            nome_entry.config(state='normal')
            
            cnpj_cpf_entry.delete(0, tk.END)
            cnpj_cpf_entry.insert(0, str(valores[0]).zfill(14))
            
            nome_entry.delete(0, tk.END)
            nome_entry.insert(0, valores[1])
            
            cnpj_cpf_entry.config(state='readonly')
            nome_entry.config(state='readonly')
            
        tree_fornecedores.bind('<Double-1>', selecionar_e_preencher)
        
        def confirmar():
            """Confirma a adição do administrador"""
            try:
                if not cnpj_cpf_entry.get() or not nome_entry.get() or not tipo_combo.get():
                    custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
                    return
                    
                # Capturar a forma de pagamento para os dados bancários
                forma_pagto_selecionada = forma_pagamento.get()
                
                # Verificar configuração específica do método
                if metodo == "Valor Fixo em Parcelas":
                    # Validar número de parcelas
                    if not num_parcelas_entry.get():
                        custom_messagebox("error", "Erro", "Informe o número de parcelas!")
                        return
                        
                    try:
                        num_parcelas = int(num_parcelas_entry.get())
                        if num_parcelas <= 0:
                            custom_messagebox("error", "Erro", "Número de parcelas deve ser maior que zero!")
                            return
                    except ValueError:
                        custom_messagebox("error", "Erro", "Número de parcelas inválido!")
                        return
                    
                    # Se tem entrada configurada, validar entrada
                    if var_tem_entrada.get():
                        if not valor_entrada_entry.get():
                            custom_messagebox("error", "Erro", "Informe o valor da entrada!")
                            return
                        
                        try:
                            valor_entrada = float(valor_entrada_entry.get().replace(',', '.'))
                            if valor_entrada <= 0:
                                custom_messagebox("error", "Erro", "Valor da entrada deve ser maior que zero!")
                                return
                        except ValueError:
                            custom_messagebox("error", "Erro", "Valor da entrada inválido!")
                            return
                
                # Verificar eventos para contratos do tipo Eventos/Fases
                if metodo == "Eventos/Fases" and not eventos:
                    custom_messagebox("error", "Erro", "Adicione pelo menos um evento para este administrador!")
                    return
                    
                # Para contratos de eventos, verificar total de percentuais
                if metodo == "Eventos/Fases":
                    total_percentual = sum(float(e[1]) for e in eventos)
                    if total_percentual < 99.99 or total_percentual > 100.01:  # Pequena margem de erro
                        if not custom_messagebox("yesno", "Confirmação", 
                                            f"O total de percentuais é {total_percentual:.2f}% ao invés de 100%. Deseja continuar mesmo assim?"):
                            return
            
                if tipo_combo.get() == 'Percentual':
                    # Validar percentual
                    if not percentual_entry.get():
                        custom_messagebox("error", "Erro", "Preencha o percentual!")
                        return
                    
                    try:
                        perc = float(percentual_entry.get().replace(',', '.'))
                        if perc <= 0 or perc > 100:
                            custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100!")
                            return
                            
                        # Configurar campos adicionais conforme método
                        if metodo == "Percentual da Quinzena":
                            # Simples percentual da quinzena
                            num_parcelas = ""
                            data_inicial = ""
                        elif metodo == "Valor Fixo em Parcelas":
                            # Número de parcelas informado no contrato
                            num_parcelas = num_parcelas_entry.get()
                            # Data inicial se houver entrada
                            data_inicial = data_entrada.get() if var_tem_entrada.get() else ""
                        else:  # Eventos/Fases
                            # Número de eventos
                            num_parcelas = str(len(eventos))
                            data_inicial = ""
                            
                        # Adicionar registro de percentual
                        valores_percentual = (
                            cnpj_cpf_entry.get(),
                            nome_entry.get(),
                            tipo_combo.get(),
                            f"{perc:.2f}%",  # Formatação com %
                            f"{valor_global_float:.2f}",  # Valor Total
                            num_parcelas,  # Número de parcelas conforme método
                            data_inicial  # Data inicial conforme método
                        )
                        
                        # Preparar tags adicionais
                        tags_extra = []
                        
                        # Adicionar descrições individuais como tag se for Valor Fixo em Parcelas
                        if metodo == "Valor Fixo em Parcelas" and descricoes_parcelas:
                            tags_extra.append(f"descricoes:{','.join(descricoes_parcelas)}")
                            
                        # Adicionar informações de entrada se necessário
                        if metodo == "Valor Fixo em Parcelas" and var_tem_entrada.get():
                            tags_extra.append(f"desc_entrada:{descricao_entrada.get()}")
                        
                        # Tags finais incluem tipo de percentual, forma de pagamento e tags extras
                        tags_finais = ['percentual', forma_pagto_selecionada, *tags_extra]
                        
                        tree.insert('', 'end', values=valores_percentual, tags=tags_finais)
                        
                    except ValueError:
                        custom_messagebox("error", "Erro", "Percentual inválido!")
                        return
                        
                elif tipo_combo.get() == 'Fixo':
                    if not valor_total_entry.get():
                        custom_messagebox("error", "Erro", "Preencha o valor total!")
                        return
                        
                    try:
                        valor_total_adm = float(valor_total_entry.get().replace(',', '.'))
                        if valor_total_adm <= 0:
                            custom_messagebox("error", "Erro", "Valor total deve ser maior que zero!")
                            return
                    except ValueError:
                        custom_messagebox("error", "Erro", "Valor total inválido!")
                        return
                    
                    # Configurar campos adicionais conforme método
                    if metodo == "Valor Fixo em Parcelas":
                        # Número de parcelas informado no contrato
                        num_parcelas = num_parcelas_entry.get()
                        # Data inicial se houver entrada
                        data_inicial = data_entrada.get() if var_tem_entrada.get() else ""
                        # Descrição da entrada
                        desc_entrada = descricao_entrada.get() if var_tem_entrada.get() else ""
                    else:  # Eventos/Fases
                        # Número de eventos
                        num_parcelas = str(len(eventos))
                        data_inicial = ""
                        desc_entrada = ""
                        
                    # Adicionar registro de valor fixo
                    valores_fixo = (
                        cnpj_cpf_entry.get(),
                        nome_entry.get(),
                        tipo_combo.get(),
                        "",  # Sem percentual para fixo
                        valor_total_entry.get(),
                        num_parcelas,
                        data_inicial
                    )
                    
                    # Adicionar tags extras para armazenar informações
                    tags_extras = []
                    
                    # Valores específicos para parcelas fixas
                    if metodo == "Valor Fixo em Parcelas":
                        # Valor e descrição de entrada se houver
                        if var_tem_entrada.get():
                            tags_extras.append(f"entrada:{valor_entrada_entry.get()}")
                            tags_extras.append(f"desc_entrada:{descricao_entrada.get()}")
                        
                        # Adicionar descrições individuais
                        if descricoes_parcelas:
                            tags_extras.append(f"descricoes:{','.join(descricoes_parcelas)}")
                    
                    # Tags completas: tipo fixo, forma de pagamento e extras
                    tags = [
                        'fixo', 
                        forma_pagto_selecionada,
                        *tags_extras
                    ]
                    
                    tree.insert('', 'end', values=valores_fixo, tags=tags)
                
                # Se tiver eventos, registrá-los na lista global
                if metodo == "Eventos/Fases":
                    # Armazenar eventos como tags adicionais no item
                    eventos_serializados = []
                    for desc, perc, valor in eventos:
                        eventos_serializados.append(f"{desc}:{perc}:{valor}")
                        
                    # Atualizar tags do item para incluir eventos
                    for item in tree.get_children():
                        # Pegar o último item inserido (mais recente)
                        if item == tree.get_children()[-1]:
                            tags_atuais = tree.item(item)['tags']
                            # Adicionar a tag com eventos
                            nova_tag = f"eventos:{'|'.join(eventos_serializados)}"
                            tree.item(item, tags=(*tags_atuais, nova_tag))
                
                # Fechar a janela
                janela_admin.destroy()
                
                # Garantir que a janela do contrato seja trazida para frente
                # Usar after para garantir que a janela tenha tempo de ser destruída primeiro
                if metodo_pagamento_combo.winfo_toplevel().winfo_exists():
                    metodo_pagamento_combo.winfo_toplevel().after(100, lambda: (
                        metodo_pagamento_combo.winfo_toplevel().lift(),
                        metodo_pagamento_combo.winfo_toplevel().focus_force()
                    ))
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao confirmar: {str(e)}")
                
        # Botões
        frame_botoes = ttk.Frame(frame_admin)
        frame_botoes.pack(fill='x', pady=10)
        ttk.Button(frame_botoes, text="Confirmar", command=confirmar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela_admin.destroy).pack(side='left', padx=5)       

    def processar_parcelas_fixas(self, ws, num_contrato, valor_global, opcoes):
        """Processa parcelas fixas para o contrato"""
        try:
            # Debug
            print("Início de processar_parcelas_fixas")
            print(f"Opções: {opcoes}")
            
            num_parcelas = int(opcoes.get('num_parcelas', 0))
            tem_entrada = opcoes.get('tem_entrada', False)
            descricoes_parcelas = opcoes.get('descricoes_parcelas', {})  # Dicionário com descrições por admin
            
            print(f"Processando {num_parcelas} parcelas, entrada: {tem_entrada}")
            
            if num_parcelas <= 0:
                print("Erro: Número de parcelas inválido")
                return
                
            # Processar cada administrador
            for item in self.tree_adm.get_children():
                valores_adm = self.tree_adm.item(item)['values']
                tags_adm = self.tree_adm.item(item)['tags']
                
                print(f"Processando administrador: {valores_adm}")
                
                cnpj_cpf_adm = str(valores_adm[0]).strip()
                cnpj_cpf_adm = formatar_cnpj_cpf(cnpj_cpf_adm)
                nome_adm = valores_adm[1]
                
                # Extrair descricoes das tags, se existirem
                descricoes_individuais = []
                for tag in tags_adm:
                    if tag.startswith('descricoes:'):
                        descricoes_individuais = tag.replace('descricoes:', '').split(',')
                        print(f"Descrições individuais: {descricoes_individuais}")
                        break
                
                # Também verificar no dicionário de descrições
                if not descricoes_individuais and cnpj_cpf_adm in descricoes_parcelas:
                    descricoes_individuais = descricoes_parcelas[cnpj_cpf_adm]
                    print(f"Usando descrições do dicionário: {descricoes_individuais}")
                
                # Extrair descrição da entrada, se existir
                descricao_entrada = "ENTRADA"
                for tag in tags_adm:
                    if tag.startswith('desc_entrada:'):
                        descricao_entrada = tag.replace('desc_entrada:', '')
                        print(f"Descrição da entrada: {descricao_entrada}")
                        break
                
                # Calcular valor por parcela para este administrador
                try:
                    if valores_adm[2] == 'Percentual':
                        # Administrador com percentual do valor total
                        perc_adm = float(str(valores_adm[3]).replace('%', '').replace(',', '.'))
                        valor_total_adm = (perc_adm / 100) * valor_global
                        print(f"Valor calculado baseado em percentual: {valor_total_adm}")
                    else:  # Fixo
                        # Valor fixo total para o administrador
                        valor_texto = str(valores_adm[4]).replace(',', '.')
                        print(f"Valor texto: {valor_texto}")
                        valor_total_adm = float(valor_texto)
                        print(f"Valor fixo: {valor_total_adm}")
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Erro ao calcular valor: {e}")
                    valores_str = ', '.join([str(v) for v in valores_adm])
                    print(f"Valores disponíveis: {valores_str}")
                    # Tentar alternativa
                    if len(valores_adm) >= 5 and valores_adm[4]:
                        try:
                            valor_total_adm = float(str(valores_adm[4]).replace(',', '.'))
                            print(f"Valor alternativo: {valor_total_adm}")
                        except (ValueError, TypeError):
                            print("Erro na alternativa também")
                            valor_total_adm = 0
                    else:
                        valor_total_adm = 0
                
                if valor_total_adm <= 0:
                    print("Valor total inválido, pulando administrador")
                    continue
                
                # Se tem entrada, tratar separadamente
                if tem_entrada:
                    valor_entrada = 0
                    # Buscar valor da entrada nas tags
                    for tag in tags_adm:
                        if tag.startswith('entrada:'):
                            try:
                                valor_entrada = float(tag.replace('entrada:', '').replace(',', '.'))
                                print(f"Valor da entrada das tags: {valor_entrada}")
                            except ValueError:
                                valor_entrada = 0
                            break
                    
                    if valor_entrada <= 0:
                        # Calcular proporcional se não estiver explícito
                        valor_entrada_opcoes = opcoes.get('valor_entrada', 0)
                        if isinstance(valor_entrada_opcoes, str):
                            valor_entrada_opcoes = float(valor_entrada_opcoes.replace(',', '.'))
                        # Proporcional da entrada para este administrador
                        proporcao_entrada = valor_entrada_opcoes / valor_global if valor_global else 0
                        valor_entrada_adm = valor_total_adm * proporcao_entrada
                        print(f"Valor da entrada calculado: {valor_entrada_adm}")
                    else:
                        # Usar o valor específico
                        valor_entrada_adm = valor_entrada
                        
                    data_entrada = opcoes.get('data_entrada')
                    print(f"Data da entrada: {data_entrada}")
                    
                    # Registrar entrada como primeira parcela
                    proxima_linha = ws.max_row + 1
                    ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                    ws.cell(row=proxima_linha, column=26, value=1)  # Número da parcela (entrada = 1)
                    ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                    ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                    ws.cell(row=proxima_linha, column=29, value=data_entrada)  # Data vencimento
                    ws.cell(row=proxima_linha, column=30, value=valor_entrada_adm)  # Valor
                    ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                    ws.cell(row=proxima_linha, column=32, value="")  # Sem evento
                    ws.cell(row=proxima_linha, column=33, value=descricao_entrada.upper())  # Descrição personalizada da entrada
                    
                    print(f"Registrada entrada com valor {valor_entrada_adm}")
                    
                    # Ajustar valor restante para as demais parcelas
                    valor_restante = valor_total_adm - valor_entrada_adm
                    valor_parcela = valor_restante / num_parcelas if num_parcelas > 0 else 0
                    
                    print(f"Valor de cada parcela após entrada: {valor_parcela}")
                    
                    # Registrar as demais parcelas
                    for i in range(1, num_parcelas + 1):
                        proxima_linha = ws.max_row + 1
                        ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                        ws.cell(row=proxima_linha, column=26, value=i + 1)  # Número da parcela (após entrada)
                        ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                        ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                        ws.cell(row=proxima_linha, column=29, value=None)  # Data vencimento (a definir)
                        ws.cell(row=proxima_linha, column=30, value=valor_parcela)  # Valor
                        ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                        ws.cell(row=proxima_linha, column=32, value="")  # Sem evento
                        
                        # Usar descrição individual se disponível
                        if i-1 < len(descricoes_individuais) and descricoes_individuais[i-1]:
                            descricao = descricoes_individuais[i-1]
                        else:
                            descricao = f"PARCELA {i}"
                            
                        ws.cell(row=proxima_linha, column=33, value=descricao.upper())  # Descrição individual ou genérica
                        print(f"Registrada parcela {i} com valor {valor_parcela} e descrição '{descricao}'")
                            
                else:
                    # Sem entrada, distribuir igualmente
                    valor_parcela = valor_total_adm / num_parcelas if num_parcelas > 0 else 0
                    print(f"Valor de cada parcela (sem entrada): {valor_parcela}")
                    
                    # Registrar parcelas
                    for i in range(1, num_parcelas + 1):
                        proxima_linha = ws.max_row + 1
                        ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                        ws.cell(row=proxima_linha, column=26, value=i)  # Número da parcela
                        ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                        ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                        ws.cell(row=proxima_linha, column=29, value=None)  # Data vencimento (a definir)
                        ws.cell(row=proxima_linha, column=30, value=valor_parcela)  # Valor
                        ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                        ws.cell(row=proxima_linha, column=32, value="")  # Sem evento
                        
                        # Usar descrição individual se disponível
                        if i-1 < len(descricoes_individuais) and descricoes_individuais[i-1]:
                            descricao = descricoes_individuais[i-1]
                        else:
                            descricao = f"PARCELA {i}"
                            
                        ws.cell(row=proxima_linha, column=33, value=descricao.upper())  # Descrição individual ou genérica
                        print(f"Registrada parcela {i} com valor {valor_parcela} e descrição '{descricao}'")
            
            print("Finalizado processamento de parcelas fixas com sucesso")
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"Erro em processar_parcelas_fixas: {str(e)}") 

    def processar_administradores(self, ws, num_contrato, valor_global, metodo_pagamento, opcoes):
        """Processa os administradores do contrato"""
        for item in self.tree_adm.get_children():
            valores = self.tree_adm.item(item)['values']
            tags = self.tree_adm.item(item)['tags']
            
            # Formatação do CNPJ/CPF
            cnpj_cpf = str(valores[0]).strip()
            cnpj_cpf = formatar_cnpj_cpf(cnpj_cpf)
            nome_admin = valores[1]
            
            # Buscar dados bancários do fornecedor
            forma_pagamento = next((tag for tag in tags if tag in ['PIX', 'TED']), 'PIX')
            dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj_cpf, forma_pagamento)

            # Registrar administrador no contrato com os dados apropriados
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=7, value=num_contrato.upper())  # Contrato
            ws.cell(row=proxima_linha, column=8, value=cnpj_cpf)              # CNPJ/CPF
            ws.cell(row=proxima_linha, column=9, value=nome_admin)            # Nome
            ws.cell(row=proxima_linha, column=10, value=valores[2])           # Tipo (Percentual/Fixo)
            ws.cell(row=proxima_linha, column=11, value=valores[3])           # Valor/Percentual
            ws.cell(row=proxima_linha, column=12, value=valores[4])           # Valor Total
            ws.cell(row=proxima_linha, column=13, value=valores[5])           # Número de parcelas
            
            # Data inicial para casos que têm entrada
            if valores[6] and metodo_pagamento == "Valor Fixo em Parcelas" and opcoes.get('tem_entrada'):
                ws.cell(row=proxima_linha, column=14, value=opcoes.get('data_entrada'))  # Data inicial

    def salvar_contrato_com_opcoes(self, num_contrato, data_inicio, data_fim, observacoes, valor_global, metodo_pagamento, opcoes, janela):
        """Salva os dados do contrato com diferentes opções de pagamento"""
        num_contrato = str(num_contrato).upper()
        
        try:
            # Adicione instruções de debug para verificar o fluxo
            print(f"Salvando contrato: {num_contrato}, método: {metodo_pagamento}")
            
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            # Verificar se o contrato já existe
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).upper() == num_contrato.upper():
                    custom_messagebox("error", "Erro", "Número de contrato já existe!")
                    return

            # Salvar dados do contrato
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1, value=num_contrato.upper())
            ws.cell(row=proxima_linha, column=2, value=data_inicio)
            ws.cell(row=proxima_linha, column=3, value=data_fim)
            ws.cell(row=proxima_linha, column=4, value='ATIVO')
            ws.cell(row=proxima_linha, column=5, value=observacoes)
            ws.cell(row=proxima_linha, column=6, value=valor_global)  # Valor global do contrato

            # Extrair dados dos administradores, incluindo descrições de parcelas
            opcoes_processadas = opcoes.copy() if opcoes else {}
            
            # Extrair informações adicionais dos administradores
            tem_entrada = False
            valor_entrada = 0
            data_entrada = None
            num_parcelas = 0
            
            # Coletar informações específicas para parcelas fixas
            if metodo_pagamento == "Valor Fixo em Parcelas":
                # Percorrer os administradores para extrair informações de parcelas
                for item in self.tree_adm.get_children():
                    valores = self.tree_adm.item(item)['values']
                    tags = self.tree_adm.item(item)['tags']
                    
                    # Extrair número de parcelas
                    if valores[5] and not num_parcelas:
                        try:
                            num_parcelas = int(valores[5])
                        except (ValueError, TypeError):
                            num_parcelas = 0
                    
                    # Verificar se tem entrada
                    for tag in tags:
                        if tag.startswith('entrada:'):
                            tem_entrada = True
                            try:
                                valor_entrada = float(tag.replace('entrada:', '').replace(',', '.'))
                            except (ValueError, TypeError):
                                valor_entrada = 0
                    
                    # Extrair data de entrada
                    if valores[6] and not data_entrada:
                        data_entrada = valores[6]

                # Adicionar ao dicionário de opções
                opcoes_processadas['num_parcelas'] = num_parcelas
                opcoes_processadas['tem_entrada'] = tem_entrada
                opcoes_processadas['valor_entrada'] = valor_entrada
                opcoes_processadas['data_entrada'] = data_entrada
                
                print(f"Configurações de parcelas: parcelas={num_parcelas}, entrada={tem_entrada}, valor_entrada={valor_entrada}")
            
            # Coletar descrições para cada administrador
            admin_descricoes = {}
            
            for item in self.tree_adm.get_children():
                tags = self.tree_adm.item(item)['tags']
                cnpj_cpf = self.tree_adm.item(item)['values'][0]
                
                # Extrair descricoes das tags, se existirem
                for tag in tags:
                    if tag.startswith('descricoes:'):
                        admin_descricoes[cnpj_cpf] = tag.replace('descricoes:', '').split(',')
                        print(f"Descrições para {cnpj_cpf}: {admin_descricoes[cnpj_cpf]}")
                        break
            
            # Adicionar ao dicionário de opções
            opcoes_processadas['descricoes_parcelas'] = admin_descricoes

            # Processar administradores baseado no método de pagamento
            self.processar_administradores(ws, num_contrato, valor_global, metodo_pagamento, opcoes_processadas)

            # Processar eventos se método for por eventos/fases
            if metodo_pagamento == "Eventos/Fases":
                # Extrair eventos dos administradores
                eventos = []
                for item in self.tree_adm.get_children():
                    tags = self.tree_adm.item(item)['tags']
                    for tag in tags:
                        if tag.startswith('eventos:'):
                            eventos_str = tag.replace('eventos:', '')
                            for evento_str in eventos_str.split('|'):
                                partes = evento_str.split(':')
                                if len(partes) == 3:
                                    desc, perc, valor = partes
                                    eventos.append((desc, float(perc), float(valor)))
                            break

                self.processar_eventos(ws, num_contrato, valor_global, eventos)
                        
            # Processar parcelas fixas se for o método apropriado
            elif metodo_pagamento == "Valor Fixo em Parcelas":
                print("Chamando processar_parcelas_fixas...")
                self.processar_parcelas_fixas(ws, num_contrato, valor_global, opcoes_processadas)

            # Salvar e fechar o arquivo explicitamente
            try:
                print(f"Salvando o arquivo {self.arquivo_cliente}")
                wb.save(self.arquivo_cliente)
                wb.close()  # Importante fechar o arquivo
            except PermissionError:
                custom_messagebox("error", "Erro", f"Não foi possível salvar a planilha. Ela pode estar aberta em outro programa.")
                return
            except Exception as e:
                import traceback
                traceback.print_exc()
                custom_messagebox("error", "Erro", f"Erro ao salvar planilha: {str(e)}")
                return
                
            # Exibir mensagem de sucesso
            custom_messagebox("info", "Sucesso", "Contrato cadastrado com sucesso!")
            
            # Fechar a janela atual
            janela.destroy()
            
            # Garantir que a janela de gestão de contratos é trazida para frente
            # após salvar o contrato e recarregar a lista
            self.carregar_contratos()
            
            # Usar after para garantir que toda a interface seja atualizada
            if self.parent and self.parent.winfo_exists():
                self.parent.after(100, lambda: (
                    self.parent.lift(),
                    self.parent.focus_force()
                ))

        except Exception as e:
            import traceback
            traceback.print_exc()  # Imprime o stack trace completo
            custom_messagebox("error", "Erro", f"Erro ao salvar contrato: {str(e)}")
            if 'wb' in locals() and wb:
                try:
                    wb.close()
                except:
                    pass

    def excluir_contrato(self):
        """Exclui o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um contrato para excluir")
            return
            
        if custom_messagebox("yesno", "Confirmação", 
                              "Deseja realmente excluir este contrato e seus administradores?"):
            try:
                num_contrato = self.tree_contratos.item(selecionado)['values'][0]
                
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']
                
                # Marcar contrato como inativo
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == num_contrato:
                        row[3].value = 'INATIVO'  # Coluna D - Status
                
                wb.save(self.arquivo_cliente)
                self.carregar_contratos()
                custom_messagebox("info", "Sucesso", "Contrato marcado como inativo")
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao excluir contrato: {str(e)})")
        
class GestaoTaxasFixas:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.gestor_parcelas = GestorParcelas(self)

    def processar_lancamentos_fixos(self, cliente, data_ref):
        """Processa os lançamentos de taxas fixas para a data de referência"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            lancamentos_gerados = []
            
            # Buscar contratos ativos com taxa fixa
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verifica se é registro de administrador e tipo fixo
                if (row[6] and  # Tem nº contrato na coluna G
                    row[9] == 'Fixo' and  # É tipo fixo
                    self.contrato_ativo(ws, row[6])):  # Contrato está ativo
                    
                    # Verificar se já tem lançamento para este período
                    if not self.tem_lancamento(ws, row[6], row[7], data_ref):
                        # Preparar dados para o lançamento
                        dados_lancamento = {
                            'data_rel': data_ref,
                            'cnpj_cpf': row[7],  # CNPJ/CPF
                            'nome': row[8],      # Nome/Razão Social
                            'referencia': f'ADM FIXA REF. {data_ref.strftime("%m/%Y")}',
                            'valor': float(row[10].replace(',', '.')),  # Valor/Parcela
                            'dt_vencto': self.calcular_vencimento(data_ref)
                        }
                        
                        # Registrar lançamento no sistema
                        self.sistema.dados_para_incluir.append(dados_lancamento)
                        lancamentos_gerados.append(dados_lancamento)
                        
                        # Registrar na aba de controle
                        self.registrar_lancamento(ws, dados_lancamento)
                        
            wb.save(arquivo_cliente)
            return lancamentos_gerados
            
        except Exception as e:
            raise Exception(f"Erro ao processar lançamentos fixos: {str(e)}")

    def contrato_ativo(self, ws, num_contrato):
        """Verifica se o contrato está ativo"""
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == num_contrato:  # Coluna A (Nº Contrato)
                return row[3] == 'ATIVO'  # Coluna D (Status)
        return False

    def tem_lancamento(self, ws, num_contrato, cnpj_cpf, data_ref):
        """Verifica se já existe lançamento para o período"""
        data_str = data_ref.strftime("%d/%m/%Y")
        for row in ws.iter_rows(min_row=3, values_only=True):
            if (row[25] and  # Tem referência na coluna PARCELAS
                row[24] == num_contrato and  # Mesmo contrato
                row[26] == cnpj_cpf and  # Mesmo CNPJ/CPF
                row[28] == data_str):  # Mesma data
                return True
        return False

    def calcular_vencimento(self, data_ref):
        """Calcula data de vencimento (dia 5 do mês seguinte)"""
        if data_ref.day == 5:
            vencto = data_ref.replace(day=20)
        else:  # day == 20
            if data_ref.month == 12:
                vencto = data_ref.replace(year=data_ref.year + 1, month=1, day=5)
            else:
                vencto = data_ref.replace(month=data_ref.month + 1, day=5)
        return vencto

    def registrar_lancamento(self, ws, dados):
        """Registra o lançamento na aba de controle"""
        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=26, value=dados['cnpj_cpf'])
        ws.cell(row=proxima_linha, column=27, value=dados['nome'])
        ws.cell(row=proxima_linha, column=28, value=dados['data_rel'])
        ws.cell(row=proxima_linha, column=29, value=dados['valor'])
        ws.cell(row=proxima_linha, column=30, value='LANÇADO')

class GestaoAdministradores:
    def __init__(self, parent):
        self.parent = parent
        self.busca_entry = None
        self.tree_fornecedores = None
        self.administradores = []  # Lista para armazenar os administradores
        
    def abrir_janela_admin(self):
        """Abre janela para gestão de administradores"""
        self.janela_admin = tk.Toplevel(self.parent)
        self.janela_admin.title("Gestão de Administradores")
        self.janela_admin.geometry("800x600")
        
        # Frame para busca de fornecedor
        frame_busca = ttk.LabelFrame(self.janela_admin, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        self.busca_entry = ttk.Entry(frame_busca, width=40)  # Definir como atributo da classe
        self.busca_entry.pack(side='left', padx=5)

        # Criar a tree de fornecedores antes de usar
        self.tree_fornecedores = ttk.Treeview(frame_busca, 
            columns=('CNPJ/CPF', 'Nome', 'Categoria'),
            show='headings')
        
        # Definir a função de busca
        def buscar():
            termo = self.busca_entry.get()
            # Implementar lógica de busca aqui
            
        # Definir a função de seleção    
        def selecionar(event):
            # Implementar lógica de seleção aqui
            pass

        self.busca_entry.bind('<Return>', lambda e: buscar())
        self.tree_fornecedores.bind('<<TreeviewSelect>>', selecionar)
        ttk.Button(frame_busca, text="Buscar", command=buscar).pack(side='left', padx=5)

        
        # Frame para lista de fornecedores
        frame_fornecedores = ttk.LabelFrame(self.janela_admin, text="Fornecedores")
        frame_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tree_fornecedores = ttk.Treeview(frame_fornecedores, 
                                             columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                             show='headings',
                                             height=5)
        self.tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_fornecedores.heading('Nome', text='Nome')
        self.tree_fornecedores.heading('Categoria', text='Categoria')
        self.tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para percentual
        frame_percentual = ttk.Frame(self.janela_admin)
        frame_percentual.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_percentual, text="Percentual (%):").pack(side='left', padx=5)
        self.percentual_entry = ttk.Entry(frame_percentual, width=10)
        self.percentual_entry.pack(side='left', padx=5)
        
        ttk.Button(frame_percentual, 
                  text="Adicionar Administrador", 
                  command=self.adicionar_administrador).pack(side='left', padx=5)
        
        # Frame para lista de administradores
        frame_lista = ttk.LabelFrame(self.janela_admin, text="Administradores Cadastrados")
        frame_lista.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tree_admin = ttk.Treeview(frame_lista, 
                                     columns=('CNPJ/CPF', 'Nome', 'Percentual'),
                                     show='headings')
        self.tree_admin.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_admin.heading('Nome', text='Nome')
        self.tree_admin.heading('Percentual', text='Percentual (%)')
        self.tree_admin.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.janela_admin)
        frame_botoes.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(frame_botoes, 
                  text="Remover Administrador", 
                  command=self.remover_administrador).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Concluir", 
                  command=self.finalizar_gestao).pack(side='right', padx=5)

    def buscar_fornecedor(self):
        termo = self.busca_entry.get()
        buscar_fornecedor(self.tree_fornecedores, termo)
            
    def adicionar_administrador(self):
        """Adiciona um fornecedor selecionado como administrador"""
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um fornecedor")
            return
            
        fornecedor = self.tree_fornecedores.item(selecionado)['values']
        percentual = self.percentual_entry.get().strip()
        
        # Validar percentual
        if not percentual:
            custom_messagebox("error", "Erro", "Informe o percentual!")
            return
            
        try:
            percentual_float = float(percentual.replace(',', '.'))
            if percentual_float <= 0 or percentual_float > 100:
                custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100!")
                return
        except ValueError:
            custom_messagebox("error", "Erro", "Percentual inválido!")
            return
            
        # Formatar CNPJ/CPF como string
        cnpj_cpf = str(fornecedor[0]).strip()  # Converter para string e remover espaços
        
        # Verificar se o fornecedor já está na lista
        for admin in self.administradores:
            if admin[0] == cnpj_cpf:  # Compara CNPJ/CPF
                custom_messagebox("error", "Erro", "Este fornecedor já está cadastrado como administrador!")
                return
                
        # Verificar se o total de percentuais não excede 100%
        total_atual = sum(float(item[2].replace(',', '.')) 
                         for item in self.administradores)
        if total_atual + percentual_float > 100:
            custom_messagebox("error", "Erro", "Soma dos percentuais excede 100%!")
            return
            
        # Adicionar à lista e à treeview usando o CNPJ/CPF como string
        self.administradores.append((cnpj_cpf, fornecedor[1], percentual))
        self.tree_admin.insert('', 'end', values=(cnpj_cpf, fornecedor[1], percentual))
        
        # Limpar campo de percentual
        self.percentual_entry.delete(0, tk.END)
        
    def remover_administrador(self):
        """Remove o administrador selecionado"""
        selecionado = self.tree_admin.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um administrador para remover")
            return
        
        self.tree_admin.delete(selecionado)
        valores = self.tree_admin.item(selecionado)['values']
        self.administradores = [(cnpj, nome, perc) for cnpj, nome, perc 
                              in self.administradores 
                              if cnpj != valores[0]]
        
    def finalizar_gestao(self):
        """Finaliza a gestão de administradores"""
        total = sum(float(perc.replace(',', '.')) 
                   for _, _, perc in self.administradores)
        if total > 100:
            custom_messagebox("error", "Erro", "Soma dos percentuais excede 100%!")
            return
        
        self.janela_admin.destroy()
        
    def get_administradores(self):
        """Retorna a lista de administradores configurados"""
        return self.administradores.copy()        

class GestorParcelas:
    # from src.combobox_autocompletar import ComboboxAutocompletar
    from src.configuracoes_sistema import GerenciadorConfiguracoes  

    def __init__(self, parent):
        print("Inicializando GestorParcelas")  # Debug
        self.parent = parent
        self.parcelas = []
        self.tipo_despesa_valor = '3'
        self.janela_parcelas = None
        self._var_tem_entrada = None  # Inicializa como None
        # Limpar referências de widgets
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None
        self.parcelas_personalizadas = []
        self.frame_parcelas_personalizadas = None
        self.valor_total_personalizado = 0.0
        self.canvas_parcelas = None
        self.scrollbar_parcelas = None
        self.campos_parcelas = []
        

    @property
    def tem_entrada(self):
        """Getter para tem_entrada - cria apenas quando necessário"""
        if self._var_tem_entrada is None:
            self._var_tem_entrada = tk.BooleanVar(master=self.parent.root, value=False)
        return self._var_tem_entrada


    # Interface e Controles
    def abrir_janela_parcelas(self):
        print("Abrindo janela de parcelas")  # Debug
        # Criar janela como Toplevel do parent
        self.janela_parcelas = tk.Toplevel(self.parent.root)
        self.janela_parcelas.title("Configuração de Parcelas")
        self.janela_parcelas.geometry("700x800")  # Aumentado para acomodar nova opção
        
        # Garantir que a janela seja modal
        self.janela_parcelas.transient(self.parent.root)
        self.janela_parcelas.grab_set()
        
        frame = ttk.Frame(self.janela_parcelas, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame para entrada
        frame_entrada = ttk.LabelFrame(frame, text="Entrada")
        frame_entrada.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        print("Criando Checkbutton")  # Debug
        check = ttk.Checkbutton(
            frame_entrada, 
            text="Possui entrada?",
            variable=self.tem_entrada,
            command=self.atualizar_campos_entrada
        )
        check.grid(row=0, column=0, padx=5, pady=5)

        # Frame para modalidades de entrada
        print("Criando frame modalidade")  # Debug
        self.frame_modalidade = ttk.Frame(frame_entrada)
        self.frame_modalidade.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        ttk.Label(self.frame_modalidade, text="Modalidade de Entrada:").grid(row=0, column=0, padx=5, pady=2)
        self.modalidade_entrada = ttk.Combobox(self.frame_modalidade, state='readonly', width=40)
        self.modalidade_entrada['values'] = [
            "Percentual do valor total na primeira parcela",
            "Primeira parcela igual às demais (arredonda no final)",
            "Valor específico na primeira parcela"
        ]
        self.modalidade_entrada.grid(row=0, column=1, padx=5, pady=2)
        

        # Garantir que o frame modalidade começa oculto
        print("Ocultando frame modalidade inicialmente")  # Debug
        self.frame_modalidade.grid_remove()
        
        # Frame para valor da entrada (dinâmico baseado na modalidade)
        self.frame_valor_entrada = ttk.Frame(frame_entrada)
        self.frame_valor_entrada.grid(row=2, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        # Ocultar frames inicialmente
        self.frame_modalidade.grid_remove()
        self.frame_valor_entrada.grid_remove()
        
        # Tipo de Despesa
        ttk.Label(frame, text="Tipo de Despesa:").grid(row=1, column=0, padx=5, pady=5)
        self.tipo_despesa = ttk.Combobox(frame, values=['2', '3', '5', '6'], state='readonly', width=5)
        self.tipo_despesa.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        self.tipo_despesa.set('3')  # Tipo 3 como padrão

        # Tipo de Parcelamento - MODIFICADO PARA INCLUIR NOVA OPÇÃO
        ttk.Label(frame, text="Tipo de Parcelamento:").grid(row=2, column=0, padx=5, pady=5)
        self.tipo_parcelamento = ttk.Combobox(frame, values=[
            "Prazo Fixo em Dias",
            "Datas Específicas",
            "Cartão de Crédito",
            "Parcelas Personalizadas"  # NOVA OPÇÃO
        ], state="readonly")
        self.tipo_parcelamento.grid(row=2, column=1, padx=5, pady=5)
        self.tipo_parcelamento.set("Prazo Fixo em Dias")
        self.tipo_parcelamento.bind('<<ComboboxSelected>>', self.atualizar_campos_parcelamento)

        # Frame para campos dinâmicos
        self.frame_dinamico = ttk.Frame(frame)
        self.frame_dinamico.grid(row=3, column=0, columnspan=2, pady=10, sticky='ew')

        # Campos comuns
        ttk.Label(frame, text="Data da Despesa:").grid(row=4, column=0, padx=5, pady=5)
        self.data_despesa = DateEntry(
            frame,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2
        )
        
        self.data_despesa.grid(row=4, column=1, padx=5, pady=5)
        self.data_despesa.configure(state='normal')
        self.configurar_calendario(self.data_despesa)

        ttk.Label(frame, text="Valor Original:").grid(row=5, column=0, padx=5, pady=5)
        self.valor_original = ttk.Entry(frame)
        self.valor_original.grid(row=5, column=1, padx=5, pady=5)
        self.valor_original.bind('<KeyPress>', self.on_valor_original_manual_edit)


        # Alterar o label do número de parcelas para ser mais claro
        self.lbl_num_parcelas = ttk.Label(frame, text="Número de Parcelas:")
        self.lbl_num_parcelas.grid(row=6, column=0, padx=5, pady=5)
        self.num_parcelas = ttk.Entry(frame)
        self.num_parcelas.grid(row=6, column=1, padx=5, pady=5)

        # Frame específico para informação sobre parcelas
        frame_info_parcelas = ttk.Frame(frame)
        frame_info_parcelas.grid(row=7, column=0, columnspan=2, padx=5, pady=5, sticky='ew')
        
        self.lbl_info_parcelas = ttk.Label(
            frame_info_parcelas, 
            text="",
            wraplength=500,  # Permitir quebra de linha se necessário
            justify='center'
        )
        self.lbl_info_parcelas.pack(fill='x', padx=5)

        # Referência Base (já existe)
        ttk.Label(frame, text="Referência Base:").grid(row=8, column=0, padx=5, pady=5)
        self.referencia_base = ttk.Entry(frame)
        self.referencia_base.grid(row=8, column=1, padx=5, pady=5, sticky='ew')

        # Adicionar campo NF
        ttk.Label(frame, text="NF:").grid(row=9, column=0, padx=5, pady=5)
        self.campos_nf = ttk.Entry(frame)
        self.campos_nf.grid(row=9, column=1, padx=5, pady=5, sticky='ew')

        # Adicionar campos Etapa da Obra e Insumos
        from src.configuracoes_sistema import GerenciadorConfiguracoes

        ttk.Label(frame, text="Etapa da Obra:").grid(row=10, column=0, padx=5, pady=5)

        etapas_obra = GerenciadorConfiguracoes.get_etapas_obra()

        self.etapa_obra = ComboboxAutocompletar(
            frame,
            values=etapas_obra,
            config_key='etapas_obra',
            config_manager=GerenciadorConfiguracoes,
            width=37,  # Ajustado para combinar com outros campos
            state='normal'
        )
        self.etapa_obra.grid(row=10, column=1, padx=5, pady=5, sticky='ew')

        # Campo Insumos - USANDO COMBOBOX AUTOCOMPLETAR
        ttk.Label(frame, text="Insumos:").grid(row=11, column=0, padx=5, pady=5)

        insumos = GerenciadorConfiguracoes.get_insumos()

        self.insumo = ComboboxAutocompletar(
            frame,
            values=insumos,
            config_key='insumos',
            config_manager=GerenciadorConfiguracoes,
            width=37,
            state='normal'
        )
        self.insumo.grid(row=11, column=1, padx=5, pady=5, sticky='ew')

        # Ajustar row dos botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.grid(row=12, column=0, columnspan=2, pady=30)  # era row=11

        ttk.Button(frame_botoes, 
                  text="Gerar Parcelas", 
                  command=self.gerar_parcelas).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Cancelar", 
                  command=self.cancelar_parcelamento).pack(side='left', padx=5)

        # Inicializar campos do tipo padrão
        self.atualizar_campos_parcelamento(None)

        # Fazer a janela modal
        self.janela_parcelas.transient(self.parent.root)
        self.janela_parcelas.grab_set()

        # Centralizar a janela
        self.janela_parcelas.update_idletasks()
        width = self.janela_parcelas.winfo_width()
        height = self.janela_parcelas.winfo_height()
        x = (self.janela_parcelas.winfo_screenwidth() // 2) - (width // 2)
        y = (self.janela_parcelas.winfo_screenheight() // 2) - (height // 2)
        self.janela_parcelas.geometry(f'{width}x{height}+{x}+{y}')

    def atualizar_campos_entrada(self):
        """Mostra/oculta campos relacionados à entrada e atualiza labels"""
        if self.tem_entrada.get():
            # Mostrar frame modalidade
            if self.frame_modalidade:
                self.frame_modalidade.grid()
                
                # Criar campos se não existirem
                if not hasattr(self, 'valor_entrada') or not self.valor_entrada:
                    if not self.frame_valor_entrada:
                        self.frame_valor_entrada = ttk.Frame(self.frame_modalidade)
                        self.frame_valor_entrada.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
                    
                    self.lbl_entrada = ttk.Label(self.frame_valor_entrada, text="Valor:")
                    self.lbl_entrada.grid(row=0, column=0, padx=5, pady=2)
                    
                    self.valor_entrada = ttk.Entry(self.frame_valor_entrada)
                    self.valor_entrada.grid(row=0, column=1, padx=5, pady=2)
                
                if self.frame_valor_entrada:
                    self.frame_valor_entrada.grid()
        else:
            # Ocultar frames
            if self.frame_modalidade:
                self.frame_modalidade.grid_remove()
            if self.frame_valor_entrada:
                self.frame_valor_entrada.grid_remove()
            
            # Restaurar label original
            for widget in self.janela_parcelas.winfo_children():
                if isinstance(widget, ttk.Label) and widget.cget("text").startswith("Número de Parcelas"):
                    widget.config(text="Número de Parcelas:")
            self.lbl_info_parcelas.config(text="")

    def atualizar_campos_modalidade(self, event=None):
        """Atualiza campos baseado na modalidade selecionada"""
        modalidade = self.modalidade_entrada.get()
        
        if not hasattr(self, 'frame_valor_entrada') or not hasattr(self, 'lbl_entrada'):
            return
            
        self.frame_valor_entrada.grid()
        
        if modalidade == "Percentual do valor total na primeira parcela":
            self.lbl_entrada.config(text="Percentual (%): ")
            self.valor_entrada.delete(0, tk.END)
        elif modalidade == "Primeira parcela igual às demais (arredonda no final)":
            self.frame_valor_entrada.grid_remove()
        elif modalidade == "Valor específico na primeira parcela":
            self.lbl_entrada.config(text="Valor (R$): ")
            self.valor_entrada.delete(0, tk.END)
            
    def atualizar_campos_parcelamento(self, event):
        # Limpar frame dinâmico
        for widget in self.frame_dinamico.winfo_children():
            widget.destroy()

        tipo = self.tipo_parcelamento.get()
        
        # Ocultar/mostrar campos baseado no tipo
        if tipo == "Parcelas Personalizadas":
            # Ocultar campos tradicionais
            self.lbl_num_parcelas.grid_remove()
            self.num_parcelas.grid_remove()
            self.lbl_info_parcelas.config(text="Configure cada parcela individualmente")
            
            # Criar interface para parcelas personalizadas
            self.criar_interface_parcelas_personalizadas()
        else:
            # Mostrar campos tradicionais
            self.lbl_num_parcelas.grid()
            self.num_parcelas.grid()
            
            # Lógica existente para outros tipos
            if tipo == "Prazo Fixo em Dias":
                ttk.Label(self.frame_dinamico, text="Prazo entre Parcelas (dias):").grid(row=0, column=0, padx=5, pady=5)
                self.prazo_dias = ttk.Entry(self.frame_dinamico)
                self.prazo_dias.grid(row=0, column=1, padx=5, pady=5)
                self.prazo_dias.insert(0, "30")  # Valor padrão

            elif tipo == "Datas Específicas":
                num_parcelas_txt = "parcelas após a entrada" if self.tem_entrada.get() else "parcelas"
                
                ttk.Label(self.frame_dinamico, 
                         text=f"Informe as datas de vencimento das {num_parcelas_txt}:").grid(
                             row=0, column=0, columnspan=2, padx=5, pady=5)
                
                self.texto_datas = tk.Text(self.frame_dinamico, height=4, width=30)
                self.texto_datas.grid(row=1, column=0, columnspan=2, padx=5, pady=5)
                
                ttk.Label(self.frame_dinamico, 
                         text="Digite uma data por linha no formato dd/mm/aaaa\n"
                              "(não inclua a data da entrada)").grid(
                             row=2, column=0, columnspan=2, padx=5, pady=5)

            elif tipo == "Cartão de Crédito":
                ttk.Label(self.frame_dinamico, text="Dia do Vencimento:").grid(row=0, column=0, padx=5, pady=5)
                self.dia_vencimento = ttk.Entry(self.frame_dinamico, width=5)
                self.dia_vencimento.grid(row=0, column=1, padx=5, pady=5)
                self.dia_vencimento.insert(0, "10")  # Valor padrão

    def criar_interface_parcelas_personalizadas(self):
        """Cria interface específica para parcelas personalizadas"""
        
        # Frame para controles básicos
        frame_controles = ttk.Frame(self.frame_dinamico)
        frame_controles.pack(fill='x', pady=5)
        
        # Número total de parcelas
        ttk.Label(frame_controles, text="Número de Parcelas:").pack(side='left', padx=5)
        self.num_parcelas_personalizado = tk.IntVar(value=2)
        spin_parcelas = ttk.Spinbox(frame_controles, from_=2, to=12, 
                                   textvariable=self.num_parcelas_personalizado,
                                   width=5, command=self.atualizar_grid_parcelas)
        spin_parcelas.pack(side='left', padx=5)
        
        # Botão para gerar grid
        ttk.Button(frame_controles, text="Gerar Grid", 
                  command=self.atualizar_grid_parcelas).pack(side='left', padx=10)
        
        # Frame scrollável para o grid de parcelas
        self.criar_frame_scrollavel_parcelas()
        
        # Frame para controles da última parcela
        frame_ultima_parcela = ttk.LabelFrame(self.frame_dinamico, 
                                            text="Configuração da Última Parcela")
        frame_ultima_parcela.pack(fill='x', pady=10)
        
        self.condicao_ultima_parcela = tk.BooleanVar()
        ttk.Checkbutton(frame_ultima_parcela, 
                       text="Última parcela depende de entrega/condição específica",
                       variable=self.condicao_ultima_parcela,
                       command=self.toggle_condicao_ultima_parcela).pack(anchor='w', padx=5, pady=5)
        
        # Frame para data condicional (inicialmente oculto)
        self.frame_data_condicional = ttk.Frame(frame_ultima_parcela)
        
        ttk.Label(self.frame_data_condicional, 
                 text="Data estimada:").pack(side='left', padx=5)
        
        self.data_condicional = DateEntry(
            self.frame_data_condicional,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2,
            width=12
        )
        self.data_condicional.pack(side='left', padx=5)
        
        ttk.Label(self.frame_data_condicional, 
                 text="Observação:").pack(side='left', padx=(20, 5))
        
        self.obs_condicional = ttk.Entry(self.frame_data_condicional, width=25)
        self.obs_condicional.pack(side='left', padx=5)

    def criar_frame_scrollavel_parcelas(self):
        """Cria um frame scrollável para edição das parcelas - VERSÃO CORRIGIDA"""
        
        # Frame container para o canvas e scrollbar
        container_frame = ttk.Frame(self.frame_dinamico)
        container_frame.pack(fill='both', expand=True, pady=10)
        
        # CORREÇÃO 1: Definir largura mínima maior para o canvas
        self.canvas_parcelas = tk.Canvas(container_frame, height=200, width=650)  # Largura aumentada
        self.scrollbar_parcelas = ttk.Scrollbar(container_frame, orient="vertical", command=self.canvas_parcelas.yview)
        self.frame_parcelas_personalizadas = ttk.Frame(self.canvas_parcelas)
        
        # CORREÇÃO 2: Configurar o frame interno para expandir adequadamente
        self.frame_parcelas_personalizadas.bind(
            "<Configure>",
            lambda e: self.canvas_parcelas.configure(scrollregion=self.canvas_parcelas.bbox("all"))
        )
        
        # CORREÇÃO 3: Configurar o canvas para ajustar a largura do frame interno
        def configure_canvas_width(event):
            canvas_width = event.width
            self.canvas_parcelas.itemconfig(self.canvas_window_id, width=canvas_width)
        
        self.canvas_parcelas.bind('<Configure>', configure_canvas_width)
        
        # Criar a janela no canvas
        self.canvas_window_id = self.canvas_parcelas.create_window((0, 0), window=self.frame_parcelas_personalizadas, anchor="nw")
        self.canvas_parcelas.configure(yscrollcommand=self.scrollbar_parcelas.set)
        
        self.canvas_parcelas.pack(side="left", fill="both", expand=True)
        self.scrollbar_parcelas.pack(side="right", fill="y")
        
        # Bind mouse wheel para scroll
        def _on_mousewheel(event):
            self.canvas_parcelas.yview_scroll(int(-1*(event.delta/120)), "units")
        self.canvas_parcelas.bind_all("<MouseWheel>", _on_mousewheel)

    def atualizar_grid_parcelas(self):
        """Atualiza o grid de parcelas baseado no número selecionado - VERSÃO CORRIGIDA"""
        
        if not hasattr(self, 'frame_parcelas_personalizadas'):
            return
            
        # Limpar grid existente
        for widget in self.frame_parcelas_personalizadas.winfo_children():
            widget.destroy()
        
        num_parcelas = self.num_parcelas_personalizado.get()
        
        # CORREÇÃO PRINCIPAL: Configurar o grid com pesos adequados
        self.frame_parcelas_personalizadas.columnconfigure(0, weight=0, minsize=60)   # Parcela - fixo
        self.frame_parcelas_personalizadas.columnconfigure(1, weight=1, minsize=120)  # Valor - expansível
        self.frame_parcelas_personalizadas.columnconfigure(2, weight=1, minsize=130)  # Data - expansível
        self.frame_parcelas_personalizadas.columnconfigure(3, weight=2, minsize=200)  # Observação - maior espaço
        
        # Cabeçalho do grid
        ttk.Label(self.frame_parcelas_personalizadas, text="Parcela", 
                font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(self.frame_parcelas_personalizadas, text="Valor (R$)", 
                font=('Arial', 10, 'bold')).grid(row=0, column=1, padx=5, pady=5, sticky='w')
        ttk.Label(self.frame_parcelas_personalizadas, text="Data Vencimento", 
                font=('Arial', 10, 'bold')).grid(row=0, column=2, padx=5, pady=5, sticky='w')
        ttk.Label(self.frame_parcelas_personalizadas, text="Observação", 
                font=('Arial', 10, 'bold')).grid(row=0, column=3, padx=5, pady=5, sticky='w')
        
        # Criar campos para cada parcela
        self.campos_parcelas = []
        
        for i in range(num_parcelas):
            parcela_num = i + 1
            
            # Número da parcela
            ttk.Label(self.frame_parcelas_personalizadas, 
                    text=f"{parcela_num}ª").grid(row=parcela_num, column=0, padx=5, pady=2, sticky='w')
            
            # Campo valor
            valor_var = tk.StringVar()
            entry_valor = ttk.Entry(self.frame_parcelas_personalizadas, 
                                textvariable=valor_var, width=15)
            entry_valor.grid(row=parcela_num, column=1, padx=5, pady=2, sticky='ew')
            entry_valor.bind('<KeyRelease>', self.calcular_total_personalizado)
            
            # Campo data
            data_entry = DateEntry(
                self.frame_parcelas_personalizadas,
                format='dd/mm/yyyy',
                locale='pt_BR',
                background='darkblue',
                foreground='white',
                borderwidth=1,
                width=12
            )
            data_entry.grid(row=parcela_num, column=2, padx=5, pady=2, sticky='ew')
            
            # CORREÇÃO PRINCIPAL: Campo observação com configuração adequada
            obs_var = tk.StringVar()
            entry_obs = ttk.Entry(self.frame_parcelas_personalizadas, 
                                textvariable=obs_var)  # Removido width fixo
            entry_obs.grid(row=parcela_num, column=3, padx=5, pady=2, sticky='ew')  # sticky='ew' para expandir
            
            self.campos_parcelas.append({
                'valor': valor_var,
                'data': data_entry,
                'observacao': obs_var,
                'entry_valor': entry_valor
            })
        
        # Frame para total
        frame_total = ttk.Frame(self.frame_parcelas_personalizadas)
        frame_total.grid(row=num_parcelas + 1, column=0, columnspan=4, pady=10, sticky='ew')
        
        ttk.Label(frame_total, text="TOTAL:", 
                font=('Arial', 10, 'bold')).pack(side='left', padx=5)
        
        self.label_total_personalizado = ttk.Label(frame_total, text="R$ 0,00", 
                                                font=('Arial', 10, 'bold'),
                                                foreground='red')
        self.label_total_personalizado.pack(side='left', padx=5)
        
        # FORÇAR ATUALIZAÇÃO DO LAYOUT
        self.frame_parcelas_personalizadas.update_idletasks()
        
        # CORREÇÃO: Forçar o canvas a reconhecer o novo tamanho
        self.canvas_parcelas.after(100, self._update_canvas_scroll)

    def _update_canvas_scroll(self):
        """Método auxiliar para atualizar o scroll do canvas após mudanças no grid"""
        if hasattr(self, 'canvas_parcelas') and hasattr(self, 'frame_parcelas_personalizadas'):
            # Atualizar a scroll region
            self.canvas_parcelas.configure(scrollregion=self.canvas_parcelas.bbox("all"))
            
            # Forçar atualização da largura se necessário
            canvas_width = self.canvas_parcelas.winfo_width()
            if canvas_width > 1:  # Verificar se o canvas já foi renderizado
                self.canvas_parcelas.itemconfig(self.canvas_window_id, width=canvas_width)

    def calcular_total_personalizado(self, event=None):
        """Calcula o total das parcelas personalizadas e atualiza o campo Valor Original"""
        total = 0.0
        
        if hasattr(self, 'campos_parcelas'):
            for campo in self.campos_parcelas:
                try:
                    valor_str = campo['valor'].get().replace(',', '.')
                    if valor_str:
                        total += float(valor_str)
                except ValueError:
                    pass
        
        self.valor_total_personalizado = total
        
        # Atualizar label do total
        if hasattr(self, 'label_total_personalizado'):
            self.label_total_personalizado.config(text=f"R$ {total:,.2f}")
        
        # NOVA FUNCIONALIDADE: Auto-preencher o campo Valor Original
        if hasattr(self, 'valor_original') and total > 0:
            # Limpar o campo atual
            self.valor_original.delete(0, tk.END)
            # Inserir o novo valor formatado
            self.valor_original.insert(0, f"{total:.2f}".replace('.', ','))
            
            # Opcional: alterar cor do campo para indicar que foi preenchido automaticamente
            self.valor_original.configure(style='AutoFilled.TEntry')
            
            # Criar o estilo se não existir
            try:
                style = ttk.Style()
                style.configure('AutoFilled.TEntry', 
                            fieldbackground='#E8F5E8',  # Verde claro
                            bordercolor='#4CAF50')       # Verde
            except:
                pass  # Se o estilo já existir ou houver erro, ignorar

    # MÉTODO ADICIONAL: resetar estilo do campo quando editado manualmente
    def on_valor_original_manual_edit(self, event=None):
        """Reseta o estilo quando o usuário edita manualmente o Valor Original"""
        if hasattr(self, 'valor_original'):
            try:
                self.valor_original.configure(style='TEntry')  # Estilo padrão
            except:
                pass

    # NOVO MÉTODO: toggle_condicao_ultima_parcela
    def toggle_condicao_ultima_parcela(self):
        """Mostra/oculta configurações da última parcela condicional"""
        if self.condicao_ultima_parcela.get():
            self.frame_data_condicional.pack(fill='x', padx=5, pady=5)
        else:
            self.frame_data_condicional.pack_forget()

    # Métodos de geração e validação de parcelas
    def validar_dados_entrada(self, valor_original, num_parcelas, referencia_base, tipo):
        """Valida os dados básicos antes de gerar parcelas"""
        if not referencia_base or num_parcelas <= 0:
            custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
            return False

        # Validações específicas por tipo de parcelamento
        if tipo == "Prazo Fixo em Dias":
            if not hasattr(self, 'prazo_dias') or not self.prazo_dias.get():
                custom_messagebox("error", "Erro", "Informe o prazo entre as parcelas!")
                return False
        elif tipo == "Datas Específicas":
            if not hasattr(self, 'texto_datas'):
                custom_messagebox("error", "Erro", "Configure as datas específicas!")
                return False
        elif tipo == "Cartão de Crédito":
            if not hasattr(self, 'dia_vencimento') or not self.dia_vencimento.get():
                custom_messagebox("error", "Erro", "Informe o dia do vencimento!")
                return False
            try:
                dia_vencimento = int(self.dia_vencimento.get())
                if not (1 <= dia_vencimento <= 31):
                    custom_messagebox("error", "Erro", "Dia de vencimento deve estar entre 1 e 31!")
                    return False
            except ValueError:
                custom_messagebox("error", "Erro", "Dia de vencimento inválido!")
                return False

        return True

    def gerar_parcelas(self):
        """Método principal para gerar parcelas"""
        try:
            tipo = self.tipo_parcelamento.get()
            
            # Se for parcelas personalizadas, usar lógica específica
            if tipo == "Parcelas Personalizadas":
                return self.gerar_parcelas_personalizadas()
            
            # Lógica existente para outros tipos
            self.tipo_despesa_valor = self.tipo_despesa.get()
            valor_original = float(self.valor_original.get().replace(',', '.'))
            num_parcelas = int(self.num_parcelas.get())
            referencia_base = self.referencia_base.get().strip()
            nf = self.campos_nf.get().strip()

            # Validar dados
            if not self.validar_dados_entrada(valor_original, num_parcelas, referencia_base, tipo):
                return False

            # Resto da lógica existente...
            # Data base é a data da despesa
            data_base = datetime.strptime(self.data_despesa.get(), '%d/%m/%Y')
            
            # Limpar lista de parcelas anterior
            self.parcelas = []

            # Calcular valores das parcelas
            valores_parcelas = self.calcular_valores_parcelas(valor_original, num_parcelas)
            if not valores_parcelas:
                return False

            # Gerar parcelas conforme o tipo
            if tipo == "Prazo Fixo em Dias":
                self.gerar_parcelas_prazo_fixo(data_base, valores_parcelas, referencia_base, num_parcelas, nf)
            elif tipo == "Datas Específicas":
                self.gerar_parcelas_datas_especificas(data_base, valores_parcelas, referencia_base, num_parcelas, nf)
            elif tipo == "Cartão de Crédito":
                self.gerar_parcelas_cartao(data_base, valores_parcelas, referencia_base, num_parcelas, nf)

            if self.parcelas:
                custom_messagebox("info", "Sucesso", f"{len(self.parcelas)} parcela(s) gerada(s) com sucesso!")
                self.limpar_campos()
                return True

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao gerar parcelas: {str(e)}")
            return False

    # NOVO MÉTODO: gerar_parcelas_personalizadas
    def gerar_parcelas_personalizadas(self):
        """Gera as parcelas personalizadas baseadas nos dados inseridos"""
        
        if not hasattr(self, 'campos_parcelas') or not self.campos_parcelas:
            custom_messagebox("error", "Erro", "Configure as parcelas primeiro!")
            return False
        
        try:
            # Obter dados básicos
            self.tipo_despesa_valor = self.tipo_despesa.get()
            referencia_base = self.referencia_base.get().strip()
            nf_base = self.campos_nf.get().strip()
            data_base = datetime.strptime(self.data_despesa.get(), '%d/%m/%Y')
            
            if not referencia_base:
                custom_messagebox("error", "Erro", "Referência base é obrigatória!")
                return False
            
            self.parcelas = []
            
            for i, campo in enumerate(self.campos_parcelas):
                parcela_num = i + 1
                
                try:
                    valor_str = campo['valor'].get().replace(',', '.')
                    if not valor_str:
                        custom_messagebox("error", "Erro", f"Valor da {parcela_num}ª parcela não informado!")
                        return False
                    
                    valor = float(valor_str)
                    dt_vencto_obj = campo['data'].get_date()
                    data_vencto = dt_vencto_obj.strftime('%d/%m/%Y')
                    observacao = campo['observacao'].get().strip()
                    
                    # Montar referência da parcela
                    if observacao:
                        referencia = f"{referencia_base} - {parcela_num}ª PARCELA - {observacao}"
                    else:
                        referencia = f"{referencia_base} - {parcela_num}ª PARCELA"
                    
                    # NF da parcela
                    if nf_base:
                        nf_parcela = f"{nf_base}-{parcela_num:02d}"
                    else:
                        nf_parcela = f"PARC-{parcela_num:02d}"
                    
                    # Verificar se é a última parcela com condição especial
                    if (parcela_num == len(self.campos_parcelas) and 
                        hasattr(self, 'condicao_ultima_parcela') and
                        self.condicao_ultima_parcela.get()):
                        
                        if hasattr(self, 'obs_condicional'):
                            obs_condicional = self.obs_condicional.get().strip()
                            if obs_condicional:
                                referencia += f" - CONDICIONADA: {obs_condicional}"
                            else:
                                referencia += " - CONDICIONADA À ENTREGA"
                        
                        # Usar data condicional se especificada
                        if hasattr(self, 'data_condicional'):
                            data_condicional_obj = self.data_condicional.get_date()
                            if data_condicional_obj != dt_vencto_obj:
                                dt_vencto_obj = data_condicional_obj
                                data_vencto = data_condicional_obj.strftime('%d/%m/%Y')
                    
                    # CALCULAR DATA DO RELATÓRIO ESPECÍFICA PARA CADA PARCELA
                    # Debug: vamos imprimir os valores para identificar o problema
                    print(f"DEBUG - Parcela {parcela_num}:")
                    print(f"  Data vencimento: {dt_vencto_obj}")
                    print(f"  Tipo despesa: {self.tipo_despesa_valor}")
                    
                    # Para parcelas personalizadas, nenhuma é considerada "primeira parcela" com entrada
                    eh_primeira_parcela = False
                    data_rel_obj = self.calcular_data_rel_personalizada(dt_vencto_obj)
                    data_rel = data_rel_obj.strftime('%d/%m/%Y')
                    
                    print(f"  Data relatório calculada: {data_rel}")
                    print("---")
                    
                    parcela = {
                        'data_rel': data_rel,
                        'nf': nf_parcela,
                        'referencia': referencia,
                        'valor': valor,
                        'dt_vencto': data_vencto,
                        'etapa_obra': self.etapa_obra.get().strip(),
                        'insumo': self.insumo.get().strip(),
                        'observacao': observacao
                    }
                    
                    self.parcelas.append(parcela)
                    
                except ValueError:
                    custom_messagebox("error", "Erro", 
                                    f"Valor inválido na {parcela_num}ª parcela!")
                    return False
            
            if self.parcelas:
                # Mostrar resumo antes de confirmar
                self.mostrar_resumo_parcelas_personalizadas()
                return True
            else:
                custom_messagebox("error", "Erro", "Nenhuma parcela foi gerada!")
                return False
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao gerar parcelas personalizadas: {str(e)}")
            return False

    # NOVO MÉTODO: mostrar_resumo_parcelas_personalizadas
    def mostrar_resumo_parcelas_personalizadas(self):
        """Mostra resumo das parcelas personalizadas antes da confirmação"""
        
        resumo_window = tk.Toplevel(self.janela_parcelas)
        resumo_window.title("Resumo das Parcelas Personalizadas")
        resumo_window.geometry("900x600")
        resumo_window.transient(self.janela_parcelas)
        resumo_window.grab_set()
        
        # Frame principal
        main_frame = ttk.Frame(resumo_window)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Título
        ttk.Label(main_frame, text="Resumo das Parcelas Personalizadas", 
                 font=('Arial', 14, 'bold')).pack(pady=10)
        
        # Frame para Treeview com scrollbar
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True, pady=10)
        
        # Treeview para mostrar as parcelas
        columns = ('Parcela', 'Valor', 'Vencimento', 'Referência', 'Etapa', 'Insumo')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        # Configurar colunas
        tree.column('Parcela', width=80, anchor='center')
        tree.column('Valor', width=120, anchor='e')
        tree.column('Vencimento', width=100, anchor='center')
        tree.column('Referência', width=500, anchor='w')
        tree.column('Etapa', width=150, anchor='w')
        tree.column('Insumo', width=150, anchor='w')
        
        for col in columns:
            tree.heading(col, text=col)
        
        # Scrollbar para o Treeview
        scrollbar_tree = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar_tree.set)
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar_tree.pack(side="right", fill="y")
        
        # Inserir dados
        total_resumo = 0
        for i, parcela in enumerate(self.parcelas, 1):
            tree.insert('', 'end', values=(
                f"{i}ª",
                f"R$ {parcela['valor']:,.2f}",
                parcela['dt_vencto'],
                parcela['referencia'],
                parcela.get('etapa_obra', ''),
                parcela.get('insumo', '')
            ))
            total_resumo += parcela['valor']
        
        # Frame para informações finais
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill='x', pady=10)
        
        # Total
        ttk.Label(info_frame, text=f"TOTAL: R$ {total_resumo:,.2f}", 
                 font=('Arial', 12, 'bold')).pack()
        
        # # Forma de pagamento
        # ttk.Label(info_frame, text=f"Forma de Pagamento: {self.forma_pagamento_var.get()}", 
        #          font=('Arial', 10)).pack(pady=5)
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Voltar para Edição", 
                  command=resumo_window.destroy).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes, text="Confirmar e Processar Parcelas", 
                  command=lambda: self.finalizar_confirmacao_personalizada(resumo_window)).pack(side='right', padx=5)

    def calcular_data_rel_personalizada(self, dt_vencto):
        """
        Calcula a data do relatório para parcelas personalizadas.
        Sempre retorna dia 5 ou 20, anterior à data de vencimento.
        """
        try:
            print(f"  calcular_data_rel_personalizada chamado com: {dt_vencto}")
            
            # CORREÇÃO: Converter hoje para date para permitir comparação
            hoje = datetime.now().date()
            tp_desp = self.tipo_despesa_valor
            
            print(f"  Hoje: {hoje}")
            print(f"  Tipo despesa: {tp_desp}")
            print(f"  Dia do vencimento: {dt_vencto.day}")
            
            # Lógica principal baseada na data de vencimento
            if dt_vencto.day == 5:
                # Se vence dia 5, relatório é dia 20 do mês anterior
                data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
                print(f"  Vence dia 5 -> Relatório dia 20 mês anterior: {data_rel}")
            elif dt_vencto.day == 20:
                # Se vence dia 20, relatório é dia 5 do mesmo mês
                data_rel = dt_vencto.replace(day=5)
                print(f"  Vence dia 20 -> Relatório dia 5 mesmo mês: {data_rel}")
            elif tp_desp == '5':
                # Para tipo 5, usar período mais próximo
                if dt_vencto.day <= 5:
                    data_rel = dt_vencto.replace(day=5)
                    print(f"  Tipo 5, vence <= 5 -> Relatório dia 5: {data_rel}")
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=20)
                    print(f"  Tipo 5, vence <= 20 -> Relatório dia 20: {data_rel}")
                else:
                    proximo_mes = dt_vencto + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                    print(f"  Tipo 5, vence > 20 -> Relatório dia 5 próximo mês: {data_rel}")
            else:
                # Para outros tipos (2, 3, 6), usar período anterior ao vencimento
                if dt_vencto.day <= 5:
                    data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
                    print(f"  Outros tipos, vence <= 5 -> Relatório dia 20 mês anterior: {data_rel}")
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=5)
                    print(f"  Outros tipos, vence <= 20 -> Relatório dia 5 mesmo mês: {data_rel}")
                else:
                    data_rel = dt_vencto.replace(day=20)
                    print(f"  Outros tipos, vence > 20 -> Relatório dia 20 mesmo mês: {data_rel}")
            
            print(f"  Data relatório antes da verificação: {data_rel}")
            
            # Garantir que a data do relatório não seja anterior à data atual
            if data_rel < hoje:
                print(f"  Data relatório {data_rel} é anterior a hoje {hoje}, ajustando...")
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                    print(f"  Hoje <= 5 -> Ajustado para dia 5: {data_rel}")
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                    print(f"  Hoje <= 20 -> Ajustado para dia 20: {data_rel}")
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                    print(f"  Hoje > 20 -> Ajustado para dia 5 próximo mês: {data_rel}")
            
            print(f"  Data relatório final: {data_rel}")
            
            # CORREÇÃO: Retornar como datetime para manter consistência
            return datetime.combine(data_rel, datetime.min.time())
            
        except Exception as e:
            print(f"ERRO ao calcular data do relatório personalizada: {str(e)}")
            import traceback
            traceback.print_exc()
            # Em caso de erro, retornar uma data válida baseada em hoje
            hoje = datetime.now().date()
            if hoje.day <= 5:
                data_fallback = hoje.replace(day=5)
            elif hoje.day <= 20:
                data_fallback = hoje.replace(day=20)
            else:
                proximo_mes = hoje + relativedelta(months=1)
                data_fallback = proximo_mes.replace(day=5)
            
            return datetime.combine(data_fallback, datetime.min.time())

            
    # NOVO MÉTODO: finalizar_confirmacao_personalizada
    def finalizar_confirmacao_personalizada(self, resumo_window):
        """Finaliza a confirmação das parcelas personalizadas"""
        resumo_window.destroy()
        
        # Mostrar mensagem de sucesso
        custom_messagebox("info", "Sucesso", 
                        f"{len(self.parcelas)} parcela(s) personalizada(s) gerada(s) com sucesso!")
        
        # Limpar campos e fechar janela
        self.limpar_campos()

    def adicionar_parcela(self, data_rel, dt_vencto, valor_parcela, referencia_base, i, num_parcelas, eh_primeira_parcela, nf):
        """Método auxiliar para criar uma parcela com todos os dados necessários"""
        parcela = {
            'data_rel': data_rel.strftime('%d/%m/%Y'),
            'dt_vencto': dt_vencto.strftime('%d/%m/%Y'),
            'valor': valor_parcela,
            'referencia': self.gerar_referencia_parcela(referencia_base, i, num_parcelas, eh_primeira_parcela),
            'nf': nf,
            'etapa_obra': self.etapa_obra.get().strip(),
            'insumo': self.insumo.get().strip()
        }
        self.parcelas.append(parcela)
        

    def gerar_parcelas_prazo_fixo(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas com prazo fixo em dias"""
        prazo_dias = int(self.prazo_dias.get())
        
        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            if eh_primeira_parcela and self.tem_entrada.get():
                dt_vencto = data_base
                data_rel = self.calcular_data_rel(data_base, dt_vencto, True)
            else:
                dt_vencto = data_base + relativedelta(days=prazo_dias * (i + (0 if self.tem_entrada.get() else 1)))
                dt_vencto = self.proximo_dia_util(dt_vencto)
                data_rel = self.calcular_data_rel(data_base, dt_vencto, eh_primeira_parcela)
            
            self.adicionar_parcela(
                data_rel,
                dt_vencto,
                valor_parcela,
                referencia_base,
                i,
                num_parcelas,
                eh_primeira_parcela,
                nf
            )

    def gerar_parcelas_datas_especificas(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas com datas específicas"""
        datas_texto = self.texto_datas.get("1.0", tk.END).strip().split('\n')
        datas_texto = [d.strip() for d in datas_texto if d.strip()]
        
        num_datas_esperado = num_parcelas
        if len(datas_texto) != num_datas_esperado:
            custom_messagebox("error", 
                "Erro", 
                f"Para {num_parcelas} {'parcelas após a entrada' if self.tem_entrada.get() else 'parcelas'}, "
                f"é necessário informar {num_datas_esperado} data(s) de vencimento."
            )
            return

        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            try:
                if eh_primeira_parcela and self.tem_entrada.get():
                    dt_vencto = data_base
                    data_rel = self.calcular_data_rel(data_base, dt_vencto, True)
                else:
                    idx_data = i - 1 if self.tem_entrada.get() else i
                    if 0 <= idx_data < len(datas_texto):
                        dt_vencto = datetime.strptime(datas_texto[idx_data], '%d/%m/%Y')
                        dt_vencto = self.proximo_dia_util(dt_vencto)
                        data_rel = self.calcular_data_rel(data_base, dt_vencto, eh_primeira_parcela)
                    else:
                        raise ValueError(f"Índice de data inválido: {idx_data}")
                
                self.adicionar_parcela(
                    data_rel,
                    dt_vencto,
                    valor_parcela,
                    referencia_base,
                    i,
                    num_parcelas,
                    eh_primeira_parcela,
                    nf
                )
                
            except ValueError as e:
                custom_messagebox("error", "Erro", f"Erro ao processar data: {str(e)}")
                return
            except IndexError:
                custom_messagebox("error", "Erro", "Número insuficiente de datas fornecidas")
                return

    def gerar_parcelas_cartao(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas para pagamento com cartão"""
        dia_vencimento = int(self.dia_vencimento.get())
        
        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            if eh_primeira_parcela:
                data_atual = data_base + relativedelta(months=1)
            else:
                data_atual = data_base + relativedelta(months=i + 1)
            
            try:
                dt_vencto = data_atual.replace(day=dia_vencimento)
            except ValueError:
                dt_vencto = data_atual + relativedelta(day=31)
            
            dt_vencto = self.proximo_dia_util(dt_vencto)
            
            if eh_primeira_parcela:
                hoje = datetime.now()
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
            else:
                data_rel = self.calcular_data_rel(data_base, dt_vencto, False)
                
            self.adicionar_parcela(
                data_rel,
                dt_vencto,
                valor_parcela,
                referencia_base,
                i,
                num_parcelas,
                eh_primeira_parcela,
                nf
            )


    # Métodos de cálculo e utilitários
    def calcular_valores_parcelas(self, valor_original, num_parcelas):
        """Calcula os valores das parcelas considerando entrada se houver"""
        try:
            if self.tem_entrada.get():
                if not self.modalidade_entrada.get():
                    custom_messagebox("error", "Erro", "Selecione a modalidade de entrada!")
                    return None
                valores_parcelas = self.calcular_parcelas_entrada(valor_original, num_parcelas)
            else:
                valores_parcelas = self.calcular_parcelas_ajustadas(valor_original, num_parcelas)

            # Verificar se a soma está correta
            soma_parcelas = sum(valores_parcelas)
            if abs(soma_parcelas - valor_original) > 0.01:
                custom_messagebox("error", 
                    "Erro", 
                    f"Erro no cálculo das parcelas: soma ({soma_parcelas:.2f}) " 
                    f"diferente do valor original ({valor_original:.2f})!"
                )
                return None

            return valores_parcelas
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao calcular valores: {str(e)}")
            return None

    def calcular_parcelas_entrada(self, valor_total, num_parcelas):
        """Calcula valores das parcelas considerando a modalidade de entrada"""
        modalidade = self.modalidade_entrada.get()
        valores_parcelas = []
        
        # Se tem entrada, o número de parcelas informado é adicional à entrada
        num_parcelas_real = num_parcelas + 1 if self.tem_entrada.get() else num_parcelas
        
        if modalidade == "Percentual do valor total na primeira parcela":
            try:
                percentual = float(self.valor_entrada.get().replace(',', '.'))
                if not (0 < percentual < 100):
                    raise ValueError("Percentual deve estar entre 0 e 100")
                
                valor_entrada = (percentual / 100) * valor_total
                valor_restante = valor_total - valor_entrada
                
                valores_parcelas = [valor_entrada]  # Primeira parcela (entrada)
                # Distribuir o valor restante no número de parcelas informado
                demais_parcelas = self.calcular_parcelas_ajustadas(valor_restante, num_parcelas)
                valores_parcelas.extend(demais_parcelas)
                
            except ValueError as e:
                raise ValueError(f"Erro no percentual de entrada: {str(e)}")
        
        elif modalidade == "Primeira parcela igual às demais (arredonda no final)":
            # Dividir o valor total pelo número total de parcelas (incluindo entrada)
            valores_parcelas = self.calcular_parcelas_ajustadas(valor_total, num_parcelas_real)
            
        elif modalidade == "Valor específico na primeira parcela":
            try:
                valor_entrada = float(self.valor_entrada.get().replace(',', '.'))
                if valor_entrada >= valor_total:
                    raise ValueError("Valor da entrada não pode ser maior ou igual ao valor total")
                
                valor_restante = valor_total - valor_entrada
                valores_parcelas = [valor_entrada]  # Primeira parcela (entrada)
                # Distribuir o valor restante no número de parcelas informado
                demais_parcelas = self.calcular_parcelas_ajustadas(valor_restante, num_parcelas)
                valores_parcelas.extend(demais_parcelas)
                
            except ValueError as e:
                raise ValueError(f"Erro no valor da entrada: {str(e)}")
        
        return valores_parcelas

    def calcular_parcelas_ajustadas(self, valor_total, num_parcelas):
        """Calcula valores das parcelas garantindo que a soma seja igual ao valor total"""
        valor_parcela_base = valor_total / num_parcelas
        valor_parcela_round = round(valor_parcela_base, 2)
        
        # Calcular diferença total devido aos arredondamentos
        diferenca = valor_total - (valor_parcela_round * num_parcelas)
        
        # Distribuir a diferença na última parcela
        parcelas = [valor_parcela_round] * (num_parcelas - 1)
        ultima_parcela = valor_parcela_round + round(diferenca, 2)
        parcelas.append(ultima_parcela)
        
        return parcelas

    def calcular_data_rel(self, data_base, dt_vencto, eh_primeira_parcela):
        """
        Calcula a data do relatório com base na data de vencimento e tipo de despesa.
        Agora considera a data atual para não retroagir em períodos fechados.
        """
        try:
            hoje = datetime.now()
            
            # Se for entrada, calcula a partir da data atual
            if eh_primeira_parcela and self.tem_entrada.get():
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                return data_rel
                
            # Para as demais parcelas, manter a lógica existente
            tp_desp = self.tipo_despesa_valor
            
            if dt_vencto.day == 5:
                # Se vence dia 5, relatório é dia 20 do mês anterior
                data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
            elif dt_vencto.day == 20:
                # Se vence dia 20, relatório é dia 5 do mesmo mês
                data_rel = dt_vencto.replace(day=5)
            elif tp_desp == '5':
                if dt_vencto.day <= 5:
                    data_rel = dt_vencto.replace(day=5)
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=20)
                else:
                    proximo_mes = dt_vencto + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
            else:
                if dt_vencto.day <= 5:
                    data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=5)
                else:
                    data_rel = dt_vencto.replace(day=20)
                    
            # Garantir que a data do relatório não seja anterior à data atual
            if data_rel < hoje:
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                    
            return data_rel
        except Exception as e:
            print(f"Erro ao calcular data do relatório: {str(e)}")
            return dt_vencto

    def configurar_calendario(self, dateentry):
        """Configura o comportamento do calendário"""
        def on_calendar_click(event):
            # Permite cliques no calendário
            return True
            
        def on_calendar_select(event):
            dateentry._top_cal.withdraw()  # Fecha o calendário
            self.janela_parcelas.after(100, lambda: self.janela_parcelas.focus_set())  # Retorna foco
        
        def on_calendar_focus(event):
            # Mantém o foco quando o calendário está aberto
            if dateentry._top_cal:
                dateentry._top_cal.focus_set()
            return True

        # Configurar bindings
        dateentry.bind('<<DateEntrySelected>>', on_calendar_select)
        dateentry.bind('<FocusIn>', on_calendar_focus)
        
        if hasattr(dateentry, '_top_cal'):
            cal = dateentry._top_cal
            cal.bind('<Button-1>', on_calendar_click)
            for w in cal.winfo_children():
                w.bind('<Button-1>', on_calendar_click)

        
    def proximo_dia_util(self, data):
        """
        Ajusta a data para o próximo dia útil se cair em fim de semana ou feriado
        """
        # Lista de feriados nacionais fixos
        feriados_fixos = [
            (1, 1),   # Ano Novo
            (21, 4),  # Tiradentes
            (1, 5),   # Dia do Trabalho
            (7, 9),   # Independência
            (12, 10), # Nossa Senhora
            (2, 11),  # Finados
            (15, 11), # Proclamação da República
            (25, 12), # Natal
        ]

        while True:
            # Verifica se é fim de semana
            if data.weekday() >= 5:  # 5 = Sábado, 6 = Domingo
                data = data + relativedelta(days=1)
                continue

            # Verifica se é feriado fixo
            if (data.day, data.month) in feriados_fixos:
                data = data + relativedelta(days=1)
                continue

            # Se não é fim de semana nem feriado, é dia útil
            break

        return data

    def gerar_referencia_parcela(self, referencia_base, indice, num_parcelas, eh_primeira_parcela):
        """Gera a referência apropriada para a parcela"""
        if eh_primeira_parcela and self.tem_entrada.get():
            return f"{referencia_base} - ENTRADA"
        else:
            if self.tem_entrada.get():
                # Para as parcelas após a entrada
                return f"{referencia_base} - PARC. {indice}/{num_parcelas}"
            else:
                # Para parcelamento sem entrada
                return f"{referencia_base} - PARC. {indice + 1}/{num_parcelas}"
           
    # Métodos de limpeza e finalização
    def limpar_campos(self):
        """Limpa todos os campos após sucesso"""
        # Limpar referências de widgets existentes
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None
        
        # LIMPAR NOVOS CAMPOS DE PARCELAS PERSONALIZADAS
        self.parcelas_personalizadas = []
        self.frame_parcelas_personalizadas = None
        self.valor_total_personalizado = 0.0
        self.canvas_parcelas = None
        self.scrollbar_parcelas = None
        self.campos_parcelas = []
        
        # Resetar checkbox
        if self._var_tem_entrada:
            self._var_tem_entrada.set(False)
        
        # Fechar janela
        if self.janela_parcelas:
            self.janela_parcelas.destroy()
            self.janela_parcelas = None

    def cancelar_parcelamento(self):
        """Cancela o parcelamento e limpa todos os campos"""
        self.parcelas = []
        
        # Limpar referências de widgets existentes
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None
        
        # LIMPAR NOVOS CAMPOS DE PARCELAS PERSONALIZADAS
        self.parcelas_personalizadas = []
        self.frame_parcelas_personalizadas = None
        self.valor_total_personalizado = 0.0
        self.canvas_parcelas = None
        self.scrollbar_parcelas = None
        self.campos_parcelas = []
        
        # Resetar variável de entrada
        if self._var_tem_entrada:
            self._var_tem_entrada.set(False)
        
        if self.janela_parcelas:
            self.janela_parcelas.destroy()
            self.janela_parcelas = None

    # Fechando os métodos/classes anteriores
    def run(self):
        """Inicia a execução do sistema"""
        self.root.mainloop()

class ImportadorRH:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.pasta_rh = Path(BASE_PATH) / "Planilhas_RH"
        
        # Criar pasta se não existir
        os.makedirs(self.pasta_rh, exist_ok=True)
        
        # Mapeamento de cabeçalhos (original -> sistema)
        self.mapeamento_cabecalhos = {
            'Empresa': 'Cliente',
            'Data Pagto':'Data_Pagamento',
            'Nome Empregado': 'Nome_Empregado',
            'Valor Líquido': 'Valor_Líquido',
            'Tipo Folha': 'Tipo_Folha',
            'Forma Pagto': 'Forma_Pagamento',
            'Nome Banco': 'Nome_Banco',
            'Agência': 'Agencia',
            'N° Conta': 'Numero_Conta',
            'Dig. Conta': 'Digito_Conta',
            'Conta': 'Numero_Conta',
            'Dígito': 'Digito_Conta'
        }
        
        # Mapeamento para referências
        self.mapeamento_referencias = {
            '13º SALÁRIO': '13º SALÁRIO',
            'ADIANTAMENTO 13º SALÁRIO': '13º SALÁRIO',
            'ADIANTAMENTO': 'SALÁRIO',
            'MENSAL': 'SALÁRIO',
            'FÉRIAS': 'FÉRIAS',
            'RESCISÃO NORMAL': 'RESCISÃO',
            'RESCISÃO': 'RESCISÃO',
            '13 SALÁRIO': '13º SALÁRIO',
            '13': '13º SALÁRIO',
            'FERIAS': 'FÉRIAS'
        }

    def solicitar_etapa_obra(self):
        """
        Abre um diálogo modal para o usuário selecionar a Etapa da Obra
        que será aplicada a todos os lançamentos importados.
        
        Returns:
            str: A etapa selecionada ou None se o usuário cancelar
        """
        import tkinter as tk
        from tkinter import ttk
        
        # Criar janela de diálogo modal
        dialogo = tk.Toplevel()
        dialogo.title("Etapa da Obra - Importação de dados de RH")
        dialogo.geometry("550x220")
        
        # Tornar a janela modal
        if hasattr(self.sistema, 'root'):
            dialogo.transient(self.sistema.root)
        dialogo.grab_set()
        
        # Centralizar janela na tela
        dialogo.update_idletasks()
        largura = dialogo.winfo_width()
        altura = dialogo.winfo_height()
        x = (dialogo.winfo_screenwidth() // 2) - (largura // 2)
        y = (dialogo.winfo_screenheight() // 2) - (altura // 2)
        dialogo.geometry(f"{largura}x{altura}+{x}+{y}")
        
        # Variável para armazenar a etapa selecionada
        etapa_selecionada = {'valor': None}
        
        # Frame principal com padding
        frame = ttk.Frame(dialogo, padding="25")
        frame.pack(fill='both', expand=True)
        
        # Título explicativo
        ttk.Label(
            frame, 
            text="📋 Selecione a Etapa da Obra",
            font=('Arial', 12, 'bold')
        ).pack(pady=(0, 5))
        
        ttk.Label(
            frame, 
            text="Esta etapa será aplicada a TODOS os lançamentos importados:",
            font=('Arial', 9)
        ).pack(pady=(0, 20))
        
        # Frame para o combobox
        frame_combo = ttk.Frame(frame)
        frame_combo.pack(fill='x', pady=(0, 25))
        
        ttk.Label(
            frame_combo, 
            text="Etapa:",
            font=('Arial', 10)
        ).pack(side='left', padx=(0, 10))
        
        # Carregar etapas disponíveis
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        etapas_obra = GerenciadorConfiguracoes.get_etapas_obra()
        
        # Criar Combobox
        combo_etapa = ttk.Combobox(
            frame_combo,
            values=etapas_obra,
            font=('Arial', 10),
            width=45,
            state='normal'
        )
        combo_etapa.pack(side='left', fill='x', expand=True)
        combo_etapa.focus_set()
        
        # ===== FUNCIONALIDADE DE AUTOCOMPLETAR =====
        def autocompletar(event):
            """Filtra e completa as opções baseado no que o usuário digitou"""
            # Ignorar teclas especiais
            if event.keysym in ['BackSpace', 'Delete', 'Left', 'Right', 'Up', 'Down', 
                            'Home', 'End', 'Return', 'Tab', 'Escape']:
                return
            
            digitado = combo_etapa.get().upper()
            
            if digitado == '':
                combo_etapa['values'] = etapas_obra
                return
            
            # Filtrar etapas que começam com o texto digitado
            etapas_filtradas = [
                etapa for etapa in etapas_obra 
                if etapa.upper().startswith(digitado)
            ]
            
            if etapas_filtradas:
                # Atualizar valores
                combo_etapa['values'] = etapas_filtradas
                
                # Completar com a primeira opção
                primeira_opcao = etapas_filtradas[0]
                combo_etapa.set(primeira_opcao)
                
                # Selecionar apenas a parte que foi autocompletada
                combo_etapa.icursor(len(digitado))
                combo_etapa.selection_range(len(digitado), tk.END)
            else:
                # Se não houver match no início, buscar em qualquer parte
                etapas_filtradas = [
                    etapa for etapa in etapas_obra 
                    if digitado in etapa.upper()
                ]
                combo_etapa['values'] = etapas_filtradas if etapas_filtradas else etapas_obra
        
        def restaurar_lista(event):
            """Restaura a lista completa quando clicar no dropdown"""
            combo_etapa['values'] = etapas_obra
        
        # Bind dos eventos
        combo_etapa.bind('<KeyRelease>', autocompletar)
        combo_etapa.bind('<Button-1>', restaurar_lista)
        # ===== FIM DO AUTOCOMPLETAR =====
        
        # Frame para os botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack()
        
        def confirmar():
            """Valida e confirma a seleção"""
            etapa = combo_etapa.get().strip()
            
            if not etapa:
                custom_messagebox("warning", "Atenção", 
                    "Por favor, selecione ou digite uma etapa da obra.")
                combo_etapa.focus_set()
                return
            
            etapa_selecionada['valor'] = etapa
            dialogo.destroy()
        
        def cancelar():
            """Cancela a operação"""
            etapa_selecionada['valor'] = None
            dialogo.destroy()
        
        # Botões de ação
        ttk.Button(
            frame_botoes, 
            text="✓ Confirmar", 
            command=confirmar,
            width=15
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes, 
            text="✗ Cancelar", 
            command=cancelar,
            width=15
        ).pack(side='left', padx=5)
        
        # Atalhos de teclado
        combo_etapa.bind('<Return>', lambda e: confirmar())
        dialogo.bind('<Escape>', lambda e: cancelar())
        
        # Impedir fechamento pela barra de título sem passar por cancelar
        dialogo.protocol("WM_DELETE_WINDOW", cancelar)
        
        # Aguardar o fechamento da janela
        dialogo.wait_window()
        
        return etapa_selecionada['valor']
            
    def selecionar_arquivo(self):
        """Permite ao usuário selecionar um arquivo Excel de RH para importar"""
        arquivo = filedialog.askopenfilename(
            title="Selecione a planilha de RH",
            filetypes=[
                ("Todos os formatos suportados", "*.xlsx *.xls *.csv *.txt *.xlsm *.xlsb"),
                ("Arquivos Excel", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("Arquivos CSV", "*.csv *.txt")
            ],
            initialdir=str(self.pasta_rh)
        )
        
        if not arquivo:
            return None
        
        # Verificar se é formato antigo do Excel (.xls)
        extensao = os.path.splitext(arquivo)[1].lower()
        if extensao == '.xls':
            # Tentar converter automaticamente
            arquivo_convertido = self.converter_xls_para_xlsx(arquivo)
            if arquivo_convertido:
                arquivo = arquivo_convertido
            else:
                # Se a conversão falhou e o usuário cancelou, retornar None
                return None
        
        # Copiar arquivo para a pasta_rh se estiver em outro local
        nome_arquivo = os.path.basename(arquivo)
        destino = self.pasta_rh / nome_arquivo
        
        if Path(arquivo) != destino:
            try:
                import shutil
                shutil.copy2(arquivo, destino)
                print(f"Arquivo copiado para {destino}")
            except Exception as e:
                print(f"Erro ao copiar arquivo: {str(e)}")
                # Continuar usando o arquivo original
                return arquivo
            
        return destino

    def converter_xls_para_xlsx(self, arquivo_origem):
        """
        Converte um arquivo XLS (Excel 97-2003) para XLSX (Excel moderno)
        Returns:
            str: Caminho do arquivo XLSX convertido ou None se falhar
        """
        print(f"Tentando converter arquivo {arquivo_origem} para formato XLSX")
        
        # Verificar se é um arquivo XLS
        extensao = os.path.splitext(arquivo_origem)[1].lower()
        if extensao != '.xls':
            print("Arquivo não é XLS, não precisa converter")
            return arquivo_origem
        
        # Criar nome para arquivo convertido
        nome_base = os.path.splitext(os.path.basename(arquivo_origem))[0]
        arquivo_destino = self.pasta_rh / f"{nome_base}_convertido.xlsx"
        
        # Método 1: Tentar usar Excel via COM (Windows)
        try:
            print("Tentando converter usando Excel via COM...")
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(str(arquivo_origem))
            wb.SaveAs(str(arquivo_destino), FileFormat=51)  # 51 = xlOpenXMLWorkbook (*.xlsx)
            wb.Close()
            excel.Quit()
            
            print(f"Arquivo convertido com sucesso para {arquivo_destino}")
            return arquivo_destino
        except Exception as e:
            print(f"Erro ao converter usando Excel COM: {str(e)}")
        
        # Método 2: Tentar usar pandas para ler e salvar
        try:
            print("Tentando converter usando pandas...")
            df = pd.read_excel(arquivo_origem, engine='xlrd')
            df.to_excel(arquivo_destino, index=False)
            print(f"Arquivo convertido com sucesso para {arquivo_destino} usando pandas")
            return arquivo_destino
        except Exception as e:
            print(f"Erro ao converter usando pandas: {str(e)}")
        
        # Método 3: Tentar usar xlrd + openpyxl
        try:
            print("Tentando converter usando xlrd + openpyxl...")
            import xlrd
            from openpyxl import Workbook
            
            # Abrir arquivo XLS
            xls_wb = xlrd.open_workbook(arquivo_origem)
            xls_sheet = xls_wb.sheet_by_index(0)
            
            # Criar novo arquivo XLSX
            xlsx_wb = Workbook()
            xlsx_sheet = xlsx_wb.active
            
            # Copiar dados
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    xlsx_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
            
            # Salvar arquivo XLSX
            xlsx_wb.save(arquivo_destino)
            print(f"Arquivo convertido com sucesso para {arquivo_destino} usando xlrd + openpyxl")
            return arquivo_destino
        except Exception as e:
            print(f"Erro ao converter usando xlrd + openpyxl: {str(e)}")
        
        # Se todas as tentativas falharem, mostrar mensagem e perguntar ao usuário
        resposta = custom_messagebox("yesno", 
            "Formato Antigo do Excel", 
            "O arquivo selecionado está no formato antigo do Excel (97-2003) e não foi possível convertê-lo automaticamente.\n\n"
            "Sugestão: Abra o arquivo no Excel e salve-o como 'Pasta de Trabalho do Excel' (.xlsx).\n\n"
            "Deseja tentar abrir o arquivo original mesmo assim?"
        )
        
        if resposta:
            return arquivo_origem
        else:
            return None

    def tentar_converter_para_csv(self, arquivo):
        """Tenta converter o arquivo para CSV usando ferramentas externas"""
        try:
            # Perguntar ao usuário se deseja tentar converter
            if not custom_messagebox("yesno", 
                "Problema de Formato", 
                "O arquivo está em um formato difícil de ler. Deseja tentar convertê-lo para CSV?\n\n"
                "Isso pode ajudar a importar arquivos com problemas de compatibilidade."
            ):
                return None
                
            # Criar um nome para o arquivo CSV de saída
            base_nome = os.path.splitext(arquivo)[0]
            csv_arquivo = f"{base_nome}_convertido.csv"
            
            # Verificar se o Excel está instalado no sistema e usar para conversão
            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                
                wb = excel.Workbooks.Open(arquivo)
                # Salvar como CSV
                wb.SaveAs(csv_arquivo, FileFormat=6)  # 6 = CSV
                wb.Close()
                excel.Quit()
                
                custom_messagebox("info", 
                    "Conversão Realizada", 
                    f"Arquivo convertido para CSV: {csv_arquivo}\n\nTentando importar novamente."
                )
                return csv_arquivo
            except Exception as e:
                print(f"Erro ao usar Excel para conversão: {str(e)}")
                
                # Se o Excel falhar, tentar com pandas
                try:
                    # Tentar ler com xlrd, que pode funcionar para versões mais antigas
                    with open(arquivo, 'rb') as file:
                        data = file.read(8)  # Lê os primeiros 8 bytes para verificar a assinatura
                    
                    # Verificar se os primeiros bytes são consistentes com XLS
                    if data[:2] == b'\xd0\xcf':  # Possível assinatura de arquivo XLS
                        import xlrd
                        wb = xlrd.open_workbook(arquivo, formatting_info=False)
                        sheet = wb.sheet_by_index(0)
                        
                        # Extrair dados
                        dados = []
                        for row_idx in range(sheet.nrows):
                            row_data = []
                            for col_idx in range(sheet.ncols):
                                cell_value = sheet.cell_value(row_idx, col_idx)
                                row_data.append(cell_value)
                            dados.append(row_data)
                        
                        # Criar DataFrame e salvar como CSV
                        df = pd.DataFrame(dados)
                        df.to_csv(csv_arquivo, index=False, header=False)
                        return csv_arquivo
                except Exception as xlrd_error:
                    print(f"Erro ao converter manualmente: {str(xlrd_error)}")
                
                custom_messagebox("error", 
                    "Erro na Conversão", 
                    "Não foi possível converter o arquivo para CSV. Tente exportar o arquivo como CSV diretamente do Excel."
                )
                return None
        except Exception as e:
            print(f"Erro geral na conversão: {str(e)}")
            return None
    
    def montar_dados_bancarios(self, row):
        """Monta os dados bancários com base na forma de pagamento"""
        forma_pagto = self.obter_valor_coluna(row, 'Forma Pagto')
        
        # Se for dinheiro, deixar em branco
        if forma_pagto and str(forma_pagto).upper() == "DINHEIRO":
            return ''
            
        # Se for PIX, usar a chave PIX se estiver disponível
        if forma_pagto and str(forma_pagto).upper() == "PIX":
            chave_pix = self.obter_valor_coluna(row, 'Tipo Conta/Chave PIX')
            if chave_pix and not pd.isna(chave_pix):
                # Tratar a chave PIX para remover prefixos desnecessários
                chave_pix_str = str(chave_pix).strip()
                
                # Verificar se contém ":" e tratar os diferentes formatos
                if ":" in chave_pix_str:
                    # Remove prefixos comuns mantendo apenas o que vem depois do ":"
                    prefixos = ["CPF:", "Celular:", "E-mail:", "Email:"]
                    
                    for prefixo in prefixos:
                        if chave_pix_str.upper().startswith(prefixo.upper()):
                            # Extrair apenas o que vem depois do ":"
                            chave_pix_str = chave_pix_str[len(prefixo):].strip()
                            break
                    
                    # Se ainda houver outros formatos com ":" não listados acima
                    if ":" in chave_pix_str and not any(prefixo.upper() in chave_pix_str.upper() for prefixo in prefixos):
                        chave_pix_str = chave_pix_str.split(":", 1)[1].strip()
                
                return f"PIX: {chave_pix_str}"
        
        # Para Crédito CC ou outros casos, montar os dados bancários
        nome_banco = self.obter_valor_coluna(row, 'Nome Banco', '')
        # Substituir CAIXA ECONÔMICA FEDERAL por CAIXA
        nome_banco_normalizado = self.normalizar_texto(str(nome_banco))
        if nome_banco and "CAIXA ECONOMICA FEDERAL" in nome_banco_normalizado:
            nome_banco = "CAIXA"
            
        # Obter e formatar agência com zeros à esquerda
        agencia_raw = self.obter_valor_coluna(row, 'Agência', '')
        
        # Limpar o valor da agência para remover possíveis decimais
        if agencia_raw:
            try:
                # Remover possíveis casas decimais e converter para inteiro
                agencia_clean = str(agencia_raw).split('.')[0]
                # Formatar com zeros à esquerda para garantir 4 dígitos
                agencia = agencia_clean.zfill(4)
            except:
                # Em caso de erro, manter o valor original
                agencia = str(agencia_raw)
        else:
            agencia = ''
        
        # Tentar diferentes nomes para o número da conta
        numero_conta_raw = self.obter_valor_coluna(row, 'N° Conta', '')
        if not numero_conta_raw:
            numero_conta_raw = self.obter_valor_coluna(row, 'Conta', '')
        
        # CORREÇÃO: Limpar número da conta (remover .0 se vier como float)
        if numero_conta_raw:
            try:
                # Remover possíveis casas decimais
                numero_conta = str(numero_conta_raw).split('.')[0]
            except:
                numero_conta = str(numero_conta_raw)
        else:
            numero_conta = ''
        
        # Tentar diferentes nomes para o dígito da conta
        digito_conta_raw = self.obter_valor_coluna(row, 'Dig. Conta', '')
        if not digito_conta_raw:
            digito_conta_raw = self.obter_valor_coluna(row, 'Dígito', '')
        
        # CORREÇÃO: Limpar dígito da conta (remover .0 se vier como float)
        if digito_conta_raw and not pd.isna(digito_conta_raw):
            try:
                # Remover possíveis casas decimais
                digito_conta = str(digito_conta_raw).split('.')[0]
            except:
                digito_conta = str(digito_conta_raw)
        else:
            digito_conta = ''
            
        # Formatar conta com dígito
        conta = f"{numero_conta}"
        if digito_conta:  # Já verificamos se não é vazio
            conta = f"{numero_conta}-{digito_conta}"
            
        # Obter CPF para incluir nos dados bancários
        cpf = self.obter_valor_coluna(row, 'CPF', '')
        
        # Formatar o CPF se não estiver formatado
        cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
        if len(cpf_limpo) == 11:
            cpf_formatado = f"{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:9]}-{cpf_limpo[9:]}"
        else:
            cpf_formatado = cpf
        
        # Montar dados bancários
        partes = [nome_banco, agencia, conta, cpf_formatado]
        partes_filtradas = [str(p) for p in partes if p and not pd.isna(p) and str(p).strip()]
        
        if partes_filtradas:
            return ' - '.join(partes_filtradas)
        else:
            return 'DADOS BANCÁRIOS NÃO CADASTRADOS'
    
    def obter_valor_coluna(self, row, coluna_original, default=''):
        """Tenta obter o valor de uma coluna com tratamento de erro"""
        try:
            # Verificar se a coluna original existe
            if coluna_original in row:
                valor = row[coluna_original]
                if pd.isna(valor):
                    return default
                return valor
            
            # Verificar se o nome mapeado existe
            coluna_mapeada = self.mapeamento_cabecalhos.get(coluna_original)
            if coluna_mapeada and coluna_mapeada in row:
                valor = row[coluna_mapeada]
                if pd.isna(valor):
                    return default
                return valor
            
            # Verificar se alguma variação do nome da coluna existe (case insensitive)
            for col in row.index:
                if coluna_original.lower() == col.lower():
                    valor = row[col]
                    if pd.isna(valor):
                        return default
                    return valor
                    
            return default
        except:
            return default
    
    def obter_referencia(self, row):
        """Obtém a referência com base no tipo de folha"""
        tipo_folha = self.obter_valor_coluna(row, 'Tipo Folha', 'MENSAL')
        
        # Padronizar o valor (remover acentos, converter para maiúsculo)
        tipo_folha = self.normalizar_texto(tipo_folha)
        
        # Se o tipo_folha estiver no mapeamento, usá-lo
        for key, value in self.mapeamento_referencias.items():
            if self.normalizar_texto(key) == tipo_folha:
                return value
        
        # Caso contrário, usar SALÁRIO como padrão
        return 'SALÁRIO'
    
    def normalizar_texto(self, texto):
        """Normaliza um texto para facilitar comparações"""
        import unicodedata
        import re
        
        # Converter para string
        texto = str(texto).strip().upper()
        
        # Remover acentos
        texto = unicodedata.normalize('NFKD', texto)
        texto = ''.join([c for c in texto if not unicodedata.combining(c)])
        
        # Remover caracteres especiais
        texto = re.sub(r'[^\w\s]', '', texto)
        
        return texto
    
    def importar_dados(self):
        """Importa dados da planilha de RH apenas para o cliente atualmente selecionado"""
        # Verificar se há um cliente selecionado
        if not self.sistema.cliente_atual:
            custom_messagebox("error", 
                "Erro", 
                "Nenhum cliente selecionado. Por favor, selecione um cliente antes de importar dados."
            )
            return
            
        cliente_alvo = self.sistema.cliente_atual.upper()
        
        arquivo = self.selecionar_arquivo()
        if not arquivo:
            return

        dtypes_para_forcar = {
            'CPF': str,
            'N° Conta': str,
            'Dig. Conta': str,
            'Conta': str,
            'Dígito': str
        }
        
        # Tentar ler o arquivo
        try:
            # Verificar extensão do arquivo
            extensao = os.path.splitext(arquivo)[1].lower()
            
            # Tentar diferentes métodos de importação baseados na extensão e no conteúdo
            df = None
            erro_leitura = None
            
            # Se for CSV, tentar abrir diretamente
            if extensao in ['.csv', '.txt']:
                # Tentar diferentes delimitadores e encodings
                delimitadores = [',', ';', '\t']
                encodings = ['utf-8', 'latin-1', 'cp1252']
                
                for encoding in encodings:
                    for delim in delimitadores:
                        try:
                            print(f"Tentando abrir CSV com delimitador: {delim}, encoding: {encoding}")
                            df = pd.read_csv(
                                arquivo,
                                dtype=dtypes_para_forcar,  # CORREÇÃO: Adicionar dtype
                                delimiter=delim,
                                encoding=encoding
                            )
                            print(f"Sucesso ao abrir CSV com delimitador: {delim}, encoding: {encoding}")
                            break
                        except Exception as e:
                            print(f"Erro ao abrir CSV com delimitador {delim}, encoding {encoding}: {str(e)}")
                    
                    if df is not None:
                        break
            
            # Se não for CSV ou não conseguiu abrir, tentar como Excel
            if df is None and extensao in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
                engines = ['openpyxl', 'xlrd']
                
                for engine in engines:
                    try:
                        print(f"Tentando abrir com engine: {engine}")
                        df = pd.read_excel(
                            arquivo,
                            dtype=dtypes_para_forcar,  # CORREÇÃO: Adicionar dtype
                            engine=engine
                        )
                        print(f"Sucesso ao abrir com engine: {engine}")
                        break
                    except Exception as e:
                        print(f"Erro ao abrir com engine {engine}: {str(e)}")
                        erro_leitura = str(e)
            
            # Se ainda não conseguiu abrir, tentar converter para CSV
            if df is None:
                csv_arquivo = self.tentar_converter_para_csv(arquivo)
                if csv_arquivo:
                    # Tentar ler o CSV convertido
                    delimitadores = [',', ';', '\t']
                    for delim in delimitadores:
                        try:
                            print(f"Tentando abrir o CSV convertido com delimitador: {delim}")
                            df = pd.read_csv(
                                csv_arquivo,
                                dtype={'CPF': str},
                                delimiter=delim,
                                encoding='utf-8'
                            )
                            print(f"Sucesso ao abrir o CSV convertido")
                            break
                        except Exception as e:
                            print(f"Erro ao abrir o CSV convertido com delimitador {delim}: {str(e)}")
                            
                    # Se ainda não funcionou, tentar com encoding latin-1
                    if df is None:
                        for delim in delimitadores:
                            try:
                                df = pd.read_csv(
                                    csv_arquivo,
                                    dtype={'CPF': str},
                                    delimiter=delim,
                                    encoding='latin-1'
                                )
                                print(f"Sucesso ao abrir o CSV convertido com encoding latin-1")
                                break
                            except Exception as e:
                                print(f"Erro ao abrir o CSV convertido com encoding latin-1: {str(e)}")
            if extensao == '.xls' and df is None:
                custom_messagebox("info", 
                    "Sugestão", 
                    "Parece que você está tentando importar um arquivo no formato Excel 97-2003 (.xls).\n\n"
                    "Este formato pode causar problemas de compatibilidade. Sugerimos abrir o arquivo no Excel e "
                    "salvá-lo como 'Pasta de Trabalho do Excel' (.xlsx) antes de importar."
                )

            # Se ainda não conseguiu abrir o arquivo, oferecer opção para selecionar outro
            if df is None:
                if custom_messagebox("yesno", 
                    "Erro de Formato", 
                    "Não foi possível abrir o arquivo. Tente salvar o arquivo como CSV no Excel e importar novamente.\n\nDeseja selecionar outro arquivo?"
                ):
                    return self.importar_dados()  # Reiniciar o processo
                else:
                    return  # Cancelar importação
            
            # Mostrar informações sobre as colunas disponíveis
            print(f"Colunas disponíveis na planilha: {df.columns.tolist()}")
            
            # Pedir ao usuário que confirme o mapeamento de colunas se necessário
            if not any(col in df.columns for col in ['Empresa', 'Cliente']):
                # Nenhuma coluna de cliente encontrada, perguntar ao usuário
                custom_messagebox("info", 
                    "Configuração Manual",
                    "Não foi possível identificar automaticamente a coluna que contém o nome do cliente.\n"
                    "Na próxima tela, selecione a coluna que contém o nome do cliente/empresa."
                )
                
                # Solicitar ao usuário que escolha a coluna para o cliente
                from tkinter import simpledialog
                coluna_cliente = simpledialog.askstring(
                    "Selecionar Coluna", 
                    f"Selecione a coluna que contém o nome do cliente entre as opções:\n\n{', '.join(df.columns.tolist())}"
                )
                
                if coluna_cliente and coluna_cliente in df.columns:
                    # Adicionar ao mapeamento temporariamente
                    self.mapeamento_cabecalhos['Empresa'] = coluna_cliente
                else:
                    custom_messagebox("info",  
                        "Coluna não encontrada",
                        "Coluna selecionada não encontrada ou inválida. Continuando com as colunas padrão."
                    )
            
            registros_processados = 0
            erros = []
            
            # Processar linha por linha, mantendo o cliente atual
            cliente_atual = None
            cliente_encontrado = False
            
            # Determinar o número total de linhas para exibir progresso
            total_linhas = len(df)
            linhas_processadas = 0
            
            for idx, row in df.iterrows():
                linhas_processadas += 1
                
                if linhas_processadas % 10 == 0:
                    print(f"Processando linha {linhas_processadas}/{total_linhas} ({linhas_processadas/total_linhas*100:.1f}%)")
                
                # Verificar se esta linha tem um valor na coluna Empresa/Cliente (possível nome de cliente)
                empresa = self.obter_valor_coluna(row, 'Empresa')
                
                if empresa and not pd.isna(empresa):
                    # Corrigir possíveis espaços extras e converter para maiúsculas
                    empresa_limpa = str(empresa).strip().upper()
                    
                    # Ignorar linhas vazias ou com 'EMPRESA' (cabeçalho)
                    if empresa_limpa and empresa_limpa != 'EMPRESA':
                        cliente_atual = empresa_limpa
                        
                        # Verificar se é o cliente que estamos procurando
                        # Fazer comparação mais flexível, usando apenas parte do nome se necessário
                        if cliente_atual == cliente_alvo:
                            cliente_encontrado = True
                            print(f"Cliente alvo encontrado (match exato): {cliente_atual}")
                        elif cliente_atual in cliente_alvo or cliente_alvo in cliente_atual:
                            cliente_encontrado = True
                            print(f"Cliente alvo encontrado (match parcial): {cliente_atual} ≈ {cliente_alvo}")
                        else:
                            cliente_encontrado = False
                            print(f"Cliente diferente encontrado: {cliente_atual}, ignorando")
                
                # Se não for o cliente alvo, pular esta linha
                if not cliente_encontrado:
                    continue
                    
                # Obter os valores necessários
                nome = self.obter_valor_coluna(row, 'Nome Empregado', '').strip().upper()
                cpf = self.obter_valor_coluna(row, 'CPF', '').strip()
                valor_liquido = self.obter_valor_coluna(row, 'Valor Líquido')
                dt_pagto = self.obter_valor_coluna(row, 'Data Pagto')
                
                # Verificar se parece ser uma linha de cabeçalho
                if "NOME EMPREGADO" in nome or "FUNCIONARIO" in nome or "FUNCIONÁRIO" in nome:
                    print(f"Ignorando provável linha de cabeçalho: {nome}")
                    continue
                # Tentar outras colunas comuns para nome se não encontrar
                if not nome:
                    nome = self.obter_valor_coluna(row, 'Nome', '').strip().upper()
                    if not nome:
                        nome = self.obter_valor_coluna(row, 'Funcionário', '').strip().upper()
                
                # Tentar outras colunas comuns para valor líquido se não encontrar
                if pd.isna(valor_liquido):
                    valor_liquido = self.obter_valor_coluna(row, 'Líquido')
                    if pd.isna(valor_liquido):
                        valor_liquido = self.obter_valor_coluna(row, 'Valor')
                
                # Verificar se esta linha representa um funcionário (tem nome, CPF e valor)
                if not nome or not cpf or pd.isna(valor_liquido):
                    # Se falta apenas o CPF, mas tem nome e valor, gerar um aviso
                    if nome and not pd.isna(valor_liquido) and not cpf:
                        erros.append(f"Funcionário {nome} sem CPF definido")
                    continue
                    
               # Verificar se o valor é numérico
                try:
                    valor = float(str(valor_liquido).replace(',', '.'))
                except (ValueError, TypeError):
                    erros.append(f"Valor inválido para {nome}: {valor_liquido}")
                    continue
                
                # Definir data do relatório atual
                data_rel = self.sistema.data_rel_entry.get()
                if not data_rel:
                    hoje = datetime.now()
                    if 6 <= hoje.day <= 20:
                        data_rel = hoje.replace(day=20).strftime('%d/%m/%Y')
                    else:
                        if hoje.day > 20:
                            proximo_mes = (hoje.replace(day=1) + relativedelta(months=1))
                            data_rel = proximo_mes.replace(day=5).strftime('%d/%m/%Y')
                        else:
                            data_rel = hoje.replace(day=5).strftime('%d/%m/%Y')
                
                # Calcular data de vencimento
                dt_vencto = dt_pagto.strftime('%d/%m/%Y')
                
                # Obter referência com base no tipo de folha
                referencia = self.obter_referencia(row)
                
                # Montar dados bancários
                dados_bancarios = self.montar_dados_bancarios(row)
                # Agora verificamos explicitamente se está vazio
                if dados_bancarios == '':
                    dados_bancarios = ''  # Manter vazio quando for pagamento em dinheiro
                elif not dados_bancarios:
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
                
                # Determinar forma de pagamento
                forma_pagto = self.obter_valor_coluna(row, 'Forma Pagto', 'PIX')
                forma_pagamento = 'PIX' if forma_pagto.upper() == 'PIX' else 'TED'
                
                # Formatar o CPF
                cpf_numerico = ''.join(filter(str.isdigit, str(cpf)))
                if len(cpf_numerico) == 11:
                    cpf_formatado = f"{cpf_numerico[:3]}.{cpf_numerico[3:6]}.{cpf_numerico[6:9]}-{cpf_numerico[9:]}"
                else:
                    cpf_formatado = cpf
                
                # Criar registro
                registro = {
                    'data': data_rel,
                    'cnpj_cpf': cpf_formatado,
                    'nome': nome,
                    'categoria': 'MO',
                    'tp_desp': '1',
                    'referencia': referencia,
                    'etapa_obra': '',  
                    'insumo': '',
                    'nf': '',
                    'nf': '',
                    'vr_unit': f"{valor:.2f}",
                    'dias': 1,
                    'valor': f"{valor:.2f}",
                    'dt_vencto': dt_vencto,
                    'dados_bancarios': dados_bancarios,
                    'observacao': f"IMPORTADO RH - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                    'forma_pagamento': forma_pagamento
                }
                
                # Adicionar à lista para incluir
                self.sistema.dados_para_incluir.append(registro)
                registros_processados += 1
            
            # Relatório final
            if registros_processados > 0:
                # Solicitar etapa da obra antes de finalizar
                etapa_obra = self.solicitar_etapa_obra()
                
                if etapa_obra is None:
                    # Usuário cancelou - remover todos os registros adicionados
                    self.sistema.dados_para_incluir = [
                        r for r in self.sistema.dados_para_incluir 
                        if r.get('observacao', '').find('IMPORTADO RH') == -1
                    ]
                    custom_messagebox("info", "Importação Cancelada", 
                        "A importação foi cancelada. Nenhum dado foi salvo.")
                    return
                
                # Aplicar etapa_obra a todos os registros importados
                for registro in self.sistema.dados_para_incluir:
                    if 'IMPORTADO RH' in registro.get('observacao', ''):
                        registro['etapa_obra'] = etapa_obra
                
                mensagem = (
                    f"Importação concluída!\n\n"
                    f"Registros processados: {registros_processados} para {cliente_alvo}\n"
                    f"Etapa da Obra: {etapa_obra}\n"
                )
                
                if erros:
                    mensagem += f"\nAdvertências ({len(erros)}):\n"
                    for erro in erros[:5]:  # Limitar a 5 para não sobrecarregar
                        mensagem += f"- {erro}\n"
                    if len(erros) > 5:
                        mensagem += f"- ...e mais {len(erros) - 5} advertências\n"
                
                custom_messagebox("info", "Resultado da Importação", mensagem)
                
                # Perguntar se deseja visualizar
                if custom_messagebox("yesno", 
                    "Importação RH", 
                    "Deseja visualizar os lançamentos antes de salvar?"
                ):
                    self.sistema.visualizar_lancamentos()
            else:
                custom_messagebox("warning",  
                    "Aviso",
                    f"Nenhum registro foi processado para o cliente {cliente_alvo}. Verifique se este cliente está presente na planilha."
                )
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao importar dados: {str(e)}")
            import traceback
            traceback.print_exc()

    def importar_transporte_cafe(self):
        """
        Importa dados específicos de transporte e gera café automaticamente
        Estrutura esperada da planilha:
        - Coluna B: NOME
        - Coluna C: CPF  
        - Coluna D: DIAS
        - Coluna E: VALOR TRANSPORTE
        """
        # Verificar se há um cliente selecionado
        if not self.sistema.cliente_atual:
            custom_messagebox("error", 
                "Erro", 
                "Nenhum cliente selecionado. Por favor, selecione um cliente antes de importar dados de transporte."
            )
            return
            
        cliente_alvo = self.sistema.cliente_atual.upper()
        
        # Selecionar arquivo específico de transporte
        arquivo = filedialog.askopenfilename(
            title="Selecione a planilha de TRANSPORTE",
            filetypes=[
                ("Arquivos Excel", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("Arquivos CSV", "*.csv *.txt"),
                ("Todos os formatos suportados", "*.xlsx *.xls *.csv *.txt *.xlsm *.xlsb")
            ],
            initialdir=str(self.pasta_rh)
        )
        
        if not arquivo:
            return

        try:
            # Tentar ler o arquivo
            df = None
            extensao = os.path.splitext(arquivo)[1].lower()
            
            if extensao in ['.csv', '.txt']:
                # Para CSV, tentar diferentes delimitadores
                for delim in [',', ';', '\t']:
                    try:
                        df = pd.read_csv(arquivo, dtype={'CPF': str}, delimiter=delim, encoding='utf-8')
                        break
                    except:
                        try:
                            df = pd.read_csv(arquivo, dtype={'CPF': str}, delimiter=delim, encoding='latin-1')
                            break
                        except:
                            continue
            else:
                # Para Excel
                df = pd.read_excel(arquivo, dtype={'CPF': str})
            
            if df is None:
                custom_messagebox("error", "Erro", "Não foi possível ler o arquivo. Verifique o formato.")
                return
            
            print(f"Arquivo lido com sucesso. Colunas disponíveis: {df.columns.tolist()}")
            print(f"Primeiras linhas:\n{df.head()}")
            
            # Processar dados de transporte
            registros_processados = self.processar_dados_transporte(df, cliente_alvo)
            
            if registros_processados > 0:
                # Solicitar etapa da obra antes de finalizar
                etapa_obra = self.solicitar_etapa_obra()
                
                if etapa_obra is None:
                    # Usuário cancelou - remover todos os registros adicionados
                    self.sistema.dados_para_incluir = [
                        r for r in self.sistema.dados_para_incluir 
                        if (r.get('observacao', '').find('IMPORTADO TRANSPORTE') == -1 and
                            r.get('observacao', '').find('IMPORTADO CAFÉ') == -1)
                    ]
                    custom_messagebox("info", "Importação Cancelada", 
                        "A importação foi cancelada. Nenhum dado foi salvo.")
                    return
                
                # Aplicar etapa_obra a todos os registros de transporte e café
                for registro in self.sistema.dados_para_incluir:
                    obs = registro.get('observacao', '')
                    if 'IMPORTADO TRANSPORTE' in obs or 'IMPORTADO CAFÉ' in obs:
                        registro['etapa_obra'] = etapa_obra
                
                # Calcular totais
                total_transporte = len([r for r in self.sistema.dados_para_incluir 
                                     if r.get('referencia') == 'TRANSPORTE'])
                total_cafe = len([r for r in self.sistema.dados_para_incluir 
                                if r.get('referencia') == 'CAFÉ'])
                
                mensagem = (
                    f"🚛 Importação de TRANSPORTE concluída!\n\n"
                    f"📊 Resultados:\n"
                    f"• Lançamentos de TRANSPORTE: {total_transporte}\n"
                    f"• Lançamentos de CAFÉ (automático): {total_cafe}\n"
                    f"• Total de registros: {total_transporte + total_cafe}\n\n"
                    f"👤 Cliente: {cliente_alvo}\n"
                    f"🏗️ Etapa: {etapa_obra}"
                )
                
                custom_messagebox("info", "Sucesso na Importação", mensagem)
                
                # Perguntar se deseja visualizar
                if custom_messagebox("yesno", 
                    "Visualizar Lançamentos", 
                    "Deseja visualizar os lançamentos antes de salvar?"
                ):
                    self.sistema.visualizar_lancamentos()

            else:
                custom_messagebox("warning", 
                    "Nenhum Registro", 
                    f"Nenhum registro de transporte foi processado.\n\n"
                    f"Verifique se:\n"
                    f"• As colunas estão corretas (B=NOME, C=CPF, D=DIAS, E=VALOR)\n"
                    f"• Os dados não estão vazios\n"
                    f"• O formato do arquivo está correto"
                )
                    
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao importar dados de transporte: {str(e)}")
            import traceback
            traceback.print_exc()

    def processar_dados_transporte(self, df, cliente_alvo):
        """
        Processa os dados da planilha de transporte
        Retorna o número de registros processados
        """
        registros_processados = 0
        erros = []
        
        print(f"Iniciando processamento de {len(df)} linhas para transporte...")
        
        # Calcular data de referência baseada na regra do sistema
        data_rel = self.calcular_data_referencia_transporte()
        print(f"Data de referência calculada: {data_rel}")
        
        # Processar linha por linha
        for idx, row in df.iterrows():
            try:
                # Extrair dados das colunas específicas
                nome = self.extrair_valor_coluna_transporte(row, 'B', 1)  # Coluna B (índice 1)
                cpf = self.extrair_valor_coluna_transporte(row, 'C', 2)   # Coluna C (índice 2)
                dias = self.extrair_valor_coluna_transporte(row, 'D', 3)  # Coluna D (índice 3)
                valor_transporte = self.extrair_valor_coluna_transporte(row, 'E', 4)  # Coluna E (índice 4)
                
                print(f"Linha {idx}: Nome={nome}, CPF={cpf}, Dias={dias}, Valor={valor_transporte}")
                
                # Validar dados obrigatórios
                if not nome or not cpf or not dias or not valor_transporte:
                    if nome:  # Se tem nome mas falta outros dados
                        erros.append(f"Linha {idx+1}: {nome} - dados incompletos")
                    continue
                
                # Limpar e validar nome
                nome_limpo = str(nome).strip().upper()
                if not nome_limpo or nome_limpo in ['NOME', 'FUNCIONARIO', 'FUNCIONÁRIO']:
                    continue  # Pular cabeçalhos
                
                # Limpar e validar CPF
                cpf_numeros = ''.join(filter(str.isdigit, str(cpf)))
                if len(cpf_numeros) != 11:
                    erros.append(f"Linha {idx+1}: {nome_limpo} - CPF inválido: {cpf}")
                    continue
                
                # Formatar CPF
                cpf_formatado = f"{cpf_numeros[:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:]}"
                
                # Validar dias
                try:
                    dias_int = int(float(str(dias).replace(',', '.')))
                    if dias_int <= 0:
                        erros.append(f"Linha {idx+1}: {nome_limpo} - Dias inválido: {dias}")
                        continue
                except (ValueError, TypeError):
                    erros.append(f"Linha {idx+1}: {nome_limpo} - Dias não numérico: {dias}")
                    continue
                
                # Validar valor do transporte (valor unitário)
                try:
                    vr_unit_float = float(str(valor_transporte).replace(',', '.'))
                    if vr_unit_float <= 0:
                        erros.append(f"Linha {idx+1}: {nome_limpo} - Valor unitário inválido: {valor_transporte}")
                        continue
                except (ValueError, TypeError):
                    erros.append(f"Linha {idx+1}: {nome_limpo} - Valor unitário não numérico: {valor_transporte}")
                    continue
                
                # CORREÇÃO: Calcular valor total = vr_unit * dias
                valor_total = vr_unit_float * dias_int
                
                # Buscar dados bancários do funcionário
                dados_bancarios = self.buscar_dados_bancarios_funcionario(cpf_formatado)
                
                # Criar registro de TRANSPORTE
                registro_transporte = {
                    'data': data_rel,
                    'cnpj_cpf': cpf_formatado,
                    'nome': nome_limpo,
                    'categoria': 'MO',
                    'tp_desp': '1',
                    'referencia': 'TRANSPORTE',
                    'etapa_obra': '',  # Será preenchido após seleção do usuário
                    'insumo': '',
                    'nf': '',
                    'vr_unit': f"{vr_unit_float:.2f}",
                    'dias': dias_int,
                    'valor': f"{valor_total:.2f}",
                    'dt_vencto': data_rel,
                    'dados_bancarios': '',
                    'observacao': f"IMPORTADO TRANSPORTE - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                    'forma_pagamento': 'PIX'
                }
                
                # Adicionar registro de transporte
                self.sistema.dados_para_incluir.append(registro_transporte)
                registros_processados += 1
                
                print(f"✅ Registro de TRANSPORTE criado para {nome_limpo}")
                
                # CORREÇÃO: Criar registro de CAFÉ manualmente (replicando a lógica existente)
                try:
                    # Buscar valor do café nas configurações (igual ao método adicionar_dados)
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    config = GerenciadorConfiguracoes.carregar_configuracoes()
                    
                    if config and 'cafe' in config and 'valor_atual' in config['cafe']:
                        vr_unit_cafe = float(config['cafe']['valor_atual'])
                    else:
                        vr_unit_cafe = 4.0  # Valor padrão caso não encontre configuração
                        
                    valor_cafe_total = vr_unit_cafe * dias_int
                    
                    # Criar dados do lançamento do CAFÉ (copiando estrutura do transporte)
                    dados_cafe = registro_transporte.copy()
                    dados_cafe.update({
                        'referencia': 'CAFÉ',
                        'vr_unit': f"{vr_unit_cafe:.2f}",
                        'valor': f"{valor_cafe_total:.2f}",
                        'observacao': f"IMPORTADO CAFÉ (AUTO) - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                    })
                    
                    self.sistema.dados_para_incluir.append(dados_cafe)
                    print(f"✅ Registro de CAFÉ criado para {nome_limpo} - {dias_int} dias × R$ {vr_unit_cafe:.2f} = R$ {valor_cafe_total:.2f}")
                    
                except Exception as e:
                    print(f"⚠️ Erro ao criar CAFÉ para {nome_limpo}: {str(e)}")
                    # Continuar processamento mesmo se café falhar
                
                # IMPORTANTE: Agora tanto TRANSPORTE quanto CAFÉ são criados na importação
                
            except Exception as e:
                erro_msg = f"Linha {idx+1}: Erro ao processar - {str(e)}"
                erros.append(erro_msg)
                print(f"❌ {erro_msg}")
                continue
        
        # Mostrar erros se houver
        if erros:
            print(f"\n⚠️ Encontrados {len(erros)} erros:")
            for erro in erros[:10]:  # Mostrar apenas os primeiros 10
                print(f"  - {erro}")
            if len(erros) > 10:
                print(f"  - ... e mais {len(erros) - 10} erros")
        
        print(f"✅ Processamento concluído: {registros_processados} registros de transporte criados")
        return registros_processados

    def extrair_valor_coluna_transporte(self, row, letra_coluna, indice):
        """
        Extrai valor de uma coluna específica por letra (B, C, D, E) ou índice
        """
        try:
            # Primeiro tentar por letra da coluna se o DataFrame tem colunas nomeadas assim
            if hasattr(row, 'index') and letra_coluna in row.index:
                valor = row[letra_coluna]
            # Tentar por índice numérico
            elif hasattr(row, 'iloc') and len(row) > indice:
                valor = row.iloc[indice]
            # Se row for uma Series com índices numéricos
            elif len(row) > indice:
                valor = row.iloc[indice] if hasattr(row, 'iloc') else row[indice]
            else:
                return None
            
            # Tratar valores NaN ou vazios
            if pd.isna(valor) or valor == '':
                return None
                
            return valor
            
        except Exception as e:
            print(f"Erro ao extrair coluna {letra_coluna} (índice {indice}): {str(e)}")
            return None

    def calcular_data_referencia_transporte(self):
        """
        Calcula a data de referência baseada na regra: dia 5 ou 20
        """
        hoje = datetime.now()
        
        if 6 <= hoje.day <= 20:
            # Se estivermos entre dia 6 e 20, a referência é dia 20 do mesmo mês
            data_rel = hoje.replace(day=20)
        else:
            if hoje.day > 20:
                # Se estivermos após dia 20, a referência é dia 5 do próximo mês
                proximo_mes = (hoje.replace(day=1) + relativedelta(months=1))
                data_rel = proximo_mes.replace(day=5)
            else:
                # Se estivermos antes do dia 6, a referência é dia 5 do mesmo mês
                data_rel = hoje.replace(day=5)
        
        return data_rel.strftime('%d/%m/%Y')

    def obter_valor_cafe_configurado(self):
        """Busca o valor do café nas configurações do sistema"""
        try:
            # Método 1: Tentar buscar nas configurações
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            config = GerenciadorConfiguracoes.carregar_configuracoes()
            
            if config and 'cafe' in config and 'valor_atual' in config['cafe']:
                valor_cafe = float(config['cafe']['valor_atual'])
                print(f"📋 Valor do café encontrado nas configurações: R$ {valor_cafe:.2f}")
                return valor_cafe
                
        except Exception as e:
            print(f"⚠️ Erro ao buscar configuração de café: {str(e)}")
        
        # Método 2: Tentar buscar no sistema principal
        try:
            if hasattr(self.sistema, 'gestao_taxas') and hasattr(self.sistema.gestao_taxas, 'configuracoes'):
                config = self.sistema.gestao_taxas.configuracoes
                if config and 'cafe' in config:
                    valor_cafe = float(config['cafe'].get('valor_atual', 4.0))
                    print(f"📋 Valor do café encontrado no sistema: R$ {valor_cafe:.2f}")
                    return valor_cafe
        except Exception as e:
            print(f"⚠️ Erro ao buscar no sistema: {str(e)}")
        
        # Valor padrão
        valor_padrao = 4.0
        print(f"📋 Usando valor padrão do café: R$ {valor_padrao:.2f}")
        return valor_padrao


    def buscar_dados_bancarios_funcionario(self, cpf):
        """
        Busca dados bancários do funcionário na base de fornecedores
        """
        try:
            # Usar a função existente do sistema
            if hasattr(self.sistema, 'buscar_fornecedor_completo'):
                fornecedor = self.sistema.buscar_fornecedor_completo(cpf)
                if fornecedor:
                    # Preferir PIX se disponível
                    if fornecedor.get('chave_pix'):
                        return f"PIX: {fornecedor['chave_pix']}"
                    else:
                        # Montar dados bancários para TED
                        partes = []
                        if fornecedor.get('banco'): partes.append(str(fornecedor['banco']))
                        if fornecedor.get('op'): partes.append(str(fornecedor['op']))
                        if fornecedor.get('agencia'): partes.append(str(fornecedor['agencia']))
                        if fornecedor.get('conta'): partes.append(str(fornecedor['conta']))
                        partes.append(cpf)  # Sempre adicionar CPF
                        
                        return ' - '.join(partes) if partes else 'DADOS BANCÁRIOS NÃO CADASTRADOS'
            
            return 'DADOS BANCÁRIOS NÃO CADASTRADOS'
            
        except Exception as e:
            print(f"Erro ao buscar dados bancários para {cpf}: {str(e)}")
            return 'DADOS BANCÁRIOS NÃO CADASTRADOS'

class GestorTaxasAdministracao:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
                
    def recalcular_taxas_afetadas(self, data_referencia, cliente=None, mostrar_detalhes=True):
        """
        VERSÃO CORRIGIDA do recálculo de taxas usando a lógica validada do finalizacao_quinzena.py
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual
            
            if not cliente:
                return {"sucesso": False, "mensagem": "Nenhum cliente especificado"}
            
            print(f"DEBUG: Iniciando recálculo de taxas para {cliente} em {data_referencia}")
            
            # CORREÇÃO 1: Chamar o método corretamente (sem parâmetro self extra)
            novo_valor_base = self.calcular_base_calculo_taxa(data_referencia)
            print(f"DEBUG: Nova base calculada: R$ {novo_valor_base:.2f}")
            
            if novo_valor_base == 0:
                return {"sucesso": True, "mensagem": "Sem lançamentos base para recálculo"}
            
            # 2. Obter percentual usando método corrigido
            percentual_taxa = self.obter_percentual_taxa_cliente(cliente)
            print(f"DEBUG: Percentual encontrado: {percentual_taxa}%")
            
            if percentual_taxa == 0:
                return {"sucesso": True, "mensagem": "Sem taxa percentual configurada"}
            
            # 3. Calcular novo valor da taxa
            novo_valor_taxa = novo_valor_base * (percentual_taxa / 100)
            print(f"DEBUG: Novo valor da taxa: R$ {novo_valor_taxa:.2f}")
            
            # 4. Verificar valor atual das taxas na planilha
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_dados = wb["Dados"]
            
            # Converter data
            if isinstance(data_referencia, str):
                data_ref = datetime.strptime(data_referencia, '%d/%m/%Y')
            else:
                data_ref = data_referencia
            
            # Buscar taxas existentes (tipo 7) na data
            valor_atual_total = 0
            linhas_taxa = []
            
            for idx, row in enumerate(ws_dados.iter_rows(min_row=2, values_only=True), start=2):
                data_lancamento = row[0]
                if isinstance(data_lancamento, datetime):
                    if (data_lancamento.day == data_ref.day and 
                        data_lancamento.month == data_ref.month and 
                        data_lancamento.year == data_ref.year):
                        
                        tipo_desp = row[1]
                        if tipo_desp == 7:  # Taxa ADM
                            # CORREÇÃO 2: Verificar status antes de incluir no valor atual
                            status = row[13] if len(row) > 13 else "ATIVO"  # Coluna N (STATUS)
                            
                            if status == "ATIVO":  # Só considerar taxas ativas
                                valor = row[8]  # Coluna I
                                if valor:
                                    valor_numeric = float(str(valor).replace(',', '.'))
                                    valor_atual_total += valor_numeric
                                    linhas_taxa.append(idx)
            
            print(f"DEBUG: Valor atual total das taxas ATIVAS: R$ {valor_atual_total:.2f}")
            
            # 5. Verificar se precisa recalcular
            diferenca = abs(novo_valor_taxa - valor_atual_total)
            
            if diferenca < 0.01:  # Diferença menor que 1 centavo
                wb.close()
                return {"sucesso": True, "mensagem": f"Taxas já estão corretas (R$ {valor_atual_total:.2f})"}
            
            # 6. Se chegou aqui, precisa recalcular
            print(f"DEBUG: Diferença detectada: R$ {diferenca:.2f}")
            
            # CORREÇÃO 3: Marcar como excluído ao invés de deletar fisicamente
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            
            for linha in linhas_taxa:
                # Marcar como EXCLUIDO
                ws_dados.cell(row=linha, column=14, value='EXCLUIDO')  # STATUS
                
                # Atualizar histórico
                historico_atual = ws_dados.cell(row=linha, column=16).value or ""
                novo_historico = f"{historico_atual} | EXCLUÍDA P/ RECÁLCULO EM: {timestamp}" if historico_atual else f"EXCLUÍDA P/ RECÁLCULO EM: {timestamp}"
                ws_dados.cell(row=linha, column=16, value=novo_historico)
            
            print(f"DEBUG: {len(linhas_taxa)} linhas de taxa marcadas como excluídas")
            
            # 7. Obter administradores e lançar novas taxas
            administradores = self.sistema.obter_administradores_cliente_CORRIGIDO(cliente)
            
            if not administradores:
                wb.close()
                return {"sucesso": False, "mensagem": "Nenhum administrador encontrado"}
            
            # 8. Lançar novas taxas (usar a mesma lógica do finalizacao_quinzena.py)
            taxa_total_percentual = sum(adm['percentual'] for adm in administradores)
            
            for adm in administradores:
                valor_adm = (novo_valor_taxa * adm['percentual']) / taxa_total_percentual
                
                # Determinar data de vencimento e quinzena
                if data_ref.day == 5:
                    dt_vencto = data_ref
                    while dt_vencto.weekday() >= 5:  # Ajustar fim de semana
                        dt_vencto += relativedelta(days=1)
                    quinzena = "1ª"
                else:
                    dt_vencto = data_ref
                    quinzena = "2ª"
                
                # Inserir nova linha
                proxima_linha = ws_dados.max_row + 1
                
                # CORREÇÃO 4: Gerar ID sequencial consistente
                id_lancamento = self._obter_proximo_id_sequencial(ws_dados)
                
                # Preencher dados (mesma estrutura do finalizacao_quinzena.py)
                ws_dados.cell(row=proxima_linha, column=1, value=data_ref)
                ws_dados.cell(row=proxima_linha, column=1).number_format = 'DD/MM/YYYY'
                ws_dados.cell(row=proxima_linha, column=2, value=7)  # Tipo taxa ADM
                ws_dados.cell(row=proxima_linha, column=3, value=adm['cnpj_cpf'])
                ws_dados.cell(row=proxima_linha, column=4, value=adm['nome'])
                
                referencia = f"ADM. OBRA REF. {quinzena} QUINZ. {data_ref.strftime('%m/%Y')}"
                ws_dados.cell(row=proxima_linha, column=5, value=referencia)
                ws_dados.cell(row=proxima_linha, column=6, value='')  # NF
                
                ws_dados.cell(row=proxima_linha, column=7, value=valor_adm)
                ws_dados.cell(row=proxima_linha, column=7).number_format = '#,##0.00'
                ws_dados.cell(row=proxima_linha, column=8, value=1)  # Dias
                ws_dados.cell(row=proxima_linha, column=9, value=valor_adm)
                ws_dados.cell(row=proxima_linha, column=9).number_format = '#,##0.00'
                
                ws_dados.cell(row=proxima_linha, column=10, value=dt_vencto)
                ws_dados.cell(row=proxima_linha, column=10).number_format = 'DD/MM/YYYY'
                ws_dados.cell(row=proxima_linha, column=11, value='ADM')
                
                # Buscar dados bancários
                from src.config.utils import buscar_dados_bancarios_fornecedor
                dados_bancarios = buscar_dados_bancarios_fornecedor(adm['cnpj_cpf'])
                ws_dados.cell(row=proxima_linha, column=12, value=dados_bancarios)
                
                # CORREÇÃO 5: Observação mais detalhada
                obs = f"RECÁLCULO AUTO - BASE: R$ {novo_valor_base:.2f} - {timestamp}"
                ws_dados.cell(row=proxima_linha, column=13, value=obs)
                
                # CORREÇÃO 6: Status e ID
                ws_dados.cell(row=proxima_linha, column=14, value='ATIVO')  # STATUS
                ws_dados.cell(row=proxima_linha, column=15, value=id_lancamento)  # ID_LANCAMENTO
                
                # Histórico inicial
                historico_inicial = f"CRIADO POR RECÁLCULO EM: {timestamp}"
                ws_dados.cell(row=proxima_linha, column=16, value=historico_inicial)
                
                print(f"DEBUG: Taxa lançada para {adm['nome']}: R$ {valor_adm:.2f} (ID: {id_lancamento})")
            
            # Salvar arquivo
            wb.save(arquivo_cliente)
            
            mensagem = f"Taxas recalculadas com sucesso! "
            mensagem += f"Base: R$ {novo_valor_base:.2f} | "
            mensagem += f"Taxa: {percentual_taxa}% | "
            mensagem += f"Valor total: R$ {novo_valor_taxa:.2f}"
            
            return {"sucesso": True, "mensagem": mensagem}
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            print(f"DEBUG: Erro no recálculo: {str(e)}")
            import traceback
            print(f"DEBUG: Traceback completo: {traceback.format_exc()}")
            return {"sucesso": False, "mensagem": f"Erro no recálculo: {str(e)}"}
 
    def _obter_proximo_id_sequencial(self, worksheet):
        """
        Obtém o próximo ID sequencial disponível (compatível com sistema principal)
        """
        try:
            max_id = 0
            
            # Percorrer coluna 15 (ID_LANCAMENTO) para encontrar o maior ID
            for row in range(2, worksheet.max_row + 1):
                id_valor = worksheet.cell(row=row, column=15).value
                if id_valor is not None:
                    try:
                        id_int = int(float(id_valor))
                        if id_int > max_id:
                            max_id = id_int
                    except (ValueError, TypeError):
                        continue
            
            return max_id + 1
            
        except Exception as e:
            print(f"DEBUG: Erro ao obter próximo ID: {str(e)}")
            # Fallback: usar número da linha como ID
            return worksheet.max_row
    
    def identificar_lancamentos_taxa_admin(self, df):
        """
        Identifica lançamentos de taxa de administração com padrões mais amplos
        """
        if df.empty:
            return pd.DataFrame()
            
        mask_taxa = df['TP_DESP'] == 7
    
        taxas = df[mask_taxa].copy()
        print(f"DEBUG: Taxas encontradas (tp_desp=7): {len(taxas)} registros")
        
        if not taxas.empty:
            print(f"DEBUG: Referências das taxas: {taxas['REFERÊNCIA'].tolist()}")
        
        return taxas

    def calcular_base_calculo_taxa(self, data_referencia, df=None):
        """
        VERSÃO UNIFICADA - Calcula valor base seguindo a lógica corrigida do finalizacao_quinzena.py
        
        Parâmetros:
        - data_referencia: Data para cálculo (str ou datetime)
        - df: DataFrame opcional (para compatibilidade com código existente)
            Se não fornecido, lê diretamente da planilha
        """
        try:
            print(f"DEBUG: Calculando valor base para {data_referencia}")
            
            # Se DataFrame foi fornecido, usar lógica compatível
            if df is not None:
                return self._calcular_base_por_dataframe(df, data_referencia)
            
            # Caso contrário, usar lógica corrigida da planilha
            return self._calcular_base_por_planilha(data_referencia)
            
        except Exception as e:
            print(f"DEBUG: Erro ao calcular valor base: {str(e)}")
            return 0

    def _calcular_base_por_planilha(self, data_referencia):
        """
        Método interno - Calcula base lendo diretamente da planilha
        (Lógica corrigida do finalizacao_quinzena.py)
        """
        try:
            cliente_atual = self.sistema.cliente_atual
            
            arquivo_cliente = PASTA_CLIENTES / f"{cliente_atual}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_dados = wb["Dados"]
            
            # Converter data de referência se necessário
            if isinstance(data_referencia, str):
                data_ref = datetime.strptime(data_referencia, '%d/%m/%Y')
            else:
                data_ref = data_referencia
            
            print(f"DEBUG: Data de referência (planilha): {data_ref.strftime('%d/%m/%Y')}")
            
            valor_base = 0
            lancamentos_encontrados = 0
            
            # Usar a mesma lógica do finalizacao_quinzena.py
            for row in ws_dados.iter_rows(min_row=2, values_only=True):
                data_lancamento = row[0]  # Coluna A
                
                if isinstance(data_lancamento, datetime):
                    # Verificar se é da mesma data (dia, mês, ano)
                    if (data_lancamento.day == data_ref.day and 
                        data_lancamento.month == data_ref.month and 
                        data_lancamento.year == data_ref.year):
                        
                        tipo_desp = row[1]  # Coluna B (TP_DESP)
                        status = row[13] if len(row) > 13 else "ATIVO"  # Coluna N (STATUS)
                        
                        # Incluir apenas tipos 1 a 6 e status ATIVO
                        if (isinstance(tipo_desp, (int, float)) and 1 <= tipo_desp <= 6 and 
                            status == "ATIVO"):
                            valor = row[8]  # Coluna I (VALOR)
                            
                            if valor:
                                try:
                                    valor_numeric = float(str(valor).replace(',', '.'))
                                    valor_base += valor_numeric
                                    lancamentos_encontrados += 1
                                    
                                    print(f"DEBUG: Lançamento incluído - Tipo: {tipo_desp}, Valor: R$ {valor_numeric:.2f}")
                                    
                                except (ValueError, TypeError) as e:
                                    print(f"DEBUG: Erro ao processar valor '{valor}': {e}")
                                    continue
            
            print(f"DEBUG: Valor base total (planilha): R$ {valor_base:.2f}")
            print(f"DEBUG: Total de lançamentos incluídos: {lancamentos_encontrados}")
            
            wb.close()
            return valor_base
            
        except Exception as e:
            print(f"DEBUG: Erro ao calcular valor base por planilha: {str(e)}")
            if 'wb' in locals():
                wb.close()
            return 0

    def _calcular_base_por_dataframe(self, df, data_referencia):
        """
        Método interno - Calcula base usando DataFrame fornecido
        (Para compatibilidade com verificações existentes)
        """
        try:
            # Converter data de referência
            if isinstance(data_referencia, str):
                data_ref = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref = pd.to_datetime(data_referencia)
            
            print(f"DEBUG: Data de referência (DataFrame): {data_ref.strftime('%d/%m/%Y')}")
            
            # Garantir que DATA_REL existe e está em formato datetime
            if 'DATA_REL' not in df.columns:
                print("DEBUG: Coluna DATA_REL não encontrada no DataFrame")
                return 0
            
            df = df.copy()
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            
            # Filtrar dados para a data específica
            df_data = df[df['DATA_REL'].dt.date == data_ref.date()].copy()
            
            if df_data.empty:
                print(f"DEBUG: Nenhum lançamento encontrado para {data_ref.strftime('%d/%m/%Y')}")
                return 0
            
            # Garantir que STATUS existe
            if 'STATUS' not in df_data.columns:
                df_data['STATUS'] = 'ATIVO'
            
            # Filtrar apenas lançamentos ativos e tipos 1-6
            df_base = df_data[
                (df_data['STATUS'] == 'ATIVO') & 
                (df_data['TP_DESP'].isin([1, 2, 3, 4, 5, 6]))
            ].copy()
            
            if df_base.empty:
                print("DEBUG: Nenhum lançamento ativo dos tipos 1-6 encontrado")
                return 0
            
            # Converter valores para numérico
            df_base['VALOR_NUM'] = pd.to_numeric(
                df_base['VALOR'].astype(str).str.replace('R$', '').str.replace(',', '.'),
                errors='coerce'
            ).fillna(0)
            
            valor_base = df_base['VALOR_NUM'].sum()
            
            print(f"DEBUG: Valor base total (DataFrame): R$ {valor_base:.2f}")
            print(f"DEBUG: Lançamentos incluídos: {len(df_base)}")
            
            return valor_base
            
        except Exception as e:
            print(f"DEBUG: Erro ao calcular valor base por DataFrame: {str(e)}")
            return 0
    
    def excluir_taxas_base_zerada(self, arquivo_cliente, taxas_existentes):
        """
        Marca taxas como excluídas quando a base for zerada
        """
        try:
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            taxas_excluidas = []
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            
            for _, taxa in taxas_existentes.iterrows():
                id_taxa = taxa.get('ID_LANCAMENTO')
                if pd.isna(id_taxa):
                    continue
                
                # Encontrar linha na planilha
                for row_num in range(2, ws.max_row + 1):
                    if ws.cell(row=row_num, column=15).value == id_taxa:
                        # Marcar como excluído
                        ws.cell(row=row_num, column=14, value='EXCLUIDO')  # STATUS
                        
                        # Atualizar histórico
                        historico_atual = ws.cell(row=row_num, column=16).value or ""
                        novo_historico = f"{historico_atual} | EXCLUÍDA (BASE ZERADA) EM: {timestamp}" if historico_atual else f"EXCLUÍDA (BASE ZERADA) EM: {timestamp}"
                        ws.cell(row=row_num, column=16, value=novo_historico)
                        
                        taxas_excluidas.append({
                            'id': id_taxa,
                            'referencia': taxa.get('REFERÊNCIA', ''),
                            'valor': taxa.get('VALOR', 0)
                        })
                        break
            
            wb.save(arquivo_cliente)
            
            return {
                "sucesso": True,
                "mensagem": f"Taxas excluídas por base zerada: {len(taxas_excluidas)} itens",
                "detalhes": taxas_excluidas,
                "nova_base": 0,
                "novo_valor_total": 0
            }
            
        except Exception as e:
            return {"sucesso": False, "mensagem": f"Erro ao excluir taxas: {str(e)}"}
    
    def atualizar_taxas_na_planilha(self, arquivo_cliente, taxas_existentes, novo_valor, nova_base, percentual):
        """
        Atualiza os valores das taxas EXISTENTES na planilha com histórico detalhado
        
        IMPORTANTE: Este método ATUALIZA taxas já lançadas, não cria novas!
        Quando uma taxa já foi lançada e a base muda, ajustamos o valor da taxa existente.
        """
        try:
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            taxas_atualizadas = []
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            
            print(f"DEBUG: Atualizando {len(taxas_existentes)} taxas já lançadas")
            print(f"DEBUG: Novo valor total a distribuir: R$ {novo_valor:,.2f}")
            
            # Se há múltiplas taxas, distribuir proporcionalmente
            if len(taxas_existentes) > 1:
                print(f"DEBUG: Distribuindo entre {len(taxas_existentes)} taxas existentes")
                
                # Calcular total atual das taxas ATIVAS para proporção
                total_atual = 0
                taxas_ativas = []
                
                for _, taxa in taxas_existentes.iterrows():
                    status = taxa.get('STATUS', 'ATIVO')
                    if status != 'EXCLUIDO':
                        try:
                            valor_atual = float(str(taxa.get('VALOR', 0)).replace(',', '.'))
                            total_atual += valor_atual
                            taxas_ativas.append((taxa, valor_atual))
                        except:
                            taxas_ativas.append((taxa, 0))
                
                if total_atual == 0:
                    # Se total atual é zero, dividir igualmente entre taxas ativas
                    valor_por_taxa = novo_valor / len(taxas_ativas) if taxas_ativas else 0
                    proporcoes = [valor_por_taxa] * len(taxas_ativas)
                    print(f"DEBUG: Divisão igual: R$ {valor_por_taxa:,.2f} por taxa")
                else:
                    # Calcular proporcionalmente ao valor atual
                    proporcoes = []
                    for taxa, valor_atual in taxas_ativas:
                        proporcao = (valor_atual / total_atual) * novo_valor
                        proporcoes.append(proporcao)
                        print(f"DEBUG: Taxa {taxa.get('ID_LANCAMENTO')}: R$ {valor_atual:,.2f} → R$ {proporcao:,.2f}")
                
                # Usar apenas taxas ativas para atualização
                taxas_para_processar = [(taxa, prop) for (taxa, _), prop in zip(taxas_ativas, proporcoes)]
            else:
                # Apenas uma taxa - usar valor total
                taxa_unica = taxas_existentes.iloc[0]
                if taxa_unica.get('STATUS', 'ATIVO') != 'EXCLUIDO':
                    taxas_para_processar = [(taxa_unica, novo_valor)]
                    print(f"DEBUG: Taxa única: R$ {novo_valor:,.2f}")
                else:
                    taxas_para_processar = []
                    print(f"DEBUG: Taxa única está excluída, não atualizando")
            
            # Atualizar cada taxa EXISTENTE na planilha
            for taxa, valor_novo in taxas_para_processar:
                id_taxa = taxa.get('ID_LANCAMENTO')
                if pd.isna(id_taxa):
                    print(f"DEBUG: Taxa sem ID, pulando")
                    continue
                
                print(f"DEBUG: Procurando taxa ID {id_taxa} na planilha")
                
                # Encontrar linha na planilha pelo ID
                linha_encontrada = False
                for row_num in range(2, ws.max_row + 1):
                    id_na_planilha = ws.cell(row=row_num, column=15).value  # ID_LANCAMENTO
                    
                    if id_na_planilha == id_taxa:
                        linha_encontrada = True
                        valor_antigo = ws.cell(row=row_num, column=9).value or 0  # VALOR
                        
                        print(f"DEBUG: Encontrada linha {row_num}, atualizando valor: R$ {valor_antigo:,.2f} → R$ {valor_novo:,.2f}")
                        
                        # ATUALIZAR O VALOR DA TAXA EXISTENTE
                        ws.cell(row=row_num, column=9, value=round(valor_novo, 2))
                        
                        # Se for tipo 1 (com dias), atualizar valor unitário também
                        tp_desp = ws.cell(row=row_num, column=2).value
                        if tp_desp == 1:
                            dias = ws.cell(row=row_num, column=8).value or 1
                            if dias > 0:
                                vr_unit_novo = round(valor_novo / dias, 2)
                                ws.cell(row=row_num, column=7, value=vr_unit_novo)
                                print(f"DEBUG: Valor unitário atualizado: R$ {vr_unit_novo:,.2f}")
                        
                        # Garantir que status seja ATIVO (caso tenha sido excluído por engano)
                        status_atual = ws.cell(row=row_num, column=14).value
                        if status_atual != 'ATIVO':
                            ws.cell(row=row_num, column=14, value='ATIVO')
                            print(f"DEBUG: Status corrigido de {status_atual} para ATIVO")
                        
                        # Atualizar observação com informações detalhadas do recálculo
                        obs_atual = ws.cell(row=row_num, column=13).value or ""
                        # Limpar observações de recálculos anteriores para evitar texto muito longo
                        if "RECALC:" in obs_atual:
                            obs_base = obs_atual.split(" - RECALC:")[0]
                        else:
                            obs_base = obs_atual
                        
                        nova_obs = f"{obs_base} - TAXA ADM {percentual}% - BASE: R$ {nova_base:,.2f} - RECALC: {timestamp}".strip()
                        ws.cell(row=row_num, column=13, value=nova_obs)
                        
                        # Atualizar histórico de alterações
                        historico_atual = ws.cell(row=row_num, column=16).value or ""
                        acao = f"RECALC AUTO: R$ {valor_antigo:,.2f} → R$ {valor_novo:,.2f} (Base: R$ {nova_base:,.2f}) - {timestamp}"
                        
                        if historico_atual:
                            # Limitar histórico para não ficar muito longo (manter últimas 5 ações)
                            historico_partes = historico_atual.split(' | ')
                            if len(historico_partes) >= 5:
                                historico_partes = historico_partes[-4:]  # Manter últimas 4
                            novo_historico = ' | '.join(historico_partes) + ' | ' + acao
                        else:
                            novo_historico = acao
                        
                        ws.cell(row=row_num, column=16, value=novo_historico)
                        
                        taxas_atualizadas.append({
                            'id': id_taxa,
                            'linha': row_num,
                            'referencia': taxa.get('REFERÊNCIA', ''),
                            'valor_antigo': valor_antigo,
                            'valor_novo': valor_novo,
                            'diferenca': valor_novo - float(valor_antigo),
                        })
                        
                        print(f"✅ Taxa ID {id_taxa} atualizada com sucesso na linha {row_num}")
                        break
                
                if not linha_encontrada:
                    print(f"❌ ERRO: Taxa ID {id_taxa} não encontrada na planilha!")
                    # Isso é um problema - taxa existe no DataFrame mas não na planilha
                    # Pode indicar inconsistência nos dados
            
            # Salvar alterações na planilha
            wb.save(arquivo_cliente)
            print(f"✅ Planilha salva com {len(taxas_atualizadas)} taxas atualizadas")
            
            return {
                "sucesso": True,
                "mensagem": f"Taxas EXISTENTES recalculadas: {len(taxas_atualizadas)} itens atualizados",
                "detalhes": taxas_atualizadas,
                "nova_base": nova_base,
                "novo_valor_total": novo_valor,
                "percentual": percentual,
                "observacao": "ATUALIZAÇÃO de taxas já lançadas, não criação de novas taxas"
            }
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro ao atualizar taxas existentes: {traceback.format_exc()}")
            return {"sucesso": False, "mensagem": f"Erro ao atualizar taxas na planilha: {str(e)}"}

    def criar_nova_taxa_se_necessario(self, data_referencia, cliente=None):
        """
        MÉTODO SEPARADO: Cria nova taxa apenas quando não existe nenhuma para a data
        
        Este método deve ser usado apenas quando:
        1. Não existe nenhuma taxa para a data/quinzena
        2. O usuário está finalizando a quinzena pela primeira vez
        
        NÃO usar este método quando já existem taxas lançadas!
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual
                
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            
            # Verificar se já existem taxas para esta data
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            
            if isinstance(data_referencia, str):
                data_ref_dt = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref_dt = pd.to_datetime(data_referencia)
            
            df_data = df[df['DATA_REL'].dt.date == data_ref_dt.date()]
            taxas_existentes = self.identificar_lancamentos_taxa_admin(df_data)
            
            if not taxas_existentes.empty:
                return {
                    "sucesso": False, 
                    "mensagem": f"Já existem {len(taxas_existentes)} taxa(s) para esta data. Use o recálculo ao invés de criar nova."
                }
            
            # Calcular base e valor da nova taxa
            base_calculo = self.calcular_base_calculo_taxa(df, data_ref_dt.date())
            
            if base_calculo <= 0:
                return {
                    "sucesso": False,
                    "mensagem": "Base de cálculo zerada. Não é possível criar taxa de administração."
                }
            
            percentual = self.obter_percentual_taxa_cliente(cliente) # or self.percentual_padrao
            valor_taxa = base_calculo * (percentual / 100)
            
            # Aqui você implementaria a lógica para criar um novo lançamento de taxa
            # (Similar ao que já existe no sistema para criar lançamentos normais)
            
            return {
                "sucesso": True,
                "mensagem": f"Nova taxa criada: R$ {valor_taxa:,.2f} ({percentual}% de R$ {base_calculo:,.2f})",
                "valor_taxa": valor_taxa,
                "base_calculo": base_calculo,
                "percentual": percentual
            }
            
        except Exception as e:
            return {"sucesso": False, "mensagem": f"Erro ao criar nova taxa: {str(e)}"}

    def distinguir_cenarios_taxa(self, data_referencia, cliente=None):
        """
        MÉTODO UTILITÁRIO: Distingue entre diferentes cenários de taxa
        
        Retorna:
        - "sem_taxa": Não há taxa para esta data (primeira finalização)
        - "taxa_existente": Há taxa que pode ser recalculada
        - "taxa_excluida": Há taxa mas está excluída
        - "multiplas_taxas": Há múltiplas taxas (situação complexa)
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual
                
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            
            if isinstance(data_referencia, str):
                data_ref_dt = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref_dt = pd.to_datetime(data_referencia)
            
            df_data = df[df['DATA_REL'].dt.date == data_ref_dt.date()]
            taxas_todas = self.identificar_lancamentos_taxa_admin(df_data)
            
            if taxas_todas.empty:
                return "sem_taxa", "Nenhuma taxa encontrada para esta data"
            
            # Adicionar coluna STATUS se não existir
            if 'STATUS' not in taxas_todas.columns:
                taxas_todas['STATUS'] = 'ATIVO'
            
            taxas_ativas = taxas_todas[taxas_todas['STATUS'] != 'EXCLUIDO']
            taxas_excluidas = taxas_todas[taxas_todas['STATUS'] == 'EXCLUIDO']
            
            if len(taxas_ativas) > 1:
                return "multiplas_taxas", f"{len(taxas_ativas)} taxas ativas encontradas"
            elif len(taxas_ativas) == 1:
                return "taxa_existente", f"1 taxa ativa encontrada (ID: {taxas_ativas.iloc[0].get('ID_LANCAMENTO', 'N/A')})"
            elif len(taxas_excluidas) > 0:
                return "taxa_excluida", f"{len(taxas_excluidas)} taxa(s) excluída(s) encontrada(s)"
            else:
                return "sem_taxa", "Nenhuma taxa ativa encontrada"
                
        except Exception as e:
            return "erro", f"Erro ao analisar cenário: {str(e)}"

    def obter_percentual_taxa_cliente(self, cliente):
        """
        VERSÃO CORRIGIDA - Busca percentual de taxa seguindo a mesma lógica do finalizacao_quinzena.py
        """
        try:
            print(f"DEBUG: Buscando percentual da taxa para cliente {cliente}")
            
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            
            if 'Contratos_ADM' not in wb.sheetnames:
                print("DEBUG: Aba 'Contratos_ADM' não encontrada")
                wb.close()
                return 0
            
            ws_contratos = wb['Contratos_ADM']
            print(f"DEBUG: Aba 'Contratos_ADM' carregada")
            
            # CORREÇÃO 1: Usar a mesma lógica do finalizacao_quinzena.py
            # 1º Passo: Encontrar contratos ativos
            contratos_ativos = set()
            for row in ws_contratos.iter_rows(min_row=3, values_only=True):  # Começar da linha 3
                if row[0] and row[3] == 'ATIVO':  # Coluna A (Nº Contrato) e Coluna D (Status)
                    contratos_ativos.add(row[0])
                    print(f"DEBUG: Contrato ativo encontrado: {row[0]}")
            
            print(f"DEBUG: Contratos ativos: {contratos_ativos}")
            
            if not contratos_ativos:
                print("DEBUG: Nenhum contrato ativo encontrado")
                wb.close()
                return 0
            
            # CORREÇÃO 2: Para cada contrato ativo, buscar administradores com taxa percentual
            taxa_total = 0
            administradores_encontrados = []
            
            for num_contrato in contratos_ativos:
                print(f"DEBUG: Verificando administradores do contrato {num_contrato}")
                
                for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                    # CORREÇÃO 3: Verificar se pertence ao contrato (coluna G) e é do tipo Percentual (coluna J)
                    if (row[6] == num_contrato and          # Coluna G (Nº Contrato)
                        row[9] == 'Percentual'):            # Coluna J (Tipo)
                        
                        # CORREÇÃO 4: Extrair percentual da coluna K
                        percentual_raw = row[10]  # Coluna K (Valor/Percentual)
                        
                        print(f"DEBUG: Administrador encontrado:")
                        print(f"  - CNPJ/CPF: {row[7]}")     # Coluna H
                        print(f"  - Nome: {row[8]}")         # Coluna I
                        print(f"  - Tipo: {row[9]}")         # Coluna J
                        print(f"  - Percentual bruto: '{percentual_raw}'")  # Coluna K
                        
                        try:
                            # CORREÇÃO 5: Processar o percentual corretamente
                            if percentual_raw:
                                # Converter para string e limpar
                                percentual_str = str(percentual_raw).strip()
                                
                                # Remover % se existir e converter vírgula para ponto
                                percentual_limpo = percentual_str.replace('%', '').replace(',', '.')
                                
                                percentual_float = float(percentual_limpo)
                                taxa_total += percentual_float
                                
                                administradores_encontrados.append({
                                    'cnpj_cpf': row[7],
                                    'nome': row[8],
                                    'percentual': percentual_float
                                })
                                
                                print(f"DEBUG: Percentual processado: {percentual_float}%")
                            
                        except (ValueError, TypeError) as e:
                            print(f"DEBUG: Erro ao processar percentual '{percentual_raw}': {e}")
                            continue
            
            print(f"DEBUG: Taxa total encontrada: {taxa_total}%")
            print(f"DEBUG: Administradores encontrados: {len(administradores_encontrados)}")
            
            wb.close()
            return taxa_total
            
        except Exception as e:
            print(f"DEBUG: Erro ao obter percentual: {str(e)}")
            if 'wb' in locals():
                wb.close()
            return 0
        
    def verificar_necessidade_recalculo(self, data_referencia, cliente=None):
        """
        VERSÃO CORRIGIDA - Verifica se há necessidade de recálculo usando os métodos unificados
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual
                
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            
            if not os.path.exists(arquivo_cliente):
                return False, "Arquivo do cliente não encontrado"
            
            print(f"DEBUG: Verificando necessidade de recálculo para {cliente} em {data_referencia}")
            
            # Ler dados da planilha
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")
            
            # Converter data
            if isinstance(data_referencia, str):
                data_ref_dt = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref_dt = pd.to_datetime(data_referencia)
            
            # Filtrar para a data específica
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            df_data = df[df['DATA_REL'].dt.date == data_ref_dt.date()].copy()
            
            # Verificar se há taxas existentes
            taxas_existentes = self.identificar_lancamentos_taxa_admin(df_data)
            
            if taxas_existentes.empty:
                return False, "Nenhuma taxa encontrada para esta data"
            
            print(f"DEBUG: {len(taxas_existentes)} taxa(s) encontrada(s)")
            
            # CORREÇÃO: Usar o método unificado para calcular base
            # Primeiro tentar com DataFrame (mais rápido)
            base_atual = self.calcular_base_calculo_taxa(data_referencia, df=df)
            
            print(f"DEBUG: Base atual calculada: R$ {base_atual:.2f}")
            
            # Obter percentual da taxa
            percentual = self.obter_percentual_taxa_cliente(cliente)
            
            if percentual == 0:
                return False, "Percentual de taxa não configurado"
            
            print(f"DEBUG: Percentual de taxa: {percentual}%")
            
            # Calcular valor esperado da taxa
            valor_esperado = base_atual * (percentual / 100)
            print(f"DEBUG: Valor esperado da taxa: R$ {valor_esperado:.2f}")
            
            # Somar valor atual das taxas ATIVAS
            valor_atual_taxas = 0
            taxas_ativas = 0
            
            for _, taxa in taxas_existentes.iterrows():
                status = taxa.get('STATUS', 'ATIVO')
                if status != 'EXCLUIDO':
                    try:
                        valor_taxa = float(str(taxa.get('VALOR', 0)).replace(',', '.'))
                        valor_atual_taxas += valor_taxa
                        taxas_ativas += 1
                        print(f"DEBUG: Taxa ativa ID {taxa.get('ID_LANCAMENTO', 'N/A')}: R$ {valor_taxa:.2f}")
                    except (ValueError, TypeError):
                        print(f"DEBUG: Erro ao processar valor da taxa: {taxa.get('VALOR', 'N/A')}")
                        pass
            
            print(f"DEBUG: Valor atual total das taxas ativas: R$ {valor_atual_taxas:.2f}")
            print(f"DEBUG: Taxas ativas encontradas: {taxas_ativas}")
            
            # Calcular diferença
            diferenca = abs(valor_esperado - valor_atual_taxas)
            tolerancia = 0.01 # R$ 0,01
            
            print(f"DEBUG: Diferença: R$ {diferenca:.2f} (tolerância: R$ {tolerancia:.2f})")
            
            if diferenca > tolerancia:
                mensagem = f"Recálculo necessário - Base: R$ {base_atual:.2f} ({percentual}%) = R$ {valor_esperado:.2f}, Atual: R$ {valor_atual_taxas:.2f}, Diferença: R$ {diferenca:.2f}"
                return True, mensagem
            
            mensagem = f"Taxas consistentes - Base: R$ {base_atual:.2f} ({percentual}%) = R$ {valor_esperado:.2f}"
            return False, mensagem
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro na verificação: {traceback.format_exc()}")
            return False, f"Erro na verificação: {str(e)}"

class GerenciadorLancamentos:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.janela = None
        self.tree_lancamentos = None
        self.dados_originais = []

        self.gestor_taxas = GestorTaxasAdministracao(sistema_principal)
        
    def abrir_gerenciador(self):
        """Abre a janela de gerenciamento de lançamentos"""
        if not self.sistema.cliente_atual:
            custom_messagebox("error", "Erro", "Selecione um cliente primeiro!")
            return
            
        self.janela = tk.Toplevel(self.sistema.root)
        self.janela.title(f"Gerenciar Lançamentos - {self.sistema.cliente_atual}")
        self.janela.geometry("1000x700")
        
        # Configurar janela
        self.janela.transient(self.sistema.root)
        self.janela.grab_set()
        
        self.criar_interface()
        self.carregar_lancamentos()
        
    def criar_interface(self):
        """Cria a interface do gerenciador - VERSÃO COM EXCLUSÃO EM LOTE"""
        #  Frame principal
        main_frame = ttk.Frame(self.janela, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Frame de filtros
        frame_filtros = ttk.LabelFrame(main_frame, text="Filtros")
        frame_filtros.pack(fill='x', pady=(0, 10))
        
        # Filtros por data
        ttk.Label(frame_filtros, text="Data Início:").grid(row=0, column=0, padx=5, pady=5)
        self.data_inicio = DateEntry(frame_filtros, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_inicio.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(frame_filtros, text="Data Fim:").grid(row=0, column=2, padx=5, pady=5)
        self.data_fim = DateEntry(frame_filtros, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_fim.grid(row=0, column=3, padx=5, pady=5)
        
        self.inicializar_datas_padrao()
        
        # Filtro por status
        ttk.Label(frame_filtros, text="Status:").grid(row=0, column=4, padx=5, pady=5)
        self.combo_status = ttk.Combobox(frame_filtros, values=['Todos', 'Ativos', 'Excluídos'], 
                                    state='readonly', width=10)
        self.combo_status.set('Ativos')
        self.combo_status.grid(row=0, column=5, padx=5, pady=5)
        
        # Botão filtrar
        ttk.Button(frame_filtros, text="Filtrar", 
                command=self.aplicar_filtros).grid(row=0, column=6, padx=10, pady=5)
        
        # Frame da lista de lançamentos
        frame_lista = ttk.Frame(main_frame)
        frame_lista.pack(fill='both', expand=True)
        
        # Treeview para lançamentos - MODIFICAÇÃO: Adicionar selectmode para múltipla seleção
        colunas = ('Data', 'Tipo', 'Nome', 'Referência', 'NF', 'Valor', 'Vencimento', 'Status', 'ID')
        self.tree_lancamentos = ttk.Treeview(frame_lista, columns=colunas, show='headings', 
                                        height=20, selectmode='extended')  # CHAVE: selectmode='extended'
        
        # Configurar cabeçalhos (mantém igual)
        for col in colunas:
            self.tree_lancamentos.heading(col, text=col)
            if col == 'ID':
                self.tree_lancamentos.column(col, width=0, stretch=False)
            elif col in ['Data', 'Vencimento']:
                self.tree_lancamentos.column(col, width=60)
            elif col == 'Tipo':
                self.tree_lancamentos.column(col, width=30, anchor='center')
            elif col == 'Valor':
                self.tree_lancamentos.column(col, width=100, anchor='e')
            elif col in ['NF', 'Status']:
                self.tree_lancamentos.column(col, width=70, anchor='center')
            else:
                self.tree_lancamentos.column(col, width=200)
        
        # Scrollbars (mantém igual)
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_lancamentos.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient='horizontal', command=self.tree_lancamentos.xview)
        self.tree_lancamentos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionar elementos
        self.tree_lancamentos.grid(row=0, column=0, sticky='nsew')
        scrolly.grid(row=0, column=1, sticky='ns')
        scrollx.grid(row=1, column=0, sticky='ew')

        # Configurar peso das linhas/colunas para expansão
        frame_lista.grid_rowconfigure(0, weight=1)
        frame_lista.grid_columnconfigure(0, weight=1)

        # Frame de botões - VERSÃO EXPANDIDA
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=(10, 0))
        
        # === SELEÇÃO E INFORMAÇÕES ===
        # Frame para informações de seleção
        frame_selecao = ttk.Frame(frame_botoes)
        frame_selecao.pack(fill='x', pady=(0, 5))
        
        # Label para mostrar quantidade selecionada
        self.label_selecao = ttk.Label(frame_selecao, text="Nenhum item selecionado", 
                                    font=('TkDefaultFont', 9, 'italic'))
        self.label_selecao.pack(side='left')
        
        # Botões de seleção
        ttk.Button(frame_selecao, text="Selecionar Todos Visíveis", 
                command=self.selecionar_todos_visiveis).pack(side='right', padx=2)
        ttk.Button(frame_selecao, text="Limpar Seleção", 
                command=self.limpar_selecao).pack(side='right', padx=2)
        
        # === BOTÕES DE AÇÃO PRINCIPAL ===
        frame_acoes = ttk.Frame(frame_botoes)
        frame_acoes.pack(fill='x', pady=(5, 0))
        
        # Grupo 1: Ações individuais
        ttk.Button(frame_acoes, text="Editar", command=self.editar_lancamento).pack(side='left', padx=5)
        ttk.Button(frame_acoes, text="Ver Histórico", 
                command=self.visualizar_historico_lancamento).pack(side='left', padx=5)
        
        # Separador
        ttk.Separator(frame_acoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # Grupo 2: Ações em lote - NOVOS BOTÕES
        self.btn_excluir_individual = ttk.Button(frame_acoes, text="Excluir", 
                                            command=self.excluir_lancamento)
        self.btn_excluir_individual.pack(side='left', padx=2)
        
        self.btn_excluir_lote = ttk.Button(frame_acoes, text="Excluir Selecionados", 
                                        command=self.excluir_lote, state='disabled')
        self.btn_excluir_lote.pack(side='left', padx=2)
        
        self.btn_restaurar_individual = ttk.Button(frame_acoes, text="Restaurar", 
                                                command=self.restaurar_lancamento)
        self.btn_restaurar_individual.pack(side='left', padx=2)
        
        self.btn_restaurar_lote = ttk.Button(frame_acoes, text="Restaurar Selecionados", 
                                        command=self.restaurar_lote, state='disabled')
        self.btn_restaurar_lote.pack(side='left', padx=2)
        
        # Separador
        ttk.Separator(frame_acoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # Grupo 3: Ações gerais
        ttk.Button(frame_acoes, text="Atualizar", command=self.carregar_lancamentos).pack(side='left', padx=5)
        ttk.Button(frame_acoes, text="Fechar", command=self.janela.destroy).pack(side='right', padx=5)

        # Configurar tags para cores
        self.tree_lancamentos.tag_configure('excluido', background='#ffcccc')
        self.tree_lancamentos.tag_configure('normal', background='white')
        self.tree_lancamentos.tag_configure('selecionado', background='#e6f3ff')  # Nova tag para seleção

        # Configurar eventos
        self.configurar_atalhos()
        self.configurar_eventos_selecao()

    def configurar_eventos_selecao(self):
        """Configura eventos para controle de seleção múltipla"""
        try:
            # Evento quando seleção muda
            def on_selection_change(event=None):
                self.atualizar_interface_selecao()
            
            # Bind no evento de seleção
            self.tree_lancamentos.bind('<<TreeviewSelect>>', on_selection_change)
            
            # Evento de clique com Ctrl para seleção múltipla
            def on_ctrl_click(event):
                # O Treeview já gerencia Ctrl+Click automaticamente com selectmode='extended'
                self.tree_lancamentos.after_idle(self.atualizar_interface_selecao)
            
            self.tree_lancamentos.bind('<Control-Button-1>', on_ctrl_click)
            
            # Evento de clique com Shift para seleção em intervalo
            def on_shift_click(event):
                # O Treeview já gerencia Shift+Click automaticamente com selectmode='extended'
                self.tree_lancamentos.after_idle(self.atualizar_interface_selecao)
            
            self.tree_lancamentos.bind('<Shift-Button-1>', on_shift_click)
            
            print("DEBUG: Eventos de seleção configurados")
            
        except Exception as e:
            print(f"Erro ao configurar eventos de seleção: {str(e)}")

    def atualizar_interface_selecao(self):
        """Atualiza a interface baseada na seleção atual"""
        try:
            items_selecionados = self.tree_lancamentos.selection()
            qtd_selecionados = len(items_selecionados)
            
            # Atualizar label de seleção
            if qtd_selecionados == 0:
                self.label_selecao.config(text="Nenhum item selecionado")
            elif qtd_selecionados == 1:
                self.label_selecao.config(text="1 item selecionado")
            else:
                self.label_selecao.config(text=f"{qtd_selecionados} itens selecionados")
            
            # Controlar estado dos botões
            if qtd_selecionados == 0:
                # Nenhum selecionado - desabilitar todos
                self.btn_excluir_individual.config(state='disabled')
                self.btn_excluir_lote.config(state='disabled')
                self.btn_restaurar_individual.config(state='disabled')
                self.btn_restaurar_lote.config(state='disabled')
                
            elif qtd_selecionados == 1:
                # Um selecionado - habilitar individuais, desabilitar lote
                self.btn_excluir_individual.config(state='normal')
                self.btn_excluir_lote.config(state='disabled')
                self.btn_restaurar_individual.config(state='normal')
                self.btn_restaurar_lote.config(state='disabled')
                
            else:
                # Múltiplos selecionados - desabilitar individuais, habilitar lote
                self.btn_excluir_individual.config(state='disabled')
                self.btn_excluir_lote.config(state='normal')
                self.btn_restaurar_individual.config(state='disabled')
                self.btn_restaurar_lote.config(state='normal')
            
        except Exception as e:
            print(f"Erro ao atualizar interface de seleção: {str(e)}")

    def selecionar_todos_visiveis(self):
        """Seleciona todos os itens visíveis na lista"""
        try:
            # Obter todos os itens filhos visíveis
            items_visiveis = self.tree_lancamentos.get_children()
            
            if not items_visiveis:
                custom_messagebox("info", "Seleção", "Nenhum item visível para selecionar")
                return
            
            # Selecionar todos os itens visíveis
            self.tree_lancamentos.selection_set(items_visiveis)
            
            # Atualizar interface
            self.atualizar_interface_selecao()
            
            print(f"DEBUG: Selecionados {len(items_visiveis)} itens visíveis")
            
        except Exception as e:
            print(f"Erro ao selecionar todos os itens: {str(e)}")
            custom_messagebox("error", "Erro", f"Erro ao selecionar itens: {str(e)}")

    def limpar_selecao(self):
        """Limpa a seleção atual"""
        try:
            self.tree_lancamentos.selection_remove(self.tree_lancamentos.selection())
            self.atualizar_interface_selecao()
            print("DEBUG: Seleção limpa")
            
        except Exception as e:
            print(f"Erro ao limpar seleção: {str(e)}")

    def obter_dados_selecionados(self):
        """Obtém dados dos itens selecionados para processamento"""
        try:
            items_selecionados = self.tree_lancamentos.selection()
            
            if not items_selecionados:
                return []
            
            dados_selecionados = []
            
            for item in items_selecionados:
                valores = self.tree_lancamentos.item(item)['values']
                
                # Extrair informações principais
                dados_item = {
                    'item_id': item,  # ID do item no Treeview
                    'data': valores[0],
                    'tp_desp': valores[1],
                    'nome': valores[2],
                    'referencia': valores[3],
                    'nf': valores[4],
                    'valor': valores[5],
                    'vencimento': valores[6],
                    'status': valores[7],
                    'id_lancamento': valores[8]
                }
                
                dados_selecionados.append(dados_item)
            
            return dados_selecionados
            
        except Exception as e:
            print(f"Erro ao obter dados selecionados: {str(e)}")
            return []

    def excluir_lote(self):
        """Executa exclusão em lote dos itens selecionados"""
        try:
            dados_selecionados = self.obter_dados_selecionados()
            
            if not dados_selecionados:
                custom_messagebox("warning", "Aviso", "Nenhum item selecionado para exclusão")
                return
            
            qtd_selecionados = len(dados_selecionados)
            
            # Verificar se há itens já excluídos
            ja_excluidos = [item for item in dados_selecionados if item['status'] == 'EXCLUIDO']
            ativos = [item for item in dados_selecionados if item['status'] != 'EXCLUIDO']
            
            if not ativos:
                custom_messagebox("info", "Informação", 
                                f"Todos os {qtd_selecionados} itens selecionados já estão excluídos")
                return
            
            # Verificar se há taxas de administração
            taxas_admin = [item for item in ativos if str(item['tp_desp']) == '7']
            qtd_taxas = len(taxas_admin)
            qtd_normais = len(ativos) - qtd_taxas
            
            # Montar mensagem de confirmação detalhada
            mensagem = f"EXCLUSÃO EM LOTE\n\n"
            mensagem += f"📊 RESUMO DA SELEÇÃO:\n"
            mensagem += f"• {len(ativos)} itens serão excluídos\n"
            
            if ja_excluidos:
                mensagem += f"• {len(ja_excluidos)} já estavam excluídos (ignorados)\n"
            
            if qtd_taxas > 0:
                mensagem += f"• ⚠️ {qtd_taxas} TAXA(S) DE ADMINISTRAÇÃO\n"
            
            if qtd_normais > 0:
                mensagem += f"• {qtd_normais} lançamento(s) normal(is)\n"
            
            mensagem += f"\n🎯 PERÍODO AFETADO:\n"
            
            # Obter intervalo de datas
            datas = [item['data'] for item in ativos if item['data']]
            if datas:
                datas_ordenadas = sorted(set(datas))
                if len(datas_ordenadas) == 1:
                    mensagem += f"• {datas_ordenadas[0]}\n"
                else:
                    mensagem += f"• De {datas_ordenadas[0]} até {datas_ordenadas[-1]}\n"
            
            # Calcular valor total
            valor_total = 0
            for item in ativos:
                try:
                    # Limpar formatação do valor
                    valor_str = str(item['valor']).replace('.', '').replace(',', '.')
                    valor_total += float(valor_str)
                except:
                    pass
            
            if valor_total > 0:
                mensagem += f"\n💰 VALOR TOTAL: R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            if qtd_taxas > 0:
                mensagem += f"\n\n⚠️ ATENÇÃO ESPECIAL:\n"
                mensagem += f"Esta operação inclui TAXAS DE ADMINISTRAÇÃO!\n"
                mensagem += f"Verifique se não há duplicação antes de prosseguir.\n"
            
            mensagem += f"\n🔄 As taxas restantes serão verificadas automaticamente.\n"
            mensagem += f"\n❓ Deseja realmente continuar com a exclusão em lote?"
            
            # Confirmar operação
            if not custom_messagebox("yesno", "Confirmação - Exclusão em Lote", mensagem):
                return
            
            # Executar exclusões
            sucessos = 0
            erros = []
            datas_afetadas = set()
            
            progress_window = self.criar_janela_progresso("Excluindo lançamentos...", len(ativos))
            
            try:
                for i, item in enumerate(ativos):
                    try:
                        # Atualizar progresso
                        self.atualizar_progresso(progress_window, i + 1, 
                                            f"Excluindo: {item['nome'][:30]}...")
                        
                        # Executar exclusão
                        self.atualizar_status_lancamento(item['id_lancamento'], 'EXCLUIDO')
                        
                        # Coletar data para verificação posterior
                        if item['data']:
                            try:
                                data_obj = datetime.strptime(item['data'], '%d/%m/%Y').date()
                                datas_afetadas.add(data_obj)
                            except:
                                pass
                        
                        sucessos += 1
                        
                    except Exception as e:
                        erros.append(f"ID {item['id_lancamento']}: {str(e)}")
                        continue
                
            finally:
                progress_window.destroy()
            
            # Verificar recálculo de taxas para as datas afetadas
            self.verificar_recalculo_datas_afetadas(datas_afetadas, "EXCLUSAO")
            
            # Recarregar interface
            self.carregar_lancamentos()
            
            # Mostrar resultado
            mensagem_resultado = f"EXCLUSÃO EM LOTE CONCLUÍDA\n\n"
            mensagem_resultado += f"✅ {sucessos} lançamentos excluídos com sucesso\n"
            
            if erros:
                mensagem_resultado += f"❌ {len(erros)} erros encontrados:\n"
                for erro in erros[:5]:  # Mostrar apenas os primeiros 5 erros
                    mensagem_resultado += f"• {erro}\n"
                if len(erros) > 5:
                    mensagem_resultado += f"• ... e mais {len(erros) - 5} erros\n"
            
            if datas_afetadas:
                mensagem_resultado += f"\n🔄 {len(datas_afetadas)} data(s) verificada(s) para recálculo de taxas"
            
            custom_messagebox("info", "Resultado da Exclusão em Lote", mensagem_resultado)
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro na exclusão em lote: {traceback.format_exc()}")
            custom_messagebox("error", "Erro", f"Erro na exclusão em lote: {str(e)}")

    def restaurar_lote(self):
        """Executa restauração em lote dos itens selecionados"""
        try:
            dados_selecionados = self.obter_dados_selecionados()
            
            if not dados_selecionados:
                custom_messagebox("warning", "Aviso", "Nenhum item selecionado para restauração")
                return
            
            qtd_selecionados = len(dados_selecionados)
            
            # Verificar se há itens já ativos
            ja_ativos = [item for item in dados_selecionados if item['status'] != 'EXCLUIDO']
            excluidos = [item for item in dados_selecionados if item['status'] == 'EXCLUIDO']
            
            if not excluidos:
                custom_messagebox("info", "Informação", 
                                f"Todos os {qtd_selecionados} itens selecionados já estão ativos")
                return
            
            # Verificar se há taxas de administração
            taxas_admin = [item for item in excluidos if str(item['tp_desp']) == '7']
            qtd_taxas = len(taxas_admin)
            qtd_normais = len(excluidos) - qtd_taxas
            
            # Montar mensagem de confirmação
            mensagem = f"RESTAURAÇÃO EM LOTE\n\n"
            mensagem += f"📊 RESUMO DA SELEÇÃO:\n"
            mensagem += f"• {len(excluidos)} itens serão restaurados\n"
            
            if ja_ativos:
                mensagem += f"• {len(ja_ativos)} já estavam ativos (ignorados)\n"
            
            if qtd_taxas > 0:
                mensagem += f"• ⚠️ {qtd_taxas} TAXA(S) DE ADMINISTRAÇÃO\n"
            
            if qtd_normais > 0:
                mensagem += f"• {qtd_normais} lançamento(s) normal(is)\n"
            
            if qtd_taxas > 0:
                mensagem += f"\n⚠️ ATENÇÃO ESPECIAL:\n"
                mensagem += f"Esta operação inclui TAXAS DE ADMINISTRAÇÃO!\n"
                mensagem += f"Verifique se não haverá duplicação.\n"
            
            mensagem += f"\n🔄 As taxas serão verificadas automaticamente.\n"
            mensagem += f"\n❓ Deseja realmente continuar com a restauração em lote?"
            
            # Confirmar operação
            if not custom_messagebox("yesno", "Confirmação - Restauração em Lote", mensagem):
                return
            
            # Executar restaurações
            sucessos = 0
            erros = []
            datas_afetadas = set()
            
            progress_window = self.criar_janela_progresso("Restaurando lançamentos...", len(excluidos))
            
            try:
                for i, item in enumerate(excluidos):
                    try:
                        # Atualizar progresso
                        self.atualizar_progresso(progress_window, i + 1, 
                                            f"Restaurando: {item['nome'][:30]}...")
                        
                        # Executar restauração
                        self.atualizar_status_lancamento(item['id_lancamento'], 'ATIVO')
                        
                        # Coletar data para verificação posterior
                        if item['data']:
                            try:
                                data_obj = datetime.strptime(item['data'], '%d/%m/%Y').date()
                                datas_afetadas.add(data_obj)
                            except:
                                pass
                        
                        sucessos += 1
                        
                    except Exception as e:
                        erros.append(f"ID {item['id_lancamento']}: {str(e)}")
                        continue
                
            finally:
                progress_window.destroy()
            
            # Verificar recálculo de taxas para as datas afetadas
            self.verificar_recalculo_datas_afetadas(datas_afetadas, "ALTERACAO")
            
            # Recarregar interface
            self.carregar_lancamentos()
            
            # Mostrar resultado
            mensagem_resultado = f"RESTAURAÇÃO EM LOTE CONCLUÍDA\n\n"
            mensagem_resultado += f"✅ {sucessos} lançamentos restaurados com sucesso\n"
            
            if erros:
                mensagem_resultado += f"❌ {len(erros)} erros encontrados:\n"
                for erro in erros[:5]:
                    mensagem_resultado += f"• {erro}\n"
                if len(erros) > 5:
                    mensagem_resultado += f"• ... e mais {len(erros) - 5} erros\n"
            
            if datas_afetadas:
                mensagem_resultado += f"\n🔄 {len(datas_afetadas)} data(s) verificada(s) para recálculo de taxas"
            
            custom_messagebox("info", "Resultado da Restauração em Lote", mensagem_resultado)
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro na restauração em lote: {traceback.format_exc()}")
            custom_messagebox("error", "Erro", f"Erro na restauração em lote: {str(e)}")

    # def verificar_recalculo_datas_afetadas(self, datas_afetadas, tipo_operacao):
    #     """Verifica recálculo de taxas para múltiplas datas afetadas"""
    #     try:
    #         if not datas_afetadas:
    #             return
            
    #         print(f"DEBUG: Verificando recálculo para {len(datas_afetadas)} datas afetadas")
            
    #         # Aguardar um pouco para garantir que as operações foram salvas
    #         import time
    #         time.sleep(0.5)
            
    #         resultados = []
            
    #         for data_afetada in sorted(datas_afetadas):
    #             try:
    #                 print(f"DEBUG: Verificando recálculo para {data_afetada}")
                    
    #                 # Usar o método unificado do sistema
    #                 resultado = self.sistema.chamar_apos_operacao_lancamento(data_afetada, tipo_operacao)
                    
    #                 resultados.append({
    #                     'data': data_afetada,
    #                     'resultado': resultado
    #                 })
                    
    #                 if resultado["sucesso"]:
    #                     print(f"✅ Verificação para {data_afetada}: {resultado['mensagem']}")
    #                 else:
    #                     print(f"⚠️ Problema na verificação para {data_afetada}: {resultado['mensagem']}")
                        
    #             except Exception as e:
    #                 print(f"❌ Erro ao verificar {data_afetada}: {str(e)}")
    #                 resultados.append({
    #                     'data': data_afetada,
    #                     'resultado': {"sucesso": False, "mensagem": f"Erro: {str(e)}"}
    #                 })
    #                 continue
            
    #         # Log consolidado
    #         verificacoes_ok = sum(1 for r in resultados if r['resultado']['sucesso'])
    #         print(f"DEBUG: Verificações concluídas: {verificacoes_ok}/{len(resultados)} OK")
            
    #     except Exception as e:
    #         print(f"DEBUG: Erro geral na verificação de múltiplas datas: {str(e)}")

    def criar_janela_progresso(self, titulo, total_items):
        """Cria janela de progresso para operações em lote"""
        try:
            janela_progress = tk.Toplevel(self.janela)
            janela_progress.title(titulo)
            janela_progress.geometry("400x120")
            janela_progress.transient(self.janela)
            janela_progress.grab_set()
            
            # Centralizar janela
            janela_progress.update_idletasks()
            x = (janela_progress.winfo_screenwidth() // 2) - (400 // 2)
            y = (janela_progress.winfo_screenheight() // 2) - (120 // 2)
            janela_progress.geometry(f"400x120+{x}+{y}")
            
            frame = ttk.Frame(janela_progress, padding="20")
            frame.pack(fill='both', expand=True)
            
            # Label de status
            label_status = ttk.Label(frame, text="Preparando...", font=('TkDefaultFont', 10))
            label_status.pack(pady=(0, 10))
            
            # Barra de progresso
            progress_var = tk.DoubleVar()
            progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=total_items, 
                                        mode='determinate', length=350)
            progress_bar.pack(pady=(0, 10))
            
            # Label de contagem
            label_count = ttk.Label(frame, text=f"0 / {total_items}", font=('TkDefaultFont', 9))
            label_count.pack()
            
            # Armazenar referências para atualização
            janela_progress.label_status = label_status
            janela_progress.progress_var = progress_var
            janela_progress.label_count = label_count
            janela_progress.total_items = total_items
            
            # Forçar atualização da interface
            janela_progress.update()
            
            return janela_progress
            
        except Exception as e:
            print(f"Erro ao criar janela de progresso: {str(e)}")
            return None

    def atualizar_progresso(self, janela_progress, item_atual, mensagem=""):
        """Atualiza a janela de progresso"""
        try:
            if not janela_progress:
                return
            
            # Atualizar barra de progresso
            janela_progress.progress_var.set(item_atual)
            
            # Atualizar status
            if mensagem:
                janela_progress.label_status.config(text=mensagem)
            
            # Atualizar contagem
            janela_progress.label_count.config(
                text=f"{item_atual} / {janela_progress.total_items}"
            )
            
            # Forçar atualização da interface
            janela_progress.update_idletasks()
            
        except Exception as e:
            print(f"Erro ao atualizar progresso: {str(e)}")
    
    def configurar_atalhos(self):
        """Configura atalhos de teclado - VERSÃO EXPANDIDA"""
        try:
            # ATALHO ORIGINAL: Duplo clique para histórico
            def on_double_click(event):
                if self.tree_lancamentos.selection():
                    self.visualizar_historico_lancamento()
            
            self.tree_lancamentos.bind('<Double-1>', on_double_click)
            
            # ATALHO ORIGINAL: Tecla H para histórico
            def on_key_h(event):
                if self.tree_lancamentos.selection():
                    self.visualizar_historico_lancamento()
                else:
                    custom_messagebox("info", "Atalho H", "Selecione um lançamento primeiro para ver o histórico")
            
            self.janela.bind('<Key-h>', on_key_h)
            self.janela.bind('<Key-H>', on_key_h)
            self.tree_lancamentos.bind('<Key-h>', on_key_h)
            self.tree_lancamentos.bind('<Key-H>', on_key_h)
            
            # NOVOS ATALHOS PARA SELEÇÃO EM LOTE
            
            # Ctrl+A: Selecionar todos os itens visíveis
            def on_ctrl_a(event):
                self.selecionar_todos_visiveis()
                return "break"  # Impede comportamento padrão
            
            self.janela.bind('<Control-a>', on_ctrl_a)
            self.tree_lancamentos.bind('<Control-a>', on_ctrl_a)
            
            # Ctrl+D: Limpar seleção
            def on_ctrl_d(event):
                self.limpar_selecao()
                return "break"
            
            self.janela.bind('<Control-d>', on_ctrl_d)
            self.tree_lancamentos.bind('<Control-d>', on_ctrl_d)
            
            # Delete: Excluir selecionados
            def on_delete(event):
                items_selecionados = self.tree_lancamentos.selection()
                if len(items_selecionados) == 1:
                    self.excluir_lancamento()
                elif len(items_selecionados) > 1:
                    self.excluir_lote()
                else:
                    custom_messagebox("info", "Atalho Delete", "Selecione um ou mais lançamentos para excluir")
                return "break"
            
            self.janela.bind('<Delete>', on_delete)
            self.tree_lancamentos.bind('<Delete>', on_delete)
            
            # Ctrl+R: Restaurar selecionados
            def on_ctrl_r(event):
                items_selecionados = self.tree_lancamentos.selection()
                if len(items_selecionados) == 1:
                    self.restaurar_lancamento()
                elif len(items_selecionados) > 1:
                    self.restaurar_lote()
                else:
                    custom_messagebox("info", "Atalho Ctrl+R", "Selecione um ou mais lançamentos para restaurar")
                return "break"
            
            self.janela.bind('<Control-r>', on_ctrl_r)
            self.tree_lancamentos.bind('<Control-r>', on_ctrl_r)
            
            # F5: Atualizar lista
            def on_f5(event):
                self.carregar_lancamentos()
                return "break"
            
            self.janela.bind('<F5>', on_f5)
            self.tree_lancamentos.bind('<F5>', on_f5)
            
            # Escape: Limpar seleção
            def on_escape(event):
                self.limpar_selecao()
                return "break"
            
            self.janela.bind('<Escape>', on_escape)
            self.tree_lancamentos.bind('<Escape>', on_escape)
            
            # Tornar a janela focável para receber eventos de teclado
            self.janela.focus_set()
            
            print("DEBUG: Atalhos configurados (incluindo seleção em lote)")
            print("       Ctrl+A: Selecionar todos visíveis")
            print("       Ctrl+D: Limpar seleção")
            print("       Delete: Excluir selecionados")
            print("       Ctrl+R: Restaurar selecionados")
            print("       F5: Atualizar")
            print("       Escape: Limpar seleção")
            
        except Exception as e:
            print(f"Erro ao configurar atalhos: {str(e)}")


    def formatar_tipo_despesa(self, tp_desp):
        """Formata tipo de despesa como inteiro"""
        try:
            if pd.isna(tp_desp) or tp_desp == "":
                return ""
            
            # Converter para float primeiro, depois para int para remover decimais
            valor_numerico = float(tp_desp)
            return str(int(valor_numerico))
            
        except (ValueError, TypeError):
            # Se não conseguir converter, retornar como string
            return str(tp_desp)

    def inicializar_datas_padrao(self):
        """Inicializa as datas padrão dos filtros baseado no sistema (dias 5 e 20)"""
        data_inicio_padrao = None
        data_fim_padrao = None

        try:
            from datetime import datetime, timedelta
            from calendar import monthrange
            
            # Data de hoje
            hoje = datetime.now().date()
            dia_atual = hoje.day
            mes_atual = hoje.month
            ano_atual = hoje.year
            
            # LÓGICA DO SISTEMA: Data fim baseada nos dias 5 e 20
            if dia_atual <= 5:
                # Do dia 1 ao 5: data fim = dia 5 do mês atual
                data_fim_padrao = hoje.replace(day=5)
            elif dia_atual <= 20:
                # Do dia 6 ao 20: data fim = dia 20 do mês atual
                data_fim_padrao = hoje.replace(day=20)
            else:
                # Do dia 21 em diante: data fim = dia 5 do próximo mês
                if mes_atual == 12:
                    # Se dezembro, vai para janeiro do próximo ano
                    data_fim_padrao = datetime(ano_atual + 1, 1, 5).date()
                else:
                    # Senão, próximo mês do mesmo ano
                    data_fim_padrao = datetime(ano_atual, mes_atual + 1, 5).date()
            
            # Data de início: 30 dias antes da data fim (mais lógico para o sistema)
            data_inicio_padrao = data_fim_padrao - timedelta(days=30)

            # Verificar se os widgets existem antes de definir as datas
            if hasattr(self, 'data_inicio') and self.data_inicio and data_inicio_padrao:
                self.data_inicio.set_date(data_inicio_padrao)
            
            if hasattr(self, 'data_fim') and self.data_fim and data_fim_padrao:
                self.data_fim.set_date(data_fim_padrao)
            
            print(f"DEBUG: Datas padrão definidas - Início: {data_inicio_padrao}, Fim: {data_fim_padrao}")
            
            # Definir as datas nos controles
            self.data_inicio.set_date(data_inicio_padrao)
            self.data_fim.set_date(data_fim_padrao)
            
            print(f"DEBUG: Datas padrão definidas (sistema dias 5/20):")
            print(f"       Hoje: {hoje} (dia {dia_atual})")
            print(f"       Data início: {data_inicio_padrao}")
            print(f"       Data fim: {data_fim_padrao}")
            
        except Exception as e:
            print(f"DEBUG: Erro ao inicializar datas padrão: {str(e)}")
            import traceback
            traceback.print_exc()
       
    def carregar_lancamentos(self):
        """Carrega os lançamentos da planilha com correção de IDs duplicados"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            
            # Carregar dados
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")
            
            # Adicionar coluna de status se não existir
            if 'STATUS' not in df.columns:
                df['STATUS'] = 'ATIVO'
            
            # CORREÇÃO: Preencher status em branco com 'ATIVO'
            df['STATUS'] = df['STATUS'].replace('', 'ATIVO')
            df['STATUS'] = df['STATUS'].fillna('ATIVO')
            
            # CORREÇÃO PRINCIPAL: Gerenciar IDs de forma mais robusta
            if 'ID_LANCAMENTO' not in df.columns:
                # Se não existe coluna ID, criar sequencialmente
                df['ID_LANCAMENTO'] = range(1, len(df) + 1)
                print("DEBUG: Criada coluna ID_LANCAMENTO sequencial")
            else:
                print("DEBUG: Verificando e corrigindo IDs duplicados/inválidos")
                
                # PASSO 1: Converter todos os IDs para numérico, transformando inválidos em NaN
                df['ID_LANCAMENTO'] = pd.to_numeric(df['ID_LANCAMENTO'], errors='coerce')
                
                # PASSO 2: Identificar IDs duplicados
                ids_duplicados = df[df.duplicated(subset=['ID_LANCAMENTO'], keep=False) & df['ID_LANCAMENTO'].notna()]
                if not ids_duplicados.empty:
                    print(f"DEBUG: Encontrados {len(ids_duplicados)} lançamentos com IDs duplicados")
                    
                    # Mostrar quais IDs estão duplicados
                    ids_problema = ids_duplicados['ID_LANCAMENTO'].unique()
                    for id_dup in ids_problema:
                        linhas_dup = ids_duplicados[ids_duplicados['ID_LANCAMENTO'] == id_dup].index.tolist()
                        print(f"DEBUG: ID {id_dup} duplicado nas linhas: {[i+2 for i in linhas_dup]}")  # +2 para contar header
                
                # PASSO 3: Encontrar o próximo ID disponível
                ids_validos = df['ID_LANCAMENTO'].dropna()
                if len(ids_validos) > 0:
                    proximo_id = int(ids_validos.max()) + 1
                else:
                    proximo_id = 1
                
                print(f"DEBUG: Próximo ID disponível: {proximo_id}")
                
                # PASSO 4: Corrigir IDs inválidos (NaN) primeiro
                mask_nan = df['ID_LANCAMENTO'].isna()
                indices_nan = df.index[mask_nan].tolist()
                
                for idx in indices_nan:
                    df.loc[idx, 'ID_LANCAMENTO'] = proximo_id
                    print(f"DEBUG: Atribuído ID {proximo_id} para linha {idx+2} (era NaN)")
                    proximo_id += 1
                
                # PASSO 5: Corrigir IDs duplicados
                # Primeiro, identificar novamente após correção dos NaN
                duplicados_restantes = df[df.duplicated(subset=['ID_LANCAMENTO'], keep='first')]
                
                for idx in duplicados_restantes.index:
                    id_original = df.loc[idx, 'ID_LANCAMENTO']
                    df.loc[idx, 'ID_LANCAMENTO'] = proximo_id
                    
                    # Informações do lançamento para debug
                    nome = df.loc[idx, 'NOME'] if 'NOME' in df.columns else 'N/A'
                    referencia = df.loc[idx, 'REFERÊNCIA'] if 'REFERÊNCIA' in df.columns else 'N/A'
                    nf = df.loc[idx, 'NF'] if 'NF' in df.columns else 'N/A'
                    
                    print(f"DEBUG: Corrigido ID duplicado {id_original} → {proximo_id} para linha {idx+2}")
                    print(f"       Lançamento: {nome} - {referencia}")
                    proximo_id += 1
                
                # PASSO 6: Verificação final
                ids_finais = df['ID_LANCAMENTO']
                duplicados_finais = df[df.duplicated(subset=['ID_LANCAMENTO'], keep=False)]
                
                if not duplicados_finais.empty:
                    print(f"DEBUG: ERRO - Ainda existem {len(duplicados_finais)} IDs duplicados após correção!")
                    # Se ainda há duplicados, forçar sequência completa
                    df['ID_LANCAMENTO'] = range(1, len(df) + 1)
                    print("DEBUG: Forçada sequência completa de IDs")
                else:
                    print("DEBUG: Todos os IDs estão únicos agora")
            
            # Converter para int (agora que não há mais NaN nem duplicados)
            df['ID_LANCAMENTO'] = df['ID_LANCAMENTO'].astype(int)
            
            # CORREÇÃO: Salvar a planilha se houve mudanças nos IDs
            ids_mudaram = not df['ID_LANCAMENTO'].equals(pd.read_excel(arquivo_cliente, sheet_name='Dados')['ID_LANCAMENTO']) if 'ID_LANCAMENTO' in pd.read_excel(arquivo_cliente, sheet_name='Dados').columns else True
            
            if ids_mudaram:
                print("DEBUG: Salvando correções de ID na planilha")
                self.salvar_correcoes_ids(arquivo_cliente, df)
            
            # Salvar dados originais
            self.dados_originais = df.copy()
            
            # Corrigir planilha se necessário (status)
            self.corrigir_planilha_status(arquivo_cliente, df)
            
            # Limpar tree
            for item in self.tree_lancamentos.get_children():
                self.tree_lancamentos.delete(item)
            
            print(f"DEBUG: Tree limpo, iniciando inserção de {len(df)} lançamentos")
            
            # Preencher tree
            items_inseridos = 0
            for idx, row in df.iterrows():
                status = row.get('STATUS', 'ATIVO')
                if status == '' or pd.isna(status):
                    status = 'ATIVO'
                    
                tag = 'excluido' if status == 'EXCLUIDO' else 'normal'
                
                # Formatar valores
                data_rel = self.formatar_data(row['DATA_REL'])
                data_vencto = self.formatar_data(row['DT_VENCTO'])
                valor = self.formatar_valor(row['VALOR'])
                
                # CORREÇÃO: Formatar TP_DESP como inteiro
                tp_desp = self.formatar_tipo_despesa(row['TP_DESP'])
                
                # CORREÇÃO: Garantir que o ID seja um inteiro único
                id_lancamento = int(row['ID_LANCAMENTO'])
                
                valores_tree = (data_rel, tp_desp, row['NOME'], 
                            row['REFERÊNCIA'], row['NF'], valor, data_vencto, status, id_lancamento)
                
                # DEBUG: Mostrar alguns lançamentos inseridos
                if items_inseridos < 3:
                    print(f"DEBUG: Inserindo item {items_inseridos + 1}: {valores_tree}")
                
                self.tree_lancamentos.insert('', 'end', 
                    values=valores_tree,
                    tags=(tag,))
                
                items_inseridos += 1
            
            print(f"DEBUG: {items_inseridos} itens inseridos no tree")
            
            # Verificar quantos itens estão no tree antes dos filtros
            itens_antes_filtro = len(self.tree_lancamentos.get_children())
            print(f"DEBUG: Itens no tree ANTES do filtro: {itens_antes_filtro}")
            
            # if hasattr(self, 'data_inicio') and self.data_inicio:
            #     self.data_inicio.set_date(data_inicio_padrao)

            # Aplicar filtros
            self.aplicar_filtros()
            
            # Verificar quantos itens estão no tree depois dos filtros
            itens_depois_filtro = len(self.tree_lancamentos.get_children())
            print(f"DEBUG: Itens no tree DEPOIS do filtro: {itens_depois_filtro}")
            
            print(f"DEBUG: Carregamento concluído. Total de lançamentos: {len(df)}")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao carregar lançamentos: {str(e)}")

    def salvar_correcoes_ids(self, arquivo_cliente, df_corrigido):
        """
        Salva as correções de ID na planilha Excel
        """
        try:
            from openpyxl import load_workbook
            
            # Carregar workbook existente
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            # Encontrar coluna ID_LANCAMENTO
            id_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'ID_LANCAMENTO':
                    id_col = col
                    break
            
            if id_col is None:
                print("DEBUG: Coluna ID_LANCAMENTO não encontrada na planilha")
                return
            
            # Atualizar IDs na planilha
            for idx, row in df_corrigido.iterrows():
                excel_row = idx + 2  # +2 porque pandas usa índice 0 e Excel começa na linha 1, plus header
                id_value = int(row['ID_LANCAMENTO'])
                ws.cell(row=excel_row, column=id_col).value = id_value
            
            # Salvar workbook
            wb.save(arquivo_cliente)
            print("DEBUG: IDs corrigidos salvos na planilha")
            
        except Exception as e:
            print(f"DEBUG: Erro ao salvar correções de ID: {str(e)}")


    def verificar_integridade_ids(self, arquivo_cliente):
        """
        Método auxiliar para verificar a integridade dos IDs na planilha
        """
        try:
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            
            if 'ID_LANCAMENTO' not in df.columns:
                return False, "Coluna ID_LANCAMENTO não existe"
            
            # Converter para numérico
            ids = pd.to_numeric(df['ID_LANCAMENTO'], errors='coerce')
            
            # Verificar NaN
            ids_nan = ids.isna().sum()
            if ids_nan > 0:
                return False, f"{ids_nan} IDs inválidos encontrados"
            
            # Verificar duplicados
            ids_validos = ids.dropna()
            duplicados = ids_validos.duplicated().sum()
            if duplicados > 0:
                return False, f"{duplicados} IDs duplicados encontrados"
            
            # Verificar sequência
            ids_ordenados = sorted(ids_validos.unique())
            esperado = list(range(1, len(ids_validos) + 1))
            
            if len(ids_ordenados) != len(esperado):
                return False, f"Quantidade de IDs únicos ({len(ids_ordenados)}) não confere com total de linhas ({len(esperado)})"
            
            return True, "IDs íntegros"
            
        except Exception as e:
            return False, f"Erro na verificação: {str(e)}"


    def debug_ids_duplicados(self, arquivo_cliente):
        """
        Método para debug específico de IDs duplicados
        """
        try:
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            
            if 'ID_LANCAMENTO' not in df.columns:
                print("DEBUG: Coluna ID_LANCAMENTO não existe")
                return
            
            print("=== DEBUG IDs DUPLICADOS ===")
            
            # Converter para numérico
            df['ID_LANCAMENTO'] = pd.to_numeric(df['ID_LANCAMENTO'], errors='coerce')
            
            # Encontrar duplicados
            duplicados = df[df.duplicated(subset=['ID_LANCAMENTO'], keep=False) & df['ID_LANCAMENTO'].notna()]
            
            if duplicados.empty:
                print("Nenhum ID duplicado encontrado")
                return
            
            print(f"Encontrados {len(duplicados)} lançamentos com IDs duplicados:")
            
            for id_dup in duplicados['ID_LANCAMENTO'].unique():
                print(f"\nID {id_dup}:")
                linhas_dup = duplicados[duplicados['ID_LANCAMENTO'] == id_dup]
                
                for idx, row in linhas_dup.iterrows():
                    linha_excel = idx + 2
                    nome = row.get('NOME', 'N/A')
                    referencia = row.get('REFERÊNCIA', 'N/A')
                    valor = row.get('VALOR', 'N/A')
                    observacao = row.get('OBSERVAÇÃO', '')
                    
                    tipo_lanc = "TAXA ADM" if 'TAXA ADM' in str(observacao) else "NORMAL"
                    print(f"  Linha {linha_excel}: {nome} - {referencia} | R$ {valor} | {tipo_lanc}")
            
        except Exception as e:
            print(f"Erro no debug: {str(e)}")

    def corrigir_planilha_status(self, arquivo_cliente, df):
        """Corrige o status na planilha para dados antigos"""
        try:
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            # Adicionar cabeçalho STATUS se não existir
            if ws.cell(row=1, column=14).value != 'STATUS':
                ws.cell(row=1, column=14, value='STATUS')
            
            # Adicionar cabeçalho ID_LANCAMENTO se não existir
            if ws.cell(row=1, column=15).value != 'ID_LANCAMENTO':
                ws.cell(row=1, column=15, value='ID_LANCAMENTO')
            
            # Adicionar cabeçalho HISTORICO_ALTERACAO se não existir
            if ws.cell(row=1, column=16).value != 'HISTORICO_ALTERACAO':
                ws.cell(row=1, column=16, value='HISTORICO_ALTERACAO')
            
            # Preencher STATUS e ID para todas as linhas
            for idx, (df_idx, row) in enumerate(df.iterrows(), start=2):
                # STATUS
                status_atual = ws.cell(row=idx, column=14).value
                if not status_atual or status_atual == '':
                    ws.cell(row=idx, column=14, value='ATIVO')
                
                # ID_LANCAMENTO - CORREÇÃO: usar o ID do DataFrame
                ws.cell(row=idx, column=15, value=int(row['ID_LANCAMENTO']))
            
            wb.save(arquivo_cliente)
                        
        except Exception as e:
            print(f"DEBUG: Erro ao corrigir status na planilha: {str(e)}")
            # Não levantar exceção aqui para não interromper o carregamento
    
    def aplicar_filtros(self):
        """Aplica os filtros selecionados"""
        try:
            # Obter filtros
            status_filtro = self.combo_status.get()
            data_inicio = self.data_inicio.get_date()
            data_fim = self.data_fim.get_date()
            
            print(f"DEBUG: Aplicando filtros - Status: {status_filtro}, Data início: {data_inicio}, Data fim: {data_fim}")
            
            itens_visiveis = 0
            itens_ocultos = 0
            
            # Filtrar itens na tree
            for item in self.tree_lancamentos.get_children():
                valores = self.tree_lancamentos.item(item, 'values')
                
                # CORREÇÃO: Verificar os índices corretos das colunas
                # Colunas: ('Data', 'Tipo', 'Nome', 'Referência', 'NF', 'Valor', 'Vencimento', 'Status', 'ID')
                #           0       1       2       3            4      5        6             7         8
                
                data_rel_str = valores[0]  # Data (índice 0)
                status_item = valores[7]   # Status (índice 7, não 6!)
                
                mostrar = True
                
                # Filtro por status
                if status_filtro == 'Ativos' and status_item != 'ATIVO':
                    mostrar = False
                elif status_filtro == 'Excluídos' and status_item != 'EXCLUIDO':
                    mostrar = False
                
                # Filtro por data
                if mostrar and data_rel_str and data_rel_str.strip():
                    try:
                        # Converter data da string para datetime.date
                        data_rel = datetime.strptime(data_rel_str, '%d/%m/%Y').date()
                        
                        # Verificar se está no intervalo
                        if data_rel < data_inicio or data_rel > data_fim:
                            mostrar = False
                            
                    except Exception as e:
                        print(f"DEBUG: Erro ao processar data '{data_rel_str}': {str(e)}")
                        # Em caso de erro na data, não filtrar por data para este item
                
                # Mostrar/ocultar item
                if mostrar:
                    # Verificar se o item já está visível
                    try:
                        self.tree_lancamentos.item(item)  # Testa se está visível
                        itens_visiveis += 1
                    except:
                        # Se não está visível, reattach
                        self.tree_lancamentos.reattach(item, '', tk.END)
                        itens_visiveis += 1
                else:
                    # Ocultar item
                    self.tree_lancamentos.detach(item)
                    itens_ocultos += 1
            
            print(f"DEBUG: Filtros aplicados - {itens_visiveis} visíveis, {itens_ocultos} ocultos")
                        
        except Exception as e:
            print(f"DEBUG: Erro ao aplicar filtros: {str(e)}")
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao aplicar filtros: {str(e)}")

    def aplicar_filtros_melhorados(self):
        """Versão melhorada do filtro que preserva seleção quando possível"""
        try:
            # Salvar seleção atual
            selecao_anterior = []
            for item in self.tree_lancamentos.selection():
                valores = self.tree_lancamentos.item(item, 'values')
                if len(valores) >= 9:
                    selecao_anterior.append(valores[8])  # ID do lançamento
            
            # Aplicar filtros (método original)
            self.aplicar_filtros()
            
            # Tentar restaurar seleção
            if selecao_anterior:
                self.restaurar_selecao_por_ids(selecao_anterior)
            
            # Atualizar interface
            self.atualizar_interface_selecao()
            
        except Exception as e:
            print(f"Erro ao aplicar filtros melhorados: {str(e)}")
            # Fallback para método original
            self.aplicar_filtros()

    def restaurar_selecao_por_ids(self, ids_para_selecionar):
        """Restaura seleção baseada nos IDs dos lançamentos"""
        try:
            items_para_selecionar = []
            
            # Buscar itens que correspondem aos IDs
            for item in self.tree_lancamentos.get_children():
                valores = self.tree_lancamentos.item(item, 'values')
                if len(valores) >= 9:
                    id_item = valores[8]
                    if id_item in ids_para_selecionar:
                        items_para_selecionar.append(item)
            
            # Selecionar itens encontrados
            if items_para_selecionar:
                self.tree_lancamentos.selection_set(items_para_selecionar)
                print(f"DEBUG: Restaurada seleção de {len(items_para_selecionar)} itens")
            
        except Exception as e:
            print(f"Erro ao restaurar seleção: {str(e)}")

    def editar_lancamento(self):
        """Abre editor para o lançamento selecionado - VERSÃO ATUALIZADA"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para editar")
            return
        
        try:
            valores = self.tree_lancamentos.item(item_selecionado[0])['values']
            
            # Verificar se temos valores suficientes
            if len(valores) < 9:
                custom_messagebox("error", "Erro", "Dados insuficientes no lançamento selecionado")
                return
                
            id_lancamento = valores[8]  # ID do lançamento (9ª coluna, índice 8)
            
            # Verificar se o ID é válido
            if not id_lancamento or pd.isna(id_lancamento):
                custom_messagebox("error", "Erro", "ID do lançamento não encontrado")
                return
            
            # Buscar dados completos
            try:
                # Converter ID para o tipo correto se necessário
                if isinstance(id_lancamento, str):
                    try:
                        id_lancamento = int(float(id_lancamento))
                    except ValueError:
                        custom_messagebox("error", "Erro", f"ID inválido: {id_lancamento}")
                        return
                
                # Buscar no DataFrame
                lancamento = None
                mask = self.dados_originais['ID_LANCAMENTO'] == id_lancamento
                dados_filtrados = self.dados_originais[mask]
                
                if not dados_filtrados.empty:
                    lancamento = dados_filtrados.iloc[0]
                else:
                    # Busca alternativa se não encontrou
                    for idx, row in self.dados_originais.iterrows():
                        row_id = row.get('ID_LANCAMENTO')
                        if pd.notna(row_id):
                            try:
                                if int(float(row_id)) == int(float(id_lancamento)):
                                    lancamento = row
                                    break
                            except (ValueError, TypeError):
                                continue
                
                if lancamento is None:
                    custom_messagebox("error", "Erro", 
                        f"Lançamento com ID {id_lancamento} não encontrado nos dados carregados")
                    return
                
                # CORREÇÃO: Abrir editor com callback correto
                editor = EditorLancamentoCompleto(self.janela, lancamento, self.salvar_edicao)
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                custom_messagebox("error", "Erro", f"Erro ao abrir editor: {str(e)}")
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao editar lançamento: {str(e)}")
    
    def excluir_lancamento(self):
        """
        Versão corrigida que usa o novo sistema de verificação após exclusão
        """
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para excluir")
            return
        
        valores = self.tree_lancamentos.item(item_selecionado[0])['values']
        tp_desp = valores[1]         # Tipo de despesa
        nome_lancamento = valores[2]  # Nome
        referencia = valores[3]      # Referência
        valor = valores[5]           # Valor
        data_lancamento = valores[0] # Data
        status_atual = valores[7]    # Status
        id_lancamento = valores[8]   # ID
        
        # Verificar se já está excluído
        if status_atual == 'EXCLUIDO':
            custom_messagebox("info", "Informação", "Este lançamento já está excluído")
            return
        
        # Verificar se é uma taxa de administração
        eh_taxa = (tp_desp == 7)
        
        if eh_taxa:
            if not custom_messagebox("yesno", "Confirmação - Taxa de Administração", 
                                f"Você está excluindo uma TAXA DE ADMINISTRAÇÃO:\n\n"
                                f"📋 {referencia}\n"
                                f"💰 {valor}\n"
                                f"📅 {data_lancamento}\n\n"
                                f"⚠️ ATENÇÃO: A exclusão de uma taxa pode afetar os cálculos!\n\n"
                                f"Deseja realmente continuar?"):
                return
        else:
            if not custom_messagebox("yesno", "Confirmação", 
                                f"Deseja realmente excluir este lançamento?\n\n"
                                f"👤 {nome_lancamento}\n"
                                f"📋 {referencia}\n"
                                f"💰 {valor}\n"
                                f"📅 {data_lancamento}\n\n"
                                f"🔄 As taxas de administração do tipo % deverão ser verificadas."):
                return
        
        try:
            id_lancamento = valores[8]
            
            print(f"DEBUG: Excluindo lançamento ID {id_lancamento}")
            
            # Atualizar status para EXCLUIDO
            self.atualizar_status_lancamento(id_lancamento, 'EXCLUIDO')
            print(f"DEBUG: Status atualizado para EXCLUIDO")

            # ===== NOVA INTEGRAÇÃO: Usar método unificado =====
            print(f"DEBUG: Iniciando verificação de recálculo para data {data_lancamento}")
            data_para_recalculo = datetime.strptime(data_lancamento, '%d/%m/%Y').date()
            
            # Aguardar um pouco para garantir que a exclusão foi salva
            import time
            time.sleep(0.5)
            
            # CORREÇÃO: Usar o novo método de verificação
            # resultado_verificacao = self.sistema.chamar_apos_operacao_lancamento(data_para_recalculo, "EXCLUSAO")
            
            # Recarregar lista
            self.carregar_lancamentos()
            
            # Mostrar resultado
            if eh_taxa:
                mensagem = "Taxa de administração excluída com sucesso!"
                mensagem += "\n\n⚠️ IMPORTANTE: Verifique se há outras taxas para esta data que precisam de ajuste."
            else:
                mensagem = "Lançamento excluído com sucesso!"
                
                # if resultado_verificacao["sucesso"]:
                #     # Verificar se houve recálculo automático baseado no resultado
                #     if ("recalculadas" in resultado_verificacao["mensagem"] or 
                #         "Recálculo Concluído" in str(resultado_verificacao)):
                #         mensagem += f"\n\n✅ Taxas foram recalculadas automaticamente!"
                #     elif "corretas" in resultado_verificacao["mensagem"]:
                #         mensagem += f"\n\n✅ Taxas verificadas - estão corretas"
                #     elif "cancelado" in resultado_verificacao["mensagem"]:
                #         mensagem += f"\n\n⚠️ Recálculo foi oferecido mas cancelado pelo usuário"
                #     else:
                #         mensagem += f"\n\n{resultado_verificacao['mensagem']}"
                # else:
                #     mensagem += f"\n\n⚠️ AVISO: {resultado_verificacao['mensagem']}"
            
            custom_messagebox("info", "Sucesso", mensagem)
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro ao excluir: {traceback.format_exc()}")
            custom_messagebox("error", "Erro", f"Erro ao excluir lançamento: {str(e)}")

    def restaurar_lancamento(self):
        """
        Versão corrigida que usa o novo sistema de verificação após restauração
        """
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para restaurar")
            return
        
        valores = self.tree_lancamentos.item(item_selecionado[0])['values']
        tp_desp = valores[1]         # Tipo de despesa
        nome_lancamento = valores[2]  # Nome
        referencia = valores[3]      # Referência
        valor = valores[5]           # Valor
        data_lancamento = valores[0] # Data
        status_atual = valores[7]    # Status
        id_lancamento = valores[8]   # ID
        
        # Verificar se já está ativo
        if status_atual != 'EXCLUIDO':
            custom_messagebox("info", "Informação", "Este lançamento já está ativo")
            return
        
        # Verificar se é uma taxa de administração
        eh_taxa = (tp_desp == 7)
        
        # Confirmar restauração
        if eh_taxa:
            if not custom_messagebox("yesno", "Confirmação - Taxa de Administração", 
                                f"Você está restaurando uma TAXA DE ADMINISTRAÇÃO:\n\n"
                                f"📋 {referencia}\n"
                                f"💰 {valor}\n"
                                f"📅 {data_lancamento}\n\n"
                                f"⚠️ ATENÇÃO: Verifique se não há duplicação de taxas!\n\n"
                                f"Deseja realmente continuar?"):
                return
        else:
            if not custom_messagebox("yesno", "Confirmação", 
                                f"Deseja realmente restaurar este lançamento?\n\n"
                                f"👤 {nome_lancamento}\n"
                                f"📋 {referencia}\n"
                                f"💰 {valor}\n"
                                f"📅 {data_lancamento}\n\n"
                                f"🔄 As taxas de administração do tipo % deverão ser verificadas."):
                return
        
        try:
            id_lancamento = valores[8]
            
            print(f"DEBUG: Restaurando lançamento ID {id_lancamento}")
            
            # Atualizar status para ATIVO
            self.atualizar_status_lancamento(id_lancamento, 'ATIVO')
            print(f"DEBUG: Status atualizado para ATIVO")

            # ===== NOVA INTEGRAÇÃO: Usar método unificado =====
            print(f"DEBUG: Iniciando verificação de recálculo para data {data_lancamento}")
            data_para_recalculo = datetime.strptime(data_lancamento, '%d/%m/%Y').date()
            
            # Aguardar um pouco para garantir que a restauração foi salva
            import time
            time.sleep(0.5)
            
            # CORREÇÃO: Usar o novo método de verificação
            # resultado_verificacao = self.sistema.chamar_apos_operacao_lancamento(data_para_recalculo, "ALTERACAO")
            
            # Recarregar lista
            self.carregar_lancamentos()
            
            # Mostrar resultado
            if eh_taxa:
                mensagem = "Taxa de administração restaurada com sucesso!"
                mensagem += "\n\n⚠️ IMPORTANTE: Verifique se não há duplicação de taxas para esta data."
            else:
                mensagem = "Lançamento restaurado com sucesso!"
                
                # if resultado_verificacao["sucesso"]:
                #     # Verificar se houve recálculo automático baseado no resultado
                #     if ("recalculadas" in resultado_verificacao["mensagem"] or 
                #         "Recálculo Concluído" in str(resultado_verificacao)):
                #         mensagem += f"\n\n✅ Taxas foram recalculadas automaticamente!"
                #     elif "corretas" in resultado_verificacao["mensagem"]:
                #         mensagem += f"\n\n✅ Taxas verificadas - estão corretas"
                #     elif "cancelado" in resultado_verificacao["mensagem"]:
                #         mensagem += f"\n\n⚠️ Recálculo foi oferecido mas cancelado pelo usuário"
                #     else:
                #         mensagem += f"\n\n{resultado_verificacao['mensagem']}"
                # else:
                #     mensagem += f"\n\n⚠️ AVISO: {resultado_verificacao['mensagem']}"
            
            custom_messagebox("info", "Sucesso", mensagem)
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro ao restaurar: {traceback.format_exc()}")
            custom_messagebox("error", "Erro", f"Erro ao restaurar lançamento: {str(e)}")        

    def visualizar_historico_lancamento(self):
        """Visualiza o histórico de alterações de um lançamento com correção robusta"""
        try:
            # Verificar se há item selecionado
            selected_items = self.tree_lancamentos.selection()
            if not selected_items:
                custom_messagebox("info", "Seleção", "Selecione um lançamento para ver o histórico!")
                return
            
            # Obter o item selecionado
            item = selected_items[0]
            valores = self.tree_lancamentos.item(item, 'values')
            
            # CORREÇÃO: O ID está na última posição (índice 8)
            # Colunas: ('Data', 'Tipo', 'Nome', 'Referência', 'NF', 'Valor', 'Vencimento', 'Status', 'ID')
            id_lancamento = valores[8]  # ID é o último elemento
            
            print(f"DEBUG: Buscando histórico para ID: {id_lancamento} (tipo: {type(id_lancamento)})")
            
            # CORREÇÃO: Verificar se dados_originais existe e não está vazio
            if not hasattr(self, 'dados_originais') or self.dados_originais.empty:
                custom_messagebox("error", "Erro", "Dados não carregados. Clique em 'Atualizar' primeiro!")
                return
            
            # CORREÇÃO: Converter ID para o mesmo tipo usado no DataFrame
            try:
                # Primeiro, verificar qual tipo está sendo usado na coluna ID_LANCAMENTO
                id_col_dtype = self.dados_originais['ID_LANCAMENTO'].dtype
                print(f"DEBUG: Tipo da coluna ID_LANCAMENTO: {id_col_dtype}")
                
                # Converter o ID para o tipo correto
                if 'int' in str(id_col_dtype):
                    id_busca = int(float(str(id_lancamento)))  # Conversão robusta via float primeiro
                else:
                    id_busca = str(id_lancamento)
                    
                print(f"DEBUG: ID convertido para busca: {id_busca} (tipo: {type(id_busca)})")
                
            except (ValueError, TypeError) as e:
                print(f"DEBUG: Erro na conversão do ID: {e}")
                custom_messagebox("error", "Erro", f"ID inválido: {id_lancamento}")
                return
            
            # CORREÇÃO: Busca mais robusta com verificação de resultado
            filtro = self.dados_originais['ID_LANCAMENTO'] == id_busca
            lancamentos_encontrados = self.dados_originais[filtro]
            
            print(f"DEBUG: Lançamentos encontrados: {len(lancamentos_encontrados)}")
            
            if lancamentos_encontrados.empty:
                # DEBUG: Mostrar alguns IDs disponíveis para comparação
                ids_disponiveis = self.dados_originais['ID_LANCAMENTO'].head(10).tolist()
                print(f"DEBUG: Primeiros 10 IDs disponíveis: {ids_disponiveis}")
                
                custom_messagebox("error", "Erro", 
                    f"Lançamento com ID {id_busca} não encontrado!\n"
                    f"Clique em 'Atualizar' para recarregar os dados.")
                return
            
            # Obter o lançamento (agora sabemos que existe)
            lancamento = lancamentos_encontrados.iloc[0]
            
            # Criar janela de histórico - TAMANHO REDUZIDO
            janela_historico = tk.Toplevel(self.janela)
            janela_historico.title(f"Histórico do Lançamento - ID {id_busca}")
            janela_historico.geometry("700x300")  # Reduzido de 800x600 para 700x400
            janela_historico.transient(self.janela)
            janela_historico.grab_set()
            
            # Frame principal
            frame_principal = ttk.Frame(janela_historico, padding="10")
            frame_principal.pack(fill='both', expand=True)
            
            # Informações do lançamento - FORMATO MAIS COMPACTO
            frame_info = ttk.LabelFrame(frame_principal, text="Informações do Lançamento")
            frame_info.pack(fill='x', pady=(0, 10))
            
            # LAYOUT EM DUAS COLUNAS para economizar espaço vertical
            info_frame_interno = ttk.Frame(frame_info)
            info_frame_interno.pack(fill='x', padx=10, pady=5)
            
            # Coluna esquerda
            frame_esq = ttk.Frame(info_frame_interno)
            frame_esq.pack(side='left', fill='x', expand=True)
            
            ttk.Label(frame_esq, text=f"ID: {lancamento['ID_LANCAMENTO']}", font=('TkDefaultFont', 9, 'bold')).pack(anchor='w')
            ttk.Label(frame_esq, text=f"Nome: {lancamento['NOME']}", font=('TkDefaultFont', 9)).pack(anchor='w')
            ttk.Label(frame_esq, text=f"Referência: {lancamento['REFERÊNCIA']}", font=('TkDefaultFont', 9)).pack(anchor='w')
            
            # Coluna direita
            frame_dir = ttk.Frame(info_frame_interno)
            frame_dir.pack(side='right', fill='x', expand=True)
            
            ttk.Label(frame_dir, text=f"Valor: R$ {lancamento['VALOR']:,.2f}", font=('TkDefaultFont', 9)).pack(anchor='w')
            ttk.Label(frame_dir, text=f"Status: {lancamento['STATUS']}", font=('TkDefaultFont', 9)).pack(anchor='w')
            
            # Frame do histórico - ALTURA FIXA E MENOR
            frame_historico = ttk.LabelFrame(frame_principal, text="Histórico de Alterações")
            frame_historico.pack(fill='x', pady=(0, 10))  # fill='x' em vez de fill='both', expand=True
            
            # Text widget para mostrar o histórico - ALTURA FIXA
            text_historico = tk.Text(frame_historico, wrap='word', font=('Consolas', 9), height=8)  # height=8 linhas fixas
            scrollbar_hist = ttk.Scrollbar(frame_historico, orient='vertical', command=text_historico.yview)
            text_historico.configure(yscrollcommand=scrollbar_hist.set)
            
            # Obter histórico
            historico = lancamento.get('HISTORICO_ALTERACAO', '')
            if historico and str(historico) not in ['', 'nan', 'None']:
                text_historico.insert('1.0', str(historico))
            else:
                text_historico.insert('1.0', "Nenhum histórico de alterações registrado.")
            
            text_historico.config(state='disabled')  # Apenas leitura
            
            # Posicionar elementos com padding reduzido
            text_historico.pack(side='left', fill='both', expand=True, padx=(10, 0), pady=5)
            scrollbar_hist.pack(side='right', fill='y', pady=5)
            
            # Frame de botões para melhor organização
            frame_botoes = ttk.Frame(frame_principal)
            frame_botoes.pack(fill='x', pady=(5, 0))
            
            # Botão fechar centralizado
            ttk.Button(frame_botoes, text="Fechar", 
                    command=janela_historico.destroy).pack(side='right')
            
            print(f"DEBUG: Histórico exibido com sucesso para ID {id_busca}")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao visualizar histórico: {str(e)}")
            print(f"DEBUG: Erro completo: {str(e)}")

    def salvar_edicao(self, id_lancamento, dados_editados):
        """
        Callback chamado pelo EditorLancamentoCompleto após edição
        
        Args:
            id_lancamento: ID do lançamento sendo editado
            dados_editados: Novos dados do lançamento
        
        Returns:
            bool: True se salvou com sucesso, False caso contrário
        """
        try:
            print(f"DEBUG: Salvando edição do lançamento ID {id_lancamento}")
            
            # Buscar dados originais para comparação
            dados_originais = None
            data_original = None
            
            # Buscar no DataFrame atual
            if hasattr(self, 'dados_originais') and not self.dados_originais.empty:
                mask = self.dados_originais['ID_LANCAMENTO'] == id_lancamento
                dados_filtrados = self.dados_originais[mask]
                
                if not dados_filtrados.empty:
                    dados_originais = dados_filtrados.iloc[0]
                    data_original = dados_originais.get('DATA_REL')
                    if pd.notna(data_original):
                        if isinstance(data_original, str):
                            data_original = datetime.strptime(data_original, '%d/%m/%Y').date()
                        else:
                            data_original = data_original.date()
                    print(f"DEBUG: Data original encontrada: {data_original}")
            
            # Obter data editada
            data_editada = None
            if dados_editados.get('data'):
                if isinstance(dados_editados['data'], str):
                    data_editada = datetime.strptime(dados_editados['data'], '%d/%m/%Y').date()
                else:
                    data_editada = dados_editados['data']
                print(f"DEBUG: Data editada: {data_editada}")
            
            # ===== SALVAR A EDIÇÃO NA PLANILHA =====
            sucesso_salvamento = self._executar_salvamento_edicao(id_lancamento, dados_editados, dados_originais)
            
            if not sucesso_salvamento:
                return False
            
            # ===== NOVA INTEGRAÇÃO: Verificar recálculo após edição =====
            print("DEBUG: Iniciando verificação de recálculo após edição")
            
            # Determinar quais datas precisam ser verificadas
            datas_para_verificar = set()
            
            if data_original:
                datas_para_verificar.add(data_original)
                print(f"DEBUG: Adicionada data original para verificação: {data_original}")
            
            if data_editada and data_editada != data_original:
                datas_para_verificar.add(data_editada)
                print(f"DEBUG: Adicionada data editada para verificação: {data_editada}")
            elif data_editada and not data_original:
                datas_para_verificar.add(data_editada)
                print(f"DEBUG: Adicionada apenas data editada: {data_editada}")
            
            print(f"DEBUG: Total de datas para verificar: {len(datas_para_verificar)}")
            
            # Aguardar um pouco para garantir que a edição foi salva
            import time
            time.sleep(0.5)
            
            # Verificar cada data afetada usando o novo sistema
            # resultados_verificacao = []
            
            # for data_verificar in datas_para_verificar:
            #     try:
            #         print(f"DEBUG: Verificando recálculo para {data_verificar}")
                    
            #         # INTEGRAÇÃO: Usar o método unificado
            #         resultado = self.sistema.chamar_apos_operacao_lancamento(data_verificar, "ALTERACAO")
                    
            #         resultados_verificacao.append({
            #             'data': data_verificar,
            #             'resultado': resultado
            #         })
                    
            #         if resultado["sucesso"]:
            #             print(f"✅ Verificação concluída para {data_verificar}: {resultado['mensagem']}")
            #         else:
            #             print(f"⚠️ Problema na verificação para {data_verificar}: {resultado['mensagem']}")
                        
            #     except Exception as e:
            #         print(f"❌ Erro ao verificar {data_verificar}: {str(e)}")
            #         resultados_verificacao.append({
            #             'data': data_verificar,
            #             'resultado': {"sucesso": False, "mensagem": f"Erro: {str(e)}"}
            #         })
            #         continue
            
            # Recarregar a visualização se existir
            if hasattr(self, 'carregar_lancamentos'):
                self.carregar_lancamentos()
            
            # Log do resultado final
            # verificacoes_ok = sum(1 for r in resultados_verificacao if r['resultado']['sucesso'])
            # print(f"DEBUG: Edição salva. Verificações: {verificacoes_ok}/{len(resultados_verificacao)} OK")
            
            return True
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro geral ao salvar edição: {traceback.format_exc()}")
            return False

    def _executar_salvamento_edicao(self, id_lancamento, dados_editados, dados_originais):
        """
        Executa o salvamento físico da edição na planilha
        
        Args:
            id_lancamento: ID do lançamento
            dados_editados: Novos dados
            dados_originais: Dados originais (para histórico)
        
        Returns:
            bool: True se salvou com sucesso
        """
        try:
            # Abrir arquivo
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb["Dados"]
            
            # Encontrar a linha do lançamento
            linha_encontrada = None
            for row_num in range(2, ws.max_row + 1):
                id_na_planilha = ws.cell(row=row_num, column=15).value  # Coluna O (ID_LANCAMENTO)
                
                if str(id_na_planilha) == str(id_lancamento):
                    linha_encontrada = row_num
                    break
            
            if not linha_encontrada:
                print(f"DEBUG: Lançamento ID {id_lancamento} não encontrado na planilha")
                wb.close()
                return False
            
            print(f"DEBUG: Editando lançamento na linha {linha_encontrada}")
            
            # ===== ATUALIZAR OS DADOS NA PLANILHA =====
            
            # Data do relatório (Coluna A)
            if dados_editados.get('data'):
                data_rel = datetime.strptime(dados_editados['data'], '%d/%m/%Y') if isinstance(dados_editados['data'], str) else dados_editados['data']
                ws.cell(row=linha_encontrada, column=1, value=data_rel)
                ws.cell(row=linha_encontrada, column=1).number_format = 'DD/MM/YYYY'
            
            # Tipo de despesa (Coluna B)
            if dados_editados.get('tp_desp'):
                ws.cell(row=linha_encontrada, column=2, value=int(dados_editados['tp_desp']))
            
            # CNPJ/CPF (Coluna C)
            ws.cell(row=linha_encontrada, column=3, value=dados_editados.get('cnpj_cpf', ''))
            
            # Nome (Coluna D)
            ws.cell(row=linha_encontrada, column=4, value=dados_editados.get('nome', ''))
            
            # Referência (Coluna E)
            ws.cell(row=linha_encontrada, column=5, value=dados_editados.get('referencia', ''))
            
            # Verificar se os cabeçalhos existem e criar se necessário
            if ws.cell(row=1, column=17).value != 'ETAPA_OBRA':
                ws.cell(row=1, column=17, value='ETAPA_OBRA')
            
            if ws.cell(row=1, column=18).value != 'INSUMO':
                ws.cell(row=1, column=18, value='INSUMO')
            
            # Etapa da Obra (Coluna Q)
            ws.cell(row=linha_encontrada, column=17, value=dados_editados.get('etapa_obra', ''))
            
            # Insumo (Coluna R)
            ws.cell(row=linha_encontrada, column=18, value=dados_editados.get('insumo', ''))
            # === FIM DOS NOVOS CAMPOS ===
            
            # NF (Coluna F)
            ws.cell(row=linha_encontrada, column=6, value=dados_editados.get('nf', ''))
            
            # Valor Unitário (Coluna G)
            if dados_editados.get('vr_unit'):
                vr_unit = float(dados_editados['vr_unit'])
                ws.cell(row=linha_encontrada, column=7, value=vr_unit)
                ws.cell(row=linha_encontrada, column=7).number_format = '#,##0.00'
            
            # Dias (Coluna H)
            if dados_editados.get('dias'):
                ws.cell(row=linha_encontrada, column=8, value=int(dados_editados['dias']))
            
            # Valor Total (Coluna I)
            if dados_editados.get('valor'):
                valor = float(dados_editados['valor'])
                ws.cell(row=linha_encontrada, column=9, value=valor)
                ws.cell(row=linha_encontrada, column=9).number_format = '#,##0.00'
            
            # Data de Vencimento (Coluna J)
            if dados_editados.get('dt_vencto'):
                dt_vencto = datetime.strptime(dados_editados['dt_vencto'], '%d/%m/%Y') if isinstance(dados_editados['dt_vencto'], str) else dados_editados['dt_vencto']
                ws.cell(row=linha_encontrada, column=10, value=dt_vencto)
                ws.cell(row=linha_encontrada, column=10).number_format = 'DD/MM/YYYY'
            
            # Categoria (Coluna K)
            ws.cell(row=linha_encontrada, column=11, value=dados_editados.get('categoria', ''))
            
            # Dados Bancários (Coluna L)
            ws.cell(row=linha_encontrada, column=12, value=dados_editados.get('dados_bancarios', ''))
            
            # ===== OBSERVAÇÃO COM HISTÓRICO DE EDIÇÃO (Coluna M) =====
            observacao_atual = ws.cell(row=linha_encontrada, column=13).value or ""
            observacao_editada = dados_editados.get('observacao', '')
            
            # Limpar edições anteriores da observação
            if 'EDITADO EM:' in observacao_atual:
                observacao_base = observacao_atual.split(' - EDITADO EM:')[0]
            else:
                observacao_base = observacao_atual
            
            # Se a observação mudou, usar a nova; senão manter a base
            if observacao_editada.strip() and observacao_editada.strip() != observacao_base.strip():
                observacao_final = observacao_editada
            else:
                observacao_final = observacao_base
            
            # Adicionar timestamp de edição
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            observacao_com_historico = f"{observacao_final} - EDITADO EM: {timestamp}"
            
            ws.cell(row=linha_encontrada, column=13, value=observacao_com_historico)
            
            # ===== HISTÓRICO DE ALTERAÇÕES (Coluna P) =====
            historico_atual = ws.cell(row=linha_encontrada, column=16).value or ""
            
            # Criar resumo das principais alterações
            alteracoes = []
            
            if dados_originais is not None:
                # Verificar principais campos que mudaram
                if dados_editados.get('valor') and str(dados_editados['valor']) != str(dados_originais.get('VALOR', '')).replace(',', '.'):
                    valor_antigo = dados_originais.get('VALOR', 0)
                    valor_novo = dados_editados['valor']
                    alteracoes.append(f"VALOR: {valor_antigo} → {valor_novo}")
                
                if dados_editados.get('data'):
                    data_antiga = dados_originais.get('DATA_REL')
                    if pd.notna(data_antiga):
                        if isinstance(data_antiga, str):
                            data_antiga_str = data_antiga
                        else:
                            data_antiga_str = data_antiga.strftime('%d/%m/%Y')
                        
                        data_nova_str = dados_editados['data'] if isinstance(dados_editados['data'], str) else dados_editados['data'].strftime('%d/%m/%Y')
                        
                        if data_antiga_str != data_nova_str:
                            alteracoes.append(f"DATA: {data_antiga_str} → {data_nova_str}")
                
                # === VERIFICAR ALTERAÇÕES NOS NOVOS CAMPOS ===
                if dados_editados.get('etapa_obra') != str(dados_originais.get('ETAPA_OBRA', '')):
                    etapa_antiga = dados_originais.get('ETAPA_OBRA', '')
                    etapa_nova = dados_editados.get('etapa_obra', '')
                    if etapa_antiga != etapa_nova:
                        alteracoes.append(f"ETAPA_OBRA: {etapa_antiga} → {etapa_nova}")
                
                if dados_editados.get('insumo') != str(dados_originais.get('INSUMO', '')):
                    insumo_antigo = dados_originais.get('INSUMO', '')
                    insumo_novo = dados_editados.get('insumo', '')
                    if insumo_antigo != insumo_novo:
                        alteracoes.append(f"INSUMO: {insumo_antigo} → {insumo_novo}")
                # === FIM DA VERIFICAÇÃO DOS NOVOS CAMPOS ===
            
            # Se não conseguiu detectar alterações específicas, registrar edição geral
            if not alteracoes:
                alteracoes = ["EDITADO"]
            
            # Adicionar ao histórico
            nova_entrada = f"EDIÇÃO: {', '.join(alteracoes)} - {timestamp}"
            
            if historico_atual:
                # Limitar histórico para não ficar muito longo (manter últimas 5 entradas)
                historico_partes = historico_atual.split(' | ')
                if len(historico_partes) >= 5:
                    historico_partes = historico_partes[-4:]  # Manter últimas 4
                novo_historico = ' | '.join(historico_partes) + ' | ' + nova_entrada
            else:
                novo_historico = nova_entrada
            
            ws.cell(row=linha_encontrada, column=16, value=novo_historico)
            
            # Salvar arquivo
            wb.save(arquivo_cliente)
            wb.close()
            
            print(f"✅ Edição salva com sucesso na linha {linha_encontrada}")
            return True
            
        except Exception as e:
            import traceback
            print(f"DEBUG: Erro ao salvar edição na planilha: {traceback.format_exc()}")
            if 'wb' in locals():
                wb.close()
            return False
    
    def atualizar_status_lancamento(self, id_lancamento, novo_status):
        """Atualiza o status de um lançamento específico"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            
            # Carregar workbook
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            # Verificar se colunas existem, se não, criar
            if ws.cell(row=1, column=14).value != 'STATUS':
                ws.cell(row=1, column=14, value='STATUS')
            
            if ws.cell(row=1, column=15).value != 'ID_LANCAMENTO':
                ws.cell(row=1, column=15, value='ID_LANCAMENTO')
                
            if ws.cell(row=1, column=16).value != 'HISTORICO_ALTERACAO':
                ws.cell(row=1, column=16, value='HISTORICO_ALTERACAO')
            
            # Encontrar e atualizar linha
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=15).value == id_lancamento:
                    # Atualizar status
                    ws.cell(row=row, column=14, value=novo_status)
                    
                    # Adicionar ao histórico
                    timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    historico_atual = ws.cell(row=row, column=16).value or ""
                    
                    if novo_status == 'EXCLUIDO':
                        acao = f"EXCLUÍDO EM: {timestamp}"
                    elif novo_status == 'ATIVO':
                        acao = f"RESTAURADO EM: {timestamp}"
                    else:
                        acao = f"STATUS ALTERADO PARA {novo_status} EM: {timestamp}"
                    
                    if historico_atual:
                        novo_historico = f"{historico_atual} | {acao}"
                    else:
                        novo_historico = acao
                        
                    ws.cell(row=row, column=16, value=novo_historico)
                    break
            
            wb.save(arquivo_cliente)
            
        except Exception as e:
            raise Exception(f"Erro ao atualizar status: {str(e)}")
    
    def formatar_data(self, data):
        """Formata data para exibição"""
        if pd.isna(data) or data == "":
            return ""
        try:
            if isinstance(data, str):
                return data
            return data.strftime('%d/%m/%Y')
        except:
            return str(data)
    
    def formatar_valor(self, valor):
        """Formata valor para exibição"""
        try:
            if isinstance(valor, str):
                valor = float(valor.replace(',', '.'))
            return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return str(valor)

class EditorLancamentoCompleto:
    def __init__(self, parent, lancamento, callback_salvar):
        self.janela = tk.Toplevel(parent)
        self.janela.title("Editar Lançamento")
        self.janela.geometry("700x600")
        self.lancamento = lancamento
        self.callback_salvar = callback_salvar
        
        # Configurar janela
        self.janela.transient(parent)
        self.janela.grab_set()
        
        self.criar_interface()
        self.preencher_dados()
        
    def criar_interface(self):
        """Cria a interface do editor"""
        # Frame principal com scroll
        main_frame = ttk.Frame(self.janela, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Dados básicos
        frame_basicos = ttk.LabelFrame(main_frame, text="Dados Básicos")
        frame_basicos.pack(fill='x', pady=5)
        
        # Data do Relatório
        ttk.Label(frame_basicos, text="Data do Relatório:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.data_rel = DateEntry(frame_basicos, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_rel.grid(row=0, column=1, padx=5, pady=2, sticky='w')
        
        # Tipo de Despesa
        ttk.Label(frame_basicos, text="Tipo Despesa:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.tp_desp = ttk.Entry(frame_basicos, width=5)
        self.tp_desp.grid(row=0, column=3, padx=5, pady=2, sticky='w')
        
        # Dados do Fornecedor
        frame_fornecedor = ttk.LabelFrame(main_frame, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', pady=5)
        
        ttk.Label(frame_fornecedor, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.cnpj_cpf = ttk.Entry(frame_fornecedor, width=20)
        self.cnpj_cpf.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_fornecedor, text="Nome:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.nome = ttk.Entry(frame_fornecedor, width=50)
        self.nome.grid(row=1, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_fornecedor, text="Categoria:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.categoria = ttk.Entry(frame_fornecedor, width=10)
        self.categoria.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        
        # Dados da Despesa
        frame_despesa = ttk.LabelFrame(main_frame, text="Dados da Despesa")
        frame_despesa.pack(fill='x', pady=5)
        
        ttk.Label(frame_despesa, text="Referência:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.referencia = ttk.Entry(frame_despesa, width=40)
        self.referencia.grid(row=0, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_despesa, text="Etapa da Obra:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        
        # Obter lista de etapas das configurações
        try:
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            etapas_obra = GerenciadorConfiguracoes.get_etapas_obra()
        except ImportError:
            etapas_obra = []  # Lista vazia se não conseguir importar
        
        etapas_obra = GerenciadorConfiguracoes.get_etapas_obra()
        self.etapa_obra = ComboboxAutocompletar(
            frame_despesa,
            values=etapas_obra,
            config_key='etapas_obra',
            config_manager=GerenciadorConfiguracoes,
            width=30,
            state='normal'
        )
        self.etapa_obra.grid(row=1, column=1, columnspan=2, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_despesa, text="Insumo:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        
        # Obter lista de insumos das configurações
        try:
            insumos = GerenciadorConfiguracoes.get_insumos()
        except (ImportError, AttributeError):
            insumos = []  # Lista vazia se não conseguir importar
        
        insumos = GerenciadorConfiguracoes.get_insumos()
        self.insumo = ComboboxAutocompletar(
            frame_despesa,
            values=insumos,
            config_key='insumos',
            config_manager=GerenciadorConfiguracoes,
            width=30,
            state='normal'
        )
        self.insumo.grid(row=2, column=1, columnspan=2, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_despesa, text="NF:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.nf = ttk.Entry(frame_despesa, width=15)
        self.nf.grid(row=3, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Valor Unitário:").grid(row=4, column=0, padx=5, pady=2, sticky='w')
        self.vr_unit = ttk.Entry(frame_despesa, width=15)
        self.vr_unit.grid(row=4, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Dias:").grid(row=4, column=2, padx=5, pady=2, sticky='w')
        self.dias = ttk.Entry(frame_despesa, width=8)
        self.dias.grid(row=4, column=3, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Valor Total:").grid(row=5, column=0, padx=5, pady=2, sticky='w')
        self.valor = ttk.Entry(frame_despesa, width=15)
        self.valor.grid(row=5, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Data Vencimento:").grid(row=5, column=2, padx=5, pady=2, sticky='w')
        self.dt_vencto = DateEntry(frame_despesa, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.dt_vencto.grid(row=5, column=3, padx=5, pady=2, sticky='w')
        
        # Dados Bancários e Observação
        ttk.Label(frame_despesa, text="Dados Bancários:").grid(row=6, column=0, padx=5, pady=2, sticky='w')
        self.dados_bancarios = ttk.Entry(frame_despesa, width=50)
        self.dados_bancarios.grid(row=6, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_despesa, text="Observação:").grid(row=7, column=0, padx=5, pady=2, sticky='w')
        self.observacao = ttk.Entry(frame_despesa, width=50)
        self.observacao.grid(row=7, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        # Configurar expansão
        frame_fornecedor.columnconfigure(1, weight=1)
        frame_despesa.columnconfigure(1, weight=1)
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Salvar", command=self.salvar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=self.janela.destroy).pack(side='left', padx=5)
        
        # Bindings para cálculo automático
        self.vr_unit.bind('<KeyRelease>', self.calcular_total)
        self.dias.bind('<KeyRelease>', self.calcular_total)
        
    def preencher_dados(self):
        """Preenche os campos com os dados do lançamento"""
        try:
            # Datas
            if pd.notna(self.lancamento['DATA_REL']):
                self.data_rel.set_date(pd.to_datetime(self.lancamento['DATA_REL']))
            if pd.notna(self.lancamento['DT_VENCTO']):
                self.dt_vencto.set_date(pd.to_datetime(self.lancamento['DT_VENCTO']))
            
            # Campos de texto
            tp_desp_valor = self.lancamento.get('TP_DESP', '')
            if pd.notna(tp_desp_valor):
                self.tp_desp.insert(0, str(int(float(tp_desp_valor))))
            
            self.cnpj_cpf.insert(0, str(self.lancamento.get('CNPJ_CPF', '')))
            self.nome.insert(0, str(self.lancamento.get('NOME', '')))
            self.categoria.insert(0, str(self.lancamento.get('CATEGORIA', '')))
            self.referencia.insert(0, str(self.lancamento.get('REFERÊNCIA', '')))
            
            etapa_obra_valor = str(self.lancamento.get('ETAPA_OBRA', ''))
            if etapa_obra_valor and etapa_obra_valor != 'nan':
                self.etapa_obra.set(etapa_obra_valor)
            
            insumo_valor = str(self.lancamento.get('INSUMO', ''))
            if insumo_valor and insumo_valor != 'nan':
                self.insumo.set(insumo_valor)
                
            self.nf.insert(0, str(self.lancamento.get('NF', '')))
            
            # Valores numéricos
            vr_unit_valor = self.lancamento.get('VR_UNIT', '')
            if pd.notna(vr_unit_valor) and vr_unit_valor != '':
                self.vr_unit.insert(0, str(float(vr_unit_valor)).replace('.', ','))
            
            dias_valor = self.lancamento.get('DIAS', '')
            if pd.notna(dias_valor) and dias_valor != '':
                self.dias.insert(0, str(int(float(dias_valor))))
                
            valor_valor = self.lancamento.get('VALOR', '')
            if pd.notna(valor_valor) and valor_valor != '':
                self.valor.insert(0, str(float(valor_valor)).replace('.', ','))
            
            self.dados_bancarios.insert(0, str(self.lancamento.get('DADOS_BANCARIOS', '')))
            
            # CORREÇÃO: Preencher apenas a observação original (sem histórico de edições)
            observacao_original = str(self.lancamento.get('OBSERVAÇÃO', ''))
            # Remover informações de edição anteriores se existirem
            if 'EDITADO EM:' in observacao_original:
                observacao_original = observacao_original.split(' - EDITADO EM:')[0]
            
            self.observacao.insert(0, observacao_original)
            
        except Exception as e:
            print(f"Erro ao preencher dados: {str(e)}")
    
    def calcular_total(self, event=None):
        """Calcula o valor total automaticamente"""
        try:
            vr_unit = float(self.vr_unit.get().replace(',', '.'))
            dias = float(self.dias.get() or 1)
            total = vr_unit * dias
            
            self.valor.delete(0, tk.END)
            self.valor.insert(0, f"{total:.2f}")
        except:
            pass
   

    def salvar(self):
        """Salva as alterações"""
        try:
            # Validações básicas
            if not self.nome.get().strip():
                custom_messagebox("error", "Erro", "Nome é obrigatório!")
                return
            
            if not self.valor.get().strip():
                custom_messagebox("error", "Erro", "Valor é obrigatório!")
                return
            
            # Coletar dados
            dados_editados = {
                'data': self.data_rel.get(),
                'tp_desp': self.tp_desp.get(),
                'cnpj_cpf': self.cnpj_cpf.get(),
                'nome': self.nome.get().upper(),
                'categoria': self.categoria.get().upper(),
                'referencia': self.referencia.get().upper(),
                'etapa_obra': self.etapa_obra.get(), 
                'insumo': self.insumo.get(),
                'nf': self.nf.get().upper(),
                'vr_unit': self.vr_unit.get().replace(',', '.'),
                'dias': self.dias.get() or '1',
                'valor': self.valor.get().replace(',', '.'),
                'dt_vencto': self.dt_vencto.get(),
                'dados_bancarios': self.dados_bancarios.get(),
                'observacao': self.observacao.get().upper()
            }
            
            # Chamar callback para salvar
            id_lancamento = self.lancamento.get('ID_LANCAMENTO')
            if self.callback_salvar(id_lancamento, dados_editados):
                custom_messagebox("info", "Sucesso", "Lançamento atualizado com sucesso!")
                self.janela.destroy()
            else:
                custom_messagebox("error", "Erro", "Erro ao salvar alterações!")
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar: {str(e)}")                

class VisualizadorLancamentosFornecedor:
    def __init__(self, parent, sistema_principal):
        self.parent = parent
        self.sistema = sistema_principal
        self.janela = None
        self.tree_lancamentos = None
        self.dados_originais = []  # Para armazenar dados completos
        
    def formatar_cnpj_cpf(self, cnpj_cpf):
        """Formata CNPJ/CPF mantendo zeros à esquerda"""
        try:
            numeros = ''.join(filter(str.isdigit, str(cnpj_cpf)))
            
            if len(numeros) == 11:  # CPF
                cpf = numeros.zfill(11)
                return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
            elif len(numeros) == 14:  # CNPJ
                cnpj = numeros.zfill(14)
                return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            else:
                return str(cnpj_cpf)
        except:
            return str(cnpj_cpf)
        
    def abrir_visualizador(self, cnpj_cpf_fornecedor, nome_fornecedor):
        """Abre o visualizador para um fornecedor específico"""
        if not self.sistema.cliente_atual:
            custom_messagebox("error", "Erro", "Nenhum cliente selecionado!")
            return
            
        # Normalizar CNPJ/CPF com zeros à esquerda ANTES de formatar
        cnpj_cpf_str = str(cnpj_cpf_fornecedor)
        cnpj_cpf_numeros = ''.join(filter(str.isdigit, cnpj_cpf_str))
        
        if len(cnpj_cpf_numeros) <= 11:
            cnpj_cpf_normalizado = cnpj_cpf_numeros.zfill(11)  # CPF
        else:
            cnpj_cpf_normalizado = cnpj_cpf_numeros.zfill(14)  # CNPJ
            
        cnpj_cpf_formatado = self.formatar_cnpj_cpf(cnpj_cpf_normalizado)
        
        self.dados_fornecedor = {
            'cnpj_cpf': cnpj_cpf_normalizado,
            'cnpj_cpf_formatado': cnpj_cpf_formatado,
            'nome': nome_fornecedor
        }
        
        self.criar_janela()
        self.carregar_lancamentos()
        
    def criar_janela(self):
        """Cria a janela do visualizador com funcionalidades estendidas"""
        self.janela = tk.Toplevel(self.parent)
        self.janela.title(f"Lançamentos - {self.dados_fornecedor['nome']}")
        self.janela.geometry("1400x900")  # Aumentado para acomodar novos botões
        
        # Configurar janela
        self.janela.transient(self.parent)
        self.janela.grab_set()
        
        self.criar_interface()
        
    def criar_interface(self):
        """Cria a interface do visualizador com funcionalidades de edição"""
        # Frame principal
        main_frame = ttk.Frame(self.janela, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Cabeçalho com informações do fornecedor
        frame_cabecalho = ttk.LabelFrame(main_frame, text="Informações do Fornecedor")
        frame_cabecalho.pack(fill='x', pady=(0, 10))
        
        info_frame = ttk.Frame(frame_cabecalho)
        info_frame.pack(fill='x', padx=10, pady=8)
        
        # Informações em duas colunas
        ttk.Label(info_frame, text="Nome:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, padx=5, pady=2, sticky='w')
        ttk.Label(info_frame, text=self.dados_fornecedor['nome'], 
                 font=('Arial', 10)).grid(row=0, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(info_frame, text="CNPJ/CPF:", font=('Arial', 10, 'bold')).grid(
            row=0, column=2, padx=20, pady=2, sticky='w')
        ttk.Label(info_frame, text=self.dados_fornecedor['cnpj_cpf_formatado'], 
                 font=('Arial', 10)).grid(row=0, column=3, padx=5, pady=2, sticky='w')
        
        ttk.Label(info_frame, text="Cliente:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, padx=5, pady=2, sticky='w')
        ttk.Label(info_frame, text=self.sistema.cliente_atual, 
                 font=('Arial', 10)).grid(row=1, column=1, padx=5, pady=2, sticky='w')
        
        # Frame de filtros
        frame_filtros = ttk.LabelFrame(main_frame, text="Filtros")
        frame_filtros.pack(fill='x', pady=(0, 10))
        
        filtros_frame = ttk.Frame(frame_filtros)
        filtros_frame.pack(fill='x', padx=10, pady=8)
        
        # Filtros por período
        ttk.Label(filtros_frame, text="Período:").grid(row=0, column=0, padx=5, pady=5)
        self.data_inicio = DateEntry(filtros_frame, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_inicio.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(filtros_frame, text="até").grid(row=0, column=2, padx=5, pady=5)
        self.data_fim = DateEntry(filtros_frame, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_fim.grid(row=0, column=3, padx=5, pady=5)
    
    # Filtro por referência
        
        # Filtro por referência
        ttk.Label(filtros_frame, text="Referência:").grid(row=0, column=4, padx=15, pady=5)
        self.filtro_referencia = ttk.Entry(filtros_frame, width=20)
        self.filtro_referencia.grid(row=0, column=5, padx=5, pady=5)
        
        # Filtro por status
        ttk.Label(filtros_frame, text="Status:").grid(row=0, column=6, padx=15, pady=5)
        self.combo_status = ttk.Combobox(filtros_frame, 
                                       values=['Todos', 'Ativos', 'Excluídos'], 
                                       state='readonly', width=10)
        self.combo_status.set('Ativos')
        self.combo_status.grid(row=0, column=7, padx=5, pady=5)
        
        # Botões de filtro
        ttk.Button(filtros_frame, text="Filtrar", 
                  command=self.aplicar_filtros).grid(row=0, column=8, padx=10, pady=5)
        ttk.Button(filtros_frame, text="Limpar", 
                  command=self.limpar_filtros).grid(row=0, column=9, padx=5, pady=5)
        
        # Campo de busca rápida
        frame_busca = ttk.Frame(filtros_frame)
        frame_busca.grid(row=1, column=0, columnspan=10, pady=10, sticky='ew')
        
        ttk.Label(frame_busca, text="Busca rápida por NF ou Observação:").pack(side='left', padx=5)
        self.busca_rapida = ttk.Entry(frame_busca, width=30)
        self.busca_rapida.pack(side='left', padx=5)
        self.busca_rapida.bind('<KeyRelease>', self.busca_incremental)
        
        # Frame da lista de lançamentos
        frame_lista = ttk.LabelFrame(main_frame, text="Lançamentos Encontrados")
        frame_lista.pack(fill='both', expand=True)
        
        # Treeview para lançamentos - ADICIONADA SELEÇÃO MÚLTIPLA
        colunas = ('Data Rel.', 'Tipo', 'Referência', 'NF', 'Valor', 'Vencimento', 'Status', 'Observação', 'ID')
        self.tree_lancamentos = ttk.Treeview(frame_lista, columns=colunas, show='headings', 
                                           height=18, selectmode='extended')  # Seleção múltipla
        
        # Configurar cabeçalhos e larguras
        larguras = {'Data Rel.': 90, 'Tipo': 50, 'Referência': 200, 'NF': 100, 
                   'Valor': 100, 'Vencimento': 90, 'Status': 80, 'Observação': 250, 'ID': 0}
        
        for col in colunas:
            self.tree_lancamentos.heading(col, text=col)
            self.tree_lancamentos.column(col, width=larguras.get(col, 100))
            if col == 'Valor':
                self.tree_lancamentos.column(col, anchor='e')
            elif col == 'ID':
                self.tree_lancamentos.column(col, width=0, stretch=False)  # Oculto
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_lancamentos.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient='horizontal', command=self.tree_lancamentos.xview)
        self.tree_lancamentos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionar elementos
        self.tree_lancamentos.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Frame de informações de seleção (NOVO)
        frame_selecao = ttk.Frame(main_frame)
        frame_selecao.pack(fill='x', pady=(5, 0))
        
        self.label_selecao = ttk.Label(frame_selecao, text="Nenhum item selecionado", 
                                     font=('TkDefaultFont', 9, 'italic'))
        self.label_selecao.pack(side='left')
        
        ttk.Button(frame_selecao, text="Selecionar Todos Visíveis", 
                  command=self.selecionar_todos_visiveis).pack(side='right', padx=2)
        ttk.Button(frame_selecao, text="Limpar Seleção", 
                  command=self.limpar_selecao).pack(side='right', padx=2)
        
        # Frame de resumo
        frame_resumo = ttk.LabelFrame(main_frame, text="Resumo")
        frame_resumo.pack(fill='x', pady=(10, 0))
        
        resumo_frame = ttk.Frame(frame_resumo)
        resumo_frame.pack(fill='x', padx=10, pady=8)
        
        self.lbl_total_lancamentos = ttk.Label(resumo_frame, text="Total de Lançamentos: 0", 
                                             font=('Arial', 10, 'bold'))
        self.lbl_total_lancamentos.pack(side='left', padx=10)
        
        self.lbl_valor_total = ttk.Label(resumo_frame, text="Valor Total: R$ 0,00", 
                                       font=('Arial', 10, 'bold'))
        self.lbl_valor_total.pack(side='left', padx=10)
        
        self.lbl_ultimo_lancamento = ttk.Label(resumo_frame, text="Último Lançamento: -", 
                                             font=('Arial', 10))
        self.lbl_ultimo_lancamento.pack(side='left', padx=10)
        
        # Frame de botões - ESTENDIDO COM FUNCIONALIDADES DE EDIÇÃO
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=(10, 0))
        
        # === BOTÕES DE AÇÃO INDIVIDUAL ===
        self.btn_editar = ttk.Button(frame_botoes, text="Editar", 
                                    command=self.editar_lancamento)
        self.btn_editar.pack(side='left', padx=5)
        
        self.btn_ver_historico = ttk.Button(frame_botoes, text="Ver Histórico", 
                                          command=self.visualizar_historico_lancamento)
        self.btn_ver_historico.pack(side='left', padx=5)
        
        # Separador visual
        ttk.Separator(frame_botoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # === BOTÕES DE EXCLUSÃO/RESTAURAÇÃO ===
        self.btn_excluir_individual = ttk.Button(frame_botoes, text="Excluir", 
                                                command=self.excluir_lancamento)
        self.btn_excluir_individual.pack(side='left', padx=2)
        
        self.btn_excluir_lote = ttk.Button(frame_botoes, text="Excluir Selecionados", 
                                         command=self.excluir_lote, state='disabled')
        self.btn_excluir_lote.pack(side='left', padx=2)
        
        self.btn_restaurar_individual = ttk.Button(frame_botoes, text="Restaurar", 
                                                 command=self.restaurar_lancamento)
        self.btn_restaurar_individual.pack(side='left', padx=2)
        
        self.btn_restaurar_lote = ttk.Button(frame_botoes, text="Restaurar Selecionados", 
                                           command=self.restaurar_lote, state='disabled')
        self.btn_restaurar_lote.pack(side='left', padx=2)
        
        # Separador visual
        ttk.Separator(frame_botoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # === BOTÕES GERAIS ===
        ttk.Button(frame_botoes, text="Exportar Lista", 
                  command=self.exportar_lista).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Atualizar", 
                  command=self.carregar_lancamentos).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Ver Estatísticas", 
                  command=self.mostrar_estatisticas).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes, text="Fechar", 
                  command=self.janela.destroy).pack(side='right', padx=5)
        
        # Configurar tags para cores
        self.tree_lancamentos.tag_configure('excluido', background='#ffcccc')
        self.tree_lancamentos.tag_configure('normal', background='white')
        self.tree_lancamentos.tag_configure('recente', background='#e8f5e8')
        
        # Configurar eventos
        self.configurar_eventos_selecao()
        self.configurar_atalhos()
        
        # Inicializar estado dos botões
        self.atualizar_interface_selecao()
        self.inicializar_datas_padrao()

    def configurar_eventos_selecao(self):
        """Configura eventos para controle de seleção múltipla"""
        try:
            def on_selection_change(event=None):
                self.atualizar_interface_selecao()
            
            self.tree_lancamentos.bind('<<TreeviewSelect>>', on_selection_change)
            self.tree_lancamentos.bind('<Double-1>', self.ver_detalhes_lancamento)
            
        except Exception as e:
            print(f"Erro ao configurar eventos de seleção: {str(e)}")

    def configurar_atalhos(self):
        """Configura atalhos de teclado"""
        try:
            # Duplo clique para ver detalhes
            self.tree_lancamentos.bind('<Double-1>', self.ver_detalhes_lancamento)
            
            # Tecla H para histórico
            def on_key_h(event):
                if self.tree_lancamentos.selection():
                    self.visualizar_historico_lancamento()
            
            self.janela.bind('<Key-h>', on_key_h)
            self.janela.bind('<Key-H>', on_key_h)
            
            # Delete para excluir
            def on_delete(event):
                items_selecionados = self.tree_lancamentos.selection()
                if len(items_selecionados) == 1:
                    self.excluir_lancamento()
                elif len(items_selecionados) > 1:
                    self.excluir_lote()
            
            self.janela.bind('<Delete>', on_delete)
            
            # Ctrl+A para selecionar todos
            def on_ctrl_a(event):
                self.selecionar_todos_visiveis()
                return "break"
            
            self.janela.bind('<Control-a>', on_ctrl_a)
            
            # Escape para limpar seleção
            def on_escape(event):
                self.limpar_selecao()
                return "break"
            
            self.janela.bind('<Escape>', on_escape)
            
            self.janela.focus_set()
            
        except Exception as e:
            print(f"Erro ao configurar atalhos: {str(e)}")

    def atualizar_interface_selecao(self):
        """Atualiza a interface baseada na seleção atual"""
        try:
            items_selecionados = self.tree_lancamentos.selection()
            qtd_selecionados = len(items_selecionados)
            
            # Atualizar label de seleção
            if qtd_selecionados == 0:
                self.label_selecao.config(text="Nenhum item selecionado")
            elif qtd_selecionados == 1:
                self.label_selecao.config(text="1 item selecionado")
            else:
                self.label_selecao.config(text=f"{qtd_selecionados} itens selecionados")
            
            # Controlar estado dos botões
            if qtd_selecionados == 0:
                # Nenhum selecionado - desabilitar todos
                self.btn_editar.config(state='disabled')
                self.btn_ver_historico.config(state='disabled')
                self.btn_excluir_individual.config(state='disabled')
                self.btn_excluir_lote.config(state='disabled')
                self.btn_restaurar_individual.config(state='disabled')
                self.btn_restaurar_lote.config(state='disabled')
                
            elif qtd_selecionados == 1:
                # Um selecionado - habilitar individuais
                self.btn_editar.config(state='normal')
                self.btn_ver_historico.config(state='normal')
                self.btn_excluir_individual.config(state='normal')
                self.btn_excluir_lote.config(state='disabled')
                self.btn_restaurar_individual.config(state='normal')
                self.btn_restaurar_lote.config(state='disabled')
                
            else:
                # Múltiplos selecionados - habilitar lote, desabilitar individuais
                self.btn_editar.config(state='disabled')
                self.btn_ver_historico.config(state='disabled')
                self.btn_excluir_individual.config(state='disabled')
                self.btn_excluir_lote.config(state='normal')
                self.btn_restaurar_individual.config(state='disabled')
                self.btn_restaurar_lote.config(state='normal')
            
        except Exception as e:
            print(f"Erro ao atualizar interface de seleção: {str(e)}")

    def selecionar_todos_visiveis(self):
        """Seleciona todos os itens visíveis na lista"""
        try:
            items_visiveis = self.tree_lancamentos.get_children()
            
            if not items_visiveis:
                custom_messagebox("info", "Seleção", "Nenhum item visível para selecionar")
                return
            
            self.tree_lancamentos.selection_set(items_visiveis)
            self.atualizar_interface_selecao()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao selecionar itens: {str(e)}")

    def limpar_selecao(self):
        """Limpa a seleção atual"""
        try:
            self.tree_lancamentos.selection_remove(self.tree_lancamentos.selection())
            self.atualizar_interface_selecao()
            
        except Exception as e:
            print(f"Erro ao limpar seleção: {str(e)}")

    def obter_dados_selecionados(self):
        """Obtém dados dos itens selecionados para processamento"""
        try:
            items_selecionados = self.tree_lancamentos.selection()
            
            if not items_selecionados:
                return []
            
            dados_selecionados = []
            
            for item in items_selecionados:
                valores = self.tree_lancamentos.item(item)['values']
                
                dados_item = {
                    'item_id': item,
                    'data': valores[0],
                    'tp_desp': valores[1],
                    'referencia': valores[2],
                    'nf': valores[3],
                    'valor': valores[4],
                    'vencimento': valores[5],
                    'status': valores[6],
                    'observacao': valores[7],
                    'id_lancamento': valores[8]
                }
                
                dados_selecionados.append(dados_item)
            
            return dados_selecionados
            
        except Exception as e:
            print(f"Erro ao obter dados selecionados: {str(e)}")
            return []

    def editar_lancamento(self):
        """Edita o lançamento selecionado usando o editor do GerenciadorLancamentos"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para editar")
            return
        
        try:
            valores = self.tree_lancamentos.item(item_selecionado[0])['values']
            id_lancamento = valores[8]  # ID do lançamento
            
            # Buscar dados completos do lançamento
            lancamento = None
            for _, row in self.dados_originais.iterrows():
                if row.get('ID_LANCAMENTO') == id_lancamento:
                    lancamento = row
                    break
            
            if lancamento is None:
                custom_messagebox("error", "Erro", f"Lançamento com ID {id_lancamento} não encontrado")
                return
            
            # Abrir editor usando a classe do GerenciadorLancamentos
            # from src.sistema_entrada_dados import EditorLancamentoCompleto
            editor = EditorLancamentoCompleto(self.janela, lancamento, self.salvar_edicao)
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao editar lançamento: {str(e)}")

    def salvar_edicao(self, id_lancamento, dados_editados):
        """Callback para salvar edições (usa lógica do GerenciadorLancamentos)"""
        try:
            # Importar e usar o método de salvamento do GerenciadorLancamentos
            from openpyxl import load_workbook
            from datetime import datetime
            import pandas as pd
            
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb["Dados"]
            
            # Encontrar a linha do lançamento
            linha_encontrada = None
            for row_num in range(2, ws.max_row + 1):
                id_na_planilha = ws.cell(row=row_num, column=15).value
                
                if str(id_na_planilha) == str(id_lancamento):
                    linha_encontrada = row_num
                    break
            
            if not linha_encontrada:
                wb.close()
                return False
            
            # Atualizar os dados na planilha (usar mesmo código do GerenciadorLancamentos)
            if dados_editados.get('data'):
                data_rel = datetime.strptime(dados_editados['data'], '%d/%m/%Y') if isinstance(dados_editados['data'], str) else dados_editados['data']
                ws.cell(row=linha_encontrada, column=1, value=data_rel)
                ws.cell(row=linha_encontrada, column=1).number_format = 'DD/MM/YYYY'
            
            if dados_editados.get('tp_desp'):
                ws.cell(row=linha_encontrada, column=2, value=int(dados_editados['tp_desp']))
            
            ws.cell(row=linha_encontrada, column=3, value=dados_editados.get('cnpj_cpf', ''))
            ws.cell(row=linha_encontrada, column=4, value=dados_editados.get('nome', ''))
            ws.cell(row=linha_encontrada, column=5, value=dados_editados.get('referencia', ''))
            ws.cell(row=linha_encontrada, column=6, value=dados_editados.get('nf', ''))
            
            if dados_editados.get('vr_unit'):
                vr_unit = float(dados_editados['vr_unit'])
                ws.cell(row=linha_encontrada, column=7, value=vr_unit)
                ws.cell(row=linha_encontrada, column=7).number_format = '#,##0.00'
            
            if dados_editados.get('dias'):
                ws.cell(row=linha_encontrada, column=8, value=int(dados_editados['dias']))
            
            if dados_editados.get('valor'):
                valor = float(dados_editados['valor'])
                ws.cell(row=linha_encontrada, column=9, value=valor)
                ws.cell(row=linha_encontrada, column=9).number_format = '#,##0.00'
            
            if dados_editados.get('dt_vencto'):
                dt_vencto = datetime.strptime(dados_editados['dt_vencto'], '%d/%m/%Y') if isinstance(dados_editados['dt_vencto'], str) else dados_editados['dt_vencto']
                ws.cell(row=linha_encontrada, column=10, value=dt_vencto)
                ws.cell(row=linha_encontrada, column=10).number_format = 'DD/MM/YYYY'
            
            ws.cell(row=linha_encontrada, column=11, value=dados_editados.get('categoria', ''))
            ws.cell(row=linha_encontrada, column=12, value=dados_editados.get('dados_bancarios', ''))
            
            # Observação com timestamp de edição
            observacao_editada = dados_editados.get('observacao', '')
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            observacao_com_historico = f"{observacao_editada} - EDITADO EM: {timestamp}"
            ws.cell(row=linha_encontrada, column=13, value=observacao_com_historico)
            
            # Adicionar aos campos novos se existirem
            if dados_editados.get('etapa_obra'):
                ws.cell(row=linha_encontrada, column=17, value=dados_editados['etapa_obra'])
            if dados_editados.get('insumo'):
                ws.cell(row=linha_encontrada, column=18, value=dados_editados['insumo'])
            
            wb.save(arquivo_cliente)
            wb.close()
            
            # Recarregar dados e atualizar visualização
            self.carregar_lancamentos()
            
            return True
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            print(f"Erro ao salvar edição: {str(e)}")
            return False

    def excluir_lancamento(self):
        """Exclui um lançamento individual"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para excluir")
            return
        
        valores = self.tree_lancamentos.item(item_selecionado[0])['values']
        nome_fornecedor = self.dados_fornecedor['nome']
        referencia = valores[2]
        valor = valores[4]
        data_lancamento = valores[0]
        status_atual = valores[6]
        
        if status_atual == 'EXCLUIDO':
            custom_messagebox("info", "Informação", "Este lançamento já está excluído")
            return
        
        if custom_messagebox("yesno", "Confirmação", 
                            f"Deseja realmente excluir este lançamento?\n\n"
                            f"Fornecedor: {nome_fornecedor}\n"
                            f"Referência: {referencia}\n"
                            f"Valor: {valor}\n"
                            f"Data: {data_lancamento}"):
            try:
                id_lancamento = valores[8]
                self.atualizar_status_lancamento(id_lancamento, 'EXCLUIDO')
                
                # Verificar recálculo de taxas se necessário
                self.verificar_recalculo_apos_alteracao(data_lancamento, "EXCLUSÃO")
                
                self.carregar_lancamentos()
                custom_messagebox("info", "Sucesso", "Lançamento excluído com sucesso!")
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao excluir lançamento: {str(e)}")

    def excluir_lote(self):
        """Exclui múltiplos lançamentos selecionados"""
        try:
            dados_selecionados = self.obter_dados_selecionados()
            
            if not dados_selecionados:
                custom_messagebox("warning", "Aviso", "Nenhum item selecionado para exclusão")
                return
            
            # Filtrar apenas lançamentos ativos
            ativos = [item for item in dados_selecionados if item['status'] != 'EXCLUIDO']
            
            if not ativos:
                custom_messagebox("info", "Informação", "Todos os itens selecionados já estão excluídos")
                return
            
            if custom_messagebox("yesno", "Confirmação de Exclusão em Lote", 
                                f"Deseja realmente excluir {len(ativos)} lançamento(s)?\n\n"
                                f"Fornecedor: {self.dados_fornecedor['nome']}\n"
                                f"Esta operação não pode ser desfeita facilmente."):
                
                sucesso = 0
                datas_afetadas = set()
                
                for item in ativos:
                    try:
                        self.atualizar_status_lancamento(item['id_lancamento'], 'EXCLUIDO')
                        # Coletar datas para verificação posterior
                        if item['data']:
                            data_obj = datetime.strptime(item['data'], '%d/%m/%Y').date()
                            datas_afetadas.add(data_obj)
                        sucesso += 1
                    except Exception as e:
                        print(f"Erro ao excluir item {item['id_lancamento']}: {str(e)}")
                        continue
                
                # Verificar recálculo para as datas afetadas
                for data_afetada in datas_afetadas:
                    self.verificar_recalculo_apos_alteracao(data_afetada.strftime('%d/%m/%Y'), "EXCLUSÃO")
                
                self.carregar_lancamentos()
                custom_messagebox("info", "Sucesso", f"{sucesso} lançamentos excluídos com sucesso!")
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro na exclusão em lote: {str(e)}")

    def restaurar_lancamento(self):
        """Restaura um lançamento individual"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para restaurar")
            return
        
        valores = self.tree_lancamentos.item(item_selecionado[0])['values']
        nome_fornecedor = self.dados_fornecedor['nome']
        referencia = valores[2]
        valor = valores[4]
        data_lancamento = valores[0]
        status_atual = valores[6]
        
        if status_atual != 'EXCLUIDO':
            custom_messagebox("info", "Informação", "Este lançamento já está ativo")
            return
        
        if custom_messagebox("yesno", "Confirmação", 
                            f"Deseja realmente restaurar este lançamento?\n\n"
                            f"Fornecedor: {nome_fornecedor}\n"
                            f"Referência: {referencia}\n"
                            f"Valor: {valor}\n"
                            f"Data: {data_lancamento}"):
            try:
                id_lancamento = valores[8]
                self.atualizar_status_lancamento(id_lancamento, 'ATIVO')
                
                # Verificar recálculo de taxas se necessário
                self.verificar_recalculo_apos_alteracao(data_lancamento, "RESTAURAÇÃO")
                
                self.carregar_lancamentos()
                custom_messagebox("info", "Sucesso", "Lançamento restaurado com sucesso!")
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao restaurar lançamento: {str(e)}")

    def restaurar_lote(self):
        """Restaura múltiplos lançamentos selecionados"""
        try:
            dados_selecionados = self.obter_dados_selecionados()
            
            if not dados_selecionados:
                custom_messagebox("warning", "Aviso", "Nenhum item selecionado para restauração")
                return
            
            # Filtrar apenas lançamentos excluídos
            excluidos = [item for item in dados_selecionados if item['status'] == 'EXCLUIDO']
            
            if not excluidos:
                custom_messagebox("info", "Informação", "Todos os itens selecionados já estão ativos")
                return
            
            if custom_messagebox("yesno", "Confirmação de Restauração em Lote", 
                                f"Deseja realmente restaurar {len(excluidos)} lançamento(s)?\n\n"
                                f"Fornecedor: {self.dados_fornecedor['nome']}"):
                
                sucesso = 0
                datas_afetadas = set()
                
                for item in excluidos:
                    try:
                        self.atualizar_status_lancamento(item['id_lancamento'], 'ATIVO')
                        # Coletar datas para verificação posterior
                        if item['data']:
                            data_obj = datetime.strptime(item['data'], '%d/%m/%Y').date()
                            datas_afetadas.add(data_obj)
                        sucesso += 1
                    except Exception as e:
                        print(f"Erro ao restaurar item {item['id_lancamento']}: {str(e)}")
                        continue
                
                # Verificar recálculo para as datas afetadas
                for data_afetada in datas_afetadas:
                    self.verificar_recalculo_apos_alteracao(data_afetada.strftime('%d/%m/%Y'), "RESTAURAÇÃO")
                
                self.carregar_lancamentos()
                custom_messagebox("info", "Sucesso", f"{sucesso} lançamentos restaurados com sucesso!")
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro na restauração em lote: {str(e)}")

    def atualizar_status_lancamento(self, id_lancamento, novo_status):
        """Atualiza o status de um lançamento específico"""
        try:
            from openpyxl import load_workbook
            
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            # Verificar se colunas existem
            if ws.cell(row=1, column=14).value != 'STATUS':
                ws.cell(row=1, column=14, value='STATUS')
            if ws.cell(row=1, column=16).value != 'HISTORICO_ALTERACAO':
                ws.cell(row=1, column=16, value='HISTORICO_ALTERACAO')
            
            # Encontrar e atualizar linha
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=15).value == id_lancamento:
                    # Atualizar status
                    ws.cell(row=row, column=14, value=novo_status)
                    
                    # Adicionar ao histórico
                    timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    historico_atual = ws.cell(row=row, column=16).value or ""
                    
                    if novo_status == 'EXCLUIDO':
                        acao = f"EXCLUÍDO EM: {timestamp}"
                    elif novo_status == 'ATIVO':
                        acao = f"RESTAURADO EM: {timestamp}"
                    else:
                        acao = f"STATUS ALTERADO PARA {novo_status} EM: {timestamp}"
                    
                    if historico_atual:
                        novo_historico = f"{historico_atual} | {acao}"
                    else:
                        novo_historico = acao
                        
                    ws.cell(row=row, column=16, value=novo_historico)
                    break
            
            wb.save(arquivo_cliente)
            wb.close()
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Erro ao atualizar status: {str(e)}")

    def visualizar_historico_lancamento(self):
        """Visualiza o histórico de alterações de um lançamento"""
        try:
            selected_items = self.tree_lancamentos.selection()
            if not selected_items:
                custom_messagebox("info", "Seleção", "Selecione um lançamento para ver o histórico!")
                return
            
            item = selected_items[0]
            valores = self.tree_lancamentos.item(item, 'values')
            id_lancamento = valores[8]
            
            # Buscar dados completos do lançamento
            if not hasattr(self, 'dados_originais') or self.dados_originais.empty:
                custom_messagebox("error", "Erro", "Dados não carregados. Clique em 'Atualizar' primeiro!")
                return
            
            try:
                id_busca = int(float(str(id_lancamento)))
            except (ValueError, TypeError):
                custom_messagebox("error", "Erro", f"ID inválido: {id_lancamento}")
                return
            
            filtro = self.dados_originais['ID_LANCAMENTO'] == id_busca
            lancamentos_encontrados = self.dados_originais[filtro]
            
            if lancamentos_encontrados.empty:
                custom_messagebox("error", "Erro", f"Lançamento com ID {id_busca} não encontrado!")
                return
            
            lancamento = lancamentos_encontrados.iloc[0]
            
            # Criar janela de histórico
            janela_historico = tk.Toplevel(self.janela)
            janela_historico.title(f"Histórico do Lançamento - ID {id_busca}")
            janela_historico.geometry("700x400")
            janela_historico.transient(self.janela)
            janela_historico.grab_set()
            
            frame_principal = ttk.Frame(janela_historico, padding="10")
            frame_principal.pack(fill='both', expand=True)
            
            # Informações do lançamento
            frame_info = ttk.LabelFrame(frame_principal, text="Informações do Lançamento")
            frame_info.pack(fill='x', pady=(0, 10))
            
            info_frame_interno = ttk.Frame(frame_info)
            info_frame_interno.pack(fill='x', padx=10, pady=5)
            
            frame_esq = ttk.Frame(info_frame_interno)
            frame_esq.pack(side='left', fill='x', expand=True)
            
            ttk.Label(frame_esq, text=f"ID: {lancamento['ID_LANCAMENTO']}", font=('TkDefaultFont', 9, 'bold')).pack(anchor='w')
            ttk.Label(frame_esq, text=f"Fornecedor: {self.dados_fornecedor['nome']}", font=('TkDefaultFont', 9)).pack(anchor='w')
            ttk.Label(frame_esq, text=f"Referência: {lancamento['REFERÊNCIA']}", font=('TkDefaultFont', 9)).pack(anchor='w')
            
            frame_dir = ttk.Frame(info_frame_interno)
            frame_dir.pack(side='right', fill='x', expand=True)
            
            ttk.Label(frame_dir, text=f"Valor: R$ {lancamento['VALOR']:,.2f}", font=('TkDefaultFont', 9)).pack(anchor='w')
            ttk.Label(frame_dir, text=f"Status: {lancamento['STATUS']}", font=('TkDefaultFont', 9)).pack(anchor='w')
            
            # Histórico
            frame_historico = ttk.LabelFrame(frame_principal, text="Histórico de Alterações")
            frame_historico.pack(fill='x', pady=(0, 10))
            
            text_historico = tk.Text(frame_historico, wrap='word', font=('Consolas', 9), height=12)
            scrollbar_hist = ttk.Scrollbar(frame_historico, orient='vertical', command=text_historico.yview)
            text_historico.configure(yscrollcommand=scrollbar_hist.set)
            
            historico = lancamento.get('HISTORICO_ALTERACAO', '')
            if historico and str(historico) not in ['', 'nan', 'None']:
                text_historico.insert('1.0', str(historico))
            else:
                text_historico.insert('1.0', "Nenhum histórico de alterações registrado.")
            
            text_historico.config(state='disabled')
            
            text_historico.pack(side='left', fill='both', expand=True, padx=(10, 0), pady=5)
            scrollbar_hist.pack(side='right', fill='y', pady=5)
            
            # Botão fechar
            frame_botoes = ttk.Frame(frame_principal)
            frame_botoes.pack(fill='x', pady=(5, 0))
            
            ttk.Button(frame_botoes, text="Fechar", command=janela_historico.destroy).pack(side='right')
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao visualizar histórico: {str(e)}")

    def carregar_lancamentos(self):
        """Carrega os lançamentos do fornecedor - VERSÃO CORRIGIDA"""
        try:
            from openpyxl import load_workbook
            import pandas as pd
            
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            
            # Carregar dados preservando tipos
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados', dtype={'CNPJ_CPF': str})
            df = df.fillna("")
            
            # Adicionar coluna de status se não existir
            if 'STATUS' not in df.columns:
                df['STATUS'] = 'ATIVO'
            df['STATUS'] = df['STATUS'].replace('', 'ATIVO').fillna('ATIVO')
            
            # Adicionar ID_LANCAMENTO se não existir
            if 'ID_LANCAMENTO' not in df.columns:
                df['ID_LANCAMENTO'] = range(1, len(df) + 1)
            
            # Converter ID_LANCAMENTO para int
            df['ID_LANCAMENTO'] = pd.to_numeric(df['ID_LANCAMENTO'], errors='coerce').fillna(0).astype(int)
            
            # === BUSCA CORRIGIDA POR FORNECEDOR ===
            
            # Normalizar CNPJ/CPF do fornecedor buscado
            cnpj_cpf_original = self.dados_fornecedor['cnpj_cpf']
            cnpj_cpf_str = str(cnpj_cpf_original)
            cnpj_cpf_numeros = ''.join(filter(str.isdigit, cnpj_cpf_str))
            
            if len(cnpj_cpf_numeros) <= 11:
                cnpj_cpf_normalizado = cnpj_cpf_numeros.zfill(11)
            else:
                cnpj_cpf_normalizado = cnpj_cpf_numeros.zfill(14)
            
            print(f"DEBUG: Buscando fornecedor:")
            print(f"       Nome: {self.dados_fornecedor['nome']}")
            print(f"       CNPJ/CPF normalizado: {cnpj_cpf_normalizado}")
            
            # Função para normalizar CNPJ/CPF da planilha
            def normalizar_cnpj_cpf_planilha(valor):
                if pd.isna(valor) or valor == '':
                    return ''
                numeros = ''.join(filter(str.isdigit, str(valor)))
                if len(numeros) <= 11:
                    return numeros.zfill(11)
                else:
                    return numeros.zfill(14)
            
            df['CNPJ_CPF_NORMALIZADO'] = df['CNPJ_CPF'].apply(normalizar_cnpj_cpf_planilha)
            
            # === BUSCA PRINCIPAL: Por CNPJ/CPF exato ===
            mask_cnpj = df['CNPJ_CPF_NORMALIZADO'] == cnpj_cpf_normalizado
            lancamentos_por_cnpj = df[mask_cnpj]
            
            print(f"DEBUG: Encontrados {len(lancamentos_por_cnpj)} lançamentos por CNPJ/CPF")
            
            # === BUSCA SECUNDÁRIA: Por nome EXATO (apenas se não encontrou por CNPJ) ===
            if lancamentos_por_cnpj.empty:
                nome_fornecedor_normalizado = str(self.dados_fornecedor['nome']).upper().strip()
                
                # CORREÇÃO: Busca por nome EXATO, não parcial
                mask_nome_exato = (
                    df['NOME'].astype(str).str.upper().str.strip() == nome_fornecedor_normalizado
                )
                lancamentos_por_nome = df[mask_nome_exato]
                
                print(f"DEBUG: Encontrados {len(lancamentos_por_nome)} lançamentos por nome exato")
                
                # === VALIDAÇÃO CRUZADA (NOVO) ===
                # Se encontrou por nome, verificar se o CNPJ/CPF bate
                if not lancamentos_por_nome.empty:
                    cnpj_encontrado = lancamentos_por_nome.iloc[0]['CNPJ_CPF_NORMALIZADO']
                    
                    if cnpj_encontrado and cnpj_encontrado != cnpj_cpf_normalizado:
                        print(f"AVISO: Nome encontrado mas CNPJ/CPF não confere!")
                        print(f"       Esperado: {cnpj_cpf_normalizado}")
                        print(f"       Encontrado: {cnpj_encontrado}")
                        
                        # Mostrar aviso ao usuário
                        custom_messagebox("warning", "Divergência de Dados", 
                                        f"ATENÇÃO: Encontrado fornecedor com nome igual mas CNPJ/CPF diferente:\n\n"
                                        f"Nome buscado: {self.dados_fornecedor['nome']}\n"
                                        f"CNPJ/CPF buscado: {self.dados_fornecedor['cnpj_cpf_formatado']}\n"
                                        f"CNPJ/CPF encontrado na base: {self.formatar_cnpj_cpf(cnpj_encontrado)}\n\n"
                                        f"Os dados podem estar inconsistentes. Verifique o cadastro do fornecedor.")
                
                self.df_fornecedor = lancamentos_por_nome.copy()
            else:
                self.df_fornecedor = lancamentos_por_cnpj.copy()
            
            # === VALIDAÇÃO FINAL (NOVO) ===
            if self.df_fornecedor.empty:
                print(f"DEBUG: Nenhum lançamento encontrado para o fornecedor")
                
                # Mostrar mensagem clara ao usuário
                custom_messagebox("info", "Nenhum Lançamento Encontrado", 
                                f"Não foram encontrados lançamentos para:\n\n"
                                f"Fornecedor: {self.dados_fornecedor['nome']}\n"
                                f"CNPJ/CPF: {self.dados_fornecedor['cnpj_cpf_formatado']}\n"
                                f"Cliente: {self.sistema.cliente_atual}\n\n"
                                f"Verifique se:\n"
                                f"• O fornecedor possui lançamentos neste cliente\n"
                                f"• Os dados do fornecedor estão corretos\n"
                                f"• O período de busca está adequado")
            else:
                # Verificar se encontrou o fornecedor correto
                primeiro_lancamento = self.df_fornecedor.iloc[0]
                nome_encontrado = str(primeiro_lancamento['NOME']).upper().strip()
                cnpj_encontrado = primeiro_lancamento['CNPJ_CPF_NORMALIZADO']
                
                print(f"DEBUG: Fornecedor encontrado:")
                print(f"       Nome na base: {nome_encontrado}")
                print(f"       CNPJ/CPF na base: {cnpj_encontrado}")
                print(f"       Total de lançamentos: {len(self.df_fornecedor)}")
                
                # Validação adicional: verificar se realmente é o fornecedor correto
                nome_buscado = str(self.dados_fornecedor['nome']).upper().strip()
                
                if (nome_encontrado != nome_buscado and 
                    cnpj_encontrado != cnpj_cpf_normalizado):
                    
                    # ERRO CRÍTICO: dados não conferem
                    print(f"ERRO: Dados não conferem!")
                    custom_messagebox("error", "Erro Crítico", 
                                    f"ERRO: Os dados encontrados não conferem com o fornecedor buscado!\n\n"
                                    f"BUSCADO:\n"
                                    f"Nome: {self.dados_fornecedor['nome']}\n"
                                    f"CNPJ/CPF: {self.dados_fornecedor['cnpj_cpf_formatado']}\n\n"
                                    f"ENCONTRADO:\n"
                                    f"Nome: {primeiro_lancamento['NOME']}\n"
                                    f"CNPJ/CPF: {self.formatar_cnpj_cpf(cnpj_encontrado)}\n\n"
                                    f"Possível erro no sistema. Contate o suporte técnico.")
                    
                    # Limpar dados para evitar mostrar informações incorretas
                    self.df_fornecedor = pd.DataFrame()
            
            # Salvar dados originais para uso posterior (apenas se houver dados válidos)
            if not self.df_fornecedor.empty:
                self.dados_originais = self.df_fornecedor.copy()
                
                # Ordenar por data (mais recente primeiro)
                self.df_fornecedor['DATA_REL'] = pd.to_datetime(self.df_fornecedor['DATA_REL'], errors='coerce')
                self.df_fornecedor = self.df_fornecedor.sort_values('DATA_REL', ascending=False)
            else:
                self.dados_originais = pd.DataFrame()
            
            # Aplicar filtros iniciais
            self.aplicar_filtros()
            
        except Exception as e:
            import traceback
            print(f"Erro ao carregar lançamentos: {traceback.format_exc()}")
            custom_messagebox("error", "Erro", f"Erro ao carregar lançamentos: {str(e)}")

    def debug_busca_fornecedor(self, cnpj_cpf_fornecedor, nome_fornecedor):
        """Método auxiliar para debug da busca por fornecedor"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados', dtype={'CNPJ_CPF': str})
            
            print("=== DEBUG BUSCA POR FORNECEDOR ===")
            print(f"Fornecedor buscado: {nome_fornecedor}")
            print(f"CNPJ/CPF buscado: {cnpj_cpf_fornecedor}")
            
            # Mostrar todos os fornecedores únicos na base
            fornecedores_unicos = df[['NOME', 'CNPJ_CPF']].drop_duplicates()
            print(f"\nFornecedores na base do cliente {self.sistema.cliente_atual}:")
            print(f"Total: {len(fornecedores_unicos)}")
            
            for _, row in fornecedores_unicos.head(10).iterrows():  # Mostrar apenas os primeiros 10
                nome_base = str(row['NOME']).strip()
                cnpj_base = str(row['CNPJ_CPF']).strip()
                print(f"  - {nome_base} | {cnpj_base}")
            
            if len(fornecedores_unicos) > 10:
                print(f"  ... e mais {len(fornecedores_unicos) - 10} fornecedores")
            
            # Buscar fornecedores com nomes similares
            nome_busca = str(nome_fornecedor).upper()
            nomes_similares = fornecedores_unicos[
                fornecedores_unicos['NOME'].astype(str).str.upper().str.contains('ANTONIO', na=False)
            ]
            
            if not nomes_similares.empty:
                print(f"\nFornecedores com 'ANTONIO' no nome:")
                for _, row in nomes_similares.iterrows():
                    print(f"  - {row['NOME']} | {row['CNPJ_CPF']}")
            
            return df, fornecedores_unicos
            
        except Exception as e:
            print(f"Erro no debug: {str(e)}")
            return None, None

    def inicializar_datas_padrao(self):
        """Inicializa as datas padrão dos filtros baseado no sistema (dias 5 e 20)"""
        try:
            from datetime import datetime, timedelta
            from calendar import monthrange
            
            # Data de hoje
            hoje = datetime.now().date()
            dia_atual = hoje.day
            mes_atual = hoje.month
            ano_atual = hoje.year
            
            # LÓGICA DO SISTEMA: Data fim baseada nos dias 5 e 20
            if dia_atual <= 5:
                # Do dia 1 ao 5: data fim = dia 5 do mês atual
                data_fim_padrao = hoje.replace(day=5)
            elif dia_atual <= 20:
                # Do dia 6 ao 20: data fim = dia 20 do mês atual
                data_fim_padrao = hoje.replace(day=20)
            else:
                # Do dia 21 em diante: data fim = dia 5 do próximo mês
                if mes_atual == 12:
                    # Se dezembro, vai para janeiro do próximo ano
                    data_fim_padrao = datetime(ano_atual + 1, 1, 5).date()
                else:
                    # Senão, próximo mês do mesmo ano
                    data_fim_padrao = datetime(ano_atual, mes_atual + 1, 5).date()
            
            # Data de início: 30 dias antes da data fim (mais lógico para o sistema)
            data_inicio_padrao = data_fim_padrao - timedelta(days=185)  # Aproximadamente 6 meses
            
            # Definir as datas nos controles
            self.data_inicio.set_date(data_inicio_padrao)
            self.data_fim.set_date(data_fim_padrao)
            
            print(f"DEBUG: Datas padrão definidas (sistema dias 5/20):")
            print(f"       Hoje: {hoje} (dia {dia_atual})")
            print(f"       Data início: {data_inicio_padrao}")
            print(f"       Data fim: {data_fim_padrao}")
            
        except Exception as e:
            print(f"DEBUG: Erro ao inicializar datas padrão: {str(e)}")
            import traceback
            traceback.print_exc()
            # Fallback para comportamento anterior
            try:
                from dateutil.relativedelta import relativedelta
                data_padrao = datetime.now() - relativedelta(months=6)
                self.data_inicio.set_date(data_padrao.date())
                self.data_fim.set_date(datetime.now().date())
            except:
                pass        
    def aplicar_filtros(self):
        """Aplica os filtros selecionados"""
        try:
            # Limpar tree
            for item in self.tree_lancamentos.get_children():
                self.tree_lancamentos.delete(item)
                
            if self.df_fornecedor.empty:
                self.atualizar_resumo(pd.DataFrame())
                return
            
            df_filtrado = self.df_fornecedor.copy()
            
            # Filtro por período
            data_inicio = self.data_inicio.get_date()
            data_fim = self.data_fim.get_date()
            
            df_filtrado = df_filtrado[
                (df_filtrado['DATA_REL'].dt.date >= data_inicio) &
                (df_filtrado['DATA_REL'].dt.date <= data_fim)
            ]
            
            # Filtro por status
            status_filtro = self.combo_status.get()
            if status_filtro == 'Ativos':
                df_filtrado = df_filtrado[df_filtrado['STATUS'] != 'EXCLUIDO']
            elif status_filtro == 'Excluídos':
                df_filtrado = df_filtrado[df_filtrado['STATUS'] == 'EXCLUIDO']
            
            # Filtro por referência
            referencia_filtro = self.filtro_referencia.get().strip().upper()
            if referencia_filtro:
                df_filtrado = df_filtrado[
                    df_filtrado['REFERÊNCIA'].astype(str).str.upper().str.contains(referencia_filtro, na=False)
                ]
            
            # Preencher tree
            hoje = datetime.now().date()
            limite_recente = hoje - relativedelta(days=30)
            
            for idx, row in df_filtrado.iterrows():
                status = row.get('STATUS', 'ATIVO')
                tag = 'excluido' if status == 'EXCLUIDO' else 'normal'
                
                # Marcar lançamentos recentes
                if pd.notna(row['DATA_REL']) and row['DATA_REL'].date() >= limite_recente:
                    tag = 'recente'
                
                valores = (
                    self.formatar_data(row['DATA_REL']),
                    self.formatar_tipo_despesa(row['TP_DESP']),
                    row.get('REFERÊNCIA', ''),
                    row.get('NF', ''),
                    self.formatar_valor(row['VALOR']),
                    self.formatar_data(row['DT_VENCTO']),
                    status,
                    row.get('OBSERVAÇÃO', ''),
                    row.get('ID_LANCAMENTO', 0)  # ID oculto
                )
                
                self.tree_lancamentos.insert('', 'end', values=valores, tags=(tag,))
            
            # Atualizar resumo
            self.atualizar_resumo(df_filtrado)
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao aplicar filtros: {str(e)}")
            
    def busca_incremental(self, event=None):
        """Busca incremental conforme o usuário digita"""
        termo = self.busca_rapida.get().strip().upper()
        
        if not termo:
            self.aplicar_filtros()
            return
            
        # Filtrar dados
        df_busca = self.df_fornecedor.copy()
        
        # Aplicar filtros básicos primeiro
        data_inicio = self.data_inicio.get_date()
        data_fim = self.data_fim.get_date()
        
        df_busca = df_busca[
            (df_busca['DATA_REL'].dt.date >= data_inicio) &
            (df_busca['DATA_REL'].dt.date <= data_fim)
        ]
        
        # Buscar em NF, Observação ou Referência
        mask_busca = (
            df_busca['NF'].astype(str).str.upper().str.contains(termo, na=False) |
            df_busca['OBSERVAÇÃO'].astype(str).str.upper().str.contains(termo, na=False) |
            df_busca['REFERÊNCIA'].astype(str).str.upper().str.contains(termo, na=False)
        )
        
        df_resultado = df_busca[mask_busca]
        
        # Limpar e preencher tree
        for item in self.tree_lancamentos.get_children():
            self.tree_lancamentos.delete(item)
            
        for idx, row in df_resultado.iterrows():
            status = row.get('STATUS', 'ATIVO')
            tag = 'excluido' if status == 'EXCLUIDO' else 'normal'
            
            valores = (
                self.formatar_data(row['DATA_REL']),
                self.formatar_tipo_despesa(row['TP_DESP']),
                row.get('REFERÊNCIA', ''),
                row.get('NF', ''),
                self.formatar_valor(row['VALOR']),
                self.formatar_data(row['DT_VENCTO']),
                status,
                row.get('OBSERVAÇÃO', ''),
                row.get('ID_LANCAMENTO', 0)
            )
            
            self.tree_lancamentos.insert('', 'end', values=valores, tags=(tag,))
        
        self.atualizar_resumo(df_resultado)
        
    def limpar_filtros(self):
        """Limpa todos os filtros e recarrega"""
        self.inicializar_datas_padrao()
    
        self.filtro_referencia.delete(0, tk.END)
        self.combo_status.set('Ativos')
        self.busca_rapida.delete(0, tk.END)
        
        self.aplicar_filtros()
        
    def atualizar_resumo(self, df_filtrado):
        """Atualiza o resumo com estatísticas"""
        if df_filtrado.empty:
            self.lbl_total_lancamentos.config(text="Total de Lançamentos: 0")
            self.lbl_valor_total.config(text="Valor Total: R$ 0,00")
            self.lbl_ultimo_lancamento.config(text="Último Lançamento: -")
            return
            
        total_lancamentos = len(df_filtrado)
        df_ativos = df_filtrado[df_filtrado['STATUS'] != 'EXCLUIDO']
        valor_total = 0
        
        for _, row in df_ativos.iterrows():
            try:
                valor_raw = row['VALOR']
                if isinstance(valor_raw, str):
                    valor_limpo = valor_raw.replace('R$', '').replace(' ', '').replace(',', '.').strip()
                    if valor_limpo:
                        valor = float(valor_limpo)
                    else:
                        valor = 0
                elif isinstance(valor_raw, (int, float)):
                    valor = float(valor_raw)
                else:
                    valor = float(str(valor_raw).replace(',', '.') if str(valor_raw) else 0)
                
                valor_total += valor
            except (ValueError, TypeError, AttributeError):
                continue
        
        # Último lançamento
        if not df_filtrado.empty:
            ultimo = df_filtrado.iloc[0]
            data_ultimo = self.formatar_data(ultimo['DATA_REL'])
            ref_ultimo = ultimo.get('REFERÊNCIA', '')[:20] + ('...' if len(str(ultimo.get('REFERÊNCIA', ''))) > 20 else '')
            ultimo_texto = f"{data_ultimo} - {ref_ultimo}"
        else:
            ultimo_texto = "-"
        
        self.lbl_total_lancamentos.config(text=f"Total de Lançamentos: {total_lancamentos}")
        self.lbl_valor_total.config(text=f"Valor Total: R$ {valor_total:,.2f}")
        self.lbl_ultimo_lancamento.config(text=f"Último Lançamento: {ultimo_texto}")
        
    def ver_detalhes_lancamento(self, event=None):
        """Mostra detalhes completos do lançamento selecionado"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            return
            
        valores = self.tree_lancamentos.item(item_selecionado[0])['values']
        
        janela_detalhes = tk.Toplevel(self.janela)
        janela_detalhes.title("Detalhes do Lançamento")
        janela_detalhes.geometry("600x450")
        janela_detalhes.transient(self.janela)
        
        frame = ttk.Frame(janela_detalhes, padding="15")
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text="Detalhes Completos do Lançamento", 
                 font=('Arial', 14, 'bold')).pack(pady=(0, 15))
        
        info_frame = ttk.Frame(frame)
        info_frame.pack(fill='both', expand=True)
        
        detalhes = [
            ("Fornecedor:", self.dados_fornecedor['nome']),
            ("CNPJ/CPF:", self.dados_fornecedor['cnpj_cpf_formatado']),
            ("Data do Relatório:", valores[0]),
            ("Tipo de Despesa:", valores[1]),
            ("Referência:", valores[2]),
            ("Número da NF:", valores[3]),
            ("Valor:", valores[4]),
            ("Data de Vencimento:", valores[5]),
            ("Status:", valores[6]),
            ("Observação:", valores[7])
        ]
        
        for i, (label, valor) in enumerate(detalhes):
            ttk.Label(info_frame, text=label, font=('Arial', 10, 'bold')).grid(
                row=i, column=0, padx=5, pady=5, sticky='w')
            ttk.Label(info_frame, text=str(valor), font=('Arial', 10)).grid(
                row=i, column=1, padx=15, pady=5, sticky='w')
        
        ttk.Button(frame, text="Fechar", command=janela_detalhes.destroy).pack(pady=15)
        
    def mostrar_estatisticas(self):
        """Mostra estatísticas detalhadas do fornecedor"""
        if self.df_fornecedor.empty:
            custom_messagebox("info", "Estatísticas", "Nenhum lançamento encontrado para este fornecedor.")
            return
            
        janela_stats = tk.Toplevel(self.janela)
        janela_stats.title(f"Estatísticas - {self.dados_fornecedor['nome']}")
        janela_stats.geometry("700x500")
        janela_stats.transient(self.janela)
        
        frame = ttk.Frame(janela_stats, padding="15")
        frame.pack(fill='both', expand=True)
        
        # Calcular estatísticas
        df_ativos = self.df_fornecedor[self.df_fornecedor['STATUS'] != 'EXCLUIDO']
        
        # Por período
        hoje = datetime.now().date()
        ultimos_30 = hoje - relativedelta(days=30)
        ultimos_90 = hoje - relativedelta(days=90)
        ultimo_ano = hoje - relativedelta(years=1)
        
        stats_30 = df_ativos[df_ativos['DATA_REL'].dt.date >= ultimos_30]
        stats_90 = df_ativos[df_ativos['DATA_REL'].dt.date >= ultimos_90]
        stats_ano = df_ativos[df_ativos['DATA_REL'].dt.date >= ultimo_ano]
        
        # Por referência
        ref_counts = df_ativos['REFERÊNCIA'].value_counts().head(5)
        
        # Criar interface
        ttk.Label(frame, text="Estatísticas do Fornecedor", 
                 font=('Arial', 16, 'bold')).pack(pady=(0, 20))
        
        # Notebook para organizar
        notebook = ttk.Notebook(frame)
        notebook.pack(fill='both', expand=True)
        
        # Aba Geral
        aba_geral = ttk.Frame(notebook)
        notebook.add(aba_geral, text="Resumo Geral")
        
        stats_frame = ttk.Frame(aba_geral, padding="10")
        stats_frame.pack(fill='both', expand=True)
        
        stats_info = [
            ("Total de Lançamentos:", len(df_ativos)),
            ("Últimos 30 dias:", len(stats_30)),
            ("Últimos 90 dias:", len(stats_90)),
            ("Último ano:", len(stats_ano)),
            ("", ""),
            ("Valor Total (histórico):", f"R$ {df_ativos['VALOR'].astype(float).sum():,.2f}"),
            ("Valor (últimos 30 dias):", f"R$ {stats_30['VALOR'].astype(float).sum():,.2f}"),
            ("Valor médio por lançamento:", f"R$ {df_ativos['VALOR'].astype(float).mean():,.2f}"),
        ]
        
        for i, (label, valor) in enumerate(stats_info):
            if label:  # Pular linhas vazias
                ttk.Label(stats_frame, text=label, font=('Arial', 11, 'bold')).grid(
                    row=i, column=0, padx=10, pady=5, sticky='w')
                ttk.Label(stats_frame, text=str(valor), font=('Arial', 11)).grid(
                    row=i, column=1, padx=20, pady=5, sticky='w')
        
        # Aba Referências
        aba_ref = ttk.Frame(notebook)
        notebook.add(aba_ref, text="Por Referência")
        
        ref_frame = ttk.Frame(aba_ref, padding="10")
        ref_frame.pack(fill='both', expand=True)
        
        ttk.Label(ref_frame, text="Top 5 Referências:", 
                 font=('Arial', 12, 'bold')).pack(pady=(0, 10))
        
        for ref, count in ref_counts.items():
            valor_ref = df_ativos[df_ativos['REFERÊNCIA'] == ref]['VALOR'].astype(float).sum()
            ttk.Label(ref_frame, 
                     text=f"{ref}: {count} lançamentos (R$ {valor_ref:,.2f})",
                     font=('Arial', 10)).pack(anchor='w', pady=2)
        
        ttk.Button(frame, text="Fechar", command=janela_stats.destroy).pack(pady=15)
        
    def exportar_lista(self):
        """Exporta a lista atual para Excel"""
        try:
            from tkinter import filedialog
            
            # Obter dados visíveis
            dados_exportar = []
            for item in self.tree_lancamentos.get_children():
                valores = self.tree_lancamentos.item(item)['values']
                # Remover ID (último elemento) da exportação
                dados_exportar.append(valores[:-1])
            
            if not dados_exportar:
                custom_messagebox("warning", "Aviso", "Nenhum dado para exportar!")
                return
            
            # Solicitar arquivo
            arquivo = filedialog.asksaveasfilename(
                title="Salvar Lista de Lançamentos",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"lancamentos_{self.dados_fornecedor['nome'].replace(' ', '_')}.xlsx"
            )
            
            if arquivo:
                # Criar DataFrame
                colunas = ['Data Rel.', 'Tipo', 'Referência', 'NF', 'Valor', 'Vencimento', 'Status', 'Observação']
                df_export = pd.DataFrame(dados_exportar, columns=colunas)
                
                # Criar workbook com informações detalhadas
                with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
                    # Aba principal com dados
                    df_export.to_excel(writer, sheet_name='Lançamentos', index=False, startrow=6)
                    
                    # Obter worksheet para adicionar cabeçalho
                    worksheet = writer.sheets['Lançamentos']
                    
                    # Adicionar cabeçalho com informações
                    worksheet['A1'] = f"LANÇAMENTOS DO FORNECEDOR: {self.dados_fornecedor['nome']}"
                    worksheet['A2'] = f"CNPJ/CPF: {self.dados_fornecedor['cnpj_cpf_formatado']}"
                    worksheet['A3'] = f"Cliente: {self.sistema.cliente_atual}"
                    worksheet['A4'] = f"Período: {self.data_inicio.get()} até {self.data_fim.get()}"
                    worksheet['A5'] = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                    
                    # Formatação do cabeçalho
                    from openpyxl.styles import Font
                    
                    title_font = Font(bold=True, size=14)
                    info_font = Font(bold=True, size=10)
                    
                    worksheet['A1'].font = title_font
                    for row in range(2, 6):
                        worksheet[f'A{row}'].font = info_font
                    
                    # Ajustar largura das colunas
                    column_widths = {'A': 12, 'B': 8, 'C': 25, 'D': 15, 'E': 12, 'F': 12, 'G': 10, 'H': 30}
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width
                    
                    # Criar aba de resumo se houver dados
                    if not self.df_fornecedor.empty:
                        resumo_data = []
                        
                        # Estatísticas básicas
                        df_ativos = self.df_fornecedor[self.df_fornecedor['STATUS'] != 'EXCLUIDO']
                        resumo_data.append(['Estatística', 'Valor'])
                        resumo_data.append(['Total de Lançamentos', len(df_ativos)])
                        resumo_data.append(['Valor Total', f"R$ {df_ativos['VALOR'].astype(float).sum():,.2f}"])
                        resumo_data.append(['Valor Médio', f"R$ {df_ativos['VALOR'].astype(float).mean():,.2f}"])
                        resumo_data.append(['Primeiro Lançamento', df_ativos['DATA_REL'].min().strftime('%d/%m/%Y') if not df_ativos.empty else 'N/A'])
                        resumo_data.append(['Último Lançamento', df_ativos['DATA_REL'].max().strftime('%d/%m/%Y') if not df_ativos.empty else 'N/A'])
                        
                        # Por referência
                        resumo_data.append(['', ''])
                        resumo_data.append(['TOP REFERÊNCIAS', 'QUANTIDADE'])
                        ref_counts = df_ativos['REFERÊNCIA'].value_counts().head(5)
                        for ref, count in ref_counts.items():
                            resumo_data.append([ref, count])
                        
                        df_resumo = pd.DataFrame(resumo_data)
                        df_resumo.to_excel(writer, sheet_name='Resumo', index=False, header=False)
                
                custom_messagebox("info", "Sucesso", f"Lista exportada com sucesso!\n\nArquivo: {arquivo}")
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao exportar: {str(e)}")
    
    def formatar_data(self, data):
        """Formata data para exibição"""
        if pd.isna(data) or data == "":
            return ""
        try:
            if isinstance(data, str):
                return data
            return data.strftime('%d/%m/%Y')
        except:
            return str(data)
    
    def formatar_valor(self, valor):
        """Formata valor para exibição"""
        try:
            if isinstance(valor, str):
                valor_limpo = valor_raw.replace('R$', '').replace(' ', '').replace(',', '.').strip()
                if valor_limpo:
                    valor = float(valor_limpo.replace(',', '.'))
                else:
                    return "0,00"
            elif isinstance(valor, (int, float)):
                valor = float(valor)
            else:
                valor = float(str(valor).replace(',', '.') if str(valor) else 0)
            
            return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except (ValueError, TypeError, AttributeError):
            return "0,00"
    
    def formatar_tipo_despesa(self, tp_desp):
        """Formata tipo de despesa como inteiro"""
        try:
            if pd.isna(tp_desp) or tp_desp == "":
                return ""
            valor_numerico = float(tp_desp)
            return str(int(valor_numerico))
        except (ValueError, TypeError):
            return str(tp_desp)

class ComboboxAutocompletar(ttk.Combobox):
    """Combobox personalizada com funcionalidade de autocompletar e adição dinâmica"""
    
    def __init__(self, parent, values=None, config_key=None, config_manager=None, **kwargs):
        """
        Inicializa o Combobox com autocompletar
        
        Args:
            parent: Widget pai
            values: Lista inicial de valores
            config_key: Chave no arquivo de configuração ('etapas_obra' ou 'insumos')
            config_manager: Instância do GerenciadorConfiguracoes
            **kwargs: Argumentos adicionais para ttk.Combobox
        """
        super().__init__(parent, **kwargs)
        
        self.config_key = config_key
        self.config_manager = config_manager
        self.valores_originais = values or []
        
        # Configurar valores iniciais
        self['values'] = self.valores_originais
        
        # Bind de eventos
        self.bind('<KeyRelease>', self.on_keyrelease)
        self.bind('<Button-1>', self.on_click)
        self.bind('<FocusOut>', self.on_focus_out)
        self.bind('<Return>', self.on_enter)
        
        # Variável para controlar se está filtrando
        self.filtrando = False
    
    def on_keyrelease(self, event):
        """Evento chamado quando uma tecla é liberada"""
        if event.keysym in ['Up', 'Down', 'Left', 'Right', 'Return', 'Tab']:
            return
        
        texto_digitado = self.get().upper()
        
        if not texto_digitado:
            # Se não há texto, mostrar todos os valores
            self['values'] = self.valores_originais
            self.filtrando = False
            return
        
        # Filtrar valores que começam com o texto digitado
        valores_filtrados = [v for v in self.valores_originais if v.upper().startswith(texto_digitado)]
        
        if valores_filtrados:
            # Atualizar lista com valores filtrados
            self['values'] = valores_filtrados
            self.filtrando = True
            
            # Autocompletar com o primeiro resultado
            if len(valores_filtrados) == 1:
                valor_completo = valores_filtrados[0]
                self.delete(0, tk.END)
                self.insert(0, valor_completo)
                self.selection_range(len(texto_digitado), len(valor_completo))
        else:
            # Nenhuma correspondência encontrada
            self['values'] = []
            self.filtrando = True
    
    def on_click(self, event):
        """Evento de clique - mostrar todos os valores"""
        if not self.filtrando:
            self['values'] = self.valores_originais
    
    def on_focus_out(self, event):
        """Evento quando perde o foco - verificar se precisa adicionar novo item"""
        texto_digitado = self.get().strip().upper()
        
        if texto_digitado and texto_digitado not in [v.upper() for v in self.valores_originais]:
            # Perguntar se quer adicionar o novo item
            self.confirmar_adicao_item(texto_digitado)
        
        # Restaurar lista completa
        self['values'] = self.valores_originais
        self.filtrando = False
    
    def on_enter(self, event):
        """Evento Enter - confirmar seleção ou adicionar novo item"""
        texto_digitado = self.get().strip().upper()
        
        if texto_digitado and texto_digitado not in [v.upper() for v in self.valores_originais]:
            self.confirmar_adicao_item(texto_digitado)
        
        # Restaurar lista completa
        self['values'] = self.valores_originais
        self.filtrando = False
    
    def confirmar_adicao_item(self, novo_item):
        """Confirma se o usuário quer adicionar um novo item"""
        tipo_item = "Etapa da Obra" if self.config_key == "etapas_obra" else "Insumo"
        
        resposta = messagebox.askyesno(
            "Adicionar Novo Item",
            f"O {tipo_item.lower()} '{novo_item}' não existe na lista.\n\n"
            f"Deseja adicioná-lo aos parâmetros do sistema?",
            parent=self.winfo_toplevel()
        )
        
        if resposta:
            self.adicionar_novo_item(novo_item)
    
    def adicionar_novo_item(self, novo_item):
        """Adiciona um novo item ao arquivo de configurações"""
        try:
            # Carregar configurações atuais
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            config = GerenciadorConfiguracoes.carregar_configuracoes()
            
            if not config:
                messagebox.showerror("Erro", "Não foi possível carregar as configurações.")
                return
            
            # Verificar se a seção existe
            if self.config_key not in config:
                config[self.config_key] = {'lista': [], 'historico_alteracoes': []}
            
            # Adicionar o novo item
            if novo_item not in config[self.config_key]['lista']:
                config[self.config_key]['lista'].append(novo_item)
                config[self.config_key]['lista'].sort()
                
                # Salvar no arquivo
                config_path = GerenciadorConfiguracoes.CONFIG_PATH
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, indent=4, ensure_ascii=False)
                
                # Atualizar cache
                GerenciadorConfiguracoes._atualizar_cache(config)
                
                # Atualizar valores do combobox
                self.valores_originais = config[self.config_key]['lista']
                self['values'] = self.valores_originais
                
                # Definir o novo valor no combobox
                self.set(novo_item)
                
                tipo_item = "Etapa da Obra" if self.config_key == "etapas_obra" else "Insumo"
                messagebox.showinfo(
                    "Sucesso", 
                    f"{tipo_item} '{novo_item}' adicionado com sucesso!",
                    parent=self.winfo_toplevel()
                )
            
        except Exception as e:
            messagebox.showerror(
                "Erro", 
                f"Erro ao adicionar item: {str(e)}",
                parent=self.winfo_toplevel()
            )
    
    def atualizar_valores(self, novos_valores):
        """Atualiza a lista de valores do combobox"""
        self.valores_originais = novos_valores
        self['values'] = self.valores_originais

class GerenciadorAgenda:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.janela = None
        self.tree_agenda = None
        self.dados_agenda = []
        self.filtro_periodo = "quinzena_atual"  # quinzena_atual, mes_atual, periodo_personalizado
        
    def abrir_agenda(self):
        """Abre a agenda com carregamento otimizado"""
        if not self.sistema.cliente_atual:
            custom_messagebox("error", "Erro", "Selecione um cliente primeiro!")
            return
            
        # Mostrar janela de loading para clientes com muitos dados
        loading_window = self.criar_janela_loading()
        
        try:
            self.janela = tk.Toplevel(self.sistema.root)
            self.janela.title(f"Agenda - {self.sistema.cliente_atual}")
            self.janela.geometry("1200x800")
            
            self.janela.transient(self.sistema.root)
            self.janela.grab_set()
            
            # Atualizar loading
            self.atualizar_loading(loading_window, "Criando interface...")
            self.criar_interface()
            
            # Atualizar loading
            self.atualizar_loading(loading_window, "Carregando dados...")
            self.carregar_dados_agenda()
            
            # Fechar loading
            loading_window.destroy()
            
            # Mostrar alertas se necessário
            self.janela.after(500, self.mostrar_alertas_se_necessario)
            
        except Exception as e:
            if 'loading_window' in locals():
                loading_window.destroy()
            custom_messagebox("error", "Erro", f"Erro ao abrir agenda: {str(e)}")

    def criar_janela_loading(self):
        """Cria janela de loading simples"""
        loading = tk.Toplevel(self.sistema.root)
        loading.title("Carregando Agenda")
        loading.geometry("300x100")
        loading.transient(self.sistema.root)
        loading.grab_set()
        
        # Centralizar
        loading.update_idletasks()
        x = (loading.winfo_screenwidth() // 2) - (300 // 2)
        y = (loading.winfo_screenheight() // 2) - (100 // 2)
        loading.geometry(f"300x100+{x}+{y}")
        
        frame = ttk.Frame(loading, padding="20")
        frame.pack(fill='both', expand=True)
        
        loading.label_status = ttk.Label(frame, text="Iniciando...", font=('TkDefaultFont', 10))
        loading.label_status.pack(pady=10)
        
        # Barra de progresso indeterminada
        progress = ttk.Progressbar(frame, mode='indeterminate')
        progress.pack(fill='x', pady=10)
        progress.start()
        
        loading.update()
        return loading

    def atualizar_loading(self, loading_window, mensagem):
        """Atualiza mensagem da janela de loading"""
        try:
            if loading_window and loading_window.winfo_exists():
                loading_window.label_status.config(text=mensagem)
                loading_window.update()
        except:
            pass
    
    def abrir_configuracoes_agenda(self):
        """Abre configurações específicas da agenda"""
        janela_config = tk.Toplevel(self.janela)
        janela_config.title("Configurações da Agenda")
        janela_config.geometry("500x400")
        janela_config.transient(self.janela)
        janela_config.grab_set()
        
        main_frame = ttk.Frame(janela_config, padding="15")
        main_frame.pack(fill='both', expand=True)
        
        # Configurações de exibição
        frame_exibicao = ttk.LabelFrame(main_frame, text="Configurações de Exibição")
        frame_exibicao.pack(fill='x', pady=(0, 10))
        
        self.var_mostrar_apenas_futuros = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_exibicao, text="Mostrar apenas itens futuros por padrão",
                    variable=self.var_mostrar_apenas_futuros).pack(anchor='w', padx=10, pady=5)
        
        self.var_agrupar_por_semana = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame_exibicao, text="Agrupar itens por semana",
                    variable=self.var_agrupar_por_semana).pack(anchor='w', padx=10, pady=5)
        
        # Configurações de alertas
        frame_alertas = ttk.LabelFrame(main_frame, text="Configurações de Alertas")
        frame_alertas.pack(fill='x', pady=(0, 10))
        
        ttk.Label(frame_alertas, text="Alertar com quantos dias de antecedência:").pack(anchor='w', padx=10, pady=5)
        self.entry_dias_alerta = ttk.Entry(frame_alertas, width=5)
        self.entry_dias_alerta.insert(0, "3")
        self.entry_dias_alerta.pack(anchor='w', padx=10, pady=5)
        
        # Configurações de dados
        frame_dados = ttk.LabelFrame(main_frame, text="Configurações de Dados")
        frame_dados.pack(fill='x', pady=(0, 10))
        
        ttk.Label(frame_dados, text="Carregar últimos X dias do passado:").pack(anchor='w', padx=10, pady=5)
        self.entry_dias_passado = ttk.Entry(frame_dados, width=5)
        self.entry_dias_passado.insert(0, "30")
        self.entry_dias_passado.pack(anchor='w', padx=10, pady=5)
        
        self.var_excluir_mao_obra = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_dados, text="Excluir mão de obra (TP_DESP = 1)",
                    variable=self.var_excluir_mao_obra).pack(anchor='w', padx=10, pady=5)
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=(10, 0))
        
        def aplicar_configuracoes():
            try:
                # Aplicar configurações e recarregar
                self.carregar_dados_agenda()
                custom_messagebox("info", "Sucesso", "Configurações aplicadas!")
                janela_config.destroy()
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao aplicar configurações: {str(e)}")
        
        ttk.Button(frame_botoes, text="Aplicar", command=aplicar_configuracoes).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela_config.destroy).pack(side='left', padx=5)

    def criar_interface(self):
        """Cria a interface da agenda"""
        # Frame principal
        main_frame = ttk.Frame(self.janela, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # === FRAME SUPERIOR: FILTROS E CONTROLES ===
        frame_controles = ttk.Frame(main_frame)
        frame_controles.pack(fill='x', pady=(0, 10))
        
        frame_periodo = ttk.LabelFrame(frame_controles, text="Período (por Data de Relatório)")
        frame_periodo.pack(side='left', padx=(0, 10), fill='y')

        self.var_periodo = tk.StringVar(value="quinzena_atual")

        ttk.Radiobutton(frame_periodo, text="Próximo Relatório", 
                    variable=self.var_periodo, value="quinzena_atual",
                    command=self.aplicar_filtro_periodo).pack(anchor='w', padx=5)
        ttk.Radiobutton(frame_periodo, text="Próximos 2 Relatórios", 
                    variable=self.var_periodo, value="mes_atual",
                    command=self.aplicar_filtro_periodo).pack(anchor='w', padx=5)
        ttk.Radiobutton(frame_periodo, text="Período Personalizado", 
                    variable=self.var_periodo, value="personalizado",
                    command=self.aplicar_filtro_periodo).pack(anchor='w', padx=5)
        
        # Datas personalizadas (inicialmente ocultas)
        self.frame_datas_personalizado = ttk.Frame(frame_periodo)
        
        ttk.Label(self.frame_datas_personalizado, text="De:").grid(row=0, column=0, padx=2)
        self.data_inicio_personalizada = DateEntry(self.frame_datas_personalizado, width=10, 
                                                  date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_inicio_personalizada.grid(row=0, column=1, padx=2)
        
        ttk.Label(self.frame_datas_personalizado, text="Até:").grid(row=0, column=2, padx=2)
        self.data_fim_personalizada = DateEntry(self.frame_datas_personalizado, width=10, 
                                                date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_fim_personalizada.grid(row=0, column=3, padx=2)
        
        ttk.Button(self.frame_datas_personalizado, text="Aplicar", 
                  command=self.aplicar_filtro_periodo).grid(row=0, column=4, padx=5)
        
        # Filtros de status
        frame_origem = ttk.LabelFrame(frame_controles, text="Origem dos Dados")
        frame_origem.pack(side='left', padx=(0, 10), fill='y')
        
        self.var_mostrar_existentes = tk.BooleanVar(value=True)
        self.var_mostrar_pendentes_config = tk.BooleanVar(value=True)
        # self.var_mostrar_vencidos_apenas = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(frame_origem, text="Já Lançados", 
                    variable=self.var_mostrar_existentes,
                    command=self.aplicar_filtros).pack(anchor='w', padx=5)
        ttk.Checkbutton(frame_origem, text="Compromissos Configurados", 
                    variable=self.var_mostrar_pendentes_config,
                    command=self.aplicar_filtros).pack(anchor='w', padx=5)
        # ttk.Checkbutton(frame_origem, text="Apenas Vencidos/Hoje", 
        #             variable=self.var_mostrar_vencidos_apenas,
        #             command=self.aplicar_filtros).pack(anchor='w', padx=5)
        
        # Resumo rápido
        frame_resumo = ttk.LabelFrame(frame_controles, text="Resumo Rápido")
        frame_resumo.pack(side='left', padx=(0, 10), fill='y')
        
        self.label_total_periodo = ttk.Label(frame_resumo, text="Total Período: R$ 0,00", 
                                           font=('TkDefaultFont', 9, 'bold'))
        self.label_total_periodo.pack(padx=5, pady=2)
        
        self.label_pendentes = ttk.Label(frame_resumo, text="Pendentes: 0 (R$ 0,00)", 
                                        foreground='orange')
        self.label_pendentes.pack(padx=5, pady=2)
        
        self.label_vencidos = ttk.Label(frame_resumo, text="Vencidos: 0 (R$ 0,00)", 
                                       foreground='red')
        self.label_vencidos.pack(padx=5, pady=2)
        
        # === FRAME PRINCIPAL: LISTA DA AGENDA ===
        frame_lista = ttk.Frame(main_frame)
        frame_lista.pack(fill='both', expand=True)
        
        # Treeview - COLUNAS CORRIGIDAS (SEM CLIENTE)
        colunas = ('Vencimento', 'Status', 'Fornecedor', 'Referência', 
                'Valor', 'Tipo', 'Observação', 'ID_Origem')
        self.tree_agenda = ttk.Treeview(frame_lista, columns=colunas, show='headings', 
                                    height=25, selectmode='extended')
        
        # Configurar cabeçalhos
        self.tree_agenda.heading('Vencimento', text='Vencimento')
        self.tree_agenda.heading('Status', text='Status')
        self.tree_agenda.heading('Fornecedor', text='Fornecedor')
        self.tree_agenda.heading('Referência', text='Referência')
        self.tree_agenda.heading('Valor', text='Valor')
        self.tree_agenda.heading('Tipo', text='Tipo')
        self.tree_agenda.heading('Observação', text='Observação')
        self.tree_agenda.heading('ID_Origem', text='ID')
        
        # Configurar larguras - AJUSTADAS SEM CLIENTE
        self.tree_agenda.column('Vencimento', width=80, anchor='center')
        self.tree_agenda.column('Status', width=100, anchor='center')
        self.tree_agenda.column('Fornecedor', width=200)  # Mais espaço sem cliente
        self.tree_agenda.column('Referência', width=280)  # Mais espaço
        self.tree_agenda.column('Valor', width=100, anchor='e')
        self.tree_agenda.column('Tipo', width=80, anchor='center')
        self.tree_agenda.column('Observação', width=200)
        self.tree_agenda.column('ID_Origem', width=0, stretch=False)  # Oculto
    
        # Scrollbars
        scrolly_agenda = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_agenda.yview)
        scrollx_agenda = ttk.Scrollbar(frame_lista, orient='horizontal', command=self.tree_agenda.xview)
        self.tree_agenda.configure(yscrollcommand=scrolly_agenda.set, xscrollcommand=scrollx_agenda.set)
        
        # Posicionar elementos
        self.tree_agenda.grid(row=0, column=0, sticky='nsew')
        scrolly_agenda.grid(row=0, column=1, sticky='ns')
        scrollx_agenda.grid(row=1, column=0, sticky='ew')
        
        # Configurar expansão
        frame_lista.grid_rowconfigure(0, weight=1)
        frame_lista.grid_columnconfigure(0, weight=1)
        
        # Tags para cores
        self.tree_agenda.tag_configure('lancado', background='#e8f5e8')      # Verde claro
        self.tree_agenda.tag_configure('pendente', background='#fff8dc')     # Bege claro
        self.tree_agenda.tag_configure('vencido', background='#ffe4e1')      # Rosa claro
        self.tree_agenda.tag_configure('condicionado', background='#e6f3ff') # Azul claro
        self.tree_agenda.tag_configure('hoje', background='#ffffe0')         # Amarelo claro
        
        # === FRAME INFERIOR: BOTÕES DE AÇÃO ===
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=(10, 0))
        
        # Botões principais
        ttk.Button(frame_botoes, text="Novo Lançamento", 
                  command=self.novo_lancamento).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Editar Selecionado", 
                  command=self.editar_selecionado).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Confirmar Lançamento", 
                  command=self.confirmar_lancamento).pack(side='left', padx=5)
        
        # Separador
        ttk.Separator(frame_botoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # NOVO BOTÃO: Gerenciar Compromissos
        ttk.Button(frame_botoes, text="📅 Gerenciar Compromissos", 
                command=self.abrir_gerenciador_compromissos).pack(side='left', padx=5)
    
        # Separador
        ttk.Separator(frame_botoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # Botões de importação/exportação
        ttk.Button(frame_botoes, text="Importar de Excel", 
                  command=self.importar_excel).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Exportar Período", 
                  command=self.exportar_periodo).pack(side='left', padx=5)
        
        # Separador
        ttk.Separator(frame_botoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # Botões de controle
        ttk.Button(frame_botoes, text="Atualizar", 
                  command=self.carregar_dados_agenda).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Fechar", 
                  command=self.janela.destroy).pack(side='right', padx=5)
        
        # Separador
        ttk.Separator(frame_botoes, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # Botões de configuração e limpeza
        ttk.Button(frame_botoes, text="🧹 Limpar Duplicações", 
                command=self.limpar_duplicacoes_inteligente).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="⚙️ Configurações", 
                command=self.abrir_configuracoes_agenda).pack(side='left', padx=5)
    
        # Configurar eventos
        self.tree_agenda.bind('<Double-1>', self.on_double_click)
        self.configurar_atalhos()
    
    def carregar_dados_agenda(self):
        """Carregamento da agenda baseado em DATA_REL"""
        try:
            print("=" * 80)
            print("DEBUG: Iniciando carregamento da agenda")
            print("IMPORTANTE: Filtragem por DATA_REL (dia do relatório)")
            print("=" * 80)
            self.dados_agenda = []
            
            # 1. Carregar lançamentos existentes (filtrados por DATA_REL)
            self.carregar_lancamentos_existentes()
            
            # 2. Carregar compromissos das configurações (gerados por DATA_REL)
            self.carregar_compromissos_futuros()
            
            # 3. Identificar condicionados
            self.carregar_lancamentos_condicionados()
            
            # 4. Aplicar filtros e atualizar
            self.aplicar_filtros()
            self.atualizar_resumo()
            
            total_items = len(self.dados_agenda)
            existentes = len([d for d in self.dados_agenda if d['origem'] == 'EXISTENTE'])
            pendentes = len([d for d in self.dados_agenda if d['origem'] in ['CONFIGURACAO', 'BASICO']])
            
            print("-" * 80)
            print(f"RESUMO: {total_items} itens ({existentes} existentes, {pendentes} pendentes)")
            print("=" * 80)
            
        except Exception as e:
            print(f"DEBUG: Erro ao carregar agenda: {str(e)}")
            custom_messagebox("error", "Erro", f"Erro ao carregar agenda: {str(e)}")
    
    def carregar_lancamentos_existentes(self):
        """Carrega lançamentos baseados na DATA_REL (data do relatório)"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            if not arquivo_cliente.exists():
                print("DEBUG: Arquivo do cliente não existe")
                return

            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")

            hoje = datetime.now().date()
            
            # Obter período de filtro baseado em DATA_REL
            data_inicio, data_fim = self.calcular_periodo_filtro()
            
            print(f"DEBUG: Carregando lançamentos com DATA_REL entre {data_inicio} e {data_fim}")

            for idx, row in df.iterrows():
                if row.get('STATUS', 'ATIVO') == 'EXCLUIDO':
                    continue
                
                # FILTRAR POR DATA_REL (data do relatório)
                try:
                    data_rel = pd.to_datetime(row['DATA_REL'], dayfirst=True).date()
                    
                    # MUDANÇA CRÍTICA: Verificar se DATA_REL está no período desejado
                    if not (data_inicio <= data_rel <= data_fim):
                        continue
                        
                except:
                    print(f"DEBUG: Data_rel inválida na linha {idx}")
                    continue
                
                # FILTRAR POR TIPO DE DESPESA - EXCLUIR MÃO DE OBRA (tp_desp == 1)
                try:
                    tp_desp = int(float(row.get('TP_DESP', 0)))
                    if tp_desp == 1:  # Pular mão de obra
                        continue
                except (ValueError, TypeError):
                    continue

                # Verificar e validar valor
                try:
                    valor = row.get('VALOR', 0)
                    if pd.isna(valor) or valor == '' or valor == 0:
                        continue  # Pular lançamentos sem valor
                    valor = float(valor)
                except (ValueError, TypeError):
                    print(f"DEBUG: Valor inválido na linha {idx}: {valor}")
                    continue

                # Determinar status baseado na data de vencimento (para exibição visual)
                try:
                    dt_vencto = pd.to_datetime(row['DT_VENCTO'], dayfirst=True).date()
                    
                    if dt_vencto < hoje:
                        status = "VENCIDO"
                    elif dt_vencto == hoje:
                        status = "VENCE HOJE"
                    else:
                        status = "LANÇADO"
                        
                except:
                    print(f"DEBUG: Data vencimento inválida na linha {idx}")
                    continue

                item_agenda = {
                    'vencimento': dt_vencto,
                    'data_rel': data_rel,  # ADICIONAR data_rel para comparações
                    'status': status,
                    'fornecedor': row.get('NOME', ''),
                    'referencia': row.get('REFERÊNCIA', ''),
                    'valor': valor,
                    'tipo': tp_desp,
                    'observacao': row.get('OBSERVAÇÃO', ''),
                    'id_origem': row.get('ID_LANCAMENTO', ''),
                    'origem': 'EXISTENTE',
                    'dados_completos': row.to_dict()
                }
                
                self.dados_agenda.append(item_agenda)
                
            print(f"DEBUG: {len([d for d in self.dados_agenda if d['origem'] == 'EXISTENTE'])} lançamentos existentes carregados")
                    
        except Exception as e:
            print(f"DEBUG: Erro ao carregar lançamentos existentes: {str(e)}")
    
    def carregar_compromissos_futuros(self):
        """Carrega apenas compromissos das configurações (sem arquivos externos)"""
        try:
            print("DEBUG: Carregando apenas compromissos das configurações")
            
            # APENAS usar compromissos recorrentes das configurações
            self.gerar_compromissos_recorrentes_config()
            
            # NÃO importar de Agenda.xlsx nem fazer análise histórica
            # Isso simplifica e torna mais previsível
            
        except Exception as e:
            print(f"DEBUG: Erro ao carregar compromissos futuros: {str(e)}")

    def gerar_compromissos_recorrentes_config(self):
        """Gera compromissos recorrentes baseados em DATA_REL (dias de relatório)"""
        try:
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            
            compromissos_config = GerenciadorConfiguracoes.get_compromissos_recorrentes()
            
            if not compromissos_config:
                print("DEBUG: Nenhum compromisso configurado")
                return
            
            hoje = datetime.now().date()
            fim_periodo = hoje + relativedelta(months=3)  # Próximos 3 meses
            
            print(f"DEBUG: Gerando compromissos para {len(compromissos_config)} itens configurados")
            
            for compromisso in compromissos_config:
                try:
                    # Calcular datas de DATA_REL (dias de relatório: 5 e 20)
                    datas_relatorio = self.calcular_datas_relatorio_recorrencia(
                        compromisso, hoje, fim_periodo
                    )
                    
                    for data_rel in datas_relatorio:
                        # Calcular data de vencimento sugerida (pode ser editada depois)
                        dia_venc_config = compromisso.get('dia_vencimento', 5)
                        try:
                            data_vencimento_sugerida = data_rel.replace(day=dia_venc_config)
                        except ValueError:
                            ultimo_dia = calendar.monthrange(data_rel.year, data_rel.month)[1]
                            data_vencimento_sugerida = data_rel.replace(day=min(dia_venc_config, ultimo_dia))
                        
                        # VERIFICAÇÃO CORRIGIDA: Verificar por DATA_REL, não por data de vencimento
                        ja_existe = any(
                            item.get('data_rel') == data_rel and 
                            compromisso['nome'].upper() in item['fornecedor'].upper()
                            for item in self.dados_agenda
                        )
                        
                        if not ja_existe:
                            item_agenda = {
                                'vencimento': data_vencimento_sugerida,
                                'data_rel': data_rel,  # CRÍTICO: Adicionar data_rel
                                'status': 'PENDENTE',
                                'cliente': self.sistema.cliente_atual,
                                'fornecedor': compromisso['nome'],
                                'referencia': f"{compromisso['nome']}",
                                'valor': compromisso.get('valor_estimado', 0.0),
                                'tipo': f"TD{compromisso.get('tipo_despesa', 3)}",
                                'observacao': f"{compromisso.get('observacao', 'Compromisso recorrente')}",
                                'id_origem': f"CONFIG_{compromisso['nome'].replace(' ', '_')}_{data_rel.strftime('%Y%m%d')}",
                                'origem': 'CONFIGURACAO',
                                'dados_originais': compromisso
                            }
                            
                            self.dados_agenda.append(item_agenda)
                            
                except Exception as e:
                    print(f"DEBUG: Erro ao processar {compromisso.get('nome', 'N/A')}: {str(e)}")
                    continue
            
            print(f"DEBUG: Compromissos das configurações gerados")
            
        except ImportError:
            print("DEBUG: Configurações não disponíveis - usando lista básica")
            self.gerar_compromissos_basicos()
        except Exception as e:
            print(f"DEBUG: Erro ao gerar compromissos das configurações: {str(e)}")
    
    def calcular_datas_relatorio_recorrencia(self, compromisso, data_inicio, data_fim):
        """
        Calcula datas de DATA_REL (relatórios) para compromissos recorrentes
        Sempre retorna dias 5 e 20 de cada mês baseado na recorrência
        """
        datas_relatorio = []
        recorrencia = compromisso.get('recorrencia', 'mensal').lower()
        
        # Começar do próximo relatório a partir de hoje
        hoje = data_inicio
        data_atual = hoje.replace(day=1)  # Primeiro dia do mês atual
        
        # Determinar qual relatório usar baseado no dia de vencimento configurado
        dia_vencimento = compromisso.get('dia_vencimento', 5)
        
        # Se vencimento é até dia 5, usar relatório do dia 5
        # Se vencimento é após dia 5, usar relatório do dia 20
        if dia_vencimento <= 5:
            dias_relatorio = [5]  # Apenas relatório do dia 5
        elif dia_vencimento <= 20:
            dias_relatorio = [20]  # Apenas relatório do dia 20
        else:
            dias_relatorio = [5]  # Default: relatório do dia 5 do próximo mês
        
        meses_processados = 0
        max_meses = 12  # Limite de segurança
        
        while data_atual <= data_fim and meses_processados < max_meses:
            for dia_rel in dias_relatorio:
                try:
                    data_relatorio = data_atual.replace(day=dia_rel)
                    
                    # Só adicionar se for data futura e dentro do período
                    if data_relatorio > hoje and data_relatorio <= data_fim:
                        datas_relatorio.append(data_relatorio)
                        
                except ValueError:
                    continue
            
            # Avançar para próximo mês baseado na recorrência
            if recorrencia == 'mensal':
                data_atual = data_atual + relativedelta(months=1)
            elif recorrencia == 'bimestral':
                data_atual = data_atual + relativedelta(months=2)
            elif recorrencia == 'trimestral':
                data_atual = data_atual + relativedelta(months=3)
            elif recorrencia == 'semestral':
                data_atual = data_atual + relativedelta(months=6)
            elif recorrencia == 'anual':
                data_atual = data_atual + relativedelta(months=12)
            else:
                data_atual = data_atual + relativedelta(months=1)  # Default mensal
            
            meses_processados += 1
        
        return sorted(datas_relatorio)

    def gerar_compromissos_basicos(self):
        """Fallback com compromissos básicos baseados em DATA_REL"""
        compromissos_basicos = [
            {
                'nome': 'MOTOBOY', 
                'dia_vencimento': 5, 
                'valor_estimado': 0.0,
                'categoria': 'DIV',
                'tipo_despesa': 2,
                'recorrencia': 'mensal',
                'observacao': 'Serviço de motoboy obra'
            },
            {
                'nome': 'FOLHA DP', 
                'dia_vencimento': 5, 
                'valor_estimado': 0.0,
                'categoria': 'MO',
                'tipo_despesa': 3,
                'recorrencia': 'mensal',
                'observacao': 'Gestão de folha de pagamento'
            }
        ]
        
        hoje = datetime.now().date()
        fim_periodo = hoje + relativedelta(months=3)
        
        print("DEBUG: Usando compromissos básicos (fallback)")
        
        for compromisso in compromissos_basicos:
            try:
                # Usar o mesmo método de cálculo de datas de relatório
                datas_relatorio = self.calcular_datas_relatorio_recorrencia(
                    compromisso, hoje, fim_periodo
                )
                
                for data_rel in datas_relatorio:
                    # Calcular data de vencimento sugerida
                    dia_venc = compromisso['dia_vencimento']
                    try:
                        data_vencimento_sugerida = data_rel.replace(day=dia_venc)
                    except ValueError:
                        ultimo_dia = calendar.monthrange(data_rel.year, data_rel.month)[1]
                        data_vencimento_sugerida = data_rel.replace(day=min(dia_venc, ultimo_dia))
                    
                    # Verificar duplicação por DATA_REL
                    ja_existe = any(
                        item.get('data_rel') == data_rel and 
                        compromisso['nome'].upper() in item['fornecedor'].upper()
                        for item in self.dados_agenda
                    )
                    
                    if not ja_existe:
                        item_agenda = {
                            'vencimento': data_vencimento_sugerida,
                            'data_rel': data_rel,
                            'status': 'PENDENTE',
                            'cliente': self.sistema.cliente_atual,
                            'fornecedor': compromisso['nome'],
                            'referencia': f"{compromisso['nome']} - REL {data_rel.strftime('%d/%m/%Y')}",
                            'valor': compromisso['valor_estimado'],
                            'tipo': f"TD{compromisso['tipo_despesa']}",
                            'observacao': f"{compromisso['observacao']} - Relatório {data_rel.strftime('%d/%m')}",
                            'id_origem': f"BASICO_{compromisso['nome'].replace(' ', '_')}_{data_rel.strftime('%Y%m%d')}",
                            'origem': 'BASICO',
                            'categoria': compromisso['categoria'],
                            'dados_originais': compromisso
                        }
                        
                        self.dados_agenda.append(item_agenda)
                        print(f"DEBUG: Adicionado compromisso básico: {compromisso['nome']} - REL {data_rel}")
                    
            except Exception as e:
                print(f"DEBUG: Erro ao processar compromisso básico {compromisso['nome']}: {str(e)}")
                continue
        
        print("DEBUG: Compromissos básicos gerados com sucesso")

    def calcular_datas_recorrencia_simples(self, compromisso, data_inicio, data_fim):
        """Versão simplificada do cálculo de recorrência"""
        datas = []
        dia_vencimento = compromisso.get('dia_vencimento', 5)
        recorrencia = compromisso.get('recorrencia', 'mensal').lower()
        
        # Começar do mês atual
        data_atual = data_inicio.replace(day=1)
        
        while data_atual <= data_fim:
            try:
                data_vencimento = data_atual.replace(day=dia_vencimento)
            except ValueError:
                # Dia não existe no mês
                ultimo_dia = calendar.monthrange(data_atual.year, data_atual.month)[1]
                data_vencimento = data_atual.replace(day=min(dia_vencimento, ultimo_dia))
            
            if data_vencimento > data_inicio:
                datas.append(data_vencimento)
            
            # Próxima data
            if recorrencia == 'mensal':
                data_atual = data_atual + relativedelta(months=1)
            elif recorrencia == 'trimestral':
                data_atual = data_atual + relativedelta(months=3)
            else:
                data_atual = data_atual + relativedelta(months=1)  # Default mensal
        
        return datas

    def gerar_compromissos_recorrentes(self):
        """Gera compromissos recorrentes comuns (salários, seguros, etc.)"""
        try:
            hoje = datetime.now().date()
            fim_periodo = hoje + relativedelta(months=2)  # Próximos 2 meses
            
            # Definir compromissos padrão baseados na imagem da agenda
            compromissos_padrao = [
                {
                    'nome': 'FOLHA DP',
                    'dia_vencimento': 5,
                    'recorrencia': 'mensal',
                    'valor_estimado': 0,
                    'categoria': 'MO'
                },
                {
                    'nome': 'ADMINISTRAÇÃO',
                    'dia_vencimento': 20,
                    'recorrencia': 'mensal',
                    'valor_estimado': 0,
                    'categoria': 'ADM'
                },
                {
                    'nome': 'MHS EVENTO SST ESOCIAL',
                    'dia_vencimento': 20,
                    'recorrencia': 'mensal',
                    'valor_estimado': 0,
                    'categoria': 'MO'
                },
                {
                    'nome': 'FGTS',
                    'dia_vencimento': 20,
                    'recorrencia': 'mensal',
                    'valor_estimado': 0,
                    'categoria': 'MO'
                }
            ]
            
            # Gerar ocorrências futuras
            data_atual = hoje
            while data_atual <= fim_periodo:
                for compromisso in compromissos_padrao:
                    try:
                        # Calcular data de vencimento para o mês atual
                        data_vencimento = data_atual.replace(day=compromisso['dia_vencimento'])
                        
                        # Só adicionar se for data futura
                        if data_vencimento > hoje:
                            # Verificar se já não existe lançamento para esta data
                            ja_existe = any(
                                item['vencimento'] == data_vencimento and 
                                compromisso['nome'].upper() in item['fornecedor'].upper()
                                for item in self.dados_agenda
                            )
                            
                            if not ja_existe:
                                item_agenda = {
                                    'vencimento': data_vencimento,
                                    'status': 'PENDENTE',
                                    'cliente': self.sistema.cliente_atual,
                                    'fornecedor': compromisso['nome'],
                                    'referencia': f"{compromisso['nome']} - {data_vencimento.strftime('%m/%Y')}",
                                    'valor': compromisso['valor_estimado'],
                                    'tipo': 'RECORRENTE',
                                    'observacao': f"Recorrente - {compromisso['categoria']}",
                                    'id_origem': f"REC_{compromisso['categoria']}_{data_vencimento.strftime('%Y%m%d')}",
                                    'origem': 'RECORRENTE',
                                    'categoria': compromisso['categoria']
                                }
                                
                                self.dados_agenda.append(item_agenda)
                    
                    except ValueError:
                        # Se o dia não existir no mês (ex: 31 em fevereiro), usar último dia
                        ultimo_dia = calendar.monthrange(data_atual.year, data_atual.month)[1]
                        dia_ajustado = min(compromisso['dia_vencimento'], ultimo_dia)
                        data_vencimento = data_atual.replace(day=dia_ajustado)
                        
                        if data_vencimento > hoje:
                            # Mesmo processo de verificação e adição
                            pass
                
                # Avançar para próximo mês
                data_atual = data_atual + relativedelta(months=1)
                data_atual = data_atual.replace(day=1)  # Primeiro dia do próximo mês
                
        except Exception as e:
            print(f"DEBUG: Erro ao gerar compromissos recorrentes: {str(e)}")
    
    def carregar_lancamentos_condicionados(self):
        """Versão simplificada - apenas identificar condicionados nos existentes"""
        try:
            for item in self.dados_agenda:
                if item['origem'] == 'EXISTENTE':
                    observacao = item['observacao'].upper()
                    
                    # Identificar se é condicionado
                    if any(palavra in observacao for palavra in ['CONDICIONADA', 'ENTREGA', 'APROVAÇÃO']):
                        item['status'] = 'CONDICIONADO'
                        item['observacao'] = f"🔒 {item['observacao']}"
                            
        except Exception as e:
            print(f"DEBUG: Erro ao carregar lançamentos condicionados: {str(e)}")
    
    def aplicar_filtro_periodo(self):
        """Aplica filtro de período e preenche datas personalizadas quando necessário"""
        # Lógica de preenchimento automático para período personalizado
        if self.var_periodo.get() == "personalizado":
            hoje = datetime.now().date()
            data_fim_padrao = hoje + timedelta(days=60)
            
            try:
                # Verificar se os campos estão vazios ou com valores muito antigos
                try:
                    inicio_atual = self.data_inicio_personalizada.get_date()
                    fim_atual = self.data_fim_personalizada.get_date()
                    
                    # Se as datas são muito antigas ou iguais, atualizar
                    if (inicio_atual < hoje - timedelta(days=30) or 
                        inicio_atual == fim_atual):
                        self.data_inicio_personalizada.set_date(hoje)
                        self.data_fim_personalizada.set_date(data_fim_padrao)
                except:
                    # Se der erro ao ler, preencher com valores padrão
                    self.data_inicio_personalizada.set_date(hoje)
                    self.data_fim_personalizada.set_date(data_fim_padrao)
                    
            except Exception as e:
                print(f"DEBUG: Erro ao preencher datas personalizadas: {e}")
            
            # Mostrar frame de datas personalizadas
            self.frame_datas_personalizado.pack(pady=5)
        else:
            # Ocultar frame de datas personalizadas
            self.frame_datas_personalizado.pack_forget()
        
        # Aplicar o filtro (sua lógica existente aqui)
        self.aplicar_filtros()  # ou qualquer método que já existe

    def calcular_periodo_filtro(self):
        """
        Calcula períodos de filtro baseados nas DATAS DE RELATÓRIO (DATA_REL)
        Próximos relatórios: dias 5 e 20 de cada mês
        """
        hoje = datetime.now().date()
        periodo = self.var_periodo.get()
        
        if periodo == "quinzena_atual":
            # Determinar o próximo relatório (dia 5 ou 20)
            if hoje.day >= 21 or hoje.day <= 5:
                # Período para o relatório do dia 5
                if hoje.day >= 21:
                    # Após dia 20: próximo relatório é dia 5 do próximo mês
                    if hoje.month == 12:
                        proximo_relatorio = hoje.replace(year=hoje.year + 1, month=1, day=5)
                    else:
                        proximo_relatorio = hoje.replace(month=hoje.month + 1, day=5)
                else:
                    # Até dia 5: relatório é dia 5 do mês atual
                    proximo_relatorio = hoje.replace(day=5)
                
                data_inicio = hoje
                data_fim = proximo_relatorio
                
            else:
                # Entre dia 6 e 20: próximo relatório é dia 20 do mês atual
                proximo_relatorio = hoje.replace(day=20)
                data_inicio = hoje
                data_fim = proximo_relatorio
        
        elif periodo == "mes_atual":
            # Ambos relatórios do próximo mês (dia 5 e dia 20)
            if hoje.month == 12:
                proximo_mes = hoje.replace(year=hoje.year + 1, month=1)
            else:
                proximo_mes = hoje.replace(month=hoje.month + 1)
            
            # Do início do mês ao dia 20 do próximo mês
            data_inicio = proximo_mes.replace(day=1)
            data_fim = proximo_mes.replace(day=20)
            
        else:  # personalizado
            # Padrão: até o segundo relatório futuro (aproximadamente 60 dias)
            data_inicio = hoje
            data_fim = hoje + timedelta(days=60)
            
            # Sobrescrever se datas personalizadas foram definidas
            try:
                data_inicio_personalizada = self.data_inicio_personalizada.get_date()
                data_fim_personalizada = self.data_fim_personalizada.get_date()
                
                if (data_inicio_personalizada != hoje or 
                    data_fim_personalizada != hoje + timedelta(days=60)):
                    data_inicio = data_inicio_personalizada
                    data_fim = data_fim_personalizada
            except:
                pass
        
        return data_inicio, data_fim

    def aplicar_filtros(self):
        """Filtros baseados em DATA_REL (data do relatório)"""
        try:
            # Limpar tree
            for item in self.tree_agenda.get_children():
                self.tree_agenda.delete(item)
            
            # Determinar período baseado em DATA_REL
            data_inicio, data_fim = self.calcular_periodo_filtro()
            
            print(f"DEBUG: Aplicando filtros para DATA_REL entre {data_inicio} e {data_fim}")
            
            items_mostrados = 0
            
            for item in self.dados_agenda:
                # MUDANÇA CRÍTICA: Filtrar por DATA_REL ao invés de data de vencimento
                data_comparacao = item.get('data_rel', item['vencimento'])
                
                if not (data_inicio <= data_comparacao <= data_fim):
                    continue
                
                # Filtros por origem
                if item['origem'] == 'EXISTENTE' and not self.var_mostrar_existentes.get():
                    continue
                if item['origem'] in ['CONFIGURACAO', 'BASICO'] and not self.var_mostrar_pendentes_config.get():
                    continue
                
                # Determinar tag para cor
                if item['status'] == 'LANÇADO':
                    tag = 'lancado'
                elif item['status'] == 'VENCIDO':
                    tag = 'vencido'
                elif item['status'] == 'VENCE HOJE':
                    tag = 'hoje'
                elif item['status'] == 'CONDICIONADO':
                    tag = 'condicionado'
                else:
                    tag = 'pendente'
                
                # Adicionar indicador de origem e DATA_REL
                tipo_display = str(item['tipo'])
                if item['origem'] == 'CONFIGURACAO':
                    tipo_display += " (CFG)"
                elif item['origem'] == 'BASICO':
                    tipo_display += " (BAS)"
                
                # Inserir no tree
                valores = (
                    item['vencimento'].strftime('%d/%m/%Y'),
                    item['status'],
                    item['fornecedor'],
                    item['referencia'],
                    f"R$ {item['valor']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    tipo_display,
                    item['observacao'],
                    item['id_origem']
                )
                
                self.tree_agenda.insert('', 'end', values=valores, tags=(tag,))
                items_mostrados += 1
            
            print(f"DEBUG: {items_mostrados} itens mostrados após filtros")
            
            # Atualizar resumo
            self.atualizar_resumo()
            
        except Exception as e:
            print(f"DEBUG: Erro ao aplicar filtros: {str(e)}")
            import traceback
            traceback.print_exc()

    def atualizar_resumo(self):
        """Resumo mais claro e útil"""
        try:
            # Contadores
            total_valor = 0
            pendentes_count = 0
            pendentes_valor = 0
            vencidos_count = 0
            vencidos_valor = 0
            lancados_count = 0
            lancados_valor = 0

            for child in self.tree_agenda.get_children():
                valores = self.tree_agenda.item(child, 'values')
                status = valores[1]
                valor_str = valores[5].replace('R$ ', '').replace('.', '').replace(',', '.')
                try:
                    valor = float(valor_str)
                except:
                    valor = 0.0

                total_valor += valor

                if status == 'PENDENTE':
                    pendentes_count += 1
                    pendentes_valor += valor
                elif status in ['VENCIDO', 'VENCE HOJE']:
                    vencidos_count += 1
                    vencidos_valor += valor
                elif status == 'LANÇADO':
                    lancados_count += 1
                    lancados_valor += valor

            # Atualizar labels com informações mais claras
            self.label_total_periodo.config(
                text=f"Total Período: R$ {total_valor:,.2f} ({lancados_count + pendentes_count + vencidos_count} itens)".replace(',', 'X').replace('.', ',').replace('X', '.')
            )

            self.label_pendentes.config(
                text=f"Pendentes: {pendentes_count} (R$ {pendentes_valor:,.2f})".replace(',', 'X').replace('.', ',').replace('X', '.')
            )

            self.label_vencidos.config(
                text=f"Vencidos: {vencidos_count} (R$ {vencidos_valor:,.2f})".replace(',', 'X').replace('.', ',').replace('X', '.')
            )

        except Exception as e:
            print(f"DEBUG: Erro ao atualizar resumo: {str(e)}")
    
    def calcular_data_rel(self):
        """
        Calcula a data de referência (relatório) seguindo a regra dos dias 5 e 20.
        Esta é a data que determina EM QUAL RELATÓRIO o lançamento aparecerá.
        
        Regra:
        - Dia 1 a 5: Relatório do dia 5 do mês atual
        - Dia 6 a 20: Relatório do dia 20 do mês atual  
        - Dia 21 a 31: Relatório do dia 5 do próximo mês
        
        A data de vencimento (DT_VENCTO) pode ser diferente e editável.
        """
        hoje = datetime.now()
        if 6 <= hoje.day <= 20:
            data_rel = hoje.replace(day=20)
        else:
            if hoje.day > 20:
                data_rel = (hoje + relativedelta(months=1)).replace(day=5)
            else:
                data_rel = hoje.replace(day=5)
        return data_rel

    def novo_lancamento(self):
        """Abre interface para novo lançamento diretamente da agenda"""
        try:
            # Abrir janela de novo lançamento direto
            self.abrir_janela_novo_lancamento()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir novo lançamento: {str(e)}")
            print(f"DEBUG: Erro ao abrir novo lançamento: {str(e)}")

    def abrir_janela_novo_lancamento(self):
        """Cria janela para novo lançamento diretamente da agenda"""
        try:
            # Janela de novo lançamento
            janela_novo = tk.Toplevel(self.janela)
            janela_novo.title("Novo Lançamento")
            janela_novo.geometry("700x600")
            janela_novo.transient(self.janela)
            janela_novo.grab_set()
            
            # Frame principal
            main_frame = ttk.Frame(janela_novo, padding="15")
            main_frame.pack(fill='both', expand=True)
            
            # Título
            ttk.Label(main_frame, text=f"Novo Lançamento - {self.sistema.cliente_atual}", 
                    font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 15))
            
            # === SEÇÃO: DATA ===
            frame_data = ttk.LabelFrame(main_frame, text="Data de Referência")
            frame_data.pack(fill='x', pady=(0, 10))
            
            ttk.Label(frame_data, text="Data do Relatório:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            data_rel = DateEntry(frame_data, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_rel.set_date(self.calcular_data_rel())  # PRÉ-PREENCHER
            data_rel.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            
            # === SEÇÃO: FORNECEDOR ===
            frame_fornecedor = ttk.LabelFrame(main_frame, text="Dados do Fornecedor")
            frame_fornecedor.pack(fill='x', pady=(0, 10))
            
            # CNPJ/CPF com busca automática
            ttk.Label(frame_fornecedor, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            cnpj_cpf = ttk.Entry(frame_fornecedor, width=20)
            cnpj_cpf.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            
            # Nome com busca por parte do nome
            ttk.Label(frame_fornecedor, text="Nome:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
            nome = ttk.Entry(frame_fornecedor, width=40)
            nome.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
            
            # Lista de sugestões para o nome
            lista_sugestoes = tk.Listbox(frame_fornecedor, height=4)
            lista_sugestoes.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
            lista_sugestoes.grid_remove()  # Inicialmente oculta
            
            def buscar_fornecedor_por_cnpj(event=None):
                """Busca fornecedor pelo CNPJ/CPF"""
                cnpj_digitado = cnpj_cpf.get().strip()
                if len(cnpj_digitado) >= 11:  # CNPJ/CPF completo
                    try:
                        fornecedor_dados = self.sistema.buscar_fornecedor_por_cnpj_agenda(cnpj_digitado)
                        if fornecedor_dados:
                            preencher_dados_fornecedor(fornecedor_dados)
                            print(f"DEBUG: Fornecedor encontrado por CNPJ: {fornecedor_dados.get('nome', '')}")
                    except Exception as e:
                        print(f"DEBUG: Erro ao buscar fornecedor por CNPJ: {str(e)}")
            
            def buscar_fornecedor_por_nome(event=None):
                """Busca fornecedor por parte do nome"""
                nome_digitado = nome.get().strip()
                
                if len(nome_digitado) < 3:  # Só buscar com 3+ caracteres
                    lista_sugestoes.grid_remove()
                    return
                    
                try:
                    fornecedores_encontrados = self.sistema.buscar_fornecedores_por_nome_parcial(nome_digitado)
                    
                    if fornecedores_encontrados:
                        # Mostrar lista de sugestões
                        lista_sugestoes.delete(0, tk.END)
                        for fornecedor in fornecedores_encontrados[:10]:  # Máximo 10 sugestões
                            texto = f"{fornecedor['nome']} - {fornecedor['cnpj_cpf']}"
                            lista_sugestoes.insert(tk.END, texto)
                        
                        lista_sugestoes.grid()
                        print(f"DEBUG: {len(fornecedores_encontrados)} fornecedores encontrados para '{nome_digitado}'")
                    else:
                        lista_sugestoes.grid_remove()
                        print(f"DEBUG: Nenhum fornecedor encontrado para '{nome_digitado}'")
                        
                except Exception as e:
                    print(f"DEBUG: Erro ao buscar fornecedor por nome: {str(e)}")
                    lista_sugestoes.grid_remove()
            
            def selecionar_fornecedor_da_lista(event=None):
                """Seleciona fornecedor da lista de sugestões - CORRIGIDO"""
                try:
                    selection = lista_sugestoes.curselection()
                    if selection:
                        texto_selecionado = lista_sugestoes.get(selection[0])
                        print(f"DEBUG: Texto selecionado: {texto_selecionado}")
                        
                        # Extrair CNPJ do texto: "NOME - CNPJ"
                        cnpj_extraido = texto_selecionado.split(' - ')[-1]
                        print(f"DEBUG: CNPJ extraído: {cnpj_extraido}")
                        
                        # Buscar dados completos do fornecedor
                        fornecedor_dados = self.sistema.buscar_fornecedor_por_cnpj_agenda(cnpj_extraido)
                        if fornecedor_dados:
                            preencher_dados_fornecedor(fornecedor_dados)
                            lista_sugestoes.grid_remove()
                            print(f"DEBUG: Fornecedor selecionado: {fornecedor_dados.get('nome', '')}")
                        else:
                            print(f"DEBUG: Dados do fornecedor não encontrados para CNPJ: {cnpj_extraido}")
                    else:
                        print("DEBUG: Nenhuma seleção na lista")
                            
                except Exception as e:
                    print(f"DEBUG: Erro ao selecionar fornecedor: {str(e)}")
                    import traceback
                    traceback.print_exc()
            
            def selecionar_por_clique(event=None):
                """Seleciona fornecedor por clique simples"""
                try:
                    # Pequeno delay para garantir que a seleção foi processada
                    janela_novo.after(50, selecionar_fornecedor_da_lista)
                except Exception as e:
                    print(f"DEBUG: Erro no clique: {str(e)}")
            
            def preencher_dados_fornecedor(fornecedor_dados):
                """Preenche todos os campos com dados do fornecedor"""
                try:
                    print(f"DEBUG: Preenchendo dados do fornecedor: {fornecedor_dados.get('nome', '')}")
                    
                    # Preencher CNPJ/CPF
                    cnpj_cpf.delete(0, tk.END)
                    cnpj_cpf.insert(0, fornecedor_dados.get('cnpj_cpf', ''))
                    
                    # Preencher nome
                    nome.delete(0, tk.END)
                    nome.insert(0, fornecedor_dados.get('nome', ''))
                    
                    # Preencher categoria
                    categoria.delete(0, tk.END)
                    categoria.insert(0, fornecedor_dados.get('categoria', 'FORNECEDOR'))
                    
                    # CORREÇÃO: Preencher dados bancários usando o método do sistema com forma de pagamento
                    dados_bancarios_texto = self.sistema.obter_dados_bancarios_fornecedor(
                        fornecedor_dados.get('cnpj_cpf', ''),
                        forma_pagamento_preferida='PIX'
                    )
                    dados_bancarios_entry.config(state='normal')
                    dados_bancarios_entry.delete(0, tk.END)
                    dados_bancarios_entry.insert(0, dados_bancarios_texto)
                    dados_bancarios_entry.config(state='readonly')
                    
                    print(f"DEBUG: Dados bancários preenchidos: {dados_bancarios_texto}")
                    
                except Exception as e:
                    print(f"DEBUG: Erro ao preencher dados: {str(e)}")
                    import traceback
                    traceback.print_exc()
            
            def navegar_lista_com_teclado(event):
                """Navega na lista com teclado e seleciona com Enter"""
                try:
                    if event.keysym == 'Down':
                        current = lista_sugestoes.curselection()
                        if current:
                            next_index = min(current[0] + 1, lista_sugestoes.size() - 1)
                        else:
                            next_index = 0
                        lista_sugestoes.selection_clear(0, tk.END)
                        lista_sugestoes.selection_set(next_index)
                        lista_sugestoes.see(next_index)
                        return "break"
                        
                    elif event.keysym == 'Up':
                        current = lista_sugestoes.curselection()
                        if current:
                            next_index = max(current[0] - 1, 0)
                        else:
                            next_index = 0
                        lista_sugestoes.selection_clear(0, tk.END)
                        lista_sugestoes.selection_set(next_index)
                        lista_sugestoes.see(next_index)
                        return "break"
                        
                    elif event.keysym == 'Return':
                        selecionar_fornecedor_da_lista()
                        return "break"
                        
                except Exception as e:
                    print(f"DEBUG: Erro na navegação: {str(e)}")
            
            def ocultar_sugestoes_ao_sair_foco(event=None):
                """Oculta sugestões quando sai do foco - MELHORADO"""
                # Verificar se o foco foi para a lista de sugestões
                try:
                    widget_foco = janela_novo.focus_get()
                    if widget_foco == lista_sugestoes:
                        return  # Não ocultar se o foco foi para a lista
                    
                    # Delay maior para permitir clique
                    janela_novo.after(300, lambda: lista_sugestoes.grid_remove())
                except:
                    janela_novo.after(300, lambda: lista_sugestoes.grid_remove())
            
            # Bindings para busca
            cnpj_cpf.bind('<KeyRelease>', buscar_fornecedor_por_cnpj)
            cnpj_cpf.bind('<FocusOut>', buscar_fornecedor_por_cnpj)
            
            nome.bind('<KeyRelease>', buscar_fornecedor_por_nome)
            nome.bind('<FocusOut>', ocultar_sugestoes_ao_sair_foco)
            
            # Navegação por teclado no campo nome quando há sugestões
            nome.bind('<Down>', lambda e: lista_sugestoes.focus() if lista_sugestoes.winfo_viewable() else None)
            
            # Bindings da lista - CORRIGIDOS
            lista_sugestoes.bind('<Button-1>', selecionar_por_clique)  # Clique simples
            lista_sugestoes.bind('<Double-Button-1>', selecionar_fornecedor_da_lista)  # Duplo clique
            lista_sugestoes.bind('<KeyPress>', navegar_lista_com_teclado)  # Navegação por teclado
            
            # Permitir que a lista receba foco
            lista_sugestoes.bind('<FocusIn>', lambda e: print("DEBUG: Lista recebeu foco"))
            lista_sugestoes.bind('<FocusOut>', ocultar_sugestoes_ao_sair_foco)
            
            # Atualizar dados bancários quando CNPJ for alterado E quando forma de pagamento mudar
            # cnpj_cpf.bind('<FocusOut>', lambda e: janela_confirm.after(200, atualizar_dados_bancarios))

            # Categoria
            ttk.Label(frame_fornecedor, text="Categoria:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
            categoria = ttk.Entry(frame_fornecedor, width=20)
            categoria.grid(row=2, column=1, padx=5, pady=5, sticky='w')
            
            # Dados bancários
            ttk.Label(frame_fornecedor, text="Dados Bancários:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
            dados_bancarios_entry = ttk.Entry(frame_fornecedor, width=40, state='readonly')
            dados_bancarios_entry.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
            
            frame_fornecedor.columnconfigure(1, weight=1)
            
            # === SEÇÃO: DESPESA ===
            frame_despesa = ttk.LabelFrame(main_frame, text="Dados da Despesa")
            frame_despesa.pack(fill='x', pady=(0, 10))
            
            # Tipo de despesa
            ttk.Label(frame_despesa, text="Tipo Despesa:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            tp_desp = ttk.Combobox(frame_despesa, values=['1', '2', '3', '4', '5', '6'], state='readonly', width=5)
            tp_desp.set('3')  # Padrão
            tp_desp.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            
            # Referência
            ttk.Label(frame_despesa, text="Referência:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
            referencia = ttk.Entry(frame_despesa, width=30)
            referencia.grid(row=0, column=3, padx=5, pady=5, sticky='ew')
            
            # Valor unitário
            ttk.Label(frame_despesa, text="Valor Unitário:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
            vr_unit = ttk.Entry(frame_despesa, width=15)
            vr_unit.grid(row=1, column=1, padx=5, pady=5, sticky='w')
            
            # Dias
            ttk.Label(frame_despesa, text="Dias:").grid(row=1, column=2, padx=5, pady=5, sticky='w')
            dias = ttk.Entry(frame_despesa, width=8)
            dias.insert(0, "1")
            dias.grid(row=1, column=3, padx=5, pady=5, sticky='w')
            
            # Valor total (calculado automaticamente)
            ttk.Label(frame_despesa, text="Valor Total:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
            valor_total = ttk.Entry(frame_despesa, width=15, state='readonly')
            valor_total.grid(row=2, column=1, padx=5, pady=5, sticky='w')
            
            def calcular_valor_total_automatico(event=None):
                try:
                    vr_unit_val = float(vr_unit.get().replace(',', '.')) if vr_unit.get() else 0
                    dias_val = float(dias.get().replace(',', '.')) if dias.get() else 1
                    total = vr_unit_val * dias_val
                    
                    valor_total.config(state='normal')
                    valor_total.delete(0, tk.END)
                    valor_total.insert(0, f"{total:.2f}".replace('.', ','))
                    valor_total.config(state='readonly')
                except ValueError:
                    pass
            
            vr_unit.bind('<KeyRelease>', calcular_valor_total_automatico)
            dias.bind('<KeyRelease>', calcular_valor_total_automatico)
            
            # Data de vencimento
            ttk.Label(frame_despesa, text="Data Vencimento:").grid(row=2, column=2, padx=5, pady=5, sticky='w')
            dt_vencto = DateEntry(frame_despesa, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
            dt_vencto.grid(row=2, column=3, padx=5, pady=5, sticky='w')
            
            # NF
            ttk.Label(frame_despesa, text="NF:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
            nf = ttk.Entry(frame_despesa, width=15)
            nf.grid(row=3, column=1, padx=5, pady=5, sticky='w')
            
            # Observação
            ttk.Label(frame_despesa, text="Observação:").grid(row=3, column=2, padx=5, pady=5, sticky='w')
            observacao = ttk.Entry(frame_despesa, width=30)
            observacao.grid(row=3, column=3, padx=5, pady=5, sticky='ew')
            
            frame_despesa.columnconfigure(3, weight=1)
            
            # === BOTÕES ===
            frame_botoes = ttk.Frame(main_frame)
            frame_botoes.pack(fill='x', pady=(15, 0))
            
            def salvar_lancamento():
                try:
                    # Validações básicas
                    if not cnpj_cpf.get().strip():
                        custom_messagebox("error", "Erro", "CNPJ/CPF é obrigatório!")
                        cnpj_cpf.focus()
                        return
                    
                    if not nome.get().strip():
                        custom_messagebox("error", "Erro", "Nome é obrigatório!")
                        nome.focus()
                        return
                    
                    if not vr_unit.get().strip():
                        custom_messagebox("error", "Erro", "Valor unitário é obrigatório!")
                        vr_unit.focus()
                        return
                    
                    # Preparar dados do lançamento
                    dados_lancamento = {
                        'data_rel': data_rel.get_date(),
                        'tp_desp': tp_desp.get(),
                        'cnpj_cpf': cnpj_cpf.get().strip(),
                        'nome': nome.get().strip().upper(),
                        'categoria': categoria.get().strip().upper() or 'FORNECEDOR',
                        'referencia': referencia.get().strip().upper(),
                        'nf': nf.get().strip().upper(),
                        'valor': float(valor_total.get().replace(',', '.')),
                        'vr_unit': float(vr_unit.get().replace(',', '.')),
                        'dias': float(dias.get().replace(',', '.')) if dias.get() else 1,
                        'dt_vencto': dt_vencto.get_date(),
                        'observacao': observacao.get().strip().upper(),
                        'etapa_obra': '',  # Valores padrão para campos não presentes
                        'insumo': '',
                        'dados_bancarios': dados_bancarios_entry.get(),
                        'forma_pagamento': 'PIX'  # Padrão
                    }
                    
                    # Inserir lançamento
                    sucesso = self.sistema.inserir_lancamento_completo(dados_lancamento)
                    
                    if sucesso:
                        custom_messagebox("info", "Sucesso", "Lançamento criado com sucesso!")
                        janela_novo.destroy()
                        self.carregar_dados_agenda()  # Recarregar agenda
                    else:
                        custom_messagebox("error", "Erro", "Erro ao criar lançamento!")
                        
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao salvar lançamento: {str(e)}")
                    print(f"DEBUG: Erro ao salvar lançamento direto: {str(e)}")
            
            ttk.Button(frame_botoes, text="Salvar Lançamento", 
                    command=salvar_lancamento).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Cancelar", 
                    command=janela_novo.destroy).pack(side='left', padx=5)
            
            # Focar no primeiro campo
            cnpj_cpf.focus()
            
            print("DEBUG: Janela de novo lançamento aberta")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao criar janela de lançamento: {str(e)}")
            print(f"DEBUG: Erro ao criar janela de lançamento: {str(e)}")
        
    def editar_selecionado(self):
        """Edita apenas a data de vencimento do item selecionado"""
        try:
            selected = self.tree_agenda.selection()
            if not selected:
                custom_messagebox("warning", "Aviso", "Selecione um item para editar")
                return
            
            valores = self.tree_agenda.item(selected[0], 'values')
            
            if len(valores) < 8:
                custom_messagebox("error", "Erro", "Dados incompletos no item selecionado")
                return
            
            id_origem = valores[7]  # ID_Origem
            status = valores[1]     # Status
            fornecedor = valores[2] # Fornecedor
            vencimento_atual = valores[0]  # Data de vencimento atual
            
            # Verificar se é um lançamento existente (ID numérico)
            if not id_origem or id_origem.strip() == "" or id_origem.startswith(('CONFIG_', 'BASICO_', 'REC_', 'TPL_', 'HIST_')):
                custom_messagebox("info", "Edição", 
                    "Este é um compromisso configurado.\n"
                    "Use 'Confirmar Lançamento' para criar o lançamento real primeiro.")
                return
            
            try:
                # Verificar se é um ID numérico (lançamento real)
                id_numerico = int(float(id_origem))
            except (ValueError, TypeError):
                custom_messagebox("info", "Edição", 
                    "Este item não pode ser editado.\n"
                    "Use 'Confirmar Lançamento' para criar o lançamento real.")
                return
            
            # Abrir dialog de edição da data de vencimento
            self.abrir_editor_vencimento(id_numerico, fornecedor, vencimento_atual, selected[0])
            
        except Exception as e:
            print(f"DEBUG: Erro ao editar selecionado: {str(e)}")
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao editar item: {str(e)}")

    def abrir_editor_vencimento(self, id_lancamento, fornecedor, vencimento_atual, tree_item):
        """Versão simplificada e robusta do editor de vencimento"""
        try:
            print(f"DEBUG: Abrindo editor para ID {id_lancamento}")
            
            # Criar janela simples
            dialog = tk.Toplevel(self.janela)
            dialog.title("Editar Data de Vencimento")
            dialog.geometry("350x320")
            dialog.transient(self.janela)
            dialog.grab_set()
            
            # Frame principal
            frame = tk.Frame(dialog, padx=20, pady=20)
            frame.pack(fill='both', expand=True)
            
            # Informações
            tk.Label(frame, text=f"ID: {id_lancamento}", font=('Arial', 11, 'bold')).pack(anchor='w', pady=2)
            tk.Label(frame, text=f"Fornecedor: {fornecedor}").pack(anchor='w', pady=2)
            
            # Separador
            tk.Label(frame, text="").pack(pady=5)
            
            # Data atual
            tk.Label(frame, text=f"Data atual: {vencimento_atual}", font=('Arial', 10, 'bold')).pack(anchor='w')
            
            # Espaço
            tk.Label(frame, text="").pack(pady=5)
            
            # Nova data
            tk.Label(frame, text="Nova data (dd/mm/yyyy):").pack(anchor='w')
            entry_data = tk.Entry(frame, font=('Arial', 12), width=12)
            entry_data.pack(anchor='w', pady=5)
            entry_data.insert(0, vencimento_atual)
            
            # Instrução
            tk.Label(frame, text="Exemplo: 25/12/2025", font=('Arial', 9), fg='gray').pack(anchor='w')
            
            # Espaço
            tk.Label(frame, text="").pack(pady=10)
            
            # Botões
            frame_botoes = tk.Frame(frame)
            frame_botoes.pack(fill='x')
            
            def salvar_simples():
                try:
                    nova_data_str = entry_data.get().strip()
                    print(f"DEBUG: Data digitada: {nova_data_str}")
                    
                    # Validar data
                    try:
                        from datetime import datetime
                        nova_data_obj = datetime.strptime(nova_data_str, '%d/%m/%Y').date()
                        nova_data_formatada = nova_data_obj.strftime('%d/%m/%Y')
                        print(f"DEBUG: Data validada: {nova_data_formatada}")
                    except ValueError:
                        import tkinter.messagebox as msg
                        msg.showerror("Erro", "Data inválida! Use o formato dd/mm/yyyy")
                        return
                    
                    # Verificar se mudou
                    if nova_data_formatada == vencimento_atual:
                        import tkinter.messagebox as msg
                        msg.showinfo("Informação", "A data não foi alterada.")
                        return
                    
                    # Confirmar
                    import tkinter.messagebox as msg
                    resposta = msg.askyesno("Confirmar", 
                        f"Alterar data de vencimento:\n\n"
                        f"DE: {vencimento_atual}\n"
                        f"PARA: {nova_data_formatada}\n\n"
                        f"Confirma?")
                    
                    print(f"DEBUG: Resposta: {resposta}")
                    
                    if resposta:
                        # Salvar
                        if self.salvar_nova_data_vencimento(id_lancamento, nova_data_obj):
                            # CORREÇÃO: Atualizar tree ANTES de recarregar
                            valores = list(self.tree_agenda.item(tree_item, 'values'))
                            valores[0] = nova_data_formatada
                            self.tree_agenda.item(tree_item, values=valores)
                            
                            msg.showinfo("Sucesso", f"Data alterada para {nova_data_formatada}")
                            dialog.destroy()
                            
                            # CORREÇÃO: Recarregar dados da agenda para sincronizar
                            self.carregar_dados_agenda()  # Em vez de aplicar_filtros()
                        else:
                            msg.showerror("Erro", "Falha ao salvar na planilha")
                            
                except Exception as e:
                    print(f"DEBUG: Erro ao salvar: {e}")
                    import tkinter.messagebox as msg
                    msg.showerror("Erro", f"Erro: {str(e)}")
            
            def cancelar_simples():
                dialog.destroy()
            
            # Botões
            tk.Button(frame_botoes, text="Cancelar", command=cancelar_simples).pack(side='left', padx=5)
            tk.Button(frame_botoes, text="Salvar", command=salvar_simples).pack(side='left', padx=5)
            
            # Focar no campo
            entry_data.focus()
            entry_data.select_range(0, tk.END)
            
            # Bindings
            entry_data.bind('<Return>', lambda e: salvar_simples())
            dialog.bind('<Escape>', lambda e: cancelar_simples())
            
            print("DEBUG: Editor simples criado")
            
        except Exception as e:
            print(f"DEBUG: Erro ao criar editor: {e}")
            import traceback
            traceback.print_exc()

    def salvar_nova_data_vencimento(self, id_lancamento, nova_data):
        """Salva a nova data de vencimento usando o fluxo correto do sistema"""
        print(f"DEBUG: ==> MÉTODO SALVAR_NOVA_DATA_VENCIMENTO CHAMADO <==")
        print(f"DEBUG: ID: {id_lancamento}, Nova Data: {nova_data}")
        
        try:
            if not self.sistema.cliente_atual:
                print("DEBUG: ERRO - Nenhum cliente selecionado")
                return False
            
            # Importações necessárias
            import os
            from pathlib import Path
            from openpyxl import load_workbook
            from datetime import datetime
            
            # Caminho da planilha
            caminho_planilha = Path(PASTA_CLIENTES) / f"{self.sistema.cliente_atual}.xlsx"
            print(f"DEBUG: Caminho da planilha: {caminho_planilha}")
            
            if not caminho_planilha.exists():
                print(f"DEBUG: ERRO - Planilha não encontrada: {caminho_planilha}")
                return False
            
            print("DEBUG: Planilha encontrada, abrindo...")
            
            # Abrir workbook
            wb = load_workbook(caminho_planilha)
            print(f"DEBUG: Workbook aberto. Abas disponíveis: {wb.sheetnames}")
            
            if 'Dados' not in wb.sheetnames:
                print("DEBUG: ERRO - Aba 'Dados' não encontrada")
                wb.close()
                return False
            
            ws = wb['Dados']
            print(f"DEBUG: Aba 'Dados' selecionada. Max row: {ws.max_row}")
            
            # Procurar o ID na coluna O (posição 15)
            linha_encontrada = None
            print(f"DEBUG: Procurando ID {id_lancamento} na coluna O...")
            
            for row in range(2, ws.max_row + 1):  # Começar da linha 2
                valor_id = ws[f'O{row}'].value  # Coluna O = ID_LANCAMENTO
                
                if valor_id is not None:
                    try:
                        id_convertido = int(valor_id)
                        id_procurado = int(id_lancamento)
                        
                        if id_convertido == id_procurado:
                            linha_encontrada = row
                            print(f"DEBUG: *** ID {id_lancamento} ENCONTRADO na linha {row} ***")
                            break
                            
                    except (ValueError, TypeError) as e:
                        continue
            
            if not linha_encontrada:
                print(f"DEBUG: ERRO - ID {id_lancamento} NÃO ENCONTRADO")
                wb.close()
                return False
            
            # Mostrar valor atual da coluna J antes da alteração
            valor_atual = ws[f'J{linha_encontrada}'].value
            print(f"DEBUG: Valor atual da célula J{linha_encontrada}: {valor_atual}")
            
            # CORREÇÃO: Salvar como objeto datetime, não como string
            if isinstance(nova_data, str):
                nova_data_obj = datetime.strptime(nova_data, '%d/%m/%Y')
            else:
                nova_data_obj = datetime.combine(nova_data, datetime.min.time())
            
            print(f"DEBUG: Atualizando J{linha_encontrada} com objeto datetime: {nova_data_obj}")
            
            # Salvar como datetime e aplicar formato
            celula = ws[f'J{linha_encontrada}']
            celula.value = nova_data_obj
            celula.number_format = 'DD/MM/YYYY'  # Aplicar formato de data
            
            # Verificar se a atualização funcionou
            valor_depois = ws[f'J{linha_encontrada}'].value
            print(f"DEBUG: Valor após atualização: {valor_depois}")
            
            # Salvar o arquivo
            print("DEBUG: Salvando workbook...")
            wb.save(caminho_planilha)
            wb.close()
            
            print("DEBUG: *** SUCESSO - Data de vencimento alterada com sucesso ***")
            return True
            
        except Exception as e:
            print(f"DEBUG: *** ERRO CRÍTICO ao salvar nova data: {str(e)} ***")
            import traceback
            traceback.print_exc()
            return False

    def confirmar_lancamento(self):
        """Confirma um lançamento pendente, convertendo-o em lançamento real"""
        selected = self.tree_agenda.selection()
        if not selected:
            custom_messagebox("warning", "Aviso", "Selecione um item para confirmar")
            return
        
        valores = self.tree_agenda.item(selected[0], 'values')
        status = valores[1]
        
        if status != 'PENDENTE':
            custom_messagebox("info", "Informação", "Apenas itens pendentes podem ser confirmados")
            return
        
        # Abrir interface de confirmação/criação de lançamento
        self.abrir_confirmacao_lancamento(valores)
    
    def limpar_duplicacoes_inteligente(self):
        """Remove compromissos configurados que já foram lançados (baseado em DATA_REL)"""
        try:
            if not custom_messagebox("yesno", "Limpeza Inteligente", 
                "Esta função vai remover da visualização os compromissos configurados "
                "que já possuem lançamentos correspondentes no mesmo relatório (DATA_REL).\n\n"
                "Isso deixará a agenda mais limpa, mostrando apenas:\n"
                "• Lançamentos reais\n" 
                "• Compromissos que ainda precisam ser lançados\n\n"
                "Continuar?"):
                return
            
            # Identificar duplicações baseadas em DATA_REL
            removidos = 0
            agenda_limpa = []
            
            for item in self.dados_agenda:
                if item['origem'] in ['CONFIGURACAO', 'BASICO']:
                    # VERIFICAÇÃO CORRIGIDA: Verificar se existe lançamento real 
                    # para a mesma DATA_REL e fornecedor
                    data_rel_item = item.get('data_rel')
                    
                    existe_real = any(
                        outro['origem'] == 'EXISTENTE' and
                        outro.get('data_rel') == data_rel_item and
                        item['fornecedor'].upper() in outro['fornecedor'].upper()
                        for outro in self.dados_agenda
                    )
                    
                    if existe_real:
                        removidos += 1
                        print(f"DEBUG: Removido compromisso duplicado: {item['fornecedor']} - REL {data_rel_item}")
                        continue
                
                agenda_limpa.append(item)
            
            # Atualizar dados
            self.dados_agenda = agenda_limpa
            
            # Reaplicar filtros
            self.aplicar_filtros()
            
            custom_messagebox("info", "Limpeza Concluída", 
                f"Foram removidos {removidos} compromissos configurados que "
                f"já possuem lançamentos no mesmo relatório (DATA_REL).\n\n"
                f"A agenda agora mostra apenas itens únicos por relatório.")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro na limpeza: {str(e)}")

    def resetar_agenda(self):
        """Reseta e recarrega todos os dados da agenda"""
        try:
            if custom_messagebox("yesno", "Resetar Agenda", 
                "Isso vai recarregar todos os dados da agenda do zero.\n"
                "Útil se os dados estiverem inconsistentes.\n\n"
                "Continuar?"):
                
                # Limpar cache
                self.dados_agenda = []
                
                # Recarregar tudo
                self.carregar_dados_agenda()
                
                custom_messagebox("info", "Sucesso", "Agenda resetada e recarregada!")
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao resetar agenda: {str(e)}")

    def abrir_gerenciador_compromissos(self):
        """Abre o gerenciador de compromissos recorrentes integrado à agenda"""
        try:
            from src.config.utils import custom_messagebox
            # Criar janela de gerenciamento
            janela_compromissos = tk.Toplevel(self.janela)
            janela_compromissos.title("Gerenciar Compromissos Recorrentes")
            janela_compromissos.geometry("900x700")
            janela_compromissos.transient(self.janela)
            janela_compromissos.grab_set()
            
            # Frame principal
            main_frame = ttk.Frame(janela_compromissos, padding="10")
            main_frame.pack(fill='both', expand=True)
            
            # Título
            ttk.Label(main_frame, text="Gerenciador de Compromissos Recorrentes", 
                    font=('TkDefaultFont', 14, 'bold')).pack(pady=(0, 15))
            
            # Frame dividido em duas colunas
            frame_conteudo = ttk.Frame(main_frame)
            frame_conteudo.pack(fill='both', expand=True)
            
            # === COLUNA ESQUERDA: LISTA DE COMPROMISSOS ===
            frame_esquerda = ttk.LabelFrame(frame_conteudo, text="Compromissos Cadastrados")
            frame_esquerda.pack(side='left', fill='both', expand=True, padx=(0, 10))
            
            # Treeview para compromissos
            colunas = ('Nome', 'Dia', 'Recorrência', 'Valor', 'Status')
            tree_compromissos = ttk.Treeview(frame_esquerda, columns=colunas, show='headings', height=20)
            
            # Configurar cabeçalhos
            for col in colunas:
                tree_compromissos.heading(col, text=col)
            
            # Configurar larguras
            tree_compromissos.column('Nome', width=180)
            tree_compromissos.column('Dia', width=60, anchor='center')
            tree_compromissos.column('Recorrência', width=100, anchor='center')
            tree_compromissos.column('Valor', width=100, anchor='e')
            tree_compromissos.column('Status', width=80, anchor='center')
            
            # Scrollbar
            scrollbar = ttk.Scrollbar(frame_esquerda, orient='vertical', command=tree_compromissos.yview)
            tree_compromissos.configure(yscrollcommand=scrollbar.set)
            
            tree_compromissos.pack(side='left', fill='both', expand=True, padx=5, pady=5)
            scrollbar.pack(side='right', fill='y', pady=5)
            
            # === COLUNA DIREITA: FORMULÁRIOS ===
            frame_direita = ttk.Frame(frame_conteudo)
            frame_direita.pack(side='right', fill='y', padx=(10, 0))
            
            # Frame para novo compromisso
            frame_novo = ttk.LabelFrame(frame_direita, text="Novo Compromisso")
            frame_novo.pack(fill='x', pady=(0, 10))
            
            # Campos do formulário
            campos_novo = {}
            
            # Nome
            ttk.Label(frame_novo, text="Nome:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            campos_novo['nome'] = ttk.Entry(frame_novo, width=25)
            campos_novo['nome'].grid(row=0, column=1, padx=5, pady=5)
            
            # Dia vencimento
            ttk.Label(frame_novo, text="Dia Vencimento:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
            campos_novo['dia'] = ttk.Spinbox(frame_novo, from_=1, to=31, width=10)
            campos_novo['dia'].set('5')
            campos_novo['dia'].grid(row=1, column=1, padx=5, pady=5, sticky='w')
            
            # Recorrência
            ttk.Label(frame_novo, text="Recorrência:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
            campos_novo['recorrencia'] = ttk.Combobox(frame_novo, 
                                                    values=['mensal', 'trimestral', 'semestral', 'anual'],
                                                    state='readonly', width=22)
            campos_novo['recorrencia'].set('mensal')
            campos_novo['recorrencia'].grid(row=2, column=1, padx=5, pady=5)
            
            # Categoria
            ttk.Label(frame_novo, text="Categoria:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
            categorias = ['ADM', 'DIV', 'LOC', 'MAT', 'MO', 'SERV', 'TP']
            campos_novo['categoria'] = ttk.Combobox(frame_novo, values=categorias, state='readonly', width=22)
            campos_novo['categoria'].set('MO')
            campos_novo['categoria'].grid(row=3, column=1, padx=5, pady=5)
            
            # Tipo despesa
            ttk.Label(frame_novo, text="Tipo Despesa:").grid(row=4, column=0, padx=5, pady=5, sticky='w')
            campos_novo['tipo_despesa'] = ttk.Combobox(frame_novo, values=['2', '3', '5', '6', '7'], 
                                                    state='readonly', width=10)
            campos_novo['tipo_despesa'].set('3')
            campos_novo['tipo_despesa'].grid(row=4, column=1, padx=5, pady=5, sticky='w')
            
            # Valor estimado
            ttk.Label(frame_novo, text="Valor Estimado:").grid(row=5, column=0, padx=5, pady=5, sticky='w')
            campos_novo['valor'] = ttk.Entry(frame_novo, width=15)
            campos_novo['valor'].insert(0, "0,00")
            campos_novo['valor'].grid(row=5, column=1, padx=5, pady=5, sticky='w')
            
            # Observação
            ttk.Label(frame_novo, text="Observação:").grid(row=6, column=0, padx=5, pady=5, sticky='w')
            campos_novo['observacao'] = ttk.Entry(frame_novo, width=25)
            campos_novo['observacao'].grid(row=6, column=1, padx=5, pady=5)
            
            # Frame para editar compromisso
            frame_editar = ttk.LabelFrame(frame_direita, text="Editar Compromisso Selecionado")
            frame_editar.pack(fill='x', pady=(0, 10))
            
            # Campos de edição (similares aos de criação)
            campos_editar = {}
            
            ttk.Label(frame_editar, text="Nome:").grid(row=0, column=0, padx=5, pady=3, sticky='w')
            campos_editar['nome'] = ttk.Entry(frame_editar, width=25)
            campos_editar['nome'].grid(row=0, column=1, padx=5, pady=3)
            
            ttk.Label(frame_editar, text="Dia:").grid(row=1, column=0, padx=5, pady=3, sticky='w')
            campos_editar['dia'] = ttk.Spinbox(frame_editar, from_=1, to=31, width=10)
            campos_editar['dia'].grid(row=1, column=1, padx=5, pady=3, sticky='w')
            
            ttk.Label(frame_editar, text="Valor:").grid(row=2, column=0, padx=5, pady=3, sticky='w')
            campos_editar['valor'] = ttk.Entry(frame_editar, width=15)
            campos_editar['valor'].grid(row=2, column=1, padx=5, pady=3, sticky='w')
            
            # === FUNÇÕES LOCAIS ===
            
            def carregar_compromissos():
                """Carrega compromissos no treeview"""
                # Limpar tree
                for item in tree_compromissos.get_children():
                    tree_compromissos.delete(item)
                
                # Carregar configurações
                try:
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    compromissos = GerenciadorConfiguracoes.get_compromissos_recorrentes_todos()
                    
                    for comp in compromissos:
                        status = "ATIVO" if comp.get('ativo', True) else "INATIVO"
                        tag = 'ativo' if comp.get('ativo', True) else 'inativo'
                        
                        tree_compromissos.insert('', 'end',
                            values=(
                                comp['nome'],
                                comp['dia_vencimento'],
                                comp['recorrencia'],
                                f"R$ {comp['valor_estimado']:.2f}",
                                status
                            ),
                            tags=(tag,)
                        )
                    
                    # Configurar cores
                    tree_compromissos.tag_configure('ativo', background='#e8f5e8')
                    tree_compromissos.tag_configure('inativo', background='#ffe4e1')
                    
                except Exception as e:
                    print(f"Erro ao carregar compromissos: {e}")
            
            def on_select_compromisso(event):
                """Evento de seleção de compromisso"""
                selecionado = tree_compromissos.selection()
                if not selecionado:
                    return
                
                valores = tree_compromissos.item(selecionado[0])['values']
                
                # Preencher campos de edição
                campos_editar['nome'].delete(0, tk.END)
                campos_editar['nome'].insert(0, valores[0])
                
                campos_editar['dia'].delete(0, tk.END)
                campos_editar['dia'].insert(0, valores[1])
                
                # Extrair valor numérico
                valor_str = valores[3].replace('R$ ', '').replace('.', '').replace(',', '.')
                campos_editar['valor'].delete(0, tk.END)
                campos_editar['valor'].insert(0, f"{float(valor_str):.2f}".replace('.', ','))
            
            def adicionar_compromisso():
                """Adiciona novo compromisso"""
                try:
                    # Validações
                    nome = campos_novo['nome'].get().strip().upper()
                    if not nome:
                        custom_messagebox("error", "Erro", "Nome é obrigatório!")
                        return
                    
                    # Coletar dados
                    dia = int(campos_novo['dia'].get())
                    recorrencia = campos_novo['recorrencia'].get()
                    categoria = campos_novo['categoria'].get()
                    tipo_despesa = int(campos_novo['tipo_despesa'].get())
                    
                    valor_str = campos_novo['valor'].get().replace(',', '.')
                    valor = float(valor_str) if valor_str else 0.0
                    
                    observacao = campos_novo['observacao'].get().strip()
                    
                    # Salvar na configuração
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    config = GerenciadorConfiguracoes.carregar_configuracoes()
                    
                    if 'compromissos_recorrentes' not in config:
                        config['compromissos_recorrentes'] = {'lista': [], 'historico_alteracoes': []}
                    
                    # Verificar duplicatas
                    if any(c['nome'] == nome for c in config['compromissos_recorrentes']['lista']):
                        custom_messagebox("error", "Erro", "Já existe um compromisso com este nome!")
                        return
                    
                    # Adicionar compromisso
                    novo_compromisso = {
                        'nome': nome,
                        'dia_vencimento': dia,
                        'recorrencia': recorrencia,
                        'valor_estimado': valor,
                        'categoria': categoria,
                        'tipo_despesa': tipo_despesa,
                        'ativo': True,
                        'observacao': observacao
                    }
                    
                    config['compromissos_recorrentes']['lista'].append(novo_compromisso)
                    
                    # Salvar configuração
                    from pathlib import Path
                    import json
                    config_path = GerenciadorConfiguracoes.CONFIG_PATH
                    with open(config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=4, ensure_ascii=False)
                    
                    # Atualizar cache
                    GerenciadorConfiguracoes._atualizar_cache(config)
                    
                    # Limpar campos
                    for campo in campos_novo.values():
                        if hasattr(campo, 'delete'):
                            campo.delete(0, tk.END)
                        elif hasattr(campo, 'set'):
                            campo.set('')
                    
                    # Recarregar lista
                    carregar_compromissos()
                    
                    custom_messagebox("info", "Sucesso", "Compromisso adicionado com sucesso!")
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao adicionar compromisso: {str(e)}")
            
            def salvar_alteracoes():
                """Salva alterações no compromisso selecionado"""
                selecionado = tree_compromissos.selection()
                if not selecionado:
                    custom_messagebox("warning", "Aviso", "Selecione um compromisso para editar!")
                    return
                
                try:
                    nome_original = tree_compromissos.item(selecionado[0])['values'][0]
                    
                    # Carregar configuração
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    config = GerenciadorConfiguracoes.carregar_configuracoes()
                    
                    # Encontrar e atualizar compromisso
                    for comp in config['compromissos_recorrentes']['lista']:
                        if comp['nome'] == nome_original:
                            comp['nome'] = campos_editar['nome'].get().strip().upper()
                            comp['dia_vencimento'] = int(campos_editar['dia'].get())
                            
                            valor_str = campos_editar['valor'].get().replace(',', '.')
                            comp['valor_estimado'] = float(valor_str) if valor_str else 0.0
                            break
                    
                    # Salvar configuração
                    import json
                    config_path = GerenciadorConfiguracoes.CONFIG_PATH
                    with open(config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=4, ensure_ascii=False)
                    
                    # Atualizar cache
                    GerenciadorConfiguracoes._atualizar_cache(config)
                    
                    # Recarregar
                    carregar_compromissos()
                    
                    custom_messagebox("info", "Sucesso", "Compromisso atualizado com sucesso!")
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao salvar: {str(e)}")
            
            def toggle_status():
                """Ativa/desativa compromisso selecionado"""
                selecionado = tree_compromissos.selection()
                if not selecionado:
                    custom_messagebox("warning", "Aviso", "Selecione um compromisso!")
                    return
                
                try:
                    nome = tree_compromissos.item(selecionado[0])['values'][0]
                    
                    # Carregar e alterar configuração
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    config = GerenciadorConfiguracoes.carregar_configuracoes()
                    
                    for comp in config['compromissos_recorrentes']['lista']:
                        if comp['nome'] == nome:
                            comp['ativo'] = not comp.get('ativo', True)
                            status = "ativado" if comp['ativo'] else "desativado"
                            break
                    
                    # Salvar
                    import json
                    config_path = GerenciadorConfiguracoes.CONFIG_PATH
                    with open(config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=4, ensure_ascii=False)
                    
                    GerenciadorConfiguracoes._atualizar_cache(config)
                    carregar_compromissos()
                    
                    custom_messagebox("info", "Sucesso", f"Compromisso {status}!")
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao alterar status: {str(e)}")
            
            def remover_compromisso():
                """Remove compromisso selecionado"""
                selecionado = tree_compromissos.selection()
                if not selecionado:
                    custom_messagebox("warning", "Aviso", "Selecione um compromisso para remover!")
                    return
                
                nome = tree_compromissos.item(selecionado[0])['values'][0]
                
                if not custom_messagebox("yesno", "Confirmar", f"Deseja remover o compromisso '{nome}'?"):
                    return
                
                try:
                    # Remover da configuração
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    config = GerenciadorConfiguracoes.carregar_configuracoes()
                    
                    config['compromissos_recorrentes']['lista'] = [
                        c for c in config['compromissos_recorrentes']['lista'] 
                        if c['nome'] != nome
                    ]
                    
                    # Salvar
                    import json
                    config_path = GerenciadorConfiguracoes.CONFIG_PATH
                    with open(config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=4, ensure_ascii=False)
                    
                    GerenciadorConfiguracoes._atualizar_cache(config)
                    carregar_compromissos()
                    
                    # Limpar campos de edição
                    for campo in campos_editar.values():
                        campo.delete(0, tk.END)
                    
                    custom_messagebox("info", "Sucesso", "Compromisso removido com sucesso!")
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao remover: {str(e)}")
            
            def fechar_e_recarregar():
                """Fecha o gerenciador e recarrega a agenda"""
                janela_compromissos.destroy()
                # Recarregar dados da agenda para refletir alterações
                self.carregar_dados_agenda()
            
            # === BOTÕES ===
            
            # Botões para novo compromisso
            ttk.Button(frame_novo, text="Adicionar Compromisso",
                    command=adicionar_compromisso).grid(row=7, column=0, columnspan=2, pady=10)
            
            # Botões para edição
            frame_botoes_edicao = ttk.Frame(frame_editar)
            frame_botoes_edicao.grid(row=3, column=0, columnspan=2, pady=10)
            
            ttk.Button(frame_botoes_edicao, text="Salvar",
                    command=salvar_alteracoes).pack(side='left', padx=2)
            ttk.Button(frame_botoes_edicao, text="Ativar/Desativar",
                    command=toggle_status).pack(side='left', padx=2)
            ttk.Button(frame_botoes_edicao, text="Remover",
                    command=remover_compromisso).pack(side='left', padx=2)
            
            # Botões inferiores
            frame_botoes_inferior = ttk.Frame(main_frame)
            frame_botoes_inferior.pack(fill='x', pady=(15, 0))
            
            ttk.Button(frame_botoes_inferior, text="Fechar e Recarregar Agenda",
                    command=fechar_e_recarregar).pack(side='right', padx=5)
            
            # === EVENTOS ===
            tree_compromissos.bind('<<TreeviewSelect>>', on_select_compromisso)
            
            # Carregar dados iniciais
            carregar_compromissos()
            
            print("DEBUG: Gerenciador de compromissos aberto com sucesso")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir gerenciador: {str(e)}")
            print(f"DEBUG: Erro ao abrir gerenciador de compromissos: {str(e)}")

    def importar_excel(self):
        """Importa agenda de arquivo Excel"""
        try:
            arquivo = filedialog.askopenfilename(
                title="Selecionar arquivo de agenda",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not arquivo:
                return
            
            # Implementar lógica de importação
            custom_messagebox("info", "Importação", "Funcionalidade de importação será implementada")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao importar: {str(e)}")
    
    def exportar_periodo(self):
        """Exporta período atual para Excel com mais detalhes"""
        try:
            # Preparar dados para exportação
            dados_exportacao = []
            
            for child in self.tree_agenda.get_children():
                valores = self.tree_agenda.item(child, 'values')
                
                # Buscar dados originais para mais detalhes
                item_original = None
                for item in self.dados_agenda:
                    if item['vencimento'].strftime('%d/%m/%Y') == valores[0] and item['fornecedor'] == valores[3]:
                        item_original = item
                        break
                
                linha_excel = {
                    'Vencimento': valores[0],
                    'Status': valores[1],
                    'Cliente': valores[2],
                    'Fornecedor': valores[3],
                    'Referência': valores[4],
                    'Valor': valores[5],
                    'Tipo': valores[6],
                    'Observação': valores[7],
                    'Origem': item_original['origem'] if item_original else 'N/A',
                    'ID_Sistema': valores[8] if len(valores) > 8 else ''
                }
                
                dados_exportacao.append(linha_excel)
            
            if not dados_exportacao:
                custom_messagebox("warning", "Aviso", "Nenhum dado para exportar no período atual")
                return
            
            # Salvar arquivo
            from tkinter import filedialog
            data_atual = datetime.now().strftime('%Y%m%d_%H%M')
            nome_padrao = f"Agenda_{self.sistema.cliente_atual}_{data_atual}.xlsx"
            
            arquivo = filedialog.asksaveasfilename(
                title="Salvar agenda",
                defaultextension=".xlsx",
                initialvalue=nome_padrao,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if arquivo:
                df_export = pd.DataFrame(dados_exportacao)
                
                # Adicionar resumo
                total_valor = sum(float(d['Valor'].replace('R$ ', '').replace('.', '').replace(',', '.')) 
                                for d in dados_exportacao)
                
                resumo = {
                    'Total_Itens': len(dados_exportacao),
                    'Total_Valor': f"R$ {total_valor:,.2f}",
                    'Periodo_Inicio': self.data_inicio_personalizada.get(),
                    'Periodo_Fim': self.data_fim_personalizada.get(),
                    'Data_Exportacao': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    'Cliente': self.sistema.cliente_atual
                }
                
                with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
                    df_export.to_excel(writer, sheet_name='Agenda', index=False)
                    
                    # Adicionar resumo em aba separada
                    df_resumo = pd.DataFrame([resumo])
                    df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                
                custom_messagebox("info", "Sucesso", f"Agenda exportada para:\n{arquivo}")
                    
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao exportar: {str(e)}")
    
    def verificar_alertas_inteligentes(self):
        """Verifica alertas baseados em padrões reais"""
        alertas = []
        hoje = datetime.now().date()
        
        try:
            # Alertar sobre vencimentos de hoje
            venc_hoje = [item for item in self.dados_agenda if item['vencimento'] == hoje]
            if venc_hoje:
                valor_hoje = sum(item['valor'] for item in venc_hoje)
                alertas.append({
                    'tipo': 'URGENTE',
                    'titulo': f'{len(venc_hoje)} vencimento(s) HOJE',
                    'descricao': f'Total: R$ {valor_hoje:,.2f}',
                    'itens': [f"• {item['fornecedor']}" for item in venc_hoje]
                })
            
            # Alertar sobre vencidos
            vencidos = [item for item in self.dados_agenda if item['vencimento'] < hoje and item['status'] != 'LANÇADO']
            if vencidos:
                valor_vencido = sum(item['valor'] for item in vencidos)
                alertas.append({
                    'tipo': 'ATENCAO',
                    'titulo': f'{len(vencidos)} item(ns) vencido(s)',
                    'descricao': f'Total: R$ {valor_vencido:,.2f}',
                    'itens': [f"• {item['fornecedor']} ({item['vencimento'].strftime('%d/%m')})" for item in vencidos[:5]]
                })
            
            # Alertar sobre alto volume nos próximos 7 dias
            proximos_7_dias = hoje + timedelta(days=7)
            proximos = [item for item in self.dados_agenda 
                    if hoje < item['vencimento'] <= proximos_7_dias]
            
            if proximos:
                valor_proximo = sum(item['valor'] for item in proximos)
                if valor_proximo > 50000:  # Configurável
                    alertas.append({
                        'tipo': 'INFO',
                        'titulo': f'Alto volume próximos 7 dias',
                        'descricao': f'R$ {valor_proximo:,.2f} em {len(proximos)} itens',
                        'itens': []
                    })
            
            return alertas
            
        except Exception as e:
            print(f"DEBUG: Erro ao verificar alertas: {str(e)}")
            return []

    def mostrar_alertas_se_necessario(self):
        """Mostra alertas inteligentes se houver itens importantes"""
        alertas = self.verificar_alertas_inteligentes()
        
        if not alertas:
            return
        
        # Mostrar apenas se houver alertas urgentes ou de atenção
        alertas_importantes = [a for a in alertas if a['tipo'] in ['URGENTE', 'ATENCAO']]
        
        if not alertas_importantes:
            return
        
        # Criar janela de alerta discreta
        janela_alerta = tk.Toplevel(self.janela)
        janela_alerta.title("Alertas da Agenda")
        janela_alerta.geometry("400x300")
        janela_alerta.attributes('-topmost', True)
        
        frame_main = ttk.Frame(janela_alerta, padding="10")
        frame_main.pack(fill='both', expand=True)
        
        ttk.Label(frame_main, text="Alertas da Agenda", 
                font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 10))
        
        for alerta in alertas_importantes:
            frame_alerta = ttk.LabelFrame(frame_main, text=alerta['titulo'])
            frame_alerta.pack(fill='x', pady=5)
            
            ttk.Label(frame_alerta, text=alerta['descricao']).pack(anchor='w', padx=5, pady=2)
            
            for item in alerta.get('itens', [])[:3]:  # Máximo 3 itens
                ttk.Label(frame_alerta, text=item, font=('TkDefaultFont', 8)).pack(anchor='w', padx=15, pady=1)
        
        # Botões
        frame_botoes = ttk.Frame(frame_main)
        frame_botoes.pack(fill='x', pady=(10, 0))
        
        ttk.Button(frame_botoes, text="OK", command=janela_alerta.destroy).pack(side='right')

    def abrir_confirmacao_lancamento(self, valores_item):
        """Janela de confirmação com DATA_REL preservada"""
        
        # Buscar dados do item na agenda
        item_agenda = None
        for item in self.dados_agenda:
            if (item['fornecedor'] == valores_item[2] and
                item['vencimento'].strftime('%d/%m/%Y') == valores_item[0]):
                item_agenda = item
                break
        
        # Janela de confirmação
        janela_confirm = tk.Toplevel(self.janela)
        janela_confirm.title("Confirmar Lançamento")
        janela_confirm.geometry("700x650")
        janela_confirm.transient(self.janela)
        janela_confirm.grab_set()
        
        # Frame principal
        main_frame = ttk.Frame(janela_confirm, padding="15")
        main_frame.pack(fill='both', expand=True)
        
        # Informações do item
        frame_info = ttk.LabelFrame(main_frame, text="Dados da Agenda")
        frame_info.pack(fill='x', pady=(0, 10))
        
        ttk.Label(frame_info, text=f"Vencimento: {valores_item[0]}", 
                font=('TkDefaultFont', 10, 'bold')).pack(anchor='w', padx=10, pady=2)
        ttk.Label(frame_info, text=f"Fornecedor: {valores_item[2]}").pack(anchor='w', padx=10, pady=2)
        ttk.Label(frame_info, text=f"Referência: {valores_item[3]}").pack(anchor='w', padx=10, pady=2)
        ttk.Label(frame_info, text=f"Valor: {valores_item[4]}").pack(anchor='w', padx=10, pady=2)
        
        # Formulário de lançamento
        frame_form = ttk.LabelFrame(main_frame, text="Dados do Lançamento")
        frame_form.pack(fill='both', expand=True, pady=(0, 10))
        
        # === DATA DO RELATÓRIO - DESTACADA ===
        frame_data_rel = ttk.Frame(frame_form)
        frame_data_rel.grid(row=0, column=0, columnspan=4, padx=5, pady=10, sticky='ew')
        
        ttk.Label(frame_data_rel, text="📋 Data do Relatório:", 
                font=('TkDefaultFont', 10, 'bold'), 
                foreground='#0066cc').pack(side='left', padx=(0, 5))
        data_rel = DateEntry(frame_data_rel, width=12, date_pattern='dd/mm/yyyy', 
                        locale='pt_BR', state='readonly')
        
        # Pré-preencher com DATA_REL da agenda
        if item_agenda and 'data_rel' in item_agenda:
            data_rel.set_date(item_agenda['data_rel'])
        else:
            data_rel.set_date(self.calcular_data_rel())
        
        data_rel.pack(side='left')
        
        ttk.Label(frame_data_rel, text="(Dia do relatório - fixo)", 
                font=('TkDefaultFont', 8), foreground='gray').pack(side='left', padx=(5, 0))
        
        # Tipo de despesa
        ttk.Label(frame_form, text="Tipo Despesa:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        tp_desp = ttk.Combobox(frame_form, values=['2', '3', '5', '6', '7'], state='readonly', width=5)
        
        if item_agenda and item_agenda.get('dados_originais'):
            tp_desp.set(str(item_agenda['dados_originais'].get('tipo_despesa', 3)))
        else:
            tp_desp.set('3')
        tp_desp.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        # === DATA DE VENCIMENTO - EDITÁVEL ===
        frame_dt_vencto = ttk.Frame(frame_form)
        frame_dt_vencto.grid(row=1, column=2, columnspan=2, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame_dt_vencto, text="📅 Data Vencimento Real:", 
                font=('TkDefaultFont', 9)).pack(side='left', padx=(0, 5))
        dt_vencto = DateEntry(frame_dt_vencto, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        dt_vencto.set_date(datetime.strptime(valores_item[0], '%d/%m/%Y').date())
        dt_vencto.pack(side='left')
        
        ttk.Label(frame_dt_vencto, text="(editável)", 
                font=('TkDefaultFont', 8), foreground='green').pack(side='left', padx=(5, 0))
        
        # === SEÇÃO DE FORNECEDOR COM BUSCA ===
        
        # CNPJ/CPF
        ttk.Label(frame_form, text="CNPJ/CPF:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        cnpj_cpf = ttk.Entry(frame_form, width=20)
        cnpj_cpf.grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky='ew')
        
        # Nome/Fornecedor
        ttk.Label(frame_form, text="Nome:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
        nome = ttk.Entry(frame_form, width=40)
        nome.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky='ew')
        
        # Lista de sugestões
        lista_sugestoes = tk.Listbox(frame_form, height=4)
        lista_sugestoes.grid(row=4, column=1, columnspan=3, padx=5, pady=5, sticky='ew')
        lista_sugestoes.grid_remove()
        
        # === FUNÇÕES DE BUSCA (manter as mesmas do código original) ===
        
        def buscar_fornecedor_por_cnpj(event=None):
            cnpj_digitado = cnpj_cpf.get().strip()
            if len(cnpj_digitado) >= 11:
                try:
                    fornecedor_dados = self.sistema.buscar_fornecedor_por_cnpj_agenda(cnpj_digitado)
                    if fornecedor_dados:
                        preencher_dados_fornecedor(fornecedor_dados)
                except Exception as e:
                    print(f"DEBUG: Erro ao buscar fornecedor: {e}")
        
        def buscar_fornecedor_por_nome(event=None):
            nome_digitado = nome.get().strip()
            if len(nome_digitado) < 3:
                lista_sugestoes.grid_remove()
                return
            try:
                fornecedores = self.sistema.buscar_fornecedores_por_nome_parcial(nome_digitado)
                if fornecedores:
                    lista_sugestoes.delete(0, tk.END)
                    for f in fornecedores[:10]:
                        lista_sugestoes.insert(tk.END, f"{f['nome']} - {f['cnpj_cpf']}")
                    lista_sugestoes.grid()
                else:
                    lista_sugestoes.grid_remove()
            except Exception as e:
                print(f"DEBUG: Erro: {e}")
                lista_sugestoes.grid_remove()
        
        def selecionar_fornecedor_da_lista(event=None):
            try:
                selection = lista_sugestoes.curselection()
                if selection:
                    texto = lista_sugestoes.get(selection[0])
                    cnpj_extraido = texto.split(' - ')[-1]
                    fornecedor_dados = self.sistema.buscar_fornecedor_por_cnpj_agenda(cnpj_extraido)
                    if fornecedor_dados:
                        preencher_dados_fornecedor(fornecedor_dados)
                        lista_sugestoes.grid_remove()
            except Exception as e:
                print(f"DEBUG: Erro: {e}")
        
        def preencher_dados_fornecedor(fornecedor_dados):
            """Preenche campos com dados do fornecedor - SEM duplicação"""
            try:
                print(f"DEBUG preencher_dados_fornecedor CHAMADA")
                print(f"DEBUG: Dados recebidos: {fornecedor_dados}")
                
                # CRÍTICO: Limpar campos ANTES de preencher
                cnpj_cpf.delete(0, tk.END)
                nome.delete(0, tk.END)
                
                # Preencher CNPJ/CPF
                cnpj_valor = fornecedor_dados.get('cnpj_cpf', '').strip()
                print(f"DEBUG: Preenchendo CNPJ: '{cnpj_valor}'")
                cnpj_cpf.insert(0, cnpj_valor)
                
                # Preencher Nome
                nome_valor = fornecedor_dados.get('nome', '').strip()
                print(f"DEBUG: Preenchendo Nome: '{nome_valor}'")
                nome.insert(0, nome_valor)
                
                # Aguardar um momento e então atualizar dados bancários
                janela_confirm.after(200, atualizar_dados_bancarios)
                
                print(f"DEBUG: Campos preenchidos com sucesso")
                
            except Exception as e:
                print(f"DEBUG: ERRO em preencher_dados_fornecedor: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # Bindings
        cnpj_cpf.bind('<KeyRelease>', buscar_fornecedor_por_cnpj)
        nome.bind('<KeyRelease>', buscar_fornecedor_por_nome)
        lista_sugestoes.bind('<Button-1>', lambda e: janela_confirm.after(50, selecionar_fornecedor_da_lista))
        lista_sugestoes.bind('<Double-Button-1>', selecionar_fornecedor_da_lista)
        
        # Preencher dados iniciais
        print(f"DEBUG: Tentando buscar fornecedor: {valores_item[2]}")

        # Limpar campos antes de preencher
        cnpj_cpf.delete(0, tk.END)
        nome.delete(0, tk.END)

        try:
            # Tentar buscar fornecedor pelo nome da agenda
            fornecedor_dados = self.sistema.buscar_fornecedor_por_nome_agenda(valores_item[2])
            
            if fornecedor_dados:
                print(f"DEBUG: Fornecedor encontrado: {fornecedor_dados}")
                preencher_dados_fornecedor(fornecedor_dados)
            else:
                print(f"DEBUG: Fornecedor não encontrado, preenchendo com dados da agenda")
                # Se não encontrar, preencher apenas o nome (sem CNPJ)
                nome.insert(0, valores_item[2])
                # Deixar CNPJ vazio para o usuário preencher
                custom_messagebox("warning", "Atenção", 
                    f"Fornecedor '{valores_item[2]}' não encontrado no cadastro.\n"
                    f"Por favor, preencha o CNPJ/CPF manualmente.")
                cnpj_cpf.focus()
                
        except Exception as e:
            print(f"DEBUG: Erro ao buscar fornecedor: {str(e)}")
            import traceback
            traceback.print_exc()
            # Em caso de erro, preencher apenas o nome
            nome.insert(0, valores_item[2])
            custom_messagebox("error", "Erro", 
                f"Erro ao buscar dados do fornecedor: {str(e)}\n"
                f"Preencha os dados manualmente.")
            cnpj_cpf.focus()
        
        # Referência
        ttk.Label(frame_form, text="Referência:").grid(row=5, column=0, padx=5, pady=5, sticky='w')
        referencia = ttk.Entry(frame_form, width=40)
        referencia.insert(0, valores_item[3])
        referencia.grid(row=5, column=1, columnspan=3, padx=5, pady=5, sticky='ew')
        
        # NF
        ttk.Label(frame_form, text="NF:").grid(row=6, column=0, padx=5, pady=5, sticky='w')
        nf = ttk.Entry(frame_form, width=15)
        nf.grid(row=6, column=1, padx=5, pady=5, sticky='w')
        
        # Valor
        ttk.Label(frame_form, text="Valor:").grid(row=6, column=2, padx=5, pady=5, sticky='w')
        valor = ttk.Entry(frame_form, width=15)
        valor_numerico = valores_item[4].replace('R$ ', '').replace('.', '').replace(',', '.')
        valor.insert(0, valor_numerico)
        valor.grid(row=6, column=3, padx=5, pady=5, sticky='w')
        
        # Observação
        ttk.Label(frame_form, text="Observação:").grid(row=7, column=0, padx=5, pady=5, sticky='w')
        observacao = ttk.Entry(frame_form, width=40)
        if item_agenda and item_agenda.get('dados_originais'):
            obs = item_agenda['dados_originais'].get('observacao', '')
            observacao.insert(0, f"{obs} - CONFIRMADO DA AGENDA")
        else:
            observacao.insert(0, "CONFIRMADO DA AGENDA")
        observacao.grid(row=7, column=1, columnspan=3, padx=5, pady=5, sticky='ew')
        
        # Forma de pagamento
        ttk.Label(frame_form, text="Forma Pagamento:").grid(row=8, column=0, padx=5, pady=5, sticky='w')
        forma_pagamento = ttk.Combobox(frame_form, values=['PIX', 'TED', 'DINHEIRO'], 
                                    state='readonly', width=15)
        forma_pagamento.set('PIX')
        forma_pagamento.grid(row=8, column=1, padx=5, pady=5, sticky='w')

        # Dados bancários (atualizado dinamicamente)
        ttk.Label(frame_form, text="Dados Bancários:").grid(row=8, column=2, padx=5, pady=5, sticky='w')
        dados_bancarios_entry = ttk.Entry(frame_form, width=40, state='readonly')
        dados_bancarios_entry.grid(row=8, column=3, padx=5, pady=5, sticky='ew')

        # Função para atualizar dados bancários quando mudar fornecedor ou forma de pagamento
        def atualizar_dados_bancarios():
            """Atualiza dados bancários baseado no CNPJ e forma de pagamento"""
            try:
                cnpj_atual = cnpj_cpf.get().strip()
                forma_atual = forma_pagamento.get()
                
                print(f"DEBUG: Atualizando dados bancários - CNPJ: {cnpj_atual}, Forma: {forma_atual}")
                
                # Validar CNPJ antes de buscar
                if not cnpj_atual or len(cnpj_atual) < 11:
                    print(f"DEBUG: CNPJ inválido ou vazio: {cnpj_atual}")
                    dados_bancarios_entry.config(state='normal')
                    dados_bancarios_entry.delete(0, tk.END)
                    dados_bancarios_entry.insert(0, "PREENCHA O CNPJ/CPF PRIMEIRO")
                    dados_bancarios_entry.config(state='readonly')
                    return
                
                # Buscar dados bancários usando o método do sistema
                dados_banc = self.sistema.obter_dados_bancarios_fornecedor(
                    cnpj_atual, 
                    forma_pagamento_preferida=forma_atual
                )
                
                print(f"DEBUG: Dados bancários obtidos: {dados_banc}")
                
                # Atualizar campo
                dados_bancarios_entry.config(state='normal')
                dados_bancarios_entry.delete(0, tk.END)
                dados_bancarios_entry.insert(0, dados_banc)
                dados_bancarios_entry.config(state='readonly')
                
            except Exception as e:
                print(f"DEBUG: Erro ao atualizar dados bancários: {str(e)}")
                import traceback
                traceback.print_exc()
                dados_bancarios_entry.config(state='normal')
                dados_bancarios_entry.delete(0, tk.END)
                dados_bancarios_entry.insert(0, f"ERRO: {str(e)}")
                dados_bancarios_entry.config(state='readonly')

        # Atualizar quando mudar a forma de pagamento
        forma_pagamento.bind('<<ComboboxSelected>>', lambda e: atualizar_dados_bancarios())

        frame_form.columnconfigure(1, weight=1)
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x')
        
        def confirmar():
            try:
                if not nome.get().strip() or not valor.get().strip() or not cnpj_cpf.get().strip():
                    custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
                    return
                
                # CRÍTICO: Usar DATA_REL do item da agenda
                data_rel_usar = data_rel.get_date()
                if item_agenda and 'data_rel' in item_agenda:
                    data_rel_usar = item_agenda['data_rel']
                    print(f"DEBUG: Usando DATA_REL da agenda: {data_rel_usar}")
                
                dados_lancamento = {
                    'data_rel': data_rel_usar,
                    'tp_desp': tp_desp.get(),
                    'cnpj_cpf': cnpj_cpf.get().strip(),
                    'nome': nome.get().strip().upper(),
                    'referencia': referencia.get().strip().upper(),
                    'nf': nf.get().strip().upper(),
                    'valor': float(valor.get().replace(',', '.')),
                    'dt_vencto': dt_vencto.get_date(),
                    'observacao': observacao.get().strip().upper(),
                    'forma_pagamento': forma_pagamento.get(), 
                    'dados_bancarios': dados_bancarios_entry.get()
                }
                
                print(f"DEBUG: Lançamento - DATA_REL: {data_rel_usar} | DT_VENCTO: {dt_vencto.get_date()}")
                
                sucesso = self.sistema.inserir_lancamento_completo(dados_lancamento)
                
                if sucesso:
                    custom_messagebox("info", "Sucesso", 
                        f"Lançamento confirmado!\n"
                        f"Relatório: {data_rel_usar.strftime('%d/%m/%Y')}\n"
                        f"Vencimento: {dt_vencto.get_date().strftime('%d/%m/%Y')}")
                    janela_confirm.destroy()
                    self.carregar_dados_agenda()
                else:
                    custom_messagebox("error", "Erro", "Erro ao inserir lançamento!")
                    
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro: {str(e)}")
        
        ttk.Button(frame_botoes, text="Confirmar e Lançar", command=confirmar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela_confirm.destroy).pack(side='left', padx=5)
  
    def importar_template_agenda(self):
        """Importa compromissos de um template/arquivo de agenda"""
        try:
            # Verificar se existe arquivo Agenda.xlsx
            arquivo_agenda = PASTA_CLIENTES / "Agenda.xlsx"
            
            if not arquivo_agenda.exists():
                print("DEBUG: Arquivo Agenda.xlsx não encontrado")
                return
            
            # Carregar dados do template
            df_agenda = pd.read_excel(arquivo_agenda)
            df_agenda = df_agenda.fillna("")
            
            hoje = datetime.now().date()
            
            # Processar cada linha do template
            for idx, row in df_agenda.iterrows():
                try:
                    nome = row.get('FORNECEDOR', '').strip()
                    if not nome:
                        continue
                    
                    # Extrair informações de vencimento (assumindo colunas como RELATÓRIO DIA 05, DIA 20, etc.)
                    for coluna in df_agenda.columns:
                        if 'DIA' in str(coluna).upper() and row.get(coluna) == 'X':
                            # Extrair dia do nome da coluna
                            dia_vencimento = None
                            if 'DIA 05' in str(coluna).upper() or 'DIA.05' in str(coluna).upper():
                                dia_vencimento = 5
                            elif 'DIA 20' in str(coluna).upper() or 'DIA.20' in str(coluna).upper():
                                dia_vencimento = 20
                            
                            if dia_vencimento:
                                # Gerar compromissos futuros para os próximos meses
                                for mes_futuro in range(3):  # Próximos 3 meses
                                    data_compromisso = hoje + relativedelta(months=mes_futuro)
                                    try:
                                        data_compromisso = data_compromisso.replace(day=dia_vencimento)
                                    except ValueError:
                                        # Se dia não existe no mês, usar último dia
                                        ultimo_dia = calendar.monthrange(data_compromisso.year, data_compromisso.month)[1]
                                        data_compromisso = data_compromisso.replace(day=min(dia_vencimento, ultimo_dia))
                                    
                                    if data_compromisso > hoje:
                                        # Verificar se já não existe
                                        ja_existe = any(
                                            item['vencimento'] == data_compromisso and 
                                            nome.upper() in item['fornecedor'].upper()
                                            for item in self.dados_agenda
                                        )
                                        
                                        if not ja_existe:
                                            item_agenda = {
                                                'vencimento': data_compromisso,
                                                'status': 'PENDENTE',
                                                'cliente': self.sistema.cliente_atual,
                                                'fornecedor': nome,
                                                'referencia': f"{nome} - {data_compromisso.strftime('%m/%Y')}",
                                                'valor': 0,
                                                'tipo': 'TEMPLATE',
                                                'observacao': f"Template - Venc. dia {dia_vencimento}",
                                                'id_origem': f"TPL_{nome.replace(' ', '_')}_{data_compromisso.strftime('%Y%m%d')}",
                                                'origem': 'TEMPLATE'
                                            }
                                            
                                            self.dados_agenda.append(item_agenda)
                
                except Exception as e:
                    print(f"DEBUG: Erro ao processar linha do template: {str(e)}")
                    continue
                    
            print(f"DEBUG: Template de agenda processado com sucesso")
            
        except Exception as e:
            print(f"DEBUG: Erro ao importar template de agenda: {str(e)}")
    
    def analisar_padroes_historicos(self):
        """Analisa padrões históricos para sugerir compromissos futuros"""
        try:
            # Esta função pode analisar o histórico de lançamentos do cliente
            # para identificar padrões recorrentes e sugerir compromissos futuros
            
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            if not arquivo_cliente.exists():
                return
            
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")
            
            # Analisar últimos 6 meses para identificar padrões
            seis_meses_atras = datetime.now().date() - relativedelta(months=6)
            
            # Agrupar por fornecedor e mês para identificar recorrências
            padroes_encontrados = {}
            
            for idx, row in df.iterrows():
                try:
                    dt_vencto = pd.to_datetime(row['DT_VENCTO']).date()
                    if dt_vencto < seis_meses_atras:
                        continue
                    
                    fornecedor = row.get('NOME', '').strip().upper()
                    if not fornecedor:
                        continue
                    
                    mes_ano = dt_vencto.strftime('%Y-%m')
                    dia_vencimento = dt_vencto.day
                    valor = float(row.get('VALOR', 0))
                    
                    chave = f"{fornecedor}_{dia_vencimento}"
                    
                    if chave not in padroes_encontrados:
                        padroes_encontrados[chave] = {
                            'fornecedor': fornecedor,
                            'dia_vencimento': dia_vencimento,
                            'ocorrencias': [],
                            'valores': []
                        }
                    
                    padroes_encontrados[chave]['ocorrencias'].append(mes_ano)
                    padroes_encontrados[chave]['valores'].append(valor)
                    
                except Exception as e:
                    continue
            
            # Identificar padrões realmente recorrentes (pelo menos 3 ocorrências)
            hoje = datetime.now().date()
            
            for chave, padrao in padroes_encontrados.items():
                if len(padrao['ocorrencias']) >= 3:
                    # É um padrão recorrente, gerar compromissos futuros
                    valor_medio = sum(padrao['valores']) / len(padrao['valores'])
                    
                    # Gerar para próximos 2 meses
                    for mes_futuro in range(1, 3):
                        data_futura = hoje + relativedelta(months=mes_futuro)
                        try:
                            data_compromisso = data_futura.replace(day=padrao['dia_vencimento'])
                        except ValueError:
                            ultimo_dia = calendar.monthrange(data_futura.year, data_futura.month)[1]
                            data_compromisso = data_futura.replace(day=min(padrao['dia_vencimento'], ultimo_dia))
                        
                        # Verificar se já não existe
                        ja_existe = any(
                            item['vencimento'] == data_compromisso and 
                            padrao['fornecedor'] in item['fornecedor'].upper()
                            for item in self.dados_agenda
                        )
                        
                        if not ja_existe:
                            item_agenda = {
                                'vencimento': data_compromisso,
                                'status': 'PENDENTE',
                                'cliente': self.sistema.cliente_atual,
                                'fornecedor': padrao['fornecedor'],
                                'referencia': f"{padrao['fornecedor']} - {data_compromisso.strftime('%m/%Y')}",
                                'valor': valor_medio,
                                'tipo': 'HISTÓRICO',
                                'observacao': f"Baseado em padrão histórico ({len(padrao['ocorrencias'])} ocorrências)",
                                'id_origem': f"HIST_{chave}_{data_compromisso.strftime('%Y%m%d')}",
                                'origem': 'HISTÓRICO'
                            }
                            
                            self.dados_agenda.append(item_agenda)
            
            print(f"DEBUG: Análise de padrões históricos concluída")
            
        except Exception as e:
            print(f"DEBUG: Erro na análise de padrões históricos: {str(e)}")
    
    def on_double_click(self, event):
        """Trata duplo clique nos itens"""
        selected = self.tree_agenda.selection()
        if selected:
            valores = self.tree_agenda.item(selected[0], 'values')
            status = valores[1]
            
            if status == 'PENDENTE':
                self.confirmar_lancamento()
            elif status in ['LANÇADO', 'VENCIDO', 'VENCE HOJE']:
                self.editar_selecionado()
    
    def configurar_atalhos(self):
        """Configura atalhos de teclado para a agenda"""
        try:
            # Enter: Confirmar lançamento (se pendente) ou editar (se lançado)
            def on_enter(event):
                selected = self.tree_agenda.selection()
                if selected:
                    valores = self.tree_agenda.item(selected[0], 'values')
                    status = valores[1]
                    
                    if status == 'PENDENTE':
                        self.confirmar_lancamento()
                    else:
                        self.editar_selecionado()
                return "break"
            
            self.tree_agenda.bind('<Return>', on_enter)
            self.janela.bind('<Return>', on_enter)
            
            # F5: Atualizar
            def on_f5(event):
                self.carregar_dados_agenda()
                return "break"
            
            self.janela.bind('<F5>', on_f5)
            self.tree_agenda.bind('<F5>', on_f5)
            
            # Ctrl+N: Novo lançamento
            def on_ctrl_n(event):
                self.novo_lancamento()
                return "break"
            
            self.janela.bind('<Control-n>', on_ctrl_n)
            
            # Escape: Fechar
            def on_escape(event):
                self.janela.destroy()
                return "break"
            
            self.janela.bind('<Escape>', on_escape)
            
            print("DEBUG: Atalhos da agenda configurados")
            print("       Enter: Confirmar/Editar item selecionado")
            print("       F5: Atualizar agenda")
            print("       Ctrl+N: Novo lançamento")
            print("       Escape: Fechar")
            
        except Exception as e:
            print(f"Erro ao configurar atalhos da agenda: {str(e)}")

class ConfiguradorAgenda:
    """Classe para configurar templates e padrões da agenda"""
    
    @staticmethod
    def criar_template_agenda():
        """Cria template inicial da agenda baseado na imagem"""
        try:
            arquivo_template = PASTA_CLIENTES / "Agenda.xlsx"
            
            # Dados baseados na imagem fornecida
            template_data = [
                {'FORNECEDOR': 'FOLHA DP', 'DIA_05': 'X', 'DIA_20': '', 'CATEGORIA': 'FOLHA_PAGAMENTO'},
                {'FORNECEDOR': 'MOTORISTA', 'DIA_05': 'X', 'DIA_20': '', 'CATEGORIA': 'FOLHA_PAGAMENTO'},
                {'FORNECEDOR': 'ELETRICISTA', 'DIA_05': 'X', 'DIA_20': '', 'CATEGORIA': 'FOLHA_PAGAMENTO'},
                {'FORNECEDOR': 'MISS SST', 'DIA_05': '', 'DIA_20': 'X', 'CATEGORIA': 'SERVICOS'},
                {'FORNECEDOR': 'FGTS', 'DIA_05': 'X', 'DIA_20': '', 'CATEGORIA': 'TRIBUTOS'},
                {'FORNECEDOR': 'DETRAN', 'DIA_05': '', 'DIA_20': 'X', 'CATEGORIA': 'SERVICOS'},
                {'FORNECEDOR': 'ADMINISTRAÇÃO', 'DIA_05': '', 'DIA_20': 'X', 'CATEGORIA': 'TAXA_ADM'},
                {'FORNECEDOR': 'FESTA SAFRA', 'DIA_05': '', 'DIA_20': '', 'CATEGORIA': 'EVENTOS'},
                {'FORNECEDOR': 'PROJ SEGURANÇA', 'DIA_05': '', 'DIA_20': 'X', 'CATEGORIA': 'PROJETOS'},
                {'FORNECEDOR': 'MAQUINAS', 'DIA_05': '', 'DIA_20': 'X', 'CATEGORIA': 'EQUIPAMENTOS'},
            ]
            
            df_template = pd.DataFrame(template_data)
            df_template.to_excel(arquivo_template, index=False)
            
            print(f"✅ Template da agenda criado: {arquivo_template}")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao criar template: {str(e)}")
            return False
    
    @staticmethod
    def configurar_compromissos_personalizados():
        """Interface para configurar compromissos personalizados por cliente"""
        
        janela_config = tk.Toplevel()
        janela_config.title("Configurar Compromissos Personalizados")
        janela_config.geometry("800x600")
        
        # Frame principal
        main_frame = ttk.Frame(janela_config, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Lista de compromissos
        frame_lista = ttk.LabelFrame(main_frame, text="Compromissos Configurados")
        frame_lista.pack(fill='both', expand=True, pady=(0, 10))
        
        # Treeview
        colunas = ('Fornecedor', 'Dia_Vencimento', 'Recorrencia', 'Valor_Estimado', 'Categoria')
        tree_config = ttk.Treeview(frame_lista, columns=colunas, show='headings', height=15)
        
        for col in colunas:
            tree_config.heading(col, text=col.replace('_', ' '))
            tree_config.column(col, width=150)
        
        tree_config.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Frame de edição
        frame_edicao = ttk.LabelFrame(main_frame, text="Adicionar/Editar Compromisso")
        frame_edicao.pack(fill='x', pady=(0, 10))
        
        # Campos de entrada
        ttk.Label(frame_edicao, text="Fornecedor:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        entry_fornecedor = ttk.Entry(frame_edicao, width=30)
        entry_fornecedor.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        
        ttk.Label(frame_edicao, text="Dia Vencimento:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        combo_dia = ttk.Combobox(frame_edicao, values=[str(i) for i in range(1, 32)], width=5)
        combo_dia.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame_edicao, text="Categoria:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        combo_categoria = ttk.Combobox(frame_edicao, values=[
            'FOLHA_PAGAMENTO', 'TRIBUTOS', 'SERVICOS', 'EQUIPAMENTOS', 
            'PROJETOS', 'EVENTOS', 'TAXA_ADM', 'OUTROS'
        ], width=20)
        combo_categoria.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        ttk.Label(frame_edicao, text="Valor Estimado:").grid(row=1, column=2, padx=5, pady=5, sticky='w')
        entry_valor = ttk.Entry(frame_edicao, width=15)
        entry_valor.grid(row=1, column=3, padx=5, pady=5, sticky='w')
        
        frame_edicao.columnconfigure(1, weight=1)
        
        # Botões
        frame_botoes_config = ttk.Frame(frame_edicao)
        frame_botoes_config.grid(row=2, column=0, columnspan=4, pady=10)
        
        ttk.Button(frame_botoes_config, text="Adicionar", 
                  command=lambda: None).pack(side='left', padx=5)
        ttk.Button(frame_botoes_config, text="Atualizar", 
                  command=lambda: None).pack(side='left', padx=5)
        ttk.Button(frame_botoes_config, text="Remover", 
                  command=lambda: None).pack(side='left', padx=5)
        
        # Botões principais
        frame_botoes_main = ttk.Frame(main_frame)
        frame_botoes_main.pack(fill='x')
        
        ttk.Button(frame_botoes_main, text="Salvar Configurações", 
                  command=lambda: None).pack(side='left', padx=5)
        ttk.Button(frame_botoes_main, text="Fechar", 
                  command=janela_config.destroy).pack(side='right', padx=5)

class NotificadorAgenda:
    """Sistema de notificações para a agenda"""
    
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.notificacoes_ativas = []
    
    def verificar_vencimentos_hoje(self):
        """Verifica vencimentos do dia atual"""
        try:
            if not self.sistema.cliente_atual:
                return []
            
            hoje = datetime.now().date()
            vencimentos_hoje = []
            
            # Carregar dados do cliente
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            if arquivo_cliente.exists():
                df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
                df = df.fillna("")
                
                for idx, row in df.iterrows():
                    if row.get('STATUS', 'ATIVO') == 'EXCLUIDO':
                        continue
                    
                    try:
                        dt_vencto = pd.to_datetime(row['DT_VENCTO']).date()
                        if dt_vencto == hoje:
                            vencimentos_hoje.append({
                                'fornecedor': row.get('NOME', ''),
                                'referencia': row.get('REFERÊNCIA', ''),
                                'valor': float(row.get('VALOR', 0)),
                                'observacao': row.get('OBSERVAÇÃO', '')
                            })
                    except:
                        continue
            
            return vencimentos_hoje
            
        except Exception as e:
            print(f"DEBUG: Erro ao verificar vencimentos: {str(e)}")
            return []
    
    def verificar_vencimentos_proximos(self, dias_antecedencia=3):
        """Verifica vencimentos nos próximos dias"""
        try:
            if not self.sistema.cliente_atual:
                return []
            
            hoje = datetime.now().date()
            data_limite = hoje + timedelta(days=dias_antecedencia)
            vencimentos_proximos = []
            
            # Carregar dados do cliente
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            if arquivo_cliente.exists():
                df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
                df = df.fillna("")
                
                for idx, row in df.iterrows():
                    if row.get('STATUS', 'ATIVO') == 'EXCLUIDO':
                        continue
                    
                    try:
                        dt_vencto = pd.to_datetime(row['DT_VENCTO']).date()
                        if hoje < dt_vencto <= data_limite:
                            dias_restantes = (dt_vencto - hoje).days
                            vencimentos_proximos.append({
                                'fornecedor': row.get('NOME', ''),
                                'referencia': row.get('REFERÊNCIA', ''),
                                'valor': float(row.get('VALOR', 0)),
                                'vencimento': dt_vencto,
                                'dias_restantes': dias_restantes
                            })
                    except:
                        continue
            
            return sorted(vencimentos_proximos, key=lambda x: x['vencimento'])
            
        except Exception as e:
            print(f"DEBUG: Erro ao verificar vencimentos próximos: {str(e)}")
            return []
    
    def mostrar_notificacao_vencimentos(self):
        """Mostra notificação com vencimentos do dia e próximos"""
        vencimentos_hoje = self.verificar_vencimentos_hoje()
        vencimentos_proximos = self.verificar_vencimentos_proximos()
        
        if not vencimentos_hoje and not vencimentos_proximos:
            return
        
        # Janela de notificação
        janela_notif = tk.Toplevel(self.sistema.root)
        janela_notif.title("Notificações de Vencimento")
        janela_notif.geometry("500x400")
        janela_notif.attributes('-topmost', True)
        
        # Frame principal
        main_frame = ttk.Frame(janela_notif, padding="15")
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(main_frame, text=f"📅 Agenda - {self.sistema.cliente_atual}", 
                 font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 10))
        
        # Vencimentos de hoje
        if vencimentos_hoje:
            frame_hoje = ttk.LabelFrame(main_frame, text="🚨 Vencimentos de HOJE")
            frame_hoje.pack(fill='x', pady=(0, 10))
            
            valor_total_hoje = sum(v['valor'] for v in vencimentos_hoje)
            
            ttk.Label(frame_hoje, text=f"Total: R$ {valor_total_hoje:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                     font=('TkDefaultFont', 10, 'bold'), foreground='red').pack(anchor='w', padx=10, pady=5)
            
            for venc in vencimentos_hoje[:5]:  # Mostrar no máximo 5
                texto = f"• {venc['fornecedor']} - R$ {venc['valor']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ttk.Label(frame_hoje, text=texto).pack(anchor='w', padx=20, pady=1)
            
            if len(vencimentos_hoje) > 5:
                ttk.Label(frame_hoje, text=f"... e mais {len(vencimentos_hoje) - 5} vencimentos",
                         font=('TkDefaultFont', 9, 'italic')).pack(anchor='w', padx=20, pady=1)
        
        # Vencimentos próximos
        if vencimentos_proximos:
            frame_proximos = ttk.LabelFrame(main_frame, text="⏰ Próximos Vencimentos")
            frame_proximos.pack(fill='x', pady=(0, 10))
            
            for venc in vencimentos_proximos[:5]:  # Mostrar no máximo 5
                dias_texto = "amanhã" if venc['dias_restantes'] == 1 else f"em {venc['dias_restantes']} dias"
                texto = f"• {venc['fornecedor']} - {dias_texto} (R$ {venc['valor']:,.2f})".replace(',', 'X').replace('.', ',').replace('X', '.')
                ttk.Label(frame_proximos, text=texto).pack(anchor='w', padx=20, pady=1)
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=(10, 0))
        
        ttk.Button(frame_botoes, text="Abrir Agenda", 
                  command=lambda: [janela_notif.destroy(), self.sistema.abrir_agenda()]).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Lembrar depois", 
                  command=janela_notif.destroy).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Não mostrar hoje", 
                  command=self.desativar_notificacoes_hoje).pack(side='right', padx=5)
        
        # Posicionar no canto inferior direito
        janela_notif.update_idletasks()
        largura = janela_notif.winfo_width()
        altura = janela_notif.winfo_height()
        pos_x = janela_notif.winfo_screenwidth() - largura - 50
        pos_y = janela_notif.winfo_screenheight() - altura - 100
        janela_notif.geometry(f"+{pos_x}+{pos_y}")
    
    def desativar_notificacoes_hoje(self):
        """Desativa notificações para o dia atual"""
        # Implementar lógica para não mostrar mais notificações hoje
        pass

class RelatorioAgenda:
    """Gerador de relatórios da agenda"""
    
    @staticmethod
    def gerar_relatorio_mensal(cliente, mes, ano):
        """Gera relatório mensal da agenda"""
        try:
            # Carregar dados
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            if not arquivo_cliente.exists():
                return None
            
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")
            
            # Filtrar por mês/ano
            inicio_mes = datetime(ano, mes, 1).date()
            if mes == 12:
                fim_mes = datetime(ano + 1, 1, 1).date() - timedelta(days=1)
            else:
                fim_mes = datetime(ano, mes + 1, 1).date() - timedelta(days=1)
            
            dados_mes = []
            
            for idx, row in df.iterrows():
                try:
                    dt_vencto = pd.to_datetime(row['DT_VENCTO']).date()
                    if inicio_mes <= dt_vencto <= fim_mes:
                        dados_mes.append({
                            'Data_Vencimento': dt_vencto.strftime('%d/%m/%Y'),
                            'Fornecedor': row.get('NOME', ''),
                            'Referencia': row.get('REFERÊNCIA', ''),
                            'Valor': float(row.get('VALOR', 0)),
                            'Status': row.get('STATUS', 'ATIVO'),
                            'Observacao': row.get('OBSERVAÇÃO', '')
                        })
                except:
                    continue
            
            # Gerar relatório Excel
            if dados_mes:
                df_relatorio = pd.DataFrame(dados_mes)
                df_relatorio = df_relatorio.sort_values('Data_Vencimento')
                
                nome_arquivo = f"Agenda_{cliente}_{mes:02d}_{ano}.xlsx"
                caminho_arquivo = PASTA_CLIENTES / nome_arquivo
                
                with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                    df_relatorio.to_excel(writer, sheet_name='Agenda_Mensal', index=False)
                    
                    # Adicionar resumo
                    total_mes = df_relatorio['Valor'].sum()
                    ativos = df_relatorio[df_relatorio['Status'] == 'ATIVO']
                    
                    resumo = pd.DataFrame([
                        {'Metrica': 'Total de Lançamentos', 'Valor': len(df_relatorio)},
                        {'Metrica': 'Lançamentos Ativos', 'Valor': len(ativos)},
                        {'Metrica': 'Valor Total (R$)', 'Valor': f"{total_mes:,.2f}"},
                        {'Metrica': 'Valor Ativos (R$)', 'Valor': f"{ativos['Valor'].sum():,.2f}"}
                    ])
                    
                    resumo.to_excel(writer, sheet_name='Resumo', index=False)
                
                return caminho_arquivo
            
            return None
            
        except Exception as e:
            print(f"DEBUG: Erro ao gerar relatório mensal: {str(e)}")
            return None

class CacheFornecedores:
    """Cache para otimizar buscas de fornecedores"""
    
    def __init__(self):
        self.cache_fornecedores = None
        self.cache_timestamp = None
        self.cache_duracao = 300  # 5 minutos
    
    def carregar_cache_se_necessario(self, arquivo_fornecedores):
        """Carrega cache se necessário ou se arquivo foi modificado"""
        try:
            import os
            from datetime import datetime
            
            agora = datetime.now()
            arquivo_modificado = os.path.getmtime(arquivo_fornecedores)
            
            # Verificar se precisa recarregar
            precisa_recarregar = (
                self.cache_fornecedores is None or
                self.cache_timestamp is None or
                (agora - self.cache_timestamp).seconds > self.cache_duracao or
                arquivo_modificado > self.cache_timestamp.timestamp()
            )
            
            if precisa_recarregar:
                print("DEBUG: Recarregando cache de fornecedores...")
                self.cache_fornecedores = self._carregar_fornecedores(arquivo_fornecedores)
                self.cache_timestamp = agora
                print(f"DEBUG: Cache carregado com {len(self.cache_fornecedores)} fornecedores")
            
            return self.cache_fornecedores
            
        except Exception as e:
            print(f"DEBUG: Erro ao carregar cache: {str(e)}")
            return []
    
    def _carregar_fornecedores(self, arquivo_fornecedores):
        """Carrega todos os fornecedores em memória"""
        fornecedores = []
        
        try:
            wb = load_workbook(arquivo_fornecedores, data_only=True)
            ws = wb['Fornecedores']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[3]:  # Pular se não tem CNPJ ou nome
                    continue
                
                fornecedor = {
                    'cnpj_cpf': str(row[0]).strip(),
                    'nome': str(row[3]).strip().upper(),
                    'categoria': str(row[11] or '').strip(),
                    'banco': str(row[4] or '').strip(),
                    'op': str(row[5] or '').strip(),
                    'agencia': str(row[6] or '').strip(),
                    'conta': str(row[7] or '').strip(),
                    'chave_pix': str(row[8] or '').strip()
                }
                
                fornecedores.append(fornecedor)
            
            wb.close()
            
        except Exception as e:
            print(f"DEBUG: Erro ao carregar fornecedores: {str(e)}")
        
        return fornecedores

class ConfiguracaoTaxas:
    """
    Classe para configurações do sistema de taxas
    """
    # Modo de tratamento de diferenças históricas
    MODO_HISTORICO = "COMPENSATORIO"  # ou "RECALCULO_FORCADO"
    
    # Permitir recálculo de períodos fechados (não recomendado)
    PERMITIR_ALTERACAO_HISTORICO = False
    
    # Dias para considerar uma quinzena como "fechada"
    DIAS_PARA_FECHAR_QUINZENA = 5  # Após 5 dias, considera fechada
    
    @staticmethod
    def quinzena_esta_fechada(data_quinzena):
        """
        Verifica se uma quinzena deve ser considerada fechada
        """
        hoje = datetime.now().date()
        dias_passados = (hoje - data_quinzena).days
        return dias_passados > ConfiguracaoTaxas.DIAS_PARA_FECHAR_QUINZENA
     
if __name__ == "__main__":
    print("Iniciando aplicação...")
    app = SistemaEntradaDados()
    print("Atualizando interface...")
    app.root.update_idletasks()
    print("Iniciando mainloop...")
    app.root.mainloop()
