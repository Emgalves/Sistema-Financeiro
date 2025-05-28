import os
import sys
import traceback

# Configurar diretório de log no desktop do usuário
log_dir = os.path.join(os.path.expanduser("~"), "Desktop")
log_file = os.path.join(log_dir, "sistema_log.txt")

def log_error(message):
    """Grava mensagens de erro em arquivo no desktop"""
    try:
        with open(log_file, "a") as f:
            f.write(f"{message}\n")
    except:
        pass

# Capturar e registrar exceções não tratadas
def handle_exception(exc_type, exc_value, exc_traceback):
    """Manipulador para exceções não capturadas"""
    error_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    log_error(f"\n\n--- ERRO NÃO TRATADO: {error_msg}")
    # Mostrar mensagem simples para o usuário
    import tkinter.messagebox as msgbox
    msgbox.showerror("Erro", f"Ocorreu um erro: {str(exc_value)}\nDetalhes foram registrados em: {log_file}")
    
# Substituir o manipulador de exceções padrão
sys.excepthook = handle_exception

log_error(f"\n\n--- INICIANDO APLICAÇÃO: {sys.argv[0]} ---")

# Imports da biblioteca padrão Python
import os
import sys
from pathlib import Path
import re
from datetime import datetime
from decimal import Decimal

# Imports relacionados ao Tkinter (MUITO IMPORTANTE - NÃO REMOVER)
import tkinter as tk
from tkinter import ttk, messagebox, StringVar
from tkinter import *
from tkcalendar import DateEntry, Calendar
from tkinter import filedialog, messagebox

# Imports para manipulação de dados e Excel
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import openpyxl
import babel
from dateutil.relativedelta import relativedelta

# Detectar modo PyInstaller e ajustar paths
if getattr(sys, 'frozen', False):
    # Estamos em um executável criado pelo PyInstaller
    base_dir = Path(sys._MEIPASS)
    # Garantir que src e src/config estão no path
    for subdir in ['src', os.path.join('src', 'config')]:
        path = os.path.join(base_dir, subdir)
        if path not in sys.path:
            sys.path.insert(0, path)
            print(f"PyInstaller: Adicionando {path} ao sys.path")

# Configurar caminhos de importação
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    config_dir = current_dir / 'config'
    
    # Adicionar todos os caminhos necessários
    for path in [str(current_dir), str(project_root), str(config_dir)]:
        if path not in sys.path:
            sys.path.insert(0, path)
            print(f"Adicionado ao path: {path}")

add_project_root()

# Configurar logging básico para diagnóstico
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger("sistema")

# Importações centralizadas com tratamento de erro
try:
    # Primeiro, tentar importação com 'src.config'
    from src.config.utils import *
    from src.config.dialogs import custom_messagebox
    from src.config.logger_config import system_logger, log_action
    from src.config.window_config import configurar_janela
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH,
        ARQUIVO_FORNECEDORES
    )
    logger.info("Configurações importadas com sucesso via src.config")
except ImportError:
    # Se falhar, tentar importação direta de 'config'
    try:
        from config.utils import *
        from config.logger_config import system_logger, log_action
        from config.window_config import configurar_janela
        from config.config import (
            ARQUIVO_CLIENTES,
            ARQUIVO_MODELO,
            PASTA_CLIENTES,
            BASE_PATH,
            ARQUIVO_FORNECEDORES
        )
        logger.info("Configurações importadas com sucesso via config")
    except ImportError as e:
        # Último recurso - importações relativas
        try:
            from src.config.utils import *
            from src.config.logger_config import system_logger, log_action
            from src.config.window_config import configurar_janela
            from src.config.config import (
                ARQUIVO_CLIENTES,
                ARQUIVO_MODELO,
                PASTA_CLIENTES,
                BASE_PATH,
                ARQUIVO_FORNECEDORES
            )
            logger.info("Configurações importadas com sucesso via .config")
        except Exception as e:
            logger.error(f"Erro ao importar configurações: {str(e)}")
            # Não raise aqui para permitir definições alternativas


# Definir funções de compatibilidade, caso a importação das configurações falhe
def get_categorias_fornecedor():
    """Retorna a lista de categorias de fornecedor"""
    try:
        # Tentar importar do arquivo de configurações
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        return GerenciadorConfiguracoes.get_categorias_fornecedor()
    except:
        # Valores padrão como fallback
        return ['MO', 'MAT', 'SERV', 'DIV', 'ADM', 'LOC', 'TP']

@log_action("Carregar configurações")
def carregar_configuracoes():
    """Carrega as configurações do sistema"""
    try:
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        return GerenciadorConfiguracoes.carregar_configuracoes()
    except:
        # Configurações padrão como fallback
        return {
            'cafe': {'valor_atual': 4.0}
        }

def get_bancos():
    """Retorna a lista de bancos"""
    try:
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        return GerenciadorConfiguracoes.get_bancos()
    except:
        # Valores padrão como fallback
        return [
            '104 - CAIXA ECONÔMICA FEDERAL',
            '001 - BANCO DO BRASIL',
            '033 - BANCO SANTANDER',
            '237 - BRADESCO',
            '341 - ITAÚ',
            '077 - BANCO INTER',
            '260 - NUBANK'
        ]

# Importar configurações
try:
    from src.config.utils import *
    from src.configuracoes_sistema import GerenciadorConfiguracoes
    logger.info("Configurações importadas com sucesso")
except Exception as e:
    logger.error(f"Erro ao importar configurações: {str(e)}")
    raise

# Importar configurações do sistema
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH
    )
    logger.info("Configurações do sistema importadas com sucesso")
except Exception as e:
    logger.error(f"Erro ao importar configurações do sistema: {str(e)}")
    raise

from src.config.window_config import configurar_janela
    

# Modificação para usar o método de utils.py
from src.config.utils import buscar_dados_bancarios_fornecedor


class VisualizadorLancamentos:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal  # referência ao sistema principal
        self.janela = tk.Toplevel(sistema_principal.root)  # usar .root para o Toplevel
        configurar_janela(self.janela, "Visualização de Lançamentos Pendentes", 1000, 400)

        # Registrar esta janela para os diálogos
        try:
            from src.config.dialogs import set_main_window
            set_main_window(self.janela)
        except ImportError:
            pass

        self.alteracoes = False
        self.dados_para_incluir = []
        
        # Adicionar um flag para controlar comportamento de foco
        self.focus_locked = False
        
        # Configurar comportamento quando a janela é fechada
        self.janela.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # Vincular evento de foco para manter a janela na frente quando necessário
        self.janela.bind("<FocusIn>", self.on_focus_in)
        self.janela.bind("<Map>", self.on_map)  # Quando janela é mapeada (mostrada)

        # Frame principal
        self.frame_principal = ttk.Frame(self.janela)
        self.frame_principal.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar Treeview para visualização
        self.criar_treeview()
        
        # Frame para resumo
        self.frame_resumo = ttk.LabelFrame(self.frame_principal, text="Resumo")
        self.frame_resumo.pack(fill='x', pady=5)
        
        self.lbl_total_lancamentos = ttk.Label(self.frame_resumo, text="Total de Lançamentos: 0")
        self.lbl_total_lancamentos.pack(side='left', padx=5)
        
        self.lbl_valor_total = ttk.Label(self.frame_resumo, text="Valor Total: R$ 0,00")
        self.lbl_valor_total.pack(side='left', padx=5)
        
        # Frame para botões
        self.frame_botoes = ttk.Frame(self.frame_principal)
        self.frame_botoes.pack(fill='x', pady=5)
        
        ttk.Button(self.frame_botoes, text="Editar", command=self.editar_lancamento).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="Remover", command=self.remover_lancamento).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="Salvar na Planilha", command=self.salvar_na_planilha).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="Fechar", command=self.janela.destroy).pack(side='right', padx=5)

    
        # Variável para rastrear se houve alterações
        self.alteracoes = False
    
    def on_close(self):
        """Manipula o fechamento da janela"""
        if hasattr(self.sistema, 'on_visualizador_close'):
            self.sistema.on_visualizador_close()
        else:
            self.janela.destroy()
    
    def on_focus_in(self, event=None):
        """Quando a janela recebe foco"""
        if self.focus_locked:
            self.janela.lift()
    
    def on_map(self, event=None):
        """Quando a janela é mapeada (tornada visível)"""
        # Trazer para frente e forçar foco
        self.janela.lift()
        self.janela.focus_force()
    
    def travar_foco(self, travar=True):
        """Define se a janela deve ser mantida à frente"""
        self.focus_locked = travar
        if travar:
            # Trazer para frente imediatamente
            self.janela.lift()
            self.janela.focus_force()        

    def criar_treeview(self):
        # Alterar a lista de colunas para incluir forma de pagamento
        colunas = ('Data', 'Tipo', 'CNPJ/CPF', 'Nome', 'Referência', 'NF', 'Vr. Unit.', 
                   'Dias', 'Valor', 'Vencimento', 'Categoria', 'Forma Pagamento', 'Dados Bancários', 'Observação')
        
        self.tree = ttk.Treeview(self.frame_principal, columns=colunas, show='headings')
        
        # Configurar cabeçalhos
        for col in colunas:
            self.tree.heading(col, text=col)
            # Ajustar largura baseado no conteúdo
            if col in ['CNPJ/CPF', 'Nome', 'Referência', 'Dados Bancários', 'Observação']:
                width = 150
            elif col in ['Data', 'Vencimento']:
                width = 100
            elif col in ['Vr. Unit.', 'Valor', 'NF']:
                width = 100
            elif col == 'Forma Pagamento':
                width = 80
            else:
                width = 80
            self.tree.column(col, width=width)

        # Adicionar scrollbars
        scrolly = ttk.Scrollbar(self.frame_principal, orient='vertical', command=self.tree.yview)
        scrollx = ttk.Scrollbar(self.frame_principal, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionar elementos
        self.tree.pack(fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
    
    def atualizar_dados(self, dados):
        """Atualiza os dados na visualização"""
        self.dados_para_incluir = dados.copy() if dados else []
        
        # Limpar dados existentes
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        if not dados:
            # Atualizar resumo para zero quando não houver dados
            self.lbl_total_lancamentos.config(text="Total de Lançamentos: 0")
            self.lbl_valor_total.config(text="Valor Total: R$ 0,00")
            return
            
        # Inserir novos dados
        valor_total = 0
        for lancamento in self.dados_para_incluir:
            valores = (
                lancamento['data'],
                lancamento['tp_desp'],
                lancamento['cnpj_cpf'],
                lancamento['nome'],
                lancamento['referencia'],
                lancamento.get('nf', ''),
                lancamento['vr_unit'],
                lancamento['dias'],
                lancamento['valor'],
                lancamento['dt_vencto'],
                lancamento['categoria'],
                lancamento.get('forma_pagamento', ''),  
                lancamento['dados_bancarios'],
                lancamento['observacao']
            )
            self.tree.insert('', 'end', values=valores)
            valor_total += float(lancamento['valor'])
        
        # Atualizar resumo
        self.lbl_total_lancamentos.config(text=f"Total de Lançamentos: {len(dados)}")
        self.lbl_valor_total.config(text=f"Valor Total: R$ {valor_total:,.2f}")

    def editar_lancamento(self):
        """Abre a janela de edição para o lançamento selecionado"""
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um lançamento para editar")
            return

        # Obter índice do item selecionado
        todos_items = self.tree.get_children()
        indice = todos_items.index(item_selecionado[0])
        
        # Obter valores atuais
        valores = self.tree.item(item_selecionado)['values']
        dados = {
            'data': valores[0],
            'tp_desp': valores[1],
            'cnpj_cpf': valores[2],
            'nome': valores[3],
            'referencia': valores[4],
            'nf': valores[5],
            'vr_unit': valores[6],
            'dias': valores[7],
            'valor': valores[8],
            'dt_vencto': valores[9],
            'categoria': valores[10],
            'forma_pagamento': valores[11],
        'dados_bancarios': valores[12],
        'observacao': valores[13] if len(valores) > 13 else ''
    }
        
        # Criar editor
        editor = EditorLancamento(self.janela, dados, indice, self.atualizar_lancamento)

    def atualizar_lancamento(self, indice, novos_dados):
        """Atualiza os dados de um lançamento específico"""
        try:
            # Formatar CNPJ/CPF baseado no número de dígitos
            cnpj_cpf = str(novos_dados['cnpj_cpf']).replace('.', '').replace('-', '').replace('/', '')
            novos_dados['cnpj_cpf'] = formatar_cnpj_cpf(cnpj_cpf)

            # Converter observação para maiúsculas
            novos_dados['observacao'] = novos_dados['observacao'].upper()

            # Atualizar na treeview
            item = self.tree.get_children()[indice]
            valores = (
                novos_dados['data'],
                novos_dados['tp_desp'],
                novos_dados['cnpj_cpf'],
                novos_dados['nome'],
                novos_dados['referencia'],
                novos_dados['nf'],
                novos_dados['vr_unit'],
                novos_dados['dias'],
                novos_dados['valor'],
                novos_dados['dt_vencto'],
                novos_dados['categoria'],
                novos_dados['dados_bancarios'],
                novos_dados['observacao']
            )
            
            # Atualizar dados na lista
            self.dados_para_incluir[indice] = novos_dados.copy()
            
            # Atualizar treeview
            self.tree.item(item, values=valores)
            
            # Atualizar resumo
            self.atualizar_resumo()
            
            return True
        except Exception as e:
            print(f"Erro ao atualizar lançamento: {str(e)}")
            return False


    def remover_lancamento(self):
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um lançamento para remover")
            return
        
        # Travar o foco durante a operação
        self.travar_foco(True)
            
        if custom_messagebox("yesno", "Confirmação", "Deseja realmente remover este lançamento?"):
            # Obter índice do item selecionado
            todos_items = self.tree.get_children()
            indice = todos_items.index(item_selecionado[0])
            
            # Remover da lista de dados
            if 0 <= indice < len(self.dados_para_incluir):
                self.dados_para_incluir.pop(indice)
            
            # Remover da visualização
            self.tree.delete(item_selecionado)
            
            # Atualizar contadores e totais
            self.atualizar_resumo()

        # Destravar o foco após a operação
        self.travar_foco(False)
        
        # Garantir que a janela volte a ter foco após o diálogo
        self.janela.after(100, lambda: self.janela.focus_force())

    def salvar_na_planilha(self):
        """Salva os dados diretamente na planilha"""
        try:
            if not self.dados_para_incluir:
                custom_messagebox("warning",  "Aviso", "Não há dados para salvar!")
                return

            # Travar o foco durante a operação
            self.travar_foco(True)

            # Atualizar dados do sistema principal
            self.sistema.dados_para_incluir = self.dados_para_incluir.copy()
            
            # Chamar o método enviar_dados do sistema principal
            if self.sistema:
                self.sistema.enviar_dados()
                self.janela.destroy()  # Fecha o visualizador após salvar
            else:
                custom_messagebox("error", "Erro", "Referência ao sistema principal não encontrada")

            # Destravar o foco após a operação
            self.travar_foco(False)

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar dados: {str(e)}")
            print(f"Erro detalhado ao salvar: {str(e)}")  # Log para debug

        

    def atualizar_resumo(self):
        items = self.tree.get_children()
        total_lancamentos = len(items)
        valor_total = sum(float(self.tree.item(item)['values'][8]) for item in items)
        
        self.lbl_total_lancamentos.config(text=f"Total de Lançamentos: {total_lancamentos}")
        self.lbl_valor_total.config(text=f"Valor Total: R$ {valor_total:,.2f}")


    def get_dados_atualizados(self):
        """Retorna todos os dados atualizados"""
        return self.dados_para_incluir.copy()


class EditorLancamento:
    def __init__(self, parent, dados, indice, callback_atualizacao):
        self.janela = tk.Toplevel(parent)
        self.janela.title("Editar Lançamento")
        self.janela.geometry("600x500")
        
        self.dados = dados
        self.indice = indice
        self.callback_atualizacao = callback_atualizacao
        
        # Frame principal
        frame = ttk.Frame(self.janela, padding="10")
        frame.pack(fill='both', expand=True)
        
        # Frame para dados do fornecedor (não editáveis)
        frame_fornecedor = ttk.LabelFrame(frame, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', pady=5)
        
        # CNPJ/CPF
        ttk.Label(frame_fornecedor, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2)
        self.cnpj_cpf = ttk.Entry(frame_fornecedor, state='readonly')
        self.cnpj_cpf.grid(row=0, column=1, padx=5, pady=2)
        
        # Nome
        ttk.Label(frame_fornecedor, text="Nome:").grid(row=1, column=0, padx=5, pady=2)
        self.nome = ttk.Entry(frame_fornecedor, state='readonly')
        self.nome.grid(row=1, column=1, padx=5, pady=2)
        
        # Frame para dados da despesa
        frame_despesa = ttk.LabelFrame(frame, text="Dados da Despesa")
        frame_despesa.pack(fill='x', pady=5)
        
        # Data de Referência
        ttk.Label(frame_despesa, text="Data do Relatório:").grid(row=0, column=0, padx=5, pady=2)
        self.data_rel = DateEntry(frame_despesa, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_rel.grid(row=0, column=1, padx=5, pady=2)
        
        # Tipo de Despesa
        ttk.Label(frame_despesa, text="Tipo Despesa:").grid(row=1, column=0, padx=5, pady=2)
        self.tp_desp = ttk.Entry(frame_despesa)
        self.tp_desp.grid(row=1, column=1, padx=5, pady=2)
        
        # Referência
        ttk.Label(frame_despesa, text="Referência:").grid(row=2, column=0, padx=5, pady=2)
        self.referencia = ttk.Entry(frame_despesa)
        self.referencia.grid(row=2, column=1, padx=5, pady=2)
        
        # NF
        ttk.Label(frame_despesa, text="NF:").grid(row=3, column=0, padx=5, pady=2)
        self.nf = ttk.Entry(frame_despesa)
        self.nf.grid(row=3, column=1, padx=5, pady=2)
        
        # Valor Unitário
        ttk.Label(frame_despesa, text="Valor Unitário:").grid(row=4, column=0, padx=5, pady=2)
        self.vr_unit = ttk.Entry(frame_despesa)
        self.vr_unit.grid(row=4, column=1, padx=5, pady=2)
        
        # Dias
        ttk.Label(frame_despesa, text="Dias:").grid(row=5, column=0, padx=5, pady=2)
        self.dias = ttk.Entry(frame_despesa)
        self.dias.grid(row=5, column=1, padx=5, pady=2)
        
        # Valor Total
        ttk.Label(frame_despesa, text="Valor Total:").grid(row=6, column=0, padx=5, pady=2)
        self.valor = ttk.Entry(frame_despesa, state='readonly')
        self.valor.grid(row=6, column=1, padx=5, pady=2)
        
        # Data de Vencimento
        ttk.Label(frame_despesa, text="Data Vencimento:").grid(row=7, column=0, padx=5, pady=2)
        self.dt_vencto = DateEntry(frame_despesa, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.dt_vencto.grid(row=7, column=1, padx=5, pady=2)
        
        # Configurar o calendário para permitir navegação
        def configurar_calendario(event=None):
            if hasattr(self.dt_vencto, '_top_cal'):
                cal = self.dt_vencto._top_cal
                if cal:
                    def permitir_navegacao(event):
                        return "break"
                    
                    # Permitir cliques nas setas e mês/ano
                    for widget in cal.winfo_children():
                        if isinstance(widget, tk.Button):
                            widget.unbind('<Button-1>')
                            widget.bind('<Button-1>', permitir_navegacao)
                        
        self.dt_vencto.bind('<<DateEntryPopup>>', configurar_calendario)
        
        # Forma de Pagamento
        ttk.Label(frame_despesa, text="Forma de Pagamento:").grid(row=8, column=0, padx=5, pady=2)
        self.forma_pagamento = ttk.Combobox(frame_despesa, values=['PIX', 'TED'], state='readonly')
        self.forma_pagamento.grid(row=8, column=1, padx=5, pady=2)
        
        # Observação
        ttk.Label(frame_despesa, text="Observação:").grid(row=9, column=0, padx=5, pady=2)
        self.observacao = ttk.Entry(frame_despesa)
        self.observacao.grid(row=9, column=1, padx=5, pady=2)
        
        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Salvar", command=self.salvar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=self.janela.destroy).pack(side='left', padx=5)
        
        # Preencher dados existentes
        self.preencher_dados()
        
        # Vincular eventos
        self.vr_unit.bind('<KeyRelease>', self.calcular_valor_total)
        self.dias.bind('<KeyRelease>', self.calcular_valor_total)
        
    def preencher_dados(self):
        """Preenche os campos com os dados atuais"""
        self.cnpj_cpf.config(state='normal')
        self.cnpj_cpf.insert(0, self.dados['cnpj_cpf'])
        self.cnpj_cpf.config(state='readonly')
        
        self.nome.config(state='normal')
        self.nome.insert(0, self.dados['nome'])
        self.nome.config(state='readonly')
        
        self.data_rel.set_date(datetime.strptime(self.dados['data'], '%d/%m/%Y'))
        self.tp_desp.insert(0, self.dados['tp_desp'])
        self.referencia.insert(0, self.dados['referencia'])
        self.nf.insert(0, self.dados.get('nf', ''))
        self.vr_unit.insert(0, self.dados['vr_unit'])
        self.dias.insert(0, str(self.dados['dias']))
        
        self.valor.config(state='normal')
        self.valor.insert(0, self.dados['valor'])
        self.valor.config(state='readonly')
        
        self.dt_vencto.set_date(datetime.strptime(self.dados['dt_vencto'], '%d/%m/%Y'))
        self.observacao.insert(0, self.dados.get('observacao', ''))

        self.forma_pagamento.set(self.dados.get('forma_pagamento', ''))


    def atualizar_dados_bancarios(self, event=None):
        """Atualiza os dados bancários baseado no tipo de despesa e forma de pagamento"""
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get().strip()

        if not cnpj_cpf:  # Se não houver fornecedor selecionado
            return
        
        forma_pagamento = self.forma_pagamento_var.get()
        
        try:
            # Usar a função centralizada em utils
            from src.config.utils import buscar_dados_bancarios_fornecedor
            dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj_cpf, forma_pagamento)
        except ImportError:
            # Implementação alternativa se a função não estiver disponível
            fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
            if not fornecedor_completo:
                return
            
            if forma_pagamento == "PIX" and fornecedor_completo['chave_pix']:
                dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
            else:
                # Estrutura para TED
                dados_ted = []
                if fornecedor_completo['banco']: dados_ted.append(str(fornecedor_completo['banco']))
                if fornecedor_completo['op']: dados_ted.append(str(fornecedor_completo['op']))
                if fornecedor_completo['agencia']: dados_ted.append(str(fornecedor_completo['agencia']))
                if fornecedor_completo['conta']: dados_ted.append(str(fornecedor_completo['conta']))
                # SEMPRE adicionar o CNPJ/CPF para TED
                dados_ted.append(str(fornecedor_completo['cnpj_cpf']))
                
                dados_bancarios = ' - '.join(filter(None, dados_ted))

            if dados_bancarios.strip() in ['', ' - ']:
                dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'

        # Atualizar o campo
        self.campos_fornecedor['dados_bancarios'].config(state='normal')
        self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
        self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
        self.campos_fornecedor['dados_bancarios'].config(state='readonly')
        
        
    def calcular_valor_total(self, event=None):
        """Calcula o valor total baseado no valor unitário e dias"""
        try:
            vr_unit = float(self.vr_unit.get().replace(',', '.'))
            dias = float(self.dias.get() or 1)
            valor_total = vr_unit * dias
            
            self.valor.config(state='normal')
            self.valor.delete(0, tk.END)
            self.valor.insert(0, f"{valor_total:.2f}")
            self.valor.config(state='readonly')
            
        except (ValueError, AttributeError):
            self.valor.config(state='normal')
            self.valor.delete(0, tk.END)
            self.valor.config(state='readonly')
            
    def salvar(self):
        """Salva as alterações e fecha a janela"""
        try:
            # Validar campos obrigatórios
            if not all([self.tp_desp.get(), self.referencia.get(), self.vr_unit.get()]):
                custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
                return
            
            # Validar datas
            for data_entry in [self.data_rel, self.dt_vencto]:
                data_str = data_entry.get()
                try:
                    datetime.strptime(data_str, '%d/%m/%Y')
                except ValueError:
                    custom_messagebox("error", "Erro", "Data inválida!")
                    return
            
            # Atualizar dados
            dados_atualizados = {
                'data': self.data_rel.get(),
                'tp_desp': self.tp_desp.get(),
                'cnpj_cpf': self.dados['cnpj_cpf'],
                'nome': self.dados['nome'],
                'forma_pagamento': self.forma_pagamento.get(),
                'referencia': self.referencia.get(),
                'nf': self.nf.get(),
                'vr_unit': self.vr_unit.get(),
                'dias': float(self.dias.get().replace(',', '.') if self.dias.get() else 1),
                'valor': self.valor.get(),
                'dt_vencto': self.dt_vencto.get(),
                'categoria': self.dados['categoria'],
                'dados_bancarios': self.dados['dados_bancarios'],
                'observacao': self.observacao.get()
            }
            
            # Chamar callback de atualização e verificar sucesso
            if self.callback_atualizacao(self.indice, dados_atualizados):
                custom_messagebox("info", "Sucesso", "Alterações salvas com sucesso!")
                self.janela.destroy()
            else:
                custom_messagebox("error", "Erro", "Não foi possível salvar as alterações!")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar alterações: {str(e)}")

        

class SistemaEntradaDados:

    def atualizar_combos_configuracoes(self):
        """Atualiza os valores das Comboboxes baseados nas configurações"""
        if 'categoria' in self.campos_fornecedor and isinstance(self.campos_fornecedor['categoria'], ttk.Combobox):
            categorias = get_categorias_fornecedor()
            self.campos_fornecedor['categoria']['values'] = categorias
            # Define o primeiro valor como padrão se houver categorias
            if categorias:
                self.campos_fornecedor['categoria'].set(categorias[0])
            
    def __init__(self, parent=None):
        print("Inicializando SistemaEntradaDados...")
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Sistema de Entrada de Dados")
        self.dados_para_incluir = []
        self.data_rel = None
        self.cliente_atual = None
        self.visualizador = None
        self._gestor_parcelas = None  # Inicializa como None
        self.gerenciador_lancamentos = None

        # Inicializar a variável de forma de pagamento
        self.forma_pagamento_var = tk.StringVar(value="")
            
        # Frame temporário para criar os entries
        temp_frame = ttk.Frame(self.root)

        # Criação dos campos_fornecedor e campos_despesa
        self.campos_fornecedor = {
            'cnpj_cpf': tk.Entry(temp_frame),
            'nome': tk.Entry(temp_frame),
            'categoria': tk.Entry(temp_frame),
            'dados_bancarios': tk.Entry(temp_frame)
        }

        self.campos_despesa = {
            'tp_desp': tk.Entry(temp_frame),
            'referencia': tk.Entry(temp_frame),
            'nf': tk.Entry(temp_frame),
            'vr_unit': tk.Entry(temp_frame),
            'dias': tk.Entry(temp_frame),
            'valor': tk.Entry(temp_frame),
            'dt_vencto': tk.Entry(temp_frame),
            'observacao': tk.Entry(temp_frame)
        }

        self.gestao_taxas = GestaoTaxasFixas(self)

        self.atualizar_combos_configuracoes()
            
        # Configurar interface
        self.setup_gui()
        self.configurar_todos_calendarios()

        # Adicionar estas linhas para configurar cada aba explicitamente
        print("Configurando aba de seleção...")
        self.setup_aba_selecao()
        print("Configurando aba de fornecedor...")
        self.setup_aba_fornecedor()
        print("Configurando aba de dados...")
        self.setup_aba_dados()
        
        print("Finalizada inicialização do sistema")

    def setup_gui(self):
        print("Iniciando setup_gui...")
        
        # Remover notebook existente se houver
        for widget in self.root.winfo_children():
            if isinstance(widget, ttk.Notebook):
                print("Notebook existente encontrado e será removido")
                widget.destroy()
        
        # Frame principal com abas
        self.notebook = ttk.Notebook(self.root)
        print("Novo Notebook criado")
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar abas
        self.aba_selecao = ttk.Frame(self.notebook)
        self.aba_fornecedor = ttk.Frame(self.notebook)
        self.aba_dados = ttk.Frame(self.notebook)
        
        print("Adicionando abas ao Notebook")
        self.notebook.add(self.aba_selecao, text='Seleção de Cliente')
        self.notebook.add(self.aba_fornecedor, text='Fornecedor')
        self.notebook.add(self.aba_dados, text='Entrada de Dados')

        print("Setup_gui concluído")

    @property
    def gestor_parcelas(self):
        """Getter para gestor_parcelas - cria apenas quando necessário"""
        if self._gestor_parcelas is None:
            print("Criando nova instância do GestorParcelas")  # Debug
            self._gestor_parcelas = GestorParcelas(self)
        return self._gestor_parcelas

    @gestor_parcelas.setter
    def gestor_parcelas(self, valor):
        """Setter para gestor_parcelas"""
        self._gestor_parcelas = valor        

    def voltar_menu(self):
        """Retorna ao menu principal verificando dados não salvos"""
        if self.dados_para_incluir and custom_messagebox("yesno", 
            "Confirmação", 
            "Existem dados não salvos. Deseja salvá-los antes de sair?"):
            self.enviar_dados()
        
        self.root.destroy()  # Fecha a janela atual
        
        # Se tiver referência ao menu principal, mostra ele
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()

    def sair_sistema(self):
        """Fecha o sistema verificando dados não salvos"""
        if self.dados_para_incluir and custom_messagebox("yesno", 
            "Confirmação", 
            "Existem dados não salvos. Deseja salvá-los antes de sair?"):
            self.enviar_dados()
        self.root.destroy()
        sys.exit()    
    
    def configurar_todos_calendarios(self):
        """Configura a navegação para todos os calendários do sistema"""
        # Lista de todos os campos de data que usam DateEntry
        date_entries = []
        
        # Adicionar campos da interface principal
        if hasattr(self, 'data_rel_entry'):
            date_entries.append(self.data_rel_entry)
        
        # Adicionar campos da aba de dados
        if hasattr(self, 'campos_despesa') and 'dt_vencto' in self.campos_despesa:
            date_entries.append(self.campos_despesa['dt_vencto'])
        
        # Configurar cada DateEntry encontrado
        for date_entry in date_entries:
            if isinstance(date_entry, DateEntry):
                configurar_navegacao_calendario(date_entry)
        
        print("Calendários configurados para permitir navegação livre.")


    def setup_aba_selecao(self):
        """Configura a aba de seleção de cliente"""
        # Frame principal para organização
        frame_principal = ttk.Frame(self.aba_selecao)
        frame_principal.pack(expand=True, fill='both', padx=10, pady=5)

        # Frame para seleção de cliente
        frame_selecao = ttk.LabelFrame(frame_principal, text="Seleção do Cliente")
        frame_selecao.pack(fill='x', pady=10)

        # Container para label e combobox
        frame_cliente = ttk.Frame(frame_selecao)
        frame_cliente.pack(fill='x', padx=10, pady=10)

        # Label alinhado à esquerda
        ttk.Label(frame_cliente, text="Selecione o Cliente:", font=('Arial', 11)).pack(side='left', pady=5)
        
        # Combobox com largura aumentada
        self.cliente_combobox = ttk.Combobox(frame_cliente, width=60, font=('Arial', 11))  # Aumentado a fonte
        self.cliente_combobox.pack(side='left', padx=5, fill='x', expand=True)
        
        # Botão para mostrar todos os clientes (incluindo finalizados)
        ttk.Button(frame_cliente, text="Ver Todos", 
                command=self.mostrar_todos_clientes).pack(side='right', padx=5)
        
        # Frame para botões de gerenciamento de clientes
        frame_gerenciar = ttk.Frame(frame_principal)
        frame_gerenciar.pack(pady=15)
        
        # Estilo para botões maiores
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 12, 'bold'), padding=(15, 10))
        
        # Primeira linha de botões
        frame_botoes_linha1 = ttk.Frame(frame_gerenciar)
        frame_botoes_linha1.pack(fill='x', pady=5)
        
        ttk.Button(frame_botoes_linha1, 
                text="Novo Cliente", 
                command=self.criar_novo_cliente,
                style='Big.TButton').pack(side='left', padx=10)
                
        ttk.Button(frame_botoes_linha1,
                text="Editar Cliente",
                command=self.editar_cliente,
                style='Big.TButton').pack(side='left', padx=10)

        ttk.Button(frame_botoes_linha1, 
                text="Gerir Contratos",
                command=self.abrir_gestao_contratos,
                style='Big.TButton').pack(side='left', padx=10)

        # Segunda linha de botões (apenas para o Continuar)
        frame_botoes_linha2 = ttk.Frame(frame_gerenciar)
        frame_botoes_linha2.pack(fill='x', pady=10)
        
        # Botão continuar (inicialmente desabilitado)
        self.btn_continuar = ttk.Button(frame_botoes_linha2,
                                    text="Continuar →",
                                    command=self.continuar_para_fornecedor,
                                    state='disabled',
                                    style='Big.TButton')
        self.btn_continuar.pack(side='right', padx=10)
        
        # Carregar clientes existentes
        self.atualizar_lista_clientes()
        
        # Binding para seleção de cliente
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)

        # Frame de botões
        frame_botoes_selecao = ttk.Frame(frame_principal)
        frame_botoes_selecao.pack(fill='x', side='bottom', pady=10)

        ttk.Button(frame_botoes_selecao, 
                text="Voltar ao Menu", 
                command=self.voltar_menu,
                style='Big.TButton').pack(side='left', padx=10)
        ttk.Button(frame_botoes_selecao, 
                text="Sair", 
                command=self.sair_sistema,
                style='Big.TButton').pack(side='left', padx=10)

    def abrir_gestao_contratos(self):
        """Abre a gestão de contratos para o cliente atual"""
        if not self.cliente_atual:
            custom_messagebox("warning", "Aviso", "Selecione um cliente primeiro!")
            return
        
        # Ocultar temporariamente a janela principal
        self.root.withdraw()
        
        # Criar e configurar a janela de gestão de contratos diretamente aqui
        # em vez de delegar para outra classe/método
        janela_gestao = tk.Toplevel(self.root)
        janela_gestao.title(f"Gestão de Contratos - {self.cliente_atual}")
        janela_gestao.geometry("800x750")
        
        # Centralizar a janela (sem depender de um método da classe GestaoContratos)
        janela_gestao.update_idletasks()
        width = janela_gestao.winfo_width()
        height = janela_gestao.winfo_height()
        x = (janela_gestao.winfo_screenwidth() // 2) - (width // 2)
        y = (janela_gestao.winfo_screenheight() // 2) - (height // 2)
        janela_gestao.geometry(f'{width}x{height}+{x}+{y}')
        
        # Colocar a janela em primeiro plano
        janela_gestao.attributes('-topmost', True)
        janela_gestao.after(100, lambda: janela_gestao.attributes('-topmost', False))
        
        # Definir comportamento quando a janela for fechada
        def on_close():
            janela_gestao.destroy()
            self.root.deiconify()  # Mostrar a janela principal novamente
            self.root.lift()
            self.root.focus_force()
        
        # Configurar protocolo de fechamento
        janela_gestao.protocol("WM_DELETE_WINDOW", on_close)
        
        # Criar o restante da interface usando o gestor de contratos
        gestor = GestaoContratos(janela_gestao)  # Passamos a janela_gestao como parent
        gestor.cliente_atual = self.cliente_atual
        gestor.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
        
        # Criar e preencher a interface dentro da janela_gestao
        gestor.criar_interface_contratos(janela_gestao, on_close)

    def abrir_controle_pagamentos(self):
        """Abre o módulo de controle de pagamentos"""
        try:
            # Importar módulo
            from controle_pagamentos import ControlePagamentos
            
            # Instanciar e abrir janela de controle
            controle = ControlePagamentos(self.root)
            controle.abrir_janela_controle()
        except ImportError as e:
            custom_messagebox("error", "Erro", f"Não foi possível importar o módulo de Controle de Pagamentos: {str(e)}")
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir controle de pagamentos: {str(e)}")

    def selecionar_cliente(self, event):
        """Atualiza seleção de cliente e habilita botão de continuar"""
        self.cliente_atual = self.cliente_combobox.get()
        # Atualiza label na aba de dados
        self.cliente_label.config(text=f"Cliente: {self.cliente_atual}")
        
        # Atualiza também o label na aba de fornecedor
        if hasattr(self, 'lbl_cliente_fornecedor'):
            self.lbl_cliente_fornecedor.config(text=f"Cliente: {self.cliente_atual}")
        
        # Habilita o botão continuar
        self.btn_continuar.config(state='normal')
        # Não muda de aba automaticamente


    def continuar_para_fornecedor(self):
        """Avança para a aba de fornecedor após confirmar seleção"""
        if self.cliente_atual:
            self.notebook.select(1)  # Vai para aba de fornecedor
        else:
            custom_messagebox("warning",  "Aviso", "Selecione um cliente primeiro!")



    def criar_arquivo_clientes(self):
        """Cria arquivo base de clientes se não existir"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Clientes'
            
            # Adicionar cabeçalhos - somente campos básicos agora
            headers = ['Nome', 'Endereco', 'Data_Inicial', 'Observacoes']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            caminho_base = ARQUIVO_CLIENTES
            workbook.save(caminho_base)
            custom_messagebox("info", "Informação", "Arquivo de clientes criado com sucesso!")
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao criar arquivo de clientes: {str(e)}")


 
    def criar_novo_cliente(self):
        """Abre janela para cadastro de novo cliente"""
        janela_cliente = tk.Toplevel(self.root)
        janela_cliente.title("Novo Cliente")
        janela_cliente.geometry("700x400")  # Reduzido pois terá menos campos

        # Campos do formulário
        tk.Label(janela_cliente, text="Nome do Cliente:*").pack(pady=5)
        nome_entry = tk.Entry(janela_cliente, width=80)
        nome_entry.pack(pady=5)

        tk.Label(janela_cliente, text="Endereço:*").pack(pady=5)
        endereco_entry = tk.Entry(janela_cliente, width=80)
        endereco_entry.pack(pady=5)

        tk.Label(janela_cliente, text="Data Inicial:* (Dia 5 ou 20)").pack(pady=5)
        data_entry = DateEntry(
            janela_cliente,
            width=20,
            date_pattern='yyyy-mm-dd',
            locale='pt_BR'
        )
        data_entry.pack(pady=5)

        def validar_data(*args):
            """Valida se a data selecionada é dia 5 ou 20"""
            data = data_entry.get_date()
            if data.day not in [5, 20]:
                custom_messagebox("info",  
                    "Data Inválida",
                    "A data inicial deve ser dia 5 ou 20 do mês.\n"
                    "Por favor, selecione uma data válida."
                )
                # Encontrar o próximo dia 5 ou 20
                if data.day < 5:
                    data = data.replace(day=5)
                elif data.day < 20:
                    data = data.replace(day=20)
                else:
                    if data.month == 12:
                        data = data.replace(year=data.year + 1, month=1, day=5)
                    else:
                        data = data.replace(month=data.month + 1, day=5)
                data_entry.set_date(data)

        # Adicionar validação quando a data é alterada
        data_entry.bind("<<DateEntrySelected>>", validar_data)

        tk.Label(janela_cliente, text="Observações:").pack(pady=5)
        obs_entry = tk.Entry(janela_cliente, width=80)
        obs_entry.pack(pady=5)
        

        def salvar_cliente():
            nome = nome_entry.get().strip()
            endereco = endereco_entry.get().strip()
            data = data_entry.get()
            observacoes = obs_entry.get().strip()
            
            if not nome or not endereco:
                custom_messagebox("error", "Erro", "Nome e Endereço são obrigatórios!")
                return

            # Verificar se a data é válida
            try:
                data = datetime.strptime(data, '%Y-%m-%d').date()
                if data.day not in [5, 20]:
                    custom_messagebox("error", "Erro", "A data inicial deve ser dia 5 ou 20 do mês!")
                    return
            except ValueError:
                custom_messagebox("error", "Erro", "Data inválida!")
                return

            try:
                # Salvar no arquivo de clientes
                caminho_base = ARQUIVO_CLIENTES
                workbook = load_workbook(caminho_base)
                sheet = workbook['Clientes']

                # Verificar se cliente já existe
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[0].upper() == nome.upper():
                        custom_messagebox("error", "Erro", "Cliente já cadastrado!")
                        return

                # Adicionar novo cliente
                proxima_linha = sheet.max_row + 1
                sheet.cell(row=proxima_linha, column=1, value=nome.upper())
                sheet.cell(row=proxima_linha, column=2, value=endereco.upper())
                sheet.cell(row=proxima_linha, column=3, value=data)
                sheet.cell(row=proxima_linha, column=4, value=observacoes.upper())

                workbook.save(caminho_base)

                # Criar arquivo do cliente baseado no modelo
                if self.criar_arquivo_cliente(nome.upper(), endereco.upper()):
                    custom_messagebox("info", "Sucesso", "Cliente cadastrado com sucesso!")
                    self.atualizar_lista_clientes()
                    janela_cliente.destroy()

            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao cadastrar cliente: {str(e)}")

        tk.Button(janela_cliente, text="Salvar", command=salvar_cliente).pack(pady=10)
        tk.Button(janela_cliente, text="Cancelar", 
                 command=janela_cliente.destroy).pack(pady=5)

    def criar_arquivo_clientes(self):
        """Cria arquivo base de clientes se não existir"""
        try:
            print(f"Tentando criar arquivo de clientes em: {ARQUIVO_CLIENTES}")
            print(f"Diretório existe? {os.path.exists(os.path.dirname(ARQUIVO_CLIENTES))}")
            
            # Garantir que o diretório existe
            os.makedirs(os.path.dirname(ARQUIVO_CLIENTES), exist_ok=True)
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Clientes'
            
            # Adicionar cabeçalhos - agora incluindo Data_Final
            headers = ['Nome', 'Endereco', 'Data_Inicial', 'Observacoes', 'Data_Final']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            print(f"Tentando salvar arquivo em: {ARQUIVO_CLIENTES}")
            workbook.save(ARQUIVO_CLIENTES)
            custom_messagebox("info", "Informação", "Arquivo de clientes criado com sucesso!")
            
        except Exception as e:
            print(f"Erro detalhado ao criar arquivo de clientes: {str(e)}")
            print(f"Tipo do erro: {type(e)}")
            custom_messagebox("error", "Erro", f"Erro ao criar arquivo de clientes: {str(e)}")

    def mostrar_todos_clientes(self):
        """Mostra todos os clientes, incluindo os que têm data final"""
        try:
            # Carregar arquivo de clientes
            caminho_base = ARQUIVO_CLIENTES
            workbook = load_workbook(caminho_base)
            sheet = workbook['Clientes']
            
            # Pegar todos os clientes
            clientes = []
            clientes_finalizados = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome do cliente está na primeira coluna
                    # Verificar se tem data final
                    data_final = row[4] if len(row) > 4 else None
                    
                    if data_final:
                        clientes_finalizados.append(row[0])
                    else:
                        clientes.append(row[0])
            
            workbook.close()
            
            # Mostrar janela com todos os clientes
            janela_todos = tk.Toplevel(self.root)
            janela_todos.title("Todos os Clientes")
            janela_todos.geometry("600x500")
            
            frame = ttk.Frame(janela_todos, padding="10")
            frame.pack(fill='both', expand=True)
            
            # Lista de clientes ativos
            frame_ativos = ttk.LabelFrame(frame, text="Clientes Ativos")
            frame_ativos.pack(fill='both', expand=True, pady=5)
            
            lista_ativos = tk.Listbox(frame_ativos, width=50, height=13)
            lista_ativos.pack(fill='both', expand=True, padx=5, pady=5)
            
            for cliente in sorted(clientes):
                lista_ativos.insert(tk.END, cliente)
            
            # Lista de clientes finalizados
            frame_finalizados = ttk.LabelFrame(frame, text="Clientes Finalizados")
            frame_finalizados.pack(fill='both', expand=True, pady=5)
            
            lista_finalizados = tk.Listbox(frame_finalizados, width=50, height=7)
            lista_finalizados.pack(fill='both', expand=True, padx=5, pady=5)
            
            for cliente in sorted(clientes_finalizados):
                lista_finalizados.insert(tk.END, cliente)
            
            # Botão para selecionar cliente finalizado
            def selecionar_finalizado():
                selected = lista_finalizados.curselection()
                if not selected:
                    custom_messagebox("warning",  "Aviso", "Selecione um cliente finalizado")
                    return
                    
                cliente = lista_finalizados.get(selected[0])
                self.cliente_combobox.set(cliente)
                self.selecionar_cliente(None)
                janela_todos.destroy()
            
            ttk.Button(frame, text="Selecionar Cliente Finalizado", 
                    command=selecionar_finalizado).pack(side='left', pady=10)
            
            ttk.Button(frame, text="Fechar", 
                    command=janela_todos.destroy).pack(side='right', pady=10)
                    
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar todos os clientes: {str(e)}")

    def criar_arquivo_cliente(self, nome_cliente, endereco):
        """Cria um novo arquivo Excel para o cliente baseado no MODELO.xlsx"""
        try:
            print(f"\nTentando criar arquivo para cliente: {nome_cliente}")
            print(f"ARQUIVO_MODELO: {ARQUIVO_MODELO}")
            print(f"ARQUIVO_MODELO existe? {os.path.exists(ARQUIVO_MODELO)}")
            print(f"PASTA_CLIENTES: {PASTA_CLIENTES}")
            print(f"PASTA_CLIENTES existe? {os.path.exists(PASTA_CLIENTES)}")
            
            modelo_path = ARQUIVO_MODELO
            novo_arquivo = PASTA_CLIENTES / f"{nome_cliente}.xlsx"
            
            print(f"Novo arquivo será criado em: {novo_arquivo}")
            print(f"Diretório do novo arquivo existe? {os.path.exists(os.path.dirname(novo_arquivo))}")
                
            if os.path.exists(novo_arquivo):
                print(f"Arquivo {novo_arquivo} já existe!")
                raise Exception("Arquivo do cliente já existe!")
                    
            # Garantir que o diretório existe
            os.makedirs(os.path.dirname(novo_arquivo), exist_ok=True)
                
            print(f"Tentando copiar de {modelo_path} para {novo_arquivo}")
            
            # Copiar o arquivo modelo
            from shutil import copy2
            copy2(modelo_path, novo_arquivo)
            
            print("Arquivo copiado com sucesso")
                
            # Buscar data inicial do arquivo clientes.xlsx
            wb_clientes = load_workbook(ARQUIVO_CLIENTES)
            ws_clientes = wb_clientes['Clientes']
            
            data_inicial = None
            # Procurar o cliente e sua data inicial
            for row in range(2, ws_clientes.max_row + 1):
                if ws_clientes.cell(row=row, column=1).value == nome_cliente:
                    data_valor = ws_clientes.cell(row=row, column=3).value  # Coluna C
                    if not data_valor:
                        raise Exception("Data inicial não informada no cadastro do cliente")
                        
                    if isinstance(data_valor, datetime):
                        data_inicial = data_valor.date()
                    else:
                        try:
                            data_inicial = datetime.strptime(str(data_valor), '%Y-%m-%d').date()
                        except ValueError:
                            raise Exception("Data inicial deve estar no formato AAAA-MM-DD")
                    break
            
            if not data_inicial:
                raise Exception("Cliente não encontrado no cadastro")
                
            # Validar se é dia 5 ou 20
            if data_inicial.day not in [5, 20]:
                raise Exception("A data inicial deve ser dia 5 ou 20 do mês")
                
            # Abrir o novo arquivo para edição
            workbook = load_workbook(novo_arquivo)
            
            # Atualizar planilha RESUMO
            resumo_sheet = workbook["RESUMO"]
            
            # Informações básicas
            resumo_sheet["A3"] = nome_cliente
            resumo_sheet["A4"] = endereco
            
            # Descrições das células
            resumo_sheet["K3"] = "Data Inicial"
            
            # Adicionar data inicial
            resumo_sheet["L3"] = data_inicial
            resumo_sheet["L3"].number_format = 'dd/mm/yyyy'
            
            # Gerar as 96 datas quinzenais
            data_atual = data_inicial
            datas_geradas = []
            
            for i in range(96):  # 4 anos = 96 relatórios
                row = i + 9  # Começar na linha 9
                
                # Verificar se a data já foi usada
                if data_atual in datas_geradas:
                    raise Exception(f"Data duplicada detectada: {data_atual.strftime('%d/%m/%Y')}")
                datas_geradas.append(data_atual)
                
                # Adicionar data e número do relatório
                resumo_sheet.cell(row=row, column=1, value=data_atual)
                resumo_sheet.cell(row=row, column=1).number_format = 'dd/mm/yyyy'
                resumo_sheet.cell(row=row, column=2, value=i + 1)
                
                # Próxima data
                if data_atual.day == 5:
                    data_atual = data_atual.replace(day=20)
                else:  # day == 20
                    if data_atual.month == 12:
                        data_atual = data_atual.replace(year=data_atual.year + 1, month=1, day=5)
                    else:
                        data_atual = data_atual.replace(month=data_atual.month + 1, day=5)

            # Criar aba Contratos_ADM
            contratos_sheet = workbook.create_sheet("Contratos_ADM")
            
            # Definir os blocos na linha 1
            blocos = ["CONTRATOS", "", "", "", "", "",
                     "ADMINISTRADORES_CONTRATO", "", "", "", "", "", "",
                     "ADITIVOS", "", "", "",
                     "ADMINISTRADORES_ADITIVO", "", "", "", "", "", "",
                     "PARCELAS", "", "", "", "", "", "", ""]
            
            for col, valor in enumerate(blocos, 1):
                contratos_sheet.cell(row=1, column=col, value=valor)
            
            # Definir cabeçalhos na linha 2
            headers = [
                # CONTRATOS
                "Nº Contrato", "Data Início", "Data Fim", "Status", "Observações", "",
                # ADMINISTRADORES_CONTRATO
                "Nº Contrato", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total", "Nº Parcelas", 
                # ADITIVOS
                "Nº Contrato", "Nº Aditivo", "Data Início", "Data Fim",
                # ADMINISTRADORES_ADITIVO
                "Nº Contrato", "Nº Aditivo", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total",
                # PARCELAS
                "Referência", "Número", "CNPJ/CPF", "Nome", "Data Vencimento", "Valor", "Status", "Data Pagamento", "Eventos/Fases", "Percentual"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = contratos_sheet.cell(row=2, column=col, value=header)
                # Formatação do cabeçalho
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Ajustar largura das colunas
            for col in range(1, len(headers) + 1):
                contratos_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # Salvar alterações
            workbook.save(novo_arquivo)
            wb_clientes.close()
            
            return True
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao criar arquivo do cliente: {str(e)}")
            if 'wb_clientes' in locals():
                wb_clientes.close()
            return False



    def editar_cliente(self):
        """Edita o cliente selecionado"""
        cliente_selecionado = self.cliente_combobox.get()
        
        try:
            if not cliente_selecionado:
                from src.config.dialogs import custom_messagebox
                custom_messagebox("warning", "Aviso", "Selecione um cliente para editar")
                return
        except Exception as e:
            # Fallback em caso de erro
            import tkinter.messagebox as msgbox
            msgbox.showwarning("Aviso", "Selecione um cliente para editar")
            print(f"Erro ao mostrar diálogo personalizado: {str(e)}")
            return

        try:
            # Carregar dados do cliente
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            dados_cliente = None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == cliente_selecionado:
                    dados_cliente = {
                        'nome': row[0],
                        'endereco': row[1],
                        'data_inicial': row[2],
                        'observacoes': row[3],
                        'data_final': row[4] if len(row) > 4 else None  # Campo para Data Final
                    }
                    break
            
            wb.close()
            
            if not dados_cliente:
                custom_messagebox("error", "Erro", "Cliente não encontrado!")
                return
                
            # Criar janela de edição
            janela_edicao = tk.Toplevel(self.root)
            janela_edicao.title(f"Editar Cliente - {cliente_selecionado}")
            janela_edicao.geometry("760x250")  # Aumentado para acomodar o novo campo

            # Frame principal
            frame = ttk.Frame(janela_edicao, padding="10")
            frame.pack(fill='both', expand=True)

            # Frame para dados básicos
            frame_dados = ttk.LabelFrame(frame, text="Dados do Cliente")
            frame_dados.pack(fill='x', pady=5)

            # Nome
            ttk.Label(frame_dados, text="Nome do Cliente:*").grid(row=0, column=0, padx=5, pady=2)
            nome_entry = ttk.Entry(frame_dados, width=100)
            nome_entry.insert(0, dados_cliente['nome'])
            nome_entry.grid(row=0, column=1, padx=5, pady=2)

            # Endereço
            ttk.Label(frame_dados, text="Endereço:*").grid(row=1, column=0, padx=5, pady=2)
            endereco_entry = ttk.Entry(frame_dados, width=100)
            endereco_entry.insert(0, dados_cliente['endereco'])
            endereco_entry.grid(row=1, column=1, padx=5, pady=2)

            # Data Inicial
            ttk.Label(frame_dados, text="Data Inicial:*").grid(row=2, column=0, padx=5, pady=2)
            data_inicial_entry = DateEntry(
                frame_dados,
                width=20,
                date_pattern='yyyy-mm-dd',
                locale='pt_BR'
            )
            if dados_cliente['data_inicial']:
                data_inicial_entry.set_date(dados_cliente['data_inicial'])
            data_inicial_entry.grid(row=2, column=1, padx=5, pady=2)

            # Data Final (NOVO CAMPO)
            ttk.Label(frame_dados, text="Data Final:").grid(row=3, column=0, padx=5, pady=2)
            data_final_entry = DateEntry(
                frame_dados,
                width=20,
                date_pattern='yyyy-mm-dd',
                locale='pt_BR'
            )
            data_final_entry.delete(0, tk.END)  # Limpa a data padrão
            
            # Checkbox para ativar/desativar a data final
            tem_data_final = tk.BooleanVar(value=False)
            if dados_cliente['data_final']:
                tem_data_final.set(True)
                data_final_entry.set_date(dados_cliente['data_final'])
            
            def toggle_data_final():
                if tem_data_final.get():
                    data_final_entry.config(state='normal')
                    if not data_final_entry.get():
                        data_final_entry.set_date(datetime.now().date())
                else:
                    data_final_entry.delete(0, tk.END)
                    data_final_entry.config(state='disabled')
            
            check_data_final = ttk.Checkbutton(
                frame_dados, 
                text="Obra finalizada",
                variable=tem_data_final,
                command=toggle_data_final
            )
            check_data_final.grid(row=3, column=1, padx=5, pady=2)
            
            # Configurar estado inicial do campo data final
            if not dados_cliente['data_final']:
                data_final_entry.config(state='disabled')
            
            data_final_entry.grid(row=4, column=1, padx=5, pady=2)

            # Observações
            ttk.Label(frame_dados, text="Observações:").grid(row=5, column=0, padx=5, pady=2)
            obs_entry = ttk.Entry(frame_dados, width=100)
            obs_entry.insert(0, dados_cliente['observacoes'] or '')
            obs_entry.grid(row=5, column=1, padx=5, pady=2)

            def salvar_alteracoes():
                try:
                    nome = nome_entry.get().strip()
                    endereco = endereco_entry.get().strip()
                    
                    if not nome or not endereco:
                        custom_messagebox("error", "Erro", "Nome e Endereço são obrigatórios!")
                        return

                    wb = load_workbook(ARQUIVO_CLIENTES)
                    ws = wb['Clientes']

                    # Remover registros antigos do cliente
                    linhas_para_remover = []
                    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if row[0].value == cliente_selecionado:
                            linhas_para_remover.append(idx)

                    for linha in reversed(linhas_para_remover):
                        ws.delete_rows(linha)

                    # Adicionar novo registro
                    proxima_linha = ws.max_row + 1
                    ws.cell(row=proxima_linha, column=1, value=nome.upper())
                    ws.cell(row=proxima_linha, column=2, value=endereco.upper())
                    ws.cell(row=proxima_linha, column=3, value=data_inicial_entry.get_date())
                    ws.cell(row=proxima_linha, column=4, value=obs_entry.get().upper())
                    
                    # Salvar Data Final se checkbox estiver marcado
                    if tem_data_final.get():
                        ws.cell(row=proxima_linha, column=5, value=data_final_entry.get_date())
                    else:
                        ws.cell(row=proxima_linha, column=5, value=None)

                    wb.save(ARQUIVO_CLIENTES)
                    
                    # Atualizar nome do arquivo do cliente se mudou
                    if nome.upper() != cliente_selecionado:
                        caminho_antigo = PASTA_CLIENTES / f"{cliente_selecionado}.xlsx"
                        caminho_novo = PASTA_CLIENTES / f"{nome.upper()}.xlsx"
                        if os.path.exists(caminho_antigo):
                            os.rename(caminho_antigo, caminho_novo)

                    custom_messagebox("info", "Sucesso", "Cliente atualizado com sucesso!")
                    self.atualizar_lista_clientes()
                    janela_edicao.destroy()

                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao salvar alterações: {str(e)}")

            # Frame para botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)

            ttk.Button(frame_botoes, 
                    text="Salvar", 
                    command=salvar_alteracoes).pack(side='left', padx=5)
            ttk.Button(frame_botoes, 
                    text="Cancelar", 
                    command=janela_edicao.destroy).pack(side='left', padx=5)

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir editor: {str(e)}")   
    

    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes baseado nos arquivos Excel disponíveis, filtrando apenas os ativos"""
        try:
            # Carregar arquivo de clientes
            caminho_base = ARQUIVO_CLIENTES
            workbook = load_workbook(caminho_base)
            sheet = workbook['Clientes']  # Assumindo que existe uma aba chamada 'Clientes'
            
            # Limpar lista atual
            self.cliente_combobox['values'] = []
            
            # Pegar clientes ativos (sem data final ou com data final vazia)
            clientes = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome do cliente está na primeira coluna
                    # Verificar se a data final está vazia (cliente ativo)
                    data_final = row[4] if len(row) > 4 else None
                    
                    if not data_final:  # Se não tiver data final, é um cliente ativo
                        clientes.append(row[0])
            
            # Atualizar combobox com lista ordenada
            self.cliente_combobox['values'] = sorted(clientes)
            workbook.close()
            
        except FileNotFoundError:
            # Se o arquivo não existir, criar novo
            self.criar_arquivo_clientes()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar clientes: {str(e)}")


    def setup_aba_fornecedor(self):
        """Configura a aba de fornecedor com layout elegante e botões de tamanho médio"""
        # Criar estilo para botões médios
        style = ttk.Style()
        style.configure('Medium.TButton', font=('Arial', 11), padding=(10, 6))
        
        # Frame para exibir o cliente selecionado
        frame_cliente = ttk.Frame(self.aba_fornecedor)
        frame_cliente.pack(fill='x', padx=10, pady=5)
        
        # Label para mostrar o cliente selecionado
        self.lbl_cliente_fornecedor = ttk.Label(
            frame_cliente, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_fornecedor.pack(anchor='w', padx=5)
        
        # Frame de busca com tamanho reduzido
        frame_busca = ttk.LabelFrame(self.aba_fornecedor, text="Busca de Fornecedor")
        frame_busca.pack(fill='x', padx=10, pady=5)

        # Frame interno para organizar os elementos de busca
        busca_interno = ttk.Frame(frame_busca)
        busca_interno.pack(fill='x', padx=5, pady=5)

        # Campo de busca
        ttk.Label(busca_interno, text="Nome:", font=('Arial', 10)).pack(side='left', padx=5)
        self.busca_entry = ttk.Entry(busca_interno, font=('Arial', 10), width=40)
        self.busca_entry.pack(side='left', padx=5)
        self.busca_entry.bind('<Return>', lambda e: self.buscar_fornecedor())

        # Botão de busca
        ttk.Button(busca_interno, 
                text="Buscar", 
                command=self.buscar_fornecedor,
                style='Medium.TButton').pack(side='left', padx=10)

        # Frame principal para resultados
        frame_resultados = ttk.Frame(self.aba_fornecedor)
        frame_resultados.pack(fill='both', expand=True, padx=10, pady=5)

        # Lista de resultados com scrollbar
        frame_tree = ttk.Frame(frame_resultados)
        frame_tree.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Scrollbar vertical
        scroll_y = ttk.Scrollbar(frame_tree, orient='vertical')
        scroll_y.pack(side='right', fill='y')
        
        # Treeview para resultados
        self.tree_fornecedores = ttk.Treeview(frame_tree, 
                                            columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                            show='headings',
                                            yscrollcommand=scroll_y.set) 
        
        self.tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_fornecedores.heading('Nome', text='Nome')
        self.tree_fornecedores.heading('Categoria', text='Categoria')
        
        # Configurar larguras das colunas
        self.tree_fornecedores.column('CNPJ/CPF', width=150)
        self.tree_fornecedores.column('Nome', width=300)
        self.tree_fornecedores.column('Categoria', width=100)
        
        self.tree_fornecedores.pack(side='left', fill='both', expand=True)
        scroll_y.config(command=self.tree_fornecedores.yview)
        
        # Adicionar evento de duplo clique para selecionar fornecedor
        self.tree_fornecedores.bind('<Double-1>', lambda e: self.selecionar_fornecedor())

        # Frame para botões de ação do fornecedor
        frame_acoes = ttk.Frame(self.aba_fornecedor)
        frame_acoes.pack(fill='x', padx=10, pady=5)

        ttk.Button(frame_acoes, 
                text="Novo Fornecedor", 
                command=self.novo_fornecedor,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_acoes, 
                text="Editar Fornecedor", 
                command=self.editar_fornecedor,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_acoes, 
                text="Selecionar", 
                command=self.selecionar_fornecedor,
                style='Medium.TButton').pack(side='left', padx=5)

        self.adicionar_botao_gerenciar_lancamentos()

        # Separador para dividir visualmente as seções
        ttk.Separator(self.aba_fornecedor, orient='horizontal').pack(fill='x', padx=10, pady=5)

        # Frame para taxas e processamento
        frame_taxas = ttk.LabelFrame(self.aba_fornecedor, text="Funções Administrativas")
        frame_taxas.pack(fill='x', padx=10, pady=5)

        # Container para botões de taxas
        frame_botoes_taxas = ttk.Frame(frame_taxas)
        frame_botoes_taxas.pack(fill='x', padx=5, pady=8)

        ttk.Button(
            frame_botoes_taxas, 
            text="Controle de Pagamentos de Taxa",
            command=self.abrir_controle_pagamentos,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_taxas, 
            text="Finalização de Quinzena",
            command=self.abrir_finalizacao_quinzena,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        # Separador para dividir visualmente as seções
        ttk.Separator(self.aba_fornecedor, orient='horizontal').pack(fill='x', padx=10, pady=5)

        # Frame de botões gerais na parte inferior
        frame_botoes_fornecedor = ttk.Frame(self.aba_fornecedor)
        frame_botoes_fornecedor.pack(fill='x', padx=10, pady=10, side='bottom')
        
        ttk.Button(frame_botoes_fornecedor, 
                text="Visualizar Lançamentos", 
                command=self.visualizar_lancamentos,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_botoes_fornecedor, 
                text="Enviar Registros", 
                command=self.enviar_dados,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(
            frame_botoes_fornecedor, 
            text="Importar Folha RH", 
            command=self.importar_folha_rh,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(frame_botoes_fornecedor, 
                text="Voltar ao Menu", 
                command=self.voltar_menu,
                style='Medium.TButton').pack(side='right', padx=5)
        ttk.Button(frame_botoes_fornecedor, 
                text="Sair", 
                command=self.sair_sistema,
                style='Medium.TButton').pack(side='right', padx=5)



    def validar_tipo_despesa(self, P):
        """
        Valida entrada do tipo de despesa
        Args:
            P: valor proposto após a modificação
        """
        if P == "": return True  # Permite campo vazio
        if not P.isdigit(): return False  # Permite apenas dígitos
        return 1 <= int(P) <= 6  # Permite apenas valores entre 1 e 6


    def setup_aba_dados(self):
        """Configura a aba de entrada de dados com layout aprimorado e ordem de campos otimizada"""
        # Verificar se o estilo Medium.TButton já existe
        style = ttk.Style()
        if not style.lookup('Medium.TButton', 'font'):
            style.configure('Medium.TButton', font=('Arial', 11), padding=(10, 6))

        # Frame para cabeçalho com informações do cliente
        frame_cabecalho = ttk.Frame(self.aba_dados)
        frame_cabecalho.pack(fill='x', padx=10, pady=5)
        
        # Label do cliente destacado
        self.cliente_label = ttk.Label(frame_cabecalho, 
                                    text="Cliente: Nenhum selecionado", 
                                    font=('Arial', 12, 'bold'),
                                    foreground='#0056b3')
        self.cliente_label.pack(side='left', padx=5)
        
        # Frame para data de referência
        frame_data = ttk.LabelFrame(self.aba_dados, text="Data de Referência")
        frame_data.pack(fill='x', padx=10, pady=8)
        
        # Container interno para organização da data
        frame_data_interno = ttk.Frame(frame_data)
        frame_data_interno.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_data_interno, text="Data do Relatório:", font=('Arial', 10)).pack(side='left', padx=5)
        
        def calcular_data_rel():
            hoje = datetime.now()
            if 6 <= hoje.day <= 20:
                data_rel = hoje.replace(day=20)
            else:
                if hoje.day > 20:
                    data_rel = (hoje + relativedelta(months=1)).replace(day=5)
                else:
                    data_rel = hoje.replace(day=5)
            return data_rel

        self.data_rel_entry = DateEntry(
            frame_data_interno,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2,
            font=('Arial', 10),
        )
        self.data_rel_entry.pack(side='left', padx=5, pady=5)
        
        # Definir data de referência inicial
        data_rel_inicial = calcular_data_rel()
        self.data_rel_entry.set_date(data_rel_inicial)
        
        def validar_entrada_data(event=None):
            data = self.data_rel_entry.get()
            if not validar_data(data):
                custom_messagebox("error", "Erro", "Data inválida! Use o formato dd/mm/aaaa")
                self.data_rel_entry.delete(0, tk.END)
                self.data_rel_entry.insert(0, datetime.now().strftime('%d/%m/%Y'))
                return False
            return True
        
        self.data_rel_entry.bind('<FocusOut>', validar_entrada_data)  # Valida quando perde o foco
        
        # Frame para dados do fornecedor
        frame_fornecedor = ttk.LabelFrame(self.aba_dados, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', padx=10, pady=8)
        
        # Adicione esta linha para tornar frame_fornecedor um atributo da classe
        self.frame_fornecedor = frame_fornecedor
        
        # Grid para organizar os campos de fornecedor de forma mais equilibrada
        self.campos_fornecedor = {}
        campos = [('cnpj_cpf', 'CNPJ/CPF:'), 
                ('nome', 'Nome:'), 
                ('categoria', 'Categoria:')]
        
        for row, (campo, label) in enumerate(campos):
            ttk.Label(frame_fornecedor, text=label, font=('Arial', 10)).grid(row=row, column=0, padx=5, pady=5, sticky='e')
            entry = ttk.Entry(frame_fornecedor, width=40, font=('Arial', 10))
            entry.grid(row=row, column=1, padx=5, pady=5, sticky='ew')
            if campo != 'categoria':
                entry.config(state='readonly')
            self.campos_fornecedor[campo] = entry
        
        # Frame para forma de pagamento
        frame_pagamento = ttk.Frame(frame_fornecedor)
        frame_pagamento.grid(row=len(campos), column=0, columnspan=2, pady=5, sticky='ew')
        
        ttk.Label(frame_pagamento, text="Forma de Pagamento:", font=('Arial', 10)).pack(side='left', padx=5)
        self.forma_pagamento_combo = ttk.Combobox(
            frame_pagamento,
            textvariable=self.forma_pagamento_var,
            values=["PIX", "TED"],
            state="readonly",
            width=10,
            font=('Arial', 10)
        )
        self.forma_pagamento_combo.pack(side='left', padx=5)
        self.forma_pagamento_combo.bind('<<ComboboxSelected>>', self.atualizar_dados_bancarios)
        
        # Dados Bancários (agora após a forma de pagamento)
        ttk.Label(frame_fornecedor, text="Dados Bancários:", font=('Arial', 10)).grid(row=len(campos) + 1, column=0, padx=5, pady=5, sticky='e')
        entry = ttk.Entry(frame_fornecedor, width=40, state='readonly', font=('Arial', 10))
        entry.grid(row=len(campos) + 1, column=1, padx=5, pady=5, sticky='ew')
        self.campos_fornecedor['dados_bancarios'] = entry
        
        # Configure expandability of columns
        frame_fornecedor.columnconfigure(1, weight=1)
        
        # Frame para botões de parcelamento
        frame_parcelamento = ttk.Frame(self.aba_dados)
        frame_parcelamento.pack(fill='x', padx=10, pady=5)
        
        # Inicializar o gestor de parcelas com a janela root
        self.gestor_parcelas = GestorParcelas(self)
        
        ttk.Button(
            frame_parcelamento,
            text="Parcelar Despesa",
            command=self.abrir_parcelamento,
            style='Medium.TButton'
        ).pack(side='left', padx=5)
        
        # Frame para dados da despesa com layout em grid otimizado
        frame_despesa = ttk.LabelFrame(self.aba_dados, text="Dados da Despesa")
        frame_despesa.pack(fill='both', expand=True, padx=10, pady=8)
        
        # Adicionar as opções de referência para tipo 1
        self.opcoes_referencia_tipo1 = [
            'DIÁRIA', 'SALÁRIO', 'TRANSPORTE', 
            'FÉRIAS', '13º SALÁRIO', 'RESCISÃO', 'CAFÉ'
        ]
        
        self.campos_despesa = {}
        
        # Criar grid layout de 4x4 para os campos da despesa
        # Coluna 0 e 1: Labels e campos da esquerda
        # Coluna 2 e 3: Labels e campos da direita
        
        # ===== LADO ESQUERDO (valores numéricos) =====
        
        # Tipo Despesa
        ttk.Label(frame_despesa, text="Tipo Despesa (1-7):", font=('Arial', 10)).grid(
            row=0, column=0, padx=5, pady=5, sticky='e')
        vcmd = (frame_despesa.register(self.validar_tipo_despesa), '%P')
        self.campos_despesa['tp_desp'] = ttk.Entry(
            frame_despesa, validate='key', validatecommand=vcmd, font=('Arial', 10), width=10)
        self.campos_despesa['tp_desp'].grid(row=0, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # Valor Unitário
        ttk.Label(frame_despesa, text="Valor Unitário:", font=('Arial', 10)).grid(
            row=1, column=0, padx=5, pady=5, sticky='e')
        self.campos_despesa['vr_unit'] = ttk.Entry(frame_despesa, font=('Arial', 10), width=15)
        self.campos_despesa['vr_unit'].grid(row=1, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # Dias
        ttk.Label(frame_despesa, text="Dias:", font=('Arial', 10)).grid(
            row=2, column=0, padx=5, pady=5, sticky='e')
        self.campos_despesa['dias'] = ttk.Entry(frame_despesa, font=('Arial', 10), width=8)
        self.campos_despesa['dias'].grid(row=2, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # Valor Total
        ttk.Label(frame_despesa, text="Valor Total:", font=('Arial', 10)).grid(
            row=3, column=0, padx=5, pady=5, sticky='e')
        self.campos_despesa['valor'] = ttk.Entry(
            frame_despesa, state='readonly', font=('Arial', 10), width=15)
        self.campos_despesa['valor'].grid(row=3, column=1, padx=(5, 20), pady=5, sticky='w')
        
        # ===== LADO DIREITO (texto) =====
        
        # Referência
        ttk.Label(frame_despesa, text="Referência:", font=('Arial', 10)).grid(
            row=0, column=2, padx=5, pady=5, sticky='e')
        self.campos_despesa['referencia'] = ttk.Combobox(
            frame_despesa, font=('Arial', 10), width=40)
        self.campos_despesa['referencia']['values'] = self.opcoes_referencia_tipo1
        self.campos_despesa['referencia'].grid(row=0, column=3, padx=5, pady=5, sticky='ew')
        self.campos_despesa['referencia'].bind(
            '<<ComboboxSelected>>', lambda e: self.calcular_valor_total())
        
        # NF
        ttk.Label(frame_despesa, text="NF:", font=('Arial', 10)).grid(
            row=1, column=2, padx=5, pady=5, sticky='e')
        self.campos_despesa['nf'] = ttk.Entry(frame_despesa, font=('Arial', 10), width=15)
        self.campos_despesa['nf'].grid(row=1, column=3, padx=5, pady=5, sticky='w')
        
        # Data Vencimento
        ttk.Label(frame_despesa, text="Data Vencimento:", font=('Arial', 10)).grid(
            row=2, column=2, padx=5, pady=5, sticky='e')
        self.campos_despesa['dt_vencto'] = DateEntry(
            frame_despesa,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2,
            font=('Arial', 10),
            width=15
        )
        self.campos_despesa['dt_vencto'].grid(row=2, column=3, padx=5, pady=5, sticky='w')
        # Inicializa o campo vazio
        self.campos_despesa['dt_vencto'].delete(0, tk.END)
        
        # Observação
        ttk.Label(frame_despesa, text="Observação:", font=('Arial', 10)).grid(
            row=3, column=2, padx=5, pady=5, sticky='e')
        self.campos_despesa['observacao'] = ttk.Entry(frame_despesa, font=('Arial', 10), width=40)
        self.campos_despesa['observacao'].grid(row=3, column=3, padx=5, pady=5, sticky='ew')
        
        # Configurar peso da coluna para expandir apenas os campos de referência e observação
        frame_despesa.columnconfigure(3, weight=1)  # Apenas a coluna 3 (campos expansíveis) cresce
        
        # Inserir valor padrão para dias
        self.campos_despesa['dias'].insert(0, "1")
        
        # Bindings
        self.campos_despesa['vr_unit'].bind('<KeyRelease>', self.calcular_valor_total)
        self.campos_despesa['dias'].bind('<KeyRelease>', self.calcular_valor_total)
        self.campos_despesa['tp_desp'].bind('<KeyRelease>', self.verificar_tipo_despesa)
        
        # Configurar a ordem de tab para seguir o fluxo de trabalho natural
        self.campos_despesa['tp_desp'].bind('<Return>', lambda e: self.campos_despesa['referencia'].focus())
        self.campos_despesa['referencia'].bind('<Return>', lambda e: self.campos_despesa['vr_unit'].focus())
        self.campos_despesa['vr_unit'].bind('<Return>', lambda e: self.campos_despesa['dias'].focus())
        self.campos_despesa['dias'].bind('<Return>', lambda e: self.campos_despesa['nf'].focus())
        self.campos_despesa['nf'].bind('<Return>', lambda e: self.campos_despesa['dt_vencto'].focus())
        self.campos_despesa['dt_vencto'].bind('<Return>', lambda e: self.campos_despesa['observacao'].focus())
        
        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.aba_dados)
        frame_botoes.pack(fill='x', padx=10, pady=10, side='bottom')
        
        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.aba_dados)
        frame_botoes.pack(fill='x', padx=10, pady=10, side='bottom')
        
        # Organizar botões com Adicionar em destaque à direita
        ttk.Button(frame_botoes, text="Cancelar", 
                command=self.cancelar_entrada,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Visualizar Lançamentos", 
                command=self.visualizar_lancamentos,
                style='Medium.TButton').pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Enviar", 
                command=self.enviar_dados,
                style='Medium.TButton').pack(side='left', padx=5)
        
        # Botão Adicionar em destaque (lado direito)
        adicionar_btn = ttk.Button(frame_botoes, text="Adicionar", 
                                command=self.adicionar_dados,
                                style='Medium.TButton')
        adicionar_btn.pack(side='right', padx=5)
        
        # Configurar um estilo especial para o botão Adicionar (opcional)
        style = ttk.Style()
        style.configure('Destaque.TButton', 
                    background='#0056b3',  # Esta propriedade pode não ter efeito em todos os temas
                    font=('Arial', 11, 'bold'))
        adicionar_btn.configure(style='Destaque.TButton')



    def visualizar_lancamentos(self):
        """Abre a janela de visualização de lançamentos pendentes"""
        if hasattr(self, 'visualizador') and self.visualizador and hasattr(self.visualizador, 'janela') and self.visualizador.janela.winfo_exists():
            # Se o visualizador já existir, apenas trazê-lo para frente
            self.visualizador.janela.lift()
            self.visualizador.janela.focus_force()
            return
        
        # Criar nova instância do visualizador
        self.visualizador = VisualizadorLancamentos(self)
        
        # Configurar callback para quando a janela for fechada
        self.visualizador.janela.protocol("WM_DELETE_WINDOW", self.on_visualizador_close)
        
        # Atualizar dados
        self.visualizador.dados_para_incluir = self.dados_para_incluir.copy()
        self.visualizador.atualizar_dados(self.dados_para_incluir)
        
        # Garantir que a janela fique na frente
        self.visualizador.janela.lift()
        self.visualizador.janela.focus_force()

    def on_visualizador_close(self):
        """Manipula o fechamento da janela do visualizador"""
        # Atualizar dados_para_incluir com os dados mais recentes do visualizador
        if self.visualizador:
            self.dados_para_incluir = self.visualizador.get_dados_atualizados()
            self.visualizador.janela.destroy()
            self.visualizador = None

    def adicionar_botao_gerenciar_lancamentos(self):
        """Adiciona botão para gerenciar lançamentos na aba de fornecedor"""
        frame_gerenciamento = ttk.LabelFrame(self.aba_fornecedor, text="Gerenciamento de Dados")
        frame_gerenciamento.pack(fill='x', padx=10, pady=5)
        
        frame_botoes_ger = ttk.Frame(frame_gerenciamento)
        frame_botoes_ger.pack(fill='x', padx=5, pady=8)
        
        ttk.Button(
            frame_botoes_ger, 
            text="Gerenciar Lançamentos",
            command=self.abrir_gerenciador_lancamentos,
            style='Medium.TButton'
        ).pack(side='left', padx=5)

    def abrir_gerenciador_lancamentos(self):
        """Abre o gerenciador de lançamentos"""
        if not hasattr(self, 'gerenciador_lancamentos') or self.gerenciador_lancamentos is None:
            self.gerenciador_lancamentos = GerenciadorLancamentos(self)
        self.gerenciador_lancamentos.abrir_gerenciador()

    def processar_parcelas(self):
        """Processa as parcelas geradas mantendo os dados do fornecedor"""
        print("Iniciando processamento de parcelas...")
        
        # Verificar se há parcelas para processar
        if not hasattr(self, 'gestor_parcelas') or not self.gestor_parcelas.parcelas:
            print("Nenhuma parcela para processar")
            return False
            
        # Validar se há fornecedor selecionado
        if not self.campos_fornecedor['cnpj_cpf'].get():
            custom_messagebox("error", "Erro", "Selecione um fornecedor antes de processar as parcelas!")
            return False
            
        # Guardar dados do fornecedor atual
        dados_fornecedor = {
            'cnpj_cpf': self.campos_fornecedor['cnpj_cpf'].get(),
            'nome': self.campos_fornecedor['nome'].get(),
            'categoria': self.campos_fornecedor['categoria'].get(),
            'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get()
        }
        
        print(f"Dados do fornecedor capturados: {dados_fornecedor}")
        total_parcelas = len(self.gestor_parcelas.parcelas)
        print(f"Total de parcelas a processar: {total_parcelas}")
        
        try:
            processadas = 0
            for i, parcela in enumerate(self.gestor_parcelas.parcelas, 1):
                print(f"\nProcessando parcela {i} de {total_parcelas}")
                
                # Restaurar dados do fornecedor antes de cada parcela
                for campo, valor in dados_fornecedor.items():
                    entry = self.campos_fornecedor[campo]
                    entry.config(state='normal')
                    entry.delete(0, tk.END)
                    entry.insert(0, valor)
                    if campo != 'categoria':
                        entry.config(state='readonly')
                
                print(f"Dados do fornecedor restaurados para parcela {i}")

                
                # Preencher dados da parcela
                self.data_rel_entry.set_date(datetime.strptime(parcela['data_rel'], '%d/%m/%Y'))
                self.campos_despesa['tp_desp'].delete(0, tk.END)
                self.campos_despesa['tp_desp'].insert(0, self.gestor_parcelas.tipo_despesa_valor)
                self.campos_despesa['nf'].delete(0, tk.END)
                self.campos_despesa['nf'].insert(0, parcela['nf'])
                
                if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                    self.campos_despesa['referencia'].set(parcela['referencia'])
                else:
                    self.campos_despesa['referencia'].delete(0, tk.END)
                    self.campos_despesa['referencia'].insert(0, parcela['referencia'])
                
                self.campos_despesa['vr_unit'].delete(0, tk.END)
                self.campos_despesa['vr_unit'].insert(0, f"{parcela['valor']:.2f}")
                
                self.campos_despesa['valor'].config(state='normal')
                self.campos_despesa['valor'].delete(0, tk.END)
                self.campos_despesa['valor'].insert(0, f"{parcela['valor']:.2f}")
                self.campos_despesa['valor'].config(state='readonly')
                
                self.campos_despesa['dt_vencto'].set_date(
                    datetime.strptime(parcela['dt_vencto'], '%d/%m/%Y')
                )
                
                # Adicionar à lista de dados e verificar sucesso
                if self.adicionar_dados(eh_parcelamento=True):
                    processadas += 1
                    print(f"Parcela {i} processada com sucesso")
                else:
                    print(f"Falha ao processar parcela {i}")
            
            # Relatório final
            if processadas == total_parcelas:
                custom_messagebox("info", "Sucesso", 
                                  f"Todas as {total_parcelas} parcelas foram processadas com sucesso!")
            else:
                custom_messagebox("warning",  "Aviso", 
                                     f"Apenas {processadas} de {total_parcelas} parcelas foram processadas.")
            
            return processadas == total_parcelas
            
        except Exception as e:
            erro_msg = f"Erro ao processar parcelas: {str(e)}"
            print(erro_msg)
            custom_messagebox("error", "Erro", erro_msg)
            return False
            
        finally:
            self.limpar_campos_despesa()
            print("Processamento de parcelas finalizado")



    def abrir_parcelamento(self):
        """Abre a janela de parcelamento e processa os dados após o fechamento"""
        print("\nIniciando processo de parcelamento...")
        
        # Verificar se há fornecedor selecionado
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get()
        if not cnpj_cpf:
            print("Erro: Fornecedor não selecionado")
            custom_messagebox("error", "Erro", "Selecione um fornecedor antes de criar parcelas!")
            return

        print("\nCapturando dados do fornecedor...")
        dados_fornecedor = {
            'cnpj_cpf': cnpj_cpf,
            'nome': self.campos_fornecedor['nome'].get(),
            'categoria': self.campos_fornecedor['categoria'].get(),
            'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get()
        }
        print(f"Dados capturados: {dados_fornecedor}")
        
        # Validar se todos os campos do fornecedor estão preenchidos
        if not all(dados_fornecedor.values()):
            print("Erro: Dados do fornecedor incompletos")
            custom_messagebox("error", "Erro", "Dados do fornecedor incompletos!")
            return

        print("Abrindo janela de parcelamento...")
        self.gestor_parcelas.abrir_janela_parcelas()
        self.root.wait_window(self.gestor_parcelas.janela_parcelas)

        if hasattr(self.gestor_parcelas, 'parcelas') and self.gestor_parcelas.parcelas:
            print(f"Processando {len(self.gestor_parcelas.parcelas)} parcelas...")
            
            success = True
            for i, parcela in enumerate(self.gestor_parcelas.parcelas, 1):
                try:
                    print(f"\nProcessando parcela {i}")
                    
                    # Restaurar dados do fornecedor
                    for campo, valor in dados_fornecedor.items():
                        entry = self.campos_fornecedor[campo]
                        entry.config(state='normal')
                        entry.delete(0, tk.END)
                        entry.insert(0, valor)
                        if campo != 'categoria':
                            entry.config(state='readonly')
                    
                    # Preencher dados da parcela
                    self.data_rel_entry.set_date(datetime.strptime(parcela['data_rel'], '%d/%m/%Y'))
                    
                    self.campos_despesa['tp_desp'].delete(0, tk.END)
                    self.campos_despesa['tp_desp'].insert(0, self.gestor_parcelas.tipo_despesa_valor)
                    self.campos_despesa['nf'].delete(0, tk.END)
                    self.campos_despesa['nf'].insert(0, parcela['nf'])
                    
                    if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                        self.campos_despesa['referencia'].set(parcela['referencia'])
                    else:
                        self.campos_despesa['referencia'].delete(0, tk.END)
                        self.campos_despesa['referencia'].insert(0, parcela['referencia'])
                    
                    self.campos_despesa['vr_unit'].delete(0, tk.END)
                    self.campos_despesa['vr_unit'].insert(0, f"{parcela['valor']:.2f}")
                    
                    self.campos_despesa['dias'].delete(0, tk.END)
                    self.campos_despesa['dias'].insert(0, '1')
                    
                    self.campos_despesa['valor'].config(state='normal')
                    self.campos_despesa['valor'].delete(0, tk.END)
                    self.campos_despesa['valor'].insert(0, f"{parcela['valor']:.2f}")
                    self.campos_despesa['valor'].config(state='readonly')
                    
                    self.campos_despesa['dt_vencto'].set_date(
                        datetime.strptime(parcela['dt_vencto'], '%d/%m/%Y')
                    )
                    
                    # Adicionar à lista de dados
                    if not self.adicionar_dados(eh_parcelamento=True):
                        print(f"Falha ao adicionar parcela {i}")
                        success = False
                        break
                    
                    print(f"Parcela {i} processada com sucesso")
                    
                except Exception as e:
                    success = False
                    print(f"Erro ao processar parcela {i}: {str(e)}")
                    custom_messagebox("error", "Erro", f"Erro ao processar parcela {i}: {str(e)}")
                    break
            
            if success:
                custom_messagebox("info", "Sucesso", 
                                  f"Todas as {len(self.gestor_parcelas.parcelas)} parcelas foram processadas!")
                # Calcular a data de referência padrão
                hoje = datetime.now()
                if 6 <= hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    if hoje.day > 20:
                        data_rel = (hoje + relativedelta(months=1)).replace(day=5)
                    else:
                        data_rel = hoje.replace(day=5)
                
                # Restaurar a data de referência padrão
                self.data_rel_entry.set_date(data_rel)
                
                self.limpar_campos_despesa()
                self.notebook.select(1)  # Volta para aba fornecedor
            else:
                custom_messagebox("error", "Erro", "Houve um erro no processamento das parcelas.")
        else:
            print("Nenhuma parcela para processar")



    def abrir_calendario(self):
        try:
            top = Toplevel(self.root)
            top.title("Selecionar Data")
            top.geometry("300x250")
            top.grab_set()  # Torna a janela modal
        
            cal = Calendar(top, selectmode='day', 
                          date_pattern='dd/mm/yyyy',
                          locale='pt_BR')
            cal.pack(padx=10, pady=10)
        
            def selecionar_data():
                data = cal.get_date()
                self.data_rel_entry.delete(0, tk.END)
                self.data_rel_entry.insert(0, data)
                top.destroy()
        
            ttk.Button(top, text="OK", command=selecionar_data).pack(pady=5)
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir calendário: {str(e)}")

    def atualizar_campo_referencia(self, event=None):
        """Atualiza o campo de referência baseado no tipo de despesa"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
    
        try:
            if tp_desp == '1':
                # Redefine as opções e configura como readonly
                self.campos_despesa['referencia']['values'] = self.opcoes_referencia_tipo1
                self.campos_despesa['referencia'].config(state='readonly')
                # Seleciona o primeiro item como padrão
                if self.opcoes_referencia_tipo1:
                    self.campos_despesa['referencia'].set(self.opcoes_referencia_tipo1[0])
            else:
                # Para outros tipos, limpa a seleção e permite digitação
                self.campos_despesa['referencia'].set('')
                self.campos_despesa['referencia']['values'] = []
                self.campos_despesa['referencia'].config(state='normal')
            
        except Exception as e:
            print(f"Erro ao atualizar campo referência: {str(e)}")

    def atualizar_dados_bancarios(self, event=None):
        """Atualiza os dados bancários baseado no tipo de despesa"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get().strip()
    
        if not cnpj_cpf:  # Se não houver fornecedor selecionado
            return
        
        fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
        if not fornecedor_completo:
            return
        
        self.campos_fornecedor['dados_bancarios'].config(state='normal')
        self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
        
        # Construir dados bancários baseado na forma de pagamento
        forma_pagamento = self.forma_pagamento_var.get()
        
        if forma_pagamento == "PIX" and fornecedor_completo['chave_pix']:
            dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
        else:
            # Estrutura para TED
            dados_ted = []
            if fornecedor_completo['banco']: dados_ted.append(str(fornecedor_completo['banco']))
            if fornecedor_completo['op']: dados_ted.append(str(fornecedor_completo['op']))
            if fornecedor_completo['agencia']: dados_ted.append(str(fornecedor_completo['agencia']))
            if fornecedor_completo['conta']: dados_ted.append(str(fornecedor_completo['conta']))
            # SEMPRE adicionar o CNPJ/CPF para TED
            dados_ted.append(str(fornecedor_completo['cnpj_cpf']))
            
            dados_bancarios = ' - '.join(filter(None, dados_ted))

        if dados_bancarios.strip() in ['', ' - ']:
            dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
            
        self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
        self.campos_fornecedor['dados_bancarios'].config(state='readonly')
    

    def cancelar_entrada(self):
        """Cancela a entrada de dados atual e retorna à aba fornecedor"""
        if any(self.campos_despesa[campo].get() for campo in ['tp_desp', 'referencia', 'vr_unit']):
            if custom_messagebox("yesno", "Confirmação", "Deseja descartar os dados atuais?"):
                self.limpar_campos_despesa()
                self.notebook.select(1)  # Volta para aba fornecedor
        else:
            self.notebook.select(1)  # Volta para aba fornecedor


    def buscar_fornecedor(self):
        termo = self.busca_entry.get()
        buscar_fornecedor(self.tree_fornecedores, termo)

    def novo_fornecedor(self):
        """Abre janela para cadastro de novo fornecedor"""
        self.janela_fornecedor = tk.Toplevel(self.root)
        self.janela_fornecedor.title("Novo Fornecedor")
        self.setup_formulario_fornecedor()

    def editar_fornecedor(self):
        """Abre janela para edição de fornecedor existente"""
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um fornecedor para editar")
            return

        # Buscar dados completos do fornecedor
        fornecedor = self.buscar_fornecedor_completo(
            self.tree_fornecedores.item(selecionado)['values'][0]
        )
        if not fornecedor:
            custom_messagebox("error", "Erro", "Fornecedor não encontrado")
            return

        # Criar janela de edição
        self.janela_fornecedor = tk.Toplevel(self.root)
        self.janela_fornecedor.title("Editar Fornecedor")
        self.setup_formulario_fornecedor(modo_edicao=True)

        try:
            # Determinar tipo de pessoa baseado no tamanho do CNPJ/CPF
            cnpj_cpf = str(fornecedor['cnpj_cpf']).strip()
            tipo_pessoa = 'PJ' if len(cnpj_cpf) > 11 else 'PF'

            # Preencher e configurar campos não editáveis
            # CNPJ/CPF
            self.campos_form['cnpj_cpf'].insert(0, cnpj_cpf.zfill(14 if tipo_pessoa == 'PJ' else 11))
            self.campos_form['cnpj_cpf'].config(state='readonly')
            
            # Tipo Pessoa
            self.campos_form['tipo_pessoa'].set(tipo_pessoa)
            self.campos_form['tipo_pessoa'].config(state='disabled')
            
            # Razão Social
            self.campos_form['razao_social'].insert(0, fornecedor['razao_social'] or '')
            self.campos_form['razao_social'].config(state='readonly')
            
            # Preencher campos editáveis
            self.campos_form['nome'].insert(0, fornecedor['nome'] or '')
            self.campos_form['telefone'].insert(0, fornecedor['telefone'] or '')
            self.campos_form['email'].insert(0, fornecedor['email'] or '')
            self.campos_form['banco'].insert(0, fornecedor['banco'] or '')
            self.campos_form['op'].insert(0, fornecedor['op'] or '')
            self.campos_form['agencia'].insert(0, fornecedor['agencia'] or '')
            self.campos_form['conta'].insert(0, fornecedor['conta'] or '')
            self.campos_form['chave_pix'].insert(0, fornecedor['chave_pix'] or '')
            
            # Categoria (pode ser combobox)
            if isinstance(self.campos_form['categoria'], ttk.Combobox):
                self.campos_form['categoria'].set(fornecedor['categoria'] or '')
            else:
                self.campos_form['categoria'].insert(0, fornecedor['categoria'] or '')
                
            self.campos_form['especificacao'].insert(0, fornecedor['especificacao'] or '')
            self.campos_form['vinculo'].insert(0, fornecedor['vinculo'] or '')

            # Centralizar a janela
            self.janela_fornecedor.update_idletasks()
            width = self.janela_fornecedor.winfo_width()
            height = self.janela_fornecedor.winfo_height()
            x = (self.janela_fornecedor.winfo_screenwidth() // 2) - (width // 2)
            y = (self.janela_fornecedor.winfo_screenheight() // 2) - (height // 2)
            self.janela_fornecedor.geometry('{}x{}+{}+{}'.format(width, height, x, y))
            
            # Tornar a janela modal
            self.janela_fornecedor.transient(self.root)
            self.janela_fornecedor.grab_set()

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar dados do fornecedor: {str(e)}")
            self.janela_fornecedor.destroy()

    def selecionar_fornecedor(self):
        """Seleciona o fornecedor e preenche seus dados"""
        fornecedor = selecionar_fornecedor(
            self.tree_fornecedores, 
            self.campos_fornecedor,
            self.campos_despesa,
            self.notebook,
            self.buscar_fornecedor_completo
        )
        if fornecedor:
            # Formatar CNPJ/CPF
            cnpj_cpf = str(fornecedor[0]).strip()
            self.campos_fornecedor['cnpj_cpf'].config(state='normal')
            self.campos_fornecedor['cnpj_cpf'].delete(0, tk.END)
            self.campos_fornecedor['cnpj_cpf'].insert(0, formatar_cnpj_cpf(cnpj_cpf))
            self.campos_fornecedor['cnpj_cpf'].config(state='readonly')
            
            # Carregar dados completos do fornecedor
            fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
            if fornecedor_completo:
                # Substituir o campo de categoria por Combobox
                self.campos_fornecedor['categoria'] = ttk.Combobox(
                    self.frame_fornecedor,  # Usando o atributo da classe
                    values=get_categorias_fornecedor(),
                    state='readonly',
                    width=30
                )
                self.campos_fornecedor['categoria'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')
        
                    
                # Definir categoria do fornecedor
                self.campos_fornecedor['categoria'].set(fornecedor_completo['categoria'])
                
                self.campos_fornecedor['dados_bancarios'].config(state='normal')
                self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
                
                # Construir dados bancários
                if fornecedor_completo['chave_pix']:
                    dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
                else:
                    dados_bancarios = (f"{fornecedor_completo['banco'] or ''} "
                                    f"{fornecedor_completo['op'] or ''} - "
                                    f"{fornecedor_completo['agencia'] or ''} "
                                    f"{fornecedor_completo['conta'] or ''}").strip()
                    
                if dados_bancarios.strip() in ['', ' - ']:
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
                
                self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
                self.campos_fornecedor['dados_bancarios'].config(state='readonly')
                
                # NOVO: Preencher campo de referência com a especificação do fornecedor, se disponível
                if fornecedor_completo['especificacao'] and hasattr(self, 'campos_despesa') and 'referencia' in self.campos_despesa:
                    if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                        # Para Combobox, verificamos se o valor está nas opções
                        especificacao = fornecedor_completo['especificacao'].strip()
                        valores = self.campos_despesa['referencia']['values']
                        
                        # Deixamos o campo livre para edição quando não for tipo 1
                        self.campos_despesa['referencia'].config(state='normal')
                        self.campos_despesa['referencia'].delete(0, tk.END)
                        self.campos_despesa['referencia'].insert(0, especificacao)
                    else:
                        # Para Entry normal
                        self.campos_despesa['referencia'].delete(0, tk.END)
                        self.campos_despesa['referencia'].insert(0, fornecedor_completo['especificacao'].strip())
                
                self.notebook.select(2)  # Vai para aba de dados
            

    def buscar_dados_bancarios(self, cnpj_cpf):
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
        
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == cnpj_cpf:
                    print(f"CNPJ/CPF encontrado: {cnpj_cpf}")
                    print(f"Dados da linha: {row}")
                    if row[14]:  # coluna O com dados bancários consolidados
                        return row[14]
                    return ""
            return ""
        except Exception as e:
            print(f"Erro ao buscar dados bancários: {e}")
            return ""

    # Adicionar os métodos de acesso
    def abrir_controle_pagamentos(self):
        """Abre o controle de pagamentos de taxa"""
        try:
            from gestao_taxas import GestaoTaxasAdministracao
            gestao = GestaoTaxasAdministracao(self.root)
            gestao.abrir_controle_pagamentos()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir controle de pagamentos: {str(e)}")

    def abrir_finalizacao_quinzena(self):
        """Abre a finalização de quinzena"""
        try:
            from gestao_taxas import GestaoTaxasAdministracao
            gestao = GestaoTaxasAdministracao(self.root)
            gestao.abrir_finalizacao_quinzena()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir finalização de quinzena: {str(e)}")    

    def buscar_fornecedor_completo(self, cnpj_cpf):
        """Busca todos os dados de um fornecedor"""
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
        
            cnpj_cpf = str(cnpj_cpf).zfill(14)  # Preenche com zeros à esquerda
            for row in ws.iter_rows(min_row=2):
                # Garante que o CNPJ/CPF da planilha também tenha 14 dígitos
                row_cnpj = str(row[0].value or '').zfill(14)
                if row_cnpj == cnpj_cpf:
                    fornecedor = {
                        'cnpj_cpf': row[0].value,
                        'tipo_pessoa': row[1].value,
                        'razao_social': row[2].value,
                        'nome': row[3].value,
                        'telefone': row[4].value,
                        'email': row[5].value,
                        'banco': row[6].value,
                        'op': row[7].value,
                        'agencia': row[8].value,
                        'conta': row[9].value,
                        'chave_pix': row[10].value,
                        'categoria': row[11].value,
                        'especificacao': row[12].value,
                        'vinculo': row[13].value,
                    }
                    return fornecedor
            return None
        except Exception as e:
            print(f"Erro ao buscar fornecedor: {e}")
            return None

        
    def setup_formulario_fornecedor(self, modo_edicao=False):
        """Configura o formulário de cadastro/edição de fornecedor"""
        formulario = ttk.Frame(self.janela_fornecedor)
        formulario.pack(padx=10, pady=5, fill='both', expand=True)

        # Campos principais
        campos_principais = ttk.LabelFrame(formulario, text="Dados Principais")
        campos_principais.pack(fill='x', pady=5)

        self.campos_form = {}

        # CNPJ/CPF na mesma linha
        tk.Label(campos_principais, text="CNPJ/CPF:*").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['cnpj_cpf'] = tk.Entry(campos_principais)
        self.campos_form['cnpj_cpf'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        self.campos_form['cnpj_cpf'].bind('<FocusOut>', self.atualizar_tipo_pessoa)
        
        tk.Label(campos_principais, text="Tipo:*").grid(row=0, column=2, padx=5, pady=2)
        self.campos_form['tipo_pessoa'] = ttk.Combobox(campos_principais, 
                                                    values=['PF', 'PJ'],
                                                    state='readonly',
                                                    width=5)
        self.campos_form['tipo_pessoa'].grid(row=0, column=3, padx=5, pady=2, sticky='w')
        
        # Razão Social e Nome
        tk.Label(campos_principais, text="Razão Social:*").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['razao_social'] = tk.Entry(campos_principais)
        self.campos_form['razao_social'].grid(row=1, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        self.campos_form['razao_social'].bind('<FocusOut>', self.copiar_para_nome)
        
        tk.Label(campos_principais, text="Nome Fantasia:*").grid(row=2, column=0, padx=5, pady=2)
        self.campos_form['nome'] = tk.Entry(campos_principais)
        self.campos_form['nome'].grid(row=2, column=1, columnspan=3, padx=5, pady=2, sticky='ew')


        # Contatos
        campos_contato = ttk.LabelFrame(formulario, text="Contato")
        campos_contato.pack(fill='x', pady=5)

        tk.Label(campos_contato, text="Telefone:").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['telefone'] = tk.Entry(campos_contato)
        self.campos_form['telefone'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_contato, text="Email:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['email'] = tk.Entry(campos_contato)
        self.campos_form['email'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Dados Bancários
        campos_bancarios = ttk.LabelFrame(formulario, text="Dados Bancários")
        campos_bancarios.pack(fill='x', pady=5)

        # Carregar configurações
        try:
            # Primeiro, garantir que as configurações foram carregadas
            carregar_configuracoes()  
            
            # Obter lista de bancos
            lista_bancos = get_bancos()
            
            # Modificação: Adicionar log para depuração
            print(f"Lista de bancos carregada: {lista_bancos}")
        except Exception as e:
            print(f"Erro ao carregar bancos: {str(e)}")
            lista_bancos = []  # Lista vazia como fallback

        # No frame de dados bancários, configurar o campo de banco para Combobox
        tk.Label(campos_bancarios, text="Banco:").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['banco'] = ttk.Combobox(
            campos_bancarios,
            values=lista_bancos,  # Usar a lista que carregamos
            state='readonly'  # Garantir que é readonly
        )
        self.campos_form['banco'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Operação:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['op'] = tk.Entry(campos_bancarios)
        self.campos_form['op'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Agência:").grid(row=2, column=0, padx=5, pady=2)
        self.campos_form['agencia'] = tk.Entry(campos_bancarios)
        self.campos_form['agencia'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Conta:").grid(row=3, column=0, padx=5, pady=2)
        self.campos_form['conta'] = tk.Entry(campos_bancarios)
        self.campos_form['conta'].grid(row=3, column=1, padx=5, pady=2, sticky='ew')

        # PIX
        campos_pix = ttk.LabelFrame(formulario, text="Chave PIX")
        campos_pix.pack(fill='x', pady=5)

        # Tipo de chave PIX
        ttk.Label(campos_pix, text="Tipo de Chave:").grid(row=0, column=0, padx=5, pady=2)
        self.tipo_pix = ttk.Combobox(
            campos_pix, 
            values=['Selecione', 'CNPJ/CPF', 'Telefone', 'Email'],
            state='readonly'
        )
        self.tipo_pix.grid(row=0, column=1, padx=5, pady=2)
        self.tipo_pix.set('Selecione')

        ttk.Label(campos_pix, text="Chave:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['chave_pix'] = ttk.Entry(campos_pix)
        self.campos_form['chave_pix'].grid(row=1, column=1, padx=5, pady=2)

        # Adicionar binding para atualização automática
        self.tipo_pix.bind('<<ComboboxSelected>>', self.atualizar_chave_pix)

        # Classificação
        campos_class = ttk.LabelFrame(formulario, text="Classificação")
        campos_class.pack(fill='x', pady=5)

        # Carregar categorias
        try:
            categorias = get_categorias_fornecedor()
            print(f"Categorias carregadas: {categorias}")
        except Exception as e:
            print(f"Erro ao carregar categorias: {str(e)}")
            categorias = ['ADM', 'DIV', 'LOC', 'MAT', 'MO', 'SERV', 'TP']  # Valores padrão

        # Categoria
        tk.Label(campos_class, text="Categoria:*").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['categoria'] = ttk.Combobox(campos_class, 
                                                    values=categorias,
                                                    state='readonly')  # Modificado para readonly
        self.campos_form['categoria'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        # Especificação
        tk.Label(campos_class, text="Especificação:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['especificacao'] = tk.Entry(campos_class)
        self.campos_form['especificacao'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Vínculo
        tk.Label(campos_class, text="Vínculo:").grid(row=2, column=0, padx=5, pady=2)
        self.campos_form['vinculo'] = tk.Entry(campos_class)
        self.campos_form['vinculo'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')

        # Botões de ação
        frame_botoes = ttk.Frame(formulario)
        frame_botoes.pack(fill='x', pady=10)

        ttk.Button(frame_botoes, 
                text="Salvar", 
                command=self.salvar_fornecedor).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                text="Cancelar", 
                command=self.janela_fornecedor.destroy).pack(side='left', padx=5)

    def atualizar_chave_pix(self, event=None):
        """Atualiza o campo de chave PIX baseado no tipo selecionado"""
        tipo_selecionado = self.tipo_pix.get()
        self.campos_form['chave_pix'].delete(0, tk.END)
        
        if tipo_selecionado == 'CNPJ/CPF':
            self.campos_form['chave_pix'].insert(0, self.campos_form['cnpj_cpf'].get())
        elif tipo_selecionado == 'Telefone':
            self.campos_form['chave_pix'].insert(0, self.campos_form['telefone'].get())
        elif tipo_selecionado == 'Email':
            self.campos_form['chave_pix'].insert(0, self.campos_form['email'].get())


    def atualizar_tipo_pessoa(self, event=None):
        """Determina automaticamente o tipo de pessoa baseado no CNPJ/CPF"""
        cnpj_cpf = self.campos_form['cnpj_cpf'].get().strip()
        # Remove caracteres não numéricos
        cnpj_cpf = ''.join(filter(str.isdigit, cnpj_cpf))
        
        if len(cnpj_cpf) <= 11:
            self.campos_form['tipo_pessoa'].set('PF')
        else:
            self.campos_form['tipo_pessoa'].set('PJ')

    def copiar_para_nome(self, event=None):
        """Copia a razão social para o nome se este estiver vazio"""
        razao_social = self.campos_form['razao_social'].get().strip()
        nome_atual = self.campos_form['nome'].get().strip()
        
        if razao_social and not nome_atual:
            self.campos_form['nome'].insert(0, razao_social)


    def salvar_fornecedor(self):
        """Salva os dados do fornecedor"""
        # Validar campos obrigatórios
        campos_obrigatorios = ['tipo_pessoa', 'cnpj_cpf', 'razao_social', 'nome', 'categoria']
        for campo in campos_obrigatorios:
            if not self.campos_form[campo].get().strip():
                custom_messagebox("error", "Erro", f"O campo {campo} é obrigatório!")
                return

        # Validar CNPJ/CPF
        tipo_pessoa = self.campos_form['tipo_pessoa'].get()
        cnpj_cpf = self.campos_form['cnpj_cpf'].get().strip()
        
        if not validar_cnpj_cpf(cnpj_cpf):
            custom_messagebox("error", "Erro", f"{'CPF' if tipo_pessoa == 'PF' else 'CNPJ'} inválido!")
            return
            
            cnpj_cpf = formatar_cnpj_cpf(cnpj_cpf)

        # Montar dados bancários
        if self.campos_form['chave_pix'].get():
            dados_bancarios = f"PIX: {self.campos_form['chave_pix'].get()}"
        else:
            dados_bancarios = (f"{self.campos_form['banco'].get()} "
                             f"{self.campos_form['op'].get()} - "
                             f"{self.campos_form['agencia'].get()} "
                             f"{self.campos_form['conta'].get()}").strip()

        # Preparar dados para salvar
        dados = {
            'tipo_pessoa': tipo_pessoa,
            'cnpj_cpf': cnpj_cpf,
            'razao_social': self.campos_form['razao_social'].get().upper(),
            'nome': self.campos_form['nome'].get().upper(),
            'telefone': self.campos_form['telefone'].get(),
            'email': self.campos_form['email'].get(),
            'banco': self.campos_form['banco'].get(),
            'op': self.campos_form['op'].get(),
            'agencia': self.campos_form['agencia'].get(),
            'conta': self.campos_form['conta'].get(),
            'chave_pix': self.campos_form['chave_pix'].get(),
            'categoria': self.campos_form['categoria'].get().upper(),
            'especificacao': self.campos_form['especificacao'].get().upper(),
            'vinculo': self.campos_form['vinculo'].get().upper(),
            'dados_bancarios': dados_bancarios
        }

        try:
            self.salvar_na_base_fornecedores(dados)
            custom_messagebox("info", "Sucesso", "Fornecedor salvo com sucesso!")
            self.janela_fornecedor.destroy()
            self.buscar_fornecedor()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar fornecedor: {str(e)}")

    def salvar_na_base_fornecedores(self, dados):
        """Salva os dados na planilha de fornecedores"""
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Coletar todos os dados existentes e o novo
            fornecedores = []
            
            # Converter dados existentes mantendo formato original da planilha
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tem CNPJ/CPF
                    fornecedor = {
                        'cnpj_cpf': row[0],
                        'tipo_pessoa': row[1],
                        'razao_social': row[2],
                        'nome': row[3],
                        'telefone': row[4],
                        'email': row[5],
                        'banco': row[6],
                        'op': row[7],
                        'agencia': row[8],
                        'conta': row[9],
                        'chave_pix': row[10],
                        'categoria': row[11],
                        'especificacao': row[12],
                        'vinculo': row[13],
                        'dados_bancarios': row[14]
                    }
                    fornecedores.append(fornecedor)
            
            # Adicionar novo fornecedor ou atualizar existente
            fornecedor_encontrado = False
            for i, fornecedor in enumerate(fornecedores):
                if fornecedor['cnpj_cpf'] == dados['cnpj_cpf']:
                    fornecedores[i] = dados.copy()
                    fornecedor_encontrado = True
                    break
            
            if not fornecedor_encontrado:
                fornecedores.append(dados.copy())
            
            # Ordenar por nome e CNPJ/CPF
            fornecedores_ordenados = sorted(
                fornecedores,
                key=lambda x: (x['nome'].upper(), x['cnpj_cpf'])
            )
            
            # Limpar planilha existente
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            
            # Reescrever dados ordenados
            for i, fornecedor in enumerate(fornecedores_ordenados, start=2):
                ws.cell(row=i, column=1, value=fornecedor['cnpj_cpf'])
                ws.cell(row=i, column=2, value=fornecedor['tipo_pessoa'])
                ws.cell(row=i, column=3, value=fornecedor['razao_social'])
                ws.cell(row=i, column=4, value=fornecedor['nome'])
                ws.cell(row=i, column=5, value=fornecedor['telefone'])
                ws.cell(row=i, column=6, value=fornecedor['email'])
                ws.cell(row=i, column=7, value=fornecedor['banco'])
                ws.cell(row=i, column=8, value=fornecedor['op'])
                ws.cell(row=i, column=9, value=fornecedor['agencia'])
                ws.cell(row=i, column=10, value=fornecedor['conta'])
                ws.cell(row=i, column=11, value=fornecedor['chave_pix'])
                ws.cell(row=i, column=12, value=fornecedor['categoria'])
                ws.cell(row=i, column=13, value=fornecedor['especificacao'])
                ws.cell(row=i, column=14, value=fornecedor['vinculo'])
                ws.cell(row=i, column=15, value=fornecedor['dados_bancarios'])
            
            wb.save(ARQUIVO_FORNECEDORES)
            
        except Exception as e:
            raise Exception(f"Erro ao salvar na planilha: {str(e)}")

    def atualizar_linha_fornecedor(self, row, dados):
        """Atualiza uma linha existente com novos dados"""
        row[0].value = dados['cnpj_cpf']
        row[1].value = dados['tipo_pessoa']  # Nova coluna para tipo de pessoa
        row[2].value = dados['razao_social']
        row[3].value = dados['nome']
        row[4].value = dados['telefone']
        row[5].value = dados['email']
        row[6].value = dados['banco']
        row[7].value = dados['op']
        row[8].value = dados['agencia']
        row[9].value = dados['conta']
        row[10].value = dados['chave_pix']
        row[11].value = dados['categoria']
        row[12].value = dados['especificacao']
        row[13].value = dados['vinculo']
        row[14].value = dados['dados_bancarios']

    def adicionar_linha_fornecedor(self, ws, linha, dados):
        """Adiciona uma nova linha com os dados do fornecedor"""
        ws.cell(row=linha, column=1, value=dados['cnpj_cpf'])
        ws.cell(row=linha, column=2, value=dados['tipo_pessoa'])
        ws.cell(row=linha, column=3, value=dados['razao_social'])
        ws.cell(row=linha, column=4, value=dados['nome'])
        ws.cell(row=linha, column=5, value=dados['telefone'])
        ws.cell(row=linha, column=6, value=dados['email'])
        ws.cell(row=linha, column=7, value=dados['banco'])
        ws.cell(row=linha, column=8, value=dados['op'])
        ws.cell(row=linha, column=9, value=dados['agencia'])
        ws.cell(row=linha, column=10, value=dados['conta'])
        ws.cell(row=linha, column=11, value=dados['chave_pix'])
        ws.cell(row=linha, column=12, value=dados['categoria'])
        ws.cell(row=linha, column=13, value=dados['especificacao'])
        ws.cell(row=linha, column=14, value=dados['vinculo'])
        ws.cell(row=linha, column=15, value=dados['dados_bancarios'])

        
    def atualizar_fornecedor(self):
        """Atualiza dados do fornecedor existente"""
        # Validações semelhantes ao salvar_fornecedor
        campos_obrigatorios = ['razao_social', 'nome', 'categoria']
        for campo in campos_obrigatorios:
            if not self.campos_form[campo].get().strip():
                custom_messagebox("error", "Erro", f"O campo {campo} é obrigatório!")
                return

        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            cnpj_cpf = self.campos_form['cnpj_cpf'].get()
            for row in ws.iter_rows(min_row=2):
                if row[0].value == cnpj_cpf:
                    # Atualizar dados na linha existente
                    row[1].value = self.campos_form['tipo_pessoa'].get().upper()
                    row[2].value = self.campos_form['razao_social'].get().upper()
                    row[3].value = self.campos_form['nome'].get().upper()
                    row[4].value = self.campos_form['telefone'].get()
                    row[5].value = self.campos_form['email'].get()
                    row[6].value = self.campos_form['banco'].get()
                    row[7].value = self.campos_form['op'].get()
                    row[8].value = self.campos_form['agencia'].get()
                    row[9].value = self.campos_form['conta'].get()
                    row[10].value = self.campos_form['chave_pix'].get()
                    row[11].value = self.campos_form['categoria'].get()
                    row[12].value = self.campos_form['especificacao'].get().upper()
                    row[13].value = self.campos_form['vinculo'].get().upper()
                    row[14].value = self.campos_form['dados_bancarios'].get().upper()
                    break

            wb.save(ARQUIVO_FORNECEDORES)
            custom_messagebox("info", "Sucesso", "Fornecedor atualizado com sucesso!")
            self.janela_fornecedor.destroy()
            self.buscar_fornecedor()  # Atualiza a lista
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao atualizar fornecedor: {str(e)}")


    def preencher_dados_fornecedor(self, dados):
        """Preenche os campos do fornecedor na aba de entrada"""
        self.campos_fornecedor['cnpj_cpf'].delete(0, tk.END)
        self.campos_fornecedor['cnpj_cpf'].insert(0, dados[0])
        
        self.campos_fornecedor['nome'].delete(0, tk.END)
        self.campos_fornecedor['nome'].insert(0, dados[1])
        
        self.campos_fornecedor['categoria'].delete(0, tk.END)
        self.campos_fornecedor['categoria'].insert(0, dados[2])

    def calcular_valor_total(self, event=None):
        """Calcula o valor total baseado no tipo de despesa"""
        try:
            # Pegar valor unitário
            vr_unit_str = self.campos_despesa['vr_unit'].get().strip()
            if not vr_unit_str:
                self.campos_despesa['valor'].config(state='normal')
                self.campos_despesa['valor'].delete(0, tk.END)
                self.campos_despesa['valor'].config(state='readonly')
                return
                
            vr_unit = float(vr_unit_str.replace(',', '.'))
            
            # Pegar tipo de despesa
            tp_desp = self.campos_despesa['tp_desp'].get()
            
            # Calcular com base no tipo
            if tp_desp == '1':  # Tipo que usa dias
                dias_str = self.campos_despesa['dias'].get().strip()
                dias = float(dias_str.replace(',', '.')) if dias_str else 1
                valor_total = vr_unit * dias
            else:
                valor_total = vr_unit
                
            # Atualizar campo de valor
            self.campos_despesa['valor'].config(state='normal')
            self.campos_despesa['valor'].delete(0, tk.END)
            self.campos_despesa['valor'].insert(0, f"{valor_total:.2f}")
            self.campos_despesa['valor'].config(state='readonly')
            
        except ValueError:
            # Em caso de erro, limpa o campo valor
            self.campos_despesa['valor'].config(state='normal')
            self.campos_despesa['valor'].delete(0, tk.END)
            self.campos_despesa['valor'].config(state='readonly')


    def verificar_tipo_despesa(self, event=None):
        """Verifica o tipo de despesa e ajusta campos conforme necessário"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        
        # Salvar a referência atual antes de qualquer modificação
        referencia_atual = ""
        if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
            referencia_atual = self.campos_despesa['referencia'].get()
        else:
            referencia_atual = self.campos_despesa['referencia'].get() 

        if not tp_desp.isdigit():
            self.campos_despesa['tp_desp'].delete(0, tk.END)
            return

        tp_desp_num = int(tp_desp)
        if not (1 <= tp_desp_num <= 6):
            self.campos_despesa['tp_desp'].delete(0, tk.END)
            return
            
        # Configura o campo dias
        if tp_desp == '1':
            self.campos_despesa['dias'].config(state='normal')
        else:
            self.campos_despesa['dias'].config(state='disabled')
            self.campos_despesa['dias'].delete(0, tk.END)
            self.campos_despesa['dias'].insert(0, '1')

        # Configura o campo nf
        if tp_desp != '1':
            self.campos_despesa['nf'].config(state='normal')
        else:
            self.campos_despesa['nf'].config(state='disabled')
            self.campos_despesa['nf'].delete(0, tk.END)
            
        # Atualiza o campo referência
        self.atualizar_campo_referencia(event)
        
        # NOVO: Restaurar o valor da referência se for uma especificação personalizada
        # e não uma das opções padrão do tipo 1
        if tp_desp != '1' and referencia_atual and referencia_atual not in self.opcoes_referencia_tipo1:
            if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                self.campos_despesa['referencia'].delete(0, tk.END)
                self.campos_despesa['referencia'].insert(0, referencia_atual)
            else:
                self.campos_despesa['referencia'].delete(0, tk.END)
                self.campos_despesa['referencia'].insert(0, referencia_atual)

        # Move para o campo referência
        self.campos_despesa['referencia'].focus()
        

    def adicionar_dados(self, eh_parcelamento=False):
        """Adiciona dados à lista temporária e retorna à aba fornecedor"""
        logger = system_logger.get_logger()
        logger.info(f"Iniciando adição de dados - Cliente: {self.cliente_atual}")

        if not self.validar_campos():
            logger.warning(f"Falha na validação dos campos")
            return False
        try:
            # Coleta do primeiro conjunto de dados
            vr_unit_str = self.campos_despesa['vr_unit'].get().strip()
            if not vr_unit_str:
                custom_messagebox("error", "Erro", "Valor unitário é obrigatório!")
                return
            vr_unit = float(vr_unit_str.replace(',', '.'))
        
            valor_str = self.campos_despesa['valor'].get().strip()
            if not valor_str:
                custom_messagebox("error", "Erro", "Valor total não foi calculado!")
                return
            valor = float(valor_str.replace(',', '.'))

            # Coletar dados do lançamento
            dados = {
                'data': self.data_rel_entry.get(),
                'cnpj_cpf': self.campos_fornecedor['cnpj_cpf'].get(),
                'nome': self.campos_fornecedor['nome'].get(),
                'categoria': self.campos_fornecedor['categoria'].get().upper(),
                'tp_desp': self.campos_despesa['tp_desp'].get(),
                'referencia': self.campos_despesa['referencia'].get().upper(),
                'nf': self.campos_despesa['nf'].get().upper(),
                'vr_unit': f"{vr_unit:.2f}",
                'dias': float(self.campos_despesa['dias'].get().replace(',', '.')) if self.campos_despesa['dias'].get() else 1,
                'valor': f"{valor:.2f}",
                'dt_vencto': self.campos_despesa['dt_vencto'].get(),
                'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get(),
                'observacao': self.campos_despesa['observacao'].get().upper(),
                'forma_pagamento': self.forma_pagamento_var.get()  # Adicionado campo forma_pagamento
            }
            self.dados_para_incluir.append(dados)

            # Verificar se é um lançamento de TRANSPORTE e criar lançamento automático de CAFÉ
            if dados['tp_desp'] == '1' and dados['referencia'] == 'TRANSPORTE':
                try:
                    # Buscar valor do café nas configurações
                    from src.configuracoes_sistema import GerenciadorConfiguracoes
                    config = GerenciadorConfiguracoes.carregar_configuracoes()
                    
                    if config and 'cafe' in config and 'valor_atual' in config['cafe']:
                        vr_unit_cafe = float(config['cafe']['valor_atual'])
                    else:
                        vr_unit_cafe = 4.0  # Valor padrão caso não encontre configuração
                        
                    dias_cafe = int(dados['dias'])
                    valor_cafe = vr_unit_cafe * dias_cafe
                    
                    # Criar dados do lançamento do CAFÉ
                    dados_cafe = dados.copy()
                    dados_cafe.update({
                        'referencia': 'CAFÉ',
                        'vr_unit': f"{vr_unit_cafe:.2f}",
                        'valor': f"{valor_cafe:.2f}"
                    })
                    self.dados_para_incluir.append(dados_cafe)
                    custom_messagebox("info", "Informação", 
                        f"Lançamento de CAFÉ adicionado automaticamente com valor de R$ {vr_unit_cafe:.2f} por dia!")
                except Exception as e:
                    custom_messagebox("warning",  "Aviso", 
                        f"Erro ao processar lançamento automático do café: {str(e)}\n"
                        "O lançamento principal foi salvo, mas o café não foi gerado.")

            # Só limpa os campos e mostra mensagem se não for parcelamento
            if not eh_parcelamento:
                self.limpar_campos_despesa()
                
                # Limpar campos do fornecedor
                for campo, entry in self.campos_fornecedor.items():
                    entry.config(state='normal')
                    entry.delete(0, tk.END)
                    if campo != 'categoria':
                        entry.config(state='readonly')
                
                custom_messagebox("info", "Sucesso", "Dados adicionados com sucesso!")
                
                # Voltar para a aba fornecedor
                self.notebook.select(1)
                self.tree_fornecedores.selection_remove(self.tree_fornecedores.selection())
                self.busca_entry.delete(0, tk.END)
            
            logger.info(f"Dados adicionados com sucesso - Cliente: {self.cliente_atual}, Total: {len(self.dados_para_incluir)}")
            return True
            
        except ValueError as e:
            logger.error(f"Erro ao processar valores: {str(e)}")
            custom_messagebox("error", "Erro", f"Erro ao processar valores: {str(e)}")
            return False
    
    def validar_campos(self):
        """Valida os campos antes de adicionar/enviar dados"""
        # Validar data
        if not self.data_rel_entry.get():
            custom_messagebox("error", "Erro", "Data de referência é obrigatória!")
            return False

        # Validar fornecedor
        if not self.campos_fornecedor['cnpj_cpf'].get():
            custom_messagebox("error", "Erro", "Selecione um fornecedor!")
            return False


        # Validar tipo de despesa
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        if not tp_desp or not tp_desp.isdigit() or not (1 <= int(tp_desp) <= 7):
            custom_messagebox("error", "Erro", "Tipo de despesa deve ser um número entre 1 e 7!")
            return False

        # Validar valor unitário
        vr_unit = self.campos_despesa['vr_unit'].get().strip()
        if not vr_unit:
            custom_messagebox("error", "Erro", "Valor unitário é obrigatório!")
            return False
        try:
            float(vr_unit.replace(',', '.'))
        except ValueError:
            custom_messagebox("error", "Erro", "Valor unitário inválido!")
            return False

        # Validar dias para tipo de despesa 1
        if tp_desp == '1':
            dias = self.campos_despesa['dias'].get().strip()
            if not dias:
                custom_messagebox("error", "Erro", "Quantidade de dias é obrigatória para tipo 1!")
                return False
            try:
                dias_float = float(dias.replace(',', '.'))
                if dias_float <= 0:
                    custom_messagebox("error", "Erro", "Quantidade de dias deve ser maior que zero!")
                    return False
            except ValueError:
                custom_messagebox("error", "Erro", "Quantidade de dias inválida!")
                return False

        # Validar referência
        if not self.campos_despesa['referencia'].get().strip():
            custom_messagebox("error", "Erro", "Referência é obrigatória!")
            return False

        # Validar data de vencimento
        if not self.campos_despesa['dt_vencto'].get():
            custom_messagebox("error", "Erro", "Data de vencimento é obrigatória!")
            return False

        return True

    
    def importar_folha_rh(self):
        """Inicia o processo de importação de dados da folha de RH"""
        # Verificar se um cliente está selecionado
        if not self.cliente_atual:
            if custom_messagebox("yesno", 
                "Importação RH",
                "Nenhum cliente está selecionado. A importação será feita baseada nos dados da planilha RH.\n\n"
                "Deseja continuar?"
            ):
                importador = ImportadorRH(self)
                importador.importar_dados()
        else:
            importador = ImportadorRH(self)
            importador.importar_dados()

    def limpar_campos_despesa(self):
        """Limpa os campos de despesa mantendo alguns valores padrão"""
        # Limpar todos os campos
        self.campos_despesa['tp_desp'].delete(0, tk.END)
        self.campos_despesa['referencia'].set('')  # Para Combobox
        self.campos_despesa['nf'].delete(0, tk.END)  # Novo campo
        self.campos_despesa['vr_unit'].delete(0, tk.END)
        self.campos_despesa['dias'].delete(0, tk.END)
        self.campos_despesa['dias'].insert(0, '1')  # Valor padrão
        self.campos_despesa['valor'].config(state='normal')
        self.campos_despesa['valor'].delete(0, tk.END)
        self.campos_despesa['valor'].config(state='readonly')
        self.campos_despesa['observacao'].delete(0, tk.END)
        
        # Definir data de vencimento igual à data de referência
        self.campos_despesa['dt_vencto'].set_date(self.data_rel_entry.get_date())

        # Resetar estado do campo referência
        self.campos_despesa['referencia'].config(state='normal')
        self.campos_despesa['referencia']['values'] = []

    def verificar_duplicidade_antes_salvar(self, sheet, dados):
        """
        Verifica se um lançamento similar já existe na planilha usando critérios mais precisos
        incluindo número da nota fiscal e data de vencimento.
        """
        logger = system_logger.get_logger()
        logger.debug(f"Verificando duplicidade para: {dados['nome']} - {dados['referencia']} - Valor: {dados['valor']}")
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Pular linhas vazias
            if not row[0]:
                continue

            # Verificar status - pular se excluído (coluna 14 = índice 13)
            status = row[13] if len(row) > 13 else 'ATIVO'
            if status == 'EXCLUIDO':
                continue
                
            # Verificar se os critérios principais correspondem
            try:
                valor_planilha = float(str(row[8]).replace(',', '.'))
                valor_novo = float(str(dados['valor']).replace(',', '.'))
                diferenca_valor = abs(valor_planilha - valor_novo)
                
                if (row[3] == dados['nome'] and                  # Nome do fornecedor
                    row[4] == dados['referencia'] and            # Referência 
                    diferenca_valor < 0.01):                     # Valor com tolerância
                    
                    # Critérios adicionais para distinguir lançamentos legítimos repetidos
                    nf_corresponde = True
                    dt_vencto_corresponde = True
                    
                    # Verificar NF se estiver presente no lançamento e na planilha
                    if 'nf' in dados and dados['nf'] and row[5]:
                        nf_corresponde = (str(dados['nf']).strip().upper() == str(row[5]).strip().upper())
                    
                    # Verificar data de vencimento
                    if 'dt_vencto' in dados and row[9]:
                        # Normalizar formato de data para comparação
                        if isinstance(row[9], datetime):
                            data_planilha = row[9].strftime('%d/%m/%Y')
                        else:
                            data_planilha = str(row[9])
                        
                        dt_vencto_corresponde = (str(dados['dt_vencto']).strip() == data_planilha.strip())
                    
                    # Se todos os critérios correspondem, consideramos um potencial duplicado
                    if nf_corresponde and dt_vencto_corresponde:
                        logger.warning(f"Duplicata detectada: {dados['nome']} - {dados['referencia']} - {dados['valor']} - NF: {dados.get('nf', 'N/A')} - Vencto: {dados.get('dt_vencto', 'N/A')}")
                        return True
            except Exception as e:
                logger.error(f"Erro ao verificar duplicidade: {str(e)}")
                # Continuar verificando outras linhas em caso de erro
        
        return False

    def enviar_dados(self):
        """Salva os dados na planilha existente do cliente"""
        logger = system_logger.get_logger()
        logger.info(f"Iniciando envio de dados - Cliente: {self.cliente_atual}, Registros: {len(self.dados_para_incluir) if self.dados_para_incluir else 0}")
        
        # Desabilitar botão para evitar múltiplos cliques - CORRIGIDO
        btn_enviar = None
        
        # Buscar o botão "Enviar" em todas as abas do notebook
        for aba in [self.aba_dados, self.aba_fornecedor]:
            for child in aba.winfo_children():
                # Verificar se é um frame 
                if isinstance(child, ttk.Frame):
                    # Procurar o botão dentro deste frame
                    for widget in child.winfo_children():
                        if isinstance(widget, ttk.Button) and widget['text'] == "Enviar":
                            btn_enviar = widget
                            break
        
        if btn_enviar:
            btn_enviar.config(state='disabled')
        
        try:
            # Verificar se há cliente selecionado
            if not self.cliente_atual:
                custom_messagebox("error", "Erro", "Selecione um cliente!")
                return
            
            # Verificar se existem dados para processar
            dados_para_processar = []
            if hasattr(self, 'visualizador') and self.visualizador and self.visualizador.tree.winfo_exists():
                logger.info("Obtendo dados do visualizador")
                dados_para_processar = self.visualizador.get_dados_atualizados()
            elif self.dados_para_incluir:
                logger.info("Usando dados_para_incluir existentes")
                dados_para_processar = self.dados_para_incluir.copy()
                    
            if not dados_para_processar:
                custom_messagebox("warning",  "Aviso", "Não há dados para enviar!")
                return

            logger.info(f"Total de registros a processar: {len(dados_para_processar)}")

            # Adicionar identificadores únicos para cada lançamento
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            for i, lancamento in enumerate(dados_para_processar):
                lancamento['id'] = f"{timestamp}-{i}"

            # Atualizar lista principal com dados mais recentes
            self.dados_para_incluir = dados_para_processar

            arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
            logger.info(f"Salvando em: {arquivo_cliente}")
            
            # Abrir workbook, verificar duplicidade e salvar os dados
            try:
                # Testar se o arquivo pode ser aberto para escrita
                with open(arquivo_cliente, 'a+b') as test_file:
                    pass  # Só tentamos abrir e fechar para testar acesso
            except PermissionError:
                logger.error(f"Permissão negada ao abrir arquivo: {arquivo_cliente}")
                custom_messagebox("error", 
                    "Erro", 
                    f"A planilha '{self.cliente_atual}.xlsx' está aberta!\n\n"
                    "Por favor:\n"
                    "1. Feche a planilha\n"
                    "2. Clique em OK\n"
                    "3. Tente enviar novamente"
                )
                self._is_saving = False
                return
            except Exception as e:
                logger.error(f"Erro ao verificar acesso ao arquivo: {str(e)}")
                self.custom_messagebox("error","Erro", f"Erro ao verificar acesso ao arquivo: {str(e)}")
                self._is_saving = False
                return
            
            # Agora tente abrir com openpyxl
            try:
                workbook = load_workbook(arquivo_cliente)
                sheet = workbook["Dados"]
                
                # Verificar duplicatas antes de salvar
                lancamentos_duplicados = []
                for dados in dados_para_processar:
                    if self.verificar_duplicidade_antes_salvar(sheet, dados):
                        lancamentos_duplicados.append(dados)

                # Inicializar variável para controle do fluxo
                continuar_operacao = True  # Por padrão, permite continuar

                if lancamentos_duplicados:
                    logger.warning(f"Detectados {len(lancamentos_duplicados)} possíveis lançamentos duplicados")
                    
                    # Criar mensagem com os duplicados
                    msg_duplicados = f"Foram detectados {len(lancamentos_duplicados)} possíveis lançamentos duplicados:\n\n"
                    for i, dados in enumerate(lancamentos_duplicados[:8], 1):  # Mostrar no máximo 3
                        msg_duplicados += f"{i}. {dados['nome']} - {dados['referencia']} - R$ {dados['valor']}\n"
                    
                    if len(lancamentos_duplicados) > 8:
                        msg_duplicados += f"... e mais {len(lancamentos_duplicados) - 8} lançamentos.\n"
                    
                    msg_duplicados += "\nDeseja continuar mesmo assim?"
                    
                    # Usar messagebox simples em vez de janela complexa
                    if not custom_messagebox("yesno", "Duplicatas Detectadas", msg_duplicados):
                        logger.info("Operação cancelada pelo usuário devido a duplicatas")
                        return
                    
                    logger.info("Usuário optou por continuar apesar das duplicatas")

                if sheet.tables:
                    table_name = list(sheet.tables.keys())[0]
                    sheet.tables.pop(table_name)
                    
                # Processar registros
                for dados in dados_para_processar:
                    proxima_linha = sheet.max_row + 1
                    
                    # Converter e salvar data de referência
                    data_rel = datetime.strptime(dados['data'], '%d/%m/%Y')
                    data_cell = sheet.cell(row=proxima_linha, column=1, value=data_rel)
                    data_cell.number_format = 'DD/MM/YYYY'

                    # Converter tipo de despesa para número
                    tp_desp_cell = sheet.cell(row=proxima_linha, column=2, value=int(dados['tp_desp']))
                    tp_desp_cell.number_format = '0'

                    sheet.cell(row=proxima_linha, column=3, value=dados['cnpj_cpf'])
                    sheet.cell(row=proxima_linha, column=4, value=dados['nome'])
                    sheet.cell(row=proxima_linha, column=5, value=dados['referencia'])
                    sheet.cell(row=proxima_linha, column=6, value=dados['nf'])

                    # No método enviar_dados
                    vr_unit = float(dados['vr_unit'].replace(',', '.'))
                    vr_unit_cell = sheet.cell(row=proxima_linha, column=7, value=vr_unit)
                    aplicar_formatacao_celula(vr_unit_cell)

                    sheet.cell(row=proxima_linha, column=8, value=int(dados.get('dias', 1)))

                    valor = float(dados['valor'].replace(',', '.'))
                    valor_cell = sheet.cell(row=proxima_linha, column=9, value=valor)
                    aplicar_formatacao_celula(valor_cell)

                    dt_vencto = datetime.strptime(dados['dt_vencto'], '%d/%m/%Y')
                    dt_vencto_cell = sheet.cell(row=proxima_linha, column=10, value=dt_vencto)
                    dt_vencto_cell.number_format = 'DD/MM/YYYY'

                    sheet.cell(row=proxima_linha, column=11, value=dados['categoria'])
                    sheet.cell(row=proxima_linha, column=12, value=dados['dados_bancarios'])
                    sheet.cell(row=proxima_linha, column=13, value=dados['observacao'])
                    
                    # Adicionar as novas colunas
                    sheet.cell(row=proxima_linha, column=14, value='ATIVO')  # STATUS
                    sheet.cell(row=proxima_linha, column=15, value=int(proxima_linha - 1))  # ID_LANCAMENTO

                # Salvar o arquivo com tratamento de erro aprimorado
                try:
                    # Tente salvar com tratamento explícito para PermissionError
                    workbook.save(arquivo_cliente)
                    custom_messagebox("info", "Sucesso", "Dados salvos com sucesso!")
                    
                    # Após salvar com sucesso
                    logger.info(f"Dados salvos com sucesso - Cliente: {self.cliente_atual}, Registros: {len(dados_para_processar)}")
                    
                    # Limpar dados após salvar com sucesso
                    self.dados_para_incluir.clear()
                    if hasattr(self, 'visualizador') and self.visualizador:
                        self.visualizador.janela.destroy()
                        self.visualizador = None
                    
                except PermissionError:
                    logger.error("Permissão negada ao salvar arquivo - provável arquivo aberto")
                    custom_messagebox("error", 
                        "Erro", 
                        f"Não foi possível salvar! A planilha '{self.cliente_atual}.xlsx' está aberta.\n\n"
                        "Por favor:\n"
                        "1. Feche a planilha\n"
                        "2. Clique em OK\n"
                        "3. Tente enviar novamente"
                    )
                except Exception as e:
                    logger.error(f"Erro ao salvar arquivo: {str(e)}")
                    custom_messagebox("error", "Erro", f"Erro ao salvar arquivo: {str(e)}")
            
            except Exception as e:
                logger.error(f"Erro ao processar dados: {str(e)}", exc_info=True)
                custom_messagebox("error", "Erro", f"Erro ao processar dados: {str(e)}")
            
        except Exception as e:
            logger.error(f"Erro geral no método enviar_dados: {str(e)}", exc_info=True)
            custom_messagebox("error", "Erro", f"Erro ao enviar dados: {str(e)}")
            
        finally:
            # Desmarcar flag de processamento
            self._is_saving = False

class EditorCliente:
    def __init__(self, parent): 
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("Editor de Clientes")
        self.root.geometry("800x600")
        
        self.setup_gui()
        self.carregar_clientes()

    def setup_gui(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Lista de clientes
        frame_clientes = ttk.LabelFrame(main_frame, text="Clientes")
        frame_clientes.pack(fill='both', expand=True, pady=5)

        self.tree_clientes = ttk.Treeview(frame_clientes, 
                                        columns=('Nome', 'Endereço', 'Taxa ADM'),
                                        show='headings')
        self.tree_clientes.heading('Nome', text='Nome')
        self.tree_clientes.heading('Endereço', text='Endereço')
        self.tree_clientes.heading('Taxa ADM', text='Taxa ADM (%)')
        self.tree_clientes.pack(fill='both', expand=True, padx=5, pady=5)

        # Frame para edição
        frame_edicao = ttk.LabelFrame(main_frame, text="Edição")
        frame_edicao.pack(fill='x', pady=5)

        ttk.Label(frame_edicao, text="Taxa de Administração (%):").pack(side='left', padx=5)
        self.taxa_entry = ttk.Entry(frame_edicao, width=10)
        self.taxa_entry.pack(side='left', padx=5)

        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(frame_botoes, 
                  text="Atualizar Taxa", 
                  command=self.atualizar_taxa).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Remover Taxa", 
                  command=self.remover_taxa).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Fechar", 
                  command=self.root.destroy).pack(side='right', padx=5)

    def carregar_clientes(self):
        """Carrega a lista de clientes do arquivo Excel"""
        try:
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            # Limpar lista atual
            for item in self.tree_clientes.get_children():
                self.tree_clientes.delete(item)
            
            # Adicionar clientes
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome não vazio
                    self.tree_clientes.insert('', 'end', values=(
                        row[0],  # Nome
                        row[1],  # Endereço
                        row[6] if row[6] else "0.00"  # Taxa ADM
                    ))
            
            wb.close()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar clientes: {str(e)}")

    def atualizar_taxa(self):
        """Atualiza a taxa de administração do cliente selecionado"""
        selecionado = self.tree_clientes.selection()
        if not selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um cliente")
            return

        try:
            taxa = float(self.taxa_entry.get().replace(',', '.'))
            if not (0 <= taxa <= 100):
                custom_messagebox("error", "Erro", "Taxa deve estar entre 0 e 100")
                return
                
            cliente = self.tree_clientes.item(selecionado)['values'][0]
            
            # Atualizar no arquivo
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == cliente:
                    row[6].value = taxa  # Coluna da taxa de administração
                    
            wb.save(ARQUIVO_CLIENTES)
            
            # Atualizar na treeview
            self.tree_clientes.set(selecionado, 'Taxa ADM', f"{taxa:.2f}")
            custom_messagebox("info", "Sucesso", "Taxa atualizada com sucesso!")
            
        except ValueError:
            custom_messagebox("error", "Erro", "Taxa inválida")
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao atualizar taxa: {str(e)}")

    def remover_taxa(self):
        """Remove a taxa de administração do cliente selecionado"""
        selecionado = self.tree_clientes.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um cliente")
            return

        if custom_messagebox("yesno", "Confirmar", "Deseja remover a taxa de administração?"):
            try:
                cliente = self.tree_clientes.item(selecionado)['values'][0]
                
                # Atualizar no arquivo
                wb = load_workbook(ARQUIVO_CLIENTES)
                ws = wb['Clientes']
                
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == cliente:
                        row[6].value = None  # Remover taxa
                        
                wb.save(ARQUIVO_CLIENTES)
                
                # Atualizar na treeview
                self.tree_clientes.set(selecionado, 'Taxa ADM', "0.00")
                custom_messagebox("info", "Sucesso", "Taxa removida com sucesso!")
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao remover taxa: {str(e)}")



class GestaoContratos:
    def __init__(self, parent):
        self.parent = parent
        self.arquivo_cliente = None
        self.cliente_atual = None

    def centralizar_janela(self, janela, largura=800, altura=600, parent=None):
        """
        Centraliza uma janela na tela ou relativa ao parent se fornecido.
        Também define o tamanho padrão da janela.
        """
        # Definir tamanho
        janela.geometry(f"{largura}x{altura}")
        
        # Atualizar a janela para garantir que as dimensões sejam aplicadas
        janela.update_idletasks()
        
        # Se parent for fornecido, centralize em relação a ele
        if parent and parent.winfo_exists():
            # Calcular o centro da janela pai
            x_parent = parent.winfo_x() + parent.winfo_width() // 2
            y_parent = parent.winfo_y() + parent.winfo_height() // 2
            
            # Calcular a posição da nova janela
            x = x_parent - largura // 2
            y = y_parent - altura // 2
        else:
            # Centralizar na tela
            x = (janela.winfo_screenwidth() // 2) - (largura // 2)
            y = (janela.winfo_screenheight() // 2) - (altura // 2)
        
        # Definir posição
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        
        # Tornar a janela modal (quando aplicável)
        if parent and hasattr(janela, 'transient') and hasattr(janela, 'grab_set'):
            janela.transient(parent)
            janela.grab_set()
            
        # Trazer para frente
        janela.lift()
        janela.focus_force()

    def criar_interface_contratos(self, janela, on_close_callback):
        """Cria a interface de gestão de contratos em uma janela já existente"""
        try:
            # Verificar se o arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                custom_messagebox("error", "Erro", f"Arquivo do cliente {self.cliente_atual} não encontrado!")
                on_close_callback()  # Fechar a janela em caso de erro
                return

            # Abrir arquivo e verificar aba
            wb = load_workbook(self.arquivo_cliente)
            if 'Contratos_ADM' not in wb.sheetnames:
                # Se não existir a aba, criar
                print(f"Criando aba Contratos_ADM para {self.cliente_atual}")
                ws = wb.create_sheet("Contratos_ADM")
                
                # Definir os blocos na linha 1
                blocos = ["CONTRATOS", "", "", "", "", "",
                        "ADMINISTRADORES_CONTRATO", "", "", "", "", "", "",
                        "ADITIVOS", "", "", "",
                        "ADMINISTRADORES_ADITIVO", "", "", "", "", "", "",
                        "PARCELAS", "", "", "", "", "", "", "", ""]
                
                for col, valor in enumerate(blocos, 1):
                    ws.cell(row=1, column=col, value=valor)
                
                # Definir cabeçalhos na linha 2
                headers = [
                    # CONTRATOS
                    "Nº Contrato", "Data Início", "Data Fim", "Status", "Observações", "",
                    # ADMINISTRADORES_CONTRATO
                    "Nº Contrato", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total", "Nº Parcelas", 
                    # ADITIVOS
                    "Nº Contrato", "Nº Aditivo", "Data Início", "Data Fim",
                    # ADMINISTRADORES_ADITIVO
                    "Nº Contrato", "Nº Aditivo", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total",
                    # PARCELAS
                    "Referência", "Número", "CNPJ/CPF", "Nome", "Data Vencimento", "Valor", "Status", "Data Pagamento", "Eventos/Fases", "Percentual %"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                # Salvar as alterações
                wb.save(self.arquivo_cliente)

            # Frame principal
            frame_principal = ttk.Frame(janela, padding="10")
            frame_principal.pack(fill='both', expand=True)

            # Frame para lista de contratos existentes
            frame_contratos = ttk.LabelFrame(frame_principal, text="Contratos Existentes")
            frame_contratos.pack(fill='both', expand=True, pady=5)

            # Treeview para contratos
            colunas = ('Nº Contrato', 'Data Início', 'Data Fim', 'Status')
            self.tree_contratos = ttk.Treeview(frame_contratos, columns=colunas, show='headings')
            for col in colunas:
                self.tree_contratos.heading(col, text=col)
                self.tree_contratos.column(col, width=100)
            
            # Adicionar scrollbars
            scroll_y = ttk.Scrollbar(frame_contratos, orient='vertical', command=self.tree_contratos.yview)
            scroll_x = ttk.Scrollbar(frame_contratos, orient='horizontal', command=self.tree_contratos.xview)
            self.tree_contratos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            
            self.tree_contratos.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y.pack(side='right', fill='y')
            scroll_x.pack(side='bottom', fill='x')

            # Frame para lista de administradores do contrato selecionado
            frame_admins = ttk.LabelFrame(frame_principal, text="Administradores do Contrato")
            frame_admins.pack(fill='both', expand=True, pady=5)

            # Treeview para administradores
            colunas_adm = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial Pagamento')
            self.tree_adm_contrato = ttk.Treeview(frame_admins, columns=colunas_adm, show='headings')
            for col in colunas_adm:
                self.tree_adm_contrato.heading(col, text=col)
                self.tree_adm_contrato.column(col, width=100)
            
            # Adicionar scrollbars para administradores
            scroll_y_adm = ttk.Scrollbar(frame_admins, orient='vertical', command=self.tree_adm_contrato.yview)
            scroll_x_adm = ttk.Scrollbar(frame_admins, orient='horizontal', command=self.tree_adm_contrato.xview)
            self.tree_adm_contrato.configure(yscrollcommand=scroll_y_adm.set, xscrollcommand=scroll_x_adm.set)
            
            self.tree_adm_contrato.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y_adm.pack(side='right', fill='y')
            scroll_x_adm.pack(side='bottom', fill='x')

            # Botões de ação
            frame_botoes = ttk.Frame(frame_principal)
            frame_botoes.pack(fill='x', pady=5)

            ttk.Button(frame_botoes, text="Novo Contrato", 
                    command=lambda: self.criar_novo_contrato(janela)).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Editar Contrato", 
                    command=self.editar_contrato).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Excluir Contrato", 
                    command=self.excluir_contrato).pack(side='left', padx=5)
            
            # Botão Fechar com callback personalizado
            ttk.Button(frame_botoes, text="Fechar", 
                    command=on_close_callback).pack(side='right', padx=5)

            # Carregar contratos existentes
            self.carregar_contratos()

            # Binding para atualizar administradores quando selecionar contrato
            self.tree_contratos.bind('<<TreeviewSelect>>', self.mostrar_administradores)

        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao abrir janela de contratos: {str(e)}")
            if 'wb' in locals():
                wb.close()
            # Garantir que a janela principal seja restaurada em caso de erro
            on_close_callback()

    
    def carregar_contratos(self):
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            for item in self.tree_contratos.get_children():
                self.tree_contratos.delete(item)
            
            contratos_processados = set()
            for row in ws.iter_rows(min_row=3, values_only=True):
                num_contrato = row[0]
                if num_contrato and num_contrato not in contratos_processados:
                    # Processar datas
                    data_inicio = ''
                    if row[1]:
                        try:
                            if isinstance(row[1], datetime):
                                data_inicio = row[1].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[1]), '%Y-%m-%d')
                                data_inicio = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_inicio = str(row[1])

                    data_fim = ''
                    if row[2]:
                        try:
                            if isinstance(row[2], datetime):
                                data_fim = row[2].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[2]), '%Y-%m-%d')
                                data_fim = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_fim = str(row[2])

                    
                    self.tree_contratos.insert('', 'end', values=(
                        num_contrato,
                        data_inicio,
                        data_fim,
                        row[3] or ''
                    ))
                    contratos_processados.add(num_contrato)
            
            wb.close()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar contratos: {str(e)}")

    def mostrar_administradores(self, event=None):
        """Mostra administradores do contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            return
            
        try:
            # Limpar lista atual
            for item in self.tree_adm_contrato.get_children():
                self.tree_adm_contrato.delete(item)
                
            num_contrato = self.tree_contratos.item(selecionado)['values'][0]
            
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[6] == num_contrato:  # Coluna G - Nº Contrato
                    if row[26]:  # Data Inicial de Pagamento
                        data_inicial = row[26].strftime('%d/%m/%Y') if isinstance(row[26], datetime) else str(row[26])
                    else:
                        data_inicial = ''
                        
                    self.tree_adm_contrato.insert('', 'end', values=(
                        row[7],   # CNPJ/CPF
                        row[8],   # Nome
                        row[9],   # Tipo
                        row[10],  # Valor/Percentual
                        row[11],  # Valor Total
                        row[12],  # Nº Parcelas
                        data_inicial  # Data Inicial de Pagamento
                    ))
            
            wb.close()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar administradores: {str(e)}")
  

    def criar_novo_contrato(self, janela_principal):
        """Abre janela para criar novo contrato com suporte a parcelas fixas ou eventos"""
        janela = tk.Toplevel(self.parent)
        janela.title(f"Novo Contrato - {self.cliente_atual}")
        janela.geometry("800x650")
        
        # Frame principal com scrollbar para garantir acesso a todos os campos
        main_frame = ttk.Frame(janela)
        main_frame.pack(fill='both', expand=True)
        
        # Adicionar canvas com scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)
        
        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Frame principal dentro do scrollable frame
        frame = ttk.Frame(scroll_frame, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame para dados do contrato
        frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
        frame_contrato.pack(fill='x', pady=5)

        # Número do Contrato
        ttk.Label(frame_contrato, text="Nº Contrato:*", width=15).grid(row=0, column=0, padx=5, pady=5, sticky='w')
        num_contrato = ttk.Entry(frame_contrato, width=20)
        num_contrato.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Datas 
        ttk.Label(frame_contrato, text="Data Início:*", width=15).grid(row=1, column=0, padx=5, pady=5, sticky='w')
        data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_inicio.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame_contrato, text="Data Fim:*", width=15).grid(row=2, column=0, padx=5, pady=5, sticky='w')
        data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_fim.grid(row=2, column=1, padx=5, pady=5, sticky='w')

        # Observações
        ttk.Label(frame_contrato, text="Observações:", width=15).grid(row=3, column=0, padx=5, pady=5, sticky='nw')
        observacoes = ttk.Entry(frame_contrato, width=25)
        observacoes.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        
        # Adicionar campo para valor global do contrato
        ttk.Label(frame_contrato, text="Valor Global:*", width=15).grid(row=4, column=0, padx=5, pady=5, sticky='w')
        valor_global = ttk.Entry(frame_contrato, width=20)
        valor_global.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        
        # Tipo de pagamento (metodo)
        ttk.Label(frame_contrato, text="Método de Pagamento:*", width=22).grid(row=5, column=0, padx=5, pady=5, sticky='w')
        metodo_pagamento = ttk.Combobox(frame_contrato, values=[
            "Percentual da Quinzena", 
            "Valor Fixo em Parcelas", 
            "Eventos/Fases"
        ], state='readonly', width=20)
        metodo_pagamento.grid(row=5, column=1, padx=5, pady=5, sticky='w')
        metodo_pagamento.current(0)  # Valor padrão
        
        # Frame para Administradores
        frame_adm = ttk.LabelFrame(frame, text="Administradores")
        frame_adm.pack(fill='both', expand=True, pady=5)

        # Lista de Administradores
        colunas = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial')
        self.tree_adm = ttk.Treeview(frame_adm, columns=colunas, show='headings', height=5)
        
        for col in colunas:
            self.tree_adm.heading(col, text=col)
            self.tree_adm.column(col, width=100)
        
        # Adicionar scrollbars
        scroll_y = ttk.Scrollbar(frame_adm, orient='vertical', command=self.tree_adm.yview)
        scroll_x = ttk.Scrollbar(frame_adm, orient='horizontal', command=self.tree_adm.xview)
        self.tree_adm.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_adm.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')
        
        # Botões para administradores
        frame_botoes_adm = ttk.Frame(frame_adm)
        frame_botoes_adm.pack(fill='x', pady=5)

        # Botões para administradores - explicitamente configurados
        ttk.Button(
            frame_botoes_adm, 
            text="Adicionar Administrador",
            command=lambda: self.adicionar_administrador_modificado(self.tree_adm, valor_global, metodo_pagamento)
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_adm, 
            text="Remover Administrador",
            command=lambda: self.remover_administrador(self.tree_adm)
        ).pack(side='left', padx=5)

        def salvar():
            # Validar campos obrigatórios
            if not num_contrato.get() or not data_inicio.get() or not data_fim.get() or not valor_global.get():
                custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios do contrato!")
                return
                
            # Validar valor global
            try:
                valor_global_float = float(valor_global.get().replace(',', '.'))
                if valor_global_float <= 0:
                    custom_messagebox("error", "Erro", "Valor global deve ser maior que zero!")
                    return
            except ValueError:
                custom_messagebox("error", "Erro", "Valor global inválido!")
                return
                
            # Validar administradores
            if not self.tree_adm.get_children():
                custom_messagebox("error", "Erro", "Adicione pelo menos um administrador!")
                return
                
            # Criar contrato
            self.salvar_contrato_com_opcoes(
                num_contrato.get(),
                data_inicio.get_date(),
                data_fim.get_date(),
                observacoes.get(),
                valor_global_float,
                metodo_pagamento.get(),
                {},  # Opções simplificadas pois foram movidas para o administrador
                janela
            )
            
            janela_principal.focus_set()
            self.carregar_contratos()

        ttk.Button(frame, text="Salvar", command=salvar).pack(side='left', padx=5, pady=10)
        ttk.Button(frame, text="Cancelar", command=janela.destroy).pack(side='left', padx=5, pady=10)                          

    
    def processar_eventos(self, ws, num_contrato, valor_global, eventos):
        """Processa os eventos do contrato e cria parcelas vinculadas"""
        for i, (descricao, percentual, valor_evento) in enumerate(eventos, 1):
            # Para cada administrador, criar um registro de parcela vinculada ao evento
            for item in self.tree_adm.get_children():
                valores_adm = self.tree_adm.item(item)['values']
                cnpj_cpf_adm = str(valores_adm[0]).strip()
                cnpj_cpf_adm = formatar_cnpj_cpf(cnpj_cpf_adm)
                nome_adm = valores_adm[1]
                
                # Calcular valor para este administrador (proporcional ao percentual definido)
                if valores_adm[2] == 'Percentual':
                    perc_adm = float(str(valores_adm[3]).replace('%', '').replace(',', '.'))
                    valor_admin_evento = (perc_adm / 100) * valor_evento
                else:  # Fixo
                    # Distribuir o valor total entre os eventos conforme percentuais
                    valor_total_adm = float(str(valores_adm[4]).replace('.', '').replace(',', '.'))
                    valor_admin_evento = (percentual / 100) * valor_total_adm
                
                # Registrar parcela vinculada ao evento, combinando as informações de evento e parcela
                proxima_linha = ws.max_row + 1
                ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                ws.cell(row=proxima_linha, column=26, value=i)  # Número do evento como número da parcela
                ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                ws.cell(row=proxima_linha, column=29, value=None)  # Data vencimento (vazio)
                ws.cell(row=proxima_linha, column=30, value=valor_admin_evento)  # Valor
                ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                ws.cell(row=proxima_linha, column=32, value=i)  # ID do evento vinculado
                ws.cell(row=proxima_linha, column=33, value=descricao.upper())  # Descrição do evento
                ws.cell(row=proxima_linha, column=34, value=f"{percentual:.2f}%")  # Percentual do evento
                
        
    def editar_contrato(self):
        """Edita o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um contrato para editar")
            return

        try:
            dados_contrato = self.tree_contratos.item(selecionado)['values']
            
            janela = tk.Toplevel(self.parent)
            janela.title(f"Editar Contrato - {self.cliente_atual}")
            janela.geometry("600x500")

            # Frame principal
            frame = ttk.Frame(janela, padding="10")
            frame.pack(fill='both', expand=True)

            # Dados do Contrato
            frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
            frame_contrato.pack(fill='x', pady=5)

            # Número do Contrato (readonly)
            ttk.Label(frame_contrato, text="Nº Contrato:").grid(row=0, column=0, padx=5, pady=2)
            num_contrato = ttk.Entry(frame_contrato, state='readonly')
            num_contrato.grid(row=0, column=1, padx=5, pady=2)
            num_contrato.insert(0, dados_contrato[0])

            # Datas
            ttk.Label(frame_contrato, text="Data Início:").grid(row=1, column=0, padx=5, pady=2)
            data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_inicio.grid(row=1, column=1, padx=5, pady=2)
            data_inicio.set_date(datetime.strptime(dados_contrato[1], '%d/%m/%Y'))

            ttk.Label(frame_contrato, text="Data Fim:").grid(row=2, column=0, padx=5, pady=2)
            data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_fim.grid(row=2, column=1, padx=5, pady=2)
            data_fim.set_date(datetime.strptime(dados_contrato[2], '%d/%m/%Y'))

            # Status
            ttk.Label(frame_contrato, text="Status:").grid(row=3, column=0, padx=5, pady=2)
            status_combo = ttk.Combobox(frame_contrato, values=['ATIVO', 'INATIVO'], state='readonly')
            status_combo.grid(row=3, column=1, padx=5, pady=2)
            status_combo.set(dados_contrato[3])

            def salvar_alteracoes():
                try:
                    wb = load_workbook(self.arquivo_cliente)
                    ws = wb['Contratos_ADM']
                    
                    # Atualizar dados do contrato
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value == dados_contrato[0]:
                            row[1].value = data_inicio.get_date()
                            row[2].value = data_fim.get_date()
                            row[3].value = status_combo.get()
                    
                    wb.save(self.arquivo_cliente)
                    custom_messagebox("info", "Sucesso", "Contrato atualizado com sucesso!")
                    janela.destroy()
                    self.carregar_contratos()
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao salvar alterações: {str(e)}")

            # Botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)

            ttk.Button(frame_botoes, text="Salvar", command=salvar_alteracoes).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao abrir edição: {str(e)}")



    def adicionar_administrador_modificado(self, tree, valor_global_entry, metodo_pagamento_combo):
        """Versão modificada para incluir os detalhes de parcelas/eventos na tela do administrador"""
        # Verificar se valor global foi informado
        if not valor_global_entry.get():
            custom_messagebox("error", "Erro", "Informe o valor global do contrato primeiro")
            return
            
        try:
            valor_global_float = float(valor_global_entry.get().replace(',', '.'))
            if valor_global_float <= 0:
                custom_messagebox("error", "Erro", "Valor global deve ser maior que zero")
                return
        except ValueError:
            custom_messagebox("error", "Erro", "Valor global inválido")
            return
            
        # Obter o método de pagamento selecionado
        metodo = metodo_pagamento_combo.get()
        
        # Chamar método para abrir janela de administrador
        janela_admin = tk.Toplevel(self.parent)
        janela_admin.title("Adicionar Administrador")
        
        # Ajustar tamanho baseado no método (maior para eventos)
        if metodo == "Eventos/Fases":
            janela_admin.geometry("800x700")
        else:
            janela_admin.geometry("600x650")
        
        # Frame principal com scrollbar para permitir mais conteúdo
        main_frame = ttk.Frame(janela_admin)
        main_frame.pack(fill='both', expand=True)
        
        # Adicionar canvas com scrollbar se for Eventos/Fases
        if metodo == "Eventos/Fases":
            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            frame_admin = ttk.Frame(canvas)
            
            frame_admin.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=frame_admin, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
        else:
            frame_admin = ttk.Frame(main_frame, padding="10")
            frame_admin.pack(fill='both', expand=True)
        
        # Frame de busca
        frame_busca = ttk.LabelFrame(frame_admin, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        busca_entry = ttk.Entry(frame_busca, width=40)
        busca_entry.pack(side='left', padx=5)
        
        # Lista de fornecedores
        frame_lista = ttk.LabelFrame(frame_admin, text="Fornecedores")
        frame_lista.pack(fill='x', padx=5, pady=5)
        
        tree_fornecedores = ttk.Treeview(frame_lista,
                                    columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                    show='headings',
                                    height=3)
        tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        tree_fornecedores.heading('Nome', text='Nome')
        tree_fornecedores.heading('Categoria', text='Categoria')
        tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para dados do administrador
        frame_dados = ttk.LabelFrame(frame_admin, text="Dados do Administrador")
        frame_dados.pack(fill='x', padx=5, pady=5)
        
        # CNPJ/CPF
        ttk.Label(frame_dados, text="CNPJ/CPF:*").grid(row=0, column=0, padx=5, pady=2)
        cnpj_cpf_entry = ttk.Entry(frame_dados, state='readonly')
        cnpj_cpf_entry.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        
        # Nome
        ttk.Label(frame_dados, text="Nome/Razão Social:*").grid(row=1, column=0, padx=5, pady=2)
        nome_entry = ttk.Entry(frame_dados, state='readonly')
        nome_entry.grid(row=1, column=1, padx=5, pady=2, sticky='ew')
        
        # Mostrar informações do contrato (somente leitura)
        ttk.Label(frame_dados, text="Valor Global:").grid(row=2, column=0, padx=5, pady=2)
        valor_global_label = ttk.Label(frame_dados, text=f"R$ {valor_global_float:,.2f}")
        valor_global_label.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_dados, text="Método de Pagamento:").grid(row=3, column=0, padx=5, pady=2)
        metodo_label = ttk.Label(frame_dados, text=metodo)
        metodo_label.grid(row=3, column=1, padx=5, pady=2, sticky='w')
        
        # Tipo de remuneração
        ttk.Label(frame_dados, text="Tipo de Remuneração:*").grid(row=4, column=0, padx=5, pady=2)
        
        if metodo == "Percentual da Quinzena":
            tipo_combo = ttk.Combobox(frame_dados, values=['Percentual'], state='readonly')
            tipo_combo.grid(row=4, column=1, padx=5, pady=2, sticky='w')
            tipo_combo.set('Percentual')
            
            # Percentual
            ttk.Label(frame_dados, text="Percentual (%):*").grid(row=5, column=0, padx=5, pady=2)
            percentual_entry = ttk.Entry(frame_dados)
            percentual_entry.grid(row=5, column=1, padx=5, pady=2, sticky='w')
            
        else:  # Valor Fixo em Parcelas ou Eventos/Fases
            tipo_valores = ['Percentual', 'Fixo']
            tipo_combo = ttk.Combobox(frame_dados, values=tipo_valores, state='readonly')
            tipo_combo.grid(row=4, column=1, padx=5, pady=2)
            tipo_combo.set('Fixo')  # Padrão para eventos/parcelas fixas
            
            # Frame para valores percentuais
            frame_percentual_admin = ttk.Frame(frame_dados)
            frame_percentual_admin.grid(row=5, column=0, columnspan=2, pady=5)
            
            # Percentual
            ttk.Label(frame_percentual_admin, text="Percentual do Contrato (%):*").grid(row=0, column=0, padx=5, pady=2)
            percentual_entry = ttk.Entry(frame_percentual_admin)
            percentual_entry.grid(row=0, column=1, padx=5, pady=2)
            
            # Frame para valores fixos
            frame_fixo = ttk.Frame(frame_dados)
            frame_fixo.grid(row=6, column=0, columnspan=2, pady=5)
            
            # Valor Total
            ttk.Label(frame_fixo, text="Valor Total:*").grid(row=0, column=0, padx=5, pady=2)
            valor_total_entry = ttk.Entry(frame_fixo)
            valor_total_entry.grid(row=0, column=1, padx=5, pady=2)
            
            def atualizar_campos_tipo(*args):
                """Atualiza campos baseado no tipo selecionado"""
                if tipo_combo.get() == 'Percentual':
                    frame_percentual_admin.grid()
                    frame_fixo.grid_remove()
                elif tipo_combo.get() == 'Fixo':
                    frame_percentual_admin.grid_remove()
                    frame_fixo.grid()
                
            # Configurar evento
            tipo_combo.bind('<<ComboboxSelected>>', atualizar_campos_tipo)
            
            # Configurar interface inicial
            atualizar_campos_tipo()
        
        # Forma de pagamento para dados bancários
        ttk.Label(frame_dados, text="Forma de Pagamento:").grid(row=7, column=0, padx=5, pady=2)
        forma_pagamento = ttk.Combobox(frame_dados, values=['PIX', 'TED'], state='readonly')
        forma_pagamento.grid(row=7, column=1, padx=5, pady=2)
        forma_pagamento.set('PIX')  # Valor padrão
        
        # Área para configurações específicas do método de pagamento
        # ===============================================================
        # Frame para configurações específicas de método de pagamento
        frame_config_metodo = ttk.LabelFrame(frame_admin, text="Configuração de Pagamento")
        
        if metodo in ["Valor Fixo em Parcelas", "Eventos/Fases"]:
            frame_config_metodo.pack(fill='x', padx=5, pady=5, after=frame_dados)
        
        # 1. Frame para Parcelas Fixas
        if metodo == "Valor Fixo em Parcelas":
            frame_parcelas = ttk.Frame(frame_config_metodo)
            frame_parcelas.pack(fill='x', padx=5, pady=5)
            
            # Número de parcelas
            ttk.Label(frame_parcelas, text="Número de Parcelas:*").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            num_parcelas_entry = ttk.Entry(frame_parcelas, width=10)
            num_parcelas_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            
            # Checkbox para entrada
            var_tem_entrada = tk.BooleanVar(value=False)
            check_entrada = ttk.Checkbutton(frame_parcelas, text="Possui entrada?", variable=var_tem_entrada)
            check_entrada.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky='w')
            
            # Frame para entrada
            frame_entrada = ttk.Frame(frame_parcelas)
            frame_entrada.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky='w')
            
            ttk.Label(frame_entrada, text="Valor da Entrada:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            valor_entrada_entry = ttk.Entry(frame_entrada, width=15)
            valor_entrada_entry.grid(row=0, column=1, padx=5, pady=2, sticky='w')
            
            ttk.Label(frame_entrada, text="Data da Entrada:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
            data_entrada = DateEntry(frame_entrada, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_entrada.grid(row=1, column=1, padx=5, pady=2, sticky='w')
            
            ttk.Label(frame_entrada, text="Descrição da Entrada:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
            descricao_entrada = ttk.Entry(frame_entrada, width=40)
            descricao_entrada.grid(row=2, column=1, padx=5, pady=2, sticky='w')
            descricao_entrada.insert(0, "ENTRADA")  # Valor padrão
            
            # Ocultar frame de entrada inicialmente
            frame_entrada.grid_remove()
            
            # Função para mostrar/ocultar frame de entrada
            def toggle_entrada():
                if var_tem_entrada.get():
                    frame_entrada.grid()
                else:
                    frame_entrada.grid_remove()
            
            # Configurar checkbox para chamar a função
            check_entrada.config(command=toggle_entrada)
            
            # Frame para gerenciar descrições individuais das parcelas
            frame_descricoes = ttk.LabelFrame(frame_parcelas, text="Descrições Individuais das Parcelas")
            frame_descricoes.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky='ew')
            
            ttk.Label(frame_descricoes, text="Para configurar descrições individuais, primeiro defina o número de parcelas e clique em:").grid(
                row=0, column=0, columnspan=2, padx=5, pady=2, sticky='w')
            
            # Lista para armazenar descrições individuais
            descricoes_parcelas = []
            
            def configurar_descricoes_parcelas():
                try:
                    # Validar número de parcelas
                    if not num_parcelas_entry.get():
                        custom_messagebox("error", "Erro", "Informe o número de parcelas primeiro")
                        return
                        
                    num_parcelas = int(num_parcelas_entry.get())
                    if num_parcelas <= 0:
                        custom_messagebox("error", "Erro", "Número de parcelas deve ser maior que zero")
                        return
                    
                    # Criar janela para configurar descrições
                    janela_descricoes = tk.Toplevel(janela_admin)
                    janela_descricoes.title("Descrições Individuais das Parcelas")
                    janela_descricoes.geometry("500x600")
                    
                    # Frame com scrollbar
                    frame_scroll = ttk.Frame(janela_descricoes)
                    frame_scroll.pack(fill='both', expand=True, padx=10, pady=10)
                    
                    canvas = tk.Canvas(frame_scroll)
                    scrollbar = ttk.Scrollbar(frame_scroll, orient="vertical", command=canvas.yview)
                    frame_content = ttk.Frame(canvas)
                    
                    frame_content.bind(
                        "<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                    )
                    
                    canvas.create_window((0, 0), window=frame_content, anchor="nw")
                    canvas.configure(yscrollcommand=scrollbar.set)
                    
                    canvas.pack(side="left", fill="both", expand=True)
                    scrollbar.pack(side="right", fill="y")
                    
                    # Inicializar ou redimensionar a lista de descrições
                    if len(descricoes_parcelas) < num_parcelas:
                        # Adicionar novas entradas para as parcelas adicionais
                        for _ in range(num_parcelas - len(descricoes_parcelas)):
                            descricoes_parcelas.append("")
                    else:
                        # Truncar a lista se o número de parcelas diminuiu
                        del descricoes_parcelas[num_parcelas:]
                    
                    # Criar campos para cada parcela
                    for i in range(num_parcelas):
                        ttk.Label(frame_content, text=f"Parcela {i+1}:").grid(
                            row=i, column=0, padx=5, pady=5, sticky='w')
                        
                        desc_entry = ttk.Entry(frame_content, width=40)
                        desc_entry.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                        
                        # Preencher com valor existente, se houver
                        if i < len(descricoes_parcelas) and descricoes_parcelas[i]:
                            desc_entry.insert(0, descricoes_parcelas[i])
                        else:
                            desc_entry.insert(0, f"PARCELA {i+1}")
                        
                        # Armazenar referência à entrada para recuperar valores depois
                        desc_entry.idx = i
                    
                    def salvar_descricoes():
                        # Coletar todas as descrições dos campos
                        for child in frame_content.winfo_children():
                            if isinstance(child, ttk.Entry):
                                idx = getattr(child, 'idx', -1)
                                if 0 <= idx < len(descricoes_parcelas):
                                    descricoes_parcelas[idx] = child.get().strip()
                        
                        # Confirmar para o usuário
                        custom_messagebox("info", "Sucesso", "Descrições salvas!")
                        janela_descricoes.destroy()
                    
                    # Botões
                    frame_botoes = ttk.Frame(janela_descricoes)
                    frame_botoes.pack(fill='x', pady=10)
                    
                    ttk.Button(frame_botoes, text="Salvar Descrições", 
                            command=salvar_descricoes).pack(side='right', padx=10)
                    
                    ttk.Button(frame_botoes, text="Cancelar", 
                            command=janela_descricoes.destroy).pack(side='right', padx=10)
                    
                    # Centralizar a janela
                    janela_descricoes.update_idletasks()
                    w = janela_descricoes.winfo_width()
                    h = janela_descricoes.winfo_height()
                    x = (janela_descricoes.winfo_screenwidth() // 2) - (w // 2)
                    y = (janela_descricoes.winfo_screenheight() // 2) - (h // 2)
                    janela_descricoes.geometry(f'{w}x{h}+{x}+{y}')
                    
                    # Tornar a janela modal
                    janela_descricoes.transient(janela_admin)
                    janela_descricoes.grab_set()
                    
                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao configurar descrições: {str(e)}")
                    
            # Botão para configurar descrições individuais
            btn_config_descricoes = ttk.Button(frame_descricoes, 
                                            text="Configurar Descrições", 
                                            command=configurar_descricoes_parcelas)
            btn_config_descricoes.grid(row=1, column=0, padx=5, pady=5, sticky='w')
        
        # 2. Frame para Eventos/Fases
        elif metodo == "Eventos/Fases":
            frame_eventos = ttk.Frame(frame_config_metodo)
            frame_eventos.pack(fill='x', padx=5, pady=5)
            
            # Lista de eventos
            colunas_evento = ('Nº', 'Descrição', 'Percentual', 'Valor')
            tree_eventos = ttk.Treeview(frame_eventos, columns=colunas_evento, show='headings', height=5)
            tree_eventos.heading('Nº', text='Nº')
            tree_eventos.heading('Descrição', text='Descrição')
            tree_eventos.heading('Percentual', text='Percentual (%)')
            tree_eventos.heading('Valor', text='Valor (R$)')
            
            tree_eventos.column('Nº', width=50, anchor='center')
            tree_eventos.column('Descrição', width=300)
            tree_eventos.column('Percentual', width=100, anchor='e')
            tree_eventos.column('Valor', width=100, anchor='e')
            
            # Adicionar scrollbars para eventos
            scroll_y_eventos = ttk.Scrollbar(frame_eventos, orient='vertical', command=tree_eventos.yview)
            scroll_x_eventos = ttk.Scrollbar(frame_eventos, orient='horizontal', command=tree_eventos.xview)
            tree_eventos.configure(yscrollcommand=scroll_y_eventos.set, xscrollcommand=scroll_x_eventos.set)
            
            tree_eventos.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y_eventos.pack(side='right', fill='y')
            scroll_x_eventos.pack(side='bottom', fill='x')
            
            # Frame para adicionar evento
            frame_add_evento = ttk.Frame(frame_eventos)
            frame_add_evento.pack(fill='x', padx=5, pady=5)
            
            ttk.Label(frame_add_evento, text="Descrição:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            evento_descricao = ttk.Entry(frame_add_evento, width=40)
            evento_descricao.grid(row=0, column=1, padx=5, pady=2, sticky='w')
            
            ttk.Label(frame_add_evento, text="Percentual (%):").grid(row=0, column=2, padx=5, pady=2, sticky='w')
            evento_percentual = ttk.Entry(frame_add_evento, width=10)
            evento_percentual.grid(row=0, column=3, padx=5, pady=2, sticky='w')
            
            # Botões para eventos
            frame_botoes_evento = ttk.Frame(frame_eventos)
            frame_botoes_evento.pack(fill='x', pady=5)
            
            # Variável para rastrear o total de percentuais
            total_percentual_var = tk.StringVar(value="Total: 0%")
            lbl_total_percentual = ttk.Label(frame_botoes_evento, textvariable=total_percentual_var)
            lbl_total_percentual.pack(side='left', padx=5)
            
            # Lista para armazenar eventos
            eventos = []
            
            def calcular_valor_evento(percentual, valor_total_str):
                """Calcula o valor do evento baseado no percentual e valor total"""
                try:
                    percentual_float = float(percentual.replace(',', '.'))
                    valor_float = float(valor_total_str.replace(',', '.'))
                    return (percentual_float / 100) * valor_float
                except (ValueError, AttributeError):
                    return 0
            
            def adicionar_evento():
                """Adiciona um evento à lista"""
                if not valor_global_entry.get():
                    custom_messagebox("error", "Erro", "Informe o valor global do contrato primeiro")
                    return
                    
                descricao = evento_descricao.get().strip()
                percentual_str = evento_percentual.get().strip()
                
                if not descricao:
                    custom_messagebox("error", "Erro", "Informe a descrição do evento")
                    return
                    
                try:
                    percentual = float(percentual_str.replace(',', '.'))
                    if percentual <= 0 or percentual > 100:
                        custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100")
                        return
                except ValueError:
                    custom_messagebox("error", "Erro", "Percentual inválido")
                    return
                    
                # Calcular total atual
                total_atual = sum(float(e[1]) for e in eventos)
                
                # Verificar se ultrapassa 100%
                if total_atual + percentual > 100:
                    custom_messagebox("error", "Erro", "Total de percentual não pode exceder 100%")
                    return
                    
                # Calcular valor baseado no percentual
                valor_total = valor_global_entry.get().replace(',', '.')
                try:
                    valor_total_float = float(valor_total)
                    valor_evento = (percentual / 100) * valor_total_float
                except (ValueError, TypeError):
                    valor_evento = 0
                    
                # Adicionar à lista
                eventos.append((descricao, percentual, valor_evento))
                
                # Adicionar ao treeview
                tree_eventos.insert('', 'end', values=(
                    len(eventos),  # Número sequencial
                    descricao, 
                    f"{percentual:.2f}", 
                    f"R$ {valor_evento:.2f}"
                ))
                
                # Atualizar total
                total_percentual_var.set(f"Total: {total_atual + percentual:.2f}%")
                
                # Limpar campos
                evento_descricao.delete(0, tk.END)
                evento_percentual.delete(0, tk.END)
                
            def remover_evento():
                """Remove o evento selecionado"""
                selecionado = tree_eventos.selection()
                if not selecionado:
                    custom_messagebox("warning",  "Aviso", "Selecione um evento para remover")
                    return
                    
                # Obter valores
                valores = tree_eventos.item(selecionado)['values']
                indice = int(valores[0]) - 1  # Ajusta para índice 0-based
                
                if 0 <= indice < len(eventos):
                    # Remover da lista
                    eventos.pop(indice)
                    
                    # Limpar e recriar treeview para atualizar numeração
                    for item in tree_eventos.get_children():
                        tree_eventos.delete(item)
                        
                    for i, (desc, perc, valor) in enumerate(eventos, 1):
                        tree_eventos.insert('', 'end', values=(i, desc, f"{perc:.2f}", f"R$ {valor:.2f}"))
                    
                    # Atualizar total
                    total_atual = sum(float(e[1]) for e in eventos)
                    total_percentual_var.set(f"Total: {total_atual:.2f}%")
            
            # Configurar botões de eventos
            ttk.Button(frame_botoes_evento, text="Adicionar Evento", command=adicionar_evento).pack(side='right', padx=5)
            ttk.Button(frame_botoes_evento, text="Remover Evento", command=remover_evento).pack(side='right', padx=5)
        
        # Função de busca para fornecedores
        def busca_local():
            """Função de busca"""
            termo = busca_entry.get()
            buscar_fornecedor(tree_fornecedores, termo)
            
        ttk.Button(frame_busca, text="Buscar", command=busca_local).pack(side='left', padx=5)
        busca_entry.bind('<Return>', lambda e: busca_local())
        
        def selecionar_e_preencher(event=None):
            """Seleciona fornecedor e preenche campos"""
            selecionado = tree_fornecedores.selection()
            if not selecionado:
                return
                
            valores = tree_fornecedores.item(selecionado)['values']
            cnpj_cpf_entry.config(state='normal')
            nome_entry.config(state='normal')
            
            cnpj_cpf_entry.delete(0, tk.END)
            cnpj_cpf_entry.insert(0, str(valores[0]).zfill(14))
            
            nome_entry.delete(0, tk.END)
            nome_entry.insert(0, valores[1])
            
            cnpj_cpf_entry.config(state='readonly')
            nome_entry.config(state='readonly')
            
        tree_fornecedores.bind('<Double-1>', selecionar_e_preencher)
        
        def confirmar():
            """Confirma a adição do administrador"""
            try:
                if not cnpj_cpf_entry.get() or not nome_entry.get() or not tipo_combo.get():
                    custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
                    return
                    
                # Capturar a forma de pagamento para os dados bancários
                forma_pagto_selecionada = forma_pagamento.get()
                
                # Verificar configuração específica do método
                if metodo == "Valor Fixo em Parcelas":
                    # Validar número de parcelas
                    if not num_parcelas_entry.get():
                        custom_messagebox("error", "Erro", "Informe o número de parcelas!")
                        return
                        
                    try:
                        num_parcelas = int(num_parcelas_entry.get())
                        if num_parcelas <= 0:
                            custom_messagebox("error", "Erro", "Número de parcelas deve ser maior que zero!")
                            return
                    except ValueError:
                        custom_messagebox("error", "Erro", "Número de parcelas inválido!")
                        return
                    
                    # Se tem entrada configurada, validar entrada
                    if var_tem_entrada.get():
                        if not valor_entrada_entry.get():
                            custom_messagebox("error", "Erro", "Informe o valor da entrada!")
                            return
                        
                        try:
                            valor_entrada = float(valor_entrada_entry.get().replace(',', '.'))
                            if valor_entrada <= 0:
                                custom_messagebox("error", "Erro", "Valor da entrada deve ser maior que zero!")
                                return
                        except ValueError:
                            custom_messagebox("error", "Erro", "Valor da entrada inválido!")
                            return
                
                # Verificar eventos para contratos do tipo Eventos/Fases
                if metodo == "Eventos/Fases" and not eventos:
                    custom_messagebox("error", "Erro", "Adicione pelo menos um evento para este administrador!")
                    return
                    
                # Para contratos de eventos, verificar total de percentuais
                if metodo == "Eventos/Fases":
                    total_percentual = sum(float(e[1]) for e in eventos)
                    if total_percentual < 99.99 or total_percentual > 100.01:  # Pequena margem de erro
                        if not custom_messagebox("yesno", "Confirmação", 
                                            f"O total de percentuais é {total_percentual:.2f}% ao invés de 100%. Deseja continuar mesmo assim?"):
                            return
            
                if tipo_combo.get() == 'Percentual':
                    # Validar percentual
                    if not percentual_entry.get():
                        custom_messagebox("error", "Erro", "Preencha o percentual!")
                        return
                    
                    try:
                        perc = float(percentual_entry.get().replace(',', '.'))
                        if perc <= 0 or perc > 100:
                            custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100!")
                            return
                            
                        # Configurar campos adicionais conforme método
                        if metodo == "Percentual da Quinzena":
                            # Simples percentual da quinzena
                            num_parcelas = ""
                            data_inicial = ""
                        elif metodo == "Valor Fixo em Parcelas":
                            # Número de parcelas informado no contrato
                            num_parcelas = num_parcelas_entry.get()
                            # Data inicial se houver entrada
                            data_inicial = data_entrada.get() if var_tem_entrada.get() else ""
                        else:  # Eventos/Fases
                            # Número de eventos
                            num_parcelas = str(len(eventos))
                            data_inicial = ""
                            
                        # Adicionar registro de percentual
                        valores_percentual = (
                            cnpj_cpf_entry.get(),
                            nome_entry.get(),
                            tipo_combo.get(),
                            f"{perc:.2f}%",  # Formatação com %
                            f"{valor_global_float:.2f}",  # Valor Total
                            num_parcelas,  # Número de parcelas conforme método
                            data_inicial  # Data inicial conforme método
                        )
                        
                        # Preparar tags adicionais
                        tags_extra = []
                        
                        # Adicionar descrições individuais como tag se for Valor Fixo em Parcelas
                        if metodo == "Valor Fixo em Parcelas" and descricoes_parcelas:
                            tags_extra.append(f"descricoes:{','.join(descricoes_parcelas)}")
                            
                        # Adicionar informações de entrada se necessário
                        if metodo == "Valor Fixo em Parcelas" and var_tem_entrada.get():
                            tags_extra.append(f"desc_entrada:{descricao_entrada.get()}")
                        
                        # Tags finais incluem tipo de percentual, forma de pagamento e tags extras
                        tags_finais = ['percentual', forma_pagto_selecionada, *tags_extra]
                        
                        tree.insert('', 'end', values=valores_percentual, tags=tags_finais)
                        
                    except ValueError:
                        custom_messagebox("error", "Erro", "Percentual inválido!")
                        return
                        
                elif tipo_combo.get() == 'Fixo':
                    if not valor_total_entry.get():
                        custom_messagebox("error", "Erro", "Preencha o valor total!")
                        return
                        
                    try:
                        valor_total_adm = float(valor_total_entry.get().replace(',', '.'))
                        if valor_total_adm <= 0:
                            custom_messagebox("error", "Erro", "Valor total deve ser maior que zero!")
                            return
                    except ValueError:
                        custom_messagebox("error", "Erro", "Valor total inválido!")
                        return
                    
                    # Configurar campos adicionais conforme método
                    if metodo == "Valor Fixo em Parcelas":
                        # Número de parcelas informado no contrato
                        num_parcelas = num_parcelas_entry.get()
                        # Data inicial se houver entrada
                        data_inicial = data_entrada.get() if var_tem_entrada.get() else ""
                        # Descrição da entrada
                        desc_entrada = descricao_entrada.get() if var_tem_entrada.get() else ""
                    else:  # Eventos/Fases
                        # Número de eventos
                        num_parcelas = str(len(eventos))
                        data_inicial = ""
                        desc_entrada = ""
                        
                    # Adicionar registro de valor fixo
                    valores_fixo = (
                        cnpj_cpf_entry.get(),
                        nome_entry.get(),
                        tipo_combo.get(),
                        "",  # Sem percentual para fixo
                        valor_total_entry.get(),
                        num_parcelas,
                        data_inicial
                    )
                    
                    # Adicionar tags extras para armazenar informações
                    tags_extras = []
                    
                    # Valores específicos para parcelas fixas
                    if metodo == "Valor Fixo em Parcelas":
                        # Valor e descrição de entrada se houver
                        if var_tem_entrada.get():
                            tags_extras.append(f"entrada:{valor_entrada_entry.get()}")
                            tags_extras.append(f"desc_entrada:{descricao_entrada.get()}")
                        
                        # Adicionar descrições individuais
                        if descricoes_parcelas:
                            tags_extras.append(f"descricoes:{','.join(descricoes_parcelas)}")
                    
                    # Tags completas: tipo fixo, forma de pagamento e extras
                    tags = [
                        'fixo', 
                        forma_pagto_selecionada,
                        *tags_extras
                    ]
                    
                    tree.insert('', 'end', values=valores_fixo, tags=tags)
                
                # Se tiver eventos, registrá-los na lista global
                if metodo == "Eventos/Fases":
                    # Armazenar eventos como tags adicionais no item
                    eventos_serializados = []
                    for desc, perc, valor in eventos:
                        eventos_serializados.append(f"{desc}:{perc}:{valor}")
                        
                    # Atualizar tags do item para incluir eventos
                    for item in tree.get_children():
                        # Pegar o último item inserido (mais recente)
                        if item == tree.get_children()[-1]:
                            tags_atuais = tree.item(item)['tags']
                            # Adicionar a tag com eventos
                            nova_tag = f"eventos:{'|'.join(eventos_serializados)}"
                            tree.item(item, tags=(*tags_atuais, nova_tag))
                
                # Fechar a janela
                janela_admin.destroy()
                
                # Garantir que a janela do contrato seja trazida para frente
                # Usar after para garantir que a janela tenha tempo de ser destruída primeiro
                if metodo_pagamento_combo.winfo_toplevel().winfo_exists():
                    metodo_pagamento_combo.winfo_toplevel().after(100, lambda: (
                        metodo_pagamento_combo.winfo_toplevel().lift(),
                        metodo_pagamento_combo.winfo_toplevel().focus_force()
                    ))
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao confirmar: {str(e)}")
                
        # Botões
        frame_botoes = ttk.Frame(frame_admin)
        frame_botoes.pack(fill='x', pady=10)
        ttk.Button(frame_botoes, text="Confirmar", command=confirmar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela_admin.destroy).pack(side='left', padx=5)       

    def processar_parcelas_fixas(self, ws, num_contrato, valor_global, opcoes):
        """Processa parcelas fixas para o contrato"""
        try:
            # Debug
            print("Início de processar_parcelas_fixas")
            print(f"Opções: {opcoes}")
            
            num_parcelas = int(opcoes.get('num_parcelas', 0))
            tem_entrada = opcoes.get('tem_entrada', False)
            descricoes_parcelas = opcoes.get('descricoes_parcelas', {})  # Dicionário com descrições por admin
            
            print(f"Processando {num_parcelas} parcelas, entrada: {tem_entrada}")
            
            if num_parcelas <= 0:
                print("Erro: Número de parcelas inválido")
                return
                
            # Processar cada administrador
            for item in self.tree_adm.get_children():
                valores_adm = self.tree_adm.item(item)['values']
                tags_adm = self.tree_adm.item(item)['tags']
                
                print(f"Processando administrador: {valores_adm}")
                
                cnpj_cpf_adm = str(valores_adm[0]).strip()
                cnpj_cpf_adm = formatar_cnpj_cpf(cnpj_cpf_adm)
                nome_adm = valores_adm[1]
                
                # Extrair descricoes das tags, se existirem
                descricoes_individuais = []
                for tag in tags_adm:
                    if tag.startswith('descricoes:'):
                        descricoes_individuais = tag.replace('descricoes:', '').split(',')
                        print(f"Descrições individuais: {descricoes_individuais}")
                        break
                
                # Também verificar no dicionário de descrições
                if not descricoes_individuais and cnpj_cpf_adm in descricoes_parcelas:
                    descricoes_individuais = descricoes_parcelas[cnpj_cpf_adm]
                    print(f"Usando descrições do dicionário: {descricoes_individuais}")
                
                # Extrair descrição da entrada, se existir
                descricao_entrada = "ENTRADA"
                for tag in tags_adm:
                    if tag.startswith('desc_entrada:'):
                        descricao_entrada = tag.replace('desc_entrada:', '')
                        print(f"Descrição da entrada: {descricao_entrada}")
                        break
                
                # Calcular valor por parcela para este administrador
                try:
                    if valores_adm[2] == 'Percentual':
                        # Administrador com percentual do valor total
                        perc_adm = float(str(valores_adm[3]).replace('%', '').replace(',', '.'))
                        valor_total_adm = (perc_adm / 100) * valor_global
                        print(f"Valor calculado baseado em percentual: {valor_total_adm}")
                    else:  # Fixo
                        # Valor fixo total para o administrador
                        valor_texto = str(valores_adm[4]).replace(',', '.')
                        print(f"Valor texto: {valor_texto}")
                        valor_total_adm = float(valor_texto)
                        print(f"Valor fixo: {valor_total_adm}")
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Erro ao calcular valor: {e}")
                    valores_str = ', '.join([str(v) for v in valores_adm])
                    print(f"Valores disponíveis: {valores_str}")
                    # Tentar alternativa
                    if len(valores_adm) >= 5 and valores_adm[4]:
                        try:
                            valor_total_adm = float(str(valores_adm[4]).replace(',', '.'))
                            print(f"Valor alternativo: {valor_total_adm}")
                        except (ValueError, TypeError):
                            print("Erro na alternativa também")
                            valor_total_adm = 0
                    else:
                        valor_total_adm = 0
                
                if valor_total_adm <= 0:
                    print("Valor total inválido, pulando administrador")
                    continue
                
                # Se tem entrada, tratar separadamente
                if tem_entrada:
                    valor_entrada = 0
                    # Buscar valor da entrada nas tags
                    for tag in tags_adm:
                        if tag.startswith('entrada:'):
                            try:
                                valor_entrada = float(tag.replace('entrada:', '').replace(',', '.'))
                                print(f"Valor da entrada das tags: {valor_entrada}")
                            except ValueError:
                                valor_entrada = 0
                            break
                    
                    if valor_entrada <= 0:
                        # Calcular proporcional se não estiver explícito
                        valor_entrada_opcoes = opcoes.get('valor_entrada', 0)
                        if isinstance(valor_entrada_opcoes, str):
                            valor_entrada_opcoes = float(valor_entrada_opcoes.replace(',', '.'))
                        # Proporcional da entrada para este administrador
                        proporcao_entrada = valor_entrada_opcoes / valor_global if valor_global else 0
                        valor_entrada_adm = valor_total_adm * proporcao_entrada
                        print(f"Valor da entrada calculado: {valor_entrada_adm}")
                    else:
                        # Usar o valor específico
                        valor_entrada_adm = valor_entrada
                        
                    data_entrada = opcoes.get('data_entrada')
                    print(f"Data da entrada: {data_entrada}")
                    
                    # Registrar entrada como primeira parcela
                    proxima_linha = ws.max_row + 1
                    ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                    ws.cell(row=proxima_linha, column=26, value=1)  # Número da parcela (entrada = 1)
                    ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                    ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                    ws.cell(row=proxima_linha, column=29, value=data_entrada)  # Data vencimento
                    ws.cell(row=proxima_linha, column=30, value=valor_entrada_adm)  # Valor
                    ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                    ws.cell(row=proxima_linha, column=32, value="")  # Sem evento
                    ws.cell(row=proxima_linha, column=33, value=descricao_entrada.upper())  # Descrição personalizada da entrada
                    
                    print(f"Registrada entrada com valor {valor_entrada_adm}")
                    
                    # Ajustar valor restante para as demais parcelas
                    valor_restante = valor_total_adm - valor_entrada_adm
                    valor_parcela = valor_restante / num_parcelas if num_parcelas > 0 else 0
                    
                    print(f"Valor de cada parcela após entrada: {valor_parcela}")
                    
                    # Registrar as demais parcelas
                    for i in range(1, num_parcelas + 1):
                        proxima_linha = ws.max_row + 1
                        ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                        ws.cell(row=proxima_linha, column=26, value=i + 1)  # Número da parcela (após entrada)
                        ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                        ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                        ws.cell(row=proxima_linha, column=29, value=None)  # Data vencimento (a definir)
                        ws.cell(row=proxima_linha, column=30, value=valor_parcela)  # Valor
                        ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                        ws.cell(row=proxima_linha, column=32, value="")  # Sem evento
                        
                        # Usar descrição individual se disponível
                        if i-1 < len(descricoes_individuais) and descricoes_individuais[i-1]:
                            descricao = descricoes_individuais[i-1]
                        else:
                            descricao = f"PARCELA {i}"
                            
                        ws.cell(row=proxima_linha, column=33, value=descricao.upper())  # Descrição individual ou genérica
                        print(f"Registrada parcela {i} com valor {valor_parcela} e descrição '{descricao}'")
                            
                else:
                    # Sem entrada, distribuir igualmente
                    valor_parcela = valor_total_adm / num_parcelas if num_parcelas > 0 else 0
                    print(f"Valor de cada parcela (sem entrada): {valor_parcela}")
                    
                    # Registrar parcelas
                    for i in range(1, num_parcelas + 1):
                        proxima_linha = ws.max_row + 1
                        ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())  # Contrato
                        ws.cell(row=proxima_linha, column=26, value=i)  # Número da parcela
                        ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                        ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                        ws.cell(row=proxima_linha, column=29, value=None)  # Data vencimento (a definir)
                        ws.cell(row=proxima_linha, column=30, value=valor_parcela)  # Valor
                        ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                        ws.cell(row=proxima_linha, column=32, value="")  # Sem evento
                        
                        # Usar descrição individual se disponível
                        if i-1 < len(descricoes_individuais) and descricoes_individuais[i-1]:
                            descricao = descricoes_individuais[i-1]
                        else:
                            descricao = f"PARCELA {i}"
                            
                        ws.cell(row=proxima_linha, column=33, value=descricao.upper())  # Descrição individual ou genérica
                        print(f"Registrada parcela {i} com valor {valor_parcela} e descrição '{descricao}'")
            
            print("Finalizado processamento de parcelas fixas com sucesso")
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"Erro em processar_parcelas_fixas: {str(e)}") 

    def processar_administradores(self, ws, num_contrato, valor_global, metodo_pagamento, opcoes):
        """Processa os administradores do contrato"""
        for item in self.tree_adm.get_children():
            valores = self.tree_adm.item(item)['values']
            tags = self.tree_adm.item(item)['tags']
            
            # Formatação do CNPJ/CPF
            cnpj_cpf = str(valores[0]).strip()
            cnpj_cpf = formatar_cnpj_cpf(cnpj_cpf)
            nome_admin = valores[1]
            
            # Buscar dados bancários do fornecedor
            forma_pagamento = next((tag for tag in tags if tag in ['PIX', 'TED']), 'PIX')
            dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj_cpf, forma_pagamento)

            # Registrar administrador no contrato com os dados apropriados
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=7, value=num_contrato.upper())  # Contrato
            ws.cell(row=proxima_linha, column=8, value=cnpj_cpf)              # CNPJ/CPF
            ws.cell(row=proxima_linha, column=9, value=nome_admin)            # Nome
            ws.cell(row=proxima_linha, column=10, value=valores[2])           # Tipo (Percentual/Fixo)
            ws.cell(row=proxima_linha, column=11, value=valores[3])           # Valor/Percentual
            ws.cell(row=proxima_linha, column=12, value=valores[4])           # Valor Total
            ws.cell(row=proxima_linha, column=13, value=valores[5])           # Número de parcelas
            
            # Data inicial para casos que têm entrada
            if valores[6] and metodo_pagamento == "Valor Fixo em Parcelas" and opcoes.get('tem_entrada'):
                ws.cell(row=proxima_linha, column=14, value=opcoes.get('data_entrada'))  # Data inicial


    def salvar_contrato_com_opcoes(self, num_contrato, data_inicio, data_fim, observacoes, valor_global, metodo_pagamento, opcoes, janela):
        """Salva os dados do contrato com diferentes opções de pagamento"""
        num_contrato = str(num_contrato).upper()
        
        try:
            # Adicione instruções de debug para verificar o fluxo
            print(f"Salvando contrato: {num_contrato}, método: {metodo_pagamento}")
            
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            # Verificar se o contrato já existe
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).upper() == num_contrato.upper():
                    custom_messagebox("error", "Erro", "Número de contrato já existe!")
                    return

            # Salvar dados do contrato
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1, value=num_contrato.upper())
            ws.cell(row=proxima_linha, column=2, value=data_inicio)
            ws.cell(row=proxima_linha, column=3, value=data_fim)
            ws.cell(row=proxima_linha, column=4, value='ATIVO')
            ws.cell(row=proxima_linha, column=5, value=observacoes)
            ws.cell(row=proxima_linha, column=6, value=valor_global)  # Valor global do contrato

            # Extrair dados dos administradores, incluindo descrições de parcelas
            opcoes_processadas = opcoes.copy() if opcoes else {}
            
            # Extrair informações adicionais dos administradores
            tem_entrada = False
            valor_entrada = 0
            data_entrada = None
            num_parcelas = 0
            
            # Coletar informações específicas para parcelas fixas
            if metodo_pagamento == "Valor Fixo em Parcelas":
                # Percorrer os administradores para extrair informações de parcelas
                for item in self.tree_adm.get_children():
                    valores = self.tree_adm.item(item)['values']
                    tags = self.tree_adm.item(item)['tags']
                    
                    # Extrair número de parcelas
                    if valores[5] and not num_parcelas:
                        try:
                            num_parcelas = int(valores[5])
                        except (ValueError, TypeError):
                            num_parcelas = 0
                    
                    # Verificar se tem entrada
                    for tag in tags:
                        if tag.startswith('entrada:'):
                            tem_entrada = True
                            try:
                                valor_entrada = float(tag.replace('entrada:', '').replace(',', '.'))
                            except (ValueError, TypeError):
                                valor_entrada = 0
                    
                    # Extrair data de entrada
                    if valores[6] and not data_entrada:
                        data_entrada = valores[6]

                # Adicionar ao dicionário de opções
                opcoes_processadas['num_parcelas'] = num_parcelas
                opcoes_processadas['tem_entrada'] = tem_entrada
                opcoes_processadas['valor_entrada'] = valor_entrada
                opcoes_processadas['data_entrada'] = data_entrada
                
                print(f"Configurações de parcelas: parcelas={num_parcelas}, entrada={tem_entrada}, valor_entrada={valor_entrada}")
            
            # Coletar descrições para cada administrador
            admin_descricoes = {}
            
            for item in self.tree_adm.get_children():
                tags = self.tree_adm.item(item)['tags']
                cnpj_cpf = self.tree_adm.item(item)['values'][0]
                
                # Extrair descricoes das tags, se existirem
                for tag in tags:
                    if tag.startswith('descricoes:'):
                        admin_descricoes[cnpj_cpf] = tag.replace('descricoes:', '').split(',')
                        print(f"Descrições para {cnpj_cpf}: {admin_descricoes[cnpj_cpf]}")
                        break
            
            # Adicionar ao dicionário de opções
            opcoes_processadas['descricoes_parcelas'] = admin_descricoes

            # Processar administradores baseado no método de pagamento
            self.processar_administradores(ws, num_contrato, valor_global, metodo_pagamento, opcoes_processadas)

            # Processar eventos se método for por eventos/fases
            if metodo_pagamento == "Eventos/Fases":
                # Extrair eventos dos administradores
                eventos = []
                for item in self.tree_adm.get_children():
                    tags = self.tree_adm.item(item)['tags']
                    for tag in tags:
                        if tag.startswith('eventos:'):
                            eventos_str = tag.replace('eventos:', '')
                            for evento_str in eventos_str.split('|'):
                                partes = evento_str.split(':')
                                if len(partes) == 3:
                                    desc, perc, valor = partes
                                    eventos.append((desc, float(perc), float(valor)))
                            break

                self.processar_eventos(ws, num_contrato, valor_global, eventos)
                        
            # Processar parcelas fixas se for o método apropriado
            elif metodo_pagamento == "Valor Fixo em Parcelas":
                print("Chamando processar_parcelas_fixas...")
                self.processar_parcelas_fixas(ws, num_contrato, valor_global, opcoes_processadas)

            # Salvar e fechar o arquivo explicitamente
            try:
                print(f"Salvando o arquivo {self.arquivo_cliente}")
                wb.save(self.arquivo_cliente)
                wb.close()  # Importante fechar o arquivo
            except PermissionError:
                custom_messagebox("error", "Erro", f"Não foi possível salvar a planilha. Ela pode estar aberta em outro programa.")
                return
            except Exception as e:
                import traceback
                traceback.print_exc()
                custom_messagebox("error", "Erro", f"Erro ao salvar planilha: {str(e)}")
                return
                
            # Exibir mensagem de sucesso
            custom_messagebox("info", "Sucesso", "Contrato cadastrado com sucesso!")
            
            # Fechar a janela atual
            janela.destroy()
            
            # Garantir que a janela de gestão de contratos é trazida para frente
            # após salvar o contrato e recarregar a lista
            self.carregar_contratos()
            
            # Usar after para garantir que toda a interface seja atualizada
            if self.parent and self.parent.winfo_exists():
                self.parent.after(100, lambda: (
                    self.parent.lift(),
                    self.parent.focus_force()
                ))

        except Exception as e:
            import traceback
            traceback.print_exc()  # Imprime o stack trace completo
            custom_messagebox("error", "Erro", f"Erro ao salvar contrato: {str(e)}")
            if 'wb' in locals() and wb:
                try:
                    wb.close()
                except:
                    pass

    def excluir_contrato(self):
        """Exclui o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um contrato para excluir")
            return
            
        if custom_messagebox("yesno", "Confirmação", 
                              "Deseja realmente excluir este contrato e seus administradores?"):
            try:
                num_contrato = self.tree_contratos.item(selecionado)['values'][0]
                
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']
                
                # Marcar contrato como inativo
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == num_contrato:
                        row[3].value = 'INATIVO'  # Coluna D - Status
                
                wb.save(self.arquivo_cliente)
                self.carregar_contratos()
                custom_messagebox("info", "Sucesso", "Contrato marcado como inativo")
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao excluir contrato: {str(e)})")

            
class GestaoTaxasFixas:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.gestor_parcelas = GestorParcelas(self)

    def processar_lancamentos_fixos(self, cliente, data_ref):
        """Processa os lançamentos de taxas fixas para a data de referência"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            lancamentos_gerados = []
            
            # Buscar contratos ativos com taxa fixa
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verifica se é registro de administrador e tipo fixo
                if (row[6] and  # Tem nº contrato na coluna G
                    row[9] == 'Fixo' and  # É tipo fixo
                    self.contrato_ativo(ws, row[6])):  # Contrato está ativo
                    
                    # Verificar se já tem lançamento para este período
                    if not self.tem_lancamento(ws, row[6], row[7], data_ref):
                        # Preparar dados para o lançamento
                        dados_lancamento = {
                            'data_rel': data_ref,
                            'cnpj_cpf': row[7],  # CNPJ/CPF
                            'nome': row[8],      # Nome/Razão Social
                            'referencia': f'ADM FIXA REF. {data_ref.strftime("%m/%Y")}',
                            'valor': float(row[10].replace(',', '.')),  # Valor/Parcela
                            'dt_vencto': self.calcular_vencimento(data_ref)
                        }
                        
                        # Registrar lançamento no sistema
                        self.sistema.dados_para_incluir.append(dados_lancamento)
                        lancamentos_gerados.append(dados_lancamento)
                        
                        # Registrar na aba de controle
                        self.registrar_lancamento(ws, dados_lancamento)
                        
            wb.save(arquivo_cliente)
            return lancamentos_gerados
            
        except Exception as e:
            raise Exception(f"Erro ao processar lançamentos fixos: {str(e)}")

    def contrato_ativo(self, ws, num_contrato):
        """Verifica se o contrato está ativo"""
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == num_contrato:  # Coluna A (Nº Contrato)
                return row[3] == 'ATIVO'  # Coluna D (Status)
        return False

    def tem_lancamento(self, ws, num_contrato, cnpj_cpf, data_ref):
        """Verifica se já existe lançamento para o período"""
        data_str = data_ref.strftime("%d/%m/%Y")
        for row in ws.iter_rows(min_row=3, values_only=True):
            if (row[25] and  # Tem referência na coluna PARCELAS
                row[24] == num_contrato and  # Mesmo contrato
                row[26] == cnpj_cpf and  # Mesmo CNPJ/CPF
                row[28] == data_str):  # Mesma data
                return True
        return False

    def calcular_vencimento(self, data_ref):
        """Calcula data de vencimento (dia 5 do mês seguinte)"""
        if data_ref.day == 5:
            vencto = data_ref.replace(day=20)
        else:  # day == 20
            if data_ref.month == 12:
                vencto = data_ref.replace(year=data_ref.year + 1, month=1, day=5)
            else:
                vencto = data_ref.replace(month=data_ref.month + 1, day=5)
        return vencto

    def registrar_lancamento(self, ws, dados):
        """Registra o lançamento na aba de controle"""
        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=26, value=dados['cnpj_cpf'])
        ws.cell(row=proxima_linha, column=27, value=dados['nome'])
        ws.cell(row=proxima_linha, column=28, value=dados['data_rel'])
        ws.cell(row=proxima_linha, column=29, value=dados['valor'])
        ws.cell(row=proxima_linha, column=30, value='LANÇADO')


class GestaoAdministradores:
    def __init__(self, parent):
        self.parent = parent
        self.busca_entry = None
        self.tree_fornecedores = None
        self.administradores = []  # Lista para armazenar os administradores
        
    def abrir_janela_admin(self):
        """Abre janela para gestão de administradores"""
        self.janela_admin = tk.Toplevel(self.parent)
        self.janela_admin.title("Gestão de Administradores")
        self.janela_admin.geometry("800x600")
        
        # Frame para busca de fornecedor
        frame_busca = ttk.LabelFrame(self.janela_admin, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        self.busca_entry = ttk.Entry(frame_busca, width=40)  # Definir como atributo da classe
        self.busca_entry.pack(side='left', padx=5)

        # Criar a tree de fornecedores antes de usar
        self.tree_fornecedores = ttk.Treeview(frame_busca, 
            columns=('CNPJ/CPF', 'Nome', 'Categoria'),
            show='headings')
        
        # Definir a função de busca
        def buscar():
            termo = self.busca_entry.get()
            # Implementar lógica de busca aqui
            
        # Definir a função de seleção    
        def selecionar(event):
            # Implementar lógica de seleção aqui
            pass

        self.busca_entry.bind('<Return>', lambda e: buscar())
        self.tree_fornecedores.bind('<<TreeviewSelect>>', selecionar)
        ttk.Button(frame_busca, text="Buscar", command=buscar).pack(side='left', padx=5)

        
        # Frame para lista de fornecedores
        frame_fornecedores = ttk.LabelFrame(self.janela_admin, text="Fornecedores")
        frame_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tree_fornecedores = ttk.Treeview(frame_fornecedores, 
                                             columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                             show='headings',
                                             height=5)
        self.tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_fornecedores.heading('Nome', text='Nome')
        self.tree_fornecedores.heading('Categoria', text='Categoria')
        self.tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para percentual
        frame_percentual = ttk.Frame(self.janela_admin)
        frame_percentual.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_percentual, text="Percentual (%):").pack(side='left', padx=5)
        self.percentual_entry = ttk.Entry(frame_percentual, width=10)
        self.percentual_entry.pack(side='left', padx=5)
        
        ttk.Button(frame_percentual, 
                  text="Adicionar Administrador", 
                  command=self.adicionar_administrador).pack(side='left', padx=5)
        
        # Frame para lista de administradores
        frame_lista = ttk.LabelFrame(self.janela_admin, text="Administradores Cadastrados")
        frame_lista.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tree_admin = ttk.Treeview(frame_lista, 
                                     columns=('CNPJ/CPF', 'Nome', 'Percentual'),
                                     show='headings')
        self.tree_admin.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_admin.heading('Nome', text='Nome')
        self.tree_admin.heading('Percentual', text='Percentual (%)')
        self.tree_admin.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.janela_admin)
        frame_botoes.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(frame_botoes, 
                  text="Remover Administrador", 
                  command=self.remover_administrador).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Concluir", 
                  command=self.finalizar_gestao).pack(side='right', padx=5)

    def buscar_fornecedor(self):
        termo = self.busca_entry.get()
        buscar_fornecedor(self.tree_fornecedores, termo)
            
    def adicionar_administrador(self):
        """Adiciona um fornecedor selecionado como administrador"""
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um fornecedor")
            return
            
        fornecedor = self.tree_fornecedores.item(selecionado)['values']
        percentual = self.percentual_entry.get().strip()
        
        # Validar percentual
        if not percentual:
            custom_messagebox("error", "Erro", "Informe o percentual!")
            return
            
        try:
            percentual_float = float(percentual.replace(',', '.'))
            if percentual_float <= 0 or percentual_float > 100:
                custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100!")
                return
        except ValueError:
            custom_messagebox("error", "Erro", "Percentual inválido!")
            return
            
        # Formatar CNPJ/CPF como string
        cnpj_cpf = str(fornecedor[0]).strip()  # Converter para string e remover espaços
        
        # Verificar se o fornecedor já está na lista
        for admin in self.administradores:
            if admin[0] == cnpj_cpf:  # Compara CNPJ/CPF
                custom_messagebox("error", "Erro", "Este fornecedor já está cadastrado como administrador!")
                return
                
        # Verificar se o total de percentuais não excede 100%
        total_atual = sum(float(item[2].replace(',', '.')) 
                         for item in self.administradores)
        if total_atual + percentual_float > 100:
            custom_messagebox("error", "Erro", "Soma dos percentuais excede 100%!")
            return
            
        # Adicionar à lista e à treeview usando o CNPJ/CPF como string
        self.administradores.append((cnpj_cpf, fornecedor[1], percentual))
        self.tree_admin.insert('', 'end', values=(cnpj_cpf, fornecedor[1], percentual))
        
        # Limpar campo de percentual
        self.percentual_entry.delete(0, tk.END)
        
    def remover_administrador(self):
        """Remove o administrador selecionado"""
        selecionado = self.tree_admin.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um administrador para remover")
            return
        
        self.tree_admin.delete(selecionado)
        valores = self.tree_admin.item(selecionado)['values']
        self.administradores = [(cnpj, nome, perc) for cnpj, nome, perc 
                              in self.administradores 
                              if cnpj != valores[0]]
        
    def finalizar_gestao(self):
        """Finaliza a gestão de administradores"""
        total = sum(float(perc.replace(',', '.')) 
                   for _, _, perc in self.administradores)
        if total > 100:
            custom_messagebox("error", "Erro", "Soma dos percentuais excede 100%!")
            return
        
        self.janela_admin.destroy()
        
    def get_administradores(self):
        """Retorna a lista de administradores configurados"""
        return self.administradores.copy()        



class GestorParcelas:
    def __init__(self, parent):
        print("Inicializando GestorParcelas")  # Debug
        self.parent = parent
        self.parcelas = []
        self.tipo_despesa_valor = '3'
        self.janela_parcelas = None
        self._var_tem_entrada = None  # Inicializa como None
        # Limpar referências de widgets
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None

    @property
    def tem_entrada(self):
        """Getter para tem_entrada - cria apenas quando necessário"""
        if self._var_tem_entrada is None:
            self._var_tem_entrada = tk.BooleanVar(master=self.parent.root, value=False)
        return self._var_tem_entrada


    # Interface e Controles
    def abrir_janela_parcelas(self):
        print("Abrindo janela de parcelas")  # Debug
        # Criar janela como Toplevel do parent
        self.janela_parcelas = tk.Toplevel(self.parent.root)
        self.janela_parcelas.title("Configuração de Parcelas")
        self.janela_parcelas.geometry("600x700")
        
        # Garantir que a janela seja modal
        self.janela_parcelas.transient(self.parent.root)
        self.janela_parcelas.grab_set()
        
        frame = ttk.Frame(self.janela_parcelas, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame para entrada
        frame_entrada = ttk.LabelFrame(frame, text="Entrada")
        frame_entrada.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        print("Criando Checkbutton")  # Debug
        check = ttk.Checkbutton(
            frame_entrada, 
            text="Possui entrada?",
            variable=self.tem_entrada,
            command=self.atualizar_campos_entrada
        )
        check.grid(row=0, column=0, padx=5, pady=5)

        # Frame para modalidades de entrada
        print("Criando frame modalidade")  # Debug
        self.frame_modalidade = ttk.Frame(frame_entrada)
        self.frame_modalidade.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        ttk.Label(self.frame_modalidade, text="Modalidade de Entrada:").grid(row=0, column=0, padx=5, pady=2)
        self.modalidade_entrada = ttk.Combobox(self.frame_modalidade, state='readonly', width=40)
        self.modalidade_entrada['values'] = [
            "Percentual do valor total na primeira parcela",
            "Primeira parcela igual às demais (arredonda no final)",
            "Valor específico na primeira parcela"
        ]
        self.modalidade_entrada.grid(row=0, column=1, padx=5, pady=2)
        

        # Garantir que o frame modalidade começa oculto
        print("Ocultando frame modalidade inicialmente")  # Debug
        self.frame_modalidade.grid_remove()
        
        # Frame para valor da entrada (dinâmico baseado na modalidade)
        self.frame_valor_entrada = ttk.Frame(frame_entrada)
        self.frame_valor_entrada.grid(row=2, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        # Ocultar frames inicialmente
        self.frame_modalidade.grid_remove()
        self.frame_valor_entrada.grid_remove()
        
        # Tipo de Despesa
        ttk.Label(frame, text="Tipo de Despesa:").grid(row=1, column=0, padx=5, pady=5)
        self.tipo_despesa = ttk.Combobox(frame, values=['2', '3', '5', '6'], state='readonly', width=5)
        self.tipo_despesa.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        self.tipo_despesa.set('3')  # Tipo 3 como padrão

        # Tipo de Parcelamento
        ttk.Label(frame, text="Tipo de Parcelamento:").grid(row=2, column=0, padx=5, pady=5)
        self.tipo_parcelamento = ttk.Combobox(frame, values=[
            "Prazo Fixo em Dias",
            "Datas Específicas",
            "Cartão de Crédito"
        ], state="readonly")
        self.tipo_parcelamento.grid(row=2, column=1, padx=5, pady=5)
        self.tipo_parcelamento.set("Prazo Fixo em Dias")
        self.tipo_parcelamento.bind('<<ComboboxSelected>>', self.atualizar_campos_parcelamento)

        # Frame para campos dinâmicos
        self.frame_dinamico = ttk.Frame(frame)
        self.frame_dinamico.grid(row=3, column=0, columnspan=2, pady=10)

        # Campos comuns
        ttk.Label(frame, text="Data da Despesa:").grid(row=4, column=0, padx=5, pady=5)
        self.data_despesa = DateEntry(
            frame,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2
        )
        
        self.data_despesa.grid(row=4, column=1, padx=5, pady=5)
        self.data_despesa.configure(state='normal')
        self.configurar_calendario(self.data_despesa)

        ttk.Label(frame, text="Valor Original:").grid(row=5, column=0, padx=5, pady=5)
        self.valor_original = ttk.Entry(frame)
        self.valor_original.grid(row=5, column=1, padx=5, pady=5)

        # Alterar o label do número de parcelas para ser mais claro
        if self.tem_entrada.get():
            ttk.Label(frame, text="Número de Parcelas (além da entrada):").grid(row=6, column=0, padx=5, pady=5)
        else:
            ttk.Label(frame, text="Número de Parcelas:").grid(row=6, column=0, padx=5, pady=5)
        self.num_parcelas = ttk.Entry(frame)
        self.num_parcelas.grid(row=6, column=1, padx=5, pady=5)

        # Adicionar um label informativo
        self.lbl_info_parcelas = ttk.Label(frame, text="")
        self.lbl_info_parcelas.grid(row=7, column=0, columnspan=2, padx=5, pady=2)

        # Frame específico para informação sobre parcelas
        frame_info_parcelas = ttk.Frame(frame)
        frame_info_parcelas.grid(row=7, column=0, columnspan=2, padx=5, pady=5, sticky='ew')
        
        self.lbl_info_parcelas = ttk.Label(
            frame_info_parcelas, 
            text="",
            wraplength=500,  # Permitir quebra de linha se necessário
            justify='center'
        )
        self.lbl_info_parcelas.pack(fill='x', padx=5)

        # Referência Base (já existe)
        ttk.Label(frame, text="Referência Base:").grid(row=8, column=0, padx=5, pady=5)
        self.referencia_base = ttk.Entry(frame)
        self.referencia_base.grid(row=8, column=1, padx=5, pady=5, sticky='ew')

        # Adicionar campo NF
        ttk.Label(frame, text="NF:").grid(row=9, column=0, padx=5, pady=5)
        self.campos_nf = ttk.Entry(frame)
        self.campos_nf.grid(row=9, column=1, padx=5, pady=5, sticky='ew')


        # Adicionar seleção de forma de pagamento (após campo NF)
        ttk.Label(frame, text="Forma de Pagamento:").grid(row=10, column=0, padx=5, pady=5)
        self.forma_pagamento_var = tk.StringVar(value="PIX")
        self.forma_pagamento_combo = ttk.Combobox(
            frame,
            textvariable=self.forma_pagamento_var,
            values=["PIX", "TED"],
            state="readonly",
            width=10
        )
        self.forma_pagamento_combo.grid(row=10, column=1, padx=5, pady=5, sticky='w')

        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.grid(row=11, column=0, columnspan=2, pady=20)

        ttk.Button(frame_botoes, 
                  text="Gerar Parcelas", 
                  command=self.gerar_parcelas).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Cancelar", 
                  command=self.cancelar_parcelamento).pack(side='left', padx=5)

        # Inicializar campos do tipo padrão
        self.atualizar_campos_parcelamento(None)

        # Fazer a janela modal
        self.janela_parcelas.transient(self.parent.root)
        self.janela_parcelas.grab_set()

        # Centralizar a janela
        self.janela_parcelas.update_idletasks()
        width = self.janela_parcelas.winfo_width()
        height = self.janela_parcelas.winfo_height()
        x = (self.janela_parcelas.winfo_screenwidth() // 2) - (width // 2)
        y = (self.janela_parcelas.winfo_screenheight() // 2) - (height // 2)
        self.janela_parcelas.geometry(f'{width}x{height}+{x}+{y}')


    def atualizar_campos_entrada(self):
        """Mostra/oculta campos relacionados à entrada e atualiza labels"""
        if self.tem_entrada.get():
            # Mostrar frame modalidade
            if self.frame_modalidade:
                self.frame_modalidade.grid()
                
                # Criar campos se não existirem
                if not hasattr(self, 'valor_entrada') or not self.valor_entrada:
                    if not self.frame_valor_entrada:
                        self.frame_valor_entrada = ttk.Frame(self.frame_modalidade)
                        self.frame_valor_entrada.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
                    
                    self.lbl_entrada = ttk.Label(self.frame_valor_entrada, text="Valor:")
                    self.lbl_entrada.grid(row=0, column=0, padx=5, pady=2)
                    
                    self.valor_entrada = ttk.Entry(self.frame_valor_entrada)
                    self.valor_entrada.grid(row=0, column=1, padx=5, pady=2)
                
                if self.frame_valor_entrada:
                    self.frame_valor_entrada.grid()
        else:
            # Ocultar frames
            if self.frame_modalidade:
                self.frame_modalidade.grid_remove()
            if self.frame_valor_entrada:
                self.frame_valor_entrada.grid_remove()
            
            # Restaurar label original
            for widget in self.janela_parcelas.winfo_children():
                if isinstance(widget, ttk.Label) and widget.cget("text").startswith("Número de Parcelas"):
                    widget.config(text="Número de Parcelas:")
            self.lbl_info_parcelas.config(text="")

    def atualizar_campos_modalidade(self, event=None):
        """Atualiza campos baseado na modalidade selecionada"""
        modalidade = self.modalidade_entrada.get()
        
        if not hasattr(self, 'frame_valor_entrada') or not hasattr(self, 'lbl_entrada'):
            return
            
        self.frame_valor_entrada.grid()
        
        if modalidade == "Percentual do valor total na primeira parcela":
            self.lbl_entrada.config(text="Percentual (%): ")
            self.valor_entrada.delete(0, tk.END)
        elif modalidade == "Primeira parcela igual às demais (arredonda no final)":
            self.frame_valor_entrada.grid_remove()
        elif modalidade == "Valor específico na primeira parcela":
            self.lbl_entrada.config(text="Valor (R$): ")
            self.valor_entrada.delete(0, tk.END)
            
    def atualizar_campos_parcelamento(self, event):
        # Limpar frame dinâmico
        for widget in self.frame_dinamico.winfo_children():
            widget.destroy()

        tipo = self.tipo_parcelamento.get()
        
        if tipo == "Prazo Fixo em Dias":
            ttk.Label(self.frame_dinamico, text="Prazo entre Parcelas (dias):").grid(row=0, column=0, padx=5, pady=5)
            self.prazo_dias = ttk.Entry(self.frame_dinamico)
            self.prazo_dias.grid(row=0, column=1, padx=5, pady=5)
            self.prazo_dias.insert(0, "30")  # Valor padrão

        elif tipo == "Datas Específicas":
            num_parcelas_txt = "parcelas após a entrada" if self.tem_entrada.get() else "parcelas"
            
            ttk.Label(self.frame_dinamico, 
                     text=f"Informe as datas de vencimento das {num_parcelas_txt}:").grid(
                         row=0, column=0, columnspan=2, padx=5, pady=5)
            
            self.texto_datas = tk.Text(self.frame_dinamico, height=4, width=30)
            self.texto_datas.grid(row=1, column=0, columnspan=2, padx=5, pady=5)
            
            ttk.Label(self.frame_dinamico, 
                     text="Digite uma data por linha no formato dd/mm/aaaa\n"
                          "(não inclua a data da entrada)").grid(
                         row=2, column=0, columnspan=2, padx=5, pady=5)

        elif tipo == "Cartão de Crédito":
            ttk.Label(self.frame_dinamico, text="Dia do Vencimento:").grid(row=0, column=0, padx=5, pady=5)
            self.dia_vencimento = ttk.Entry(self.frame_dinamico, width=5)
            self.dia_vencimento.grid(row=0, column=1, padx=5, pady=5)
            self.dia_vencimento.insert(0, "10")  # Valor padrão


    # Métodos de geração e validação de parcelas
    def validar_dados_entrada(self, valor_original, num_parcelas, referencia_base, tipo):
        """Valida os dados básicos antes de gerar parcelas"""
        if not referencia_base or num_parcelas <= 0:
            custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
            return False

        # Validações específicas por tipo de parcelamento
        if tipo == "Prazo Fixo em Dias":
            if not hasattr(self, 'prazo_dias') or not self.prazo_dias.get():
                custom_messagebox("error", "Erro", "Informe o prazo entre as parcelas!")
                return False
        elif tipo == "Datas Específicas":
            if not hasattr(self, 'texto_datas'):
                custom_messagebox("error", "Erro", "Configure as datas específicas!")
                return False
        elif tipo == "Cartão de Crédito":
            if not hasattr(self, 'dia_vencimento') or not self.dia_vencimento.get():
                custom_messagebox("error", "Erro", "Informe o dia do vencimento!")
                return False
            try:
                dia_vencimento = int(self.dia_vencimento.get())
                if not (1 <= dia_vencimento <= 31):
                    custom_messagebox("error", "Erro", "Dia de vencimento deve estar entre 1 e 31!")
                    return False
            except ValueError:
                custom_messagebox("error", "Erro", "Dia de vencimento inválido!")
                return False

        return True

    def gerar_parcelas(self):
        """Método principal para gerar parcelas"""
        try:
            # Coletar dados básicos
            self.tipo_despesa_valor = self.tipo_despesa.get()
            valor_original = float(self.valor_original.get().replace(',', '.'))
            num_parcelas = int(self.num_parcelas.get())
            referencia_base = self.referencia_base.get().strip()
            nf = self.campos_nf.get().strip()

            tipo = self.tipo_parcelamento.get()

            # Atualizar dados bancários com base na forma de pagamento
            fornecedor_completo = self.parent.buscar_fornecedor_completo(
                self.parent.campos_fornecedor['cnpj_cpf'].get()
            )
            if fornecedor_completo:
                forma_pagamento = self.forma_pagamento_var.get()
                if forma_pagamento == "PIX" and fornecedor_completo['chave_pix']:
                    dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
                else:
                    # Construir dados para TED
                    partes_dados = []
                    if fornecedor_completo['banco']: partes_dados.append(fornecedor_completo['banco'])
                    if fornecedor_completo['op']: partes_dados.append(fornecedor_completo['op'])
                    if fornecedor_completo['agencia']: partes_dados.append(fornecedor_completo['agencia'])
                    if fornecedor_completo['conta']: partes_dados.append(fornecedor_completo['conta'])
                    
                    # SEMPRE adicionar CNPJ/CPF para TED, independente da forma selecionada
                    partes_dados.append(fornecedor_completo['cnpj_cpf'])
                    
                    dados_bancarios = ' - '.join(partes_dados)
                    
                    if not dados_bancarios.strip():
                        dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
                
                # Armazenar para uso nas parcelas
                self.dados_bancarios = dados_bancarios

            # Validar dados
            if not self.validar_dados_entrada(valor_original, num_parcelas, referencia_base, tipo):
                return False

            # Data base é a data da despesa
            data_base = datetime.strptime(self.data_despesa.get(), '%d/%m/%Y')
            
            # Limpar lista de parcelas anterior
            self.parcelas = []

            # Calcular valores das parcelas
            valores_parcelas = self.calcular_valores_parcelas(valor_original, num_parcelas)
            if not valores_parcelas:
                return False

            # Gerar parcelas conforme o tipo
            if tipo == "Prazo Fixo em Dias":
                self.gerar_parcelas_prazo_fixo(data_base, valores_parcelas, referencia_base, num_parcelas, nf)
            elif tipo == "Datas Específicas":
                self.gerar_parcelas_datas_especificas(data_base, valores_parcelas, referencia_base, num_parcelas, nf)
            elif tipo == "Cartão de Crédito":
                self.gerar_parcelas_cartao(data_base, valores_parcelas, referencia_base, num_parcelas, nf)

            if self.parcelas:
                custom_messagebox("info", "Sucesso", f"{len(self.parcelas)} parcela(s) gerada(s) com sucesso!")
                self.limpar_campos()
                return True

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao gerar parcelas: {str(e)}")
            return False

    def adicionar_parcela(self, data_rel, dt_vencto, valor_parcela, referencia_base, i, num_parcelas, eh_primeira_parcela, nf):
        """Método auxiliar para criar uma parcela com todos os dados necessários"""
        parcela = {
            'data_rel': data_rel.strftime('%d/%m/%Y'),
            'dt_vencto': dt_vencto.strftime('%d/%m/%Y'),
            'valor': valor_parcela,
            'referencia': self.gerar_referencia_parcela(referencia_base, i, num_parcelas, eh_primeira_parcela),
            'nf': nf,
            'forma_pagamento': self.forma_pagamento_var.get()  # Adicionado forma de pagamento
        }
        self.parcelas.append(parcela)
        

    def gerar_parcelas_prazo_fixo(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas com prazo fixo em dias"""
        prazo_dias = int(self.prazo_dias.get())
        
        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            if eh_primeira_parcela and self.tem_entrada.get():
                dt_vencto = data_base
                data_rel = self.calcular_data_rel(data_base, dt_vencto, True)
            else:
                dt_vencto = data_base + relativedelta(days=prazo_dias * (i + (0 if self.tem_entrada.get() else 1)))
                dt_vencto = self.proximo_dia_util(dt_vencto)
                data_rel = self.calcular_data_rel(data_base, dt_vencto, eh_primeira_parcela)
            
            self.adicionar_parcela(
                data_rel,
                dt_vencto,
                valor_parcela,
                referencia_base,
                i,
                num_parcelas,
                eh_primeira_parcela,
                nf
            )

    def gerar_parcelas_datas_especificas(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas com datas específicas"""
        datas_texto = self.texto_datas.get("1.0", tk.END).strip().split('\n')
        datas_texto = [d.strip() for d in datas_texto if d.strip()]
        
        num_datas_esperado = num_parcelas
        if len(datas_texto) != num_datas_esperado:
            custom_messagebox("error", 
                "Erro", 
                f"Para {num_parcelas} {'parcelas após a entrada' if self.tem_entrada.get() else 'parcelas'}, "
                f"é necessário informar {num_datas_esperado} data(s) de vencimento."
            )
            return

        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            try:
                if eh_primeira_parcela and self.tem_entrada.get():
                    dt_vencto = data_base
                    data_rel = self.calcular_data_rel(data_base, dt_vencto, True)
                else:
                    idx_data = i - 1 if self.tem_entrada.get() else i
                    if 0 <= idx_data < len(datas_texto):
                        dt_vencto = datetime.strptime(datas_texto[idx_data], '%d/%m/%Y')
                        dt_vencto = self.proximo_dia_util(dt_vencto)
                        data_rel = self.calcular_data_rel(data_base, dt_vencto, eh_primeira_parcela)
                    else:
                        raise ValueError(f"Índice de data inválido: {idx_data}")
                
                self.adicionar_parcela(
                    data_rel,
                    dt_vencto,
                    valor_parcela,
                    referencia_base,
                    i,
                    num_parcelas,
                    eh_primeira_parcela,
                    nf
                )
                
            except ValueError as e:
                custom_messagebox("error", "Erro", f"Erro ao processar data: {str(e)}")
                return
            except IndexError:
                custom_messagebox("error", "Erro", "Número insuficiente de datas fornecidas")
                return

    def gerar_parcelas_cartao(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas para pagamento com cartão"""
        dia_vencimento = int(self.dia_vencimento.get())
        
        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            if eh_primeira_parcela:
                data_atual = data_base + relativedelta(months=1)
            else:
                data_atual = data_base + relativedelta(months=i + 1)
            
            try:
                dt_vencto = data_atual.replace(day=dia_vencimento)
            except ValueError:
                dt_vencto = data_atual + relativedelta(day=31)
            
            dt_vencto = self.proximo_dia_util(dt_vencto)
            
            if eh_primeira_parcela:
                hoje = datetime.now()
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
            else:
                data_rel = self.calcular_data_rel(data_base, dt_vencto, False)
                
            self.adicionar_parcela(
                data_rel,
                dt_vencto,
                valor_parcela,
                referencia_base,
                i,
                num_parcelas,
                eh_primeira_parcela,
                nf
            )


    # Métodos de cálculo e utilitários
    def calcular_valores_parcelas(self, valor_original, num_parcelas):
        """Calcula os valores das parcelas considerando entrada se houver"""
        try:
            if self.tem_entrada.get():
                if not self.modalidade_entrada.get():
                    custom_messagebox("error", "Erro", "Selecione a modalidade de entrada!")
                    return None
                valores_parcelas = self.calcular_parcelas_entrada(valor_original, num_parcelas)
            else:
                valores_parcelas = self.calcular_parcelas_ajustadas(valor_original, num_parcelas)

            # Verificar se a soma está correta
            soma_parcelas = sum(valores_parcelas)
            if abs(soma_parcelas - valor_original) > 0.01:
                custom_messagebox("error", 
                    "Erro", 
                    f"Erro no cálculo das parcelas: soma ({soma_parcelas:.2f}) " 
                    f"diferente do valor original ({valor_original:.2f})!"
                )
                return None

            return valores_parcelas
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao calcular valores: {str(e)}")
            return None

    def calcular_parcelas_entrada(self, valor_total, num_parcelas):
        """Calcula valores das parcelas considerando a modalidade de entrada"""
        modalidade = self.modalidade_entrada.get()
        valores_parcelas = []
        
        # Se tem entrada, o número de parcelas informado é adicional à entrada
        num_parcelas_real = num_parcelas + 1 if self.tem_entrada.get() else num_parcelas
        
        if modalidade == "Percentual do valor total na primeira parcela":
            try:
                percentual = float(self.valor_entrada.get().replace(',', '.'))
                if not (0 < percentual < 100):
                    raise ValueError("Percentual deve estar entre 0 e 100")
                
                valor_entrada = (percentual / 100) * valor_total
                valor_restante = valor_total - valor_entrada
                
                valores_parcelas = [valor_entrada]  # Primeira parcela (entrada)
                # Distribuir o valor restante no número de parcelas informado
                demais_parcelas = self.calcular_parcelas_ajustadas(valor_restante, num_parcelas)
                valores_parcelas.extend(demais_parcelas)
                
            except ValueError as e:
                raise ValueError(f"Erro no percentual de entrada: {str(e)}")
        
        elif modalidade == "Primeira parcela igual às demais (arredonda no final)":
            # Dividir o valor total pelo número total de parcelas (incluindo entrada)
            valores_parcelas = self.calcular_parcelas_ajustadas(valor_total, num_parcelas_real)
            
        elif modalidade == "Valor específico na primeira parcela":
            try:
                valor_entrada = float(self.valor_entrada.get().replace(',', '.'))
                if valor_entrada >= valor_total:
                    raise ValueError("Valor da entrada não pode ser maior ou igual ao valor total")
                
                valor_restante = valor_total - valor_entrada
                valores_parcelas = [valor_entrada]  # Primeira parcela (entrada)
                # Distribuir o valor restante no número de parcelas informado
                demais_parcelas = self.calcular_parcelas_ajustadas(valor_restante, num_parcelas)
                valores_parcelas.extend(demais_parcelas)
                
            except ValueError as e:
                raise ValueError(f"Erro no valor da entrada: {str(e)}")
        
        return valores_parcelas

    def calcular_parcelas_ajustadas(self, valor_total, num_parcelas):
        """Calcula valores das parcelas garantindo que a soma seja igual ao valor total"""
        valor_parcela_base = valor_total / num_parcelas
        valor_parcela_round = round(valor_parcela_base, 2)
        
        # Calcular diferença total devido aos arredondamentos
        diferenca = valor_total - (valor_parcela_round * num_parcelas)
        
        # Distribuir a diferença na última parcela
        parcelas = [valor_parcela_round] * (num_parcelas - 1)
        ultima_parcela = valor_parcela_round + round(diferenca, 2)
        parcelas.append(ultima_parcela)
        
        return parcelas

    def calcular_data_rel(self, data_base, dt_vencto, eh_primeira_parcela):
        """
        Calcula a data do relatório com base na data de vencimento e tipo de despesa.
        Agora considera a data atual para não retroagir em períodos fechados.
        """
        try:
            hoje = datetime.now()
            
            # Se for entrada, calcula a partir da data atual
            if eh_primeira_parcela and self.tem_entrada.get():
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                return data_rel
                
            # Para as demais parcelas, manter a lógica existente
            tp_desp = self.tipo_despesa_valor
            
            if dt_vencto.day == 5:
                # Se vence dia 5, relatório é dia 20 do mês anterior
                data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
            elif dt_vencto.day == 20:
                # Se vence dia 20, relatório é dia 5 do mesmo mês
                data_rel = dt_vencto.replace(day=5)
            elif tp_desp == '5':
                if dt_vencto.day <= 5:
                    data_rel = dt_vencto.replace(day=5)
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=20)
                else:
                    proximo_mes = dt_vencto + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
            else:
                if dt_vencto.day <= 5:
                    data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=5)
                else:
                    data_rel = dt_vencto.replace(day=20)
                    
            # Garantir que a data do relatório não seja anterior à data atual
            if data_rel < hoje:
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                    
            return data_rel
        except Exception as e:
            print(f"Erro ao calcular data do relatório: {str(e)}")
            return dt_vencto

    def configurar_calendario(self, dateentry):
        """Configura o comportamento do calendário"""
        def on_calendar_click(event):
            # Permite cliques no calendário
            return True
            
        def on_calendar_select(event):
            dateentry._top_cal.withdraw()  # Fecha o calendário
            self.janela_parcelas.after(100, lambda: self.janela_parcelas.focus_set())  # Retorna foco
        
        def on_calendar_focus(event):
            # Mantém o foco quando o calendário está aberto
            if dateentry._top_cal:
                dateentry._top_cal.focus_set()
            return True

        # Configurar bindings
        dateentry.bind('<<DateEntrySelected>>', on_calendar_select)
        dateentry.bind('<FocusIn>', on_calendar_focus)
        
        if hasattr(dateentry, '_top_cal'):
            cal = dateentry._top_cal
            cal.bind('<Button-1>', on_calendar_click)
            for w in cal.winfo_children():
                w.bind('<Button-1>', on_calendar_click)

        
    def proximo_dia_util(self, data):
        """
        Ajusta a data para o próximo dia útil se cair em fim de semana ou feriado
        """
        # Lista de feriados nacionais fixos
        feriados_fixos = [
            (1, 1),   # Ano Novo
            (21, 4),  # Tiradentes
            (1, 5),   # Dia do Trabalho
            (7, 9),   # Independência
            (12, 10), # Nossa Senhora
            (2, 11),  # Finados
            (15, 11), # Proclamação da República
            (25, 12), # Natal
        ]

        while True:
            # Verifica se é fim de semana
            if data.weekday() >= 5:  # 5 = Sábado, 6 = Domingo
                data = data + relativedelta(days=1)
                continue

            # Verifica se é feriado fixo
            if (data.day, data.month) in feriados_fixos:
                data = data + relativedelta(days=1)
                continue

            # Se não é fim de semana nem feriado, é dia útil
            break

        return data

    def gerar_referencia_parcela(self, referencia_base, indice, num_parcelas, eh_primeira_parcela):
        """Gera a referência apropriada para a parcela"""
        if eh_primeira_parcela and self.tem_entrada.get():
            return f"{referencia_base} - ENTRADA"
        else:
            if self.tem_entrada.get():
                # Para as parcelas após a entrada
                return f"{referencia_base} - PARC. {indice}/{num_parcelas}"
            else:
                # Para parcelamento sem entrada
                return f"{referencia_base} - PARC. {indice + 1}/{num_parcelas}"
           
    # Métodos de limpeza e finalização
    def limpar_campos(self):
        """Limpa todos os campos após sucesso"""
        # Limpar referências de widgets
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None
        
        # Resetar checkbox
        if self._var_tem_entrada:
            self._var_tem_entrada.set(False)
        
        # Fechar janela
        if self.janela_parcelas:
            self.janela_parcelas.destroy()
            self.janela_parcelas = None

    def cancelar_parcelamento(self):
        """Cancela o parcelamento e limpa todos os campos"""
        self.parcelas = []
        
        # Limpar referências de widgets
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None
        
        # Resetar variável de entrada
        if self._var_tem_entrada:
            self._var_tem_entrada.set(False)
        
        if self.janela_parcelas:
            self.janela_parcelas.destroy()
            self.janela_parcelas = None

            
        


    # Fechando os métodos/classes anteriores
    def run(self):
        """Inicia a execução do sistema"""
        self.root.mainloop()

    def __del__(self):
        """Destrutor da classe"""
        if hasattr(self, 'root'):
            self.root.destroy()


class ImportadorRH:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.pasta_rh = Path(BASE_PATH) / "Planilhas_RH"
        
        # Criar pasta se não existir
        os.makedirs(self.pasta_rh, exist_ok=True)
        
        # Mapeamento de cabeçalhos (original -> sistema)
        self.mapeamento_cabecalhos = {
            'Empresa': 'Cliente',
            'Nome Empregado': 'Nome_Empregado',
            'Valor Líquido': 'Valor_Líquido',
            'Tipo Folha': 'Tipo_Folha',
            'Forma Pagto': 'Forma_Pagamento',
            'Nome Banco': 'Nome_Banco',
            'Agência': 'Agencia',
            'N° Conta': 'Numero_Conta',
            'Dig. Conta': 'Digito_Conta',
            'Conta': 'Numero_Conta',
            'Dígito': 'Digito_Conta'
        }
        
        # Mapeamento para referências
        self.mapeamento_referencias = {
            '13º SALÁRIO': '13º SALÁRIO',
            'ADIANTAMENTO 13º SALÁRIO': '13º SALÁRIO',
            'ADIANTAMENTO': 'SALÁRIO',
            'MENSAL': 'SALÁRIO',
            'FÉRIAS': 'FÉRIAS',
            'RESCISÃO NORMAL': 'RESCISÃO',
            'RESCISÃO': 'RESCISÃO',
            '13 SALÁRIO': '13º SALÁRIO',
            '13': '13º SALÁRIO',
            'FERIAS': 'FÉRIAS'
        }
        
    def selecionar_arquivo(self):
        """Permite ao usuário selecionar um arquivo Excel de RH para importar"""
        arquivo = filedialog.askopenfilename(
            title="Selecione a planilha de RH",
            filetypes=[
                ("Todos os formatos suportados", "*.xlsx *.xls *.csv *.txt *.xlsm *.xlsb"),
                ("Arquivos Excel", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("Arquivos CSV", "*.csv *.txt")
            ],
            initialdir=str(self.pasta_rh)
        )
        
        if not arquivo:
            return None
        
        # Verificar se é formato antigo do Excel (.xls)
        extensao = os.path.splitext(arquivo)[1].lower()
        if extensao == '.xls':
            # Tentar converter automaticamente
            arquivo_convertido = self.converter_xls_para_xlsx(arquivo)
            if arquivo_convertido:
                arquivo = arquivo_convertido
            else:
                # Se a conversão falhou e o usuário cancelou, retornar None
                return None
        
        # Copiar arquivo para a pasta_rh se estiver em outro local
        nome_arquivo = os.path.basename(arquivo)
        destino = self.pasta_rh / nome_arquivo
        
        if Path(arquivo) != destino:
            try:
                import shutil
                shutil.copy2(arquivo, destino)
                print(f"Arquivo copiado para {destino}")
            except Exception as e:
                print(f"Erro ao copiar arquivo: {str(e)}")
                # Continuar usando o arquivo original
                return arquivo
            
        return destino

    def converter_xls_para_xlsx(self, arquivo_origem):
        """
        Converte um arquivo XLS (Excel 97-2003) para XLSX (Excel moderno)
        Returns:
            str: Caminho do arquivo XLSX convertido ou None se falhar
        """
        print(f"Tentando converter arquivo {arquivo_origem} para formato XLSX")
        
        # Verificar se é um arquivo XLS
        extensao = os.path.splitext(arquivo_origem)[1].lower()
        if extensao != '.xls':
            print("Arquivo não é XLS, não precisa converter")
            return arquivo_origem
        
        # Criar nome para arquivo convertido
        nome_base = os.path.splitext(os.path.basename(arquivo_origem))[0]
        arquivo_destino = self.pasta_rh / f"{nome_base}_convertido.xlsx"
        
        # Método 1: Tentar usar Excel via COM (Windows)
        try:
            print("Tentando converter usando Excel via COM...")
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(str(arquivo_origem))
            wb.SaveAs(str(arquivo_destino), FileFormat=51)  # 51 = xlOpenXMLWorkbook (*.xlsx)
            wb.Close()
            excel.Quit()
            
            print(f"Arquivo convertido com sucesso para {arquivo_destino}")
            return arquivo_destino
        except Exception as e:
            print(f"Erro ao converter usando Excel COM: {str(e)}")
        
        # Método 2: Tentar usar pandas para ler e salvar
        try:
            print("Tentando converter usando pandas...")
            df = pd.read_excel(arquivo_origem, engine='xlrd')
            df.to_excel(arquivo_destino, index=False)
            print(f"Arquivo convertido com sucesso para {arquivo_destino} usando pandas")
            return arquivo_destino
        except Exception as e:
            print(f"Erro ao converter usando pandas: {str(e)}")
        
        # Método 3: Tentar usar xlrd + openpyxl
        try:
            print("Tentando converter usando xlrd + openpyxl...")
            import xlrd
            from openpyxl import Workbook
            
            # Abrir arquivo XLS
            xls_wb = xlrd.open_workbook(arquivo_origem)
            xls_sheet = xls_wb.sheet_by_index(0)
            
            # Criar novo arquivo XLSX
            xlsx_wb = Workbook()
            xlsx_sheet = xlsx_wb.active
            
            # Copiar dados
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    xlsx_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
            
            # Salvar arquivo XLSX
            xlsx_wb.save(arquivo_destino)
            print(f"Arquivo convertido com sucesso para {arquivo_destino} usando xlrd + openpyxl")
            return arquivo_destino
        except Exception as e:
            print(f"Erro ao converter usando xlrd + openpyxl: {str(e)}")
        
        # Se todas as tentativas falharem, mostrar mensagem e perguntar ao usuário
        resposta = custom_messagebox("yesno", 
            "Formato Antigo do Excel", 
            "O arquivo selecionado está no formato antigo do Excel (97-2003) e não foi possível convertê-lo automaticamente.\n\n"
            "Sugestão: Abra o arquivo no Excel e salve-o como 'Pasta de Trabalho do Excel' (.xlsx).\n\n"
            "Deseja tentar abrir o arquivo original mesmo assim?"
        )
        
        if resposta:
            return arquivo_origem
        else:
            return None

    def tentar_converter_para_csv(self, arquivo):
        """Tenta converter o arquivo para CSV usando ferramentas externas"""
        try:
            # Perguntar ao usuário se deseja tentar converter
            if not custom_messagebox("yesno", 
                "Problema de Formato", 
                "O arquivo está em um formato difícil de ler. Deseja tentar convertê-lo para CSV?\n\n"
                "Isso pode ajudar a importar arquivos com problemas de compatibilidade."
            ):
                return None
                
            # Criar um nome para o arquivo CSV de saída
            base_nome = os.path.splitext(arquivo)[0]
            csv_arquivo = f"{base_nome}_convertido.csv"
            
            # Verificar se o Excel está instalado no sistema e usar para conversão
            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                
                wb = excel.Workbooks.Open(arquivo)
                # Salvar como CSV
                wb.SaveAs(csv_arquivo, FileFormat=6)  # 6 = CSV
                wb.Close()
                excel.Quit()
                
                custom_messagebox("info", 
                    "Conversão Realizada", 
                    f"Arquivo convertido para CSV: {csv_arquivo}\n\nTentando importar novamente."
                )
                return csv_arquivo
            except Exception as e:
                print(f"Erro ao usar Excel para conversão: {str(e)}")
                
                # Se o Excel falhar, tentar com pandas
                try:
                    # Tentar ler com xlrd, que pode funcionar para versões mais antigas
                    with open(arquivo, 'rb') as file:
                        data = file.read(8)  # Lê os primeiros 8 bytes para verificar a assinatura
                    
                    # Verificar se os primeiros bytes são consistentes com XLS
                    if data[:2] == b'\xd0\xcf':  # Possível assinatura de arquivo XLS
                        import xlrd
                        wb = xlrd.open_workbook(arquivo, formatting_info=False)
                        sheet = wb.sheet_by_index(0)
                        
                        # Extrair dados
                        dados = []
                        for row_idx in range(sheet.nrows):
                            row_data = []
                            for col_idx in range(sheet.ncols):
                                cell_value = sheet.cell_value(row_idx, col_idx)
                                row_data.append(cell_value)
                            dados.append(row_data)
                        
                        # Criar DataFrame e salvar como CSV
                        df = pd.DataFrame(dados)
                        df.to_csv(csv_arquivo, index=False, header=False)
                        return csv_arquivo
                except Exception as xlrd_error:
                    print(f"Erro ao converter manualmente: {str(xlrd_error)}")
                
                custom_messagebox("error", 
                    "Erro na Conversão", 
                    "Não foi possível converter o arquivo para CSV. Tente exportar o arquivo como CSV diretamente do Excel."
                )
                return None
        except Exception as e:
            print(f"Erro geral na conversão: {str(e)}")
            return None
    
    def montar_dados_bancarios(self, row):
        """Monta os dados bancários com base na forma de pagamento"""
        forma_pagto = self.obter_valor_coluna(row, 'Forma Pagto')
        
        # Se for dinheiro, deixar em branco
        if forma_pagto and str(forma_pagto).upper() == "DINHEIRO":
            return ''
            
         # Se for PIX, usar a chave PIX se estiver disponível
        if forma_pagto and str(forma_pagto).upper() == "PIX":
            chave_pix = self.obter_valor_coluna(row, 'Tipo Conta/Chave PIX')
            if chave_pix and not pd.isna(chave_pix):
                # Tratar a chave PIX para remover prefixos desnecessários
                chave_pix_str = str(chave_pix).strip()
                
                # Verificar se contém ":" e tratar os diferentes formatos
                if ":" in chave_pix_str:
                    # Remove prefixos comuns mantendo apenas o que vem depois do ":"
                    prefixos = ["CPF:", "Celular:", "E-mail:", "Email:"]
                    
                    for prefixo in prefixos:
                        if chave_pix_str.upper().startswith(prefixo.upper()):
                            # Extrair apenas o que vem depois do ":"
                            chave_pix_str = chave_pix_str[len(prefixo):].strip()
                            break
                    
                    # Se ainda houver outros formatos com ":" não listados acima
                    if ":" in chave_pix_str and not any(prefixo.upper() in chave_pix_str.upper() for prefixo in prefixos):
                        chave_pix_str = chave_pix_str.split(":", 1)[1].strip()
                
                return f"PIX: {chave_pix_str}"
        
        # Para Crédito CC ou outros casos, montar os dados bancários
        nome_banco = self.obter_valor_coluna(row, 'Nome Banco', '')
        # Substituir CAIXA ECONÔMICA FEDERAL por CAIXA
        nome_banco_normalizado = self.normalizar_texto(str(nome_banco))
        if nome_banco and "CAIXA ECONOMICA FEDERAL" in nome_banco_normalizado:
            nome_banco = "CAIXA"
            
        # Obter e formatar agência com zeros à esquerda
        agencia_raw = self.obter_valor_coluna(row, 'Agência', '')
        
        # Limpar o valor da agência para remover possíveis decimais
        if agencia_raw:
            try:
                # Remover possíveis casas decimais e converter para inteiro
                agencia_clean = str(agencia_raw).split('.')[0]
                # Formatar com zeros à esquerda para garantir 4 dígitos
                agencia = agencia_clean.zfill(4)
            except:
                # Em caso de erro, manter o valor original
                agencia = str(agencia_raw)
        else:
            agencia = ''
        
        # Tentar diferentes nomes para o número da conta
        numero_conta = self.obter_valor_coluna(row, 'N° Conta', '')
        if not numero_conta:
            numero_conta = self.obter_valor_coluna(row, 'Conta', '')
        
        # Tentar diferentes nomes para o dígito da conta
        digito_conta = self.obter_valor_coluna(row, 'Dig. Conta', '')
        if not digito_conta:
            digito_conta = self.obter_valor_coluna(row, 'Dígito', '')
            
        # Formatar conta com dígito
        conta = f"{numero_conta}"
        if digito_conta and not pd.isna(digito_conta):
            conta = f"{numero_conta}-{digito_conta}"
            
        # Obter CPF para incluir nos dados bancários
        cpf = self.obter_valor_coluna(row, 'CPF', '')
        
        # Formatar o CPF se não estiver formatado
        cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
        if len(cpf_limpo) == 11:
            cpf_formatado = f"{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:9]}-{cpf_limpo[9:]}"
        else:
            cpf_formatado = cpf
        
        # Montar dados bancários
        partes = [nome_banco, agencia, conta, cpf_formatado]
        partes_filtradas = [str(p) for p in partes if p and not pd.isna(p) and str(p).strip()]
        
        if partes_filtradas:
            return ' - '.join(partes_filtradas)
        else:
            return 'DADOS BANCÁRIOS NÃO CADASTRADOS'
    
    def obter_valor_coluna(self, row, coluna_original, default=''):
        """Tenta obter o valor de uma coluna com tratamento de erro"""
        try:
            # Verificar se a coluna original existe
            if coluna_original in row:
                valor = row[coluna_original]
                if pd.isna(valor):
                    return default
                return valor
            
            # Verificar se o nome mapeado existe
            coluna_mapeada = self.mapeamento_cabecalhos.get(coluna_original)
            if coluna_mapeada and coluna_mapeada in row:
                valor = row[coluna_mapeada]
                if pd.isna(valor):
                    return default
                return valor
            
            # Verificar se alguma variação do nome da coluna existe (case insensitive)
            for col in row.index:
                if coluna_original.lower() == col.lower():
                    valor = row[col]
                    if pd.isna(valor):
                        return default
                    return valor
                    
            return default
        except:
            return default
    
    def obter_referencia(self, row):
        """Obtém a referência com base no tipo de folha"""
        tipo_folha = self.obter_valor_coluna(row, 'Tipo Folha', 'MENSAL')
        
        # Padronizar o valor (remover acentos, converter para maiúsculo)
        tipo_folha = self.normalizar_texto(tipo_folha)
        
        # Se o tipo_folha estiver no mapeamento, usá-lo
        for key, value in self.mapeamento_referencias.items():
            if self.normalizar_texto(key) == tipo_folha:
                return value
        
        # Caso contrário, usar SALÁRIO como padrão
        return 'SALÁRIO'
    
    def normalizar_texto(self, texto):
        """Normaliza um texto para facilitar comparações"""
        import unicodedata
        import re
        
        # Converter para string
        texto = str(texto).strip().upper()
        
        # Remover acentos
        texto = unicodedata.normalize('NFKD', texto)
        texto = ''.join([c for c in texto if not unicodedata.combining(c)])
        
        # Remover caracteres especiais
        texto = re.sub(r'[^\w\s]', '', texto)
        
        return texto
    
    def importar_dados(self):
        """Importa dados da planilha de RH apenas para o cliente atualmente selecionado"""
        # Verificar se há um cliente selecionado
        if not self.sistema.cliente_atual:
            custom_messagebox("error", 
                "Erro", 
                "Nenhum cliente selecionado. Por favor, selecione um cliente antes de importar dados."
            )
            return
            
        cliente_alvo = self.sistema.cliente_atual.upper()
        
        arquivo = self.selecionar_arquivo()
        if not arquivo:
            return

        # Tentar ler o arquivo
        try:
            # Verificar extensão do arquivo
            extensao = os.path.splitext(arquivo)[1].lower()
            
            # Tentar diferentes métodos de importação baseados na extensão e no conteúdo
            df = None
            erro_leitura = None
            
            # Se for CSV, tentar abrir diretamente
            if extensao in ['.csv', '.txt']:
                # Tentar diferentes delimitadores e encodings
                delimitadores = [',', ';', '\t']
                encodings = ['utf-8', 'latin-1', 'cp1252']
                
                for encoding in encodings:
                    for delim in delimitadores:
                        try:
                            print(f"Tentando abrir CSV com delimitador: {delim}, encoding: {encoding}")
                            df = pd.read_csv(
                                arquivo,
                                dtype={'CPF': str},
                                delimiter=delim,
                                encoding=encoding
                            )
                            print(f"Sucesso ao abrir CSV com delimitador: {delim}, encoding: {encoding}")
                            break
                        except Exception as e:
                            print(f"Erro ao abrir CSV com delimitador {delim}, encoding {encoding}: {str(e)}")
                    
                    if df is not None:
                        break
            
            # Se não for CSV ou não conseguiu abrir, tentar como Excel
            if df is None and extensao in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
                engines = ['openpyxl', 'xlrd']
                
                for engine in engines:
                    try:
                        print(f"Tentando abrir com engine: {engine}")
                        df = pd.read_excel(
                            arquivo,
                            dtype={'CPF': str},  # Forçar leitura do CPF como string
                            engine=engine
                        )
                        print(f"Sucesso ao abrir com engine: {engine}")
                        break
                    except Exception as e:
                        print(f"Erro ao abrir com engine {engine}: {str(e)}")
                        erro_leitura = str(e)
            
            # Se ainda não conseguiu abrir, tentar converter para CSV
            if df is None:
                csv_arquivo = self.tentar_converter_para_csv(arquivo)
                if csv_arquivo:
                    # Tentar ler o CSV convertido
                    delimitadores = [',', ';', '\t']
                    for delim in delimitadores:
                        try:
                            print(f"Tentando abrir o CSV convertido com delimitador: {delim}")
                            df = pd.read_csv(
                                csv_arquivo,
                                dtype={'CPF': str},
                                delimiter=delim,
                                encoding='utf-8'
                            )
                            print(f"Sucesso ao abrir o CSV convertido")
                            break
                        except Exception as e:
                            print(f"Erro ao abrir o CSV convertido com delimitador {delim}: {str(e)}")
                            
                    # Se ainda não funcionou, tentar com encoding latin-1
                    if df is None:
                        for delim in delimitadores:
                            try:
                                df = pd.read_csv(
                                    csv_arquivo,
                                    dtype={'CPF': str},
                                    delimiter=delim,
                                    encoding='latin-1'
                                )
                                print(f"Sucesso ao abrir o CSV convertido com encoding latin-1")
                                break
                            except Exception as e:
                                print(f"Erro ao abrir o CSV convertido com encoding latin-1: {str(e)}")
            if extensao == '.xls' and df is None:
                custom_messagebox("info", 
                    "Sugestão", 
                    "Parece que você está tentando importar um arquivo no formato Excel 97-2003 (.xls).\n\n"
                    "Este formato pode causar problemas de compatibilidade. Sugerimos abrir o arquivo no Excel e "
                    "salvá-lo como 'Pasta de Trabalho do Excel' (.xlsx) antes de importar."
                )

            # Se ainda não conseguiu abrir o arquivo, oferecer opção para selecionar outro
            if df is None:
                if custom_messagebox("yesno", 
                    "Erro de Formato", 
                    "Não foi possível abrir o arquivo. Tente salvar o arquivo como CSV no Excel e importar novamente.\n\nDeseja selecionar outro arquivo?"
                ):
                    return self.importar_dados()  # Reiniciar o processo
                else:
                    return  # Cancelar importação
            
            # Mostrar informações sobre as colunas disponíveis
            print(f"Colunas disponíveis na planilha: {df.columns.tolist()}")
            
            # Pedir ao usuário que confirme o mapeamento de colunas se necessário
            if not any(col in df.columns for col in ['Empresa', 'Cliente']):
                # Nenhuma coluna de cliente encontrada, perguntar ao usuário
                custom_messagebox("info", 
                    "Configuração Manual",
                    "Não foi possível identificar automaticamente a coluna que contém o nome do cliente.\n"
                    "Na próxima tela, selecione a coluna que contém o nome do cliente/empresa."
                )
                
                # Solicitar ao usuário que escolha a coluna para o cliente
                from tkinter import simpledialog
                coluna_cliente = simpledialog.askstring(
                    "Selecionar Coluna", 
                    f"Selecione a coluna que contém o nome do cliente entre as opções:\n\n{', '.join(df.columns.tolist())}"
                )
                
                if coluna_cliente and coluna_cliente in df.columns:
                    # Adicionar ao mapeamento temporariamente
                    self.mapeamento_cabecalhos['Empresa'] = coluna_cliente
                else:
                    custom_messagebox("info",  
                        "Coluna não encontrada",
                        "Coluna selecionada não encontrada ou inválida. Continuando com as colunas padrão."
                    )
            
            registros_processados = 0
            erros = []
            
            # Processar linha por linha, mantendo o cliente atual
            cliente_atual = None
            cliente_encontrado = False
            
            # Determinar o número total de linhas para exibir progresso
            total_linhas = len(df)
            linhas_processadas = 0
            
            for idx, row in df.iterrows():
                linhas_processadas += 1
                
                if linhas_processadas % 10 == 0:
                    print(f"Processando linha {linhas_processadas}/{total_linhas} ({linhas_processadas/total_linhas*100:.1f}%)")
                
                # Verificar se esta linha tem um valor na coluna Empresa/Cliente (possível nome de cliente)
                empresa = self.obter_valor_coluna(row, 'Empresa')
                
                if empresa and not pd.isna(empresa):
                    # Corrigir possíveis espaços extras e converter para maiúsculas
                    empresa_limpa = str(empresa).strip().upper()
                    
                    # Ignorar linhas vazias ou com 'EMPRESA' (cabeçalho)
                    if empresa_limpa and empresa_limpa != 'EMPRESA':
                        cliente_atual = empresa_limpa
                        
                        # Verificar se é o cliente que estamos procurando
                        # Fazer comparação mais flexível, usando apenas parte do nome se necessário
                        if cliente_atual == cliente_alvo:
                            cliente_encontrado = True
                            print(f"Cliente alvo encontrado (match exato): {cliente_atual}")
                        elif cliente_atual in cliente_alvo or cliente_alvo in cliente_atual:
                            cliente_encontrado = True
                            print(f"Cliente alvo encontrado (match parcial): {cliente_atual} ≈ {cliente_alvo}")
                        else:
                            cliente_encontrado = False
                            print(f"Cliente diferente encontrado: {cliente_atual}, ignorando")
                
                # Se não for o cliente alvo, pular esta linha
                if not cliente_encontrado:
                    continue
                    
                # Obter os valores necessários
                nome = self.obter_valor_coluna(row, 'Nome Empregado', '').strip().upper()
                cpf = self.obter_valor_coluna(row, 'CPF', '').strip()
                valor_liquido = self.obter_valor_coluna(row, 'Valor Líquido')
                
                # Verificar se parece ser uma linha de cabeçalho
                if "NOME EMPREGADO" in nome or "FUNCIONARIO" in nome or "FUNCIONÁRIO" in nome:
                    print(f"Ignorando provável linha de cabeçalho: {nome}")
                    continue
                # Tentar outras colunas comuns para nome se não encontrar
                if not nome:
                    nome = self.obter_valor_coluna(row, 'Nome', '').strip().upper()
                    if not nome:
                        nome = self.obter_valor_coluna(row, 'Funcionário', '').strip().upper()
                
                # Tentar outras colunas comuns para valor líquido se não encontrar
                if pd.isna(valor_liquido):
                    valor_liquido = self.obter_valor_coluna(row, 'Líquido')
                    if pd.isna(valor_liquido):
                        valor_liquido = self.obter_valor_coluna(row, 'Valor')
                
                # Verificar se esta linha representa um funcionário (tem nome, CPF e valor)
                if not nome or not cpf or pd.isna(valor_liquido):
                    # Se falta apenas o CPF, mas tem nome e valor, gerar um aviso
                    if nome and not pd.isna(valor_liquido) and not cpf:
                        erros.append(f"Funcionário {nome} sem CPF definido")
                    continue
                    
               # Verificar se o valor é numérico
                try:
                    valor = float(str(valor_liquido).replace(',', '.'))
                except (ValueError, TypeError):
                    erros.append(f"Valor inválido para {nome}: {valor_liquido}")
                    continue
                
                # Definir data do relatório atual
                data_rel = self.sistema.data_rel_entry.get()
                if not data_rel:
                    hoje = datetime.now()
                    if 6 <= hoje.day <= 20:
                        data_rel = hoje.replace(day=20).strftime('%d/%m/%Y')
                    else:
                        if hoje.day > 20:
                            proximo_mes = (hoje.replace(day=1) + relativedelta(months=1))
                            data_rel = proximo_mes.replace(day=5).strftime('%d/%m/%Y')
                        else:
                            data_rel = hoje.replace(day=5).strftime('%d/%m/%Y')
                
                # Calcular data de vencimento
                dt_vencto = data_rel
                
                # Obter referência com base no tipo de folha
                referencia = self.obter_referencia(row)
                
                # Montar dados bancários
                dados_bancarios = self.montar_dados_bancarios(row)
                # Agora verificamos explicitamente se está vazio
                if dados_bancarios == '':
                    dados_bancarios = ''  # Manter vazio quando for pagamento em dinheiro
                elif not dados_bancarios:
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
                
                # Determinar forma de pagamento
                forma_pagto = self.obter_valor_coluna(row, 'Forma Pagto', 'PIX')
                forma_pagamento = 'PIX' if forma_pagto.upper() == 'PIX' else 'TED'
                
                # Formatar o CPF
                cpf_numerico = ''.join(filter(str.isdigit, str(cpf)))
                if len(cpf_numerico) == 11:
                    cpf_formatado = f"{cpf_numerico[:3]}.{cpf_numerico[3:6]}.{cpf_numerico[6:9]}-{cpf_numerico[9:]}"
                else:
                    cpf_formatado = cpf
                
                # Criar registro
                registro = {
                    'data': data_rel,
                    'cnpj_cpf': cpf_formatado,
                    'nome': nome,
                    'categoria': 'MO',
                    'tp_desp': '1',
                    'referencia': referencia,
                    'nf': '',
                    'vr_unit': f"{valor:.2f}",
                    'dias': 1,
                    'valor': f"{valor:.2f}",
                    'dt_vencto': dt_vencto,
                    'dados_bancarios': dados_bancarios,
                    'observacao': f"IMPORTADO RH - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                    'forma_pagamento': forma_pagamento
                }
                
                # Adicionar à lista para incluir
                self.sistema.dados_para_incluir.append(registro)
                registros_processados += 1
            
            # Relatório final
            if registros_processados > 0:
                mensagem = (
                    f"Importação concluída!\n\n"
                    f"Registros processados: {registros_processados} para {cliente_alvo}\n"
                )
                
                if erros:
                    mensagem += f"\nAdvertências ({len(erros)}):\n"
                    for erro in erros[:5]:  # Limitar a 5 para não sobrecarregar
                        mensagem += f"- {erro}\n"
                    if len(erros) > 5:
                        mensagem += f"- ...e mais {len(erros) - 5} advertências\n"
                
                custom_messagebox("info", "Resultado da Importação", mensagem)
                
                # Perguntar se deseja visualizar
                if custom_messagebox("yesno", 
                    "Importação RH", 
                    "Deseja visualizar os lançamentos antes de salvar?"
                ):
                    self.sistema.visualizar_lancamentos()
            else:
                custom_messagebox("warning",  
                    "Aviso",
                    f"Nenhum registro foi processado para o cliente {cliente_alvo}. Verifique se este cliente está presente na planilha."
                )
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao importar dados: {str(e)}")
            import traceback
            traceback.print_exc()

class GestorTaxasAdministracao:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        
    def recalcular_taxas_afetadas(self, data_referencia, cliente=None):
        """Recalcula taxas afetadas por exclusões/alterações"""
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual
                
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            
            # Carregar dados
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")
            
            # Identificar taxas existentes para a data
            taxas_existentes = self.identificar_lancamentos_taxa_admin(
                df[df['DATA_REL'] == pd.to_datetime(data_referencia)]
            )
            
            if taxas_existentes.empty:
                return {"sucesso": True, "mensagem": "Nenhuma taxa encontrada para recalcular"}
            
            # Calcular nova base (sem excluídos)
            nova_base = self.calcular_base_calculo_taxa(df, data_referencia, incluir_excluidos=False)
            
            # Buscar percentual da taxa no cadastro do cliente
            percentual_taxa = self.obter_percentual_taxa_cliente(cliente)
            
            if percentual_taxa is None:
                return {"sucesso": False, "mensagem": "Percentual de taxa não configurado para o cliente"}
            
            # Calcular novo valor da taxa
            novo_valor_taxa = nova_base * (percentual_taxa / 100)
            
            # Atualizar na planilha
            resultado = self.atualizar_taxas_na_planilha(
                arquivo_cliente, taxas_existentes, novo_valor_taxa, nova_base
            )
            
            return resultado
            
        except Exception as e:
            return {"sucesso": False, "mensagem": f"Erro ao recalcular taxas: {str(e)}"}
    
    def obter_percentual_taxa_cliente(self, cliente):
        """Obtém o percentual de taxa configurado para o cliente"""
        try:
            wb_clientes = load_workbook(ARQUIVO_CLIENTES)
            ws = wb_clientes['Clientes']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == cliente:  # Nome do cliente
                    # Assumindo que a taxa está na coluna 6 (ajustar conforme sua estrutura)
                    return float(row[6]) if row[6] else None
            
            return None
            
        except Exception as e:
            print(f"Erro ao obter percentual da taxa: {str(e)}")
            return None
    
    def atualizar_taxas_na_planilha(self, arquivo_cliente, taxas_existentes, novo_valor, nova_base):
        """Atualiza os valores das taxas na planilha"""
        try:
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            taxas_atualizadas = []
            
            # Procurar e atualizar cada taxa
            for _, taxa in taxas_existentes.iterrows():
                id_taxa = taxa.get('ID_LANCAMENTO')
                if pd.isna(id_taxa):
                    continue
                
                # Encontrar linha na planilha
                for row_num in range(2, ws.max_row + 1):
                    if ws.cell(row=row_num, column=15).value == id_taxa:  # Coluna ID_LANCAMENTO
                        valor_antigo = ws.cell(row=row_num, column=9).value  # VALOR
                        
                        # Calcular valor proporcional se houver múltiplas taxas
                        if len(taxas_existentes) > 1:
                            proporcao = float(valor_antigo) / taxas_existentes['VALOR'].astype(float).sum()
                            valor_proporcional = novo_valor * proporcao
                        else:
                            valor_proporcional = novo_valor
                        
                        # Atualizar valor
                        ws.cell(row=row_num, column=9, value=float(valor_proporcional))
                        
                        # Atualizar observação com histórico
                        obs_atual = ws.cell(row=row_num, column=13).value or ""
                        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M')
                        nova_obs = f"{obs_atual} - RECALCULADA EM {timestamp} (Base: R$ {nova_base:,.2f})"
                        ws.cell(row=row_num, column=13, value=nova_obs)
                        
                        taxas_atualizadas.append({
                            'id': id_taxa,
                            'valor_antigo': valor_antigo,
                            'valor_novo': valor_proporcional
                        })
                        break
            
            wb.save(arquivo_cliente)
            
            return {
                "sucesso": True,
                "mensagem": f"Taxas recalculadas: {len(taxas_atualizadas)} itens atualizados",
                "detalhes": taxas_atualizadas,
                "nova_base": nova_base,
                "novo_valor_total": novo_valor
            }
            
        except Exception as e:
            return {"sucesso": False, "mensagem": f"Erro ao atualizar planilha: {str(e)}"}

class GerenciadorLancamentos:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.janela = None
        self.tree_lancamentos = None
        self.dados_originais = []
        
    def abrir_gerenciador(self):
        """Abre a janela de gerenciamento de lançamentos"""
        if not self.sistema.cliente_atual:
            custom_messagebox("error", "Erro", "Selecione um cliente primeiro!")
            return
            
        self.janela = tk.Toplevel(self.sistema.root)
        self.janela.title(f"Gerenciar Lançamentos - {self.sistema.cliente_atual}")
        self.janela.geometry("1200x700")
        
        # Configurar janela
        self.janela.transient(self.sistema.root)
        self.janela.grab_set()
        
        self.criar_interface()
        self.carregar_lancamentos()
        
    def criar_interface(self):
        """Cria a interface do gerenciador"""
        # Frame principal
        main_frame = ttk.Frame(self.janela, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Frame de filtros
        frame_filtros = ttk.LabelFrame(main_frame, text="Filtros")
        frame_filtros.pack(fill='x', pady=(0, 10))
        
        # Filtros por data
        ttk.Label(frame_filtros, text="Data Início:").grid(row=0, column=0, padx=5, pady=5)
        self.data_inicio = DateEntry(frame_filtros, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_inicio.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(frame_filtros, text="Data Fim:").grid(row=0, column=2, padx=5, pady=5)
        self.data_fim = DateEntry(frame_filtros, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_fim.grid(row=0, column=3, padx=5, pady=5)
        
        # Filtro por status
        ttk.Label(frame_filtros, text="Status:").grid(row=0, column=4, padx=5, pady=5)
        self.combo_status = ttk.Combobox(frame_filtros, values=['Todos', 'Ativos', 'Excluídos'], 
                                       state='readonly', width=10)
        self.combo_status.set('Ativos')
        self.combo_status.grid(row=0, column=5, padx=5, pady=5)
        
        # Botão filtrar
        ttk.Button(frame_filtros, text="Filtrar", 
                  command=self.aplicar_filtros).grid(row=0, column=6, padx=10, pady=5)
        
        # Frame da lista de lançamentos
        frame_lista = ttk.Frame(main_frame)
        frame_lista.pack(fill='both', expand=True)
        
        # Treeview para lançamentos
        colunas = ('Data', 'Tipo', 'Nome', 'Referência', 'Valor', 'Vencimento', 'Status', 'ID')
        self.tree_lancamentos = ttk.Treeview(frame_lista, columns=colunas, show='headings', height=20)
        
        # Configurar cabeçalhos
        for col in colunas:
            self.tree_lancamentos.heading(col, text=col)
            if col == 'ID':
                self.tree_lancamentos.column(col, width=0, stretch=False)  # Ocultar coluna ID
            elif col in ['Data', 'Vencimento']:
                self.tree_lancamentos.column(col, width=100)
            elif col == 'Valor':
                self.tree_lancamentos.column(col, width=100, anchor='e')
            elif col == 'Status':
                self.tree_lancamentos.column(col, width=80)
            else:
                self.tree_lancamentos.column(col, width=150)
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_lancamentos.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient='horizontal', command=self.tree_lancamentos.xview)
        self.tree_lancamentos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionar elementos
        self.tree_lancamentos.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Frame de botões (existente)
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=(10, 0))
        
        ttk.Button(frame_botoes, text="Editar", command=self.editar_lancamento).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Excluir", command=self.excluir_lancamento).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Restaurar", command=self.restaurar_lancamento).pack(side='left', padx=5)
        
        # ADICIONE ESTA LINHA:
        ttk.Button(frame_botoes, text="Recalcular Taxas", 
                command=self.recalcular_taxas_manual).pack(side='left', padx=5)
                
        ttk.Button(frame_botoes, text="Atualizar", command=self.carregar_lancamentos).pack(side='left', padx=20)
        ttk.Button(frame_botoes, text="Fechar", command=self.janela.destroy).pack(side='right', padx=5)
        
        # Configurar tags para cores
        self.tree_lancamentos.tag_configure('excluido', background='#ffcccc')
        self.tree_lancamentos.tag_configure('normal', background='white')
        
    def carregar_lancamentos(self):
        """Carrega os lançamentos da planilha"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            
            # Carregar dados
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")
            
            # Adicionar coluna de status se não existir
            if 'STATUS' not in df.columns:
                df['STATUS'] = 'ATIVO'
            
            # CORREÇÃO: Preencher status em branco com 'ATIVO'
            df['STATUS'] = df['STATUS'].replace('', 'ATIVO')
            df['STATUS'] = df['STATUS'].fillna('ATIVO')
            
            # Adicionar ID único se não existir
            if 'ID_LANCAMENTO' not in df.columns:
                df['ID_LANCAMENTO'] = range(1, len(df) + 1)
            
            # CORREÇÃO: Preencher IDs em branco
            mask_id_vazio = (df['ID_LANCAMENTO'].isna()) | (df['ID_LANCAMENTO'] == '')
            df.loc[mask_id_vazio, 'ID_LANCAMENTO'] = range(1, mask_id_vazio.sum() + 1)
            
            # Salvar dados originais
            self.dados_originais = df.copy()
            
            # Salvar as correções na planilha se necessário
            if 'STATUS' not in pd.read_excel(arquivo_cliente, sheet_name='Dados').columns or \
            df['STATUS'].isna().any() or (df['STATUS'] == '').any():
                self.corrigir_planilha_status(arquivo_cliente, df)
            
            # Limpar tree
            for item in self.tree_lancamentos.get_children():
                self.tree_lancamentos.delete(item)
            
            # Preencher tree
            for idx, row in df.iterrows():
                status = row.get('STATUS', 'ATIVO')
                if status == '' or pd.isna(status):
                    status = 'ATIVO'
                    
                tag = 'excluido' if status == 'EXCLUIDO' else 'normal'
                
                # Formatar valores
                data_rel = self.formatar_data(row['DATA_REL'])
                data_vencto = self.formatar_data(row['DT_VENCTO'])
                valor = self.formatar_valor(row['VALOR'])
                
                self.tree_lancamentos.insert('', 'end', 
                    values=(data_rel, row['TP_DESP'], row['NOME'], 
                        row['REFERÊNCIA'], valor, data_vencto, status, row.get('ID_LANCAMENTO', idx)),
                    tags=(tag,))
            
            # Aplicar filtros se existirem
            self.aplicar_filtros()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar lançamentos: {str(e)}")

    def corrigir_planilha_status(self, arquivo_cliente, df):
        """Corrige o status na planilha para dados antigos"""
        try:
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            # Adicionar cabeçalho STATUS se não existir
            if ws.cell(row=1, column=14).value != 'STATUS':
                ws.cell(row=1, column=14, value='STATUS')
            
            # Adicionar cabeçalho ID_LANCAMENTO se não existir
            if ws.cell(row=1, column=15).value != 'ID_LANCAMENTO':
                ws.cell(row=1, column=15, value='ID_LANCAMENTO')
            
            # Adicionar cabeçalho ID_LANCAMENTO se não existir
            if ws.cell(row=1, column=16).value != 'HISTORICO_ALTERACAO':
                ws.cell(row=1, column=16, value='HISTORICO_ALTERACAO')
            
            # Preencher STATUS e ID para todas as linhas
            for idx, (_, row) in enumerate(df.iterrows(), start=2):
                # STATUS
                status_atual = ws.cell(row=idx, column=14).value
                if not status_atual or status_atual == '':
                    ws.cell(row=idx, column=14, value='ATIVO')
                
                # ID_LANCAMENTO
                id_atual = ws.cell(row=idx, column=15).value
                if not id_atual or id_atual == '':
                    ws.cell(row=idx, column=15, value=int(idx - 1))
            
            wb.save(arquivo_cliente)
            print("STATUS e IDs corrigidos na planilha")
            
        except Exception as e:
            print(f"Erro ao corrigir status na planilha: {str(e)}")
    
    def aplicar_filtros(self):
        """Aplica os filtros selecionados"""
        try:
            # Obter filtros
            status_filtro = self.combo_status.get()
            data_inicio = self.data_inicio.get_date()
            data_fim = self.data_fim.get_date()
            
            print(f"DEBUG: Aplicando filtros - Status: {status_filtro}, Data início: {data_inicio}, Data fim: {data_fim}")
            
            # Filtrar itens na tree
            for item in self.tree_lancamentos.get_children():
                valores = self.tree_lancamentos.item(item, 'values')
                
                # Obter dados da linha
                data_rel_str = valores[0]  # Data
                status_item = valores[6]   # Status
                
                mostrar = True
                
                # Filtro por status
                if status_filtro == 'Ativos' and status_item != 'ATIVO':
                    mostrar = False
                elif status_filtro == 'Excluídos' and status_item != 'EXCLUIDO':
                    mostrar = False
                
                # Filtro por data
                if mostrar and data_rel_str:
                    try:
                        # Converter data da string para datetime.date
                        data_rel = datetime.strptime(data_rel_str, '%d/%m/%Y').date()
                        
                        # Verificar se está no intervalo
                        if data_rel < data_inicio or data_rel > data_fim:
                            mostrar = False
                            
                    except Exception as e:
                        print(f"Erro ao processar data {data_rel_str}: {str(e)}")
                        # Em caso de erro na data, não filtrar por data
                
                # Mostrar/ocultar item
                if mostrar:
                    # Verificar se o item já está visível
                    try:
                        self.tree_lancamentos.item(item)  # Testa se está visível
                    except:
                        # Se não está visível, reattach
                        self.tree_lancamentos.reattach(item, '', tk.END)
                else:
                    # Ocultar item
                    self.tree_lancamentos.detach(item)
                        
        except Exception as e:
            print(f"Erro ao aplicar filtros: {str(e)}")
            custom_messagebox("error", "Erro", f"Erro ao aplicar filtros: {str(e)}")
    
    def editar_lancamento(self):
        """Abre editor para o lançamento selecionado"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para editar")
            return
        
        valores = self.tree_lancamentos.item(item_selecionado[0])['values']
        id_lancamento = valores[7]  # ID do lançamento
        
        # Buscar dados completos
        lancamento = self.dados_originais[self.dados_originais['ID_LANCAMENTO'] == id_lancamento].iloc[0]
        
        # Abrir editor
        editor = EditorLancamentoCompleto(self.janela, lancamento, self.salvar_edicao)
    
    def excluir_lancamento(self):
        """Marca lançamento como excluído"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para excluir")
            return
        
        if not custom_messagebox("yesno", "Confirmação", 
                                "Deseja realmente excluir este lançamento?\n\n"
                                "ATENÇÃO: As taxas de administração serão recalculadas automaticamente."):
            return
        
        try:
            valores = self.tree_lancamentos.item(item_selecionado[0])['values']
            id_lancamento = valores[7]
            data_lancamento = valores[0]  # Data do lançamento
            
            # Atualizar status na planilha
            self.atualizar_status_lancamento(id_lancamento, 'EXCLUIDO')

            # Recalcular taxas para a data do lançamento
            resultado_recalculo = self.gestor_taxas.recalcular_taxas_afetadas(
                datetime.strptime(data_lancamento, '%d/%m/%Y').date()
            )
            
            # Recarregar lista
            self.carregar_lancamentos()
            
            # Mostrar resultado
            mensagem = "Lançamento excluído com sucesso!"
            if resultado_recalculo["sucesso"]:
                if "taxas recalculadas" in resultado_recalculo["mensagem"]:
                    mensagem += f"\n\n{resultado_recalculo['mensagem']}"
                    mensagem += f"\nNova base de cálculo: R$ {resultado_recalculo.get('nova_base', 0):,.2f}"
                else:
                    mensagem += f"\n{resultado_recalculo['mensagem']}"
            else:
                mensagem += f"\n\nAVISO: {resultado_recalculo['mensagem']}"
            
            custom_messagebox("info", "Sucesso", mensagem)
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao excluir lançamento: {str(e)}")
    
    def restaurar_lancamento(self):
        """Restaura lançamento excluído"""
        item_selecionado = self.tree_lancamentos.selection()
        if not item_selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um lançamento para restaurar")
            return
        
        valores = self.tree_lancamentos.item(item_selecionado[0])['values']
        if valores[6] != 'EXCLUIDO':  # Status
            custom_messagebox("info", "Informação", "Este lançamento já está ativo")
            return
        
        try:
            id_lancamento = valores[7]
            data_lancamento = valores[0]
            
            # Atualizar status na planilha
            self.atualizar_status_lancamento(id_lancamento, 'ATIVO')

            # Recalcular taxas
            resultado_recalculo = self.gestor_taxas.recalcular_taxas_afetadas(
                datetime.strptime(data_lancamento, '%d/%m/%Y').date()
            )
            
            # Recarregar lista
            self.carregar_lancamentos()
            
            # Mostrar resultado
            mensagem = "Lançamento restaurado com sucesso!"
            if resultado_recalculo["sucesso"] and "taxas recalculadas" in resultado_recalculo["mensagem"]:
                mensagem += f"\n\n{resultado_recalculo['mensagem']}"
                mensagem += f"\nNova base de cálculo: R$ {resultado_recalculo.get('nova_base', 0):,.2f}"
            
            custom_messagebox("info", "Sucesso", mensagem)
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao restaurar lançamento: {str(e)}")
    
    def identificar_lancamentos_taxa_admin(self, df):
        """Identifica lançamentos de taxa de administração"""
        # Padrões comuns para identificar taxas de administração
        padroes_taxa = [
            'TAXA DE ADMINISTRAÇÃO',
            'ADM',
            'ADMINISTRAÇÃO', 
            'TAXA ADM',
            'TX ADM',
            'PERCENTUAL ADM'
        ]
        
        mask_taxa = df['REFERÊNCIA'].str.contains('|'.join(padroes_taxa), case=False, na=False)
        return df[mask_taxa].copy()

    def calcular_base_calculo_taxa(self, df, data_relatorio, incluir_excluidos=False):
        """Calcula a base para cálculo da taxa de administração"""
        # Filtrar dados para a quinzena
        df_quinzena = df[df['DATA_REL'] == pd.to_datetime(data_relatorio)].copy()
        
        # Se não incluir excluídos, filtrar apenas ativos
        if not incluir_excluidos:
            df_quinzena = df_quinzena[df_quinzena.get('STATUS', 'ATIVO') != 'EXCLUIDO']
        
        # Excluir as próprias taxas do cálculo para evitar recursão
        df_base = df_quinzena[~df_quinzena['REFERÊNCIA'].str.contains(
            'TAXA|ADM|ADMINISTRAÇÃO', case=False, na=False
        )].copy()
        
        # Converter valores para numérico
        df_base['VALOR_NUM'] = pd.to_numeric(
            df_base['VALOR'].astype(str).str.replace('R$', '').str.replace(',', '.'),
            errors='coerce'
        ).fillna(0)
        
        return df_base['VALOR_NUM'].sum()

    def salvar_edicao(self, id_lancamento, dados_editados):
        """Salva as edições do lançamento"""
        try:
            print(f"DEBUG: Tentando salvar ID {id_lancamento}")
            print(f"DEBUG: Dados editados: {dados_editados}")
        
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            
            # Carregar workbook
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']

            print(f"DEBUG: Procurando linha com ID {id_lancamento}")
                    
            # Encontrar linha do lançamento
            linha_encontrada = False
            for row in range(2, ws.max_row + 1):
                # CORREÇÃO: A coluna ID_LANCAMENTO está na coluna 15, não na última
                id_na_planilha = ws.cell(row=row, column=15).value  # Coluna 15 = ID_LANCAMENTO
                print(f"DEBUG: Linha {row}, ID na planilha: {id_na_planilha}")
                
                if id_na_planilha == id_lancamento:  
                    print(f"DEBUG: Encontrou linha {row} para o ID {id_lancamento}")
                    linha_encontrada = True
                    
                    # Atualizar dados
                    ws.cell(row=row, column=1, value=datetime.strptime(dados_editados['data'], '%d/%m/%Y'))
                    ws.cell(row=row, column=2, value=int(float(dados_editados['tp_desp'])))
                    ws.cell(row=row, column=3, value=dados_editados['cnpj_cpf'])
                    ws.cell(row=row, column=4, value=dados_editados['nome'])
                    ws.cell(row=row, column=5, value=dados_editados['referencia'])
                    ws.cell(row=row, column=6, value=dados_editados['nf'])
                    ws.cell(row=row, column=7, value=float(dados_editados['vr_unit']))
                    ws.cell(row=row, column=8, value=int(dados_editados['dias']))
                    ws.cell(row=row, column=9, value=float(dados_editados['valor']))
                    ws.cell(row=row, column=10, value=datetime.strptime(dados_editados['dt_vencto'], '%d/%m/%Y'))
                    ws.cell(row=row, column=11, value=dados_editados['categoria'])
                    ws.cell(row=row, column=12, value=dados_editados['dados_bancarios'])
                    
                    # Adicionar timestamp de edição
                    timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    obs_atual = dados_editados['observacao']
                    if 'EDITADO EM:' not in obs_atual:
                        nova_obs = f"{obs_atual} - EDITADO EM: {timestamp}"
                        ws.cell(row=row, column=13, value=nova_obs)
                    else:
                        ws.cell(row=row, column=13, value=obs_atual)
                    
                    print(f"DEBUG: Dados atualizados na linha {row}")
                    break
            
            if not linha_encontrada:
                print(f"DEBUG: ERRO - Linha com ID {id_lancamento} não encontrada!")
                return False
            
            wb.save(arquivo_cliente)
            print("DEBUG: Arquivo salvo com sucesso")
            
            # Recarregar lista
            self.carregar_lancamentos()
            
            return True
            
        except Exception as e:
            print(f"DEBUG: Erro ao salvar edição: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def atualizar_status_lancamento(self, id_lancamento, novo_status):
        """Atualiza o status de um lançamento específico"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
            
            # Carregar workbook
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']
            
            # Verificar se coluna STATUS existe, se não, criar
            if ws.cell(row=1, column=14).value != 'STATUS':
                ws.cell(row=1, column=14, value='STATUS')
            
            # Verificar se coluna ID_LANCAMENTO existe, se não, criar
            if ws.cell(row=1, column=15).value != 'ID_LANCAMENTO':
                ws.cell(row=1, column=15, value='ID_LANCAMENTO')
                # Adicionar IDs para linhas existentes
                for row in range(2, ws.max_row + 1):
                    if not ws.cell(row=row, column=15).value:
                        ws.cell(row=row, column=15, value=int(row - 1))
            
            # Encontrar e atualizar linha
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=15).value == id_lancamento:
                    ws.cell(row=row, column=14, value=novo_status)
                    # Adicionar timestamp
                    timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    if novo_status == 'EXCLUIDO':
                        ws.cell(row=row, column=16, value=f'EXCLUIDO EM: {timestamp}')
                    break
            
            wb.save(arquivo_cliente)
            
        except Exception as e:
            raise Exception(f"Erro ao atualizar status: {str(e)}")
    
    def formatar_data(self, data):
        """Formata data para exibição"""
        if pd.isna(data) or data == "":
            return ""
        try:
            if isinstance(data, str):
                return data
            return data.strftime('%d/%m/%Y')
        except:
            return str(data)
    
    def formatar_valor(self, valor):
        """Formata valor para exibição"""
        try:
            if isinstance(valor, str):
                valor = float(valor.replace(',', '.'))
            return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return str(valor)


class EditorLancamentoCompleto:
    def __init__(self, parent, lancamento, callback_salvar):
        self.janela = tk.Toplevel(parent)
        self.janela.title("Editar Lançamento")
        self.janela.geometry("700x600")
        self.lancamento = lancamento
        self.callback_salvar = callback_salvar
        
        # Configurar janela
        self.janela.transient(parent)
        self.janela.grab_set()
        
        self.criar_interface()
        self.preencher_dados()
        
    def criar_interface(self):
        """Cria a interface do editor"""
        # Frame principal com scroll
        main_frame = ttk.Frame(self.janela, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Dados básicos
        frame_basicos = ttk.LabelFrame(main_frame, text="Dados Básicos")
        frame_basicos.pack(fill='x', pady=5)
        
        # Data do Relatório
        ttk.Label(frame_basicos, text="Data do Relatório:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.data_rel = DateEntry(frame_basicos, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_rel.grid(row=0, column=1, padx=5, pady=2, sticky='w')
        
        # Tipo de Despesa
        ttk.Label(frame_basicos, text="Tipo Despesa:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.tp_desp = ttk.Entry(frame_basicos, width=5)
        self.tp_desp.grid(row=0, column=3, padx=5, pady=2, sticky='w')
        
        # Dados do Fornecedor
        frame_fornecedor = ttk.LabelFrame(main_frame, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', pady=5)
        
        ttk.Label(frame_fornecedor, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.cnpj_cpf = ttk.Entry(frame_fornecedor, width=20)
        self.cnpj_cpf.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_fornecedor, text="Nome:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.nome = ttk.Entry(frame_fornecedor, width=50)
        self.nome.grid(row=1, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_fornecedor, text="Categoria:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.categoria = ttk.Entry(frame_fornecedor, width=10)
        self.categoria.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        
        # Dados da Despesa
        frame_despesa = ttk.LabelFrame(main_frame, text="Dados da Despesa")
        frame_despesa.pack(fill='x', pady=5)
        
        ttk.Label(frame_despesa, text="Referência:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.referencia = ttk.Entry(frame_despesa, width=40)
        self.referencia.grid(row=0, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_despesa, text="NF:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.nf = ttk.Entry(frame_despesa, width=15)
        self.nf.grid(row=1, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Valor Unitário:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.vr_unit = ttk.Entry(frame_despesa, width=15)
        self.vr_unit.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Dias:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.dias = ttk.Entry(frame_despesa, width=8)
        self.dias.grid(row=2, column=3, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Valor Total:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.valor = ttk.Entry(frame_despesa, width=15)
        self.valor.grid(row=3, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(frame_despesa, text="Data Vencimento:").grid(row=3, column=2, padx=5, pady=2, sticky='w')
        self.dt_vencto = DateEntry(frame_despesa, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.dt_vencto.grid(row=3, column=3, padx=5, pady=2, sticky='w')
        
        # Dados Bancários e Observação
        ttk.Label(frame_despesa, text="Dados Bancários:").grid(row=4, column=0, padx=5, pady=2, sticky='w')
        self.dados_bancarios = ttk.Entry(frame_despesa, width=50)
        self.dados_bancarios.grid(row=4, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        ttk.Label(frame_despesa, text="Observação:").grid(row=5, column=0, padx=5, pady=2, sticky='w')
        self.observacao = ttk.Entry(frame_despesa, width=50)
        self.observacao.grid(row=5, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        
        # Configurar expansão
        frame_fornecedor.columnconfigure(1, weight=1)
        frame_despesa.columnconfigure(1, weight=1)
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Salvar", command=self.salvar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=self.janela.destroy).pack(side='left', padx=5)
        
        # Bindings para cálculo automático
        self.vr_unit.bind('<KeyRelease>', self.calcular_total)
        self.dias.bind('<KeyRelease>', self.calcular_total)
        
    def preencher_dados(self):
        """Preenche os campos com os dados do lançamento"""
        try:
            # Datas
            if pd.notna(self.lancamento['DATA_REL']):
                self.data_rel.set_date(pd.to_datetime(self.lancamento['DATA_REL']))
            if pd.notna(self.lancamento['DT_VENCTO']):
                self.dt_vencto.set_date(pd.to_datetime(self.lancamento['DT_VENCTO']))
            
            # Campos de texto - VERSÃO CORRIGIDA
            tp_desp_valor = self.lancamento.get('TP_DESP', '')
            if pd.notna(tp_desp_valor):
                self.tp_desp.insert(0, str(int(float(tp_desp_valor))))  # CORRIGIDO
            
            self.cnpj_cpf.insert(0, str(self.lancamento.get('CNPJ_CPF', '')))
            self.nome.insert(0, str(self.lancamento.get('NOME', '')))
            self.categoria.insert(0, str(self.lancamento.get('CATEGORIA', '')))
            self.referencia.insert(0, str(self.lancamento.get('REFERÊNCIA', '')))
            self.nf.insert(0, str(self.lancamento.get('NF', '')))
            
            # Valores numéricos - CORRIGIDOS
            vr_unit_valor = self.lancamento.get('VR_UNIT', '')
            if pd.notna(vr_unit_valor) and vr_unit_valor != '':
                self.vr_unit.insert(0, str(float(vr_unit_valor)).replace('.', ','))
            
            dias_valor = self.lancamento.get('DIAS', '')
            if pd.notna(dias_valor) and dias_valor != '':
                self.dias.insert(0, str(int(float(dias_valor))))
                
            valor_valor = self.lancamento.get('VALOR', '')
            if pd.notna(valor_valor) and valor_valor != '':
                self.valor.insert(0, str(float(valor_valor)).replace('.', ','))
            
            self.dados_bancarios.insert(0, str(self.lancamento.get('DADOS_BANCARIOS', '')))
            self.observacao.insert(0, str(self.lancamento.get('OBSERVAÇÃO', '')))
            
        except Exception as e:
            print(f"Erro ao preencher dados: {str(e)}")
    
    def calcular_total(self, event=None):
        """Calcula o valor total automaticamente"""
        try:
            vr_unit = float(self.vr_unit.get().replace(',', '.'))
            dias = float(self.dias.get() or 1)
            total = vr_unit * dias
            
            self.valor.delete(0, tk.END)
            self.valor.insert(0, f"{total:.2f}")
        except:
            pass
    
    def recalcular_taxas_manual(self):
        """Permite recálculo manual das taxas"""
        if not custom_messagebox("yesno", "Recálculo de Taxas",
                                "Deseja recalcular as taxas de administração?\n\n"
                                "Isso irá atualizar todas as taxas baseado nos lançamentos ativos."):
            return
        
        try:
            # Obter data de referência do usuário
            data_ref = self.data_inicio.get_date()  # Usar data do filtro
            
            resultado = self.gestor_taxas.recalcular_taxas_afetadas(data_ref)
            
            if resultado["sucesso"]:
                custom_messagebox("info", "Sucesso", resultado["mensagem"])
                self.carregar_lancamentos()  # Recarregar para mostrar atualizações
            else:
                custom_messagebox("error", "Erro", resultado["mensagem"])
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro no recálculo: {str(e)}")

    def salvar(self):
        """Salva as alterações"""
        try:
            # Validações básicas
            if not self.nome.get().strip():
                custom_messagebox("error", "Erro", "Nome é obrigatório!")
                return
            
            if not self.valor.get().strip():
                custom_messagebox("error", "Erro", "Valor é obrigatório!")
                return
            
            # Coletar dados
            dados_editados = {
                'data': self.data_rel.get(),
                'tp_desp': self.tp_desp.get(),
                'cnpj_cpf': self.cnpj_cpf.get(),
                'nome': self.nome.get().upper(),
                'categoria': self.categoria.get().upper(),
                'referencia': self.referencia.get().upper(),
                'nf': self.nf.get().upper(),
                'vr_unit': self.vr_unit.get().replace(',', '.'),
                'dias': self.dias.get() or '1',
                'valor': self.valor.get().replace(',', '.'),
                'dt_vencto': self.dt_vencto.get(),
                'dados_bancarios': self.dados_bancarios.get(),
                'observacao': self.observacao.get().upper()
            }
            
            # Chamar callback para salvar
            id_lancamento = self.lancamento.get('ID_LANCAMENTO')
            if self.callback_salvar(id_lancamento, dados_editados):
                custom_messagebox("info", "Sucesso", "Lançamento atualizado com sucesso!")
                self.janela.destroy()
            else:
                custom_messagebox("error", "Erro", "Erro ao salvar alterações!")
                
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar: {str(e)}")                

if __name__ == "__main__":
    print("Iniciando aplicação...")
    app = SistemaEntradaDados()
    print("Atualizando interface...")
    app.root.update_idletasks()
    print("Iniciando mainloop...")
    app.root.mainloop()
