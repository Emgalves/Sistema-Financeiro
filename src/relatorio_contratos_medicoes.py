import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from dateutil.relativedelta import relativedelta
from tkcalendar import DateEntry
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Adicionar diretório raiz ao path para importar módulos corretamente
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        PASTA_CLIENTES,
        BASE_PATH
    )
    print("Configurações importadas com sucesso")
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Definir valores padrão em caso de falha
    BASE_PATH = Path(".")
    ARQUIVO_CLIENTES = BASE_PATH / "dados" / "clientes.xlsx"
    PASTA_CLIENTES = BASE_PATH / "dados" / "clientes"

# Não é mais necessário importar utils pois implementamos tudo aqui


try:
    from src.config.window_config import configurar_janela
    print("window_config importado com sucesso")
except ImportError as e:
    print(f"Erro ao importar window_config: {str(e)}")
    # Implementação simples de configurar_janela como fallback
    def configurar_janela(janela, titulo, largura=900, altura=1000):
        """
        Configura o posicionamento e dimensionamento padrão de uma janela
        
        Args:
            janela: Instância de tk.Tk ou tk.Toplevel
            titulo: Título da janela
            largura: Largura desejada (default 900)
            altura: Altura desejada (default 1000)
        """
        janela.title(titulo)
        
        # Obter dimensões da tela
        screen_width = janela.winfo_screenwidth()
        screen_height = janela.winfo_screenheight()
        
        # Ajustar dimensões para não exceder o tamanho da tela
        largura = min(largura, screen_width)
        altura = min(altura, screen_height)
        
        # Definir posição (sempre no topo esquerdo)
        x = 0
        y = 0
        
        # Configurar geometria
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        
        # Permitir redimensionamento
        janela.resizable(True, True)
        
        # Configurar peso das linhas/colunas para redimensionamento proporcional
        janela.grid_rowconfigure(0, weight=1)
        janela.grid_columnconfigure(0, weight=1)
        
        # Trazer janela para frente
        janela.lift()
        janela.focus_force()
        
# Importar funções auxiliares ou definir aqui
def formatar_moeda_br(valor):
    """Formata um valor numérico como moeda brasileira"""
    try:
        valor_float = float(valor)
        return f"R$ {valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return f"R$ 0,00"

class RelatorioContratos:
    """Classe para geração de relatórios de posição de contratos"""
    def __init__(self, parent=None):
        """Inicializa a interface do relatório de contratos"""
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Relatório de Contratos por Medição", 1000, 950)
        
        # Configuração de variáveis
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.data_referencia = datetime.now()
        self.contratos = []
        self.medicoes = []
        
        # NOVO: Variável para filtro de status
        self.filtro_status = tk.StringVar(value="TODOS")
        
        # Configurar interface
        self.setup_gui()
        
    def setup_gui(self):
        """Configuração da interface gráfica principal"""
        # Frame principal
        self.frame_principal = ttk.Frame(self.root, padding=10)
        self.frame_principal.pack(fill='both', expand=True)
        
        # Frame para seleção
        self.frame_selecao = ttk.LabelFrame(self.frame_principal, text="Seleção de Cliente e Data")
        self.frame_selecao.pack(fill='x', pady=10)
        
        # Container para cliente
        frame_cliente = ttk.Frame(self.frame_selecao)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Selecione o Cliente:", font=('Arial', 11)).pack(side='left', pady=5)
        self.cliente_combobox = ttk.Combobox(frame_cliente, width=40, font=('Arial', 11))
        self.cliente_combobox.pack(side='left', padx=5)
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)
        
        # NOVO: Container para filtro de status e data na mesma linha
        frame_filtros = ttk.Frame(self.frame_selecao)
        frame_filtros.pack(fill='x', padx=10, pady=10)
        
        # NOVO: Filtro de Status
        ttk.Label(frame_filtros, text="Filtro de Status:", font=('Arial', 11)).pack(side='left', pady=5)
        self.status_combobox = ttk.Combobox(
            frame_filtros,
            textvariable=self.filtro_status,
            values=["TODOS", "ATIVO", "CONCLUÍDO"],
            state='readonly',
            width=15,
            font=('Arial', 11)
        )
        self.status_combobox.pack(side='left', padx=5)
        self.status_combobox.bind('<<ComboboxSelected>>', self.on_filtro_status_changed)
        
        # Espaçador
        ttk.Label(frame_filtros, text="     ").pack(side='left')
        
        # Data de referência
        ttk.Label(frame_filtros, text="Data de Referência:", font=('Arial', 11)).pack(side='left', pady=5)
        self.data_entry = DateEntry(
            frame_filtros, 
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='pt_BR',
            font=('Arial', 11)
        )
        self.data_entry.pack(side='left', padx=5)
        self.data_entry.set_date(datetime.now())
        
        # Botão de gerar relatório
        ttk.Button(
            frame_filtros,
            text="Gerar Relatório",
            command=self.gerar_relatorio,
            style='Big.TButton'
        ).pack(side='left', padx=20)
        
        # Frame para resultados - com notebook para separar visões
        self.frame_resultados = ttk.LabelFrame(self.frame_principal, text="Resultados")
        self.frame_resultados.pack(fill='both', expand=True, pady=10)
        
        # Notebook (abas)
        self.notebook = ttk.Notebook(self.frame_resultados)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Abas
        self.aba_resumo = ttk.Frame(self.notebook)
        self.aba_detalhes = ttk.Frame(self.notebook)
        self.aba_grafico = ttk.Frame(self.notebook)
        
        self.notebook.add(self.aba_resumo, text='Resumo')
        self.notebook.add(self.aba_detalhes, text='Detalhes')
        self.notebook.add(self.aba_grafico, text='Gráfico')
        
        # Configurar cada aba
        self.setup_aba_resumo()
        self.setup_aba_detalhes()
        self.setup_aba_grafico()
        
        # Botões na parte inferior
        frame_botoes = ttk.Frame(self.frame_principal)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(
            frame_botoes,
            text="Exportar para Excel",
            command=self.exportar_excel
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes,
            text="Voltar ao Menu",
            command=self.voltar_menu
        ).pack(side='right', padx=5)
        
        # Estilo para botões grandes
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 11, 'bold'), padding=(10, 5))
        
        # Carregar lista de clientes
        self.atualizar_lista_clientes()
        
    def on_filtro_status_changed(self, event=None):
        """Chamado quando o filtro de status é alterado"""
        # Se já tem dados carregados, atualiza a visualização
        if self.contratos:
            self.aplicar_filtro_e_atualizar()
    
    def aplicar_filtro_e_atualizar(self):
        """Aplica o filtro de status e atualiza todas as visualizações"""
        # Aplicar filtro
        self.contratos_filtrados = self.filtrar_contratos_por_status(self.contratos)
        
        # Atualizar visualizações com contratos filtrados
        self.preencher_resumo()
        self.preencher_detalhes()
        self.preencher_grafico()
    
    def filtrar_contratos_por_status(self, contratos_lista):
        """
        Filtra a lista de contratos baseado no filtro de status selecionado
        
        Args:
            contratos_lista: Lista de contratos a ser filtrada
            
        Returns:
            Lista filtrada de contratos
        """
        status_filtro = self.filtro_status.get()
        
        if status_filtro == "TODOS":
            return contratos_lista
        elif status_filtro == "ATIVO":
            return [c for c in contratos_lista if c.get('status', '').upper() == 'ATIVO']
        elif status_filtro == "CONCLUÍDO":
            return [c for c in contratos_lista if c.get('status', '').upper() == 'CONCLUÍDO']
        else:
            return contratos_lista
    
    def setup_aba_resumo(self):
        """Configura a aba de resumo do relatório"""
        # Frame para informações do cliente
        frame_info = ttk.Frame(self.aba_resumo, padding=5)
        frame_info.pack(fill='x', pady=5)
        
        self.lbl_cliente_resumo = ttk.Label(
            frame_info, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_resumo.pack(side='left', padx=10)
        
        self.lbl_data_resumo = ttk.Label(
            frame_info, 
            text=f"Data: {datetime.now().strftime('%d/%m/%Y')}", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_data_resumo.pack(side='left', padx=10)
        
        # Frame para totais
        frame_totais = ttk.LabelFrame(self.aba_resumo, text="Totais Consolidados")
        frame_totais.pack(fill='x', pady=5)
        
        # Grid para os campos de totais
        frame_grid_totais = ttk.Frame(frame_totais, padding=10)
        frame_grid_totais.pack(fill='x')
        
        # Labels da primeira linha
        ttk.Label(frame_grid_totais, text="Total de Contratos:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.lbl_qtd_contratos = ttk.Label(frame_grid_totais, text="0", font=('Arial', 10))
        self.lbl_qtd_contratos.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_grid_totais, text="Contratos em Andamento:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.lbl_qtd_em_andamento = ttk.Label(frame_grid_totais, text="0", font=('Arial', 10))
        self.lbl_qtd_em_andamento.grid(row=0, column=3, sticky='w', padx=5, pady=5)
        
        # Labels da segunda linha
        ttk.Label(frame_grid_totais, text="Valor Total dos Contratos:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.lbl_valor_total = ttk.Label(frame_grid_totais, text="R$ 0,00", font=('Arial', 10))
        self.lbl_valor_total.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_grid_totais, text="Valor já Pago:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.lbl_valor_pago = ttk.Label(frame_grid_totais, text="R$ 0,00", font=('Arial', 10))
        self.lbl_valor_pago.grid(row=1, column=3, sticky='w', padx=5, pady=5)
        
        # Labels da terceira linha
        ttk.Label(frame_grid_totais, text="Saldo a Pagar:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.lbl_saldo = ttk.Label(frame_grid_totais, text="R$ 0,00", font=('Arial', 10))
        self.lbl_saldo.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_grid_totais, text="Percentual Executado:", font=('Arial', 10, 'bold')).grid(row=2, column=2, sticky='w', padx=5, pady=5)
        self.lbl_percentual = ttk.Label(frame_grid_totais, text="0%", font=('Arial', 10))
        self.lbl_percentual.grid(row=2, column=3, sticky='w', padx=5, pady=5)
        
        # Tree para tabela resumo
        frame_tabela = ttk.Frame(self.aba_resumo, padding=5)
        frame_tabela.pack(fill='both', expand=True, pady=5)
        
        colunas = ('ID', 'Fornecedor', 'Descrição', 'Valor Global', 'Valor Pago', 'Saldo', '% Executado', 'Status')
        self.tree_resumo = ttk.Treeview(frame_tabela, columns=colunas, show='headings', height=15)
        
        # Configurar colunas
        self.tree_resumo.heading('ID', text='ID')
        self.tree_resumo.heading('Fornecedor', text='Fornecedor')
        self.tree_resumo.heading('Descrição', text='Descrição')
        self.tree_resumo.heading('Valor Global', text='Valor Global')
        self.tree_resumo.heading('Valor Pago', text='Valor Pago')
        self.tree_resumo.heading('Saldo', text='Saldo')
        self.tree_resumo.heading('% Executado', text='% Executado')
        self.tree_resumo.heading('Status', text='Status')
        
        # Ajustar larguras das colunas
        self.tree_resumo.column('ID', width=40, anchor='center')
        self.tree_resumo.column('Fornecedor', width=150)
        self.tree_resumo.column('Descrição', width=200)
        self.tree_resumo.column('Valor Global', width=100, anchor='e')
        self.tree_resumo.column('Valor Pago', width=100, anchor='e')
        self.tree_resumo.column('Saldo', width=100, anchor='e')
        self.tree_resumo.column('% Executado', width=80, anchor='center')
        self.tree_resumo.column('Status', width=80, anchor='center')
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_tabela, orient='vertical', command=self.tree_resumo.yview)
        scrollx = ttk.Scrollbar(frame_tabela, orient='horizontal', command=self.tree_resumo.xview)
        self.tree_resumo.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        self.tree_resumo.pack(side='left', fill='both', expand=True)
        
        # Bind para seleção de linha
        self.tree_resumo.bind('<<TreeviewSelect>>', self.selecionar_contrato_resumo)
    
    def setup_aba_detalhes(self):
        """Configura a aba de detalhes do contrato selecionado"""
        # Frame para informações do contrato
        frame_info = ttk.LabelFrame(self.aba_detalhes, text="Informações do Contrato", padding=10)
        frame_info.pack(fill='x', pady=5, padx=5)
        
        # Grid para informações
        ttk.Label(frame_info, text="ID do Contrato:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=3)
        self.lbl_id_contrato = ttk.Label(frame_info, text="-", font=('Arial', 10))
        self.lbl_id_contrato.grid(row=0, column=1, sticky='w', padx=5, pady=3)
        
        ttk.Label(frame_info, text="Fornecedor:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=3)
        self.lbl_fornecedor = ttk.Label(frame_info, text="-", font=('Arial', 10))
        self.lbl_fornecedor.grid(row=1, column=1, sticky='w', padx=5, pady=3)
        
        ttk.Label(frame_info, text="Descrição:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=3)
        self.lbl_descricao = ttk.Label(frame_info, text="-", font=('Arial', 10), wraplength=600)
        self.lbl_descricao.grid(row=2, column=1, sticky='w', padx=5, pady=3)
        
        ttk.Label(frame_info, text="Data de Início:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky='w', padx=5, pady=3)
        self.lbl_data_inicio = ttk.Label(frame_info, text="-", font=('Arial', 10))
        self.lbl_data_inicio.grid(row=3, column=1, sticky='w', padx=5, pady=3)
        
        ttk.Label(frame_info, text="Data Final:", font=('Arial', 10, 'bold')).grid(row=3, column=2, sticky='w', padx=5, pady=3)
        self.lbl_data_final = ttk.Label(frame_info, text="-", font=('Arial', 10))
        self.lbl_data_final.grid(row=3, column=3, sticky='w', padx=5, pady=3)
        
        ttk.Label(frame_info, text="Valor Global:", font=('Arial', 10, 'bold')).grid(row=4, column=0, sticky='w', padx=5, pady=3)
        self.lbl_valor_global = ttk.Label(frame_info, text="R$ 0,00", font=('Arial', 10))
        self.lbl_valor_global.grid(row=4, column=1, sticky='w', padx=5, pady=3)
        
        ttk.Label(frame_info, text="Valor Pago:", font=('Arial', 10, 'bold')).grid(row=4, column=2, sticky='w', padx=5, pady=3)
        self.lbl_valor_pago_contrato = ttk.Label(frame_info, text="R$ 0,00", font=('Arial', 10))
        self.lbl_valor_pago_contrato.grid(row=4, column=3, sticky='w', padx=5, pady=3)
        
        ttk.Label(frame_info, text="Saldo:", font=('Arial', 10, 'bold')).grid(row=5, column=0, sticky='w', padx=5, pady=3)
        self.lbl_saldo_contrato = ttk.Label(frame_info, text="R$ 0,00", font=('Arial', 10))
        self.lbl_saldo_contrato.grid(row=5, column=1, sticky='w', padx=5, pady=3)
        
        # Frame para medições
        frame_medicoes = ttk.LabelFrame(self.aba_detalhes, text="Medições do Contrato", padding=5)
        frame_medicoes.pack(fill='both', expand=True, pady=5, padx=5)
        
        # Treeview para medições
        colunas_med = ('ID', 'Data Medição', 'Data Pagamento', 'Referência', 'Valor', 'Status')
        self.tree_medicoes = ttk.Treeview(frame_medicoes, columns=colunas_med, show='headings', height=15)
        
        # Configurar colunas
        self.tree_medicoes.heading('ID', text='ID')
        self.tree_medicoes.heading('Data Medição', text='Data Medição')
        self.tree_medicoes.heading('Data Pagamento', text='Data Pagamento')
        self.tree_medicoes.heading('Referência', text='Referência')
        self.tree_medicoes.heading('Valor', text='Valor')
        self.tree_medicoes.heading('Status', text='Status')
        
        # Ajustar larguras
        self.tree_medicoes.column('ID', width=50, anchor='center')
        self.tree_medicoes.column('Data Medição', width=100, anchor='center')
        self.tree_medicoes.column('Data Pagamento', width=100, anchor='center')
        self.tree_medicoes.column('Referência', width=300)
        self.tree_medicoes.column('Valor', width=100, anchor='e')
        self.tree_medicoes.column('Status', width=80, anchor='center')
        
        # Scrollbars
        scrolly_med = ttk.Scrollbar(frame_medicoes, orient='vertical', command=self.tree_medicoes.yview)
        scrollx_med = ttk.Scrollbar(frame_medicoes, orient='horizontal', command=self.tree_medicoes.xview)
        self.tree_medicoes.configure(yscrollcommand=scrolly_med.set, xscrollcommand=scrollx_med.set)
        
        scrolly_med.pack(side='right', fill='y')
        scrollx_med.pack(side='bottom', fill='x')
        self.tree_medicoes.pack(side='left', fill='both', expand=True)
    
    def setup_aba_grafico(self):
        """Configura a aba de gráfico"""
        # Frame para o gráfico
        self.frame_grafico = ttk.Frame(self.aba_grafico)
        self.frame_grafico.pack(fill='both', expand=True, padx=5, pady=5)
        
    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes no combobox - apenas clientes ativos (sem Data Final)"""
        try:
            # Carregar arquivo de clientes
            if not ARQUIVO_CLIENTES.exists():
                messagebox.showerror("Erro", f"Arquivo de clientes não encontrado: {ARQUIVO_CLIENTES}")
                return
            
            # Ler planilha de clientes
            df_clientes = pd.read_excel(ARQUIVO_CLIENTES)
            
            # Filtrar apenas clientes SEM Data Final (ativos)
            df_ativos = df_clientes[df_clientes['Data Final'].isna()]
            
            # Ordenar por nome
            df_ativos = df_ativos.sort_values('Nome')
            
            # Atualizar combobox
            nomes_clientes = df_ativos['Nome'].tolist()
            self.cliente_combobox['values'] = nomes_clientes
            
            if nomes_clientes:
                self.cliente_combobox.current(0)
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar lista de clientes: {str(e)}")
    
    def selecionar_cliente(self, event=None):
        """Quando um cliente é selecionado no combobox"""
        nome_cliente = self.cliente_combobox.get()
        if not nome_cliente:
            return
        
        self.cliente_atual = nome_cliente
        
        # Construir caminho do arquivo do cliente
        self.arquivo_cliente = PASTA_CLIENTES / f"{nome_cliente}.xlsx"
        
        if not self.arquivo_cliente.exists():
            messagebox.showerror("Erro", f"Arquivo do cliente não encontrado: {self.arquivo_cliente}")
            return
    
    def gerar_relatorio(self):
        """Gera o relatório baseado no cliente e data selecionados"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
            
        # Obter data de referência e converter para datetime
        data_selecionada = self.data_entry.get_date()
        self.data_referencia = datetime.combine(data_selecionada, datetime.min.time())
        
        # Carregar dados
        if not self.carregar_dados():
            return
        
        # NOVO: Aplicar filtro e atualizar
        self.aplicar_filtro_e_atualizar()
        
        # Atualizar labels de informação
        self.lbl_cliente_resumo.config(text=f"Cliente: {self.cliente_atual}")
        self.lbl_data_resumo.config(text=f"Data: {self.data_referencia.strftime('%d/%m/%Y')}")
    
    def carregar_dados(self):
        """Carrega os dados do cliente do arquivo Excel"""
        try:
            wb = load_workbook(self.arquivo_cliente, data_only=True)
            
            # Verificar se as abas existem
            if "Contratos_Medicao" not in wb.sheetnames:
                messagebox.showerror("Erro", "Aba de contratos não encontrada!")
                wb.close()
                return False
                
            if "Medicoes" not in wb.sheetnames:
                messagebox.showerror("Erro", "Aba de medições não encontrada!")
                wb.close()
                return False
                
            # Carregar contratos
            ws_contratos = wb["Contratos_Medicao"]
            self.contratos = []
            
            for row in ws_contratos.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tem ID do contrato
                    # Verificar se o contrato já existia na data de referência
                    data_inicio = row[4]
                    if isinstance(data_inicio, datetime) and data_inicio <= self.data_referencia:
                        self.contratos.append({
                            'id': row[0],
                            'cnpj': row[1],
                            'nome': row[2],
                            'descricao': row[3],
                            'data_inicio': row[4],
                            'data_final': row[5],
                            'valor_global': row[6] or 0,
                            'valor_pago': row[7] or 0,
                            'saldo': row[8] or 0,
                            'status': row[9] or 'ATIVO',
                            'observacao': row[10]
                        })
            
            # Carregar medições
            ws_medicoes = wb["Medicoes"]
            self.medicoes = []
            
            for row in ws_medicoes.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tem ID do contrato
                    # Verificar se a medição já existia na data de referência
                    data_medicao = row[4]
                    if isinstance(data_medicao, datetime) and data_medicao <= self.data_referencia:
                        self.medicoes.append({
                            'id_contrato': row[0],
                            'id_medicao': row[1],
                            'cnpj': row[2],
                            'nome': row[3],
                            'data_medicao': row[4],
                            'data_pagamento': row[5],
                            'referencia': row[6],
                            'valor': row[7] or 0,
                            'status': row[8] or 'PENDENTE',
                            'data_lancamento': row[9],
                            'observacao': row[10]
                        })
            
            wb.close()
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
            try:
                wb.close()
            except:
                pass
            return False
    
    def preencher_resumo(self):
        """Preenche a tabela de resumo e os totais"""
        # Limpar treeview
        for item in self.tree_resumo.get_children():
            self.tree_resumo.delete(item)
        
        # MODIFICADO: Usar contratos filtrados ao invés de self.contratos
        contratos_exibir = self.filtrar_contratos_por_status(self.contratos)
        
        if not contratos_exibir:
            # Zerar totais
            self.lbl_qtd_contratos.config(text="0")
            self.lbl_qtd_em_andamento.config(text="0")
            self.lbl_valor_total.config(text="R$ 0,00")
            self.lbl_valor_pago.config(text="R$ 0,00")
            self.lbl_saldo.config(text="R$ 0,00")
            self.lbl_percentual.config(text="0%")
            return
            
        # Variáveis para totais
        total_contratos = len(contratos_exibir)
        contratos_andamento = 0
        valor_total = 0
        valor_pago = 0
        
        # Preencher tabela
        for contrato in contratos_exibir:
            # Calcular percentual executado
            valor_global = float(contrato['valor_global']) if contrato['valor_global'] else 0
            valor_pago_contrato = float(contrato['valor_pago']) if contrato['valor_pago'] else 0
            
            if valor_global > 0:
                percentual = (valor_pago_contrato / valor_global) * 100
            else:
                percentual = 0
                
            # Verificar se está em andamento
            if contrato['status'] == 'ATIVO':
                contratos_andamento += 1
                
            # Acumular totais
            valor_total += valor_global
            valor_pago += valor_pago_contrato
            
            # Inserir na treeview
            self.tree_resumo.insert('', 'end', values=(
                contrato['id'],
                contrato['nome'],
                contrato['descricao'],
                formatar_moeda_br(valor_global),
                formatar_moeda_br(valor_pago_contrato),
                formatar_moeda_br(valor_global - valor_pago_contrato),
                f"{percentual:.1f}%",
                contrato['status']
            ))
            
        # Calcular saldo total
        saldo_total = valor_total - valor_pago
        percentual_total = (valor_pago / valor_total) * 100 if valor_total > 0 else 0
        
        # Atualizar labels de totais
        self.lbl_qtd_contratos.config(text=str(total_contratos))
        self.lbl_qtd_em_andamento.config(text=str(contratos_andamento))
        self.lbl_valor_total.config(text=formatar_moeda_br(valor_total))
        self.lbl_valor_pago.config(text=formatar_moeda_br(valor_pago))
        self.lbl_saldo.config(text=formatar_moeda_br(saldo_total))
        self.lbl_percentual.config(text=f"{percentual_total:.1f}%")
    
    def selecionar_contrato_resumo(self, event=None):
        """Quando um contrato é selecionado na aba resumo, atualiza a aba de detalhes"""
        selecionado = self.tree_resumo.selection()
        if not selecionado:
            return
            
        # Obter ID do contrato selecionado
        valores = self.tree_resumo.item(selecionado)['values']
        id_contrato = valores[0]
        
        # Buscar dados do contrato
        contrato = next((c for c in self.contratos if c['id'] == id_contrato), None)
        if not contrato:
            return
            
        # Atualizar informações do contrato na aba de detalhes
        self.lbl_id_contrato.config(text=str(contrato['id']))
        self.lbl_fornecedor.config(text=contrato['nome'])
        self.lbl_descricao.config(text=contrato['descricao'])
        self.lbl_valor_global.config(text=formatar_moeda_br(contrato['valor_global']))
        self.lbl_valor_pago_contrato.config(text=formatar_moeda_br(contrato['valor_pago']))
        self.lbl_saldo_contrato.config(text=formatar_moeda_br(contrato['valor_global'] - contrato['valor_pago']))
        
        # Atualizar datas
        if isinstance(contrato.get('data_inicio'), datetime):
            self.lbl_data_inicio.config(text=contrato['data_inicio'].strftime('%d/%m/%Y'))
        else:
            self.lbl_data_inicio.config(text=str(contrato.get('data_inicio') if contrato.get('data_inicio') else ""))
        
        if isinstance(contrato.get('data_final'), datetime):
            self.lbl_data_final.config(text=contrato['data_final'].strftime('%d/%m/%Y'))
        else:
            self.lbl_data_final.config(text=str(contrato.get('data_final') if contrato.get('data_final') else "Não definida"))
        
        # Limpar treeview de medições
        for item in self.tree_medicoes.get_children():
            self.tree_medicoes.delete(item)
            
        # Filtrar medições deste contrato
        medicoes_contrato = [m for m in self.medicoes if m['id_contrato'] == id_contrato]
        
        # Ordenar por data
        medicoes_contrato.sort(key=lambda x: x['data_medicao'] if isinstance(x['data_medicao'], datetime) else datetime.min)
        
        # Preencher medições
        for medicao in medicoes_contrato:
            # Formatar datas
            data_med = medicao['data_medicao'].strftime('%d/%m/%Y') if isinstance(medicao['data_medicao'], datetime) else str(medicao['data_medicao'] if medicao['data_medicao'] else "")
            data_pag = medicao['data_pagamento'].strftime('%d/%m/%Y') if isinstance(medicao['data_pagamento'], datetime) else str(medicao['data_pagamento'] if medicao['data_pagamento'] else "")
            
            self.tree_medicoes.insert('', 'end', values=(
                medicao['id_medicao'],
                data_med,
                data_pag,
                medicao['referencia'],
                formatar_moeda_br(medicao['valor']),
                medicao['status']
            ))
        
        # Mudar para a aba de detalhes
        self.notebook.select(self.aba_detalhes)
    
    def preencher_detalhes(self):
        """Atualiza a aba de detalhes (se houver algum contrato selecionado)"""
        # Esta função pode ser expandida se necessário
        pass
    
    def preencher_grafico(self):
        """Preenche o gráfico com os dados dos contratos"""
        # Limpar frame do gráfico
        for widget in self.frame_grafico.winfo_children():
            widget.destroy()
        
        # MODIFICADO: Usar contratos filtrados
        contratos_exibir = self.filtrar_contratos_por_status(self.contratos)
        
        if not contratos_exibir:
            ttk.Label(
                self.frame_grafico, 
                text="Nenhum contrato para exibir no gráfico",
                font=('Arial', 12)
            ).pack(expand=True)
            return
        
        # Criar figura
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        
        # Gráfico 1: Valores por contrato
        contratos_ids = [f"ID {c['id']}" for c in contratos_exibir]
        valores_globais = [float(c['valor_global']) for c in contratos_exibir]
        valores_pagos = [float(c['valor_pago']) for c in contratos_exibir]
        
        x = range(len(contratos_ids))
        width = 0.35
        
        ax1.bar([i - width/2 for i in x], valores_globais, width, label='Valor Global', color='#1f77b4')
        ax1.bar([i + width/2 for i in x], valores_pagos, width, label='Valor Pago', color='#2ca02c')
        
        ax1.set_xlabel('Contratos')
        ax1.set_ylabel('Valor (R$)')
        ax1.set_title('Valores por Contrato')
        ax1.set_xticks(x)
        ax1.set_xticklabels(contratos_ids, rotation=45, ha='right')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        
        # Gráfico 2: Status dos contratos (Pizza)
        status_counts = {}
        for c in contratos_exibir:
            status = c['status']
            status_counts[status] = status_counts.get(status, 0) + 1
        
        if status_counts:
            ax2.pie(
                status_counts.values(),
                labels=status_counts.keys(),
                autopct='%1.1f%%',
                startangle=90,
                colors=['#2ca02c', '#d62728', '#ff7f0e']
            )
            ax2.set_title('Distribuição por Status')
        
        plt.tight_layout()
        
        # Adicionar ao frame
        canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
    
    def exportar_excel(self):
        """Exporta o relatório para um arquivo Excel"""
        if not self.contratos:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return
        
        # MODIFICADO: Usar contratos filtrados
        contratos_exportar = self.filtrar_contratos_por_status(self.contratos)
        
        if not contratos_exportar:
            messagebox.showwarning("Aviso", "Nenhum contrato corresponde ao filtro selecionado!")
            return
        
        # Solicitar local para salvar
        data_str = self.data_referencia.strftime('%d-%m-%Y')
        nome_padrao = f"Relatorio_{self.cliente_atual}_{data_str}.xlsx"
        
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nome_padrao
        )
        
        if not arquivo:
            return
        
        try:
            # Criar workbook
            wb = Workbook()
            
            # Remover sheet padrão
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Criar aba de resumo
            ws_resumo = wb.create_sheet('Resumo Contratos')
            
            # Estilos
            titulo_font = Font(size=14, bold=True)
            cabecalho_font = Font(size=11, bold=True)
            cabecalho_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            borda = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws_resumo['A1'] = f"Relatório de Contratos - {self.cliente_atual}"
            ws_resumo['A1'].font = titulo_font
            ws_resumo.merge_cells('A1:H1')
            
            ws_resumo['A2'] = f"Data de Referência: {self.data_referencia.strftime('%d/%m/%Y')}"
            ws_resumo.merge_cells('A2:H2')
            
            # NOVO: Adicionar informação sobre o filtro aplicado
            filtro_aplicado = self.filtro_status.get()
            ws_resumo['A3'] = f"Filtro de Status: {filtro_aplicado}"
            ws_resumo.merge_cells('A3:H3')
            
            # Totais - ajustado para linha 5
            ws_resumo['A5'] = "Total de Contratos:"
            ws_resumo['B5'] = self.lbl_qtd_contratos.cget("text")
            
            ws_resumo['D5'] = "Contratos em Andamento:"
            ws_resumo['E5'] = self.lbl_qtd_em_andamento.cget("text")
            
            ws_resumo['A6'] = "Valor Total dos Contratos:"
            ws_resumo['B6'] = self.lbl_valor_total.cget("text")
            
            ws_resumo['D6'] = "Valor já Pago:"
            ws_resumo['E6'] = self.lbl_valor_pago.cget("text")
            
            ws_resumo['A7'] = "Saldo a Pagar:"
            ws_resumo['B7'] = self.lbl_saldo.cget("text")
            
            ws_resumo['D7'] = "Percentual Executado:"
            ws_resumo['E7'] = self.lbl_percentual.cget("text")
            
            # Cabeçalho da tabela - ajustado para linha 9
            cabecalhos = ['ID', 'Fornecedor', 'Descrição', 'Valor Global', 'Valor Pago', 'Saldo', '% Executado', 'Status']
            for col, texto in enumerate(cabecalhos, 1):
                celula = ws_resumo.cell(row=9, column=col, value=texto)
                celula.font = cabecalho_font
                celula.fill = cabecalho_fill
                celula.border = borda
                celula.alignment = Alignment(horizontal='center')
            
            # Dados dos contratos - MODIFICADO: usar contratos_exportar
            linha = 10
            for contrato in contratos_exportar:
                valor_global = float(contrato['valor_global']) if contrato['valor_global'] else 0
                valor_pago = float(contrato['valor_pago']) if contrato['valor_pago'] else 0
                saldo = valor_global - valor_pago
                percentual = (valor_pago / valor_global) * 100 if valor_global > 0 else 0
                
                ws_resumo.cell(row=linha, column=1, value=contrato['id'])
                ws_resumo.cell(row=linha, column=2, value=contrato['nome'])
                ws_resumo.cell(row=linha, column=3, value=contrato['descricao'])
                ws_resumo.cell(row=linha, column=4, value=valor_global)
                ws_resumo.cell(row=linha, column=5, value=valor_pago)
                ws_resumo.cell(row=linha, column=6, value=saldo)
                ws_resumo.cell(row=linha, column=7, value=f"{percentual:.1f}%")
                ws_resumo.cell(row=linha, column=8, value=contrato['status'])
                
                # Formatar células de valor como moeda
                for col in [4, 5, 6]:
                    ws_resumo.cell(row=linha, column=col).number_format = '#,##0.00'
                
                linha += 1
            
            # Ajustar larguras das colunas
            for col in range(1, len(cabecalhos) + 1):
                ws_resumo.column_dimensions[get_column_letter(col)].width = 15
            ws_resumo.column_dimensions['B'].width = 25
            ws_resumo.column_dimensions['C'].width = 40
            
            # Adicionar aba para cada contrato - MODIFICADO: usar contratos_exportar
            for contrato in contratos_exportar:
                # Limitar o nome da aba para 31 caracteres (limite do Excel)
                nome_aba = f"Contrato {contrato['id']}"
                ws_contrato = wb.create_sheet(nome_aba)
                
                # Titulo
                ws_contrato['A1'] = f"Detalhamento do Contrato {contrato['id']}"
                ws_contrato['A1'].font = titulo_font
                ws_contrato.merge_cells('A1:F1')
                
                # Informações do contrato
                ws_contrato['A3'] = "ID do Contrato:"
                ws_contrato['B3'] = contrato['id']
                
                ws_contrato['A4'] = "Fornecedor:"
                ws_contrato['B4'] = contrato['nome']
                
                ws_contrato['A5'] = "Descrição:"
                ws_contrato['B5'] = contrato['descricao']
                ws_contrato.merge_cells('B5:F5')
                
                ws_contrato['A6'] = "Data de Início:"
                if isinstance(contrato['data_inicio'], datetime):
                    ws_contrato['B6'] = contrato['data_inicio'].strftime('%d/%m/%Y')
                else:
                    ws_contrato['B6'] = str(contrato['data_inicio'] if contrato['data_inicio'] else "")
                
                
                ws_contrato['D6'] = "Data Final:"
                if isinstance(contrato.get('data_final'), datetime):
                    ws_contrato['E6'] = contrato['data_final'].strftime('%d/%m/%Y')
                else:
                    ws_contrato['E6'] = str(contrato.get('data_final') if contrato.get('data_final') else "Não definida")
                
                ws_contrato['A7'] = "Valor Global:"
                ws_contrato['B7'] = float(contrato['valor_global']) if contrato['valor_global'] else 0
                ws_contrato['B7'].number_format = '#,##0.00'
                
                ws_contrato['D7'] = "Valor Pago:"
                ws_contrato['E7'] = float(contrato['valor_pago']) if contrato['valor_pago'] else 0
                ws_contrato['E7'].number_format = '#,##0.00'
                
                ws_contrato['A8'] = "Saldo:"
                saldo = float(contrato['valor_global'] or 0) - float(contrato['valor_pago'] or 0)
                ws_contrato['B8'] = saldo
                ws_contrato['B8'].number_format = '#,##0.00'
                if saldo <= 0:
                    ws_contrato['B8'].fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
                    ws_contrato['B8'].font = Font(bold=True, color='FFFFFF')
                    ws_contrato['B8'].comment = Comment(
                        "Saldo esgotado.\nSe houver aditivo, preencha B9 com o novo Valor Global do contrato.",
                        "Sistema"
                    )

                ws_contrato['D8'] = "Status:"
                ws_contrato['E8'] = contrato['status']

                ws_contrato['A9'] = "Novo Valor Global (aditivo):"
                ws_contrato['A9'].font = Font(italic=True, color='666666')
                ws_contrato['B9'].number_format = '#,##0.00'
                ws_contrato['B9'].fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
                ws_contrato['B9'].comment = Comment(
                    "Preencha apenas se houver aditivo.\nO sistema atualizará o Valor Global do contrato ao importar.",
                    "Sistema"
                )

                # Totalizadores dinâmicos — atualizados pelo Excel ao editar a tabela
                ws_contrato['D9'] = "Total novas medições:"
                ws_contrato['D9'].font = Font(italic=True, color='666666')
                # Soma coluna D (Valor) onde E (Status) está em branco e B (Data) não está vazia
                ws_contrato['E9'] = '=SUMPRODUCT((LEN(TRIM(E13:E1000))=0)*(B13:B1000<>"")*D13:D1000)'
                ws_contrato['E9'].number_format = '#,##0.00'

                ws_contrato['D10'] = "Saldo após importação:"
                ws_contrato['D10'].font = Font(italic=True, color='666666')
                ws_contrato['E10'] = "=B8-E9"
                ws_contrato['E10'].number_format = '#,##0.00'
                ws_contrato['E10'].comment = Comment(
                    "Vermelho = novas medições excedem o saldo disponível.\n"
                    "Se houver aditivo, preencha B9 com o novo Valor Global do contrato.",
                    "Sistema"
                )

                from openpyxl.formatting.rule import CellIsRule
                ws_contrato.conditional_formatting.add(
                    'E10',
                    CellIsRule(
                        operator='lessThan', formula=['0'],
                        fill=PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid'),
                        font=Font(bold=True, color='FFFFFF')
                    )
                )

                # Cabeçalho da tabela de medições
                ws_contrato['A11'] = "Medições do Contrato"
                ws_contrato['A11'].font = titulo_font
                ws_contrato.merge_cells('A11:E11')

                cabecalhos_medicoes = ['ID', 'Data Medição', 'Referência', 'Valor', 'Status']
                for col, texto in enumerate(cabecalhos_medicoes, 1):
                    celula = ws_contrato.cell(row=12, column=col, value=texto)
                    celula.font = cabecalho_font
                    celula.fill = cabecalho_fill
                    celula.border = borda
                    celula.alignment = Alignment(horizontal='center')

                ws_contrato.cell(row=12, column=1).comment = Comment(
                    "Preenchido automaticamente.\nDeixe em branco nas novas linhas.",
                    "Sistema"
                )

                # Filtrar medições deste contrato
                medicoes_contrato = [m for m in self.medicoes if m['id_contrato'] == contrato['id']]

                # Ordenar por data
                medicoes_contrato.sort(key=lambda x: x['data_medicao'] if isinstance(x['data_medicao'], datetime) else datetime.min)

                cinza = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

                # Adicionar medições existentes (linha 13+)
                linha = 13
                for medicao in medicoes_contrato:
                    id_cell = ws_contrato.cell(row=linha, column=1, value=medicao['id_medicao'])
                    id_cell.fill = cinza

                    if isinstance(medicao['data_medicao'], datetime):
                        ws_contrato.cell(row=linha, column=2, value=medicao['data_medicao'])
                    else:
                        ws_contrato.cell(row=linha, column=2, value=str(medicao['data_medicao'] or ""))

                    ws_contrato.cell(row=linha, column=3, value=medicao['referencia'])
                    ws_contrato.cell(row=linha, column=4, value=float(medicao['valor']) if medicao['valor'] else 0)
                    ws_contrato.cell(row=linha, column=5, value=medicao['status'])

                    cell_dt = ws_contrato.cell(row=linha, column=2)
                    if isinstance(cell_dt.value, datetime):
                        cell_dt.number_format = 'DD/MM/YYYY'

                    ws_contrato.cell(row=linha, column=4).number_format = '#,##0.00'

                    linha += 1

                # Ajustar larguras das colunas
                ws_contrato.column_dimensions['A'].width = 10
                ws_contrato.column_dimensions['B'].width = 15
                ws_contrato.column_dimensions['C'].width = 40
                ws_contrato.column_dimensions['D'].width = 15
                ws_contrato.column_dimensions['E'].width = 12

            from openpyxl.worksheet.datavalidation import DataValidation

            # ── Coletar fornecedores únicos já presentes no Resumo Contratos ──────────
            # A coluna B do Resumo (a partir da linha 10) contém o nome do fornecedor
            fornecedores_unicos = []
            for row in ws_resumo.iter_rows(min_row=10, max_col=2, values_only=True):
                nome = row[1]  # coluna B
                if nome and str(nome).strip() and str(nome).strip() not in fornecedores_unicos:
                    fornecedores_unicos.append(str(nome).strip())

            # ── Criar aba de serviços adicionais ─────────────────────────────────────
            ws_adicional = wb.create_sheet('Servicos_Adicionais')

            ws_adicional['A1'] = "SERVIÇOS ADICIONAIS — Preencher quando o serviço não tem contrato"
            ws_adicional['A1'].font = Font(size=12, bold=True, color="FF0000")
            ws_adicional.merge_cells('A1:G1')

            ws_adicional['A2'] = "ATENÇÃO: Selecione o fornecedor na lista suspensa da coluna A."
            ws_adicional.merge_cells('A2:G2')

            # Colunas: A=Fornecedor, B=Descrição, C=Data Início, D=Data Fim,
            #          E=Valor Contrato, F=Data Medição, G=Data Pagamento,
            #          H=Referência, I=Valor Medição (R$), J=Observação
            cabecalhos_adic = [
                'Fornecedor',
                'Descrição Serviço',
                'Data Início',
                'Data Fim',
                'Valor Contrato (R$)',
                'Data Medição',
                'Data Pagamento',
                'Referência',
                'Valor Medição (R$)',
                'Observação'
            ]
            for col, texto in enumerate(cabecalhos_adic, 1):
                celula = ws_adicional.cell(row=4, column=col, value=texto)
                celula.font = cabecalho_font
                celula.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                celula.border = borda
                celula.alignment = Alignment(horizontal='center')

            # Mesclar cabeçalho para abranger as 10 colunas
            ws_adicional.merge_cells('A1:J1')
            ws_adicional.merge_cells('A2:J2')

            # ── Data validation: lista suspensa na coluna A (linhas 5 a 14) ──────────
            if fornecedores_unicos:
                lista_inline = ','.join(f'"{n}"' for n in fornecedores_unicos)

                if len(lista_inline) <= 255:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(fornecedores_unicos)}"',
                        allow_blank=True,
                        showDropDown=False
                    )
                    dv.sqref = "A5:A14"
                    ws_adicional.add_data_validation(dv)
                else:
                    ws_listas = wb.create_sheet('_Listas')
                    ws_listas.sheet_state = 'hidden'

                    for i, nome in enumerate(fornecedores_unicos, start=1):
                        ws_listas.cell(row=i, column=1, value=nome)

                    ultima_linha = len(fornecedores_unicos)
                    formula_ref = f"_Listas!$A$1:$A${ultima_linha}"

                    dv = DataValidation(
                        type="list",
                        formula1=formula_ref,
                        allow_blank=True,
                        showDropDown=False
                    )
                    dv.sqref = "A5:A14"
                    ws_adicional.add_data_validation(dv)

            # ── Linhas em branco formatadas para preenchimento ───────────────────────
            for linha in range(5, 15):
                for col in range(1, 11):
                    celula = ws_adicional.cell(row=linha, column=col, value=None)
                    celula.border = borda

            # ── Larguras das colunas ─────────────────────────────────────────────────
            ws_adicional.column_dimensions['A'].width = 30   # Fornecedor
            ws_adicional.column_dimensions['B'].width = 45   # Descrição
            ws_adicional.column_dimensions['C'].width = 14   # Data Início
            ws_adicional.column_dimensions['D'].width = 14   # Data Fim
            ws_adicional.column_dimensions['E'].width = 18   # Valor Contrato
            ws_adicional.column_dimensions['F'].width = 14   # Data Medição
            ws_adicional.column_dimensions['G'].width = 14   # Data Pagamento
            ws_adicional.column_dimensions['H'].width = 30   # Referência
            ws_adicional.column_dimensions['I'].width = 16   # Valor Medição
            ws_adicional.column_dimensions['J'].width = 35   # Observação     
            
            # MODIFICADO: Mensagem mais informativa
            filtro_info = f"\n\nFiltro aplicado: {self.filtro_status.get()}"
            qtd_contratos = len(contratos_exportar)
            messagebox.showinfo(
                "Sucesso", 
                f"Relatório exportado com sucesso!{filtro_info}\n{qtd_contratos} contrato(s) exportado(s).\n\nArquivo:\n{arquivo}"
            )
            
            # Salvar o arquivo
            wb.save(arquivo)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel: {str(e)}")
    
    def voltar_menu(self):
        """Volta ao menu principal"""
        self.root.destroy()
        
        # Mostrar janela principal
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()

def main():
    """Função principal para executar o módulo de forma independente"""
    app = RelatorioContratos()
    app.root.mainloop()
    
if __name__ == "__main__":
    main()
