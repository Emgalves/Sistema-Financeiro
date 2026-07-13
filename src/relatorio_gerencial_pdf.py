"""
RELATÓRIO GERENCIAL EM PDF
==========================

Gera relatório hierárquico em PDF com estrutura:
Grupo → Cliente → Contratos → Medições

Layout profissional similar ao relatório de despesas quinzenais
"""

import os
import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry

# Importações para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Frame, PageTemplate
)
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Adicionar diretório raiz ao path
def add_project_root():
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        PASTA_CLIENTES,
        BASE_PATH
    )
except ImportError:
    BASE_PATH = Path(".")
    ARQUIVO_CLIENTES = BASE_PATH / "dados" / "clientes.xlsx"
    PASTA_CLIENTES = BASE_PATH / "dados" / "clientes"

# Caminho do logo (deve estar na mesma pasta do script ou configurado)
LOGO_PATH = Path(__file__).parent / "logo3.png"
if not LOGO_PATH.exists():
    # Tentar na pasta de saída
    LOGO_PATH = BASE_PATH / "outputs" / "logo3.png"

try:
    from src.config.window_config import configurar_janela
except ImportError:
    def configurar_janela(janela, titulo, largura=800, altura=600):
        janela.title(titulo)
        screen_width = janela.winfo_screenwidth()
        screen_height = janela.winfo_screenheight()
        largura = min(largura, screen_width)
        altura = min(altura, screen_height)
        x = 0
        y = 0
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        janela.resizable(True, True)
        janela.lift()
        janela.focus_force()


def formatar_moeda_br(valor):
    """Formata valor como moeda brasileira"""
    try:
        valor_float = float(valor)
        return f"R$ {valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return "R$ 0,00"


def formatar_data_br(data):
    """Formata data para padrão brasileiro"""
    try:
        # Verifica se é NaT ou None
        if pd.isna(data):
            return ""
    except:
        if data is None:
            return ""
    
    # Se for datetime válido, formata
    if isinstance(data, datetime):
        return data.strftime('%d/%m/%Y')
    return str(data) if data else ""


class NumberedCanvas(pdf_canvas.Canvas):
    """Canvas customizado para numeração de páginas"""
    
    def __init__(self, *args, **kwargs):
        pdf_canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        
    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()
        
    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            pdf_canvas.Canvas.showPage(self)
        pdf_canvas.Canvas.save(self)
        
    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(
            200*mm, 10*mm,
            f"Página {self._pageNumber} de {page_count}"
        )


class RelatorioGerencialPDF:
    """
    Gerador de relatório gerencial em PDF
    Layout hierárquico e profissional
    """
    
    def __init__(self, parent=None):
        """Inicializa interface"""
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Relatório Gerencial em PDF - v2.0", 750, 650)
        
        # Variáveis
        self.grupo_selecionado = None
        self.data_referencia = datetime.now()
        self.dados_consolidados = {}
        
        # Variáveis de filtro
        self.filtro_status = tk.StringVar(value="ativos")
        self.usar_filtro_periodo = tk.BooleanVar(value=False)
        self.incluir_logo = tk.BooleanVar(value=True)
        
        # Setup interface
        self.setup_gui()
        self.carregar_grupos()
        
    def setup_gui(self):
        """Configura interface gráfica"""
        # Frame principal com scrollbar
        main_container = ttk.Frame(self.root)
        main_container.pack(fill='both', expand=True)
        
        # Canvas para scroll
        canvas = tk.Canvas(main_container)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        frame_principal = ttk.Frame(scrollable_frame, padding=20)
        frame_principal.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            frame_principal,
            text="Relatório Gerencial de Medições",
            font=('Arial', 16, 'bold')
        ).pack(pady=(0, 15))
        
        # ===== FILTROS BÁSICOS =====
        frame_basico = ttk.LabelFrame(frame_principal, text="📋 Filtros Básicos", padding=15)
        frame_basico.pack(fill='x', pady=10)
        
        # Grupo
        frame_grupo = ttk.Frame(frame_basico)
        frame_grupo.pack(fill='x', pady=5)
        ttk.Label(frame_grupo, text="Grupo:", font=('Arial', 10, 'bold'), width=18, anchor='w').pack(side='left', padx=5)
        self.combo_grupo = ttk.Combobox(frame_grupo, width=30, font=('Arial', 10), state='readonly')
        self.combo_grupo.pack(side='left', padx=5)
        
        # Data de referência
        frame_data = ttk.Frame(frame_basico)
        frame_data.pack(fill='x', pady=5)
        ttk.Label(frame_data, text="Data de Referência:", font=('Arial', 10, 'bold'), width=18, anchor='w').pack(side='left', padx=5)
        self.data_entry = DateEntry(
            frame_data,
            width=20,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='pt_BR',
            font=('Arial', 10)
        )
        self.data_entry.pack(side='left', padx=5)
        self.data_entry.set_date(datetime.now())
        
        # ===== FILTROS AVANÇADOS =====
        frame_avancado = ttk.LabelFrame(frame_principal, text="🔍 Filtros Avançados", padding=15)
        frame_avancado.pack(fill='x', pady=10)
        
        # Status
        ttk.Label(frame_avancado, text="Status das Obras:", font=('Arial', 10, 'bold')).pack(anchor='w', pady=(0, 5))
        frame_status = ttk.Frame(frame_avancado)
        frame_status.pack(fill='x', padx=20, pady=5)
        
        ttk.Radiobutton(frame_status, text="Todos", variable=self.filtro_status, value="todos").pack(side='left', padx=10)
        ttk.Radiobutton(frame_status, text="Apenas Ativos", variable=self.filtro_status, value="ativos").pack(side='left', padx=10)
        ttk.Radiobutton(frame_status, text="Apenas Concluídos", variable=self.filtro_status, value="concluidos").pack(side='left', padx=10)
        
        # Separador
        ttk.Separator(frame_avancado, orient='horizontal').pack(fill='x', pady=10)
        
        # Período
        check_periodo = ttk.Checkbutton(
            frame_avancado,
            text="Filtrar por Período",
            variable=self.usar_filtro_periodo,
            command=self.toggle_periodo
        )
        check_periodo.pack(anchor='w', pady=(0, 5))
        
        self.frame_periodo = ttk.Frame(frame_avancado)
        self.frame_periodo.pack(fill='x', padx=20, pady=5)
        
        # Data início
        frame_dt_inicio = ttk.Frame(self.frame_periodo)
        frame_dt_inicio.pack(fill='x', pady=3)
        ttk.Label(frame_dt_inicio, text="De:", width=8).pack(side='left')
        self.data_inicio = DateEntry(
            frame_dt_inicio,
            width=15,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='pt_BR',
            font=('Arial', 9),
            state='disabled'
        )
        self.data_inicio.pack(side='left', padx=5)
        
        # Data fim
        frame_dt_fim = ttk.Frame(self.frame_periodo)
        frame_dt_fim.pack(fill='x', pady=3)
        ttk.Label(frame_dt_fim, text="Até:", width=8).pack(side='left')
        self.data_fim = DateEntry(
            frame_dt_fim,
            width=15,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='pt_BR',
            font=('Arial', 9),
            state='disabled'
        )
        self.data_fim.pack(side='left', padx=5)
        
        # Atalhos de período
        frame_atalhos = ttk.Frame(self.frame_periodo)
        frame_atalhos.pack(fill='x', pady=5)
        ttk.Label(frame_atalhos, text="Atalhos:", width=8).pack(side='left')
        ttk.Button(frame_atalhos, text="Último Mês", command=lambda: self.set_periodo_atalho(1), width=12).pack(side='left', padx=2)
        ttk.Button(frame_atalhos, text="Trimestre", command=lambda: self.set_periodo_atalho(3), width=12).pack(side='left', padx=2)
        ttk.Button(frame_atalhos, text="Ano", command=lambda: self.set_periodo_atalho(12), width=12).pack(side='left', padx=2)
        
        # ===== OPÇÕES DO RELATÓRIO =====
        frame_opcoes = ttk.LabelFrame(frame_principal, text="⚙️ Opções do Relatório", padding=15)
        frame_opcoes.pack(fill='x', pady=10)
        
        ttk.Checkbutton(
            frame_opcoes,
            text="Incluir logo da empresa no cabeçalho",
            variable=self.incluir_logo
        ).pack(anchor='w', pady=3)
        
        # Frame de botões
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(pady=15)
        
        # Botão gerar
        btn_gerar = ttk.Button(
            frame_botoes,
            text="📄 Gerar Relatório PDF",
            command=self.gerar_relatorio_pdf,
            width=30
        )
        btn_gerar.pack(side='left', padx=10)
        
        # Botão voltar
        btn_voltar = ttk.Button(
            frame_botoes,
            text="⬅ Voltar",
            command=self.voltar_menu,
            width=15
        )
        btn_voltar.pack(side='left', padx=10)
        
        # Informações
        frame_info = ttk.LabelFrame(frame_principal, text="ℹ️ Informações", padding=10)
        frame_info.pack(fill='both', expand=True, pady=10)
        
        texto_info = """Relatório gerencial consolidado com estrutura hierárquica:

• Filtro por grupo ou todos os grupos
• Filtro por status (ativos/concluídos/todos)
• Filtro opcional por período
• Logo da empresa no cabeçalho

Estrutura: Grupo → Cliente → Contratos → Medições"""
        
        ttk.Label(frame_info, text=texto_info, justify='left', font=('Arial', 9)).pack(anchor='w')
        
    def carregar_grupos(self):
        """Carrega grupos disponíveis"""
        try:
            import pandas as pd
            
            if not ARQUIVO_CLIENTES.exists():
                messagebox.showerror("Erro", f"Arquivo de clientes não encontrado:\n{ARQUIVO_CLIENTES}")
                return
                
            # Carregar planilha especificando a aba 'Clientes'
            df = pd.read_excel(ARQUIVO_CLIENTES, sheet_name='Clientes')
            
            # Verificar se coluna 'Grupo' existe (case-sensitive)
            if 'Grupo' not in df.columns:
                messagebox.showwarning(
                    "Aviso", 
                    f"Coluna 'Grupo' não encontrada.\n\n"
                    f"Colunas disponíveis: {', '.join(df.columns.tolist())}"
                )
                return
            
            # Filtrar apenas valores não vazios e válidos
            grupos_serie = df['Grupo'].dropna()
            
            # Extrair números dos grupos (ex: "Grupo 1" -> 1)
            grupos_validos = []
            for valor in grupos_serie:
                valor_str = str(valor).strip()
                if valor_str and valor_str.lower().startswith('grupo'):
                    try:
                        # Extrair número após "Grupo"
                        numero = int(valor_str.lower().replace('grupo', '').strip())
                        if numero not in grupos_validos:
                            grupos_validos.append(numero)
                    except ValueError:
                        continue
                elif valor_str.isdigit():
                    # Se for só número
                    numero = int(valor_str)
                    if numero not in grupos_validos:
                        grupos_validos.append(numero)
            
            if grupos_validos:
                grupos_validos.sort()
                # Adicionar opção "Todos os Grupos" no início
                opcoes = ['Todos os Grupos'] + [f"Grupo {g}" for g in grupos_validos]
                self.combo_grupo['values'] = opcoes
                self.combo_grupo.current(0)
            else:
                messagebox.showwarning(
                    "Aviso",
                    "Nenhum grupo válido encontrado na coluna 'Grupo'.\n\n"
                    "Formato esperado: 'Grupo 1', 'Grupo 2', etc."
                )
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar grupos:\n{str(e)}")
            import traceback
            traceback.print_exc()
            
    def coletar_dados_grupo(self):
        """Coleta dados de todos os clientes do grupo selecionado"""
        if not self.combo_grupo.get():
            messagebox.showwarning("Aviso", "Selecione um grupo!")
            return False
            
        try:
            import pandas as pd
            from openpyxl import load_workbook
            
            # Verificar se é "Todos os Grupos" ou um grupo específico
            selecao = self.combo_grupo.get()
            
            if selecao == "Todos os Grupos":
                self.grupo_selecionado = "Todos"
                processar_todos = True
            else:
                # Extrair número do grupo
                grupo_num = int(selecao.replace("Grupo ", ""))
                self.grupo_selecionado = grupo_num
                processar_todos = False
            
            self.data_referencia = self.data_entry.get_date()
            
            # Carregar clientes do grupo
            df_clientes = pd.read_excel(ARQUIVO_CLIENTES, sheet_name='Clientes')
            
            if processar_todos:
                # Filtrar apenas clientes com grupo definido
                clientes_grupo = df_clientes[df_clientes['Grupo'].notna()]
            else:
                # Filtrar por grupo específico
                def pertence_ao_grupo(valor):
                    if pd.isna(valor):
                        return False
                    valor_str = str(valor).strip().lower()
                    return (
                        valor_str == f"grupo {grupo_num}" or 
                        valor_str == str(grupo_num)
                    )
                
                clientes_grupo = df_clientes[df_clientes['Grupo'].apply(pertence_ao_grupo)]
            
            if clientes_grupo.empty:
                if processar_todos:
                    messagebox.showinfo(
                        "Informação", 
                        "Nenhum cliente com grupo definido encontrado."
                    )
                else:
                    messagebox.showinfo(
                        "Informação", 
                        f"Nenhum cliente encontrado no Grupo {grupo_num}.\n\n"
                        f"Verifique se há clientes com 'Grupo {grupo_num}' na coluna Grupo."
                    )
                return False
            
            self.dados_consolidados = {
                'grupo': self.grupo_selecionado,
                'data_referencia': self.data_referencia,
                'clientes': []
            }
            
            # Para cada cliente do grupo
            for _, row in clientes_grupo.iterrows():
                nome_cliente = row['Nome']
                
                # === APLICAR FILTRO DE STATUS ===
                data_final = row.get('Data Final')
                status_filtro = self.filtro_status.get()
                
                cliente_ativo = pd.isna(data_final)
                
                # Aplicar filtro
                if status_filtro == "ativos" and not cliente_ativo:
                    continue  # Pular clientes inativos
                elif status_filtro == "concluidos" and cliente_ativo:
                    continue  # Pular clientes ativos
                # "todos" não pula ninguém
                
                # Identificar CPF ou CNPJ baseado no tamanho
                cpf_cnpj_raw = row.get('CPF', '')
                documento = ""
                
                if pd.notna(cpf_cnpj_raw) and str(cpf_cnpj_raw).strip():
                    # Converter para string e remover .0 se vier como float
                    cpf_cnpj_str = str(cpf_cnpj_raw).replace('.0', '').strip()
                    
                    # Remover formatação existente
                    apenas_numeros = ''.join(filter(str.isdigit, cpf_cnpj_str))
                    
                    # Preencher com zeros à esquerda se necessário
                    if len(apenas_numeros) <= 11 and len(apenas_numeros) > 0:
                        # Pode ser CPF - preencher até 11 dígitos
                        apenas_numeros = apenas_numeros.zfill(11)
                    
                    if len(apenas_numeros) == 11:
                        # Formatar como CPF: 000.000.000-00
                        cpf_formatado = f"{apenas_numeros[:3]}.{apenas_numeros[3:6]}.{apenas_numeros[6:9]}-{apenas_numeros[9:]}"
                        documento = f"CPF: {cpf_formatado}"
                    elif len(apenas_numeros) == 14:
                        # Formatar como CNPJ: 00.000.000/0000-00
                        cnpj_formatado = f"{apenas_numeros[:2]}.{apenas_numeros[2:5]}.{apenas_numeros[5:8]}/{apenas_numeros[8:12]}-{apenas_numeros[12:]}"
                        documento = f"CNPJ: {cnpj_formatado}"
                    elif apenas_numeros:
                        # Número de dígitos inválido, mostrar sem formatação
                        documento = f"Doc: {apenas_numeros}"
                
                cno = row.get('CNO', '')
                endereco = row.get('Endereço', '')
                
                # Caminho do arquivo do cliente
                arquivo_cliente = PASTA_CLIENTES / f"{nome_cliente}.xlsx"
                
                if not arquivo_cliente.exists():
                    continue
                    
                # Coletar dados do cliente
                dados_cliente = self.extrair_dados_cliente(
                    arquivo_cliente,
                    nome_cliente,
                    documento,
                    cno,
                    endereco
                )
                
                if dados_cliente:
                    self.dados_consolidados['clientes'].append(dados_cliente)
                    
            if not self.dados_consolidados['clientes']:
                messagebox.showinfo(
                    "Informação", 
                    f"Nenhum cliente ativo com dados encontrado no Grupo {grupo_num}."
                )
                return False
                
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao coletar dados:\n{str(e)}")
            import traceback
            traceback.print_exc()
            return False
            
    def extrair_dados_cliente(self, arquivo, nome, documento, cno, endereco):
        """Extrai dados de contratos e medições de um cliente"""
        try:
            from openpyxl import load_workbook
            import pandas as pd
            
            wb = load_workbook(arquivo, data_only=True)
            
            dados = {
                'nome': nome,
                'documento': documento,  # Pode ser CPF ou CNPJ
                'cno': cno,
                'endereco': endereco,
                'contratos': [],
                'totais': {
                    'valor_total': 0,
                    'valor_executado': 0,
                    'saldo': 0,
                    'qtd_contratos': 0,
                    'qtd_medicoes': 0
                }
            }
            
            # Carregar contratos
            if 'Contratos_Medicao' in wb.sheetnames:
                df_contratos = pd.read_excel(arquivo, sheet_name='Contratos_Medicao')
                
                for idx, row in df_contratos.iterrows():
                    if pd.isna(row.get('CNPJ_Fornecedor')):
                        continue
                        
                    id_contrato = row.get('ID_Contrato')
                    fornecedor = row.get('Nome_Fornecedor', '')
                    descricao = row.get('Descricao', '')
                    valor_global = float(row.get('Valor_Global', 0) or 0)
                    data_inicio = row.get('Data_Inicio')
                    data_final = row.get('Data_Final')  # ✅ Corrigido: era Data_Fim
                    status = row.get('Status', 'ATIVO')
                    
                    # === APLICAR FILTRO DE PERÍODO ===
                    if self.usar_filtro_periodo.get():
                        periodo_inicio = self.data_inicio.get_date()
                        periodo_fim = self.data_fim.get_date()
                        
                        # Converter data_inicio para datetime se necessário
                        if isinstance(data_inicio, datetime):
                            # Contrato deve ter iniciado no período
                            if not (periodo_inicio <= data_inicio.date() <= periodo_fim):
                                continue
                        # Se não tem data de início, pula o contrato com filtro ativo
                        elif pd.notna(data_inicio):
                            try:
                                data_inicio_dt = pd.to_datetime(data_inicio).date()
                                if not (periodo_inicio <= data_inicio_dt <= periodo_fim):
                                    continue
                            except:
                                continue
                    
                    contrato = {
                        'id': id_contrato,
                        'fornecedor': fornecedor,
                        'descricao': descricao,
                        'valor_global': valor_global,
                        'data_inicio': data_inicio,
                        'data_final': data_final,  # ✅ Corrigido: era data_fim
                        'status': status,
                        'medicoes': [],
                        'valor_executado': 0,
                        'saldo': valor_global,
                        'percentual': 0
                    }
                    
                    dados['contratos'].append(contrato)
                    dados['totais']['qtd_contratos'] += 1
                    dados['totais']['valor_total'] += valor_global
                    
            # Carregar medições
            if 'Medicoes' in wb.sheetnames:
                df_medicoes = pd.read_excel(arquivo, sheet_name='Medicoes')
                
                for idx, row in df_medicoes.iterrows():
                    if pd.isna(row.get('ID_Contrato')):
                        continue
                        
                    id_contrato = row.get('ID_Contrato')
                    id_medicao = row.get('ID_Medicao')
                    data_medicao = row.get('Data_Medicao')
                    data_pagamento = row.get('Data_Pagamento')
                    referencia = row.get('Referencia', '')
                    valor = float(row.get('Valor', 0) or 0)
                    status_medicao = row.get('Status', '')
                    
                    # Encontrar contrato correspondente
                    for contrato in dados['contratos']:
                        if contrato['id'] == id_contrato:
                            medicao = {
                                'id': id_medicao,
                                'data_medicao': data_medicao,
                                'data_pagamento': data_pagamento,
                                'referencia': referencia,
                                'valor': valor,
                                'status': status_medicao
                            }
                            contrato['medicoes'].append(medicao)
                            contrato['valor_executado'] += valor
                            break
                            
                    dados['totais']['qtd_medicoes'] += 1
                    dados['totais']['valor_executado'] += valor
                    
            # Calcular saldos e percentuais
            dados['totais']['saldo'] = dados['totais']['valor_total'] - dados['totais']['valor_executado']
            
            for contrato in dados['contratos']:
                contrato['saldo'] = contrato['valor_global'] - contrato['valor_executado']
                if contrato['valor_global'] > 0:
                    contrato['percentual'] = (contrato['valor_executado'] / contrato['valor_global']) * 100
                    
            wb.close()
            return dados
            
        except Exception as e:
            print(f"Erro ao extrair dados de {nome}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
            
    def gerar_relatorio_pdf(self):
        """Gera relatório em PDF"""
        if not self.coletar_dados_grupo():
            return
            
        try:
            # Definir nome do arquivo
            from tkinter import filedialog
            
            nome_sugerido = f"Relatorio_Gerencial_Grupo_{self.grupo_selecionado}_{self.data_referencia.strftime('%Y%m%d')}.pdf"
            
            arquivo_pdf = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=nome_sugerido,
                title="Salvar Relatório PDF"
            )
            
            if not arquivo_pdf:
                return
                
            # Gerar PDF
            self.criar_pdf(arquivo_pdf)
            
            messagebox.showinfo(
                "Sucesso",
                f"Relatório gerado com sucesso!\n\n{arquivo_pdf}"
            )
            
            # Abrir arquivo
            import subprocess
            import platform
            
            if platform.system() == 'Windows':
                os.startfile(arquivo_pdf)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', arquivo_pdf))
            else:  # Linux
                subprocess.call(('xdg-open', arquivo_pdf))
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar PDF:\n{str(e)}")
            import traceback
            traceback.print_exc()
            
    def criar_pdf(self, arquivo):
        """Cria o arquivo PDF com layout profissional"""
        doc = SimpleDocTemplate(
            arquivo,
            pagesize=A4,
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        
        # Customizar estilos
        style_titulo = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        style_cliente = ParagraphStyle(
            'ClienteTitle',
            parent=styles['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#2c3e50'),
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold',
            backColor=colors.HexColor('#ecf0f1'),
            borderPadding=5
        )
        
        style_contrato = ParagraphStyle(
            'ContratoTitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#34495e'),
            spaceBefore=10,
            spaceAfter=5,
            fontName='Helvetica-Bold',
            leftIndent=10
        )
        
        style_normal = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            leading=12
        )
        
        # Elementos do PDF
        elementos = []
        
        # === CABEÇALHO COM LOGO ===
        # Criar tabela para título e logo lado a lado
        if self.incluir_logo.get() and LOGO_PATH.exists():
            from reportlab.platypus import Image
            
            # Logo — logo3.png tem proporção real ~2:1 (medida do conteúdo real
            # do arquivo, não é quadrada); width=88/height=43 preserva a proporção
            logo_img = Image(str(LOGO_PATH),  width=88, height=43)
            
            # Título
            titulo_texto = Paragraph("RELATÓRIO GERENCIAL DE OBRAS", style_titulo)
            
            # Tabela com logo à esquerda e título ao centro
            cabecalho_data = [[logo_img, titulo_texto, ""]]
            table_cabecalho = Table(cabecalho_data, colWidths=[40*mm, 140*mm, 10*mm])
            table_cabecalho.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),   # Logo à esquerda
                ('ALIGN', (1, 0), (1, 0), 'CENTER'), # Título ao centro
            ]))
            elementos.append(table_cabecalho)
        else:
            # Sem logo, apenas título
            elementos.append(Paragraph(
                "RELATÓRIO GERENCIAL DE OBRAS",
                style_titulo
            ))
        
        elementos.append(Spacer(1, 5*mm))
        
        # Informações do grupo - usando Paragraph para renderizar HTML
        data_hoje = datetime.now().strftime('%d/%m/%Y')
        data_ref = self.data_referencia.strftime('%d/%m/%Y')
        
        # Criar estilo para info do grupo
        style_info = ParagraphStyle(
            'InfoGrupo',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#2c3e50'),
            leading=14
        )
        
        # Criar tabela com Paragraphs para renderizar HTML
        info_grupo_paragraphs = [
            [
                Paragraph(f"<b>Grupo:</b> {self.grupo_selecionado}", style_info),
                Paragraph(f"<b>Data Emissão:</b> {data_hoje}", style_info)
            ],
            [
                Paragraph(f"<b>Data Referência:</b> {data_ref}", style_info),
                Paragraph(f"<b>Total Clientes:</b> {len(self.dados_consolidados['clientes'])}", style_info)
            ]
        ]
        
        table_info = Table(info_grupo_paragraphs, colWidths=[90*mm, 90*mm])
        table_info.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        elementos.append(table_info)
        
        # === FILTROS APLICADOS ===
        filtros_texto = []
        if self.filtro_status.get() != "todos":
            status_map = {
                "ativos": "Apenas Ativos",
                "concluidos": "Apenas Concluídos"
            }
            filtros_texto.append(f"Status: {status_map[self.filtro_status.get()]}")
        
        if self.usar_filtro_periodo.get():
            dt_inicio = self.data_inicio.get_date().strftime('%d/%m/%Y')
            dt_fim = self.data_fim.get_date().strftime('%d/%m/%Y')
            filtros_texto.append(f"Período: {dt_inicio} a {dt_fim}")
        
        if filtros_texto:
            style_filtros = ParagraphStyle(
                'Filtros',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#7f8c8d'),
                leading=10,
                alignment=TA_CENTER
            )
            elementos.append(Spacer(1, 2*mm))
            elementos.append(Paragraph(
                f"<i>Filtros aplicados: {' | '.join(filtros_texto)}</i>",
                style_filtros
            ))
        
        elementos.append(Spacer(1, 8*mm))
        
        # === RESUMO GERAL ===
        total_geral_contratos = sum(c['totais']['valor_total'] for c in self.dados_consolidados['clientes'])
        total_geral_executado = sum(c['totais']['valor_executado'] for c in self.dados_consolidados['clientes'])
        total_geral_saldo = total_geral_contratos - total_geral_executado
        perc_geral = (total_geral_executado / total_geral_contratos * 100) if total_geral_contratos > 0 else 0
        
        resumo_data = [
            ['RESUMO GERAL DO GRUPO', '', '', ''],
            ['Total Contratado', 'Total Executado', 'Saldo a Executar', '% Executado'],
            [
                formatar_moeda_br(total_geral_contratos),
                formatar_moeda_br(total_geral_executado),
                formatar_moeda_br(total_geral_saldo),
                f"{perc_geral:.1f}%"
            ]
        ]
        
        table_resumo = Table(resumo_data, colWidths=[45*mm, 45*mm, 45*mm, 45*mm])
        table_resumo.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 11),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONT', (0, 1), (-1, 1), 'Helvetica-Bold', 9),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#bdc3c7')),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONT', (0, 2), (-1, -1), 'Helvetica', 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('SPAN', (0, 0), (-1, 0)),
        ]))
        
        elementos.append(table_resumo)
        elementos.append(Spacer(1, 10*mm))
        
        # === DETALHAMENTO POR CLIENTE ===
        for indice_cliente, cliente_dados in enumerate(self.dados_consolidados['clientes']):
            # Salto de página a cada novo cliente, exceto o primeiro
            # (que já começa logo após o resumo geral do grupo)
            if indice_cliente > 0:
                elementos.append(PageBreak())

            # Cabeçalho do cliente
            texto_cliente = f"<b>{cliente_dados['nome']}</b>"
            if cliente_dados['documento']:
                texto_cliente += f" | {cliente_dados['documento']}"
            # if cliente_dados['cno']:
            #     texto_cliente += f" | CNO: {cliente_dados['cno']}"
                
            elementos.append(Paragraph(texto_cliente, style_cliente))
            
            if cliente_dados['endereco']:
                elementos.append(Paragraph(
                    f"<i>{cliente_dados['endereco']}</i>",
                    style_normal
                ))
                elementos.append(Spacer(1, 3*mm))
                
            # Resumo do cliente
            totais = cliente_dados['totais']
            perc_cliente = (totais['valor_executado'] / totais['valor_total'] * 100) if totais['valor_total'] > 0 else 0
            
            resumo_cliente_data = [
                ['Contratos', 'Medições', 'Valor Total', 'Executado', 'Saldo', '% Exec'],
                [
                    str(totais['qtd_contratos']),
                    str(totais['qtd_medicoes']),
                    formatar_moeda_br(totais['valor_total']),
                    formatar_moeda_br(totais['valor_executado']),
                    formatar_moeda_br(totais['saldo']),
                    f"{perc_cliente:.1f}%"
                ]
            ]
            
            table_cliente = Table(resumo_cliente_data, colWidths=[20*mm, 20*mm, 35*mm, 35*mm, 35*mm, 20*mm])
            table_cliente.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#95a5a6')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elementos.append(table_cliente)
            elementos.append(Spacer(1, 5*mm))
            
            # Contratos do cliente
            if cliente_dados['contratos']:
                for contrato in cliente_dados['contratos']:
                    # Título do contrato
                    texto_contrato = f"<b>Contrato #{contrato['id']}</b> - {contrato['fornecedor']}"
                    elementos.append(Paragraph(texto_contrato, style_contrato))
                    
                    # Detalhes do contrato - usando Paragraph para descrição longa
                    style_desc = ParagraphStyle(
                        'Descricao',
                        parent=styles['Normal'],
                        fontSize=8,
                        leading=10,
                        wordWrap='CJK'
                    )
                    
                    # Descrição em linha separada (ocupando largura total)
                    contrato_descricao = [
                        ['Descrição:', Paragraph(contrato['descricao'], style_desc)]
                    ]
                    
                    table_descricao = Table(contrato_descricao, colWidths=[35*mm, 125*mm])
                    table_descricao.setStyle(TableStyle([
                        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 8),
                        ('FONT', (1, 0), (1, -1), 'Helvetica', 8),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 15),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]))
                    
                    elementos.append(table_descricao)
                    
                    # Informações em 2 colunas para melhor aproveitamento de espaço
                    contrato_info_2col = [
                        # Linha 1: Período | Valor Global
                        [
                            f"Período: {formatar_data_br(contrato['data_inicio'])} a {formatar_data_br(contrato['data_final']) if contrato['data_final'] else 'Não definida'}",
                            f"Valor Global: {formatar_moeda_br(contrato['valor_global'])}"
                        ],
                        # Linha 2: % Realizado | Executado
                        [
                            f"% Realizado: {contrato['percentual']:.1f}%",
                            f"Executado: {formatar_moeda_br(contrato['valor_executado'])}"
                        ],
                        # Linha 3: Status | Saldo
                        [
                            f"Status: {contrato['status']}",
                            f"Saldo: {formatar_moeda_br(contrato['saldo'])}"
                        ]
                    ]
                    
                    table_contrato_info = Table(contrato_info_2col, colWidths=[80*mm, 80*mm])
                    table_contrato_info.setStyle(TableStyle([
                        ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 15),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]))
                    
                    elementos.append(table_contrato_info)
                    
                    # Medições do contrato
                    if contrato['medicoes']:
                        elementos.append(Spacer(1, 3*mm))
                        
                        # Estilo para referências na tabela
                        style_ref = ParagraphStyle(
                            'Referencia',
                            parent=styles['Normal'],
                            fontSize=7,
                            leading=8,
                            wordWrap='CJK'
                        )
                        
                        medicoes_data = [['ID', 'Data Medição', 'Dt. Pagto', 'Referência', 'Valor', 'Status']]
                        
                        for medicao in contrato['medicoes']:
                            medicoes_data.append([
                                str(medicao['id']),
                                formatar_data_br(medicao['data_medicao']),
                                formatar_data_br(medicao['data_pagamento']),
                                Paragraph(medicao['referencia'], style_ref),
                                formatar_moeda_br(medicao['valor']),
                                medicao['status']
                            ])
                            
                        table_medicoes = Table(
                            medicoes_data,
                            colWidths=[10*mm, 20*mm, 20*mm, 60*mm, 25*mm, 20*mm]
                        )
                        table_medicoes.setStyle(TableStyle([
                            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 7),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d5dbdb')),
                            ('FONT', (0, 1), (-1, -1), 'Helvetica', 7),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('ALIGN', (3, 1), (3, -1), 'LEFT'),  # Referência alinhada à esquerda
                            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),  # Cabeçalho centralizado
                            ('VALIGN', (0, 1), (-1, -1), 'TOP'),    # Dados alinhados ao topo
                            ('TOPPADDING', (0, 0), (-1, -1), 3),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                            ('LEFTPADDING', (0, 0), (-1, -1), 3),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                        ]))
                        
                        elementos.append(table_medicoes)
                        
                    elementos.append(Spacer(1, 5*mm))
                    
            else:
                elementos.append(Paragraph(
                    "<i>Nenhum contrato cadastrado para este cliente.</i>",
                    style_normal
                ))
                elementos.append(Spacer(1, 5*mm))
                
            # Separador entre clientes
            elementos.append(Spacer(1, 3*mm))
            
        # === RODAPÉ ===
        elementos.append(Spacer(1, 10*mm))
        
        rodape_style = ParagraphStyle(
            'Rodape',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        
        elementos.append(Paragraph(
            f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
            rodape_style
        ))
        
        # Construir PDF
        doc.build(elementos, canvasmaker=NumberedCanvas)
        
    
    def toggle_periodo(self):
        """Habilita/desabilita campos de período"""
        if self.usar_filtro_periodo.get():
            self.data_inicio.config(state='normal')
            self.data_fim.config(state='normal')
            # Definir valores padrão (último mês)
            self.set_periodo_atalho(1)
        else:
            self.data_inicio.config(state='disabled')
            self.data_fim.config(state='disabled')
    
    def set_periodo_atalho(self, meses):
        """Define período com base em atalho"""
        hoje = datetime.now()
        self.data_fim.set_date(hoje)
        data_inicio = hoje - relativedelta(months=meses)
        self.data_inicio.set_date(data_inicio)
        
        # Habilitar checkbox se não estiver
        if not self.usar_filtro_periodo.get():
            self.usar_filtro_periodo.set(True)
            self.toggle_periodo()
    
    def voltar_menu(self):
        """Volta ao menu principal"""
        self.root.destroy()
        
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()


def main():
    """Função principal"""
    app = RelatorioGerencialPDF()
    app.root.mainloop()


if __name__ == "__main__":
    main()