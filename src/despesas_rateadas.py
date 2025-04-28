# Imports da biblioteca padrão Python
import os
import sys
from pathlib import Path
import re
from datetime import datetime
from decimal import Decimal

# Imports relacionados ao Tkinter
import tkinter as tk
from tkinter import ttk, messagebox, StringVar
from tkinter import *
from tkcalendar import DateEntry, Calendar

# Imports para manipulação de dados e Excel
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import openpyxl
import babel
from dateutil.relativedelta import relativedelta

# Imports para validação
from validate_docbr import CPF, CNPJ

def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar logger

from src.config.logger_config import system_logger, log_action
   

from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH,
        ARQUIVO_FORNECEDORES
 )


from src.config.utils import (
        PASTA_CLIENTES,
        validar_data,
        validar_data_quinzena,
        calcular_proxima_data_quinzena,
        formatar_moeda,
        formatar_valor_excel,
        aplicar_formatacao_celula,
        buscar_dados_bancarios_fornecedor
    )

from src.config.window_config import configurar_janela
    
# Para criação de planilhas
from openpyxl import Workbook

class InterfaceDespesasRateadas:
    def __init__(self, parent):
        self.parent = parent
        self.root = parent
        self.menu_principal = None  # Será definido pelo sistema principal
        
        configurar_janela(self.root, "Gestão de Despesas Rateadas", 900, 1000)

        # Ajustar altura máxima com base na resolução da tela
        altura_tela = self.root.winfo_screenheight()
        altura_maxima = min(1000, altura_tela - 100)  # 100 pixels de margem

        self.root.geometry(f"900x{altura_maxima}")

        # Importar funções necessárias
        from src.config.utils import validar_data_quinzena, calcular_proxima_data_quinzena
        
        # Calcular a próxima data de relatório válida (dia 5 ou 20)
        data_atual = datetime.now().date()
        self.proxima_data_quinzena, _ = calcular_proxima_data_quinzena(data_atual)
        
        # Variáveis
        self.clientes = []
        self.modo_rateio = StringVar(value="percentual") # percentual ou valor
        self.tipo_despesa = StringVar(value="3")  # Padrão: tipo 3
        
        # Configurar interface
        self.setup_gui()
        self.carregar_clientes()
        self.mostrar_historico_rateios()

        self.fornecedor_selecionado = None
    
    def setup_gui(self):
        """Configura a interface gráfica principal"""
        # Frame principal com notebook para abas
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Aba de Rateio
        self.aba_rateio = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_rateio, text="Novo Rateio")
        
        # Aba de Histórico
        self.aba_historico = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_historico, text="Histórico de Rateios")
        
        # Configurar aba de rateio
        self.setup_aba_rateio()
        
        # Configurar aba de histórico
        self.setup_aba_historico()
    
    def setup_aba_rateio(self):
        """Configura a interface da aba de novo rateio"""
        # Frame para seleção de fornecedor (NOVO)
        frame_fornecedor = ttk.LabelFrame(self.aba_rateio, text="Seleção de Fornecedor")
        frame_fornecedor.pack(fill='x', padx=10, pady=5)
        
        # Frame de busca com tamanho reduzido
        frame_busca = ttk.Frame(frame_fornecedor)
        frame_busca.pack(fill='x', padx=5, pady=5)

        # Campo de busca
        ttk.Label(frame_busca, text="Nome:", font=('Arial', 10)).pack(side='left', padx=5)
        self.busca_entry = ttk.Entry(frame_busca, font=('Arial', 10), width=40)
        self.busca_entry.pack(side='left', padx=5)
        self.busca_entry.bind('<Return>', lambda e: self.buscar_fornecedor())

        # Botão de busca
        ttk.Button(frame_busca, 
                text="Buscar", 
                command=self.buscar_fornecedor).pack(side='left', padx=10)

        # Frame para resultados
        frame_resultados = ttk.Frame(frame_fornecedor)
        frame_resultados.pack(fill='x', expand=False, padx=5, pady=5)

        # Lista de resultados com scrollbar
        frame_tree = ttk.Frame(frame_resultados)
        frame_tree.pack(fill='x', expand=False, padx=5, pady=5)
        
        # Scrollbar vertical
        scroll_y = ttk.Scrollbar(frame_tree, orient='vertical')
        scroll_y.pack(side='right', fill='y')
        
        # Treeview para resultados
        self.tree_fornecedores = ttk.Treeview(frame_tree, 
                                            columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                            show='headings',
                                            yscrollcommand=scroll_y.set,
                                            height=3)  # Altura fixa para não ocupar muito espaço
        
        self.tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_fornecedores.heading('Nome', text='Nome')
        self.tree_fornecedores.heading('Categoria', text='Categoria')
        
        # Configurar larguras das colunas
        self.tree_fornecedores.column('CNPJ/CPF', width=150)
        self.tree_fornecedores.column('Nome', width=300)
        self.tree_fornecedores.column('Categoria', width=100)
        
        self.tree_fornecedores.pack(side='left', fill='x', expand=True)
        scroll_y.config(command=self.tree_fornecedores.yview)
        
        # Adicionar evento de duplo clique para selecionar fornecedor
        self.tree_fornecedores.bind('<Double-1>', lambda e: self.selecionar_fornecedor())

        # Frame para dados do fornecedor selecionado
        self.frame_fornecedor_dados = ttk.Frame(frame_fornecedor)
        self.frame_fornecedor_dados.pack(fill='x', pady=5)

        # CNPJ/CPF e Nome
        ttk.Label(self.frame_fornecedor_dados, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2, sticky='e')
        self.cnpj_cpf_fornecedor = ttk.Entry(self.frame_fornecedor_dados, width=20, state='readonly')
        self.cnpj_cpf_fornecedor.grid(row=0, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(self.frame_fornecedor_dados, text="Nome:").grid(row=0, column=2, padx=5, pady=2, sticky='e')
        self.nome_fornecedor = ttk.Entry(self.frame_fornecedor_dados, width=40, state='readonly')
        self.nome_fornecedor.grid(row=0, column=3, padx=5, pady=2, sticky='w')

        # Botão para selecionar fornecedor
        ttk.Button(self.frame_fornecedor_dados, 
                text="Selecionar", 
                command=self.selecionar_fornecedor).grid(row=0, column=4, padx=10, pady=2)

        # Frame para dados da despesa
        frame_despesa = ttk.LabelFrame(self.aba_rateio, text="Dados da Despesa")
        frame_despesa.pack(fill='x', padx=10, pady=5)
        
        # Grid para organizar os campos de forma mais equilibrada
        # Linha 0: Descrição (ocupa 2 colunas)
        ttk.Label(frame_despesa, text="Descrição:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.descricao = ttk.Entry(frame_despesa, width=80)
        self.descricao.grid(row=0, column=1, columnspan=3, padx=5, pady=5, sticky='ew')
        
        # Linha 1: Valor Total e Data de Referência
        ttk.Label(frame_despesa, text="Valor Total (R$):").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.valor_total = ttk.Entry(frame_despesa, width=15)
        self.valor_total.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame_despesa, text="Data de Referência:").grid(row=1, column=2, padx=5, pady=5, sticky='e')
        self.data_rel = DateEntry(frame_despesa, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_rel.set_date(self.proxima_data_quinzena)  # Definir a data calculada
        self.data_rel.grid(row=1, column=3, padx=5, pady=5, sticky='w')

        ttk.Label(frame_despesa, text="Data de Vencimento:").grid(row=1, column=4, padx=5, pady=5, sticky='e')
        self.data_vencto = DateEntry(frame_despesa, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_vencto.grid(row=1, column=5, padx=5, pady=5, sticky='w')
        
        # Linha 2: Tipo de Despesa e Observações
        ttk.Label(frame_despesa, text="Tipo de Despesa:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        tipo_combo = ttk.Combobox(frame_despesa, textvariable=self.tipo_despesa, values=['2', '3'], state='readonly', width=5)
        tipo_combo.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame_despesa, text="Observações:").grid(row=2, column=2, padx=5, pady=5, sticky='e')
        self.observacao = ttk.Entry(frame_despesa, width=40)
        self.observacao.grid(row=2, column=3, padx=5, pady=5, sticky='ew')
        
        # Frame para modo de rateio
        frame_modo = ttk.LabelFrame(self.aba_rateio, text="Modo de Rateio")
        frame_modo.pack(fill='x', padx=10, pady=5)
        
        # Radiobuttons para selecionar o modo
        ttk.Radiobutton(frame_modo, text="Por Percentual (%)", variable=self.modo_rateio, value="percentual",
                    command=self.atualizar_modo_rateio).pack(side='left', padx=20, pady=5)
        ttk.Radiobutton(frame_modo, text="Por Valor (R$)", variable=self.modo_rateio, value="valor",
                    command=self.atualizar_modo_rateio).pack(side='left', padx=20, pady=5)
        
        # Frame para resumo
        self.frame_resumo = ttk.LabelFrame(self.aba_rateio, text="Resumo do Rateio")
        self.frame_resumo.pack(fill='x', padx=10, pady=5)
        
        # Grid para resumo
        self.lbl_total_clientes = ttk.Label(self.frame_resumo, text="Total de Clientes: 0")
        self.lbl_total_clientes.pack(side='left', padx=10, pady=5)
        
        self.lbl_total_valor = ttk.Label(self.frame_resumo, text="Valor Total: R$ 0,00")
        self.lbl_total_valor.pack(side='left', padx=10, pady=5)
        
        self.lbl_total_rateio = ttk.Label(self.frame_resumo, text="Total Rateado: 0%")
        self.lbl_total_rateio.pack(side='left', padx=10, pady=5)
        
        # Frame para lista de clientes com botões de controle
        frame_clientes_header = ttk.Frame(self.aba_rateio)
        frame_clientes_header.pack(fill='x', padx=10, pady=0)
        
        ttk.Label(frame_clientes_header, text="Clientes:", font=('Arial', 10, 'bold')).pack(side='left')
        
        # Botões para marcar/desmarcar todos
        ttk.Button(frame_clientes_header, 
                text="Marcar Todos", 
                command=self.marcar_todos_clientes).pack(side='left', padx=10)
                
        ttk.Button(frame_clientes_header, 
                text="Desmarcar Todos", 
                command=self.desmarcar_todos_clientes).pack(side='left', padx=10)
        
        ttk.Button(frame_clientes_header, 
                text="Ver Todos os Clientes", 
                command=self.mostrar_todos_clientes).pack(side='right', padx=5)
        
        # Frame para lista de clientes
        frame_clientes = ttk.LabelFrame(self.aba_rateio, text="")
        frame_clientes.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar Treeview para clientes
        colunas = ('Cliente', 'Percentual', 'Valor')
        self.tree_clientes = ttk.Treeview(frame_clientes, columns=colunas, show='headings', height=11)
        for col in colunas:
            self.tree_clientes.heading(col, text=col)
        
        self.tree_clientes.column('Cliente', width=300)
        self.tree_clientes.column('Percentual', width=100, anchor='e')
        self.tree_clientes.column('Valor', width=100, anchor='e')
        
        # Adicionar scrollbars
        scrolly = ttk.Scrollbar(frame_clientes, orient='vertical', command=self.tree_clientes.yview)
        scrollx = ttk.Scrollbar(frame_clientes, orient='horizontal', command=self.tree_clientes.xview)
        self.tree_clientes.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_clientes.pack(fill='both', expand=True, side='left')
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Frame para ajuste rápido
        frame_ajuste = ttk.LabelFrame(self.aba_rateio, text="Ajuste Rápido")
        frame_ajuste.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(frame_ajuste, text="Distribuir Igualmente:").pack(side='left', padx=5, pady=5)
        ttk.Button(frame_ajuste, text="Aplicar", command=self.distribuir_igualmente).pack(side='left', padx=5, pady=5)
        
        ttk.Separator(frame_ajuste, orient='vertical').pack(side='left', padx=10, fill='y', pady=5)
        
        ttk.Label(frame_ajuste, text="Definir Percentual:").pack(side='left', padx=5, pady=5)
        self.percentual_selecionado = ttk.Entry(frame_ajuste, width=8)
        self.percentual_selecionado.pack(side='left', padx=5, pady=5)
        ttk.Button(frame_ajuste, text="Aplicar ao Selecionado", 
                command=self.aplicar_percentual_selecionado).pack(side='left', padx=5, pady=5)
        
        ttk.Separator(frame_ajuste, orient='vertical').pack(side='left', padx=10, fill='y', pady=5)

        ttk.Button(frame_ajuste, text="Editar Valores Individuais", 
                    command=self.janela_editar_valores_individuais).pack(side='left', padx=5, pady=5)

        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.aba_rateio)
        frame_botoes.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(frame_botoes, text="Calcular Rateio", 
                command=self.calcular_rateio_modo_atual).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Aplicar Rateio", 
                command=self.aplicar_rateio_clientes).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Voltar ao Menu", 
                command=self.voltar_menu).pack(side='right', padx=5)
    
    def setup_aba_historico(self):
        """Configura a interface da aba de histórico"""
        # Frame para filtros
        frame_filtros = ttk.LabelFrame(self.aba_historico, text="Filtros")
        frame_filtros.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(frame_filtros, text="Data Inicial:").grid(row=0, column=0, padx=5, pady=5)
        self.data_inicial = DateEntry(frame_filtros, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_inicial.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(frame_filtros, text="Data Final:").grid(row=0, column=2, padx=5, pady=5)
        self.data_final = DateEntry(frame_filtros, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_final.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Label(frame_filtros, text="Descrição:").grid(row=1, column=0, padx=5, pady=5)
        self.filtro_descricao = ttk.Entry(frame_filtros, width=40)
        self.filtro_descricao.grid(row=1, column=1, columnspan=2, padx=5, pady=5)
        
        ttk.Button(frame_filtros, text="Filtrar", 
                 command=self.filtrar_historico).grid(row=1, column=3, padx=5, pady=5)
        
        # Frame para histórico
        frame_historico = ttk.LabelFrame(self.aba_historico, text="Registros de Rateios")
        frame_historico.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar Treeview para histórico
        colunas = ('Data Registro', 'Data Relatório', 'Descrição', 'Valor Total', 
                  'Tipo Despesa', 'Qtd Clientes', 'Status')
        self.tree_historico = ttk.Treeview(frame_historico, columns=colunas, show='headings', height=10)
        
        for col in colunas:
            self.tree_historico.heading(col, text=col)
            if col in ['Data Registro', 'Data Relatório']:
                self.tree_historico.column(col, width=150)
            elif col == 'Descrição':
                self.tree_historico.column(col, width=300)
            else:
                self.tree_historico.column(col, width=100)
        
        # Adicionar scrollbars
        scrolly = ttk.Scrollbar(frame_historico, orient='vertical', command=self.tree_historico.yview)
        scrollx = ttk.Scrollbar(frame_historico, orient='horizontal', command=self.tree_historico.xview)
        self.tree_historico.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_historico.pack(fill='both', expand=True, side='left')
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Frame para detalhes do rateio selecionado
        frame_detalhes = ttk.LabelFrame(self.aba_historico, text="Detalhes do Rateio Selecionado")
        frame_detalhes.pack(fill='x', padx=10, pady=5)
        
        # Criar Treeview para detalhes
        colunas_det = ('Cliente', 'Valor', 'Status')
        self.tree_detalhes = ttk.Treeview(frame_detalhes, columns=colunas_det, show='headings', height=5)
        
        for col in colunas_det:
            self.tree_detalhes.heading(col, text=col)
            if col == 'Cliente':
                self.tree_detalhes.column(col, width=300)
            elif col == 'Valor':
                self.tree_detalhes.column(col, width=100, anchor='e')
            else:
                self.tree_detalhes.column(col, width=200)
        
        # Adicionar scrollbars
        scrolly_det = ttk.Scrollbar(frame_detalhes, orient='vertical', command=self.tree_detalhes.yview)
        scrollx_det = ttk.Scrollbar(frame_detalhes, orient='horizontal', command=self.tree_detalhes.xview)
        self.tree_detalhes.configure(yscrollcommand=scrolly_det.set, xscrollcommand=scrollx_det.set)
        
        self.tree_detalhes.pack(fill='both', expand=True, side='left')
        scrolly_det.pack(side='right', fill='y')
        scrollx_det.pack(side='bottom', fill='x')
        
        # Binding para mostrar detalhes
        self.tree_historico.bind('<<TreeviewSelect>>', self.mostrar_detalhes_rateio)
        
        # Botões
        frame_botoes = ttk.Frame(self.aba_historico)
        frame_botoes.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(frame_botoes, text="Atualizar", 
                 command=self.mostrar_historico_rateios).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Voltar ao Menu", 
                 command=self.voltar_menu).pack(side='right', padx=5)
    
    def buscar_fornecedor(self):
        """Busca fornecedores baseado no termo informado"""
        from openpyxl import load_workbook
        
        termo = self.busca_entry.get().strip().upper()
        if not termo:
            messagebox.showwarning("Aviso", "Informe um termo para a busca")
            return
        
        wb = None
        try:
            # Limpar resultados anteriores
            for item in self.tree_fornecedores.get_children():
                self.tree_fornecedores.delete(item)
            
            # Carregar arquivo de fornecedores
            from src.config.config import ARQUIVO_FORNECEDORES
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Realizar a busca
            encontrados = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                    
                # Verificar se o termo está em qualquer campo relevante
                cnpj_cpf = str(row[0]) if row[0] else ""
                nome = str(row[3]).upper() if row[3] else ""
                razao_social = str(row[2]).upper() if row[2] else ""
                
                if (termo in cnpj_cpf or termo in nome or termo in razao_social):
                    # Adicionar à treeview - garantir que todos são strings
                    self.tree_fornecedores.insert('', 'end', values=(
                        str(row[0]),  # CNPJ/CPF como string
                        str(row[3]),  # Nome como string
                        str(row[11]) if row[11] else ""  # Categoria como string
                    ))
                    encontrados += 1
                    
                    # Limitar a 50 resultados
                    if encontrados >= 50:
                        break
            
            if encontrados == 0:
                messagebox.showinfo("Informação", "Nenhum fornecedor encontrado com este termo")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar fornecedores: {str(e)}")
            print(f"Erro detalhado: {e}")
        finally:
            # Garantir que o workbook seja fechado
            if wb:
                wb.close()

    def selecionar_fornecedor(self):
        """Seleciona o fornecedor para o rateio"""
        # Obter item selecionado
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um fornecedor da lista")
            return
        
        try:
            # Obter dados do fornecedor
            valores = self.tree_fornecedores.item(selecionado)['values']
            
            # Verificar se temos todos os valores necessários
            if len(valores) < 3:
                messagebox.showerror("Erro", "Dados do fornecedor incompletos")
                return
            
            # Extrair valores
            cnpj_cpf_raw = str(valores[0]).strip()
            nome = str(valores[1]).strip()
            categoria = str(valores[2]).strip()
            
            # Formatar CNPJ/CPF - usar versão mais simples de formatação
            if len(cnpj_cpf_raw) <= 11:
                cnpj_cpf = cnpj_cpf_raw.zfill(11)  # CPF
            else:
                cnpj_cpf = cnpj_cpf_raw.zfill(14)  # CNPJ
            
            # Atualizar campos
            self.cnpj_cpf_fornecedor.config(state='normal')
            self.cnpj_cpf_fornecedor.delete(0, tk.END)
            self.cnpj_cpf_fornecedor.insert(0, cnpj_cpf)
            self.cnpj_cpf_fornecedor.config(state='readonly')
            
            self.nome_fornecedor.config(state='normal')
            self.nome_fornecedor.delete(0, tk.END)
            self.nome_fornecedor.insert(0, nome)
            self.nome_fornecedor.config(state='readonly')
            
            # Armazenar dados do fornecedor para uso posterior
            self.fornecedor_selecionado = {
                'cnpj_cpf': cnpj_cpf,
                'nome': nome,
                'categoria': categoria
            }
            
            # Informar ao usuário que o fornecedor foi selecionado
            # messagebox.showinfo("Sucesso", f"Fornecedor {nome} selecionado")
            
            return True
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao selecionar fornecedor: {str(e)}")
            print(f"Erro detalhado: {e}")
            return False

    def carregar_clientes(self):
        """Carrega a lista de clientes disponíveis com opção de seleção, apenas clientes ativos (sem data final)"""
        try:
            self.clientes = []
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            # Lista temporária para armazenar clientes antes de ordenar
            clientes_temp = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome não vazio
                    # Verificar se a data final está vazia (cliente ativo)
                    data_final = row[4] if len(row) > 4 else None
                    
                    if not data_final:  # Se não tiver data final, é um cliente ativo
                        clientes_temp.append({
                            'nome': row[0],
                            'percentual': 0,
                            'valor': 0,
                            'arquivo': PASTA_CLIENTES / f"{row[0]}.xlsx",
                            'ativo': True  # Começa como ativo para seleção
                        })
            
            # Ordenar a lista por nome antes de atribuir a self.clientes
            self.clientes = sorted(clientes_temp, key=lambda x: x['nome'])
            
            # Limpar a treeview
            for item in self.tree_clientes.get_children():
                self.tree_clientes.delete(item)
            
            # Preencher a treeview com os clientes e opção de seleção
            # Vamos adicionar uma coluna para selecionar o cliente
            self.tree_clientes.configure(columns=('Ativo', 'Cliente', 'Percentual', 'Valor'))
            self.tree_clientes.heading('Ativo', text='Ativo')
            self.tree_clientes.heading('Cliente', text='Cliente')
            self.tree_clientes.heading('Percentual', text='Percentual (%)' if self.modo_rateio.get() == "percentual" else 'Valor (R$)')
            self.tree_clientes.heading('Valor', text='Valor (R$)')
            
            # Ajustar larguras
            self.tree_clientes.column('Ativo', width=50, anchor='center')
            self.tree_clientes.column('Cliente', width=300)
            self.tree_clientes.column('Percentual', width=100, anchor='e')
            self.tree_clientes.column('Valor', width=100, anchor='e')
            
            # Preencher dados
            for cliente in self.clientes:
                vals = (
                    "✓",  # Marca de ativo
                    cliente['nome'], 
                    f"{cliente['percentual']:.2f}%" if self.modo_rateio.get() == "percentual" else f"R$ {cliente['valor']:.2f}", 
                    f"R$ {cliente['valor']:.2f}"
                )
                item = self.tree_clientes.insert('', 'end', values=vals)
                # Armazenar referência para o cliente no item
                self.tree_clientes.item(item, tags=(cliente['nome'],))
            
            # Adicionar binding para alternar ativo/inativo
            self.tree_clientes.bind('<Button-1>', self.toggle_cliente_ativo)
            
            # Atualizar resumo
            self.atualizar_resumo()
            
            # Fechar workbook
            wb.close()
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")

    # Adicionar método para ver todos os clientes (incluindo os finalizados)
    def mostrar_todos_clientes(self):
        """Abre uma janela mostrando todos os clientes, incluindo os finalizados, permitindo selecionar"""
        try:
            # Carregar arquivo de clientes
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            # Pegar todos os clientes
            clientes_ativos = []
            clientes_finalizados = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome do cliente está na primeira coluna
                    # Verificar se tem data final
                    data_final = row[4] if len(row) > 4 else None
                    
                    if data_final:
                        clientes_finalizados.append(row[0])
                    else:
                        clientes_ativos.append(row[0])
            
            wb.close()
            
            # Mostrar janela com todos os clientes
            janela_todos = tk.Toplevel(self.root)
            janela_todos.title("Todos os Clientes")
            janela_todos.geometry("600x500")
            janela_todos.transient(self.root)
            janela_todos.grab_set()
            
            frame = ttk.Frame(janela_todos, padding="10")
            frame.pack(fill='both', expand=True)
            
            # Lista de clientes ativos
            frame_ativos = ttk.LabelFrame(frame, text="Clientes Ativos")
            frame_ativos.pack(fill='both', expand=True, pady=5)
            
            # Com scrollbar
            frame_ativos_scroll = ttk.Frame(frame_ativos)
            frame_ativos_scroll.pack(fill='both', expand=True, padx=5, pady=5)
            
            scrollbar_ativos = ttk.Scrollbar(frame_ativos_scroll, orient="vertical")
            lista_ativos = tk.Listbox(frame_ativos_scroll, width=50, height=10, yscrollcommand=scrollbar_ativos.set)
            scrollbar_ativos.config(command=lista_ativos.yview)
            
            lista_ativos.pack(side='left', fill='both', expand=True)
            scrollbar_ativos.pack(side='right', fill='y')
            
            for cliente in sorted(clientes_ativos):
                lista_ativos.insert(tk.END, cliente)
            
            # Lista de clientes finalizados
            frame_finalizados = ttk.LabelFrame(frame, text="Clientes Finalizados")
            frame_finalizados.pack(fill='both', expand=True, pady=5)
            
            # Com scrollbar
            frame_finalizados_scroll = ttk.Frame(frame_finalizados)
            frame_finalizados_scroll.pack(fill='both', expand=True, padx=5, pady=5)
            
            scrollbar_finalizados = ttk.Scrollbar(frame_finalizados_scroll, orient="vertical")
            lista_finalizados = tk.Listbox(frame_finalizados_scroll, width=50, height=10, yscrollcommand=scrollbar_finalizados.set)
            scrollbar_finalizados.config(command=lista_finalizados.yview)
            
            lista_finalizados.pack(side='left', fill='both', expand=True)
            scrollbar_finalizados.pack(side='right', fill='y')
            
            for cliente in sorted(clientes_finalizados):
                lista_finalizados.insert(tk.END, cliente)
            
            # Frame para botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)
            
            # Função para incluir cliente selecionado nos rateios
            def incluir_cliente_finalizado():
                selected = lista_finalizados.curselection()
                if not selected:
                    messagebox.showwarning("Aviso", "Selecione um cliente finalizado")
                    return
                    
                cliente_nome = lista_finalizados.get(selected[0])
                
                # Verificar se já existe na lista atual
                for cliente in self.clientes:
                    if cliente['nome'] == cliente_nome:
                        messagebox.showinfo("Informação", "Cliente já está na lista de rateio")
                        janela_todos.destroy()
                        return
                
                # Adicionar à lista
                novo_cliente = {
                    'nome': cliente_nome,
                    'percentual': 0,
                    'valor': 0,
                    'arquivo': PASTA_CLIENTES / f"{cliente_nome}.xlsx",
                    'ativo': True  # Começa como ativo para seleção
                }
                self.clientes.append(novo_cliente)
                
                # Adicionar à treeview
                vals = (
                    "✓",  # Marca de ativo
                    novo_cliente['nome'], 
                    f"{novo_cliente['percentual']:.2f}%" if self.modo_rateio.get() == "percentual" else f"R$ {novo_cliente['valor']:.2f}", 
                    f"R$ {novo_cliente['valor']:.2f}"
                )
                item = self.tree_clientes.insert('', 'end', values=vals)
                self.tree_clientes.item(item, tags=(novo_cliente['nome'],))
                
                # Atualizar resumo
                self.atualizar_resumo()
                
                messagebox.showinfo("Sucesso", f"Cliente {cliente_nome} adicionado à lista de rateio")
                janela_todos.destroy()
            
            ttk.Button(frame_botoes, 
                    text="Incluir Cliente Finalizado no Rateio", 
                    command=incluir_cliente_finalizado).pack(side='left', padx=5)
                    
            ttk.Button(frame_botoes, 
                    text="Fechar", 
                    command=janela_todos.destroy).pack(side='right', padx=5)
            
            # Centralizar a janela
            janela_todos.update_idletasks()
            width = janela_todos.winfo_width()
            height = janela_todos.winfo_height()
            x = (janela_todos.winfo_screenwidth() // 2) - (width // 2)
            y = (janela_todos.winfo_screenheight() // 2) - (height // 2)
            janela_todos.geometry(f'{width}x{height}+{x}+{y}')
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao mostrar todos os clientes: {str(e)}")

    def marcar_todos_clientes(self):
        """Marca todos os clientes como ativos"""
        # Atualizar os dados dos clientes
        for cliente in self.clientes:
            cliente['ativo'] = True
            
        # Atualizar a treeview
        for item in self.tree_clientes.get_children():
            valores = self.tree_clientes.item(item)['values']
            self.tree_clientes.item(item, values=(
                "✓",  # Marca de ativo
                valores[1], 
                valores[2], 
                valores[3]
            ))
        
        # Atualizar resumo
        self.atualizar_resumo()
        messagebox.showinfo("Informação", "Todos os clientes foram marcados")

    def desmarcar_todos_clientes(self):
        """Desmarca todos os clientes"""
        # Atualizar os dados dos clientes
        for cliente in self.clientes:
            cliente['ativo'] = False
            
        # Atualizar a treeview
        for item in self.tree_clientes.get_children():
            valores = self.tree_clientes.item(item)['values']
            self.tree_clientes.item(item, values=(
                " ",  # Marca de inativo
                valores[1], 
                valores[2], 
                valores[3]
            ))
        
        # Atualizar resumo
        self.atualizar_resumo()
        messagebox.showinfo("Informação", "Todos os clientes foram desmarcados")

    def toggle_cliente_ativo(self, event):
        """Alterna o status ativo/inativo do cliente na coluna de checkbox"""
        # Verificar se clicou na coluna "Ativo"
        region = self.tree_clientes.identify_region(event.x, event.y)
        if region != "cell":
            return
            
        coluna = self.tree_clientes.identify_column(event.x)
        coluna_idx = int(coluna[1:]) - 1  # Converter #1, #2, etc. para 0, 1, etc.
        
        # Só processa se clicou na coluna Ativo (índice 0)
        if coluna_idx != 0:
            return
            
        # Obter o item clicado
        item = self.tree_clientes.identify_row(event.y)
        if not item:
            return
            
        # Obter o cliente desta linha
        tags = self.tree_clientes.item(item)['tags']
        if not tags:
            return
            
        nome_cliente = tags[0]
        
        # Encontrar o cliente nos dados
        for cliente in self.clientes:
            if cliente['nome'] == nome_cliente:
                # Alternar status ativo
                cliente['ativo'] = not cliente['ativo']
                
                # Atualizar visualização
                valores = self.tree_clientes.item(item)['values']
                novo_status = "✓" if cliente['ativo'] else " "
                
                # Usar o nome do cliente da tag, não dos valores
                self.tree_clientes.item(item, values=(
                    novo_status,
                    nome_cliente,  # Usar o nome das tags em vez de valores[1]
                    valores[2],
                    valores[3]
                ))
                
                break
        
        # Atualizar resumo considerando apenas clientes ativos
        self.atualizar_resumo()
    
    def atualizar_modo_rateio(self):
        """Atualiza a interface baseado no modo de rateio selecionado"""
        modo = self.modo_rateio.get()
        
        # Atualizar cabeçalho da coluna editável
        if modo == "percentual":
            self.tree_clientes.heading('Percentual', text='Percentual (%)')
            self.lbl_total_rateio.config(text=f"Total Rateado: {self.calcular_total_percentual():.2f}%")
        else:  # modo == "valor"
            self.tree_clientes.heading('Percentual', text='Valor (R$)')
            self.lbl_total_rateio.config(text=f"Total Rateado: R$ {self.calcular_total_valor():.2f}")
        
        # Limpar os valores atuais preservando o status ativo e o nome do cliente
        for item in self.tree_clientes.get_children():
            valores = self.tree_clientes.item(item)['values']
            ativo = valores[0]  # Status ativo
            nome_cliente = valores[1]  # Nome do cliente
            
            if modo == "percentual":
                self.tree_clientes.item(item, values=(ativo, nome_cliente, "0.00%", "R$ 0.00"))
            else:
                self.tree_clientes.item(item, values=(ativo, nome_cliente, "R$ 0.00", "R$ 0.00"))
        
        # Resetar valores nos dados
        for cliente in self.clientes:
            cliente['percentual'] = 0
            cliente['valor'] = 0
        
        # Permitir edição direta na célula
        self.tree_clientes.bind('<Double-1>', self.editar_celula)
    
    def editar_celula(self, event):
        """Permite edição direta na célula após duplo clique"""
        # Identificar a coluna clicada
        region = self.tree_clientes.identify_region(event.x, event.y)
        if region != "cell":
            return
            
        # Obter informações da célula
        coluna = self.tree_clientes.identify_column(event.x)
        coluna_idx = int(coluna[1:]) - 1  # Converter #1, #2, etc. para 0, 1, etc.
        
        # No modo "valor", permitir edição na coluna 2 (Valor)
        # No modo "percentual", permitir edição na coluna 2 (Percentual)
        if (self.modo_rateio.get() == "valor" and coluna_idx == 2) or \
        (self.modo_rateio.get() == "percentual" and coluna_idx == 2):
            pass
        else:
            return
            
        # Obter o item selecionado
        item = self.tree_clientes.identify_row(event.y)
        if not item:
            return
            
        # Obter o valor atual
        valores = self.tree_clientes.item(item)['values']
        nome_cliente = valores[1]  # Nome do cliente está no índice 1
        valor_texto = valores[2]   # Percentual ou Valor está no índice 2
        
        # Obter o tag para identificar o cliente
        tags = self.tree_clientes.item(item)['tags']
        if not tags:
            return
        
        # Limpar símbolos e formatos
        if self.modo_rateio.get() == "percentual":
            valor_atual = valor_texto.replace('%', '').strip()
        else:
            valor_atual = valor_texto.replace('R$', '').strip()
        
        # Criar uma entrada temporária para edição
        x, y, width, height = self.tree_clientes.bbox(item, coluna)
        
        # Criar um frame para a entrada
        entry_frame = ttk.Frame(self.tree_clientes)
        entry_frame.place(x=x, y=y, width=width, height=height)
        
        # Criar a entrada
        entry = ttk.Entry(entry_frame)
        entry.pack(fill='both', expand=True)
        entry.insert(0, valor_atual)
        entry.select_range(0, tk.END)
        entry.focus_set()
        
        # Função para finalizar a edição
        def finalizar_edicao(event=None):
            try:
                # Obter o novo valor
                novo_valor = entry.get().strip().replace(',', '.')
                if not novo_valor:
                    novo_valor = "0"
                    
                novo_valor_float = float(novo_valor)
                
                # Atualizar o valor no cliente correspondente
                for cliente in self.clientes:
                    if cliente['nome'] == tags[0]:  # Usando a tag como identificador
                        if self.modo_rateio.get() == "percentual":
                            cliente['percentual'] = novo_valor_float
                            # Atualizar a treeview
                            status = "✓" if cliente['ativo'] else " "
                            self.tree_clientes.item(item, values=(
                                status,
                                cliente['nome'], 
                                f"{novo_valor_float:.2f}%", 
                                f"R$ {cliente['valor']:.2f}"
                            ))
                        else:  # modo == "valor"
                            cliente['valor'] = novo_valor_float
                            # Atualizar a treeview
                            status = "✓" if cliente['ativo'] else " "
                            self.tree_clientes.item(item, values=(
                                status,
                                cliente['nome'], 
                                f"R$ {novo_valor_float:.2f}", 
                                f"R$ {cliente['valor']:.2f}"
                            ))
                        break
                
                # Atualizar resumo
                self.atualizar_resumo()
                
            except ValueError:
                messagebox.showerror("Erro", "Valor inválido!")
            finally:
                # Destruir o frame de edição
                entry_frame.destroy()
        
        # Eventos para finalizar a edição
        entry.bind("<Return>", finalizar_edicao)
        entry.bind("<FocusOut>", finalizar_edicao)
    
    def janela_editar_valores_individuais(self):
        """Abre uma janela para editar valores individuais para cada cliente"""
        # Verificar se estamos no modo valor
        if self.modo_rateio.get() != "valor":
            messagebox.showinfo("Informação", "Esta funcionalidade está disponível apenas no modo de rateio por valor.")
            return
        
        # Verificar se há um valor total definido
        try:
            valor_total = float(self.valor_total.get().replace(',', '.'))
            if valor_total <= 0:
                messagebox.showerror("Erro", "Informe um valor total válido primeiro!")
                return
        except ValueError:
            messagebox.showerror("Erro", "Informe um valor total válido primeiro!")
            return
        
        # Criar janela para edição de valores
        janela = tk.Toplevel(self.root)
        janela.title("Editar Valores Individuais")
        janela.geometry("500x550")
        janela.transient(self.root)
        janela.grab_set()
        
        # Frame principal
        frame = ttk.Frame(janela, padding="10")
        frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(frame, text="Editar Valores para Cada Cliente", font=('Helvetica', 12, 'bold')).pack(pady=10)
        ttk.Label(frame, text=f"Valor Total: R$ {valor_total:.2f}").pack(pady=5)
        
        # Criar frame com scroll para listar clientes
        frame_clientes = ttk.Frame(frame)
        frame_clientes.pack(fill='both', expand=True, pady=10)
        
        # Canvas e scrollbar
        canvas = tk.Canvas(frame_clientes)
        scrollbar = ttk.Scrollbar(frame_clientes, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)
        
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Variáveis para armazenar os valores temporariamente
        valor_entries = {}
        
        # Adicionar campos para cada cliente
        row = 0
        clientes_ativos = [c for c in self.clientes if c['ativo']]
        
        for cliente in clientes_ativos:
            ttk.Label(scroll_frame, text=cliente['nome'], width=30, anchor='w').grid(row=row, column=0, padx=5, pady=2, sticky='w')
            
            valor_var = tk.StringVar(value=f"{cliente['valor']:.2f}")
            valor_entries[cliente['nome']] = valor_var
            
            ttk.Entry(scroll_frame, textvariable=valor_var, width=15).grid(row=row, column=1, padx=5, pady=2)
            ttk.Label(scroll_frame, text="R$").grid(row=row, column=2, padx=0, pady=2, sticky='w')
            
            row += 1
        
        # Mostrar soma atual
        def atualizar_soma(*args):
            try:
                soma = sum(float(valor_entries[c['nome']].get().replace(',', '.')) for c in clientes_ativos)
                diferenca = valor_total - soma
                lbl_soma.config(text=f"Total Atual: R$ {soma:.2f}")
                lbl_diferenca.config(text=f"Diferença: R$ {diferenca:.2f}")
                
                # Mudar cor do texto da diferença
                if abs(diferenca) < 0.01:  # Tolerância de 1 centavo
                    lbl_diferenca.config(foreground="green")
                else:
                    lbl_diferenca.config(foreground="red")
            except ValueError:
                lbl_soma.config(text="Total Atual: Erro")
                lbl_diferenca.config(text="Diferença: Erro", foreground="red")
        
        # Vincular eventos de mudança
        for var in valor_entries.values():
            var.trace_add("write", atualizar_soma)
        
        # Frame para mostrar resumo
        frame_resumo = ttk.Frame(frame)
        frame_resumo.pack(fill='x', pady=5)
        
        lbl_soma = ttk.Label(frame_resumo, text="Total Atual: R$ 0.00")
        lbl_soma.pack(side='left', padx=10)
        
        lbl_diferenca = ttk.Label(frame_resumo, text="Diferença: R$ 0.00")
        lbl_diferenca.pack(side='left', padx=10)
        
        # Atualizar soma inicial
        atualizar_soma()
        
        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)
        
        def aplicar_valores():
            try:
                # Verificar se a soma está correta
                soma = sum(float(valor_entries[c['nome']].get().replace(',', '.')) for c in clientes_ativos)
                diferenca = valor_total - soma
                
                if abs(diferenca) > 0.01:  # Tolerância de 1 centavo
                    if not messagebox.askyesno("Confirmação", 
                                            f"A soma dos valores (R$ {soma:.2f}) não corresponde ao valor total (R$ {valor_total:.2f}).\n"
                                            f"Deseja continuar mesmo assim?"):
                        return
                
                # Aplicar valores
                for cliente in self.clientes:
                    if cliente['nome'] in valor_entries:
                        cliente['valor'] = float(valor_entries[cliente['nome']].get().replace(',', '.'))
                
                # Atualizar a treeview
                for item in self.tree_clientes.get_children():
                    tags = self.tree_clientes.item(item)['tags']
                    if tags:
                        nome_cliente = tags[0]
                        for cliente in self.clientes:
                            if cliente['nome'] == nome_cliente:
                                status = "✓" if cliente['ativo'] else " "
                                self.tree_clientes.item(item, values=(
                                    status,
                                    cliente['nome'], 
                                    f"R$ {cliente['valor']:.2f}", 
                                    f"R$ {cliente['valor']:.2f}"
                                ))
                                break
                
                # Atualizar resumo
                self.atualizar_resumo()
                
                # Fechar janela
                janela.destroy()
                messagebox.showinfo("Sucesso", "Valores aplicados com sucesso!")
                
            except ValueError as e:
                messagebox.showerror("Erro", f"Erro ao aplicar valores: {str(e)}")
        
        ttk.Button(frame_botoes, text="Aplicar", command=aplicar_valores).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela.destroy).pack(side='right', padx=5)
        
        # Centralizar a janela
        janela.update_idletasks()
        width = janela.winfo_width()
        height = janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (width // 2)
        y = (janela.winfo_screenheight() // 2) - (height // 2)
        janela.geometry(f'{width}x{height}+{x}+{y}')

    def atualizar_resumo(self):
        """Atualiza as informações de resumo"""
        # Atualizar o número de clientes
        self.lbl_total_clientes.config(text=f"Total de Clientes: {len(self.clientes)}")
        
        # Atualizar o valor total
        try:
            valor_total = float(self.valor_total.get().replace(',', '.'))
            self.lbl_total_valor.config(text=f"Valor Total: R$ {valor_total:,.2f}")
        except (ValueError, AttributeError):
            self.lbl_total_valor.config(text="Valor Total: R$ 0,00")
        
        # Atualizar o total rateado
        if self.modo_rateio.get() == "percentual":
            total_percentual = self.calcular_total_percentual()
            self.lbl_total_rateio.config(text=f"Total Rateado: {total_percentual:.2f}%")
        else:  # modo == "valor"
            total_valor = self.calcular_total_valor()
            self.lbl_total_rateio.config(text=f"Total Rateado: R$ {total_valor:.2f}")
    
    def calcular_total_percentual(self):
        """Calcula o total de percentual atribuído apenas para clientes ativos"""
        return sum(cliente['percentual'] for cliente in self.clientes if cliente['ativo'])

    def calcular_total_valor(self):
        """Calcula o total de valor atribuído apenas para clientes ativos"""
        return sum(cliente['valor'] for cliente in self.clientes if cliente['ativo'])

    def distribuir_igualmente(self):
        """Distribui os valores igualmente entre os clientes ativos"""
        # Filtrar apenas clientes ativos
        clientes_ativos = [cliente for cliente in self.clientes if cliente['ativo']]
        num_clientes = len(clientes_ativos)
        
        if num_clientes == 0:
            messagebox.showwarning("Aviso", "Selecione pelo menos um cliente ativo!")
            return
                
        if self.modo_rateio.get() == "percentual":
            # Calcular o percentual igual para cada cliente
            percentual_igual = 100 / num_clientes
            
            # Atualizar os clientes
            for cliente in self.clientes:
                # Definir percentual apenas para clientes ativos
                if cliente['ativo']:
                    cliente['percentual'] = percentual_igual
                else:
                    cliente['percentual'] = 0
                    
                # Atualizar a treeview
                for item in self.tree_clientes.get_children():
                    if self.tree_clientes.item(item)['tags'][0] == cliente['nome']:
                        status = "✓" if cliente['ativo'] else " "
                        self.tree_clientes.item(item, values=(
                            status,
                            cliente['nome'], 
                            f"{cliente['percentual']:.2f}%", 
                            f"R$ {cliente['valor']:.2f}"
                        ))
                        break
        else:  # modo == "valor"
            try:
                valor_total = float(self.valor_total.get().replace(',', '.'))
                valor_igual = valor_total / num_clientes
                
                # Atualizar os clientes
                for cliente in self.clientes:
                    # Definir valor apenas para clientes ativos
                    if cliente['ativo']:
                        cliente['valor'] = valor_igual
                    else:
                        cliente['valor'] = 0
                    
                    # Atualizar a treeview
                    for item in self.tree_clientes.get_children():
                        if self.tree_clientes.item(item)['tags'][0] == cliente['nome']:
                            status = "✓" if cliente['ativo'] else " "
                            self.tree_clientes.item(item, values=(
                                status,
                                cliente['nome'], 
                                f"R$ {cliente['valor']:.2f}", 
                                f"R$ {cliente['valor']:.2f}"
                            ))
                            break
            except (ValueError, AttributeError):
                messagebox.showerror("Erro", "Informe um valor total válido!")
                return
        
        # Atualizar resumo
        self.atualizar_resumo()
    
    def aplicar_percentual_selecionado(self):
        """Aplica o percentual informado ao cliente selecionado"""
        selecionado = self.tree_clientes.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um cliente!")
            return
        
        try:
            if self.modo_rateio.get() == "percentual":
                novo_percentual = float(self.percentual_selecionado.get().replace(',', '.'))
                
                # Identificar o cliente selecionado
                item = selecionado[0]
                valores = self.tree_clientes.item(item)['values']
                cliente_nome = valores[0]
                
                # Atualizar o cliente correspondente
                for i, cliente in enumerate(self.clientes):
                    if cliente['nome'] == cliente_nome:
                        cliente['percentual'] = novo_percentual
                        
                        # Atualizar a treeview
                        self.tree_clientes.item(item, values=(
                            cliente_nome, 
                            f"{novo_percentual:.2f}%", 
                            f"R$ {cliente['valor']:.2f}"
                        ))
                        break
            else:  # modo == "valor"
                novo_valor = float(self.percentual_selecionado.get().replace(',', '.'))
                
                # Identificar o cliente selecionado
                item = selecionado[0]
                valores = self.tree_clientes.item(item)['values']
                cliente_nome = valores[0]
                
                # Atualizar o cliente correspondente
                for i, cliente in enumerate(self.clientes):
                    if cliente['nome'] == cliente_nome:
                        cliente['valor'] = novo_valor
                        
                        # Atualizar a treeview
                        self.tree_clientes.item(item, values=(
                            cliente_nome, 
                            f"R$ {novo_valor:.2f}", 
                            f"R$ {cliente['valor']:.2f}"
                        ))
                        break
            
            # Atualizar resumo
            self.atualizar_resumo()
            
        except ValueError:
            messagebox.showerror("Erro", "Valor inválido!")
    
    def calcular_rateio_modo_atual(self):
        """Calcula o rateio baseado no modo atual"""
        try:
            # Validar valor total
            if not self.valor_total.get():
                messagebox.showerror("Erro", "Informe o valor total!")
                return
                
            valor_total = float(self.valor_total.get().replace(',', '.'))
            if valor_total <= 0:
                messagebox.showerror("Erro", "Valor total deve ser maior que zero!")
                return
                
            if self.modo_rateio.get() == "percentual":
                # Verificar se o total é 100%
                total_percentual = self.calcular_total_percentual()
                if not (99.9 <= total_percentual <= 100.1):  # Tolerância para arredondamentos
                    messagebox.showerror("Erro", f"O total de percentuais deve ser 100%. Atual: {total_percentual:.2f}%")
                    return
                    
                # Calcular valores baseados nos percentuais
                for i, cliente in enumerate(self.clientes):
                    cliente['valor'] = (cliente['percentual'] / 100) * valor_total
                    
                    # Atualizar a treeview
                    item = self.tree_clientes.get_children()[i]
                    self.tree_clientes.item(item, values=(
                        cliente['nome'], 
                        f"{cliente['percentual']:.2f}%", 
                        f"R$ {cliente['valor']:.2f}"
                    ))
            else:  # modo == "valor"
                # Verificar se o total corresponde ao valor da despesa
                total_valores = self.calcular_total_valor()
                
                if abs(total_valores - valor_total) > 0.01:  # Tolerância de 1 centavo
                    messagebox.showerror("Erro", 
                                        f"O total dos valores ({total_valores:.2f}) não corresponde ao valor da despesa ({valor_total:.2f})")
                    return
                
                # Os valores já estão definidos, apenas atualizar a treeview
                for i, cliente in enumerate(self.clientes):
                    # Atualizar a treeview
                    item = self.tree_clientes.get_children()[i]
                    self.tree_clientes.item(item, values=(
                        cliente['nome'], 
                        f"R$ {cliente['valor']:.2f}", 
                        f"R$ {cliente['valor']:.2f}"
                    ))
            
            messagebox.showinfo("Sucesso", "Rateio calculado com sucesso!")
            
        except ValueError as e:
            messagebox.showerror("Erro", f"Erro ao calcular rateio: {str(e)}")
    
    def aplicar_rateio_clientes(self):
        """Aplica o rateio aos arquivos dos clientes"""
        # Validar se temos dados básicos preenchidos
        if not self.descricao.get():
            messagebox.showerror("Erro", "Informe a descrição da despesa!")
            return
                
        if not self.valor_total.get():
            messagebox.showerror("Erro", "Informe o valor total da despesa!")
            return
        
        # Validar fornecedor
        if not hasattr(self, 'fornecedor_selecionado') or not self.fornecedor_selecionado:
            messagebox.showerror("Erro", "Selecione um fornecedor!")
            return
        
        # Verificar se o rateio foi calculado
        tem_valores = any(cliente['valor'] > 0 and cliente['ativo'] for cliente in self.clientes)
        if not tem_valores:
            messagebox.showerror("Erro", "Calcule o rateio antes de aplicar!")
            return
                
        # Pedir confirmação
        if not messagebox.askyesno("Confirmação", 
                                "Deseja realmente aplicar este rateio aos clientes?"):
            return
                
        try:
            # Obter datas do relatório e de vencimento
            data_rel_obj = self.data_rel.get_date()
            data_vencto_obj = self.data_vencto.get_date()
                        
            # Converter datas para string formatada
            data_rel_str = data_rel_obj.strftime('%d/%m/%Y')
            data_vencto_str = data_vencto_obj.strftime('%d/%m/%Y')
            
            # Outros dados do lançamento
            tipo_despesa = self.tipo_despesa.get()
            descricao = self.descricao.get().upper()
            
            # Obter observação do usuário (se houver)
            observacao_usuario = self.observacao.get().strip()
            
            # Construir observação final
            if observacao_usuario:
                observacao = f"LANÇAMENTO AUTOMÁTICO - {observacao_usuario}"
            else:
                observacao = "LANÇAMENTO AUTOMÁTICO"
            
            # Lista para registrar resultados
            registros = []
            
            # Filtrar apenas clientes ativos com valor > 0
            clientes_ativos = [cliente for cliente in self.clientes if cliente['ativo'] and cliente['valor'] > 0]
            
            # Buscar dados bancários do fornecedor
            dados_bancarios = buscar_dados_bancarios_fornecedor(
                self.fornecedor_selecionado['cnpj_cpf'],
                "PIX"  # Por padrão, usar PIX como forma de pagamento
            )
            
            for cliente in clientes_ativos:
                try:
                    wb = load_workbook(cliente['arquivo'])
                    ws = wb["Dados"]
                    
                    # Preparar dados do lançamento
                    proxima_linha = ws.max_row + 1
                    
                    # Data do Relatório (formatada)
                    data_rel_date = datetime.strptime(data_rel_str, '%d/%m/%Y').date()
                    ws.cell(row=proxima_linha, column=1, value=data_rel_date)
                    ws.cell(row=proxima_linha, column=1).number_format = 'DD/MM/YYYY'
                    
                    # Tipo de Despesa
                    ws.cell(row=proxima_linha, column=2, value=int(tipo_despesa))
                    
                    # CNPJ/CPF do fornecedor
                    ws.cell(row=proxima_linha, column=3, value=self.fornecedor_selecionado['cnpj_cpf'])
                    
                    # Nome do fornecedor
                    ws.cell(row=proxima_linha, column=4, value=self.fornecedor_selecionado['nome'])
                    
                    # Referência
                    ws.cell(row=proxima_linha, column=5, value=f"{descricao}")
                    
                    # NF (vazio para rateios)
                    ws.cell(row=proxima_linha, column=6, value="")
                    
                    # Valor Unitário
                    ws.cell(row=proxima_linha, column=7, value=cliente['valor'])
                    aplicar_formatacao_celula(ws.cell(row=proxima_linha, column=7))
                    
                    # Dias (1 para despesas rateadas)
                    ws.cell(row=proxima_linha, column=8, value=1)
                    
                    # Valor Total
                    ws.cell(row=proxima_linha, column=9, value=cliente['valor'])
                    aplicar_formatacao_celula(ws.cell(row=proxima_linha, column=9))
                    
                    # Data de Vencimento
                    data_vencto_date = datetime.strptime(data_vencto_str, '%d/%m/%Y').date()
                    ws.cell(row=proxima_linha, column=10, value=data_vencto_date)
                    ws.cell(row=proxima_linha, column=10).number_format = 'DD/MM/YYYY'
                    
                    # Categoria
                    ws.cell(row=proxima_linha, column=11, value=self.fornecedor_selecionado['categoria'])
                    
                    # Dados Bancários
                    ws.cell(row=proxima_linha, column=12, value=dados_bancarios)
                    
                    # Observação - Sempre incluir LANÇAMENTO AUTOMÁTICO
                    ws.cell(row=proxima_linha, column=13, value=observacao)
                    
                    # Salvar planilha
                    wb.save(cliente['arquivo'])
                    
                    # Registrar sucesso
                    registros.append({
                        'cliente': cliente['nome'],
                        'valor': cliente['valor'],
                        'status': 'SUCESSO'
                    })
                    
                except Exception as e:
                    # Registrar falha
                    registros.append({
                        'cliente': cliente['nome'],
                        'valor': cliente['valor'],
                        'status': f'FALHA: {str(e)}'
                    })
            
            # Registrar o rateio no histórico
            self.registrar_historico(registros)
            
            # Exibir resultados
            self.mostrar_resultado_rateio(registros)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao aplicar rateio: {str(e)}")
    
    def registrar_historico(self, registros):
        """Registra o rateio no histórico"""
        try:
            from src.config.config import BASE_PATH  # Importar o caminho base
            
            data_atual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            data_rel_obj = self.data_ref.get_date()
            data_rel_str = data_rel_obj.strftime('%d/%m/%Y')
            descricao = self.descricao.get()
            valor_total = float(self.valor_total.get().replace(',', '.'))
            tipo_despesa = self.tipo_despesa.get()
            
            # Caminho para salvar no Google Drive
            drive_path = Path(BASE_PATH) / "Financeiro" / "Planilhas_Base" / "historico_rateios.xlsx"
            
            # Certificar-se de que a pasta existe
            drive_dir = drive_path.parent
            if not drive_dir.exists():
                drive_dir.mkdir(parents=True, exist_ok=True)
            
            # Criar arquivo de histórico se não existir
            if not drive_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.title = "Histórico"
                
                # Cabeçalhos
                headers = ['Data Registro', 'Data Relatório', 'Descrição', 'Valor Total', 
                        'Tipo Despesa', 'Qtd Clientes', 'Status']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    
                wb.save(drive_path)
            
            # Abrir arquivo de histórico
            wb = load_workbook(drive_path)
            ws = wb["Histórico"]
            
            # Adicionar registro principal
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1, value=data_atual)
            ws.cell(row=proxima_linha, column=2, value=data_rel_str)
            ws.cell(row=proxima_linha, column=3, value=descricao)
            
            # Aplicar formato monetário à coluna de Valor Total
            cell_valor = ws.cell(row=proxima_linha, column=4, value=valor_total)
            aplicar_formatacao_celula(cell_valor)
            
            ws.cell(row=proxima_linha, column=5, value=tipo_despesa)
            ws.cell(row=proxima_linha, column=6, value=len(registros))
            
            # Verificar status geral
            falhas = [r for r in registros if r['status'].startswith('FALHA')]
            status = "SUCESSO" if not falhas else f"PARCIAL ({len(falhas)} falhas)"
            ws.cell(row=proxima_linha, column=7, value=status)
            
            # Adicionar detalhes em outra aba se não existir
            if "Detalhes" not in wb.sheetnames:
                ws_details = wb.create_sheet("Detalhes")
                # Cabeçalhos
                headers = ['ID Rateio', 'Cliente', 'Valor', 'Status']
                for col, header in enumerate(headers, 1):
                    ws_details.cell(row=1, column=col, value=header)
            else:
                ws_details = wb["Detalhes"]
            
            # Adicionar detalhes de cada cliente
            id_rateio = proxima_linha - 1  # Usar a linha como ID do rateio
            for registro in registros:
                proxima_linha_det = ws_details.max_row + 1
                ws_details.cell(row=proxima_linha_det, column=1, value=id_rateio)
                ws_details.cell(row=proxima_linha_det, column=2, value=registro['cliente'])
                
                # Aplicar formato monetário à coluna de valor dos detalhes
                cell_valor_det = ws_details.cell(row=proxima_linha_det, column=3, value=registro['valor'])
                aplicar_formatacao_celula(cell_valor_det)
                
                ws_details.cell(row=proxima_linha_det, column=4, value=registro['status'])
            
            # Salvar histórico no Google Drive
            wb.save(drive_path)
            
            # Registrar log da ação
            log_action(f"Rateio de despesa aplicado: {descricao} - R$ {valor_total:.2f} - {len(registros)} clientes")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao registrar histórico: {str(e)}")
    
    def mostrar_resultado_rateio(self, registros):
        """Mostra o resultado do rateio aplicado"""
        # Contar sucessos e falhas
        sucessos = [r for r in registros if r['status'] == 'SUCESSO']
        falhas = [r for r in registros if r['status'].startswith('FALHA')]
        
        # Criar janela de resultados
        janela_resultado = tk.Toplevel(self.root)
        janela_resultado.title("Resultado do Rateio")
        janela_resultado.geometry("600x400")
        janela_resultado.transient(self.root)
        janela_resultado.grab_set()
        
        # Frame principal
        frame = ttk.Frame(janela_resultado, padding="10")
        frame.pack(fill='both', expand=True)
        
        # Resumo
        ttk.Label(frame, text="Resumo do Rateio", font=('Helvetica', 12, 'bold')).pack(pady=10)
        ttk.Label(frame, text=f"Total de Clientes: {len(registros)}").pack(pady=2)
        ttk.Label(frame, text=f"Sucessos: {len(sucessos)}").pack(pady=2)
        ttk.Label(frame, text=f"Falhas: {len(falhas)}").pack(pady=2)
        
        # Frame para lista de resultados
        frame_lista = ttk.Frame(frame)
        frame_lista.pack(fill='both', expand=True, pady=10)
        
        # Criar Treeview para resultados
        colunas = ('Cliente', 'Valor', 'Status')
        tree_resultados = ttk.Treeview(frame_lista, columns=colunas, show='headings', height=8)
        
        for col in colunas:
            tree_resultados.heading(col, text=col)
            if col == 'Cliente':
                tree_resultados.column(col, width=250)
            elif col == 'Valor':
                tree_resultados.column(col, width=100, anchor='e')
            else:
                tree_resultados.column(col, width=150)
        
        # Adicionar scrollbars
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', command=tree_resultados.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient='horizontal', command=tree_resultados.xview)
        tree_resultados.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        tree_resultados.pack(fill='both', expand=True, side='left')
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Preencher a treeview com os resultados
        for registro in registros:
            tree_resultados.insert('', 'end', values=(
                registro['cliente'], 
                f"R$ {registro['valor']:.2f}", 
                registro['status']
            ))
        
        # Botão fechar
        ttk.Button(frame, text="Fechar", command=janela_resultado.destroy).pack(pady=10)
        
        # Centralizar a janela
        janela_resultado.update_idletasks()
        width = janela_resultado.winfo_width()
        height = janela_resultado.winfo_height()
        x = (janela_resultado.winfo_screenwidth() // 2) - (width // 2)
        y = (janela_resultado.winfo_screenheight() // 2) - (height // 2)
        janela_resultado.geometry(f'{width}x{height}+{x}+{y}')
        
        # Atualizar lista de histórico
        self.mostrar_historico_rateios()
        
        # Limpar os campos após aplicação bem-sucedida
        if not falhas:
            self.limpar_campos()
    
    def limpar_campos(self):
        """Limpa os campos de entrada após operação bem-sucedida"""
        self.descricao.delete(0, tk.END)
        self.valor_total.delete(0, tk.END)
        self.observacao.delete(0, tk.END)
        
        # Resetar valores nos clientes
        for cliente in self.clientes:
            cliente['percentual'] = 0
            cliente['valor'] = 0
            
        # Limpar a treeview
        for i, cliente in enumerate(self.clientes):
            item = self.tree_clientes.get_children()[i]
            if self.modo_rateio.get() == "percentual":
                self.tree_clientes.item(item, values=(
                    cliente['nome'], 
                    "0.00%", 
                    "R$ 0.00"
                ))
            else:
                self.tree_clientes.item(item, values=(
                    cliente['nome'], 
                    "R$ 0.00", 
                    "R$ 0.00"
                ))
        
        # Atualizar resumo
        self.atualizar_resumo()
    
    def mostrar_historico_rateios(self):
        """Carrega e mostra o histórico de rateios"""
        try:
            from src.config.config import BASE_PATH  # Importar o caminho base
            
            # Limpar a treeview
            for item in self.tree_historico.get_children():
                self.tree_historico.delete(item)
                    
            # Caminho do arquivo no Google Drive
            drive_path = Path(BASE_PATH) / "Financeiro" / "Planilhas_Base" / "historico_rateios.xlsx"
            
            # Verificar se o arquivo de histórico existe
            if not drive_path.exists():
                return
                    
            # Abrir arquivo de histórico
            wb = load_workbook(drive_path)
            ws = wb["Histórico"]
            
            # Preencher a treeview com os registros
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tiver data de registro
                    # Formatar a coluna de valor total
                    if row[3] is not None:
                        try:
                            valor_total = float(row[3])
                            valor_formatado = f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                        except:
                            valor_formatado = str(row[3])
                    else:
                        valor_formatado = "R$ 0,00"
                    
                    # Inserir na treeview com o valor formatado
                    valores = list(row)
                    valores[3] = valor_formatado
                    self.tree_historico.insert('', 'end', values=valores)
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao mostrar histórico: {str(e)}")
    
    def filtrar_historico(self):
        """Filtra o histórico de rateios"""
        try:
            from src.config.config import BASE_PATH  # Importar o caminho base
            
            # Obter dados do filtro
            data_inicial = self.data_inicial.get_date()
            data_final = self.data_final.get_date()
            descricao = self.filtro_descricao.get().strip().upper()
            
            # Limpar a treeview
            for item in self.tree_historico.get_children():
                self.tree_historico.delete(item)
                
            # Caminho do arquivo no Google Drive
            drive_path = Path(BASE_PATH) / "Financeiro" / "Planilhas_Base" / "historico_rateios.xlsx"
            
            # Verificar se o arquivo de histórico existe
            if not drive_path.exists():
                return
                
            # Abrir arquivo de histórico
            wb = load_workbook(drive_path)
            ws = wb["Histórico"]
            
            # Preencher a treeview com os registros filtrados
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Se não tiver data de registro
                    continue
                    
                # Processar data (formato: dd/mm/yyyy hh:mm:ss)
                try:
                    data_registro = datetime.strptime(row[0], '%d/%m/%Y %H:%M:%S').date()
                except ValueError:
                    # Tentar outro formato
                    try:
                        data_registro = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S').date()
                    except ValueError:
                        # Se não conseguir interpretar, pular
                        continue
                
                # Verificar filtro de data
                if data_registro < data_inicial.date() or data_registro > data_final.date():
                    continue
                    
                # Verificar filtro de descrição
                if descricao and descricao not in str(row[2]).upper():
                    continue
                    
                # Adicionar à treeview
                self.tree_historico.insert('', 'end', values=row)
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao filtrar histórico: {str(e)}")
    
    def mostrar_detalhes_rateio(self, event):
        """Mostra os detalhes do rateio selecionado"""
        try:
            from src.config.config import BASE_PATH  # Importar o caminho base
            
            # Obter item selecionado
            selecionado = self.tree_historico.selection()
            if not selecionado:
                return
                    
            # Obter ID do rateio (linha)
            linha = self.tree_historico.index(selecionado[0]) + 2  # +2 pois o índice começa em 0 e temos o cabeçalho
            
            # Limpar a treeview de detalhes
            for item in self.tree_detalhes.get_children():
                self.tree_detalhes.delete(item)
                    
            # Caminho do arquivo no Google Drive
            drive_path = Path(BASE_PATH) / "Financeiro" / "Planilhas_Base" / "historico_rateios.xlsx"
            
            # Verificar se o arquivo de histórico existe
            if not drive_path.exists():
                return
                    
            # Abrir arquivo de histórico
            wb = load_workbook(drive_path)
            if "Detalhes" not in wb.sheetnames:
                return
                    
            ws = wb["Detalhes"]
            
            # Preencher a treeview com os detalhes do rateio
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == linha - 1:  # ID do rateio
                    # Formatar valor
                    if row[2] is not None:
                        try:
                            valor = float(row[2])
                            valor_formatado = f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                        except:
                            valor_formatado = str(row[2])
                    else:
                        valor_formatado = "R$ 0,00"
                    
                    # Inserir na treeview com o valor formatado
                    self.tree_detalhes.insert('', 'end', values=(
                        row[1],  # Cliente
                        valor_formatado,  # Valor formatado
                        row[3]   # Status
                    ))
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao mostrar detalhes do rateio: {str(e)}")
    
    def voltar_menu(self):
        """Volta ao menu principal"""
        if self.menu_principal:
            self.root.destroy()
            self.menu_principal.deiconify()
            self.menu_principal.lift()

class GerenciadorDespesasRateadas:

    def carregar_clientes_ativos(self):
        """Carrega todos os clientes ativos do sistema (sem data final)"""
        clientes = []
        try:
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            # Lista temporária para armazenar clientes antes de ordenar
            clientes_temp = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome não vazio
                    # Verificar se a data final está vazia (cliente ativo)
                    data_final = row[4] if len(row) > 4 else None
                    
                    if not data_final:  # Se não tiver data final, é um cliente ativo
                        clientes_temp.append({
                            'nome': row[0],
                            'percentual': 0,
                            'valor': 0,
                            'arquivo': PASTA_CLIENTES / f"{row[0]}.xlsx"
                        })
            
            # Ordenar a lista por nome antes de retornar
            clientes = sorted(clientes_temp, key=lambda x: x['nome'])
            
            wb.close()
            return clientes
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
            return []
        
    def calcular_rateio(self):
        """Calcula o rateio baseado nos percentuais ou valores definidos"""
        if self.modo_rateio.get() == "percentual":
            # Verificar se o total é 100%
            total_percentual = sum(cliente['percentual'] for cliente in self.clientes)
            if not (99.9 <= total_percentual <= 100.1):  # Tolerância para arredondamentos
                messagebox.showerror("Erro", f"O total de percentuais deve ser 100%. Atual: {total_percentual}%")
                return False
                
            # Calcular valores baseados nos percentuais
            valor_total = float(self.valor_total.get().replace(',', '.'))
            for cliente in self.clientes:
                cliente['valor'] = (cliente['percentual'] / 100) * valor_total
        else:  # modo = valor
            # Verificar se o total corresponde ao valor da despesa
            total_valores = sum(cliente['valor'] for cliente in self.clientes)
            valor_total = float(self.valor_total.get().replace(',', '.'))
            
            if abs(total_valores - valor_total) > 0.01:  # Tolerância de 1 centavo
                messagebox.showerror("Erro", 
                                    f"O total dos valores ({total_valores:.2f}) não corresponde ao valor da despesa ({valor_total:.2f})")
                return False
                
        return True

    def aplicar_rateio(self):
        """Aplica o rateio nos arquivos de cada cliente"""
        data_rel = self.data_rel.get_date()
        descricao = self.descricao.get()
        tipo_despesa = self.tipo_despesa.get()
        observacao = self.observacao.get()
        
        # Lista para registrar resultados
        registros = []
        
        for cliente in self.clientes:
            if cliente['valor'] <= 0:
                continue  # Pular clientes sem valor
                
            try:
                wb = load_workbook(cliente['arquivo'])
                ws = wb["Dados"]
                
                # Preparar dados do lançamento
                proxima_linha = ws.max_row + 1
                
                # Data do Relatório (formatada)
                ws.cell(row=proxima_linha, column=1, value=data_rel)
                ws.cell(row=proxima_linha, column=1).number_format = 'DD/MM/YYYY'
                
                # Tipo de Despesa
                ws.cell(row=proxima_linha, column=2, value=int(tipo_despesa))
                
                # CNPJ/CPF do sistema (se aplicável)
                ws.cell(row=proxima_linha, column=3, value=cnpj_cpf)
                
                # Nome do sistema
                ws.cell(row=proxima_linha, column=4, value=nome)
                
                # Referência
                ws.cell(row=proxima_linha, column=5, value=f"{descricao}")
                
                # NF (vazio para rateios)
                ws.cell(row=proxima_linha, column=6, value="")
                
                # Valor Unitário
                ws.cell(row=proxima_linha, column=7, value=cliente['valor'])
                ws.cell(row=proxima_linha, column=7).number_format = '#,##0.00'
                
                # Dias (1 para despesas rateadas)
                ws.cell(row=proxima_linha, column=8, value=1)
                
                # Valor Total
                ws.cell(row=proxima_linha, column=9, value=cliente['valor'])
                ws.cell(row=proxima_linha, column=9).number_format = '#,##0.00'
                
                # Data de Vencimento (mesma data do relatório por padrão)
                ws.cell(row=proxima_linha, column=10, value=data_rel)
                ws.cell(row=proxima_linha, column=10).number_format = 'DD/MM/YYYY'
                
                # Categoria
                ws.cell(row=proxima_linha, column=11, value="MO")
                
                # Dados Bancários (vazio para rateios)
                dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj_cpf)
                ws.cell(row=proxima_linha, column=12, value=dados_bancarios)
                
                # Observação
                ws.cell(row=proxima_linha, column=13, value="LANÇAMENTO AUTOMÁTICO")
                
                # Salvar planilha
                wb.save(cliente['arquivo'])
                
                # Registrar sucesso
                registros.append({
                    'cliente': cliente['nome'],
                    'valor': cliente['valor'],
                    'status': 'SUCESSO'
                })
                
            except Exception as e:
                # Registrar falha
                registros.append({
                    'cliente': cliente['nome'],
                    'valor': cliente['valor'],
                    'status': f'FALHA: {str(e)}'
                })
        
        # Registrar o rateio no histórico
        self.registrar_historico(registros)
        
        # Exibir resultados
        self.mostrar_resultado_rateio(registros)

    def registrar_historico(self, registros):
        """Registra o rateio no histórico"""
        try:
            data_atual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            data_rel = self.data_rel.get()
            descricao = self.descricao.get()
            valor_total = float(self.valor_total.get().replace(',', '.'))
            tipo_despesa = self.tipo_despesa.get()
            
            # Criar arquivo de histórico se não existir
            historico_path = Path('historico_rateios.xlsx')
            if not historico_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.title = "Histórico"
                
                # Cabeçalhos
                headers = ['Data Registro', 'Data Referência', 'Descrição', 'Valor Total', 
                        'Tipo Despesa', 'Qtd Clientes', 'Status']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    
                wb.save(historico_path)
            
            # Abrir arquivo de histórico
            wb = load_workbook(historico_path)
            ws = wb["Histórico"]
            
            # Adicionar registro principal
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1, value=data_atual)
            ws.cell(row=proxima_linha, column=2, value=data_rel)
            ws.cell(row=proxima_linha, column=3, value=descricao)
            ws.cell(row=proxima_linha, column=4, value=valor_total)
            ws.cell(row=proxima_linha, column=5, value=tipo_despesa)
            ws.cell(row=proxima_linha, column=6, value=len(registros))
            
            # Verificar status geral
            falhas = [r for r in registros if r['status'].startswith('FALHA')]
            status = "SUCESSO" if not falhas else f"PARCIAL ({len(falhas)} falhas)"
            ws.cell(row=proxima_linha, column=7, value=status)
            
            # Adicionar detalhes em outra aba se não existir
            if "Detalhes" not in wb.sheetnames:
                ws_details = wb.create_sheet("Detalhes")
                # Cabeçalhos
                headers = ['ID Rateio', 'Cliente', 'Valor', 'Status']
                for col, header in enumerate(headers, 1):
                    ws_details.cell(row=1, column=col, value=header)
            else:
                ws_details = wb["Detalhes"]
            
            # Adicionar detalhes de cada cliente
            id_rateio = proxima_linha - 1  # Usar a linha como ID do rateio
            for registro in registros:
                proxima_linha_det = ws_details.max_row + 1
                ws_details.cell(row=proxima_linha_det, column=1, value=id_rateio)
                ws_details.cell(row=proxima_linha_det, column=2, value=registro['cliente'])
                ws_details.cell(row=proxima_linha_det, column=3, value=registro['valor'])
                ws_details.cell(row=proxima_linha_det, column=4, value=registro['status'])
            
            # Salvar histórico
            wb.save(historico_path)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao registrar histórico: {str(e)}")