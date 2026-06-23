# utils.py

from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from validate_docbr import CPF, CNPJ
import os
from openpyxl import load_workbook
import re
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime

from pathlib import Path


from src.config.config import (
    BASE_PATH,
    PASTA_CLIENTES,
    ARQUIVO_CLIENTES,
    ARQUIVO_FORNECEDORES,
    ARQUIVO_MODELO,
    ARQUIVO_CONTROLE
)


# === VALIDAÇÃO DE DOCUMENTOS ===
def validar_cpf(cpf):
    """
    Valida CPF usando algoritmo oficial dos dígitos verificadores
    
    Args:
        cpf: string com 11 dígitos numéricos (com ou sem formatação)
    
    Returns:
        bool: True se CPF válido, False caso contrário
    """
    try:
        # Remover caracteres não numéricos
        cpf_numeros = ''.join(filter(str.isdigit, str(cpf)))
        
        # Verificar se tem 11 dígitos
        if len(cpf_numeros) != 11:
            return False
        
        # Verificar se todos os dígitos são iguais (CPF inválido)
        if cpf_numeros == cpf_numeros[0] * 11:
            return False
        
        # Calcular primeiro dígito verificador
        soma = 0
        for i in range(9):
            soma += int(cpf_numeros[i]) * (10 - i)
        
        resto = soma % 11
        digito1 = 0 if resto < 2 else 11 - resto
        
        if int(cpf_numeros[9]) != digito1:
            return False
        
        # Calcular segundo dígito verificador
        soma = 0
        for i in range(10):
            soma += int(cpf_numeros[i]) * (11 - i)
        
        resto = soma % 11
        digito2 = 0 if resto < 2 else 11 - resto
        
        return int(cpf_numeros[10]) == digito2
        
    except Exception as e:
        print(f"Erro ao validar CPF: {str(e)}")
        return False


def validar_cnpj(cnpj):
    """
    Valida CNPJ usando algoritmo oficial dos dígitos verificadores
    
    Args:
        cnpj: string com 14 dígitos numéricos (com ou sem formatação)
    
    Returns:
        bool: True se CNPJ válido, False caso contrário
    """
    try:
        # Remover caracteres não numéricos
        cnpj_numeros = ''.join(filter(str.isdigit, str(cnpj)))
        
        # Verificar se tem 14 dígitos
        if len(cnpj_numeros) != 14:
            return False
        
        # Verificar se todos os dígitos são iguais (CNPJ inválido)
        if cnpj_numeros == cnpj_numeros[0] * 14:
            return False
        
        # Calcular primeiro dígito verificador
        peso = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        soma = sum(int(cnpj_numeros[i]) * peso[i] for i in range(12))
        resto = soma % 11
        digito1 = 0 if resto < 2 else 11 - resto
        
        if int(cnpj_numeros[12]) != digito1:
            return False
        
        # Calcular segundo dígito verificador
        peso = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        soma = sum(int(cnpj_numeros[i]) * peso[i] for i in range(13))
        resto = soma % 11
        digito2 = 0 if resto < 2 else 11 - resto
        
        return int(cnpj_numeros[13]) == digito2
        
    except Exception as e:
        print(f"Erro ao validar CNPJ: {str(e)}")
        return False


def validar_documento(documento):
    """
    Valida CPF ou CNPJ automaticamente baseado no tamanho
    
    Args:
        documento: string com CPF (11 dígitos) ou CNPJ (14 dígitos)
    
    Returns:
        bool: True se documento válido, False caso contrário
    """
    try:
        # Remover caracteres não numéricos
        doc_numeros = ''.join(filter(str.isdigit, str(documento)))
        
        if len(doc_numeros) == 11:
            return validar_cpf(doc_numeros)
        elif len(doc_numeros) == 14:
            return validar_cnpj(doc_numeros)
        else:
            return False
            
    except Exception as e:
        print(f"Erro ao validar documento: {str(e)}")
        return False


# === CLIENTES ===
def obter_clientes_ativos(mostrar_inativos=False):
    """
    Obtém a lista de clientes ativos (sem data final ou com data final futura)
    
    Args:
        mostrar_inativos (bool): Se True, retorna todos os clientes, incluindo inativos
                                 Se False, retorna apenas clientes ativos
    
    Returns:
        list: Lista de nomes dos clientes ativos (ou todos os clientes se mostrar_inativos=True)
        dict: Dicionário com informações adicionais de cada cliente (pode ser útil para exibir status)
    """
    try:
        # Verifica se o arquivo de clientes existe
        if not os.path.exists(ARQUIVO_CLIENTES):
            print(f"Arquivo de clientes não encontrado: {ARQUIVO_CLIENTES}")
            return [], {}
        
        # Carrega os dados do arquivo de clientes
        df_clientes = pd.read_excel(ARQUIVO_CLIENTES, sheet_name='Clientes')
        
        # Data atual para comparação
        data_atual = datetime.now().date()
        
        # Dicionário para armazenar informações adicionais de clientes
        info_clientes = {}
        
        # Lista de clientes ativos
        clientes_ativos = []
        
        # Itera sobre os clientes (assumindo que a coluna A tem o nome e a coluna E tem a data final)
        for _, row in df_clientes.iterrows():
            # Pular linhas sem nome de cliente
            if pd.isna(row.iloc[0]) or not row.iloc[0]:
                continue
            
            nome_cliente = str(row.iloc[0]).strip()
            data_final = None
            
            # Verifica se a coluna E (índice 4) existe e tem uma data final
            if len(row) > 4 and not pd.isna(row.iloc[4]):
                try:
                    # Tenta converter para data
                    data_final = pd.to_datetime(row.iloc[4]).date()
                except:
                    # Se falhar, tenta outros formatos ou deixa como None
                    try:
                        if isinstance(row.iloc[4], str):
                            data_final = datetime.strptime(row.iloc[4], '%d/%m/%Y').date()
                    except:
                        data_final = None
            
            # Verifica se o cliente está ativo (sem data final ou data final futura)
            cliente_ativo = data_final is None or data_final > data_atual
            
            # Armazenar informações adicionais
            info_clientes[nome_cliente] = {
                'ativo': cliente_ativo,
                'data_final': data_final,
                'arquivo': os.path.join(PASTA_CLIENTES, f"{nome_cliente}.xlsx")
            }
            
            # Adicionar à lista de clientes ativos ou todos, dependendo do parâmetro
            if mostrar_inativos or cliente_ativo:
                clientes_ativos.append(nome_cliente)
        
        # Retorna a lista ordenada e o dicionário de informações
        return sorted(clientes_ativos), info_clientes
        
    except Exception as e:
        print(f"Erro ao obter clientes ativos: {str(e)}")
        import traceback
        traceback.print_exc()
        return [], {}


def atualizar_combobox_clientes(combobox, mostrar_inativos=False):
    """
    Atualiza um combobox com a lista de clientes ativos
    
    Args:
        combobox: O widget Combobox a ser atualizado
        mostrar_inativos (bool): Se True, inclui clientes inativos no combobox
    
    Returns:
        dict: Dicionário com informações dos clientes
    """
    clientes, info_clientes = obter_clientes_ativos(mostrar_inativos)
    
    # Limpar lista atual
    combobox['values'] = []
    
    # Atualizar combobox
    combobox['values'] = clientes
    
    # Se houver clientes, selecionar o primeiro
    if clientes:
        combobox.current(0)
    
    return info_clientes


def cliente_esta_ativo(nome_cliente):
    """
    Verifica se um cliente específico está ativo
    
    Args:
        nome_cliente (str): Nome do cliente a verificar
    
    Returns:
        bool: True se o cliente estiver ativo, False caso contrário
    """
    _, info_clientes = obter_clientes_ativos(mostrar_inativos=True)
    
    if nome_cliente in info_clientes:
        return info_clientes[nome_cliente]['ativo']
    
    # Se o cliente não for encontrado, retorna False
    return False


def obter_info_cliente(nome_cliente):
    """
    Obtém informações detalhadas sobre um cliente específico
    
    Args:
        nome_cliente (str): Nome do cliente
    
    Returns:
        dict: Dicionário com informações do cliente ou None se não encontrado
    """
    _, info_clientes = obter_clientes_ativos(mostrar_inativos=True)
    
    if nome_cliente in info_clientes:
        return info_clientes[nome_cliente]
    
    return None


# === DATA VALIDATION AND FORMATTING ===
def validar_data(data_str):
    """Valida se uma string está no formato de data correto"""
    try:
        if not data_str:
            return False
        if not re.match(r'^\d{2}/\d{2}/\d{4}$', data_str):
            return False
        datetime.strptime(data_str, '%d/%m/%Y')
        return True
    except ValueError:
        return False

def validar_data_quinzena(data):
    """Valida se a data é dia 5 ou 20 e ajusta se necessário"""
    if not (data.day == 5 or data.day == 20):
        if data.day < 5:
            data_ajustada = data.replace(day=5)
            msg = f"Data ajustada para dia 5: {data_ajustada.strftime('%d/%m/%Y')}"
        elif data.day < 20:
            data_ajustada = data.replace(day=20)
            msg = f"Data ajustada para dia 20: {data_ajustada.strftime('%d/%m/%Y')}"
        else:
            if data.month == 12:
                data_ajustada = data.replace(year=data.year + 1, month=1, day=5)
            else:
                data_ajustada = data.replace(month=data.month + 1, day=5)
            msg = f"Data ajustada para dia 5 do próximo mês: {data_ajustada.strftime('%d/%m/%Y')}"
        return data_ajustada, msg
    return data, None

def calcular_proxima_data_quinzena(data):
    """Calcula a próxima data quinzenal"""
    if data.day == 5:
        return data.replace(day=20)
    else:
        if data.month == 12:
            return data.replace(year=data.year + 1, month=1, day=5)
        else:
            return data.replace(month=data.month + 1, day=5)

# === DOCUMENT VALIDATION ===
# def validar_cnpj_cpf(documento):
#     """Valida CNPJ ou CPF"""
#     if len(documento) <= 11:
#         cpf = CPF()
#         return cpf.validate(documento)
#     else:
#         cnpj = CNPJ()
#         return cnpj.validate(documento)

def formatar_cnpj_cpf(documento):
    """Formata CNPJ/CPF com zeros à esquerda"""
    if len(documento) <= 11:
        return documento.zfill(11)
    return documento.zfill(14)

def normalizar_documento(valor_excel, tipo_pessoa):
    """
    Normaliza CNPJ ou CPF baseado no tipo de pessoa (PJ/PF).
    
    Esta é a função CENTRAL para normalização. Use sempre que possível.
    A partir de junho/2026, CNPJ poderá conter letras (IN RFB 2.229/2024).
    
    Args:
        valor_excel: Valor do documento (pode ser int, float ou string)
        tipo_pessoa: 'PF' para CPF ou 'PJ' para CNPJ
    
    Returns:
        String com documento normalizado (11 dígitos para CPF, 14 para CNPJ)
    
    Examples:
        normalizar_documento(12345678901, 'PF') -> '12345678901'
        normalizar_documento(123456, 'PF') -> '00000123456'
        normalizar_documento(12345678000190, 'PJ') -> '12345678000190'
    """
    # Converter para string preservando dígitos E letras (futuro CNPJ alfanumérico)
    if isinstance(valor_excel, (int, float)):
        texto = str(int(valor_excel))
    else:
        texto = str(valor_excel).strip().upper()
    
    # Para CPF: extrair apenas números (CPF permanece numérico)
    if tipo_pessoa == 'PF':
        numeros = ''.join(filter(str.isdigit, texto))
        return numeros.zfill(11) if numeros else ""
    
    # Para CNPJ: extrair números E letras (preparado para IN 2.229/2024)
    # Mantém alfanuméricos, remove pontuação/espaços
    alfanumerico = ''.join(c for c in texto if c.isalnum())
    return alfanumerico.zfill(14) if alfanumerico else ""


def formatar_documento(documento, tipo_pessoa):
    """
    Formata documento COM MÁSCARA para exibição visual.
    
    Args:
        documento: Documento já normalizado ou valor bruto
        tipo_pessoa: 'PF' para CPF ou 'PJ' para CNPJ
    
    Returns:
        String formatada com máscara:
        - CPF: 000.000.000-00
        - CNPJ: 00.000.000/0000-00
    
    Examples:
        formatar_documento_com_mascara('12345678901', 'PF') -> '123.456.789-01'
        formatar_documento_com_mascara(12345678000190, 'PJ') -> '12.345.678/0001-90'
    """
    # Normalizar primeiro
    doc_normalizado = normalizar_documento(documento, tipo_pessoa)
    
    if not doc_normalizado:
        return ""
    
    # Aplicar máscara
    if tipo_pessoa == 'PF':
        # CPF: 000.000.000-00
        return f"{doc_normalizado[:3]}.{doc_normalizado[3:6]}.{doc_normalizado[6:9]}-{doc_normalizado[9:11]}"
    else:  # PJ
        # CNPJ: 00.000.000/0000-00
        # Suporta alfanuméricos (letras ficam visíveis na máscara)
        return f"{doc_normalizado[:2]}.{doc_normalizado[2:5]}.{doc_normalizado[5:8]}/{doc_normalizado[8:12]}-{doc_normalizado[12:14]}"


def validar_documento(documento, tipo_pessoa):
    """
    Valida CNPJ ou CPF usando biblioteca validate-docbr.
    
    ATENÇÃO: Esta validação usa algoritmos atuais (pré-2026).
    A partir de junho/2026, CNPJs alfanuméricos exigirão novo algoritmo da RFB.
    
    Args:
        documento: Documento a validar (normalizado ou não)
        tipo_pessoa: 'PF' para CPF ou 'PJ' para CNPJ
    
    Returns:
        bool: True se válido, False caso contrário
    """
    # Normalizar primeiro
    doc_normalizado = normalizar_documento(documento, tipo_pessoa)
    
    if not doc_normalizado:
        return False
    
    try:
        if tipo_pessoa == 'PF':
            cpf = CPF()
            return cpf.validate(doc_normalizado)
        else:  # PJ
            # TODO: Atualizar validação quando RFB publicar algoritmo para CNPJ alfanumérico
            cnpj = CNPJ()
            # Validar apenas se for numérico (CNPJs alfanuméricos ainda não têm validação)
            if doc_normalizado.isdigit():
                return cnpj.validate(doc_normalizado)
            else:
                # CNPJ alfanumérico: aceitar formato mas sem validação de dígito
                return len(doc_normalizado) == 14
    except:
        return False

# === FORMATADORES AUTOMÁTICOS PARA CAMPOS DE ENTRADA ===

def formatar_cpf_campo(event):
    """
    Formata CPF automaticamente durante a digitação em campos Entry: 999.999.999-99
    
    Args:
        event: Evento do Tkinter (KeyRelease)
    
    Uso:
        cpf_entry.bind('<KeyRelease>', formatar_cpf_campo)
    """
    widget = event.widget
    
    # Obter valor atual
    valor = widget.get()
    
    # Remover tudo que não é número
    apenas_numeros = re.sub(r'\D', '', valor)
    
    # Limitar a 11 dígitos
    apenas_numeros = apenas_numeros[:11]
    
    # Aplicar formatação
    if len(apenas_numeros) <= 3:
        formatado = apenas_numeros
    elif len(apenas_numeros) <= 6:
        formatado = f"{apenas_numeros[:3]}.{apenas_numeros[3:]}"
    elif len(apenas_numeros) <= 9:
        formatado = f"{apenas_numeros[:3]}.{apenas_numeros[3:6]}.{apenas_numeros[6:]}"
    else:
        formatado = f"{apenas_numeros[:3]}.{apenas_numeros[3:6]}.{apenas_numeros[6:9]}-{apenas_numeros[9:]}"
    
    # Atualizar campo se mudou
    if formatado != valor:
        # Salvar posição do cursor
        pos_cursor = widget.index(tk.INSERT)
        
        # Calcular nova posição (ajustar por caracteres inseridos)
        diff = len(formatado) - len(valor)
        nova_pos = pos_cursor + diff
        
        # Atualizar valor
        widget.delete(0, tk.END)
        widget.insert(0, formatado)
        
        # Restaurar cursor (limitar à posição válida)
        nova_pos = max(0, min(nova_pos, len(formatado)))
        widget.icursor(nova_pos)


def formatar_cno_campo(event):
    """
    Formata CNO automaticamente durante a digitação: 99.999.99999/99
    
    Formato: XX.XXX.XXXXX/XX (12 dígitos)
    Exemplo: 02.043.11375/74
    
    Args:
        event: Evento do Tkinter (KeyRelease)
    
    Uso:
        cno_entry.bind('<KeyRelease>', formatar_cno_campo)
    """
    widget = event.widget
    
    # Obter valor atual
    valor = widget.get()
    
    # Remover tudo que não é número
    apenas_numeros = re.sub(r'\D', '', valor)
    
    # Limitar a 12 dígitos
    apenas_numeros = apenas_numeros[:12]
    
    # Aplicar formatação
    if len(apenas_numeros) <= 2:
        formatado = apenas_numeros
    elif len(apenas_numeros) <= 5:
        formatado = f"{apenas_numeros[:2]}.{apenas_numeros[2:]}"
    elif len(apenas_numeros) <= 10:
        formatado = f"{apenas_numeros[:2]}.{apenas_numeros[2:5]}.{apenas_numeros[5:]}"
    else:
        formatado = f"{apenas_numeros[:2]}.{apenas_numeros[2:5]}.{apenas_numeros[5:10]}/{apenas_numeros[10:]}"
    
    # Atualizar campo se mudou
    if formatado != valor:
        # Salvar posição do cursor
        pos_cursor = widget.index(tk.INSERT)
        
        # Calcular nova posição (ajustar por caracteres inseridos)
        diff = len(formatado) - len(valor)
        nova_pos = pos_cursor + diff
        
        # Atualizar valor
        widget.delete(0, tk.END)
        widget.insert(0, formatado)
        
        # Restaurar cursor (limitar à posição válida)
        nova_pos = max(0, min(nova_pos, len(formatado)))
        widget.icursor(nova_pos)


def formatar_cep_campo(event):
    """
    Formata CEP automaticamente durante a digitação: 99999-999
    
    Args:
        event: Evento do Tkinter (KeyRelease)
    
    Uso:
        cep_entry.bind('<KeyRelease>', formatar_cep_campo)
    """
    widget = event.widget
    
    # Obter valor atual
    valor = widget.get()
    
    # Remover tudo que não é número
    apenas_numeros = re.sub(r'\D', '', valor)
    
    # Limitar a 8 dígitos
    apenas_numeros = apenas_numeros[:8]
    
    # Aplicar formatação
    if len(apenas_numeros) <= 5:
        formatado = apenas_numeros
    else:
        formatado = f"{apenas_numeros[:5]}-{apenas_numeros[5:]}"
    
    # Atualizar campo se mudou
    if formatado != valor:
        # Salvar posição do cursor
        pos_cursor = widget.index(tk.INSERT)
        
        # Calcular nova posição
        diff = len(formatado) - len(valor)
        nova_pos = pos_cursor + diff
        
        # Atualizar valor
        widget.delete(0, tk.END)
        widget.insert(0, formatado)
        
        # Restaurar cursor
        nova_pos = max(0, min(nova_pos, len(formatado)))
        widget.icursor(nova_pos)


def formatar_telefone_campo(event):
    """
    Formata telefone automaticamente: (99) 99999-9999 ou (99) 9999-9999
    
    Args:
        event: Evento do Tkinter (KeyRelease)
    
    Uso:
        telefone_entry.bind('<KeyRelease>', formatar_telefone_campo)
    """
    widget = event.widget
    
    # Obter valor atual
    valor = widget.get()
    
    # Remover tudo que não é número
    apenas_numeros = re.sub(r'\D', '', valor)
    
    # Limitar a 11 dígitos
    apenas_numeros = apenas_numeros[:11]
    
    # Aplicar formatação
    if len(apenas_numeros) <= 2:
        formatado = apenas_numeros
    elif len(apenas_numeros) <= 6:
        formatado = f"({apenas_numeros[:2]}) {apenas_numeros[2:]}"
    elif len(apenas_numeros) <= 10:
        # Telefone fixo: (XX) XXXX-XXXX
        formatado = f"({apenas_numeros[:2]}) {apenas_numeros[2:6]}-{apenas_numeros[6:]}"
    else:
        # Celular: (XX) XXXXX-XXXX
        formatado = f"({apenas_numeros[:2]}) {apenas_numeros[2:7]}-{apenas_numeros[7:]}"
    
    # Atualizar campo se mudou
    if formatado != valor:
        pos_cursor = widget.index(tk.INSERT)
        diff = len(formatado) - len(valor)
        nova_pos = pos_cursor + diff
        
        widget.delete(0, tk.END)
        widget.insert(0, formatado)
        
        nova_pos = max(0, min(nova_pos, len(formatado)))
        widget.icursor(nova_pos)


def limpar_formatacao(texto):
    """
    Remove toda formatação de um texto, deixando apenas números
    
    Args:
        texto (str): Texto formatado
    
    Returns:
        str: Apenas números
    
    Exemplo:
        >>> limpar_formatacao("123.456.789-10")
        "12345678910"
        >>> limpar_formatacao("02.043.11375/74")
        "020431137574"
    """
    if not texto:
        return ""
    return re.sub(r'\D', '', str(texto))


def validar_cpf_completo(cpf):
    """
    Valida CPF usando algoritmo oficial
    
    Args:
        cpf (str): CPF formatado ou não
    
    Returns:
        bool: True se válido, False se inválido
    
    Exemplo:
        >>> validar_cpf_completo("123.456.789-10")
        False
        >>> validar_cpf_completo("11144477735")
        True
    """
    # Limpar formatação
    cpf_limpo = limpar_formatacao(cpf)
    
    # Validar tamanho
    if len(cpf_limpo) != 11:
        return False
    
    # Verificar se todos os dígitos são iguais
    if cpf_limpo == cpf_limpo[0] * 11:
        return False
    
    # Calcular primeiro dígito verificador
    soma = sum(int(cpf_limpo[i]) * (10 - i) for i in range(9))
    digito1 = 11 - (soma % 11)
    digito1 = 0 if digito1 > 9 else digito1
    
    # Verificar primeiro dígito
    if int(cpf_limpo[9]) != digito1:
        return False
    
    # Calcular segundo dígito verificador
    soma = sum(int(cpf_limpo[i]) * (11 - i) for i in range(10))
    digito2 = 11 - (soma % 11)
    digito2 = 0 if digito2 > 9 else digito2
    
    # Verificar segundo dígito
    return int(cpf_limpo[10]) == digito2


def validar_cno(cno):
    """
    Valida CNO (validação básica de formato)
    
    Args:
        cno (str): CNO formatado ou não
    
    Returns:
        bool: True se tem 12 dígitos, False caso contrário
    
    Nota: CNO não possui dígito verificador padrão,
    apenas validamos se tem 12 dígitos numéricos
    """
    cno_limpo = limpar_formatacao(cno)
    return len(cno_limpo) == 12


def validar_cep(cep):
    """
    Valida CEP (validação básica de formato)
    
    Args:
        cep (str): CEP formatado ou não
    
    Returns:
        bool: True se tem 8 dígitos, False caso contrário
    """
    cep_limpo = limpar_formatacao(cep)
    return len(cep_limpo) == 8

# === FILE OPERATIONS ===
def verificar_arquivo_excel(caminho):
    """Verifica se arquivo Excel existe e pode ser aberto"""
    try:
        if not os.path.exists(caminho):
            return False
        wb = load_workbook(caminho)
        wb.close()
        return True
    except Exception:
        return False

# === VALUE FORMATTING ===
def formatar_moeda(valor):
    """Formata valor para moeda brasileira"""
    try:
        return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return "R$ 0,00"


# === FIND SUPPLIER ===
def buscar_fornecedor(tree_fornecedores, termo_busca='', categoria_filtro=None):
    """
    Busca fornecedores na base e atualiza o treeview
    
    Parâmetros:
    - tree_fornecedores: Treeview onde serão exibidos os resultados
    - termo_busca: Termo para filtrar por nome/razão social (opcional)
    - categoria_filtro: Filtrar por categoria específica, ex: 'TAX' (opcional)
    """
    # Limpar resultados anteriores
    for item in tree_fornecedores.get_children():
        tree_fornecedores.delete(item)
    
    try:
        # Verificar se arquivo existe
        if not ARQUIVO_FORNECEDORES.exists():
            messagebox.showerror("Erro", f"Arquivo não encontrado: {ARQUIVO_FORNECEDORES}")
            return
        
        wb = load_workbook(ARQUIVO_FORNECEDORES)
        ws = wb['Fornecedores']
        
        # Normalizar termo de busca
        termo = termo_busca.lower().strip() if termo_busca else ''
        categoria_upper = categoria_filtro.upper().strip() if categoria_filtro else None
        
        resultados_encontrados = 0
        
        # Iterar pelas linhas (começando da linha 2)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # Estrutura esperada da planilha:
                # row[0] = CNPJ/CPF (coluna A)
                # row[1] = tipo_pessoa (coluna B)
                # row[2] = RAZÃO SOCIAL (coluna C)
                # row[3] = NOME (coluna D)
                # row[11] = CATEGORIA: (coluna L)
                
                cnpj_cpf = str(row[0]).strip() if row[0] else ''
                tipo_pessoa = str(row[1]).strip().upper() if row[1] else 'PJ'  # Default PJ
                razao_social = str(row[2]).strip() if row[2] else ''
                nome = str(row[3]).strip() if row[3] else ''
                categoria = str(row[11]).strip().upper() if row[11] else ''
                
                # ✅ FILTRO 1: Categoria (se especificado)
                if categoria_upper and categoria != categoria_upper:
                    continue
                
                # ✅ FILTRO 2: Termo de busca (busca em NOME, RAZÃO SOCIAL e CNPJ/CPF)
                if termo:
                    if not (termo in nome.lower() or 
                           termo in razao_social.lower() or 
                           termo in cnpj_cpf.lower()):
                        continue
                
                # Definir qual nome exibir (priorizar RAZÃO SOCIAL)
                nome_exibir = razao_social if razao_social else nome
                
                # ✅ Inserir com tipo_pessoa nas TAGS (oculto mas acessível)
                tree_fornecedores.insert('', 'end', 
                                       values=(cnpj_cpf, nome_exibir, categoria),
                                       tags=(tipo_pessoa,))
                
                resultados_encontrados += 1
                
            except Exception as e_row:
                # Logar erro mas continuar processando outras linhas
                # logger.debug(f"Erro ao processar linha: {e_row}")
                continue
        
        wb.close()
        
        # Mensagem se não encontrou resultados
        if resultados_encontrados == 0:
            mensagem = "Nenhum fornecedor encontrado"
            if categoria_upper:
                mensagem += f" na categoria '{categoria_filtro}'"
            if termo:
                mensagem += f" com o termo '{termo_busca}'"
            messagebox.showinfo("Aviso", mensagem)
        
    except Exception as e:
        import traceback
        # logger.debug(traceback.format_exc())
        messagebox.showerror("Erro", f"Erro ao buscar fornecedores: {str(e)}")

def selecionar_fornecedor(tree_fornecedores, campos_fornecedor, campos_despesa=None, notebook=None, buscar_fornecedor_completo=None):
    """Preenche campos com o fornecedor selecionado"""
    selecionado = tree_fornecedores.selection()
    if not selecionado:
        messagebox.showwarning("Aviso", "Selecione um fornecedor")
        return None

    fornecedor = tree_fornecedores.item(selecionado)['values']

    # Para o caso de campos simplificados (apenas cnpj_cpf e nome)
    if 'cnpj_cpf' in campos_fornecedor and 'nome' in campos_fornecedor:
        for campo in ['cnpj_cpf', 'nome']:
            campos_fornecedor[campo].config(state='normal')
            campos_fornecedor[campo].delete(0, tk.END)
            idx = 0 if campo == 'cnpj_cpf' else 1
            campos_fornecedor[campo].insert(0, str(fornecedor[idx]))
            campos_fornecedor[campo].config(state='readonly')
        return fornecedor
    
    # Limpar e preencher campos básicos
    for entry in campos_fornecedor.values():
        entry.config(state='normal')
        entry.delete(0, tk.END)

    campos_fornecedor['cnpj_cpf'].insert(0, str(fornecedor[0]).zfill(14))
    campos_fornecedor['nome'].insert(0, fornecedor[1])
    campos_fornecedor['categoria'].insert(0, fornecedor[2])

    # Tratamento de dados bancários se necessário
    if 'dados_bancarios' in campos_fornecedor and buscar_fornecedor_completo:
        campos_fornecedor['dados_bancarios'].config(state='normal')
        campos_fornecedor['dados_bancarios'].delete(0, tk.END)

        tp_desp = campos_despesa['tp_desp'].get().strip() if campos_despesa else ''
        fornecedor_completo = buscar_fornecedor_completo(fornecedor[0])

        if fornecedor_completo:
            if tp_desp not in ['3', '5']:
                if fornecedor_completo['chave_pix']:
                    dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
                else:
                    dados_bancarios = (f"{fornecedor_completo['banco'] or ''} "
                                     f"{fornecedor_completo['op'] or ''} - "
                                     f"{fornecedor_completo['agencia'] or ''} "
                                     f"{fornecedor_completo['conta'] or ''}").strip()
                if dados_bancarios.strip() in ['', ' - ']:
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
            else:
                dados_bancarios = ''
            
            campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)

    # Configurar estados finais
    for campo, entry in campos_fornecedor.items():
        if campo != 'categoria':
            entry.config(state='readonly')

    # Mudar para próxima aba se necessário
    if notebook:
        notebook.select(2)

    return fornecedor


# === VALUE EXCEL FORMATTING ===

def formatar_valor_excel(valor):
    """
    Formata um valor numérico para o Excel, garantindo que seja um float
    com exatamente duas casas decimais
    
    Args:
        valor: número ou string representando um valor monetário
        
    Returns:
        float: valor formatado como float com 2 casas decimais
    """
    try:
        # Se for string, converter para float
        if isinstance(valor, str):
            # Remover todos os separadores de milhar e trocar vírgula por ponto
            valor_limpo = valor.replace('.', '').replace(',', '.')
            valor_float = float(valor_limpo)
        else:
            valor_float = float(valor)
            
        # Arredondar para duas casas decimais - sem converter para centavos
        valor_formatado = round(valor_float, 2)
        
        return valor_formatado
        
    except (ValueError, TypeError) as e:
        print(f"Erro ao formatar valor '{valor}': {str(e)}")
        return 0.0

def aplicar_formatacao_celula(cell):
    """
    Aplica a formatação correta para células de valor no Excel
    
    Args:
        cell: célula do openpyxl
    """
    cell.number_format = '#,##0.00'
    return cell

def formatar_moeda_br(valor):
    """Formata um valor para o padrão monetário brasileiro (R$ 1.234,56)"""
    try:
        # Converte para float e formata
        valor_float = float(str(valor).replace(',', '.'))
        # Usa locale para formatação brasileira
        import locale
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        return locale.currency(valor_float, symbol=True, grouping=True)
    except (ValueError, TypeError, locale.Error):
        # Método alternativo se locale falhar
        try:
            valor_float = float(str(valor).replace(',', '.'))
            # Formata manualmente
            texto = f"R$ {valor_float:,.2f}"
            # Substitui ponto por X temporário, depois vírgula por ponto, depois X por vírgula
            return texto.replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return f"R$ 0,00"

def formatar_valor_br(self, valor):
    """Formata um valor numérico para o padrão brasileiro (com vírgula)"""
    try:
        if isinstance(valor, str):
            valor = float(valor.replace(',', '.'))
        return f"{valor:.2f}".replace('.', ',')
    except (ValueError, TypeError):
        return valor

# === DADOS BANCARIOS ===

def buscar_dados_bancarios_fornecedor(cnpj_cpf, forma_pagamento="PIX", arquivo_fornecedores=None):
    """
    Busca os dados bancários do fornecedor conforme a forma de pagamento
    
    Args:
        cnpj_cpf (str): CNPJ ou CPF do fornecedor (com ou sem máscara)
        forma_pagamento (str): Forma de pagamento (PIX ou TED)
        arquivo_fornecedores (str, optional): Caminho para o arquivo de fornecedores
            Se não informado, usa o ARQUIVO_FORNECEDORES da configuração
            
    Returns:
        str: Dados bancários formatados para o fornecedor
    """
    try:
        # Se arquivo_fornecedores não foi informado, usar o da configuração
        if not arquivo_fornecedores:
            from src.config.config import ARQUIVO_FORNECEDORES
            arquivo_fornecedores = ARQUIVO_FORNECEDORES
            
        from openpyxl import load_workbook
        
        # Limpar CNPJ/CPF (remover máscara: pontos, traços, barras)
        cnpj_cpf_limpo = ''.join(filter(str.isdigit, str(cnpj_cpf).strip()))
        
        # Abrir arquivo de fornecedores
        wb = load_workbook(arquivo_fornecedores, data_only=True)
        ws = wb['Fornecedores']
    
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Pular se não tem CNPJ/CPF
                continue
                
            # Limpar CNPJ/CPF da planilha também para comparação
            cnpj_cpf_planilha = ''.join(filter(str.isdigit, str(row[0]).strip()))
            
            if cnpj_cpf_planilha == cnpj_cpf_limpo:
                # Encontrou o fornecedor
                dados_bancarios = None
                
                # PRIMEIRO: Verificar se existe dado na coluna O (DADOS BANCÁRIOS) - índice 14
                if len(row) > 14 and row[14] and str(row[14]).strip():
                    dados_bancarios = str(row[14]).strip()
                    wb.close()
                    return dados_bancarios
                
                # SEGUNDO: Se coluna O estiver em branco, continuar com o processo existente
                if forma_pagamento == "PIX" and row[10]:  # Chave PIX está na coluna K
                    dados_bancarios = f"PIX: {row[10]}"
                else:
                    # Construir dados para TED, SEMPRE incluindo CNPJ/CPF
                    partes_dados = []
                    if row[6]: partes_dados.append(str(row[6]))  # Banco
                    if row[7]: partes_dados.append(str(row[7]))  # OP
                    if row[8]: partes_dados.append(str(row[8]))  # Agência
                    if row[9]: partes_dados.append(str(row[9]))  # Conta
                    
                    # SEMPRE incluir CNPJ/CPF para TED, independente da forma de pagamento selecionada
                    if row[0]: partes_dados.append(str(row[0]))
                    
                    dados_bancarios = ' - '.join(filter(None, partes_dados))
                    
                # Se não encontrou dados bancários
                if not dados_bancarios or dados_bancarios.strip() == '-':
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
                    
                wb.close()
                return dados_bancarios
        
        wb.close()
        return 'DADOS BANCÁRIOS NÃO CADASTRADOS'
        
    except Exception as e:
        print(f"Erro ao buscar dados bancários: {str(e)}")
        if 'wb' in locals():
            wb.close()
        return 'ERRO AO BUSCAR DADOS BANCÁRIOS'

def custom_messagebox(tipo, titulo, mensagem):
    """
    Função de messagebox personalizada que padroniza as caixas de diálogo do sistema
    
    Args:
        tipo (str): Tipo da mensagem - "info", "error", "warning", "yesno", "question"
        titulo (str): Título da janela
        mensagem (str): Texto da mensagem
        
    Returns:
        bool: Para tipos "yesno" e "question", retorna True/False
        None: Para outros tipos
    """
    import tkinter.messagebox as msg
    
    try:
        if tipo.lower() == "info":
            msg.showinfo(titulo, mensagem)
            return None
            
        elif tipo.lower() == "error":
            msg.showerror(titulo, mensagem)
            return None
            
        elif tipo.lower() == "warning":
            msg.showwarning(titulo, mensagem)
            return None
            
        elif tipo.lower() == "yesno":
            return msg.askyesno(titulo, mensagem)
            
        elif tipo.lower() == "question":
            return msg.askyesno(titulo, mensagem)
            
        else:
            # Tipo desconhecido, usar info como padrão
            msg.showinfo(titulo, mensagem)
            return None
            
    except Exception as e:
        print(f"Erro no custom_messagebox: {str(e)}")
        # Fallback para messagebox padrão
        msg.showinfo("Erro", f"Erro ao exibir mensagem: {mensagem}")
        return None


def mostrar_alerta(mensagem, titulo="Alerta"):
    """
    Função helper para mostrar alertas de forma simplificada
    
    Args:
        mensagem (str): Mensagem a ser exibida
        titulo (str): Título da janela (padrão: "Alerta")
    """
    return custom_messagebox("warning", titulo, mensagem)


def mostrar_erro(mensagem, titulo="Erro"):
    """
    Função helper para mostrar erros de forma simplificada
    
    Args:
        mensagem (str): Mensagem de erro
        titulo (str): Título da janela (padrão: "Erro")
    """
    return custom_messagebox("error", titulo, mensagem)


def mostrar_info(mensagem, titulo="Informação"):
    """
    Função helper para mostrar informações de forma simplificada
    
    Args:
        mensagem (str): Mensagem informativa
        titulo (str): Título da janela (padrão: "Informação")
    """
    return custom_messagebox("info", titulo, mensagem)


def confirmar_acao(mensagem, titulo="Confirmar"):
    """
    Função helper para confirmações de forma simplificada
    
    Args:
        mensagem (str): Mensagem de confirmação
        titulo (str): Título da janela (padrão: "Confirmar")
        
    Returns:
        bool: True se usuário confirmar, False caso contrário
    """
    return custom_messagebox("yesno", titulo, mensagem)


def validar_entrada_obrigatoria(valor, nome_campo):
    """
    Valida se um campo obrigatório foi preenchido
    
    Args:
        valor (str): Valor do campo
        nome_campo (str): Nome do campo para a mensagem de erro
        
    Returns:
        bool: True se válido, False se inválido (e mostra erro)
    """
    if not valor or not str(valor).strip():
        mostrar_erro(f"O campo '{nome_campo}' é obrigatório!")
        return False
    return True


def validar_valor_numerico(valor, nome_campo, permitir_zero=True):
    """
    Valida se um valor é numérico válido
    
    Args:
        valor (str): Valor a ser validado
        nome_campo (str): Nome do campo para mensagem de erro
        permitir_zero (bool): Se permite valor zero
        
    Returns:
        tuple: (bool, float) - (é_válido, valor_convertido)
    """
    try:
        if isinstance(valor, str):
            valor_limpo = valor.replace(',', '.')
        else:
            valor_limpo = str(valor)
            
        valor_float = float(valor_limpo)
        
        if not permitir_zero and valor_float <= 0:
            mostrar_erro(f"O campo '{nome_campo}' deve ser maior que zero!")
            return False, 0.0
            
        return True, valor_float
        
    except (ValueError, TypeError):
        mostrar_erro(f"O campo '{nome_campo}' deve conter um valor numérico válido!")
        return False, 0.0

def calcular_numero_relatorio(self, data_relatorio):
        """
        Calcula o número do relatório quinzenal para uma data específica.
        Usa self.df_completo (todos os lançamentos do cliente) para encontrar
        a primeira data, garantindo consistência com RelatorioHandler.
        df_completo: DataFrame completo da aba Dados do cliente
        data_relatorio: data do lançamento (datetime, date ou string)
        
        PARA USAR:
        from src.config.utils import calcular_numero_relatorio

        numero = calcular_numero_relatorio(self.df_completo, row['DATA_REL'])

        VERIFICAR:
        trocar self.df_completo por um parâmetro — e o restante da lógica permanece idêntico
        
        """
        try:
            import pandas as pd

            data_ref = pd.to_datetime(data_relatorio).date()

            # Encontrar primeira data em TODOS os lançamentos do cliente
            df_temp = self.df_completo.copy()
            df_temp['DATA_REL'] = pd.to_datetime(df_temp['DATA_REL'], errors='coerce')
            df_temp = df_temp.dropna(subset=['DATA_REL'])

            if df_temp.empty:
                return 1

            if not hasattr(self, '_primeira_data_cache') or self._primeira_data_cache is None:
                self._primeira_data_cache = df_temp['DATA_REL'].min().date()
            primeira_data = self._primeira_data_cache

            # Replicar lógica exata de RelatorioHandler.obter_numero_relatorio
            numero = 1
            data_atual = primeira_data

            while data_atual <= data_ref:
                if data_atual == data_ref:
                    return numero

                if data_atual.day == 5:
                    data_atual = data_atual.replace(day=20)
                else:  # day == 20
                    if data_atual.month == 12:
                        data_atual = data_atual.replace(
                            year=data_atual.year + 1, month=1, day=5
                        )
                    else:
                        data_atual = data_atual.replace(
                            month=data_atual.month + 1, day=5
                        )
                numero += 1

            # Data não cai exatamente em dia 5 ou 20 — retorna o período mais próximo
            return numero

        except Exception as e:
            logger.debug(f"Erro ao calcular número do relatório: {str(e)}")
            return ""

# === CONSTANTS ===
DIAS_QUINZENA = [5, 20]
TIPOS_DESPESA = {
    1: "Despesas com Colaboradores",
    2: "Transferências Programadas",
    3: "Boletos",
    4: "Ressarcimentos",
    5: "Despesas Pagas pelo Cliente",
    6: "Pagamentos Caixa",
    7: "Administração"
}