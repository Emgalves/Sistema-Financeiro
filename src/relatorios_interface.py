import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import importlib
import logging
import pandas as pd
import psutil  # Para verificação de memória
import gc
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from pathlib import Path

from relatorio_despesas_service import RelatoriosDespesasService
from config_relatorio_quinzenal import configurar_relatorio_quinzenal
from src.config.config import PASTA_CLIENTES

# from correcoes_emergenciais import aplicar_todas_correcoes 
# aplicar_todas_correcoes()

# Adicionar diretório raiz ao path ANTES de qualquer importação
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# SISTEMA DE LOGGING ROBUSTO usando o sistema existente
def setup_logging_safe():
    """
    Configura logging usando o sistema existente com fallback seguro
    """
    try:
        # Tentar usar o sistema de logging existente
        from src.config.logger_config import system_logger, log_action
        
        # Configurar usuário padrão se não estiver definido
        system_logger.set_user('sistema_relatorios')
        
        # Obter logger
        logger = system_logger.get_logger()
        logger.info("Sistema de relatórios inicializando usando logger configurado")
        
        return logger, log_action
        
    except ImportError as e:
        print(f"Aviso: Não foi possível importar sistema de logging configurado: {str(e)}")
        
        # Fallback: criar sistema de logging simples
        import logging
        
        # Configurar logger básico
        logger = logging.getLogger("sistema_relatorios")
        logger.setLevel(logging.INFO)
        
        # Evitar handlers duplicados
        if not logger.handlers:
            # Handler para console
            console_handler = logging.StreamHandler()
            console_handler.setFormatter(
                logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            )
            logger.addHandler(console_handler)
            
            # Tentar handler para arquivo
            try:
                # Determinar diretório base
                if getattr(sys, 'frozen', False):
                    base_dir = os.path.dirname(sys.executable)
                else:
                    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                
                logs_dir = os.path.join(base_dir, 'logs')
                os.makedirs(logs_dir, exist_ok=True)
                
                log_file = os.path.join(logs_dir, f"sistema_relatorios_{datetime.now().strftime('%Y%m%d')}.log")
                file_handler = logging.FileHandler(log_file, encoding='utf-8')
                file_handler.setFormatter(
                    logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
                )
                logger.addHandler(file_handler)
                logger.info(f"Log de fallback criado: {log_file}")
                
            except Exception as file_error:
                logger.warning(f"Não foi possível criar log em arquivo: {str(file_error)}")
        
        # Criar decorator simples para compatibilidade
        def log_action_fallback(description):
            def decorator(func):
                def wrapper(*args, **kwargs):
                    logger.info(f"Executando: {description}")
                    try:
                        result = func(*args, **kwargs)
                        logger.info(f"Concluído: {description}")
                        return result
                    except Exception as e:
                        logger.error(f"Erro em {description}: {str(e)}")
                        raise
                return wrapper
            return decorator
        
        logger.info("Sistema de logging fallback configurado")
        return logger, log_action_fallback
    
    except Exception as e:
        print(f"Erro crítico ao configurar logging: {str(e)}")
        
        # Último recurso: logging mínimo para console
        import logging
        logger = logging.getLogger("sistema_relatorios")
        logger.setLevel(logging.INFO)
        
        if not logger.handlers:
            handler = logging.StreamHandler()
            handler.setFormatter(logging.Formatter('%(levelname)s - %(message)s'))
            logger.addHandler(handler)
        
        def no_op_decorator(description):
            def decorator(func):
                return func
            return decorator
        
        logger.warning("Usando sistema de logging mínimo")
        return logger, no_op_decorator

# Configurar logging
logger, log_action = setup_logging_safe()

# Importar configurações (com fallback)
try:
    from src.config.window_config import configurar_janela
    logger.info("Configurações de janela importadas com sucesso")
except ImportError:
    logger.warning("Usando configuração de janela fallback")
    # Implementação básica caso o módulo não seja encontrado

    def configurar_janela(janela, titulo, largura=700, altura=1000):
        """
        Configura o posicionamento e dimensionamento padrão de uma janela
        
        Args:
            janela: Instância de tk.Tk ou tk.Toplevel
            titulo: Título da janela
            largura: Largura desejada (default 900)
            altura: Altura desejada (default 1000)
        """
        janela.title(titulo)
        
        # Obter dimensões da tela
        screen_width = janela.winfo_screenwidth()
        screen_height = janela.winfo_screenheight()
        
        # Ajustar dimensões para não exceder o tamanho da tela
        largura = min(largura, screen_width)
        altura = min(altura, screen_height)
        
        # Definir posição (sempre no topo esquerdo)
        x = 0
        y = 0
        
        # Configurar geometria
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        
        # Permitir redimensionamento
        janela.resizable(True, True)
        
        # Configurar peso das linhas/colunas para redimensionamento proporcional
        janela.grid_rowconfigure(0, weight=1)
        janela.grid_columnconfigure(0, weight=1)
        
        # Trazer janela para frente
        janela.lift()
        janela.focus_force()

# Log de inicialização
logger.info("=== Sistema de Relatórios Inicializando ===")

class SistemaRelatorios:
    """Interface centralizada para todos os relatórios do sistema"""
    
    def __init__(self, parent=None):
        """Inicializa a interface do sistema de relatórios"""
        self.parent = parent
        
        # Configurar janela principal
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        # Configurar a janela
        configurar_janela(self.root, "Sistema Integrado de Relatórios", 900, 1000)
        
        # Acompanhar quais módulos foram carregados
        self.modulos_carregados = {}
        
        # Inicializar os atributos para os comboboxes
        self.cliente_combobox = None
        self.cliente_contratos = None

        # Inicializar variáveis para controle do relatório de despesas
        self.arquivo_cliente_selecionado = None
        self.arquivos_lote = []
        self.pasta_lancamentos = None
        
        self.despesas_service = RelatoriosDespesasService()

        # Configurar interface
        self.setup_ui()

    def setup_ui(self):
        """Configura a interface gráfica do sistema"""
        # Frame principal dividido em esquerda e direita
        self.main_frame = ttk.Frame(self.root, padding=10)
        self.main_frame.pack(fill='both', expand=True)
        
        # Frame esquerdo para lista de relatórios
        self.left_frame = ttk.LabelFrame(self.main_frame, text="Tipos de Relatórios")
        self.left_frame.pack(side='left', fill='both', padx=10, pady=10)
        
        # Frame direito para opções do relatório selecionado
        self.right_frame = ttk.LabelFrame(self.main_frame, text="Configurações do Relatório")
        self.right_frame.pack(side='right', fill='both', expand=True, padx=10, pady=10)
        
        # Lista de relatórios disponíveis
        self.setup_relatorios_list()
        
        # Frame inferior para botões de ação
        self.bottom_frame = ttk.Frame(self.root, padding=10)
        self.bottom_frame.pack(side='bottom', fill='x')
        
        # Botão para voltar ao menu principal
        ttk.Button(
            self.bottom_frame, 
            text="Voltar ao Menu Principal", 
            command=self.voltar_menu
        ).pack(side='right', padx=5)

        # Carregar lista de clientes
        self.atualizar_lista_clientes()
        
        # Configurar período inicial
        # self.alterar_periodo()

        # Configurar validações
        self.backup_metodo_original()
        
        # Forçar atualização da interface para garantir que todos os widgets estejam prontos
        self.root.update_idletasks()

    def gerar_relatorio(self, relatorio):
        """VERSÃO LIMPA - Remove toda a complexidade anterior"""
        try:
            logger.info(f"🔍 INICIANDO gerar_relatorio para: {relatorio['id']}")
            
            # Verificar disponibilidade
            if not relatorio["disponivel"]:
                messagebox.showinfo("Em desenvolvimento", "Este relatório ainda está em desenvolvimento.")
                return
            
            # === TRATAR DESPESAS COM NOVA ARQUITETURA ===
            if relatorio["id"] == "despesas":
                logger.info("🎯 PROCESSANDO: Relatório de despesas")
                self._processar_despesas_limpo()
                return  # Para aqui para despesas
            
            # === OUTROS RELATÓRIOS (mantém como estava) ===
            logger.info(f"📋 Processando outros relatórios: {relatorio['id']}")
            
            if relatorio["id"] == "lancamentos_pendentes":
                self.processar_lancamentos_pendentes()
            elif relatorio["id"] == "fornecedores":
                self.processar_fornecedores()
            
            elif relatorio["id"] == "gerencial_engenheiro":
                self.processar_gerencial_engenheiro()

            elif relatorio["id"] == "gerencial_pdf":
                self.processar_gerencial_pdf()

            elif relatorio["id"] == "consistencia_dados":
                self.processar_consistencia_dados()

            else:
                self.processar_outros_relatorios(relatorio)
                
        except Exception as e:
            logger.error(f"💥 ERRO em gerar_relatorio: {str(e)}", exc_info=True)
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {str(e)}")

    def _processar_despesas_limpo(self):
        """Processamento de despesas LIMPO - apenas orquestração"""
        try:
            logger.info("🎯 PROCESSANDO DESPESAS - ARQUITETURA LIMPA")
            
            # 1. Validar configurações (responsabilidade da UI)
            if not self.validar_configuracoes_despesas():
                logger.warning("❌ Validação de configurações falhou")
                return
            
            # 2. Coletar configurações (responsabilidade da UI)
            configuracoes = self.coletar_configuracoes_completas()
            logger.info(f"✅ Configurações coletadas")
            
            # 3. Verificar arquivo
            if not configuracoes.get('arquivo'):
                logger.error("❌ ERRO: Arquivo não encontrado nas configurações!")
                messagebox.showerror(
                    "Erro", 
                    "Arquivo não encontrado. Verifique se um cliente foi selecionado."
                )
                return
            
            logger.info(f"✅ Arquivo confirmado: {os.path.basename(configuracoes['arquivo'])}")
            
            # 4. Confirmar geração (responsabilidade da UI)
            if not self.confirmar_geracao_relatorio():
                logger.info("❌ Geração cancelada pelo usuário")
                return
            
            # 5. Verificar modo selecionado
            usar_preview = hasattr(self, 'modo_visualizacao') and self.modo_visualizacao.get() == "preview"
            logger.info(f"Modo selecionado: {'Preview' if usar_preview else 'Direto'}")
            
            # 6. Executar conforme modo - DELEGAÇÃO PARA SERVIÇO
            if usar_preview:
                self._executar_com_preview_limpo(configuracoes)
            else:
                self._executar_direto_limpo(configuracoes)
                
        except Exception as e:
            logger.error(f"💥 ERRO no processamento: {str(e)}", exc_info=True)
            messagebox.showerror("Erro", f"Erro no processamento: {str(e)}")

    def _executar_com_preview_limpo(self, configuracoes):
        """Execução com preview DIRETO - pula interface e abre PDF temporário"""
        try:
            logger.info("🎯 EXECUTANDO COM PREVIEW DIRETO - PDF TEMPORÁRIO")
            
            # 1. Mostrar progresso simples
            progress_label = tk.Label(
                self.root, 
                text="Gerando PDF temporário para análise...", 
                font=('Arial', 12), 
                bg='lightblue', 
                relief='raised', 
                padx=20, 
                pady=10
            )
            progress_label.place(relx=0.5, rely=0.5, anchor='center')
            self.root.update()
            
            try:
                # 2. PROCESSAR dados através do serviço
                dados_processados = self.despesas_service.processar_para_preview(configuracoes)
                
                # 3. GERAR PDF temporário através do serviço
                pdf_temp_path = self.despesas_service.gerar_pdf_temporario(
                    dados_processados, 
                    configuracoes['arquivo']
                )
                
                # 4. Remover progresso
                progress_label.destroy()
                
                # 5. ABRIR PDF temporário automaticamente
                self.abrir_arquivo(pdf_temp_path)
                
                # 6. MOSTRAR janela de decisão
                self._mostrar_janela_decisao_pdf(dados_processados, configuracoes, pdf_temp_path)
                
            except Exception as e:
                try:
                    progress_label.destroy()
                except:
                    pass
                raise e
                    
        except Exception as e:
            logger.error(f"💥 ERRO no preview direto: {str(e)}")
            messagebox.showerror("Erro", f"Erro: {str(e)}")

    def _mostrar_janela_decisao_pdf(self, dados_processados, configuracoes, pdf_temp_path):
        """Janela simples de decisão após abrir PDF temporário - VERSÃO OTIMIZADA"""
        try:
            # Criar janela de decisão
            decisao_window = tk.Toplevel(self.root)
            decisao_window.title("🔍 Análise do Relatório - Escolha uma Ação")
            decisao_window.geometry("550x350")
            decisao_window.transient(self.root)
            decisao_window.grab_set()
            
            # Posicionamento estratégico no canto superior direito
            decisao_window.update_idletasks()
            screen_width = decisao_window.winfo_screenwidth()
            x = screen_width - 570
            y = 50
            decisao_window.geometry(f"550x350+{x}+{y}")
            
            # Garantir visibilidade
            decisao_window.attributes('-topmost', True)
            decisao_window.lift()
            decisao_window.focus_force()
            
            # Reforçar após delay
            def reativar_janela():
                try:
                    decisao_window.lift()
                    decisao_window.focus_force()
                    decisao_window.bell()  # Som de notificação
                    # Remover topmost após 3s para não incomodar
                    decisao_window.after(3000, lambda: decisao_window.attributes('-topmost', False))
                except:
                    pass
            
            decisao_window.after(1200, reativar_janela)
            
            # Frame principal
            main_frame = ttk.Frame(decisao_window, padding=20)
            main_frame.pack(fill='both', expand=True)
            
            # Título
            ttk.Label(
                main_frame, 
                text="📄 PDF Temporário Aberto para Análise", 
                font=('Arial', 14, 'bold'),
                foreground='darkgreen'
            ).pack(pady=(0, 20))
            
            # Informações
            info_text = f"""
    Cliente: {dados_processados.get('nome_cliente', 'N/A')}
    Data: {dados_processados.get('data_relatorio', 'N/A')}
    Relatório nº: {dados_processados.get('numero_relatorio', 'N/A')}

    O PDF temporário foi aberto para sua análise.
    Após revisar o conteúdo, escolha uma das opções abaixo:
            """
            
            ttk.Label(main_frame, text=info_text, font=('Arial', 10)).pack(pady=(0, 30))
            
            # Frame para botões
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(fill='x', pady=20)
            
            # Função para gerar PDF definitivo
            def gerar_definitivo():
                try:
                    decisao_window.destroy()
                    
                    # Mostrar progresso
                    progress_window = self.criar_progress_window()
                    self.atualizar_progresso_seguro(progress_window, "Gerando PDF definitivo...", 50)
                    
                    # Gerar PDF definitivo
                    caminho_final, nome_arquivo = self.despesas_service.gerar_pdf_definitivo(
                        dados_processados, 
                        configuracoes['arquivo']
                    )
                    
                    progress_window.destroy()
                    
                    # Limpar PDF temporário
                    self._limpar_pdf_temporario(pdf_temp_path)
                    
                    # Mostrar resultado
                    resposta = messagebox.askyesno(
                        "PDF Definitivo Gerado!",
                        f"✅ PDF definitivo gerado com sucesso!\n\n"
                        f"Cliente: {dados_processados['nome_cliente']}\n"
                        f"Arquivo: {nome_arquivo}\n\n"
                        f"Deseja abrir o PDF definitivo?"
                    )
                    
                    if resposta:
                        self.abrir_arquivo(caminho_final)
                    
                except Exception as e:
                    logger.error(f"Erro ao gerar PDF definitivo: {str(e)}")
                    messagebox.showerror("Erro", f"Erro: {str(e)}")
            
            # Função para voltar ao menu
            def voltar_menu():
                try:
                    # Limpar PDF temporário
                    self._limpar_pdf_temporario(pdf_temp_path)
                    
                    decisao_window.destroy()
                    # Manter na interface atual para novo relatório
                    
                except Exception as e:
                    logger.error(f"Erro ao voltar: {str(e)}")
            
            # Botões
            ttk.Button(
                btn_frame,
                text="🚀 Gerar PDF Definitivo",
                command=gerar_definitivo,
                style='Accentuated.TButton'
            ).pack(side='left', padx=(0, 20), fill='x', expand=True)
            
            ttk.Button(
                btn_frame,
                text="⬅️ Voltar ao Menu Anterior",
                command=voltar_menu
            ).pack(side='right', fill='x', expand=True)
            
            # Configurar fechamento
            decisao_window.protocol("WM_DELETE_WINDOW", voltar_menu)
            
            logger.info("✅ Janela de decisão criada")
            
        except Exception as e:
            logger.error(f"💥 ERRO na janela de decisão: {str(e)}")
            self._limpar_pdf_temporario(pdf_temp_path)
            messagebox.showerror("Erro", f"Erro: {str(e)}")

    def _limpar_pdf_temporario(self, pdf_path):
        """Limpa PDF temporário de forma segura"""
        try:
            if pdf_path and os.path.exists(pdf_path):
                # Aguardar um pouco para garantir que o PDF não está sendo usado
                self.root.after(2000, lambda: self.limpar_arquivo_temporario(pdf_path))
                logger.info(f"🗑️ PDF temporário agendado para remoção: {os.path.basename(pdf_path)}")
        except Exception as e:
            logger.warning(f"Aviso ao limpar PDF temporário: {str(e)}")

    def _executar_direto_limpo(self, configuracoes):
        """Execução direta - DELEGAÇÃO para serviço"""
        try:
            logger.info("🎯 EXECUTANDO DIRETO - DELEGAÇÃO PARA SERVIÇO")
            
            # 1. Mostrar progresso
            progress_window = self.criar_progress_window()
            self.atualizar_progresso_seguro(progress_window, "Processando através do serviço...", 10)
            
            # 2. DELEGAR processamento para serviço
            dados_processados = self.despesas_service.processar_para_preview(configuracoes)
            self.atualizar_progresso_seguro(progress_window, "Gerando PDF definitivo...", 70)
            
            # 3. DELEGAR geração de PDF para serviço
            caminho_final, nome_arquivo = self.despesas_service.gerar_pdf_definitivo(
                dados_processados, 
                configuracoes['arquivo']
            )
            self.atualizar_progresso_seguro(progress_window, "PDF gerado com sucesso!", 100)
            
            # 4. Fechar progresso
            progress_window.destroy()
            
            # 5. Mostrar resultado
            self._mostrar_resultado_geracao_limpo(
                dados_processados['nome_cliente'],
                nome_arquivo, 
                caminho_final
            )
            
            logger.info("✅ Execução direta concluída")
            
        except Exception as e:
            try:
                progress_window.destroy()
            except:
                pass
            logger.error(f"💥 ERRO no executar_direto_limpo: {str(e)}")
            messagebox.showerror("Erro", f"Erro: {str(e)}")


    def _mostrar_resultado_geracao_limpo(self, nome_cliente, nome_arquivo, caminho_final):
        """Mostra resultado da geração de forma limpa"""
        try:
            resposta = messagebox.askyesnocancel(
                "Relatório Gerado!",
                f"✅ Relatório gerado com sucesso!\n\n"
                f"Cliente: {nome_cliente}\n"
                f"Arquivo: {nome_arquivo}\n\n"
                f"🔄 Opções:\n"
                f"• Sim: Abrir PDF\n"
                f"• Não: Continuar sem abrir\n"
                f"• Cancelar: Gerar outro relatório",
                icon='question'
            )
            
            if resposta is True:  # Abrir PDF
                self.abrir_arquivo(caminho_final)
            elif resposta is False:  # Não abrir
                pass  # Continua na interface
            # resposta is None = Cancelar = continua na interface
            
            logger.info(f"✅ Resultado mostrado para: {nome_cliente}")
            
        except Exception as e:
            logger.error(f"Erro ao mostrar resultado: {str(e)}")


    def setup_relatorios_list(self):
        """Configura a lista de relatórios disponíveis"""
        # Definir os relatórios disponíveis
        self.relatorios = [
            {
                "id": "despesas",
                "nome": "Relatório de Despesas",
                "descricao": "Relatório financeiro de despesas por cliente",
                "modulo": "relatorio_despesas_aprimorado",
                "classe": "RelatorioHandler",
                "disponivel": True
            },
            {
                "id": "contratos",
                "nome": "Relatório de Contratos e Medições",
                "descricao": "Relatório de contratos por medição e status",
                "modulo": "relatorio_contratos_medicoes",
                "classe": "RelatorioContratos",
                "disponivel": True
            },
            {
                "id": "medicoes_quinzenal",
                "nome": "Relatório Quinzenal de Medições (PDF)",
                "descricao": "Relatório PDF de medições da quinzena (dias 5 e 20)",
                "modulo": None,
                "classe": None,
                "disponivel": True
            },
            {
                "id": "categoria",
                "nome": "Relatório por Categoria",
                "descricao": "Análise de despesas agrupadas por categoria",
                "modulo": "relatorio_categoria",
                "classe": "RelatorioCategoria",
                "disponivel": True
            },
            {
                "id": "tipo_despesa",
                "nome": "Relatório por Tipo de Despesa",
                "descricao": "Análise detalhada por tipo de despesa",
                "modulo": "relatorio_tipo_despesa",
                "classe": "RelatorioTipoDespesa",
                "disponivel": True
            },
            {
                "id": "fornecedores",
                "nome": "Relatório de Principais Fornecedores",
                "descricao": "Resumo de fornecedores por cliente e global",
                "modulo": "relatorio_fornecedores",
                "classe": "RelatorioFornecedores",
                "disponivel": True
            },
            {
                "id": "gerencial_engenheiro",
                "nome": "Relatório Gerencial Medições",
                "descricao": "Visão consolidada de todas as obras por grupo/engenheiro",
                "modulo": "relatorio_gerencial_engenheiro",
                "classe": "RelatorioGerencialEngenheiro",
                "disponivel": True
            },
            {
                "id": "gerencial_pdf",
                "nome": "Relatório Gerencial Medições (PDF)",
                "descricao": "Visão consolidada em PDF - Layout profissional para apresentações",
                "modulo": "relatorio_gerencial_pdf",
                "classe": "RelatorioGerencialPDF",
                "disponivel": True
            },
            {
                "id": "administracao",
                "nome": "Relatório de Contratos de Administração",
                "descricao": "Relatório de contratos de administração de obra",
                "modulo": "relatorio_administracao",
                "classe": "RelatorioAdministracao",
                "disponivel": False
            },
            {
                "id": "lancamentos_pendentes",
                "nome": "Relatório de Lançamentos Pendentes",
                "descricao": "Relatório de lançamentos pendentes de múltiplos clientes",
                "modulo": "relatorio_despesas_aprimorado",
                "classe": "RelatorioLancamentosPendentes",
                "disponivel": True
            },
            {
                "id": "consistencia_dados",
                "nome": "Verificação de Consistência de Dados",
                "descricao": "Verifica registros em 'Dados' sem correspondência em Medições/Contratos ADM e vice-versa",
                "modulo": "relatorio_consistencia_dados",
                "classe": "RelatorioConsistenciaDados",
                "disponivel": True
            }
        ]
        
        # Criar o Treeview para a lista de relatórios
        columns = ('nome', 'status')
        self.tree_relatorios = ttk.Treeview(self.left_frame, columns=columns, show='headings', height=15)
        
        # Configurar cabeçalhos
        self.tree_relatorios.heading('nome', text='Relatório')
        self.tree_relatorios.heading('status', text='Status')
        
        # Configurar colunas
        self.tree_relatorios.column('nome', width=200)
        self.tree_relatorios.column('status', width=100, anchor='center')
        
        # Preencher a treeview
        for relatorio in self.relatorios:
            status = "Disponível" if relatorio["disponivel"] else "Em Desenvolvimento"
            self.tree_relatorios.insert('', 'end', iid=relatorio["id"], values=(relatorio["nome"], status))
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.left_frame, orient="vertical", command=self.tree_relatorios.yview)
        self.tree_relatorios.configure(yscrollcommand=scrollbar.set)
        
        # Colocar widgets na tela
        self.tree_relatorios.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Bind para seleção
        self.tree_relatorios.bind('<<TreeviewSelect>>', self.mostrar_opcoes_relatorio)
    
    def mostrar_configuracoes_quinzenal(self):
        """Mostra configurações do relatório quinzenal no painel direito"""
        try:
            logger.info("📄 Carregando configurações do Relatório Quinzenal")
            
            # Limpar painel direito
            for widget in self.right_frame.winfo_children():
                widget.destroy()
            
            # Criar frame de conteúdo com scroll
            canvas = tk.Canvas(self.right_frame)
            scrollbar = ttk.Scrollbar(self.right_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Pack canvas e scrollbar
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Carregar configurações do relatório quinzenal
            from config_relatorio_quinzenal import configurar_relatorio_quinzenal
            configurar_relatorio_quinzenal(scrollable_frame, self)
            
            logger.info("✅ Configurações do Relatório Quinzenal carregadas")
            
        except Exception as e:
            logger.error(f"Erro ao carregar configurações quinzenal: {str(e)}", exc_info=True)
            messagebox.showerror("Erro", f"Erro ao carregar configurações: {str(e)}")

    def mostrar_opcoes_relatorio(self, event=None):
        """Versão corrigida que usa estrutura original"""
        
        # Limpar frame direito
        for widget in self.right_frame.winfo_children():
            widget.destroy()
        
        # Obter relatório selecionado
        selecao = self.tree_relatorios.selection()
        if not selecao:
            return
            
        rel_id = selecao[0]
        relatorio = next((r for r in self.relatorios if r["id"] == rel_id), None)
        
        if not relatorio:
            return
        
        # Mostrar informações do relatório
        ttk.Label(
            self.right_frame, 
            text=relatorio["nome"], 
            font=('Arial', 14, 'bold')
        ).pack(pady=(10,5), anchor='w')
        
        ttk.Label(
            self.right_frame, 
            text=relatorio["descricao"],
            wraplength=400
        ).pack(pady=(0,20), anchor='w')
        
        # Se o relatório não estiver disponível
        if not relatorio["disponivel"]:
            ttk.Label(
                self.right_frame,
                text="Este relatório está em desenvolvimento e ainda não está disponível.",
                foreground='red'
            ).pack(pady=20)
            return
        
        # Adicionar separador visual
        ttk.Separator(self.right_frame, orient='horizontal').pack(fill='x', pady=(0, 15))
        
        # Configurar opções específicas diretamente no right_frame (sem frame duplicado)
        if relatorio["id"] == "despesas":
            self.setup_opcoes_despesas(self.right_frame)
        elif relatorio["id"] == "contratos":
            self.setup_opcoes_contratos(self.right_frame)
        elif relatorio["id"] == "categoria":
            self.setup_opcoes_categoria(self.right_frame)
        elif relatorio["id"] == "tipo_despesa":
            self.setup_opcoes_tipo_despesa(self.right_frame)
        elif relatorio["id"] == "fornecedores":
            self.setup_opcoes_fornecedores(self.right_frame)
        elif relatorio["id"] == "gerencial_engenheiro":
            self.setup_opcoes_gerencial_engenheiro(self.right_frame)
        elif relatorio["id"] == "gerencial_pdf":
            self.setup_opcoes_gerencial_pdf(self.right_frame)
        elif relatorio["id"] == "lancamentos_pendentes":
            self.setup_opcoes_lancamentos_pendentes(self.right_frame)
        elif relatorio["id"] == "medicoes_quinzenal":
            self.setup_opcoes_quinzenal(self.right_frame)
        elif relatorio["id"] == "consistencia_dados":
            self.setup_opcoes_consistencia_dados(self.right_frame)
        else:
            ttk.Label(
                self.right_frame,
                text="Opções específicas para este relatório serão implementadas em breve."
            ).pack(pady=20)
        
        # === ÚNICA MUDANÇA: BOTÃO PERSONALIZADO APENAS PARA DESPESAS ===
        btn_frame = ttk.Frame(self.right_frame)
        btn_frame.pack(fill='x', pady=20)
        
        if relatorio["id"] == "despesas":
                       
            # Botão de validação (opcional)
            ttk.Button(
                btn_frame,
                text="✅ Validar Configurações",
                command=lambda: self.validar_e_mostrar_resumo(),
                style='TButton'
            ).pack(side='left', padx=5)
            
            # Botão principal otimizado
            ttk.Button(
                btn_frame,
                text="🚀 Processar e Gerar Relatório",
                command=lambda: self.gerar_relatorio(relatorio),
                style='Accentuated.TButton'
            ).pack(side='right', padx=5)
            
        else:
            # ORIGINAL: Botão padrão para outros relatórios
            ttk.Button(
                btn_frame,
                text="Gerar Relatório",
                command=lambda: self.gerar_relatorio(relatorio),
                style='Accentuated.TButton'
            ).pack(side='right', padx=5)

    def criar_botao_despesas_otimizado(self, btn_frame, relatorio):
        """Cria botão otimizado específico para relatório de despesas"""
        
        # Label explicativo
        info_label = ttk.Label(
            btn_frame,
            text="💡 O sistema processará os dados e abrirá diretamente o preview ou gerará o PDF conforme configurado.",
            font=('Arial', 9),
            foreground='blue',
            wraplength=400
        )
        info_label.pack(pady=(0, 10))
        
        # Botão principal otimizado
        botao_principal = ttk.Button(
            btn_frame,
            text="🚀 Processar e Gerar Relatório",
            command=lambda: self.gerar_relatorio(relatorio),
            style='Accentuated.TButton'
        )
        botao_principal.pack(side='right', padx=5)
        
        # OPCIONAL: Botão de validação prévia
        botao_validar = ttk.Button(
            btn_frame,
            text="✅ Validar Configurações",
            command=lambda: self.validar_e_mostrar_resumo(),
            style='TButton'
        )
        botao_validar.pack(side='left', padx=5)

    def criar_botao_padrao(self, btn_frame, relatorio):
        """Cria botão padrão para outros tipos de relatório"""
        
        # Botão padrão (comportamento original)
        ttk.Button(
            btn_frame,
            text="Gerar Relatório",
            command=lambda: self.gerar_relatorio(relatorio),
            style='Accentuated.TButton'
        ).pack(side='right', padx=5)

    def validar_e_mostrar_resumo(self):
        """Valida configurações e mostra resumo antes da geração"""
        try:
            # Validar configurações
            if not self.validar_configuracoes_despesas():
                return
            
            # Coletar configurações
            configuracoes = self.coletar_configuracoes_completas()
            
            # Gerar resumo
            resumo = self.gerar_resumo_configuracoes(configuracoes)
            
            # Mostrar resumo em janela separada
            self.mostrar_janela_resumo(resumo, configuracoes)
            
        except Exception as e:
            logger.error(f"Erro na validação prévia: {str(e)}")
            messagebox.showerror("Erro", f"Erro na validação: {str(e)}")

    def mostrar_janela_resumo(self, resumo, configuracoes):
        """Mostra janela com resumo das configurações"""
        
        # Criar janela
        resumo_window = tk.Toplevel(self.root)
        resumo_window.title("Resumo das Configurações")
        resumo_window.geometry("500x400")
        resumo_window.transient(self.root)
        resumo_window.grab_set()
        
        # Frame principal
        main_frame = ttk.Frame(resumo_window, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            main_frame,
            text="📋 Resumo das Configurações",
            font=('Arial', 14, 'bold')
        ).pack(pady=(0, 20))
        
        # Área de texto com scroll
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill='both', expand=True)
        
        text_widget = tk.Text(
            text_frame,
            wrap='word',
            font=('Courier', 10),
            state='disabled'
        )
        
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Inserir resumo
        text_widget.config(state='normal')
        text_widget.insert('1.0', resumo)
        text_widget.config(state='disabled')
        
        # Frame para botões
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=(20, 0))
        
        # Botões
        ttk.Button(
            btn_frame,
            text="❌ Cancelar",
            command=resumo_window.destroy
        ).pack(side='left', padx=5)
        
        ttk.Button(
            btn_frame,
            text="✏️ Editar Configurações",
            command=resumo_window.destroy
        ).pack(side='left', padx=5)
        
        ttk.Button(
            btn_frame,
            text="🚀 Continuar com Geração",
            command=lambda: self.continuar_geracao_apos_resumo(resumo_window, configuracoes)
        ).pack(side='right', padx=5)

    def continuar_geracao_apos_resumo(self, resumo_window, configuracoes):
        """Continua com a geração após confirmação do resumo"""
        try:
            resumo_window.destroy()
            
            # Criar o relatório mock para compatibilidade
            relatorio_mock = {
                "id": "despesas",
                "nome": "Relatório de Despesas",
                "disponivel": True
            }
            
            # Proceder com geração otimizada
            self.gerar_relatorio(relatorio_mock)
            
        except Exception as e:
            logger.error(f"Erro ao continuar geração: {str(e)}")
            messagebox.showerror("Erro", f"Erro: {str(e)}")

    def carregar_clientes(self):
        """Carrega a lista de clientes ativos do arquivo de clientes"""
        try:
            # Importar bibliotecas necessárias
            import pandas as pd
            from openpyxl import load_workbook
            
            # Caminho para o arquivo de clientes
            try:
                from src.config.config import ARQUIVO_CLIENTES
                logger.info(f"Carregando clientes de: {ARQUIVO_CLIENTES}")
            except ImportError:
                # Caminho padrão se não conseguir importar das configurações
                ARQUIVO_CLIENTES = "dados/clientes.xlsx"
                logger.warning(f"Usando caminho padrão para clientes: {ARQUIVO_CLIENTES}")
            
            # Verificar se o arquivo existe
            if not os.path.exists(ARQUIVO_CLIENTES):
                logger.warning(f"Arquivo de clientes não encontrado: {ARQUIVO_CLIENTES}")
                return ['']
            
            # Carregar o arquivo usando pandas
            try:
                # Ler o arquivo Excel
                df = pd.read_excel(ARQUIVO_CLIENTES, sheet_name='Clientes')
                
                # Debug: mostrar as colunas disponíveis
                logger.info(f"Colunas disponíveis: {df.columns.tolist()}")
                
                # Verificar se a coluna E existe (coluna 4 em índice baseado em 0)
                # Ou verificar pelo nome da coluna se existir
                if len(df.columns) >= 5:  # Verifica se tem pelo menos 5 colunas (A-E)
                    # Filtrar clientes ativos (coluna E vazia)
                    coluna_status = df.columns[4]  # Coluna E (índice 4)
                    logger.info(f"Coluna de status: {coluna_status}")
                    
                    # Considera como vazio: None, NaN, '', etc.
                    df_ativos = df[df[coluna_status].isna() | (df[coluna_status] == '')]
                    
                    # Verificar se a primeira coluna contém os nomes dos clientes
                    coluna_nome = df.columns[0]  # Coluna A
                    logger.info(f"Coluna de nome: {coluna_nome}")
                    
                    # Extrair nomes dos clientes ativos (assumindo que estão na primeira coluna)
                    clientes_ativos = df_ativos[coluna_nome].dropna().tolist()
                    
                    logger.info(f"Total de clientes ativos encontrados: {len(clientes_ativos)}")
                    
                    # Ordenar alfabeticamente
                    clientes_ativos.sort()
                    
                    # Adicionar "Todos os Clientes" no início
                    clientes = ['Todos os Clientes'] + clientes_ativos
                    
                    return clientes
                else:
                    logger.warning("Arquivo não tem colunas suficientes (precisa de pelo menos 5 colunas - A até E)")
                    return ['Todos os Clientes']
                
            except Exception as e:
                logger.error(f"Erro ao ler arquivo Excel com pandas: {str(e)}")
                # Tentar com openpyxl como fallback
                try:
                    workbook = load_workbook(ARQUIVO_CLIENTES)
                    sheet = workbook['Clientes']
                    
                    clientes = ['Todos os Clientes']
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Verifica se a coluna E (índice 4) está vazia
                        if row[0] and (len(row) < 5 or not row[4]):
                            clientes.append(row[0])
                    
                    workbook.close()
                    clientes.sort()  # Ordenar alfabeticamente (mantendo "Todos os Clientes" primeiro)
                    return clientes
                    
                except Exception as inner_e:
                    logger.error(f"Erro ao ler arquivo Excel com openpyxl: {str(inner_e)}")
                    return ['Todos os Clientes']
                
        except Exception as e:
            logger.error(f"Erro ao carregar clientes: {str(e)}", exc_info=True)
            return ['Todos os Clientes']

    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes na combobox"""
        try:
            clientes = self.carregar_clientes()
            
            # Atualizar todos os comboboxes que mostram clientes
            if hasattr(self, 'cliente_combobox') and self.cliente_combobox is not None:
                self.cliente_combobox['values'] = clientes
                self.cliente_combobox.current(0)  # Selecionar "Todos os Clientes"
            
            if hasattr(self, 'cliente_contratos') and self.cliente_contratos is not None:
                self.cliente_contratos['values'] = clientes
                self.cliente_contratos.current(0)
                
            logger.info(f"Lista de clientes atualizada com {len(clientes)} clientes")
            
        except Exception as e:
            logger.error(f"Erro ao atualizar lista de clientes: {str(e)}")
    
    def preencher_combobox_clientes(self, combobox):
        """Preenche um combobox com a lista de clientes ativos"""
        try:
            if hasattr(self, 'lista_clientes') and self.lista_clientes:
                clientes = self.lista_clientes
            else:
                clientes = self.carregar_clientes()
                self.lista_clientes = clientes  # Cache da lista
            
            combobox['values'] = clientes
            combobox.current(0)  # Selecionar "Todos os Clientes"
            
        except Exception as e:
            logger.error(f"Erro ao preencher combobox de clientes: {str(e)}")
            combobox['values'] = ['Todos os Clientes']
            combobox.current(0)

    def calcular_data_rel_automatica(self):
        """Calcula automaticamente a data do relatório baseado na regra dos dias 5 e 20"""
        try:
            hoje = datetime.now()
            
            if 6 <= hoje.day <= 20:
                # Entre dia 6 e 20: relatório do dia 20 do mês atual
                data_rel = hoje.replace(day=20)
            else:
                if hoje.day > 20:
                    # Após dia 20: relatório do dia 5 do próximo mês
                    data_rel = (hoje + relativedelta(months=1)).replace(day=5)
                else:
                    # Antes do dia 6: relatório do dia 5 do mês atual
                    data_rel = hoje.replace(day=5)
            
            logger.info(f"Data calculada automaticamente: {data_rel.strftime('%d/%m/%Y')}")
            return data_rel
            
        except Exception as e:
            logger.error(f"Erro ao calcular data automática: {str(e)}")
            # Fallback: retorna data atual
            return datetime.now()

    def explicar_regra_data(self):
        """Retorna explicação da regra de cálculo de data"""
        hoje = datetime.now()
        data_calculada = self.calcular_data_rel_automatica()
        
        if 6 <= hoje.day <= 20:
            explicacao = f"📅 Relatório do dia 20"
        elif hoje.day > 20:
            explicacao = f"📅 Relatório do dia 5 do próximo mês"
        else:
            explicacao = f"📅 Relatório do dia 5"
        
        return f"{explicacao}\n🎯 Data: {data_calculada.strftime('%d/%m/%Y')}"

    def validar_data_relatorio(self, data_selecionada):
        """Valida se a data selecionada está correta conforme a regra"""
        try:
            if isinstance(data_selecionada, str):
                data_selecionada = datetime.strptime(data_selecionada, '%d/%m/%Y')
            
            # Verificar se é dia 5 ou 20
            if data_selecionada.day not in [5, 20]:
                return False, f"❌ Data deve ser dia 5 ou 20 do mês.\nData selecionada: {data_selecionada.strftime('%d/%m/%Y')}"
            
            # Verificar se está no período correto
            data_automatica = self.calcular_data_rel_automatica()
            
            if data_selecionada.date() == data_automatica.date():
                return True, f"✅ Data correta para o período atual"
            else:
                return True, f"⚠️ Data válida, mas não é a sugerida para hoje.\nSugerida: {data_automatica.strftime('%d/%m/%Y')}"
            
        except Exception as e:
            return False, f"❌ Erro ao validar data: {str(e)}"

    def setup_opcoes_despesas(self, parent_frame):
        """Versão otimizada com seleção de cliente via combobox"""
        
        # Frame para data com cálculo automático
        frame_data = ttk.LabelFrame(parent_frame, text="Data do Relatório")
        frame_data.pack(fill='x', padx=10, pady=10)
        
        # Calcular data automática
        data_automatica = self.calcular_data_rel_automatica()
        
        # Área de informações sobre a regra
        info_frame = ttk.Frame(frame_data)
        info_frame.pack(fill='x', padx=10, pady=5)
        
        # Label explicativa
        explicacao = self.explicar_regra_data()
        ttk.Label(info_frame, text=explicacao, font=('Arial', 9), foreground='blue').pack(anchor='w')
        
        # Frame para seleção de data
        selecao_frame = ttk.Frame(frame_data)
        selecao_frame.pack(fill='x', padx=10, pady=5)
        
        # Opção de usar data automática (padrão)
        self.usar_data_automatica = tk.BooleanVar(master=self.root, value=True)
        
        ttk.Checkbutton(
            selecao_frame,
            text="Usar data calculada automaticamente",
            variable=self.usar_data_automatica,
            command=self.alternar_modo_data
        ).pack(anchor='w', pady=2)
        
        # Frame para data manual (inicialmente oculto)
        self.frame_data_manual = ttk.Frame(frame_data)
        
        ttk.Label(self.frame_data_manual, text="Data manual:").pack(side='left', padx=5)
        
        # DateEntry para seleção manual
        try:
            from tkcalendar import DateEntry
            self.data_entry = DateEntry(
                self.frame_data_manual,
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd/mm/yyyy',
                locale='pt_BR'
            )
            self.data_entry.pack(side='left', padx=5)
            
            # Botão para validar data manual
            ttk.Button(
                self.frame_data_manual,
                text="Validar Data",
                command=self.validar_data_manual
            ).pack(side='left', padx=5)
            
        except ImportError:
            ttk.Label(
                self.frame_data_manual, 
                text="Módulo tkcalendar não encontrado"
            ).pack(side='left')
        
        # Configurar data inicial
        self.data_automatica_calculada = data_automatica
        if hasattr(self, 'data_entry'):
            self.data_entry.set_date(data_automatica)

        # === NOVA SEÇÃO: SELEÇÃO DE CLIENTE ===
        frame_cliente = ttk.LabelFrame(parent_frame, text="Seleção de Cliente")
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        # Frame interno para organizar melhor
        cliente_inner_frame = ttk.Frame(frame_cliente)
        cliente_inner_frame.pack(fill='x', padx=10, pady=10)
        
        # Label e Combobox de cliente
        ttk.Label(cliente_inner_frame, text="Cliente:", font=('Arial', 10, 'bold')).pack(anchor='w', pady=(0, 5))
        
        self.cliente_combobox = ttk.Combobox(
            cliente_inner_frame, 
            width=50,
            state='readonly',  # Apenas seleção, não digitação
            font=('Arial', 10)
        )
        self.cliente_combobox.pack(fill='x', pady=(0, 10))
        
        # Preencher combobox com clientes
        self.preencher_combobox_clientes(self.cliente_combobox)
        
        # Bind para evento de seleção
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.on_cliente_selecionado)
        
        # Label para mostrar status da seleção
        self.status_cliente_label = ttk.Label(
            cliente_inner_frame, 
            text="Selecione um cliente para continuar",
            font=('Arial', 9),
            foreground='gray'
        )
        self.status_cliente_label.pack(anchor='w', pady=(0, 10))
        
        # Frame para botões adicionais de cliente
        botoes_cliente_frame = ttk.Frame(cliente_inner_frame)
        botoes_cliente_frame.pack(fill='x')
        
        # Botão para atualizar lista de clientes
        ttk.Button(
            botoes_cliente_frame,
            text="🔄 Atualizar Lista",
            command=self.atualizar_lista_clientes_despesas,
            width=15
        ).pack(side='left', padx=(0, 10))
        
        # Botão para seleção manual de arquivo (fallback)
        ttk.Button(
            botoes_cliente_frame,
            text="📁 Selecionar Arquivo Manual",
            command=self.selecionar_arquivo_manual_despesas,
            width=25
        ).pack(side='left')
        
        # === OPÇÕES DE PROCESSAMENTO ===
        frame_opcoes = ttk.LabelFrame(parent_frame, text="Opções de Processamento")
        frame_opcoes.pack(fill='x', padx=10, pady=10)
        
        # Checkbox para incluir lançamentos futuros
        self.incluir_futuros = tk.BooleanVar(master=self.root, value=False)
        ttk.Checkbutton(
            frame_opcoes,
            text="Incluir lançamentos futuros",
            variable=self.incluir_futuros
        ).pack(anchor='w', padx=15, pady=5)
        
        # Checkbox para incluir lançamentos excluídos
        self.incluir_excluidos = tk.BooleanVar(master=self.root, value=False)
        ttk.Checkbutton(
            frame_opcoes,
            text="Incluir lançamentos excluídos no relatório",
            variable=self.incluir_excluidos
        ).pack(anchor='w', padx=15, pady=5)

        # Checkbox para incluir notas no relatório
        self.incluir_notas = tk.BooleanVar(master=self.root, value=False)
        ttk.Checkbutton(
            frame_opcoes,
            text="Incluir notas no relatório",
            variable=self.incluir_notas,
            command=self.abrir_janela_notas_despesas
        ).pack(anchor='w', padx=15, pady=5)

        # Label de status das notas
        self.label_notas_status = ttk.Label(
            frame_opcoes,
            text="",
            foreground='green',
            font=('Arial', 9)
        )
        self.label_notas_status.pack(anchor='w', padx=30, pady=2)

        # Variável para armazenar o texto das notas
        self.texto_notas = tk.StringVar(master=self.root, value="")
        
        # === TIPO DE GERAÇÃO ===
        frame_tipo = ttk.LabelFrame(parent_frame, text="Tipo de Geração")
        frame_tipo.pack(fill='x', padx=10, pady=10)
        
        self.tipo_geracao = tk.StringVar(master=self.root, value="individual")
        
        ttk.Radiobutton(
            frame_tipo,
            text="Relatório Individual",
            variable=self.tipo_geracao,
            value="individual",
            command=self.alternar_tipo_geracao
        ).pack(anchor='w', padx=15, pady=5)
        
        ttk.Radiobutton(
            frame_tipo,
            text="Relatório em Lote",
            variable=self.tipo_geracao,
            value="lote",
            command=self.alternar_tipo_geracao
        ).pack(anchor='w', padx=15, pady=5)
        
        # === FRAMES PARA TIPOS ESPECÍFICOS ===
        
        # Frame para seleção individual (já preenchido com cliente selecionado)
        self.frame_individual = ttk.Frame(parent_frame)
        self.frame_individual.pack(fill='x', padx=10, pady=10)
        
        # Label de status para individual
        # self.status_individual_label = ttk.Label(
        #     self.frame_individual,
        #     text="Cliente será selecionado através da combobox acima",
        #     font=('Arial', 9),
        #     foreground='blue'
        # )
        # self.status_individual_label.pack(anchor='w', padx=15, pady=5)
        
        # Frame para seleção em lote
        self.frame_lote = ttk.Frame(parent_frame)
        
        ttk.Button(
            self.frame_lote,
            text="Selecionar Arquivos para Lote",
            command=self.selecionar_arquivos_lote
        ).pack(anchor='w', padx=15, pady=10)
        
        self.lbl_arquivos_lote = ttk.Label(self.frame_lote, text="")
        self.lbl_arquivos_lote.pack(anchor='w', padx=15, pady=5)
        
        self.arquivos_lote = []
        
        # === MODO DE VISUALIZAÇÃO ===
        frame_visualizacao = ttk.LabelFrame(parent_frame, text="Modo de Visualização")
        frame_visualizacao.pack(fill='x', padx=10, pady=10)
        
        self.modo_visualizacao = tk.StringVar(master=self.root, value="preview")
        ttk.Radiobutton(
            frame_visualizacao,
            text="Gerar com Preview",
            variable=self.modo_visualizacao,
            value="preview"
        ).pack(side='left', padx=20, pady=5)
        
        ttk.Radiobutton(
            frame_visualizacao,
            text="Gerar Direto",
            variable=self.modo_visualizacao,
            value="direto"
        ).pack(side='left', padx=20, pady=5)
        
        # === FORMATO DE SAÍDA ===
        frame_formato = ttk.LabelFrame(parent_frame, text="Formato de Saída")
        frame_formato.pack(fill='x', padx=10, pady=10)
        
        self.formato_saida = tk.StringVar(master=self.root, value="pdf")
        ttk.Radiobutton(
            frame_formato,
            text="PDF",
            variable=self.formato_saida,
            value="pdf"
        ).pack(side='left', padx=20, pady=5)
        
        ttk.Radiobutton(
            frame_formato,
            text="Excel",
            variable=self.formato_saida,
            value="excel"
        ).pack(side='left', padx=20, pady=5)
        
        # Inicializar mostrando apenas a opção individual
        self.frame_lote.pack_forget()
        
        # Configurar variáveis de controle
        self.arquivo_cliente_selecionado = None
        self.cliente_atual = None

    def on_cliente_selecionado(self, event=None):
        """Trata a seleção de um cliente na combobox"""
        try:
            cliente_selecionado = self.cliente_combobox.get()
            logger.info(f"Cliente selecionado: {cliente_selecionado}")
            
            if not cliente_selecionado or cliente_selecionado == 'Todos os Clientes':
                self.limpar_selecao_cliente()
                return
            
            # Buscar arquivo do cliente
            caminho_arquivo = self.buscar_arquivo_cliente(cliente_selecionado)
            
            if caminho_arquivo and os.path.exists(caminho_arquivo):
                self.arquivo_cliente_selecionado = caminho_arquivo
                self.cliente_atual = cliente_selecionado
                
                # Atualizar status
                self.status_cliente_label.config(
                    text=f"✅ Arquivo: {os.path.basename(caminho_arquivo)}",
                    foreground='green'
                )
                
                # Atualizar status individual
                if hasattr(self, 'status_individual_label'):
                    self.status_individual_label.config(
                        text=f"✅ Arquivo selecionado: {os.path.basename(caminho_arquivo)}",
                        foreground='green'
                    )
                
                logger.info(f"Arquivo encontrado: {caminho_arquivo}")
                
            else:
                self.status_cliente_label.config(
                    text=f"❌ Arquivo não encontrado para {cliente_selecionado}",
                    foreground='red'
                )
                
                # Oferecer seleção manual
                resposta = messagebox.askyesno(
                    "Arquivo não encontrado",
                    f"Não foi encontrado arquivo para o cliente '{cliente_selecionado}'.\n\n"
                    f"Deseja selecionar manualmente o arquivo deste cliente?"
                )
                
                if resposta:
                    self.selecionar_arquivo_manual_despesas()
                else:
                    self.limpar_selecao_cliente()
                    
        except Exception as e:
            logger.error(f"Erro ao selecionar cliente: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar cliente: {str(e)}")

    def buscar_arquivo_cliente(self, nome_cliente):
        """Busca o arquivo Excel do cliente especificado"""
        try:
            # Importar configurações de pasta
            try:
                from src.config.config import PASTA_CLIENTES
            except ImportError:
                try:
                    from config.config import PASTA_CLIENTES
                except ImportError:
                    # Pasta padrão
                    PASTA_CLIENTES = "clientes"
            
            # Verificar se PASTA_CLIENTES existe
            if not os.path.exists(PASTA_CLIENTES):
                logger.warning(f"Pasta de clientes não encontrada: {PASTA_CLIENTES}")
                # Tentar pasta relativa
                pasta_alternativa = os.path.join(os.path.dirname(__file__), "..", "clientes")
                if os.path.exists(pasta_alternativa):
                    PASTA_CLIENTES = pasta_alternativa
                else:
                    return None
            
            # Possíveis nomes de arquivo
            possíveis_nomes = [
                f"{nome_cliente}.xlsx",
                f"{nome_cliente}.xls",
                f"{nome_cliente.upper()}.xlsx",
                f"{nome_cliente.lower()}.xlsx",
                f"{nome_cliente.replace(' ', '_')}.xlsx",
                f"{nome_cliente.replace(' ', '')}.xlsx"
            ]
            
            # Buscar arquivo
            for nome_arquivo in possíveis_nomes:
                caminho_completo = os.path.join(PASTA_CLIENTES, nome_arquivo)
                if os.path.exists(caminho_completo):
                    logger.info(f"Arquivo encontrado: {caminho_completo}")
                    return caminho_completo
            
            # Se não encontrou, listar arquivos na pasta para debug
            try:
                arquivos_existentes = os.listdir(PASTA_CLIENTES)
                logger.debug(f"Arquivos na pasta {PASTA_CLIENTES}: {arquivos_existentes}")
            except:
                pass
            
            logger.warning(f"Arquivo não encontrado para cliente: {nome_cliente}")
            return None
            
        except Exception as e:
            logger.error(f"Erro ao buscar arquivo do cliente: {str(e)}")
            return None

    def selecionar_arquivo_manual_despesas(self):
        """Permite seleção manual de arquivo (fallback)"""
        try:
            arquivo = filedialog.askopenfilename(
                title="Selecione o arquivo Excel do cliente",
                filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
                initialdir=self.obter_pasta_clientes()
            )
            
            if arquivo:
                # Verificar se o arquivo é válido
                if not os.path.exists(arquivo):
                    messagebox.showerror("Erro", "Arquivo não encontrado.")
                    return
                    
                try:
                    # Tentar abrir o arquivo para verificar se é válido
                    from openpyxl import load_workbook
                    wb = load_workbook(arquivo, data_only=True)
                    
                    # Tentar obter nome do cliente do arquivo
                    try:
                        ws_resumo = wb['RESUMO']
                        nome_cliente_arquivo = ws_resumo['A3'].value
                        if nome_cliente_arquivo:
                            self.cliente_atual = nome_cliente_arquivo
                            # Atualizar combobox para mostrar o cliente correto
                            self.cliente_combobox.set(nome_cliente_arquivo)
                    except:
                        # Se não conseguir obter nome, usar nome do arquivo
                        self.cliente_atual = os.path.splitext(os.path.basename(arquivo))[0]
                    
                    wb.close()
                    
                    # Configurar arquivo selecionado
                    self.arquivo_cliente_selecionado = arquivo
                    
                    # Atualizar status
                    self.status_cliente_label.config(
                        text=f"✅ Arquivo selecionado manualmente: {os.path.basename(arquivo)}",
                        foreground='blue'
                    )
                    
                    if hasattr(self, 'status_individual_label'):
                        self.status_individual_label.config(
                            text=f"✅ Arquivo: {os.path.basename(arquivo)}",
                            foreground='blue'
                        )
                    
                    logger.info(f"Arquivo selecionado manualmente: {arquivo}")
                    
                except Exception as e:
                    messagebox.showerror(
                        "Erro", 
                        f"Arquivo inválido ou corrompido.\nErro: {str(e)}"
                    )
                    
        except Exception as e:
            logger.error(f"Erro na seleção manual: {str(e)}")
            messagebox.showerror("Erro", f"Erro na seleção manual: {str(e)}")

    def limpar_selecao_cliente(self):
        """Limpa a seleção de cliente atual"""
        self.arquivo_cliente_selecionado = None
        self.cliente_atual = None
        self.cliente_combobox.set('Todos os Clientes')
        
        self.status_cliente_label.config(
            text="Selecione um cliente para continuar",
            foreground='gray'
        )
        
        if hasattr(self, 'status_individual_label'):
            self.status_individual_label.config(
                text="Cliente será selecionado através da combobox acima",
                foreground='blue'
            )

    def atualizar_lista_clientes_despesas(self):
        """Atualiza a lista de clientes especificamente para despesas"""
        try:
            # Salvar seleção atual
            cliente_atual = self.cliente_combobox.get()
            
            # Recarregar lista
            self.atualizar_lista_clientes()
            
            # Tentar restaurar seleção
            if cliente_atual and cliente_atual in self.cliente_combobox['values']:
                self.cliente_combobox.set(cliente_atual)
            else:
                self.cliente_combobox.set('Todos os Clientes')
            
            messagebox.showinfo("Sucesso", "Lista de clientes atualizada!")
            logger.info("Lista de clientes atualizada na interface de despesas")
            
        except Exception as e:
            logger.error(f"Erro ao atualizar lista: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao atualizar lista: {str(e)}")

    def obter_pasta_clientes(self):
        """Obtém o caminho da pasta de clientes"""
        try:
            from src.config.config import PASTA_CLIENTES
            return PASTA_CLIENTES
        except ImportError:
            try:
                from config.config import PASTA_CLIENTES
                return PASTA_CLIENTES
            except ImportError:
                return "clientes"

    def alternar_modo_data(self):
        """Alterna entre data automática e manual"""
        try:
            if self.usar_data_automatica.get():
                # Usar data automática - ocultar seleção manual
                self.frame_data_manual.pack_forget()
                
                # Recalcular data automática
                data_auto = self.calcular_data_rel_automatica()
                self.data_automatica_calculada = data_auto
                
                if hasattr(self, 'data_entry'):
                    self.data_entry.set_date(data_auto)
                    
                logger.info(f"Modo automático ativado: {data_auto.strftime('%d/%m/%Y')}")
                
            else:
                # Usar data manual - mostrar seleção
                self.frame_data_manual.pack(fill='x', padx=10, pady=5)
                logger.info("Modo manual ativado")
                
        except Exception as e:
            logger.error(f"Erro ao alternar modo de data: {str(e)}")

    def validar_data_manual(self):
        """Valida a data inserida manualmente"""
        try:
            if hasattr(self, 'data_entry'):
                data_selecionada = self.data_entry.get_date()
                valida, mensagem = self.validar_data_relatorio(data_selecionada)
                
                if valida:
                    messagebox.showinfo("Validação de Data", mensagem)
                else:
                    messagebox.showerror("Data Inválida", mensagem)
                    
        except Exception as e:
            logger.error(f"Erro ao validar data manual: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao validar data: {str(e)}")

    def obter_data_relatorio_final(self):
        """Versão corrigida que retorna data sem hora"""
        try:
            if self.usar_data_automatica.get():
                data = self.data_automatica_calculada
            else:
                if hasattr(self, 'data_entry'):
                    data = self.data_entry.get_date()
                else:
                    data = self.data_automatica_calculada
            
            # CORREÇÃO: Garantir que retorna apenas a data sem hora
            from datetime import datetime, date
            
            if isinstance(data, datetime):
                # Se é datetime, pegar apenas a parte da data
                data = data.date()
            
            # Converter para datetime no início do dia para processamento
            if isinstance(data, date):
                data = datetime.combine(data, datetime.min.time())
            
            logger.info(f"Data final obtida: {data} (tipo: {type(data)})")
            return data
            
        except Exception as e:
            logger.error(f"Erro ao obter data final: {str(e)}")
            from datetime import datetime
            return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    def alternar_tipo_geracao(self):
        """Alterna entre opções de geração individual e em lote"""
        try:
            if self.tipo_geracao.get() == "individual":
                self.frame_lote.pack_forget()
                self.frame_individual.pack(fill='x', padx=10, pady=10)
            else:
                self.frame_individual.pack_forget()
                self.frame_lote.pack(fill='x', padx=10, pady=10)
                
            # NOVO: Atualizar botão de geração
            self.atualizar_botao_geracao()
            
        except Exception as e:
            logger.error(f"Erro ao alternar tipo de geração: {str(e)}")

    def selecionar_arquivos_lote(self):
        """Abre diálogo para selecionar múltiplos arquivos para geração em lote"""
        try:
            arquivos = filedialog.askopenfilenames(
                title="Selecione os arquivos Excel",
                filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
            )
            if arquivos:
                # Validar todos os arquivos
                arquivos_validos = []
                arquivos_invalidos = []
                
                for arquivo in arquivos:
                    if os.path.exists(arquivo):
                        try:
                            # Verificar se é um arquivo Excel válido
                            from openpyxl import load_workbook
                            wb = load_workbook(arquivo, data_only=True)
                            wb.close()
                            arquivos_validos.append(arquivo)
                        except:
                            arquivos_invalidos.append(os.path.basename(arquivo))
                    else:
                        arquivos_invalidos.append(os.path.basename(arquivo))
                
                if arquivos_invalidos:
                    messagebox.showwarning(
                        "Arquivos Inválidos",
                        f"Os seguintes arquivos não puderam ser carregados:\n" +
                        "\n".join(arquivos_invalidos)
                    )
                
                if arquivos_validos:
                    self.arquivos_lote = arquivos_validos
                    self.lbl_arquivos_lote.config(
                        text=f"{len(arquivos_validos)} arquivos válidos selecionados"
                    )
                    logger.info(f"Selecionados {len(arquivos_validos)} arquivos para lote")
                else:
                    messagebox.showerror("Erro", "Nenhum arquivo válido foi selecionado.")
                    
        except Exception as e:
            logger.error(f"Erro ao selecionar arquivos em lote: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar arquivos: {str(e)}")
    
    def abrir_janela_notas_despesas(self):
        """Abre janela para edição de notas do relatório de despesas"""
        if not self.incluir_notas.get():
            # Se desmarcou o checkbox, limpar as notas
            self.texto_notas.set("")
            self.label_notas_status.config(text="")
            return
        
        # Criar janela de edição
        janela_notas = tk.Toplevel(self.root)
        janela_notas.title("Notas do Relatório")
        janela_notas.geometry("700x600")
        janela_notas.transient(self.root)
        janela_notas.grab_set()
        
        # Centralizar janela
        janela_notas.update_idletasks()
        x = (janela_notas.winfo_screenwidth() // 2) - (janela_notas.winfo_width() // 2)
        y = (janela_notas.winfo_screenheight() // 2) - (janela_notas.winfo_height() // 2)
        janela_notas.geometry(f"+{x}+{y}")
        
        # Frame principal com padding
        frame_principal = ttk.Frame(janela_notas, padding="20")
        frame_principal.pack(fill='both', expand=True)
        
        # Título com instrução
        label_titulo = ttk.Label(
            frame_principal, 
            text="Digite as notas que aparecerão no relatório:",
            font=('Helvetica', 11, 'bold')
        )
        label_titulo.pack(pady=(0, 5))
        
        # Informação adicional
        label_info = ttk.Label(
            frame_principal,
            text="As notas serão exibidas na seção 'NOTAS:' do relatório PDF",
            font=('Helvetica', 9),
            foreground='gray'
        )
        label_info.pack(pady=(0, 15))
        
        # Frame para texto com scrollbar
        frame_texto = ttk.Frame(frame_principal)
        frame_texto.pack(fill='both', expand=True, pady=(0, 15))
        
        # Área de texto com scrollbar
        texto_widget = tk.Text(
            frame_texto, 
            wrap='word', 
            font=('Arial', 10), 
            relief='solid', 
            borderwidth=1,
            padx=10,
            pady=10
        )
        scrollbar = tk.Scrollbar(frame_texto, orient='vertical', command=texto_widget.yview)
        texto_widget.configure(yscrollcommand=scrollbar.set)
        
        texto_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Carregar texto existente se houver
        if self.texto_notas.get():
            texto_widget.insert('1.0', self.texto_notas.get())
        else:
            # Texto de exemplo/placeholder
            exemplo = (
                "Exemplo de notas:\n\n"
                "Aguardando cópia das Notas Fiscais para incluir no caderno físico da quinzena:\n\n"
                "- LOJA DO PAULO\n"
                "- LOJA ELÉTRICA\n"
            )
            texto_widget.insert('1.0', exemplo)
            texto_widget.tag_add('exemplo', '1.0', 'end')
            texto_widget.tag_config('exemplo', foreground='gray')
            
            # Remover placeholder ao clicar
            def limpar_placeholder(event):
                if texto_widget.tag_ranges('exemplo'):
                    texto_widget.delete('1.0', 'end')
                    texto_widget.tag_remove('exemplo', '1.0', 'end')
                    texto_widget.config(foreground='black')
            
            texto_widget.bind('<FocusIn>', limpar_placeholder)
        
        # Focar no texto
        texto_widget.focus_set()
        
        # Frame para botões
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x')
        
        # Função para salvar notas
        def salvar_notas():
            """Salva o texto das notas"""
            texto = texto_widget.get('1.0', 'end-1c').strip()
            
            # Ignorar texto de exemplo
            if texto_widget.tag_ranges('exemplo'):
                texto = ""
            
            self.texto_notas.set(texto)
            
            if texto:
                # Mostrar preview curto no label
                preview = texto[:60] + "..." if len(texto) > 60 else texto
                self.label_notas_status.config(
                    text=f"✓ Notas adicionadas",
                    foreground='green'
                )
                logger.info(f"Notas salvas: {len(texto)} caracteres")
            else:
                self.label_notas_status.config(text="")
                self.incluir_notas.set(False)
            
            janela_notas.destroy()
        
        # Função para cancelar
        def cancelar():
            """Cancela e desmarca o checkbox"""
            self.incluir_notas.set(False)
            self.texto_notas.set("")
            self.label_notas_status.config(text="")
            janela_notas.destroy()
        
        # Botões com ícones
        btn_salvar = ttk.Button(
            frame_botoes, 
            text="✓ Salvar Notas", 
            command=salvar_notas
        )
        btn_salvar.pack(side='left', padx=(0, 10))
        
        btn_cancelar = ttk.Button(
            frame_botoes, 
            text="✗ Cancelar", 
            command=cancelar
        )
        btn_cancelar.pack(side='left')
        
        # Label com dicas de atalhos
        label_atalhos = ttk.Label(
            frame_botoes,
            text="Dica: Ctrl+Enter para salvar | Esc para cancelar",
            font=('Arial', 8),
            foreground='gray'
        )
        label_atalhos.pack(side='right')
        
        # Atalhos de teclado
        janela_notas.bind('<Control-Return>', lambda e: salvar_notas())
        janela_notas.bind('<Escape>', lambda e: cancelar())
        
        # Configurar fechamento da janela
        janela_notas.protocol("WM_DELETE_WINDOW", cancelar)

    def setup_opcoes_contratos(self, parent_frame):
        """Configura as opções específicas para relatório de contratos e medições"""
        # Frame para data
        frame_data = ttk.Frame(parent_frame)
        frame_data.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_data, text="Data de Referência:").pack(side='left', padx=5)
        
        # Importar DateEntry apenas quando necessário
        try:
            from tkcalendar import DateEntry
            self.data_referencia = DateEntry(
                frame_data,
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd/mm/yyyy',
                locale='pt_BR'
            )
            self.data_referencia.pack(side='left', padx=5)
        except ImportError:
            # Fallback se tkcalendar não estiver instalado
            ttk.Label(frame_data, text="Módulo tkcalendar não encontrado. Data atual será usada.").pack(side='left')
        
        # Frame para seleção de cliente
        frame_cliente = ttk.Frame(parent_frame)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Cliente:").pack(side='left', padx=5)
        
        # Combobox para seleção de cliente
        self.cliente_contratos = ttk.Combobox(frame_cliente, width=40)
        self.cliente_contratos.pack(side='left', padx=5)
        
        # Preencher com clientes reais
        self.preencher_combobox_clientes(self.cliente_contratos)
        
        # Opções de visualização
        frame_visualizacao = ttk.LabelFrame(parent_frame, text="Opções de Visualização")
        frame_visualizacao.pack(fill='x', padx=10, pady=10)
        
        # Checkboxes para diferentes visualizações
        self.mostrar_resumo = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Resumo",
            variable=self.mostrar_resumo
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_detalhes = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Detalhes",
            variable=self.mostrar_detalhes
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_grafico = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Incluir Gráficos",
            variable=self.mostrar_grafico
        ).pack(anchor='w', padx=15, pady=5)
        
        # Frame para formato de saída
        frame_formato = ttk.LabelFrame(parent_frame, text="Formato de Saída")
        frame_formato.pack(fill='x', padx=10, pady=10)
        
        self.formato_contratos = tk.StringVar(master=self.root, value="excel")
        ttk.Radiobutton(
            frame_formato,
            text="Excel",
            variable=self.formato_contratos,
            value="excel"
        ).pack(side='left', padx=20, pady=5)
        
        ttk.Radiobutton(
            frame_formato,
            text="PDF",
            variable=self.formato_contratos,
            value="pdf"
        ).pack(side='left', padx=20, pady=5)

    def setup_opcoes_quinzenal(self, parent_frame):
        """
        Configura as opções do Relatório Quinzenal de Medições
        
        Este método substitui a mensagem genérica amarela por configurações reais
        """
        try:
            # Limpar frame
            for widget in parent_frame.winfo_children():
                widget.destroy()
            
            # Usar a configuração completa do relatório quinzenal
            from config_relatorio_quinzenal import configurar_relatorio_quinzenal
            configurar_relatorio_quinzenal(parent_frame, self)
            
            logger.info("✅ Opções do Relatório Quinzenal configuradas")
            
        except Exception as e:
            logger.error(f"Erro ao configurar opções quinzenal: {str(e)}", exc_info=True)
            # Se falhar, mostra mensagem de erro ao invés da mensagem genérica
            ttk.Label(
                parent_frame,
                text=f"Erro ao carregar configurações: {str(e)}",
                font=('Arial', 10),
                foreground='red',
                wraplength=400
            ).pack(pady=20)

    def setup_opcoes_categoria(self, parent_frame):
        """Configura as opções específicas para relatório por tipo de despesa"""
        # Frame para seleção de cliente
        frame_cliente = ttk.Frame(parent_frame)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Cliente:").pack(side='left', padx=5)
        
        # Combobox para seleção de cliente
        self.cliente_categoria = ttk.Combobox(frame_cliente, width=40)
        self.cliente_categoria.pack(side='left', padx=5)
        
        # Preencher com clientes reais
        self.preencher_combobox_clientes(self.cliente_categoria)
        
        # Descrição do relatório
        ttk.Label(
            parent_frame,
            text="Este relatório mostra os dados agrupados por data,\n" +
                "com colunas para cada tipo de categoria e seus totais.",
            justify='center',
            font=('Arial', 10),
            foreground='gray'
        ).pack(pady=10)
        
        # Opções de visualização (opcional)
        frame_visualizacao = ttk.LabelFrame(parent_frame, text="Opções de Visualização")
        frame_visualizacao.pack(fill='x', padx=10, pady=10)
        
        # Checkboxes para diferentes visualizações
        self.mostrar_resumo_td = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Resumo",
            variable=self.mostrar_resumo_td
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_detalhes_td = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Detalhes",
            variable=self.mostrar_detalhes_td
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_grafico_td = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Incluir Gráficos",
            variable=self.mostrar_grafico_td
        ).pack(anchor='w', padx=15, pady=5)
        
        # Frame para formato de saída
        frame_formato = ttk.LabelFrame(parent_frame, text="Formato de Saída")
        frame_formato.pack(fill='x', padx=10, pady=10)
        
        self.formato_categoria = tk.StringVar(master=self.root, value="excel")
        ttk.Radiobutton(
            frame_formato,
            text="Excel",
            variable=self.formato_categoria,
            value="excel"
        ).pack(side='left', padx=20, pady=5)
        
        ttk.Radiobutton(
            frame_formato,
            text="PDF",
            variable=self.formato_categoria,
            value="pdf"
        ).pack(side='left', padx=20, pady=5)

    def setup_opcoes_tipo_despesa(self, parent_frame):
        """Configura as opções específicas para relatório por tipo de despesa"""
        # Frame para seleção de cliente
        frame_cliente = ttk.Frame(parent_frame)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Cliente:").pack(side='left', padx=5)
        
        # Combobox para seleção de cliente
        self.cliente_tipo_despesa = ttk.Combobox(frame_cliente, width=40)
        self.cliente_tipo_despesa.pack(side='left', padx=5)
        
        # Preencher com clientes reais
        self.preencher_combobox_clientes(self.cliente_tipo_despesa)
        
        # Descrição do relatório
        ttk.Label(
            parent_frame,
            text="Este relatório mostra os dados agrupados por data, \n" +
                "com colunas para cada tipo de despesa e seus totais.",
            justify='center',
            font=('Arial', 10),
            foreground='gray'
        ).pack(pady=10)
        
        # Opções de visualização (opcional)
        frame_visualizacao = ttk.LabelFrame(parent_frame, text="Opções de Visualização")
        frame_visualizacao.pack(fill='x', padx=10, pady=10)
        
        # Checkboxes para diferentes visualizações
        self.mostrar_resumo_td = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Resumo",
            variable=self.mostrar_resumo_td
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_detalhes_td = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Detalhes",
            variable=self.mostrar_detalhes_td
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_grafico_td = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Incluir Gráficos",
            variable=self.mostrar_grafico_td
        ).pack(anchor='w', padx=15, pady=5)
        
        # Frame para formato de saída
        frame_formato = ttk.LabelFrame(parent_frame, text="Formato de Saída")
        frame_formato.pack(fill='x', padx=10, pady=10)
        
        self.formato_tipo_despesa = tk.StringVar(master=self.root, value="excel")
        ttk.Radiobutton(
            frame_formato,
            text="Excel",
            variable=self.formato_tipo_despesa,
            value="excel"
        ).pack(side='left', padx=20, pady=5)
        
        ttk.Radiobutton(
            frame_formato,
            text="PDF",
            variable=self.formato_tipo_despesa,
            value="pdf"
        ).pack(side='left', padx=20, pady=5)

    def setup_opcoes_fornecedores(self, parent_frame):
        """Configura as opções específicas para relatório de fornecedores"""
        # Frame para data
        frame_data = ttk.Frame(parent_frame)
        frame_data.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_data, text="Data de Referência:").pack(side='left', padx=5)
        
        # Importar DateEntry apenas quando necessário
        try:
            from tkcalendar import DateEntry
            self.data_referencia = DateEntry(
                frame_data,
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd/mm/yyyy',
                locale='pt_BR'
            )
            self.data_referencia.pack(side='left', padx=5)
        except ImportError:
            # Fallback se tkcalendar não estiver instalado
            ttk.Label(frame_data, text="Módulo tkcalendar não encontrado. Data atual será usada.").pack(side='left')
       # Frame para seleção de cliente
        frame_cliente = ttk.Frame(parent_frame)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Cliente:").pack(side='left', padx=5)
        
        # Combobox para seleção de cliente
        self.cliente_contratos = ttk.Combobox(frame_cliente, width=40)
        self.cliente_contratos.pack(side='left', padx=5)
        
        # Preencher com clientes reais
        self.preencher_combobox_clientes(self.cliente_contratos)
        
        # Opções de visualização
        frame_visualizacao = ttk.LabelFrame(parent_frame, text="Opções de Visualização")
        frame_visualizacao.pack(fill='x', padx=10, pady=10)
        
        # Checkboxes para diferentes visualizações
        self.mostrar_resumo = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Resumo",
            variable=self.mostrar_resumo
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_detalhes = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Mostrar Detalhes",
            variable=self.mostrar_detalhes
        ).pack(anchor='w', padx=15, pady=5)
        
        self.mostrar_grafico = tk.BooleanVar(master=self.root, value=True)
        ttk.Checkbutton(
            frame_visualizacao,
            text="Incluir Gráficos",
            variable=self.mostrar_grafico
        ).pack(anchor='w', padx=15, pady=5)
        
        # Frame para formato de saída
        frame_formato = ttk.LabelFrame(parent_frame, text="Formato de Saída")
        frame_formato.pack(fill='x', padx=10, pady=10)
        
        self.formato_contratos = tk.StringVar(master=self.root, value="excel")
        ttk.Radiobutton(
            frame_formato,
            text="Excel",
            variable=self.formato_contratos,
            value="excel"
        ).pack(side='left', padx=20, pady=5)
        
        ttk.Radiobutton(
            frame_formato,
            text="PDF",
            variable=self.formato_contratos,
            value="pdf"
        ).pack(side='left', padx=20, pady=5)

    def setup_opcoes_gerencial_engenheiro(self, parent):
        """
        Configura opções para o relatório gerencial por engenheiro
        
        Este relatório não precisa de muitas configurações pois trabalha
        com dados consolidados de todos os clientes de um grupo.
        """
        try:
            # Frame principal
            main_frame = ttk.Frame(parent, padding=10)
            main_frame.pack(fill='both', expand=True)
            
            # Descrição
            desc_text = """
    Este relatório consolida informações de TODAS as obras de um grupo/engenheiro:

    📊 O que mostra:
    • Resumo executivo do grupo (obras, contratos, valores)
    • Visão por obra (lista consolidada)
    • Todos os contratos do grupo
    • Todas as medições do grupo
    • Gráficos de acompanhamento

    🎯 Ideal para:
    • Engenheiros acompanharem múltiplas obras
    • Gestores visualizarem performance de grupos
    • Identificação rápida de problemas
    • Relatórios executivos para diretoria

    💡 O relatório é gerado em tempo real a partir dos dados
    das planilhas individuais de cada cliente do grupo.
            """
            
            ttk.Label(
                main_frame,
                text=desc_text,
                justify='left',
                wraplength=450,
                font=('Arial', 10)
            ).pack(pady=10, anchor='w')
            
            # Separador
            ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=15)
            
            # Informações importantes
            info_frame = ttk.LabelFrame(main_frame, text="ℹ️ Informações Importantes", padding=10)
            info_frame.pack(fill='x', pady=10)
            
            info_text = """
    ✓ Os grupos são definidos na planilha clientes.xlsx (coluna "Grupo")
    ✓ Valores válidos: Grupo 1, Grupo 2, Grupo 3, Grupo 4
    ✓ O relatório filtrará automaticamente os clientes do grupo selecionado
    ✓ Você pode exportar os resultados para Excel
    ✓ Gráficos são gerados automaticamente
            """
            
            ttk.Label(
                info_frame,
                text=info_text,
                justify='left',
                wraplength=430,
                font=('Arial', 9)
            ).pack(anchor='w')
            
            # Nota sobre configurações
            nota_frame = ttk.Frame(main_frame)
            nota_frame.pack(fill='x', pady=(20, 10))
            
            ttk.Label(
                nota_frame,
                text="💡 Este relatório não requer configurações adicionais.",
                font=('Arial', 9, 'italic'),
                foreground='blue'
            ).pack(anchor='w')
            
            ttk.Label(
                nota_frame,
                text="   Ao clicar em 'Gerar Relatório', a interface será aberta para seleção do grupo.",
                font=('Arial', 9, 'italic'),
                foreground='blue'
            ).pack(anchor='w')
            
            logger.info("✅ Opções do relatório gerencial configuradas")
            
        except Exception as e:
            logger.error(f"Erro ao configurar opções gerencial: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao configurar opções: {str(e)}")

    def setup_opcoes_gerencial_pdf(self, parent):
            """Configura opções para o relatório gerencial em PDF"""
            try:
                main_frame = ttk.Frame(parent, padding=10)
                main_frame.pack(fill='both', expand=True)
                
                desc_text = """
    Este relatório gera um PDF profissional com layout hierárquico:

    📊 Estrutura do Documento:
    • Resumo executivo do grupo (consolidado)
    • Detalhamento por cliente (nome, CNPJ, CNO, endereço)
    • Contratos (descrição, valores, status, período)
    • Medições (datas, referências, valores, status)

    🎨 Características:
    • Layout profissional similar ao relatório de despesas
    • Hierarquia visual clara (cores e indentação)
    • Tabelas formatadas com bordas e cores
    • Numeração automática de páginas
    • Pronto para impressão e apresentação

    🎯 Ideal para:
    • Apresentações para diretoria
    • Envio a clientes (formato não editável)
    • Documentação formal de projetos
    • Reuniões executivas
                """
                
                ttk.Label(
                    main_frame,
                    text=desc_text,
                    justify='left',
                    wraplength=450,
                    font=('Arial', 10)
                ).pack(pady=10, anchor='w')
                
                ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=15)
                
                info_frame = ttk.LabelFrame(main_frame, text="ℹ️ Informações", padding=10)
                info_frame.pack(fill='x', pady=10)
                
                info_text = """
    ✓ Usa os mesmos dados do relatório gerencial em Excel
    ✓ Formato PDF universal - não requer programas específicos
    ✓ Formato não editável - garante integridade dos dados
    ✓ Arquivo menor que Excel
    ✓ O PDF abre automaticamente após geração
                """
                
                ttk.Label(
                    info_frame,
                    text=info_text,
                    justify='left',
                    wraplength=430,
                    font=('Arial', 9)
                ).pack(anchor='w')
                
                nota_frame = ttk.Frame(main_frame)
                nota_frame.pack(fill='x', pady=(20, 10))
                
                ttk.Label(
                    nota_frame,
                    text="📦 Requer biblioteca: reportlab",
                    font=('Arial', 9, 'bold'),
                    foreground='darkblue'
                ).pack(anchor='w')
                
                ttk.Label(
                    nota_frame,
                    text="   Instalação: pip install reportlab",
                    font=('Arial', 9, 'italic'),
                    foreground='darkblue'
                ).pack(anchor='w')
                
                logger.info("✅ Opções do relatório gerencial PDF configuradas")
                
            except Exception as e:
                logger.error(f"Erro ao configurar opções gerencial PDF: {str(e)}")

    def setup_opcoes_lancamentos_pendentes(self, parent_frame):
        """
        Configura as opções específicas para relatório de lançamentos pendentes
        """
        # Frame para data de referência
        frame_data = ttk.Frame(parent_frame)
        frame_data.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_data, text="Data de Referência:").pack(side='left', padx=5)
        
        try:
            from tkcalendar import DateEntry
            self.data_referencia_pendentes = DateEntry(
                frame_data,
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd/mm/yyyy',
                locale='pt_BR'
            )
            self.data_referencia_pendentes.pack(side='left', padx=5)
        except ImportError:
            ttk.Label(frame_data, text="Módulo tkcalendar não encontrado.").pack(side='left')
        
        # Frame para seleção de pasta
        frame_pasta = ttk.Frame(parent_frame)
        frame_pasta.pack(fill='x', padx=10, pady=10)
        
        # Botão para selecionar pasta
        ttk.Button(
            frame_pasta,
            text="Selecionar Pasta com Arquivos",
            command=self.selecionar_pasta_lancamentos
        ).pack(side='left', padx=5)
        
        # Label para mostrar pasta selecionada
        self.pasta_selecionada_label = ttk.Label(
            frame_pasta, 
            text="Nenhuma pasta selecionada",
            wraplength=400
        )
        self.pasta_selecionada_label.pack(side='left', padx=5)
        
        # Descrição do processo
        ttk.Label(
            parent_frame,
            text="Este relatório processará todos os arquivos Excel \n"
                "na pasta selecionada e gerará um relatório consolidado\n"
                "em HTML com os lançamentos pendentes.",
            justify='center',
            font=('Arial', 10),
            foreground='gray'
        ).pack(pady=20)

    def selecionar_pasta_lancamentos(self):
        """
        Seleciona pasta com arquivos para relatório de lançamentos pendentes
        """
        pasta = filedialog.askdirectory(
            title="Selecione a pasta com os arquivos dos clientes"
        )
        if pasta:
            self.pasta_lancamentos = pasta
            # Verificar se o label existe antes de tentar atualizar
            if hasattr(self, 'pasta_selecionada_label'):
                # Mostrar apenas o nome da pasta, não o caminho completo para melhor visualização
                nome_pasta = os.path.basename(pasta) or pasta
                self.pasta_selecionada_label.config(text=f"Pasta: {nome_pasta}")
            else:
                print(f"Pasta selecionada: {pasta}")  # Fallback caso o label não exista
                messagebox.showinfo("Pasta Selecionada", f"Pasta selecionada: {pasta}")
      
    def carregar_modulo(self, nome_modulo):
        """Carrega ou recarrega um módulo e retorna a classe especificada"""
        try:
            print(f"Tentando carregar módulo: {nome_modulo}")
            # Se o módulo já foi carregado, recarregá-lo
            if nome_modulo in sys.modules:
                print(f"Recarregando módulo existente: {nome_modulo}")
                modulo = importlib.reload(sys.modules[nome_modulo])
            else:
                # Tentar importar do caminho atual
                try:
                    print(f"Tentando importar direto: {nome_modulo}")
                    modulo = importlib.import_module(nome_modulo)
                except ImportError as e1:
                    print(f"Erro importando direto: {str(e1)}")
                    # Tentar importar de src
                    try:
                        print(f"Tentando importar de src: src.{nome_modulo}")
                        modulo = importlib.import_module(f"src.{nome_modulo}")
                    except ImportError as e2:
                        print(f"Erro importando de src: {str(e2)}")
                        raise ImportError(f"Não foi possível importar {nome_modulo}: {str(e1)}, {str(e2)}")
            
            # Armazenar módulo carregado
            self.modulos_carregados[nome_modulo] = modulo
            print(f"Módulo carregado com sucesso: {nome_modulo}")
            return modulo
            
        except Exception as e:
            print(f"Erro ao carregar módulo {nome_modulo}: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror(
                "Erro ao carregar módulo", 
                f"Não foi possível carregar o módulo {nome_modulo}.\nErro: {str(e)}"
            )
            return None
    
    def limpar_arquivo_temporario(self, caminho_temp):
        """Remove arquivo temporário após uso"""
        try:
            if os.path.exists(caminho_temp):
                os.remove(caminho_temp)
                logger.info(f"🗑️ Arquivo temporário removido: {os.path.basename(caminho_temp)}")
        except Exception as e:
            logger.warning(f"Aviso: Não foi possível remover arquivo temporário: {str(e)}")

    def executar_relatorio_direto(self, configuracoes):
        """Executa relatório direto sem preview"""
        try:
            logger.info("🎯 EXECUTANDO RELATÓRIO DIRETO")
            
            # Processar dados
            handler = self.obter_handler_despesas_limpo()
            dados_processados = self.processar_dados_completo(handler, configuracoes)
            
            # Gerar nome do arquivo
            data_formatada = configuracoes['data'].strftime('%d-%m-%Y')
            nome_arquivo = f"REL - {dados_processados['nome_cliente']} - {data_formatada}.pdf"
            
            if configuracoes['incluir_excluidos']:
                nome_arquivo = nome_arquivo.replace('.pdf', ' (com excluídos).pdf')
            
            caminho_output = os.path.join(os.path.dirname(configuracoes['arquivo']), nome_arquivo)
            
            # Gerar PDF
            handler.gerar_relatorio_pdf(dados_processados, caminho_output, configuracoes['arquivo'])
            
            # Mostrar resultado
            resposta = messagebox.askyesno(
                "Relatório Gerado!",
                f"Relatório gerado com sucesso!\n\n"
                f"Cliente: {dados_processados['nome_cliente']}\n"
                f"Arquivo: {nome_arquivo}\n\n"
                f"Deseja abrir o PDF?"
            )
            
            if resposta:
                self.abrir_arquivo(caminho_output)
            
            logger.info("✅ Relatório direto concluído")
            
        except Exception as e:
            logger.error(f"💥 ERRO no relatório direto: {str(e)}")
            messagebox.showerror("Erro", f"Erro: {str(e)}")

    def criar_janela_progresso_simples(self):
        """Cria janela de progresso simples"""
        try:
            window = tk.Toplevel(self.root)
            window.title("Processando...")
            window.geometry("400x120")
            window.transient(self.root)
            window.grab_set()
            window.resizable(False, False)
            
            # Centralizar
            window.update_idletasks()
            x = (window.winfo_screenwidth() // 2) - 200
            y = (window.winfo_screenheight() // 2) - 60
            window.geometry(f"400x120+{x}+{y}")
            
            # Widgets
            frame = ttk.Frame(window, padding=20)
            frame.pack(fill='both', expand=True)
            
            ttk.Label(frame, text="Processando Relatório...", font=('Arial', 12)).pack(pady=10)
            
            window.status_label = ttk.Label(frame, text="Iniciando...")
            window.status_label.pack(pady=5)
            
            window.progress_bar = ttk.Progressbar(frame, length=300, mode='determinate')
            window.progress_bar.pack(pady=10)
            
            return window
            
        except Exception as e:
            logger.error(f"💥 ERRO ao criar progresso: {str(e)}")
            return None

    def formatar_numero(self, valor):
        """Formata número de forma segura"""
        try:
            import pandas as pd
            
            if valor is None or pd.isna(valor):
                return "0,00"
            
            # Converter para float se for string
            if isinstance(valor, str):
                valor = valor.replace('R$', '').replace(' ', '').replace(',', '.')
                valor = float(valor)
            
            # Formatar no padrão brasileiro
            return f"{float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return "0,00"            

    def limpar_data(self, data_input):
        """Limpa e normaliza data de forma definitiva"""
        try:
            from datetime import datetime, date
            
            logger.info(f"🔧 Limpando data: {data_input} (tipo: {type(data_input)})")
            
            # Converter conforme tipo
            if isinstance(data_input, str):
                data_limpa = datetime.strptime(data_input, '%d/%m/%Y')
            elif isinstance(data_input, datetime):
                data_limpa = data_input
            elif isinstance(data_input, date):
                data_limpa = datetime.combine(data_input, datetime.min.time())
            else:
                logger.warning(f"⚠️ Tipo de data não reconhecido: {type(data_input)}")
                data_limpa = datetime.now()
            
            # Normalizar para início do dia
            data_final = data_limpa.replace(hour=0, minute=0, second=0, microsecond=0)
            
            logger.info(f"✅ Data limpa: {data_final}")
            return data_final
            
        except Exception as e:
            logger.error(f"💥 ERRO ao limpar data: {str(e)}")
            return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    def criar_progress_window(self):
        """Cria janela de progresso simples"""
        try:
            window = tk.Toplevel(self.root)
            window.title("Processando...")
            window.geometry("400x150")
            window.transient(self.root)
            window.grab_set()
            
            # Centralizar
            window.update_idletasks()
            x = (window.winfo_screenwidth() // 2) - 200
            y = (window.winfo_screenheight() // 2) - 75
            window.geometry(f"400x150+{x}+{y}")
            
            # Widgets
            frame = ttk.Frame(window, padding=20)
            frame.pack(fill='both', expand=True)
            
            ttk.Label(frame, text="Processando Relatório...", font=('Arial', 12, 'bold')).pack(pady=10)
            
            window.status_label = ttk.Label(frame, text="Iniciando...")
            window.status_label.pack(pady=5)
            
            window.progress_bar = ttk.Progressbar(frame, length=300, mode='determinate')
            window.progress_bar.pack(pady=10)
            
            return window
            
        except Exception as e:
            logger.error(f"💥 ERRO ao criar progress: {str(e)}")
            return None

    def processar_lancamentos_pendentes(self):
        """Processa lançamentos pendentes - mantém original"""
        try:
            if not hasattr(self, 'pasta_lancamentos'):
                messagebox.showerror("Erro", "Selecione uma pasta primeiro.")
                return
            
            data_ref = self.data_referencia_pendentes.get_date() if hasattr(self, 'data_referencia_pendentes') else datetime.now()
            
            from src.relatorio_despesas_aprimorado import RelatorioLancamentosPendentes
            relatorio = RelatorioLancamentosPendentes()
            arquivo_saida = os.path.join(self.pasta_lancamentos, "relatorio_lancamentos_pendentes.html")
            
            if relatorio.gerar_relatorio_pendentes(self.pasta_lancamentos, arquivo_saida, data_ref):
                messagebox.showinfo("Sucesso", f"Relatório gerado: {arquivo_saida}")
            else:
                messagebox.showwarning("Aviso", "Nenhum lançamento pendente encontrado.")
                
        except Exception as e:
            logger.error(f"💥 ERRO lançamentos pendentes: {str(e)}")
            messagebox.showerror("Erro", f"Erro: {str(e)}")

    def processar_fornecedores(self):
        """Processa fornecedores - mantém original simplificado"""
        try:
            self.root.withdraw()
            
            from src.relatorio_fornecedores import RelatorioFornecedores
            app = RelatorioFornecedores(parent=self.root)
            app.menu_principal = self.root
            
            app.root.protocol("WM_DELETE_WINDOW", lambda: self.finalizar_sistema(app.root))
            app.root.lift()
            app.root.focus_force()
            app.root.mainloop()
            
        except Exception as e:
            logger.error(f"💥 ERRO fornecedores: {str(e)}")
            messagebox.showerror("Erro", f"Erro: {str(e)}")
            self.root.deiconify()

    def setup_opcoes_consistencia_dados(self, parent_frame):
        """Configura as opções para o relatório de consistência de dados"""
        frame_cliente = ttk.Frame(parent_frame)
        frame_cliente.pack(fill='x', padx=10, pady=10)

        ttk.Label(frame_cliente, text="Cliente:").pack(side='left', padx=5)

        self.cliente_consistencia = ttk.Combobox(frame_cliente, width=40)
        self.cliente_consistencia.pack(side='left', padx=5)

        self.preencher_combobox_clientes(self.cliente_consistencia)

        ttk.Label(
            parent_frame,
            text=(
                "Verifica registros em 'Dados' cujo CNPJ possui contrato ativo,\n"
                "mas sem correspondência nas abas Medições ou Contratos ADM, e vice-versa."
            ),
            foreground='#555',
            wraplength=420
        ).pack(anchor='w', padx=10, pady=(0, 10))

    def processar_consistencia_dados(self):
        """Abre o relatório de consistência entre Dados e Medições/Contratos ADM"""
        try:
            cliente_sel = getattr(self, 'cliente_consistencia', None)
            cliente_nome = cliente_sel.get() if cliente_sel else None
            if cliente_nome == 'Todos os Clientes':
                cliente_nome = None

            self.root.withdraw()

            from src.relatorio_consistencia_dados import RelatorioConsistenciaDados
            RelatorioConsistenciaDados(parent=self.root, cliente_inicial=cliente_nome)

        except Exception as e:
            logger.error(f"Erro ao abrir consistência de dados: {str(e)}", exc_info=True)
            messagebox.showerror("Erro", f"Erro ao abrir relatório:\n{str(e)}")
            self.root.deiconify()

    def processar_outros_relatorios(self, relatorio):
        """Processa outros relatórios - mantém original"""
        try:
            modulo = self.carregar_modulo(relatorio["modulo"])
            if not modulo:
                return
            
            classe_relatorio = getattr(modulo, relatorio["classe"])
            
            if relatorio["id"] == "contratos":
                self.iniciar_relatorio_contratos(classe_relatorio)
            elif relatorio["id"] == "categoria":
                self.iniciar_relatorio_categoria(classe_relatorio)
            elif relatorio["id"] == "tipo_despesa":
                self.iniciar_relatorio_tipo_despesa(classe_relatorio)
            else:
                messagebox.showinfo("Em desenvolvimento", "Em desenvolvimento.")
                
        except Exception as e:
            logger.error(f"💥 ERRO outros relatórios: {str(e)}")
            messagebox.showerror("Erro", f"Erro: {str(e)}")

    def processar_gerencial_engenheiro(self):
        """
        Processa o relatório gerencial por engenheiro
        
        Este relatório abre uma janela independente que permite:
        1. Selecionar o grupo (1-4)
        2. Visualizar dados consolidados
        3. Gerar gráficos
        4. Exportar para Excel
        """
        try:
            logger.info("🚀 Iniciando Relatório Gerencial por Engenheiro")
            
            # Importar o módulo
            try:
                from relatorio_gerencial_engenheiro import RelatorioGerencialEngenheiro
                logger.info("✅ Módulo RelatorioGerencialEngenheiro importado")
            except ImportError as e:
                logger.error(f"❌ Erro ao importar módulo: {str(e)}")
                messagebox.showerror(
                    "Erro de Importação",
                    f"Não foi possível carregar o módulo do relatório gerencial.\n\n"
                    f"Erro: {str(e)}\n\n"
                    f"Certifique-se de que o arquivo 'relatorio_gerencial_engenheiro.py' "
                    f"está na pasta correta e que todas as dependências estão instaladas."
                )
                return
            
            # Ocultar janela principal temporariamente
            self.root.withdraw()
            
            # Criar e abrir o relatório gerencial
            try:
                relatorio = RelatorioGerencialEngenheiro(parent=self.root)
                logger.info("✅ Interface do relatório gerencial aberta")
                
                # Aguardar fechamento da janela do relatório
                self.root.wait_window(relatorio.root)
                
            except Exception as e:
                logger.error(f"❌ Erro ao criar relatório gerencial: {str(e)}")
                messagebox.showerror(
                    "Erro",
                    f"Erro ao criar relatório gerencial:\n\n{str(e)}"
                )
            
            finally:
                # Restaurar janela principal
                self.root.deiconify()
                self.root.lift()
                self.root.focus_force()
                logger.info("✅ Retornado à interface principal")
                
        except Exception as e:
            logger.error(f"💥 ERRO no processamento gerencial: {str(e)}", exc_info=True)
            messagebox.showerror("Erro", f"Erro ao processar relatório: {str(e)}")
            
            # Garantir que janela principal volte
            try:
                self.root.deiconify()
            except:
                pass

    def processar_gerencial_pdf(self):
            """Processa o relatório gerencial em PDF"""
            try:
                logger.info("🚀 Iniciando Relatório Gerencial PDF")
                
                try:
                    from relatorio_gerencial_pdf import RelatorioGerencialPDF
                    logger.info("✅ Módulo RelatorioGerencialPDF importado")
                except ImportError as e:
                    logger.error(f"❌ Erro ao importar módulo: {str(e)}")
                    messagebox.showerror(
                        "Erro de Importação",
                        f"Não foi possível carregar o módulo do relatório gerencial PDF.\n\n"
                        f"Erro: {str(e)}\n\n"
                        f"Certifique-se de que:\n"
                        f"1. O arquivo 'relatorio_gerencial_pdf.py' está na pasta correta\n"
                        f"2. A biblioteca 'reportlab' está instalada: pip install reportlab"
                    )
                    return
                
                self.root.withdraw()
                
                try:
                    relatorio = RelatorioGerencialPDF(parent=self.root)
                    logger.info("✅ Interface do relatório gerencial PDF aberta")
                    self.root.wait_window(relatorio.root)
                    
                except Exception as e:
                    logger.error(f"❌ Erro ao criar relatório gerencial PDF: {str(e)}")
                    messagebox.showerror(
                        "Erro",
                        f"Erro ao criar relatório gerencial PDF:\n\n{str(e)}"
                    )
                
                finally:
                    self.root.deiconify()
                    self.root.lift()
                    self.root.focus_force()
                    logger.info("✅ Retornado à interface principal")
                    
            except Exception as e:
                logger.error(f"💥 ERRO no processamento gerencial PDF: {str(e)}", exc_info=True)
                messagebox.showerror("Erro", f"Erro ao processar relatório: {str(e)}")
                try:
                    self.root.deiconify()
                except:
                    pass

    def atualizar_progresso_seguro(self, progress_window, mensagem, porcentagem):
        """Atualiza progresso de forma segura"""
        try:
            if progress_window and hasattr(progress_window, 'winfo_exists'):
                if progress_window.winfo_exists():
                    progress_window.status_label.config(text=mensagem)
                    progress_window.progress_bar['value'] = porcentagem
                    progress_window.percent_label.config(text=f"{porcentagem}%")
                    progress_window.update()
                    
                    # Log do progresso
                    logger.info(f"Progresso: {porcentagem}% - {mensagem}")
                    
        except Exception as e:
            logger.error(f"Erro ao atualizar progresso: {str(e)}")

    def abrir_arquivo(self, caminho):
        """Abre arquivo com programa padrão do sistema"""
        try:
            import platform
            import subprocess
            
            if platform.system() == 'Darwin':       # macOS
                subprocess.run(['open', caminho])
            elif platform.system() == 'Windows':    # Windows
                os.startfile(caminho)
            else:                                   # Linux
                subprocess.run(['xdg-open', caminho])
                
        except Exception as e:
            logger.error(f"Erro ao abrir arquivo: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao abrir arquivo: {str(e)}")
    
    def coletar_configuracoes_completas(self):
        """Versão corrigida que GARANTE que arquivo esteja nas configurações"""
        config = {
            'data': datetime.now(),
            'incluir_futuros': True,
            'incluir_excluidos': False,
            'incluir_notas': False,
            'texto_notas': False,
            'arquivo': None,  # IMPORTANTE: Inicializar como None
            'tipo_geracao': 'individual',
            'arquivos_lote': [],
            'formato_saida': 'pdf',
            'cliente_selecionado': None,
            'data_automatica': True
        }
        
        try:
            # Data - usar o método que considera automático/manual
            config['data'] = self.obter_data_relatorio_final()
            config['data_automatica'] = self.usar_data_automatica.get() if hasattr(self, 'usar_data_automatica') else True
        except Exception as e:
            logger.debug(f"Erro ao coletar data: {str(e)}")
            config['data'] = self.calcular_data_rel_automatica()
        
        try:
            # Flags
            if hasattr(self, 'incluir_futuros'):
                config['incluir_futuros'] = self.incluir_futuros.get()
        except Exception as e:
            logger.debug(f"Erro ao coletar incluir_futuros: {str(e)}")
        
        try:
            if hasattr(self, 'incluir_excluidos'):
                config['incluir_excluidos'] = self.incluir_excluidos.get()
        except Exception as e:
            logger.debug(f"Erro ao coletar incluir_excluidos: {str(e)}")

        try:
            if hasattr(self, 'incluir_notas'):
                config['incluir_notas'] = self.incluir_notas.get()
        except Exception as e:
            logger.debug(f"Erro ao coletar incluir_notas: {str(e)}")

        try:
            if hasattr(self, 'texto_notas'):
                config['texto_notas'] = self.texto_notas.get()
        except Exception as e:
            logger.debug(f"Erro ao coletar texto_notas: {str(e)}")
        
        try:
            # CORREÇÃO: Arquivo individual - verificar múltiplas fontes
            if hasattr(self, 'arquivo_cliente_selecionado') and self.arquivo_cliente_selecionado:
                config['arquivo'] = self.arquivo_cliente_selecionado
                logger.info(f"✅ Arquivo encontrado em arquivo_cliente_selecionado: {config['arquivo']}")
            elif hasattr(self, 'arquivo_path') and self.arquivo_path:
                config['arquivo'] = self.arquivo_path
                logger.info(f"✅ Arquivo encontrado em arquivo_path: {config['arquivo']}")
            else:
                logger.warning("❌ Arquivo não encontrado em nenhuma variável")
                
        except Exception as e:
            logger.error(f"Erro ao coletar arquivo: {str(e)}")
        
        try:
            # Tipo de geração
            if hasattr(self, 'tipo_geracao'):
                config['tipo_geracao'] = self.tipo_geracao.get()
        except Exception as e:
            logger.debug(f"Erro ao coletar tipo_geracao: {str(e)}")
        
        try:
            # Arquivos em lote
            if hasattr(self, 'arquivos_lote') and self.arquivos_lote:
                config['arquivos_lote'] = self.arquivos_lote
        except Exception as e:
            logger.debug(f"Erro ao coletar arquivos_lote: {str(e)}")
        
        try:
            # Formato de saída
            if hasattr(self, 'formato_saida'):
                config['formato_saida'] = self.formato_saida.get()
        except Exception as e:
            logger.debug(f"Erro ao coletar formato_saida: {str(e)}")
        
        try:
            # Cliente selecionado
            if hasattr(self, 'cliente_combobox'):
                config['cliente_selecionado'] = self.cliente_combobox.get()
        except Exception as e:
            logger.debug(f"Erro ao coletar cliente: {str(e)}")
        
        # CORREÇÃO: Verificação final crítica
        if not config['arquivo']:
            logger.error("❌ ERRO CRÍTICO: Nenhum arquivo foi encontrado nas configurações!")
            logger.error("Variáveis verificadas:")
            logger.error(f"  - hasattr(self, 'arquivo_cliente_selecionado'): {hasattr(self, 'arquivo_cliente_selecionado')}")
            if hasattr(self, 'arquivo_cliente_selecionado'):
                logger.error(f"  - self.arquivo_cliente_selecionado: {self.arquivo_cliente_selecionado}")
            logger.error(f"  - hasattr(self, 'arquivo_path'): {hasattr(self, 'arquivo_path')}")
            if hasattr(self, 'arquivo_path'):
                logger.error(f"  - self.arquivo_path: {self.arquivo_path}")
        else:
            logger.info(f"✅ Arquivo final nas configurações: {config['arquivo']}")
        
        return config

    def gerar_resumo_configuracoes(self, config):
        """Gera resumo legível das configurações"""
        try:
            resumo_lines = []
            
            # Data
            resumo_lines.append(f"📅 Data: {config['data'].strftime('%d/%m/%Y')}")
            
            # Arquivo/Cliente
            if config['arquivo']:
                nome_arquivo = os.path.basename(config['arquivo'])
                resumo_lines.append(f"📁 Arquivo: {nome_arquivo}")
            
            if config['cliente_selecionado'] and "Arquivo:" not in config['cliente_selecionado']:
                resumo_lines.append(f"👤 Cliente: {config['cliente_selecionado']}")
            
            # Tipo de geração
            tipo_texto = "Individual" if config['tipo_geracao'] == 'individual' else "Lote"
            resumo_lines.append(f"🔄 Tipo: {tipo_texto}")
            
            if config['arquivos_lote']:
                resumo_lines.append(f"📂 Arquivos em lote: {len(config['arquivos_lote'])} arquivos")
            
            # Opções
            opcoes = []
            if config['incluir_futuros']:
                opcoes.append("Lançamentos futuros")
            if config['incluir_excluidos']:
                opcoes.append("Lançamentos excluídos")
            
            if opcoes:
                resumo_lines.append(f"⚙️ Incluir: {', '.join(opcoes)}")
            
            # Formato
            resumo_lines.append(f"📄 Formato: {config['formato_saida'].upper()}")
            
            return "\n".join(resumo_lines)
            
        except Exception as e:
            logger.error(f"Erro ao gerar resumo: {str(e)}")
            return "Erro ao gerar resumo das configurações"

    def validar_configuracoes_despesas(self):
        """Versão atualizada da validação que considera a nova seleção"""
        try:
            # Verificar data
            if hasattr(self, 'data_entry'):
                try:
                    data = self.data_entry.get_date()
                    if not data:
                        messagebox.showerror("Erro", "Por favor, selecione uma data válida.")
                        return False
                except Exception:
                    messagebox.showerror("Erro", "Data selecionada é inválida.")
                    return False
            
            # Verificar tipo de geração
            if hasattr(self, 'tipo_geracao'):
                tipo = self.tipo_geracao.get()
                
                if tipo == "individual":
                    # Verificar se há cliente/arquivo selecionado
                    if not hasattr(self, 'arquivo_cliente_selecionado') or not self.arquivo_cliente_selecionado:
                        messagebox.showerror(
                            "Erro", 
                            "Por favor, selecione um cliente na combobox ou use a seleção manual de arquivo."
                        )
                        return False
                        
                    # Verificar se o arquivo existe
                    if not os.path.exists(self.arquivo_cliente_selecionado):
                        messagebox.showerror(
                            "Erro", 
                            "O arquivo do cliente selecionado não existe ou não pode ser acessado.\n"
                            "Tente atualizar a lista de clientes ou selecionar manualmente."
                        )
                        return False
                        
                elif tipo == "lote":
                    # Verificar se há arquivos selecionados para lote
                    if not hasattr(self, 'arquivos_lote') or not self.arquivos_lote:
                        messagebox.showerror(
                            "Erro", 
                            "Por favor, selecione arquivos para processamento em lote."
                        )
                        return False
                        
                    # Verificar se todos os arquivos existem
                    arquivos_inexistentes = []
                    for arquivo in self.arquivos_lote:
                        if not os.path.exists(arquivo):
                            arquivos_inexistentes.append(os.path.basename(arquivo))
                    
                    if arquivos_inexistentes:
                        messagebox.showerror(
                            "Erro", 
                            f"Os seguintes arquivos não existem:\n" + 
                            "\n".join(arquivos_inexistentes)
                        )
                        return False
            
            return True
            
        except Exception as e:
            logger.error(f"Erro ao validar configurações: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao validar configurações: {str(e)}")
            return False

    def mostrar_resumo_configuracoes(self):
        """Mostra um resumo das configurações antes de gerar o relatório"""
        try:
            resumo = []
            
            # Data
            if hasattr(self, 'data_entry'):
                try:
                    data = self.data_entry.get_date().strftime('%d/%m/%Y')
                    resumo.append(f"Data do relatório: {data}")
                except:
                    resumo.append("Data do relatório: Data atual")
            
            # Lançamentos futuros
            if hasattr(self, 'incluir_futuros'):
                status = "Sim" if self.incluir_futuros.get() else "Não"
                resumo.append(f"Incluir lançamentos futuros: {status}")
            
            # Lançamentos excluídos
            if hasattr(self, 'incluir_excluidos'):
                status = "Sim" if self.incluir_excluidos.get() else "Não"
                resumo.append(f"Incluir lançamentos excluídos: {status}")
            
            # Tipo de geração
            if hasattr(self, 'tipo_geracao'):
                tipo = self.tipo_geracao.get()
                resumo.append(f"Tipo de geração: {tipo.title()}")
                
                if tipo == "individual" and hasattr(self, 'arquivo_cliente_selecionado'):
                    nome_arquivo = os.path.basename(self.arquivo_cliente_selecionado)
                    resumo.append(f"Arquivo: {nome_arquivo}")
                elif tipo == "lote" and hasattr(self, 'arquivos_lote'):
                    resumo.append(f"Arquivos em lote: {len(self.arquivos_lote)} arquivos")
            
            # Modo de visualização
            if hasattr(self, 'modo_visualizacao'):
                modo = self.modo_visualizacao.get()
                modo_texto = "Com Preview" if modo == "preview" else "Direto"
                resumo.append(f"Modo de visualização: {modo_texto}")
            
            # Formato de saída
            if hasattr(self, 'formato_saida'):
                formato = self.formato_saida.get().upper()
                resumo.append(f"Formato de saída: {formato}")
            
            return "\n".join(resumo)
            
        except Exception as e:
            logger.error(f"Erro ao gerar resumo: {str(e)}")
            return "Erro ao gerar resumo das configurações"

    def confirmar_geracao_relatorio(self):
        """Confirma a geração do relatório mostrando um resumo"""
        try:
            resumo = self.mostrar_resumo_configuracoes()
            
            resposta = messagebox.askyesno(
                "Confirmar Geração",
                f"Confirma a geração do relatório com as seguintes configurações?\n\n{resumo}",
                icon='question'
            )
            
            return resposta
            
        except Exception as e:
            logger.error(f"Erro ao confirmar geração: {str(e)}")
            return True  # Em caso de erro, prosseguir

    def atualizar_botao_geracao(self):
        """Atualiza o texto e estado do botão de geração conforme as configurações"""
        try:
            # Este método pode ser chamado quando há mudanças nas configurações
            # para atualizar dinamicamente a interface
            
            if hasattr(self, 'tipo_geracao'):
                tipo = self.tipo_geracao.get()
                
                if tipo == "individual":
                    texto_botao = "Gerar Relatório Individual"
                else:
                    texto_botao = "Gerar Relatórios em Lote"
                    
                # Se houver um botão específico, atualizar seu texto
                # (Este código pode ser adaptado conforme a estrutura real da interface)
                
            logger.debug(f"Botão atualizado para: {texto_botao}")
            
        except Exception as e:
            logger.debug(f"Erro ao atualizar botão: {str(e)}")

    def limpar_selecoes(self):
        """Limpa as seleções de arquivos"""
        try:
            if hasattr(self, 'arquivo_cliente_selecionado'):
                delattr(self, 'arquivo_cliente_selecionado')
                
            if hasattr(self, 'arquivos_lote'):
                self.arquivos_lote = []
                
            if hasattr(self, 'lbl_arquivos_lote'):
                self.lbl_arquivos_lote.config(text="")
                
            if hasattr(self, 'cliente_combobox'):
                self.cliente_combobox.set("Todos os Clientes")
                
            logger.info("Seleções de arquivos limpas")
            
        except Exception as e:
            logger.error(f"Erro ao limpar seleções: {str(e)}")

    def resetar_configuracoes_despesas(self):
        """Reseta todas as configurações para valores padrão"""
        try:
            # Data atual
            if hasattr(self, 'data_entry'):
                from datetime import datetime
                self.data_entry.set_date(datetime.now())
            
            # Lançamentos futuros: True
            if hasattr(self, 'incluir_futuros'):
                self.incluir_futuros.set(True)
                
            # Lançamentos excluídos: False
            if hasattr(self, 'incluir_excluidos'):
                self.incluir_excluidos.set(False)
                
            # Tipo individual
            if hasattr(self, 'tipo_geracao'):
                self.tipo_geracao.set("individual")
                self.alternar_tipo_geracao()
                
            # Modo preview
            if hasattr(self, 'modo_visualizacao'):
                self.modo_visualizacao.set("preview")
                
            # Formato PDF
            if hasattr(self, 'formato_saida'):
                self.formato_saida.set("pdf")
                
            # Limpar seleções
            self.limpar_selecoes()
            
            logger.info("Configurações resetadas para padrão")
            
        except Exception as e:
            logger.error(f"Erro ao resetar configurações: {str(e)}")

    def adicionar_botoes_auxiliares(self, parent_frame):
        """Adiciona botões auxiliares para gerenciar configurações"""
        try:
            # Frame para botões auxiliares
            frame_botoes = ttk.LabelFrame(parent_frame, text="Ações Auxiliares")
            frame_botoes.pack(fill='x', padx=10, pady=10)
            
            # Botão para limpar seleções
            ttk.Button(
                frame_botoes,
                text="Limpar Seleções",
                command=self.limpar_selecoes
            ).pack(side='left', padx=5, pady=5)
            
            # Botão para resetar configurações
            ttk.Button(
                frame_botoes,
                text="Resetar Configurações",
                command=self.resetar_configuracoes_despesas
            ).pack(side='left', padx=5, pady=5)
            
            # Botão para mostrar resumo
            ttk.Button(
                frame_botoes,
                text="Ver Resumo",
                command=lambda: messagebox.showinfo(
                    "Resumo das Configurações", 
                    self.mostrar_resumo_configuracoes()
                )
            ).pack(side='left', padx=5, pady=5)
            
            logger.debug("Botões auxiliares adicionados")
            
        except Exception as e:
            logger.error(f"Erro ao adicionar botões auxiliares: {str(e)}")

    def mostrar_opcoes_relatorio_com_validacao(self, event=None):
        """Versão melhorada do mostrar_opcoes_relatorio com validação"""
        try:
            # Chamar o método original
            self.mostrar_opcoes_relatorio_original(event)
            
            # Adicionar validações específicas após mostrar as opções
            selecao = self.tree_relatorios.selection()
            if selecao:
                rel_id = selecao[0]
                
                if rel_id == "despesas":
                    # Adicionar botões auxiliares para relatório de despesas
                    if hasattr(self, 'right_frame'):
                        # Verificar se já foi adicionado
                        botoes_existem = any(
                            isinstance(widget, ttk.LabelFrame) and 
                            "Ações Auxiliares" in str(widget.cget('text', ''))
                            for widget in self.right_frame.winfo_children()
                        )
                        
                        if not botoes_existem:
                            self.adicionar_botoes_auxiliares(self.right_frame)
                            
        except Exception as e:
            logger.error(f"Erro na validação de opções: {str(e)}")

    def backup_metodo_original(self):
        """Cria backup do método original se necessário"""
        if not hasattr(self, 'mostrar_opcoes_relatorio_original'):
            self.mostrar_opcoes_relatorio_original = self.mostrar_opcoes_relatorio
            self.mostrar_opcoes_relatorio = self.mostrar_opcoes_relatorio_com_validacao
 
    def iniciar_relatorio_contratos(self, classe_relatorio):
        """Inicia a geração do relatório de contratos e medições"""
        # Esconder a janela atual
        self.root.withdraw()
        
        # Inicializar o relatório passando a janela atual como parent
        app_relatorio = classe_relatorio(self.root)
        
        # Verificar se app_relatorio tem os atributos esperados
        if not hasattr(app_relatorio, 'root'):
            messagebox.showerror(
                "Erro", 
                "Erro ao inicializar relatório. A classe do relatório não retornou o objeto esperado."
            )
            self.root.deiconify()
            return
        
        # Configurar menu principal para retornar
        app_relatorio.menu_principal = self.root
        
        # Se houver cliente selecionado e o método adequado existir, selecioná-lo
        if hasattr(app_relatorio, 'cliente_combobox') and self.cliente_contratos.get() != 'Todos os Clientes':
            try:
                app_relatorio.cliente_combobox.set(self.cliente_contratos.get())
                # Se existir um método específico para selecionar o cliente, chamá-lo
                if hasattr(app_relatorio, 'selecionar_cliente'):
                    app_relatorio.selecionar_cliente()
            except Exception as e:
                logger.warning(f"Não foi possível selecionar o cliente: {str(e)}")
        
        # Se houver data selecionada, configurá-la
        if hasattr(self, 'data_referencia') and hasattr(app_relatorio, 'data_entry'):
            try:
                app_relatorio.data_entry.set_date(self.data_referencia.get_date())
            except Exception as e:
                logger.warning(f"Não foi possível configurar a data: {str(e)}")
        
        # Configurar comportamento ao fechar
        app_relatorio.root.protocol("WM_DELETE_WINDOW", lambda: self.finalizar_sistema(app_relatorio.root))
        
        # Exibir janela
        app_relatorio.root.lift()
        app_relatorio.root.focus_force()
        app_relatorio.root.mainloop()

    def iniciar_relatorio_categoria(self, classe_relatorio):
        """Inicia a geração do relatório por tipo de despesa"""
        # Esconder a janela atual
        self.root.withdraw()
        
        # Inicializar o relatório passando a janela atual como parent
        app_relatorio = classe_relatorio(self.root)
        
        # Verificar se app_relatorio tem os atributos esperados
        if not hasattr(app_relatorio, 'root'):
            messagebox.showerror(
                "Erro", 
                "Erro ao inicializar relatório. A classe do relatório não retornou o objeto esperado."
            )
            self.root.deiconify()
            return
        
        # Configurar menu principal para retornar
        app_relatorio.menu_principal = self.root
        
        # Se houver cliente selecionado, configurá-lo
        if hasattr(app_relatorio, 'cliente_combobox') and hasattr(self, 'cliente_categoria'):
            cliente_selecionado = self.cliente_categoria.get()
            if cliente_selecionado and cliente_selecionado != 'Todos os Clientes':
                try:
                    # Atualizar lista de clientes primeiro
                    app_relatorio.atualizar_lista_clientes()
                    
                    # Configurar o cliente no combobox
                    app_relatorio.cliente_combobox.set(cliente_selecionado)
                    
                    # Chamar o método para selecionar cliente
                    if hasattr(app_relatorio, 'selecionar_cliente'):
                        app_relatorio.selecionar_cliente()
                except Exception as e:
                    logger.warning(f"Não foi possível selecionar o cliente: {str(e)}")
        
        # Configurar comportamento ao fechar
        app_relatorio.root.protocol("WM_DELETE_WINDOW", lambda: self.finalizar_sistema(app_relatorio.root))
        
        # Exibir janela
        app_relatorio.root.lift()
        app_relatorio.root.focus_force()
        app_relatorio.root.mainloop()

    def iniciar_relatorio_tipo_despesa(self, classe_relatorio):
        """Inicia a geração do relatório por tipo de despesa"""
        # Esconder a janela atual
        self.root.withdraw()
        
        # Inicializar o relatório passando a janela atual como parent
        app_relatorio = classe_relatorio(self.root)
        
        # Verificar se app_relatorio tem os atributos esperados
        if not hasattr(app_relatorio, 'root'):
            messagebox.showerror(
                "Erro", 
                "Erro ao inicializar relatório. A classe do relatório não retornou o objeto esperado."
            )
            self.root.deiconify()
            return
        
        # Configurar menu principal para retornar
        app_relatorio.menu_principal = self.root
        
        # Se houver cliente selecionado, configurá-lo
        if hasattr(app_relatorio, 'cliente_combobox') and hasattr(self, 'cliente_tipo_despesa'):
            cliente_selecionado = self.cliente_tipo_despesa.get()
            if cliente_selecionado and cliente_selecionado != 'Todos os Clientes':
                try:
                    # Atualizar lista de clientes primeiro
                    app_relatorio.atualizar_lista_clientes()
                    
                    # Configurar o cliente no combobox
                    app_relatorio.cliente_combobox.set(cliente_selecionado)
                    
                    # Chamar o método para selecionar cliente
                    if hasattr(app_relatorio, 'selecionar_cliente'):
                        app_relatorio.selecionar_cliente()
                except Exception as e:
                    logger.warning(f"Não foi possível selecionar o cliente: {str(e)}")
        
        # Configurar comportamento ao fechar
        app_relatorio.root.protocol("WM_DELETE_WINDOW", lambda: self.finalizar_sistema(app_relatorio.root))
        
        # Exibir janela
        app_relatorio.root.lift()
        app_relatorio.root.focus_force()
        app_relatorio.root.mainloop()

    def iniciar_relatorio_fornecedores(self, classe_relatorio):
        """Inicia a geração do relatório de fornecedores"""
        try:
            print("Iniciando método iniciar_relatorio_fornecedores")
            # Esconder a janela atual
            self.root.withdraw()
            
            # Criar uma nova janela para o relatório
            print("Criando instância do relatório de fornecedores")
            app_relatorio = classe_relatorio(self.root)
            
            # Configurar menu principal para retornar
            app_relatorio.menu_principal = self.root
            
            # IMPORTANTE: Verificar se há um cliente selecionado e configurá-lo
            if hasattr(self, 'cliente_contratos') and self.cliente_contratos:
                cliente_selecionado = self.cliente_contratos.get()
                if cliente_selecionado and cliente_selecionado != 'Todos os Clientes':
                    try:
                        # Configurar o cliente no relatório de fornecedores
                        print(f"Configurando cliente: {cliente_selecionado}")
                        app_relatorio.cliente_combobox.set(cliente_selecionado)
                        
                        # Chamar o método selecionar_cliente diretamente
                        app_relatorio.cliente_atual = cliente_selecionado
                        app_relatorio.arquivo_cliente = PASTA_CLIENTES / f"{cliente_selecionado}.xlsx"
                        app_relatorio.lbl_cliente_resumo.config(text=f"Cliente: {cliente_selecionado}")
                        
                        # Desmarcar checkbox de todos os clientes
                        app_relatorio.var_todos_clientes.set(False)
                        app_relatorio.todos_clientes = False
                        
                    except Exception as e:
                        print(f"Erro ao configurar cliente: {str(e)}")
            
            # Configurar comportamento ao fechar
            app_relatorio.root.protocol("WM_DELETE_WINDOW", lambda: self.finalizar_sistema(app_relatorio.root))
            
            # Exibir janela
            app_relatorio.root.lift()
            app_relatorio.root.focus_force()
            print("Iniciando mainloop do relatório de fornecedores")
            app_relatorio.root.mainloop()
        except Exception as e:
            import traceback
            print(f"Erro em iniciar_relatorio_fornecedores: {str(e)}")
            traceback.print_exc()
            messagebox.showerror(
                "Erro", 
                f"Ocorreu um erro ao iniciar o relatório de fornecedores.\nErro: {str(e)}"
            )
            self.root.deiconify()

    def iniciar_relatorio_lancamentos_pendentes(self, classe_relatorio):
        """
        Inicia a geração do relatório de lançamentos pendentes
        """
        try:
            # Verificar se pasta foi selecionada
            if not hasattr(self, 'pasta_lancamentos'):
                messagebox.showerror("Erro", "Por favor, selecione uma pasta primeiro.")
                return
            
            # Data de referência
            data_ref = self.data_referencia_pendentes.get_date() if hasattr(self, 'data_referencia_pendentes') else datetime.now()
        
            # Garantir que data_ref é datetime e não apenas date
            if isinstance(data_ref, date) and not isinstance(data_ref, datetime):
                data_ref = datetime.combine(data_ref, datetime.min.time())
            
            # Instanciar relatório
            relatorio = classe_relatorio()
            
            # Gerar relatório
            arquivo_saida = os.path.join(self.pasta_lancamentos, "relatorio_lancamentos_pendentes.html")
            
            # Usar o método gerar_relatorio_pendentes que já existe na classe
            if relatorio.gerar_relatorio_pendentes(self.pasta_lancamentos, arquivo_saida, data_ref):
                messagebox.showinfo(
                    "Sucesso",
                    f"Relatório gerado com sucesso!\nSalvo em: {arquivo_saida}"
                )
            else:
                messagebox.showwarning(
                    "Aviso",
                    "Nenhum lançamento pendente encontrado."
                )
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {str(e)}")
        
    def finalizar_sistema(self, janela):
        """Fecha a janela do sistema e mostra a janela principal"""
        janela.destroy()
        self.root.deiconify()
        self.root.lift()

    def voltar_menu(self):
        """Volta ao menu principal de forma segura"""
        try:
            logger.info("Solicitado retorno ao menu principal")
            
            # Verificar se existe menu principal para retornar
            if self.menu_principal and hasattr(self.menu_principal, 'winfo_exists'):
                try:
                    # Verificar se a janela do menu principal ainda existe
                    if self.menu_principal.winfo_exists():
                        logger.info("Retornando ao menu principal existente")
                        
                        # Destruir janela atual
                        self.root.destroy()
                        
                        # Restaurar e focar no menu principal
                        self.menu_principal.deiconify()
                        self.menu_principal.lift()
                        self.menu_principal.focus_force()
                        
                        logger.info("Retorno ao menu principal concluído")
                        return
                    else:
                        logger.warning("Menu principal não existe mais")
                except Exception as e:
                    logger.error(f"Erro ao verificar menu principal: {str(e)}")
            
            # Se não há menu principal válido, fechar aplicação completamente
            logger.info("Não há menu principal válido, fechando aplicação")
            
            # Tentar fechar de forma segura
            try:
                self.root.quit()
                self.root.destroy()
            except:
                pass
            
            # Forçar saída se necessário
            import sys
            import os
            os._exit(0)
            
        except Exception as e:
            logger.error(f"Erro crítico no voltar_menu: {str(e)}")
            # Último recurso: forçar saída
            try:
                import os
                os._exit(0)
            except:
                pass
    
    def _criar_preview_alternativo(self, dados_processados, configuracoes):
        """Preview com PDF temporário real para análise detalhada"""
        try:
            logger.info("🔧 CRIANDO PREVIEW COM PDF TEMPORÁRIO")
            
            # Criar janela de preview
            preview_window = tk.Toplevel(self.root)
            preview_window.title("Preview do Relatório - PDF Temporário")
            preview_window.geometry("800x600")
            preview_window.transient(self.root)
            
            # Centralizar
            preview_window.update_idletasks()
            x = (preview_window.winfo_screenwidth() // 2) - 400
            y = (preview_window.winfo_screenheight() // 2) - 300
            preview_window.geometry(f"800x600+{x}+{y}")
            
            # Frame principal
            main_frame = ttk.Frame(preview_window, padding=15)
            main_frame.pack(fill='both', expand=True)
            
            # Título
            title_label = ttk.Label(
                main_frame, 
                text="📄 PREVIEW - PDF TEMPORÁRIO DO RELATÓRIO", 
                font=('Arial', 16, 'bold'),
                foreground='darkgreen'
            )
            title_label.pack(pady=(0, 15))
            
            # Informações
            info_frame = ttk.LabelFrame(main_frame, text="Informações do Relatório", padding=10)
            info_frame.pack(fill='x', pady=(0, 15))
            
            info_text = f"""
    Cliente: {dados_processados.get('nome_cliente', 'N/A')}
    Data: {dados_processados.get('data_relatorio', 'N/A')}
    Relatório nº: {dados_processados.get('numero_relatorio', 'N/A')}
    Total Acumulado: R$ {self._formatar_numero_preview(dados_processados.get('acumulado', 0))}

    📊 Registros processados:
    • Despesas principais: {len(dados_processados.get('df_filtrado', []))} registros
    • Colaboradores (sal/transp): {len(dados_processados.get('df_tp_desp_1', []))} registros  
    • Colaboradores (13º/fér): {len(dados_processados.get('df_tp_desp_2', []))} registros
    • Diárias: {len(dados_processados.get('df_diaria', []))} registros
            """
            
            ttk.Label(info_frame, text=info_text, font=('Arial', 10)).pack(anchor='w')
            
            # Status do PDF temporário
            status_frame = ttk.LabelFrame(main_frame, text="Status do PDF Temporário", padding=10)
            status_frame.pack(fill='x', pady=(0, 15))
            
            status_label = ttk.Label(status_frame, text="⏳ Gerando PDF temporário...", 
                                font=('Arial', 11, 'bold'), foreground='orange')
            status_label.pack()
            
            # Frame de ações
            action_frame = ttk.LabelFrame(main_frame, text="Ações Disponíveis", padding=10)
            action_frame.pack(fill='both', expand=True)
            
            # Variável para armazenar caminho do PDF temporário
            self.pdf_temporario_path = None
            
            # Função para gerar PDF temporário
            def gerar_pdf_temp():
                try:
                    status_label.config(text="⏳ Gerando PDF temporário...", foreground='orange')
                    preview_window.update()
                    
                    # USAR O SERVIÇO para gerar PDF temporário
                    self.pdf_temporario_path = self.despesas_service.gerar_pdf_temporario(
                        dados_processados, 
                        configuracoes['arquivo']
                    )
                    
                    status_label.config(text="✅ PDF temporário gerado!", foreground='green')
                    btn_abrir_temp.config(state='normal')
                    
                    logger.info(f"✅ PDF temporário: {self.pdf_temporario_path}")
                    
                except Exception as e:
                    logger.error(f"Erro ao gerar PDF temporário: {str(e)}")
                    status_label.config(text="❌ Erro ao gerar PDF temporário", foreground='red')
                    messagebox.showerror("Erro", f"Erro ao gerar PDF temporário: {str(e)}")
            
            # Função para abrir PDF temporário
            def abrir_pdf_temporario():
                try:
                    if self.pdf_temporario_path and os.path.exists(self.pdf_temporario_path):
                        self.abrir_arquivo(self.pdf_temporario_path)
                        logger.info(f"📖 Abrindo PDF temporário: {self.pdf_temporario_path}")
                    else:
                        messagebox.showerror("Erro", "PDF temporário não encontrado!")
                except Exception as e:
                    logger.error(f"Erro ao abrir PDF temporário: {str(e)}")
                    messagebox.showerror("Erro", f"Erro ao abrir PDF: {str(e)}")
            
            # Função para gerar PDF definitivo
            def gerar_pdf_definitivo():
                try:
                    logger.info("🚀 Gerando PDF definitivo...")
                    
                    caminho_final, nome_arquivo = self.despesas_service.gerar_pdf_definitivo(
                        dados_processados, 
                        configuracoes['arquivo']
                    )
                    
                    resposta = messagebox.askyesno(
                        "PDF Definitivo Gerado!",
                        f"✅ PDF definitivo gerado!\n\n"
                        f"Arquivo: {nome_arquivo}\n\n"
                        f"Deseja abrir o arquivo definitivo?"
                    )
                    
                    if resposta:
                        self.abrir_arquivo(caminho_final)
                    
                except Exception as e:
                    logger.error(f"Erro ao gerar PDF definitivo: {str(e)}")
                    messagebox.showerror("Erro", f"Erro ao gerar PDF definitivo: {str(e)}")
            
            # Função para voltar
            def voltar():
                try:
                    # Limpar PDF temporário
                    if self.pdf_temporario_path and os.path.exists(self.pdf_temporario_path):
                        try:
                            os.remove(self.pdf_temporario_path)
                            logger.info(f"🗑️ PDF temporário removido: {self.pdf_temporario_path}")
                        except:
                            pass
                    
                    preview_window.destroy()
                    self.root.deiconify()
                    self.root.lift()
                    self.root.focus_force()
                    
                except Exception as e:
                    logger.error(f"Erro ao voltar: {str(e)}")
            
            # Botões organizados
            btn_frame = ttk.Frame(action_frame)
            btn_frame.pack(fill='x', pady=10)
            
            # Primeira linha de botões
            btn_frame1 = ttk.Frame(btn_frame)
            btn_frame1.pack(fill='x', pady=(0, 5))
            
            ttk.Button(btn_frame1, text="🔧 Gerar PDF Temporário", 
                    command=gerar_pdf_temp).pack(side='left', padx=(0, 10))
            
            btn_abrir_temp = ttk.Button(btn_frame1, text="📖 Abrir PDF Temporário", 
                                    command=abrir_pdf_temporario, state='disabled')
            btn_abrir_temp.pack(side='left', padx=(0, 10))
            
            # Segunda linha de botões
            btn_frame2 = ttk.Frame(btn_frame)
            btn_frame2.pack(fill='x')
            
            ttk.Button(btn_frame2, text="🚀 Gerar PDF Definitivo", 
                    command=gerar_pdf_definitivo).pack(side='left', padx=(0, 10))
            
            ttk.Button(btn_frame2, text="⬅️ Voltar", 
                    command=voltar).pack(side='right')
            
            # Configurar fechamento
            preview_window.protocol("WM_DELETE_WINDOW", voltar)
            
            # Ocultar interface principal e focar preview
            self.root.withdraw()
            preview_window.deiconify()
            preview_window.lift()
            preview_window.focus_force()
            
            # Gerar PDF temporário automaticamente
            preview_window.after(500, gerar_pdf_temp)
            
            logger.info("✅ Preview com PDF temporário criado")
            
        except Exception as e:
            logger.error(f"💥 ERRO no preview com PDF temporário: {str(e)}")
            messagebox.showerror("Erro", f"Erro no preview: {str(e)}")
            self.root.deiconify()

    def _formatar_numero_preview(self, valor):
        """Formata número para preview"""
        try:
            if valor is None or pd.isna(valor):
                return "0,00"
            return f"{float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return "0,00"

    def run(self):
        """Inicia a execução do sistema"""
        try:
            # Pré-carregar lista de clientes
            self.lista_clientes = self.carregar_clientes()
            logger.info(f"Lista de clientes carregada com {len(self.lista_clientes)} itens")
            
        except Exception as e:
            logger.error(f"Erro ao carregar lista de clientes: {str(e)}", exc_info=True)
            self.lista_clientes = ['Todos os Clientes']
        
        # Configurar estilos
        style = ttk.Style()
        style.configure('Accentuated.TButton', font=('Arial', 11, 'bold'))
        
        # Iniciar mainloop
        self.root.mainloop()

# Função para executar o sistema como módulo independente
def main():
    app = SistemaRelatorios()
    app.run()

if __name__ == "__main__":
    main()
    
RelatoriosInterface = SistemaRelatorios

__all__ = ['SistemaRelatorios', 'RelatoriosInterface']