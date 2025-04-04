import tkinter as tk
from tkinter import ttk, messagebox
import os
from pathlib import Path
from openpyxl import load_workbook

# Adicionar diretório raiz ao path
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações mínimas necessárias
try:
    from src.config.config import ARQUIVO_CLIENTES, PASTA_CLIENTES
except ImportError:
    try:
        from config.config import ARQUIVO_CLIENTES, PASTA_CLIENTES
    except ImportError:
        messagebox.showerror("Erro", "Não foi possível importar as configurações.")
        raise

def carregar_lista_clientes():
    """Carrega a lista de clientes disponíveis"""
    try:
        print("Carregando lista de clientes...")
        # Verificar se o arquivo existe
        if not os.path.exists(ARQUIVO_CLIENTES):
            print(f"Arquivo não encontrado: {ARQUIVO_CLIENTES}")
            return []
            
        workbook = load_workbook(ARQUIVO_CLIENTES)
        sheet = workbook['Clientes']
        print(f"Planilha aberta, abas disponíveis: {workbook.sheetnames}")
        print(f"Número de linhas na planilha: {sheet.max_row}")
        
        clientes = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Nome do cliente
                print(f"Cliente encontrado: {row[0]}")
                clientes.append(row[0])
                
        workbook.close()
        print(f"Total de clientes carregados: {len(clientes)}")
        return sorted(clientes)
        
    except Exception as e:
        print(f"Erro ao carregar clientes: {str(e)}")
        return []

def aplicacao_emergencia():
    # Função principal que será executada
    print("Iniciando aplicação de emergência")
    
    # Criando janela básica
    root = tk.Tk()
    root.title("Aplicação de Emergência")
    root.geometry("600x400")
    
    # Certifique-se de que a janela está visível
    root.attributes('-topmost', True)
    root.update()
    root.attributes('-topmost', False)
    
    # Frame principal com fundo colorido para ser facilmente visível
    frame = ttk.Frame(root, padding=20)
    frame.pack(fill='both', expand=True)
    
    # Título
    tk.Label(
        frame, 
        text="Sistema de Gestão de Eventos - MODO DE EMERGÊNCIA",
        font=("Arial", 16, "bold"),
        bg="yellow"
    ).pack(pady=20)
    
    # Carregar clientes
    clientes = carregar_lista_clientes()
    
    # Seleção de cliente
    ttk.Label(frame, text="Selecione um cliente:").pack(pady=(10, 5))
    
    cliente_selecionado = tk.StringVar()
    combo = ttk.Combobox(frame, textvariable=cliente_selecionado, values=clientes, state="readonly", width=40)
    combo.pack(pady=5)
    
    # Função para processar a seleção
    def processar_selecao():
        cliente = cliente_selecionado.get()
        if not cliente:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro")
            return
            
        messagebox.showinfo("Cliente Selecionado", 
                           f"Cliente: {cliente}\nArquivo: {PASTA_CLIENTES / f'{cliente}.xlsx'}")
    
    ttk.Button(frame, text="Confirmar Seleção", command=processar_selecao).pack(pady=20)
    
    # Garantir que a janela seja encerrada corretamente
    def on_close():
        print("Fechando aplicação")
        root.quit()
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_close)
    
    # Forçar visibilidade
    root.lift()
    root.focus_force()
    
    print("Iniciando mainloop")
    root.mainloop()
    print("Aplicação encerrada")

# Executar a aplicação de emergência
if __name__ == "__main__":
    print("Iniciando modo de emergência")
    try:
        aplicacao_emergencia()
    except Exception as e:
        print(f"Erro na aplicação de emergência: {str(e)}")
        import traceback
        traceback.print_exc()