import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.ticker as mticker

# Adicionar diretório raiz ao path para importar módulos corretamente
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        PASTA_CLIENTES,
        BASE_PATH
    )
    print("Configurações importadas com sucesso")
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Definir valores padrão em caso de falha
    BASE_PATH = Path(".")
    ARQUIVO_CLIENTES = BASE_PATH / "dados" / "clientes.xlsx"
    PASTA_CLIENTES = BASE_PATH / "dados" / "clientes"

# Importar o utils.py
from src.config.utils import atualizar_combobox_clientes, cliente_esta_ativo, obter_info_cliente

try:
    from src.config.window_config import configurar_janela
    print("window_config importado com sucesso")
except ImportError as e:
    print(f"Erro ao importar window_config: {str(e)}")
    # Implementação simples de configurar_janela como fallback
    def configurar_janela(janela, titulo="Janela", largura=800, altura=600):
        janela.title(titulo)
        janela.geometry(f"{largura}x{altura}")
        janela.resizable(True, True)
        
        # Centralizar na tela
        janela.update_idletasks()
        width = janela.winfo_width()
        height = janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (width // 2)
        y = (janela.winfo_screenheight() // 2) - (height // 2)
        janela.geometry(f'{width}x{height}+{x}+{y}')

# Funções auxiliares
def formatar_moeda_br(valor):
    """Formata um valor numérico como moeda brasileira"""
    try:
        valor_float = float(valor)
        return f"R$ {valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return f"R$ 0,00"

class RelatorioTipoDespesa:
    """Classe para geração de relatórios por tipo de despesa"""
    
    def __init__(self, parent=None):
        """Inicializa a interface do relatório"""
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Relatório por Tipo de Despesa", 1200, 1000)
        
        # Definição dos tipos de despesa
        self.tipos_despesas = {
            1: "1) DESP. C/COLABORADORES",
            2: "2) TRANSF. PROGRAMADAS",
            3: "3) BOLETOS",
            4: "4) REEMBOLSOS",
            5: "5) DESP. PAGAS P/CLIENTE",
            6: "6) PAGTOS CAIXA DE OBRA",
            7: "7) ADMINISTRAÇÃO DA OBRA"
        }
        
        # Configuração de variáveis
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.data_referencia = datetime.now()
        self.df_despesas = None
        self.df_por_data = None
        self.data_selecionada = None
        self.dados_grafico = {}
        
        # Configurar interface
        self.setup_gui()
    
    def setup_gui(self):
        """Configuração da interface gráfica principal"""
        # Frame principal
        self.frame_principal = ttk.Frame(self.root, padding=10)
        self.frame_principal.pack(fill='both', expand=True)
        
        # Frame para seleção
        self.frame_selecao = ttk.LabelFrame(self.frame_principal, text="Seleção de Cliente e Data")
        self.frame_selecao.pack(fill='x', pady=10)
        
        # Container para cliente
        frame_cliente = ttk.Frame(self.frame_selecao)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Selecione o Cliente:", font=('Arial', 11)).pack(side='left', pady=5)
        self.cliente_combobox = ttk.Combobox(frame_cliente, width=40, font=('Arial', 11))
        self.cliente_combobox.pack(side='left', padx=5)
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)
        
        # Eliminamos o container para data de referência já que não é necessário neste relatório
        frame_data = ttk.Frame(self.frame_selecao)
        frame_data.pack(fill='x', padx=10, pady=10)
        
        # Botão de gerar relatório
        ttk.Button(
            frame_data,
            text="Gerar Relatório",
            command=self.gerar_relatorio,
            style='Big.TButton'
        ).pack(side='left', padx=20)
        
        # Frame para resultados - com notebook para separar visões
        self.frame_resultados = ttk.LabelFrame(self.frame_principal, text="Resultados")
        self.frame_resultados.pack(fill='both', expand=True, pady=10)
        
        # Notebook (abas)
        self.notebook = ttk.Notebook(self.frame_resultados)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Abas
        self.aba_resumo = ttk.Frame(self.notebook)
        self.aba_detalhes = ttk.Frame(self.notebook)
        self.aba_grafico = ttk.Frame(self.notebook)
        
        self.notebook.add(self.aba_resumo, text='Resumo')
        self.notebook.add(self.aba_detalhes, text='Detalhes')
        self.notebook.add(self.aba_grafico, text='Gráfico')
        
        # Configurar cada aba
        self.setup_aba_resumo()
        self.setup_aba_detalhes()
        self.setup_aba_grafico()
        
        # Botões na parte inferior
        frame_botoes = ttk.Frame(self.frame_principal)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(
            frame_botoes,
            text="Exportar para Excel",
            command=self.exportar_excel
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes,
            text="Exportar para PDF",
            command=self.exportar_pdf
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes,
            text="Voltar ao Menu",
            command=self.voltar_menu
        ).pack(side='right', padx=5)
        
        # Estilo para botões grandes
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 11, 'bold'), padding=(10, 5))
        
        # Carregar lista de clientes
        self.atualizar_lista_clientes()
    
    def setup_aba_resumo(self):
        """Configura a aba de resumo do relatório por data e tipo de despesa"""
        # Frame para informações do cliente
        frame_info = ttk.Frame(self.aba_resumo, padding=5)
        frame_info.pack(fill='x', pady=5)
        
        self.lbl_cliente_resumo = ttk.Label(
            frame_info, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_resumo.pack(side='left', padx=10)
        
        self.lbl_data_resumo = ttk.Label(
            frame_info, 
            text=f"Data: {datetime.now().strftime('%d/%m/%Y')}", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_data_resumo.pack(side='left', padx=10)
        
        # Frame para o TreeView com os dados por data
        frame_resumo = ttk.Frame(self.aba_resumo, padding=5)
        frame_resumo.pack(fill='both', expand=True, pady=5)
        
        # Criar TreeView para os dados por data
        # Colunas: 'data', tipo1, tipo2, ..., tipo7, 'total'
        colunas = ['data']
        for i in range(1, 8):
            colunas.append(f'tipo_{i}')
        colunas.append('total')
        
        self.tv_resumo = ttk.Treeview(frame_resumo, columns=colunas, show='headings', height=15)
        
        # Configurar cabeçalhos
        self.tv_resumo.heading('data', text='Data')
        for i in range(1, 8):
            # Usar uma versão abreviada do nome do tipo para o cabeçalho
            nome_tipo = self.tipos_despesas[i].split(')')[0] + ')'
            self.tv_resumo.heading(f'tipo_{i}', text=nome_tipo)
            
        self.tv_resumo.heading('total', text='Total (R$)')
        
        # Configurar colunas
        self.tv_resumo.column('data', width=100, anchor='center')
        for i in range(1, 8):
            # Permitir quebra de texto nos tipos de despesa
            self.tv_resumo.column(f'tipo_{i}', width=120, anchor='e', stretch=True)
        self.tv_resumo.column('total', width=120, anchor='e')
        
        # Configurar scrollbars
        scrollbar_y = ttk.Scrollbar(frame_resumo, orient='vertical', command=self.tv_resumo.yview)
        scrollbar_x = ttk.Scrollbar(frame_resumo, orient='horizontal', command=self.tv_resumo.xview)
        self.tv_resumo.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Adicionar à tela
        self.tv_resumo.pack(side='top', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        
        # Adicionar evento de seleção
        self.tv_resumo.bind('<<TreeviewSelect>>', self.selecionar_data)
        
        # Frame para resumo de totais
        frame_totais = ttk.LabelFrame(self.aba_resumo, text="Resumo Financeiro", padding=10)
        frame_totais.pack(fill='x', pady=10, padx=10)
        
        # Adicionar labels para total geral
        ttk.Label(frame_totais, text="Total de Despesas:", font=('Arial', 11, 'bold')).grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.lbl_total_geral = ttk.Label(frame_totais, text="R$ 0,00", font=('Arial', 11))
        self.lbl_total_geral.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_totais, text="Número de Datas:", font=('Arial', 11, 'bold')).grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.lbl_num_datas = ttk.Label(frame_totais, text="0", font=('Arial', 11))
        self.lbl_num_datas.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_totais, text="Média por Data:", font=('Arial', 11, 'bold')).grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.lbl_media_data = ttk.Label(frame_totais, text="R$ 0,00", font=('Arial', 11))
        self.lbl_media_data.grid(row=0, column=3, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_totais, text="Data de Maior Valor:", font=('Arial', 11, 'bold')).grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.lbl_maior_data = ttk.Label(frame_totais, text="Nenhuma", font=('Arial', 11))
        self.lbl_maior_data.grid(row=1, column=3, sticky='w', padx=5, pady=5)
    
    def setup_aba_detalhes(self):
        """Configura a aba de detalhes do relatório para a data selecionada"""
        # Frame para informações da data selecionada
        frame_info_data = ttk.Frame(self.aba_detalhes, padding=5)
        frame_info_data.pack(fill='x', pady=5)
        
        self.lbl_data_detalhe = ttk.Label(
            frame_info_data, 
            text="Data Selecionada: Nenhuma", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_data_detalhe.pack(side='left', padx=10)
        
        self.lbl_total_data_detalhe = ttk.Label(
            frame_info_data, 
            text="Total: R$ 0,00", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_total_data_detalhe.pack(side='left', padx=10)
        
        # Frame para a tabela de detalhes
        frame_tabela = ttk.Frame(self.aba_detalhes, padding=5)
        frame_tabela.pack(fill='both', expand=True, pady=5)
        
        # Criar TreeView para os lançamentos da data selecionada
        colunas = ('data', 'tipo', 'nome', 'referencia', 'dt_vencto', 'valor', 'observacao')
        self.tv_detalhes = ttk.Treeview(frame_tabela, columns=colunas, show='headings', height=15)
        
        # Configurar cabeçalhos
        self.tv_detalhes.heading('data', text='Data')
        self.tv_detalhes.heading('tipo', text='Tipo de Despesa')
        self.tv_detalhes.heading('nome', text='Nome')
        self.tv_detalhes.heading('referencia', text='Referência')
        self.tv_detalhes.heading('dt_vencto', text='Data Vencto')
        self.tv_detalhes.heading('valor', text='Valor (R$)')
        self.tv_detalhes.heading('observacao', text='Observação')
        
        # Configurar colunas
        self.tv_detalhes.column('data', width=80, anchor='center')
        self.tv_detalhes.column('tipo', width=50, anchor='center')  # Reduzido para mostrar apenas número
        self.tv_detalhes.column('nome', width=210, anchor='w')
        self.tv_detalhes.column('referencia', width=250, anchor='w')
        self.tv_detalhes.column('dt_vencto', width=80, anchor='center')
        self.tv_detalhes.column('valor', width=120, anchor='e')
        self.tv_detalhes.column('observacao', width=180, anchor='w')
        
        # Configurar scrollbars
        scrollbar_y = ttk.Scrollbar(frame_tabela, orient='vertical', command=self.tv_detalhes.yview)
        scrollbar_x = ttk.Scrollbar(frame_tabela, orient='horizontal', command=self.tv_detalhes.xview)
        self.tv_detalhes.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Adicionar à tela
        self.tv_detalhes.pack(side='top', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        
        # Frame para estatísticas da data
        frame_stats = ttk.LabelFrame(self.aba_detalhes, text="Estatísticas", padding=10)
        frame_stats.pack(fill='x', pady=10, padx=10)
        
        # Adicionar labels para estatísticas
        ttk.Label(frame_stats, text="Número de Lançamentos:", font=('Arial', 11, 'bold')).grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.lbl_num_lancamentos = ttk.Label(frame_stats, text="0", font=('Arial', 11))
        self.lbl_num_lancamentos.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_stats, text="Média por Lançamento:", font=('Arial', 11, 'bold')).grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.lbl_media_lancamento = ttk.Label(frame_stats, text="R$ 0,00", font=('Arial', 11))
        self.lbl_media_lancamento.grid(row=0, column=3, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_stats, text="Maior Lançamento:", font=('Arial', 11, 'bold')).grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.lbl_maior_lancamento = ttk.Label(frame_stats, text="R$ 0,00", font=('Arial', 11))
        self.lbl_maior_lancamento.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frame_stats, text="Menor Lançamento:", font=('Arial', 11, 'bold')).grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.lbl_menor_lancamento = ttk.Label(frame_stats, text="R$ 0,00", font=('Arial', 11))
        self.lbl_menor_lancamento.grid(row=1, column=3, sticky='w', padx=5, pady=5)
    
    def setup_aba_grafico(self):
        """Configura a aba de gráficos para a data selecionada"""
        # Frame para controles do gráfico
        frame_controles = ttk.Frame(self.aba_grafico, padding=5)
        frame_controles.pack(fill='x', pady=5)
        
        ttk.Label(frame_controles, text="Tipo de Gráfico:").pack(side='left', padx=5)
        self.combo_tipo_grafico = ttk.Combobox(frame_controles, values=[
            "Gráfico de Pizza",
            "Gráfico de Barras"
        ], state='readonly', width=30)
        self.combo_tipo_grafico.pack(side='left', padx=5)
        self.combo_tipo_grafico.current(0)
        
        ttk.Button(frame_controles, text="Atualizar Gráfico", command=self.atualizar_grafico).pack(side='left', padx=20)
        
        # Frame para informações da data no gráfico
        frame_info_grafico = ttk.Frame(self.aba_grafico, padding=5)
        frame_info_grafico.pack(fill='x', pady=5)
        
        self.lbl_data_grafico = ttk.Label(
            frame_info_grafico, 
            text="Data Selecionada: Nenhuma", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_data_grafico.pack(side='left', padx=10)
        
        # Frame para o gráfico
        self.frame_grafico = ttk.Frame(self.aba_grafico)
        self.frame_grafico.pack(fill='both', expand=True, pady=5)
    
    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes no combobox usando a função centralizada"""
        try:
            # Usar a função centralizada (apenas clientes ativos)
            self.info_clientes = atualizar_combobox_clientes(self.cliente_combobox, mostrar_inativos=False)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")

    # E modifique o método selecionar_cliente:

    def selecionar_cliente(self, event=None):
        """Atualiza o cliente selecionado"""
        self.cliente_atual = self.cliente_combobox.get()
        
        if self.cliente_atual:
            # Verificar se o cliente está ativo (extra proteção)
            if not cliente_esta_ativo(self.cliente_atual):
                messagebox.showwarning(
                    "Cliente Inativo", 
                    f"O cliente '{self.cliente_atual}' está inativo (contrato finalizado). " +
                    "Os dados serão mostrados somente para consulta."
                )
            
            # Obter informações do cliente
            info_cliente = obter_info_cliente(self.cliente_atual)
            
            # Atualizar label
            if hasattr(self, 'lbl_cliente_resumo'):
                texto_cliente = f"Cliente: {self.cliente_atual}"
                if info_cliente and not info_cliente['ativo']:
                    texto_cliente += " (INATIVO)"
                self.lbl_cliente_resumo.config(text=texto_cliente)
            
            # Definir o caminho do arquivo
            if info_cliente and 'arquivo' in info_cliente:
                self.arquivo_cliente = info_cliente['arquivo']
            else:
                self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
    
    def gerar_relatorio(self):
        """Gera o relatório com base nos dados selecionados"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
            
        # Definir data de referência como data atual (apenas para o relatório exportado)
        self.data_referencia = datetime.now()
            
        # Carregar dados
        if not self.carregar_dados():
            return
        
        # Preencher resumo
        self.preencher_resumo()
        
        # Limpar detalhes (pois ainda não há data selecionada)
        if hasattr(self, 'tv_detalhes'):
            for item in self.tv_detalhes.get_children():
                self.tv_detalhes.delete(item)
        
        # Resetar gráfico
        self.limpar_grafico()
        
        # Selecionar aba de resumo
        self.notebook.select(0)  # Índice 0 corresponde à primeira aba (resumo)
    
    def carregar_dados(self):
        """Carrega os dados para o relatório a partir da aba 'Dados'"""
        try:
            if not os.path.exists(self.arquivo_cliente):
                messagebox.showerror("Erro", f"Arquivo do cliente '{self.cliente_atual}' não encontrado!")
                return False
            
            # Carregar dados do Excel - da aba 'Dados'
            try:
                # Usar pandas para ler a aba Dados
                self.df_despesas = pd.read_excel(self.arquivo_cliente, sheet_name='Dados')
                print(f"Colunas do DataFrame: {self.df_despesas.columns.tolist()}")
                
                # Verificar se as colunas necessárias existem
                colunas_necessarias = ['TP_DESP', 'VALOR', 'DATA_REL']
                for coluna in colunas_necessarias:
                    if coluna not in self.df_despesas.columns:
                        messagebox.showerror("Erro", f"A coluna '{coluna}' não foi encontrada na aba Dados!")
                        return False
                
                # Converter DATA_REL para datetime
                self.df_despesas['DATA_REL'] = pd.to_datetime(self.df_despesas['DATA_REL'], errors='coerce')
                
                # Converter DT_VENCTO para datetime (se existir)
                if 'DT_VENCTO' in self.df_despesas.columns:
                    self.df_despesas['DT_VENCTO'] = pd.to_datetime(self.df_despesas['DT_VENCTO'], errors='coerce')
                else:
                    # Se não existir, criar coluna vazia
                    self.df_despesas['DT_VENCTO'] = pd.NaT

                # Garantir valores numéricos para a coluna valor
                self.df_despesas['VALOR'] = pd.to_numeric(self.df_despesas['VALOR'], errors='coerce').fillna(0)
                
                # Garantir que TP_DESP seja numérico (usar o primeiro número encontrado)
                def extrair_numero(valor):
                    if pd.isna(valor):
                        return 0
                    try:
                        # Tentar converter diretamente para inteiro
                        return int(valor)
                    except (ValueError, TypeError):
                        # Se falhar, tentar extrair o primeiro número
                        import re
                        match = re.search(r'(\d+)', str(valor))
                        if match:
                            return int(match.group(1))
                        return 0
                
                self.df_despesas['TP_DESP_NUM'] = self.df_despesas['TP_DESP'].apply(extrair_numero)
                
                # Ordenar dados por data
                self.df_despesas = self.df_despesas.sort_values(by='DATA_REL')
                
                # Agrupar dados por data e tipo de despesa
                self.preparar_dados_por_data()
                
                return True
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao carregar dados do Excel: {str(e)}")
                import traceback
                traceback.print_exc()
                return False
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
            return False
    
    def preparar_dados_por_data(self):
        """Prepara os dados agrupados por data"""
        try:
            # Criar um DataFrame para armazenar os dados agrupados por data e tipo
            # Agrupar por data e tipo de despesa
            df_pivot = self.df_despesas.pivot_table(
                index='DATA_REL', 
                columns='TP_DESP_NUM', 
                values='VALOR', 
                aggfunc='sum'
            ).fillna(0)
            
            # Resetar o índice para tornar a data uma coluna
            df_pivot = df_pivot.reset_index()
            
            # Criar colunas para cada tipo de despesa se não existirem
            for i in range(1, 8):
                if i not in df_pivot.columns:
                    df_pivot[i] = 0.0
            
            # Calcular total por data
            df_pivot['total'] = df_pivot[[i for i in range(1, 8) if i in df_pivot.columns]].sum(axis=1)
            
            # Ordenar por data (ascendente)
            df_pivot = df_pivot.sort_values(by='DATA_REL')
            
            # Armazenar o DataFrame para uso posterior
            self.df_por_data = df_pivot
            
            # Preparar dados para gráficos
            self.preparar_dados_grafico()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao preparar dados por data: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preparar_dados_grafico(self):
        """Prepara os dados para os gráficos"""
        try:
            # Inicializar dicionário de dados para gráficos
            self.dados_grafico = {}
            
            # Se não temos dados ou data selecionada, não há o que fazer
            if not hasattr(self, 'df_por_data') or self.df_por_data.empty:
                return
                
            # Dados para gráfico de pizza e barras por data selecionada
            # Serão preenchidos quando uma data for selecionada
            self.dados_grafico['pizza'] = None
            self.dados_grafico['barras'] = None
            
        except Exception as e:
            print(f"Erro ao preparar dados para gráfico: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preencher_resumo(self):
        """Preenche os dados da aba de resumo"""
        try:
            # Limpar TreeView
            for item in self.tv_resumo.get_children():
                self.tv_resumo.delete(item)
            
            # Adicionar dados à TreeView
            for _, row in self.df_por_data.iterrows():
                # Formatar data
                data_str = row['DATA_REL'].strftime('%d/%m/%Y')
                
                # Preparar valores para cada tipo de despesa
                valores = []
                for i in range(1, 8):
                    valor_formatado = formatar_moeda_br(row[i]) if i in row else "R$ 0,00"
                    valores.append(valor_formatado)
                
                # Adicionar total
                total_formatado = formatar_moeda_br(row['total'])
                
                # Inserir na treeview
                self.tv_resumo.insert(
                    '', 'end', 
                    values=[data_str] + valores + [total_formatado]
                )
            
            # Atualizar labels de totais
            total_geral = self.df_por_data['total'].sum()
            num_datas = len(self.df_por_data)
            
            self.lbl_total_geral.config(text=formatar_moeda_br(total_geral))
            self.lbl_num_datas.config(text=str(num_datas))
            
            # Calcular média por data
            if num_datas > 0:
                media_data = total_geral / num_datas
                self.lbl_media_data.config(text=formatar_moeda_br(media_data))
            else:
                self.lbl_media_data.config(text="R$ 0,00")
            
            # Identificar data de maior valor
            if not self.df_por_data.empty:
                idx_maior = self.df_por_data['total'].idxmax()
                data_maior = self.df_por_data.loc[idx_maior, 'DATA_REL'].strftime('%d/%m/%Y')
                self.lbl_maior_data.config(text=data_maior)
            else:
                self.lbl_maior_data.config(text="Nenhuma")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao preencher resumo: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def selecionar_data(self, event=None):
        """Atualiza a data selecionada e preenche as abas de detalhes e gráfico"""
        try:
            # Obter seleção atual
            selecao = self.tv_resumo.selection()
            if not selecao:
                return
                
            # Obter data selecionada
            item = self.tv_resumo.item(selecao[0])
            data_str = item['values'][0]  # Primeira coluna é a data
            
            # Converter string de data para datetime
            try:
                self.data_selecionada = datetime.strptime(data_str, '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Erro", f"Formato de data inválido: {data_str}")
                return
            
            # Atualizar label na aba de detalhes
            self.lbl_data_detalhe.config(text=f"Data Selecionada: {data_str}")
            
            # Atualizar label na aba de gráfico
            self.lbl_data_grafico.config(text=f"Data Selecionada: {data_str}")
            
            # Encontrar o total da data no DataFrame
            df_data = self.df_por_data[self.df_por_data['DATA_REL'] == self.data_selecionada]
            if not df_data.empty:
                total_data = df_data.iloc[0]['total']
                self.lbl_total_data_detalhe.config(text=f"Total: {formatar_moeda_br(total_data)}")
            
            # Filtrar dados para a data selecionada
            df_filtrado = self.df_despesas[self.df_despesas['DATA_REL'].dt.date == self.data_selecionada.date()].copy()
            
            # Preencher detalhes
            self.preencher_detalhes(df_filtrado)
            
            # Preparar dados para gráfico
            self.preparar_grafico_data_selecionada(df_filtrado)
            
            # Atualizar gráfico
            self.atualizar_grafico()
            
            # Alternar para a aba de detalhes
            self.notebook.select(1)  # Índice 1 corresponde à aba de detalhes
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao selecionar data: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preparar_grafico_data_selecionada(self, df_filtrado):
        """Prepara dados para os gráficos da data selecionada"""
        try:
            if df_filtrado.empty:
                self.dados_grafico['pizza'] = None
                self.dados_grafico['barras'] = None
                return
                
            # Agrupar por tipo de despesa
            df_agrupado = df_filtrado.groupby('TP_DESP_NUM')['VALOR'].sum().reset_index()
            
            # Adicionar nome completo do tipo de despesa
            df_agrupado['tipo_nome'] = df_agrupado['TP_DESP_NUM'].apply(
                lambda x: self.tipos_despesas.get(x, f"Tipo {x}") if pd.notna(x) else "Não classificado"
            )
            
            # Ordenar por tipo
            df_agrupado = df_agrupado.sort_values(by='TP_DESP_NUM')
            
            # Armazenar para gráficos
            self.dados_grafico['pizza'] = df_agrupado.copy()
            self.dados_grafico['barras'] = df_agrupado.copy()
            
        except Exception as e:
            print(f"Erro ao preparar gráfico para data selecionada: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preencher_detalhes(self, df_filtrado):
        """Preenche os detalhes para a data selecionada"""
        try:
            # Limpar tabela
            for item in self.tv_detalhes.get_children():
                self.tv_detalhes.delete(item)
            
            # Verificar se o DataFrame está vazio
            if df_filtrado.empty:
                return
            
            # Ordenar o DataFrame por Tipo de Despesa (ascendente), Nome (ascendente) e Valor (descendente)
            df_ordenado = df_filtrado.copy()
            
            # Garantir que todos os campos necessários existam
            if 'TP_DESP_NUM' not in df_ordenado.columns:
                df_ordenado['TP_DESP_NUM'] = df_ordenado['TP_DESP'].apply(lambda x: 
                    int(x) if isinstance(x, (int, float)) else (
                        int(x.split(')')[0]) if isinstance(x, str) and ')' in x else 0
                    )
                )
            
            if 'NOME' not in df_ordenado.columns:
                df_ordenado['NOME'] = ''
                
            # Ordenar primeiro por tipo, depois por nome (asc) e finalmente por valor (desc)
            df_ordenado = df_ordenado.sort_values(
                by=['TP_DESP_NUM', 'NOME', 'VALOR'], 
                ascending=[True, True, False]
            )
            
            # Adicionar dados à tabela
            for _, row in df_ordenado.iterrows():
                # Formatar data
                data_str = row['DATA_REL'].strftime('%d/%m/%Y') if pd.notna(row['DATA_REL']) else ''
                
                # Obter tipo de despesa
                tipo_num = row['TP_DESP_NUM'] if 'TP_DESP_NUM' in row else None
                tipo_str = str(tipo_num) if tipo_num is not None else "?"
                
                # Obter nome e referência
                nome = row.get('NOME', '') if pd.notna(row.get('NOME', '')) else ''
                
                # Obter referência e NF (juntar referência e NF)
                referencia = row.get('REFERÊNCIA', '') if pd.notna(row.get('REFERÊNCIA', '')) else ''
                nf = row.get('NF', '') if pd.notna(row.get('NF', '')) else ''

                # Concatenar referência e NF se ambos existirem
                if referencia and nf:
                    referencia = f"{referencia} - NF: {nf}"
                elif nf:
                    referencia = f"NF: {nf}"

                # Data de vencimento
                dt_vencto_str = ''
                if 'DT_VENCTO' in row and pd.notna(row['DT_VENCTO']):
                    dt_vencto_str = row['DT_VENCTO'].strftime('%d/%m/%Y')
                
                # Obter valor
                valor = formatar_moeda_br(row['VALOR'])
                
                # Obter observação
                observacao = row.get('OBSERVAÇÃO', '') if pd.notna(row.get('OBSERVAÇÃO', '')) else ''
                
                # Inserir na tabela
                self.tv_detalhes.insert(
                    '', 'end', 
                    values=(
                        data_str,
                        tipo_str,
                        nome,
                        referencia,
                        dt_vencto_str,
                        valor,
                        observacao
                    )
                )
            
            # Atualizar estatísticas
            num_lancamentos = len(df_filtrado)
            total_data = df_filtrado['VALOR'].sum()
            
            self.lbl_num_lancamentos.config(text=str(num_lancamentos))
            
            # Média por lançamento
            if num_lancamentos > 0:
                media_lancamento = total_data / num_lancamentos
                self.lbl_media_lancamento.config(text=formatar_moeda_br(media_lancamento))
            else:
                self.lbl_media_lancamento.config(text="R$ 0,00")
            
            # Maior e menor lançamento
            if not df_filtrado.empty:
                maior_valor = df_filtrado['VALOR'].max()
                menor_valor = df_filtrado['VALOR'].min()
                
                self.lbl_maior_lancamento.config(text=formatar_moeda_br(maior_valor))
                self.lbl_menor_lancamento.config(text=formatar_moeda_br(menor_valor))
            else:
                self.lbl_maior_lancamento.config(text="R$ 0,00")
                self.lbl_menor_lancamento.config(text="R$ 0,00")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao preencher detalhes: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def limpar_grafico(self):
        """Limpa o gráfico atual"""
        for widget in self.frame_grafico.winfo_children():
            widget.destroy()
    
    def atualizar_grafico(self, event=None):
        """Atualiza o gráfico com base na data selecionada"""
        try:
            tipo_grafico = self.combo_tipo_grafico.get()
            
            # Limpar frame do gráfico
            self.limpar_grafico()
                
            # Verificar se há dados para gerar o gráfico
            if not hasattr(self, 'dados_grafico') or not self.dados_grafico:
                return
                
            # Verificar se temos uma data selecionada
            if not self.data_selecionada:
                return
                
            # Criar figura
            fig, ax = plt.subplots(figsize=(10, 6))
            
            if tipo_grafico == "Gráfico de Pizza":
                self.criar_grafico_pizza(fig, ax)
            elif tipo_grafico == "Gráfico de Barras":
                self.criar_grafico_barras(fig, ax)
                
            # Exibir o gráfico
            canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar gráfico: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def criar_grafico_pizza(self, fig, ax):
        """Cria um gráfico de pizza com os tipos de despesa da data selecionada"""
        try:
            # Usar os dados para gráfico de pizza
            df = self.dados_grafico.get('pizza')
            
            if df is None or df.empty:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Cores para o gráfico
            colors = plt.cm.tab10.colors
            
            # Criar o gráfico de pizza
            wedges, texts, autotexts = ax.pie(
                df['VALOR'], 
                labels=df['tipo_nome'], 
                autopct='%1.1f%%',
                startangle=90,
                colors=colors,
                wedgeprops={'edgecolor': 'w', 'linewidth': 1}
            )
            
            # Melhorar aparência
            for text in texts:
                text.set_fontsize(9)
            
            for autotext in autotexts:
                autotext.set_fontsize(9)
                autotext.set_fontweight('bold')
            
            # Adicionar título
            data_str = self.data_selecionada.strftime('%d/%m/%Y')
            ax.set_title(f'Distribuição por Tipo de Despesa - {data_str}', fontsize=14, pad=20)
            
            # Ajustar layout
            fig.tight_layout()
            
        except Exception as e:
            print(f"Erro ao criar gráfico de pizza: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Mostrar erro no gráfico
            ax.text(0.5, 0.5, f"Erro ao gerar gráfico: {str(e)}", 
                horizontalalignment='center', verticalalignment='center',
                transform=ax.transAxes, fontsize=12, color='red')
    
    def criar_grafico_barras(self, fig, ax):
        """Cria um gráfico de barras com os tipos de despesa da data selecionada"""
        try:
            # Usar os dados para gráfico de barras
            df = self.dados_grafico.get('barras')
            
            if df is None or df.empty:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Ordenar por valor para melhor visualização
            df = df.sort_values(by='VALOR', ascending=True)
            
            # Cores para o gráfico
            colors = plt.cm.tab10.colors[:len(df)]
            
            # Criar o gráfico de barras
            bars = ax.barh(df['tipo_nome'], df['VALOR'], color=colors)
            
            # Adicionar valores nas barras
            for bar in bars:
                width = bar.get_width()
                label_x_pos = width + width * 0.01
                ax.text(label_x_pos, bar.get_y() + bar.get_height()/2, f'R$ {width:,.2f}'.replace(',', '.'),
                       va='center', fontsize=9)
            
            # Ajustar formatação do eixo x (valores)
            def format_real(x, pos):
                return f'R$ {x:,.0f}'.replace(',', '.')
            
            ax.xaxis.set_major_formatter(mticker.FuncFormatter(format_real))
            
            # Adicionar títulos e labels
            data_str = self.data_selecionada.strftime('%d/%m/%Y')
            ax.set_title(f'Valores por Tipo de Despesa - {data_str}', fontsize=14)
            ax.set_xlabel('Valor (R$)', fontsize=11)
            ax.set_ylabel('Tipo de Despesa', fontsize=11)
            
            # Adicionar grid
            ax.grid(axis='x', linestyle='--', alpha=0.7)
            
            # Ajustar layout
            fig.tight_layout()
            
        except Exception as e:
            print(f"Erro ao criar gráfico de barras: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Mostrar erro no gráfico
            ax.text(0.5, 0.5, f"Erro ao gerar gráfico: {str(e)}", 
                horizontalalignment='center', verticalalignment='center',
                transform=ax.transAxes, fontsize=12, color='red')
    
    def exportar_excel(self):
        """Exporta o relatório para um arquivo Excel"""
        if not hasattr(self, 'cliente_atual') or not self.cliente_atual:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return
            
        # Solicitar nome do arquivo ao usuário
        data_str = self.data_referencia.strftime('%d-%m-%Y')
        nome_padrao = f"Relatorio_{self.__class__.__name__}_{self.cliente_atual}_{data_str}.xlsx"
        
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            initialfile=nome_padrao
        )
        
        if not arquivo:
            return
            
        try:
            # Criar workbook
            wb = Workbook()
            
            # Remover a aba padrão
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Criar aba de resumo
            ws_resumo = wb.create_sheet("Resumo")
            
            # Adicionar cabeçalho
            ws_resumo['A1'] = "Relatório por Tipo de Despesa"
            ws_resumo['A1'].font = Font(size=14, bold=True)
            ws_resumo.merge_cells('A1:J1')
            ws_resumo['A1'].alignment = Alignment(horizontal='center')
            
            ws_resumo['A2'] = f"Cliente: {self.cliente_atual}"
            ws_resumo['A2'].font = Font(size=12, bold=True)
            ws_resumo.merge_cells('A2:J2')
            
            ws_resumo['A3'] = f"Data do relatório: {data_str}"
            ws_resumo['A3'].font = Font(size=12)
            ws_resumo.merge_cells('A3:J3')
            
            # Adicionar cabeçalhos da tabela
            headers = ["Data"] + [f"{i}) " for i in range(1, 8)] + ["Total (R$)"]
            for col, header in enumerate(headers, start=1):
                cell = ws_resumo.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(fgColor="DDDDDD", fill_type="solid")
            
            # Adicionar dados
            for i, (_, row) in enumerate(self.df_por_data.iterrows(), start=6):
                # Data formatada
                ws_resumo.cell(row=i, column=1, value=row['DATA_REL'])
                ws_resumo.cell(row=i, column=1).number_format = "dd/mm/yyyy"
                
                # Valores por tipo de despesa
                for j in range(1, 8):
                    ws_resumo.cell(row=i, column=j+1, value=row[j] if j in row else 0)
                    ws_resumo.cell(row=i, column=j+1).number_format = "R$ #,##0.00"
                
                # Total
                ws_resumo.cell(row=i, column=9, value=row['total'])
                ws_resumo.cell(row=i, column=9).number_format = "R$ #,##0.00"
            
            # Ajustar largura das colunas
            ws_resumo.column_dimensions['A'].width = 15
            for col in range(1, 9):
                ws_resumo.column_dimensions[get_column_letter(col+1)].width = 15
            
            # Adicionar totais
            total_row = 6 + len(self.df_por_data)
            
            ws_resumo.cell(row=total_row, column=1, value="TOTAL")
            ws_resumo.cell(row=total_row, column=1).font = Font(bold=True)
            
            # Totais por tipo de despesa
            for j in range(1, 8):
                formula = f"=SUM({get_column_letter(j+1)}6:{get_column_letter(j+1)}{total_row-1})"
                ws_resumo.cell(row=total_row, column=j+1, value=formula)
                ws_resumo.cell(row=total_row, column=j+1).font = Font(bold=True)
                ws_resumo.cell(row=total_row, column=j+1).number_format = "R$ #,##0.00"
            
            # Total geral
            total_formula = f"=SUM(I6:I{total_row-1})"
            ws_resumo.cell(row=total_row, column=9, value=total_formula)
            ws_resumo.cell(row=total_row, column=9).font = Font(bold=True)
            ws_resumo.cell(row=total_row, column=9).number_format = "R$ #,##0.00"
            
            # Criar aba de detalhes se tivermos uma data selecionada
            if hasattr(self, 'data_selecionada') and self.data_selecionada:
                ws_detalhes = wb.create_sheet("Detalhes")
                
                # Adicionar cabeçalho
                data_str_detalhe = self.data_selecionada.strftime('%d/%m/%Y')
                ws_detalhes['A1'] = f"Detalhes da Data: {data_str_detalhe}"
                ws_detalhes['A1'].font = Font(size=14, bold=True)
                ws_detalhes.merge_cells('A1:F1')
                ws_detalhes['A1'].alignment = Alignment(horizontal='center')
                
                # Filtrando dados para a data selecionada
                df_filtrado = self.df_despesas[self.df_despesas['DATA_REL'].dt.date == self.data_selecionada.date()].copy()
                
                # Adicionar cabeçalhos da tabela
                headers = ["Data", "Tipo", "Nome", "Referência", "Valor (R$)", "Observação"]
                for col, header in enumerate(headers, start=1):
                    cell = ws_detalhes.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(fgColor="DDDDDD", fill_type="solid")
                
                # Adicionar dados
                for i, (_, row) in enumerate(df_filtrado.iterrows(), start=4):
                    # Data formatada
                    if pd.notna(row['DATA_REL']):
                        ws_detalhes.cell(row=i, column=1, value=row['DATA_REL'])
                        ws_detalhes.cell(row=i, column=1).number_format = "dd/mm/yyyy"
                    
                    # Tipo de despesa
                    tipo_num = row['TP_DESP_NUM'] if 'TP_DESP_NUM' in row else None
                    tipo_nome = self.tipos_despesas.get(tipo_num, row.get('TP_DESP', 'Não classificado'))
                    ws_detalhes.cell(row=i, column=2, value=tipo_nome)
                    
                    # Nome
                    if 'NOME' in row and pd.notna(row['NOME']):
                        ws_detalhes.cell(row=i, column=3, value=row['NOME'])
                    
                    # Referência
                    if 'REFERÊNCIA' in row and pd.notna(row['REFERÊNCIA']):
                        ws_detalhes.cell(row=i, column=4, value=row['REFERÊNCIA'])
                    
                    # Valor
                    ws_detalhes.cell(row=i, column=5, value=row['VALOR'])
                    ws_detalhes.cell(row=i, column=5).number_format = "R$ #,##0.00"
                    
                    # Observação
                    if 'OBSERVAÇÃO' in row and pd.notna(row['OBSERVAÇÃO']):
                        ws_detalhes.cell(row=i, column=6, value=row['OBSERVAÇÃO'])
                
                # Ajustar largura das colunas
                ws_detalhes.column_dimensions['A'].width = 15
                ws_detalhes.column_dimensions['B'].width = 30
                ws_detalhes.column_dimensions['C'].width = 30
                ws_detalhes.column_dimensions['D'].width = 30
                ws_detalhes.column_dimensions['E'].width = 15
                ws_detalhes.column_dimensions['F'].width = 40
                
                # Adicionar total
                total_row = 4 + len(df_filtrado)
                
                ws_detalhes.cell(row=total_row, column=4, value="TOTAL")
                ws_detalhes.cell(row=total_row, column=4).font = Font(bold=True)
                
                # Total em R$
                total_formula = f"=SUM(E4:E{total_row-1})"
                ws_detalhes.cell(row=total_row, column=5, value=total_formula)
                ws_detalhes.cell(row=total_row, column=5).font = Font(bold=True)
                ws_detalhes.cell(row=total_row, column=5).number_format = "R$ #,##0.00"
            
            # Criar aba para todos os dados
            ws_dados = wb.create_sheet("Todos os Dados")
            
            # Adicionar cabeçalho
            ws_dados['A1'] = "Todos os Lançamentos"
            ws_dados['A1'].font = Font(size=14, bold=True)
            ws_dados.merge_cells('A1:F1')
            ws_dados['A1'].alignment = Alignment(horizontal='center')
            
            # Adicionar cabeçalhos da tabela
            headers = ["Data", "Tipo", "Nome", "Referência", "Valor (R$)", "Observação"]
            for col, header in enumerate(headers, start=1):
                cell = ws_dados.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(fgColor="DDDDDD", fill_type="solid")
            
            # Ordenar todos os dados por data
            df_todos = self.df_despesas.copy()
            df_todos = df_todos.sort_values(by='DATA_REL')
            
            # Adicionar todos os dados
            for i, (_, row) in enumerate(df_todos.iterrows(), start=4):
                # Data formatada
                if pd.notna(row['DATA_REL']):
                    ws_dados.cell(row=i, column=1, value=row['DATA_REL'])
                    ws_dados.cell(row=i, column=1).number_format = "dd/mm/yyyy"
                
                # Tipo de despesa
                tipo_num = row['TP_DESP_NUM'] if 'TP_DESP_NUM' in row else None
                tipo_nome = self.tipos_despesas.get(tipo_num, row.get('TP_DESP', 'Não classificado'))
                ws_dados.cell(row=i, column=2, value=tipo_nome)
                
                # Nome
                if 'NOME' in row and pd.notna(row['NOME']):
                    ws_dados.cell(row=i, column=3, value=row['NOME'])
                
                # Referência
                if 'REFERÊNCIA' in row and pd.notna(row['REFERÊNCIA']):
                    ws_dados.cell(row=i, column=4, value=row['REFERÊNCIA'])
                
                # Valor
                ws_dados.cell(row=i, column=5, value=row['VALOR'])
                ws_dados.cell(row=i, column=5).number_format = "R$ #,##0.00"
                
                # Observação
                if 'OBSERVAÇÃO' in row and pd.notna(row['OBSERVAÇÃO']):
                    ws_dados.cell(row=i, column=6, value=row['OBSERVAÇÃO'])
            
            # Ajustar largura das colunas
            ws_dados.column_dimensions['A'].width = 15
            ws_dados.column_dimensions['B'].width = 30
            ws_dados.column_dimensions['C'].width = 30
            ws_dados.column_dimensions['D'].width = 30
            ws_dados.column_dimensions['E'].width = 15
            ws_dados.column_dimensions['F'].width = 40
            
            # Adicionar total
            total_row = 4 + len(df_todos)
            
            ws_dados.cell(row=total_row, column=4, value="TOTAL")
            ws_dados.cell(row=total_row, column=4).font = Font(bold=True)
            
            # Total em R$
            total_formula = f"=SUM(E4:E{total_row-1})"
            ws_dados.cell(row=total_row, column=5, value=total_formula)
            ws_dados.cell(row=total_row, column=5).font = Font(bold=True)
            ws_dados.cell(row=total_row, column=5).number_format = "R$ #,##0.00"
            
            # Salvar o arquivo
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso para:\n{arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def exportar_pdf(self):
        """Exporta o relatório para um arquivo PDF"""
        if not hasattr(self, 'cliente_atual') or not self.cliente_atual:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return
            
        # Solicitar nome do arquivo ao usuário
        data_str = self.data_referencia.strftime('%d-%m-%Y')
        nome_padrao = f"Relatorio_{self.__class__.__name__}_{self.cliente_atual}_{data_str}.pdf"
        
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Arquivos PDF", "*.pdf")],
            initialfile=nome_padrao
        )
        
        if not arquivo:
            return
            
        try:
            # Verificar se temos a biblioteca reportlab disponível
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                import matplotlib.pyplot as plt
                import io
            except ImportError:
                messagebox.showerror("Erro", "Biblioteca ReportLab não encontrada. Por favor instale usando 'pip install reportlab'.")
                return
            
            # Criar documento PDF
            doc = SimpleDocTemplate(arquivo, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            heading1_style = styles['Heading1']
            heading2_style = styles['Heading2']
            normal_style = styles['Normal']
            
            # Título
            story.append(Paragraph(f"Relatório por Tipo de Despesa", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Informações do cliente
            story.append(Paragraph(f"Cliente: {self.cliente_atual}", heading1_style))
            story.append(Paragraph(f"Data do relatório: {data_str}", normal_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Resumo por data
            story.append(Paragraph("Resumo por Data", heading2_style))
            story.append(Spacer(1, 0.1*inch))
            
            # Cabeçalhos
            headers = ["Data"]
            for i in range(1, 8):
                headers.append(f"{i})")
            headers.append("Total (R$)")
            
            # Dados da tabela de resumo
            table_data = [headers]
            
            for _, row in self.df_por_data.iterrows():
                # Formatar data
                data_str = row['DATA_REL'].strftime('%d/%m/%Y')
                
                # Preparar valores para cada tipo de despesa
                valores = [data_str]
                for i in range(1, 8):
                    valor_formatado = f"R$ {row[i]:,.2f}".replace(',', '.').replace('.', ',') if i in row else "R$ 0,00"
                    valores.append(valor_formatado)
                
                # Adicionar total
                total_formatado = f"R$ {row['total']:,.2f}".replace(',', '.').replace('.', ',')
                valores.append(total_formatado)
                
                table_data.append(valores)
            
            # Adicionar linha de total
            if not self.df_por_data.empty:
                total_row = ["TOTAL"]
                for i in range(1, 8):
                    total_tipo = self.df_por_data[i].sum() if i in self.df_por_data.columns else 0
                    total_formatado = f"R$ {total_tipo:,.2f}".replace(',', '.').replace('.', ',')
                    total_row.append(total_formatado)
                
                # Total geral
                total_geral = self.df_por_data['total'].sum()
                total_geral_formatado = f"R$ {total_geral:,.2f}".replace(',', '.').replace('.', ',')
                total_row.append(total_geral_formatado)
                
                table_data.append(total_row)
            
            # Criar tabela de resumo
            # Calcular larguras de colunas (data é maior, valores são menores)
            col_widths = [1.0*inch]  # Data
            for _ in range(1, 8):
                col_widths.append(0.8*inch)  # Tipos de despesa
            col_widths.append(1.0*inch)  # Total
            
            resumo_table = Table(table_data, colWidths=col_widths)
            
            # Estilo da tabela
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),  # Fonte menor para caber
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ])
            
            # Destacar a linha de total
            if not self.df_por_data.empty:
                table_style.add('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
                table_style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
            
            resumo_table.setStyle(table_style)
            story.append(resumo_table)
            story.append(Spacer(1, 0.2*inch))
            
            # Se tiver uma data selecionada, adicionar detalhes
            if hasattr(self, 'data_selecionada') and self.data_selecionada:
                # Detalhes da data selecionada
                data_str_detalhe = self.data_selecionada.strftime('%d/%m/%Y')
                story.append(Paragraph(f"Detalhes - Data: {data_str_detalhe}", heading2_style))
                story.append(Spacer(1, 0.1*inch))
                
                # Filtrar dados para a data selecionada
                df_filtrado = self.df_despesas[self.df_despesas['DATA_REL'].dt.date == self.data_selecionada.date()].copy()
                
                if not df_filtrado.empty:
                    # Cabeçalhos
                    headers = ["Tipo", "Nome", "Referência", "Valor (R$)"]
                    
                    # Dados da tabela de detalhes
                    table_data = [headers]
                    
                    for _, row in df_filtrado.iterrows():
                        # Obter tipo de despesa
                        tipo_num = row['TP_DESP_NUM'] if 'TP_DESP_NUM' in row else None
                        tipo_nome = self.tipos_despesas.get(tipo_num, row.get('TP_DESP', 'Não classificado'))
                        
                        # Obter nome e referência
                        nome = row.get('NOME', '') if pd.notna(row.get('NOME', '')) else ''
                        referencia = row.get('REFERÊNCIA', '') if pd.notna(row.get('REFERÊNCIA', '')) else ''
                        
                        # Formatar valor
                        valor = f"R$ {row['VALOR']:,.2f}".replace(',', '.').replace('.', ',')
                        
                        # Adicionar linha
                        table_data.append([tipo_nome, nome, referencia, valor])
                    
                    # Adicionar linha de total
                    total_data = df_filtrado['VALOR'].sum()
                    total_formatado = f"R$ {total_data:,.2f}".replace(',', '.').replace('.', ',')
                    table_data.append(["TOTAL", "", "", total_formatado])
                    
                    # Criar tabela de detalhes
                    col_widths = [1.5*inch, 2.0*inch, 2.0*inch, 1.0*inch]
                    detalhes_table = Table(table_data, colWidths=col_widths)
                    
                    # Estilo da tabela
                    table_style = TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (2, -1), 'LEFT'),
                        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ])
                    
                    # Destacar a linha de total
                    table_style.add('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
                    table_style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
                    
                    detalhes_table.setStyle(table_style)
                    story.append(detalhes_table)
                    story.append(Spacer(1, 0.2*inch))
                    
                    # Adicionar gráfico
                    story.append(Paragraph("Gráfico de Distribuição por Tipo de Despesa", heading2_style))
                    story.append(Spacer(1, 0.1*inch))
                    
                    # Gerar gráfico para incluir no PDF
                    plt.figure(figsize=(7, 5))
                    
                    # Preparar dados para o gráfico
                    df_grafico = df_filtrado.groupby('TP_DESP_NUM')['VALOR'].sum().reset_index()
                    
                    # Adicionar nome do tipo
                    df_grafico['tipo_nome'] = df_grafico['TP_DESP_NUM'].apply(
                        lambda x: self.tipos_despesas.get(x, f"Tipo {x}")
                    )
                    
                    # Criar gráfico de pizza
                    if not df_grafico.empty:
                        plt.pie(
                            df_grafico['VALOR'], 
                            labels=df_grafico['tipo_nome'], 
                            autopct='%1.1f%%',
                            startangle=90,
                            colors=plt.cm.tab10.colors,
                            wedgeprops={'edgecolor': 'w', 'linewidth': 1}
                        )
                        
                        plt.title(f'Distribuição por Tipo de Despesa - {data_str_detalhe}', fontsize=12, pad=20)
                        plt.tight_layout()
                        
                        # Salvar o gráfico em um buffer
                        img_buffer = io.BytesIO()
                        plt.savefig(img_buffer, format='png', dpi=100)
                        img_buffer.seek(0)
                        plt.close()
                        
                        # Adicionar o gráfico ao PDF
                        img = Image(img_buffer, width=6*inch, height=4*inch)
                        story.append(img)
                    else:
                        story.append(Paragraph("Não há dados suficientes para gerar o gráfico.", normal_style))
                else:
                    story.append(Paragraph("Não há lançamentos para esta data.", normal_style))
            
            # Construir o PDF
            doc.build(story)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso para:\n{arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para PDF: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def voltar_menu(self):
        """Volta ao menu principal"""
        self.root.destroy()
        
        # Mostrar janela principal
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()

def main():
    """Função principal para executar o módulo de forma independente"""
    app = RelatorioTipoDespesa()
    app.root.mainloop()
    
if __name__ == "__main__":
    main()