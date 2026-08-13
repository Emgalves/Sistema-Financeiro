# src/comprovantes_beneficios/controle_registros.py
"""
Controle de emissões — substitui a versão anterior baseada em SQLite
(controle_db.py, descontinuada).

Decisão (registrada na especificação, seção 6): usar uma aba própria
("Controle_Comprovantes") dentro da MESMA planilha do cliente que já
guarda Dados, RESUMO, Contratos_ADM etc. — mantém coerência com o
resto do sistema e elimina a necessidade de um arquivo/pasta novos.

Por que isso é seguro (testado com arquivo real antes de adotar):
  - round-trip de abrir+salvar com openpyxl preserva fórmulas, imagens,
    mesclagens e formatação das demais abas (nenhuma perda observada).
  - abrir + acrescentar linha + salvar leva < 1s no arquivo real.

Princípio de segurança contra corrupção por sincronização (Drive):
  - NUNCA se edita uma linha já gravada. Toda operação só ACRESCENTA
    linhas novas ao final da aba. Uma reemissão (2ª via) ou um
    cancelamento também são novas linhas, nunca uma edição da linha
    original — isso preserva o histórico e minimiza (mas não elimina
    por completo — isso exigiria um servidor central) a janela de
    conflito em caso de duas máquinas gravando quase ao mesmo tempo.
  - Emissões em lote abrem e salvam o arquivo UMA VEZ só (não uma vez
    por colaborador), reduzindo tanto o tempo total quanto o número de
    janelas de gravação.
"""

from dataclasses import dataclass
from datetime import datetime
from typing import Optional
import openpyxl

NOME_ABA = 'Controle_Comprovantes'

_CABECALHO = [
    'DATA_EMISSAO', 'BENEFICIO', 'COMPETENCIA', 'CPF', 'NOME',
    'VALOR', 'DIAS', 'DATA_VENCIMENTO', 'USUARIO', 'MAQUINA',
    'CAMINHO_PDF', 'STATUS', 'OBSERVACAO',
]


@dataclass
class RegistroRecibo:
    data_emissao: str
    beneficio: str
    competencia: str
    cpf: str
    nome: str
    valor: Optional[float]
    dias: Optional[int]
    data_vencimento: Optional[str]
    usuario: Optional[str]
    maquina: Optional[str]
    caminho_pdf: str
    status: str
    observacao: Optional[str]


@dataclass
class NovoRegistro:
    """O que o chamador (interface) monta para cada colaborador emitido, antes de gravar."""
    beneficio: str
    competencia: str
    cpf: str
    nome: str
    caminho_pdf: str
    valor: Optional[float] = None
    dias: Optional[int] = None
    data_vencimento: Optional[str] = None
    usuario: Optional[str] = None
    maquina: Optional[str] = None
    observacao: Optional[str] = None


class ErroControleDuplicidade(RuntimeError):
    pass


# ============================================================================
# Leitura
# ============================================================================

def _linha_para_registro(linha) -> RegistroRecibo:
    return RegistroRecibo(
        data_emissao=linha[0], beneficio=linha[1], competencia=linha[2],
        cpf=linha[3], nome=linha[4], valor=linha[5], dias=linha[6],
        data_vencimento=linha[7], usuario=linha[8], maquina=linha[9],
        caminho_pdf=linha[10], status=linha[11], observacao=linha[12],
    )


def _ler_todos_registros(caminho_planilha_cliente: str) -> list[RegistroRecibo]:
    wb = openpyxl.load_workbook(caminho_planilha_cliente, read_only=True, data_only=True)
    if NOME_ABA not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[NOME_ABA]
    registros = [
        _linha_para_registro(linha)
        for linha in ws.iter_rows(min_row=2, values_only=True)
        if linha and linha[0] is not None
    ]
    wb.close()
    return registros


def _status_efetivo(registros: list[RegistroRecibo], beneficio: str, competencia: str, cpf: str) -> Optional[str]:
    """
    Status da última linha para essa combinação (chave natural), na ordem
    em que foram gravadas (== ordem no arquivo, já que só se acrescenta).
    None se nunca houve registro para essa combinação.
    """
    ultimo = None
    for r in registros:
        if r.beneficio == beneficio and r.competencia == competencia and r.cpf == cpf:
            ultimo = r
    return ultimo.status if ultimo else None


def ja_emitido(caminho_planilha_cliente: str, beneficio: str, competencia: str, cpf: str) -> bool:
    registros = _ler_todos_registros(caminho_planilha_cliente)
    return _status_efetivo(registros, beneficio, competencia, cpf) == 'EMITIDO'


def listar_emitidos(caminho_planilha_cliente: str, competencia: str, beneficio: str) -> list[RegistroRecibo]:
    """Último status por CPF, para essa competência/benefício — usado pela interface."""
    registros = [
        r for r in _ler_todos_registros(caminho_planilha_cliente)
        if r.competencia == competencia and r.beneficio == beneficio
    ]
    por_cpf: dict[str, RegistroRecibo] = {}
    for r in registros:  # ordem cronológica -> o último grava por cima no dict
        por_cpf[r.cpf] = r
    return list(por_cpf.values())


def buscar_por_cpf(caminho_planilha_cliente: str, cpf: str) -> list[RegistroRecibo]:
    return [r for r in _ler_todos_registros(caminho_planilha_cliente) if r.cpf == cpf]


# ============================================================================
# Escrita (sempre em lote — um único open/save por chamada)
# ============================================================================

def registrar_lote(
    caminho_planilha_cliente: str,
    novos_registros: list[NovoRegistro],
    permitir_reemissao: bool = False,
) -> tuple[list[NovoRegistro], list[NovoRegistro]]:
    """
    Grava vários registros de uma vez (uma única abertura/gravação do
    arquivo). Retorna (gravados, pulados):
      - gravados: registros efetivamente escritos.
      - pulados: registros que já tinham status EMITIDO e
        permitir_reemissao=False (não escritos).

    Não levanta erro para duplicidade — quem decide se um "pulado" é
    problema é a interface (mostra no resumo). Isso evita que uma
    emissão em lote pare no meio por causa de 1 colaborador já emitido.
    """
    if not novos_registros:
        return [], []

    wb = openpyxl.load_workbook(caminho_planilha_cliente)
    if NOME_ABA not in wb.sheetnames:
        ws = wb.create_sheet(NOME_ABA)
        ws.append(_CABECALHO)
    else:
        ws = wb[NOME_ABA]

    registros_existentes = [
        _linha_para_registro(linha)
        for linha in ws.iter_rows(min_row=2, values_only=True)
        if linha and linha[0] is not None
    ]

    gravados, pulados = [], []
    agora = datetime.now().isoformat(timespec='seconds')

    for novo in novos_registros:
        status_atual = _status_efetivo(
            registros_existentes, novo.beneficio, novo.competencia, novo.cpf,
        )
        if status_atual == 'EMITIDO' and not permitir_reemissao:
            pulados.append(novo)
            continue

        linha = [
            agora, novo.beneficio, novo.competencia, novo.cpf, novo.nome,
            novo.valor, novo.dias, novo.data_vencimento, novo.usuario,
            novo.maquina, novo.caminho_pdf, 'EMITIDO', novo.observacao,
        ]
        ws.append(linha)
        # Atualiza a lista em memória para que, dentro do MESMO lote, um
        # segundo registro para a mesma pessoa/benefício já veja o
        # anterior como emitido (evita duas linhas 'EMITIDO' se o
        # chamador mandar entradas repetidas por engano).
        registros_existentes.append(_linha_para_registro(linha))
        gravados.append(novo)

    if gravados:
        wb.save(caminho_planilha_cliente)
    wb.close()

    return gravados, pulados


def cancelar_emissao(
    caminho_planilha_cliente: str, beneficio: str, competencia: str,
    cpf: str, motivo: str, usuario: Optional[str] = None,
) -> bool:
    """
    Acrescenta uma linha de cancelamento (nunca edita a original).
    Retorna False se não havia emissão ativa pra cancelar.
    """
    registros = _ler_todos_registros(caminho_planilha_cliente)
    ultimo = None
    for r in registros:
        if r.beneficio == beneficio and r.competencia == competencia and r.cpf == cpf:
            ultimo = r
    if ultimo is None or ultimo.status != 'EMITIDO':
        return False

    wb = openpyxl.load_workbook(caminho_planilha_cliente)
    ws = wb[NOME_ABA] if NOME_ABA in wb.sheetnames else wb.create_sheet(NOME_ABA)
    if ws.max_row == 0 or ws.cell(1, 1).value != _CABECALHO[0]:
        ws.append(_CABECALHO)

    ws.append([
        datetime.now().isoformat(timespec='seconds'), beneficio, competencia,
        cpf, ultimo.nome, ultimo.valor, ultimo.dias, ultimo.data_vencimento,
        usuario, None, ultimo.caminho_pdf, 'CANCELADO', motivo,
    ])
    wb.save(caminho_planilha_cliente)
    wb.close()
    return True
