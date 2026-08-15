# src/comprovantes_beneficios/dados_candidatos.py
"""
Leitura das planilhas de origem e aplicação das regras de elegibilidade
definidas na especificação (seção 2).

Duas fontes:
  - Planilha do cliente (aba 'Dados'): quem tem direito a quê, e a valor de quê.
  - Base_Fornecedores.xlsx (aba 'Fornecedores', tipo_pessoa == 'PF'):
    fonte canônica de nome por CPF, usada para corrigir divergências de
    grafia do mesmo colaborador entre lançamentos.

Este módulo não depende de UI — recebe caminhos de arquivo e devolve
estruturas de dados (dataclasses) prontas para a interface consumir.
"""

from dataclasses import dataclass, field
from datetime import date
from typing import Optional
import openpyxl

from .normalizacao import normalizar_cpf, normalizar_nome, formatar_cpf


# ============================================================================
# TIPOS
# ============================================================================

BENEFICIO_TRANSPORTE = 'TRANSPORTE'
BENEFICIO_CAFE = 'CAFE'
BENEFICIO_CESTA_BASICA = 'CESTA_BASICA'
BENEFICIO_CESTA_NATAL = 'CESTA_NATAL'
BENEFICIO_TRANSPORTE_CAFE = 'TRANSPORTE_CAFE'  # impressão combinada — não é um benefício novo, é um MODO de emissão

# Valores aceitos na coluna REFERÊNCIA da planilha do cliente para cada
# benefício com valor. Comparação sempre por igualdade exata após
# normalizar (strip + upper) — 'VT E CAFÉ' (linha combinada) nunca cai
# aqui de propósito: é ignorada conforme decisão registrada na especificação.
_REFERENCIAS_TRANSPORTE = {'TRANSPORTE'}
_REFERENCIAS_CAFE = {'CAFÉ', 'CAFE'}
_REFERENCIAS_DIARIA = {'DIÁRIA', 'DIARIA'}


@dataclass
class Candidato:
    cpf: str                       # normalizado, 11 dígitos
    nome: str                      # já resolvido (canônico se encontrado)
    nome_divergente: bool          # True se a planilha do cliente grafou diferente da base de fornecedores
    nome_planilha: str             # nome como veio na planilha do cliente (para o aviso)
    fonte_nome: str                # 'BASE_FORNECEDORES' | 'PLANILHA_CLIENTE'
    valor: Optional[float] = None  # None para cestas
    dias: Optional[int] = None
    data_vencimento: Optional[date] = None
    competencia: str = ''          # 'AAAA-MM'


@dataclass
class DadosPagador:
    nome: str
    endereco: str
    cidade_emissao: str = 'NOVA LIMA'  # fallback só se Clientes.xlsx não tiver Cidade


@dataclass
class AvisoElegibilidade:
    cpf: str
    mensagem: str


def obter_dados_pagador(
    nome_cliente: str,
    caminho_clientes_xlsx: str,
    cidade_padrao: str = 'NOVA LIMA',
) -> DadosPagador:
    """
    Lê nome, endereço e cidade do pagador a partir de `Clientes.xlsx`
    (colunas 'Nome', 'Endereço', 'Cidade'), casando pelo nome do cliente
    selecionado na interface.

    Substitui a versão anterior, que lia da aba 'RESUMO' da própria
    planilha do cliente — descontinuada e ausente em vários clientes
    reais (ex.: BRUNO_AUGUSTO_DELLI_ZOTTI_SOUZA.xlsx só tem 'Dados' e
    'Contratos_ADM'). `Clientes.xlsx` é a fonte confiável e já é lida
    de qualquer forma para montar a lista de clientes ativos.

    cidade_emissao usa **só a coluna 'Cidade'**, sem UF — a coluna
    'Endereço' já é montada juntando Logradouro/Número/Localidade/CEP/
    Estado (inclusive em cadastros antigos, com formatação inconsistente
    de UF: " - MG", ", MG", " / MG"), então incluir UF de novo em
    cidade_emissao duplicaria a informação. Se 'Cidade' estiver vazia,
    usa `cidade_padrao`.
    """
    wb = openpyxl.load_workbook(caminho_clientes_xlsx, data_only=True, read_only=True)
    ws = wb['Clientes']

    cabecalho = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_nome = cabecalho.index('Nome')
    idx_endereco = cabecalho.index('Endereço')
    idx_cidade = cabecalho.index('Cidade')

    chave_buscada = normalizar_nome(nome_cliente)

    for row in ws.iter_rows(min_row=2, values_only=True):
        nome_linha = row[idx_nome]
        if not nome_linha or normalizar_nome(nome_linha) != chave_buscada:
            continue

        endereco = str(row[idx_endereco] or '').strip()
        cidade = row[idx_cidade]
        cidade_emissao = str(cidade).strip().upper() if cidade else cidade_padrao

        wb.close()
        return DadosPagador(
            nome=normalizar_nome(nome_linha),
            endereco=endereco,
            cidade_emissao=cidade_emissao,
        )

    wb.close()
    raise ValueError(
        f"Cliente '{nome_cliente}' não encontrado em Clientes.xlsx — "
        f"não foi possível montar os dados do pagador."
    )


# ============================================================================
# LEITURA DA BASE DE FORNECEDORES (fonte canônica de nome)
# ============================================================================

def carregar_nomes_canonicos(caminho_base_fornecedores: str) -> dict[str, str]:
    """
    Retorna um dict {cpf_normalizado: nome_canonico} a partir da aba
    'Fornecedores' do Base_Fornecedores.xlsx, restrito a tipo_pessoa == 'PF'.
    """
    wb = openpyxl.load_workbook(caminho_base_fornecedores, data_only=True, read_only=True)
    ws = wb['Fornecedores']

    cabecalho = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_cpf = cabecalho.index('CNPJ/CPF')
    idx_tipo = cabecalho.index('tipo_pessoa')
    idx_nome = cabecalho.index('NOME')
    idx_razao = cabecalho.index('RAZÃO SOCIAL')

    canonicos: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_tipo] != 'PF':
            continue
        cpf = normalizar_cpf(row[idx_cpf])
        if cpf is None:
            continue
        nome = row[idx_nome] or row[idx_razao]
        if nome:
            canonicos[cpf] = normalizar_nome(nome)

    wb.close()
    return canonicos


# ============================================================================
# LEITURA DA PLANILHA DO CLIENTE (aba 'Dados')
# ============================================================================

def _competencia_do_vencimento(dt_vencto) -> tuple[str, date]:
    """
    Extrai (competencia 'AAAA-MM', date) a partir do campo DT_VENCTO, que
    na planilha real aparece ora como datetime, ora como string 'dd/mm/aaaa'.
    """
    if isinstance(dt_vencto, str):
        dia, mes, ano = dt_vencto.strip().split('/')
        d = date(int(ano), int(mes), int(dia))
    else:
        d = dt_vencto.date() if hasattr(dt_vencto, 'date') else dt_vencto

    return f"{d.year:04d}-{d.month:02d}", d


def _ler_linhas_dados(caminho_planilha_cliente: str):
    """Gera cada linha da aba 'Dados' como dict indexado pelo cabeçalho."""
    wb = openpyxl.load_workbook(caminho_planilha_cliente, data_only=True, read_only=True)
    ws = wb['Dados']

    cabecalho = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        yield dict(zip(cabecalho, row))

    wb.close()



# ============================================================================
# MONTAGEM DA LISTA DE CANDIDATOS
# ============================================================================

def _resolver_candidato(
    cpf_bruto,
    nome_bruto,
    nomes_canonicos: dict[str, str],
    valor=None,
    dias=None,
    data_vencimento=None,
    competencia='',
) -> Optional[Candidato]:
    cpf = normalizar_cpf(cpf_bruto)
    if cpf is None:
        return None

    nome_planilha = normalizar_nome(nome_bruto)
    nome_canonico = nomes_canonicos.get(cpf)

    if nome_canonico:
        nome_final = nome_canonico
        divergente = nome_canonico != nome_planilha
        fonte = 'BASE_FORNECEDORES'
    else:
        nome_final = nome_planilha
        divergente = False
        fonte = 'PLANILHA_CLIENTE'

    return Candidato(
        cpf=cpf,
        nome=nome_final,
        nome_divergente=divergente,
        nome_planilha=nome_planilha,
        fonte_nome=fonte,
        valor=valor,
        dias=dias,
        data_vencimento=data_vencimento,
        competencia=competencia,
    )


def obter_competencias_disponiveis(caminho_planilha_cliente: str) -> list[str]:
    """
    Lista TODAS as competências ('AAAA-MM') presentes na planilha do
    cliente, ordenadas da mais recente para a mais antiga (a mais
    recente fica selecionada por padrão na interface).

    Revisão: emissão retroativa passou a ser permitida (recibo
    extraviado ou não emitido a tempo por outro motivo) — a restrição
    anterior (só mês corrente/futuro) foi removida a pedido do usuário.
    O controle de duplicidade (controle_registros.py) continua sendo o
    que evita emissão repetida, independente da competência ser passada,
    presente ou futura.
    """
    competencias = set()
    for linha in _ler_linhas_dados(caminho_planilha_cliente):
        dt_vencto = linha.get('DT_VENCTO')
        if not dt_vencto:
            continue
        try:
            competencia, _ = _competencia_do_vencimento(dt_vencto)
        except (ValueError, AttributeError):
            continue
        competencias.add(competencia)
    return sorted(competencias, reverse=True)


def obter_candidatos(
    caminho_planilha_cliente: str,
    caminho_base_fornecedores: str,
    beneficio: str,
    competencia: str,
) -> tuple[list[Candidato], list[AvisoElegibilidade]]:
    """
    Retorna (candidatos, avisos) para o benefício e competência pedidos.

    - beneficio: uma de BENEFICIO_TRANSPORTE / BENEFICIO_CAFE /
      BENEFICIO_CESTA_BASICA / BENEFICIO_CESTA_NATAL.
    - Não restringe mais competência passada (emissão retroativa
      permitida — ver obter_competencias_disponiveis).
    """
    nomes_canonicos = carregar_nomes_canonicos(caminho_base_fornecedores)

    candidatos: dict[str, Candidato] = {}
    avisos: list[AvisoElegibilidade] = []

    for linha in _ler_linhas_dados(caminho_planilha_cliente):
        referencia = (linha.get('REFERÊNCIA') or '').strip().upper()
        dt_vencto = linha.get('DT_VENCTO')
        if not dt_vencto:
            continue
        try:
            comp_linha, data_venc = _competencia_do_vencimento(dt_vencto)
        except (ValueError, AttributeError):
            continue
        if comp_linha != competencia:
            continue

        if beneficio == BENEFICIO_TRANSPORTE and referencia in _REFERENCIAS_TRANSPORTE:
            c = _resolver_candidato(
                linha.get('CNPJ_CPF'), linha.get('NOME'), nomes_canonicos,
                valor=linha.get('VALOR'), dias=linha.get('DIAS'),
                data_vencimento=data_venc, competencia=comp_linha,
            )
        elif beneficio == BENEFICIO_CAFE and referencia in _REFERENCIAS_CAFE:
            c = _resolver_candidato(
                linha.get('CNPJ_CPF'), linha.get('NOME'), nomes_canonicos,
                valor=linha.get('VALOR'), dias=linha.get('DIAS'),
                data_vencimento=data_venc, competencia=comp_linha,
            )
        elif beneficio in (BENEFICIO_CESTA_BASICA, BENEFICIO_CESTA_NATAL) and (
            referencia in _REFERENCIAS_TRANSPORTE
            or referencia in _REFERENCIAS_CAFE
            or referencia in _REFERENCIAS_DIARIA
        ):
            c = _resolver_candidato(
                linha.get('CNPJ_CPF'), linha.get('NOME'), nomes_canonicos,
                valor=None, dias=None,
                data_vencimento=data_venc, competencia=comp_linha,
            )
        else:
            c = None

        if c is None:
            continue

        if c.cpf not in candidatos:
            candidatos[c.cpf] = c

        if c.nome_divergente:
            avisos.append(AvisoElegibilidade(
                cpf=formatar_cpf(c.cpf),
                mensagem=(
                    f"Nome na planilha ('{c.nome_planilha}') diverge do "
                    f"cadastro em Base_Fornecedores.xlsx ('{c.nome}'). "
                    f"Usando o nome da base de fornecedores."
                ),
            ))
        if c.fonte_nome == 'PLANILHA_CLIENTE':
            avisos.append(AvisoElegibilidade(
                cpf=formatar_cpf(c.cpf),
                mensagem=(
                    f"CPF não encontrado em Base_Fornecedores.xlsx — "
                    f"usando nome da planilha do cliente ('{c.nome}'). "
                    f"Considere cadastrar este colaborador na base."
                ),
            ))

    return sorted(candidatos.values(), key=lambda c: c.nome), avisos


# ============================================================================
# MODO COMBINADO — Transporte + Café na mesma página (opção, não substitui
# a emissão separada). Só faz sentido para quem tem os DOIS lançamentos na
# mesma competência.
# ============================================================================

def obter_candidatos_transporte_cafe(
    caminho_planilha_cliente: str,
    caminho_base_fornecedores: str,
    competencia: str,
) -> tuple[list[tuple[Candidato, Candidato]], list[AvisoElegibilidade]]:
    """
    Retorna pares (candidato_transporte, candidato_cafe) — só para CPFs
    que têm os DOIS lançamentos na competência pedida. Quem só tem um
    dos dois não aparece aqui (continua sendo emitido normalmente pelo
    fluxo separado de Transporte ou Café).
    """
    cands_transporte, avisos_t = obter_candidatos(
        caminho_planilha_cliente, caminho_base_fornecedores, BENEFICIO_TRANSPORTE, competencia,
    )
    cands_cafe, avisos_c = obter_candidatos(
        caminho_planilha_cliente, caminho_base_fornecedores, BENEFICIO_CAFE, competencia,
    )

    mapa_cafe = {c.cpf: c for c in cands_cafe}
    pares = [
        (ct, mapa_cafe[ct.cpf])
        for ct in cands_transporte
        if ct.cpf in mapa_cafe
    ]
    pares.sort(key=lambda par: par[0].nome)

    # Dedup de avisos (mesmo CPF pode gerar o mesmo aviso nas duas buscas)
    avisos_vistos = set()
    avisos = []
    for a in avisos_t + avisos_c:
        chave = (a.cpf, a.mensagem)
        if chave not in avisos_vistos:
            avisos_vistos.add(chave)
            avisos.append(a)

    return pares, avisos
