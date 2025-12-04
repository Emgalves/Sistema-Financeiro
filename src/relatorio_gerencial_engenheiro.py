"""
RELATÓRIO GERENCIAL POR ENGENHEIRO/GRUPO
=========================================

Novo módulo para visualização global das obras sob administração de cada engenheiro.

FUNCIONALIDADES:
1. Filtro por grupo (1 a 4)
2. Visão consolidada de todos os clientes do grupo
3. Status de contratos e medições por obra
4. Resumo executivo por grupo
5. Exportação para Excel
6. Gráficos de acompanhamento

DIFERENÇA DO RELATÓRIO ATUAL:
- Atual: Foco em UM cliente específico
- Novo: Visão de TODOS os clientes de um grupo/engenheiro
"""

import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from dateutil.relativedelta import relativedelta
from tkcalendar import DateEntry
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Adicionar diretório raiz ao path
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        PASTA_CLIENTES,
        BASE_PATH
    )
except ImportError:
    BASE_PATH = Path(".")
    ARQUIVO_CLIENTES = BASE_PATH / "dados" / "clientes.xlsx"
    PASTA_CLIENTES = BASE_PATH / "dados" / "clientes"

# Importar utils
from src.config.utils import cliente_esta_ativo

try:
    from src.config.window_config import configurar_janela
except ImportError:
    def configurar_janela(janela, titulo, largura=1200, altura=950):
        janela.title(titulo)
        screen_width = janela.winfo_screenwidth()
        screen_height = janela.winfo_screenheight()
        largura = min(largura, screen_width)
        altura = min(altura, screen_height)
        x = 0
        y = 0
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        janela.resizable(True, True)
        janela.lift()
        janela.focus_force()

def formatar_moeda_br(valor):
    """Formata um valor numérico como moeda brasileira"""
    try:
        valor_float = float(valor)
        return f"R$ {valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return f"R$ 0,00"


class RelatorioGerencialEngenheiro:
    """Relatório gerencial para acompanhamento de obras por engenheiro/grupo"""
    
    def __init__(self, parent=None):
        """Inicializa a interface do relatório gerencial"""
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Relatório Gerencial por Engenheiro/Grupo", 1400, 950)
        
        # Variáveis
        self.grupo_selecionado = None
        self.data_referencia = datetime.now()
        self.dados_obras = []  # Lista com dados consolidados
        
        # Setup interface
        self.setup_gui()
        
        # Carregar grupos disponíveis
        self.carregar_grupos()
        
    def setup_gui(self):
        """Configuração da interface gráfica"""
        # Frame principal
        self.frame_principal = ttk.Frame(self.root, padding=10)
        self.frame_principal.pack(fill='both', expand=True)
        
        # === FRAME DE SELEÇÃO ===
        self.frame_selecao = ttk.LabelFrame(self.frame_principal, text="Filtros", padding=10)
        self.frame_selecao.pack(fill='x', pady=10)
        
        # Linha 1: Grupo e Data
        frame_filtros = ttk.Frame(self.frame_selecao)
        frame_filtros.pack(fill='x')
        
        # Grupo
        ttk.Label(frame_filtros, text="Selecione o Grupo:", font=('Arial', 11, 'bold')).pack(side='left', padx=5)
        self.combo_grupo = ttk.Combobox(frame_filtros, width=20, font=('Arial', 11), state='readonly')
        self.combo_grupo.pack(side='left', padx=5)
        self.combo_grupo.bind('<<ComboboxSelected>>', self.on_grupo_selecionado)
        
        # Data de referência
        ttk.Label(frame_filtros, text="Data de Referência:", font=('Arial', 11)).pack(side='left', padx=(30, 5))
        self.data_entry = DateEntry(
            frame_filtros,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='pt_BR',
            font=('Arial', 11)
        )
        self.data_entry.pack(side='left', padx=5)
        self.data_entry.set_date(datetime.now())
        
        # Botão gerar
        ttk.Button(
            frame_filtros,
            text="🔍 Gerar Relatório",
            command=self.gerar_relatorio,
            style='Big.TButton'
        ).pack(side='left', padx=20)
        
        # Botão exportar
        ttk.Button(
            frame_filtros,
            text="📊 Exportar Excel",
            command=self.exportar_excel,
            style='Big.TButton'
        ).pack(side='left', padx=5)
        
        # === FRAME DE RESUMO ===
        self.frame_resumo = ttk.LabelFrame(self.frame_principal, text="Resumo do Grupo", padding=10)
        self.frame_resumo.pack(fill='x', pady=10)
        
        # Grid de resumo
        resumo_grid = ttk.Frame(self.frame_resumo)
        resumo_grid.pack(fill='x')
        
        # Linha 1
        ttk.Label(resumo_grid, text="Total de Obras:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5)
        self.lbl_total_obras = ttk.Label(resumo_grid, text="0", font=('Arial', 10))
        self.lbl_total_obras.grid(row=0, column=1, sticky='w', padx=5)
        
        ttk.Label(resumo_grid, text="Obras Ativas:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky='w', padx=(20, 5))
        self.lbl_obras_ativas = ttk.Label(resumo_grid, text="0", font=('Arial', 10))
        self.lbl_obras_ativas.grid(row=0, column=3, sticky='w', padx=5)
        
        ttk.Label(resumo_grid, text="Total Contratos:", font=('Arial', 10, 'bold')).grid(row=0, column=4, sticky='w', padx=(20, 5))
        self.lbl_total_contratos = ttk.Label(resumo_grid, text="0", font=('Arial', 10))
        self.lbl_total_contratos.grid(row=0, column=5, sticky='w', padx=5)
        
        # Linha 2
        ttk.Label(resumo_grid, text="Valor Total Contratado:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.lbl_valor_total = ttk.Label(resumo_grid, text="R$ 0,00", font=('Arial', 10))
        self.lbl_valor_total.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(resumo_grid, text="Valor Executado:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky='w', padx=(20, 5), pady=5)
        self.lbl_valor_executado = ttk.Label(resumo_grid, text="R$ 0,00", font=('Arial', 10))
        self.lbl_valor_executado.grid(row=1, column=3, sticky='w', padx=5, pady=5)
        
        ttk.Label(resumo_grid, text="Saldo a Executar:", font=('Arial', 10, 'bold')).grid(row=1, column=4, sticky='w', padx=(20, 5), pady=5)
        self.lbl_saldo = ttk.Label(resumo_grid, text="R$ 0,00", font=('Arial', 10))
        self.lbl_saldo.grid(row=1, column=5, sticky='w', padx=5, pady=5)
        
        # Linha 3
        ttk.Label(resumo_grid, text="% Executado Médio:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.lbl_percentual = ttk.Label(resumo_grid, text="0%", font=('Arial', 10))
        self.lbl_percentual.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(resumo_grid, text="Total de Medições:", font=('Arial', 10, 'bold')).grid(row=2, column=2, sticky='w', padx=(20, 5), pady=5)
        self.lbl_total_medicoes = ttk.Label(resumo_grid, text="0", font=('Arial', 10))
        self.lbl_total_medicoes.grid(row=2, column=3, sticky='w', padx=5, pady=5)
        
        ttk.Label(resumo_grid, text="Medições Pendentes:", font=('Arial', 10, 'bold')).grid(row=2, column=4, sticky='w', padx=(20, 5), pady=5)
        self.lbl_medicoes_pendentes = ttk.Label(resumo_grid, text="0", font=('Arial', 10))
        self.lbl_medicoes_pendentes.grid(row=2, column=5, sticky='w', padx=5, pady=5)
        
        # === NOTEBOOK (ABAS) ===
        self.frame_resultados = ttk.LabelFrame(self.frame_principal, text="Detalhamento", padding=5)
        self.frame_resultados.pack(fill='both', expand=True, pady=10)
        
        self.notebook = ttk.Notebook(self.frame_resultados)
        self.notebook.pack(fill='both', expand=True)
        
        # Aba 1: Visão por Obra
        self.aba_obras = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_obras, text='Visão por Obra')
        self.setup_aba_obras()
        
        # Aba 2: Visão por Contrato
        self.aba_contratos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_contratos, text='Todos os Contratos')
        self.setup_aba_contratos()
        
        # Aba 3: Visão por Medição
        self.aba_medicoes = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_medicoes, text='Todas as Medições')
        self.setup_aba_medicoes()
        
        # Aba 4: Gráficos
        self.aba_graficos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_graficos, text='Gráficos')
        self.setup_aba_graficos()
        
        # === BOTÃO VOLTAR ===
        ttk.Button(
            self.frame_principal,
            text="⬅ Voltar ao Menu",
            command=self.voltar_menu
        ).pack(pady=10)
        
    def setup_aba_obras(self):
        """Configura aba de visão por obra"""
        # Frame com scrollbar
        frame_scroll = ttk.Frame(self.aba_obras)
        frame_scroll.pack(fill='both', expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(frame_scroll)
        scrollbar.pack(side='right', fill='y')
        
        # Treeview
        colunas = ('Cliente', 'Contratos', 'Valor Total', 'Executado', 'Saldo', '% Exec', 'Status Obra')
        self.tree_obras = ttk.Treeview(
            frame_scroll,
            columns=colunas,
            show='headings',
            yscrollcommand=scrollbar.set,
            height=20
        )
        
        # Configurar colunas
        self.tree_obras.heading('Cliente', text='Cliente')
        self.tree_obras.heading('Contratos', text='Qtd Contratos')
        self.tree_obras.heading('Valor Total', text='Valor Total')
        self.tree_obras.heading('Executado', text='Executado')
        self.tree_obras.heading('Saldo', text='Saldo')
        self.tree_obras.heading('% Exec', text='% Executado')
        self.tree_obras.heading('Status Obra', text='Status da Obra')
        
        self.tree_obras.column('Cliente', width=300, anchor='w')
        self.tree_obras.column('Contratos', width=100, anchor='center')
        self.tree_obras.column('Valor Total', width=120, anchor='e')
        self.tree_obras.column('Executado', width=120, anchor='e')
        self.tree_obras.column('Saldo', width=120, anchor='e')
        self.tree_obras.column('% Exec', width=100, anchor='center')
        self.tree_obras.column('Status Obra', width=150, anchor='center')
        
        self.tree_obras.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.tree_obras.yview)
        
        # Bind duplo clique
        self.tree_obras.bind('<Double-1>', self.on_obra_duplo_clique)
        
    def setup_aba_contratos(self):
        """Configura aba de todos os contratos"""
        frame_scroll = ttk.Frame(self.aba_contratos)
        frame_scroll.pack(fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(frame_scroll)
        scrollbar.pack(side='right', fill='y')
        
        colunas = ('Cliente', 'ID Contrato', 'Fornecedor', 'Descrição', 'Valor Global', 'Executado', 'Saldo', '% Exec', 'Status')
        self.tree_contratos = ttk.Treeview(
            frame_scroll,
            columns=colunas,
            show='headings',
            yscrollcommand=scrollbar.set,
            height=20
        )
        
        # Cabeçalhos
        for col in colunas:
            self.tree_contratos.heading(col, text=col)
        
        # Larguras
        self.tree_contratos.column('Cliente', width=200, anchor='w')
        self.tree_contratos.column('ID Contrato', width=80, anchor='center')
        self.tree_contratos.column('Fornecedor', width=200, anchor='w')
        self.tree_contratos.column('Descrição', width=250, anchor='w')
        self.tree_contratos.column('Valor Global', width=120, anchor='e')
        self.tree_contratos.column('Executado', width=120, anchor='e')
        self.tree_contratos.column('Saldo', width=120, anchor='e')
        self.tree_contratos.column('% Exec', width=100, anchor='center')
        self.tree_contratos.column('Status', width=100, anchor='center')
        
        self.tree_contratos.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.tree_contratos.yview)
        
    def setup_aba_medicoes(self):
        """Configura aba de todas as medições"""
        frame_scroll = ttk.Frame(self.aba_medicoes)
        frame_scroll.pack(fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(frame_scroll)
        scrollbar.pack(side='right', fill='y')
        
        colunas = ('Cliente', 'ID Contrato', 'ID Medição', 'Data Medição', 'Data Pagamento', 'Referência', 'Valor', 'Status')
        self.tree_medicoes = ttk.Treeview(
            frame_scroll,
            columns=colunas,
            show='headings',
            yscrollcommand=scrollbar.set,
            height=20
        )
        
        # Cabeçalhos
        for col in colunas:
            self.tree_medicoes.heading(col, text=col)
        
        # Larguras
        self.tree_medicoes.column('Cliente', width=200, anchor='w')
        self.tree_medicoes.column('ID Contrato', width=80, anchor='center')
        self.tree_medicoes.column('ID Medição', width=80, anchor='center')
        self.tree_medicoes.column('Data Medição', width=100, anchor='center')
        self.tree_medicoes.column('Data Pagamento', width=110, anchor='center')
        self.tree_medicoes.column('Referência', width=300, anchor='w')
        self.tree_medicoes.column('Valor', width=120, anchor='e')
        self.tree_medicoes.column('Status', width=100, anchor='center')
        
        self.tree_medicoes.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.tree_medicoes.yview)
        
    def setup_aba_graficos(self):
        """Configura aba de gráficos"""
        # Frame para gráficos
        self.frame_graficos = ttk.Frame(self.aba_graficos)
        self.frame_graficos.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Placeholder
        ttk.Label(
            self.frame_graficos,
            text="Gráficos serão gerados após selecionar um grupo e gerar o relatório",
            font=('Arial', 12),
            foreground='gray'
        ).pack(expand=True)
        
    def carregar_grupos(self):
        """Carrega lista de grupos disponíveis (apenas de obras ativas)"""
        try:
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb.active
            
            grupos = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_final = row[4]  # Coluna E (índice 4) - Data Final
                grupo = row[10]  # Coluna K (índice 10) - Grupo
                
                # Só adiciona grupos de obras ativas (sem Data Final)
                if grupo and not data_final:
                    grupos.add(str(grupo))
            
            wb.close()
            
            # Ordenar grupos
            grupos_ordenados = sorted(grupos)
            
            # Adicionar opção "Todos"
            opcoes = ['Todos os Grupos'] + grupos_ordenados
            
            self.combo_grupo['values'] = opcoes
            if opcoes:
                self.combo_grupo.current(0)
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar grupos: {str(e)}")
            
    def on_grupo_selecionado(self, event=None):
        """Callback quando grupo é selecionado"""
        self.grupo_selecionado = self.combo_grupo.get()
        
    def gerar_relatorio(self):
        """Gera o relatório completo"""
        try:
            if not self.grupo_selecionado:
                messagebox.showwarning("Aviso", "Selecione um grupo primeiro!")
                return
            
            print("\n" + "="*80)
            print(f"🚀 INICIANDO GERAÇÃO DE RELATÓRIO - Grupo: {self.grupo_selecionado}")
            print("="*80)
            
            # Limpar dados anteriores
            self.dados_obras = []
            
            # Carregar clientes do grupo
            clientes = self.carregar_clientes_grupo()
            
            if not clientes:
                mensagem = f"Nenhum cliente encontrado para {self.grupo_selecionado}"
                print(f"⚠️ {mensagem}")
                messagebox.showinfo("Info", mensagem)
                return
            
            print(f"\n📋 Processando {len(clientes)} clientes...")
            
            # Processar cada cliente
            clientes_processados = 0
            clientes_com_erro = 0
            clientes_sem_dados = 0
            
            for i, cliente in enumerate(clientes, 1):
                print(f"\n[{i}/{len(clientes)}] {cliente}")
                dados_cliente = self.processar_cliente(cliente)
                
                if dados_cliente:
                    self.dados_obras.append(dados_cliente)
                    clientes_processados += 1
                elif dados_cliente is None:
                    clientes_sem_dados += 1
                else:
                    clientes_com_erro += 1
            
            print("\n" + "="*80)
            print("📊 RESULTADO DO PROCESSAMENTO:")
            print(f"   ✅ Processados com sucesso: {clientes_processados}")
            print(f"   ⚠️ Sem dados/arquivo: {clientes_sem_dados}")
            print(f"   ❌ Com erro: {clientes_com_erro}")
            print("="*80 + "\n")
            
            if not self.dados_obras:
                mensagem = f"Nenhuma obra com dados válidos encontrada para {self.grupo_selecionado}"
                print(f"⚠️ {mensagem}")
                messagebox.showwarning("Aviso", mensagem)
                return
            
            # Atualizar interface
            self.atualizar_resumo()
            self.atualizar_aba_obras()
            self.atualizar_aba_contratos()
            self.atualizar_aba_medicoes()
            self.gerar_graficos()
            
            mensagem = f"Relatório gerado com sucesso!\n\nObras processadas: {len(self.dados_obras)}/{len(clientes)}"
            print(f"✅ {mensagem}")
            messagebox.showinfo("Sucesso", mensagem)
            
        except Exception as e:
            mensagem = f"Erro ao gerar relatório: {str(e)}"
            print(f"❌ {mensagem}")
            messagebox.showerror("Erro", mensagem)
            import traceback
            traceback.print_exc()
            
    def carregar_clientes_grupo(self):
        """Carrega clientes do grupo selecionado (exclui obras finalizadas)"""
        try:
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb.active
            
            clientes = []
            total_processados = 0
            finalizados = 0
            sem_grupo_ativos = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                nome = row[0]
                data_final = row[4]  # Coluna E (índice 4) - Data Final
                grupo = str(row[10]) if row[10] else None
                
                if nome:
                    total_processados += 1
                    
                    # Verificar se está finalizado (Data Final preenchida)
                    if data_final:
                        finalizados += 1
                        continue  # Pula clientes finalizados
                    
                    # Contar clientes ativos sem grupo
                    if not grupo:
                        sem_grupo_ativos += 1
                        print(f"⚠️ Cliente ativo sem grupo: {nome}")
                    
                    # Filtrar por grupo (apenas ativos)
                    if self.grupo_selecionado == "Todos os Grupos":
                        if grupo:  # Só inclui se tiver grupo definido
                            clientes.append(nome)
                    elif grupo and grupo == self.grupo_selecionado:
                        clientes.append(nome)
            
            wb.close()
            
            print(f"\n📊 ANÁLISE DE CLIENTES:")
            print(f"   Total no arquivo: {total_processados}")
            print(f"   Finalizados (Data Final preenchida): {finalizados}")
            print(f"   Ativos sem grupo: {sem_grupo_ativos}")
            print(f"   Selecionados para '{self.grupo_selecionado}': {len(clientes)}")
            
            return clientes
            
        except Exception as e:
            print(f"❌ Erro ao carregar clientes: {e}")
            import traceback
            traceback.print_exc()
            return []
            
    def processar_cliente(self, nome_cliente):
        """Processa dados de um cliente"""
        try:
            # Caminho do arquivo do cliente
            nome_arquivo = nome_cliente.replace('/', '_').replace('\\', '_')
            arquivo_cliente = PASTA_CLIENTES / f"{nome_arquivo}.xlsx"
            
            if not arquivo_cliente.exists():
                print(f"⚠️ Arquivo não encontrado: {arquivo_cliente}")
                return None
            
            # Carregar planilha
            wb = load_workbook(arquivo_cliente)
            
            # Identificar aba de contratos (APENAS Contratos_Medicao)
            # Contratos_ADM é para taxa de administração e NÃO deve ser usado aqui
            aba_contratos = None
            if 'Contratos_Medicao' in wb.sheetnames:
                aba_contratos = 'Contratos_Medicao'
            elif 'Contratos' in wb.sheetnames:
                aba_contratos = 'Contratos'
            
            # Verificar se tem as abas necessárias
            if not aba_contratos:
                print(f"⚠️ Aba de contratos não encontrada em {nome_cliente}")
                print(f"   Abas disponíveis: {wb.sheetnames}")
                wb.close()
                return None
            
            ws_contratos = wb[aba_contratos]
            
            # Aba Medicoes é opcional
            ws_medicoes = None
            if 'Medicoes' in wb.sheetnames:
                ws_medicoes = wb['Medicoes']
            else:
                print(f"   ⚠️ Aba 'Medicoes' não encontrada (opcional), processando sem medições")
            
            print(f"✅ Processando {nome_cliente} (Aba: {aba_contratos})")
            
            # Processar contratos da aba Contratos_Medicao
            # Estrutura: [ID, CNPJ, Nome, Descrição, Data_Inicio, Valor_Global, Valor_Pago, Saldo, Status, Obs]
            contratos = []
            
            for row in ws_contratos.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tem ID
                    try:
                        # Converter data_inicio para string se for datetime
                        data_inicio = row[4]
                        if hasattr(data_inicio, 'strftime'):
                            data_inicio = data_inicio.strftime('%Y-%m-%d')
                        
                        contrato = {
                            'id': row[0],
                            'nome': row[2],  # Nome_Fornecedor (coluna 2)
                            'cnpj': row[1],  # CNPJ_Fornecedor (coluna 1)
                            'descricao': row[3],  # Descrição (coluna 3)
                            'data_inicio': data_inicio,  # Data_Inicio (coluna 4)
                            'valor_global': float(row[5]) if row[5] else 0,  # Valor_Global (coluna 5)
                            'valor_pago': float(row[6]) if row[6] else 0,  # Valor_Pago (coluna 6)
                            'status': row[8] if len(row) > 8 and row[8] else 'EM ANDAMENTO'  # Status (coluna 8)
                        }
                        contratos.append(contrato)
                    except (ValueError, TypeError, IndexError) as e:
                        print(f"   ⚠️ Erro ao processar contrato linha {row[0]}: {e}")
                        continue
            
            # Processar medições (se a aba existir)
            medicoes = []
            if ws_medicoes:
                for row in ws_medicoes.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Se tem ID do contrato
                        try:
                            medicao = {
                                'id_contrato': row[0],
                                'id_medicao': row[1],
                                'data_medicao': row[4] if len(row) > 4 else None,
                                'data_pagamento': row[5] if len(row) > 5 else None,
                                'referencia': row[6] if len(row) > 6 else '',
                                'valor': float(row[7]) if len(row) > 7 and row[7] else 0,
                                'status': row[8] if len(row) > 8 and row[8] else 'PENDENTE'
                            }
                            medicoes.append(medicao)
                        except (ValueError, TypeError, IndexError) as e:
                            print(f"   ⚠️ Erro ao processar medição, pulando: {e}")
                            continue
            
            wb.close()
            
            # Calcular totais
            valor_total = sum(c['valor_global'] for c in contratos)
            valor_executado = sum(c['valor_pago'] for c in contratos)
            saldo = valor_total - valor_executado
            perc_exec = (valor_executado / valor_total * 100) if valor_total > 0 else 0
            
            # Determinar status da obra
            contratos_ativos = sum(1 for c in contratos if c['status'] != 'CONCLUÍDO')
            if contratos_ativos == 0 and len(contratos) > 0:
                status_obra = 'CONCLUÍDA'
            elif contratos_ativos > 0:
                status_obra = 'EM ANDAMENTO'
            else:
                status_obra = 'SEM CONTRATOS'
            
            return {
                'cliente': nome_cliente,
                'contratos': contratos,
                'medicoes': medicoes,
                'totais': {
                    'qtd_contratos': len(contratos),
                    'valor_total': valor_total,
                    'valor_executado': valor_executado,
                    'saldo': saldo,
                    'perc_exec': perc_exec,
                    'status_obra': status_obra,
                    'qtd_medicoes': len(medicoes),
                    'medicoes_pendentes': sum(1 for m in medicoes if m['status'] == 'PENDENTE')
                }
            }
            
        except Exception as e:
            print(f"Erro ao processar cliente {nome_cliente}: {e}")
            return None
            
    def atualizar_resumo(self):
        """Atualiza painel de resumo"""
        try:
            # Calcular totais
            total_obras = len(self.dados_obras)
            obras_ativas = sum(1 for o in self.dados_obras if o['totais']['status_obra'] == 'EM ANDAMENTO')
            total_contratos = sum(o['totais']['qtd_contratos'] for o in self.dados_obras)
            
            valor_total = sum(o['totais']['valor_total'] for o in self.dados_obras)
            valor_executado = sum(o['totais']['valor_executado'] for o in self.dados_obras)
            saldo = valor_total - valor_executado
            
            perc_medio = (valor_executado / valor_total * 100) if valor_total > 0 else 0
            
            total_medicoes = sum(o['totais']['qtd_medicoes'] for o in self.dados_obras)
            medicoes_pendentes = sum(o['totais']['medicoes_pendentes'] for o in self.dados_obras)
            
            # Atualizar labels
            self.lbl_total_obras.config(text=str(total_obras))
            self.lbl_obras_ativas.config(text=str(obras_ativas))
            self.lbl_total_contratos.config(text=str(total_contratos))
            
            self.lbl_valor_total.config(text=formatar_moeda_br(valor_total))
            self.lbl_valor_executado.config(text=formatar_moeda_br(valor_executado))
            self.lbl_saldo.config(text=formatar_moeda_br(saldo))
            
            self.lbl_percentual.config(text=f"{perc_medio:.1f}%")
            self.lbl_total_medicoes.config(text=str(total_medicoes))
            self.lbl_medicoes_pendentes.config(text=str(medicoes_pendentes))
            
        except Exception as e:
            print(f"Erro ao atualizar resumo: {e}")
            
    def atualizar_aba_obras(self):
        """Atualiza árvore de obras"""
        # Limpar
        for item in self.tree_obras.get_children():
            self.tree_obras.delete(item)
        
        # Adicionar dados
        for obra in self.dados_obras:
            totais = obra['totais']
            self.tree_obras.insert('', 'end', values=(
                obra['cliente'],
                totais['qtd_contratos'],
                formatar_moeda_br(totais['valor_total']),
                formatar_moeda_br(totais['valor_executado']),
                formatar_moeda_br(totais['saldo']),
                f"{totais['perc_exec']:.1f}%",
                totais['status_obra']
            ))
            
    def atualizar_aba_contratos(self):
        """Atualiza árvore de contratos"""
        # Limpar
        for item in self.tree_contratos.get_children():
            self.tree_contratos.delete(item)
        
        # Adicionar dados
        for obra in self.dados_obras:
            for contrato in obra['contratos']:
                valor_global = contrato['valor_global']
                valor_pago = contrato['valor_pago']
                saldo = valor_global - valor_pago
                perc = (valor_pago / valor_global * 100) if valor_global > 0 else 0
                
                self.tree_contratos.insert('', 'end', values=(
                    obra['cliente'],
                    contrato['id'],
                    contrato['nome'],
                    contrato['descricao'],
                    formatar_moeda_br(valor_global),
                    formatar_moeda_br(valor_pago),
                    formatar_moeda_br(saldo),
                    f"{perc:.1f}%",
                    contrato['status']
                ))
                
    def atualizar_aba_medicoes(self):
        """Atualiza árvore de medições"""
        # Limpar
        for item in self.tree_medicoes.get_children():
            self.tree_medicoes.delete(item)
        
        # Adicionar dados
        for obra in self.dados_obras:
            for medicao in obra['medicoes']:
                data_med = medicao['data_medicao'].strftime('%d/%m/%Y') if isinstance(medicao['data_medicao'], datetime) else str(medicao['data_medicao'])
                data_pag = medicao['data_pagamento'].strftime('%d/%m/%Y') if isinstance(medicao['data_pagamento'], datetime) else str(medicao['data_pagamento'] or '')
                
                self.tree_medicoes.insert('', 'end', values=(
                    obra['cliente'],
                    medicao['id_contrato'],
                    medicao['id_medicao'],
                    data_med,
                    data_pag,
                    medicao['referencia'],
                    formatar_moeda_br(medicao['valor']),
                    medicao['status']
                ))
                
    def gerar_graficos(self):
        """Gera gráficos de acompanhamento"""
        try:
            # Limpar frame
            for widget in self.frame_graficos.winfo_children():
                widget.destroy()
            
            # Criar figura com subplots
            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
            fig.suptitle(f'Análise do {self.grupo_selecionado}', fontsize=16, fontweight='bold')
            
            # Gráfico 1: Valor por Obra (Top 10)
            obras_ordenadas = sorted(self.dados_obras, key=lambda x: x['totais']['valor_total'], reverse=True)[:10]
            nomes = [o['cliente'][:30] for o in obras_ordenadas]
            valores = [o['totais']['valor_total'] for o in obras_ordenadas]
            
            ax1.barh(nomes, valores, color='steelblue')
            ax1.set_xlabel('Valor (R$)')
            ax1.set_title('Top 10 Obras por Valor Contratado')
            ax1.grid(axis='x', alpha=0.3)
            
            # Gráfico 2: Percentual Executado por Obra
            nomes_perc = [o['cliente'][:30] for o in obras_ordenadas]
            percentuais = [o['totais']['perc_exec'] for o in obras_ordenadas]
            
            cores = ['green' if p >= 75 else 'orange' if p >= 50 else 'red' for p in percentuais]
            ax2.barh(nomes_perc, percentuais, color=cores)
            ax2.set_xlabel('% Executado')
            ax2.set_title('Percentual de Execução')
            ax2.grid(axis='x', alpha=0.3)
            
            # Gráfico 3: Status das Obras (Pizza)
            status_count = {}
            for obra in self.dados_obras:
                status = obra['totais']['status_obra']
                status_count[status] = status_count.get(status, 0) + 1
            
            ax3.pie(
                status_count.values(),
                labels=status_count.keys(),
                autopct='%1.1f%%',
                startangle=90,
                colors=['lightgreen', 'lightcoral', 'lightgray']
            )
            ax3.set_title('Distribuição por Status')
            
            # Gráfico 4: Valor Total vs Executado
            valor_total_grupo = sum(o['totais']['valor_total'] for o in self.dados_obras)
            valor_executado_grupo = sum(o['totais']['valor_executado'] for o in self.dados_obras)
            saldo_grupo = valor_total_grupo - valor_executado_grupo
            
            categorias = ['Contratado', 'Executado', 'Saldo']
            valores_resumo = [valor_total_grupo, valor_executado_grupo, saldo_grupo]
            cores_resumo = ['steelblue', 'green', 'orange']
            
            ax4.bar(categorias, valores_resumo, color=cores_resumo)
            ax4.set_ylabel('Valor (R$)')
            ax4.set_title('Resumo Financeiro do Grupo')
            ax4.grid(axis='y', alpha=0.3)
            
            # Ajustar layout
            plt.tight_layout()
            
            # Exibir no tkinter
            canvas = FigureCanvasTkAgg(fig, master=self.frame_graficos)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
            
        except Exception as e:
            print(f"Erro ao gerar gráficos: {e}")
            import traceback
            traceback.print_exc()
            
    def exportar_excel(self):
        """Exporta relatório para Excel"""
        try:
            if not self.dados_obras:
                messagebox.showwarning("Aviso", "Gere o relatório primeiro!")
                return
            
            # Solicitar local de salvamento
            arquivo = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile=f'Relatorio_Gerencial_{self.grupo_selecionado}_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
            
            if not arquivo:
                return
            
            # Criar workbook
            wb = Workbook()
            
            # Estilos
            titulo_font = Font(bold=True, size=14)
            cabecalho_font = Font(bold=True)
            cabecalho_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            borda = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # === ABA RESUMO ===
            ws_resumo = wb.active
            ws_resumo.title = "Resumo"
            
            # Título
            ws_resumo['A1'] = f"Relatório Gerencial - {self.grupo_selecionado}"
            ws_resumo['A1'].font = titulo_font
            ws_resumo.merge_cells('A1:G1')
            
            ws_resumo['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            # Totais
            linha = 4
            ws_resumo[f'A{linha}'] = "Total de Obras:"
            ws_resumo[f'B{linha}'] = self.lbl_total_obras.cget('text')
            
            ws_resumo[f'D{linha}'] = "Obras Ativas:"
            ws_resumo[f'E{linha}'] = self.lbl_obras_ativas.cget('text')
            
            linha += 1
            ws_resumo[f'A{linha}'] = "Valor Total Contratado:"
            ws_resumo[f'B{linha}'] = self.lbl_valor_total.cget('text')
            
            ws_resumo[f'D{linha}'] = "Valor Executado:"
            ws_resumo[f'E{linha}'] = self.lbl_valor_executado.cget('text')
            
            linha += 1
            ws_resumo[f'A{linha}'] = "Saldo a Executar:"
            ws_resumo[f'B{linha}'] = self.lbl_saldo.cget('text')
            
            ws_resumo[f'D{linha}'] = "% Executado Médio:"
            ws_resumo[f'E{linha}'] = self.lbl_percentual.cget('text')
            
            # Tabela de obras
            linha += 2
            cabecalhos = ['Cliente', 'Contratos', 'Valor Total', 'Executado', 'Saldo', '% Exec', 'Status']
            for col, texto in enumerate(cabecalhos, 1):
                celula = ws_resumo.cell(row=linha, column=col, value=texto)
                celula.font = cabecalho_font
                celula.fill = cabecalho_fill
                celula.border = borda
            
            linha += 1
            for obra in self.dados_obras:
                ws_resumo.cell(row=linha, column=1, value=obra['cliente'])
                ws_resumo.cell(row=linha, column=2, value=obra['totais']['qtd_contratos'])
                ws_resumo.cell(row=linha, column=3, value=obra['totais']['valor_total'])
                ws_resumo.cell(row=linha, column=4, value=obra['totais']['valor_executado'])
                ws_resumo.cell(row=linha, column=5, value=obra['totais']['saldo'])
                ws_resumo.cell(row=linha, column=6, value=f"{obra['totais']['perc_exec']:.1f}%")
                ws_resumo.cell(row=linha, column=7, value=obra['totais']['status_obra'])
                
                # Formato
                for col in [3, 4, 5]:
                    ws_resumo.cell(row=linha, column=col).number_format = '#,##0.00'
                
                linha += 1
            
            # Ajustar larguras
            ws_resumo.column_dimensions['A'].width = 40
            ws_resumo.column_dimensions['B'].width = 12
            ws_resumo.column_dimensions['C'].width = 15
            ws_resumo.column_dimensions['D'].width = 15
            ws_resumo.column_dimensions['E'].width = 15
            ws_resumo.column_dimensions['F'].width = 12
            ws_resumo.column_dimensions['G'].width = 15
            
            # === ABA CONTRATOS ===
            ws_contratos = wb.create_sheet("Todos os Contratos")
            
            ws_contratos['A1'] = "Todos os Contratos"
            ws_contratos['A1'].font = titulo_font
            ws_contratos.merge_cells('A1:I1')
            
            cabecalhos_contratos = ['Cliente', 'ID', 'Fornecedor', 'Descrição', 'Valor Global', 'Executado', 'Saldo', '% Exec', 'Status']
            for col, texto in enumerate(cabecalhos_contratos, 1):
                celula = ws_contratos.cell(row=3, column=col, value=texto)
                celula.font = cabecalho_font
                celula.fill = cabecalho_fill
                celula.border = borda
            
            linha = 4
            for obra in self.dados_obras:
                for contrato in obra['contratos']:
                    ws_contratos.cell(row=linha, column=1, value=obra['cliente'])
                    ws_contratos.cell(row=linha, column=2, value=contrato['id'])
                    ws_contratos.cell(row=linha, column=3, value=contrato['nome'])
                    ws_contratos.cell(row=linha, column=4, value=contrato['descricao'])
                    ws_contratos.cell(row=linha, column=5, value=contrato['valor_global'])
                    ws_contratos.cell(row=linha, column=6, value=contrato['valor_pago'])
                    ws_contratos.cell(row=linha, column=7, value=contrato['valor_global'] - contrato['valor_pago'])
                    
                    perc = (contrato['valor_pago'] / contrato['valor_global'] * 100) if contrato['valor_global'] > 0 else 0
                    ws_contratos.cell(row=linha, column=8, value=f"{perc:.1f}%")
                    ws_contratos.cell(row=linha, column=9, value=contrato['status'])
                    
                    for col in [5, 6, 7]:
                        ws_contratos.cell(row=linha, column=col).number_format = '#,##0.00'
                    
                    linha += 1
            
            # Ajustar larguras
            ws_contratos.column_dimensions['A'].width = 40
            ws_contratos.column_dimensions['B'].width = 8
            ws_contratos.column_dimensions['C'].width = 30
            ws_contratos.column_dimensions['D'].width = 40
            ws_contratos.column_dimensions['E'].width = 15
            ws_contratos.column_dimensions['F'].width = 15
            ws_contratos.column_dimensions['G'].width = 15
            ws_contratos.column_dimensions['H'].width = 12
            ws_contratos.column_dimensions['I'].width = 15
            
            # === ABA MEDIÇÕES ===
            ws_medicoes = wb.create_sheet("Todas as Medições")
            
            ws_medicoes['A1'] = "Todas as Medições"
            ws_medicoes['A1'].font = titulo_font
            ws_medicoes.merge_cells('A1:H1')
            
            cabecalhos_medicoes = ['Cliente', 'ID Contrato', 'ID Medição', 'Data Medição', 'Data Pagamento', 'Referência', 'Valor', 'Status']
            for col, texto in enumerate(cabecalhos_medicoes, 1):
                celula = ws_medicoes.cell(row=3, column=col, value=texto)
                celula.font = cabecalho_font
                celula.fill = cabecalho_fill
                celula.border = borda
            
            linha = 4
            for obra in self.dados_obras:
                for medicao in obra['medicoes']:
                    ws_medicoes.cell(row=linha, column=1, value=obra['cliente'])
                    ws_medicoes.cell(row=linha, column=2, value=medicao['id_contrato'])
                    ws_medicoes.cell(row=linha, column=3, value=medicao['id_medicao'])
                    
                    if isinstance(medicao['data_medicao'], datetime):
                        ws_medicoes.cell(row=linha, column=4, value=medicao['data_medicao'])
                        ws_medicoes.cell(row=linha, column=4).number_format = 'dd/mm/yyyy'
                    else:
                        ws_medicoes.cell(row=linha, column=4, value=str(medicao['data_medicao']))
                    
                    if isinstance(medicao['data_pagamento'], datetime):
                        ws_medicoes.cell(row=linha, column=5, value=medicao['data_pagamento'])
                        ws_medicoes.cell(row=linha, column=5).number_format = 'dd/mm/yyyy'
                    else:
                        ws_medicoes.cell(row=linha, column=5, value=str(medicao['data_pagamento'] or ''))
                    
                    ws_medicoes.cell(row=linha, column=6, value=medicao['referencia'])
                    ws_medicoes.cell(row=linha, column=7, value=medicao['valor'])
                    ws_medicoes.cell(row=linha, column=7).number_format = '#,##0.00'
                    ws_medicoes.cell(row=linha, column=8, value=medicao['status'])
                    
                    linha += 1
            
            # Ajustar larguras
            ws_medicoes.column_dimensions['A'].width = 40
            ws_medicoes.column_dimensions['B'].width = 12
            ws_medicoes.column_dimensions['C'].width = 12
            ws_medicoes.column_dimensions['D'].width = 15
            ws_medicoes.column_dimensions['E'].width = 15
            ws_medicoes.column_dimensions['F'].width = 50
            ws_medicoes.column_dimensions['G'].width = 15
            ws_medicoes.column_dimensions['H'].width = 12
            
            # Salvar
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso para:\n{arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
            import traceback
            traceback.print_exc()
            
    def on_obra_duplo_clique(self, event):
        """Callback para duplo clique em obra"""
        selection = self.tree_obras.selection()
        if selection:
            item = self.tree_obras.item(selection[0])
            cliente = item['values'][0]
            
            messagebox.showinfo(
                "Detalhes da Obra",
                f"Cliente: {cliente}\n\n"
                f"Duplo clique implementado!\n"
                f"Pode abrir janela detalhada ou navegar para cliente."
            )
            
    def voltar_menu(self):
        """Volta ao menu principal"""
        self.root.destroy()
        
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()


def main():
    """Função principal"""
    app = RelatorioGerencialEngenheiro()
    app.root.mainloop()


if __name__ == "__main__":
    main()