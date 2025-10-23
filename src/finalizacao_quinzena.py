def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
import os

from src.config.utils import (
    validar_data,
    ARQUIVO_CLIENTES,
    PASTA_CLIENTES,
    buscar_dados_bancarios_fornecedor
)


class FinalizacaoQuinzena:
    def __init__(self, parent=None):
        self.parent = parent
        self.root = tk.Toplevel(parent) if parent else tk.Tk()
        self.root.title("Finalização de Quinzena - Compensação Automática")
        self.root.geometry("1200x650")
        
        self.data_ref_entry = None
        self.tree_clientes = None
        self._cache_detalhes = {}
        
        self.setup_gui()
        
    def run(self):
        """Inicia a execução do sistema"""
        self.root.protocol("WM_DELETE_WINDOW", self.voltar_menu)
        
        # Centralizar janela
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        self.root.lift()
        self.root.focus_force()
        self.root.mainloop()

    def setup_gui(self):
        """Configura a interface gráfica"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Frame superior: Data e busca
        frame_topo = ttk.LabelFrame(main_frame, text="Período de Referência", padding="10")
        frame_topo.pack(fill='x', pady=(0, 10))

        # Botões para selecionar quinzena automaticamente
        ttk.Label(frame_topo, text="Selecione a Quinzena:").pack(side='left', padx=5)
        
        ttk.Button(
            frame_topo,
            text="📅 Próximo Dia 05",
            command=lambda: self._definir_proxima_quinzena(5)
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_topo,
            text="📅 Próximo Dia 20",
            command=lambda: self._definir_proxima_quinzena(20)
        ).pack(side='left', padx=5)
        
        # Ou selecionar data manualmente
        ttk.Label(frame_topo, text="ou Data Manual:").pack(side='left', padx=(20, 5))
        self.data_ref_entry = DateEntry(
            frame_topo,
            format='dd/mm/yyyy',
            locale='pt_BR'
        )
        self.data_ref_entry.pack(side='left', padx=5)
        
        # Buscar automaticamente quando data for alterada
        self.data_ref_entry.bind('<<DateEntrySelected>>', lambda e: self.carregar_clientes())

        ttk.Button(
            frame_topo, 
            text="🔍 Buscar Clientes",
            command=self.carregar_clientes
        ).pack(side='left', padx=10)

        # Frame de clientes
        frame_lista = ttk.LabelFrame(main_frame, text="Clientes Pendentes", padding="5")
        frame_lista.pack(fill='both', expand=True, pady=(0, 10))

        # Scrollbar
        scrollbar = ttk.Scrollbar(frame_lista)
        scrollbar.pack(side='right', fill='y')

        # Treeview
        self.tree_clientes = ttk.Treeview(
            frame_lista,
            columns=('Cliente', 'Base Atual', 'Taxa %', 'Taxa Quinzena', 'Compensação', 'Valor Final', 'Status'),
            show='headings',
            selectmode='extended',
            yscrollcommand=scrollbar.set
        )

        scrollbar.config(command=self.tree_clientes.yview)

        # Configurar colunas
        self.tree_clientes.heading('Cliente', text='Cliente')
        self.tree_clientes.heading('Base Atual', text='Base Quinzena')
        self.tree_clientes.heading('Taxa %', text='Taxa %')
        self.tree_clientes.heading('Taxa Quinzena', text='Taxa Quinzena')
        self.tree_clientes.heading('Compensação', text='Compensação')
        self.tree_clientes.heading('Valor Final', text='Valor Final')
        self.tree_clientes.heading('Status', text='Status')

        self.tree_clientes.column('Cliente', width=180)
        self.tree_clientes.column('Base Atual', width=120)
        self.tree_clientes.column('Taxa %', width=70)
        self.tree_clientes.column('Taxa Quinzena', width=120)
        self.tree_clientes.column('Compensação', width=120)
        self.tree_clientes.column('Valor Final', width=120)
        self.tree_clientes.column('Status', width=150)

        self.tree_clientes.pack(fill='both', expand=True)

        # Frame de botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes, 
            text="✅ Processar Selecionados",
            command=self.processar_clientes_selecionados
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="🔄 Atualizar Lista",
            command=self.carregar_clientes
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="📊 Ver Detalhes",
            command=self.ver_detalhes_compensacao
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="🔙 Voltar ao Menu",
            command=self.voltar_menu
        ).pack(side='right', padx=5)

    def _definir_proxima_quinzena(self, dia):
        """Define a próxima data de quinzena (dia 5 ou 20) e busca automaticamente"""
        hoje = datetime.now()
        
        if dia == 5:
            # Se hoje é antes do dia 5, usa dia 5 deste mês
            # Se já passou, usa dia 5 do próximo mês
            if hoje.day < 5:
                proxima = hoje.replace(day=5)
            else:
                # Próximo mês
                if hoje.month == 12:
                    proxima = datetime(hoje.year + 1, 1, 5)
                else:
                    proxima = datetime(hoje.year, hoje.month + 1, 5)
        else:  # dia == 20
            # Se hoje é antes do dia 20, usa dia 20 deste mês
            # Se já passou, usa dia 20 do próximo mês
            if hoje.day < 20:
                proxima = hoje.replace(day=20)
            else:
                # Próximo mês
                if hoje.month == 12:
                    proxima = datetime(hoje.year + 1, 1, 20)
                else:
                    proxima = datetime(hoje.year, hoje.month + 1, 20)
        
        self.data_ref_entry.set_date(proxima)
        # Buscar automaticamente após definir a data
        self.carregar_clientes()

    def carregar_clientes(self):
        """Carrega clientes com cálculo SIMPLES de compensação"""
        data_ref = self.data_ref_entry.get()
        if not validar_data(data_ref):
            messagebox.showerror("Erro", "Data inválida!")
            return

        try:
            # Limpar tree
            for item in self.tree_clientes.get_children():
                self.tree_clientes.delete(item)
            
            self._cache_detalhes = {}

            data_ref_dt = datetime.strptime(data_ref, '%d/%m/%Y')
            
            # Validar se é dia 5 ou 20
            if data_ref_dt.day not in [5, 20]:
                resposta = messagebox.askyesno(
                    "Data Inválida",
                    f"A data {data_ref} não é dia 05 ou 20.\n\n"
                    f"Deseja usar a data correta mais próxima?"
                )
                if not resposta:
                    return
                
                # Definir data mais próxima
                if data_ref_dt.day < 5:
                    data_ref_dt = data_ref_dt.replace(day=5)
                elif data_ref_dt.day < 20:
                    data_ref_dt = data_ref_dt.replace(day=20)
                else:
                    # Após dia 20, próximo é dia 5 do mês seguinte
                    if data_ref_dt.month == 12:
                        data_ref_dt = datetime(data_ref_dt.year + 1, 1, 5)
                    else:
                        data_ref_dt = datetime(data_ref_dt.year, data_ref_dt.month + 1, 5)
                
                self.data_ref_entry.set_date(data_ref_dt)
            
            wb_clientes = load_workbook(ARQUIVO_CLIENTES)
            ws_clientes = wb_clientes['Clientes']

            print(f"\n{'='*70}")
            print(f"BUSCANDO CLIENTES PARA {data_ref_dt.strftime('%d/%m/%Y')}")
            print(f"{'='*70}\n")

            # Verificar coluna Tipo Taxa
            headers = [cell.value for cell in ws_clientes[1]]
            idx_tipo_taxa = None
            for idx, header in enumerate(headers):
                if header and 'TIPO' in str(header).upper() and 'TAXA' in str(header).upper():
                    idx_tipo_taxa = idx
                    break

            clientes_encontrados = 0

            for row in ws_clientes.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                nome_cliente = row[0]
                
                # Verificar tipo de taxa
                tipo_taxa = row[idx_tipo_taxa] if idx_tipo_taxa and len(row) > idx_tipo_taxa else None
                
                # FILTRO: Só processar clientes com taxa Percentual
                if tipo_taxa and str(tipo_taxa).upper() != 'PERCENTUAL':
                    print(f"⏭️  {nome_cliente}: Taxa {tipo_taxa} - IGNORADO")
                    continue

                arquivo_cliente = PASTA_CLIENTES / f"{nome_cliente}.xlsx"
                if not os.path.exists(arquivo_cliente):
                    print(f"⚠️  {nome_cliente}: Arquivo não encontrado")
                    continue

                try:
                    print(f"\n📋 Processando: {nome_cliente}")

                    # Obter percentual
                    percentual = self._obter_percentual_cliente(nome_cliente)
                    if percentual == 0:
                        print(f"  ⚠️  Sem taxa percentual configurada - IGNORADO")
                        continue

                    # LÓGICA SIMPLES: Calcular compensação
                    resultado = self._calcular_taxa_com_compensacao_simples(
                        nome_cliente, data_ref_dt, percentual
                    )
                    
                    if resultado is None:
                        continue
                    
                    # Determinar status baseado na comparação de valores
                    valor_final = resultado['valor_final']
                    
                    if resultado['taxa_existente']:
                        valor_existente = resultado['taxa_existente_valor']
                        
                        # Verificar se valores são praticamente iguais (diferença < R$ 0,10)
                        diferenca = abs(valor_final - valor_existente)
                        
                        if diferenca < 0.10:
                            # Valores são iguais - taxa está OK!
                            status = f"✅ Processado - R$ {valor_final:,.2f}"
                            tag = 'processado'
                            print(f"  ✅ Taxa existente OK (diferença: R$ {diferenca:.2f})")
                        else:
                            # Valores diferentes - precisa recalcular
                            status = f"🔄 Recalcular (R$ {valor_existente:,.2f} → R$ {valor_final:,.2f})"
                            tag = 'recalcular'
                            print(f"  ⚠️  Taxa desatualizada (diferença: R$ {diferenca:.2f})")
                    else:
                        status = "➕ Nova taxa"
                        tag = 'nova'
                    
                    # Salvar detalhes no cache
                    self._cache_detalhes[nome_cliente] = resultado
                    
                    # Adicionar na tree
                    self.tree_clientes.insert('', 'end', values=(
                        nome_cliente,
                        self._formatar_moeda(resultado['base_atual']),
                        f"{percentual:.1f}%",
                        self._formatar_moeda(resultado['taxa_quinzena']),
                        self._formatar_moeda_com_sinal(resultado['compensacao']),
                        self._formatar_moeda(resultado['valor_final']),
                        status
                    ), tags=(tag,))
                    
                    # Configurar cores das tags
                    self.tree_clientes.tag_configure('nova', background='#fff9e6')
                    self.tree_clientes.tag_configure('recalcular', background='#fff0f0')
                    
                    clientes_encontrados += 1
                    
                    print(f"  ✅ Base Atual: R$ {resultado['base_atual']:.2f}")
                    print(f"     Taxa Quinzena: R$ {resultado['taxa_quinzena']:.2f}")
                    print(f"     Compensação: R$ {resultado['compensacao']:+.2f}")
                    print(f"     Valor Final: R$ {resultado['valor_final']:.2f}")
                    print(f"     Status: {status}")

                except Exception as e:
                    print(f"  ❌ Erro: {str(e)}")
                    import traceback
                    print(traceback.format_exc())
                    continue

            wb_clientes.close()
            
            print(f"\n{'='*70}")
            print(f"BUSCA CONCLUÍDA: {clientes_encontrados} cliente(s) encontrado(s)")
            print(f"{'='*70}\n")
            
            if clientes_encontrados == 0:
                messagebox.showinfo("Aviso", 
                    "Nenhum cliente encontrado.\n\n"
                    "Verifique se:\n"
                    "• Coluna 'Tipo Taxa' = 'Percentual'\n"
                    "• Existem lançamentos ativos na data")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
            import traceback
            print(traceback.format_exc())
            if 'wb_clientes' in locals():
                wb_clientes.close()

    def _calcular_taxa_com_compensacao_simples(self, cliente, data_atual, percentual):
        """
        LÓGICA SIMPLES E CORRETA:
        1. Somar TODAS as bases históricas (exceto tipo 7, exceto data atual)
        2. Somar TODAS as taxas pagas (tipo 7)
        3. Calcular diferença = Compensação
        4. Adicionar taxa da quinzena atual
        """
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            
            if 'DATA_REL' not in df.columns:
                df.columns.values[0] = 'DATA_REL'
            
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], format='%d/%m/%Y', errors='coerce')
            tem_status = 'STATUS' in df.columns
            
            # 1. BASE DA QUINZENA ATUAL
            df_atual = df[df['DATA_REL'].dt.date == data_atual.date()]
            
            if tem_status:
                base_atual = df_atual[
                    (df_atual['TP_DESP'] != 7) & 
                    (df_atual['STATUS'] == 'ATIVO')
                ]['VALOR'].apply(lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0).sum()
            else:
                base_atual = df_atual[
                    df_atual['TP_DESP'] != 7
                ]['VALOR'].apply(lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0).sum()
            
            if base_atual == 0:
                print(f"  ⚠️  Base zero - IGNORADO")
                return None
            
            taxa_quinzena = base_atual * (percentual / 100)
            
            # 2. VERIFICAR SE JÁ TEM TAXA LANÇADA NA DATA ATUAL
            # IMPORTANTE: Considera apenas taxas ATIVAS (EXCLUIDO não conta!)
            if tem_status:
                taxas_ativas_df = df_atual[
                    (df_atual['TP_DESP'] == 7) & 
                    (df_atual['STATUS'].str.upper() == 'ATIVO')  # Garante comparação case-insensitive
                ]
                
                # Debug: Mostrar todas as taxas encontradas
                todas_taxas = df_atual[df_atual['TP_DESP'] == 7]
                if len(todas_taxas) > 0:
                    print(f"  🔍 DEBUG: Taxas tipo 7 encontradas na data:")
                    for idx, row in todas_taxas.iterrows():
                        status_taxa = row.get('STATUS', 'N/A')
                        valor_taxa = row.get('VALOR', 0)
                        print(f"     Status: {status_taxa} | Valor: R$ {valor_taxa}")
                
                taxa_existente_valor = taxas_ativas_df['VALOR'].apply(
                    lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0
                ).sum()
            else:
                # Sem coluna STATUS, considera todas (fallback)
                taxas_ativas_df = df_atual[df_atual['TP_DESP'] == 7]
                
                taxa_existente_valor = taxas_ativas_df['VALOR'].apply(
                    lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0
                ).sum()
            
            tem_taxa_existente = taxa_existente_valor > 0
            
            if tem_taxa_existente:
                print(f"  ⚠️  Taxa ATIVA encontrada: R$ {taxa_existente_valor:.2f}")
                print(f"     Quantidade de taxas ATIVAS: {len(taxas_ativas_df)}")
            else:
                print(f"  ✅ Nenhuma taxa ATIVA encontrada")
            
            # 3. CALCULAR COMPENSAÇÃO HISTÓRICA
            # IMPORTANTE:
            # - Bases: até ANTES da data atual (< data_atual)
            # - Taxas pagas: até ANTES da data atual (< data_atual)
            #   Se tem taxa existente, ela será EXCLUÍDA e RECRIADA
            df_historico_bases = df[df['DATA_REL'].dt.date < data_atual.date()]
            df_historico_taxas = df[df['DATA_REL'].dt.date < data_atual.date()]  # < SÓ anteriores
            
            if tem_status:
                # Filtrar: TP_DESP != 7 E STATUS = ATIVO
                df_bases_filtrado = df_historico_bases[
                    (df_historico_bases['TP_DESP'] != 7) & 
                    (df_historico_bases['STATUS'] == 'ATIVO')
                ]
                
                # VERIFICAÇÃO EXTRA DE SEGURANÇA
                if 7 in df_bases_filtrado['TP_DESP'].values:
                    print(f"  ⚠️  ALERTA: Tipo 7 detectado após filtragem!")
                    # Forçar exclusão de tipo 7
                    df_bases_filtrado = df_bases_filtrado[df_bases_filtrado['TP_DESP'] != 7]
                
                total_bases_historicas = df_bases_filtrado['VALOR'].apply(
                    lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0
                ).sum()
                
                # Taxas pagas: TP_DESP == 7 E STATUS = ATIVO (SÓ ANTERIORES)
                total_taxas_pagas = df_historico_taxas[
                    (df_historico_taxas['TP_DESP'] == 7) & 
                    (df_historico_taxas['STATUS'] == 'ATIVO')
                ]['VALOR'].apply(lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0).sum()
            else:
                # Sem coluna STATUS
                # Filtrar: TP_DESP != 7
                df_bases_filtrado = df_historico_bases[df_historico_bases['TP_DESP'] != 7]
                
                # VERIFICAÇÃO EXTRA DE SEGURANÇA
                if 7 in df_bases_filtrado['TP_DESP'].values:
                    print(f"  ⚠️  ALERTA: Tipo 7 detectado após filtragem!")
                    # Forçar exclusão de tipo 7
                    df_bases_filtrado = df_bases_filtrado[df_bases_filtrado['TP_DESP'] != 7]
                
                total_bases_historicas = df_bases_filtrado['VALOR'].apply(
                    lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0
                ).sum()
                
                # Taxas pagas: TP_DESP == 7 (SÓ ANTERIORES)
                total_taxas_pagas = df_historico_taxas[
                    df_historico_taxas['TP_DESP'] == 7
                ]['VALOR'].apply(lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else 0).sum()
            
            # DEBUG: Mostrar valores calculados
            print(f"  🔍 DEBUG HISTÓRICO:")
            print(f"     Bases: < {data_atual.strftime('%d/%m/%Y')} (só anteriores)")
            print(f"     Taxas: < {data_atual.strftime('%d/%m/%Y')} (só anteriores)")
            print(f"     Total registros bases: {len(df_historico_bases)}")
            print(f"     Total registros taxas: {len(df_historico_taxas)}")
            print(f"     Registros tipo 7 (taxas): {len(df_historico_taxas[df_historico_taxas['TP_DESP'] == 7])}")
            print(f"     Registros NÃO tipo 7 (bases): {len(df_bases_filtrado)}")
            print(f"     Total bases (SEM tipo 7): R$ {total_bases_historicas:,.2f}")
            print(f"     Total taxas pagas (COM tipo 7): R$ {total_taxas_pagas:,.2f}")
            
            # Calcular quanto DEVERIA ter sido pago
            total_devido_historico = total_bases_historicas * (percentual / 100)
            
            # Compensação = Diferença entre devido e pago
            compensacao = total_devido_historico - total_taxas_pagas
            
            # 4. VALOR FINAL
            valor_final = taxa_quinzena + compensacao
            
            return {
                'base_atual': base_atual,
                'taxa_quinzena': taxa_quinzena,
                'compensacao': compensacao,
                'valor_final': valor_final,
                'taxa_existente': tem_taxa_existente,
                'taxa_existente_valor': taxa_existente_valor if tem_taxa_existente else 0,
                'total_bases_historicas': total_bases_historicas,
                'total_taxas_pagas': total_taxas_pagas,
                'total_devido_historico': total_devido_historico
            }
            
        except Exception as e:
            print(f"  ❌ Erro ao calcular: {e}")
            import traceback
            print(traceback.format_exc())
            return None

    def processar_clientes_selecionados(self):
        """Processa os clientes selecionados"""
        selecionados = self.tree_clientes.selection()
        
        if not selecionados:
            messagebox.showwarning("Aviso", "Selecione pelo menos um cliente!")
            return

        data_ref = datetime.strptime(self.data_ref_entry.get(), '%d/%m/%Y')
        processados = []
        erros = []

        for item in selecionados:
            valores = self.tree_clientes.item(item)['values']
            cliente = valores[0]

            try:
                # Buscar detalhes do cache
                if cliente not in self._cache_detalhes:
                    raise Exception("Detalhes não encontrados. Refaça a busca.")
                
                resultado = self._cache_detalhes[cliente]
                percentual = self._obter_percentual_cliente(cliente)
                
                # Preparar mensagem de confirmação
                mensagem = self._preparar_mensagem_confirmacao(
                    cliente, data_ref, resultado, percentual
                )

                # Confirmar com usuário
                if messagebox.askyesno("Confirmar Lançamento", mensagem):
                    # EXCLUIR taxa existente se houver
                    if resultado['taxa_existente']:
                        self._excluir_taxa_existente(cliente, data_ref)
                        print(f"  🗑️  Taxa anterior excluída")
                    
                    # CRIAR nova taxa
                    self._criar_lancamento_taxa(
                        cliente, data_ref, resultado['valor_final'], resultado['compensacao']
                    )
                    
                    processados.append(f"✅ {cliente} - R$ {resultado['valor_final']:.2f}")
                    print(f"✅ {cliente} processado com sucesso!")
                    
                    # ATUALIZAR STATUS NA TREE (mantém na lista)
                    self.tree_clientes.item(item, values=(
                        cliente,
                        valores[1],  # Base
                        valores[2],  # Taxa %
                        valores[3],  # Taxa Quinzena
                        valores[4],  # Compensação
                        valores[5],  # Valor Final
                        f"✅ Processado - R$ {resultado['valor_final']:,.2f}"
                    ), tags=('processado',))
                    
                else:
                    print(f"⏭️  {cliente} ignorado pelo usuário")

            except Exception as e:
                erro_msg = f"❌ {cliente} - {str(e)}"
                erros.append(erro_msg)
                print(erro_msg)
                import traceback
                print(traceback.format_exc())

        # Configurar tag de processado (cor verde claro)
        self.tree_clientes.tag_configure('processado', background='#d4edda')
        
        # Mostrar resultado
        self._mostrar_resultado_processamento(processados, erros)

    def _excluir_taxa_existente(self, cliente, data_ref):
        """Exclui taxa existente marcando como EXCLUIDO"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb["Dados"]
            
            print(f"\n  🔍 DEBUG EXCLUSÃO:")
            print(f"     Data referência: {data_ref.strftime('%d/%m/%Y')}")
            
            # Buscar taxas tipo 7 na data
            taxas_encontradas = 0
            taxas_excluidas = 0
            
            for row_idx in range(2, ws.max_row + 1):
                data_cel = ws.cell(row=row_idx, column=1).value
                tipo_cel = ws.cell(row=row_idx, column=2).value
                status_cel = ws.cell(row=row_idx, column=14).value
                valor_cel = ws.cell(row=row_idx, column=7).value
                
                # Verificar se é a data e tipo corretos
                if isinstance(data_cel, datetime):
                    if data_cel.date() == data_ref.date() and tipo_cel == 7:
                        taxas_encontradas += 1
                        print(f"     Taxa encontrada linha {row_idx}: Status={status_cel}, Valor={valor_cel}")
                        
                        if status_cel == 'ATIVO':
                            # Marcar como EXCLUIDO
                            ws.cell(row=row_idx, column=14, value='EXCLUIDO')
                            
                            # Adicionar observação sobre exclusão
                            obs_atual = ws.cell(row=row_idx, column=13).value or ''
                            nova_obs = f"{obs_atual} | EXCLUÍDO: Recálculo por mudança na base - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                            ws.cell(row=row_idx, column=13, value=nova_obs)
                            
                            # Atualizar histórico
                            hist_atual = ws.cell(row=row_idx, column=16).value or ''
                            novo_hist = f"{hist_atual} | EXCLUÍDO EM: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                            ws.cell(row=row_idx, column=16, value=novo_hist)
                            
                            taxas_excluidas += 1
                            print(f"     ✅ Taxa linha {row_idx} marcada como EXCLUIDO")
                        else:
                            print(f"     ⏭️  Taxa linha {row_idx} já estava {status_cel} - ignorada")
            
            print(f"     Total taxas encontradas: {taxas_encontradas}")
            print(f"     Total taxas excluídas: {taxas_excluidas}\n")
            
            wb.save(arquivo_cliente)
            wb.close()
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Erro ao excluir taxa: {str(e)}")

    # ============================================
    # MÉTODOS AUXILIARES
    # ============================================

    def _obter_percentual_cliente(self, cliente):
        """Obtém percentual de taxa do cliente"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente, data_only=True)
            
            if 'Contratos_ADM' not in wb.sheetnames:
                wb.close()
                return 0
            
            ws = wb['Contratos_ADM']
            
            contratos_ativos = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not row or len(row) < 4:
                    continue
                
                status = row[3]
                num_contrato = row[0]
                
                if status == 'ATIVO' and num_contrato:
                    contratos_ativos.append(row_idx)
            
            if not contratos_ativos:
                wb.close()
                return 0
            
            percentual_total = 0
            
            for linha_contrato in contratos_ativos:
                for offset in range(1, 11):
                    linha_atual = linha_contrato + offset
                    
                    if linha_atual > ws.max_row:
                        break
                    
                    row = list(ws.iter_rows(min_row=linha_atual, max_row=linha_atual, values_only=True))[0]
                    
                    if not row or len(row) < 11:
                        continue
                    
                    tipo = row[9]
                    valor_taxa = row[10]
                    
                    if tipo == 'Percentual' and valor_taxa:
                        try:
                            percentual_str = str(valor_taxa).replace('%', '').replace(',', '.').strip()
                            percentual = float(percentual_str)
                            percentual_total += percentual
                            break
                        except (ValueError, TypeError):
                            continue
                    
                    status_linha = row[3] if len(row) > 3 else None
                    if status_linha in ['ATIVO', 'INATIVO']:
                        break
            
            wb.close()
            return percentual_total
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            return 0

    def _criar_lancamento_taxa(self, cliente, data_quinzena, valor_taxa, compensacao):
        """Cria lançamento da taxa na planilha do cliente"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb["Dados"]
            
            # Obter próximo ID
            max_id = 0
            for row in range(2, ws.max_row + 1):
                id_val = ws.cell(row=row, column=15).value
                if id_val:
                    try:
                        max_id = max(max_id, int(float(id_val)))
                    except:
                        pass
            
            novo_id = max_id + 1
            proxima_linha = ws.max_row + 1
            
            # Obter administrador
            administrador = self._obter_administrador_principal(cliente)
            
            # Preparar dados
            quinzena = "1ª" if data_quinzena.day == 5 else "2ª"
            referencia = f"TAXA ADM - {quinzena} QUINZ. {data_quinzena.strftime('%m/%Y')}"
            
            # Observação com compensação
            observacao = ""
            if abs(compensacao) > 0.01:
                if compensacao > 0:
                    observacao = f"Inclui compensação de R$ {compensacao:.2f}"
                else:
                    observacao = f"Com desconto de R$ {abs(compensacao):.2f}"
            
            # Gravar lançamento
            ws.cell(row=proxima_linha, column=1, value=data_quinzena)
            ws.cell(row=proxima_linha, column=2, value=7)
            ws.cell(row=proxima_linha, column=3, value=administrador['cnpj_cpf'])
            ws.cell(row=proxima_linha, column=4, value=administrador['nome'])
            ws.cell(row=proxima_linha, column=5, value=referencia)
            ws.cell(row=proxima_linha, column=6, value='')
            ws.cell(row=proxima_linha, column=7, value=valor_taxa)
            ws.cell(row=proxima_linha, column=8, value=1)
            ws.cell(row=proxima_linha, column=9, value=valor_taxa)
            ws.cell(row=proxima_linha, column=10, value=data_quinzena)
            ws.cell(row=proxima_linha, column=11, value='TAX')
            ws.cell(row=proxima_linha, column=12, value=administrador.get('dados_bancarios', 'PIX'))
            ws.cell(row=proxima_linha, column=13, value=observacao)
            ws.cell(row=proxima_linha, column=14, value='ATIVO')
            ws.cell(row=proxima_linha, column=15, value=novo_id)
            ws.cell(row=proxima_linha, column=16, value=f"Criado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            
            # Formatar
            ws.cell(row=proxima_linha, column=1).number_format = 'DD/MM/YYYY'
            ws.cell(row=proxima_linha, column=7).number_format = '#,##0.00'
            ws.cell(row=proxima_linha, column=9).number_format = '#,##0.00'
            ws.cell(row=proxima_linha, column=10).number_format = 'DD/MM/YYYY'
            
            wb.save(arquivo_cliente)
            wb.close()
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Erro ao criar lançamento: {str(e)}")

    def _obter_administrador_principal(self, cliente):
        """Obtém dados do administrador principal"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente, data_only=True)
            
            if 'Contratos_ADM' not in wb.sheetnames:
                wb.close()
                return {'cnpj_cpf': '00000000000', 'nome': 'ADMINISTRADOR', 'dados_bancarios': 'PIX'}
            
            ws = wb['Contratos_ADM']
            
            contratos_ativos = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not row or len(row) < 4:
                    continue
                
                status = row[3]
                num_contrato = row[0]
                
                if status == 'ATIVO' and num_contrato:
                    contratos_ativos.append(row_idx)
            
            for linha_contrato in contratos_ativos:
                for offset in range(1, 11):
                    linha_atual = linha_contrato + offset
                    
                    if linha_atual > ws.max_row:
                        break
                    
                    row = list(ws.iter_rows(min_row=linha_atual, max_row=linha_atual, values_only=True))[0]
                    
                    if not row or len(row) < 11:
                        continue
                    
                    tipo = row[9]
                    
                    if tipo == 'Percentual':
                        cnpj_cpf = row[7] if len(row) > 7 else None
                        nome = row[8] if len(row) > 8 else None
                        
                        if cnpj_cpf and nome:
                            administrador = {
                                'cnpj_cpf': cnpj_cpf,
                                'nome': nome,
                                'dados_bancarios': buscar_dados_bancarios_fornecedor(cnpj_cpf) or 'PIX'
                            }
                            wb.close()
                            return administrador
                        break
                    
                    status_linha = row[3] if len(row) > 3 else None
                    if status_linha in ['ATIVO', 'INATIVO']:
                        break
            
            wb.close()
            return {'cnpj_cpf': '00000000000', 'nome': 'ADMINISTRADOR', 'dados_bancarios': 'PIX'}
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            return {'cnpj_cpf': '00000000000', 'nome': 'ADMINISTRADOR', 'dados_bancarios': 'PIX'}

    # ============================================
    # INTERFACE
    # ============================================

    def _preparar_mensagem_confirmacao(self, cliente, data_ref, resultado, percentual):
        """Prepara mensagem de confirmação"""
        mensagem = f"📊 FINALIZAÇÃO DE QUINZENA\n"
        mensagem += f"{'='*50}\n\n"
        mensagem += f"Cliente: {cliente}\n"
        mensagem += f"Data: {data_ref.strftime('%d/%m/%Y')}\n\n"
        
        mensagem += f"📅 QUINZENA ATUAL:\n"
        mensagem += f"   Base: R$ {resultado['base_atual']:,.2f}\n"
        mensagem += f"   Taxa ({percentual}%): R$ {resultado['taxa_quinzena']:,.2f}\n\n"
        
        if abs(resultado['compensacao']) > 0.01:
            mensagem += f"⚖️  COMPENSAÇÃO HISTÓRICA:\n"
            mensagem += f"   Total bases anteriores: R$ {resultado['total_bases_historicas']:,.2f}\n"
            mensagem += f"   Total devido: R$ {resultado['total_devido_historico']:,.2f}\n"
            mensagem += f"   Total pago: R$ {resultado['total_taxas_pagas']:,.2f}\n"
            mensagem += f"   Diferença: R$ {resultado['compensacao']:+,.2f}\n\n"
        
        mensagem += f"💰 VALOR FINAL: R$ {resultado['valor_final']:,.2f}\n\n"
        
        if resultado['taxa_existente']:
            mensagem += f"⚠️  AÇÃO:\n"
            mensagem += f"   Taxa anterior (R$ {resultado['taxa_existente_valor']:,.2f}) será EXCLUÍDA\n"
            mensagem += f"   Nova taxa será criada\n"
            mensagem += f"   Motivo: Mudança na base de cálculo\n\n"
        
        mensagem += "Confirma o lançamento?"
        
        return mensagem

    def ver_detalhes_compensacao(self):
        """Mostra detalhes da compensação"""
        selecionados = self.tree_clientes.selection()
        
        if not selecionados:
            messagebox.showinfo("Info", "Selecione um cliente para ver detalhes")
            return
        
        cliente = self.tree_clientes.item(selecionados[0])['values'][0]
        
        if cliente not in self._cache_detalhes:
            messagebox.showinfo("Info", "Detalhes não disponíveis. Refaça a busca.")
            return
        
        resultado = self._cache_detalhes[cliente]
        
        # Janela de detalhes
        janela = tk.Toplevel(self.root)
        janela.title(f"Detalhes da Compensação - {cliente}")
        janela.geometry("600x650")
        
        frame = ttk.Frame(janela, padding="20")
        frame.pack(fill='both', expand=True)
        
        # Texto com detalhes
        texto = tk.Text(frame, wrap=tk.WORD, font=('Arial', 10))
        texto.pack(fill='both', expand=True)
        
        detalhes = f"""
📊 DETALHAMENTO DA COMPENSAÇÃO

Cliente: {cliente}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📅 QUINZENA ATUAL:
   Base: R$ {resultado['base_atual']:,.2f}
   Taxa: R$ {resultado['taxa_quinzena']:,.2f}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚖️  HISTÓRICO (períodos anteriores):

   Total de bases: R$ {resultado['total_bases_historicas']:,.2f}
   
   Total que deveria ter sido pago:
   R$ {resultado['total_devido_historico']:,.2f}
   
   Total que foi efetivamente pago:
   R$ {resultado['total_taxas_pagas']:,.2f}
   
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   
   COMPENSAÇÃO: R$ {resultado['compensacao']:+,.2f}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💰 VALOR FINAL:
   Taxa quinzena: R$ {resultado['taxa_quinzena']:,.2f}
   Compensação: R$ {resultado['compensacao']:+,.2f}
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   TOTAL: R$ {resultado['valor_final']:,.2f}
"""
        
        texto.insert('1.0', detalhes)
        texto.config(state='disabled')
        
        ttk.Button(frame, text="Fechar", command=janela.destroy).pack(pady=10)

    def _mostrar_resultado_processamento(self, processados, erros):
        """Mostra resultado do processamento"""
        janela = tk.Toplevel(self.root)
        janela.title("Resultado do Processamento")
        janela.geometry("600x400")
        
        frame = ttk.Frame(janela, padding="10")
        frame.pack(fill='both', expand=True)
        
        if processados:
            frame_proc = ttk.LabelFrame(frame, text="Processados com Sucesso", padding="5")
            frame_proc.pack(fill='both', expand=True, pady=5)
            
            texto_proc = tk.Text(frame_proc, height=10)
            texto_proc.pack(fill='both', expand=True)
            
            for msg in processados:
                texto_proc.insert(tk.END, f"{msg}\n")
            texto_proc.config(state='disabled')
        
        if erros:
            frame_erros = ttk.LabelFrame(frame, text="Erros", padding="5")
            frame_erros.pack(fill='both', expand=True, pady=5)
            
            texto_erros = tk.Text(frame_erros, height=10)
            texto_erros.pack(fill='both', expand=True)
            
            for msg in erros:
                texto_erros.insert(tk.END, f"{msg}\n")
            texto_erros.config(state='disabled')
        
        ttk.Button(frame, text="Fechar", command=janela.destroy).pack(pady=10)

    def _formatar_moeda(self, valor):
        """Formata valor em moeda brasileira"""
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _formatar_moeda_com_sinal(self, valor):
        """Formata valor com sinal + ou -"""
        sinal = "+" if valor >= 0 else ""
        return f"{sinal}R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def voltar_menu(self):
        """Volta ao menu principal"""
        self.root.destroy()
        if self.parent:
            self.parent.deiconify()
            self.parent.lift()


# ============================================
# EXECUÇÃO
# ============================================

if __name__ == "__main__":
    app = FinalizacaoQuinzena()
    app.run()