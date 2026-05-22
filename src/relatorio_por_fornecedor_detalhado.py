import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _add_project_root():
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

_add_project_root()

try:
    from src.config.config import PASTA_CLIENTES, ARQUIVO_FORNECEDORES
except ImportError:
    PASTA_CLIENTES = Path(".")
    ARQUIVO_FORNECEDORES = Path(".")

try:
    from src.config.window_config import configurar_janela
except ImportError:
    def configurar_janela(janela, titulo, largura=1100, altura=900):
        janela.title(titulo)
        sw = janela.winfo_screenwidth()
        sh = janela.winfo_screenheight()
        x = (sw - largura) // 2
        y = (sh - altura) // 2
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        janela.resizable(True, True)


CATEGORIAS = {
    'ADM': 'Administrativo',
    'DIV': 'Diversos',
    'LOC': 'Locação',
    'MAT': 'Material',
    'MO': 'Mão-de-Obra',
    'SERV': 'Serviços',
    'TAX': 'Taxa Administração',
    'TP': 'Tarifas/Tributos Públicos',
}


def formatar_moeda_br(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return "R$ 0,00"


class RelatorioPorFornecedorDetalhado:
    """Relatório detalhado por fornecedor com filtros de período, categoria e vínculo."""

    def __init__(self, parent=None):
        self.parent = parent
        if parent:
            self.root = tk.Toplevel(parent)
            self.root.protocol("WM_DELETE_WINDOW", self.voltar_menu)
        else:
            self.root = tk.Tk()
            self.root.protocol("WM_DELETE_WINDOW", self.fechar_aplicacao)

        configurar_janela(self.root, "Relatório Detalhado por Fornecedor", 1150, 920)

        self.fornecedor_selecionado = None
        self.fornecedores_disponiveis = []   # list of dicts {nome, vinculo, cnpj}
        self._df_master = None               # cache do arquivo de fornecedores
        self.dados_carregados = False
        self._registros = []                 # registros brutos após carregar

        self.periodo_inicio = datetime.now() - timedelta(days=180)
        self.periodo_fim = datetime.now()

        self._setup_gui()

    # ------------------------------------------------------------------ #
    #  CONSTRUÇÃO DA INTERFACE                                             #
    # ------------------------------------------------------------------ #

    def _setup_gui(self):
        main = ttk.Frame(self.root, padding=8)
        main.pack(fill='both', expand=True)
        main.rowconfigure(1, weight=1)
        main.columnconfigure(0, weight=1)

        self.frame_filtros = ttk.LabelFrame(main, text="Filtros")
        self.frame_filtros.grid(row=0, column=0, sticky='ew', pady=4)

        self.frame_resultados = ttk.LabelFrame(main, text="Resultados")
        self.frame_resultados.grid(row=1, column=0, sticky='nsew', pady=4)

        frame_btns = ttk.Frame(main)
        frame_btns.grid(row=2, column=0, sticky='ew', pady=4)

        self._setup_filtros()
        self._setup_resultados()
        self._setup_botoes(frame_btns)

        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 11, 'bold'), padding=(10, 5))

    # ── Filtros ─────────────────────────────────────────────────────────

    def _setup_filtros(self):
        p = self.frame_filtros

        # --- linha 1: busca de fornecedor ---
        row1 = ttk.Frame(p)
        row1.pack(fill='x', padx=10, pady=(8, 2))

        ttk.Label(row1, text="Fornecedor:", font=('Arial', 10, 'bold'), width=12).pack(side='left')

        self.var_busca = tk.StringVar()
        self._entry_busca = ttk.Entry(row1, textvariable=self.var_busca, width=45, font=('Arial', 10))
        self._entry_busca.pack(side='left', padx=5)
        self._entry_busca.bind('<KeyRelease>', self._filtrar_lista)
        self._entry_busca.bind('<Return>', self._selecionar_primeiro)

        ttk.Button(row1, text="Carregar Lista", command=self.carregar_fornecedores).pack(side='left', padx=5)
        ttk.Button(row1, text="Limpar", command=self._limpar_selecao).pack(side='left', padx=3)

        self.lbl_sel = ttk.Label(row1, text="(nenhum selecionado – deixe em branco para todos)",
                                 foreground='gray', font=('Arial', 9))
        self.lbl_sel.pack(side='left', padx=8)

        # lista de resultados da busca
        row1b = ttk.Frame(p)
        row1b.pack(fill='x', padx=10, pady=2)

        self._tree_forn = ttk.Treeview(row1b, columns=('nome', 'vinculo'), show='headings', height=5)
        self._tree_forn.heading('nome', text='Nome do Fornecedor')
        self._tree_forn.heading('vinculo', text='Vínculo')
        self._tree_forn.column('nome', width=520)
        self._tree_forn.column('vinculo', width=180)
        self._scroll_forn = ttk.Scrollbar(row1b, orient='vertical', command=self._tree_forn.yview)
        self._tree_forn.configure(yscrollcommand=self._scroll_forn.set)
        self._tree_forn.bind('<Double-1>', self._selecionar_da_lista)
        self._tree_forn.bind('<Return>', self._selecionar_da_lista)
        # inicialmente ocultos
        self._tree_forn_visible = False

        # --- linha 2: período ---
        row2 = ttk.Frame(p)
        row2.pack(fill='x', padx=10, pady=5)

        ttk.Label(row2, text="Período:", font=('Arial', 10, 'bold'), width=12).pack(side='left')

        self.var_periodo = tk.StringVar(value='Últimos 180 dias')
        ttk.Combobox(
            row2, textvariable=self.var_periodo, width=20, state='readonly',
            values=['Últimos 30 dias', 'Últimos 90 dias', 'Últimos 180 dias',
                    'Último ano', 'Todo o período', 'Personalizado']
        ).pack(side='left', padx=5)
        self.var_periodo.trace_add('write', lambda *_: self._atualizar_periodo())

        self._frame_datas = ttk.Frame(row2)
        ttk.Label(self._frame_datas, text="De:").pack(side='left', padx=(5, 2))
        try:
            from tkcalendar import DateEntry
            self._dt_ini = DateEntry(self._frame_datas, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
            self._dt_ini.pack(side='left', padx=3)
            ttk.Label(self._frame_datas, text="Até:").pack(side='left', padx=(5, 2))
            self._dt_fim = DateEntry(self._frame_datas, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
            self._dt_fim.pack(side='left', padx=3)
            self._usar_datentry = True
        except ImportError:
            self._var_ini = tk.StringVar(value=(datetime.now() - timedelta(days=180)).strftime('%d/%m/%Y'))
            ttk.Entry(self._frame_datas, textvariable=self._var_ini, width=12).pack(side='left', padx=3)
            ttk.Label(self._frame_datas, text="Até:").pack(side='left', padx=(5, 2))
            self._var_fim = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
            ttk.Entry(self._frame_datas, textvariable=self._var_fim, width=12).pack(side='left', padx=3)
            self._usar_datentry = False

        self._atualizar_periodo()

        # --- linha 3: categorias ---
        row3 = ttk.Frame(p)
        row3.pack(fill='x', padx=10, pady=5)

        ttk.Label(row3, text="Categoria:", font=('Arial', 10, 'bold'), width=12).pack(side='left')

        self._var_todas_cat = tk.BooleanVar(value=True)
        ttk.Checkbutton(row3, text="Todas", variable=self._var_todas_cat,
                        command=self._toggle_todas_cat).pack(side='left', padx=4)

        self._vars_cat = {}
        for cod, nome in CATEGORIAS.items():
            v = tk.BooleanVar(value=True)
            self._vars_cat[cod] = v
            ttk.Checkbutton(
                row3, text=f"{cod} – {nome}", variable=v,
                command=self._sincronizar_todas_cat
            ).pack(side='left', padx=4)

        # --- linha 4: vínculo + botão ---
        row4 = ttk.Frame(p)
        row4.pack(fill='x', padx=10, pady=(5, 10))

        ttk.Label(row4, text="Vínculo:", font=('Arial', 10, 'bold'), width=12).pack(side='left')

        self.var_vinculo = tk.StringVar(value='Todos')
        self._combo_vinculo = ttk.Combobox(row4, textvariable=self.var_vinculo, width=35,
                                           state='readonly', values=['Todos'])
        self._combo_vinculo.pack(side='left', padx=5)

        ttk.Label(row4, text="(carregado automaticamente ao buscar fornecedores)",
                  foreground='gray', font=('Arial', 9)).pack(side='left', padx=5)

        ttk.Button(row4, text="Gerar Relatório", command=self.gerar_relatorio,
                   style='Big.TButton').pack(side='right', padx=10)

    # ── Resultados ──────────────────────────────────────────────────────

    def _setup_resultados(self):
        self._notebook = ttk.Notebook(self.frame_resultados)
        self._notebook.pack(fill='both', expand=True, padx=5, pady=5)

        self._aba_cliente = ttk.Frame(self._notebook)
        self._aba_registros = ttk.Frame(self._notebook)
        self._aba_categoria = ttk.Frame(self._notebook)

        self._notebook.add(self._aba_cliente, text='Por Cliente')
        self._notebook.add(self._aba_registros, text='Todos os Registros')
        self._notebook.add(self._aba_categoria, text='Por Categoria')

        self._setup_aba_cliente()
        self._setup_aba_registros()
        self._setup_aba_categoria()

    def _setup_aba_cliente(self):
        # Info
        fi = ttk.Frame(self._aba_cliente)
        fi.pack(fill='x', padx=5, pady=5)
        self._lbl_info_cliente = ttk.Label(fi, text="", font=('Arial', 11, 'bold'), foreground='#0056b3')
        self._lbl_info_cliente.pack(side='left', padx=5)
        self._lbl_total_cliente = ttk.Label(fi, text="", font=('Arial', 11, 'bold'))
        self._lbl_total_cliente.pack(side='right', padx=10)

        # Treeview
        ft = ttk.Frame(self._aba_cliente)
        ft.pack(fill='both', expand=True, padx=5, pady=5)

        cols = ('pos', 'cliente', 'total', 'pct', 'qtd', 'categorias')
        self._tree_cliente = ttk.Treeview(ft, columns=cols, show='headings', height=18)
        self._tree_cliente.heading('pos', text='#')
        self._tree_cliente.heading('cliente', text='Cliente')
        self._tree_cliente.heading('total', text='Total Gasto')
        self._tree_cliente.heading('pct', text='% do Total')
        self._tree_cliente.heading('qtd', text='Registros')
        self._tree_cliente.heading('categorias', text='Categorias')
        self._tree_cliente.column('pos', width=40, anchor='center')
        self._tree_cliente.column('cliente', width=260)
        self._tree_cliente.column('total', width=140, anchor='e')
        self._tree_cliente.column('pct', width=90, anchor='center')
        self._tree_cliente.column('qtd', width=80, anchor='center')
        self._tree_cliente.column('categorias', width=200)

        sy = ttk.Scrollbar(ft, orient='vertical', command=self._tree_cliente.yview)
        sx = ttk.Scrollbar(ft, orient='horizontal', command=self._tree_cliente.xview)
        self._tree_cliente.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self._tree_cliente.pack(side='left', fill='both', expand=True)
        sy.pack(side='right', fill='y')
        sx.pack(side='bottom', fill='x')

        self._tree_cliente.bind('<Double-1>', self._drill_down_cliente)

    def _setup_aba_registros(self):
        # controles de ordenação
        fc = ttk.Frame(self._aba_registros)
        fc.pack(fill='x', padx=5, pady=5)
        ttk.Label(fc, text="Ordenar por:").pack(side='left', padx=5)
        self.var_ordem = tk.StringVar(value='Data (mais recente)')
        ttk.Combobox(fc, textvariable=self.var_ordem, state='readonly', width=22,
                     values=['Data (mais recente)', 'Data (mais antiga)', 'Cliente A-Z',
                             'Valor (maior)', 'Valor (menor)']).pack(side='left', padx=5)
        ttk.Button(fc, text="Aplicar Ordem", command=self._aplicar_ordem_registros).pack(side='left', padx=5)

        self._lbl_qtd_registros = ttk.Label(fc, text="", font=('Arial', 10))
        self._lbl_qtd_registros.pack(side='right', padx=10)

        # Treeview
        ft = ttk.Frame(self._aba_registros)
        ft.pack(fill='both', expand=True, padx=5, pady=5)

        cols = ('data', 'cliente', 'fornecedor', 'tp_desp', 'categoria', 'referencia',
                'dt_vencto', 'valor', 'observacao')
        self._tree_regs = ttk.Treeview(ft, columns=cols, show='headings', height=20)
        self._tree_regs.heading('data', text='Data')
        self._tree_regs.heading('cliente', text='Cliente')
        self._tree_regs.heading('fornecedor', text='Fornecedor')
        self._tree_regs.heading('tp_desp', text='Tp.Desp')
        self._tree_regs.heading('categoria', text='Categoria')
        self._tree_regs.heading('referencia', text='Referência')
        self._tree_regs.heading('dt_vencto', text='Vencimento')
        self._tree_regs.heading('valor', text='Valor')
        self._tree_regs.heading('observacao', text='Observação')

        self._tree_regs.column('data', width=90, anchor='center')
        self._tree_regs.column('cliente', width=180)
        self._tree_regs.column('fornecedor', width=180)
        self._tree_regs.column('tp_desp', width=60, anchor='center')
        self._tree_regs.column('categoria', width=70, anchor='center')
        self._tree_regs.column('referencia', width=230)
        self._tree_regs.column('dt_vencto', width=90, anchor='center')
        self._tree_regs.column('valor', width=110, anchor='e')
        self._tree_regs.column('observacao', width=150)

        sy = ttk.Scrollbar(ft, orient='vertical', command=self._tree_regs.yview)
        sx = ttk.Scrollbar(ft, orient='horizontal', command=self._tree_regs.xview)
        self._tree_regs.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self._tree_regs.pack(side='left', fill='both', expand=True)
        sy.pack(side='right', fill='y')
        sx.pack(side='bottom', fill='x')

    def _setup_aba_categoria(self):
        fi = ttk.Frame(self._aba_categoria)
        fi.pack(fill='x', padx=5, pady=5)
        self._lbl_info_cat = ttk.Label(fi, text="", font=('Arial', 11, 'bold'), foreground='#0056b3')
        self._lbl_info_cat.pack(side='left', padx=5)

        ft = ttk.Frame(self._aba_categoria)
        ft.pack(fill='both', expand=True, padx=5, pady=5)

        # colunas dinâmicas – criadas ao preencher
        self._tree_cat = ttk.Treeview(ft, show='headings', height=18)
        sy = ttk.Scrollbar(ft, orient='vertical', command=self._tree_cat.yview)
        sx = ttk.Scrollbar(ft, orient='horizontal', command=self._tree_cat.xview)
        self._tree_cat.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self._tree_cat.pack(side='left', fill='both', expand=True)
        sy.pack(side='right', fill='y')
        sx.pack(side='bottom', fill='x')

    # ── Botões ──────────────────────────────────────────────────────────

    def _setup_botoes(self, frame):
        ttk.Button(frame, text="Exportar para Excel",
                   command=self.exportar_excel).pack(side='left', padx=5)
        ttk.Button(frame, text="Voltar ao Menu",
                   command=self.voltar_menu).pack(side='right', padx=5)

    # ------------------------------------------------------------------ #
    #  LÓGICA DOS FILTROS                                                  #
    # ------------------------------------------------------------------ #

    def _atualizar_periodo(self, *_):
        sel = self.var_periodo.get()
        hoje = datetime.now()
        self._frame_datas.pack_forget()
        if sel == 'Últimos 30 dias':
            self.periodo_inicio, self.periodo_fim = hoje - timedelta(days=30), hoje
        elif sel == 'Últimos 90 dias':
            self.periodo_inicio, self.periodo_fim = hoje - timedelta(days=90), hoje
        elif sel == 'Últimos 180 dias':
            self.periodo_inicio, self.periodo_fim = hoje - timedelta(days=180), hoje
        elif sel == 'Último ano':
            self.periodo_inicio, self.periodo_fim = hoje - timedelta(days=365), hoje
        elif sel == 'Todo o período':
            self.periodo_inicio, self.periodo_fim = datetime(2000, 1, 1), hoje
        elif sel == 'Personalizado':
            self._frame_datas.pack(side='left')

    def _obter_periodo(self):
        if self.var_periodo.get() != 'Personalizado':
            return self.periodo_inicio, self.periodo_fim
        try:
            if self._usar_datentry:
                ini = datetime.strptime(self._dt_ini.get(), '%d/%m/%Y')
                fim = datetime.strptime(self._dt_fim.get(), '%d/%m/%Y')
            else:
                ini = datetime.strptime(self._var_ini.get(), '%d/%m/%Y')
                fim = datetime.strptime(self._var_fim.get(), '%d/%m/%Y')
            return ini, fim
        except ValueError:
            messagebox.showerror("Erro", "Datas inválidas no período personalizado.")
            raise

    def _categorias_selecionadas(self):
        if self._var_todas_cat.get():
            return set(CATEGORIAS.keys()) | {''}
        selecionadas = {cod for cod, var in self._vars_cat.items() if var.get()}
        return selecionadas

    def _toggle_todas_cat(self):
        valor = self._var_todas_cat.get()
        for v in self._vars_cat.values():
            v.set(valor)

    def _sincronizar_todas_cat(self):
        todas = all(v.get() for v in self._vars_cat.values())
        self._var_todas_cat.set(todas)

    # ------------------------------------------------------------------ #
    #  CARREGAMENTO DE FORNECEDORES                                        #
    # ------------------------------------------------------------------ #

    def _carregar_master(self):
        """Lê base_fornecedores.xlsx e retorna dict nome_upper → vínculo."""
        if self._df_master is not None:
            return self._df_master
        try:
            df = pd.read_excel(str(ARQUIVO_FORNECEDORES))
            # descobrir coluna de nome
            nome_col = next(
                (c for c in df.columns if str(c).upper() in ('NOME', 'RAZAO_SOCIAL', 'RAZÃO SOCIAL')),
                df.columns[3] if len(df.columns) > 3 else None
            )
            # descobrir coluna de vínculo (pos 13 → índice 13)
            vinculo_col = next(
                (c for c in df.columns if str(c).upper() in ('VÍNCULO', 'VINCULO', 'VINCULO_')),
                df.columns[13] if len(df.columns) > 13 else None
            )
            if nome_col is None or vinculo_col is None:
                return {}
            result = {}
            for _, row in df.iterrows():
                n = str(row[nome_col]).strip().upper() if pd.notna(row[nome_col]) else ''
                v = str(row[vinculo_col]).strip() if pd.notna(row[vinculo_col]) else ''
                if v.lower() in ('nan', 'none', ''):
                    v = ''
                if n:
                    result[n] = v
            self._df_master = result
            return result
        except Exception as e:
            print(f"Aviso: não foi possível carregar base de fornecedores: {e}")
            return {}

    def carregar_fornecedores(self):
        """Escaneia todos os arquivos de clientes e monta lista de fornecedores."""
        try:
            master = self._carregar_master()

            prog = tk.Toplevel(self.root)
            prog.title("Aguarde...")
            prog.geometry("380x90")
            prog.transient(self.root)
            prog.grab_set()
            prog.update_idletasks()
            x = (prog.winfo_screenwidth() - 380) // 2
            y = (prog.winfo_screenheight() - 90) // 2
            prog.geometry(f"380x90+{x}+{y}")
            lbl_prog = ttk.Label(prog, text="Processando arquivos...")
            lbl_prog.pack(pady=10)
            bar = ttk.Progressbar(prog, mode='indeterminate')
            bar.pack(padx=20, fill='x')
            bar.start()
            self.root.update()

            nomes_encontrados = {}  # nome_upper → (nome_original, vinculo)
            pasta = str(PASTA_CLIENTES)

            if not os.path.exists(pasta):
                prog.destroy()
                messagebox.showwarning("Aviso", "Pasta de clientes não encontrada.")
                return

            for arq in os.listdir(pasta):
                if not arq.endswith('.xlsx'):
                    continue
                try:
                    lbl_prog.config(text=f"Processando: {os.path.splitext(arq)[0]}")
                    self.root.update()
                    df = pd.read_excel(os.path.join(pasta, arq), sheet_name='Dados')
                    if 'NOME' not in df.columns:
                        continue
                    if 'STATUS' in df.columns:
                        df = df[df['STATUS'].astype(str).str.upper().str.strip() == 'ATIVO']
                    for nome in df['NOME'].dropna().str.strip().unique():
                        if nome:
                            nu = nome.upper()
                            if nu not in nomes_encontrados:
                                vinculo = master.get(nu, '')
                                nomes_encontrados[nu] = (nome, vinculo)
                except Exception as e:
                    print(f"Erro ao ler {arq}: {e}")

            prog.destroy()

            if not nomes_encontrados:
                messagebox.showwarning("Aviso", "Nenhum fornecedor encontrado.")
                return

            self.fornecedores_disponiveis = sorted(
                [{'nome': v[0], 'vinculo': v[1]} for v in nomes_encontrados.values()],
                key=lambda x: x['nome'].upper()
            )

            # Atualizar combo de vínculo
            vinculos = sorted({f['vinculo'] for f in self.fornecedores_disponiveis if f['vinculo']})
            self._combo_vinculo['values'] = ['Todos'] + vinculos
            self.var_vinculo.set('Todos')

            # Mostrar a lista completa
            self._preencher_tree_forn(self.fornecedores_disponiveis)

            messagebox.showinfo(
                "Sucesso",
                f"{len(self.fornecedores_disponiveis)} fornecedores carregados.\n"
                f"Digite parte do nome para filtrar a lista."
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar fornecedores: {e}")

    def _preencher_tree_forn(self, lista):
        for item in self._tree_forn.get_children():
            self._tree_forn.delete(item)
        for f in lista[:200]:
            self._tree_forn.insert('', 'end', values=(f['nome'], f['vinculo'] or '–'))
        if not self._tree_forn_visible:
            self._tree_forn.pack(side='left', fill='both', expand=True)
            self._scroll_forn.pack(side='right', fill='y')
            self._tree_forn_visible = True

    def _filtrar_lista(self, event=None):
        if not self.fornecedores_disponiveis:
            return
        termo = self.var_busca.get().upper().strip()
        vinculo_sel = self.var_vinculo.get()

        filtrados = [
            f for f in self.fornecedores_disponiveis
            if (not termo or termo in f['nome'].upper()) and
               (vinculo_sel == 'Todos' or f['vinculo'] == vinculo_sel)
        ]
        self._preencher_tree_forn(filtrados)

    def _selecionar_primeiro(self, event=None):
        filhos = self._tree_forn.get_children()
        if filhos:
            self._tree_forn.selection_set(filhos[0])
            self._tree_forn.focus(filhos[0])
            self._selecionar_da_lista()

    def _selecionar_da_lista(self, event=None):
        sel = self._tree_forn.selection()
        if sel:
            valores = self._tree_forn.item(sel[0], 'values')
            nome = valores[0]
            vinculo = valores[1] if valores[1] != '–' else ''
            self.fornecedor_selecionado = {'nome': nome, 'vinculo': vinculo}
            self.var_busca.set(nome)
            self.lbl_sel.config(
                text=f"Selecionado: {nome}" + (f"  |  Vínculo: {vinculo}" if vinculo else ""),
                foreground='green'
            )
            # ocultar lista após seleção
            self._tree_forn.pack_forget()
            self._scroll_forn.pack_forget()
            self._tree_forn_visible = False

    def _limpar_selecao(self):
        self.fornecedor_selecionado = None
        self.var_busca.set('')
        self.lbl_sel.config(
            text="(nenhum selecionado – deixe em branco para todos)",
            foreground='gray'
        )
        if self._tree_forn_visible:
            self._tree_forn.pack_forget()
            self._scroll_forn.pack_forget()
            self._tree_forn_visible = False

    # ------------------------------------------------------------------ #
    #  GERAÇÃO DO RELATÓRIO                                                #
    # ------------------------------------------------------------------ #

    def gerar_relatorio(self):
        try:
            ini, fim = self._obter_periodo()
        except ValueError:
            return

        cats = self._categorias_selecionadas()
        vinculo_filtro = self.var_vinculo.get()
        nome_filtro = self.fornecedor_selecionado['nome'] if self.fornecedor_selecionado else None

        # Validação mínima
        if nome_filtro is None and vinculo_filtro == 'Todos' and not self.fornecedores_disponiveis:
            resp = messagebox.askyesno(
                "Confirmar",
                "Nenhum fornecedor específico selecionado e lista de fornecedores não carregada.\n"
                "Isso irá processar TODOS os fornecedores de TODOS os clientes no período.\n"
                "Deseja continuar?"
            )
            if not resp:
                return

        # Mostrar progresso
        prog = tk.Toplevel(self.root)
        prog.title("Gerando relatório...")
        prog.geometry("400x110")
        prog.transient(self.root)
        prog.grab_set()
        prog.update_idletasks()
        x = (prog.winfo_screenwidth() - 400) // 2
        y = (prog.winfo_screenheight() - 110) // 2
        prog.geometry(f"400x110+{x}+{y}")
        lbl_pr = ttk.Label(prog, text="Processando...", font=('Arial', 10))
        lbl_pr.pack(pady=12)
        bar = ttk.Progressbar(prog, mode='indeterminate')
        bar.pack(padx=20, fill='x')
        bar.start()
        self.root.update()

        try:
            registros = self._carregar_registros(ini, fim, nome_filtro, cats, vinculo_filtro, lbl_pr)
        except Exception as e:
            prog.destroy()
            messagebox.showerror("Erro", f"Erro ao carregar dados: {e}")
            import traceback; traceback.print_exc()
            return

        prog.destroy()

        if not registros:
            messagebox.showinfo(
                "Sem resultados",
                "Nenhum registro encontrado com os filtros selecionados."
            )
            return

        self._registros = registros
        self.dados_carregados = True

        desc_forn = f"Fornecedor: {nome_filtro}" if nome_filtro else "Todos os fornecedores"
        desc_periodo = f"{ini.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
        desc_vinculo = f"  |  Vínculo: {vinculo_filtro}" if vinculo_filtro != 'Todos' else ""
        descricao = f"{desc_forn}  |  Período: {desc_periodo}{desc_vinculo}"

        self._preencher_aba_cliente(registros, descricao)
        self._preencher_aba_registros(registros)
        self._preencher_aba_categoria(registros, descricao)
        self._notebook.select(0)

    def _carregar_registros(self, ini, fim, nome_filtro, cats, vinculo_filtro, lbl_prog=None):
        """Lê todos os arquivos de clientes e retorna lista de dicts com os registros."""
        master = self._carregar_master()
        registros = []
        pasta = str(PASTA_CLIENTES)

        if not os.path.exists(pasta):
            raise FileNotFoundError(f"Pasta de clientes não encontrada: {pasta}")

        for arq in os.listdir(pasta):
            if not arq.endswith('.xlsx'):
                continue
            nome_cliente = os.path.splitext(arq)[0]
            if lbl_prog:
                lbl_prog.config(text=f"Processando: {nome_cliente}")
                try:
                    self.root.update()
                except Exception:
                    pass
            try:
                df = pd.read_excel(os.path.join(pasta, arq), sheet_name='Dados')
            except Exception as e:
                print(f"Erro ao ler {arq}: {e}")
                continue

            colunas_req = ['DATA_REL', 'NOME', 'VALOR']
            if not all(c in df.columns for c in colunas_req):
                continue

            # Filtrar ativos
            if 'STATUS' in df.columns:
                df = df[df['STATUS'].astype(str).str.upper().str.strip() == 'ATIVO'].copy()
            if df.empty:
                continue

            # Converter datas
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            df = df.dropna(subset=['DATA_REL'])

            # Filtrar período
            df = df[(df['DATA_REL'] >= ini) & (df['DATA_REL'] <= fim)]
            if df.empty:
                continue

            # Normalizar CATEGORIA (coluna K = índice 10)
            if 'CATEGORIA' not in df.columns and len(df.columns) > 10:
                df = df.rename(columns={df.columns[10]: 'CATEGORIA'})
            if 'CATEGORIA' in df.columns:
                df['CATEGORIA'] = df['CATEGORIA'].astype(str).str.upper().str.strip()
                df['CATEGORIA'] = df['CATEGORIA'].replace(['NAN', 'NONE', ''], 'DIV')
            else:
                df['CATEGORIA'] = 'DIV'

            # Filtrar por categoria
            df = df[df['CATEGORIA'].isin(cats)]
            if df.empty:
                continue

            # Filtrar por fornecedor (nome)
            if nome_filtro:
                mask_exato = df['NOME'].astype(str).str.upper().str.strip() == nome_filtro.upper().strip()
                df_filt = df[mask_exato]
                if df_filt.empty:
                    mask_parcial = df['NOME'].astype(str).str.upper().str.contains(
                        nome_filtro.upper().strip(), na=False, regex=False
                    )
                    df_filt = df[mask_parcial]
                df = df_filt

            if df.empty:
                continue

            # Filtrar por vínculo (via master)
            if vinculo_filtro != 'Todos':
                df = df[df['NOME'].astype(str).str.upper().str.strip().map(
                    lambda n: master.get(n, '') == vinculo_filtro
                )]
            if df.empty:
                continue

            # Colunas auxiliares
            tp_desp_col = 'TP_DESP' if 'TP_DESP' in df.columns else None
            ref_col = 'REFERÊNCIA' if 'REFERÊNCIA' in df.columns else ('REFERENCIA' if 'REFERENCIA' in df.columns else None)
            nf_col = 'NF' if 'NF' in df.columns else None
            vencto_col = 'DT_VENCTO' if 'DT_VENCTO' in df.columns else None
            obs_col = next((c for c in ['OBSERVAÇÃO', 'OBSERVACAO'] if c in df.columns), None)
            cnpj_col = next((c for c in ['CNPJ_CPF', 'CNPJ/CPF'] if c in df.columns), None)

            for _, row in df.iterrows():
                try:
                    valor = float(row['VALOR']) if isinstance(row['VALOR'], (int, float)) else 0.0
                    if isinstance(row['VALOR'], str):
                        vs = row['VALOR'].replace('R$', '').replace('.', '').replace(',', '.').strip()
                        try:
                            valor = float(vs)
                        except ValueError:
                            valor = 0.0
                    if valor <= 0:
                        continue

                    nome_forn = str(row['NOME']).strip()
                    cnpj = str(row[cnpj_col]).strip() if cnpj_col and pd.notna(row[cnpj_col]) else ''
                    vinculo_forn = master.get(nome_forn.upper(), '')
                    tp = int(row[tp_desp_col]) if tp_desp_col and pd.notna(row[tp_desp_col]) else 0
                    ref = str(row[ref_col]).strip() if ref_col and pd.notna(row[ref_col]) else ''
                    nf = str(row[nf_col]).strip() if nf_col and pd.notna(row[nf_col]) else ''
                    if nf and nf.lower() not in ('nan', 'none', ''):
                        ref = f"{ref} (NF: {nf})" if ref else f"NF: {nf}"
                    vencto = None
                    if vencto_col and pd.notna(row[vencto_col]):
                        try:
                            vencto = pd.to_datetime(row[vencto_col])
                        except Exception:
                            pass
                    obs = str(row[obs_col]).strip() if obs_col and pd.notna(row[obs_col]) else ''
                    categoria = str(row['CATEGORIA']) if pd.notna(row['CATEGORIA']) else 'DIV'

                    registros.append({
                        'data': row['DATA_REL'],
                        'cliente': nome_cliente,
                        'fornecedor': nome_forn,
                        'cnpj': cnpj,
                        'vinculo': vinculo_forn,
                        'tp_desp': tp,
                        'categoria': categoria,
                        'referencia': ref,
                        'dt_vencto': vencto,
                        'valor': valor,
                        'observacao': obs,
                    })
                except Exception as e:
                    print(f"Erro no registro ({nome_cliente}): {e}")

        return registros

    # ------------------------------------------------------------------ #
    #  PREENCHIMENTO DAS ABAS                                              #
    # ------------------------------------------------------------------ #

    def _preencher_aba_cliente(self, registros, descricao):
        for item in self._tree_cliente.get_children():
            self._tree_cliente.delete(item)

        total_geral = sum(r['valor'] for r in registros)

        por_cliente = defaultdict(lambda: {'total': 0.0, 'qtd': 0, 'cats': set()})
        for r in registros:
            por_cliente[r['cliente']]['total'] += r['valor']
            por_cliente[r['cliente']]['qtd'] += 1
            por_cliente[r['cliente']]['cats'].add(r['categoria'])

        ordenado = sorted(por_cliente.items(), key=lambda x: x[1]['total'], reverse=True)

        for i, (cli, d) in enumerate(ordenado, 1):
            pct = d['total'] / total_geral * 100 if total_geral > 0 else 0
            cats_str = ', '.join(sorted(d['cats']))
            self._tree_cliente.insert('', 'end', values=(
                i, cli, formatar_moeda_br(d['total']),
                f"{pct:.1f}%", d['qtd'], cats_str
            ))

        self._lbl_info_cliente.config(text=descricao)
        self._lbl_total_cliente.config(
            text=f"Total: {formatar_moeda_br(total_geral)}  ({len(registros)} registros)"
        )

    def _preencher_aba_registros(self, registros):
        self._aplicar_ordem_registros(registros)

    def _aplicar_ordem_registros(self, registros=None):
        if registros is None:
            registros = self._registros
        if not registros:
            return

        ordem = self.var_ordem.get()
        if ordem == 'Data (mais recente)':
            r = sorted(registros, key=lambda x: x['data'], reverse=True)
        elif ordem == 'Data (mais antiga)':
            r = sorted(registros, key=lambda x: x['data'])
        elif ordem == 'Cliente A-Z':
            r = sorted(registros, key=lambda x: x['cliente'])
        elif ordem == 'Valor (maior)':
            r = sorted(registros, key=lambda x: x['valor'], reverse=True)
        else:
            r = sorted(registros, key=lambda x: x['valor'])

        for item in self._tree_regs.get_children():
            self._tree_regs.delete(item)

        for reg in r:
            self._tree_regs.insert('', 'end', values=(
                reg['data'].strftime('%d/%m/%Y'),
                reg['cliente'],
                reg['fornecedor'],
                reg['tp_desp'],
                reg['categoria'],
                reg['referencia'],
                reg['dt_vencto'].strftime('%d/%m/%Y') if reg['dt_vencto'] else '',
                formatar_moeda_br(reg['valor']),
                reg['observacao'],
            ))

        self._lbl_qtd_registros.config(text=f"{len(r)} registros encontrados")

    def _preencher_aba_categoria(self, registros, descricao):
        # Reconstruir treeview com colunas dinâmicas
        self._tree_cat.destroy()
        ft = self._tree_cat.master

        cats_presentes = sorted({r['categoria'] for r in registros})
        clientes = sorted({r['cliente'] for r in registros})

        # pivot: cliente → categoria → total
        pivot = defaultdict(lambda: defaultdict(float))
        for r in registros:
            pivot[r['cliente']][r['categoria']] += r['valor']

        cols = ('cliente',) + tuple(cats_presentes) + ('total',)
        self._tree_cat = ttk.Treeview(ft, columns=cols, show='headings', height=18)
        self._tree_cat.heading('cliente', text='Cliente')
        self._tree_cat.column('cliente', width=220)
        for cat in cats_presentes:
            nome_cat = CATEGORIAS.get(cat, cat)
            self._tree_cat.heading(cat, text=f"{cat}\n({nome_cat})")
            self._tree_cat.column(cat, width=110, anchor='e')
        self._tree_cat.heading('total', text='TOTAL')
        self._tree_cat.column('total', width=130, anchor='e')

        sy2 = ttk.Scrollbar(ft, orient='vertical', command=self._tree_cat.yview)
        sx2 = ttk.Scrollbar(ft, orient='horizontal', command=self._tree_cat.xview)
        self._tree_cat.configure(yscrollcommand=sy2.set, xscrollcommand=sx2.set)
        self._tree_cat.pack(side='left', fill='both', expand=True)
        sy2.pack(side='right', fill='y')
        sx2.pack(side='bottom', fill='x')

        totais_col = defaultdict(float)
        for cli in clientes:
            linha = [cli]
            total_cli = 0.0
            for cat in cats_presentes:
                val = pivot[cli][cat]
                linha.append(formatar_moeda_br(val) if val else '–')
                total_cli += val
                totais_col[cat] += val
            linha.append(formatar_moeda_br(total_cli))
            self._tree_cat.insert('', 'end', values=tuple(linha))

        # linha de totais
        linha_total = ['TOTAL']
        grand_total = 0.0
        for cat in cats_presentes:
            linha_total.append(formatar_moeda_br(totais_col[cat]))
            grand_total += totais_col[cat]
        linha_total.append(formatar_moeda_br(grand_total))
        item_total = self._tree_cat.insert('', 'end', values=tuple(linha_total))
        self._tree_cat.item(item_total, tags=('total_row',))
        self._tree_cat.tag_configure('total_row', background='#dce9f7', font=('Arial', 10, 'bold'))

        self._lbl_info_cat.config(text=descricao)

    def _drill_down_cliente(self, event):
        """Ao clicar duas vezes em um cliente na aba 'Por Cliente', vai para 'Todos os Registros' filtrado."""
        sel = self._tree_cliente.selection()
        if not sel:
            return
        cliente = self._tree_cliente.item(sel[0], 'values')[1]
        filtrados = [r for r in self._registros if r['cliente'] == cliente]
        self._aplicar_ordem_registros(filtrados)
        self._notebook.select(1)

    # ------------------------------------------------------------------ #
    #  EXPORTAÇÃO                                                          #
    # ------------------------------------------------------------------ #

    def exportar_excel(self):
        if not self.dados_carregados or not self._registros:
            messagebox.showwarning("Aviso", "Gere o relatório primeiro.")
            return

        nome_arq = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            title="Salvar Relatório"
        )
        if not nome_arq:
            return

        try:
            ini, fim = self._obter_periodo()
            forn = self.fornecedor_selecionado['nome'] if self.fornecedor_selecionado else 'Todos'
            vinculo = self.var_vinculo.get()

            wb = Workbook()

            # --- aba: Todos os Registros ---
            ws1 = wb.active
            ws1.title = "Registros"
            cabecalho = ['Data', 'Cliente', 'Fornecedor', 'CNPJ/CPF', 'Vínculo',
                         'Tp.Desp', 'Categoria', 'Referência', 'Vencimento', 'Valor', 'Observação']
            ws1.append(cabecalho)
            for cell in ws1[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

            for r in sorted(self._registros, key=lambda x: (x['cliente'], x['data'])):
                ws1.append([
                    r['data'].strftime('%d/%m/%Y'),
                    r['cliente'],
                    r['fornecedor'],
                    r['cnpj'],
                    r['vinculo'],
                    r['tp_desp'],
                    r['categoria'],
                    r['referencia'],
                    r['dt_vencto'].strftime('%d/%m/%Y') if r['dt_vencto'] else '',
                    r['valor'],
                    r['observacao'],
                ])
            for col in ws1.columns:
                ws1.column_dimensions[get_column_letter(col[0].column)].width = max(
                    12, max(len(str(cell.value or '')) for cell in col)
                )

            # --- aba: Por Cliente ---
            ws2 = wb.create_sheet("Por Cliente")
            ws2.append(['Fornecedor:', forn])
            ws2.append(['Período:', f"{ini.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"])
            ws2.append(['Vínculo:', vinculo])
            ws2.append([])
            cabecalho2 = ['#', 'Cliente', 'Total Gasto', '% do Total', 'Qtd. Registros', 'Categorias']
            ws2.append(cabecalho2)
            for cell in ws2[5]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='2E74B5', end_color='2E74B5', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

            total_geral = sum(r['valor'] for r in self._registros)
            por_cliente = defaultdict(lambda: {'total': 0.0, 'qtd': 0, 'cats': set()})
            for r in self._registros:
                por_cliente[r['cliente']]['total'] += r['valor']
                por_cliente[r['cliente']]['qtd'] += 1
                por_cliente[r['cliente']]['cats'].add(r['categoria'])

            for i, (cli, d) in enumerate(
                sorted(por_cliente.items(), key=lambda x: x[1]['total'], reverse=True), 1
            ):
                pct = d['total'] / total_geral * 100 if total_geral > 0 else 0
                ws2.append([i, cli, d['total'], f"{pct:.1f}%", d['qtd'], ', '.join(sorted(d['cats']))])

            ws2.append([])
            ws2.append(['', 'TOTAL GERAL', total_geral, '100%', len(self._registros), ''])
            for col in ws2.columns:
                ws2.column_dimensions[get_column_letter(col[0].column)].width = 18

            # --- aba: Por Categoria ---
            ws3 = wb.create_sheet("Por Categoria")
            cats_presentes = sorted({r['categoria'] for r in self._registros})
            clientes_sorted = sorted(por_cliente.keys())
            pivot = defaultdict(lambda: defaultdict(float))
            for r in self._registros:
                pivot[r['cliente']][r['categoria']] += r['valor']

            header3 = ['Cliente'] + [f"{c} – {CATEGORIAS.get(c, c)}" for c in cats_presentes] + ['TOTAL']
            ws3.append(header3)
            for cell in ws3[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='2E74B5', end_color='2E74B5', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

            totais_cat = defaultdict(float)
            for cli in clientes_sorted:
                linha = [cli]
                total_cli = 0.0
                for cat in cats_presentes:
                    val = pivot[cli][cat]
                    linha.append(val if val else 0)
                    total_cli += val
                    totais_cat[cat] += val
                linha.append(total_cli)
                ws3.append(linha)

            linha_total = ['TOTAL'] + [totais_cat[c] for c in cats_presentes] + [sum(totais_cat.values())]
            ws3.append(linha_total)
            for cell in ws3[ws3.max_row]:
                cell.font = Font(bold=True)

            for col in ws3.columns:
                ws3.column_dimensions[get_column_letter(col[0].column)].width = 18

            wb.save(nome_arq)
            messagebox.showinfo("Sucesso", f"Relatório exportado:\n{nome_arq}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {e}")
            import traceback; traceback.print_exc()

    # ------------------------------------------------------------------ #
    #  CONTROLE DE JANELA                                                  #
    # ------------------------------------------------------------------ #

    def voltar_menu(self):
        try:
            self.root.destroy()
            if self.parent and hasattr(self.parent, 'deiconify'):
                self.parent.deiconify()
        except Exception:
            pass

    def fechar_aplicacao(self):
        if messagebox.askyesno("Confirmar", "Deseja realmente sair?"):
            self.voltar_menu()


def main():
    app = RelatorioPorFornecedorDetalhado()
    app.root.mainloop()


if __name__ == '__main__':
    main()
