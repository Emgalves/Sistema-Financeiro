import os
import sys
from pathlib import Path
import re
from datetime import datetime
from decimal import Decimal

import tkinter as tk
from tkinter import ttk, messagebox, StringVar
from tkcalendar import DateEntry

from dateutil.relativedelta import relativedelta

import pandas as pd
from openpyxl import load_workbook

# Adicionar diretório raiz ao path
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações do sistema
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH,
        ARQUIVO_FORNECEDORES
    )
    from src.config.utils import formatar_cnpj_cpf, aplicar_formatacao_celula
    from src.config.window_config import configurar_janela
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Tentar caminho alternativo
    try:
        from config.config import (
            ARQUIVO_CLIENTES,
            ARQUIVO_MODELO,
            PASTA_CLIENTES,
            BASE_PATH,
            ARQUIVO_FORNECEDORES
        )
        from config.utils import formatar_cnpj_cpf, aplicar_formatacao_celula
        from config.window_config import configurar_janela
    except ImportError as e2:
        print(f"Erro ao importar configurações (caminho alternativo): {str(e2)}")
        raise

class GestaoEventos:
    def __init__(self, parent=None):
        self.parent = parent
        self.janela = None
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.contratos = []
        self.eventos = []
        self.administradores_contratos = {}
        self.cliente_selecionado = False  # Variável para controlar resultado da seleção
        self.notebook = None  # Para armazenar o notebook de abas
        self.aba_contratos = None
        self.aba_eventos = None
        self.aba_pagamentos = None
        self.tree_contratos = None
        self.tree_eventos = None
        self.tree_pagamentos = None
        self.tree_adm = None
        
        # Labels para detalhes do contrato
        self.lbl_contrato = None
        self.lbl_periodo = None
        self.lbl_status = None
        self.lbl_valor_total = None
        self.lbl_administradores = None
        self.lbl_eventos = None
        
        # Campos para eventos
        self.contrato_selecionado = None
        self.evento_descricao = None
        self.evento_percentual = None
        self.evento_status = None
        self.evento_data = None
        
        # Campos para pagamentos
        self.pagto_contrato = None
        self.pagto_status = None
        self.pagto_cnpj = None
        self.pagto_nome = None
        self.pagto_valor = None
        self.pagto_vencimento = None
        self.pagto_status_atual = None
        self.pagto_data = None

    def fechar_janela(self):
        """Fecha a janela corretamente"""
        print("Fechando janela")
        if self.janela:
            self.janela.destroy()
        
    def carregar_lista_clientes(self):
        """Carrega a lista de clientes disponíveis"""
        try:
            print("Carregando lista de clientes...")
            # Verificar se o arquivo existe
            if not os.path.exists(ARQUIVO_CLIENTES):
                print(f"Arquivo não encontrado: {ARQUIVO_CLIENTES}")
                messagebox.showwarning("Aviso", "Arquivo de clientes não encontrado!")
                return []
                
            workbook = load_workbook(ARQUIVO_CLIENTES)
            sheet = workbook['Clientes']
            print(f"Planilha aberta, abas disponíveis: {workbook.sheetnames}")
            print(f"Número de linhas na planilha: {sheet.max_row}")
            
            clientes = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome do cliente
                    print(f"Cliente encontrado: {row[0]}")
                    clientes.append(row[0])
                    
            workbook.close()
            print(f"Total de clientes carregados: {len(clientes)}")
            return sorted(clientes)
            
        except Exception as e:
            print(f"Erro ao carregar clientes: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
            return []
    
    def abrir_janela_eventos(self, cliente=None):
        """Abre a janela principal de gestão de eventos para o cliente selecionado"""
        print("Entrando em abrir_janela_eventos")
        
        # Se a janela já existir, apenas traz para frente
        if self.janela and self.janela.winfo_exists():
            print("Janela já existe, trazendo para frente")
            self.janela.lift()
            self.janela.focus_force()
            return

        print("Criando nova janela")
        # Cria nova janela
        self.janela = tk.Toplevel(self.parent)
        configurar_janela(self.janela, "Gestão de Pagamentos por Eventos", 900, 950)
        
        # Garantir visibilidade
        self.janela.attributes('-topmost', True)
        self.janela.update()
        self.janela.attributes('-topmost', False)

        # Define o cliente atual
        if cliente:
            print(f"Cliente fornecido: {cliente}")
            self.cliente_atual = cliente
            self.arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        else:
            print("Cliente não fornecido, solicitando seleção")
            # Se não tiver cliente selecionado, solicita seleção
            if not self.selecionar_cliente():
                print("Cliente não selecionado, fechando janela")
                self.janela.destroy()
                return

        # Configurar fechamento adequado
        def fechar_janela_local():
            print("Fechando janela")
            self.janela.destroy()
        
        # Usar a função local para fechamento
        self.janela.protocol("WM_DELETE_WINDOW", fechar_janela_local)
        
        # Frame principal
        frame_principal = ttk.Frame(self.janela, padding=10)
        frame_principal.pack(fill='both', expand=True)

        # Cabeçalho com informações do cliente
        frame_cabecalho = ttk.Frame(frame_principal)
        frame_cabecalho.pack(fill='x', pady=5)

        ttk.Label(
            frame_cabecalho, 
            text=f"Cliente: {self.cliente_atual}", 
            font=('Arial', 12, 'bold')
        ).pack(side='left')

        # Notebook para abas
        self.notebook = ttk.Notebook(frame_principal)
        self.notebook.pack(fill='both', expand=True, pady=10)

        # Aba de contratos
        self.aba_contratos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_contratos, text="Contratos")

        # Aba de eventos
        self.aba_eventos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_eventos, text="Eventos")

        # Aba de pagamentos
        self.aba_pagamentos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_pagamentos, text="Pagamentos")

        # Configurar cada aba
        self.configurar_aba_contratos()
        self.configurar_aba_eventos()
        self.configurar_aba_pagamentos()

         # Botões principais
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=10)

        ttk.Button(
            frame_botoes,
            text="Fechar",
            command=fechar_janela_local,
            width=15
        ).pack(side='right', padx=5)

        # Carregar dados iniciais
        self.carregar_contratos()
    
    def selecionar_cliente(self):
        """Seleciona um cliente usando uma caixa de diálogo simplificada"""
        # Carregar a lista de clientes
        clientes = self.carregar_lista_clientes()
        if not clientes:
            messagebox.showwarning("Aviso", "Nenhum cliente encontrado!")
            return False
        
        # Criar janela de diálogo
        dialogo = tk.Toplevel(self.janela)
        dialogo.title("Selecionar Cliente")
        dialogo.geometry("400x200")
        dialogo.resizable(False, False)
        dialogo.transient(self.janela)
        dialogo.grab_set()
        
        # Variável para resultado
        resultado = [False]  # Usamos uma lista para poder modificar de dentro da função
        
        # Frame principal
        frame = ttk.Frame(dialogo, padding=20)
        frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            frame, 
            text="Selecione um Cliente",
            font=("Arial", 12, "bold")
        ).pack(pady=10)
        
        # Combobox para seleção
        cliente_var = tk.StringVar()
        combo = ttk.Combobox(
            frame, 
            textvariable=cliente_var,
            values=clientes,
            state="readonly",
            width=40
        )
        combo.pack(pady=10)
        if clientes:
            combo.set(clientes[0])  # Selecionar o primeiro cliente por padrão
        
        # Função para confirmar
        def confirmar():
            if not cliente_var.get():
                messagebox.showwarning("Aviso", "Selecione um cliente!")
                return
            
            self.cliente_atual = cliente_var.get()
            self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
            resultado[0] = True
            dialogo.destroy()
        
        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(pady=20)
        
        ttk.Button(
            frame_botoes,
            text="Confirmar",
            command=confirmar,
            width=15
        ).pack(side='left', padx=10)
        
        ttk.Button(
            frame_botoes,
            text="Cancelar",
            command=dialogo.destroy,
            width=15
        ).pack(side='left', padx=10)
        
        # Primeiro é necessário atualizar a geometria para obter o tamanho real
        dialogo.update_idletasks()
        
        # Obter o tamanho e posição da janela pai
        width = dialogo.winfo_width()
        height = dialogo.winfo_height()
        
        # Centralizar em relação à janela pai
        parent_x = self.janela.winfo_x()
        parent_y = self.janela.winfo_y()
        parent_width = self.janela.winfo_width()
        parent_height = self.janela.winfo_height()
        
        # Calcular a posição para centralizar
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        # Garantir que a janela não fique fora da tela
        screen_width = self.janela.winfo_screenwidth()
        screen_height = self.janela.winfo_screenheight()
        
        # Ajustar se necessário
        if x < 0:
            x = 0
        elif x + width > screen_width:
            x = screen_width - width
            
        if y < 0:
            y = 0
        elif y + height > screen_height:
            y = screen_height - height
        
        # Definir a posição
        dialogo.geometry(f"{width}x{height}+{x}+{y}")
        
        # Adicionar tecla Enter para confirmar
        dialogo.bind('<Return>', lambda event: confirmar())
        
        # Trazer para frente e focar no combobox
        dialogo.lift()
        combo.focus_set()
        
        # Aguardar seleção
        self.janela.wait_window(dialogo)
        
        # Retornar resultado
        return resultado[0]
    
    def confirmar_selecao_cliente(self):
        """Confirma a seleção do cliente e inicializa a interface"""
        cliente = self.var_cliente.get()
        if not cliente:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro")
            return
            
        self.cliente_atual = cliente
        self.arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        print(f"Cliente selecionado: {cliente}")
        print(f"Arquivo: {self.arquivo_cliente}")
        
        # Verificar existência do arquivo
        if not os.path.exists(self.arquivo_cliente):
            messagebox.showerror("Erro", f"Arquivo do cliente não encontrado: {self.arquivo_cliente}")
            return
            
        # Inicializar interface principal
        self.inicializar_interface()
    
    def inicializar_interface(self):
        """Inicializa a interface principal com abas"""
        print("Inicializando interface principal")
        
        # Limpar janela atual
        for widget in self.janela.winfo_children():
            widget.destroy()
            
        # Frame principal
        frame_principal = ttk.Frame(self.janela, padding=10)
        frame_principal.pack(fill='both', expand=True)
        
        # Cabeçalho com informações do cliente
        frame_cabecalho = ttk.Frame(frame_principal)
        frame_cabecalho.pack(fill='x', pady=5)
        
        ttk.Label(
            frame_cabecalho, 
            text=f"Cliente: {self.cliente_atual}", 
            font=('Arial', 12, 'bold')
        ).pack(side='left')
        
        # Notebook para abas
        self.notebook = ttk.Notebook(frame_principal)
        self.notebook.pack(fill='both', expand=True, pady=10)
        
        # Aba de contratos
        self.aba_contratos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_contratos, text="Contratos")
        
        # Aba de eventos
        self.aba_eventos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_eventos, text="Eventos")
        
        # Aba de pagamentos
        self.aba_pagamentos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_pagamentos, text="Pagamentos")
        
        # Configurar cada aba
        self.configurar_aba_contratos()
        self.configurar_aba_eventos()
        self.configurar_aba_pagamentos()
        
        # Botões principais
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(
            frame_botoes,
            text="Fechar",
            command=self.fechar_janela,
            width=15
        ).pack(side='right', padx=5)
        
        # Carregar dados iniciais
        self.carregar_contratos()
        
    def configurar_aba_contratos(self):
        """Configura a aba de gestão de contratos"""
        print("Configurando aba de contratos")
        frame = ttk.Frame(self.aba_contratos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para lista de contratos
        frame_lista = ttk.LabelFrame(frame, text="Contratos Disponíveis")
        frame_lista.pack(fill='both', expand=True, pady=5)

        # Treeview para contratos
        colunas = ('Nº Contrato', 'Data Início', 'Data Fim', 'Status', 'Valor Total')
        self.tree_contratos = ttk.Treeview(frame_lista, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_contratos.heading(col, text=col)
            if col == 'Valor Total':
                self.tree_contratos.column(col, width=120, anchor='e')
            else:
                self.tree_contratos.column(col, width=120)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_contratos.yview)
        scroll_x = ttk.Scrollbar(frame_lista, orient='horizontal', command=self.tree_contratos.xview)
        self.tree_contratos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_contratos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do contrato selecionado
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Contrato")
        frame_detalhes.pack(fill='x', pady=10)

        frame_detalhes_grid = ttk.Frame(frame_detalhes)
        frame_detalhes_grid.pack(fill='x', padx=5, pady=5)

        # Criar labels e campos para detalhes
        ttk.Label(frame_detalhes_grid, text="Contrato:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.lbl_contrato = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_contrato.grid(row=0, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Período:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.lbl_periodo = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_periodo.grid(row=1, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Status:").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.lbl_status = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_status.grid(row=2, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Valor Total:").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.lbl_valor_total = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_valor_total.grid(row=0, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Administradores:").grid(row=1, column=2, sticky='w', padx=5, pady=2)
        self.lbl_administradores = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_administradores.grid(row=1, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Eventos:").grid(row=2, column=2, sticky='w', padx=5, pady=2)
        self.lbl_eventos = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_eventos.grid(row=2, column=3, sticky='w', padx=5, pady=2)

        # Frame para administradores do contrato
        frame_admins = ttk.LabelFrame(frame, text="Administradores do Contrato")
        frame_admins.pack(fill='both', expand=True, pady=5)

        # Treeview para administradores
        colunas_adm = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas')
        self.tree_adm = ttk.Treeview(frame_admins, columns=colunas_adm, show='headings', height=4)
        
        for col in colunas_adm:
            self.tree_adm.heading(col, text=col)
            if col in ['Valor/Percentual', 'Valor Total']:
                self.tree_adm.column(col, width=120, anchor='e')
            elif col == 'Nº Parcelas':
                self.tree_adm.column(col, width=80, anchor='center')
            else:
                self.tree_adm.column(col, width=120)

        # Scrollbars
        scroll_y_adm = ttk.Scrollbar(frame_admins, orient='vertical', command=self.tree_adm.yview)
        scroll_x_adm = ttk.Scrollbar(frame_admins, orient='horizontal', command=self.tree_adm.xview)
        self.tree_adm.configure(yscrollcommand=scroll_y_adm.set, xscrollcommand=scroll_x_adm.set)
        
        self.tree_adm.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y_adm.pack(side='right', fill='y')
        scroll_x_adm.pack(side='bottom', fill='x')

        # Botões de ação
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes,
            text="Definir Eventos",
            command=self.definir_eventos,
            width=20
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Atualizar Eventos",
            command=self.atualizar_eventos_contrato,
            width=20
        ).pack(side='left', padx=5)

        # Binding para mostrar detalhes ao selecionar contrato
        self.tree_contratos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_contrato)
        
    def configurar_aba_eventos(self):
        """Configura a aba de gestão de eventos"""
        print("Configurando aba de eventos")
        frame = ttk.Frame(self.aba_eventos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para contrato selecionado
        frame_contrato = ttk.LabelFrame(frame, text="Contrato")
        frame_contrato.pack(fill='x', pady=5)

        ttk.Label(frame_contrato, text="Contrato Selecionado:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.contrato_selecionado = ttk.Combobox(frame_contrato, state='readonly', width=40)
        self.contrato_selecionado.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.contrato_selecionado.bind('<<ComboboxSelected>>', self.carregar_eventos_contrato)

        # Frame para botão de adicionar evento
        frame_adicionar = ttk.Frame(frame)
        frame_adicionar.pack(fill='x', pady=5)

        ttk.Button(
            frame_adicionar,
            text="Adicionar Evento",
            command=self.adicionar_evento,
            width=20
        ).pack(side='left', padx=5)

        # Frame para lista de eventos
        frame_eventos = ttk.LabelFrame(frame, text="Eventos do Contrato")
        frame_eventos.pack(fill='both', expand=True, pady=5)

        # Treeview para eventos
        colunas = ('ID', 'Descrição', 'Percentual', 'Valor', 'Status', 'Data Conclusão')
        self.tree_eventos = ttk.Treeview(frame_eventos, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_eventos.heading(col, text=col)
            if col == 'ID':
                self.tree_eventos.column(col, width=50, anchor='center')
            elif col in ['Percentual', 'Valor']:
                self.tree_eventos.column(col, width=100, anchor='e')
            elif col == 'Status':
                self.tree_eventos.column(col, width=100, anchor='center')
            else:
                self.tree_eventos.column(col, width=150)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_eventos, orient='vertical', command=self.tree_eventos.yview)
        scroll_x = ttk.Scrollbar(frame_eventos, orient='horizontal', command=self.tree_eventos.xview)
        self.tree_eventos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_eventos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do evento selecionado
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Evento")
        frame_detalhes.pack(fill='x', pady=10)

        # Grid para formulário
        frame_form = ttk.Frame(frame_detalhes)
        frame_form.pack(fill='x', padx=5, pady=5)

        ttk.Label(frame_form, text="Descrição:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.evento_descricao = ttk.Entry(frame_form, width=50)
        self.evento_descricao.grid(row=0, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Percentual (%):").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.evento_percentual = ttk.Entry(frame_form, width=15)
        self.evento_percentual.grid(row=1, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Status:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.evento_status = ttk.Combobox(frame_form, values=['pendente', 'concluido'], state='readonly', width=15)
        self.evento_status.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        self.evento_status.set('pendente')

        ttk.Label(frame_form, text="Data Conclusão:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.evento_data = DateEntry(frame_form, width=15, state='normal', date_pattern='dd/mm/yyyy')
        self.evento_data.grid(row=3, column=1, padx=5, pady=2, sticky='w')

        # Botões de ação para eventos
        frame_botoes = ttk.Frame(frame_detalhes)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes,
            text="Salvar Evento",
            command=self.salvar_evento,
            width=15
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Marcar como Concluído",
            command=self.concluir_evento,
            width=20
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Remover Evento",
            command=self.remover_evento,
            width=15
        ).pack(side='left', padx=5)

        # Binding para seleção de evento
        self.tree_eventos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_evento)

    def configurar_aba_pagamentos(self):
        """Configura a aba de pagamentos de eventos"""
        print("Configurando aba de pagamentos")
        frame = ttk.Frame(self.aba_pagamentos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para controle de filtros
        frame_filtros = ttk.LabelFrame(frame, text="Filtros")
        frame_filtros.pack(fill='x', pady=5)

        ttk.Label(frame_filtros, text="Contrato:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.pagto_contrato = ttk.Combobox(frame_filtros, state='readonly', width=40)
        self.pagto_contrato.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.pagto_contrato.bind('<<ComboboxSelected>>', self.filtrar_pagamentos)

        ttk.Label(frame_filtros, text="Status:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        self.pagto_status = ttk.Combobox(frame_filtros, values=['Todos', 'Pendente', 'Pago'], state='readonly', width=15)
        self.pagto_status.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        self.pagto_status.set('Todos')
        self.pagto_status.bind('<<ComboboxSelected>>', self.filtrar_pagamentos)

        # Frame para lista de pagamentos
        frame_pagamentos = ttk.LabelFrame(frame, text="Pagamentos Previstos")
        frame_pagamentos.pack(fill='both', expand=True, pady=5)

        # Treeview para pagamentos
        colunas = ('Contrato', 'Evento', 'CNPJ/CPF', 'Nome', 'Data Vencimento', 'Valor', 'Status', 'Data Pagamento')
        self.tree_pagamentos = ttk.Treeview(frame_pagamentos, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_pagamentos.heading(col, text=col)
            if col in ['Data Vencimento', 'Data Pagamento']:
                self.tree_pagamentos.column(col, width=120, anchor='center')
            elif col == 'Valor':
                self.tree_pagamentos.column(col, width=100, anchor='e')
            elif col == 'Status':
                self.tree_pagamentos.column(col, width=100, anchor='center')
            elif col in ['Contrato', 'Evento']:
                self.tree_pagamentos.column(col, width=80, anchor='center')
            else:
                self.tree_pagamentos.column(col, width=150)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_pagamentos, orient='vertical', command=self.tree_pagamentos.yview)
        scroll_x = ttk.Scrollbar(frame_pagamentos, orient='horizontal', command=self.tree_pagamentos.xview)
        self.tree_pagamentos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_pagamentos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do pagamento
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Pagamento")
        frame_detalhes.pack(fill='x', pady=10)

        # Grid para formulário
        frame_form = ttk.Frame(frame_detalhes)
        frame_form.pack(fill='x', padx=5, pady=5)

        ttk.Label(frame_form, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.pagto_cnpj = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_cnpj.grid(row=0, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Nome:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.pagto_nome = ttk.Entry(frame_form, width=40, state='readonly')
        self.pagto_nome.grid(row=0, column=3, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Valor:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.pagto_valor = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_valor.grid(row=1, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Vencimento:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.pagto_vencimento = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_vencimento.grid(row=1, column=3, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Status:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.pagto_status_atual = ttk.Combobox(frame_form, values=['Pendente', 'Pago'], state='readonly', width=15)
        self.pagto_status_atual.grid(row=2, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Data Pagamento:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.pagto_data = DateEntry(frame_form, width=15, state='normal', date_pattern='dd/mm/yyyy')
        self.pagto_data.grid(row=2, column=3, padx=5, pady=2, sticky='w')

        # Botões de ação para pagamentos
        frame_botoes = ttk.Frame(frame_detalhes)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes,
            text="Registrar Pagamento",
            command=self.registrar_pagamento,
            width=20
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Gerar Lançamento",
            command=self.gerar_lancamento_pagamento,
            width=20
        ).pack(side='left', padx=5)

        # Binding para seleção de pagamento
        self.tree_pagamentos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_pagamento)
    
    def definir_eventos(self):
        """Abre uma janela para definir eventos para o contrato selecionado"""
        # Verificar se há um contrato selecionado
        selecao = self.tree_contratos.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione um contrato primeiro!")
            return
            
        # Obtém o número do contrato selecionado
        item = self.tree_contratos.item(selecao[0])
        num_contrato = item['values'][0]
        
        # Muda para a aba de eventos e seleciona o contrato no combobox
        self.notebook.select(self.aba_eventos)
        self.contrato_selecionado.set(num_contrato)
        
        # Carrega os eventos do contrato
        self.carregar_eventos_contrato()
        
    def carregar_eventos_contrato(self, event=None):
        """Carrega a lista de eventos do contrato selecionado"""
        contrato = self.contrato_selecionado.get()
        if not contrato:
            return
            
        # Limpar tabela
        for item in self.tree_eventos.get_children():
            self.tree_eventos.delete(item)
        
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de contratos existe
            if 'Contratos_ADM' not in workbook.sheetnames:
                messagebox.showwarning("Aviso", "Aba de contratos não encontrada!")
                workbook.close()
                return
                
            # Como não temos uma aba específica para eventos, vamos criar eventos virtuais
            # baseados nos administradores do contrato
            sheet = workbook['Contratos_ADM']
            
            # Encontrar todas as linhas do contrato selecionado
            evento_id = 1
            
            # Determinar o valor total do contrato para calcular os eventos
            valor_contrato = 0
            for row in sheet.iter_rows(min_row=3, values_only=True):
                # Verificar se é uma linha de contrato (não de administrador, aditivo ou parcela)
                if row[0] and str(row[0]) == str(contrato) and row[1] is not None and row[6] is None:
                    # Encontramos a linha do contrato
                    if row[4] and isinstance(row[4], (int, float)):
                        valor_contrato = row[4]
                    else:
                        try:
                            if row[4] and isinstance(row[4], str):
                                valor_contrato = float(row[4].replace(',', '.'))
                        except (ValueError, TypeError):
                            valor_contrato = 0
                    break
            
            # Agora procurar administradores deste contrato para criar eventos
            for row in sheet.iter_rows(min_row=3, values_only=True):
                # Verificar se é um administrador do contrato
                if row[6] and str(row[6]) == str(contrato):
                    tipo = "Percentual" if row[9] == "Percentual" else "Valor Fixo"
                    
                    # Calcular valor e percentual
                    percentual = 0
                    valor = 0
                    
                    if tipo == "Percentual":
                        try:
                            if isinstance(row[10], str):
                                percentual = float(row[10].replace('%', '').replace(',', '.'))
                            else:
                                percentual = float(row[10]) if row[10] else 0
                                
                            # Calcular valor baseado no percentual
                            valor = (percentual / 100) * valor_contrato
                        except (ValueError, TypeError):
                            percentual = 0
                            valor = 0
                    else:
                        # Valor fixo
                        try:
                            if isinstance(row[10], str):
                                valor = float(row[10].replace(',', '.'))
                            else:
                                valor = float(row[10]) if row[10] else 0
                                
                            # Calcular percentual baseado no valor
                            if valor_contrato > 0:
                                percentual = (valor / valor_contrato) * 100
                            else:
                                percentual = 0
                        except (ValueError, TypeError):
                            valor = 0
                            percentual = 0
                    
                    # Formatar valor e percentual para exibição
                    percentual_formatado = f"{percentual:.2f}%"
                    valor_formatado = f"R$ {valor:,.2f}"
                    
                    # Nome do evento
                    descricao = f"Administração: {row[8]}"
                    if not isinstance(descricao, str):
                        descricao = str(descricao)
                    
                    # Status e data (assumimos que não foram concluídos)
                    status = "Pendente"
                    data = ""
                    
                    # Adicionar à tabela
                    self.tree_eventos.insert('', 'end', values=(
                        evento_id,           # ID
                        descricao,           # Descrição
                        percentual_formatado,# Percentual
                        valor_formatado,     # Valor
                        status,              # Status
                        data                 # Data Conclusão
                    ))
                    
                    # Incrementar ID para o próximo evento
                    evento_id += 1
                
            workbook.close()
            
            # Se não encontramos nenhum evento, mostrar mensagem
            if evento_id == 1:
                messagebox.showinfo("Informação", "Não foram encontrados administradores para este contrato.")
            
        except Exception as e:
            print(f"Erro ao carregar eventos: {str(e)}")
            import traceback
            traceback.print_exc()  # Imprimir stack trace para depuração
            messagebox.showerror("Erro", f"Erro ao carregar eventos: {str(e)}")

    def adicionar_evento(self):
        """Adiciona um novo evento ao contrato selecionado"""
        contrato = self.contrato_selecionado.get()
        if not contrato:
            messagebox.showwarning("Aviso", "Selecione um contrato primeiro!")
            return
        
        # Limpar campos do formulário
        self.evento_descricao.delete(0, 'end')
        self.evento_percentual.delete(0, 'end')
        self.evento_status.set('pendente')
        self.evento_data.set_date(datetime.now())
        
        # Obter próximo ID disponível
        proximo_id = self.obter_proximo_id_evento(contrato)
        
        # Focar no campo de descrição
        self.evento_descricao.focus_set()
        
        # Atualizar mensagem para usuário
        messagebox.showinfo("Novo Evento", f"Adicione os detalhes do novo evento (ID: {proximo_id})")

    def obter_proximo_id_evento(self, contrato):
        """Obtém o próximo ID disponível para eventos do contrato"""
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Tentar acessar a aba de eventos
            if 'Eventos' in workbook.sheetnames:
                sheet = workbook['Eventos']
                
                # Encontrar maior ID existente
                maior_id = 0
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]) == str(contrato) and row[1]:
                        try:
                            id_evento = int(row[1])
                            if id_evento > maior_id:
                                maior_id = id_evento
                        except:
                            pass
                
                workbook.close()
                return maior_id + 1
                
            else:
                return 1
                
        except Exception as e:
            print(f"Erro ao obter próximo ID: {str(e)}")
            return 1

    def mostrar_detalhes_evento(self, event=None):
        """Exibe os detalhes do evento selecionado para edição"""
        selecao = self.tree_eventos.selection()
        if not selecao:
            return
            
        item = self.tree_eventos.item(selecao[0])
        valores = item['values']
        
        # Atualizar campos do formulário
        if len(valores) >= 6:
            # Descrição
            self.evento_descricao.delete(0, 'end')
            self.evento_descricao.insert(0, valores[1])
            
            # Percentual (remover % e espaços)
            self.evento_percentual.delete(0, 'end')
            percentual = valores[2].replace("%", "").strip() if valores[2] else ""
            self.evento_percentual.insert(0, percentual)
            
            # Status
            status = "concluido" if valores[4] == "Concluído" else "pendente"
            self.evento_status.set(status)
            
            # Data
            if valores[5] and valores[5].strip():
                try:
                    data = datetime.strptime(valores[5], "%d/%m/%Y")
                    self.evento_data.set_date(data)
                except:
                    self.evento_data.set_date(datetime.now())
            else:
                self.evento_data.set_date(datetime.now())

    def salvar_evento(self):
        """Salva um evento (novo ou atualizado)"""
        contrato = self.contrato_selecionado.get()
        if not contrato:
            messagebox.showwarning("Aviso", "Selecione um contrato primeiro!")
            return
        
        # Obter valores do formulário
        descricao = self.evento_descricao.get().strip()
        if not descricao:
            messagebox.showwarning("Aviso", "Informe a descrição do evento!")
            self.evento_descricao.focus_set()
            return
        
        try:
            percentual = float(self.evento_percentual.get().replace(",", "."))
            if percentual <= 0 or percentual > 100:
                messagebox.showwarning("Aviso", "Percentual deve estar entre 0 e 100!")
                self.evento_percentual.focus_set()
                return
        except:
            messagebox.showwarning("Aviso", "Percentual inválido!")
            self.evento_percentual.focus_set()
            return
        
        status = self.evento_status.get()
        data_conclusao = self.evento_data.get_date() if status == "concluido" else None
        
        # Verificar se estamos editando ou criando
        selecao = self.tree_eventos.selection()
        if selecao:
            # Editar evento existente
            item = self.tree_eventos.item(selecao[0])
            id_evento = item['values'][0]
            
            # Salvar no Excel
            self.salvar_evento_excel(contrato, id_evento, descricao, percentual, status, data_conclusao)
        else:
            # Criar novo evento
            id_evento = self.obter_proximo_id_evento(contrato)
            
            # Salvar no Excel
            self.salvar_evento_excel(contrato, id_evento, descricao, percentual, status, data_conclusao)
        
        # Recarregar eventos
        self.carregar_eventos_contrato()
        
        # Atualizar contagem de eventos na aba de contratos
        selecao_contrato = self.tree_contratos.selection()
        if selecao_contrato:
            self.mostrar_detalhes_contrato()

    def salvar_evento_excel(self, contrato, id_evento, descricao, percentual, status, data_conclusao):
        """Salva um evento no arquivo Excel"""
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba Contratos_ADM existe
            if 'Contratos_ADM' not in workbook.sheetnames:
                messagebox.showerror("Erro", "Aba de contratos não encontrada!")
                workbook.close()
                return False
                
            sheet = workbook['Contratos_ADM']
            
            # Como não temos uma aba específica para eventos, vamos salvar a informação
            # sobre a conclusão do evento na aba de contratos
            
            # Determinar qual linha de administrador corresponde ao evento
            linhas_administradores = []
            administradores = []
            evento_encontrado = False
            
            # Primeiro, encontrar todos os administradores do contrato
            evento_id = 1
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                # Verificar se é um administrador do contrato
                if row[6] and str(row[6]) == str(contrato):
                    if evento_id == int(id_evento):
                        # Este é o administrador correspondente ao evento
                        evento_encontrado = True
                        
                        # Se o status é "concluído", atualizar o administrador
                        if status == "concluido" and data_conclusao:
                            # Atualizar o status na planilha
                            # Como não temos uma coluna específica para status, podemos adicionar uma observação
                            # na coluna de observações ou criar uma nova coluna
                            
                            # Se já existir um valor nas observações
                            obs_atual = sheet.cell(row=row_idx, column=5).value or ""
                            nova_obs = f"{obs_atual}\nEvento concluído em {data_conclusao.strftime('%d/%m/%Y')}"
                            sheet.cell(row=row_idx, column=5).value = nova_obs
                            
                            # Gerar pagamentos para este administrador
                            self.gerar_pagamentos_evento_adaptado(contrato, id_evento, sheet, row_idx, data_conclusao)
                    evento_id += 1
            
            # Se o evento não foi encontrado
            if not evento_encontrado:
                messagebox.showwarning("Aviso", f"Evento ID {id_evento} não foi encontrado!")
                workbook.close()
                return False
            
            # Salvar arquivo
            workbook.save(self.arquivo_cliente)
            workbook.close()
            
            messagebox.showinfo("Sucesso", "Evento salvo com sucesso!")
            return True
            
        except Exception as e:
            print(f"Erro ao salvar evento: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao salvar evento: {str(e)}")
            return False

    def calcular_valor_evento(self, contrato, percentual):
        """Calcula o valor do evento com base no percentual e valor do contrato"""
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Buscar o valor total do contrato
            valor_contrato = 0
            
            if 'Contratos_ADM' in workbook.sheetnames:
                sheet = workbook['Contratos_ADM']
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]) == str(contrato):
                        # Tratar o valor conforme o tipo
                        if row[4]:
                            try:
                                if isinstance(row[4], str):
                                    valor_contrato = float(row[4].replace(',', '.'))
                                else:
                                    valor_contrato = float(row[4])
                            except (ValueError, TypeError):
                                valor_contrato = 0
                        break
            
            workbook.close()
            
            # Calcular o valor do evento
            valor_evento = (float(percentual) / 100) * valor_contrato
            return valor_evento
            
        except Exception as e:
            print(f"Erro ao calcular valor do evento: {str(e)}")
            return 0
            
    def concluir_evento(self):
        """Marca um evento como concluído"""
        selecao = self.tree_eventos.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione um evento primeiro!")
            return
        
        # Obter detalhes do evento
        item = self.tree_eventos.item(selecao[0])
        valores = item['values']
        
        if valores[4] == "Concluído":
            messagebox.showinfo("Aviso", "Este evento já está concluído!")
            return
        
        # Confirmar conclusão
        resposta = messagebox.askyesno(
            "Confirmar", 
            f"Deseja marcar o evento '{valores[1]}' como concluído?\n\n"
            f"Isso vai gerar os pagamentos correspondentes para o administrador."
        )
        
        if not resposta:
            return
        
        # Obter o contrato selecionado
        contrato = self.contrato_selecionado.get()
        if not contrato:
            messagebox.showwarning("Aviso", "Contrato não identificado!")
            return
        
        # Salvar o evento como concluído
        resultado = self.salvar_evento_excel(
            contrato=contrato,
            id_evento=valores[0],
            descricao=valores[1],
            percentual=valores[2].replace("%", ""),
            status="concluido",
            data_conclusao=datetime.now()
        )
        
        if resultado:
            # Atualizar a visualização do evento na tabela
            self.tree_eventos.item(selecao[0], values=(
                valores[0],
                valores[1],
                valores[2],
                valores[3],
                "Concluído",
                datetime.now().strftime("%d/%m/%Y")
            ))
            
            # Recarregar pagamentos
            self.carregar_pagamentos()

    def remover_evento(self):
        """Remove um evento do contrato"""
        selecao = self.tree_eventos.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione um evento primeiro!")
            return
        
        # Confirmar remoção
        resposta = messagebox.askyesno("Confirmar", "Tem certeza que deseja remover este evento? Esta ação não pode ser desfeita.")
        if not resposta:
            return
        
        # Obter detalhes do evento
        item = self.tree_eventos.item(selecao[0])
        valores = item['values']
        contrato = self.contrato_selecionado.get()
        id_evento = valores[0]
        
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            if 'Eventos' in workbook.sheetnames:
                sheet = workbook['Eventos']
                
                # Encontrar e remover o evento
                linhas_para_remover = []
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if row[0] and str(row[0]) == str(contrato) and row[1] and str(row[1]) == str(id_evento):
                        linhas_para_remover.append(row_idx)
                
                # Remover de baixo para cima para não afetar os índices
                for idx in sorted(linhas_para_remover, reverse=True):
                    sheet.delete_rows(idx)
            
            # Verificar e remover pagamentos associados
            if 'Pagamentos' in workbook.sheetnames:
                sheet = workbook['Pagamentos']
                
                # Encontrar e remover pagamentos do evento
                linhas_para_remover = []
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if row[0] and str(row[0]) == str(contrato) and row[1] and str(row[1]) == str(id_evento):
                        linhas_para_remover.append(row_idx)
                
                # Remover de baixo para cima
                for idx in sorted(linhas_para_remover, reverse=True):
                    sheet.delete_rows(idx)
            
            # Salvar arquivo
            workbook.save(self.arquivo_cliente)
            workbook.close()
            
            # Recarregar eventos
            self.carregar_eventos_contrato()
            
            # Atualizar contagem de eventos na aba de contratos
            selecao_contrato = self.tree_contratos.selection()
            if selecao_contrato:
                self.mostrar_detalhes_contrato()
            
            messagebox.showinfo("Sucesso", "Evento removido com sucesso!")
            
        except Exception as e:
            print(f"Erro ao remover evento: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao remover evento: {str(e)}")

    def gerar_pagamentos_evento_adaptado(self, contrato, id_evento, sheet, row_idx, data_conclusao):
        """Versão adaptada para gerar pagamentos com base na estrutura atual da planilha"""
        try:
            # Recuperar informações do administrador desta linha
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            
            # Extrair dados necessários
            cnpj_cpf = row[7]
            nome = row[8]
            tipo = row[9]
            valor_percentual = row[10]
            valor_total = row[11]
            num_parcelas = row[12] if row[12] else 1
            
            # Encontrar a última linha da planilha para adicionar os pagamentos
            ultima_linha = sheet.max_row + 1
            
            # Data de vencimento: 30 dias após conclusão
            data_vencimento = data_conclusao + relativedelta(days=30)
            
            # Se for pagamento único
            if num_parcelas <= 1:
                # Adicionar linha de pagamento
                sheet.cell(row=ultima_linha, column=25).value = "Evento"  # Referência
                sheet.cell(row=ultima_linha, column=26).value = id_evento  # Número
                sheet.cell(row=ultima_linha, column=27).value = cnpj_cpf  # CNPJ/CPF
                sheet.cell(row=ultima_linha, column=28).value = nome  # Nome
                sheet.cell(row=ultima_linha, column=29).value = data_vencimento  # Data Vencimento
                sheet.cell(row=ultima_linha, column=30).value = valor_total  # Valor
                sheet.cell(row=ultima_linha, column=31).value = "Pendente"  # Status
                sheet.cell(row=ultima_linha, column=32).value = None  # Data Pagamento
            else:
                # Para pagamento parcelado
                valor_parcela = float(valor_total) / int(num_parcelas)
                
                for i in range(int(num_parcelas)):
                    data_vencimento_parcela = data_vencimento + relativedelta(months=i)
                    
                    # Adicionar linha de pagamento para cada parcela
                    sheet.cell(row=ultima_linha+i, column=25).value = "Evento"  # Referência
                    sheet.cell(row=ultima_linha+i, column=26).value = id_evento  # Número
                    sheet.cell(row=ultima_linha+i, column=27).value = cnpj_cpf  # CNPJ/CPF
                    sheet.cell(row=ultima_linha+i, column=28).value = f"{nome} - Parcela {i+1}/{num_parcelas}"  # Nome
                    sheet.cell(row=ultima_linha+i, column=29).value = data_vencimento_parcela  # Data Vencimento
                    sheet.cell(row=ultima_linha+i, column=30).value = valor_parcela  # Valor
                    sheet.cell(row=ultima_linha+i, column=31).value = "Pendente"  # Status
                    sheet.cell(row=ultima_linha+i, column=32).value = None  # Data Pagamento
                    
            return True
        except Exception as e:
            print(f"Erro ao gerar pagamentos: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def registrar_pagamento(self):
        """Registra um pagamento como efetuado"""
        selecao = self.tree_pagamentos.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione um pagamento primeiro!")
            return
        
        # Obter detalhes do pagamento
        item = self.tree_pagamentos.item(selecao[0])
        valores = item['values']
        
        if valores[6] == "Pago":
            messagebox.showinfo("Aviso", "Este pagamento já está registrado como pago!")
            return
        
        # Preencher campos
        self.pagto_status_atual.set("Pago")
        self.pagto_data.set_date(datetime.now())
        
        # Perguntar se deseja salvar
        resposta = messagebox.askyesno("Confirmar", "Deseja registrar este pagamento como pago?")
        if not resposta:
            return
        
        # Salvar no Excel
        try:
            # Obter valores atualizados
            contrato = valores[0]
            evento = valores[1]
            status = self.pagto_status_atual.get()
            data_pagamento = self.pagto_data.get_date()
            cnpj_cpf = valores[2]  # CNPJ/CPF formatado
            
            # Abrir arquivo Excel
            workbook = load_workbook(self.arquivo_cliente)
            
            if 'Contratos_ADM' in workbook.sheetnames:
                sheet = workbook['Contratos_ADM']
                
                # Procurar o pagamento correspondente
                pagamento_encontrado = False
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                    # Verificar se é uma linha de pagamento (referência na coluna 25)
                    if row[24] is not None:  
                        # Verificar se corresponde ao pagamento selecionado
                        if (str(row[25]) == str(evento) and 
                            formatar_cnpj_cpf(row[26]) == cnpj_cpf):
                            
                            # Atualizar status e data de pagamento
                            sheet.cell(row=row_idx, column=31).value = status  # Status (coluna 31)
                            sheet.cell(row=row_idx, column=32).value = data_pagamento  # Data Pagamento (coluna 32)
                            
                            pagamento_encontrado = True
                            break
                
                if not pagamento_encontrado:
                    messagebox.showwarning("Aviso", "Pagamento não encontrado na planilha!")
                    workbook.close()
                    return
            else:
                messagebox.showwarning("Aviso", "Aba de contratos não encontrada!")
                workbook.close()
                return
            
            # Salvar arquivo
            workbook.save(self.arquivo_cliente)
            workbook.close()
            
            # Recarregar pagamentos
            self.carregar_pagamentos()
            
            messagebox.showinfo("Sucesso", "Pagamento registrado com sucesso!")
            
        except Exception as e:
            print(f"Erro ao registrar pagamento: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao registrar pagamento: {str(e)}")

    def gerar_lancamento_pagamento(self):
        """Gera um lançamento contábil para o pagamento selecionado"""
        selecao = self.tree_pagamentos.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione um pagamento primeiro!")
            return
        
        # Obter detalhes do pagamento
        item = self.tree_pagamentos.item(selecao[0])
        valores = item['values']
        
        # Informar que esta funcionalidade ainda não está implementada
        messagebox.showinfo("Informação", 
                            f"Função não implementada!\n\n"
                            f"Esta função irá gerar um lançamento contábil para o pagamento:\n"
                            f"Contrato: {valores[0]}\n"
                            f"Evento: {valores[1]}\n"
                            f"Fornecedor: {valores[3]}\n"
                            f"Valor: {valores[5]}")

    def filtrar_pagamentos(self, event=None):
        """Filtra a lista de pagamentos pelo contrato e status selecionados"""
        # Em vez de ocultar itens, vamos recarregar a tabela com os filtros aplicados
        try:
            # Limpar tabela
            for item in self.tree_pagamentos.get_children():
                self.tree_pagamentos.delete(item)
            
            # Obter filtros
            contrato_filtro = self.pagto_contrato.get()
            status_filtro = self.pagto_status.get()
            
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de contratos existe
            if 'Contratos_ADM' not in workbook.sheetnames:
                workbook.close()
                return
                
            sheet = workbook['Contratos_ADM']
            
            # Procurar por linhas de pagamento na planilha (colunas 25-32)
            for row in sheet.iter_rows(min_row=3, values_only=True):
                # Verificar se é uma linha de pagamento
                if row[24] is not None:  # Coluna 25 (Referência)
                    referencia = row[24]
                    numero = row[25]
                    cnpj_cpf = row[26]
                    nome = row[27]
                    data_vencimento = row[28]
                    valor = row[29]
                    status = row[30] if row[30] else "Pendente"
                    data_pagamento = row[31]
                    
                    # Encontrar o contrato associado
                    contrato = "Desconhecido"
                    if referencia == "Evento" and numero is not None:
                        # Aqui temos que fazer uma busca para determinar qual contrato
                        # está associado a este evento/administrador
                        evento_id = int(numero)
                        contrato_encontrado = False
                        
                        # Percorrer todas as linhas procurando administradores
                        evento_contador = 1
                        for adm_row in sheet.iter_rows(min_row=3, values_only=True):
                            if adm_row[6] is not None:  # É um administrador
                                if evento_contador == evento_id:
                                    contrato = adm_row[6]  # Contrato deste administrador
                                    contrato_encontrado = True
                                    break
                                evento_contador += 1
                        
                        if not contrato_encontrado:
                            contrato = f"Evento {numero}"
                    
                    # Aplicar filtros
                    if (contrato_filtro == "Todos" or str(contrato) == contrato_filtro) and \
                    (status_filtro == "Todos" or status == status_filtro):
                        
                        # Formatar data de vencimento
                        data_venc_formatada = ""
                        if data_vencimento:
                            try:
                                if isinstance(data_vencimento, datetime):
                                    data_venc_formatada = data_vencimento.strftime("%d/%m/%Y")
                                else:
                                    data_venc_formatada = str(data_vencimento)
                            except:
                                data_venc_formatada = str(data_vencimento)
                        
                        # Formatar data de pagamento
                        data_pagto_formatada = ""
                        if data_pagamento:
                            try:
                                if isinstance(data_pagamento, datetime):
                                    data_pagto_formatada = data_pagamento.strftime("%d/%m/%Y")
                                else:
                                    data_pagto_formatada = str(data_pagamento)
                            except:
                                data_pagto_formatada = str(data_pagamento)
                        
                        # Formatar valor
                        valor_formatado = ""
                        if valor:
                            try:
                                if isinstance(valor, str):
                                    valor_num = float(valor.replace(',', '.'))
                                    valor_formatado = f"R$ {valor_num:,.2f}"
                                else:
                                    valor_formatado = f"R$ {float(valor):,.2f}"
                            except (ValueError, TypeError):
                                valor_formatado = f"R$ {str(valor)}"
                        else:
                            valor_formatado = "R$ 0,00"
                        
                        # Formatar CNPJ/CPF
                        cnpj_cpf_formatado = formatar_cnpj_cpf(cnpj_cpf) if cnpj_cpf else ""
                        
                        # Adicionar à tabela
                        self.tree_pagamentos.insert('', 'end', values=(
                            contrato,            # Contrato
                            numero,              # Evento/Número
                            cnpj_cpf_formatado,  # CNPJ/CPF formatado
                            nome,                # Nome
                            data_venc_formatada, # Data vencimento formatada
                            valor_formatado,     # Valor formatado
                            status,              # Status
                            data_pagto_formatada # Data pagamento formatada
                        ))
            
            workbook.close()
            
        except Exception as e:
            print(f"Erro ao filtrar pagamentos: {str(e)}")
            import traceback
            traceback.print_exc()

    def mostrar_detalhes_contrato(self, event=None):
        """Exibe os detalhes do contrato selecionado"""
        selecao = self.tree_contratos.selection()
        if not selecao:
            return
            
        item = self.tree_contratos.item(selecao[0])
        valores = item['values']
        
        # Atualizar labels de detalhes
        if len(valores) >= 5:
            self.lbl_contrato.config(text=str(valores[0]))
            self.lbl_periodo.config(text=f"{valores[1]} a {valores[2]}")
            self.lbl_status.config(text=valores[3])
            self.lbl_valor_total.config(text=valores[4])
            
            # Carregar administradores do contrato
            self.mostrar_administradores_contrato(valores[0])
            
            # Contar eventos do contrato
            num_eventos = self.contar_eventos_contrato(valores[0])
            self.lbl_eventos.config(text=f"{num_eventos} evento(s)")
    
    def mostrar_administradores_contrato(self, num_contrato):
        """Exibe os administradores do contrato na tabela de administradores"""
        # Limpar tabela
        for item in self.tree_adm.get_children():
            self.tree_adm.delete(item)
        
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Tentar acessar a aba de administradores
            if 'Contratos_ADM' in workbook.sheetnames:
                sheet = workbook['Contratos_ADM']
                
                # Encontrar administradores deste contrato
                num_admin = 0
                total_adm = ""
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]) == str(num_contrato):
                        # Formatar CNPJ/CPF
                        cnpj_cpf = formatar_cnpj_cpf(row[1]) if row[1] else ""
                        
                        # Determinar tipo de valor (percentual ou fixo)
                        # Verificar se valor[3] é string antes de procurar "%" nele
                        tipo = "Percentual" if (row[3] and isinstance(row[3], str) and "%" in row[3]) else "Valor Fixo"
                        
                        # Formatar valor com tratamento de tipo
                        valor = ""
                        if tipo == "Percentual" and row[3] and isinstance(row[3], str):
                            valor = row[3].replace("%", "") + "%"
                        else:
                            if row[3]:
                                try:
                                    if isinstance(row[3], str):
                                        valor_formatado = float(row[3].replace(',', '.'))
                                        valor = f"R$ {valor_formatado:,.2f}"
                                    else:
                                        valor = f"R$ {float(row[3]):,.2f}"
                                except (ValueError, TypeError):
                                    valor = f"R$ {str(row[3])}"
                            else:
                                valor = "R$ 0,00"
                        
                        # Valor total com tratamento de tipo
                        valor_total = ""
                        if row[4]:
                            try:
                                if isinstance(row[4], str):
                                    valor_formatado = float(row[4].replace(',', '.'))
                                    valor_total = f"R$ {valor_formatado:,.2f}"
                                else:
                                    valor_total = f"R$ {float(row[4]):,.2f}"
                            except (ValueError, TypeError):
                                valor_total = f"R$ {str(row[4])}"
                        else:
                            valor_total = "R$ 0,00"
                        
                        # Tratar parcelas
                        parcelas = ""
                        if row[5]:
                            try:
                                parcelas = str(int(row[5]))
                            except (ValueError, TypeError):
                                parcelas = str(row[5])
                        else:
                            parcelas = "1"
                        
                        # Adicionar à tabela
                        self.tree_adm.insert('', 'end', values=(
                            cnpj_cpf,
                            row[2] if row[2] else "",  # Nome
                            tipo,
                            valor,
                            valor_total,
                            parcelas  # Nº Parcelas
                        ))
                        
                        num_admin += 1
                        
                        # Adicionar ao texto dos administradores
                        admin_nome = row[2] if row[2] else "Sem nome"
                        total_adm += (", " if total_adm else "") + admin_nome
                
                # Atualizar label de administradores
                self.lbl_administradores.config(text=f"{num_admin} ({total_adm})")
                
            else:
                self.lbl_administradores.config(text="Sem administradores")
                
            workbook.close()
            
        except Exception as e:
            print(f"Erro ao carregar administradores: {str(e)}")
            import traceback
            traceback.print_exc()  # Imprimir stack trace para depuração
            self.lbl_administradores.config(text="Erro ao carregar")

    def contar_eventos_contrato(self, num_contrato):
        """Conta quantos eventos estão associados ao contrato"""
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Tentar acessar a aba de eventos
            if 'Eventos' in workbook.sheetnames:
                sheet = workbook['Eventos']
                
                # Contar eventos deste contrato
                contador = 0
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]) == str(num_contrato):
                        contador += 1
                
                workbook.close()
                return contador
                
            else:
                return 0
                
        except Exception as e:
            print(f"Erro ao contar eventos: {str(e)}")
            return 0

    def atualizar_eventos_contrato(self):
        """Atualiza a lista de eventos do contrato selecionado"""
        selecao = self.tree_contratos.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione um contrato primeiro!")
            return
            
        # Obtém o número do contrato selecionado
        item = self.tree_contratos.item(selecao[0])
        num_contrato = item['values'][0]
        
        # Muda para a aba de eventos e seleciona o contrato no combobox
        self.notebook.select(self.aba_eventos)
        self.contrato_selecionado.set(num_contrato)
        
        # Carrega os eventos do contrato
        self.carregar_eventos_contrato()
        self.tree_pagamentos.tag_configure('hidden', hide=True)

    def carregar_pagamentos(self):
        """Carrega todos os pagamentos do cliente de forma otimizada para evitar recursão"""
        # Limpar tabela
        for item in self.tree_pagamentos.get_children():
            self.tree_pagamentos.delete(item)
        
        try:
            # Primeiro, carregar os contratos para o combobox
            contratos = ['Todos']
            
            try:
                # Usar pandas que é mais eficiente para arquivos grandes
                import pandas as pd
                df_contratos = pd.read_excel(self.arquivo_cliente, sheet_name='Contratos_ADM', engine='openpyxl')
                
                # Filtrar apenas linhas que são contratos (não administradores)
                for _, row in df_contratos.iterrows():
                    if not pd.isna(row.iloc[0]) and pd.isna(row.iloc[6]):
                        contratos.append(str(row.iloc[0]))
            except Exception as e:
                print(f"Erro ao carregar contratos com pandas: {str(e)}")
                
                # Abordagem alternativa com openpyxl
                try:
                    # Tentar carregar com openpyxl, mas limitando o processamento
                    from openpyxl import load_workbook
                    
                    # Carregar apenas a planilha Contratos_ADM
                    workbook = load_workbook(self.arquivo_cliente, read_only=True)
                    
                    if 'Contratos_ADM' in workbook.sheetnames:
                        sheet = workbook['Contratos_ADM']
                        
                        for row in sheet.iter_rows(min_row=3, max_row=100, values_only=True):
                            if row[0] and row[1] is not None and row[6] is None:
                                contratos.append(str(row[0]))
                                
                    workbook.close()
                except Exception as e2:
                    print(f"Erro ao carregar contratos com openpyxl: {str(e2)}")
            
            # Remover duplicatas e ordenar
            contratos = sorted(list(set(contratos)))
            
            # Atualizar combobox
            self.pagto_contrato['values'] = contratos
            self.pagto_contrato.set('Todos')
            
            # Carregar pagamentos - implementação simplificada
            try:
                # Usar pandas para carregar pagamentos
                import pandas as pd
                df = pd.read_excel(self.arquivo_cliente, sheet_name='Contratos_ADM', engine='openpyxl')
                
                # Filtrar apenas linhas que têm dados na coluna 25 (referência de pagamento)
                for _, row in df.iterrows():
                    if not pd.isna(row.iloc[24]):  # Tem referência de pagamento
                        try:
                            # Extrair valores importantes
                            referencia = row.iloc[24]
                            numero = row.iloc[25]
                            cnpj_cpf = row.iloc[26]
                            nome = row.iloc[27]
                            data_vencimento = row.iloc[28]
                            valor = row.iloc[29]
                            status = row.iloc[30] if not pd.isna(row.iloc[30]) else "Pendente"
                            data_pagamento = row.iloc[31]
                            
                            # Contrato - simplificado para evitar pesquisas complexas
                            contrato = "Evento " + str(numero) if not pd.isna(numero) else "Desconhecido"
                            
                            # Formatações
                            data_venc_formatada = ""
                            if not pd.isna(data_vencimento):
                                try:
                                    if isinstance(data_vencimento, pd._libs.tslibs.timestamps.Timestamp):
                                        data_venc_formatada = data_vencimento.strftime("%d/%m/%Y")
                                    else:
                                        data_venc_formatada = str(data_vencimento)
                                except:
                                    data_venc_formatada = str(data_vencimento)
                            
                            data_pagto_formatada = ""
                            if not pd.isna(data_pagamento):
                                try:
                                    if isinstance(data_pagamento, pd._libs.tslibs.timestamps.Timestamp):
                                        data_pagto_formatada = data_pagamento.strftime("%d/%m/%Y")
                                    else:
                                        data_pagto_formatada = str(data_pagamento)
                                except:
                                    data_pagto_formatada = str(data_pagamento)
                            
                            # Formatar valor
                            valor_formatado = "R$ 0,00"
                            if not pd.isna(valor):
                                try:
                                    valor_formatado = f"R$ {float(valor):,.2f}"
                                except:
                                    valor_formatado = f"R$ {str(valor)}"
                            
                            # Formatar CNPJ/CPF
                            cnpj_cpf_formatado = ""
                            if not pd.isna(cnpj_cpf):
                                cnpj_cpf_formatado = formatar_cnpj_cpf(str(cnpj_cpf))
                            
                            # Adicionar à tabela
                            self.tree_pagamentos.insert('', 'end', values=(
                                contrato,
                                numero,
                                cnpj_cpf_formatado,
                                nome,
                                data_venc_formatada,
                                valor_formatado,
                                status,
                                data_pagto_formatada
                            ))
                        except Exception as row_e:
                            print(f"Erro ao processar linha de pagamento: {row_e}")
                            continue
                            
            except Exception as e:
                print(f"Erro ao carregar pagamentos com pandas: {str(e)}")
                import traceback
                traceback.print_exc()
                
                # Tentativa mais simples, sem pandas
                try:
                    # Usar openpyxl diretamente, com limitações para evitar recursão
                    workbook = load_workbook(self.arquivo_cliente, read_only=True)
                    
                    if 'Contratos_ADM' in workbook.sheetnames:
                        sheet = workbook['Contratos_ADM']
                        
                        # Limitar o número de linhas processadas para evitar recursão
                        row_count = 0
                        max_rows = 200  # Limite para evitar problemas
                        
                        for row in sheet.iter_rows(min_row=3, values_only=True):
                            row_count += 1
                            if row_count > max_rows:
                                print(f"Limitando a {max_rows} linhas para evitar problemas")
                                break
                                
                            # Se referência não for nula, é um pagamento
                            if row[24] is not None:
                                try:
                                    # Adicionar à tabela com formatação mínima
                                    self.tree_pagamentos.insert('', 'end', values=(
                                        str(row[0]) if row[0] else "Desconhecido",  # Contrato
                                        str(row[25]) if row[25] else "",            # Evento
                                        str(row[26]) if row[26] else "",            # CNPJ/CPF
                                        str(row[27]) if row[27] else "",            # Nome
                                        str(row[28]) if row[28] else "",            # Data vencimento
                                        f"R$ {float(row[29]):,.2f}" if row[29] else "R$ 0,00",  # Valor
                                        str(row[30]) if row[30] else "Pendente",    # Status
                                        str(row[31]) if row[31] else ""             # Data pagamento
                                    ))
                                except Exception as row_e:
                                    print(f"Erro ao processar linha de pagamento simples: {row_e}")
                                    continue
                    
                    workbook.close()
                    
                except Exception as e3:
                    print(f"Erro na abordagem simplificada: {str(e3)}")
                    traceback.print_exc()
            
        except Exception as e:
            print(f"Erro geral ao carregar pagamentos: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao carregar pagamentos: {str(e)}")

    def carregar_contratos(self):
        """Carrega a lista de contratos do cliente"""
        # Limpar tabela
        for item in self.tree_contratos.get_children():
            self.tree_contratos.delete(item)
        
        try:
            # Abrir arquivo Excel do cliente
            workbook = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de contratos existe
            if 'Contratos_ADM' not in workbook.sheetnames:
                messagebox.showwarning("Aviso", "Este cliente não possui contratos cadastrados!")
                workbook.close()
                return
            
            sheet = workbook['Contratos_ADM']
            contratos_ids = []  # Para preencher o combobox de eventos
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular linhas vazias
                    continue
                    
                # Formatar datas
                data_inicio = ""
                if row[1]:
                    try:
                        if isinstance(row[1], datetime):
                            data_inicio = row[1].strftime("%d/%m/%Y")
                        else:
                            data_inicio = row[1]
                    except:
                        data_inicio = str(row[1])
                
                data_fim = ""
                if row[2]:
                    try:
                        if isinstance(row[2], datetime):
                            data_fim = row[2].strftime("%d/%m/%Y")
                        else:
                            data_fim = row[2]
                    except:
                        data_fim = str(row[2])
                
                # Formatar valor com tratamento de tipo
                valor_total = ""
                if row[4]:
                    try:
                        # Converter para float se for string
                        if isinstance(row[4], str):
                            valor_formatado = float(row[4].replace(',', '.'))
                            valor_total = f"R$ {valor_formatado:,.2f}"
                        else:
                            valor_total = f"R$ {float(row[4]):,.2f}"
                    except (ValueError, TypeError):
                        # Se não conseguir converter, mostra como string
                        valor_total = f"R$ {str(row[4])}"
                else:
                    valor_total = "R$ 0,00"
                
                # Adicionar à tabela
                self.tree_contratos.insert('', 'end', values=(
                    row[0],  # Nº Contrato
                    data_inicio,
                    data_fim,
                    row[3] if row[3] else "Em andamento",  # Status
                    valor_total
                ))
                
                # Adicionar ao array de contratos para combobox
                contratos_ids.append(str(row[0]))
            
            # Preencher combobox de contratos
            self.contrato_selecionado['values'] = contratos_ids
            
            # Preencher combobox de contratos para pagamentos
            contratos_pagto = ['Todos'] + contratos_ids
            self.pagto_contrato['values'] = contratos_pagto
            self.pagto_contrato.set('Todos')
            
            workbook.close()
            
            # Carregar pagamentos iniciais
            self.carregar_pagamentos()
            
        except Exception as e:
            print(f"Erro ao carregar contratos: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar contratos: {str(e)}")

    def mostrar_detalhes_pagamento(self, event=None):
        """Exibe os detalhes do pagamento selecionado para edição"""
        selecao = self.tree_pagamentos.selection()
        if not selecao:
            return
            
        item = self.tree_pagamentos.item(selecao[0])
        valores = item['values']
        
        # Atualizar campos do formulário
        if len(valores) >= 8:
            # CNPJ/CPF (modo somente leitura)
            self.pagto_cnpj.configure(state='normal')
            self.pagto_cnpj.delete(0, 'end')
            self.pagto_cnpj.insert(0, valores[2])
            self.pagto_cnpj.configure(state='readonly')
            
            # Nome (modo somente leitura)
            self.pagto_nome.configure(state='normal')
            self.pagto_nome.delete(0, 'end')
            self.pagto_nome.insert(0, valores[3])
            self.pagto_nome.configure(state='readonly')
            
            # Valor (modo somente leitura)
            self.pagto_valor.configure(state='normal')
            self.pagto_valor.delete(0, 'end')
            self.pagto_valor.insert(0, valores[5])
            self.pagto_valor.configure(state='readonly')
            
            # Vencimento (modo somente leitura)
            self.pagto_vencimento.configure(state='normal')
            self.pagto_vencimento.delete(0, 'end')
            self.pagto_vencimento.insert(0, valores[4])
            self.pagto_vencimento.configure(state='readonly')
            
            # Status
            self.pagto_status_atual.set(valores[6])
            
            # Data Pagamento
            if valores[7] and valores[7].strip():
                try:
                    data = datetime.strptime(valores[7], "%d/%m/%Y")
                    self.pagto_data.set_date(data)
                except:
                    self.pagto_data.set_date(datetime.now())
            else:
                self.pagto_data.set_date(datetime.now())