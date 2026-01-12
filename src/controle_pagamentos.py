import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import os
import sys
from pathlib import Path

# No início de controle_pagamentos.py
from src.correcao_monetaria import GerenciadorCorrecaoMonetaria
from src.config.utils import (
    PASTA_CLIENTES,
    formatar_moeda,
    validar_data,
    formatar_valor_excel,
    buscar_dados_bancarios_fornecedor
)

class ControlePagamentos:
    def __init__(self, parent=None):
        self.parent = parent
        # Se parent não for especificado, criar uma janela Tk, caso contrário criar uma Toplevel
        if parent is None:
            self.root = tk.Tk()
            self.is_independent = True  # Marcar como janela independente
        else:
            self.root = tk.Toplevel(parent)
            self.is_independent = False  # Marcar como janela secundária
        
        self.root.title("Controle de Pagamentos de Taxas")
        self.root.geometry("1400x900+50+50")  # Aumentado e reposicionado
        
        # Forçar a janela para frente
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after(100, lambda: self.root.attributes('-topmost', False))
        
        # Variáveis de controle
        self.cliente_selecionado = None
        self.parcelas_selecionadas = []
        self.scrollbar_y = None
        self.scrollbar_x = None
        self.valor_editado = {}  # Dicionário para armazenar valores editados
        
        self.setup_gui()
    
    @staticmethod
    def converter_valor_seguro(valor_celula):
        """
        Converte valor de célula Excel para float de forma segura.
        Trata strings formatadas como moeda brasileira e valores numéricos.
        
        Args:
            valor_celula: Valor da célula (pode ser int, float, string ou None)
            
        Returns:
            float: Valor convertido ou 0.0 se conversão falhar
        """
        try:
            if valor_celula is None:
                return 0.0
            
            # Se já for número, converter para float
            if isinstance(valor_celula, (int, float)):
                return float(valor_celula)
            
            # Se for string, limpar formatação brasileira
            if isinstance(valor_celula, str):
                # Remove R$, pontos de milhar, substitui vírgula decimal por ponto
                valor_limpo = valor_celula.replace('R$', '').replace('.', '').replace(',', '.').strip()
                if valor_limpo:
                    return float(valor_limpo)
                return 0.0
            
            # Outros tipos: tentar conversão direta
            return float(valor_celula)
            
        except (ValueError, TypeError, AttributeError):
            return 0.0
    
    def setup_gui(self):
        # Frame principal SEM scroll para melhor controle do layout
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Frame seleção de cliente
        frame_cliente = ttk.LabelFrame(main_frame, text="Selecione o Cliente")
        frame_cliente.pack(fill='x', pady=5)
        
        self.cliente_combo = ttk.Combobox(frame_cliente, state='readonly', width=50)
        self.cliente_combo.pack(side='left', padx=5, pady=5)
        self.cliente_combo.bind('<<ComboboxSelected>>', self.carregar_parcelas)
        
        # Frame lista de parcelas - EXPANDIDO para usar todo espaço disponível
        self.frame_parcelas = ttk.LabelFrame(main_frame, text="Parcelas Pendentes")
        self.frame_parcelas.pack(fill='both', expand=True, pady=5)
        
        # Container para treeview e scrollbars
        self.tree_container = ttk.Frame(self.frame_parcelas)
        self.tree_container.pack(fill='both', expand=True, padx=5, pady=5)
        
        # ============== MUDANÇA PRINCIPAL: ATIVAR SELEÇÃO MÚLTIPLA ==============
        # Treeview com coluna adicional para valor editado E SELECTMODE EXTENDED
        colunas = ('Nº Contrato', 'Nº Parcela', 'CNPJ', 'Adm', 'Eventos/Fases', 
                   'Valor Original', 'Valor a Pagar', 'Status', 'Data Pagamento')
        self.tree_parcelas = ttk.Treeview(self.tree_container, columns=colunas, 
                                         show='headings', selectmode='extended')  # <-- MUDANÇA AQUI
        
        # Configurar colunas com larguras proporcionais ao espaço disponível
        larguras_iniciais = {
            'Nº Contrato': 90,
            'Nº Parcela': 80,
            'CNPJ': 130,
            'Adm': 200,
            'Eventos/Fases': 350,
            'Valor Original': 120,
            'Valor a Pagar': 120,
            'Status': 100,
            'Data Pagamento': 120
        }

        for col in colunas:
            self.tree_parcelas.heading(col, text=col)
            self.tree_parcelas.column(col, width=larguras_iniciais.get(col, 100), minwidth=50)
        
        # Scrollbars
        self.scrollbar_y = ttk.Scrollbar(self.tree_container, orient='vertical',
                                       command=self.tree_parcelas.yview)
        self.scrollbar_x = ttk.Scrollbar(self.tree_container, orient='horizontal',
                                       command=self.tree_parcelas.xview)
        
        # Configurar treeview
        self.tree_parcelas.configure(yscrollcommand=self.scrollbar_y.set,
                                   xscrollcommand=self.scrollbar_x.set)
        
        # Grid layout para treeview e scrollbars
        self.tree_parcelas.grid(row=0, column=0, sticky='nsew')
        self.scrollbar_y.grid(row=0, column=1, sticky='ns')
        self.scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        # Configurar grid weights
        self.tree_container.grid_rowconfigure(0, weight=1)
        self.tree_container.grid_columnconfigure(0, weight=1)
        
        # ============== NOVA FUNCIONALIDADE: BIND PARA ATUALIZAR SELEÇÃO ==============
        self.tree_parcelas.bind('<<TreeviewSelect>>', self.atualizar_info_selecao)
        
        # Bind para duplo clique editar valor
        self.tree_parcelas.bind('<Double-Button-1>', self.editar_valor_parcela)
        
        # ============== NOVO FRAME: INFORMAÇÕES DE SELEÇÃO MÚLTIPLA ==============
        frame_info_selecao = ttk.LabelFrame(main_frame, text="Parcelas Selecionadas")
        frame_info_selecao.pack(fill='x', pady=5)
        
        info_container = ttk.Frame(frame_info_selecao)
        info_container.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(info_container, text="Quantidade:").grid(row=0, column=0, padx=5, sticky='w')
        self.label_qtd_selecionadas = ttk.Label(info_container, text="0", 
                                               font=('Arial', 10, 'bold'), foreground='blue')
        self.label_qtd_selecionadas.grid(row=0, column=1, padx=5, sticky='w')
        
        ttk.Label(info_container, text="Valor Total:").grid(row=0, column=2, padx=20, sticky='w')
        self.label_valor_total_selecionadas = ttk.Label(info_container, text="R$ 0,00", 
                                                       font=('Arial', 10, 'bold'), foreground='green')
        self.label_valor_total_selecionadas.grid(row=0, column=3, padx=5, sticky='w')
        
        ttk.Button(info_container, text="Limpar Seleção", 
                  command=self.limpar_selecao).grid(row=0, column=4, padx=20)
        
        # Frame para edição de valores selecionados
        frame_edicao = ttk.LabelFrame(main_frame, text="Edição de Valores")
        frame_edicao.pack(fill='x', pady=5)
        
        # Subframe para organizar componentes
        subframe_edicao = ttk.Frame(frame_edicao)
        subframe_edicao.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(subframe_edicao, text="Parcela Selecionada:").grid(row=0, column=0, padx=5, sticky='w')
        self.label_parcela_selecionada = ttk.Label(subframe_edicao, text="Nenhuma", font=('Arial', 9, 'bold'))
        self.label_parcela_selecionada.grid(row=0, column=1, padx=5, sticky='w')
        
        ttk.Label(subframe_edicao, text="Valor Original:").grid(row=0, column=2, padx=5, sticky='w')
        self.label_valor_original = ttk.Label(subframe_edicao, text="R$ 0,00", font=('Arial', 9))
        self.label_valor_original.grid(row=0, column=3, padx=5, sticky='w')
        
        ttk.Label(subframe_edicao, text="Novo Valor:").grid(row=0, column=4, padx=5, sticky='w')
        self.entry_novo_valor = ttk.Entry(subframe_edicao, width=15)
        self.entry_novo_valor.grid(row=0, column=5, padx=5)
        
        ttk.Button(subframe_edicao, text="Aplicar", 
                  command=self.aplicar_valor_editado).grid(row=0, column=6, padx=5)
        
        ttk.Button(subframe_edicao, text="Resetar Todos", 
                  command=self.resetar_valores).grid(row=0, column=7, padx=5)
        
        # Frame para registrar pagamento - POSICIONADO MAIS ACIMA
        frame_pagamento = ttk.LabelFrame(main_frame, text="Registrar Pagamento")
        frame_pagamento.pack(fill='x', pady=5)
        
        # Container interno para melhor organização
        container_pagamento = ttk.Frame(frame_pagamento)
        container_pagamento.pack(fill='x', padx=5, pady=10)
        
        ttk.Label(container_pagamento, text="Data do Pagamento:").pack(side='left', padx=5)
        
        # DateEntry com configuração para garantir visibilidade
        self.data_pagamento = DateEntry(container_pagamento, width=12, locale='pt_BR',
                                      background='darkblue', foreground='white',
                                      borderwidth=2, date_pattern='dd/mm/yyyy',
                                      showweeknumbers=False)
        self.data_pagamento.pack(side='left', padx=5)
        
        # ============== BOTÃO ATUALIZADO PARA MÚLTIPLAS PARCELAS ==============
        ttk.Button(container_pagamento, text="Registrar Pagamento das Parcelas Selecionadas",
                  command=self.registrar_pagamento_multiplo).pack(side='left', padx=20)
        
        # NOVO: Botão de Vincular
        ttk.Button(container_pagamento, text="Vincular a Lançamento Existente",
                  command=self.vincular_parcelas_multiplas).pack(side='left', padx=5)
        
        # Label informativo atualizado
        self.label_info = ttk.Label(container_pagamento, 
                                   text="(Dica: Use Ctrl+Click ou Shift+Click para selecionar múltiplas parcelas)",
                                   foreground='#666')
        self.label_info.pack(side='left', padx=20)
        
        # Frame de botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Voltar ao Menu", 
                  command=self.voltar_menu).pack(side='right', padx=5)
        
        self.carregar_clientes()
        self.verificar_correcoes_pendentes()
    
    # ============== NOVAS FUNÇÕES PARA SELEÇÃO MÚLTIPLA ==============
    
    def atualizar_info_selecao(self, event=None):
        """Atualiza as informações sobre as parcelas selecionadas"""
        try:
            selecionados = self.tree_parcelas.selection()
            qtd = len(selecionados)
            
            # Atualizar quantidade
            self.label_qtd_selecionadas.config(text=str(qtd))
            
            # Calcular valor total
            valor_total = 0
            for item_id in selecionados:
                valores = self.tree_parcelas.item(item_id)['values']
                
                # Pegar o valor a pagar (coluna 6)
                valor_str = str(valores[6]).replace('R$', '').replace('.', '').replace(',', '.').strip()
                try:
                    valor = float(valor_str)
                    valor_total += valor
                except:
                    pass
            
            # Atualizar label de valor total
            self.label_valor_total_selecionadas.config(text=f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
            
        except Exception as e:
            print(f"Erro ao atualizar info de seleção: {str(e)}")
    
    def limpar_selecao(self):
        """Limpa todas as seleções"""
        for item in self.tree_parcelas.selection():
            self.tree_parcelas.selection_remove(item)
        self.atualizar_info_selecao()
    
    def registrar_pagamento_multiplo(self):
        """Registra pagamento para MÚLTIPLAS parcelas selecionadas"""
        try:
            # Verificar se há parcelas selecionadas
            selecionados = self.tree_parcelas.selection()
            if not selecionados:
                messagebox.showwarning("Aviso", "Selecione ao menos uma parcela para registrar o pagamento!")
                return
            
            # Obter data do pagamento
            data_pag = self.data_pagamento.get_date()
            
            # Coletar informações de todas as parcelas selecionadas
            parcelas_info = []
            valor_total = 0
            
            for item_id in selecionados:
                valores = self.tree_parcelas.item(item_id)['values']
                
                num_contrato = str(valores[0])
                num_parcela = int(valores[1])
                cnpj = str(valores[2])
                nome_adm = str(valores[3])
                
                # Valor a pagar
                valor_str = str(valores[6]).replace('R$', '').replace('.', '').replace(',', '.').strip()
                valor_pagar = float(valor_str)
                valor_total += valor_pagar
                
                parcelas_info.append({
                    'num_contrato': num_contrato,
                    'num_parcela': num_parcela,
                    'cnpj': cnpj,
                    'nome': nome_adm,
                    'valor': valor_pagar
                })
            
            # Confirmar com o usuário
            lista_parcelas = "\n".join([
                f"• Contrato {p['num_contrato']} - Parcela {p['num_parcela']}: R$ {p['valor']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                for p in parcelas_info
            ])
            
            resposta = messagebox.askyesno(
                "Confirmar Pagamento Múltiplo",
                f"Confirma o registro de pagamento para {len(parcelas_info)} parcela(s)?\n\n"
                f"Data do Pagamento: {data_pag.strftime('%d/%m/%Y')}\n\n"
                f"Parcelas:\n{lista_parcelas}\n\n"
                f"VALOR TOTAL: R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                parent=self.root
            )
            
            if not resposta:
                return
            
            # Abrir arquivo do cliente
            arquivo_cliente = Path(PASTA_CLIENTES) / f"{self.cliente_selecionado}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_contratos = wb['Contratos_ADM']
            ws_dados = wb['Dados']
            
            # Processar cada parcela
            parcelas_processadas = 0
            
            for parcela_info in parcelas_info:
                # Atualizar status na aba Contratos_ADM
                for row_idx, row in enumerate(ws_contratos.iter_rows(min_row=3), start=3):
                    if (str(row[24].value) == parcela_info['num_contrato'] and
                        int(row[25].value) == parcela_info['num_parcela'] and
                        str(row[26].value) == parcela_info['cnpj']):
                        
                        # Atualizar status e data
                        ws_contratos.cell(row=row_idx, column=31, value="PAGO")
                        ws_contratos.cell(row=row_idx, column=32, value=data_pag)
                        # Aplicar formato de data
                        ws_contratos.cell(row=row_idx, column=32).number_format = 'DD/MM/YYYY'
                        
                        parcelas_processadas += 1
                        break
            
            # ✅ CORREÇÃO: Criar UM lançamento POR ADMINISTRADOR (agrupar por CNPJ)
            # Agrupar parcelas por administrador
            parcelas_por_admin = {}
            
            for parcela in parcelas_info:
                cnpj = parcela['cnpj']
                if cnpj not in parcelas_por_admin:
                    parcelas_por_admin[cnpj] = {
                        'nome': parcela['nome'],
                        'parcelas': [],
                        'valor_total': 0
                    }
                
                parcelas_por_admin[cnpj]['parcelas'].append(parcela)
                parcelas_por_admin[cnpj]['valor_total'] += parcela['valor']
            
            # Criar um lançamento para cada administrador
            lancamentos_criados = 0
            
            for cnpj, dados_admin in parcelas_por_admin.items():
                nova_linha = ws_dados.max_row + 1
                
                # Preencher dados do lançamento
                ws_dados.cell(row=nova_linha, column=1, value=data_pag)  # DATA_REL
                ws_dados.cell(row=nova_linha, column=1).number_format = 'DD/MM/YYYY'
                
                ws_dados.cell(row=nova_linha, column=2, value=2)  # TP_DESP (2 = Transferências e boletos)
                
                ws_dados.cell(row=nova_linha, column=3, value=cnpj)  # CNPJ_CPF
                
                # Nome SEM número de parcela
                ws_dados.cell(row=nova_linha, column=4, value=dados_admin['nome'])  # NOME
                
                # ✅ CORREÇÃO: Referência no formato "TAXA ADM - num_parcela/total"
                parcelas_list = dados_admin['parcelas']
                parcelas_nums = [str(p['num_parcela']) for p in parcelas_list]
                
                # ✅ CORREÇÃO: Descobrir o total de parcelas ÚNICO do contrato
                # Usar conjunto (set) para evitar contar duplicatas (múltiplos admins)
                num_contrato_ref = parcelas_list[0]['num_contrato']
                parcelas_unicas = set()
                
                for row in ws_contratos.iter_rows(min_row=3):
                    if str(row[24].value) == num_contrato_ref:  # Coluna Y - Referência
                        num_parcela = row[25].value  # Coluna Z - Número
                        if num_parcela:
                            parcelas_unicas.add(num_parcela)
                
                # Total = maior número de parcela única
                total_parcelas = max(parcelas_unicas) if parcelas_unicas else 0
                
                # Montar referência com formato correto
                if len(parcelas_nums) == 1:
                    referencia = f"TAXA ADM - {parcelas_nums[0]}/{total_parcelas}"
                else:
                    referencia = f"TAXA ADM - {','.join(parcelas_nums)}/{total_parcelas}"
                
                ws_dados.cell(row=nova_linha, column=5, value=referencia)  # REFERÊNCIA
                
                # Preencher VR_UNIT (coluna 7) com valor_total
                ws_dados.cell(row=nova_linha, column=7, value=dados_admin['valor_total'])  # VR_UNIT
                ws_dados.cell(row=nova_linha, column=7).number_format = '#,##0.00'
                
                # Preencher DIAS (coluna 8) com 1
                ws_dados.cell(row=nova_linha, column=8, value=1)  # DIAS
                
                # VALOR - apenas deste administrador (coluna 9)
                ws_dados.cell(row=nova_linha, column=9, value=dados_admin['valor_total'])  # VALOR
                ws_dados.cell(row=nova_linha, column=9).number_format = '#,##0.00'
                
                # Data de vencimento (mesma data do pagamento)
                ws_dados.cell(row=nova_linha, column=10, value=data_pag)  # DT_VENCTO
                ws_dados.cell(row=nova_linha, column=10).number_format = 'DD/MM/YYYY'
                
                # Categoria
                ws_dados.cell(row=nova_linha, column=11, value="TAX")  # CATEGORIA
                
                # Buscar dados bancários deste administrador específico
                dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj, self.cliente_selecionado)
                if dados_bancarios:
                    ws_dados.cell(row=nova_linha, column=12, value=dados_bancarios)  # DADOS_BANCARIOS
                
                # Observação
                ws_dados.cell(row=nova_linha, column=13, value="LANÇAMENTO AUTOMÁTICO - Pagamento Registrado")  # OBSERVAÇÃO
                
                # Status = "ATIVO"
                ws_dados.cell(row=nova_linha, column=14, value="ATIVO")  # STATUS
                
                lancamentos_criados += 1
            
            # Salvar arquivo
            wb.save(arquivo_cliente)
            wb.close()
            
            # Mensagem de sucesso detalhada
            detalhes_admins = "\n".join([
                f"  • {dados['nome']}: R$ {dados['valor_total']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                for dados in parcelas_por_admin.values()
            ])
            
            messagebox.showinfo(
                "Sucesso",
                f"Pagamento registrado com sucesso!\n\n"
                f"✓ {parcelas_processadas} parcela(s) marcada(s) como PAGO\n"
                f"✓ {lancamentos_criados} lançamento(s) criado(s) na aba Dados\n\n"
                f"Lançamentos por administrador:\n{detalhes_admins}\n\n"
                f"VALOR TOTAL: R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                parent=self.root
            )
            
            # Recarregar parcelas
            self.carregar_parcelas()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao registrar pagamento: {str(e)}")
            if 'wb' in locals():
                wb.close()
    
    def vincular_parcelas_multiplas(self):
        """Vincula MÚLTIPLAS parcelas a um lançamento existente"""
        try:
            # Verificar seleção
            selecionados = self.tree_parcelas.selection()
            if not selecionados:
                messagebox.showwarning("Aviso", "Selecione ao menos uma parcela para vincular!")
                return
            
            # Coletar dados das parcelas selecionadas
            parcelas_dados = []
            valor_total = 0
            
            for item_id in selecionados:
                valores = self.tree_parcelas.item(item_id)['values']
                
                num_contrato = str(valores[0])
                num_parcela = int(valores[1])
                cnpj = str(valores[2])
                nome_adm = str(valores[3])
                
                valor_str = str(valores[6]).replace('R$', '').replace('.', '').replace(',', '.').strip()
                valor_pagar = float(valor_str)
                valor_total += valor_pagar
                
                parcelas_dados.append({
                    'num_contrato': num_contrato,
                    'num_parcela': num_parcela,
                    'cnpj': cnpj,
                    'nome': nome_adm,
                    'valor_pagar': f"R$ {valor_pagar:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                })
            
            # Criar janela de busca
            janela_busca = tk.Toplevel(self.root)
            janela_busca.title("Buscar Lançamento para Vincular")
            janela_busca.geometry("1200x600")
            janela_busca.transient(self.root)
            janela_busca.grab_set()
            
            # Frame de busca
            frame_busca = ttk.LabelFrame(janela_busca, text="Critérios de Busca")
            frame_busca.pack(fill='x', padx=10, pady=10)
            
            busca_container = ttk.Frame(frame_busca)
            busca_container.pack(fill='x', padx=5, pady=5)
            
            ttk.Label(busca_container, text="Nome/Fornecedor:").grid(row=0, column=0, padx=5, sticky='w')
            entry_nome = ttk.Entry(busca_container, width=30)
            entry_nome.grid(row=0, column=1, padx=5)
            entry_nome.insert(0, parcelas_dados[0]['nome'])
            
            ttk.Label(busca_container, text="Valor Total:").grid(row=0, column=2, padx=5, sticky='w')
            entry_valor = ttk.Entry(busca_container, width=15)
            entry_valor.grid(row=0, column=3, padx=5)
            entry_valor.insert(0, f"{valor_total:.2f}".replace('.', ','))
            
            var_valor_aprox = tk.BooleanVar(value=True)
            ttk.Checkbutton(busca_container, text="Valor aproximado (±5%)", 
                           variable=var_valor_aprox).grid(row=0, column=4, padx=5)
            
            ttk.Button(busca_container, text="Buscar", 
                      command=lambda: self.buscar_lancamentos(entry_nome, entry_valor, 
                                                             var_valor_aprox, tree_lancamentos)
                      ).grid(row=0, column=5, padx=10)
            
            # Info sobre parcelas selecionadas
            frame_info = ttk.LabelFrame(janela_busca, text="Parcelas a Vincular")
            frame_info.pack(fill='x', padx=10, pady=5)
            
            info_text = f"Quantidade: {len(parcelas_dados)} | Valor Total: R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ttk.Label(frame_info, text=info_text, font=('Arial', 10, 'bold')).pack(padx=5, pady=5)
            
            # Frame resultados
            frame_resultados = ttk.LabelFrame(janela_busca, text="Lançamentos Encontrados")
            frame_resultados.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Treeview
            colunas_lanc = ('Linha', 'Data', 'Nome', 'CNPJ', 'Valor', 'Status')
            tree_lancamentos = ttk.Treeview(frame_resultados, columns=colunas_lanc, show='headings')
            
            for col in colunas_lanc:
                tree_lancamentos.heading(col, text=col)
                tree_lancamentos.column(col, width=150)
            
            scrollbar_lanc = ttk.Scrollbar(frame_resultados, orient='vertical', 
                                          command=tree_lancamentos.yview)
            tree_lancamentos.configure(yscrollcommand=scrollbar_lanc.set)
            
            tree_lancamentos.pack(side='left', fill='both', expand=True)
            scrollbar_lanc.pack(side='right', fill='y')
            
            # Botões
            frame_botoes_vinc = ttk.Frame(janela_busca)
            frame_botoes_vinc.pack(fill='x', padx=10, pady=10)
            
            ttk.Button(frame_botoes_vinc, text="Vincular Selecionado", 
                      command=lambda: self.confirmar_vinculacao_multipla(
                          janela_busca, parcelas_dados, tree_lancamentos)
                      ).pack(side='left', padx=5)
            
            ttk.Button(frame_botoes_vinc, text="Cancelar", 
                      command=janela_busca.destroy).pack(side='right', padx=5)
            
            # Busca automática inicial
            self.buscar_lancamentos(entry_nome, entry_valor, var_valor_aprox, tree_lancamentos)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao vincular parcelas: {str(e)}")
    
    def confirmar_vinculacao_multipla(self, janela, parcelas_dados, tree):
        """Confirma e executa a vinculação de MÚLTIPLAS parcelas ao lançamento selecionado"""
        try:
            # Verificar seleção
            selecao = tree.selection()
            if not selecao:
                messagebox.showwarning("Aviso", "Selecione um lançamento para vincular!", parent=janela)
                return
            
            # Obter dados do lançamento
            item = tree.item(selecao[0])
            valores = item['values']
            linha_lancamento = valores[0]
            nome_lancamento = valores[2]
            valor_lancamento = valores[4]
            
            # Montar texto de confirmação
            lista_parcelas = "\n".join([
                f"• Contrato {p['num_contrato']} - Parcela {p['num_parcela']} ({p['valor_pagar']})"
                for p in parcelas_dados
            ])
            
            resposta = messagebox.askyesno(
                "Confirmar Vinculação Múltipla",
                f"Confirma a vinculação de {len(parcelas_dados)} parcela(s)?\n\n"
                f"PARCELAS:\n{lista_parcelas}\n\n"
                f"SERÃO VINCULADAS AO LANÇAMENTO:\n"
                f"Linha: {linha_lancamento}\n"
                f"Nome: {nome_lancamento}\n"
                f"Valor: {valor_lancamento}",
                parent=janela
            )
            
            if not resposta:
                return
            
            # Executar vinculação
            arquivo_cliente = Path(PASTA_CLIENTES) / f"{self.cliente_selecionado}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_contratos = wb['Contratos_ADM']
            
            hoje = datetime.now()
            parcelas_vinculadas = 0
            
            # Atualizar cada parcela
            for parcela in parcelas_dados:
                for row_idx, row in enumerate(ws_contratos.iter_rows(min_row=3), start=3):
                    if (str(row[24].value) == parcela['num_contrato'] and
                        int(row[25].value) == parcela['num_parcela'] and
                        str(row[26].value) == parcela['cnpj']):
                        
                        # Atualizar status
                        ws_contratos.cell(row=row_idx, column=31, value="VINCULADO")
                        ws_contratos.cell(row=row_idx, column=32, value=hoje)
                        
                        # Adicionar observação
                        obs_atual = ws_contratos.cell(row=row_idx, column=36).value or ""
                        nova_obs = f"{obs_atual} [VINCULADO À DESPESA DA LINHA {linha_lancamento} DE DADOS]".strip()
                        ws_contratos.cell(row=row_idx, column=36, value=nova_obs)
                        
                        parcelas_vinculadas += 1
                        break
            
            # Salvar
            wb.save(arquivo_cliente)
            wb.close()
            
            # Sucesso
            messagebox.showinfo(
                "Sucesso",
                f"Vinculação concluída!\n\n"
                f"✓ {parcelas_vinculadas} parcela(s) vinculada(s)\n"
                f"✓ Status: VINCULADO\n"
                f"✓ Linha do lançamento: {linha_lancamento}",
                parent=janela
            )
            
            # Fechar e atualizar
            janela.destroy()
            self.carregar_parcelas()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao confirmar vinculação: {str(e)}", parent=janela)
            if 'wb' in locals():
                wb.close()
    
    # ============== FUNÇÕES MANTIDAS DO CÓDIGO ORIGINAL ==============
    # (As demais funções permanecem iguais ao código original)
    
    def carregar_clientes(self):
        """Carrega lista de clientes ativos usando função de utils.py"""
        try:
            # Importar função de utils se ainda não foi importada
            from src.config.utils import obter_clientes_ativos
            
            # Obter apenas clientes ativos
            clientes, info_clientes = obter_clientes_ativos(mostrar_inativos=False)
            
            if not clientes:
                messagebox.showwarning("Aviso", "Nenhum cliente ativo encontrado!")
                return
            
            # Atualizar combobox com clientes ativos
            self.cliente_combo['values'] = clientes
            
            # Armazenar informações dos clientes para referência futura
            self.info_clientes = info_clientes
            
        except ImportError:
            # Fallback para método antigo se utils.py não estiver disponível
            messagebox.showwarning("Aviso", 
                "Módulo utils.py não encontrado. Carregando todos os clientes do diretório.\n" +
                "Para filtrar apenas clientes ativos, certifique-se de que src.config.utils está acessível.")
            
            if not os.path.exists(PASTA_CLIENTES):
                messagebox.showerror("Erro", f"Pasta de clientes não encontrada: {PASTA_CLIENTES}")
                return
            
            clientes = []
            for arquivo in os.listdir(PASTA_CLIENTES):
                if arquivo.endswith('.xlsx') and not arquivo.startswith('~'):
                    cliente = arquivo.replace('.xlsx', '')
                    clientes.append(cliente)
            
            if not clientes:
                messagebox.showwarning("Aviso", "Nenhum cliente encontrado!")
                return
            
            clientes.sort()
            self.cliente_combo['values'] = clientes
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
    
    def carregar_parcelas(self, event=None):
        """Carrega parcelas pendentes do cliente selecionado (apenas de contratos ATIVOS)"""
        try:
            # Limpar treeview
            for item in self.tree_parcelas.get_children():
                self.tree_parcelas.delete(item)
            
            # Limpar valores editados
            self.valor_editado.clear()
            
            cliente = self.cliente_combo.get()
            if not cliente:
                return
            
            self.cliente_selecionado = cliente
            
            arquivo_cliente = Path(PASTA_CLIENTES) / f"{cliente}.xlsx"
            if not arquivo_cliente.exists():
                messagebox.showerror("Erro", f"Arquivo do cliente não encontrado: {arquivo_cliente}")
                return
            
            wb = load_workbook(arquivo_cliente, data_only=True)
            ws = wb['Contratos_ADM']
            
            # ✅ PASSO 1: Identificar quais contratos estão ATIVOS
            # Ler a área de CONTRATOS (colunas A-E) para saber o status
            contratos_ativos = set()
            
            for row in ws.iter_rows(min_row=3):
                num_contrato = row[0].value  # Coluna A - Nº Contrato
                status_contrato = row[3].value  # Coluna D - Status
                
                # Se tem número de contrato e status é ATIVO
                if num_contrato and status_contrato == 'ATIVO':
                    contratos_ativos.add(str(num_contrato).strip().upper())
            
            # ✅ PASSO 2: Carregar parcelas APENAS de contratos ATIVOS
            # Usar um set para evitar duplicação de parcelas
            # Chave: (num_contrato, num_parcela, cnpj)
            parcelas_processadas = set()
            
            for row in ws.iter_rows(min_row=3):
                # Área de PARCELAS (colunas Y-AH / índices 24-33)
                num_contrato = row[24].value  # Coluna Y - Referência (Contrato)
                if not num_contrato:
                    continue
                
                # ✅ FILTRO 1: Verificar se o contrato está ATIVO
                contrato_key = str(num_contrato).strip().upper()
                if contrato_key not in contratos_ativos:
                    continue  # Pular parcelas de contratos inativos
                
                num_parcela = row[25].value  # Coluna Z - Número
                cnpj = row[26].value  # Coluna AA - CNPJ/CPF
                nome_adm = row[27].value  # Coluna AB - Nome
                eventos = row[32].value or ""  # Coluna AG - Eventos/Fases
                valor_original = row[29].value  # Coluna AD - Valor
                status = row[30].value or "PENDENTE"  # Coluna AE - Status
                data_pagamento = row[31].value  # Coluna AF - Data Pagamento
                
                # ✅ FILTRO 2: Mostrar apenas parcelas PENDENTES
                if status not in ["PENDENTE", None, ""]:
                    continue
                
                # ✅ FILTRO 3: Evitar duplicação
                # Cada parcela deve aparecer UMA VEZ, mesmo que tenha múltiplos administradores
                # Chave única: contrato + número da parcela + cnpj do administrador
                chave_parcela = (contrato_key, num_parcela, str(cnpj).strip() if cnpj else "")
                
                if chave_parcela in parcelas_processadas:
                    continue  # Já foi processada
                
                parcelas_processadas.add(chave_parcela)
                
                # Formatar valores
                valor_original_fmt = f"R$ {valor_original:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_original else "R$ 0,00"
                valor_pagar_fmt = valor_original_fmt  # Inicialmente igual ao original
                
                # Formatar data de pagamento (se existir e for uma data válida)
                data_pag_fmt = ""
                if data_pagamento:
                    if isinstance(data_pagamento, (date, datetime)):
                        data_pag_fmt = data_pagamento.strftime('%d/%m/%Y')
                    elif isinstance(data_pagamento, str):
                        try:
                            data_obj = datetime.strptime(data_pagamento, '%Y-%m-%d')
                            data_pag_fmt = data_obj.strftime('%d/%m/%Y')
                        except:
                            data_pag_fmt = str(data_pagamento)
                
                # Inserir no treeview
                self.tree_parcelas.insert('', 'end', values=(
                    num_contrato,
                    num_parcela,
                    cnpj,
                    nome_adm,
                    eventos,
                    valor_original_fmt,
                    valor_pagar_fmt,
                    status,
                    data_pag_fmt
                ))
            
            wb.close()
            
            # Atualizar informações de seleção
            self.atualizar_info_selecao()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar parcelas: {str(e)}")
            if 'wb' in locals():
                wb.close()
    
    def editar_valor_parcela(self, event):
        """Permite editar o valor de uma parcela específica"""
        try:
            # Obter item selecionado
            item_id = self.tree_parcelas.focus()
            if not item_id:
                return
            
            valores = self.tree_parcelas.item(item_id)['values']
            
            num_contrato = str(valores[0])
            num_parcela = int(valores[1])
            
            # Valor original
            valor_original_str = str(valores[5]).replace('R$', '').replace('.', '').replace(',', '.').strip()
            valor_original = float(valor_original_str)
            
            # Atualizar labels
            self.label_parcela_selecionada.config(text=f"Contrato {num_contrato} - Parcela {num_parcela}")
            self.label_valor_original.config(text=valores[5])
            
            # Limpar e focar no campo de novo valor
            self.entry_novo_valor.delete(0, tk.END)
            self.entry_novo_valor.focus()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao editar valor: {str(e)}")
    
    def aplicar_valor_editado(self):
        """Aplica o novo valor editado à parcela"""
        try:
            item_id = self.tree_parcelas.focus()
            if not item_id:
                messagebox.showwarning("Aviso", "Selecione uma parcela primeiro!")
                return
            
            novo_valor_str = self.entry_novo_valor.get().replace(',', '.')
            if not novo_valor_str:
                messagebox.showwarning("Aviso", "Digite um valor!")
                return
            
            novo_valor = float(novo_valor_str)
            
            valores = self.tree_parcelas.item(item_id)['values']
            chave = f"{valores[0]}_{valores[1]}_{valores[2]}"  # contrato_parcela_cnpj
            
            # Armazenar valor editado
            self.valor_editado[chave] = novo_valor
            
            # Atualizar treeview
            novo_valor_fmt = f"R$ {novo_valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            self.tree_parcelas.item(item_id, values=(
                valores[0], valores[1], valores[2], valores[3],
                valores[4], valores[5], novo_valor_fmt, valores[7], valores[8]
            ))
            
            # Limpar campos
            self.entry_novo_valor.delete(0, tk.END)
            self.label_parcela_selecionada.config(text="Nenhuma")
            self.label_valor_original.config(text="R$ 0,00")
            
            # Atualizar soma se houver múltiplas selecionadas
            self.atualizar_info_selecao()
            
            messagebox.showinfo("Sucesso", "Valor atualizado!")
            
        except ValueError:
            messagebox.showerror("Erro", "Valor inválido!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao aplicar valor: {str(e)}")
    
    def resetar_valores(self):
        """Reseta todos os valores editados para os originais"""
        try:
            if not self.valor_editado:
                messagebox.showinfo("Info", "Não há valores editados para resetar.")
                return
            
            if messagebox.askyesno("Confirmar", "Deseja resetar todos os valores editados?"):
                self.valor_editado.clear()
                self.carregar_parcelas()
                messagebox.showinfo("Sucesso", "Valores resetados!")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao resetar valores: {str(e)}")
    
    def buscar_lancamentos_existentes(self, tree, dados_parcela, filtro_nome, valor_aproximado):
        """Busca lançamentos existentes que podem corresponder à parcela (versão para parcela única)"""
        try:
            # Limpar treeview
            for item in tree.get_children():
                tree.delete(item)
            
            # Carregar planilha
            arquivo_cliente = Path(PASTA_CLIENTES) / f"{self.cliente_selecionado}.xlsx"
            wb = load_workbook(arquivo_cliente, data_only=True)
            ws = wb['Dados']
            
            # Valor da parcela
            valor_str = dados_parcela['valor_pagar'].replace('R$', '').replace('.', '').replace(',', '.').strip()
            valor_parcela = float(valor_str)
            
            # Calcular margem de valor (±10%)
            margem = valor_parcela * 0.10
            valor_min = valor_parcela - margem
            valor_max = valor_parcela + margem
            
            # Buscar lançamentos
            encontrados = 0
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Extrair dados conforme estrutura real:
                # [0]=DATA_REL, [1]=TP_DESP, [2]=CNPJ_CPF, [3]=NOME, [4]=REFERÊNCIA,
                # [5]=NF, [6]=VR_UNIT, [7]=DIAS, [8]=VALOR, [9]=DT_VENCTO,
                # [10]=CATEGORIA, [11]=DADOS_BANCARIOS, [12]=OBSERVAÇÃO, [13]=STATUS
                
                data_rel = row[0]
                cnpj_cpf = str(row[2]) if row[2] else ""
                nome = str(row[3]) if row[3] else ""
                referencia = str(row[4]) if row[4] else ""
                valor_celula = row[8]  # COLUNA 9 (índice 8) = VALOR
                dt_vencto = row[9] if len(row) > 9 else None
                observacao = str(row[12]) if len(row) > 12 and row[12] else ""
                
                # Aplicar filtros
                # 1. Filtro de nome (case insensitive, busca parcial)
                if filtro_nome:
                    filtro_lower = filtro_nome.lower()
                    nome_lower = nome.lower()
                    
                    # Verificar se há correspondência parcial
                    if filtro_lower not in nome_lower:
                        continue
                
                # 2. Filtro de valor usando função auxiliar
                valor_float = self.converter_valor_seguro(valor_celula)
                
                if valor_aproximado:
                    # Buscar valor aproximado (±10%)
                    if not (valor_min <= valor_float <= valor_max):
                        continue
                else:
                    # Buscar valor exato
                    if abs(valor_float - valor_parcela) > 0.01:
                        continue
                
                # Formatar dados para exibição
                data_formatada = data_rel.strftime('%d/%m/%Y') if isinstance(data_rel, datetime) else str(data_rel)
                vencto_formatado = dt_vencto.strftime('%d/%m/%Y') if isinstance(dt_vencto, datetime) else str(dt_vencto) if dt_vencto else ""
                valor_formatado = f"R$ {valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                # Adicionar ao treeview
                tree.insert('', 'end', values=(
                    idx,  # Número da linha
                    data_formatada,
                    nome,
                    cnpj_cpf,
                    valor_formatado,
                    vencto_formatado,
                    referencia,
                    observacao
                ))
                encontrados += 1
            
            wb.close()
            
            # Mensagem se nada foi encontrado
            if encontrados == 0:
                # Temporariamente desabilitar topmost da janela pai para messagebox aparecer
                janela_pai = tree.master
                while janela_pai and not isinstance(janela_pai, tk.Toplevel):
                    janela_pai = janela_pai.master
                
                if janela_pai:
                    janela_pai.attributes('-topmost', False)
                
                messagebox.showinfo(
                    "Busca", 
                    "Nenhum lançamento encontrado com os critérios especificados.\n\n"
                    "Dicas:\n"
                    "• Experimente remover parte do nome\n"
                    "• Verifique se marcou 'valor aproximado'\n"
                    "• O fornecedor pode estar com nome diferente (PF/PJ)",
                    parent=janela_pai if janela_pai else None
                )
                
                if janela_pai:
                    janela_pai.attributes('-topmost', True)
                    janela_pai.lift()
            
        except Exception as e:
            # Encontrar janela pai
            janela_pai = tree.master
            while janela_pai and not isinstance(janela_pai, tk.Toplevel):
                janela_pai = janela_pai.master
            
            if janela_pai:
                janela_pai.attributes('-topmost', False)
                messagebox.showerror("Erro", f"Erro ao buscar lançamentos: {str(e)}", parent=janela_pai)
                janela_pai.attributes('-topmost', True)
                janela_pai.lift()
            else:
                messagebox.showerror("Erro", f"Erro ao buscar lançamentos: {str(e)}")
    
    def buscar_lancamentos(self, entry_nome, entry_valor, var_valor_aprox, tree):
        """Busca lançamentos na aba Dados que correspondam aos critérios"""
        try:
            # Limpar tree
            for item in tree.get_children():
                tree.delete(item)
            
            nome_busca = entry_nome.get().strip().upper()
            valor_busca_str = entry_valor.get().strip().replace(',', '.')
            
            if not nome_busca and not valor_busca_str:
                messagebox.showwarning("Aviso", "Informe ao menos um critério de busca!", 
                                     parent=tree.master.master)
                return
            
            valor_busca = None
            if valor_busca_str:
                try:
                    valor_busca = float(valor_busca_str)
                except:
                    messagebox.showerror("Erro", "Valor inválido!", parent=tree.master.master)
                    return
            
            # Abrir arquivo
            arquivo_cliente = Path(PASTA_CLIENTES) / f"{self.cliente_selecionado}.xlsx"
            wb = load_workbook(arquivo_cliente, data_only=True)
            ws_dados = wb['Dados']
            
            resultados_encontrados = 0
            
            # Buscar nas linhas
            for row_idx, row in enumerate(ws_dados.iter_rows(min_row=2), start=2):
                data_rel = row[0].value
                nome = str(row[3].value or "")
                cnpj = str(row[2].value or "")
                valor_celula = row[8].value  # COLUNA 9 (índice 8) = VALOR
                status = str(row[13].value or "")  # COLUNA 14 (índice 13) = STATUS
                
                # Aplicar filtros
                if nome_busca and nome_busca not in nome.upper():
                    continue
                
                # Converter valor usando função auxiliar
                if valor_busca is not None:
                    valor_float = self.converter_valor_seguro(valor_celula)
                    
                    # Pular se valor for zero (provavelmente inválido)
                    if valor_float == 0 and valor_celula:
                        continue
                    
                    if var_valor_aprox.get():
                        # Tolerância de 5%
                        margem = valor_busca * 0.05
                        if not (valor_busca - margem <= valor_float <= valor_busca + margem):
                            continue
                    else:
                        if abs(valor_float - valor_busca) > 0.01:
                            continue
                
                # Formatar e inserir
                data_fmt = data_rel.strftime('%d/%m/%Y') if data_rel else ""
                
                # Formatar valor usando a função auxiliar
                valor_float = self.converter_valor_seguro(valor_celula)
                valor_fmt = f"R$ {valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                tree.insert('', 'end', values=(
                    row_idx,
                    data_fmt,
                    nome,
                    cnpj,
                    valor_fmt,
                    status
                ))
                
                resultados_encontrados += 1
            
            wb.close()
            
            if resultados_encontrados == 0:
                messagebox.showinfo(
                    "Busca",
                    "Nenhum lançamento encontrado com os critérios especificados.",
                    parent=tree.master.master
                )
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar lançamentos: {str(e)}", 
                               parent=tree.master.master)
            if 'wb' in locals():
                wb.close()
    
    def verificar_correcoes_pendentes(self):
        """Verifica se há correções monetárias pendentes"""
        try:
            gerenciador_correcao = GerenciadorCorrecaoMonetaria()
            
            hoje = date.today()
            config_correcao = gerenciador_correcao.config.get('correcao_automatica', {})
            
            if (config_correcao.get('ativa', False) and 
                hoje.day == config_correcao.get('dia_calculo', 15)):
                
                if messagebox.askyesno("Correção Monetária", 
                                    "Detectamos que hoje pode ser o dia de aplicar correção monetária nos contratos.\n\n"
                                    "Deseja abrir o gerenciador de correção?"):
                    from src.correcao_monetaria import InterfaceIndicesCorrecao
                    interface = InterfaceIndicesCorrecao(self.root)
                    
        except Exception as e:
            print(f"Erro ao verificar correções: {str(e)}")
    
    def voltar_menu(self):
        """Fecha a janela e retorna ao menu principal"""
        if hasattr(self, 'controlador_principal') and self.controlador_principal:
            self.root.destroy()
            
            if hasattr(self.controlador_principal, 'janela') and self.controlador_principal.janela:
                self.controlador_principal.janela.deiconify()
                return
        
        self.root.destroy()
        
        if self.parent:
            if hasattr(self.parent, 'janela') and self.parent.janela:
                self.parent.janela.deiconify()
            else:
                self.parent.deiconify()
    
    def run(self):
        """Inicia a execução do sistema"""
        self.root.mainloop()


if __name__ == "__main__":
    app = ControlePagamentos()
    app.run()