# -*- coding: utf-8 -*-
"""
Gerenciador de Materiais 
Armazena dados por cliente em abas individuais, mantendo configurações centralizadas
"""

import pandas as pd
import json
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl

from src.config.config import BASE_PATH, PASTA_CLIENTES
from configuracoes_sistema import GerenciadorConfiguracoes

class GerenciadorMateriais:
    """Classe para gerenciar materiais por cliente usando abas nas planilhas individuais"""
    
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        
        # Usar estrutura existente - dados por cliente
        self.pasta_clientes = PASTA_CLIENTES
        print(f"✅ Usando pasta de clientes: {self.pasta_clientes}")
        
        # Configurações centralizadas
        self.carregar_parametros_centralizados()
    
    def carregar_parametros_centralizados(self):
        """Carrega parâmetros das configurações centralizadas"""
        try:
            self.parametros = GerenciadorConfiguracoes.carregar_configuracoes_materiais()
            
            if self.parametros is None:
                print("⚠️ Configurações de materiais não encontradas, usando padrão...")
                self.parametros = self._obter_parametros_padrao()
            else:
                print("✅ Parâmetros carregados das configurações centralizadas")
                
        except Exception as e:
            print(f"❌ Erro ao carregar parâmetros centralizados: {e}")
            self.parametros = self._obter_parametros_padrao()
    
    def _obter_parametros_padrao(self):
        """Parâmetros padrão caso não encontre configurações"""
        return GerenciadorConfiguracoes._obter_configuracoes_materiais_padrao()
    
    def obter_arquivo_cliente(self, cliente):
        """Retorna o caminho do arquivo do cliente"""
        return self.pasta_clientes / f"{cliente}.xlsx"
    
    def verificar_aba_materiais(self, cliente):
        """Verifica se a aba de materiais existe e cria se necessário"""
        arquivo_cliente = self.obter_arquivo_cliente(cliente)
        
        if not arquivo_cliente.exists():
            print(f"❌ Arquivo do cliente {cliente} não encontrado: {arquivo_cliente}")
            return False
        
        try:
            wb = load_workbook(arquivo_cliente)
            
            if "Materiais" not in wb.sheetnames:
                print(f"📝 Criando aba Materiais para cliente {cliente}")
                ws = wb.create_sheet("Materiais")
                
                # Definir cabeçalhos
                headers = [
                    # Identificação
                    'ID', 'Data_Cadastro', 'Data_Ultima_Atualizacao',
                    
                    # Classificação
                    'Categoria', 'Subcategoria', 'Codigo_Produto',
                    
                    # Descrição
                    'Descricao_Completa', 'Marca', 'Modelo', 'Cor_Acabamento', 
                    'Dimensoes', 'Especificacoes_Tecnicas',
                    
                    # Dados de compra
                    'Tem_Dados_Compra', 'Data_Compra', 'CNPJ_Fornecedor', 
                    'Nome_Fornecedor', 'Numero_NF', 'Item_NF',
                    
                    # Quantidades e valores
                    'Quantidade', 'Unidade', 'Valor_Unitario', 'Valor_Total',
                    
                    # Localização e instalação
                    'Ambiente_Aplicacao', 'Localizacao_Especifica',
                    'Data_Instalacao', 'Instalador', 'Status_Instalacao',
                    
                    # Garantia e manutenção
                    'Garantia_Meses', 'Data_Fim_Garantia', 'Manutencao_Preventiva',
                    
                    # Documentação
                    'Observacoes', 'Foto_Produto', 'Manual_Fabricante', 
                    'Certificados', 'Origem_Dados',
                    
                    # Referência para mediçoes/contratos
                    'ID_Contrato_Origem', 'ID_Medicao_Origem'
                ]
                
                # Criar cabeçalho
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                # Colunas específicas com larguras diferentes
                ws.column_dimensions['G'].width = 40  # Descrição
                ws.column_dimensions['J'].width = 30  # Especificações
                ws.column_dimensions['O'].width = 25  # Nome Fornecedor
                
                wb.save(arquivo_cliente)
                print(f"✅ Aba Materiais criada para {cliente}")
            
            wb.close()
            return True
            
        except Exception as e:
            print(f"❌ Erro ao verificar/criar aba materiais para {cliente}: {e}")
            return False
    
    def salvar_material(self, cliente, dados_material):
        """Salva um material na aba do cliente específico"""
        try:
            if not self.verificar_aba_materiais(cliente):
                raise Exception(f"Não foi possível preparar aba de materiais para {cliente}")
            
            arquivo_cliente = self.obter_arquivo_cliente(cliente)
            wb = load_workbook(arquivo_cliente)
            ws = wb["Materiais"]
            
            # Gerar ID único para o cliente
            ultimo_id = 0
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], int):
                    ultimo_id = max(ultimo_id, row[0])
            
            novo_id = ultimo_id + 1
            
            # Preparar dados
            dados_material['ID'] = novo_id
            dados_material['Data_Cadastro'] = datetime.now().strftime('%d/%m/%Y')
            dados_material['Data_Ultima_Atualizacao'] = datetime.now().strftime('%d/%m/%Y')
            dados_material['Origem_Dados'] = dados_material.get('Origem_Dados', 'CADASTRO_MANUAL')
            
            # Calcular data fim garantia se aplicável
            if dados_material.get('Garantia_Meses') and dados_material.get('Data_Instalacao'):
                try:
                    from dateutil.relativedelta import relativedelta
                    data_inst = datetime.strptime(dados_material['Data_Instalacao'], '%d/%m/%Y')
                    meses_garantia = int(dados_material['Garantia_Meses'])
                    data_fim = data_inst + relativedelta(months=meses_garantia)
                    dados_material['Data_Fim_Garantia'] = data_fim.strftime('%d/%m/%Y')
                except:
                    dados_material['Data_Fim_Garantia'] = ''
            
            # Validações básicas
            if not dados_material.get('Categoria'):
                dados_material['Categoria'] = 'OUTROS'
            
            if not dados_material.get('Descricao_Completa'):
                raise ValueError("Descrição é obrigatória")
            
            # Inserir na planilha
            proxima_linha = ws.max_row + 1
            
            # Mapear campos para colunas (baseado na ordem dos headers)
            campos_ordenados = [
                'ID', 'Data_Cadastro', 'Data_Ultima_Atualizacao',
                'Categoria', 'Subcategoria', 'Codigo_Produto',
                'Descricao_Completa', 'Marca', 'Modelo', 'Cor_Acabamento', 
                'Dimensoes', 'Especificacoes_Tecnicas',
                'Tem_Dados_Compra', 'Data_Compra', 'CNPJ_Fornecedor', 
                'Nome_Fornecedor', 'Numero_NF', 'Item_NF',
                'Quantidade', 'Unidade', 'Valor_Unitario', 'Valor_Total',
                'Ambiente_Aplicacao', 'Localizacao_Especifica',
                'Data_Instalacao', 'Instalador', 'Status_Instalacao',
                'Garantia_Meses', 'Data_Fim_Garantia', 'Manutencao_Preventiva',
                'Observacoes', 'Foto_Produto', 'Manual_Fabricante', 
                'Certificados', 'Origem_Dados',
                'ID_Contrato_Origem', 'ID_Medicao_Origem'
            ]
            
            for col, campo in enumerate(campos_ordenados, 1):
                valor = dados_material.get(campo, '')
                
                # Tratamento especial para alguns campos
                if campo == 'Tem_Dados_Compra':
                    valor = bool(valor)
                elif campo in ['Valor_Unitario', 'Valor_Total'] and valor:
                    try:
                        valor = float(str(valor).replace(',', '.'))
                    except:
                        valor = 0
                elif campo == 'Quantidade' and valor:
                    try:
                        valor = float(str(valor).replace(',', '.'))
                    except:
                        valor = 1
                
                cell = ws.cell(row=proxima_linha, column=col, value=valor)
                
                # Formatação específica
                if campo in ['Valor_Unitario', 'Valor_Total']:
                    cell.number_format = '#,##0.00'
                elif campo in ['Data_Cadastro', 'Data_Ultima_Atualizacao', 'Data_Compra', 'Data_Instalacao', 'Data_Fim_Garantia']:
                    if valor and isinstance(valor, str):
                        try:
                            data_obj = datetime.strptime(valor, '%d/%m/%Y')
                            cell.value = data_obj
                            cell.number_format = 'DD/MM/YYYY'
                        except:
                            pass
            
            wb.save(arquivo_cliente)
            wb.close()
            
            print(f"✅ Material {novo_id} salvo para cliente {cliente}: {dados_material.get('Descricao_Completa', '')[:50]}")
            return novo_id
            
        except Exception as e:
            print(f"❌ Erro ao salvar material para {cliente}: {e}")
            raise e
    
    def carregar_materiais_cliente(self, cliente):
        """Carrega materiais de um cliente específico"""
        try:
            if not self.verificar_aba_materiais(cliente):
                return pd.DataFrame()
            
            arquivo_cliente = self.obter_arquivo_cliente(cliente)
            wb = load_workbook(arquivo_cliente)
            
            if "Materiais" not in wb.sheetnames:
                return pd.DataFrame()
            
            # Ler dados da aba Materiais
            df = pd.read_excel(arquivo_cliente, sheet_name="Materiais")
            
            wb.close()
            print(f"📊 Carregados {len(df)} materiais do cliente {cliente}")
            return df
            
        except Exception as e:
            print(f"❌ Erro ao carregar materiais do cliente {cliente}: {e}")
            return pd.DataFrame()
    
    def atualizar_material(self, cliente, material_id, dados_atualizacao):
        """Atualiza um material específico do cliente"""
        try:
            arquivo_cliente = self.obter_arquivo_cliente(cliente)
            wb = load_workbook(arquivo_cliente)
            ws = wb["Materiais"]
            
            # Encontrar material
            material_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), 2):
                if row[0] == material_id:
                    material_row = idx
                    break
            
            if not material_row:
                raise ValueError(f"Material ID {material_id} não encontrado para cliente {cliente}")
            
            # Headers para mapear colunas
            headers = [cell.value for cell in ws[1]]
            
            # Atualizar campos
            for campo, valor in dados_atualizacao.items():
                if campo in headers:
                    col_idx = headers.index(campo) + 1
                    ws.cell(row=material_row, column=col_idx, value=valor)
            
            # Atualizar timestamp
            if 'Data_Ultima_Atualizacao' in headers:
                col_idx = headers.index('Data_Ultima_Atualizacao') + 1
                ws.cell(row=material_row, column=col_idx, value=datetime.now().strftime('%d/%m/%Y'))
            
            wb.save(arquivo_cliente)
            wb.close()
            
            print(f"✅ Material {material_id} atualizado para cliente {cliente}")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao atualizar material {material_id} do cliente {cliente}: {e}")
            return False
    
    def excluir_material(self, cliente, material_id):
        """Marca um material como excluído"""
        return self.atualizar_material(cliente, material_id, {
            'Status_Instalacao': 'CANCELADO',
            'Observacoes': f"EXCLUÍDO EM {datetime.now().strftime('%d/%m/%Y')}"
        })
    
    def buscar_materiais(self, cliente, filtros):
        """Busca materiais do cliente com filtros"""
        try:
            df = self.carregar_materiais_cliente(cliente)
            
            if len(df) == 0:
                return df
            
            # Aplicar filtros
            for campo, valor in filtros.items():
                if campo in df.columns and valor:
                    if campo in ['Categoria', 'Status_Instalacao', 'Ambiente_Aplicacao']:
                        df = df[df[campo] == valor]
                    else:
                        # Busca textual
                        df = df[df[campo].astype(str).str.contains(str(valor), case=False, na=False)]
            
            return df
            
        except Exception as e:
            print(f"❌ Erro na busca de materiais para {cliente}: {e}")
            return pd.DataFrame()
    
    def gerar_relatorio_resumo(self, cliente):
        """Gera relatório resumo dos materiais do cliente"""
        try:
            df = self.carregar_materiais_cliente(cliente)
            
            if len(df) == 0:
                return {"erro": "Nenhum material encontrado"}
            
            resumo = {
                "cliente": cliente,
                "total_materiais": len(df),
                "total_valor": df['Valor_Total'].fillna(0).sum() if 'Valor_Total' in df.columns else 0,
                "por_categoria": df['Categoria'].value_counts().to_dict() if 'Categoria' in df.columns else {},
                "por_status": df['Status_Instalacao'].value_counts().to_dict() if 'Status_Instalacao' in df.columns else {},
                "por_ambiente": df['Ambiente_Aplicacao'].value_counts().to_dict() if 'Ambiente_Aplicacao' in df.columns else {},
                "com_dados_nf": len(df[df['Tem_Dados_Compra'] == True]) if 'Tem_Dados_Compra' in df.columns else 0,
                "sem_dados_nf": len(df[df['Tem_Dados_Compra'] == False]) if 'Tem_Dados_Compra' in df.columns else 0,
                "data_atualizacao": datetime.now().strftime('%d/%m/%Y %H:%M')
            }
            
            return resumo
            
        except Exception as e:
            print(f"❌ Erro ao gerar relatório para {cliente}: {e}")
            return {"erro": str(e)}
    
    def migrar_dados_existentes(self, cliente):
        """Migra dados da planilha central para a aba do cliente (se necessário)"""
        try:
            # Verificar se existe planilha central antiga
            arquivo_central = BASE_PATH / "materiais" / "materiais_obra.xlsx"
            
            if not arquivo_central.exists():
                print(f"📝 Não há dados para migrar para {cliente}")
                return True
            
            # Carregar dados centrais
            df_central = pd.read_excel(arquivo_central)
            
            # Filtrar dados do cliente
            if 'Cliente' in df_central.columns:
                df_cliente = df_central[df_central['Cliente'] == cliente]
            else:
                print(f"📝 Nenhum dado específico para migrar para {cliente}")
                return True
            
            if len(df_cliente) == 0:
                print(f"📝 Nenhum material encontrado para migrar para {cliente}")
                return True
            
            # Migrar cada material
            for _, material in df_cliente.iterrows():
                dados_material = material.to_dict()
                
                # Remover campo Cliente pois agora é implícito
                dados_material.pop('Cliente', None)
                
                # Salvar na aba do cliente
                self.salvar_material(cliente, dados_material)
            
            print(f"✅ Migrados {len(df_cliente)} materiais para {cliente}")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao migrar dados para {cliente}: {e}")
            return False
    
    def integrar_com_medicoes(self, cliente, id_medicao, dados_instalacao):
        """Integra dados de instalação vindos das medições"""
        try:
            # Buscar materiais relacionados à medição (por referência na descrição ou observações)
            df = self.carregar_materiais_cliente(cliente)
            
            if len(df) == 0:
                return False
            
            # Atualizar materiais que referenciam esta medição
            materiais_atualizados = 0
            
            for idx, material in df.iterrows():
                # Verificar se o material está relacionado à medição
                # (implementar lógica específica conforme necessidade)
                if self._material_relacionado_medicao(material, id_medicao, dados_instalacao):
                    
                    dados_atualizacao = {
                        'Data_Instalacao': dados_instalacao.get('data_instalacao', ''),
                        'Instalador': dados_instalacao.get('instalador', ''),
                        'Status_Instalacao': 'INSTALADO',
                        'ID_Medicao_Origem': id_medicao
                    }
                    
                    if self.atualizar_material(cliente, material['ID'], dados_atualizacao):
                        materiais_atualizados += 1
            
            print(f"✅ {materiais_atualizados} materiais atualizados com dados da medição {id_medicao}")
            return materiais_atualizados > 0
            
        except Exception as e:
            print(f"❌ Erro ao integrar com medições: {e}")
            return False
    
    def _material_relacionado_medicao(self, material, id_medicao, dados_instalacao):
        """Verifica se um material está relacionado a uma medição específica"""
        # Implementar lógica específica baseada em:
        # - Descrição do material vs referência da medição
        # - Categoria/ambiente do material
        # - Período de instalação
        # etc.
        
        # Exemplo básico:
        if material.get('Status_Instalacao') in ['PENDENTE', 'EM INSTALACAO']:
            # Verificar se ambiente coincide com algum critério
            return True
        
        return False


class IntegradorMateriais:
    """Integra o gerenciador refatorado com o sistema principal"""
    
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.gerenciador = GerenciadorMateriais(sistema_principal)
        
        # Adicionar referência no sistema principal
        sistema_principal.gerenciador_materiais = self.gerenciador
    
    def adicionar_botoes_interface(self):
        """Adiciona botões de materiais na interface principal"""
        try:
            # Verificar se existe aba fornecedor
            if hasattr(self.sistema, 'aba_fornecedor'):
                self.adicionar_secao_materiais()
                print("✅ Seção de materiais (refatorada) adicionada na aba fornecedor")
                
        except Exception as e:
            print(f"❌ Erro ao adicionar botões: {e}")
    
    def adicionar_secao_materiais(self):
        """Adiciona seção de materiais"""
        # Frame principal
        frame_materiais = ttk.LabelFrame(self.sistema.aba_fornecedor, 
                                        text="🗃️ Gestão de Materiais da Obra (Por Cliente)", 
                                        padding=10)
        frame_materiais.pack(fill='x', padx=10, pady=5)
        
        # Container para botões
        frame_botoes = ttk.Frame(frame_materiais)
        frame_botoes.pack(fill='x', pady=5)
        
        # Botões principais
        botoes = [
            ("📦 Novo Material", self.abrir_cadastro_material),
            ("📋 Consultar Materiais", self.abrir_consulta_materiais),
            ("🔄 Migrar Dados", self.migrar_dados_cliente),
            ("⚙️ Configurações", self.abrir_configuracoes_materiais)
        ]
        
        for texto, comando in botoes:
            ttk.Button(frame_botoes, text=texto, command=comando).pack(side='left', padx=5)
        
        # Status
        self.label_status_materiais = tk.Label(frame_materiais, 
                                             text="Sistema de materiais por cliente carregado", 
                                             fg='green', font=('Arial', 8))
        self.label_status_materiais.pack(anchor='w', pady=2)
        
        # Atualizar status
        self.atualizar_status_materiais()
    
    def atualizar_status_materiais(self):
        """Atualiza status dos materiais na interface"""
        try:
            cliente_atual = getattr(self.sistema, 'cliente_atual', None)
            if cliente_atual:
                resumo = self.gerenciador.gerar_relatorio_resumo(cliente_atual)
                if 'erro' not in resumo:
                    texto = f"Cliente: {cliente_atual} | {resumo['total_materiais']} materiais | R$ {resumo['total_valor']:,.2f}"
                    self.label_status_materiais.config(text=texto, fg='blue')
                else:
                    self.label_status_materiais.config(text=f"Cliente: {cliente_atual} | Nenhum material cadastrado", fg='gray')
            else:
                self.label_status_materiais.config(text="Selecione um cliente para ver materiais", fg='orange')
                
        except Exception as e:
            self.label_status_materiais.config(text=f"Erro: {str(e)}", fg='red')
    
    def abrir_cadastro_material(self):
        """Abre cadastro de material"""
        cliente_atual = getattr(self.sistema, 'cliente_atual', None)
        if not cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
        
        try:
            CadastroMaterial(self.sistema, self.gerenciador, cliente_atual)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir cadastro:\n{str(e)}")
    
    def abrir_consulta_materiais(self):
        """Abre consulta de materiais"""
        cliente_atual = getattr(self.sistema, 'cliente_atual', None)
        if not cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
        
        try:
            ConsultaMateriais(self.sistema, self.gerenciador, cliente_atual)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir consulta:\n{str(e)}")
    
    def migrar_dados_cliente(self):
        """Migra dados do cliente atual"""
        cliente_atual = getattr(self.sistema, 'cliente_atual', None)
        if not cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
        
        try:
            if messagebox.askyesno("Migração", f"Deseja migrar dados de materiais para o cliente {cliente_atual}?"):
                if self.gerenciador.migrar_dados_existentes(cliente_atual):
                    messagebox.showinfo("Sucesso", "Dados migrados com sucesso!")
                    self.atualizar_status_materiais()
                else:
                    messagebox.showerror("Erro", "Erro durante a migração!")
                    
        except Exception as e:
            messagebox.showerror("Erro", f"Erro na migração:\n{str(e)}")
    
    def abrir_configuracoes_materiais(self):
        """Abre configurações de materiais"""
        try:
            config_window = GerenciadorConfiguracoes(self.sistema.root)
            # Focar na aba de materiais
            config_window.notebook.select(5)  # Índice da aba de materiais
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir configurações:\n{str(e)}")


class CadastroMaterial:
    """Janela de cadastro de material para cliente específico"""
    
    def __init__(self, sistema, gerenciador, cliente):
        self.sistema = sistema
        self.gerenciador = gerenciador
        self.cliente = cliente
        
        self.criar_janela()
        self.criar_interface()
        self.carregar_dados_iniciais()
    
    def criar_janela(self):
        """Cria janela principal"""
        self.janela = tk.Toplevel(self.sistema.root)
        self.janela.title(f"Cadastro de Material - {self.cliente}")
        self.janela.geometry("800x600")
        self.janela.grab_set()
        self.janela.transient(self.sistema.root)
    
    def criar_interface(self):
        """Cria interface básica - mesmo padrão da versão anterior"""
        # Info do cliente
        frame_info = ttk.Frame(self.janela)
        frame_info.pack(fill='x', padx=10, pady=5)
        
        tk.Label(frame_info, text=f"Cliente: {self.cliente}", 
                font=('Arial', 12, 'bold'), fg='blue').pack(anchor='w')
        
        # Notebook para organizar abas
        notebook = ttk.Notebook(self.janela)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Abas (mesmo conteúdo da versão anterior)
        self.criar_aba_dados_basicos(notebook)
        self.criar_aba_compra(notebook)
        self.criar_aba_instalacao(notebook)
        
        # Botões
        self.criar_botoes_principais()
    
    def criar_aba_dados_basicos(self, notebook):
        """Cria aba de dados básicos - implementação similar à anterior"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text="📋 Dados Básicos")
        
        # Implementação similar à classe anterior
        # Por brevidade, mantendo estrutura similar
        
        self.campos = {}
        
        # Categoria
        tk.Label(frame, text="Categoria:*").pack(anchor='w', padx=10, pady=2)
        self.campos['categoria'] = ttk.Combobox(frame, state='readonly')
        self.campos['categoria'].pack(fill='x', padx=10, pady=2)
        
        # Descrição
        tk.Label(frame, text="Descrição:*").pack(anchor='w', padx=10, pady=2)
        self.campos['descricao'] = tk.Entry(frame)
        self.campos['descricao'].pack(fill='x', padx=10, pady=2)
        
        # Outros campos conforme necessário...
        
    def criar_aba_compra(self, notebook):
        """Cria aba de compra"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text="💰 Compra")
        
        # Checkbox para habilitar dados de compra
        self.tem_dados_compra = tk.BooleanVar()
        cb_frame = ttk.Frame(frame)
        cb_frame.pack(fill='x', padx=10, pady=10)
        
        cb = tk.Checkbutton(cb_frame, text="Este material possui dados de compra/fornecedor", 
                           variable=self.tem_dados_compra, command=self.toggle_campos_compra,
                           font=('Arial', 10, 'bold'))
        cb.pack(anchor='w')
        
        # Frame para dados de compra
        self.frame_compra = ttk.LabelFrame(frame, text="Dados da Compra", padding=10)
        self.frame_compra.pack(fill='x', padx=10, pady=5)
        
        # Campos de compra
        tk.Label(self.frame_compra, text="Fornecedor:").grid(row=0, column=0, sticky='w', pady=5)
        self.campos['fornecedor'] = tk.Entry(self.frame_compra, width=40, state='disabled')
        self.campos['fornecedor'].grid(row=0, column=1, columnspan=2, sticky='ew', padx=5, pady=5)
        
        tk.Label(self.frame_compra, text="Quantidade:").grid(row=1, column=0, sticky='w', pady=5)
        self.campos['quantidade'] = tk.Entry(self.frame_compra, width=10, state='disabled')
        self.campos['quantidade'].grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        tk.Label(self.frame_compra, text="Valor Total:").grid(row=1, column=2, sticky='w', pady=5)
        self.campos['valor_total'] = tk.Entry(self.frame_compra, width=15, state='disabled')
        self.campos['valor_total'].grid(row=1, column=3, sticky='w', padx=5, pady=5)
    
    def criar_aba_instalacao(self, notebook):
        """Cria aba de instalação"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text="🔧 Instalação")
        
        # Ambiente
        tk.Label(frame, text="Ambiente:").pack(anchor='w', padx=10, pady=2)
        self.campos['ambiente'] = ttk.Combobox(frame, state='readonly')
        self.campos['ambiente'].pack(fill='x', padx=10, pady=2)
        
        # Status
        tk.Label(frame, text="Status:").pack(anchor='w', padx=10, pady=2)
        self.campos['status'] = ttk.Combobox(frame, state='readonly')
        self.campos['status'].pack(fill='x', padx=10, pady=2)
        
        # Observações
        tk.Label(frame, text="Observações:").pack(anchor='w', padx=10, pady=2)
        self.campos['observacoes'] = tk.Text(frame, height=6)
        self.campos['observacoes'].pack(fill='both', expand=True, padx=10, pady=2)
    
    def criar_botoes_principais(self):
        """Cria botões principais"""
        frame_botoes = ttk.Frame(self.janela)
        frame_botoes.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(frame_botoes, text="💾 Salvar Material", 
                  command=self.salvar_material).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes, text="🧹 Limpar Campos", 
                  command=self.limpar_campos).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes, text="❌ Cancelar", 
                  command=self.janela.destroy).pack(side='right', padx=5)
    
    def carregar_dados_iniciais(self):
        """Carrega dados iniciais nos comboboxes"""
        try:
            parametros = self.gerenciador.parametros
            
            # Categorias
            categorias = list(parametros['categorias_materiais'].keys())
            self.campos['categoria']['values'] = categorias
            
            # Ambientes
            self.campos['ambiente']['values'] = parametros['ambientes']
            
            # Status
            self.campos['status']['values'] = parametros['status_instalacao']
            self.campos['status'].set('PENDENTE')
            
        except Exception as e:
            print(f"❌ Erro ao carregar dados iniciais: {e}")
    
    def toggle_campos_compra(self):
        """Habilita/desabilita campos de compra"""
        estado = 'normal' if self.tem_dados_compra.get() else 'disabled'
        
        campos_compra = ['fornecedor', 'quantidade', 'valor_total']
        
        for campo in campos_compra:
            if campo in self.campos:
                self.campos[campo].config(state=estado)
    
    def salvar_material(self):
        """Salva o material"""
        try:
            # Validações básicas
            if not self.campos['categoria'].get().strip():
                messagebox.showerror("Erro", "Categoria é obrigatória")
                return
            
            if not self.campos['descricao'].get().strip():
                messagebox.showerror("Erro", "Descrição é obrigatória")
                return
            
            # Preparar dados
            dados_material = {
                'Categoria': self.campos['categoria'].get().strip(),
                'Descricao_Completa': self.campos['descricao'].get().strip(),
                'Ambiente_Aplicacao': self.campos['ambiente'].get().strip(),
                'Status_Instalacao': self.campos['status'].get() or 'PENDENTE',
                'Observacoes': self.campos['observacoes'].get('1.0', tk.END).strip(),
                'Tem_Dados_Compra': self.tem_dados_compra.get(),
                'Origem_Dados': 'CADASTRO_MANUAL'
            }
            
            # Adicionar dados de compra se habilitado
            if self.tem_dados_compra.get():
                dados_material.update({
                    'Nome_Fornecedor': self.campos['fornecedor'].get().strip(),
                    'Quantidade': self.campos['quantidade'].get().strip() or '1',
                    'Valor_Total': self.campos['valor_total'].get().strip() or '0'
                })
            
            # Salvar
            material_id = self.gerenciador.salvar_material(self.cliente, dados_material)
            
            messagebox.showinfo("Sucesso", 
                f"Material cadastrado com sucesso!\n\nID: {material_id}\nCliente: {self.cliente}\nDescrição: {dados_material['Descricao_Completa']}")
            
            # Perguntar se quer cadastrar outro
            if messagebox.askyesno("Novo Material", "Deseja cadastrar outro material?"):
                self.limpar_campos()
            else:
                self.janela.destroy()
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar material:\n{str(e)}")
    
    def limpar_campos(self):
        """Limpa todos os campos"""
        try:
            for campo, widget in self.campos.items():
                if isinstance(widget, tk.Entry):
                    widget.delete(0, tk.END)
                elif isinstance(widget, ttk.Combobox):
                    widget.set('')
                elif isinstance(widget, tk.Text):
                    widget.delete('1.0', tk.END)
            
            self.campos['status'].set('PENDENTE')
            self.tem_dados_compra.set(False)
            self.toggle_campos_compra()
            
        except Exception as e:
            print(f"❌ Erro ao limpar campos: {e}")


class ConsultaMateriais:
    """Janela de consulta de materiais para cliente específico"""
    
    def __init__(self, sistema, gerenciador, cliente):
        self.sistema = sistema
        self.gerenciador = gerenciador
        self.cliente = cliente
        
        self.criar_janela()
        self.criar_interface()
        self.carregar_materiais()
    
    def criar_janela(self):
        """Cria janela principal"""
        self.janela = tk.Toplevel(self.sistema.root)
        self.janela.title(f"Materiais - {self.cliente}")
        self.janela.geometry("1000x700")
        self.janela.grab_set()
        self.janela.transient(self.sistema.root)
    
    def criar_interface(self):
        """Cria interface de consulta"""
        # Info do cliente
        frame_info = ttk.Frame(self.janela)
        frame_info.pack(fill='x', padx=10, pady=5)
        
        tk.Label(frame_info, text=f"Materiais do Cliente: {self.cliente}", 
                font=('Arial', 12, 'bold'), fg='blue').pack(anchor='w')
        
        # Frame para filtros
        frame_filtros = ttk.LabelFrame(self.janela, text="Filtros", padding=10)
        frame_filtros.pack(fill='x', padx=10, pady=5)
        
        # Filtros básicos
        tk.Label(frame_filtros, text="Categoria:").grid(row=0, column=0, sticky='w', padx=5)
        self.filtro_categoria = ttk.Combobox(frame_filtros, state='readonly')
        self.filtro_categoria.grid(row=0, column=1, sticky='ew', padx=5)
        
        tk.Label(frame_filtros, text="Status:").grid(row=0, column=2, sticky='w', padx=5)
        self.filtro_status = ttk.Combobox(frame_filtros, state='readonly')
        self.filtro_status.grid(row=0, column=3, sticky='ew', padx=5)
        
        ttk.Button(frame_filtros, text="Filtrar", command=self.aplicar_filtros).grid(row=0, column=4, padx=10)
        ttk.Button(frame_filtros, text="Limpar", command=self.limpar_filtros).grid(row=0, column=5, padx=5)
        
        # Configurar grid
        frame_filtros.columnconfigure(1, weight=1)
        frame_filtros.columnconfigure(3, weight=1)
        
        # Treeview para materiais
        frame_tree = ttk.Frame(self.janela)
        frame_tree.pack(fill='both', expand=True, padx=10, pady=5)
        
        colunas = ('ID', 'Categoria', 'Descrição', 'Ambiente', 'Status', 'Valor')
        self.tree_materiais = ttk.Treeview(frame_tree, columns=colunas, show='headings')
        
        # Configurar colunas
        for col in colunas:
            self.tree_materiais.heading(col, text=col)
        
        self.tree_materiais.column('ID', width=50, anchor='center')
        self.tree_materiais.column('Categoria', width=120)
        self.tree_materiais.column('Descrição', width=300)
        self.tree_materiais.column('Ambiente', width=150)
        self.tree_materiais.column('Status', width=100, anchor='center')
        self.tree_materiais.column('Valor', width=100, anchor='e')
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_tree, orient='vertical', command=self.tree_materiais.yview)
        scrollx = ttk.Scrollbar(frame_tree, orient='horizontal', command=self.tree_materiais.xview)
        self.tree_materiais.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_materiais.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Botões
        frame_botoes = ttk.Frame(self.janela)
        frame_botoes.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(frame_botoes, text="📝 Editar", command=self.editar_material).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="🗑️ Excluir", command=self.excluir_material).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="📊 Relatório", command=self.gerar_relatorio).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="❌ Fechar", command=self.janela.destroy).pack(side='right', padx=5)
        
        # Carregar filtros
        self.carregar_filtros()
    
    def carregar_filtros(self):
        """Carrega opções dos filtros"""
        try:
            parametros = self.gerenciador.parametros
            
            # Categorias
            categorias = [''] + list(parametros['categorias_materiais'].keys())
            self.filtro_categoria['values'] = categorias
            
            # Status
            status = [''] + parametros['status_instalacao']
            self.filtro_status['values'] = status
            
        except Exception as e:
            print(f"❌ Erro ao carregar filtros: {e}")
    
    def carregar_materiais(self):
        """Carrega materiais do cliente"""
        try:
            # Limpar treeview
            for item in self.tree_materiais.get_children():
                self.tree_materiais.delete(item)
            
            # Carregar dados
            df = self.gerenciador.carregar_materiais_cliente(self.cliente)
            
            if len(df) == 0:
                return
            
            # Preencher treeview
            for _, material in df.iterrows():
                valor = material.get('Valor_Total', 0)
                if pd.isna(valor):
                    valor = 0
                
                valor_formatado = f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                self.tree_materiais.insert('', 'end', values=(
                    material.get('ID', ''),
                    material.get('Categoria', ''),
                    material.get('Descricao_Completa', '')[:50] + ('...' if len(str(material.get('Descricao_Completa', ''))) > 50 else ''),
                    material.get('Ambiente_Aplicacao', ''),
                    material.get('Status_Instalacao', ''),
                    valor_formatado
                ))
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar materiais:\n{str(e)}")
    
    def aplicar_filtros(self):
        """Aplica filtros na consulta"""
        try:
            filtros = {}
            
            if self.filtro_categoria.get():
                filtros['Categoria'] = self.filtro_categoria.get()
            
            if self.filtro_status.get():
                filtros['Status_Instalacao'] = self.filtro_status.get()
            
            # Buscar com filtros
            df = self.gerenciador.buscar_materiais(self.cliente, filtros)
            
            # Atualizar treeview
            for item in self.tree_materiais.get_children():
                self.tree_materiais.delete(item)
            
            for _, material in df.iterrows():
                valor = material.get('Valor_Total', 0)
                if pd.isna(valor):
                    valor = 0
                
                valor_formatado = f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                self.tree_materiais.insert('', 'end', values=(
                    material.get('ID', ''),
                    material.get('Categoria', ''),
                    material.get('Descricao_Completa', '')[:50] + ('...' if len(str(material.get('Descricao_Completa', ''))) > 50 else ''),
                    material.get('Ambiente_Aplicacao', ''),
                    material.get('Status_Instalacao', ''),
                    valor_formatado
                ))
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao aplicar filtros:\n{str(e)}")
    
    def limpar_filtros(self):
        """Limpa filtros e recarrega todos os materiais"""
        self.filtro_categoria.set('')
        self.filtro_status.set('')
        self.carregar_materiais()
    
    def editar_material(self):
        """Edita material selecionado"""
        selecionado = self.tree_materiais.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um material para editar!")
            return
        
        material_id = self.tree_materiais.item(selecionado)['values'][0]
        messagebox.showinfo("Em Desenvolvimento", f"Edição do material {material_id} em desenvolvimento")
    
    def excluir_material(self):
        """Exclui material selecionado"""
        selecionado = self.tree_materiais.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um material para excluir!")
            return
        
        material_id = self.tree_materiais.item(selecionado)['values'][0]
        
        if messagebox.askyesno("Confirmar", f"Deseja excluir o material {material_id}?"):
            if self.gerenciador.excluir_material(self.cliente, material_id):
                messagebox.showinfo("Sucesso", "Material excluído com sucesso!")
                self.carregar_materiais()
            else:
                messagebox.showerror("Erro", "Erro ao excluir material!")
    
    def gerar_relatorio(self):
        """Gera relatório dos materiais"""
        try:
            resumo = self.gerenciador.gerar_relatorio_resumo(self.cliente)
            
            if 'erro' in resumo:
                messagebox.showerror("Erro", resumo['erro'])
                return
            
            # Montar texto do relatório
            texto_relatorio = f"""RELATÓRIO DE MATERIAIS
Cliente: {self.cliente}
Data: {resumo['data_atualizacao']}

RESUMO GERAL:
• Total de Materiais: {resumo['total_materiais']}
• Valor Total: R$ {resumo['total_valor']:,.2f}
• Com dados de NF: {resumo['com_dados_nf']}
• Sem dados de NF: {resumo['sem_dados_nf']}

POR CATEGORIA:
{chr(10).join([f"• {cat}: {qtd}" for cat, qtd in resumo['por_categoria'].items()])}

POR STATUS:
{chr(10).join([f"• {status}: {qtd}" for status, qtd in resumo['por_status'].items()])}

POR AMBIENTE:
{chr(10).join([f"• {amb}: {qtd}" for amb, qtd in resumo['por_ambiente'].items()])}
"""
            
            # Mostrar relatório em nova janela
            janela_relatorio = tk.Toplevel(self.janela)
            janela_relatorio.title(f"Relatório de Materiais - {self.cliente}")
            janela_relatorio.geometry("600x500")
            
            text_widget = tk.Text(janela_relatorio, wrap='word', font=('Courier', 10))
            text_widget.pack(fill='both', expand=True, padx=10, pady=10)
            text_widget.insert('1.0', texto_relatorio)
            text_widget.config(state='disabled')
            
            ttk.Button(janela_relatorio, text="Fechar", 
                      command=janela_relatorio.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório:\n{str(e)}")


def inicializar_sistema_materiais_completo(sistema_principal):
    """
    Inicializa o sistema refatorado de materiais
    Armazena dados por cliente, mantém configurações centralizadas
    
    Args:
        sistema_principal: Instância da classe SistemaEntradaDados
    
    Returns:
        IntegradorMateriais: Instância do integrador
    """
    try:
        print("🚀 Inicializando Sistema de Materiais Refatorado (Por Cliente)...")
        
        # Criar integrador
        integrador = IntegradorMateriais(sistema_principal)
        
        # Adicionar interface
        integrador.adicionar_botoes_interface()
        
        print("✅ Sistema de Materiais  inicializado com sucesso!")
        print("📁 Dados serão armazenados por cliente em abas individuais")
        print("⚙️ Configurações mantidas centralizadas")
        
        return integrador
        
    except Exception as e:
        print(f"❌ Erro ao inicializar sistema de materiais: {e}")
        return None