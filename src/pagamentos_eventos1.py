import os
import sys
from pathlib import Path
import re
from datetime import datetime
from decimal import Decimal

import tkinter as tk
from tkinter import ttk, messagebox, StringVar
from tkcalendar import DateEntry

from dateutil.relativedelta import relativedelta

import pandas as pd
from openpyxl import load_workbook

# Adicionar diretório raiz ao path
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações do sistema
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH,
        ARQUIVO_FORNECEDORES
    )
    from src.config.utils import formatar_cnpj_cpf, aplicar_formatacao_celula
    from src.config.window_config import configurar_janela
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Tentar caminho alternativo
    try:
        from config.config import (
            ARQUIVO_CLIENTES,
            ARQUIVO_MODELO,
            PASTA_CLIENTES,
            BASE_PATH,
            ARQUIVO_FORNECEDORES
        )
        from config.utils import formatar_cnpj_cpf, aplicar_formatacao_celula
        from config.window_config import configurar_janela
    except ImportError as e2:
        print(f"Erro ao importar configurações (caminho alternativo): {str(e2)}")
        raise

if __name__ == "__main__":
    import tkinter as tk
    from tkinter import ttk
    
    # Criar a janela principal diretamente
    root = tk.Tk()
    root.title("Gestão de Pagamentos por Eventos")
    root.geometry("900x700")
    
    # Função para configurar a aplicação real
    def configurar_app(cliente):
        # Limpar a janela atual
        for widget in root.winfo_children():
            widget.destroy()
            
        # Inicializar o objeto GestaoEventos
        app = GestaoEventos(root)
        app.cliente_atual = cliente
        app.arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        
        # Configurar manualmente a interface na janela principal
        app.janela = root  # Usa a janela root diretamente
        
        # Frame principal
        frame_principal = ttk.Frame(root, padding=10)
        frame_principal.pack(fill='both', expand=True)
        
        # Cabeçalho
        ttk.Label(
            frame_principal, 
            text=f"Cliente: {app.cliente_atual}", 
            font=('Arial', 12, 'bold')
        ).pack(side='left')
        
        # Chamar os métodos para configurar as abas
        app.notebook = ttk.Notebook(frame_principal)
        app.notebook.pack(fill='both', expand=True, pady=10)
        
        # Configurar abas
        app.aba_contratos = ttk.Frame(app.notebook)
        app.notebook.add(app.aba_contratos, text="Contratos")
        
        app.aba_eventos = ttk.Frame(app.notebook)
        app.notebook.add(app.aba_eventos, text="Eventos")
        
        app.aba_pagamentos = ttk.Frame(app.notebook)
        app.notebook.add(app.aba_pagamentos, text="Pagamentos")
        
        # Configurar cada aba
        app.configurar_aba_contratos()
        app.configurar_aba_eventos()
        app.configurar_aba_pagamentos()
        
        # Carregar dados
        app.carregar_contratos()
    
    # Criar frame para seleção de cliente
    frame_selecao = ttk.LabelFrame(root, text="Selecione um cliente", padding=20)
    frame_selecao.pack(padx=20, pady=20, fill='both', expand=True)
    
    # Carregar lista de clientes - função simplificada
    def carregar_clientes():
        # Importa a classe para usar seu método
        app_temp = GestaoEventos()
        return app_temp.carregar_lista_clientes()
    
    clientes = carregar_clientes()
    
    # Combobox para seleção
    ttk.Label(frame_selecao, text="Cliente:").pack(pady=(0, 5))
    
    cliente_var = tk.StringVar()
    combo = ttk.Combobox(frame_selecao, textvariable=cliente_var, values=clientes, state="readonly", width=40)
    combo.pack(pady=10)
    
    # Botão de confirmação
    def confirmar_cliente():
        if not cliente_var.get():
            tk.messagebox.showwarning("Aviso", "Selecione um cliente")
            return
            
        configurar_app(cliente_var.get())
    
    ttk.Button(frame_selecao, text="Confirmar", command=confirmar_cliente, width=15).pack(pady=10)
    
    # Iniciar loop principal
    root.mainloop()

class GestaoEventos:
    def __init__(self, parent=None):
        self.parent = parent
        self.janela = None
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.contratos = []
        self.eventos = []
        self.administradores_contratos = {}
        self.cliente_selecionado = False  # Variável para controlar resultado da seleção

    def abrir_janela_eventos(self, cliente=None):
        """Abre a janela principal de gestão de eventos para o cliente selecionado"""
        print("Entrando em abrir_janela_eventos")
        
        # Se a janela já existir, apenas traz para frente
        if self.janela and self.janela.winfo_exists():
            print("Janela já existe, trazendo para frente")
            self.janela.lift()
            self.janela.focus_force()
            return

        print("Criando nova janela")
        # Cria nova janela
        self.janela = tk.Toplevel(self.parent)
        configurar_janela(self.janela, "Gestão de Pagamentos por Eventos", 900, 700)
        
        # Garantir visibilidade
        self.janela.attributes('-topmost', True)
        self.janela.update()
        self.janela.attributes('-topmost', False)

        # Define o cliente atual
        if cliente:
            print(f"Cliente fornecido: {cliente}")
            self.cliente_atual = cliente
            self.arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        else:
            print("Cliente não fornecido, solicitando seleção")
            # Se não tiver cliente selecionado, solicita seleção
            if not self.selecionar_cliente():
                print("Cliente não selecionado, fechando janela")
                self.janela.destroy()
                return

        print("Criando nova janela")
        # Cria nova janela
        self.janela = tk.Toplevel(self.parent)
        configurar_janela(self.janela, "Gestão de Pagamentos por Eventos", 1000, 800)

        print(f"Cliente selecionado: {self.cliente_atual}")
        print(f"Arquivo do cliente: {self.arquivo_cliente}")
        
        # Frame principal
        frame_principal = ttk.Frame(self.janela, padding=10)
        frame_principal.pack(fill='both', expand=True)

        # Cabeçalho com informações do cliente
        frame_cabecalho = ttk.Frame(frame_principal)
        frame_cabecalho.pack(fill='x', pady=5)

        ttk.Label(
            frame_cabecalho, 
            text=f"Cliente: {self.cliente_atual}", 
            font=('Arial', 12, 'bold')
        ).pack(side='left')

        # Notebook para abas
        self.notebook = ttk.Notebook(frame_principal)
        self.notebook.pack(fill='both', expand=True, pady=10)

        # Aba de contratos
        self.aba_contratos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_contratos, text="Contratos")

        # Aba de eventos
        self.aba_eventos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_eventos, text="Eventos")

        # Aba de pagamentos
        self.aba_pagamentos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_pagamentos, text="Pagamentos")

        # NOVO: Adicionar callback para sincronização entre abas
        def on_tab_changed(event):
            try:
                current_tab = event.widget.index('current')
                print(f"Mudou para a aba: {current_tab}")
                
                # Se mudou para a aba de eventos, sincronizar seleção
                if current_tab == 1:  # Índice 1 = Aba Eventos
                    print("Sincronizando seleção de contrato com aba de eventos")
                    selecionado = self.tree_contratos.selection()
                    if selecionado:
                        try:
                            valores = self.tree_contratos.item(selecionado)['values']
                            num_contrato = valores[0]
                            
                            if hasattr(self, 'contrato_selecionado'):
                                valores_combo = self.contrato_selecionado['values']
                                if num_contrato in valores_combo:
                                    self.contrato_selecionado.set(num_contrato)
                                    print(f"Contrato {num_contrato} selecionado na combobox de eventos")
                                    self.carregar_eventos_contrato(None)
                        except Exception as e:
                            print(f"Erro ao sincronizar com aba de eventos: {str(e)}")
                
                # Se mudou para a aba de pagamentos, carregar todos os pagamentos
                elif current_tab == 2:  # Índice 2 = Aba Pagamentos
                    print("Carregando pagamentos")
                    if hasattr(self, 'filtrar_pagamentos'):
                        self.filtrar_pagamentos(None)
                        
            except Exception as e:
                print(f"Erro no callback de mudança de aba: {str(e)}")
        
        # Vincular callback de mudança de aba
        self.notebook.bind("<<NotebookTabChanged>>", on_tab_changed)

        # Configurar cada aba
        self.configurar_aba_contratos()
        self.configurar_aba_eventos()
        self.configurar_aba_pagamentos()

        # Botões principais
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=10)

        ttk.Button(
            frame_botoes,
            text="Fechar",
            command=self.janela.destroy,
            width=15
        ).pack(side='right', padx=5)

        # Garantir que a janela apareça
        self.janela.lift()
        self.janela.focus_force()
        self.janela.update()

        # Carregar dados iniciais
        if self.cliente_selecionado:
            self.carregar_contratos()

        # Configurar fechamento adequado
        if self.parent:
            def on_close():
                # Apenas fecha esta janela, não a aplicação inteira
                self.janela.destroy()
                
            self.janela.protocol("WM_DELETE_WINDOW", on_close)

    def selecionar_cliente(self):
        """Seleciona um cliente usando uma caixa de diálogo simplificada"""
        # Carregar a lista de clientes
        clientes = self.carregar_lista_clientes()
        if not clientes:
            messagebox.showwarning("Aviso", "Nenhum cliente encontrado!")
            return False
        
        # Usar método mais simples e direto para seleção
        escolha = tk.StringVar()
        resultado = False
        
        def confirmar_selecao():
            nonlocal resultado
            if escolha.get():
                self.cliente_atual = escolha.get()
                self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
                resultado = True
                dialogo.destroy()
        
        # Criar diálogo
        dialogo = tk.Toplevel(self.parent)
        dialogo.title("Selecionar Cliente")
        dialogo.geometry("350x180")
        dialogo.resizable(False, False)
        dialogo.transient(self.parent)
        dialogo.grab_set()  # Torna o diálogo modal
        
        # Garantir que o diálogo fique visível
        dialogo.attributes('-topmost', True)
        
        # Layout
        ttk.Label(dialogo, text="Selecione um cliente:", font=("Arial", 12)).pack(pady=(15, 5))
        combo = ttk.Combobox(dialogo, textvariable=escolha, values=clientes, width=30, state="readonly")
        combo.pack(pady=10)
        
        # Botões
        frame_botoes = ttk.Frame(dialogo)
        frame_botoes.pack(pady=15)
        ttk.Button(frame_botoes, text="Confirmar", command=confirmar_selecao, width=12).pack(side="left", padx=10)
        ttk.Button(frame_botoes, text="Cancelar", command=dialogo.destroy, width=12).pack(side="left", padx=10)
        
        # Centralizar
        dialogo.update_idletasks()
        width = dialogo.winfo_width()
        height = dialogo.winfo_height()
        x = (dialogo.winfo_screenwidth() // 2) - (width // 2)
        y = (dialogo.winfo_screenheight() // 2) - (height // 2)
        dialogo.geometry(f"{width}x{height}+{x}+{y}")
        
        # Trazer para frente
        dialogo.lift()
        dialogo.focus_force()
        
        # Adicionar manipulador de protocolo para o botão X da janela
        dialogo.protocol("WM_DELETE_WINDOW", dialogo.destroy)
        
        # Aguardar seleção - modificado para garantir funcionamento
        self.parent.wait_window(dialogo)
        return resultado

    def carregar_lista_clientes(self):
        """Carrega a lista de clientes disponíveis"""
        print("Carregando lista de clientes...")
        try:
            # Verificar se o arquivo existe
            if not os.path.exists(ARQUIVO_CLIENTES):
                print(f"Arquivo de clientes não encontrado: {ARQUIVO_CLIENTES}")
                messagebox.showwarning("Aviso", "Arquivo de clientes não encontrado!")
                return []
                
            workbook = load_workbook(ARQUIVO_CLIENTES)
            print(f"Planilha aberta, abas disponíveis: {workbook.sheetnames}")
            
            if 'Clientes' not in workbook.sheetnames:
                print("Aba 'Clientes' não encontrada!")
                messagebox.showwarning("Aviso", "Aba 'Clientes' não encontrada no arquivo!")
                workbook.close()
                return []
                
            sheet = workbook['Clientes']
            clientes = []

            # Contagem de linhas para debug
            num_linhas = sheet.max_row
            print(f"Número de linhas na planilha: {num_linhas}")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Nome do cliente
                    clientes.append(row[0])
                    print(f"Cliente encontrado: {row[0]}")

            workbook.close()
            print(f"Total de clientes carregados: {len(clientes)}")
            return sorted(clientes)

        except Exception as e:
            print(f"Erro ao carregar clientes: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
            if 'workbook' in locals():
                workbook.close()
            return []


    def selecionar_cliente_terminal(self):
        """Seleciona um cliente usando o terminal (modo de emergência)"""
        print("\n" + "="*50)
        print("SELEÇÃO DE CLIENTE VIA TERMINAL (MODO DE EMERGÊNCIA)")
        print("="*50)
        
        # Carregar a lista de clientes
        clientes = self.carregar_lista_clientes()
        if not clientes:
            print("Nenhum cliente encontrado!")
            return False
        
        # Exibir a lista numerada de clientes
        print("\nClientes disponíveis:")
        for i, cliente in enumerate(clientes, 1):
            print(f"{i}. {cliente}")
        
        # Solicitar escolha do usuário
        while True:
            try:
                escolha = input("\nDigite o número do cliente (ou 'q' para cancelar): ")
                
                if escolha.lower() == 'q':
                    print("Seleção cancelada pelo usuário.")
                    return False
                
                num = int(escolha)
                if 1 <= num <= len(clientes):
                    self.cliente_atual = clientes[num-1]
                    self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
                    self.cliente_selecionado = True
                    print(f"Cliente selecionado: {self.cliente_atual}")
                    return True
                else:
                    print(f"Número inválido. Por favor, escolha entre 1 e {len(clientes)}.")
            except ValueError:
                print("Entrada inválida. Digite um número ou 'q' para cancelar.")
            except Exception as e:
                print(f"Erro: {str(e)}")

    def sincronizar_contrato_com_aba_eventos(self):
        """Sincroniza a seleção de contrato entre a aba de contratos e eventos"""
        try:
            print("Sincronizando seleção de contrato com aba de eventos")
            
            # Verificar se há um contrato selecionado na aba de contratos
            selecionado = self.tree_contratos.selection()
            if not selecionado:
                print("Nenhum contrato selecionado na árvore")
                return False
                
            # Obter o número do contrato
            valores = self.tree_contratos.item(selecionado)['values']
            if not valores or len(valores) < 1:
                print("Valores do contrato selecionado são inválidos")
                return False
                
            num_contrato = valores[0]
            print(f"Contrato selecionado na árvore: {num_contrato}")
            
            # Verificar se o contrato está disponível na combobox da aba de eventos
            if not hasattr(self, 'contrato_selecionado'):
                print("Combobox de contrato não encontrada na aba de eventos")
                return False
                
            valores_combo = list(self.contrato_selecionado['values'])
            if not valores_combo:
                print("Combobox de contrato não tem valores")
                return False
                
            print(f"Valores disponíveis na combobox: {valores_combo}")
            
            # Verificar se o contrato está na lista de valores
            if num_contrato in valores_combo:
                # Selecionar o contrato na combobox
                self.contrato_selecionado.set(num_contrato)
                print(f"Contrato {num_contrato} selecionado na combobox")
                
                # Carregar eventos deste contrato
                self.carregar_eventos_contrato(None)
                return True
            else:
                print(f"Contrato {num_contrato} não está nos valores da combobox")
                return False
                
        except Exception as e:
            print(f"Erro ao sincronizar contrato: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def configurar_aba_contratos(self):
        """Configura a aba de gestão de contratos"""
        frame = ttk.Frame(self.aba_contratos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para lista de contratos
        frame_lista = ttk.LabelFrame(frame, text="Contratos Disponíveis")
        frame_lista.pack(fill='both', expand=True, pady=5)

        # Treeview para contratos
        colunas = ('Nº Contrato', 'Data Início', 'Data Fim', 'Status', 'Valor Total')
        self.tree_contratos = ttk.Treeview(frame_lista, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_contratos.heading(col, text=col)
            if col == 'Valor Total':
                self.tree_contratos.column(col, width=120, anchor='e')
            else:
                self.tree_contratos.column(col, width=120)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_contratos.yview)
        scroll_x = ttk.Scrollbar(frame_lista, orient='horizontal', command=self.tree_contratos.xview)
        self.tree_contratos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_contratos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do contrato selecionado
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Contrato")
        frame_detalhes.pack(fill='x', pady=10)

        frame_detalhes_grid = ttk.Frame(frame_detalhes)
        frame_detalhes_grid.pack(fill='x', padx=5, pady=5)

        # Criar labels e campos para detalhes
        ttk.Label(frame_detalhes_grid, text="Contrato:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.lbl_contrato = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_contrato.grid(row=0, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Período:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.lbl_periodo = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_periodo.grid(row=1, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Status:").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.lbl_status = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_status.grid(row=2, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Valor Total:").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.lbl_valor_total = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_valor_total.grid(row=0, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Administradores:").grid(row=1, column=2, sticky='w', padx=5, pady=2)
        self.lbl_administradores = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_administradores.grid(row=1, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Eventos:").grid(row=2, column=2, sticky='w', padx=5, pady=2)
        self.lbl_eventos = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_eventos.grid(row=2, column=3, sticky='w', padx=5, pady=2)

        # Frame para administradores do contrato
        frame_admins = ttk.LabelFrame(frame, text="Administradores do Contrato")
        frame_admins.pack(fill='both', expand=True, pady=5)

        # Treeview para administradores
        colunas_adm = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas')
        self.tree_adm = ttk.Treeview(frame_admins, columns=colunas_adm, show='headings', height=4)
        
        for col in colunas_adm:
            self.tree_adm.heading(col, text=col)
            if col in ['Valor/Percentual', 'Valor Total']:
                self.tree_adm.column(col, width=120, anchor='e')
            elif col == 'Nº Parcelas':
                self.tree_adm.column(col, width=80, anchor='center')
            else:
                self.tree_adm.column(col, width=120)

        # Scrollbars
        scroll_y_adm = ttk.Scrollbar(frame_admins, orient='vertical', command=self.tree_adm.yview)
        scroll_x_adm = ttk.Scrollbar(frame_admins, orient='horizontal', command=self.tree_adm.xview)
        self.tree_adm.configure(yscrollcommand=scroll_y_adm.set, xscrollcommand=scroll_x_adm.set)
        
        self.tree_adm.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y_adm.pack(side='right', fill='y')
        scroll_x_adm.pack(side='bottom', fill='x')

        # Botões de ação
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes,
            text="Definir Eventos",
            command=self.definir_eventos,
            width=20
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Atualizar Eventos",
            command=self.atualizar_eventos_contrato,
            width=20
        ).pack(side='left', padx=5)

        # Binding para mostrar detalhes ao selecionar contrato
        self.tree_contratos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_contrato)

    def configurar_aba_eventos(self):
        """Configura a aba de gestão de eventos (com layout melhorado)"""
        frame = ttk.Frame(self.aba_eventos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para contrato selecionado
        frame_contrato = ttk.LabelFrame(frame, text="Contrato")
        frame_contrato.pack(fill='x', pady=5)

        ttk.Label(frame_contrato, text="Contrato Selecionado:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.contrato_selecionado = ttk.Combobox(frame_contrato, state='readonly', width=40)
        self.contrato_selecionado.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.contrato_selecionado.bind('<<ComboboxSelected>>', self.carregar_eventos_contrato)

        # Frame para botão de adicionar evento
        frame_adicionar = ttk.Frame(frame)
        frame_adicionar.pack(fill='x', pady=5)

        ttk.Button(
            frame_adicionar,
            text="Adicionar Evento",
            command=self.adicionar_evento,
            width=20
        ).pack(side='left', padx=5)

        # Frame para lista de eventos
        frame_eventos = ttk.LabelFrame(frame, text="Eventos do Contrato")
        frame_eventos.pack(fill='both', expand=True, pady=5)

        # Treeview para eventos
        colunas = ('ID', 'Descrição', 'Percentual', 'Valor', 'Status', 'Data Conclusão')
        self.tree_eventos = ttk.Treeview(frame_eventos, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_eventos.heading(col, text=col)
            if col == 'ID':
                self.tree_eventos.column(col, width=50, anchor='center')
            elif col in ['Percentual', 'Valor']:
                self.tree_eventos.column(col, width=100, anchor='e')
            elif col == 'Status':
                self.tree_eventos.column(col, width=100, anchor='center')
            else:
                self.tree_eventos.column(col, width=150)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_eventos, orient='vertical', command=self.tree_eventos.yview)
        scroll_x = ttk.Scrollbar(frame_eventos, orient='horizontal', command=self.tree_eventos.xview)
        self.tree_eventos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_eventos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do evento selecionado
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Evento")
        frame_detalhes.pack(fill='x', pady=10)

        # Grid para formulário
        frame_form = ttk.Frame(frame_detalhes)
        frame_form.pack(fill='x', padx=5, pady=5)

        ttk.Label(frame_form, text="Descrição:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.evento_descricao = ttk.Entry(frame_form, width=50)
        self.evento_descricao.grid(row=0, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Percentual (%):").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.evento_percentual = ttk.Entry(frame_form, width=15)
        self.evento_percentual.grid(row=1, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Status:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.evento_status = ttk.Combobox(frame_form, values=['pendente', 'concluido'], state='readonly', width=15)
        self.evento_status.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        self.evento_status.set('pendente')

        ttk.Label(frame_form, text="Data Conclusão:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.evento_data = DateEntry(frame_form, width=15, state='normal', date_pattern='dd/mm/yyyy')
        self.evento_data.grid(row=3, column=1, padx=5, pady=2, sticky='w')

        # NOVO: Usar um frame em grid para melhor organização dos botões
        frame_botoes = ttk.Frame(frame_detalhes)
        frame_botoes.pack(fill='x', pady=10, padx=5)  # Aumentado o pady

        # Criar grid para botões (2 linhas x 2 colunas)
        ttk.Button(
            frame_botoes,
            text="Salvar Evento",
            command=self.salvar_evento,
            width=20  # Aumentado a largura
        ).grid(row=0, column=0, padx=10, pady=5)

        ttk.Button(
            frame_botoes,
            text="Marcar como Concluído",
            command=self.concluir_evento,
            width=25  # Aumentado a largura
        ).grid(row=0, column=1, padx=10, pady=5)

        ttk.Button(
            frame_botoes,
            text="Remover Evento",
            command=self.remover_evento,
            width=20  # Aumentado a largura
        ).grid(row=0, column=2, padx=10, pady=5)
        
        ttk.Button(
            frame_botoes,
            text="Cancelar",
            command=lambda: self.notebook.select(self.aba_contratos),
            width=20  # Aumentado a largura
        ).grid(row=0, column=3, padx=10, pady=5)

        # Binding para seleção de evento
        self.tree_eventos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_evento)

    def configurar_aba_pagamentos(self):
        """Configura a aba de pagamentos de eventos (com layout melhorado)"""
        frame = ttk.Frame(self.aba_pagamentos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para controle de filtros
        frame_filtros = ttk.LabelFrame(frame, text="Filtros")
        frame_filtros.pack(fill='x', pady=5)

        ttk.Label(frame_filtros, text="Contrato:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.pagto_contrato = ttk.Combobox(frame_filtros, state='readonly', width=40)
        self.pagto_contrato.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.pagto_contrato.bind('<<ComboboxSelected>>', self.filtrar_pagamentos)

        ttk.Label(frame_filtros, text="Status:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        self.pagto_status = ttk.Combobox(frame_filtros, values=['Todos', 'Pendente', 'Pago'], state='readonly', width=15)
        self.pagto_status.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        self.pagto_status.set('Todos')
        self.pagto_status.bind('<<ComboboxSelected>>', self.filtrar_pagamentos)

        # Frame para lista de pagamentos
        frame_pagamentos = ttk.LabelFrame(frame, text="Pagamentos Previstos")
        frame_pagamentos.pack(fill='both', expand=True, pady=5)

        # Treeview para pagamentos
        colunas = ('Contrato', 'Evento', 'CNPJ/CPF', 'Nome', 'Data Vencimento', 'Valor', 'Status', 'Data Pagamento')
        self.tree_pagamentos = ttk.Treeview(frame_pagamentos, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_pagamentos.heading(col, text=col)
            if col in ['Data Vencimento', 'Data Pagamento']:
                self.tree_pagamentos.column(col, width=120, anchor='center')
            elif col == 'Valor':
                self.tree_pagamentos.column(col, width=100, anchor='e')
            elif col == 'Status':
                self.tree_pagamentos.column(col, width=100, anchor='center')
            elif col in ['Contrato', 'Evento']:
                self.tree_pagamentos.column(col, width=80, anchor='center')
            else:
                self.tree_pagamentos.column(col, width=150)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_pagamentos, orient='vertical', command=self.tree_pagamentos.yview)
        scroll_x = ttk.Scrollbar(frame_pagamentos, orient='horizontal', command=self.tree_pagamentos.xview)
        self.tree_pagamentos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_pagamentos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do pagamento
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Pagamento")
        frame_detalhes.pack(fill='x', pady=10)

        # Grid para formulário
        frame_form = ttk.Frame(frame_detalhes)
        frame_form.pack(fill='x', padx=5, pady=5)

        ttk.Label(frame_form, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.pagto_cnpj = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_cnpj.grid(row=0, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Nome:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.pagto_nome = ttk.Entry(frame_form, width=40, state='readonly')
        self.pagto_nome.grid(row=0, column=3, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Valor:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.pagto_valor = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_valor.grid(row=1, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Vencimento:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.pagto_vencimento = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_vencimento.grid(row=1, column=3, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Status:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.pagto_status_atual = ttk.Combobox(frame_form, values=['Pendente', 'Pago'], state='readonly', width=15)
        self.pagto_status_atual.grid(row=2, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Data Pagamento:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.pagto_data = DateEntry(frame_form, width=15, state='normal', date_pattern='dd/mm/yyyy')
        self.pagto_data.grid(row=2, column=3, padx=5, pady=2, sticky='w')

        # NOVO: Usar um frame em grid para melhor organização dos botões
        frame_botoes = ttk.Frame(frame_detalhes)
        frame_botoes.pack(fill='x', pady=10, padx=5)  # Aumentado o pady

        # Criar grid para botões (linha única, mas com espaçamento adequado)
        ttk.Button(
            frame_botoes,
            text="Registrar Pagamento",
            command=self.registrar_pagamento,
            width=25  # Aumentado a largura
        ).grid(row=0, column=0, padx=10, pady=5)

        ttk.Button(
            frame_botoes,
            text="Gerar Lançamento",
            command=self.gerar_lancamento_pagamento,
            width=25  # Aumentado a largura
        ).grid(row=0, column=1, padx=10, pady=5)
        
        ttk.Button(
            frame_botoes,
            text="Cancelar",
            command=lambda: self.notebook.select(self.aba_contratos),
            width=25  # Aumentado a largura
        ).grid(row=0, column=2, padx=10, pady=5)

        # Binding para seleção de pagamento
        self.tree_pagamentos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_pagamento)

    
        
    def carregar_contratos(self):
        """Carrega contratos do cliente atual"""
        try:
            if not self.arquivo_cliente or not self.cliente_atual:
                print("Arquivo de cliente ou cliente atual não definidos")
                return

            # Verificar se o arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                messagebox.showwarning("Aviso", f"Arquivo do cliente {self.cliente_atual} não encontrado!")
                return

            print(f"Carregando contratos do cliente: {self.cliente_atual}")
            print(f"Arquivo: {self.arquivo_cliente}")
                
            wb = load_workbook(self.arquivo_cliente, data_only=True)
            
            # Verificar se a aba Contratos_ADM existe
            if 'Contratos_ADM' not in wb.sheetnames:
                messagebox.showinfo("Informação", "Este cliente não possui contratos de administração cadastrados.")
                wb.close()
                return
                
            ws = wb['Contratos_ADM']
            
            # Limpar dados anteriores
            self.contratos = []
            self.administradores_contratos = {}
            
            for item in self.tree_contratos.get_children():
                self.tree_contratos.delete(item)

            # Processar contratos
            contratos_vistos = set()
            
            # Verificar se há dados suficientes na planilha
            if ws.max_row < 3:
                messagebox.showinfo("Informação", "Não há contratos cadastrados para este cliente.")
                wb.close()
                return
            
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verificar se a linha tem conteúdo válido
                if not row or len(row) < 4 or row[0] is None:
                    continue
                    
                num_contrato = row[0]
                if num_contrato and num_contrato not in contratos_vistos:
                    contratos_vistos.add(num_contrato)
                    
                    # Formatar datas
                    data_inicio = None
                    data_fim = None
                    
                    # Verificar se as datas são válidas
                    if len(row) > 1 and row[1]:
                        if isinstance(row[1], datetime):
                            data_inicio = row[1]
                        else:
                            try:
                                data_inicio = datetime.strptime(str(row[1]), '%Y-%m-%d')
                            except ValueError:
                                data_inicio = None
                    
                    if len(row) > 2 and row[2]:
                        if isinstance(row[2], datetime):
                            data_fim = row[2]
                        else:
                            try:
                                data_fim = datetime.strptime(str(row[2]), '%Y-%m-%d')
                            except ValueError:
                                data_fim = None
                    
                    status = row[3] if len(row) > 3 and row[3] else 'ATIVO'
                    
                    # Calcular valor total do contrato
                    valor_total = self.calcular_valor_total_contrato(ws, num_contrato)
                    
                    # Armazenar informações do contrato
                    contrato_info = {
                        'num_contrato': num_contrato,
                        'data_inicio': data_inicio,
                        'data_fim': data_fim,
                        'status': status,
                        'valor_total': valor_total,
                        'administradores': []
                    }
                    
                    self.contratos.append(contrato_info)
                    
                    # Adicionar à Treeview
                    data_inicio_str = data_inicio.strftime('%d/%m/%Y') if isinstance(data_inicio, datetime) else "-"
                    data_fim_str = data_fim.strftime('%d/%m/%Y') if isinstance(data_fim, datetime) else "-"
                    
                    self.tree_contratos.insert('', 'end', values=(
                        num_contrato,
                        data_inicio_str,
                        data_fim_str,
                        status,
                        f"R$ {valor_total:,.2f}" if valor_total else "-"
                    ))

            # Processar administradores de contratos
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verificar se a linha tem dados suficientes
                if not row or len(row) < 10:
                    continue
                    
                # Verificar se é um registro de administrador (coluna G em diante)
                if len(row) > 8 and row[6] and row[7] and row[8]:  # Tem número contrato, CNPJ/CPF e nome
                    num_contrato = row[6]
                    
                    # Adicionamos o administrador ao dicionário
                    if num_contrato not in self.administradores_contratos:
                        self.administradores_contratos[num_contrato] = []
                    
                    # Formatar CNPJ/CPF
                    cnpj_cpf = formatar_cnpj_cpf(str(row[7])) if row[7] else ""
                    
                    # Adicionar administrador
                    tipo = row[9] if len(row) > 9 and row[9] else 'Fixo'
                    valor_percentual = row[10] if len(row) > 10 and row[10] else '0'
                    valor_total = row[11] if len(row) > 11 and row[11] else '0'
                    num_parcelas = row[12] if len(row) > 12 and row[12] else '1'
                    
                    self.administradores_contratos[num_contrato].append({
                        'cnpj_cpf': cnpj_cpf,
                        'nome': row[8],
                        'tipo': tipo,
                        'valor_percentual': valor_percentual,
                        'valor_total': valor_total,
                        'num_parcelas': num_parcelas
                    })
            
            # Atualizar comboboxes
            self.atualizar_comboboxes_contratos()
            
            wb.close()
            print(f"Carregados {len(self.contratos)} contratos para o cliente {self.cliente_atual}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar contratos: {str(e)}")
            print(f"Exceção ao carregar contratos: {str(e)}")
            if 'wb' in locals():
                wb.close()

    def atualizar_comboboxes_contratos(self):
        """Atualiza as comboboxes com os contratos disponíveis (versão melhorada)"""
        try:
            print("Atualizando comboboxes de contratos")
            
            # Garantir que temos a lista de contratos
            if not hasattr(self, 'contratos') or not self.contratos:
                print("Lista de contratos vazia ou não definida")
                # Se não tiver contratos, recarregá-los
                self.carregar_contratos()
                
                if not self.contratos:
                    print("Não foi possível carregar contratos")
                    return
            
            # Filtrar contratos ativos e garantir que são únicos
            contratos_unicos = set()
            contratos_ativos = []
            
            for c in self.contratos:
                try:
                    if isinstance(c, dict) and 'num_contrato' in c and 'status' in c:
                        num_contrato = c['num_contrato']
                        
                        # Evitar duplicações
                        if num_contrato in contratos_unicos:
                            continue
                            
                        contratos_unicos.add(num_contrato)
                        
                        if c['status'] == 'ATIVO':
                            contratos_ativos.append(num_contrato)
                except Exception as e:
                    print(f"Erro ao processar contrato: {str(e)}")
            
            print(f"Contratos ativos encontrados: {len(contratos_ativos)}")
            print(f"Lista de contratos ativos: {contratos_ativos}")
            
            # Atualizar combobox na aba de eventos
            if hasattr(self, 'contrato_selecionado'):
                valor_atual = self.contrato_selecionado.get()
                self.contrato_selecionado['values'] = contratos_ativos
                
                # Manter seleção atual se possível
                if valor_atual and valor_atual in contratos_ativos:
                    self.contrato_selecionado.set(valor_atual)
                    print(f"Mantendo contrato atual selecionado: {valor_atual}")
                # Caso contrário, selecionar o primeiro contrato se houver algum
                elif contratos_ativos:
                    self.contrato_selecionado.set(contratos_ativos[0])
                    print(f"Primeiro contrato selecionado: {contratos_ativos[0]}")
                else:
                    self.contrato_selecionado.set('')
                    print("Nenhum contrato ativo encontrado")
            else:
                print("Atributo 'contrato_selecionado' não encontrado")
                
            # Atualizar combobox na aba de pagamentos
            if hasattr(self, 'pagto_contrato'):
                valor_atual = self.pagto_contrato.get()
                todos_contratos = ['Todos'] + list(contratos_unicos)
                self.pagto_contrato['values'] = todos_contratos
                
                # Manter valor atual se possível
                if valor_atual and valor_atual in todos_contratos:
                    self.pagto_contrato.set(valor_atual)
                else:
                    self.pagto_contrato.set('Todos')
                    
                print("Combobox de pagamentos atualizada")
                
        except Exception as e:
            print(f"Erro ao atualizar comboboxes: {str(e)}")
            import traceback
            traceback.print_exc()
        
    def carregar_eventos_contrato(self, event=None):
        """Carrega os eventos do contrato selecionado (versão melhorada)"""
        try:
            num_contrato = self.contrato_selecionado.get()
            if not num_contrato:
                print("Nenhum contrato selecionado na combobox")
                return
                    
            print(f"Carregando eventos do contrato: {num_contrato}")
            
            # Verificar se o arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                print(f"Arquivo não encontrado: {self.arquivo_cliente}")
                messagebox.showerror("Erro", "Arquivo do cliente não encontrado")
                return
                
            wb = load_workbook(self.arquivo_cliente, data_only=True)
            
            # Verificar se existe a aba Contratos_ADM
            if 'Contratos_ADM' not in wb.sheetnames:
                print("Aba 'Contratos_ADM' não encontrada")
                messagebox.showinfo("Informação", "Este cliente não possui contratos de administração cadastrados.")
                wb.close()
                return
                    
            ws = wb['Contratos_ADM']
            
            # Verificar dimensões da planilha
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"Dimensões da planilha: {max_row} linhas x {max_col} colunas")
            
            # Verificar se há colunas suficientes
            if max_col < 31:
                print(f"Planilha não tem colunas suficientes para eventos (tem {max_col}, precisa de pelo menos 31)")
                messagebox.showwarning("Aviso", "A estrutura da planilha não contém a área de eventos")
                wb.close()
                return
            
            # Limpar treeview
            for item in self.tree_eventos.get_children():
                self.tree_eventos.delete(item)
                    
            # Limpar lista de eventos
            self.eventos = []
            
            # Buscar valor total do contrato para cálculos percentuais
            valor_contrato = None
            for contrato in self.contratos:
                if contrato['num_contrato'] == num_contrato:
                    valor_contrato = contrato['valor_total']
                    break
            
            print(f"Valor total do contrato: {valor_contrato}")
            
            # Processar eventos com tratamento de índices
            evento_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                # Verificar se a linha tem dados suficientes
                if not row or len(row) < 31:
                    continue
                    
                # Certificar-se de que é uma linha para o contrato atual
                contrato_na_linha = row[30] if len(row) > 30 else None
                if contrato_na_linha != num_contrato:
                    continue
                    
                # Extrair dados com segurança
                evento_id = row[31] if len(row) > 31 else None
                descricao = row[32] if len(row) > 32 else None
                percentual = row[33] if len(row) > 33 else None
                status = (row[34] or 'pendente') if len(row) > 34 else 'pendente'
                
                print(f"Evento encontrado na linha {row_idx}: ID={evento_id}, Descrição={descricao}, Percentual={percentual}, Status={status}")
                
                # Se não tiver ID ou descrição, pular
                if not evento_id or not descricao:
                    print("Evento sem ID ou descrição, ignorando")
                    continue
                    
                # Calcular valor baseado no percentual
                valor = None
                if percentual and valor_contrato:
                    try:
                        perc = float(str(percentual).replace('%', '').replace(',', '.'))
                        valor = (perc / 100) * valor_contrato
                        print(f"Valor calculado: {valor}")
                    except (ValueError, TypeError) as e:
                        print(f"Erro ao calcular valor: {e}")
                        valor = 0
                
                # Formatar data de conclusão
                data_conclusao = None
                if len(row) > 35 and row[35]:  # Data conclusão
                    if isinstance(row[35], datetime):
                        data_conclusao = row[35].strftime('%d/%m/%Y')
                    else:
                        try:
                            data_conclusao = datetime.strptime(str(row[35]), '%Y-%m-%d').strftime('%d/%m/%Y')
                        except ValueError:
                            try:
                                # Tentar outros formatos comuns
                                for fmt in ['%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y']:
                                    try:
                                        data_conclusao = datetime.strptime(str(row[35]), fmt).strftime('%d/%m/%Y')
                                        break
                                    except ValueError:
                                        continue
                            except:
                                data_conclusao = str(row[35])
                
                # Adicionar à lista de eventos
                evento = {
                    'id': evento_id,
                    'descricao': descricao,
                    'percentual': percentual,
                    'valor': valor,
                    'status': status,
                    'data_conclusao': data_conclusao
                }
                self.eventos.append(evento)
                evento_count += 1
                
                # Adicionar ao treeview
                valor_fmt = f"R$ {valor:,.2f}" if valor else "-"
                percentual_fmt = f"{percentual}" if percentual else "-"
                
                self.tree_eventos.insert('', 'end', values=(
                    evento_id,
                    descricao,
                    percentual_fmt,
                    valor_fmt,
                    status.capitalize(),
                    data_conclusao or "-"
                ))
            
            print(f"Total de {evento_count} eventos carregados para o contrato {num_contrato}")
            
            # Se não encontrou eventos, mostrar mensagem
            if evento_count == 0:
                print("Nenhum evento encontrado para este contrato")
                messagebox.showinfo("Informação", f"Não há eventos cadastrados para o contrato {num_contrato}.")
            
            wb.close()
                
        except Exception as e:
            print(f"Erro ao carregar eventos: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao carregar eventos: {str(e)}")
            if 'wb' in locals():
                wb.close()
                
        except Exception as e:
            print(f"Erro ao carregar eventos: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao carregar eventos: {str(e)}")
            if 'wb' in locals():
                wb.close()
                
    def adicionar_evento(self):
        """Adiciona um novo evento ao contrato selecionado"""
        num_contrato = self.contrato_selecionado.get()
        if not num_contrato:
            messagebox.showwarning("Aviso", "Selecione um contrato primeiro")
            return
                
        # Limpar campos do formulário
        self.evento_descricao.delete(0, tk.END)
        self.evento_percentual.delete(0, tk.END)
        self.evento_status.set('pendente')
        self.evento_data.set_date(datetime.now())
        
        # Definir ID do próximo evento
        proximo_id = 1
        if self.eventos:
            ids = []
            for e in self.eventos:
                try:
                    if str(e['id']).isdigit():
                        ids.append(int(e['id']))
                except (ValueError, TypeError):
                    continue
            if ids:
                proximo_id = max(ids) + 1
        
        print(f"Próximo ID de evento: {proximo_id}")
                    
        # Mostrar janela para adicionar evento
        janela = tk.Toplevel(self.janela)
        janela.title("Adicionar Evento")
        janela.geometry("500x300")
        janela.transient(self.janela)
        janela.grab_set()
        
        frame = ttk.Frame(janela, padding=10)
        frame.pack(fill='both', expand=True)
        
        # Formulário
        ttk.Label(frame, text="Contrato:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(frame, text=num_contrato, font=('Arial', 10, 'bold')).grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame, text="ID do Evento:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        ttk.Label(frame, text=str(proximo_id), font=('Arial', 10, 'bold')).grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(frame, text="Descrição:*").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        descricao_entry = ttk.Entry(frame, width=40)
        descricao_entry.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        
        ttk.Label(frame, text="Percentual (%):*").grid(row=3, column=0, padx=5, pady=5, sticky='w')
        percentual_entry = ttk.Entry(frame, width=15)
        percentual_entry.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.grid(row=4, column=0, columnspan=2, pady=20)
        
        def confirmar():
            # Validar campos
            descricao = descricao_entry.get().strip()
            percentual_str = percentual_entry.get().strip()
            
            if not descricao:
                messagebox.showerror("Erro", "Descrição é obrigatória")
                return
                    
            if not percentual_str:
                messagebox.showerror("Erro", "Percentual é obrigatório")
                return
                    
            try:
                percentual = float(percentual_str.replace(',', '.'))
                if percentual <= 0 or percentual > 100:
                    messagebox.showerror("Erro", "Percentual deve estar entre 0 e 100")
                    return
            except ValueError:
                messagebox.showerror("Erro", "Percentual inválido")
                return
                    
            # Calcular total de percentuais já existentes
            percentual_atual = 0
            for e in self.eventos:
                try:
                    if e['percentual']:
                        perc_str = str(e['percentual']).replace('%', '').replace(',', '.')
                        if perc_str.replace('.', '').isdigit():
                            percentual_atual += float(perc_str)
                except (ValueError, TypeError, AttributeError) as err:
                    print(f"Erro ao processar percentual existente: {err}")
            
            print(f"Percentual atual: {percentual_atual}%, Novo: {percentual}%")
                
            # Verificar se excede 100%
            if percentual_atual + percentual > 100:
                messagebox.showerror(
                    "Erro", 
                    f"Total de percentuais excede 100%! Atual: {percentual_atual:.2f}%, Novo: {percentual:.2f}%"
                )
                return
                    
            # Salvar evento
            try:
                self.salvar_novo_evento(num_contrato, proximo_id, descricao, percentual)
                messagebox.showinfo("Sucesso", "Evento adicionado com sucesso!")
                janela.destroy()
                # Atualizar lista de eventos
                self.carregar_eventos_contrato(None)
            except Exception as e:
                print(f"Erro ao salvar evento: {str(e)}")
                import traceback
                traceback.print_exc()
                messagebox.showerror("Erro", f"Erro ao salvar evento: {str(e)}")
        
        ttk.Button(frame_botoes, text="Confirmar", command=confirmar, width=15).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela.destroy, width=15).pack(side='left', padx=5)
        
        # Centralizar janela
        janela.update_idletasks()
        width = janela.winfo_width()
        height = janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (width // 2)
        y = (janela.winfo_screenheight() // 2) - (height // 2)
        janela.geometry(f'{width}x{height}+{x}+{y}')

    def salvar_novo_evento(self, num_contrato, evento_id, descricao, percentual):
        """Salva um novo evento na planilha"""
        try:
            print(f"Salvando novo evento: Contrato={num_contrato}, ID={evento_id}, Descrição={descricao}, Percentual={percentual}%")
            
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            # Determinar se precisamos criar um novo registro ou atualizar um existente
            linha_existente = None
            
            # Primeiro verificar se já existe uma linha para este contrato e evento
            for row in range(3, ws.max_row + 1):
                if (ws.cell(row=row, column=31).value == num_contrato and 
                    str(ws.cell(row=row, column=32).value) == str(evento_id)):
                    linha_existente = row
                    break
            
            print(f"Linha existente encontrada: {linha_existente}")
            
            if linha_existente:
                # Atualizar linha existente
                ws.cell(row=linha_existente, column=33, value=descricao)     # Descrição
                ws.cell(row=linha_existente, column=34, value=f"{percentual:.2f}%")  # Percentual
                ws.cell(row=linha_existente, column=35, value="pendente")    # Status
            else:
                # Encontrar próxima linha disponível
                proxima_linha = ws.max_row + 1
                
                # Verificar se é a primeira linha de eventos
                primeira_linha_eventos = True
                for row in range(3, ws.max_row + 1):
                    if ws.cell(row=row, column=31).value is not None:
                        primeira_linha_eventos = False
                        break
                
                # Se for a primeira linha, adicionar cabeçalhos
                if primeira_linha_eventos:
                    headers = [
                        "Contrato", "ID Evento", "Descrição", "Percentual", "Status", "Data Conclusão"
                    ]
                    for i, header in enumerate(headers, start=31):
                        ws.cell(row=2, column=i, value=header)
                
                # Salvar evento
                ws.cell(row=proxima_linha, column=31, value=num_contrato)  # Contrato
                ws.cell(row=proxima_linha, column=32, value=evento_id)     # ID Evento
                ws.cell(row=proxima_linha, column=33, value=descricao)     # Descrição
                ws.cell(row=proxima_linha, column=34, value=f"{percentual:.2f}%")  # Percentual
                ws.cell(row=proxima_linha, column=35, value="pendente")    # Status
            
            print("Salvando planilha...")
            wb.save(self.arquivo_cliente)
            print("Planilha salva com sucesso!")
            
            # Atualizar lista interna de eventos
            valor_contrato = None
            for contrato in self.contratos:
                if contrato['num_contrato'] == num_contrato:
                    valor_contrato = contrato['valor_total']
                    break
                        
            valor = None
            if valor_contrato:
                valor = (percentual / 100) * valor_contrato
                    
            evento = {
                'id': evento_id,
                'descricao': descricao,
                'percentual': f"{percentual:.2f}%",
                'valor': valor,
                'status': 'pendente',
                'data_conclusao': None
            }
            self.eventos.append(evento)
            print("Evento adicionado à lista interna")
                
        except Exception as e:
            print(f"Erro ao salvar evento: {str(e)}")
            import traceback
            traceback.print_exc()
            raise Exception(f"Erro ao salvar evento: {str(e)}")

    def mostrar_detalhes_contrato(self, event):
        """Mostra detalhes do contrato selecionado e seus administradores (com sincronização)"""
        try:
            print("Entrando em mostrar_detalhes_contrato")
            
            # Verificar se há seleção
            selecionado = self.tree_contratos.selection()
            if not selecionado:
                print("Nenhum contrato selecionado")
                return

            # Obter valores com tratamento de erro
            try:
                valores = self.tree_contratos.item(selecionado)['values']
                if not valores or len(valores) < 1:
                    print("Valores do contrato selecionado são inválidos")
                    return
                    
                num_contrato = valores[0] if len(valores) > 0 else "N/A"
                print(f"Contrato selecionado: {num_contrato}")
            except Exception as e:
                print(f"Erro ao obter valores do contrato: {str(e)}")
                return

            # Buscar o contrato na lista
            contrato = None
            for c in self.contratos:
                try:
                    if isinstance(c, dict) and 'num_contrato' in c and c['num_contrato'] == num_contrato:
                        contrato = c
                        break
                except Exception as e:
                    print(f"Erro ao verificar contrato na lista: {str(e)}")
                    
            if not contrato:
                print(f"Contrato {num_contrato} não encontrado na lista interna")
                return

            print(f"Detalhes do contrato: {contrato}")
                
            # Atualizar labels com informações do contrato (com verificações)
            if hasattr(self, 'lbl_contrato'):
                self.lbl_contrato.config(text=str(num_contrato))
            
            # Formatação de datas com tratamento de erros
            periodo = ""
            try:
                data_inicio = contrato.get('data_inicio')
                data_fim = contrato.get('data_fim')
                
                if data_inicio and data_fim:
                    data_inicio_str = data_inicio.strftime('%d/%m/%Y') if isinstance(data_inicio, datetime) else str(data_inicio)
                    data_fim_str = data_fim.strftime('%d/%m/%Y') if isinstance(data_fim, datetime) else str(data_fim)
                    periodo = f"{data_inicio_str} a {data_fim_str}"
                elif data_inicio:
                    data_inicio_str = data_inicio.strftime('%d/%m/%Y') if isinstance(data_inicio, datetime) else str(data_inicio)
                    periodo = f"A partir de {data_inicio_str}"
                else:
                    periodo = "Sem período definido"
            except Exception as e:
                print(f"Erro ao formatar datas: {str(e)}")
                periodo = "Erro ao formatar datas"
            
            if hasattr(self, 'lbl_periodo'):
                self.lbl_periodo.config(text=periodo)
                
            if hasattr(self, 'lbl_status'):
                self.lbl_status.config(text=contrato.get('status', 'N/A'))
            
            # Formatação de valor com tratamento de erros
            valor_str = "N/A"
            try:
                valor_total = contrato.get('valor_total')
                if valor_total is not None:
                    valor_str = f"R$ {valor_total:,.2f}"
            except Exception as e:
                print(f"Erro ao formatar valor: {str(e)}")
                valor_str = "Erro ao formatar valor"
                
            if hasattr(self, 'lbl_valor_total'):
                self.lbl_valor_total.config(text=valor_str)
            
            # Mostrar administradores com tratamento de erros
            num_adm = 0
            try:
                administradores = self.administradores_contratos.get(num_contrato, [])
                num_adm = len(administradores)
            except Exception as e:
                print(f"Erro ao contar administradores: {str(e)}")
                
            if hasattr(self, 'lbl_administradores'):
                self.lbl_administradores.config(text=str(num_adm))
            
            # Mostrar número de eventos com tratamento de erros
            eventos_count = 0
            try:
                eventos_count = self.contar_eventos_contrato(num_contrato)
            except Exception as e:
                print(f"Erro ao contar eventos: {str(e)}")
                
            if hasattr(self, 'lbl_eventos'):
                self.lbl_eventos.config(text=str(eventos_count))
            
            # Limpar e preencher tree de administradores com tratamento de erros
            try:
                if hasattr(self, 'tree_adm'):
                    for item in self.tree_adm.get_children():
                        self.tree_adm.delete(item)
                    
                    administradores = self.administradores_contratos.get(num_contrato, [])
                    for adm in administradores:
                        if not isinstance(adm, dict):
                            continue
                            
                        cnpj_cpf = adm.get('cnpj_cpf', '')
                        nome = adm.get('nome', '')
                        tipo = adm.get('tipo', '')
                        
                        # Formatações com tratamento de erros
                        valor_percentual_fmt = "-"
                        try:
                            valor_percentual = adm.get('valor_percentual', '')
                            if tipo == 'Percentual' and valor_percentual:
                                valor_percentual_fmt = f"{valor_percentual}%"
                        except Exception as e:
                            print(f"Erro ao formatar percentual: {str(e)}")
                        
                        valor_total_adm_fmt = "-"
                        try:
                            valor_total_adm = adm.get('valor_total', '')
                            if valor_total_adm:
                                valor_total_adm = float(str(valor_total_adm).replace('.', '').replace(',', '.'))
                                valor_total_adm_fmt = f"R$ {valor_total_adm:,.2f}"
                        except (ValueError, TypeError, AttributeError) as e:
                            print(f"Erro ao formatar valor total do administrador: {str(e)}")
                        
                        num_parcelas = adm.get('num_parcelas', '')
                        
                        self.tree_adm.insert('', 'end', values=(
                            cnpj_cpf,
                            nome,
                            tipo,
                            valor_percentual_fmt,
                            valor_total_adm_fmt,
                            num_parcelas
                        ))
            except Exception as e:
                print(f"Erro ao preencher tree de administradores: {str(e)}")
                import traceback
                traceback.print_exc()
                
            # NOVO: Pré-selecionar este contrato na aba de eventos
            if hasattr(self, 'contrato_selecionado'):
                valores_combo = list(self.contrato_selecionado['values'])
                if num_contrato in valores_combo:
                    self.contrato_selecionado.set(num_contrato)
                    # Pré-carregar os eventos para este contrato
                    self.carregar_eventos_contrato(None)
                    print(f"Contrato {num_contrato} pré-selecionado na aba de eventos")
                
            print("Detalhes do contrato exibidos com sucesso")
                
        except Exception as e:
            print(f"Erro geral em mostrar_detalhes_contrato: {str(e)}")
            import traceback
            traceback.print_exc()


    def mostrar_detalhes_evento(self, event):
        """Mostra detalhes do evento selecionado"""
        selecionado = self.tree_eventos.selection()
        if not selecionado:
            return
            
        valores = self.tree_eventos.item(selecionado)['values']
        evento_id = valores[0]
        
        # Buscar o evento na lista
        evento = next((e for e in self.eventos if str(e['id']) == str(evento_id)), None)
        if not evento:
            return
            
        # Preencher campos
        self.evento_descricao.delete(0, tk.END)
        self.evento_descricao.insert(0, evento['descricao'])
        
        self.evento_percentual.delete(0, tk.END)
        percentual_str = str(evento['percentual']).replace('%', '').strip()
        self.evento_percentual.insert(0, percentual_str)
        
        self.evento_status.set(evento['status'])
        
        if evento['data_conclusao']:
            try:
                data = datetime.strptime(evento['data_conclusao'], '%d/%m/%Y')
                self.evento_data.set_date(data)
            except ValueError:
                self.evento_data.set_date(datetime.now())

    def calcular_valor_total_contrato(self, ws, num_contrato):
        """Calcula o valor total do contrato somando os valores dos administradores"""
        valor_total = 0
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[6] == num_contrato and row[9] == 'Fixo' and row[11]:  # Verifica se é administrador fixo com valor total
                try:
                    valor_total += float(str(row[11]).replace('.', '').replace(',', '.'))
                except (ValueError, TypeError):
                    pass
        
        return valor_total if valor_total > 0 else None
    
    def mostrar_detalhes_pagamento(self, event):
        """Mostra detalhes do pagamento selecionado"""
        selecionado = self.tree_pagamentos.selection()
        if not selecionado:
            return
            
        valores = self.tree_pagamentos.item(selecionado)['values']
        
        # Preencher campos
        self.pagto_cnpj.config(state='normal')
        self.pagto_cnpj.delete(0, tk.END)
        self.pagto_cnpj.insert(0, valores[2])
        self.pagto_cnpj.config(state='readonly')
        
        self.pagto_nome.config(state='normal')
        self.pagto_nome.delete(0, tk.END)
        self.pagto_nome.insert(0, valores[3])
        self.pagto_nome.config(state='readonly')
        
        self.pagto_valor.config(state='normal')
        self.pagto_valor.delete(0, tk.END)
        self.pagto_valor.insert(0, valores[5])
        self.pagto_valor.config(state='readonly')
        
        self.pagto_vencimento.config(state='normal')
        self.pagto_vencimento.delete(0, tk.END)
        self.pagto_vencimento.insert(0, valores[4])
        self.pagto_vencimento.config(state='readonly')
        
        # Status
        self.pagto_status_atual.set(valores[6])
        
        # Data de pagamento
        if valores[7] != "-":
            try:
                data = datetime.strptime(valores[7], '%d/%m/%Y')
                self.pagto_data.set_date(data)
            except ValueError:
                self.pagto_data.set_date(datetime.now())

    def contar_eventos_contrato(self, num_contrato):
        """Conta o número de eventos cadastrados para o contrato (versão robusta)"""
        try:
            print(f"Contando eventos para o contrato: {num_contrato}")
            
            # Verificar se o arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                print(f"Arquivo do cliente não encontrado: {self.arquivo_cliente}")
                return 0
                
            wb = load_workbook(self.arquivo_cliente, data_only=True)
            
            # Verificar se a aba existe
            if 'Contratos_ADM' not in wb.sheetnames:
                print("Aba 'Contratos_ADM' não encontrada na planilha")
                wb.close()
                return 0
                
            ws = wb['Contratos_ADM']
            
            eventos_count = 0
            
            # Verificar o máximo de linhas e colunas
            max_row = ws.max_row
            max_col = ws.max_column
            
            print(f"Dimensões da planilha: {max_row} linhas x {max_col} colunas")
            
            # Verificar se há colunas suficientes
            if max_col < 31:
                print(f"Planilha não tem colunas suficientes para eventos (tem {max_col}, precisa de pelo menos 31)")
                wb.close()
                return 0
            
            # Contar eventos com verificação de comprimento das linhas
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verificar se a linha tem elementos suficientes
                if not row or len(row) <= 30:
                    continue
                    
                # Verificar se é um evento deste contrato
                contrato_value = row[30]  # Coluna AE (31ª coluna)
                if contrato_value == num_contrato:
                    eventos_count += 1
                    
            wb.close()
            print(f"Total de {eventos_count} eventos encontrados para o contrato {num_contrato}")
            return eventos_count
            
        except Exception as e:
            print(f"Erro ao contar eventos: {str(e)}")
            import traceback
            traceback.print_exc()
            if 'wb' in locals():
                wb.close()
            return 0

    def definir_eventos(self):
        """Abre a tela para definir eventos do contrato selecionado (versão corrigida)"""
        try:
            print("Iniciando definição de eventos")
            
            # Verificar se há seleção
            selecionado = self.tree_contratos.selection()
            if not selecionado:
                print("Nenhum contrato selecionado")
                messagebox.showwarning("Aviso", "Selecione um contrato primeiro")
                return
                
            # Obter o número do contrato
            valores = self.tree_contratos.item(selecionado)['values']
            if not valores or len(valores) < 4:
                print("Valores do contrato selecionado são inválidos")
                messagebox.showwarning("Aviso", "Contrato selecionado não possui informações completas")
                return
                
            num_contrato = valores[0]
            status = valores[3] if len(valores) > 3 else None
            
            print(f"Contrato selecionado: {num_contrato}, Status: {status}")
            
            # Verificar status do contrato
            if status != 'ATIVO':
                print(f"Contrato {num_contrato} não está ativo (status: {status})")
                messagebox.showwarning("Aviso", "Apenas contratos ativos podem ter eventos definidos")
                return
            
            # Mudar para a aba de eventos
            print("Mudando para a aba de eventos")
            self.notebook.select(self.aba_eventos)
            
            # Garantir que a combobox de contrato existe
            if not hasattr(self, 'contrato_selecionado'):
                print("Combobox de contrato não encontrada na aba de eventos")
                messagebox.showerror("Erro", "Erro na interface: combobox de contrato não encontrada")
                return
                
            # Verificar se o contrato está na lista de valores da combobox
            valores_combo = list(self.contrato_selecionado['values'])
            print(f"Valores disponíveis na combobox: {valores_combo}")
            
            if not valores_combo:
                print("Atualizando comboboxes de contratos")
                self.atualizar_comboboxes_contratos()
                valores_combo = list(self.contrato_selecionado['values'])
                
            if num_contrato not in valores_combo:
                print(f"Contrato {num_contrato} não está nos valores da combobox. Atualizando valores.")
                # Adicionar o contrato à lista se não estiver lá
                self.contrato_selecionado['values'] = tuple(list(valores_combo) + [num_contrato])
            
            # Selecionar o contrato na combobox
            self.contrato_selecionado.set(num_contrato)
            print(f"Contrato {num_contrato} selecionado na combobox")
            
            # Chamar explicitamente o método para carregar eventos
            print("Carregando eventos do contrato")
            self.carregar_eventos_contrato(None)
            
            # Verificar se os eventos foram carregados
            eventos_count = len(self.eventos)
            print(f"Total de {eventos_count} eventos carregados")
            
            # Se não houver eventos, oferecer para adicionar
            if eventos_count == 0:
                if messagebox.askyesno("Eventos", "Este contrato não possui eventos definidos. Deseja adicionar um evento agora?"):
                    print("Iniciando adição de evento")
                    self.adicionar_evento()
                    
        except Exception as e:
            print(f"Erro ao definir eventos: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao definir eventos: {str(e)}")      

    def atualizar_eventos_contrato(self):
        """Atualiza a lista de eventos do contrato selecionado (versão robusta)"""
        try:
            print("Entrando em atualizar_eventos_contrato")
            
            # Verificar se há seleção
            selecionado = self.tree_contratos.selection()
            if not selecionado:
                print("Nenhum contrato selecionado")
                messagebox.showwarning("Aviso", "Selecione um contrato primeiro")
                return
                
            # Obter valores com tratamento de erro
            try:
                valores = self.tree_contratos.item(selecionado)['values']
                if not valores or len(valores) < 4:
                    print("Valores do contrato selecionado são inválidos")
                    messagebox.showwarning("Aviso", "Contrato selecionado não possui informações completas")
                    return
                    
                num_contrato = valores[0]
                status = valores[3] if len(valores) > 3 else None
                
                print(f"Contrato selecionado: {num_contrato}, Status: {status}")
            except Exception as e:
                print(f"Erro ao obter valores do contrato: {str(e)}")
                messagebox.showerror("Erro", "Erro ao obter informações do contrato")
                return
            
            # Verificar status do contrato
            if status != 'ATIVO':
                print(f"Contrato {num_contrato} não está ativo (status: {status})")
                messagebox.showwarning("Aviso", "Apenas contratos ativos podem ter eventos atualizados")
                return
                
            print("Mudando para aba de eventos")
            # Carregar eventos do contrato novamente
            try:
                # Mudar para a aba de eventos
                if hasattr(self, 'notebook') and hasattr(self, 'aba_eventos'):
                    self.notebook.select(self.aba_eventos)
                    
                # Selecionar o contrato na combobox
                if hasattr(self, 'contrato_selecionado'):
                    # Verificar se o contrato está na lista de valores
                    valores_combo = self.contrato_selecionado['values']
                    if num_contrato in valores_combo:
                        self.contrato_selecionado.set(num_contrato)
                        print(f"Contrato {num_contrato} selecionado na combobox")
                    else:
                        print(f"Contrato {num_contrato} não encontrado nos valores da combobox")
                        messagebox.showwarning("Aviso", f"Contrato {num_contrato} não está disponível para seleção")
                        return
                    
                    # Carregar eventos
                    print("Chamando carregar_eventos_contrato")
                    self.carregar_eventos_contrato(None)
                else:
                    print("Atributo 'contrato_selecionado' não encontrado")
                    messagebox.showerror("Erro", "Erro ao acessar a interface de eventos")
                    
            except Exception as e:
                print(f"Erro ao carregar eventos: {str(e)}")
                import traceback
                traceback.print_exc()
                messagebox.showerror("Erro", f"Erro ao carregar eventos: {str(e)}")
                
        except Exception as e:
            print(f"Erro geral em atualizar_eventos_contrato: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao atualizar eventos: {str(e)}")

    def salvar_evento(self):
        """Salva as alterações em um evento existente"""
        selecionado = self.tree_eventos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um evento primeiro")
            return
            
        valores = self.tree_eventos.item(selecionado)['values']
        evento_id = valores[0]
        num_contrato = self.contrato_selecionado.get()
        
        # Validar campos
        descricao = self.evento_descricao.get().strip()
        percentual_str = self.evento_percentual.get().strip()
        status = self.evento_status.get()
        
        if not descricao:
            messagebox.showerror("Erro", "Descrição é obrigatória")
            return
            
        if not percentual_str:
            messagebox.showerror("Erro", "Percentual é obrigatório")
            return
            
        try:
            percentual = float(percentual_str.replace(',', '.'))
            if percentual <= 0 or percentual > 100:
                messagebox.showerror("Erro", "Percentual deve estar entre 0 e 100")
                return
        except ValueError:
            messagebox.showerror("Erro", "Percentual inválido")
            return
            
        # Calcular total de percentuais já existentes (excluindo o evento atual)
        percentual_atual = sum([
            float(str(e['percentual']).replace('%', '').replace(',', '.'))
            for e in self.eventos
            if e['percentual'] and str(e['id']) != str(evento_id) and 
            str(e['percentual']).replace('%', '').replace(',', '.').replace('.', '').isdigit()
        ])
        
        # Verificar se excede 100%
        if percentual_atual + percentual > 100:
            messagebox.showerror(
                "Erro", 
                f"Total de percentuais excede 100%! Outros eventos: {percentual_atual:.2f}%, Este evento: {percentual:.2f}%"
            )
            return
            
        # Atualizar evento
        try:
            # Formatar data
            data_conclusao = self.evento_data.get_date().strftime('%d/%m/%Y') if status == 'concluido' else None
            
            self.atualizar_evento(num_contrato, evento_id, descricao, percentual, status, data_conclusao)
            messagebox.showinfo("Sucesso", "Evento atualizado com sucesso!")
            
            # Atualizar lista de eventos
            self.carregar_eventos_contrato(None)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar evento: {str(e)}")
            
    def atualizar_evento(self, num_contrato, evento_id, descricao, percentual, status, data_conclusao):
        """Atualiza um evento existente na planilha"""
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            # Encontrar o evento
            evento_encontrado = False
            for row in range(3, ws.max_row + 1):
                if (ws.cell(row=row, column=31).value == num_contrato and 
                    str(ws.cell(row=row, column=32).value) == str(evento_id)):
                    
                    # Atualizar valores
                    ws.cell(row=row, column=33, value=descricao)         # Descrição
                    ws.cell(row=row, column=34, value=f"{percentual:.2f}%")  # Percentual
                    ws.cell(row=row, column=35, value=status)            # Status
                    
                    # Data de conclusão apenas se status for 'concluido'
                    if status == 'concluido' and data_conclusao:
                        data_obj = datetime.strptime(data_conclusao, '%d/%m/%Y')
                        ws.cell(row=row, column=36, value=data_obj)      # Data conclusão
                    elif status != 'concluido':
                        ws.cell(row=row, column=36, value=None)          # Limpar data
                        
                    evento_encontrado = True
                    break
                    
            if not evento_encontrado:
                raise Exception(f"Evento {evento_id} não encontrado para o contrato {num_contrato}")
                
            wb.save(self.arquivo_cliente)
            
            # Se o evento foi concluído, gerar pagamento
            if status == 'concluido':
                self.gerar_pagamento_evento(num_contrato, evento_id, descricao, percentual, data_conclusao)
                
        except Exception as e:
            raise Exception(f"Erro ao atualizar evento: {str(e)}")

    def concluir_evento(self):
        """Marca um evento como concluído e gera pagamento"""
        selecionado = self.tree_eventos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um evento primeiro")
            return
            
        valores = self.tree_eventos.item(selecionado)['values']
        evento_id = valores[0]
        num_contrato = self.contrato_selecionado.get()
        
        # Verificar se o evento já está concluído
        if valores[4] == 'Concluido':
            messagebox.showinfo("Informação", "Este evento já está concluído")
            return
            
        # Confirmar conclusão
        if not messagebox.askyesno("Confirmar", "Confirma a conclusão deste evento? Isso gerará pagamentos para os administradores."):
            return
            
        # Buscar detalhes do evento
        evento = next((e for e in self.eventos if str(e['id']) == str(evento_id)), None)
        if not evento:
            messagebox.showerror("Erro", "Evento não encontrado")
            return
            
        # Marcar como concluído
        try:
            # Obter data atual
            data_conclusao = datetime.now().strftime('%d/%m/%Y')
            
            # Atualizar campos
            self.evento_status.set('concluido')
            self.evento_data.set_date(datetime.now())
            
            # Atualizar evento
            self.atualizar_evento(
                num_contrato, 
                evento_id, 
                evento['descricao'], 
                float(str(evento['percentual']).replace('%', '').replace(',', '.')), 
                'concluido', 
                data_conclusao
            )
            
            messagebox.showinfo("Sucesso", "Evento concluído com sucesso! Pagamentos foram gerados.")
            
            # Atualizar lista de eventos
            self.carregar_eventos_contrato(None)
            
            # Ir para aba de pagamentos
            self.notebook.select(self.aba_pagamentos)
            self.pagto_contrato.set(num_contrato)
            self.filtrar_pagamentos(None)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao concluir evento: {str(e)}")

    def gerar_pagamento_evento(self, num_contrato, evento_id, descricao, percentual, data_conclusao):
        """Gera pagamentos para os administradores baseado no evento concluído"""
        try:
            # Buscar valor do contrato
            valor_contrato = None
            for contrato in self.contratos:
                if contrato['num_contrato'] == num_contrato:
                    valor_contrato = contrato['valor_total']
                    break
                    
            if not valor_contrato:
                raise Exception(f"Não foi possível encontrar o valor total do contrato {num_contrato}")
                
            # Buscar administradores do contrato
            administradores = self.administradores_contratos.get(num_contrato, [])
            if not administradores:
                raise Exception(f"Não há administradores cadastrados para o contrato {num_contrato}")
                
            # Valor do evento
            valor_evento = (percentual / 100) * valor_contrato
            
            # Determinar data de vencimento (15 dias após conclusão)
            data_conclusao_obj = datetime.strptime(data_conclusao, '%d/%m/%Y')
            data_vencimento = data_conclusao_obj + relativedelta(days=15)
            
            # Salvar pagamentos para cada administrador
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            for admin in administradores:
                # Calcular valor do pagamento
                if admin['tipo'] == 'Percentual':
                    # Administrador com percentual sobre o contrato
                    try:
                        admin_percentual = float(str(admin['valor_percentual']).replace('%', '').replace(',', '.'))
                        valor_pagamento = (admin_percentual / 100) * valor_evento
                    except (ValueError, TypeError):
                        valor_pagamento = 0
                else:
                    # Administrador com valor fixo - calcula proporcional ao evento
                    try:
                        valor_total_admin = float(str(admin['valor_total']).replace('.', '').replace(',', '.'))
                        valor_pagamento = (percentual / 100) * valor_total_admin
                    except (ValueError, TypeError):
                        valor_pagamento = 0
                
                # Encontrar próxima linha disponível para pagamentos
                proxima_linha = ws.max_row + 1
                
                # Salvar pagamento
                ws.cell(row=proxima_linha, column=38, value=num_contrato)          # Contrato
                ws.cell(row=proxima_linha, column=39, value=evento_id)             # ID Evento
                ws.cell(row=proxima_linha, column=40, value=admin['cnpj_cpf'])     # CNPJ/CPF
                ws.cell(row=proxima_linha, column=41, value=admin['nome'])         # Nome
                ws.cell(row=proxima_linha, column=42, value=data_vencimento)       # Data Vencimento
                ws.cell(row=proxima_linha, column=43, value=valor_pagamento)       # Valor
                ws.cell(row=proxima_linha, column=44, value='pendente')            # Status
                ws.cell(row=proxima_linha, column=45, value=f"Pagamento referente ao evento: {descricao}")  # Observação
            
            wb.save(self.arquivo_cliente)
            
        except Exception as e:
            raise Exception(f"Erro ao gerar pagamentos: {str(e)}")
            
    def remover_evento(self):
        """Remove um evento não concluído"""
        selecionado = self.tree_eventos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um evento primeiro")
            return
            
        valores = self.tree_eventos.item(selecionado)['values']
        evento_id = valores[0]
        num_contrato = self.contrato_selecionado.get()
        
        # Verificar se o evento já está concluído
        if valores[4] == 'Concluido':
            messagebox.showerror("Erro", "Não é possível remover um evento já concluído")
            return
            
        # Confirmar remoção
        if not messagebox.askyesno("Confirmar", "Tem certeza que deseja remover este evento?"):
            return
            
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            # Encontrar o evento
            linha_evento = None
            for row in range(3, ws.max_row + 1):
                if (ws.cell(row=row, column=31).value == num_contrato and 
                    str(ws.cell(row=row, column=32).value) == str(evento_id)):
                    linha_evento = row
                    break
                    
            if not linha_evento:
                raise Exception(f"Evento {evento_id} não encontrado para o contrato {num_contrato}")
                
            # Remover linha
            ws.delete_rows(linha_evento)
            
            wb.save(self.arquivo_cliente)
            
            messagebox.showinfo("Sucesso", "Evento removido com sucesso!")
            
            # Atualizar lista de eventos
            self.carregar_eventos_contrato(None)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao remover evento: {str(e)}")

    def filtrar_pagamentos(self, event):
        """Filtra pagamentos conforme seleção"""
        contrato = self.pagto_contrato.get()
        status = self.pagto_status.get()
        
        try:
            self.carregar_pagamentos(contrato, status)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao filtrar pagamentos: {str(e)}")

    def carregar_pagamentos(self, contrato_filtro, status_filtro):
        """Carrega os pagamentos de eventos"""
        try:
            wb = load_workbook(self.arquivo_cliente, data_only=True)
            
            # Verificar se existe a aba
            if 'Contratos_ADM' not in wb.sheetnames:
                wb.close()
                return
                
            ws = wb['Contratos_ADM']
            
            # Limpar treeview
            for item in self.tree_pagamentos.get_children():
                self.tree_pagamentos.delete(item)
                
            # Processar pagamentos
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verificar se é um registro de pagamento (coluna AM ou 38 em diante)
                if len(row) >= 44 and row[37] and row[38]:
                    num_contrato = row[37]
                    evento_id = row[38]
                    cnpj_cpf = row[39]
                    nome = row[40]
                    
                    # Aplicar filtro de contrato
                    if contrato_filtro != 'Todos' and num_contrato != contrato_filtro:
                        continue
                    
                    # Formatar data de vencimento
                    data_vencimento = None
                    if row[41]:
                        if isinstance(row[41], datetime):
                            data_vencimento = row[41].strftime('%d/%m/%Y')
                        else:
                            try:
                                data_vencimento = datetime.strptime(str(row[41]), '%Y-%m-%d').strftime('%d/%m/%Y')
                            except ValueError:
                                data_vencimento = str(row[41])
                    
                    valor = row[42]
                    status = row[43] or 'pendente'
                    
                    # Aplicar filtro de status
                    if status_filtro != 'Todos' and status.capitalize() != status_filtro:
                        continue
                    
                    # Formatar data de pagamento
                    data_pagamento = None
                    if row[44]:
                        if isinstance(row[44], datetime):
                            data_pagamento = row[44].strftime('%d/%m/%Y')
                        else:
                            try:
                                data_pagamento = datetime.strptime(str(row[44]), '%Y-%m-%d').strftime('%d/%m/%Y')
                            except ValueError:
                                data_pagamento = str(row[44])
                    
                    # Adicionar ao treeview
                    valor_fmt = f"R$ {valor:,.2f}" if isinstance(valor, (int, float)) else valor
                    
                    self.tree_pagamentos.insert('', 'end', values=(
                        num_contrato,
                        evento_id,
                        str(cnpj_cpf),  # Garantir que seja string
                        str(nome),      # Garantir que seja string
                        data_vencimento or "-",
                        valor_fmt,
                        status.capitalize(),
                        data_pagamento or "-"
                    ))
            
            wb.close()
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Erro ao carregar pagamentos: {str(e)}")

    
            
    def registrar_pagamento(self):
        """Registra um pagamento como efetuado (versão corrigida)"""
        try:
            print("Iniciando registro de pagamento")
            selecionado = self.tree_pagamentos.selection()
            if not selecionado:
                messagebox.showwarning("Aviso", "Selecione um pagamento primeiro")
                return
                
            valores = self.tree_pagamentos.item(selecionado)['values']
            print(f"Valores do pagamento selecionado: {valores}")
            
            num_contrato = valores[0]
            evento_id = valores[1]
            cnpj_cpf = valores[2]
            
            # Verificar se o pagamento já está como pago
            status = valores[6] if len(valores) > 6 else "Pendente"
            if status == 'Pago':
                print("Pagamento já está registrado como pago")
                messagebox.showinfo("Informação", "Este pagamento já está registrado como pago")
                return
                
            # Confirmar registro
            if not messagebox.askyesno("Confirmar", "Confirma o registro deste pagamento?"):
                print("Registro cancelado pelo usuário")
                return
                
            print(f"Registrando pagamento: Contrato={num_contrato}, Evento={evento_id}, CNPJ/CPF={cnpj_cpf}")
            
            # Verificar se arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                print(f"Arquivo do cliente não encontrado: {self.arquivo_cliente}")
                messagebox.showerror("Erro", "Arquivo do cliente não encontrado")
                return
                
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba existe
            if 'Contratos_ADM' not in wb.sheetnames:
                print("Aba 'Contratos_ADM' não encontrada")
                messagebox.showerror("Erro", "Estrutura da planilha inválida: aba 'Contratos_ADM' não encontrada")
                wb.close()
                return
                
            ws = wb['Contratos_ADM']
            
            # Encontrar o pagamento
            pagamento_encontrado = False
            for row in range(3, ws.max_row + 1):
                # Verificar se a linha tem dados nas colunas de pagamento
                if (ws.cell(row=row, column=38).value == num_contrato and 
                    str(ws.cell(row=row, column=39).value) == str(evento_id) and
                    str(ws.cell(row=row, column=40).value) == str(cnpj_cpf)):
                    
                    print(f"Pagamento encontrado na linha {row}")
                    
                    # Atualizar status
                    ws.cell(row=row, column=44, value='pago')
                    
                    # Atualizar data de pagamento
                    data_pagamento = self.pagto_data.get_date()
                    ws.cell(row=row, column=45, value=data_pagamento)
                    
                    pagamento_encontrado = True
                    break
                    
            if not pagamento_encontrado:
                print("Pagamento não encontrado na planilha")
                messagebox.showerror("Erro", "Pagamento não encontrado na planilha")
                wb.close()
                return
            
            # Verificar se também precisamos atualizar a planilha "Dados"
            atualizar_dados = False
            if 'Dados' in wb.sheetnames:
                print("Verificando aba 'Dados' para atualização")
                ws_dados = wb['Dados']
                
                # Procurar o lançamento correspondente na aba Dados
                for row in range(2, ws_dados.max_row + 1):
                    referencia = ws_dados.cell(row=row, column=6).value  # Coluna F - Referência
                    if referencia and f"EVENTO {evento_id}" in str(referencia) and f"CONTRATO {num_contrato}" in str(referencia):
                        print(f"Lançamento encontrado na aba Dados, linha {row}")
                        
                        # Atualizar status para pago
                        ws_dados.cell(row=row, column=15, value="Pago")  # Coluna O - Status
                        
                        # Atualizar data de pagamento
                        ws_dados.cell(row=row, column=16, value=data_pagamento)  # Coluna P - Data Pagamento
                        
                        atualizar_dados = True
                        break
            
            # Salvar alterações
            print("Salvando alterações na planilha")
            wb.save(self.arquivo_cliente)
            
            mensagem = "Pagamento registrado com sucesso!"
            if atualizar_dados:
                mensagem += " A aba 'Dados' também foi atualizada."
                
            messagebox.showinfo("Sucesso", mensagem)
            
            # Atualizar lista de pagamentos
            self.filtrar_pagamentos(None)
            
        except Exception as e:
            print(f"Erro ao registrar pagamento: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao registrar pagamento: {str(e)}")
            if 'wb' in locals():
                try:
                    wb.close()
                except:
                    pass

    def gerar_lancamento_pagamento(self):
        """Gera um lançamento para o sistema de entrada de dados (versão corrigida)"""
        try:
            print("Iniciando geração de lançamento")
            selecionado = self.tree_pagamentos.selection()
            if not selecionado:
                messagebox.showwarning("Aviso", "Selecione um pagamento primeiro")
                return
                
            valores = self.tree_pagamentos.item(selecionado)['values']
            
            print(f"Valores do pagamento selecionado: {valores}")
            
            num_contrato = valores[0]
            evento_id = valores[1]
            cnpj_cpf = valores[2]
            nome = valores[3]
            data_vencimento = valores[4]
            
            # Extrair e converter o valor com tratamento de diferentes formatos
            valor_str = str(valores[5])
            print(f"Valor original: {valor_str}")
            
            # Remover caracteres não numéricos (exceto . e ,)
            valor_limpo = ""
            for char in valor_str:
                if char.isdigit() or char == '.' or char == ',':
                    valor_limpo += char
                    
            # Substituir R$ e outros caracteres
            valor_limpo = valor_limpo.replace('R$', '').replace(' ', '').strip()
            print(f"Valor limpo: {valor_limpo}")
            
            # Converter para float
            try:
                # Tratar formatação brasileira (vírgula como decimal)
                if ',' in valor_limpo and '.' in valor_limpo:
                    # Formato brasileiro com separador de milhar
                    valor_limpo = valor_limpo.replace('.', '')
                    valor_limpo = valor_limpo.replace(',', '.')
                elif ',' in valor_limpo:
                    # Vírgula como separador decimal
                    valor_limpo = valor_limpo.replace(',', '.')
                    
                valor = float(valor_limpo)
                print(f"Valor convertido: {valor}")
            except ValueError as e:
                print(f"Erro ao converter valor: {e} - Valor: '{valor_limpo}'")
                messagebox.showerror("Erro", f"Valor do pagamento inválido: {valor_str}")
                return
                
            # Converter data
            try:
                data_vencto = datetime.strptime(data_vencimento, '%d/%m/%Y')
                print(f"Data de vencimento: {data_vencto}")
            except ValueError:
                print(f"Data de vencimento inválida: {data_vencimento}")
                messagebox.showerror("Erro", "Data de vencimento inválida")
                return
            
            # Buscar administrador para obter categoria
            categoria = "ADM"
            
            # Verificar se o Sistema de Entrada de Dados está disponível no parent
            if hasattr(self.parent, 'dados_para_incluir'):
                # Preparar lançamento
                lancamento = {
                    'data': self.calcular_data_relatorio(data_vencto),
                    'tp_desp': '3',  # Tipo 3 para administração
                    'cnpj_cpf': cnpj_cpf,
                    'nome': nome,
                    'referencia': f"ADM OBRA - EVENTO {evento_id} - CONTRATO {num_contrato}",
                    'nf': '',
                    'vr_unit': f"{valor:.2f}",
                    'dias': 1,
                    'valor': f"{valor:.2f}",
                    'dt_vencto': data_vencimento,
                    'categoria': categoria,
                    'dados_bancarios': self.buscar_dados_bancarios(cnpj_cpf),
                    'observacao': f"PAGAMENTO EVENTO {evento_id} - CONTRATO {num_contrato}",
                    'forma_pagamento': 'PIX'  # Valor padrão
                }
                
                # Adicionar à lista do sistema
                self.parent.dados_para_incluir.append(lancamento)
                
                print("Lançamento gerado com sucesso")
                messagebox.showinfo(
                    "Sucesso", 
                    "Lançamento gerado com sucesso! Acesse a aba de 'Visualização de Lançamentos' para conferir."
                )
                
                # Fechar esta janela
                self.janela.destroy()
                
                # Mostrar visualizador de lançamentos
                if hasattr(self.parent, 'visualizar_lancamentos'):
                    self.parent.visualizar_lancamentos()
            else:
                print("Sistema de Entrada de Dados não disponível")
                messagebox.showwarning(
                    "Aviso", 
                    "Sistema de Entrada de Dados não está disponível. O lançamento não pôde ser gerado."
                )
        except Exception as e:
            print(f"Erro ao gerar lançamento: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao gerar lançamento: {str(e)}")

    def buscar_dados_bancarios(self, cnpj_cpf):
        """Busca dados bancários do fornecedor"""
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == str(cnpj_cpf):
                    # Verificar se tem chave PIX
                    if row[10]:  # Coluna K - Chave PIX
                        dados = f"PIX: {row[10]}"
                    else:
                        # Construir dados baseados nas informações bancárias
                        partes = []
                        if row[6]: partes.append(str(row[6]))  # Banco
                        if row[7]: partes.append(str(row[7]))  # OP
                        if row[8]: partes.append(str(row[8]))  # Agência
                        if row[9]: partes.append(str(row[9]))  # Conta
                        if row[0]: partes.append(str(row[0]))  # CNPJ/CPF
                        
                        dados = ' - '.join(partes)
                    
                    wb.close()
                    return dados or "DADOS BANCÁRIOS NÃO CADASTRADOS"
                    
            wb.close()
            return "DADOS BANCÁRIOS NÃO CADASTRADOS"
            
        except Exception as e:
            print(f"Erro ao buscar dados bancários: {str(e)}")
            if 'wb' in locals():
                wb.close()
            return "ERRO AO BUSCAR DADOS BANCÁRIOS"

    def calcular_data_relatorio(self, data_vencimento):
        """Calcula a data do relatório com base na data de vencimento"""
        try:
            hoje = datetime.now()
            
            # Para as demais parcelas, manter a lógica existente
            if data_vencimento.day <= 5:
                # Se vence até dia 5, relatório é dia 20 do mês anterior
                data_rel = (data_vencimento - relativedelta(months=1)).replace(day=20)
            elif data_vencimento.day <= 20:
                # Se vence até dia 20, relatório é dia 5 do mesmo mês
                data_rel = data_vencimento.replace(day=5)
            else:
                # Se vence após dia 20, relatório é dia 20 do mesmo mês
                data_rel = data_vencimento.replace(day=20)
                
            # Garantir que a data do relatório não seja anterior à data atual
            if data_rel < hoje:
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                    
            return data_rel.strftime('%d/%m/%Y')
            
        except Exception as e:
            print(f"Erro ao calcular data do relatório: {str(e)}")
            return hoje.strftime('%d/%m/%Y')
        

    