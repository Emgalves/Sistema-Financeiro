import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import defaultdict
import numpy as np

# Adicionar diretório raiz ao path
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        PASTA_CLIENTES,
        BASE_PATH,
        ARQUIVO_FORNECEDORES
    )
    print("Configurações importadas com sucesso")
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Definir valores padrão em caso de falha
    BASE_PATH = Path(".")
    ARQUIVO_CLIENTES = BASE_PATH / "dados" / "clientes.xlsx"
    PASTA_CLIENTES = BASE_PATH / "dados" / "clientes"
    ARQUIVO_FORNECEDORES = BASE_PATH / "dados" / "fornecedores.xlsx"

# Importar o utils.py
from src.config.utils import atualizar_combobox_clientes, cliente_esta_ativo, obter_info_cliente

try:
    from src.config.window_config import configurar_janela
    print("window_config importado com sucesso")
except ImportError as e:
    print(f"Erro ao importar window_config: {str(e)}")
    def configurar_janela(janela, titulo="Janela", largura=800, altura=950):
        janela.title(titulo)
        janela.geometry(f"{largura}x{altura}")
        janela.resizable(True, True)
        
        # Centralizar na tela
        janela.update_idletasks()
        width = janela.winfo_width()
        height = janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (width // 2)
        y = (janela.winfo_screenheight() // 2) - (height // 2)
        janela.geometry(f'{width}x{height}+{x}+{y}')

# Função para formatação de moeda
def formatar_moeda_br(valor):
    """Formata um valor numérico como moeda brasileira"""
    try:
        valor_float = float(valor)
        return f"R$ {valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return f"R$ 0,00"

class RelatorioFornecedores:
    """Classe para geração de relatórios de principais fornecedores"""
    
    def __init__(self, parent=None):
        """Inicializa a interface do relatório de fornecedores - VERSÃO CORRIGIDA"""
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
            
            # Configurar protocolo de fechamento para janela filha
            self.root.protocol("WM_DELETE_WINDOW", self.voltar_menu)
            
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
            # Configurar protocolo de fechamento para janela principal
            self.root.protocol("WM_DELETE_WINDOW", self.fechar_aplicacao)
            
        configurar_janela(self.root, "Relatório de Fornecedores", 900, 1000)
        
        # Configuração de variáveis (resto permanece igual)
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.data_referencia = datetime.now()
        self.periodo_inicio = None
        self.periodo_fim = None
        self.todos_clientes = False
        self.fornecedor_especifico = None
        self.modo_relatorio = "fornecedores"
        self.dados_fornecedores = {}
        self.dados_por_fornecedor = {}
        self.top_fornecedores = []
        self.dados_carregados = False
        
        # Configurar interface
        self.setup_gui()
        
    def setup_gui(self):
        """Configuração da interface gráfica principal"""
        # Frame principal
        self.frame_principal = ttk.Frame(self.root, padding=10)
        self.frame_principal.pack(fill='both', expand=True)

        # Dividir o frame principal em três partes usando grid
        self.frame_principal.columnconfigure(0, weight=1)
        self.frame_principal.rowconfigure(0, weight=0)  # Seleção de modo
        self.frame_principal.rowconfigure(1, weight=0)  # Seleção
        self.frame_principal.rowconfigure(2, weight=1)  # Resultados
        self.frame_principal.rowconfigure(3, weight=0)  # Botões
        
        # NOVO: Frame para seleção do modo de relatório
        self.frame_modo = ttk.LabelFrame(self.frame_principal, text="Tipo de Relatório")
        self.frame_modo.grid(row=0, column=0, sticky='ew', pady=5)
        
        frame_modo_int = ttk.Frame(self.frame_modo)
        frame_modo_int.pack(fill='x', padx=10, pady=10)
        
        self.var_modo = tk.StringVar(value="fornecedores")
        
        ttk.Radiobutton(
            frame_modo_int,
            text="Principais Fornecedores (por cliente)",
            variable=self.var_modo,
            value="fornecedores",
            command=self.alterar_modo_relatorio
        ).pack(side='left', padx=20)
        
        ttk.Radiobutton(
            frame_modo_int,
            text="Clientes de um Fornecedor Específico",
            variable=self.var_modo,
            value="por_fornecedor",
            command=self.alterar_modo_relatorio
        ).pack(side='left', padx=20)
            
        # Frame para seleção
        self.frame_selecao = ttk.LabelFrame(self.frame_principal, text="Seleção de Cliente e Período")
        self.frame_selecao.grid(row=1, column=0, sticky='ew', pady=10)
        
        # Container para cliente
        self.frame_cliente = ttk.Frame(self.frame_selecao)
        self.frame_cliente.pack(fill='x', padx=10, pady=10)
        
        self.lbl_cliente = ttk.Label(self.frame_cliente, text="Selecione o Cliente:", font=('Arial', 11))
        self.lbl_cliente.pack(side='left', pady=5)
        
        self.cliente_combobox = ttk.Combobox(self.frame_cliente, width=40, font=('Arial', 11))
        self.cliente_combobox.pack(side='left', padx=5)
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)
        
        # Checkbox para todos os clientes
        self.var_todos_clientes = tk.BooleanVar(value=False)
        self.cb_todos_clientes = ttk.Checkbutton(
            self.frame_cliente, 
            text="Analisar todos os clientes",
            variable=self.var_todos_clientes,
            command=self.alternar_todos_clientes
        )
        self.cb_todos_clientes.pack(side='left', padx=20)
        
        self.setup_frame_fornecedor_melhorado()
        
        ttk.Button(
            self.frame_fornecedor,
            text="Buscar Fornecedores",
            command=self.carregar_fornecedores
        ).pack(side='left', padx=20)
        
        # Container para período
        frame_periodo = ttk.Frame(self.frame_selecao)
        frame_periodo.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_periodo, text="Período de Análise:", font=('Arial', 11)).pack(side='left', pady=5)
        
        # Combobox para seleção de período pré-definido
        self.periodo_combobox = ttk.Combobox(
            frame_periodo, 
            width=15, 
            font=('Arial', 11),
            values=[
                "Últimos 30 dias",
                "Últimos 90 dias",
                "Últimos 180 dias",
                "Último ano",
                "Todo o período",
                "Personalizado"
            ]
        )
        self.periodo_combobox.pack(side='left', padx=5)
        self.periodo_combobox.current(2)  # Padrão: Últimos 180 dias
        self.periodo_combobox.bind('<<ComboboxSelected>>', self.alterar_periodo)
        
        # Frame para datas personalizadas (inicialmente oculto)
        self.frame_datas_personalizadas = ttk.Frame(frame_periodo)
        
        ttk.Label(self.frame_datas_personalizadas, text="De:", font=('Arial', 11)).pack(side='left', padx=(20, 5))
        
        # Usar DateEntry se disponível
        try:
            from tkcalendar import DateEntry
            self.data_inicio = DateEntry(
                self.frame_datas_personalizadas, 
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd/mm/yyyy',
                locale='pt_BR',
                font=('Arial', 11)
            )
            self.data_inicio.pack(side='left', padx=5)
            
            ttk.Label(self.frame_datas_personalizadas, text="Até:", font=('Arial', 11)).pack(side='left', padx=(20, 5))
            
            self.data_fim = DateEntry(
                self.frame_datas_personalizadas, 
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd/mm/yyyy',
                locale='pt_BR',
                font=('Arial', 11)
            )
            self.data_fim.pack(side='left', padx=5)
            
        except ImportError:
            # Fallback para Entry
            self.data_inicio_var = tk.StringVar(value=(datetime.now() - timedelta(days=180)).strftime('%d/%m/%Y'))
            ttk.Entry(
                self.frame_datas_personalizadas,
                textvariable=self.data_inicio_var,
                width=12,
                font=('Arial', 11)
            ).pack(side='left', padx=5)
            
            ttk.Label(self.frame_datas_personalizadas, text="Até:", font=('Arial', 11)).pack(side='left', padx=(20, 5))
            
            self.data_fim_var = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
            ttk.Entry(
                self.frame_datas_personalizadas,
                textvariable=self.data_fim_var,
                width=12,
                font=('Arial', 11)
            ).pack(side='left', padx=5)
        
        # Botão de gerar relatório
        ttk.Button(
            frame_periodo,
            text="Gerar Relatório",
            command=self.gerar_relatorio,
            style='Big.TButton'
        ).pack(side='right', padx=20)
        
        # Configurar opções de análise
        self.frame_opcoes = ttk.LabelFrame(self.frame_principal, text="Opções de Análise")
        self.frame_opcoes.grid(row=2, column=0, sticky='ew', pady=10)
        
        frame_opcoes_int = ttk.Frame(self.frame_opcoes)
        frame_opcoes_int.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_opcoes_int, text="Quantidade a exibir:", font=('Arial', 11)).pack(side='left', pady=5)
        
        self.top_n_var = tk.StringVar(value="10")
        self.top_n_combobox = ttk.Combobox(
            frame_opcoes_int, 
            width=5, 
            font=('Arial', 11),
            textvariable=self.top_n_var,
            values=["5", "10", "15", "20", "25", "30", "50", "100"]
        )
        self.top_n_combobox.pack(side='left', padx=5)
        
        # Checkbox para mostrar fornecedores agrupados por tipo de despesa
        self.var_agrupar_tipo = tk.BooleanVar(value=True)
        self.cb_agrupar_tipo = ttk.Checkbutton(
            frame_opcoes_int, 
            text="Agrupar por tipo de despesa",
            variable=self.var_agrupar_tipo
        )
        self.cb_agrupar_tipo.pack(side='left', padx=20)
        
        # Frame para resultados - com notebook para separar visões
        self.frame_resultados = ttk.LabelFrame(self.frame_principal, text="Resultados")
        self.frame_resultados.grid(row=2, column=0, sticky='nsew', pady=10)
        
        # Notebook (abas)
        self.notebook = ttk.Notebook(self.frame_resultados)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Abas
        self.aba_resumo = ttk.Frame(self.notebook)
        self.aba_detalhes = ttk.Frame(self.notebook)
        self.aba_grafico = ttk.Frame(self.notebook)
        
        self.notebook.add(self.aba_resumo, text='Resumo')
        self.notebook.add(self.aba_detalhes, text='Detalhes')
        self.notebook.add(self.aba_grafico, text='Gráfico')
        
        # Configurar cada aba
        self.setup_aba_resumo()
        self.setup_aba_detalhes()
        self.setup_aba_grafico()
        
        # Botões na parte inferior
        frame_botoes = ttk.Frame(self.frame_principal)
        frame_botoes.grid(row=3, column=0, sticky='ew', pady=10)

        btn_excel = ttk.Button(
            frame_botoes,
            text="Exportar para Excel",
            command=self.exportar_excel
        )
        btn_excel.grid(row=0, column=0, padx=5, sticky='w')

        btn_pdf = ttk.Button(
            frame_botoes,
            text="Exportar para PDF",
            command=self.exportar_pdf
        )
        btn_pdf.grid(row=0, column=1, padx=5, sticky='w')

        btn_voltar = ttk.Button(
            frame_botoes,
            text="Voltar ao Menu",
            command=self.voltar_menu
        )
        btn_voltar.grid(row=0, column=2, padx=5, sticky='e')

        # Configure as colunas para posicionar corretamente
        frame_botoes.columnconfigure(0, weight=0)
        frame_botoes.columnconfigure(1, weight=0)
        frame_botoes.columnconfigure(2, weight=1)
        
        # Estilo para botões grandes
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 11, 'bold'), padding=(10, 5))
        
        # Carregar lista de clientes
        self.atualizar_lista_clientes()
        
        # Configurar período inicial
        self.alterar_periodo()
        
        # Configurar modo inicial
        self.alterar_modo_relatorio()

    def setup_frame_fornecedor_melhorado(self):
        """Configura o frame de seleção de fornecedor com busca aprimorada"""
        # Container principal para fornecedor
        self.frame_fornecedor = ttk.Frame(self.frame_selecao)
        
        # Primeira linha: Label e campo de busca
        frame_busca = ttk.Frame(self.frame_fornecedor)
        frame_busca.pack(fill='x', pady=5)
        
        self.lbl_fornecedor = ttk.Label(frame_busca, text="Buscar Fornecedor:", font=('Arial', 11))
        self.lbl_fornecedor.pack(side='left', pady=5)
        
        # Campo de busca com autocomplete
        self.fornecedor_busca_var = tk.StringVar()
        self.fornecedor_busca_entry = ttk.Entry(
            frame_busca, 
            textvariable=self.fornecedor_busca_var,
            width=50, 
            font=('Arial', 11)
        )
        self.fornecedor_busca_entry.pack(side='left', padx=5)
        self.fornecedor_busca_entry.bind('<KeyRelease>', self.filtrar_fornecedores)
        self.fornecedor_busca_entry.bind('<Return>', self.selecionar_primeiro_resultado)
        
        # Botão para limpar busca
        ttk.Button(
            frame_busca,
            text="Limpar",
            command=self.limpar_busca_fornecedor,
            width=8
        ).pack(side='left', padx=5)
        
        # Botão para buscar fornecedores
        ttk.Button(
            frame_busca,
            text="Buscar Fornecedores",
            command=self.carregar_fornecedores
        ).pack(side='left', padx=20)
        
        # Segunda linha: Lista de resultados
        frame_resultados = ttk.Frame(self.frame_fornecedor)
        frame_resultados.pack(fill='both', expand=True, pady=5)
        
        # Label para status
        self.lbl_status_busca = ttk.Label(
            frame_resultados, 
            text="Use o botão 'Buscar Fornecedores' para carregar a lista",
            font=('Arial', 10),
            foreground='gray'
        )
        self.lbl_status_busca.pack(pady=5)
        
        # Treeview para mostrar resultados da busca
        self.tree_fornecedores = ttk.Treeview(
            frame_resultados, 
            columns=('nome',), 
            show='headings', 
            height=8
        )
        self.tree_fornecedores.heading('nome', text='Fornecedores Encontrados')
        self.tree_fornecedores.column('nome', width=400)
        
        # Scrollbar para a lista
        scrollbar_fornecedores = ttk.Scrollbar(
            frame_resultados, 
            orient='vertical', 
            command=self.tree_fornecedores.yview
        )
        self.tree_fornecedores.configure(yscrollcommand=scrollbar_fornecedores.set)
        
        # Eventos para seleção
        self.tree_fornecedores.bind('<Double-1>', self.selecionar_fornecedor_da_lista)
        self.tree_fornecedores.bind('<Return>', self.selecionar_fornecedor_da_lista)
        
        # Inicialmente ocultos
        self.tree_fornecedores.pack_forget()
        scrollbar_fornecedores.pack_forget()
        
        # Variáveis para controle
        self.fornecedores_disponiveis = []
        self.fornecedores_filtrados = []

    # NOVO: Método para alternar modo de relatório
    def alterar_modo_relatorio(self):
        """Alterna entre os modos de relatório com limpeza completa"""
        # Limpar todos os dados antes de alternar
        self.limpar_dados_completo()
        
        self.modo_relatorio = self.var_modo.get()
        
        if self.modo_relatorio == "fornecedores":
            # Modo original: principais fornecedores
            self.frame_cliente.pack(fill='x', padx=10, pady=10)
            self.frame_fornecedor.pack_forget()
            
            # Atualizar labels das abas
            self.notebook.tab(0, text="Resumo - Principais Fornecedores")
            self.notebook.tab(1, text="Detalhes do Fornecedor")
            self.lbl_cliente.config(text="Selecione o Cliente:")
            self.frame_opcoes.config(text="Opções de Análise - Fornecedores")
            
            # Atualizar labels específicos do modo
            self.lbl_cliente_resumo.config(text="Cliente: Nenhum selecionado")
            self.cb_todos_clientes.config(text="Analisar todos os clientes")
            
        else:
            # Novo modo: clientes de um fornecedor
            self.frame_cliente.pack_forget()
            self.frame_fornecedor.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Atualizar labels das abas
            self.notebook.tab(0, text="Resumo - Clientes do Fornecedor")
            self.notebook.tab(1, text="Detalhes do Cliente")
            self.frame_opcoes.config(text="Opções de Análise - Por Fornecedor")
            
            # Limpar seleção de fornecedor se necessário
            if hasattr(self, 'fornecedor_busca_var'):
                self.limpar_busca_fornecedor()

    def limpar_dados_completo(self):
        """Limpa completamente todos os dados e interface"""
        try:
            # Resetar variáveis de controle
            self.dados_carregados = False
            self.cliente_atual = None
            self.fornecedor_especifico = None
            self.todos_clientes = False
            
            # Limpar dados dos relatórios
            self.dados_fornecedores = {}
            self.dados_por_fornecedor = {}
            self.top_fornecedores = []
            self.total_geral = 0.0
            
            # Limpar seleções
            self.cliente_combobox.set("")
            if hasattr(self, 'fornecedor_busca_var'):
                self.fornecedor_busca_var.set("")
            
            # Resetar checkbox de todos os clientes
            self.var_todos_clientes.set(False)
            self.cliente_combobox.config(state='normal')
            
            # Limpar todas as abas
            self.limpar_aba_resumo()
            self.limpar_aba_detalhes()
            self.limpar_aba_grafico()
            
            # Resetar labels informativos
            self.lbl_periodo_resumo.config(text="Período: Não definido")
            
            print("Dados limpos completamente ao alternar modo de relatório")
            
        except Exception as e:
            print(f"Erro ao limpar dados: {str(e)}")

    # NOVO MÉTODO: Limpar aba de resumo
    def limpar_aba_resumo(self):
        """Limpa completamente a aba de resumo"""
        try:
            # Limpar tabela de resumo
            for item in self.tree_resumo.get_children():
                self.tree_resumo.delete(item)
            
            # Resetar totais
            self.lbl_total_geral.config(text="Total Geral: R$ 0,00")
            self.lbl_total_apresentado.config(text="Total Apresentado: R$ 0,00 (0%)")
            
        except Exception as e:
            print(f"Erro ao limpar aba de resumo: {str(e)}")

    # NOVO MÉTODO: Limpar aba de detalhes
    def limpar_aba_detalhes(self):
        """Limpa completamente a aba de detalhes"""
        try:
            # Limpar combobox de seleção
            self.item_detalhes_combobox['values'] = ()
            self.item_detalhes_combobox.set('')
            
            # Limpar informações do item
            self.lbl_nome_item.config(text="")
            self.lbl_total_item.config(text="")
            self.lbl_qtd_lancamentos_item.config(text="")
            self.lbl_media_lancamento_item.config(text="")
            self.lbl_tipos_despesa_item.config(text="")
            
            # Limpar tabela de lançamentos
            for item in self.tree_lancamentos.get_children():
                self.tree_lancamentos.delete(item)
                
        except Exception as e:
            print(f"Erro ao limpar aba de detalhes: {str(e)}")

    # NOVO: Método para carregar lista de fornecedores
    def carregar_fornecedores(self):
        """Carrega a lista de fornecedores encontrados nos arquivos dos clientes"""
        try:
            fornecedores_encontrados = set()
            clientes_processados = 0
            
            # Processar todos os arquivos de clientes para encontrar fornecedores
            if not os.path.exists(PASTA_CLIENTES):
                messagebox.showwarning("Aviso", "Pasta de clientes não encontrada!")
                return
            
            # Mostrar progresso
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Carregando Fornecedores...")
            progress_window.geometry("400x100")
            progress_window.resizable(False, False)
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Centralizar janela de progresso
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
            y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
            progress_window.geometry(f"+{x}+{y}")
            
            progress_label = ttk.Label(progress_window, text="Processando arquivos de clientes...")
            progress_label.pack(pady=20)
            
            progress_bar = ttk.Progressbar(progress_window, mode='indeterminate')
            progress_bar.pack(pady=10, padx=20, fill='x')
            progress_bar.start()
            
            # Forçar atualização da interface
            self.root.update()
            
            for arquivo in os.listdir(PASTA_CLIENTES):
                if arquivo.endswith('.xlsx'):
                    try:
                        caminho_arquivo = os.path.join(PASTA_CLIENTES, arquivo)
                        nome_cliente = os.path.splitext(arquivo)[0]
                        
                        # Atualizar label de progresso
                        progress_label.config(text=f"Processando: {nome_cliente}")
                        self.root.update()
                        
                        # Carregar dados do Excel
                        df = pd.read_excel(caminho_arquivo, sheet_name='Dados')
                        
                        # Verificar se existe coluna NOME
                        if 'NOME' not in df.columns:
                            continue
                        
                        # Filtrar apenas lançamentos ativos se a coluna STATUS existir
                        if 'STATUS' in df.columns:
                            df = df[df['STATUS'].str.upper().str.strip() == 'ATIVO'].copy()
                        
                        # Obter fornecedores únicos (removendo valores nulos e vazios)
                        fornecedores_cliente = df['NOME'].dropna().str.strip()
                        fornecedores_cliente = fornecedores_cliente[fornecedores_cliente != '']
                        
                        # Adicionar ao conjunto de fornecedores
                        fornecedores_encontrados.update(fornecedores_cliente.unique())
                        clientes_processados += 1
                        
                    except Exception as e:
                        print(f"Erro ao processar arquivo {arquivo}: {str(e)}")
                        continue
            
            # Fechar janela de progresso
            progress_window.destroy()
            
            if not fornecedores_encontrados:
                messagebox.showwarning("Aviso", "Nenhum fornecedor encontrado nos arquivos dos clientes!")
                return
            
            # Ordenar fornecedores alfabeticamente e armazenar
            self.fornecedores_disponiveis = sorted(list(fornecedores_encontrados))
            
            # Limpar busca atual
            self.limpar_busca_fornecedor()
            
            # Atualizar status
            self.lbl_status_busca.config(
                text=f"{len(self.fornecedores_disponiveis)} fornecedores carregados. Digite para buscar.",
                foreground='blue'
            )
            
            messagebox.showinfo(
                "Sucesso", 
                f"{len(self.fornecedores_disponiveis)} fornecedores encontrados em {clientes_processados} clientes!\n\n" +
                "Use o campo de busca para encontrar fornecedores específicos."
            )
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar fornecedores: {str(e)}")
            import traceback
            traceback.print_exc()

    def filtrar_fornecedores(self, event=None):
        """Filtra a lista de fornecedores conforme o usuário digita"""
        if not self.fornecedores_disponiveis:
            return
        
        termo_busca = self.fornecedor_busca_var.get().upper().strip()
        
        # Limpar lista atual
        for item in self.tree_fornecedores.get_children():
            self.tree_fornecedores.delete(item)
        
        if not termo_busca:
            # Se não há termo de busca, mostrar todos (limitado)
            self.fornecedores_filtrados = self.fornecedores_disponiveis[:100]  # Limitar a 100
            self.lbl_status_busca.config(
                text=f"Mostrando primeiros 100 de {len(self.fornecedores_disponiveis)} fornecedores"
            )
        else:
            # Filtrar fornecedores que contêm o termo
            self.fornecedores_filtrados = [
                f for f in self.fornecedores_disponiveis 
                if termo_busca in f.upper()
            ]
            
            # Limitar resultados para performance
            if len(self.fornecedores_filtrados) > 50:
                self.fornecedores_filtrados = self.fornecedores_filtrados[:50]
                self.lbl_status_busca.config(
                    text=f"Mostrando primeiros 50 resultados (digite mais para refinar a busca)"
                )
            else:
                self.lbl_status_busca.config(
                    text=f"{len(self.fornecedores_filtrados)} fornecedores encontrados"
                )
        
        # Adicionar resultados ao Treeview
        for fornecedor in self.fornecedores_filtrados:
            self.tree_fornecedores.insert('', 'end', values=(fornecedor,))
        
        # Mostrar a lista se há resultados
        if self.fornecedores_filtrados:
            self.tree_fornecedores.pack(side='left', fill='both', expand=True)
            # Note: scrollbar seria necessária aqui, mas simplificando
        else:
            self.tree_fornecedores.pack_forget()
            if termo_busca:
                self.lbl_status_busca.config(text="Nenhum fornecedor encontrado")

    def selecionar_primeiro_resultado(self, event=None):
        """Seleciona o primeiro resultado da lista ao pressionar Enter"""
        children = self.tree_fornecedores.get_children()
        if children:
            # Selecionar primeiro item
            self.tree_fornecedores.selection_set(children[0])
            self.tree_fornecedores.focus(children[0])
            self.selecionar_fornecedor_da_lista()

    def selecionar_fornecedor_da_lista(self, event=None):
        """Seleciona o fornecedor clicado na lista"""
        selecionado = self.tree_fornecedores.selection()
        if selecionado:
            item = self.tree_fornecedores.item(selecionado[0])
            fornecedor_selecionado = item['values'][0]
            
            # Atualizar campo de busca e variável
            self.fornecedor_busca_var.set(fornecedor_selecionado)
            self.fornecedor_especifico = fornecedor_selecionado
            
            # Ocultar lista após seleção
            self.tree_fornecedores.pack_forget()
            
            # Atualizar status
            self.lbl_status_busca.config(
                text=f"Fornecedor selecionado: {fornecedor_selecionado}",
                foreground='green'
            )

    def limpar_busca_fornecedor(self):
        """Limpa a busca de fornecedor"""
        self.fornecedor_busca_var.set('')
        self.fornecedor_especifico = None
        self.tree_fornecedores.pack_forget()
        
        if self.fornecedores_disponiveis:
            self.lbl_status_busca.config(
                text=f"{len(self.fornecedores_disponiveis)} fornecedores disponíveis. Digite para buscar.",
                foreground='gray'
            )
        else:
            self.lbl_status_busca.config(
                text="Use o botão 'Buscar Fornecedores' para carregar a lista",
                foreground='gray'
            )
               
    # NOVO: Método para selecionar fornecedor específico
    def selecionar_fornecedor_especifico(self, event=None):
        """Seleciona o fornecedor para análise"""
        self.fornecedor_especifico = self.fornecedor_especifico_combobox.get()

    def setup_aba_resumo(self):
        """Configura a aba de resumo do relatório"""
        # Frame para informações do cliente/período
        frame_info = ttk.Frame(self.aba_resumo, padding=5)
        frame_info.pack(fill='x', pady=5)
        
        self.lbl_cliente_resumo = ttk.Label(
            frame_info, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_resumo.pack(side='left', padx=10)
        
        self.lbl_periodo_resumo = ttk.Label(
            frame_info, 
            text="Período: Não definido", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_periodo_resumo.pack(side='left', padx=10)
        
        # Frame para tabela de resumo
        frame_tabela = ttk.Frame(self.aba_resumo, padding=5)
        frame_tabela.pack(fill='both', expand=True, pady=5)
        
        # Criar Treeview para listar (fornecedores ou clientes)
        self.colunas_resumo = ('posicao', 'nome', 'total_gasto', 'percentual', 'qtd_lancamentos', 'tipos_despesa')
        self.tree_resumo = ttk.Treeview(frame_tabela, columns=self.colunas_resumo, show='headings', height=20)
        
        # Configurar colunas (serão atualizadas conforme o modo)
        self.atualizar_colunas_resumo()
        
        self.tree_resumo.bind("<Double-1>", self.mostrar_detalhes_por_clique)
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_tabela, orient='vertical', command=self.tree_resumo.yview)
        scrollx = ttk.Scrollbar(frame_tabela, orient='horizontal', command=self.tree_resumo.xview)
        self.tree_resumo.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionamento
        self.tree_resumo.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Frame para totais
        frame_total = ttk.Frame(self.aba_resumo, padding=5)
        frame_total.pack(fill='x', pady=5)
        
        self.lbl_total_geral = ttk.Label(
            frame_total,
            text="Total Geral: R$ 0,00",
            font=('Arial', 12, 'bold')
        )
        self.lbl_total_geral.pack(side='left', padx=10)
        
        self.lbl_total_apresentado = ttk.Label(
            frame_total,
            text="Total Apresentado: R$ 0,00 (0%)",
            font=('Arial', 12)
        )
        self.lbl_total_apresentado.pack(side='left', padx=10)

    # NOVO: Método para atualizar colunas conforme o modo
    def atualizar_colunas_resumo(self):
        """Atualiza as colunas da tabela de resumo conforme o modo"""
        if self.modo_relatorio == "fornecedores":
            # Modo original
            self.tree_resumo.heading('posicao', text='#')
            self.tree_resumo.heading('nome', text='Fornecedor')
            self.tree_resumo.heading('total_gasto', text='Total Gasto')
            self.tree_resumo.heading('percentual', text='% do Total')
            self.tree_resumo.heading('qtd_lancamentos', text='Qtd. Lançamentos')
            self.tree_resumo.heading('tipos_despesa', text='Tipos de Despesa')
        else:
            # Novo modo: clientes do fornecedor
            self.tree_resumo.heading('posicao', text='#')
            self.tree_resumo.heading('nome', text='Cliente')
            self.tree_resumo.heading('total_gasto', text='Total Gasto')
            self.tree_resumo.heading('percentual', text='% do Total')
            self.tree_resumo.heading('qtd_lancamentos', text='Qtd. Lançamentos')
            self.tree_resumo.heading('tipos_despesa', text='Tipos de Despesa')
        
        # Definir larguras
        self.tree_resumo.column('posicao', width=50, anchor='center')
        self.tree_resumo.column('nome', width=250)
        self.tree_resumo.column('total_gasto', width=150, anchor='e')
        self.tree_resumo.column('percentual', width=100, anchor='center')
        self.tree_resumo.column('qtd_lancamentos', width=150, anchor='center')
        self.tree_resumo.column('tipos_despesa', width=150)

    def setup_aba_detalhes(self):
        """Configura a aba de detalhes do relatório"""
        # Frame para seleção
        frame_selecao = ttk.Frame(self.aba_detalhes, padding=5)
        frame_selecao.pack(fill='x', pady=5)
        
        self.lbl_selecao_detalhes = ttk.Label(frame_selecao, text="Selecione o Item:", font=('Arial', 11))
        self.lbl_selecao_detalhes.pack(side='left', pady=5)
        
        self.item_detalhes_combobox = ttk.Combobox(frame_selecao, width=40, font=('Arial', 11))
        self.item_detalhes_combobox.pack(side='left', padx=5)
        self.item_detalhes_combobox.bind('<<ComboboxSelected>>', self.carregar_detalhes_item)
        
        # Frame para informações do item
        frame_info_item = ttk.LabelFrame(self.aba_detalhes, text="Informações")
        frame_info_item.pack(fill='x', pady=5, padx=5)
        
        # Grid para informações
        frame_grid = ttk.Frame(frame_info_item, padding=10)
        frame_grid.pack(fill='x')
        
        # Primeira linha
        self.lbl_titulo_nome = ttk.Label(frame_grid, text="Nome:", font=('Arial', 10, 'bold'))
        self.lbl_titulo_nome.grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.lbl_nome_item = ttk.Label(frame_grid, text="", font=('Arial', 10))
        self.lbl_nome_item.grid(row=0, column=1, sticky='w', padx=5, pady=2)
        
        ttk.Label(frame_grid, text="Total Gasto:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.lbl_total_item = ttk.Label(frame_grid, text="", font=('Arial', 10))
        self.lbl_total_item.grid(row=0, column=3, sticky='w', padx=5, pady=2)
        
        # Segunda linha
        ttk.Label(frame_grid, text="Quantidade de Lançamentos:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.lbl_qtd_lancamentos_item = ttk.Label(frame_grid, text="", font=('Arial', 10))
        self.lbl_qtd_lancamentos_item.grid(row=1, column=1, sticky='w', padx=5, pady=2)
        
        ttk.Label(frame_grid, text="Média por Lançamento:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky='w', padx=5, pady=2)
        self.lbl_media_lancamento_item = ttk.Label(frame_grid, text="", font=('Arial', 10))
        self.lbl_media_lancamento_item.grid(row=1, column=3, sticky='w', padx=5, pady=2)
        
        # Terceira linha
        ttk.Label(frame_grid, text="Tipos de Despesa:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.lbl_tipos_despesa_item = ttk.Label(frame_grid, text="", font=('Arial', 10))
        self.lbl_tipos_despesa_item.grid(row=2, column=1, columnspan=3, sticky='w', padx=5, pady=2)
        
        # Frame para tabela de lançamentos
        frame_lancamentos = ttk.LabelFrame(self.aba_detalhes, text="Lançamentos")
        frame_lancamentos.pack(fill='both', expand=True, pady=5, padx=5)
        
        # Tree para lançamentos
        colunas = ('data', 'cliente_ou_fornecedor', 'tipo_despesa', 'referencia', 'nf', 'dt_vencto', 'valor', 'observacao')
        self.tree_lancamentos = ttk.Treeview(frame_lancamentos, columns=colunas, show='headings', height=15)
        
        # Configurar colunas (serão atualizadas conforme o modo)
        self.atualizar_colunas_detalhes()
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_lancamentos, orient='vertical', command=self.tree_lancamentos.yview)
        scrollx = ttk.Scrollbar(frame_lancamentos, orient='horizontal', command=self.tree_lancamentos.xview)
        self.tree_lancamentos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionamento
        self.tree_lancamentos.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')

    # NOVO: Método para atualizar colunas da aba de detalhes
    def atualizar_colunas_detalhes(self):
        """Atualiza as colunas da tabela de detalhes conforme o modo"""
        if self.modo_relatorio == "fornecedores":
            # Modo original: detalhes de um fornecedor
            self.tree_lancamentos.heading('data', text='Data')
            self.tree_lancamentos.heading('cliente_ou_fornecedor', text='Cliente')
            self.tree_lancamentos.heading('tipo_despesa', text='Tipo')
            self.tree_lancamentos.heading('referencia', text='Referência')
            self.tree_lancamentos.heading('nf', text='NF')
            self.tree_lancamentos.heading('dt_vencto', text='Vencimento')
            self.tree_lancamentos.heading('valor', text='Valor')
            self.tree_lancamentos.heading('observacao', text='Observação')
        else:
            # Novo modo: detalhes de um cliente (compras de um fornecedor)
            self.tree_lancamentos.heading('data', text='Data')
            self.tree_lancamentos.heading('cliente_ou_fornecedor', text='Fornecedor')
            self.tree_lancamentos.heading('tipo_despesa', text='Tipo')
            self.tree_lancamentos.heading('referencia', text='Referência')
            self.tree_lancamentos.heading('nf', text='NF')
            self.tree_lancamentos.heading('dt_vencto', text='Vencimento')
            self.tree_lancamentos.heading('valor', text='Valor')
            self.tree_lancamentos.heading('observacao', text='Observação')
        
        # Ajustar larguras
        self.tree_lancamentos.column('data', width=80, anchor='center')
        self.tree_lancamentos.column('cliente_ou_fornecedor', width=200)
        self.tree_lancamentos.column('tipo_despesa', width=50, anchor='center')
        self.tree_lancamentos.column('referencia', width=250)
        self.tree_lancamentos.column('nf', width=100)
        self.tree_lancamentos.column('dt_vencto', width=80, anchor='center')
        self.tree_lancamentos.column('valor', width=100, anchor='e')
        self.tree_lancamentos.column('observacao', width=150)

    def setup_aba_grafico(self):
        """Configura a aba de gráficos"""
        # Frame para controles do gráfico
        frame_controles = ttk.Frame(self.aba_grafico, padding=5)
        frame_controles.pack(fill='x', pady=5)
        
        ttk.Label(frame_controles, text="Tipo de Gráfico:").pack(side='left', padx=5)
        self.combo_tipo_grafico = ttk.Combobox(frame_controles, values=[
            "Pizza - Total por Item",
            "Barras - Top Items",
            "Linhas - Evolução Mensal",
            "Barras Empilhadas - Por Tipo de Despesa"
        ], state='readonly', width=30)
        self.combo_tipo_grafico.pack(side='left', padx=5)
        self.combo_tipo_grafico.current(0)
        
        ttk.Button(frame_controles, text="Atualizar Gráfico", command=self.atualizar_grafico).pack(side='left', padx=20)
        
        # Frame para o gráfico
        self.frame_grafico = ttk.Frame(self.aba_grafico)
        self.frame_grafico.pack(fill='both', expand=True, pady=5)

    def mostrar_detalhes_por_clique(self, event):
        """Abre a aba de detalhes quando o usuário dá duplo clique em um item"""
        # Obter o item selecionado
        item = self.tree_resumo.identify('item', event.x, event.y)
        if not item:
            return
            
        # Obter a posição do item (valor da primeira coluna)
        posicao = self.tree_resumo.item(item, 'values')[0]
        
        # Selecionar o item correspondente no combobox
        if self.item_detalhes_combobox['values']:
            # Os valores do combobox começam com a posição (ex: "1. Nome do Item")
            for i, valor in enumerate(self.item_detalhes_combobox['values']):
                if valor.startswith(f"{posicao}. "):
                    self.item_detalhes_combobox.current(i)
                    self.carregar_detalhes_item()
                    # Mudar para a aba de detalhes
                    self.notebook.select(1)  # Índice 1 = aba de detalhes
                    break

    def alterar_periodo(self, event=None):
        """Atualiza o período de análise com base na seleção"""
        periodo_selecionado = self.periodo_combobox.get()
        
        # Ocultar frame de datas personalizadas por padrão
        self.frame_datas_personalizadas.pack_forget()
        
        # Data de hoje
        hoje = datetime.now()
        
        # Configurar período com base na seleção
        if periodo_selecionado == "Últimos 30 dias":
            self.periodo_inicio = hoje - timedelta(days=30)
            self.periodo_fim = hoje
        elif periodo_selecionado == "Últimos 90 dias":
            self.periodo_inicio = hoje - timedelta(days=90)
            self.periodo_fim = hoje
        elif periodo_selecionado == "Últimos 180 dias":
            self.periodo_inicio = hoje - timedelta(days=180)
            self.periodo_fim = hoje
        elif periodo_selecionado == "Último ano":
            self.periodo_inicio = hoje - timedelta(days=365)
            self.periodo_fim = hoje
        elif periodo_selecionado == "Todo o período":
            # Usar um período bem amplo
            self.periodo_inicio = datetime(2000, 1, 1)
            self.periodo_fim = hoje
        elif periodo_selecionado == "Personalizado":
            # Mostrar frame de datas personalizadas
            self.frame_datas_personalizadas.pack(side='left')
            
            # Configurar valores iniciais se estiver usando DateEntry
            if hasattr(self, 'data_inicio'):
                data_inicio = hoje - timedelta(days=180)
                self.data_inicio.set_date(data_inicio)
                self.data_fim.set_date(hoje)
            
            # Datas serão definidas quando o relatório for gerado
            return
        
        # Atualizar label do período (se já tiver sido gerado algum relatório)
        if hasattr(self, 'lbl_periodo_resumo'):
            self.lbl_periodo_resumo.config(
                text=f"Período: {self.periodo_inicio.strftime('%d/%m/%Y')} a {self.periodo_fim.strftime('%d/%m/%Y')}"
            )
    
    def alternar_todos_clientes(self):
        """Alterna entre análise de um cliente específico ou todos"""
        todos = self.var_todos_clientes.get()
        
        if todos:
            self.cliente_combobox.set("TODOS OS CLIENTES")
            self.cliente_combobox.config(state='disabled')
            self.cliente_atual = None
            self.todos_clientes = True
            
            # Atualizar label
            if hasattr(self, 'lbl_cliente_resumo'):
                self.lbl_cliente_resumo.config(text="Cliente: Todos os Clientes")
        else:
            self.cliente_combobox.config(state='normal')
            self.cliente_combobox.set("")
            self.cliente_atual = None
            self.todos_clientes = False
            
            # Atualizar label
            if hasattr(self, 'lbl_cliente_resumo'):
                self.lbl_cliente_resumo.config(text="Cliente: Nenhum selecionado")
    
    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes no combobox usando a função centralizada"""
        try:
            # Usar a função centralizada (apenas clientes ativos)
            self.info_clientes = atualizar_combobox_clientes(self.cliente_combobox, mostrar_inativos=False)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")

    def selecionar_cliente(self, event=None):
        """Atualiza o cliente selecionado"""
        self.cliente_atual = self.cliente_combobox.get()
        
        if self.cliente_atual:
            # Verificar se o cliente está ativo (extra proteção)
            if not cliente_esta_ativo(self.cliente_atual):
                messagebox.showwarning(
                    "Cliente Inativo", 
                    f"O cliente '{self.cliente_atual}' está inativo (contrato finalizado). " +
                    "Os dados serão mostrados somente para consulta."
                )
            
            # Obter informações do cliente
            info_cliente = obter_info_cliente(self.cliente_atual)
            
            # Atualizar label
            if hasattr(self, 'lbl_cliente_resumo'):
                texto_cliente = f"Cliente: {self.cliente_atual}"
                if info_cliente and not info_cliente['ativo']:
                    texto_cliente += " (INATIVO)"
                self.lbl_cliente_resumo.config(text=texto_cliente)
            
            # Definir o caminho do arquivo
            if info_cliente and 'arquivo' in info_cliente:
                self.arquivo_cliente = info_cliente['arquivo']
            else:
                self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
    
    def gerar_relatorio(self):
        """Gera o relatório com validação melhorada"""
        # Validações específicas por modo
        if self.modo_relatorio == "fornecedores":
            if not self.cliente_atual and not self.todos_clientes:
                messagebox.showwarning(
                    "Aviso", 
                    "Para o relatório de 'Principais Fornecedores':\n" +
                    "• Selecione um cliente específico, OU\n" +
                    "• Marque a opção 'Analisar todos os clientes'"
                )
                return
        else:
            if not self.fornecedor_especifico:
                messagebox.showwarning(
                    "Aviso", 
                    "Para o relatório 'Clientes de um Fornecedor':\n" +
                    "• Use o botão 'Buscar Fornecedores' para carregar a lista\n" +
                    "• Digite parte do nome do fornecedor no campo de busca\n" +
                    "• Selecione o fornecedor desejado na lista"
                )
                return
        
        # Limpar dados anteriores antes de gerar novo relatório
        self.limpar_dados_relatorio_anterior()
        
        # Se o período for personalizado, obter as datas selecionadas
        if self.periodo_combobox.get() == "Personalizado":
            try:
                if hasattr(self, 'data_inicio'):  # Usando DateEntry
                    self.periodo_inicio = datetime.strptime(self.data_inicio.get(), '%d/%m/%Y')
                    self.periodo_fim = datetime.strptime(self.data_fim.get(), '%d/%m/%Y')
                else:  # Usando Entry
                    self.periodo_inicio = datetime.strptime(self.data_inicio_var.get(), '%d/%m/%Y')
                    self.periodo_fim = datetime.strptime(self.data_fim_var.get(), '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Erro", "Data inválida no período personalizado!")
                return
        
        # Atualizar label do período
        self.lbl_periodo_resumo.config(
            text=f"Período: {self.periodo_inicio.strftime('%d/%m/%Y')} a {self.periodo_fim.strftime('%d/%m/%Y')}"
        )
        
        # Mostrar progresso
        self.mostrar_progresso_geracao()
        
        try:
            # Carregar dados conforme o modo
            if self.modo_relatorio == "fornecedores":
                if not self.carregar_dados_fornecedores():
                    self.ocultar_progresso_geracao()
                    return
            else:
                if not self.carregar_dados_por_fornecedor():
                    self.ocultar_progresso_geracao()
                    return
            
            # Preencher resumo
            self.preencher_resumo()
            
            # Atualizar lista para a aba de detalhes
            self.atualizar_lista_detalhes()
            
            # Limpar detalhes
            self.limpar_detalhes()
            
            # Gerar gráfico inicial
            self.atualizar_grafico()
            
            # Marcar que os dados foram carregados
            self.dados_carregados = True
            
            # Selecionar aba de resumo
            self.notebook.select(0)
            
            # Ocultar progresso
            self.ocultar_progresso_geracao()
            
            # Mostrar resumo do que foi gerado
            self.mostrar_resumo_geracao()
            
        except Exception as e:
            self.ocultar_progresso_geracao()
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {str(e)}")

    # NOVO MÉTODO: Limpar dados do relatório anterior
    def limpar_dados_relatorio_anterior(self):
        """Limpa apenas os dados do relatório anterior mantendo as configurações"""
        try:
            # Manter: modo_relatorio, cliente_atual, fornecedor_especifico, periodo
            # Limpar: apenas os dados gerados
            
            self.dados_fornecedores = {}
            self.dados_por_fornecedor = {}
            self.top_fornecedores = []
            self.total_geral = 0.0
            self.dados_carregados = False
            
            # Limpar apenas as tabelas (não os controles)
            for item in self.tree_resumo.get_children():
                self.tree_resumo.delete(item)
            
            for item in self.tree_lancamentos.get_children():
                self.tree_lancamentos.delete(item)
            
            # Resetar totais
            self.lbl_total_geral.config(text="Total Geral: R$ 0,00")
            self.lbl_total_apresentado.config(text="Total Apresentado: R$ 0,00 (0%)")
            
        except Exception as e:
            print(f"Erro ao limpar dados do relatório anterior: {str(e)}")

    # NOVOS MÉTODOS: Controle de progresso
    def mostrar_progresso_geracao(self):
        """Mostra indicador de progresso durante a geração"""
        try:
            # Criar ou atualizar uma label de status
            if not hasattr(self, 'lbl_status_geracao'):
                self.lbl_status_geracao = ttk.Label(
                    self.frame_selecao,
                    text="Gerando relatório...",
                    font=('Arial', 10, 'italic'),
                    foreground='blue'
                )
            
            self.lbl_status_geracao.config(text="Gerando relatório, aguarde...")
            self.lbl_status_geracao.pack(pady=5)
            self.root.update()
            
        except Exception as e:
            print(f"Erro ao mostrar progresso: {str(e)}")

    def ocultar_progresso_geracao(self):
        """Oculta o indicador de progresso"""
        try:
            if hasattr(self, 'lbl_status_geracao'):
                self.lbl_status_geracao.pack_forget()
                
        except Exception as e:
            print(f"Erro ao ocultar progresso: {str(e)}")

    def mostrar_resumo_geracao(self):
        """Mostra um resumo do que foi gerado"""
        try:
            if self.modo_relatorio == "fornecedores":
                if self.todos_clientes:
                    cliente_info = "todos os clientes"
                else:
                    cliente_info = f"cliente '{self.cliente_atual}'"
                
                resumo = f"Relatório gerado para {cliente_info}:\n" + \
                        f"• {len(self.top_fornecedores)} fornecedores encontrados\n" + \
                        f"• Total geral: {formatar_moeda_br(self.total_geral)}"
            else:
                resumo = f"Relatório gerado para fornecedor '{self.fornecedor_especifico}':\n" + \
                        f"• {len(self.top_fornecedores)} clientes encontrados\n" + \
                        f"• Total geral: {formatar_moeda_br(self.total_geral)}"
            
            # Mostrar na barra de status temporariamente
            if hasattr(self, 'lbl_status_geracao'):
                self.lbl_status_geracao.config(
                    text=f"Relatório gerado com sucesso! {len(self.top_fornecedores)} registros encontrados.",
                    foreground='green'
                )
                self.lbl_status_geracao.pack(pady=5)
                
                # Ocultar após 3 segundos
                self.root.after(3000, self.ocultar_progresso_geracao)
            
        except Exception as e:
            print(f"Erro ao mostrar resumo: {str(e)}")

    # NOVO: Método para carregar dados por fornecedor específico
    def carregar_dados_por_fornecedor(self):
        """Carrega os dados de todos os clientes para um fornecedor específico"""
        try:
            # Dicionário para armazenar dados por cliente
            self.dados_por_fornecedor = defaultdict(lambda: {
                'total': 0.0,
                'lancamentos': [],
                'qtd_lancamentos': 0,
                'tipos_despesa': set(),
                'por_mes': defaultdict(float),
                'por_tipo': defaultdict(float)
            })
            
            # Variáveis para somatórios
            self.total_geral = 0.0
            
            # Processar todos os arquivos de clientes
            clientes_processados = []
            fornecedor_encontrado = False
            
            for arquivo in os.listdir(PASTA_CLIENTES):
                if arquivo.endswith('.xlsx'):
                    try:
                        caminho_arquivo = os.path.join(PASTA_CLIENTES, arquivo)
                        nome_cliente = os.path.splitext(arquivo)[0]
                        
                        # Processar arquivo do cliente para o fornecedor específico
                        if self.processar_arquivo_cliente_fornecedor(caminho_arquivo, nome_cliente):
                            clientes_processados.append(nome_cliente)
                            fornecedor_encontrado = True
                    except Exception as e:
                        print(f"Erro ao processar arquivo {arquivo}: {str(e)}")
            
            if not fornecedor_encontrado:
                messagebox.showwarning("Aviso", f"Fornecedor '{self.fornecedor_especifico}' não encontrado em nenhum cliente no período selecionado!")
                return False
            
            if self.total_geral == 0:
                messagebox.showinfo("Aviso", f"Nenhum lançamento encontrado para o fornecedor '{self.fornecedor_especifico}' no período selecionado.")
                return False
            
            # Ordenar clientes por valor total (decrescente)
            clientes_ordenados = sorted(
                self.dados_por_fornecedor.items(),
                key=lambda x: x[1]['total'],
                reverse=True
            )
            
            # Obter top N clientes
            try:
                top_n = int(self.top_n_var.get())
            except ValueError:
                top_n = 10
                
            self.top_fornecedores = clientes_ordenados[:top_n]  # Reutilizando a variável
            
            print(f"Clientes processados para {self.fornecedor_especifico}: {len(clientes_processados)}")
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
                

    # NOVO: Método para processar arquivo de cliente para um fornecedor específico
    def processar_arquivo_cliente_fornecedor(self, caminho_arquivo, nome_cliente):
        """Processa um arquivo de cliente buscando lançamentos de um fornecedor específico"""
        try:
            # Carregar dados do Excel
            df = pd.read_excel(caminho_arquivo, sheet_name='Dados')
            
            # Verificar colunas necessárias
            colunas_necessarias = ['DATA_REL', 'TP_DESP', 'NOME', 'REFERÊNCIA', 'VALOR']
            if not all(coluna in df.columns for coluna in colunas_necessarias):
                print(f"Arquivo {nome_cliente} não contém todas as colunas necessárias.")
                return False
            
            # Filtrar apenas lançamentos ativos
            if 'STATUS' in df.columns:
                df = df[df['STATUS'].str.upper().str.strip() == 'ATIVO'].copy()
                if df.empty:
                    return False
            
            # Converter DATA_REL para datetime
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            
            # Remover linhas com datas inválidas
            df = df.dropna(subset=['DATA_REL'])
            
            # Filtrar por período
            df_periodo = df[
                (df['DATA_REL'] >= self.periodo_inicio) & 
                (df['DATA_REL'] <= self.periodo_fim)
            ]
            
            if df_periodo.empty:
                return False
            
            # Busca mais precisa pelo fornecedor específico
            # Primeiro, busca exata (case-insensitive)
            fornecedor_mask_exato = df_periodo['NOME'].str.upper().str.strip() == self.fornecedor_especifico.upper().strip()
            df_fornecedor = df_periodo[fornecedor_mask_exato].copy()
            
            # Se não encontrou com busca exata, fazer busca parcial
            if df_fornecedor.empty:
                # Busca por contém (case-insensitive)
                fornecedor_mask_parcial = df_periodo['NOME'].str.upper().str.contains(
                    self.fornecedor_especifico.upper().strip(), 
                    na=False, 
                    regex=False
                )
                df_fornecedor = df_periodo[fornecedor_mask_parcial].copy()
            
            if df_fornecedor.empty:
                return False
            
            encontrou_lancamentos = False
            
            # Processar cada lançamento do fornecedor
            for _, row in df_fornecedor.iterrows():
                try:
                    # Obter valor do lançamento
                    valor = 0.0
                    if isinstance(row['VALOR'], (int, float)):
                        valor = float(row['VALOR'])
                    elif isinstance(row['VALOR'], str):
                        # Limpar string e converter para float
                        valor_str = row['VALOR'].replace('R$', '').replace('.', '').replace(',', '.').strip()
                        try:
                            valor = float(valor_str)
                        except ValueError:
                            valor = 0.0
                    
                    # Ignorar lançamentos com valor zero ou negativo
                    if valor <= 0:
                        continue
                    
                    # Obter tipo de despesa
                    tipo_despesa = int(row['TP_DESP']) if pd.notnull(row['TP_DESP']) else 0
                    
                    # Obter referência
                    referencia = str(row['REFERÊNCIA']) if pd.notnull(row['REFERÊNCIA']) else ""
                    
                    # Verificar se existe coluna 'NF' e adicionar à referência se disponível
                    nf = ""
                    if 'NF' in df.columns and pd.notnull(row['NF']) and str(row['NF']).strip():
                        nf = str(row['NF']).strip()
                        if nf and nf.lower() != 'nan':
                            if referencia:
                                referencia = f"{referencia} (NF: {nf})"
                            else:
                                referencia = f"NF: {nf}"
                    
                    # Obter data
                    data = row['DATA_REL']
                    
                    # Obter data de vencimento se disponível
                    dt_vencto = None
                    if 'DT_VENCTO' in df.columns and pd.notnull(row['DT_VENCTO']):
                        try:
                            dt_vencto = pd.to_datetime(row['DT_VENCTO'])
                        except:
                            dt_vencto = None
                    
                    # Obter observação se disponível
                    observacao = ""
                    for col_obs in ['OBSERVACAO', 'OBSERVAÇÃO']:
                        if col_obs in df.columns and pd.notnull(row[col_obs]):
                            observacao = str(row[col_obs])
                            break
                    
                    # Criar identificador do mês para análise mensal
                    mes_ano = f"{data.year}-{data.month:02d}"
                    
                    # Atualizar dados do cliente
                    self.dados_por_fornecedor[nome_cliente]['total'] += valor
                    self.dados_por_fornecedor[nome_cliente]['qtd_lancamentos'] += 1
                    self.dados_por_fornecedor[nome_cliente]['tipos_despesa'].add(tipo_despesa)
                    self.dados_por_fornecedor[nome_cliente]['por_mes'][mes_ano] += valor
                    self.dados_por_fornecedor[nome_cliente]['por_tipo'][tipo_despesa] += valor
                    
                    # Adicionar lançamento à lista de lançamentos do cliente
                    self.dados_por_fornecedor[nome_cliente]['lancamentos'].append({
                        'data': data,
                        'fornecedor': str(row['NOME']).strip(),  # Nome real do fornecedor
                        'tipo_despesa': tipo_despesa,
                        'referencia': referencia,
                        'nf': nf,
                        'dt_vencto': dt_vencto,
                        'valor': valor,
                        'observacao': observacao
                    })
                    
                    # Atualizar total geral
                    self.total_geral += valor
                    encontrou_lancamentos = True
                    
                except Exception as e:
                    print(f"Erro ao processar lançamento no cliente {nome_cliente}: {str(e)}")
                    continue
            
            if encontrou_lancamentos:
                print(f"Cliente {nome_cliente}: {len(df_fornecedor)} lançamentos encontrados para '{self.fornecedor_especifico}'")
            
            return encontrou_lancamentos
            
        except Exception as e:
            print(f"Erro ao processar arquivo {nome_cliente}: {str(e)}")
            return False
    
    def carregar_dados_fornecedores(self):
        """Carrega os dados para o relatório (modo original)"""
        try:
            # Dicionário para armazenar dados por fornecedor
            self.dados_fornecedores = defaultdict(lambda: {
                'total': 0.0,
                'lancamentos': [],
                'qtd_lancamentos': 0,
                'tipos_despesa': set(),
                'clientes': set(),
                'por_mes': defaultdict(float),
                'por_tipo': defaultdict(float)
            })
            
            # Variáveis para somatórios
            self.total_geral = 0.0
            
            if self.todos_clientes:
                # Processar todos os arquivos de clientes
                clientes_processados = []
                
                for arquivo in os.listdir(PASTA_CLIENTES):
                    if arquivo.endswith('.xlsx'):
                        try:
                            caminho_arquivo = os.path.join(PASTA_CLIENTES, arquivo)
                            nome_cliente = os.path.splitext(arquivo)[0]
                            
                            # Processar arquivo do cliente
                            self.processar_arquivo_cliente(caminho_arquivo, nome_cliente)
                            clientes_processados.append(nome_cliente)
                        except Exception as e:
                            print(f"Erro ao processar arquivo {arquivo}: {str(e)}")
                
                if not clientes_processados:
                    messagebox.showwarning("Aviso", "Nenhum arquivo de cliente encontrado!")
                    return False
                    
                print(f"Clientes processados: {len(clientes_processados)}")
                
            else:
                # Processar apenas o cliente selecionado
                if not os.path.exists(self.arquivo_cliente):
                    messagebox.showerror("Erro", f"Arquivo do cliente '{self.cliente_atual}' não encontrado!")
                    return False
                
                # Processar arquivo do cliente
                self.processar_arquivo_cliente(self.arquivo_cliente, self.cliente_atual)
            
            # Verificar se encontrou algum lançamento
            if self.total_geral == 0:
                messagebox.showinfo("Aviso", "Nenhum lançamento encontrado no período selecionado.")
                return False
            
            # Ordenar fornecedores por valor total (decrescente)
            fornecedores_ordenados = sorted(
                self.dados_fornecedores.items(),
                key=lambda x: x[1]['total'],
                reverse=True
            )
            
            # Obter top N fornecedores
            try:
                top_n = int(self.top_n_var.get())
            except ValueError:
                top_n = 10
                
            self.top_fornecedores = fornecedores_ordenados[:top_n]
            
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def processar_arquivo_cliente(self, caminho_arquivo, nome_cliente):
        """Processa um arquivo de cliente (método original)"""
        try:
            # Carregar dados do Excel
            df = pd.read_excel(caminho_arquivo, sheet_name='Dados')
            
            # Verificar colunas necessárias
            colunas_necessarias = ['DATA_REL', 'TP_DESP', 'NOME', 'REFERÊNCIA', 'VALOR']
            if not all(coluna in df.columns for coluna in colunas_necessarias):
                print(f"Arquivo {nome_cliente} não contém todas as colunas necessárias.")
                return
            
            # Filtrar apenas lançamentos ativos
            if 'STATUS' in df.columns:
                # Filtrar apenas registros com STATUS = 'ATIVO'
                df_original_len = len(df)
                df = df[df['STATUS'].str.upper().str.strip() == 'ATIVO'].copy()
                df_filtrado_len = len(df)
                
                print(f"Cliente {nome_cliente}: {df_original_len} registros totais, {df_filtrado_len} ativos processados")
                
                # Se não há registros ativos, não processar
                if df.empty:
                    print(f"Nenhum lançamento ativo encontrado para {nome_cliente}")
                    return
            else:
                # Se não existe a coluna STATUS, processar todos (compatibilidade)
                print(f"Cliente {nome_cliente}: Coluna STATUS não encontrada, processando todos os registros")

            # Converter DATA_REL para datetime
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            
            # Remover linhas com datas inválidas
            df = df.dropna(subset=['DATA_REL'])
            
            # Filtrar por período
            df_periodo = df[
                (df['DATA_REL'] >= self.periodo_inicio) & 
                (df['DATA_REL'] <= self.periodo_fim)
            ]
            
            # Verificar se há dados no período
            if df_periodo.empty:
                print(f"Nenhum lançamento encontrado para {nome_cliente} no período selecionado.")
                return
            
            # Processar cada lançamento
            for _, row in df_periodo.iterrows():
                try:
                    # Obter nome do fornecedor
                    fornecedor = row['NOME']
                    if not isinstance(fornecedor, str) or not fornecedor.strip():
                        continue
                    
                    fornecedor = fornecedor.strip()
                    
                    # Obter valor do lançamento
                    valor = 0.0
                    if isinstance(row['VALOR'], (int, float)):
                        valor = float(row['VALOR'])
                    elif isinstance(row['VALOR'], str):
                        # Limpar string e converter para float
                        valor_str = row['VALOR'].replace('R$', '').replace('.', '').replace(',', '.').strip()
                        try:
                            valor = float(valor_str)
                        except ValueError:
                            valor = 0.0
                    
                    # Ignorar lançamentos com valor zero ou negativo
                    if valor <= 0:
                        continue
                    
                    # Obter tipo de despesa
                    tipo_despesa = int(row['TP_DESP']) if pd.notnull(row['TP_DESP']) else 0
                    
                    # Obter referência
                    referencia = str(row['REFERÊNCIA']) if pd.notnull(row['REFERÊNCIA']) else ""
                    
                    # Verificar se existe coluna 'NF' e adicionar à referência se disponível
                    nf = ""
                    if 'NF' in df.columns and pd.notnull(row['NF']) and str(row['NF']).strip():
                        nf = str(row['NF']).strip()
                        if nf and nf.lower() != 'nan':
                            if referencia:
                                referencia = f"{referencia} (NF: {nf})"
                            else:
                                referencia = f"NF: {nf}"
                    
                    # Obter data
                    data = row['DATA_REL']
                    
                    # Obter data de vencimento se disponível
                    dt_vencto = None
                    if 'DT_VENCTO' in df.columns and pd.notnull(row['DT_VENCTO']):
                        try:
                            dt_vencto = pd.to_datetime(row['DT_VENCTO'])
                        except:
                            dt_vencto = None
                    
                    # Obter observação se disponível
                    observacao = ""
                    for col_obs in ['OBSERVACAO', 'OBSERVAÇÃO']:
                        if col_obs in df.columns and pd.notnull(row[col_obs]):
                            observacao = str(row[col_obs])
                            break
                    
                    # Criar identificador do mês para análise mensal
                    mes_ano = f"{data.year}-{data.month:02d}"
                    
                    # Atualizar dados do fornecedor
                    self.dados_fornecedores[fornecedor]['total'] += valor
                    self.dados_fornecedores[fornecedor]['qtd_lancamentos'] += 1
                    self.dados_fornecedores[fornecedor]['tipos_despesa'].add(tipo_despesa)
                    self.dados_fornecedores[fornecedor]['clientes'].add(nome_cliente)
                    self.dados_fornecedores[fornecedor]['por_mes'][mes_ano] += valor
                    self.dados_fornecedores[fornecedor]['por_tipo'][tipo_despesa] += valor
                    
                    # Adicionar lançamento à lista de lançamentos do fornecedor
                    self.dados_fornecedores[fornecedor]['lancamentos'].append({
                        'data': data,
                        'cliente': nome_cliente,
                        'tipo_despesa': tipo_despesa,
                        'referencia': referencia,
                        'nf': nf,
                        'dt_vencto': dt_vencto,
                        'valor': valor,
                        'observacao': observacao
                    })
                    
                    # Atualizar total geral
                    self.total_geral += valor
                    
                except Exception as e:
                    print(f"Erro ao processar lançamento: {str(e)}")
                    continue
            
            print(f"Cliente {nome_cliente} processado: {len(df_periodo)} lançamentos no período.")
            
        except Exception as e:
            print(f"Erro ao processar arquivo {nome_cliente}: {str(e)}")

    def preencher_resumo(self):
        """Preenche a aba de resumo com os dados carregados"""
        try:
            # Limpar dados anteriores
            for item in self.tree_resumo.get_children():
                self.tree_resumo.delete(item)
            
            # Atualizar colunas conforme o modo
            self.atualizar_colunas_resumo()
            
            if self.modo_relatorio == "fornecedores":
                # Preencher com dados de fornecedores
                total_apresentado = 0.0
                
                for i, (fornecedor, dados) in enumerate(self.top_fornecedores, 1):
                    percentual = (dados['total'] / self.total_geral) * 100 if self.total_geral > 0 else 0
                    tipos_despesa = ', '.join(map(str, sorted(dados['tipos_despesa'])))
                    
                    self.tree_resumo.insert('', 'end', values=(
                        i,
                        fornecedor,
                        formatar_moeda_br(dados['total']),
                        f"{percentual:.1f}%",
                        dados['qtd_lancamentos'],
                        tipos_despesa
                    ))
                    
                    total_apresentado += dados['total']
                
            else:
                # Preencher com dados de clientes
                total_apresentado = 0.0
                
                for i, (cliente, dados) in enumerate(self.top_fornecedores, 1):  # Reutilizando variável
                    percentual = (dados['total'] / self.total_geral) * 100 if self.total_geral > 0 else 0
                    tipos_despesa = ', '.join(map(str, sorted(dados['tipos_despesa'])))
                    
                    self.tree_resumo.insert('', 'end', values=(
                        i,
                        cliente,
                        formatar_moeda_br(dados['total']),
                        f"{percentual:.1f}%",
                        dados['qtd_lancamentos'],
                        tipos_despesa
                    ))
                    
                    total_apresentado += dados['total']
            
            # Atualizar totais
            self.lbl_total_geral.config(text=f"Total Geral: {formatar_moeda_br(self.total_geral)}")
            
            percentual_apresentado = (total_apresentado / self.total_geral) * 100 if self.total_geral > 0 else 0
            self.lbl_total_apresentado.config(
                text=f"Total Apresentado: {formatar_moeda_br(total_apresentado)} ({percentual_apresentado:.1f}%)"
            )
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao preencher resumo: {str(e)}")

    def atualizar_lista_detalhes(self):
        """Atualiza a lista de itens para a aba de detalhes"""
        try:
            # Limpar lista atual
            self.item_detalhes_combobox['values'] = ()
            self.item_detalhes_combobox.set('')
            
            # Atualizar label conforme o modo
            if self.modo_relatorio == "fornecedores":
                self.lbl_selecao_detalhes.config(text="Selecione o Fornecedor:")
            else:
                self.lbl_selecao_detalhes.config(text="Selecione o Cliente:")
            
            # Criar lista de itens
            if self.top_fornecedores:
                items = [f"{i}. {nome}" for i, (nome, dados) in enumerate(self.top_fornecedores, 1)]
                self.item_detalhes_combobox['values'] = items
                
        except Exception as e:
            print(f"Erro ao atualizar lista de detalhes: {str(e)}")

    def limpar_detalhes(self):
        """Limpa a aba de detalhes"""
        try:
            # Limpar informações do item
            self.lbl_nome_item.config(text="")
            self.lbl_total_item.config(text="")
            self.lbl_qtd_lancamentos_item.config(text="")
            self.lbl_media_lancamento_item.config(text="")
            self.lbl_tipos_despesa_item.config(text="")
            
            # Limpar lista de lançamentos
            for item in self.tree_lancamentos.get_children():
                self.tree_lancamentos.delete(item)
                
        except Exception as e:
            print(f"Erro ao limpar detalhes: {str(e)}")

    def carregar_detalhes_item(self, event=None):
        """Carrega os detalhes de um item selecionado"""
        try:
            selecao = self.item_detalhes_combobox.get()
            if not selecao:
                return
            
            # Extrair posição do item (formato: "1. Nome do Item")
            posicao = int(selecao.split('.')[0]) - 1
            
            if 0 <= posicao < len(self.top_fornecedores):
                nome, dados = self.top_fornecedores[posicao]
                
                # Atualizar informações
                self.lbl_nome_item.config(text=nome)
                self.lbl_total_item.config(text=formatar_moeda_br(dados['total']))
                self.lbl_qtd_lancamentos_item.config(text=str(dados['qtd_lancamentos']))
                
                media = dados['total'] / dados['qtd_lancamentos'] if dados['qtd_lancamentos'] > 0 else 0
                self.lbl_media_lancamento_item.config(text=formatar_moeda_br(media))
                
                tipos = ', '.join(map(str, sorted(dados['tipos_despesa'])))
                self.lbl_tipos_despesa_item.config(text=tipos)
                
                # Carregar lançamentos
                self.carregar_lancamentos_item(dados['lancamentos'])
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar detalhes: {str(e)}")

    def carregar_lancamentos_item(self, lancamentos):
        """Carrega os lançamentos de um item na tabela"""
        try:
            # Limpar tabela
            for item in self.tree_lancamentos.get_children():
                self.tree_lancamentos.delete(item)
            
            # Atualizar colunas conforme o modo
            self.atualizar_colunas_detalhes()
            
            # Adicionar lançamentos
            for lancamento in sorted(lancamentos, key=lambda x: x['data'], reverse=True):
                data_str = lancamento['data'].strftime('%d/%m/%Y')
                dt_vencto_str = lancamento['dt_vencto'].strftime('%d/%m/%Y') if lancamento['dt_vencto'] else ""
                
                if self.modo_relatorio == "fornecedores":
                    # Modo fornecedores: mostrar cliente
                    cliente_ou_fornecedor = lancamento.get('cliente', '')
                else:
                    # Modo por fornecedor: mostrar fornecedor
                    cliente_ou_fornecedor = lancamento.get('fornecedor', '')
                
                self.tree_lancamentos.insert('', 'end', values=(
                    data_str,
                    cliente_ou_fornecedor,
                    lancamento['tipo_despesa'],
                    lancamento['referencia'],
                    lancamento['nf'],
                    dt_vencto_str,
                    formatar_moeda_br(lancamento['valor']),
                    lancamento['observacao']
                ))
                
        except Exception as e:
            print(f"Erro ao carregar lançamentos: {str(e)}")

    def atualizar_grafico(self):
        """Atualiza o gráfico na aba correspondente"""
        if not self.dados_carregados:
            return
        
        # Limpar frame do gráfico
        for widget in self.frame_grafico.winfo_children():
            widget.destroy()
        
        tipo_grafico = self.combo_tipo_grafico.get()
        
        try:
            if tipo_grafico == "Pizza - Total por Item":
                self.criar_grafico_pizza()
            elif tipo_grafico == "Barras - Top Items":
                self.criar_grafico_barras()
            elif tipo_grafico == "Linhas - Evolução Mensal":
                self.criar_grafico_linhas()
            elif tipo_grafico == "Barras Empilhadas - Por Tipo de Despesa":
                self.criar_grafico_barras_empilhadas()
        except Exception as e:
            # Mostrar erro no frame do gráfico
            error_label = ttk.Label(
                self.frame_grafico, 
                text=f"Erro ao gerar gráfico: {str(e)}",
                font=('Arial', 12),
                foreground='red'
            )
            error_label.pack(expand=True)

    def criar_grafico_pizza(self):
        """Cria gráfico de pizza dos principais itens"""
        fig, ax = plt.subplots(figsize=(10, 8))
        
        # Preparar dados
        nomes = []
        valores = []
        for nome, dados in self.top_fornecedores[:10]:  # Limitar a 10 para melhor visualização
            nomes.append(nome[:30])  # Truncar nomes longos
            valores.append(dados['total'])
        
        # Criar gráfico de pizza
        colors = cm.Set3(np.linspace(0, 1, len(nomes)))
        wedges, texts, autotexts = ax.pie(
            valores, 
            labels=nomes, 
            autopct='%1.1f%%',
            colors=colors,
            startangle=90
        )
        
        # Melhorar aparência
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
        
        # Título
        if self.modo_relatorio == "fornecedores":
            ax.set_title('Principais Fornecedores - Distribuição dos Gastos', fontsize=14, fontweight='bold')
        else:
            ax.set_title(f'Principais Clientes - {self.fornecedor_especifico}', fontsize=14, fontweight='bold')
        
        # Adicionar ao frame
        canvas = FigureCanvasTkAgg(fig, self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def criar_grafico_barras(self):
        """Cria gráfico de barras dos principais itens"""
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Preparar dados
        nomes = []
        valores = []
        for nome, dados in self.top_fornecedores[:15]:  # Top 15
            nomes.append(nome[:25])  # Truncar nomes longos
            valores.append(dados['total'])
        
        # Criar gráfico de barras horizontais
        y_pos = np.arange(len(nomes))
        bars = ax.barh(y_pos, valores, color=cm.viridis(np.linspace(0, 1, len(nomes))))
        
        # Configurar eixos
        ax.set_yticks(y_pos)
        ax.set_yticklabels(nomes)
        ax.invert_yaxis()  # Maior valor no topo
        ax.set_xlabel('Valor Total (R$)')
        
        # Título
        if self.modo_relatorio == "fornecedores":
            ax.set_title('Principais Fornecedores - Ranking por Valor', fontsize=14, fontweight='bold')
        else:
            ax.set_title(f'Principais Clientes - {self.fornecedor_especifico}', fontsize=14, fontweight='bold')
        
        # Adicionar valores nas barras
        for i, (bar, valor) in enumerate(zip(bars, valores)):
            ax.text(valor + max(valores) * 0.01, bar.get_y() + bar.get_height()/2, 
                   f'R$ {valor:,.0f}'.replace(',', '.'), 
                   va='center', fontsize=9)
        
        # Ajustar layout
        plt.tight_layout()
        
        # Adicionar ao frame
        canvas = FigureCanvasTkAgg(fig, self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def criar_grafico_linhas(self):
        """Cria gráfico de linhas da evolução mensal"""
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Preparar dados mensais dos top 5 itens
        top_5 = self.top_fornecedores[:5]
        
        # Obter todos os meses do período
        meses = set()
        for _, dados in top_5:
            meses.update(dados['por_mes'].keys())
        
        meses_ordenados = sorted(list(meses))
        
        # Criar linhas para cada item
        colors = cm.tab10(np.linspace(0, 1, len(top_5)))
        
        for i, (nome, dados) in enumerate(top_5):
            valores_mes = []
            for mes in meses_ordenados:
                valores_mes.append(dados['por_mes'].get(mes, 0))
            
            ax.plot(meses_ordenados, valores_mes, 
                   marker='o', linewidth=2, label=nome[:20], color=colors[i])
        
        # Configurar eixos
        ax.set_xlabel('Mês')
        ax.set_ylabel('Valor (R$)')
        ax.set_title('Evolução Mensal dos Principais Itens', fontsize=14, fontweight='bold')
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        ax.grid(True, alpha=0.3)
        
        # Rotacionar labels do eixo x
        plt.setp(ax.get_xticklabels(), rotation=45)
        
        # Ajustar layout
        plt.tight_layout()
        
        # Adicionar ao frame
        canvas = FigureCanvasTkAgg(fig, self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def criar_grafico_barras_empilhadas(self):
        """Cria gráfico de barras empilhadas por tipo de despesa"""
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Preparar dados por tipo de despesa
        top_items = self.top_fornecedores[:10]
        nomes = [nome[:20] for nome, _ in top_items]
        
        # Obter todos os tipos de despesa
        todos_tipos = set()
        for _, dados in top_items:
            todos_tipos.update(dados['por_tipo'].keys())
        
        tipos_ordenados = sorted(list(todos_tipos))
        
        # Preparar dados para empilhamento
        valores_por_tipo = {}
        for tipo in tipos_ordenados:
            valores_por_tipo[tipo] = []
            for _, dados in top_items:
                valores_por_tipo[tipo].append(dados['por_tipo'].get(tipo, 0))
        
        # Criar barras empilhadas
        bottom = np.zeros(len(nomes))
        colors = cm.Set3(np.linspace(0, 1, len(tipos_ordenados)))
        
        for i, tipo in enumerate(tipos_ordenados):
            ax.bar(nomes, valores_por_tipo[tipo], bottom=bottom, 
                  label=f'Tipo {tipo}', color=colors[i])
            bottom += valores_por_tipo[tipo]
        
        # Configurar eixos
        ax.set_ylabel('Valor (R$)')
        ax.set_title('Distribuição por Tipo de Despesa', fontsize=14, fontweight='bold')
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        
        # Rotacionar labels do eixo x
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
        
        # Ajustar layout
        plt.tight_layout()
        
        # Adicionar ao frame
        canvas = FigureCanvasTkAgg(fig, self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def exportar_excel(self):
        """Exporta o relatório para Excel"""
        if not self.dados_carregados:
            messagebox.showwarning("Aviso", "Gere o relatório primeiro!")
            return
        
        # Solicitar local para salvar
        nome_arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Salvar Relatório"
        )
        
        if not nome_arquivo:
            return
        
        try:
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            
            # Título e informações do relatório
            if self.modo_relatorio == "fornecedores":
                ws.title = "Relatório Fornecedores"
                ws['A1'] = "RELATÓRIO DE PRINCIPAIS FORNECEDORES"
                
                if self.todos_clientes:
                    ws['A2'] = "Cliente: Todos os Clientes"
                else:
                    ws['A2'] = f"Cliente: {self.cliente_atual}"
            else:
                ws.title = "Relatório por Fornecedor"
                ws['A1'] = "RELATÓRIO DE CLIENTES POR FORNECEDOR"
                ws['A2'] = f"Fornecedor: {self.fornecedor_especifico}"
            
            ws['A3'] = f"Período: {self.periodo_inicio.strftime('%d/%m/%Y')} a {self.periodo_fim.strftime('%d/%m/%Y')}"
            ws['A4'] = f"Total Geral: {formatar_moeda_br(self.total_geral)}"
            
            # Estilo do cabeçalho
            header_font = Font(bold=True, size=14)
            ws['A1'].font = header_font
            
            # Cabeçalhos da tabela
            row = 6
            if self.modo_relatorio == "fornecedores":
                headers = ['#', 'Fornecedor', 'Total Gasto', '% do Total', 'Qtd. Lançamentos', 'Tipos de Despesa']
            else:
                headers = ['#', 'Cliente', 'Total Gasto', '% do Total', 'Qtd. Lançamentos', 'Tipos de Despesa']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Dados
            for i, (nome, dados) in enumerate(self.top_fornecedores, 1):
                row += 1
                percentual = (dados['total'] / self.total_geral * 100) if self.total_geral > 0 else 0
                tipos_str = ', '.join(sorted([str(t) for t in dados['tipos_despesa']]))
                
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=nome)
                ws.cell(row=row, column=3, value=dados['total'])
                ws.cell(row=row, column=4, value=f"{percentual:.1f}%")
                ws.cell(row=row, column=5, value=dados['qtd_lancamentos'])
                ws.cell(row=row, column=6, value=tipos_str)
            
            # Ajustar largura das colunas
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Criar aba de detalhes
            ws_detalhes = wb.create_sheet("Detalhes")
            
            # Cabeçalho dos detalhes
            if self.modo_relatorio == "fornecedores":
                detail_headers = ['Data', 'Cliente', 'Fornecedor', 'Tipo', 'Referência', 'NF', 'Vencimento', 'Valor', 'Observação']
            else:
                detail_headers = ['Data', 'Fornecedor', 'Cliente', 'Tipo', 'Referência', 'NF', 'Vencimento', 'Valor', 'Observação']
            
            for col, header in enumerate(detail_headers, 1):
                cell = ws_detalhes.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Adicionar todos os lançamentos
            row = 2
            for nome, dados in self.top_fornecedores:
                for lancamento in dados['lancamentos']:
                    ws_detalhes.cell(row=row, column=1, value=lancamento['data'].strftime('%d/%m/%Y'))
                    
                    if self.modo_relatorio == "fornecedores":
                        ws_detalhes.cell(row=row, column=2, value=lancamento.get('cliente', ''))
                        ws_detalhes.cell(row=row, column=3, value=nome)
                    else:
                        ws_detalhes.cell(row=row, column=2, value=lancamento.get('fornecedor', ''))
                        ws_detalhes.cell(row=row, column=3, value=nome)
                    
                    ws_detalhes.cell(row=row, column=4, value=lancamento['tipo_despesa'])
                    ws_detalhes.cell(row=row, column=5, value=lancamento['referencia'])
                    ws_detalhes.cell(row=row, column=6, value=lancamento.get('nf', ''))
                    
                    dt_vencto_str = ""
                    if lancamento['dt_vencto']:
                        dt_vencto_str = lancamento['dt_vencto'].strftime('%d/%m/%Y')
                    ws_detalhes.cell(row=row, column=7, value=dt_vencto_str)
                    
                    ws_detalhes.cell(row=row, column=8, value=lancamento['valor'])
                    ws_detalhes.cell(row=row, column=9, value=lancamento.get('observacao', ''))
                    
                    row += 1
            
            # Ajustar largura das colunas da aba de detalhes
            for col in range(1, 10):
                ws_detalhes.column_dimensions[get_column_letter(col)].width = 15
            
            # Salvar arquivo
            wb.save(nome_arquivo)
            messagebox.showinfo("Sucesso", f"Relatório exportado para: {nome_arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel: {str(e)}")

    def exportar_pdf(self):
        """Exporta o relatório para PDF"""
        if not self.dados_carregados:
            messagebox.showwarning("Aviso", "Gere o relatório primeiro!")
            return
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            # Solicitar local para salvar
            nome_arquivo = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Salvar Relatório PDF"
            )
            
            if not nome_arquivo:
                return
            
            # Criar documento
            doc = SimpleDocTemplate(nome_arquivo, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centralizado
            )
            
            # Título
            if self.modo_relatorio == "fornecedores":
                titulo = "RELATÓRIO DE PRINCIPAIS FORNECEDORES"
            else:
                titulo = "RELATÓRIO DE CLIENTES POR FORNECEDOR"
            
            story.append(Paragraph(titulo, title_style))
            story.append(Spacer(1, 12))
            
            # Informações do relatório
            if self.modo_relatorio == "fornecedores":
                if self.todos_clientes:
                    info_cliente = "Cliente: Todos os Clientes"
                else:
                    info_cliente = f"Cliente: {self.cliente_atual}"
            else:
                info_cliente = f"Fornecedor: {self.fornecedor_especifico}"
            
            story.append(Paragraph(info_cliente, styles['Normal']))
            story.append(Paragraph(
                f"Período: {self.periodo_inicio.strftime('%d/%m/%Y')} a {self.periodo_fim.strftime('%d/%m/%Y')}", 
                styles['Normal']
            ))
            story.append(Paragraph(f"Total Geral: {formatar_moeda_br(self.total_geral)}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Preparar dados da tabela
            if self.modo_relatorio == "fornecedores":
                headers = ['#', 'Fornecedor', 'Total Gasto', '% do Total', 'Qtd. Lançamentos']
            else:
                headers = ['#', 'Cliente', 'Total Gasto', '% do Total', 'Qtd. Lançamentos']
            
            data = [headers]
            
            for i, (nome, dados) in enumerate(self.top_fornecedores, 1):
                percentual = (dados['total'] / self.total_geral * 100) if self.total_geral > 0 else 0
                
                data.append([
                    str(i),
                    nome[:40],  # Truncar nomes muito longos
                    formatar_moeda_br(dados['total']),
                    f"{percentual:.1f}%",
                    str(dados['qtd_lancamentos'])
                ])
            
            # Criar tabela
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
            # Construir PDF
            doc.build(story)
            messagebox.showinfo("Sucesso", f"Relatório PDF exportado para: {nome_arquivo}")
            
        except ImportError:
            messagebox.showerror(
                "Erro", 
                "Biblioteca ReportLab não encontrada!\n" +
                "Instale com: pip install reportlab"
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para PDF: {str(e)}")

    def voltar_menu(self):
        """Volta ao menu principal ou fecha a janela adequadamente"""
        try:
            # Se foi chamado como janela filha (tem parent)
            if self.parent and self.menu_principal:
                # Destruir apenas esta janela
                self.root.destroy()
                
                # Reativar a janela pai se ela existir
                if hasattr(self.menu_principal, 'deiconify'):
                    self.menu_principal.deiconify()
                elif hasattr(self.menu_principal, 'focus_force'):
                    self.menu_principal.focus_force()
                    
            else:
                # Se foi chamado como janela principal, apenas fechar
                self.root.quit()  # Para o mainloop
                self.root.destroy()  # Destrói a janela
                
        except Exception as e:
            print(f"Erro ao voltar ao menu: {str(e)}")
            # Em caso de erro, força o fechamento
            try:
                self.root.destroy()
            except:
                pass

    def fechar_aplicacao(self):
        """Método para fechar a aplicação adequadamente"""
        try:
            # Confirmar se o usuário quer sair
            if messagebox.askyesno("Confirmar", "Deseja realmente sair?"):
                self.voltar_menu()
        except Exception as e:
            print(f"Erro ao fechar aplicação: {str(e)}")


def main():
    """Função principal para executar o módulo de forma independente"""
    app = RelatorioFornecedores()
    app.root.mainloop()
    
if __name__ == "__main__":
    main()