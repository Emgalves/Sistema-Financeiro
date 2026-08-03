# -*- coding: utf-8 -*-
r"""
regularizar_fornecedor.py
==========================
Corrige CNPJ/CPF (ou nome/razão social) de um fornecedor NA BASE MESTRE e em
TODOS os arquivos de clientes onde ele aparece, mantendo um log de auditoria
de cada célula alterada.

Por que isso é separado de `editar_fornecedor()`:
--------------------------------------------------
O editor normal trava o campo CNPJ/CPF de propósito, porque trocar o
documento de um fornecedor no dia a dia é perigoso (pode juntar dois
fornecedores diferentes por engano). Este módulo é a "válvula de escape"
formal e auditada para quando o documento realmente está errado — deve ser
disparado por um botão separado ("🔧 Regularizar Cadastro"), nunca a partir
do fluxo comum de edição.

Uso típico (a partir da sua classe GUI):
-----------------------------------------
    from regularizar_fornecedor import regularizar_fornecedor

    resultado = regularizar_fornecedor(
        base_path=ARQUIVO_FORNECEDORES,
        pasta_clientes=r"C:\...\Clientes",
        doc_antigo="00051798600",
        doc_novo="04751798610",     # opcional, só se o documento mudar
        nome_novo=None,             # opcional, só se o nome também mudar
        razao_novo=None,
        dry_run=True,               # SEMPRE rode dry_run=True primeiro!
    )
    for c in resultado.alteracoes:
        print(c)

Regras de segurança embutidas:
-------------------------------
1. Nunca decide sozinho por "match de nome" — só altera célula cujo
   documento (comparado por dígitos) bate com `doc_antigo`. Nome divergente
   sem documento batendo entra em `resultado.divergencias_nome` para revisão
   manual, nunca é alterado automaticamente.
2. Se `doc_novo` já pertencer a OUTRO fornecedor cadastrado na base (um
   fornecedor diferente, não o mesmo registro), a função recusa e retorna
   erro — isso é fusão de cadastros, não correção de digitação, e precisa de
   uma decisão humana explícita (ver `fundir_fornecedores`, não implementado
   aqui de propósito).
3. `dry_run=True` (padrão) apenas relata o que seria alterado, sem gravar
   nada. Só grava quando `dry_run=False`.
4. Arquivos abertos no Excel (PermissionError) são reportados em
   `resultado.arquivos_com_falha` e NADA nesse arquivo é alterado — evita
   estado parcialmente corrigido.
5. Arquivos de relatório por fornecedor (o .xlsx individual gerado por
   "📋 Lançamentos"/VisualizadorLancamentosFornecedor) NÃO são editados por
   este módulo — eles são derivados da aba "Dados"/"Registros" dos clientes.
   Depois de regularizar, regenere-os (rode o visualizador de novo) em vez
   de tentar corrigi-los manualmente.
6. TIPO_PESSOA (PF/PJ) nunca é inferido pela quantidade de dígitos do
   documento. É lido da coluna 'tipo_pessoa' da própria base (definida no
   cadastro original) e usado como verdade para decidir o zero-padding
   (11 dígitos para PF, 14 para PJ) antes de gravar em qualquer lugar. Isso
   evita o mesmo bug que existe em `atualizar_tipo_pessoa()` do editor atual:
   se o CNPJ/CPF chegar até aqui com zeros à esquerda perdidos (o que
   acontece quando alguém edita a célula manualmente no Excel e a formatação
   vira Número/Geral), contar dígitos dá uma classificação errada.

Changelog
---------
- v1: regularizar_fornecedor() — correção de documento/nome com propagação.
- v2: + fundir_fornecedores() e detectar_possiveis_duplicatas() — fusão de
  cadastros duplicados (ex.: CPF criado + CPF real do mesmo fornecedor).
- v3: + tipo_pessoa da base como fonte da verdade para zero-padding (nunca
  mais inferido por contagem de dígitos); + propagação de RAZÃO SOCIAL para
  a aba 'Contratos_ADM' (campo "Nome/Razão Social", usado no contrato em si
  — distinto do "nome fantasia" usado em 'Dados'/'Contratos_Medicao').
"""

import os
import glob
import json
import difflib
from datetime import datetime
from dataclasses import dataclass, field
from openpyxl import load_workbook

# Validação estrutural (dígito verificador) de CPF/CNPJ — reaproveita a
# implementação oficial já existente no sistema, em vez de duplicar o
# algoritmo aqui. Não confirma que o documento existe de fato (isso só a
# Receita Federal saberia); só garante que os dígitos são matematicamente
# possíveis, pegando erros de digitação como "trocar só o último dígito".
from src.config.utils import validar_documento

# Aba de controle de CPFs criados (GerenciadorCPFsCriados) dentro do próprio
# Base_Fornecedores.xlsx
CPF_CRIADO_SHEET = "CPF"
CPF_CRIADO_COL_DOC = 1       # coluna A: CPF gerado (só dígitos)
CPF_CRIADO_COL_STATUS = 2    # coluna B: DISPONIVEL / USADO / INVALIDO
CPF_CRIADO_COL_NOME = 3      # coluna C: nome de quem está usando
CPF_CRIADO_COL_DATA = 4      # coluna D: data de uso


# ---------------------------------------------------------------------------
# Config: onde procurar CNPJ/CPF em cada tipo de arquivo.
# Ajuste esses mapas se sua estrutura de abas mudar.
# ---------------------------------------------------------------------------

# Base mestre (Base_Fornecedores.xlsx) — documento gravado SEM formatação
BASE_SHEET = "Fornecedores"
BASE_COL_DOC = 1          # coluna A
BASE_COL_TIPO_PESSOA = 2  # coluna B: 'PF' ou 'PJ' — DEFINIDO NO CADASTRO,
                           # é a fonte da verdade. NUNCA inferir PF/PJ pela
                           # quantidade de dígitos do documento (não confiável
                           # quando zeros à esquerda se perdem).
BASE_COL_RAZAO = 3        # coluna C
BASE_COL_NOME = 4         # coluna D

# Arquivos de cliente — cada aba onde o documento pode aparecer, em que
# formato ('raw' = só dígitos, 'formatado' = com pontuação) e a que campo do
# fornecedor a coluna de nome corresponde SEMANTICAMENTE ('nome' = nome
# fantasia/curto; 'razao_social' = nome/razão social usado no texto do
# contrato). Baseado na estrutura observada em JOSE_PAULO_DE_SOUZA_PEIXE.xlsx.
CLIENTE_MAPA_ABAS = [
    {"aba": "Dados",              "col_doc": 3,  "col_nome": 4,  "formato": "formatado", "campo_nome": "nome"},
    {"aba": "Contratos_Medicao",  "col_doc": 2,  "col_nome": 3,  "formato": "formatado", "campo_nome": "nome"},
    {"aba": "Medicoes",           "col_doc": 3,  "col_nome": 4,  "formato": "formatado", "campo_nome": "nome"},
    # Contratos_ADM tem 3 blocos de colunas na mesma aba (contrato,
    # aditivo, parcelas), cada um com seu próprio par CNPJ/CPF + Nome/Razão
    # Social. É o campo que aparece de fato no texto do contrato — por isso
    # mapeado como 'razao_social', não 'nome'.
    {"aba": "Contratos_ADM",      "col_doc": 8,  "col_nome": 9,  "formato": "raw", "campo_nome": "razao_social"},
    {"aba": "Contratos_ADM",      "col_doc": 20, "col_nome": 21, "formato": "raw", "campo_nome": "razao_social"},
    {"aba": "Contratos_ADM",      "col_doc": 27, "col_nome": 28, "formato": "raw", "campo_nome": "razao_social"},
]

# Arquivos a ignorar quando varrendo a pasta de clientes (o próprio mestre e
# relatórios derivados por fornecedor não devem ser tratados como "cliente")
ARQUIVOS_IGNORAR = {"base_fornecedores.xlsx"}


def _digitos(v):
    return "".join(ch for ch in str(v) if ch.isdigit()) if v is not None else ""


class _DebugLogger:
    """Log de diagnóstico opcional: escreve uma linha por evento, ABRINDO E
    FECHANDO o arquivo a cada chamada (sem buffer) — se o processo travar ou
    for encerrado à força no meio de uma operação, as linhas já escritas
    continuam no disco, apontando exatamente até onde chegou. Ative passando
    `arquivo_log_debug='caminho/algo.log'` para regularizar_fornecedor() ou
    fundir_fornecedores()."""
    def __init__(self, caminho):
        self.caminho = caminho
        if caminho:
            with open(caminho, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*70}\n[{datetime.now().isoformat(timespec='seconds')}] INÍCIO\n")

    def log(self, msg):
        if not self.caminho:
            return
        with open(self.caminho, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().isoformat(timespec='seconds')}] {msg}\n")


def _formatar_doc(digitos, tipo_pessoa=None):
    """Formata string de dígitos como CPF ou CNPJ.

    Se `tipo_pessoa` ('PF'/'PJ') for informado, ele manda: o valor é
    zero-preenchido para 11 (PF) ou 14 (PJ) dígitos ANTES de formatar — isso
    corrige o caso de zeros à esquerda perdidos (documento com zeros que
    virou número em algum momento no Excel). Só cai de volta para "adivinhar
    pelo tamanho" quando tipo_pessoa não é passado (uso legado/depuração) —
    isso é deliberadamente menos confiável e não deve ser usado para gravar
    dados, só para exibição avulsa.
    """
    if tipo_pessoa == "PF":
        digitos = digitos.zfill(11)
    elif tipo_pessoa == "PJ":
        digitos = digitos.zfill(14)
    if len(digitos) == 11:
        return f"{digitos[0:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:11]}"
    if len(digitos) == 14:
        return f"{digitos[0:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:14]}"
    return digitos


def _tipo_pessoa_da_linha(row_values):
    """Lê e normaliza a coluna tipo_pessoa (B) de uma linha da base.
    Retorna 'PF', 'PJ' ou None se estiver vazia/inválida — nunca adivinha."""
    v = str(row_values[BASE_COL_TIPO_PESSOA - 1] or "").strip().upper()
    return v if v in ("PF", "PJ") else None


@dataclass
class Alteracao:
    arquivo: str
    aba: str
    celula: str
    campo: str
    valor_antigo: str
    valor_novo: str


@dataclass
class ResultadoRegularizacao:
    doc_antigo: str
    doc_novo: str | None
    nome_novo: str | None
    razao_novo: str | None
    dry_run: bool
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    alteracoes: list = field(default_factory=list)
    divergencias_nome: list = field(default_factory=list)   # nome não bate, não alterado
    arquivos_com_falha: list = field(default_factory=list)  # PermissionError etc.
    erro: str | None = None

    def salvar_log(self, caminho):
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump({
                "doc_antigo": self.doc_antigo,
                "doc_novo": self.doc_novo,
                "nome_novo": self.nome_novo,
                "razao_novo": self.razao_novo,
                "dry_run": self.dry_run,
                "timestamp": self.timestamp,
                "alteracoes": [a.__dict__ for a in self.alteracoes],
                "divergencias_nome": self.divergencias_nome,
                "arquivos_com_falha": self.arquivos_com_falha,
                "erro": self.erro,
            }, f, ensure_ascii=False, indent=2)

    def resumo(self):
        linhas = [
            f"Documento antigo : {self.doc_antigo}",
            f"Documento novo   : {self.doc_novo or '(inalterado)'}",
            f"Nome novo        : {self.nome_novo or '(inalterado)'}",
            f"Modo             : {'DRY-RUN (nada foi gravado)' if self.dry_run else 'APLICADO'}",
            f"Total de células alteradas: {len(self.alteracoes)}",
        ]
        if self.erro:
            linhas.append(f"ERRO: {self.erro}")
        for a in self.alteracoes:
            linhas.append(f"  [{a.arquivo}] {a.aba}!{a.celula} ({a.campo}): '{a.valor_antigo}' -> '{a.valor_novo}'")
        if self.divergencias_nome:
            linhas.append("Divergências de nome encontradas (NÃO alteradas, revisar manualmente):")
            for d in self.divergencias_nome:
                linhas.append(f"  [{d['arquivo']}] {d['aba']}!{d['celula']}: nome na planilha = '{d['nome_planilha']}'")
        if self.arquivos_com_falha:
            linhas.append("Arquivos que NÃO puderam ser abertos/salvos (feche-os e rode de novo):")
            for f_ in self.arquivos_com_falha:
                linhas.append(f"  - {f_}")
        return "\n".join(linhas)


def _doc_ja_pertence_a_outro_fornecedor(base_path, doc_novo_digitos, doc_antigo_digitos):
    """Impede fusão silenciosa: se doc_novo já está cadastrado em outra linha
    da base (que não seja a linha que estamos corrigindo), aborta."""
    wb = load_workbook(base_path, data_only=True)
    ws = wb[BASE_SHEET]
    achou = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[BASE_COL_DOC - 1]:
            continue
        d = _digitos(row[BASE_COL_DOC - 1])
        if d == doc_novo_digitos and d != doc_antigo_digitos:
            achou = True
            break
    wb.close()
    return achou


def regularizar_fornecedor(base_path, pasta_clientes, doc_antigo,
                            doc_novo=None, nome_novo=None, razao_novo=None,
                            dry_run=True, arquivos_cliente=None,
                            callback_progresso=None, arquivo_log_debug=None):
    """
    Corrige o documento/nome/razão social de um fornecedor na base mestre e
    em todos os arquivos de cliente encontrados em `pasta_clientes` (ou na
    lista explícita `arquivos_cliente`, se fornecida — útil para testar em
    um subconjunto).

    `callback_progresso(indice, total, nome_arquivo)`, se informado, é
    chamado antes de cada arquivo de cliente ser verificado — use para
    mostrar progresso na GUI (barra, label) ou para logar/depurar em qual
    arquivo o processo está, caso demore ou trave em algum específico.

    `arquivo_log_debug`, se informado (caminho de um .txt/.log), grava uma
    linha com timestamp ANTES E DEPOIS de cada operação potencialmente
    demorada (abrir/salvar a base, abrir/salvar cada arquivo de cliente).
    Cada linha é escrita e o arquivo fechado na hora — se o processo travar
    ou for encerrado à força, o log já tem tudo até o ponto exato onde
    parou. Use isso pra diagnosticar uma trava que não se repete fora do
    sistema (ex.: acontece só na sua máquina, com seus dados de verdade).

    O tipo de pessoa (PF/PJ) usado para formatar/zero-preencher `doc_novo` é
    sempre lido da coluna 'tipo_pessoa' do próprio cadastro na base — nunca
    inferido pela quantidade de dígitos. Se essa coluna estiver vazia na
    linha encontrada, a função recusa alterar o documento (para não gravar
    um zero-padding errado) e pede para preencher tipo_pessoa antes.

    Retorna um ResultadoRegularizacao.
    """
    log = _DebugLogger(arquivo_log_debug)
    log.log(f"regularizar_fornecedor: doc_antigo={doc_antigo} doc_novo={doc_novo} dry_run={dry_run}")

    doc_antigo_d = _digitos(doc_antigo)
    doc_novo_d = _digitos(doc_novo) if doc_novo else None

    resultado = ResultadoRegularizacao(
        doc_antigo=doc_antigo_d, doc_novo=doc_novo_d,
        nome_novo=nome_novo, razao_novo=razao_novo, dry_run=dry_run,
    )

    if not (doc_novo or nome_novo or razao_novo):
        resultado.erro = "Nada para alterar: informe doc_novo e/ou nome_novo/razao_novo."
        return resultado

    # --- 1) Base mestre: localizar a linha e ler tipo_pessoa (fonte da verdade) ---
    log.log(f"Abrindo base mestre (modo completo): {base_path}")
    wb = load_workbook(base_path, data_only=False)
    log.log("Base mestre aberta.")
    ws = wb[BASE_SHEET]
    linha_encontrada = None
    for row in ws.iter_rows(min_row=2):
        if _digitos(row[BASE_COL_DOC - 1].value) == doc_antigo_d:
            linha_encontrada = row
            break

    if linha_encontrada is None:
        resultado.erro = f"Documento {_formatar_doc(doc_antigo_d)} não encontrado na base mestre."
        wb.close()
        return resultado

    tipo_pessoa = _tipo_pessoa_da_linha([c.value for c in linha_encontrada])
    if doc_novo_d and tipo_pessoa is None:
        resultado.erro = (
            f"A linha de {_formatar_doc(doc_antigo_d)} não tem tipo_pessoa (PF/PJ) "
            f"preenchido na base. Preencha essa coluna antes de corrigir o documento "
            f"— sem ela não é seguro decidir se são 11 ou 14 dígitos."
        )
        wb.close()
        return resultado

    if doc_novo_d:
        doc_novo_d = doc_novo_d.zfill(11 if tipo_pessoa == "PF" else 14)

    # --- Guarda contra documento estruturalmente inválido ------------------
    # Roda ANTES da checagem de fusão (mais barata: falha sem precisar
    # varrer a base inteira de novo) e vale tanto para dry_run quanto para
    # a aplicação real — a pré-visualização já deve acusar o problema, sem
    # deixar o usuário chegar até "Confirmar e Aplicar" achando que estava
    # tudo certo.
    if doc_novo_d and not validar_documento(doc_novo_d, tipo_pessoa):
        tipo_doc = "CPF" if tipo_pessoa == "PF" else "CNPJ"
        resultado.erro = (
            f"O documento novo ({_formatar_doc(doc_novo_d, tipo_pessoa)}) não é um "
            f"{tipo_doc} estruturalmente válido — o dígito verificador não confere. "
            f"Confira a digitação antes de prosseguir. (Isso NÃO confirma que o "
            f"documento existe de fato na Receita Federal, só que os dígitos são "
            f"matematicamente possíveis.)"
        )
        wb.close()
        return resultado

    # --- Guarda contra fusão acidental de dois fornecedores distintos -----
    if doc_novo_d and _doc_ja_pertence_a_outro_fornecedor(base_path, doc_novo_d, doc_antigo_d):
        resultado.erro = (
            f"O documento novo ({_formatar_doc(doc_novo_d, tipo_pessoa)}) já pertence a "
            f"OUTRO fornecedor cadastrado na base. Isso seria uma fusão de cadastros, "
            f"não uma correção de digitação — trate com fundir_fornecedores()."
        )
        wb.close()
        return resultado
    resultado.doc_novo = doc_novo_d

    row = linha_encontrada
    cel_doc = row[BASE_COL_DOC - 1]
    if doc_novo_d:
        resultado.alteracoes.append(Alteracao(
            os.path.basename(base_path), BASE_SHEET, cel_doc.coordinate,
            "CNPJ/CPF", str(cel_doc.value), doc_novo_d))
        if not dry_run:
            cel_doc.value = doc_novo_d
    if nome_novo:
        cel_nome = row[BASE_COL_NOME - 1]
        resultado.alteracoes.append(Alteracao(
            os.path.basename(base_path), BASE_SHEET, cel_nome.coordinate,
            "NOME", str(cel_nome.value), nome_novo.upper()))
        if not dry_run:
            cel_nome.value = nome_novo.upper()
    if razao_novo:
        cel_razao = row[BASE_COL_RAZAO - 1]
        resultado.alteracoes.append(Alteracao(
            os.path.basename(base_path), BASE_SHEET, cel_razao.coordinate,
            "RAZAO_SOCIAL", str(cel_razao.value), razao_novo.upper()))
        if not dry_run:
            cel_razao.value = razao_novo.upper()

    if not dry_run:
        log.log(f"Salvando base mestre: {base_path}")
        try:
            wb.save(base_path)
            log.log("Base mestre salva com sucesso.")
        except PermissionError:
            log.log("PermissionError ao salvar base mestre.")
            resultado.arquivos_com_falha.append(base_path)
            resultado.erro = "Base mestre está aberta em outro programa. Feche e rode novamente."
            wb.close()
            return resultado
    wb.close()

    # --- 2) Arquivos de clientes ----------------------------------------
    if arquivos_cliente is None:
        arquivos_cliente = [
            f for f in glob.glob(os.path.join(pasta_clientes, "*.xlsx"))
            if os.path.basename(f).lower() not in ARQUIVOS_IGNORAR
            and not os.path.basename(f).startswith("~$")
        ]
    log.log(f"Iniciando varredura de {len(arquivos_cliente)} arquivo(s) de cliente.")

    _propagar_documento_em_clientes(
        arquivos_cliente, doc_antigo_d, doc_novo_d, nome_novo, razao_novo,
        tipo_pessoa, dry_run, resultado, callback_progresso, log)

    log.log("regularizar_fornecedor: concluído.")
    return resultado


def _arquivo_contem_documento(caminho, doc_d, mapas):
    """Passo 1 (rápido): abre o arquivo em modo streaming (read_only=True) só
    para checar se o documento aparece em alguma das colunas mapeadas. Não
    carrega o arquivo inteiro em memória — é o que evita o travamento comum
    do openpyxl em planilhas antigas com 'área usada' inflada (formatação
    aplicada muito além dos dados reais, às vezes até a última linha da
    planilha). Retorna (encontrado: bool, erro: str|None)."""
    try:
        wb = load_workbook(caminho, read_only=True, data_only=True)
    except Exception as e:
        return None, str(e)
    encontrado = False
    try:
        for mapa in mapas:
            if mapa["aba"] not in wb.sheetnames:
                continue
            ws = wb[mapa["aba"]]
            idx_doc = mapa["col_doc"] - 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                if idx_doc < len(row) and _digitos(row[idx_doc]) == doc_d:
                    encontrado = True
                    break
            if encontrado:
                break
    finally:
        wb.close()
    return encontrado, None


def _propagar_documento_em_clientes(arquivos_cliente, doc_antigo_d, doc_novo_d,
                                     nome_novo, razao_novo, tipo_pessoa,
                                     dry_run, resultado, callback_progresso=None,
                                     log=None):
    """Função compartilhada por regularizar_fornecedor() e fundir_fornecedores():
    varre os arquivos de cliente informados e troca doc_antigo_d -> doc_novo_d
    (e opcionalmente nome/razão social, cada um na coluna semanticamente
    correspondente) em todas as abas mapeadas em CLIENTE_MAPA_ABAS.
    `tipo_pessoa` ('PF'/'PJ', lido da base) garante o zero-padding correto do
    documento em todo lugar, sem depender de contar dígitos.

    Cada arquivo passa primeiro por uma checagem rápida em modo streaming
    (`_arquivo_contem_documento`) — só é reaberto no modo de edição (mais
    lento) quando de fato contém o documento procurado. Na prática, a
    esmagadora maioria dos arquivos de uma pasta de clientes não tem nada a
    ver com o fornecedor sendo corrigido, então isso evita abrir todos eles
    no modo pesado.

    `callback_progresso`, se informado, é chamado como
    `callback_progresso(indice, total, nome_arquivo)` antes de processar
    cada arquivo — use para atualizar uma barra/label de progresso na GUI,
    ou para logar em que arquivo o processo está (essencial para diagnosticar
    se ele travar em algum arquivo específico). `log`, se informado
    (instância de _DebugLogger), grava timestamp de cada sub-etapa.

    Escreve as alterações em `resultado` (Alteracao / arquivos_com_falha)."""
    if log is None:
        log = _DebugLogger(None)  # no-op

    total = len(arquivos_cliente)
    for indice, caminho in enumerate(arquivos_cliente, start=1):
        nome_arquivo = os.path.basename(caminho)
        if callback_progresso:
            callback_progresso(indice, total, nome_arquivo)

        log.log(f"[{indice}/{total}] Checagem rápida (read_only): {nome_arquivo}")
        encontrado, erro = _arquivo_contem_documento(caminho, doc_antigo_d, CLIENTE_MAPA_ABAS)
        log.log(f"[{indice}/{total}] Checagem rápida concluída: encontrado={encontrado} erro={erro}")
        if erro:
            resultado.arquivos_com_falha.append(f"{caminho} (erro ao abrir: {erro})")
            continue
        if not encontrado:
            continue

        log.log(f"[{indice}/{total}] Reabrindo em modo completo (edição): {nome_arquivo}")
        try:
            wbc = load_workbook(caminho, data_only=False)
        except Exception as e:
            log.log(f"[{indice}/{total}] Erro ao reabrir: {e}")
            resultado.arquivos_com_falha.append(f"{caminho} (erro ao abrir: {e})")
            continue
        log.log(f"[{indice}/{total}] Reaberto em modo completo com sucesso: {nome_arquivo}")

        alterou_algo = False
        for mapa in CLIENTE_MAPA_ABAS:
            if mapa["aba"] not in wbc.sheetnames:
                continue
            wsc = wbc[mapa["aba"]]
            for row in wsc.iter_rows(min_row=2):
                idx_doc = mapa["col_doc"] - 1
                idx_nome = mapa["col_nome"] - 1
                if idx_doc >= len(row):
                    continue
                cel_doc = row[idx_doc]
                if _digitos(cel_doc.value) != doc_antigo_d:
                    continue

                cel_nome = row[idx_nome] if idx_nome < len(row) else None

                if doc_novo_d:
                    novo_valor = (doc_novo_d if mapa["formato"] == "raw"
                                  else _formatar_doc(doc_novo_d, tipo_pessoa))
                    resultado.alteracoes.append(Alteracao(
                        os.path.basename(caminho), mapa["aba"], cel_doc.coordinate,
                        "CNPJ/CPF", str(cel_doc.value), novo_valor))
                    if not dry_run:
                        cel_doc.value = novo_valor
                    alterou_algo = True

                # Escolhe nome ou razão social conforme o que essa coluna
                # significa nessa aba (ex.: Contratos_ADM quer razão social).
                valor_nome = None
                if mapa.get("campo_nome") == "razao_social":
                    valor_nome = razao_novo or nome_novo
                else:
                    valor_nome = nome_novo or razao_novo

                if valor_nome and cel_nome is not None:
                    resultado.alteracoes.append(Alteracao(
                        os.path.basename(caminho), mapa["aba"], cel_nome.coordinate,
                        "NOME/RAZAO_SOCIAL", str(cel_nome.value), valor_nome.upper()))
                    if not dry_run:
                        cel_nome.value = valor_nome.upper()
                    alterou_algo = True

        if alterou_algo and not dry_run:
            log.log(f"[{indice}/{total}] Salvando: {nome_arquivo}")
            try:
                wbc.save(caminho)
                log.log(f"[{indice}/{total}] Salvo com sucesso: {nome_arquivo}")
            except PermissionError:
                log.log(f"[{indice}/{total}] PermissionError ao salvar: {nome_arquivo}")
                resultado.arquivos_com_falha.append(caminho)
        wbc.close()
        log.log(f"[{indice}/{total}] Arquivo fechado: {nome_arquivo}")


# ===========================================================================
# FUSÃO DE CADASTROS DUPLICADOS
# ===========================================================================
# Cenário: o mesmo fornecedor foi cadastrado DUAS vezes com documentos
# DIFERENTES DE VERDADE — tipicamente um CPF criado (usar_cpf_criado_auto)
# usado enquanto não se tinha o CPF real, e depois um segundo cadastro com o
# CPF real e nome completo. Ao contrário de regularizar_fornecedor(), aqui
# não existe "o documento certo e o errado digitado" — existem dois registros
# válidos que representam a mesma pessoa. Por isso a escolha de qual vence é
# sempre explícita (parâmetro `doc_vencedor`), nunca automática.

import unicodedata

# Conectores de nome em português — não contam como evidência de match,
# porque aparecem em sobrenomes comuns demais (DOS SANTOS, DE OLIVEIRA...)
# e por si só não distinguem uma pessoa de outra.
CONECTORES_NOME = {"DE", "DA", "DO", "DAS", "DOS", "E"}


def _sem_acento(s):
    """Remove acentos/cedilha para comparação (SÉRGIO == SERGIO, JOSÉ == JOSE)."""
    nfkd = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _tokens_significativos(nome):
    """Tokens do nome, maiúsculos, sem acento, sem conectores (DE/DA/DOS...)
    e sem tokens de 1 letra só (iniciais soltas não provam nada)."""
    limpo = _sem_acento(str(nome)).upper()
    return [t for t in limpo.split() if t not in CONECTORES_NOME and len(t) > 1]


def _similaridade_nomes(nome_a, nome_b, tipo_a=None, tipo_b=None):
    """
    Critério principal (só para PESSOA FÍSICA — tipo_a == tipo_b == 'PF'):
    o PRIMEIRO nome tem que bater (ou o nome inteiro, ignorando acento, ser
    idêntico) — sobrenome/conector batendo sozinho (DOS, SANTOS, PEREIRA...)
    NÃO conta como evidência, porque são comuns demais e geram falso
    positivo em cascata.

    Para PESSOA JURÍDICA (tipo 'PJ') esse bônus de token NÃO é aplicado:
    nomes de empresa compartilham prefixo genérico ('CASA', 'BH',
    'COMERCIAL') e sufixo jurídico ('LTDA', 'ME') com frequência, sem que
    isso signifique a mesma empresa — usar isso como evidência forte gera
    uma explosão de falso positivo (testado: sem essa restrição, uma base
    de ~1000 fornecedores gerou mais de 3000 "candidatos"). PJ usa só a
    razão de similaridade textual simples (difflib), mais conservadora.

    Score:
      1.0   -> nomes idênticos após remover acento/cedilha
      0.90  -> (só PF) primeiro E último token batem
      0.70  -> (só PF) só o primeiro token bate
      senão -> razão de similaridade textual simples (difflib)
    """
    a = _sem_acento(str(nome_a)).upper().strip()
    b = _sem_acento(str(nome_b)).upper().strip()
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0

    ratio = difflib.SequenceMatcher(None, a, b).ratio()

    if tipo_a == "PF" and tipo_b == "PF":
        tok_a = _tokens_significativos(nome_a)
        tok_b = _tokens_significativos(nome_b)
        if tok_a and tok_b and tok_a[0] == tok_b[0]:
            if tok_a[-1] == tok_b[-1]:
                return max(ratio, 0.90)
            return max(ratio, 0.70)
    return ratio


def detectar_possiveis_duplicatas(base_path, limiar=0.6, incluir_todos_fornecedores=True):
    """
    Procura pares de cadastros na base que provavelmente são a mesma pessoa
    ou empresa com dois documentos diferentes. Não altera nada — só relata.

    Cobre dois cenários:
    1) CPF criado em uso (aba 'CPF', status USADO) vs. resto da base —
       o caso clássico de "usei CPF provisório, depois cadastrei com o
       CPF real e nome completo".
    2) (se `incluir_todos_fornecedores=True`, padrão) TODOS os pares de
       fornecedores ativos entre si — cobre o caso de dois documentos
       REAIS cadastrados por engano (erro de digitação em um deles, sem
       nenhum CPF criado envolvido). Mais lento numa base grande (é uma
       comparação de cada fornecedor com todos os outros), mas ainda leva
       segundos, não minutos — é uma ação sob demanda, não algo rodando
       toda hora.

    Retorna (candidatos, orfaos_no_registro_cpf). Cada candidato é único
    (não aparece duas vezes com os papéis trocados).
    """
    wb = load_workbook(base_path, data_only=True)

    ws_cpf = wb[CPF_CRIADO_SHEET]
    criados_em_uso = []
    for row in ws_cpf.iter_rows(min_row=3, values_only=True):
        if row[CPF_CRIADO_COL_STATUS - 1] == "USADO":
            doc = _digitos(row[CPF_CRIADO_COL_DOC - 1])
            nome = str(row[CPF_CRIADO_COL_NOME - 1] or "").strip()
            if doc and nome:
                criados_em_uso.append((doc, nome))

    ws_forn = wb[BASE_SHEET]
    fornecedores = []
    for row in ws_forn.iter_rows(min_row=2, values_only=True):
        if not row or not row[BASE_COL_DOC - 1]:
            continue
        status = str(row[16]).strip().upper() if len(row) > 16 and row[16] else "ATIVO"
        if status == "INATIVO":
            continue
        doc = _digitos(row[BASE_COL_DOC - 1])
        nome = str(row[BASE_COL_NOME - 1] or row[BASE_COL_RAZAO - 1] or "").strip()
        tipo = _tipo_pessoa_da_linha(row)
        if doc and nome:
            fornecedores.append((doc, nome, tipo))
    wb.close()

    docs_criados = {d for d, _ in criados_em_uso}
    docs_fornecedores = {d for d, _, _ in fornecedores}
    tipo_por_doc = {d: t for d, _, t in fornecedores}

    orfaos_no_registro_cpf = [
        {"doc": d, "nome_no_registro": n}
        for d, n in criados_em_uso if d not in docs_fornecedores
    ]

    candidatos_por_par = {}  # frozenset({doc_a, doc_b}) -> dict (evita par invertido duplicado)

    def _registrar(doc_a, nome_a, tipo_a, doc_b, nome_b, tipo_b):
        if doc_a == doc_b:
            return
        # Documentos de tipo_pessoa diferente (um PF, outro PJ) não podem
        # ser o mesmo cadastro — nem vale gastar tempo comparando o nome.
        if tipo_a and tipo_b and tipo_a != tipo_b:
            return
        score = _similaridade_nomes(nome_a, nome_b, tipo_a, tipo_b)
        if score < limiar:
            return
        chave = frozenset((doc_a, doc_b))
        existente = candidatos_por_par.get(chave)
        if existente and existente["score"] >= score:
            return
        # Mantém os nomes de campo 'provisorio'/'candidato' por compatibilidade
        # com quem já consome este retorno — não significa mais estritamente
        # "provisório", é só "lado A" / "lado B" do par suspeito.
        candidatos_por_par[chave] = {
            "doc_provisorio": doc_a, "nome_provisorio": nome_a,
            "doc_candidato": doc_b, "nome_candidato": nome_b,
            "candidato_tambem_e_cpf_criado": doc_b in docs_criados,
            "score": round(score, 3),
        }

    # 1) CPF criado em uso vs. resto da base
    for doc_prov, nome_prov in criados_em_uso:
        if doc_prov not in docs_fornecedores:
            continue
        tipo_prov = tipo_por_doc.get(doc_prov)
        for doc_forn, nome_forn, tipo_forn in fornecedores:
            _registrar(doc_prov, nome_prov, tipo_prov, doc_forn, nome_forn, tipo_forn)

    # 2) todos os fornecedores PESSOA FÍSICA entre si (cobre duplicata por
    # erro de digitação, sem nenhum CPF criado envolvido — ex.: dois CPFs
    # verdadeiros cadastrados, um deles com dígito errado). Restrito a PF de
    # propósito: nome de EMPRESA compartilha palavra genérica com frequência
    # ('COMERCIAL', 'MATERIAIS DE CONSTRUÇÃO', sufixo 'LTDA'...) sem que isso
    # signifique duplicata — testado: incluir PJ aqui gerou milhares de
    # falsos candidatos por coincidência de palavra comum no ramo de negócio.
    if incluir_todos_fornecedores:
        pessoas_fisicas = [f for f in fornecedores if f[2] == "PF"]
        for i, (doc_a, nome_a, tipo_a) in enumerate(pessoas_fisicas):
            for doc_b, nome_b, tipo_b in pessoas_fisicas[i + 1:]:
                _registrar(doc_a, nome_a, tipo_a, doc_b, nome_b, tipo_b)

    candidatos = sorted(candidatos_por_par.values(), key=lambda c: -c["score"])
    return candidatos, orfaos_no_registro_cpf


def _liberar_cpf_criado(wb_base, doc_digitos):
    """Marca um CPF criado como DISPONIVEL novamente (libera para reuso),
    dentro de um workbook Base_Fornecedores já aberto. Não faz nada se o
    documento informado não estiver cadastrado na aba CPF (ou seja, é seguro
    chamar mesmo quando o documento perdedor era um CPF real, não gerado)."""
    if CPF_CRIADO_SHEET not in wb_base.sheetnames:
        return False
    ws = wb_base[CPF_CRIADO_SHEET]
    for row in ws.iter_rows(min_row=3):
        cel_doc = row[CPF_CRIADO_COL_DOC - 1]
        if _digitos(cel_doc.value) == doc_digitos:
            row[CPF_CRIADO_COL_STATUS - 1].value = "DISPONIVEL"
            row[CPF_CRIADO_COL_NOME - 1].value = None
            row[CPF_CRIADO_COL_DATA - 1].value = None
            return True
    return False


def fundir_fornecedores(base_path, pasta_clientes, doc_vencedor, docs_perdedores,
                         campos_vencedor=None, liberar_cpfs_criados=True,
                         ignorar_divergencia_tipo_pessoa=False,
                         dry_run=True, arquivos_cliente=None,
                         callback_progresso=None, arquivo_log_debug=None):
    """
    Mescla um ou mais cadastros duplicados (`docs_perdedores`) no cadastro
    escolhido como definitivo (`doc_vencedor`). Os dois já precisam existir
    na base — esta função não cria fornecedor novo.

    campos_vencedor: dict opcional {campo: valor} para forçar o valor final
        de um campo específico (ex.: {'nome': 'VALDECI VIEGAS DE AMORIM'}
        se você quiser o nome completo mesmo que o cadastro vencedor tenha
        um nome mais curto). Campos não listados aqui seguem a regra:
        valor do vencedor se não for vazio; senão, primeiro valor não-vazio
        encontrado entre os perdedores (na ordem informada).

    liberar_cpfs_criados: se True (padrão), qualquer doc perdedor que esteja
        cadastrado como CPF criado (aba 'CPF', status USADO) volta a ficar
        DISPONIVEL para reuso depois da fusão. Não afeta documentos que não
        estejam nessa aba.

    arquivo_log_debug: mesmo mecanismo de regularizar_fornecedor() — grava
        timestamp de cada etapa demorada, sobrevive a travamento/kill.

    O tipo_pessoa (PF/PJ) do registro vencedor (lido da base, nunca inferido
    por tamanho) é o que vale para formatar o documento propagado. Se algum
    perdedor tiver tipo_pessoa diferente do vencedor, a função recusa por
    padrão (PF mesclando com PJ quase sempre é engano de escolha de
    documento, não duplicata de cadastro) — passe
    ignorar_divergencia_tipo_pessoa=True só se tiver certeza.

    O(s) cadastro(s) perdedor(es) são marcados como INATIVO na base (não
    excluídos), para manter rastro de auditoria — quem consultar a lista de
    inativos verá o documento antigo ainda ali, apontando (via log) para o
    documento vencedor.
    """
    log = _DebugLogger(arquivo_log_debug)
    log.log(f"fundir_fornecedores: doc_vencedor={doc_vencedor} docs_perdedores={docs_perdedores} dry_run={dry_run}")
    doc_vencedor_d = _digitos(doc_vencedor)
    docs_perdedores_d = [_digitos(d) for d in docs_perdedores]
    campos_vencedor = campos_vencedor or {}

    resultado = ResultadoRegularizacao(
        doc_antigo="+".join(docs_perdedores_d), doc_novo=doc_vencedor_d,
        nome_novo=campos_vencedor.get("nome"), razao_novo=campos_vencedor.get("razao_social"),
        dry_run=dry_run,
    )

    if doc_vencedor_d in docs_perdedores_d:
        resultado.erro = "doc_vencedor não pode também aparecer em docs_perdedores."
        return resultado

    COLUNAS = ['cnpj_cpf', 'tipo_pessoa', 'razao_social', 'nome', 'telefone', 'email',
               'banco', 'op', 'agencia', 'conta', 'chave_pix', 'categoria',
               'especificacao', 'vinculo', 'dados_bancarios', 'endereco',
               'status', 'responsavel']

    wb = load_workbook(base_path, data_only=False)
    ws = wb[BASE_SHEET]

    def _ler_linha(doc_d):
        for row in ws.iter_rows(min_row=2):
            if _digitos(row[BASE_COL_DOC - 1].value) == doc_d:
                return row
        return None

    linha_vencedor = _ler_linha(doc_vencedor_d)
    if linha_vencedor is None:
        resultado.erro = f"doc_vencedor {_formatar_doc(doc_vencedor_d)} não encontrado na base."
        wb.close()
        return resultado

    tipo_pessoa_vencedor = _tipo_pessoa_da_linha([c.value for c in linha_vencedor])
    if tipo_pessoa_vencedor is None:
        resultado.erro = (
            f"O cadastro vencedor ({_formatar_doc(doc_vencedor_d)}) não tem "
            f"tipo_pessoa (PF/PJ) preenchido na base. Preencha antes de mesclar."
        )
        wb.close()
        return resultado

    linhas_perdedoras = []
    for d in docs_perdedores_d:
        linha = _ler_linha(d)
        if linha is None:
            resultado.erro = f"doc_perdedor {_formatar_doc(d)} não encontrado na base."
            wb.close()
            return resultado
        tipo_p = _tipo_pessoa_da_linha([c.value for c in linha])
        if tipo_p and tipo_p != tipo_pessoa_vencedor and not ignorar_divergencia_tipo_pessoa:
            resultado.erro = (
                f"doc_perdedor {_formatar_doc(d, tipo_p)} é {tipo_p}, mas o vencedor "
                f"{_formatar_doc(doc_vencedor_d, tipo_pessoa_vencedor)} é {tipo_pessoa_vencedor}. "
                f"Isso não parece duplicata de cadastro — confira antes de forçar com "
                f"ignorar_divergencia_tipo_pessoa=True."
            )
            wb.close()
            return resultado
        linhas_perdedoras.append((d, linha))

    doc_vencedor_d = doc_vencedor_d.zfill(11 if tipo_pessoa_vencedor == "PF" else 14)
    resultado.doc_novo = doc_vencedor_d

    # --- Reconciliar campos no registro vencedor -------------------------
    for i, campo in enumerate(COLUNAS):
        if campo in ("cnpj_cpf", "status"):
            continue
        cel_vencedor = linha_vencedor[i]
        valor_final = campos_vencedor.get(campo)
        if valor_final is None:
            valor_final = cel_vencedor.value
            if not str(valor_final or "").strip():
                for _, linha_p in linhas_perdedoras:
                    v = linha_p[i].value
                    if str(v or "").strip():
                        valor_final = v
                        break
        if str(valor_final or "").strip().upper() != str(cel_vencedor.value or "").strip().upper():
            resultado.alteracoes.append(Alteracao(
                os.path.basename(base_path), BASE_SHEET, cel_vencedor.coordinate,
                campo.upper(), str(cel_vencedor.value), str(valor_final)))
            if not dry_run:
                cel_vencedor.value = valor_final

    nome_final = campos_vencedor.get("nome") or linha_vencedor[COLUNAS.index("nome")].value
    razao_final = campos_vencedor.get("razao_social") or linha_vencedor[COLUNAS.index("razao_social")].value

    # --- Inativar registros perdedores ------------------------------------
    idx_status = COLUNAS.index("status")
    for d, linha_p in linhas_perdedoras:
        cel_status = linha_p[idx_status]
        resultado.alteracoes.append(Alteracao(
            os.path.basename(base_path), BASE_SHEET, cel_status.coordinate,
            "STATUS", str(cel_status.value), f"INATIVO (mesclado -> {_formatar_doc(doc_vencedor_d, tipo_pessoa_vencedor)})"))
        if not dry_run:
            cel_status.value = "INATIVO"

    # --- Liberar CPFs criados ---------------------------------------------
    if liberar_cpfs_criados:
        for d, _ in linhas_perdedoras:
            liberado = _liberar_cpf_criado(wb, d) if not dry_run else (
                CPF_CRIADO_SHEET in wb.sheetnames and any(
                    _digitos(r[CPF_CRIADO_COL_DOC - 1].value) == d
                    for r in wb[CPF_CRIADO_SHEET].iter_rows(min_row=3)
                )
            )
            if liberado:
                resultado.alteracoes.append(Alteracao(
                    os.path.basename(base_path), CPF_CRIADO_SHEET, "-", "STATUS",
                    "USADO", f"DISPONIVEL (liberado, mesclado -> {_formatar_doc(doc_vencedor_d, tipo_pessoa_vencedor)})"))

    if not dry_run:
        log.log(f"Salvando base mestre: {base_path}")
        try:
            wb.save(base_path)
            log.log("Base mestre salva com sucesso.")
        except PermissionError:
            log.log("PermissionError ao salvar base mestre.")
            resultado.arquivos_com_falha.append(base_path)
            resultado.erro = "Base mestre está aberta em outro programa. Feche e rode novamente."
            wb.close()
            return resultado
    wb.close()

    # --- Propagar nos clientes: cada doc perdedor -> doc vencedor ---------
    if arquivos_cliente is None:
        arquivos_cliente = [
            f for f in glob.glob(os.path.join(pasta_clientes, "*.xlsx"))
            if os.path.basename(f).lower() not in ARQUIVOS_IGNORAR
            and not os.path.basename(f).startswith("~$")
        ]

    for d in docs_perdedores_d:
        log.log(f"Propagando doc perdedor {d} -> vencedor {doc_vencedor_d}")
        _propagar_documento_em_clientes(
            arquivos_cliente, d, doc_vencedor_d, nome_final, razao_final,
            tipo_pessoa_vencedor, dry_run, resultado, callback_progresso, log)

    log.log("fundir_fornecedores: concluído.")
    return resultado
