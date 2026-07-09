"""
Gerenciamento de CPFs fictícios criados automaticamente para fornecedores
pessoa física sem CPF cadastrado (ex.: mão de obra avulsa, diaristas).

Extraído de Sistema_Entrada_Dados.py em [DATA_DA_EXTRACAO].
Nenhuma alteração de lógica foi feita nesta extração — apenas mudança
de localização e ajuste de imports.
"""
import random
from datetime import datetime
import logging

from openpyxl import load_workbook

from src.config.config import ARQUIVO_FORNECEDORES

# Mesmo logger usado no restante do sistema (Sistema_Entrada_Dados.py).
# Não definimos nível aqui: ele herda o nível efetivo do logger raiz,
# configurado uma única vez no arquivo principal (logging.basicConfig).
logger = logging.getLogger("sistema")


class GerenciadorCPFsCriados:
    def __init__(self):
        self.arquivo_fornecedores = ARQUIVO_FORNECEDORES

    def gerar_cpf_valido(self):
        """Gera um CPF válido seguindo EXATAMENTE o algoritmo oficial"""
        # Gerar os 9 primeiros dígitos (evitar sequências óbvias)
        while True:
            cpf = [random.randint(0, 9) for _ in range(9)]

            # Evitar CPFs com todos os dígitos iguais (000.000.000, 111.111.111, etc.)
            if len(set(cpf)) > 1:
                break

        # Calcular PRIMEIRO dígito verificador
        soma = 0
        for i in range(9):
            soma += cpf[i] * (10 - i)

        resto = soma % 11
        if resto < 2:
            primeiro_digito = 0
        else:
            primeiro_digito = 11 - resto

        cpf.append(primeiro_digito)

        # Calcular SEGUNDO dígito verificador
        soma = 0
        for i in range(10):
            soma += cpf[i] * (11 - i)

        resto = soma % 11
        if resto < 2:
            segundo_digito = 0
        else:
            segundo_digito = 11 - resto

        cpf.append(segundo_digito)

        return ''.join(map(str, cpf))

    def validar_cpf_gerado(self, cpf):
        """Valida se o CPF gerado está correto"""
        if len(cpf) != 11:
            return False

        # Verificar se não são todos iguais
        if cpf == cpf[0] * 11:
            return False

        # Calcular primeiro dígito
        soma = 0
        for i in range(9):
            soma += int(cpf[i]) * (10 - i)
        resto = soma % 11
        digito1 = 0 if resto < 2 else 11 - resto

        if int(cpf[9]) != digito1:
            return False

        # Calcular segundo dígito
        soma = 0
        for i in range(10):
            soma += int(cpf[i]) * (11 - i)
        resto = soma % 11
        digito2 = 0 if resto < 2 else 11 - resto

        return int(cpf[10]) == digito2

    def obter_proximo_cpf_disponivel(self):
        """Busca o próximo CPF disponível na aba CPF"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)

            # Verificar se a aba CPF existe
            if 'CPF' not in wb.sheetnames:
                logger.debug("Criando aba CPF...")
                # Criar a aba CPF se não existir
                ws_cpf = wb.create_sheet('CPF')
                ws_cpf.cell(row=1, column=1, value='CPF_CRIADO')
                ws_cpf.cell(row=1, column=2, value='STATUS')
                ws_cpf.cell(row=1, column=3, value='USADO_POR')
                ws_cpf.cell(row=1, column=4, value='DATA_USO')
                wb.save(self.arquivo_fornecedores)
            else:
                ws_cpf = wb['CPF']

            # Buscar primeiro CPF disponível
            cpf_disponivel = None
            linha_disponivel = None

            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                status = ws_cpf.cell(row=row, column=2).value

                if cpf_valor and (not status or status == 'DISPONIVEL'):
                    # Validar se o CPF é realmente válido
                    if self.validar_cpf_gerado(str(cpf_valor)):
                        cpf_disponivel = str(cpf_valor)
                        linha_disponivel = row
                        logger.debug(f"CPF disponível encontrado: {cpf_disponivel}")
                        break
                    else:
                        logger.debug(f"CPF inválido encontrado na planilha: {cpf_valor}, removendo...")
                        # Marcar como inválido
                        ws_cpf.cell(row=row, column=2, value='INVALIDO')

            # Se não encontrou nenhum disponível, gerar novos
            if not cpf_disponivel:
                logger.debug("Gerando novos CPFs...")
                # Gerar 20 novos CPFs válidos
                cpfs_gerados = 0
                tentativas = 0
                max_tentativas = 100

                while cpfs_gerados < 20 and tentativas < max_tentativas:
                    tentativas += 1
                    novo_cpf = self.gerar_cpf_valido()

                    # Validar o CPF gerado
                    if self.validar_cpf_gerado(novo_cpf):
                        # Verificar se já existe
                        if not self.cpf_ja_existe(ws_cpf, novo_cpf):
                            proxima_linha = ws_cpf.max_row + 1
                            ws_cpf.cell(row=proxima_linha, column=1, value=novo_cpf)
                            ws_cpf.cell(row=proxima_linha, column=2, value='DISPONIVEL')
                            cpfs_gerados += 1

                            logger.debug(f"CPF válido gerado: {novo_cpf}")

                            if not cpf_disponivel:  # Pegar o primeiro gerado
                                cpf_disponivel = novo_cpf
                                linha_disponivel = proxima_linha
                    else:
                        logger.debug(f"CPF inválido gerado (descartado): {novo_cpf}")

                if cpfs_gerados > 0:
                    wb.save(self.arquivo_fornecedores)
                    logger.debug(f"Total de CPFs válidos gerados: {cpfs_gerados}")
                else:
                    logger.debug("ERRO: Não foi possível gerar CPFs válidos")

            wb.close()

            if cpf_disponivel:
                logger.debug(f"Retornando CPF: {cpf_disponivel}")
                # Validar uma última vez antes de retornar
                if self.validar_cpf_gerado(cpf_disponivel):
                    return cpf_disponivel, linha_disponivel
                else:
                    logger.debug(f"ERRO: CPF retornado é inválido: {cpf_disponivel}")
                    return None, None
            else:
                return None, None

        except Exception as e:
            logger.debug(f"Erro ao obter CPF disponível: {str(e)}")
            import traceback
            logger.debug(traceback.format_exc())
            return None, None

    def cpf_ja_existe(self, worksheet, cpf):
        """Verifica se o CPF já existe na planilha"""
        for row in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row, column=1).value) == str(cpf):
                return True
        return False

    def marcar_cpf_como_usado(self, cpf, nome_fornecedor):
        """Marca um CPF como usado"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)
            ws_cpf = wb['CPF']

            for row in range(2, ws_cpf.max_row + 1):
                if str(ws_cpf.cell(row=row, column=1).value) == str(cpf):
                    ws_cpf.cell(row=row, column=2, value='USADO')
                    ws_cpf.cell(row=row, column=3, value=nome_fornecedor)
                    ws_cpf.cell(row=row, column=4, value=datetime.now().strftime('%d/%m/%Y %H:%M'))
                    break

            wb.save(self.arquivo_fornecedores)
            wb.close()
            return True

        except Exception as e:
            logger.debug(f"Erro ao marcar CPF como usado: {str(e)}")
            return False

    def listar_cpfs_disponiveis(self):
        """Lista todos os CPFs disponíveis"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)

            if 'CPF' not in wb.sheetnames:
                wb.close()
                return []

            ws_cpf = wb['CPF']
            cpfs_disponiveis = []

            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                status = ws_cpf.cell(row=row, column=2).value

                if cpf_valor and (not status or status == 'DISPONIVEL'):
                    # Validar antes de adicionar à lista
                    if self.validar_cpf_gerado(str(cpf_valor)):
                        cpfs_disponiveis.append(str(cpf_valor))

            wb.close()
            return cpfs_disponiveis

        except Exception as e:
            logger.debug(f"Erro ao listar CPFs disponíveis: {str(e)}")
            return []

    def marcar_cpf_como_disponivel(self, cpf):
        """Marca um CPF como disponível novamente na aba CPF"""
        try:
            wb = load_workbook(self.arquivo_fornecedores)

            if 'CPF' not in wb.sheetnames:
                wb.close()
                return False

            ws_cpf = wb['CPF']

            # Procurar o CPF e marcar como disponível
            for row in range(2, ws_cpf.max_row + 1):
                if str(ws_cpf.cell(row=row, column=1).value).strip() == str(cpf):
                    ws_cpf.cell(row=row, column=2, value='DISPONIVEL')  # Status
                    ws_cpf.cell(row=row, column=3, value='')  # Limpar nome do fornecedor
                    ws_cpf.cell(row=row, column=4, value='')  # Limpar data de uso
                    break

            wb.save(self.arquivo_fornecedores)
            wb.close()
            return True

        except Exception as e:
            logger.debug(f"Erro ao marcar CPF como disponível: {str(e)}")
            return False

    def listar_todos_cpfs_criados(self):
        """Lista todos os CPFs criados da aba CPF (disponíveis e usados)"""
        try:
            wb = load_workbook(self.arquivo_fornecedores, data_only=True)

            if 'CPF' not in wb.sheetnames:
                wb.close()
                return []

            ws_cpf = wb['CPF']
            cpfs = []

            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                if cpf_valor:  # Se tem CPF
                    cpfs.append(str(cpf_valor).strip())

            wb.close()
            return cpfs

        except Exception as e:
            logger.debug(f"Erro ao listar todos os CPFs criados: {str(e)}")
            return []

    def listar_cpfs_usados(self):
        """Lista apenas os CPFs que estão marcados como USADO na aba CPF"""
        try:
            wb = load_workbook(self.arquivo_fornecedores, data_only=True)

            if 'CPF' not in wb.sheetnames:
                wb.close()
                return []

            ws_cpf = wb['CPF']
            cpfs_usados = []

            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value
                status = ws_cpf.cell(row=row, column=2).value

                if cpf_valor and status and str(status).strip().upper() == 'USADO':
                    cpfs_usados.append(str(cpf_valor).strip())

            wb.close()
            return cpfs_usados

        except Exception as e:
            logger.debug(f"Erro ao listar CPFs usados: {str(e)}")
            return []

    def obter_detalhes_cpf_usado(self, cpf):
        """Obtém detalhes de um CPF usado (nome do fornecedor e data de uso)"""
        try:
            wb = load_workbook(self.arquivo_fornecedores, data_only=True)

            if 'CPF' not in wb.sheetnames:
                wb.close()
                return None

            ws_cpf = wb['CPF']

            for row in range(2, ws_cpf.max_row + 1):
                cpf_valor = ws_cpf.cell(row=row, column=1).value

                if str(cpf_valor).strip() == str(cpf):
                    status = ws_cpf.cell(row=row, column=2).value
                    usado_por = ws_cpf.cell(row=row, column=3).value
                    data_uso = ws_cpf.cell(row=row, column=4).value

                    wb.close()
                    return {
                        'status': str(status) if status else '',
                        'usado_por': str(usado_por) if usado_por else '',
                        'data_uso': str(data_uso) if data_uso else ''
                    }

            wb.close()
            return None

        except Exception as e:
            logger.debug(f"Erro ao obter detalhes do CPF: {str(e)}")
            return None
