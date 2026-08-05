"""
Cache em memória de fornecedores, usado para acelerar buscas repetidas
sem reabrir a planilha de fornecedores a cada consulta.

Extraído de Sistema_Entrada_Dados.py em [DATA_DA_EXTRACAO].
Nenhuma alteração de lógica foi feita nesta extração — apenas mudança
de localização e ajuste de imports.
"""
import logging
import os
from datetime import datetime

from openpyxl import load_workbook

# Mesmo logger usado no restante do sistema (Sistema_Entrada_Dados.py).
# Não definimos nível aqui: ele herda o nível efetivo do logger raiz,
# configurado uma única vez no arquivo principal (logging.basicConfig).
logger = logging.getLogger("sistema")


class CacheFornecedores:
    """Cache para otimizar buscas de fornecedores"""

    def __init__(self):
        self.cache_fornecedores = None
        self.cache_timestamp = None
        self.cache_duracao = 300  # 5 minutos

    def carregar_cache_se_necessario(self, arquivo_fornecedores):
        """Carrega cache se necessário ou se arquivo foi modificado"""
        try:
            agora = datetime.now()
            arquivo_modificado = os.path.getmtime(arquivo_fornecedores)

            # Verificar se precisa recarregar
            precisa_recarregar = (
                self.cache_fornecedores is None or
                self.cache_timestamp is None or
                (agora - self.cache_timestamp).seconds > self.cache_duracao or
                arquivo_modificado > self.cache_timestamp.timestamp()
            )

            if precisa_recarregar:
                logger.debug("DEBUG: Recarregando cache de fornecedores...")
                self.cache_fornecedores = self._carregar_fornecedores(arquivo_fornecedores)
                self.cache_timestamp = agora
                logger.debug(f"DEBUG: Cache carregado com {len(self.cache_fornecedores)} fornecedores")

            return self.cache_fornecedores

        except Exception as e:
            logger.debug(f"DEBUG: Erro ao carregar cache: {str(e)}")
            return []

    def _carregar_fornecedores(self, arquivo_fornecedores):
        """Carrega todos os fornecedores em memória"""
        fornecedores = []
    
        try:
            wb = load_workbook(arquivo_fornecedores, data_only=True)
            ws = wb['Fornecedores']
    
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[3]:  # Pular se não tem CNPJ ou nome
                    continue
    
                # Preenche com None até ter 18 colunas, para não estourar índice
                # em planilhas antigas que ainda não têm todas as colunas novas
                row_pad = list(row) + [None] * (18 - len(row))
    
                fornecedor = {
                    'cnpj_cpf': str(row_pad[0]).strip() if row_pad[0] else '',
                    'cnpj_cpf_numeros': ''.join(filter(str.isdigit, str(row_pad[0]))) if row_pad[0] else '',  # NOVO
                    'tipo_pessoa': str(row_pad[1]).strip().upper() if row_pad[1] else '',  # NOVO
                    'razao_social': str(row_pad[2]).strip() if row_pad[2] else '',  # NOVO
                    'nome': str(row_pad[3]).strip().upper() if row_pad[3] else '',
                    'telefone': str(row_pad[4]).strip() if row_pad[4] else '',  # NOVO
                    'email': str(row_pad[5]).strip() if row_pad[5] else '',  # NOVO
                    'banco': str(row_pad[6]).strip() if row_pad[6] else '',  # ALTERADO: era row[4]
                    'op': str(row_pad[7]).strip() if row_pad[7] else '',  # ALTERADO: era row[5]
                    'agencia': str(row_pad[8]).strip() if row_pad[8] else '',  # ALTERADO: era row[6]
                    'conta': str(row_pad[9]).strip() if row_pad[9] else '',  # ALTERADO: era row[7]
                    'chave_pix': str(row_pad[10]).strip() if row_pad[10] else '',  # ALTERADO: era row[8]
                    'categoria': str(row_pad[11] or '').strip(),
                    'especificacao': str(row_pad[12]).strip() if row_pad[12] else '',  # NOVO
                    'vinculo': str(row_pad[13]).strip() if row_pad[13] else '',  # NOVO
                    'dados_bancarios': str(row_pad[14]).strip() if row_pad[14] else '',  # NOVO
                    'endereco': str(row_pad[15]).strip() if row_pad[15] else '',  # NOVO
                    'status': str(row_pad[16]).strip().upper() if row_pad[16] else 'ATIVO',  # NOVO
                    'responsavel': str(row_pad[17]).strip() if row_pad[17] else '',  # NOVO
                }
    
                fornecedores.append(fornecedor)
    
            wb.close()
    
        except Exception as e:
            logger.debug(f"DEBUG: Erro ao carregar fornecedores: {str(e)}")
    
        return fornecedores