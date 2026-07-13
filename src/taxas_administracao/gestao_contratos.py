"""
Gestão de contratos de administração: cadastro, edição, exclusão de
contratos e seus administradores/gestores, incluindo os três métodos de
pagamento suportados (Percentual da Quinzena, Valor Fixo em Parcelas,
Eventos/Fases) e a geração do documento .docx do contrato.

Extraído de Sistema_Entrada_Dados.py em [DATA_DA_EXTRACAO].
Nenhuma alteração de LÓGICA foi feita nesta extração — apenas mudança de
localização, ajuste de imports, e uma correção de formatação (uma linha
que veio colada/sem quebra em salvar_contrato... ver nota abaixo).

ATENÇÃO — pendências conhecidas, não corrigidas nesta extração:
    1) Em criar_novo_contrato, existem DOIS pares de botões
       "Adicionar/Remover Administrador" (um com emoji + botão de copiar,
       outro sem emoji, chamando as mesmas funções). Provável resíduo de
       quando a função de copiar descrições foi adicionada. Vale conferir
       visualmente na tela "Novo Contrato" e remover o par redundante.
    2) GeradorContratoADM ainda não foi extraída (próxima da fila) — o
       import abaixo é tardio (dentro do método, não no topo do arquivo)
       para evitar import circular enquanto as duas migrações acontecem
       em momentos diferentes. Quando GeradorContratoADM for extraída,
       mover este import para o topo do arquivo.
"""
import logging
import os
import re
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from tkcalendar import DateEntry

from src.config.config import BASE_PATH, PASTA_CLIENTES, ARQUIVO_CLIENTES
from src.config.dialogs import custom_messagebox
from src.config.utils import (
    normalizar_documento,
    formatar_documento,
    buscar_dados_bancarios_fornecedor,
    buscar_fornecedor,
)

# Mesmo logger usado no restante do sistema (Sistema_Entrada_Dados.py).
logger = logging.getLogger("sistema")


class GestaoContratos:
    def __init__(self, parent):
        self.parent = parent
        self.arquivo_cliente = None
        self.cliente_atual = None

    def _obter_tipo_pessoa_da_base(self, cnpj_cpf_raw):
        """
        Busca o tipo_pessoa na base_fornecedores.xlsx

        Parâmetros:
        - cnpj_cpf_raw: CNPJ/CPF em qualquer formato

        Retorna:
        - 'PF' ou 'PJ' conforme cadastrado na base
        - None se não encontrar
        """
        try:
            # Normalizar para comparação (apenas números)
            apenas_numeros = ''.join(filter(str.isdigit, str(cnpj_cpf_raw)))

            # Ler base de fornecedores
            arquivo_fornecedores = BASE_PATH / 'base_fornecedores.xlsx'

            if not arquivo_fornecedores.exists():
                logger.debug(f"Arquivo base_fornecedores.xlsx não encontrado")
                return None

            df_fornecedores = pd.read_excel(arquivo_fornecedores)

            # Normalizar coluna CNPJ/CPF da base para comparação
            df_fornecedores['CNPJ_CPF_LIMPO'] = df_fornecedores['CNPJ/CPF'].astype(str).apply(
                lambda x: ''.join(filter(str.isdigit, x))
            )

            # Buscar o registro
            registro = df_fornecedores[df_fornecedores['CNPJ_CPF_LIMPO'] == apenas_numeros]

            if not registro.empty:
                tipo_pessoa = registro.iloc[0]['tipo_pessoa']
                logger.debug(f"Tipo pessoa encontrado na base: {tipo_pessoa} para {cnpj_cpf_raw}")
                return tipo_pessoa
            else:
                logger.debug(f"CNPJ/CPF {cnpj_cpf_raw} não encontrado na base")
                return None

        except Exception as e:
            logger.debug(f"Erro ao buscar tipo_pessoa na base: {e}")
            return None

    def _formatar_documento_admin(self, cnpj_cpf_raw, tipo_pessoa):
        """
        Helper simplificado - recebe tipo_pessoa como parâmetro

        Parâmetros:
        - cnpj_cpf_raw: Documento em qualquer formato
        - tipo_pessoa: 'PF' ou 'PJ' (vindo da base)

        Retorna:
        - Documento formatado corretamente
        """
        try:
            if not tipo_pessoa or tipo_pessoa not in ['PF', 'PJ']:
                raise ValueError(f"tipo_pessoa inválido: {tipo_pessoa}")

            cnpj_cpf = str(cnpj_cpf_raw).strip()

            # Normalizar COM o tipo da base
            cnpj_cpf_normalizado = normalizar_documento(cnpj_cpf, tipo_pessoa)

            # E formatar
            return formatar_documento(cnpj_cpf_normalizado, tipo_pessoa)

        except Exception as e:
            logger.debug(f"Erro ao formatar documento '{cnpj_cpf_raw}' como {tipo_pessoa}: {e}")
            raise

    def centralizar_janela(self, janela, largura=800, altura=600, parent=None):
        """
        Centraliza uma janela na tela ou relativa ao parent se fornecido.
        Também define o tamanho padrão da janela.
        """
        # Definir tamanho
        janela.geometry(f"{largura}x{altura}")

        # Atualizar a janela para garantir que as dimensões sejam aplicadas
        janela.update_idletasks()

        # Se parent for fornecido, centralize em relação a ele
        if parent and parent.winfo_exists():
            # Calcular o centro da janela pai
            x_parent = parent.winfo_x() + parent.winfo_width() // 2
            y_parent = parent.winfo_y() + parent.winfo_height() // 2

            # Calcular a posição da nova janela
            x = x_parent - largura // 2
            y = y_parent - altura // 2
        else:
            # Centralizar na tela
            x = (janela.winfo_screenwidth() // 2) - (largura // 2)
            y = (janela.winfo_screenheight() // 2) - (altura // 2)

        # Definir posição
        janela.geometry(f"{largura}x{altura}+{x}+{y}")

        # Tornar a janela modal (quando aplicável)
        if parent and hasattr(janela, 'transient') and hasattr(janela, 'grab_set'):
            janela.transient(parent)
            janela.grab_set()

        # Trazer para frente
        janela.lift()
        janela.focus_force()

    def criar_interface_contratos(self, janela, on_close_callback):
        """Cria a interface de gestão de contratos em uma janela já existente"""
        try:
            # Verificar se o arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                custom_messagebox("error", "Erro", f"Arquivo do cliente {self.cliente_atual} não encontrado!")
                on_close_callback()  # Fechar a janela em caso de erro
                return

            # Abrir arquivo e verificar aba
            wb = load_workbook(self.arquivo_cliente)
            if 'Contratos_ADM' not in wb.sheetnames:
                # Se não existir a aba, criar
                logger.debug(f"Criando aba Contratos_ADM para {self.cliente_atual}")
                ws = wb.create_sheet("Contratos_ADM")

                # Definir os blocos na linha 1
                blocos = ["CONTRATOS", "", "", "", "", "",
                        "ADMINISTRADORES_CONTRATO", "", "", "", "", "", "",
                        "ADITIVOS", "", "", "",
                        "ADMINISTRADORES_ADITIVO", "", "", "", "", "", "",
                         "PARCELAS", "", "", "", "", "", "", "", "", ""]

                for col, valor in enumerate(blocos, 1):
                    ws.cell(row=1, column=col, value=valor)

                # Definir cabeçalhos na linha 2
                headers = [
                    # CONTRATOS
                    "Nº Contrato", "Data Início", "Data Fim", "Status", "Observações", "",
                    # ADMINISTRADORES_CONTRATO
                    "Nº Contrato", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total", "Nº Parcelas",
                    # ADITIVOS
                    "Nº Contrato", "Nº Aditivo", "Data Início", "Data Fim",
                    # ADMINISTRADORES_ADITIVO
                    "Nº Contrato", "Nº Aditivo", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total",
                    # PARCELAS
                    "Referência", "Número", "CNPJ/CPF", "Nome", "Data Vencimento", "Valor", "Status", "Data Pagamento", "Eventos/Fases", "Percentual %"
                ]

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

                # Salvar as alterações
                wb.save(self.arquivo_cliente)

            # Frame principal
            frame_principal = ttk.Frame(janela, padding="10")
            frame_principal.pack(fill='both', expand=True)

            # Frame para lista de contratos existentes
            frame_contratos = ttk.LabelFrame(frame_principal, text="Contratos Existentes")
            frame_contratos.pack(fill='both', expand=True, pady=5)

            # Treeview para contratos
            colunas = ('Nº Contrato', 'Data Início', 'Data Fim', 'Status')
            self.tree_contratos = ttk.Treeview(frame_contratos, columns=colunas, show='headings')
            for col in colunas:
                self.tree_contratos.heading(col, text=col)
                self.tree_contratos.column(col, width=100)

            # Adicionar scrollbars
            scroll_y = ttk.Scrollbar(frame_contratos, orient='vertical', command=self.tree_contratos.yview)
            scroll_x = ttk.Scrollbar(frame_contratos, orient='horizontal', command=self.tree_contratos.xview)
            self.tree_contratos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

            self.tree_contratos.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y.pack(side='right', fill='y')
            scroll_x.pack(side='bottom', fill='x')

            # Frame para lista de administradores do contrato selecionado
            frame_admins = ttk.LabelFrame(frame_principal, text="Administradores do Contrato")
            frame_admins.pack(fill='both', expand=True, pady=5)

            # Treeview para administradores
            colunas_adm = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial Pagamento')
            self.tree_adm_contrato = ttk.Treeview(frame_admins, columns=colunas_adm, show='headings')
            for col in colunas_adm:
                self.tree_adm_contrato.heading(col, text=col)
                self.tree_adm_contrato.column(col, width=100)

            # Adicionar scrollbars para administradores
            scroll_y_adm = ttk.Scrollbar(frame_admins, orient='vertical', command=self.tree_adm_contrato.yview)
            scroll_x_adm = ttk.Scrollbar(frame_admins, orient='horizontal', command=self.tree_adm_contrato.xview)
            self.tree_adm_contrato.configure(yscrollcommand=scroll_y_adm.set, xscrollcommand=scroll_x_adm.set)

            self.tree_adm_contrato.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y_adm.pack(side='right', fill='y')
            scroll_x_adm.pack(side='bottom', fill='x')

            # Botões de ação
            frame_botoes = ttk.Frame(frame_principal)
            frame_botoes.pack(fill='x', pady=5)

            ttk.Button(frame_botoes, text="Novo Contrato",
                    command=lambda: self.criar_novo_contrato(janela)).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Editar Contrato",
                    command=self.editar_contrato).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Excluir Contrato",
                    command=self.excluir_contrato).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Gerar Contrato",
                   command=self.gerar_contrato_adm).pack(side='left', padx=5)

            # Botão Fechar com callback personalizado
            ttk.Button(frame_botoes, text="Fechar",
                    command=on_close_callback).pack(side='right', padx=5)

            # Carregar contratos existentes
            self.carregar_contratos()

            # Auto-selecionar o primeiro contrato para exibir seus administradores imediatamente
            primeiros = self.tree_contratos.get_children()
            if primeiros:
                self.tree_contratos.selection_set(primeiros[0])
                self.tree_contratos.focus(primeiros[0])
                self.mostrar_administradores()

            # Binding para atualizar administradores quando selecionar contrato
            self.tree_contratos.bind('<<TreeviewSelect>>', self.mostrar_administradores)

        except Exception as e:
            import traceback
            logger.debug(traceback.format_exc())
            custom_messagebox("error", "Erro", f"Erro ao abrir janela de contratos: {str(e)}")
            if 'wb' in locals():
                wb.close()
            # Garantir que a janela principal seja restaurada em caso de erro
            on_close_callback()

    def carregar_contratos(self):
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']

            for item in self.tree_contratos.get_children():
                self.tree_contratos.delete(item)

            contratos_processados = set()
            for row in ws.iter_rows(min_row=3, values_only=True):
                num_contrato = row[0]
                if num_contrato and num_contrato not in contratos_processados:
                    # Processar datas
                    data_inicio = ''
                    if row[1]:
                        try:
                            if isinstance(row[1], datetime):
                                data_inicio = row[1].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[1]), '%Y-%m-%d')
                                data_inicio = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_inicio = str(row[1])

                    data_fim = ''
                    if row[2]:
                        try:
                            if isinstance(row[2], datetime):
                                data_fim = row[2].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[2]), '%Y-%m-%d')
                                data_fim = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_fim = str(row[2])

                    self.tree_contratos.insert('', 'end', values=(
                        num_contrato,
                        data_inicio,
                        data_fim,
                        row[3] or ''
                    ))
                    contratos_processados.add(num_contrato)

            wb.close()

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar contratos: {str(e)}")

    def mostrar_administradores(self, event=None):
        """Mostra administradores do contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            return

        try:
            # Limpar lista atual
            for item in self.tree_adm_contrato.get_children():
                self.tree_adm_contrato.delete(item)

            num_contrato = str(self.tree_contratos.item(selecionado)['values'][0])

            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']

            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[6]) == num_contrato:  # Coluna G - Nº Contrato
                    if row[26]:  # Data Inicial de Pagamento
                        data_inicial = row[26].strftime('%d/%m/%Y') if isinstance(row[26], datetime) else str(row[26])
                    else:
                        data_inicial = ''

                    valor_percentual = row[10] or ''  # Coluna K
                    if valor_percentual:
                        valor_percentual_str = str(valor_percentual)
                        if '%' in valor_percentual_str:
                            # Já é percentual, manter como está
                            valor_perc_formatado = valor_percentual_str
                        else:
                            # É valor numérico, formatar como percentual
                            try:
                                perc_num = float(str(valor_percentual).replace(',', '.'))
                                valor_perc_formatado = f"{perc_num:.2f}%"
                            except:
                                valor_perc_formatado = str(valor_percentual)
                    else:
                        valor_perc_formatado = ''

                    # Formatar valor total
                    valor_total = row[11] or 0  # Coluna L
                    try:
                        valor_total_num = float(str(valor_total).replace(',', '.'))
                        valor_total_formatado = f"R$ {valor_total_num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    except:
                        valor_total_formatado = str(valor_total)

                    self.tree_adm_contrato.insert('', 'end', values=(
                        row[7],   # CNPJ/CPF
                        row[8],   # Nome
                        row[9],   # Tipo
                        valor_perc_formatado,
                        valor_total_formatado,
                        row[12],  # Nº Parcelas
                        data_inicial
                    ))

            if not self.tree_adm_contrato.get_children():
                self.tree_adm_contrato.insert('', 'end', values=(
                    '—', 'Nenhum administrador cadastrado para este contrato',
                    '', '', '', '', ''
                ))

            wb.close()

        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar administradores: {str(e)}")

    def gerar_numero_contrato(self, pasta_clientes: Path, tipo_pessoa: str = 'J') -> str:
        """
        Gera o próximo número de contrato no formato AAAA/NNNsufixo.

        Parâmetros
        ----------
        pasta_clientes : Path
            Caminho para PASTA_CLIENTES (onde ficam os .xlsx dos clientes).
        tipo_pessoa : str
            'PJ' ou 'J'  →  sufixo 'J'
            'PF' ou 'F'  →  sufixo 'F'

        Retorna
        -------
        str  ex.: '2025/001J'
        """
        sufixo = 'J' if tipo_pessoa.upper() in ('PJ', 'J') else 'F'
        ano_atual = datetime.now().year
        prefixo_ano = str(ano_atual)

        maior_seq = 0
        # Padrão: 4 dígitos de ano / 2-3 dígitos numéricos + letra
        padrao = re.compile(r'^(\d{4})/(\d{2,3})[A-Z]?$', re.IGNORECASE)

        for xlsx in pasta_clientes.glob('*.xlsx'):
            try:
                wb = load_workbook(xlsx, read_only=True, data_only=True)
                if 'Contratos_ADM' not in wb.sheetnames:
                    wb.close()
                    continue
                ws = wb['Contratos_ADM']
                for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
                    valor = row[0]
                    if not valor:
                        continue
                    m = padrao.match(str(valor).strip())
                    if m and m.group(1) == prefixo_ano:
                        seq = int(m.group(2))
                        if seq > maior_seq:
                            maior_seq = seq
                wb.close()
            except Exception:
                pass  # arquivo em uso ou corrompido — ignorar

        proximo = maior_seq + 1
        return f"{prefixo_ano}/{proximo:03d}{sufixo}"

    def criar_novo_contrato(self, janela_principal):
        """Abre janela para criar novo contrato com suporte a parcelas fixas ou eventos"""
        janela = tk.Toplevel(self.parent)
        janela.title(f"Novo Contrato - {self.cliente_atual}")
        janela.geometry("800x650")

        # Frame principal com scrollbar para garantir acesso a todos os campos
        main_frame = ttk.Frame(janela)
        main_frame.pack(fill='both', expand=True)

        # Adicionar canvas com scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame principal dentro do scrollable frame
        frame = ttk.Frame(scroll_frame, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame para dados do contrato
        frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
        frame_contrato.pack(fill='x', pady=5)

        # Nº Contrato – gerado automaticamente (somente leitura)
        ttk.Label(frame_contrato, text="Nº Contrato:", width=22).grid(
            row=0, column=0, padx=5, pady=5, sticky='w')
        num_contrato_var = tk.StringVar()
        num_contrato = ttk.Entry(frame_contrato, textvariable=num_contrato_var,
                                state='readonly', width=20)
        num_contrato.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Preenche automaticamente ao abrir — usa o tipo do 1º admin (default J)
        num_contrato_var.set(
            self.gerar_numero_contrato(
                Path(self.arquivo_cliente).parent,
                tipo_pessoa='J'
            )
        )

        numero = self.gerar_numero_contrato(PASTA_CLIENTES, tipo_pessoa='J')
        print(f"DEBUG numero gerado: '{numero}'")
        num_contrato_var.set(numero)

        # Datas
        ttk.Label(frame_contrato, text="Data Início:*", width=15).grid(row=1, column=0, padx=5, pady=5, sticky='w')
        data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_inicio.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        ttk.Label(frame_contrato, text="Data Fim:*", width=15).grid(row=2, column=0, padx=5, pady=5, sticky='w')
        data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_fim.grid(row=2, column=1, padx=5, pady=5, sticky='w')

        # Observações
        ttk.Label(frame_contrato, text="Observações:", width=15).grid(row=3, column=0, padx=5, pady=5, sticky='nw')
        observacoes = ttk.Entry(frame_contrato, width=25)
        observacoes.grid(row=3, column=1, padx=5, pady=5, sticky='ew')

        # Tipo de pagamento (metodo)
        ttk.Label(frame_contrato, text="Método de Pagamento:*", width=22).grid(
            row=4, column=0, padx=5, pady=5, sticky='w')
        metodo_pagamento = ttk.Combobox(frame_contrato, values=[
            "Percentual da Quinzena",
            "Valor Fixo em Parcelas",
            "Eventos/Fases"
        ], state='readonly', width=20)
        metodo_pagamento.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        metodo_pagamento.current(0)

        # ── Valor Global CONDICIONAL (oculto para Percentual da Quinzena) ─
        lbl_valor_global = ttk.Label(frame_contrato, text="Valor Global:*", width=22)
        lbl_valor_global.grid(row=5, column=0, padx=5, pady=5, sticky='w')
        valor_global = ttk.Entry(frame_contrato, width=20)
        valor_global.grid(row=5, column=1, padx=5, pady=5, sticky='w')

        def _on_metodo_change(event=None):
            """Oculta/exibe Valor Global conforme método selecionado."""
            if metodo_pagamento.get() == "Percentual da Quinzena":
                lbl_valor_global.grid_remove()
                valor_global.grid_remove()
                valor_global.delete(0, tk.END)
                valor_global.insert(0, "0")        # sentinela para salvar()
            else:
                lbl_valor_global.grid()
                valor_global.grid()
                if valor_global.get() == "0":
                    valor_global.delete(0, tk.END)

        metodo_pagamento.bind("<<ComboboxSelected>>", _on_metodo_change)
        # Disparar uma vez para definir estado inicial
        _on_metodo_change()

        # Frame para Administradores
        frame_adm = ttk.LabelFrame(frame, text="Administradores")
        frame_adm.pack(fill='both', expand=True, pady=5)

        # Lista de Administradores
        colunas = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial')
        self.tree_adm = ttk.Treeview(frame_adm, columns=colunas, show='headings', height=5)

        for col in colunas:
            self.tree_adm.heading(col, text=col)
            self.tree_adm.column(col, width=100)

        # Adicionar scrollbars
        scroll_y = ttk.Scrollbar(frame_adm, orient='vertical', command=self.tree_adm.yview)
        scroll_x = ttk.Scrollbar(frame_adm, orient='horizontal', command=self.tree_adm.xview)
        self.tree_adm.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        self.tree_adm.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para botões de administradores
        frame_botoes_adm = ttk.Frame(frame_adm)
        frame_botoes_adm.pack(fill='x', pady=5)

        # FUNCIONALIDADE: Copiar eventos entre gestores
        def copiar_eventos_entre_gestores():
            """Copia eventos de um gestor para outro(s) DURANTE a criação do contrato"""

            # Verificar se há gestores cadastrados
            if not self.tree_adm.get_children():
                custom_messagebox("warning", "Aviso", "Adicione pelo menos um gestor primeiro!")
                return

            metodo = metodo_pagamento.get()
            if metodo not in ["Eventos/Fases", "Valor Fixo em Parcelas"]:
                custom_messagebox("warning", "Aviso",
                                "Esta função funciona apenas para:\n" +
                                "• Eventos/Fases\n" +
                                "• Valor Fixo em Parcelas\n\n" +
                                f"Método atual: {metodo}")
                return

            # Criar janela de cópia
            janela_copia = tk.Toplevel(janela)
            janela_copia.title("Copiar Descrições Entre Gestores")
            janela_copia.geometry("700x750")
            janela_copia.transient(janela)
            janela_copia.grab_set()

            frame_copia = ttk.Frame(janela_copia, padding="10")
            frame_copia.pack(fill='both', expand=True)

            # Título
            ttk.Label(frame_copia,
                    text="Copiar Descrições de Parcelas/Eventos",
                    font=('Arial', 12, 'bold')).pack(pady=10)

            if metodo == "Eventos/Fases":
                msg_explicativa = "Esta ferramenta copia os eventos (com descrições e percentuais) de um gestor para outro(s)."
            else:
                msg_explicativa = "Esta ferramenta copia as descrições das parcelas de um gestor para outro(s)."

            ttk.Label(frame_copia,
                    text=msg_explicativa,
                    wraplength=650,
                    font=('Arial', 9, 'italic')).pack(pady=5)

            # ETAPA 1: Selecionar gestor de ORIGEM
            frame_origem = ttk.LabelFrame(frame_copia, text="1. Gestor de Origem (copiar DE)")
            frame_origem.pack(fill='x', pady=10, padx=10)

            if metodo == "Eventos/Fases":
                label_origem = "Selecione o gestor que já possui os eventos configurados:"
            else:
                label_origem = "Selecione o gestor que já possui as descrições das parcelas configuradas:"

            ttk.Label(frame_origem, text=label_origem).pack(anchor='w', padx=10, pady=5)

            # Listar gestores com descrições
            gestores_com_descricoes = []
            gestores_info = {}

            for item in self.tree_adm.get_children():
                valores = self.tree_adm.item(item)['values']
                tags = self.tree_adm.item(item)['tags']

                cnpj_cpf = valores[0]
                nome = valores[1]

                tem_descricoes = False
                descricoes_info = {}

                if metodo == "Eventos/Fases":
                    for tag in tags:
                        if tag.startswith('eventos:'):
                            tem_descricoes = True
                            eventos_str = tag.replace('eventos:', '')
                            eventos_parts = eventos_str.split('|')

                            eventos_list = []
                            for evento_str in eventos_parts:
                                partes = evento_str.split(':')
                                if len(partes) == 3:
                                    eventos_list.append({
                                        'descricao': partes[0],
                                        'percentual': float(partes[1]),
                                        'valor': float(partes[2])
                                    })

                            descricoes_info = {
                                'tipo': 'eventos',
                                'dados': eventos_list
                            }
                            break
                else:  # Valor Fixo em Parcelas
                    DELIMITADOR = "|||"
                    for tag in tags:
                        if tag.startswith('descricoes:'):
                            tem_descricoes = True
                            descricoes_str = tag.replace('descricoes:', '')
                            descricoes_list = descricoes_str.split(DELIMITADOR)

                            descricoes_info = {
                                'tipo': 'parcelas',
                                'dados': descricoes_list
                            }
                            break

                if tem_descricoes:
                    num_itens = len(descricoes_info['dados'])
                    tipo_texto = "eventos" if metodo == "Eventos/Fases" else "parcelas"

                    gestores_com_descricoes.append(
                        f"{cnpj_cpf} - {nome} ({num_itens} {tipo_texto})"
                    )

                    gestores_info[cnpj_cpf] = {
                        'nome': nome,
                        'descricoes_info': descricoes_info,
                        'item': item,
                        'valores': valores,
                        'tags': tags
                    }

            if not gestores_com_descricoes:
                tipo_config = "eventos" if metodo == "Eventos/Fases" else "descrições de parcelas"
                custom_messagebox("warning", "Aviso",
                                f"Nenhum gestor com {tipo_config} encontrado!\n\n" +
                                f"Configure {tipo_config} para pelo menos um gestor antes de copiar.")
                janela_copia.destroy()
                return

            origem_var = tk.StringVar()
            combo_origem = ttk.Combobox(frame_origem,
                                    textvariable=origem_var,
                                    values=gestores_com_descricoes,
                                    state='readonly',
                                    width=80)
            combo_origem.pack(padx=10, pady=10, fill='x')

            # ETAPA 2: Selecionar gestores de DESTINO (COM OPÇÃO DE ADICIONAR NOVO)
            frame_destino = ttk.LabelFrame(frame_copia, text="2. Gestores de Destino (copiar PARA)")
            frame_destino.pack(fill='both', expand=True, pady=10, padx=10)

            tipo_texto = "eventos" if metodo == "Eventos/Fases" else "descrições"
            ttk.Label(frame_destino,
                    text=f"Marque os gestores que receberão a cópia das {tipo_texto}:").pack(
                        anchor='w', padx=10, pady=5)

            # Frame para botão de adicionar gestor
            frame_btn_adicionar = ttk.Frame(frame_destino)
            frame_btn_adicionar.pack(fill='x', padx=10, pady=(0, 10))

            def adicionar_novo_gestor_para_copia():
                """Adiciona um novo gestor diretamente pela janela de cópia"""
                self.adicionar_administrador_modificado(
                    self.tree_adm,
                    valor_global,
                    metodo_pagamento,
                    janela_pai=janela_copia  # Passa a janela ao invés de callback
                )

            ttk.Button(
                frame_btn_adicionar,
                text="➕ Adicionar Novo Gestor",
                command=adicionar_novo_gestor_para_copia
            ).pack(side='left', padx=5)

            ttk.Label(
                frame_btn_adicionar,
                text="(Adicione um gestor sem eventos/descrições para receber a cópia)",
                font=('Arial', 8, 'italic'),
                foreground='gray'
            ).pack(side='left', padx=10)

            # Frame scrollável para checkboxes
            canvas_dest = tk.Canvas(frame_destino, height=200)
            scrollbar_dest = ttk.Scrollbar(frame_destino, orient="vertical", command=canvas_dest.yview)
            frame_checks = ttk.Frame(canvas_dest)

            frame_checks.bind(
                "<Configure>",
                lambda e: canvas_dest.configure(scrollregion=canvas_dest.bbox("all"))
            )

            canvas_dest.create_window((0, 0), window=frame_checks, anchor="nw")
            canvas_dest.configure(yscrollcommand=scrollbar_dest.set)

            canvas_dest.pack(side="left", fill="both", expand=True, padx=5)
            scrollbar_dest.pack(side="right", fill="y")

            # Criar checkbox para cada gestor
            gestores_destino_vars = {}

            def atualizar_lista_destinos():
                """Atualiza a lista de gestores de destino disponíveis"""
                # Limpar checkboxes existentes
                for widget in frame_checks.winfo_children():
                    widget.destroy()

                gestores_destino_vars.clear()

                # Recriar lista com gestores atualizados
                for item in self.tree_adm.get_children():
                    valores = self.tree_adm.item(item)['values']
                    cnpj_cpf = valores[0]
                    nome = valores[1]

                    var = tk.BooleanVar()
                    check = ttk.Checkbutton(frame_checks,
                                        text=f"{cnpj_cpf} - {nome}",
                                        variable=var)
                    check.pack(anchor='w', padx=10, pady=2)

                    gestores_destino_vars[cnpj_cpf] = {
                        'var': var,
                        'item': item,
                        'nome': nome,
                        'valores': valores
                    }

            # Criar lista inicial
            atualizar_lista_destinos()

            # ETAPA 3: Opções de cópia
            frame_opcoes = ttk.LabelFrame(frame_copia, text="3. Opções de Cópia")
            frame_opcoes.pack(fill='x', pady=10, padx=10)

            if metodo == "Eventos/Fases":
                var_ajustar_valores = tk.BooleanVar(value=True)
                ttk.Checkbutton(frame_opcoes,
                            text="Ajustar valores dos eventos proporcionalmente ao valor total de cada gestor",
                            variable=var_ajustar_valores).pack(anchor='w', padx=10, pady=5)

                ttk.Label(frame_opcoes,
                        text="📝 Os percentuais serão mantidos, mas os valores serão recalculados.",
                        font=('Arial', 8, 'italic'),
                        foreground='gray').pack(anchor='w', padx=25, pady=2)
            else:  # Valor Fixo em Parcelas
                var_copiar_entrada = tk.BooleanVar(value=True)
                ttk.Checkbutton(frame_opcoes,
                            text="Copiar também a descrição da entrada (se houver)",
                            variable=var_copiar_entrada).pack(anchor='w', padx=10, pady=5)

                ttk.Label(frame_opcoes,
                        text="📝 As descrições das parcelas serão copiadas na mesma ordem.",
                        font=('Arial', 8, 'italic'),
                        foreground='gray').pack(anchor='w', padx=25, pady=2)

            # ETAPA 4: Executar cópia
            def executar_copia_eventos():
                """Executa a cópia dos eventos/descrições entre gestores"""
                try:
                    # Validar origem
                    if not origem_var.get():
                        custom_messagebox("error", "Erro", "Selecione o gestor de origem!")
                        return

                    origem_selecionada = origem_var.get()
                    cnpj_origem_raw = origem_selecionada.split(' - ')[0].strip()
                    cnpj_origem = cnpj_origem_raw.replace('.', '').replace('/', '').replace('-', '').strip()

                    cnpj_origem_encontrado = None
                    for cnpj_key in gestores_info.keys():
                        cnpj_key_normalizado = str(cnpj_key).replace('.', '').replace('/', '').replace('-', '').strip()
                        if cnpj_key_normalizado == cnpj_origem:
                            cnpj_origem_encontrado = cnpj_key
                            break

                    if not cnpj_origem_encontrado:
                        custom_messagebox("error", "Erro",
                                        f"Gestor de origem não encontrado!\n\n" +
                                        f"CNPJ buscado: {cnpj_origem}\n" +
                                        f"CNPJs disponíveis: {list(gestores_info.keys())}")
                        return

                    # Validar destinos
                    destinos_selecionados = []
                    for cnpj, dados in gestores_destino_vars.items():
                        if dados['var'].get():
                            cnpj_normalizado = str(cnpj).replace('.', '').replace('/', '').replace('-', '').strip()
                            cnpj_origem_normalizado = str(cnpj_origem_encontrado).replace('.', '').replace('/', '').replace('-', '').strip()

                            if cnpj_normalizado != cnpj_origem_normalizado:
                                destinos_selecionados.append({
                                    'cnpj': cnpj,
                                    'item': dados['item'],
                                    'nome': dados['nome'],
                                    'valores': dados['valores']
                                })

                    if not destinos_selecionados:
                        custom_messagebox("error", "Erro",
                                        "Selecione pelo menos um gestor de destino!\n" +
                                        "(Diferente do gestor de origem)")
                        return

                    descricoes_origem = gestores_info[cnpj_origem_encontrado]['descricoes_info']

                    # Confirmar
                    num_itens = len(descricoes_origem['dados'])
                    tipo_item = "evento(s)" if metodo == "Eventos/Fases" else "descrição(ões)"

                    msg = f"Copiar {num_itens} {tipo_item} para {len(destinos_selecionados)} gestor(es)?\n\n"

                    if metodo == "Eventos/Fases" and var_ajustar_valores.get():
                        msg += "✓ Valores serão ajustados proporcionalmente\n"
                    elif metodo == "Eventos/Fases":
                        msg += "⚠️ Valores serão copiados exatamente iguais\n"

                    if not custom_messagebox("yesno", "Confirmação", msg):
                        return

                    # Executar cópia
                    valor_global_float = float(valor_global.get().replace(',', '.'))
                    DELIMITADOR = "|||"

                    for destino in destinos_selecionados:
                        item_destino = destino['item']
                        valores_destino = destino['valores']

                        tags_destino = list(self.tree_adm.item(item_destino)['tags'])

                        if metodo == "Eventos/Fases":
                            # Remover tag de eventos existente
                            tags_destino = [tag for tag in tags_destino if not tag.startswith('eventos:')]

                            eventos_destino = []

                            for evento_orig in descricoes_origem['dados']:
                                if var_ajustar_valores.get():
                                    if valores_destino[2] == 'Percentual':
                                        perc_gestor = float(str(valores_destino[3]).replace('%', '').replace(',', '.'))
                                        valor_total_gestor = (perc_gestor / 100) * valor_global_float
                                    else:  # Fixo
                                        valor_total_gestor = float(str(valores_destino[4]).replace(',', '.'))

                                    valor_evento = (evento_orig['percentual'] / 100) * valor_total_gestor
                                else:
                                    valor_evento = evento_orig['valor']

                                eventos_destino.append(
                                    f"{evento_orig['descricao']}:{evento_orig['percentual']}:{valor_evento}"
                                )

                            nova_tag_eventos = f"eventos:{'|'.join(eventos_destino)}"
                            tags_destino.append(nova_tag_eventos)

                            valores_atualizados = list(valores_destino)
                            valores_atualizados[5] = str(len(descricoes_origem['dados']))

                        else:  # Valor Fixo em Parcelas
                            tags_destino = [tag for tag in tags_destino
                                        if not tag.startswith('descricoes:')]

                            descricoes_list = descricoes_origem['dados'].copy()
                            nova_tag_descricoes = f"descricoes:{DELIMITADOR.join(descricoes_list)}"
                            tags_destino.append(nova_tag_descricoes)

                            if var_copiar_entrada.get():
                                tags_origem = gestores_info[cnpj_origem_encontrado]['tags']

                                for tag in tags_origem:
                                    if tag.startswith('desc_entrada:'):
                                        tags_destino = [t for t in tags_destino
                                                    if not t.startswith('desc_entrada:')]
                                        tags_destino.append(tag)
                                        break

                                for tag in tags_origem:
                                    if tag.startswith('entrada:'):
                                        tags_destino = [t for t in tags_destino
                                                    if not t.startswith('entrada:')]
                                        tags_destino.append(tag)
                                        break

                            valores_atualizados = list(valores_destino)

                        # Atualizar item na tree
                        self.tree_adm.item(item_destino, tags=tuple(tags_destino))
                        self.tree_adm.item(item_destino, values=tuple(valores_atualizados))

                    tipo_copiado = "Eventos" if metodo == "Eventos/Fases" else "Descrições"
                    custom_messagebox("info", "Sucesso",
                                    f"{tipo_copiado} copiado(s) com sucesso para {len(destinos_selecionados)} gestor(es)!\n\n" +
                                    f"Total de itens por gestor: {num_itens}")

                    janela_copia.destroy()

                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    custom_messagebox("error", "Erro", f"Erro ao copiar: {str(e)}")

            # Botões finais
            frame_botoes_copia = ttk.Frame(frame_copia)
            frame_botoes_copia.pack(fill='x', pady=10)

            ttk.Button(frame_botoes_copia,
                    text="✓ Executar Cópia",
                    command=executar_copia_eventos).pack(side='right', padx=5)

            ttk.Button(frame_botoes_copia,
                    text="Cancelar",
                    command=janela_copia.destroy).pack(side='right', padx=5)

        # Botões para administradores
        ttk.Button(
            frame_botoes_adm,
            text="➕ Adicionar Administrador",
            command=lambda: self.adicionar_administrador_modificado(
                self.tree_adm,
                valor_global,
                metodo_pagamento
                # SEM janela_pai (None por padrão)
            )
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_adm,
            text="➖ Remover Administrador",
            command=lambda: self.remover_administrador(self.tree_adm)
        ).pack(side='left', padx=5)

        # Copiar eventos entre gestores
        ttk.Button(
            frame_botoes_adm,
            text="📋 Copiar Descrições Entre Gestores",
            command=copiar_eventos_entre_gestores
        ).pack(side='left', padx=5)

        ttk.Label(frame_botoes_adm,
                text="💡 Configure um gestor primeiro, depois use 'Copiar' para replicar",
                font=('Arial', 8, 'italic'),
                foreground='gray').pack(side='left', padx=20)

        # NOTA DE EXTRAÇÃO: os dois botões abaixo (sem emoji) duplicam os
        # dois primeiros botões acima (com emoji), chamando exatamente as
        # mesmas funções. Mantidos como estavam no arquivo original — ver
        # docstring do módulo, pendência (1).
        ttk.Button(
            frame_botoes_adm,
            text="Adicionar Administrador",
            command=lambda: self.adicionar_administrador_modificado(self.tree_adm, valor_global, metodo_pagamento)
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes_adm,
            text="Remover Administrador",
            command=lambda: self.remover_administrador(self.tree_adm)
        ).pack(side='left', padx=5)

        def salvar():
            # Validar Nº Contrato
            if not num_contrato_var.get():
                custom_messagebox("error", "Erro", "Número do contrato não gerado. Reinicie a janela.")
                return

            # Validar datas
            if not data_inicio.get() or not data_fim.get():
                custom_messagebox("error", "Erro", "Preencha as datas do contrato!")
                return

            # Validar valor global apenas se o método exigir
            if metodo_pagamento.get() == "Percentual da Quinzena":
                valor_global_float = 0.0
            else:
                if not valor_global.get() or valor_global.get() == "0":
                    custom_messagebox("error", "Erro", "Informe o Valor Global do contrato!")
                    return
                try:
                    valor_global_float = float(valor_global.get().replace(',', '.'))
                    if valor_global_float <= 0:
                        custom_messagebox("error", "Erro", "Valor global deve ser maior que zero!")
                        return
                except ValueError:
                    custom_messagebox("error", "Erro", "Valor global inválido!")
                    return

            # Validar administradores
            if not self.tree_adm.get_children():
                custom_messagebox("error", "Erro", "Adicione pelo menos um administrador!")
                return

            self.salvar_contrato_com_opcoes(
                num_contrato_var.get(),
                data_inicio.get_date(),
                data_fim.get_date(),
                observacoes.get(),
                valor_global_float,
                metodo_pagamento.get(),
                {},
                janela
            )

            janela_principal.focus_set()
            self.carregar_contratos()

        ttk.Button(frame, text="Salvar", command=salvar).pack(side='left', padx=5, pady=10)
        ttk.Button(frame, text="Cancelar", command=janela.destroy).pack(side='left', padx=5, pady=10)

    def processar_eventos(self, ws, num_contrato, eventos_por_admin):
        """
        Processa eventos específicos de cada administrador com valores
        calculados corretamente baseados no valor total do admin.

        Args:
            ws: worksheet
            num_contrato: número do contrato
            eventos_por_admin: dict {cnpj_cpf: {
                'eventos': [(descricao, percentual, valor), ...],
                'nome': nome_do_admin,
                'valor_total': valor_total_do_admin
            }}
        """
        for cnpj_cpf, dados_admin in eventos_por_admin.items():
            eventos = dados_admin['eventos']
            nome_adm = dados_admin['nome']
            valor_total_admin = dados_admin['valor_total']

            # Criar parcelas para este administrador
            for i, (descricao, percentual, valor_original) in enumerate(eventos, 1):
                # Recalcular valor baseado no percentual e valor total DESTE admin
                valor_evento = (percentual / 100) * valor_total_admin

                proxima_linha = ws.max_row + 1
                ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())
                ws.cell(row=proxima_linha, column=26, value=i)
                ws.cell(row=proxima_linha, column=27, value=cnpj_cpf)
                ws.cell(row=proxima_linha, column=28, value=nome_adm)
                ws.cell(row=proxima_linha, column=29, value=None)
                ws.cell(row=proxima_linha, column=30, value=valor_evento)
                ws.cell(row=proxima_linha, column=31, value='PENDENTE')
                ws.cell(row=proxima_linha, column=32, value=None)
                ws.cell(row=proxima_linha, column=33, value=descricao.upper())
                ws.cell(row=proxima_linha, column=34, value=f"{percentual:.2f}%")

    def editar_contrato(self):
        """Edita o contrato com acesso completo a gestores e eventos"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um contrato para editar")
            return

        try:
            dados_contrato = self.tree_contratos.item(selecionado)['values']
            num_contrato = str(dados_contrato[0])

            janela = tk.Toplevel(self.parent)
            janela.title(f"Editar Contrato - {self.cliente_atual}")
            janela.geometry("1000x700")

            # Frame principal com scrollbar
            main_frame = ttk.Frame(janela)
            main_frame.pack(fill='both', expand=True)

            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scroll_frame = ttk.Frame(canvas)

            scroll_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            frame = ttk.Frame(scroll_frame, padding="10")
            frame.pack(fill='both', expand=True)

            # === DADOS DO CONTRATO ===
            frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
            frame_contrato.pack(fill='x', pady=5)

            # Número do Contrato (readonly)
            ttk.Label(frame_contrato, text="Nº Contrato:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            num_contrato_var = tk.StringVar(value=str(num_contrato))
            num_contrato_entry = ttk.Entry(frame_contrato, textvariable=num_contrato_var, state='readonly', width=30)
            num_contrato_entry.grid(row=0, column=1, padx=5, pady=2, sticky='w')

            # Datas
            ttk.Label(frame_contrato, text="Data Início:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
            data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_inicio.grid(row=1, column=1, padx=5, pady=2, sticky='w')
            data_inicio.set_date(datetime.strptime(dados_contrato[1], '%d/%m/%Y'))

            ttk.Label(frame_contrato, text="Data Fim:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
            data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_fim.grid(row=2, column=1, padx=5, pady=2, sticky='w')
            data_fim.set_date(datetime.strptime(dados_contrato[2], '%d/%m/%Y'))

            # Status
            ttk.Label(frame_contrato, text="Status:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
            status_combo = ttk.Combobox(frame_contrato, values=['ATIVO', 'INATIVO'], state='readonly', width=28)
            status_combo.grid(row=3, column=1, padx=5, pady=2, sticky='w')
            status_combo.set(dados_contrato[3])

            # === GESTORES/ADMINISTRADORES ===
            frame_gestores = ttk.LabelFrame(frame, text="Gestores/Administradores")
            frame_gestores.pack(fill='both', expand=True, pady=5)

            # Treeview para gestores
            colunas_gestores = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor Total')
            tree_gestores = ttk.Treeview(frame_gestores, columns=colunas_gestores, show='headings', height=4)
            for col in colunas_gestores:
                tree_gestores.heading(col, text=col)
                tree_gestores.column(col, width=150)

            tree_gestores.pack(fill='both', expand=True, padx=5, pady=5)

            # Botões para gestores
            frame_btn_gestores = ttk.Frame(frame_gestores)
            frame_btn_gestores.pack(fill='x', pady=5)

            def carregar_gestores():
                """Carrega gestores do contrato"""
                tree_gestores.delete(*tree_gestores.get_children())
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']

                for row in ws.iter_rows(min_row=3, values_only=True):
                    if str(row[6]) == str(num_contrato):  # Coluna G - Nº Contrato
                        valor_total = row[11] or 0  # Coluna L
                        try:
                            valor_total_num = float(str(valor_total).replace(',', '.'))
                            valor_total_formatado = f"R$ {valor_total_num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                        except:
                            valor_total_formatado = str(valor_total)

                        tree_gestores.insert('', 'end', values=(
                            row[7],   # CNPJ/CPF
                            row[8],   # Nome
                            row[9],   # Tipo
                            valor_total_formatado
                        ))
                wb.close()

            def excluir_gestor():
                """Exclui gestor selecionado"""
                selecionado_gest = tree_gestores.selection()
                if not selecionado_gest:
                    custom_messagebox("warning", "Aviso", "Selecione um gestor para excluir")
                    return

                if custom_messagebox("yesno", "Confirmação", "Deseja realmente excluir este gestor?"):
                    valores = tree_gestores.item(selecionado_gest)['values']
                    cnpj_cpf = valores[0]

                    wb = load_workbook(self.arquivo_cliente)
                    ws = wb['Contratos_ADM']

                    # Encontrar e deletar linha
                    linhas_deletar = []
                    for idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
                        if str(row[6].value) == str(num_contrato) and str(row[7].value) == str(cnpj_cpf):
                            linhas_deletar.append(idx)

                    for linha in reversed(linhas_deletar):
                        ws.delete_rows(linha)

                    wb.save(self.arquivo_cliente)
                    wb.close()
                    carregar_gestores()
                    carregar_eventos()  # Recarregar eventos também

            ttk.Button(frame_btn_gestores, text="Excluir Gestor",
                    command=excluir_gestor).pack(side='left', padx=5)

            # === EVENTOS/PARCELAS ===
            frame_eventos = ttk.LabelFrame(frame, text="Eventos/Parcelas")
            frame_eventos.pack(fill='both', expand=True, pady=5)

            # Treeview para eventos
            colunas_eventos = ('Nº', 'CNPJ/CPF', 'Nome', 'Valor', 'Descrição')
            tree_eventos = ttk.Treeview(frame_eventos, columns=colunas_eventos, show='headings', height=8)

            tree_eventos.column('Nº', width=50)
            tree_eventos.column('CNPJ/CPF', width=120)
            tree_eventos.column('Nome', width=150)
            tree_eventos.column('Valor', width=100)
            tree_eventos.column('Descrição', width=400)

            for col in colunas_eventos:
                tree_eventos.heading(col, text=col)

            scroll_y_eventos = ttk.Scrollbar(frame_eventos, orient='vertical', command=tree_eventos.yview)
            tree_eventos.configure(yscrollcommand=scroll_y_eventos.set)

            tree_eventos.pack(side='left', fill='both', expand=True, padx=5, pady=5)
            scroll_y_eventos.pack(side='right', fill='y')

            # Frame para botões de eventos (em coluna ao invés de linha)
            frame_btn_eventos = ttk.Frame(frame_eventos)
            frame_btn_eventos.pack(side='right', fill='y', padx=5, pady=5)

            def carregar_eventos():
                """Carrega eventos do contrato"""
                tree_eventos.delete(*tree_eventos.get_children())
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']

                for row in ws.iter_rows(min_row=3, values_only=True):
                    if str(row[24]) == str(num_contrato):  # Coluna Y - Referência (contrato)
                        num_parcela = row[25]  # Número

                        # Formatar exibição do número
                        if num_parcela == 0:
                            num_display = "ENTRADA"
                        else:
                            num_display = str(num_parcela)

                        # Formatar valor
                        valor = row[29] or 0  # Coluna AD - Valor
                        try:
                            valor_num = float(str(valor).replace(',', '.'))
                            valor_formatado = f"R$ {valor_num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                        except:
                            valor_formatado = str(valor)

                        descricao = row[32] if len(row) > 32 else ''

                        tree_eventos.insert('', 'end', values=(
                            num_display,
                            row[26],  # CNPJ/CPF
                            row[27],  # Nome
                            valor_formatado,  # Valor
                            descricao[:80] + '...' if descricao and len(str(descricao)) > 80 else descricao
                        ))
                wb.close()

            # FUNCIONALIDADE: Copiar eventos entre gestores
            def copiar_eventos_para_gestor():
                """Copia eventos de um gestor para outro(s)"""
                # Criar janela de seleção
                janela_copiar = tk.Toplevel(janela)
                janela_copiar.title("Copiar Eventos Entre Gestores")
                janela_copiar.geometry("600x500")

                frame_copiar = ttk.Frame(janela_copiar, padding="10")
                frame_copiar.pack(fill='both', expand=True)

                # Título e instruções
                ttk.Label(frame_copiar, text="Copiar Eventos Entre Gestores",
                        font=('Arial', 12, 'bold')).pack(pady=10)

                ttk.Label(frame_copiar,
                        text="Selecione o gestor de origem e os gestores de destino.",
                        wraplength=550).pack(pady=5)

                # PASSO 1: Selecionar gestor de origem
                frame_origem = ttk.LabelFrame(frame_copiar, text="Gestor de Origem (copiar DE)")
                frame_origem.pack(fill='x', pady=10, padx=10)

                # Listar gestores com eventos
                gestores_com_eventos = {}
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']

                for row in ws.iter_rows(min_row=3, values_only=True):
                    if str(row[24]) == num_contrato:  # Este contrato
                        cnpj_cpf = row[26]
                        nome = row[27]

                        if cnpj_cpf not in gestores_com_eventos:
                            gestores_com_eventos[cnpj_cpf] = {
                                'nome': nome,
                                'eventos': []
                            }

                        # Adicionar evento à lista
                        gestores_com_eventos[cnpj_cpf]['eventos'].append({
                            'numero': row[25],
                            'valor': row[29],
                            'descricao': row[32] if len(row) > 32 else '',
                            'status': row[30] if len(row) > 30 else 'PENDENTE'
                        })

                wb.close()

                if not gestores_com_eventos:
                    custom_messagebox("warning", "Aviso", "Nenhum gestor com eventos encontrado!")
                    janela_copiar.destroy()
                    return

                # ComboBox para selecionar gestor de origem
                gestor_origem_var = tk.StringVar()

                gestores_opcoes = [f"{cnpj} - {dados['nome']} ({len(dados['eventos'])} eventos)"
                                for cnpj, dados in gestores_com_eventos.items()]

                combo_origem = ttk.Combobox(frame_origem, textvariable=gestor_origem_var,
                                        values=gestores_opcoes, state='readonly', width=70)
                combo_origem.pack(padx=10, pady=10, fill='x')

                # PASSO 2: Selecionar gestores de destino
                frame_destino = ttk.LabelFrame(frame_copiar, text="Gestores de Destino (copiar PARA)")
                frame_destino.pack(fill='both', expand=True, pady=10, padx=10)

                ttk.Label(frame_destino,
                        text="Marque os gestores que receberão a cópia dos eventos:",
                        font=('Arial', 9, 'italic')).pack(pady=5)

                # Frame scrollável para checkboxes
                canvas_dest = tk.Canvas(frame_destino, height=150)
                scrollbar_dest = ttk.Scrollbar(frame_destino, orient="vertical", command=canvas_dest.yview)
                frame_checks = ttk.Frame(canvas_dest)

                frame_checks.bind(
                    "<Configure>",
                    lambda e: canvas_dest.configure(scrollregion=canvas_dest.bbox("all"))
                )

                canvas_dest.create_window((0, 0), window=frame_checks, anchor="nw")
                canvas_dest.configure(yscrollcommand=scrollbar_dest.set)

                canvas_dest.pack(side="left", fill="both", expand=True, padx=5)
                scrollbar_dest.pack(side="right", fill="y")

                # Criar checkboxes para cada gestor
                gestores_destino_vars = {}

                # Buscar TODOS os gestores do contrato
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']

                todos_gestores = set()
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if str(row[6]) == num_contrato:  # Coluna G - gestores do contrato
                        cnpj_cpf = row[7]
                        nome = row[8]
                        todos_gestores.add((cnpj_cpf, nome))

                wb.close()

                for cnpj_cpf, nome in sorted(todos_gestores):
                    var = tk.BooleanVar()
                    check = ttk.Checkbutton(frame_checks,
                                        text=f"{cnpj_cpf} - {nome}",
                                        variable=var)
                    check.pack(anchor='w', padx=10, pady=2)
                    gestores_destino_vars[cnpj_cpf] = var

                # PASSO 3: Opções de cópia
                frame_opcoes = ttk.LabelFrame(frame_copiar, text="Opções de Cópia")
                frame_opcoes.pack(fill='x', pady=10, padx=10)

                var_substituir = tk.BooleanVar(value=False)
                ttk.Checkbutton(frame_opcoes,
                            text="Substituir eventos existentes nos gestores de destino",
                            variable=var_substituir).pack(anchor='w', padx=10, pady=5)

                var_ajustar_valores = tk.BooleanVar(value=False)
                ttk.Checkbutton(frame_opcoes,
                            text="Ajustar valores proporcionalmente ao valor total de cada gestor",
                            variable=var_ajustar_valores).pack(anchor='w', padx=10, pady=5)

                # PASSO 4: Executar cópia
                def executar_copia():
                    """Executa a cópia dos eventos"""
                    try:
                        # Validar seleção
                        if not gestor_origem_var.get():
                            custom_messagebox("error", "Erro", "Selecione o gestor de origem!")
                            return

                        # Extrair CNPJ do gestor de origem
                        cnpj_origem = gestor_origem_var.get().split(' - ')[0]

                        # Verificar gestores de destino selecionados
                        destinos_selecionados = [cnpj for cnpj, var in gestores_destino_vars.items()
                                                if var.get() and cnpj != cnpj_origem]

                        if not destinos_selecionados:
                            custom_messagebox("error", "Erro", "Selecione pelo menos um gestor de destino!")
                            return

                        # Confirmar ação
                        msg_confirmacao = f"Copiar {len(gestores_com_eventos[cnpj_origem]['eventos'])} evento(s) "
                        msg_confirmacao += f"para {len(destinos_selecionados)} gestor(es)?\n\n"

                        if var_substituir.get():
                            msg_confirmacao += "⚠️ ATENÇÃO: Eventos existentes serão SUBSTITUÍDOS!"

                        if not custom_messagebox("yesno", "Confirmação", msg_confirmacao):
                            return

                        # Executar cópia
                        wb = load_workbook(self.arquivo_cliente)
                        ws = wb['Contratos_ADM']

                        eventos_origem = gestores_com_eventos[cnpj_origem]['eventos']
                        eventos_copiados = 0

                        for cnpj_destino in destinos_selecionados:
                            # Buscar nome do gestor de destino
                            nome_destino = None
                            valor_total_destino = 0

                            for row in ws.iter_rows(min_row=3, values_only=True):
                                if str(row[6]) == num_contrato and str(row[7]) == str(cnpj_destino):
                                    nome_destino = row[8]
                                    # Pegar valor total do gestor
                                    try:
                                        valor_total_destino = float(str(row[11]).replace(',', '.'))
                                    except:
                                        valor_total_destino = 0
                                    break

                            if not nome_destino:
                                continue

                            # Se deve substituir, marcar eventos existentes como excluídos
                            if var_substituir.get():
                                for idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
                                    if (str(row[24].value) == num_contrato and
                                        str(row[26].value) == str(cnpj_destino)):
                                        ws.cell(row=idx, column=31, value='EXCLUIDO')  # Status

                            # Copiar cada evento
                            for evento in eventos_origem:
                                proxima_linha = ws.max_row + 1

                                # Calcular valor (ajustar se necessário)
                                valor_evento = evento['valor']

                                if var_ajustar_valores.get() and valor_total_destino > 0:
                                    # Buscar valor total do gestor de origem
                                    valor_total_origem = 0
                                    for row in ws.iter_rows(min_row=3, values_only=True):
                                        if str(row[6]) == num_contrato and str(row[7]) == str(cnpj_origem):
                                            try:
                                                valor_total_origem = float(str(row[11]).replace(',', '.'))
                                            except:
                                                pass
                                            break

                                    if valor_total_origem > 0:
                                        # Ajustar proporcionalmente
                                        proporcao = valor_total_destino / valor_total_origem
                                        valor_evento = evento['valor'] * proporcao

                                # Inserir evento copiado
                                ws.cell(proxima_linha, 25, value=num_contrato)
                                ws.cell(proxima_linha, 26, value=evento['numero'])
                                ws.cell(proxima_linha, 27, value=cnpj_destino)
                                ws.cell(proxima_linha, 28, value=nome_destino)
                                ws.cell(proxima_linha, 29, value=None)  # Data vencimento
                                ws.cell(proxima_linha, 30, value=valor_evento)
                                ws.cell(proxima_linha, 31, value='PENDENTE')
                                ws.cell(proxima_linha, 32, value=None)  # Data pagamento
                                ws.cell(proxima_linha, 33, value=evento['descricao'])

                                eventos_copiados += 1

                        wb.save(self.arquivo_cliente)
                        wb.close()

                        custom_messagebox("info", "Sucesso",
                                        f"{eventos_copiados} evento(s) copiado(s) com sucesso!")

                        janela_copiar.destroy()
                        carregar_eventos()

                    except Exception as e:
                        import traceback
                        logger.debug(traceback.format_exc())
                        custom_messagebox("error", "Erro", f"Erro ao copiar eventos: {str(e)}")

                # Botões
                frame_botoes_copiar = ttk.Frame(frame_copiar)
                frame_botoes_copiar.pack(fill='x', pady=10)

                ttk.Button(frame_botoes_copiar, text="Executar Cópia",
                        command=executar_copia).pack(side='right', padx=5)
                ttk.Button(frame_botoes_copiar, text="Cancelar",
                        command=janela_copiar.destroy).pack(side='right', padx=5)

            def editar_evento():
                """Edita evento selecionado usando identificação completa"""
                selecionado_ev = tree_eventos.selection()
                if not selecionado_ev:
                    custom_messagebox("warning", "Aviso", "Selecione um evento para editar")
                    return

                valores = tree_eventos.item(selecionado_ev)['values']
                num_evento = valores[0]
                cnpj_cpf_evento = valores[1]
                nome_gestor_local = valores[2]

                # Buscar dados completos do evento
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']

                linha_evento = None
                for idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
                    if (str(row[24].value) == num_contrato and
                        row[25].value == num_evento and
                        str(row[26].value) == str(cnpj_cpf_evento)):  # Coluna AA - CNPJ/CPF
                        linha_evento = idx
                        break

                if linha_evento:
                    janela_editar = tk.Toplevel(janela)
                    janela_editar.title(f"Editar Evento {num_evento} - {valores[2]}")
                    janela_editar.geometry("700x450")

                    frame_ed = ttk.Frame(janela_editar, padding="10")
                    frame_ed.pack(fill='both', expand=True)

                    # Mostrar informações do gestor (readonly)
                    ttk.Label(frame_ed, text="Gestor:", font=('Arial', 10, 'bold')).grid(
                        row=0, column=0, padx=5, pady=5, sticky='w')
                    ttk.Label(frame_ed, text=f"{valores[2]} ({cnpj_cpf_evento})").grid(
                        row=0, column=1, padx=5, pady=5, sticky='w')

                    # Campos editáveis
                    ttk.Label(frame_ed, text="Valor:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
                    valor_entry = ttk.Entry(frame_ed, width=20)
                    valor_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
                    valor_entry.insert(0, ws.cell(linha_evento, 30).value or '')  # Coluna AD

                    ttk.Label(frame_ed, text="Status:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
                    status_evento = ttk.Combobox(frame_ed, values=['PENDENTE', 'PAGO', 'CANCELADO'],
                                                state='readonly', width=18)
                    status_evento.grid(row=2, column=1, padx=5, pady=5, sticky='w')
                    status_evento.set(ws.cell(linha_evento, 31).value or 'PENDENTE')  # Coluna AE

                    ttk.Label(frame_ed, text="Descrição do Evento:").grid(row=3, column=0, padx=5, pady=5, sticky='nw')
                    frame_text = ttk.Frame(frame_ed)
                    frame_text.grid(row=3, column=1, padx=5, pady=5, sticky='nsew')

                    desc_text = tk.Text(frame_text, width=60, height=12, wrap='word')
                    scroll_desc = ttk.Scrollbar(frame_text, orient='vertical', command=desc_text.yview)
                    desc_text.configure(yscrollcommand=scroll_desc.set)

                    desc_text.pack(side='left', fill='both', expand=True)
                    scroll_desc.pack(side='right', fill='y')

                    desc_atual = ws.cell(linha_evento, 33).value or ''  # Coluna AG
                    desc_text.insert('1.0', desc_atual)

                    frame_ed.grid_rowconfigure(3, weight=1)
                    frame_ed.grid_columnconfigure(1, weight=1)

                    def salvar_evento():
                        try:
                            novo_valor = float(valor_entry.get().replace(',', '.'))
                            ws.cell(linha_evento, 30, value=novo_valor)
                            ws.cell(linha_evento, 31, value=status_evento.get())

                            nova_desc = desc_text.get('1.0', 'end-1c').strip()
                            nova_desc_upper = nova_desc.upper()
                            ws.cell(linha_evento, 33, value=nova_desc_upper)

                            wb.save(self.arquivo_cliente)
                            wb.close()
                            janela_editar.destroy()
                            carregar_eventos()
                            custom_messagebox("info", "Sucesso", "Evento atualizado!")
                        except ValueError as e:
                            custom_messagebox("error", "Erro", f"Valor inválido: {str(e)}")
                        except Exception as e:
                            custom_messagebox("error", "Erro", f"Erro ao salvar: {str(e)}")

                    frame_btn = ttk.Frame(frame_ed)
                    frame_btn.grid(row=4, column=0, columnspan=2, pady=10)

                    ttk.Button(frame_btn, text="Salvar", command=salvar_evento).pack(side='left', padx=5)
                    ttk.Button(frame_btn, text="Cancelar", command=janela_editar.destroy).pack(side='left', padx=5)
                else:
                    wb.close()
                    custom_messagebox("error", "Erro", "Evento não encontrado!")

            def excluir_evento():
                """Exclui evento selecionado usando identificação completa"""
                selecionado_ev = tree_eventos.selection()
                if not selecionado_ev:
                    custom_messagebox("warning", "Aviso", "Selecione um evento para excluir")
                    return

                valores = tree_eventos.item(selecionado_ev)['values']
                num_evento = valores[0]
                cnpj_cpf_evento = valores[1]
                nome_gestor = valores[2]

                # Confirmação mais específica
                if custom_messagebox("yesno", "Confirmação",
                                    f"Deseja realmente excluir o evento {num_evento} do gestor {nome_gestor}?"):
                    wb = load_workbook(self.arquivo_cliente)
                    ws = wb['Contratos_ADM']

                    linhas_deletar = []
                    for idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
                        if (str(row[24].value) == num_contrato and
                            row[25].value == num_evento and
                            str(row[26].value) == str(cnpj_cpf_evento)):  # Coluna AA - CNPJ/CPF
                            linhas_deletar.append(idx)

                    if linhas_deletar:
                        for linha in reversed(linhas_deletar):
                            ws.delete_rows(linha)

                        wb.save(self.arquivo_cliente)
                        wb.close()
                        carregar_eventos()
                        custom_messagebox("info", "Sucesso", f"Evento {num_evento} do gestor {nome_gestor} excluído!")
                    else:
                        wb.close()
                        custom_messagebox("error", "Erro", "Evento não encontrado!")

            def adicionar_novo_evento():
                """Adiciona um novo evento ao contrato"""
                janela_novo = tk.Toplevel(janela)
                janela_novo.title(f"Adicionar Novo Evento - {num_contrato}")
                janela_novo.geometry("700x500")

                frame_novo = ttk.Frame(janela_novo, padding="10")
                frame_novo.pack(fill='both', expand=True)

                # Título
                ttk.Label(frame_novo, text="Novo Evento",
                        font=('Arial', 12, 'bold')).grid(row=0, column=0, columnspan=2, pady=10)

                # Selecionar Gestor
                ttk.Label(frame_novo, text="Gestor:*").grid(row=1, column=0, padx=5, pady=5, sticky='w')

                # Buscar gestores do contrato
                wb_temp = load_workbook(self.arquivo_cliente)
                ws_temp = wb_temp['Contratos_ADM']

                gestores_disponiveis = []
                for row in ws_temp.iter_rows(min_row=3, values_only=True):
                    if str(row[6]) == num_contrato:  # Coluna G - Nº Contrato
                        gestor_info = f"{row[8]} ({row[7]})"  # Nome (CNPJ/CPF)
                        if gestor_info not in gestores_disponiveis:
                            gestores_disponiveis.append(gestor_info)

                wb_temp.close()

                combo_gestor = ttk.Combobox(frame_novo, values=gestores_disponiveis,
                                            state='readonly', width=50)
                combo_gestor.grid(row=1, column=1, padx=5, pady=5, sticky='w')
                if gestores_disponiveis:
                    combo_gestor.current(0)

                # Número do Evento
                ttk.Label(frame_novo, text="Número da Parcela:*").grid(row=2, column=0, padx=5, pady=5, sticky='w')
                numero_entry = ttk.Entry(frame_novo, width=10)
                numero_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')

                # Calcular próximo número automaticamente
                wb_temp = load_workbook(self.arquivo_cliente)
                ws_temp = wb_temp['Contratos_ADM']
                max_num = 0
                for row in ws_temp.iter_rows(min_row=3, values_only=True):
                    if str(row[24]) == num_contrato and row[25]:  # Tem número
                        try:
                            num_atual = int(row[25])
                            if num_atual > max_num:
                                max_num = num_atual
                        except:
                            pass
                wb_temp.close()

                proximo_num = max_num + 1
                numero_entry.insert(0, str(proximo_num))

                ttk.Label(frame_novo, text=f"(Sugestão: próximo número disponível)",
                        font=('Arial', 8, 'italic')).grid(row=2, column=2, padx=5, pady=5, sticky='w')

                # Valor
                ttk.Label(frame_novo, text="Valor:*").grid(row=3, column=0, padx=5, pady=5, sticky='w')
                valor_entry = ttk.Entry(frame_novo, width=20)
                valor_entry.grid(row=3, column=1, padx=5, pady=5, sticky='w')

                # Status
                ttk.Label(frame_novo, text="Status:").grid(row=4, column=0, padx=5, pady=5, sticky='w')
                status_combo = ttk.Combobox(frame_novo, values=['PENDENTE', 'PAGO', 'CANCELADO'],
                                        state='readonly', width=18)
                status_combo.grid(row=4, column=1, padx=5, pady=5, sticky='w')
                status_combo.set('PENDENTE')

                # Descrição
                ttk.Label(frame_novo, text="Descrição:*").grid(row=5, column=0, padx=5, pady=5, sticky='nw')

                frame_text = ttk.Frame(frame_novo)
                frame_text.grid(row=5, column=1, columnspan=2, padx=5, pady=5, sticky='nsew')

                desc_text = tk.Text(frame_text, width=60, height=10, wrap='word')
                scroll_desc = ttk.Scrollbar(frame_text, orient='vertical', command=desc_text.yview)
                desc_text.configure(yscrollcommand=scroll_desc.set)

                desc_text.pack(side='left', fill='both', expand=True)
                scroll_desc.pack(side='right', fill='y')

                frame_novo.grid_rowconfigure(5, weight=1)
                frame_novo.grid_columnconfigure(1, weight=1)

                def salvar_novo_evento():
                    try:
                        # Validações
                        if not combo_gestor.get():
                            custom_messagebox("error", "Erro", "Selecione um gestor!")
                            return

                        if not numero_entry.get():
                            custom_messagebox("error", "Erro", "Informe o número da parcela!")
                            return

                        if not valor_entry.get():
                            custom_messagebox("error", "Erro", "Informe o valor!")
                            return

                        descricao = desc_text.get('1.0', 'end-1c').strip()
                        descricao_upper = descricao.upper()
                        if not descricao:
                            custom_messagebox("error", "Erro", "Informe a descrição!")
                            return

                        # Extrair CNPJ/CPF do gestor selecionado
                        gestor_texto = combo_gestor.get()
                        cnpj_cpf = gestor_texto.split('(')[1].split(')')[0]
                        nome_gestor_novo = gestor_texto.split('(')[0].strip()

                        numero = int(numero_entry.get())
                        valor = float(valor_entry.get().replace(',', '.'))

                        # Salvar na planilha
                        wb = load_workbook(self.arquivo_cliente)
                        ws = wb['Contratos_ADM']

                        # Verificar se já existe evento com mesmo número e gestor
                        existe = False
                        for row in ws.iter_rows(min_row=3, values_only=True):
                            if (str(row[24]) == num_contrato and
                                row[25] == numero and
                                str(row[26]) == str(cnpj_cpf)):
                                existe = True
                                break

                        if existe:
                            if not custom_messagebox("yesno", "Confirmação",
                                f"Já existe evento {numero} para o gestor {nome_gestor_novo}. Deseja criar mesmo assim?"):
                                wb.close()
                                return

                        # Adicionar nova linha
                        proxima_linha = ws.max_row + 1

                        # Preencher dados do evento
                        ws.cell(proxima_linha, 25, value=num_contrato)  # Coluna Y - Referência
                        ws.cell(proxima_linha, 26, value=numero)        # Coluna Z - Número
                        ws.cell(proxima_linha, 27, value=cnpj_cpf)      # Coluna AA - CNPJ/CPF
                        ws.cell(proxima_linha, 28, value=nome_gestor_novo)   # Coluna AB - Nome
                        ws.cell(proxima_linha, 29, value=None)          # Coluna AC - Data Vencimento
                        ws.cell(proxima_linha, 30, value=valor)         # Coluna AD - Valor
                        ws.cell(proxima_linha, 31, value=status_combo.get())  # Coluna AE - Status
                        ws.cell(proxima_linha, 32, value=None)          # Coluna AF - Data Pagamento
                        ws.cell(proxima_linha, 33, value=descricao_upper) # Coluna AG - Eventos/Fases

                        wb.save(self.arquivo_cliente)
                        wb.close()

                        janela_novo.destroy()
                        carregar_eventos()
                        custom_messagebox("info", "Sucesso", f"Evento {numero} adicionado com sucesso!")

                    except ValueError as e:
                        custom_messagebox("error", "Erro", f"Valor inválido: {str(e)}")
                    except Exception as e:
                        import traceback
                        logger.debug(traceback.format_exc())
                        custom_messagebox("error", "Erro", f"Erro ao salvar: {str(e)}")

                # Botões
                frame_btn = ttk.Frame(frame_novo)
                frame_btn.grid(row=6, column=0, columnspan=3, pady=10)

                ttk.Button(frame_btn, text="Salvar", command=salvar_novo_evento).pack(side='left', padx=5)
                ttk.Button(frame_btn, text="Cancelar", command=janela_novo.destroy).pack(side='left', padx=5)

            # Botões para eventos
            ttk.Button(frame_btn_eventos, text="Editar Evento",
                    command=editar_evento, width=20).pack(pady=2, fill='x')
            ttk.Button(frame_btn_eventos, text="Excluir Evento",
                    command=excluir_evento, width=20).pack(pady=2, fill='x')
            ttk.Button(frame_btn_eventos, text="➕ Adicionar Evento",
                    command=adicionar_novo_evento, width=20).pack(pady=2, fill='x')

            # Copiar eventos
            ttk.Button(frame_btn_eventos, text="📋 Copiar Eventos",
                    command=copiar_eventos_para_gestor, width=20).pack(pady=2, fill='x')

            # Separador visual
            ttk.Separator(frame_btn_eventos, orient='horizontal').pack(fill='x', pady=10)

            # Informação útil
            ttk.Label(frame_btn_eventos,
                    text="💡 Dica:\nDuplo clique\npara editar",
                    font=('Arial', 8, 'italic'),
                    justify='center').pack(pady=5)

            # DUPLO CLIQUE PARA EDITAR
            tree_eventos.bind('<Double-Button-1>', lambda e: editar_evento())

            # Carregar dados iniciais
            carregar_gestores()
            carregar_eventos()

            # === BOTÕES FINAIS ===
            def salvar_alteracoes():
                try:
                    wb = load_workbook(self.arquivo_cliente)
                    ws = wb['Contratos_ADM']

                    # Atualizar dados básicos do contrato
                    for row in ws.iter_rows(min_row=3):
                        if str(row[0].value) == num_contrato:
                            row[1].value = data_inicio.get_date()
                            row[2].value = data_fim.get_date()
                            row[3].value = status_combo.get()
                            break

                    wb.save(self.arquivo_cliente)
                    wb.close()
                    custom_messagebox("info", "Sucesso", "Contrato atualizado com sucesso!")
                    janela.destroy()
                    self.carregar_contratos()

                except Exception as e:
                    import traceback
                    logger.debug(traceback.format_exc())
                    custom_messagebox("error", "Erro", f"Erro ao salvar: {str(e)}")

            # NOTA DE EXTRAÇÃO: no texto colado pelo usuário, a linha acima
            # (mensagem de erro do except) e a criação de frame_botoes
            # abaixo vieram grudadas na mesma linha, sem quebra — corrigido
            # aqui para o padrão de indentação esperado. Vale confirmar no
            # arquivo real se essa junção existe de fato ou foi só um
            # efeito do copiar/colar na conversa.
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)

            ttk.Button(frame_botoes, text="Salvar Alterações",
                    command=salvar_alteracoes).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Fechar",
                    command=janela.destroy).pack(side='right', padx=5)

        except Exception as e:
            import traceback
            logger.debug(traceback.format_exc())
            custom_messagebox("error", "Erro", f"Erro ao abrir edição: {str(e)}")

    def adicionar_administrador_modificado(self, tree, valor_global_entry, metodo_pagamento_combo, janela_pai=None):
        """
        Versão modificada para incluir os detalhes de parcelas/eventos na tela do administrador
        """
        metodo = metodo_pagamento_combo.get()

        if metodo == "Percentual da Quinzena":
            valor_global_float = 0.0  # não usado neste método
        else:
            if not valor_global_entry.get() or valor_global_entry.get() == "0":
                custom_messagebox("error", "Erro", "Informe o valor global do contrato primeiro")
                return
            try:
                valor_global_float = float(valor_global_entry.get().replace(',', '.'))
                if valor_global_float <= 0:
                    custom_messagebox("error", "Erro", "Valor global deve ser maior que zero")
                    return
            except ValueError:
                custom_messagebox("error", "Erro", "Valor global inválido")
                return

        # Criar janela como Toplevel do parent
        janela_admin = tk.Toplevel(self.parent)
        janela_admin.title("Adicionar Administrador")

        # Ajustar tamanho baseado no método
        if metodo == "Eventos/Fases":
            altura_inicial = 750
        elif metodo == "Valor Fixo em Parcelas":
            altura_inicial = 750
        else:
            altura_inicial = 700

        janela_admin.geometry(f"700x{altura_inicial}")
        # CONFIGURAR MODAL CORRETAMENTE
        if janela_pai and janela_pai.winfo_exists():
            # Ocultar janela_pai
            janela_pai.withdraw()

            # NÃO usar grab_set quando há janela_pai (causa travamento)
            # Apenas tornar transient para manter hierarquia
            janela_admin.transient(janela_pai)

            # Garantir que janela_admin fique na frente
            janela_admin.lift()
            janela_admin.focus_force()

            # Protocolo de fechamento para restaurar janela_pai
            def ao_fechar_janela():
                """Restaura janela_pai ao fechar janela_admin"""
                janela_admin.destroy()
                if janela_pai.winfo_exists():
                    janela_pai.deiconify()
                    janela_pai.lift()
                    janela_pai.focus_force()

            janela_admin.protocol("WM_DELETE_WINDOW", ao_fechar_janela)
        else:
            # Sem janela_pai, usar grab_set normalmente
            janela_admin.transient(self.parent)
            janela_admin.grab_set()
            janela_admin.lift()
            janela_admin.focus_force()

        # Frame principal com scrollbar para garantir acesso a todos os campos
        main_frame = ttk.Frame(janela_admin)
        main_frame.pack(fill='both', expand=True)

        # Adicionar canvas com scrollbar se for Eventos/Fases
        if metodo == "Eventos/Fases":
            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            frame_admin = ttk.Frame(canvas)

            frame_admin.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=frame_admin, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
        else:
            frame_admin = ttk.Frame(main_frame, padding="10")
            frame_admin.pack(fill='both', expand=True)

        # Frame de busca
        frame_busca = ttk.LabelFrame(frame_admin, text="Buscar Administrador")
        frame_busca.pack(fill='x', padx=5, pady=5)

        # Label informativo
        ttk.Label(frame_busca,
                text="Apenas fornecedores com categoria TAX são exibidos",
                font=('Arial', 8, 'italic'),
                foreground='gray').pack(side='top', padx=5, pady=2)

        ttk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        busca_entry = ttk.Entry(frame_busca, width=40)
        busca_entry.pack(side='left', padx=5)

        # Lista de fornecedores
        frame_lista = ttk.LabelFrame(frame_admin, text="Fornecedores (Categoria TAX)")
        frame_lista.pack(fill='x', padx=5, pady=5)

        tree_fornecedores = ttk.Treeview(frame_lista,
                                    columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                    show='headings',
                                    height=3)
        tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        tree_fornecedores.heading('Nome', text='Nome')
        tree_fornecedores.heading('Categoria', text='Categoria')
        tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)

        # Frame para dados do administrador
        frame_dados = ttk.LabelFrame(frame_admin, text="Dados do Administrador")
        frame_dados.pack(fill='x', padx=5, pady=5)

        # CNPJ/CPF
        ttk.Label(frame_dados, text="CNPJ/CPF:*").grid(row=0, column=0, padx=5, pady=2)
        cnpj_cpf_entry = ttk.Entry(frame_dados, state='readonly')
        cnpj_cpf_entry.grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        # Nome
        ttk.Label(frame_dados, text="Nome/Razão Social:*").grid(row=1, column=0, padx=5, pady=2)
        nome_entry = ttk.Entry(frame_dados, state='readonly')
        nome_entry.grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Mostrar informações do contrato (somente leitura)
        ttk.Label(frame_dados, text="Valor Global:").grid(row=2, column=0, padx=5, pady=2)
        valor_global_label = ttk.Label(frame_dados, text=f"R$ {valor_global_float:,.2f}")
        valor_global_label.grid(row=2, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_dados, text="Método de Pagamento:").grid(row=3, column=0, padx=5, pady=2)
        metodo_label = ttk.Label(frame_dados, text=metodo)
        metodo_label.grid(row=3, column=1, padx=5, pady=2, sticky='w')

        # Tipo de remuneração
        ttk.Label(frame_dados, text="Tipo de Remuneração:*").grid(row=4, column=0, padx=5, pady=2)

        if metodo == "Percentual da Quinzena":
            tipo_combo = ttk.Combobox(frame_dados, values=['Percentual'], state='readonly')
            tipo_combo.grid(row=4, column=1, padx=5, pady=2, sticky='w')
            tipo_combo.set('Percentual')

            # Percentual
            ttk.Label(frame_dados, text="Percentual (%):*").grid(row=5, column=0, padx=5, pady=2)
            percentual_entry = ttk.Entry(frame_dados)
            percentual_entry.grid(row=5, column=1, padx=5, pady=2, sticky='w')

        else:  # Valor Fixo em Parcelas ou Eventos/Fases
            tipo_valores = ['Percentual', 'Fixo']
            tipo_combo = ttk.Combobox(frame_dados, values=tipo_valores, state='readonly')
            tipo_combo.grid(row=4, column=1, padx=5, pady=2)
            tipo_combo.set('Fixo')  # Padrão para eventos/parcelas fixas

            # Frame para valores percentuais
            frame_percentual_admin = ttk.Frame(frame_dados)
            frame_percentual_admin.grid(row=5, column=0, columnspan=2, pady=5)

            # Percentual
            ttk.Label(frame_percentual_admin, text="Percentual do Contrato (%):*").grid(row=0, column=0, padx=5, pady=2)
            percentual_entry = ttk.Entry(frame_percentual_admin)
            percentual_entry.grid(row=0, column=1, padx=5, pady=2)

            # Frame para valores fixos
            frame_fixo = ttk.Frame(frame_dados)
            frame_fixo.grid(row=6, column=0, columnspan=2, pady=5)

            # Valor Total
            ttk.Label(frame_fixo, text="Valor Total:*").grid(row=0, column=0, padx=5, pady=2)
            valor_total_entry = ttk.Entry(frame_fixo)
            valor_total_entry.grid(row=0, column=1, padx=5, pady=2)

            def atualizar_campos_tipo(*args):
                """Atualiza campos baseado no tipo selecionado"""
                if tipo_combo.get() == 'Percentual':
                    frame_percentual_admin.grid()
                    frame_fixo.grid_remove()
                elif tipo_combo.get() == 'Fixo':
                    frame_percentual_admin.grid_remove()
                    frame_fixo.grid()

            # Configurar evento
            tipo_combo.bind('<<ComboboxSelected>>', atualizar_campos_tipo)

            # Configurar interface inicial
            atualizar_campos_tipo()

        # Forma de pagamento para dados bancários
        ttk.Label(frame_dados, text="Forma de Pagamento:").grid(row=7, column=0, padx=5, pady=2)
        forma_pagamento = ttk.Combobox(frame_dados, values=['PIX', 'TED'], state='readonly')
        forma_pagamento.grid(row=7, column=1, padx=5, pady=2)
        forma_pagamento.set('PIX')  # Valor padrão

        # Área para configurações específicas do método de pagamento
        frame_config_metodo = ttk.LabelFrame(frame_admin, text="Configuração de Pagamento")

        if metodo in ["Valor Fixo em Parcelas", "Eventos/Fases"]:
            frame_config_metodo.pack(fill='x', padx=5, pady=5, after=frame_dados)

        # Variáveis que precisam existir em todos os casos
        eventos = []
        descricoes_parcelas = []
        var_tem_entrada = tk.BooleanVar(value=False)

        # 1. Frame para Parcelas Fixas
        if metodo == "Valor Fixo em Parcelas":
            frame_parcelas = ttk.Frame(frame_config_metodo)
            frame_parcelas.pack(fill='x', padx=5, pady=5)

            # Número de parcelas
            ttk.Label(frame_parcelas, text="Número de Parcelas:*").grid(row=0, column=0, padx=5, pady=5, sticky='w')
            num_parcelas_entry = ttk.Entry(frame_parcelas, width=10)
            num_parcelas_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')

            # Checkbox para entrada
            check_entrada = ttk.Checkbutton(frame_parcelas, text="Possui entrada?", variable=var_tem_entrada)
            check_entrada.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky='w')

            # Frame para entrada
            frame_entrada = ttk.Frame(frame_parcelas)
            frame_entrada.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky='w')

            ttk.Label(frame_entrada, text="Valor da Entrada:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            valor_entrada_entry = ttk.Entry(frame_entrada, width=15)
            valor_entrada_entry.grid(row=0, column=1, padx=5, pady=2, sticky='w')

            ttk.Label(frame_entrada, text="Data da Entrada:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
            data_entrada = DateEntry(frame_entrada, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_entrada.grid(row=1, column=1, padx=5, pady=2, sticky='w')

            ttk.Label(frame_entrada, text="Descrição da Entrada:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
            descricao_entrada = ttk.Entry(frame_entrada, width=40)
            descricao_entrada.grid(row=2, column=1, padx=5, pady=2, sticky='w')
            descricao_entrada.insert(0, "ENTRADA")  # Valor padrão

            # Ocultar frame de entrada inicialmente
            frame_entrada.grid_remove()

            def toggle_entrada():
                if var_tem_entrada.get():
                    frame_entrada.grid()
                    janela_admin.update_idletasks()
                    nova_altura = 750
                    largura_atual = janela_admin.winfo_width()
                    x_atual = janela_admin.winfo_x()
                    y_atual = janela_admin.winfo_y()
                    altura_tela = janela_admin.winfo_screenheight()
                    novo_y = max(50, min(y_atual, altura_tela - nova_altura - 50))
                    janela_admin.geometry(f"{largura_atual}x{nova_altura}+{x_atual}+{novo_y}")
                else:
                    frame_entrada.grid_remove()
                    janela_admin.update_idletasks()
                    altura_reduzida = 650
                    largura_atual = janela_admin.winfo_width()
                    x_atual = janela_admin.winfo_x()
                    y_atual = janela_admin.winfo_y()
                    janela_admin.geometry(f"{largura_atual}x{altura_reduzida}+{x_atual}+{y_atual}")

            check_entrada.config(command=toggle_entrada)

            # Frame para gerenciar descrições individuais das parcelas
            frame_descricoes = ttk.LabelFrame(frame_parcelas, text="Descrições Individuais das Parcelas")
            frame_descricoes.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky='ew')

            ttk.Label(frame_descricoes, text="Para configurar descrições individuais, primeiro defina o número de parcelas e clique em:").grid(
                row=0, column=0, columnspan=2, padx=5, pady=2, sticky='w')

            def configurar_descricoes_parcelas():
                try:
                    if not num_parcelas_entry.get():
                        custom_messagebox("error", "Erro", "Informe o número de parcelas primeiro")
                        return

                    num_parcelas = int(num_parcelas_entry.get())
                    if num_parcelas <= 0:
                        custom_messagebox("error", "Erro", "Número de parcelas deve ser maior que zero")
                        return

                    janela_descricoes = tk.Toplevel(janela_admin)
                    janela_descricoes.title("Descrições Individuais das Parcelas")
                    janela_descricoes.geometry("500x700")

                    frame_scroll = ttk.Frame(janela_descricoes)
                    frame_scroll.pack(fill='both', expand=True, padx=10, pady=10)

                    canvas = tk.Canvas(frame_scroll)
                    scrollbar = ttk.Scrollbar(frame_scroll, orient="vertical", command=canvas.yview)
                    frame_content = ttk.Frame(canvas)

                    frame_content.bind(
                        "<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                    )

                    canvas.create_window((0, 0), window=frame_content, anchor="nw")
                    canvas.configure(yscrollcommand=scrollbar.set)

                    canvas.pack(side="left", fill="both", expand=True)
                    scrollbar.pack(side="right", fill="y")

                    if len(descricoes_parcelas) < num_parcelas:
                        for _ in range(num_parcelas - len(descricoes_parcelas)):
                            descricoes_parcelas.append("")
                    else:
                        del descricoes_parcelas[num_parcelas:]

                    for i in range(num_parcelas):
                        ttk.Label(frame_content, text=f"Parcela {i+1}:").grid(
                            row=i, column=0, padx=5, pady=5, sticky='w')

                        desc_entry = ttk.Entry(frame_content, width=40)
                        desc_entry.grid(row=i, column=1, padx=5, pady=5, sticky='ew')

                        if i < len(descricoes_parcelas) and descricoes_parcelas[i]:
                            desc_entry.insert(0, descricoes_parcelas[i])
                        else:
                            desc_entry.insert(0, f"PARCELA {i+1}")

                        desc_entry.idx = i

                    def salvar_descricoes():
                        for child in frame_content.winfo_children():
                            if isinstance(child, ttk.Entry):
                                idx = getattr(child, 'idx', -1)
                                if 0 <= idx < len(descricoes_parcelas):
                                    descricoes_parcelas[idx] = child.get().strip()

                        custom_messagebox("info", "Sucesso", "Descrições salvas!")
                        janela_descricoes.destroy()

                    frame_botoes = ttk.Frame(janela_descricoes)
                    frame_botoes.pack(fill='x', pady=10)

                    ttk.Button(frame_botoes, text="Salvar Descrições",
                            command=salvar_descricoes).pack(side='right', padx=10)

                    ttk.Button(frame_botoes, text="Cancelar",
                            command=janela_descricoes.destroy).pack(side='right', padx=10)

                    janela_descricoes.update_idletasks()
                    w = janela_descricoes.winfo_width()
                    h = janela_descricoes.winfo_height()
                    x = (janela_descricoes.winfo_screenwidth() // 2) - (w // 2)
                    y = (janela_descricoes.winfo_screenheight() // 2) - (h // 2)
                    janela_descricoes.geometry(f'{w}x{h}+{x}+{y}')

                    janela_descricoes.transient(janela_admin)
                    janela_descricoes.grab_set()

                except Exception as e:
                    custom_messagebox("error", "Erro", f"Erro ao configurar descrições: {str(e)}")

            btn_config_descricoes = ttk.Button(frame_descricoes,
                                            text="Configurar Descrições",
                                            command=configurar_descricoes_parcelas)
            btn_config_descricoes.grid(row=1, column=0, padx=5, pady=5, sticky='w')

        # 2. Frame para Eventos/Fases
        elif metodo == "Eventos/Fases":
            frame_eventos = ttk.Frame(frame_config_metodo)
            frame_eventos.pack(fill='x', padx=5, pady=5)

            # Lista de eventos
            colunas_evento = ('Nº', 'Descrição', 'Percentual', 'Valor')
            tree_eventos = ttk.Treeview(frame_eventos, columns=colunas_evento, show='headings', height=5)
            tree_eventos.heading('Nº', text='Nº')
            tree_eventos.heading('Descrição', text='Descrição')
            tree_eventos.heading('Percentual', text='Percentual (%)')
            tree_eventos.heading('Valor', text='Valor (R$)')

            tree_eventos.column('Nº', width=50, anchor='center')
            tree_eventos.column('Descrição', width=300)
            tree_eventos.column('Percentual', width=100, anchor='e')
            tree_eventos.column('Valor', width=100, anchor='e')

            scroll_y_eventos = ttk.Scrollbar(frame_eventos, orient='vertical', command=tree_eventos.yview)
            scroll_x_eventos = ttk.Scrollbar(frame_eventos, orient='horizontal', command=tree_eventos.xview)
            tree_eventos.configure(yscrollcommand=scroll_y_eventos.set, xscrollcommand=scroll_x_eventos.set)

            tree_eventos.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y_eventos.pack(side='right', fill='y')
            scroll_x_eventos.pack(side='bottom', fill='x')

            # Frame para adicionar evento
            frame_add_evento = ttk.Frame(frame_eventos)
            frame_add_evento.pack(fill='x', padx=5, pady=5)

            ttk.Label(frame_add_evento, text="Descrição:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            evento_descricao = ttk.Entry(frame_add_evento, width=40)
            evento_descricao.grid(row=0, column=1, padx=5, pady=2, sticky='w')

            ttk.Label(frame_add_evento, text="Percentual (%):").grid(row=0, column=2, padx=5, pady=2, sticky='w')
            evento_percentual = ttk.Entry(frame_add_evento, width=10)
            evento_percentual.grid(row=0, column=3, padx=5, pady=2, sticky='w')

            # Botões para eventos
            frame_botoes_evento = ttk.Frame(frame_eventos)
            frame_botoes_evento.pack(fill='x', pady=5)

            total_percentual_var = tk.StringVar(value="Total: 0%")
            lbl_total_percentual = ttk.Label(frame_botoes_evento, textvariable=total_percentual_var)
            lbl_total_percentual.pack(side='left', padx=5)

            def calcular_valor_evento(percentual, valor_total_str):
                try:
                    percentual_float = float(percentual.replace(',', '.'))
                    valor_float = float(valor_total_str.replace(',', '.'))
                    return (percentual_float / 100) * valor_float
                except (ValueError, AttributeError):
                    return 0

            def adicionar_evento():
                if not valor_global_entry.get():
                    custom_messagebox("error", "Erro", "Informe o valor global do contrato primeiro")
                    return

                descricao = evento_descricao.get().strip()
                percentual_str = evento_percentual.get().strip()

                if not descricao:
                    custom_messagebox("error", "Erro", "Informe a descrição do evento")
                    return

                try:
                    percentual = float(percentual_str.replace(',', '.'))
                    if percentual <= 0 or percentual > 100:
                        custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100")
                        return
                except ValueError:
                    custom_messagebox("error", "Erro", "Percentual inválido")
                    return

                total_atual = sum(float(e[1]) for e in eventos)

                if total_atual + percentual > 100:
                    custom_messagebox("error", "Erro", "Total de percentual não pode exceder 100%")
                    return

                valor_total = valor_global_entry.get().replace(',', '.')
                try:
                    valor_total_float = float(valor_total)
                    valor_evento = (percentual / 100) * valor_total_float
                except (ValueError, TypeError):
                    valor_evento = 0

                eventos.append((descricao, percentual, valor_evento))

                tree_eventos.insert('', 'end', values=(
                    len(eventos),
                    descricao,
                    f"{percentual:.2f}",
                    f"R$ {valor_evento:.2f}"
                ))

                total_percentual_var.set(f"Total: {total_atual + percentual:.2f}%")

                evento_descricao.delete(0, tk.END)
                evento_percentual.delete(0, tk.END)

            def remover_evento():
                selecionado = tree_eventos.selection()
                if not selecionado:
                    custom_messagebox("warning", "Aviso", "Selecione um evento para remover")
                    return

                valores = tree_eventos.item(selecionado)['values']
                indice = int(valores[0]) - 1

                if 0 <= indice < len(eventos):
                    eventos.pop(indice)

                    for item in tree_eventos.get_children():
                        tree_eventos.delete(item)

                    for i, (desc, perc, valor) in enumerate(eventos, 1):
                        tree_eventos.insert('', 'end', values=(i, desc, f"{perc:.2f}", f"R$ {valor:.2f}"))

                    total_atual = sum(float(e[1]) for e in eventos)
                    total_percentual_var.set(f"Total: {total_atual:.2f}%")

            ttk.Button(frame_botoes_evento, text="Adicionar Evento", command=adicionar_evento).pack(side='right', padx=5)
            ttk.Button(frame_botoes_evento, text="Remover Evento", command=remover_evento).pack(side='right', padx=5)

        # Função de busca LOCAL que usa filtro TAX
        def busca_local():
            termo = busca_entry.get()
            buscar_fornecedor(tree_fornecedores, termo, categoria_filtro='TAX')

        ttk.Button(frame_busca, text="Buscar", command=busca_local).pack(side='left', padx=5)
        busca_entry.bind('<Return>', lambda e: busca_local())

        # Carregar automaticamente os TAX ao abrir
        janela_admin.after(100, lambda: buscar_fornecedor(tree_fornecedores, '', categoria_filtro='TAX'))

        def selecionar_e_preencher(event=None):
            """Seleciona fornecedor e preenche campos com tipo_pessoa correto"""
            selecionado = tree_fornecedores.selection()
            if not selecionado:
                return

            valores = tree_fornecedores.item(selecionado)['values']
            tags = tree_fornecedores.item(selecionado)['tags']

            cnpj_cpf_bruto = str(valores[0])
            nome = valores[1]

            tipo_pessoa = tags[0] if tags else None

            if not tipo_pessoa or tipo_pessoa not in ['PF', 'PJ']:
                custom_messagebox("error", "Erro",
                                f"Tipo de pessoa não identificado para '{nome}'.\n" +
                                "Verifique o cadastro em base_fornecedores.xlsx (coluna B)")
                return

            try:
                cnpj_cpf_normalizado = normalizar_documento(cnpj_cpf_bruto, tipo_pessoa)
                cnpj_cpf_formatado = formatar_documento(cnpj_cpf_normalizado, tipo_pessoa)

                cnpj_cpf_entry.config(state='normal')
                nome_entry.config(state='normal')

                cnpj_cpf_entry.delete(0, tk.END)
                cnpj_cpf_entry.insert(0, cnpj_cpf_formatado)

                nome_entry.delete(0, tk.END)
                nome_entry.insert(0, nome)

                cnpj_cpf_entry.config(state='readonly')
                nome_entry.config(state='readonly')

                logger.debug(f"Fornecedor selecionado: {nome} ({cnpj_cpf_formatado}) - Tipo: {tipo_pessoa}")

            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao formatar documento: {str(e)}")
                logger.debug(f"Erro ao formatar {cnpj_cpf_bruto} como {tipo_pessoa}: {e}")

        tree_fornecedores.bind('<Double-1>', selecionar_e_preencher)

        def confirmar():
            """Confirma a adição do administrador"""
            try:
                if not cnpj_cpf_entry.get() or not nome_entry.get() or not tipo_combo.get():
                    custom_messagebox("error", "Erro", "Preencha todos os campos obrigatórios!")
                    return

                forma_pagto_selecionada = forma_pagamento.get()
                tags_extra = []

                metodo = metodo_pagamento_combo.get()

                if metodo == "Valor Fixo em Parcelas":
                    if not num_parcelas_entry.get():
                        custom_messagebox("error", "Erro", "Informe o número de parcelas!")
                        return

                    try:
                        num_parcelas = int(num_parcelas_entry.get())
                        if num_parcelas <= 0:
                            custom_messagebox("error", "Erro", "Número de parcelas deve ser maior que zero!")
                            return
                    except ValueError:
                        custom_messagebox("error", "Erro", "Número de parcelas inválido!")
                        return

                    if var_tem_entrada.get():
                        if not valor_entrada_entry.get():
                            custom_messagebox("error", "Erro", "Informe o valor da entrada!")
                            return

                        try:
                            valor_entrada = float(valor_entrada_entry.get().replace(',', '.'))
                            if valor_entrada <= 0:
                                custom_messagebox("error", "Erro", "Valor da entrada deve ser maior que zero!")
                                return
                        except ValueError:
                            custom_messagebox("error", "Erro", "Valor da entrada inválido!")
                            return

                if metodo == "Eventos/Fases":
                    total_percentual = sum(float(e[1]) for e in eventos)
                    if total_percentual < 99.99 or total_percentual > 100.01:
                        if not custom_messagebox("yesno", "Confirmação",
                                            f"O total de percentuais é {total_percentual:.2f}% ao invés de 100%. Deseja continuar mesmo assim?"):
                            return

                if tipo_combo.get() == 'Percentual':
                    if not percentual_entry.get():
                        custom_messagebox("error", "Erro", "Preencha o percentual!")
                        return

                    try:
                        perc = float(percentual_entry.get().replace(',', '.'))
                        if perc <= 0 or perc > 100:
                            custom_messagebox("error", "Erro", "Percentual deve estar entre 0 e 100!")
                            return

                        valor_global_float = float(valor_global_entry.get().replace(',', '.'))

                        if metodo == "Percentual da Quinzena":
                            num_parcelas_val = ""
                            data_inicial = ""
                        elif metodo == "Valor Fixo em Parcelas":
                            num_parcelas_val = num_parcelas_entry.get()
                            data_inicial = data_entrada.get() if var_tem_entrada.get() else ""
                        else:
                            num_parcelas_val = str(len(eventos))
                            data_inicial = ""

                        valores_finais = (
                            cnpj_cpf_entry.get(),
                            nome_entry.get(),
                            tipo_combo.get(),
                            f"{perc:.2f}%",
                            f"{valor_global_float:.2f}",
                            num_parcelas_val,
                            data_inicial
                        )

                        if metodo == "Valor Fixo em Parcelas" and descricoes_parcelas:
                            DELIMITADOR = "|||"
                            tags_extra.append(f"descricoes:{DELIMITADOR.join(descricoes_parcelas)}")

                        if metodo == "Valor Fixo em Parcelas" and var_tem_entrada.get():
                            desc_entrada_safe = descricao_entrada.get().replace("|||", " ")
                            tags_extra.append(f"desc_entrada:{desc_entrada_safe}")

                        tags_finais = ['percentual', forma_pagto_selecionada, *tags_extra]

                    except ValueError:
                        custom_messagebox("error", "Erro", "Percentual inválido!")
                        return

                elif tipo_combo.get() == 'Fixo':
                    if not valor_total_entry.get():
                        custom_messagebox("error", "Erro", "Preencha o valor total!")
                        return

                    try:
                        valor_total_adm = float(valor_total_entry.get().replace(',', '.'))
                        if valor_total_adm <= 0:
                            custom_messagebox("error", "Erro", "Valor total deve ser maior que zero!")
                            return
                    except ValueError:
                        custom_messagebox("error", "Erro", "Valor total inválido!")
                        return

                    if metodo == "Valor Fixo em Parcelas":
                        num_parcelas_val = num_parcelas_entry.get()
                        data_inicial = data_entrada.get() if var_tem_entrada.get() else ""
                    else:
                        num_parcelas_val = str(len(eventos))
                        data_inicial = ""

                    valores_finais = (
                        cnpj_cpf_entry.get(),
                        nome_entry.get(),
                        tipo_combo.get(),
                        "",
                        valor_total_entry.get(),
                        num_parcelas_val,
                        data_inicial
                    )

                    if metodo == "Valor Fixo em Parcelas":
                        if var_tem_entrada.get():
                            tags_extra.append(f"entrada:{valor_entrada_entry.get()}")
                            desc_entrada_safe = descricao_entrada.get().replace("|||", " ")
                            tags_extra.append(f"desc_entrada:{desc_entrada_safe}")

                        if descricoes_parcelas:
                            DELIMITADOR = "|||"
                            tags_extra.append(f"descricoes:{DELIMITADOR.join(descricoes_parcelas)}")

                    tags_finais = [
                        'fixo',
                        forma_pagto_selecionada,
                        *tags_extra
                    ]

                if metodo == "Eventos/Fases":
                    eventos_serializados = []
                    for desc, perc, valor in eventos:
                        eventos_serializados.append(f"{desc}:{perc}:{valor}")

                    nova_tag = f"eventos:{'|'.join(eventos_serializados)}"
                    tags_finais = (*tags_finais, nova_tag)

                # Adicionar à tree
                tree.insert('', 'end', values=valores_finais, tags=tags_finais)

                janela_admin.destroy()

                # Restaurar janela_pai se fornecida
                if janela_pai and janela_pai.winfo_exists():
                    janela_pai.deiconify()
                    janela_pai.lift()
                    janela_pai.focus_force()
                else:
                    # Se não tem janela_pai, restaurar janela do contrato
                    if metodo_pagamento_combo.winfo_toplevel().winfo_exists():
                        metodo_pagamento_combo.winfo_toplevel().after(100, lambda: (
                            metodo_pagamento_combo.winfo_toplevel().lift(),
                            metodo_pagamento_combo.winfo_toplevel().focus_force()
                        ))

            except Exception as e:
                import traceback
                logger.debug(traceback.format_exc())
                custom_messagebox("error", "Erro", f"Erro ao confirmar: {str(e)}")

        def cancelar():
            """Cancela e restaura janela_pai se existir"""
            janela_admin.destroy()

            # Restaurar janela_pai
            if janela_pai and janela_pai.winfo_exists():
                janela_pai.deiconify()
                janela_pai.lift()
                janela_pai.focus_force()

        # Botões
        frame_botoes = ttk.Frame(frame_admin)
        frame_botoes.pack(fill='x', pady=10)
        ttk.Button(frame_botoes, text="Confirmar", command=confirmar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=cancelar).pack(side='left', padx=5)

    def remover_administrador(self, tree):
        """
        Remove o administrador selecionado da lista temporária (treeview)
        de um contrato ainda não salvo.

        NOTA DE EXTRAÇÃO: este método não existia em GestaoContratos no
        código original — só existia um método de mesmo nome dentro de
        GestaoAdministradores (classe confirmada como código morto, sem
        nenhuma instanciação no sistema). Ou seja: o botão "Remover
        Administrador" da tela "Novo Contrato" quebrava com AttributeError
        sempre que clicado, antes desta extração. Implementação abaixo
        adaptada ao padrão real desta classe (dados vivem na treeview,
        não em uma lista paralela self.administradores como no método
        órfão de GestaoAdministradores). Testar antes de confiar 100%.
        """
        selecionado = tree.selection()
        if not selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um administrador para remover")
            return

        if custom_messagebox("yesno", "Confirmação", "Deseja remover o administrador selecionado?"):
            tree.delete(selecionado)

    def processar_parcelas_fixas(self, ws, num_contrato, valor_global, opcoes):
        """Processa parcelas fixas para o contrato"""
        try:
            logger.debug("Início de processar_parcelas_fixas")
            logger.debug(f"Opções: {opcoes}")

            num_parcelas = int(opcoes.get('num_parcelas', 0))
            tem_entrada = opcoes.get('tem_entrada', False)
            descricoes_parcelas = opcoes.get('descricoes_parcelas', {})
            DELIMITADOR = "|||"
            logger.debug(f"Processando {num_parcelas} parcelas, entrada: {tem_entrada}")

            if num_parcelas <= 0:
                logger.debug("Erro: Número de parcelas inválido")
                return

            # Processar cada administrador
            for item in self.tree_adm.get_children():
                valores_adm = self.tree_adm.item(item)['values']
                tags_adm = self.tree_adm.item(item)['tags']

                logger.debug(f"Processando administrador: {valores_adm}")

                cnpj_cpf_adm = str(valores_adm[0]).strip()

                apenas_digitos = ''.join(filter(str.isdigit, str(cnpj_cpf_adm)))

                if len(apenas_digitos) == 11:
                    tipo_pessoa = 'PF'
                elif len(apenas_digitos) == 14:
                    tipo_pessoa = 'PJ'
                else:
                    # fallback pela pontuação
                    tipo_pessoa = 'PF' if ('.' in cnpj_cpf_adm and '/' not in cnpj_cpf_adm) else 'PJ'

                cnpj_cpf_adm = formatar_documento(
                    normalizar_documento(cnpj_cpf_adm, tipo_pessoa), tipo_pessoa
                )
                nome_adm = valores_adm[1]

                # Extrair descricoes das tags, se existirem
                descricoes_individuais = []
                for tag in tags_adm:
                    if tag.startswith('descricoes:'):
                        # Usar DELIMITADOR ao invés de vírgula
                        descricoes_str = tag.replace('descricoes:', '')
                        descricoes_individuais = descricoes_str.split(DELIMITADOR)
                        logger.debug(f"Descrições individuais: {descricoes_individuais}")
                        break

                # Também verificar no dicionário de descrições
                if not descricoes_individuais and cnpj_cpf_adm in descricoes_parcelas:
                    descricoes_individuais = descricoes_parcelas[cnpj_cpf_adm]
                    logger.debug(f"Usando descrições do dicionário: {descricoes_individuais}")

                # Extrair descrição da entrada, se existir
                descricao_entrada = "ENTRADA"
                for tag in tags_adm:
                    if tag.startswith('desc_entrada:'):
                        descricao_entrada = tag.replace('desc_entrada:', '')
                        logger.debug(f"Descrição da entrada: {descricao_entrada}")
                        break

                # Calcular valor por parcela para este administrador
                try:
                    if valores_adm[2] == 'Percentual':
                        # Administrador com percentual do valor total
                        perc_adm = float(str(valores_adm[3]).replace('%', '').replace(',', '.'))
                        valor_total_adm = (perc_adm / 100) * valor_global
                        logger.debug(f"Valor calculado baseado em percentual: {valor_total_adm}")
                    else:  # Fixo
                        # Valor fixo total para o administrador
                        valor_texto = str(valores_adm[4]).replace(',', '.')
                        logger.debug(f"Valor texto: {valor_texto}")
                        valor_total_adm = float(valor_texto)
                        logger.debug(f"Valor fixo: {valor_total_adm}")
                except (ValueError, TypeError, IndexError) as e:
                    logger.debug(f"Erro ao calcular valor: {e}")
                    valores_str = ', '.join([str(v) for v in valores_adm])
                    logger.debug(f"Valores disponíveis: {valores_str}")
                    # Tentar alternativa
                    if len(valores_adm) >= 5 and valores_adm[4]:
                        try:
                            valor_total_adm = float(str(valores_adm[4]).replace(',', '.'))
                            logger.debug(f"Valor alternativo: {valor_total_adm}")
                        except (ValueError, TypeError):
                            logger.debug("Erro na alternativa também")
                            valor_total_adm = 0
                    else:
                        valor_total_adm = 0

                if valor_total_adm <= 0:
                    logger.debug("Valor total inválido, pulando administrador")
                    continue

                # Total de lançamentos = entrada (se houver) + parcelas
                total_lancamentos = num_parcelas + (1 if tem_entrada else 0)

                # Se tem entrada, tratar separadamente
                if tem_entrada:
                    valor_entrada = 0
                    # Buscar valor da entrada nas tags
                    for tag in tags_adm:
                        if tag.startswith('entrada:'):
                            try:
                                valor_entrada = float(tag.replace('entrada:', '').replace(',', '.'))
                                logger.debug(f"Valor da entrada das tags: {valor_entrada}")
                            except ValueError:
                                valor_entrada = 0
                            break

                    if valor_entrada <= 0:
                        # Calcular proporcional se não estiver explícito
                        valor_entrada_opcoes = opcoes.get('valor_entrada', 0)
                        if isinstance(valor_entrada_opcoes, str):
                            valor_entrada_opcoes = float(valor_entrada_opcoes.replace(',', '.'))
                        # Proporcional da entrada para este administrador
                        proporcao_entrada = valor_entrada_opcoes / valor_global if valor_global else 0
                        valor_entrada_adm = valor_total_adm * proporcao_entrada
                        logger.debug(f"Valor da entrada calculado: {valor_entrada_adm}")
                    else:
                        # Usar o valor específico
                        valor_entrada_adm = valor_entrada

                    data_entrada_obj = opcoes.get('data_entrada')

                    # Garantir que é um objeto datetime sem hora
                    if isinstance(data_entrada_obj, str):
                        try:
                            data_entrada_obj = datetime.strptime(data_entrada_obj, '%d/%m/%Y')
                        except:
                            data_entrada_obj = datetime.now()

                    # Converter para date (apenas data, sem hora)
                    if isinstance(data_entrada_obj, datetime):
                        data_entrada_obj = data_entrada_obj.date()

                    # Registrar entrada como parcela 0
                    proxima_linha = ws.max_row + 1
                    ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())
                    ws.cell(row=proxima_linha, column=26, value=0)  # Número 0 para entrada
                    ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)
                    ws.cell(row=proxima_linha, column=28, value=nome_adm)

                    # Gravar data como date e aplicar formato
                    ws.cell(row=proxima_linha, column=29, value=data_entrada_obj)
                    ws.cell(row=proxima_linha, column=29).number_format = 'DD/MM/YYYY'

                    ws.cell(row=proxima_linha, column=30, value=valor_entrada_adm)
                    ws.cell(row=proxima_linha, column=31, value='PENDENTE')
                    ws.cell(row=proxima_linha, column=32, value=None)
                    ws.cell(row=proxima_linha, column=33,
                            value=f"ADM. OBRA - PARCELA 1/{total_lancamentos}")

                    logger.debug(f"Registrada entrada (parcela 0) com valor {valor_entrada_adm} e data {data_entrada_obj}")

                    # Ajustar valor restante para as demais parcelas
                    valor_restante = valor_total_adm - valor_entrada_adm
                    valor_parcela = valor_restante / num_parcelas if num_parcelas > 0 else 0

                    # Registrar parcelas começando em 1
                    for i in range(1, num_parcelas + 1):
                        proxima_linha = ws.max_row + 1
                        ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())
                        ws.cell(row=proxima_linha, column=26, value=i)  # Parcelas de 1 a N
                        ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)
                        ws.cell(row=proxima_linha, column=28, value=nome_adm)
                        ws.cell(row=proxima_linha, column=29, value=None)
                        ws.cell(row=proxima_linha, column=30, value=valor_parcela)
                        ws.cell(row=proxima_linha, column=31, value='PENDENTE')
                        ws.cell(row=proxima_linha, column=32, value=None)

                        # Usar descrição individual se disponível
                        if i-1 < len(descricoes_individuais) and descricoes_individuais[i-1]:
                            descricao = descricoes_individuais[i-1]
                        else:
                            descricao = f"PARCELA {i}"

                        pos = i + 1  # entrada é posição 1, parcelas começam em 2
                        ws.cell(row=proxima_linha, column=33,
                                value=f"ADM. OBRA - PARCELA {pos}/{total_lancamentos}")
                        logger.debug(f"Registrada parcela {i} com valor {valor_parcela} e descrição '{descricao}'")

                else:
                    # Sem entrada, parcelas começam em 1 normalmente
                    valor_parcela = valor_total_adm / num_parcelas if num_parcelas > 0 else 0

                    for i in range(1, num_parcelas + 1):
                        proxima_linha = ws.max_row + 1
                        ws.cell(row=proxima_linha, column=25, value=num_contrato.upper())
                        ws.cell(row=proxima_linha, column=26, value=i)  # Parcelas de 1 a N
                        ws.cell(row=proxima_linha, column=27, value=cnpj_cpf_adm)  # CNPJ/CPF
                        ws.cell(row=proxima_linha, column=28, value=nome_adm)  # Nome
                        ws.cell(row=proxima_linha, column=29, value=None)  # Data vencimento (a definir)
                        ws.cell(row=proxima_linha, column=30, value=valor_parcela)  # Valor
                        ws.cell(row=proxima_linha, column=31, value='PENDENTE')  # Status
                        ws.cell(row=proxima_linha, column=32, value=None)  # Sem evento

                        # Usar descrição individual se disponível
                        if i-1 < len(descricoes_individuais) and descricoes_individuais[i-1]:
                            descricao = descricoes_individuais[i-1]
                        else:
                            descricao = f"PARCELA {i}"

                        ws.cell(row=proxima_linha, column=33,
                                value=f"ADM. OBRA - PARCELA {i}/{total_lancamentos}")
                        logger.debug(f"Registrada parcela {i} com valor {valor_parcela} e descrição '{descricao}'")

            logger.debug("Finalizado processamento de parcelas fixas com sucesso")
        except Exception as e:
            import traceback
            logger.debug(traceback.format_exc())
            logger.debug(f"Erro em processar_parcelas_fixas: {str(e)}")

    def processar_administradores(self, ws, num_contrato, valor_global, metodo_pagamento, opcoes):
        """Processa os administradores do contrato"""
        for item in self.tree_adm.get_children():
            valores = self.tree_adm.item(item)['values']
            tags = self.tree_adm.item(item)['tags']

            # Formatação do CNPJ/CPF
            cnpj_cpf = str(valores[0]).strip()

            tags_adm = self.tree_adm.item(item)['tags']
            tipo_pessoa = next((tag for tag in tags_adm if tag in ['PF', 'PJ']), None)

            if tipo_pessoa:
                cnpj_cpf_adm = self._formatar_documento_admin(valores[0], tipo_pessoa)
            else:
                # Fallback: buscar na base
                tipo_pessoa = self._obter_tipo_pessoa_da_base(valores[0])
                cnpj_cpf_adm = self._formatar_documento_admin(valores[0], tipo_pessoa)

            nome_admin = valores[1]

            # Buscar dados bancários do fornecedor
            forma_pagamento = next((tag for tag in tags if tag in ['PIX', 'TED']), 'PIX')
            dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj_cpf, forma_pagamento)

            # Registrar administrador no contrato com os dados apropriados
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=7, value=num_contrato.upper())  # Contrato
            ws.cell(row=proxima_linha, column=8, value=cnpj_cpf)              # CNPJ/CPF
            ws.cell(row=proxima_linha, column=9, value=nome_admin)            # Nome
            ws.cell(row=proxima_linha, column=10, value=valores[2])           # Tipo (Percentual/Fixo)
            ws.cell(row=proxima_linha, column=11, value=valores[3])           # Valor/Percentual
            ws.cell(row=proxima_linha, column=12, value=valores[4])           # Valor Total
            ws.cell(row=proxima_linha, column=13, value=valores[5])           # Número de parcelas

            # Usar data da entrada das opções, não data atual
            data_inicial_gravacao = None

            if valores[6] and metodo_pagamento == "Valor Fixo em Parcelas" and opcoes.get('tem_entrada'):
                # Pegar a data informada pelo usuário
                data_entrada_usuario = opcoes.get('data_entrada')

                if data_entrada_usuario:
                    # Se for string, converter para datetime
                    if isinstance(data_entrada_usuario, str):
                        try:
                            data_inicial_gravacao = data_entrada_usuario.date()
                        except:
                            data_inicial_gravacao = data_entrada_usuario
                    else:
                        data_inicial_gravacao = data_entrada_usuario

                if data_inicial_gravacao:
                    ws.cell(row=proxima_linha, column=14, value=data_inicial_gravacao)
                    ws.cell(row=proxima_linha, column=14).number_format = 'DD/MM/YYYY'

    def salvar_contrato_com_opcoes(self, num_contrato, data_inicio, data_fim, observacoes, valor_global, metodo_pagamento, opcoes, janela):
        """Salva os dados do contrato com diferentes opções de pagamento"""
        num_contrato = str(num_contrato).upper()

        try:
            logger.debug(f"Salvando contrato: {num_contrato}, método: {metodo_pagamento}")

            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']

            # Verificar se o contrato já existe
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).upper() == num_contrato.upper():
                    custom_messagebox("error", "Erro", "Número de contrato já existe!")
                    return

            # Salvar dados do contrato
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1, value=num_contrato.upper())

            # Data de início como date
            data_inicio_date = data_inicio.date() if isinstance(data_inicio, datetime) else data_inicio
            ws.cell(row=proxima_linha, column=2, value=data_inicio_date)
            ws.cell(row=proxima_linha, column=2).number_format = 'DD/MM/YYYY'

            # Data de fim como date
            data_fim_date = data_fim.date() if isinstance(data_fim, datetime) else data_fim
            ws.cell(row=proxima_linha, column=3, value=data_fim_date)
            ws.cell(row=proxima_linha, column=3).number_format = 'DD/MM/YYYY'

            ws.cell(row=proxima_linha, column=4, value='ATIVO')
            ws.cell(row=proxima_linha, column=5, value=observacoes)
            ws.cell(row=proxima_linha, column=6, value=valor_global)  # Valor global do contrato

            # Extrair dados dos administradores, incluindo descrições de parcelas
            opcoes_processadas = opcoes.copy() if opcoes else {}

            # Extrair informações adicionais dos administradores
            tem_entrada = False
            valor_entrada = 0
            data_entrada = None
            num_parcelas = 0

            # Coletar informações específicas para parcelas fixas
            if metodo_pagamento == "Valor Fixo em Parcelas":
                # Percorrer os administradores para extrair informações de parcelas
                for item in self.tree_adm.get_children():
                    valores = self.tree_adm.item(item)['values']
                    tags = self.tree_adm.item(item)['tags']

                    # Extrair número de parcelas
                    if valores[5] and not num_parcelas:
                        try:
                            num_parcelas = int(valores[5])
                        except (ValueError, TypeError):
                            num_parcelas = 0

                    # Verificar se tem entrada
                    for tag in tags:
                        if tag.startswith('entrada:'):
                            tem_entrada = True
                            try:
                                valor_entrada = float(tag.replace('entrada:', '').replace(',', '.'))
                            except (ValueError, TypeError):
                                valor_entrada = 0

                    # Extrair data de entrada
                    if valores[6] and not data_entrada:
                        data_entrada_str = valores[6]

                        # Converter para datetime se necessário
                        if isinstance(data_entrada_str, str):
                            try:
                                # Tentar formato dd/mm/yyyy
                                data_entrada = datetime.strptime(data_entrada_str, '%d/%m/%Y').date()
                            except ValueError:
                                try:
                                    # Tentar outros formatos possíveis
                                    data_entrada = datetime.strptime(data_entrada_str, '%Y-%m-%d').date()
                                except:
                                    # Fallback: data atual
                                    data_entrada = datetime.now().date()
                        elif isinstance(data_entrada_str, datetime):
                            # Se já for datetime, extrair apenas a data
                            data_entrada = data_entrada_str.date()
                        else:
                            # Se já for date, usar diretamente
                            data_entrada = data_entrada_str

                # Adicionar ao dicionário de opções
                opcoes_processadas['num_parcelas'] = num_parcelas
                opcoes_processadas['tem_entrada'] = tem_entrada
                opcoes_processadas['valor_entrada'] = valor_entrada
                opcoes_processadas['data_entrada'] = data_entrada

                logger.debug(f"Configurações de parcelas: parcelas={num_parcelas}, entrada={tem_entrada}, valor_entrada={valor_entrada}")

            # Coletar descrições para cada administrador
            admin_descricoes = {}

            for item in self.tree_adm.get_children():
                tags = self.tree_adm.item(item)['tags']
                cnpj_cpf = self.tree_adm.item(item)['values'][0]

                # Extrair descricoes das tags, se existirem
                for tag in tags:
                    if tag.startswith('descricoes:'):
                        admin_descricoes[cnpj_cpf] = tag.replace('descricoes:', '').split('|||')
                        logger.debug(f"Descrições para {cnpj_cpf}: {admin_descricoes[cnpj_cpf]}")
                        break

            # Adicionar ao dicionário de opções
            opcoes_processadas['descricoes_parcelas'] = admin_descricoes

            # Processar administradores baseado no método de pagamento
            self.processar_administradores(ws, num_contrato, valor_global, metodo_pagamento, opcoes_processadas)

            # Processar eventos se método for por eventos/fases
            if metodo_pagamento == "Eventos/Fases":
                # Criar dicionário com eventos E valor total por administrador
                eventos_por_admin = {}

                for item in self.tree_adm.get_children():
                    valores_adm = self.tree_adm.item(item)['values']
                    tags_adm = self.tree_adm.item(item)['tags']
                    cnpj_cpf = str(valores_adm[0]).strip()
                    nome_adm = valores_adm[1]

                    # Extrair valor total deste administrador
                    if valores_adm[2] == 'Percentual':
                        # Calcular baseado no percentual
                        perc_adm = float(str(valores_adm[3]).replace('%', '').replace(',', '.'))
                        valor_total_admin = (perc_adm / 100) * valor_global
                    else:  # Fixo
                        # Usar valor fixo informado
                        try:
                            valor_total_admin = float(str(valores_adm[4]).replace('.', '').replace(',', '.'))
                        except (ValueError, TypeError, IndexError):
                            valor_total_admin = 0

                    # Extrair eventos DESTE administrador específico
                    eventos_list = []
                    for tag in tags_adm:
                        if tag.startswith('eventos:'):
                            eventos_str = tag.replace('eventos:', '')

                            for evento_str in eventos_str.split('|'):
                                partes = evento_str.split(':')
                                if len(partes) == 3:
                                    desc, perc, valor = partes
                                    # Valor aqui é apenas referência, será recalculado
                                    eventos_list.append((desc, float(perc), float(valor)))
                            break

                    # Armazenar eventos com informações completas do admin
                    if eventos_list:
                        eventos_por_admin[cnpj_cpf] = {
                            'eventos': eventos_list,
                            'nome': nome_adm,
                            'valor_total': valor_total_admin
                        }

                # Processar eventos por administrador (sem duplicação e com valores corretos)
                self.processar_eventos(ws, num_contrato, eventos_por_admin)

            # Processar parcelas fixas se for o método apropriado
            elif metodo_pagamento == "Valor Fixo em Parcelas":
                logger.debug("Chamando processar_parcelas_fixas...")
                self.processar_parcelas_fixas(ws, num_contrato, valor_global, opcoes_processadas)

            # Salvar e fechar o arquivo explicitamente
            try:
                logger.debug(f"Salvando o arquivo {self.arquivo_cliente}")
                wb.save(self.arquivo_cliente)
                wb.close()  # Importante fechar o arquivo
            except PermissionError:
                custom_messagebox("error", "Erro", f"Não foi possível salvar a planilha. Ela pode estar aberta em outro programa.")
                return
            except Exception as e:
                import traceback
                logger.debug(traceback.format_exc())
                custom_messagebox("error", "Erro", f"Erro ao salvar planilha: {str(e)}")
                return

            # Exibir mensagem de sucesso
            custom_messagebox("info", "Sucesso", "Contrato cadastrado com sucesso!")

            # Fechar a janela atual
            janela.destroy()

            # Garantir que a janela de gestão de contratos é trazida para frente
            # após salvar o contrato e recarregar a lista
            self.carregar_contratos()

            # Usar after para garantir que toda a interface seja atualizada
            if self.parent and self.parent.winfo_exists():
                self.parent.after(100, lambda: (
                    self.parent.lift(),
                    self.parent.focus_force()
                ))

        except Exception as e:
            import traceback
            traceback.print_exc()  # Imprime o stack trace completo
            custom_messagebox("error", "Erro", f"Erro ao salvar contrato: {str(e)}")
            if 'wb' in locals() and wb:
                try:
                    wb.close()
                except:
                    pass

    def excluir_contrato(self):
        """Exclui o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            custom_messagebox("warning",  "Aviso", "Selecione um contrato para excluir")
            return

        if custom_messagebox("yesno", "Confirmação",
                              "Deseja realmente excluir este contrato e seus administradores?"):
            try:
                num_contrato = self.tree_contratos.item(selecionado)['values'][0]

                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']

                # Marcar contrato como inativo
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == num_contrato:
                        row[3].value = 'INATIVO'  # Coluna D - Status

                wb.save(self.arquivo_cliente)
                self.carregar_contratos()
                custom_messagebox("info", "Sucesso", "Contrato marcado como inativo")

            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao excluir contrato: {str(e)})")

    def gerar_contrato_adm(self):
        """Gera o contrato Word do contrato selecionado na lista."""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            custom_messagebox("warning", "Aviso", "Selecione um contrato para gerar!")
            return

        num_contrato = self.tree_contratos.item(selecionado)['values'][0]
        raiz = num_contrato.rstrip('JF')

        try:
            # Import tardio: GeradorContratoADM ainda não foi extraída para
            # seu próprio módulo (será a próxima extração). Ver nota no
            # topo deste arquivo.
            from src.taxas_administracao.gerador_contrato_adm import GeradorContratoADM

            gerador = GeradorContratoADM()
            gerador.PASTA_CONTRATOS_ADM = PASTA_CLIENTES / "Contratos_ADM"

            paths = gerador.gerar_contratos_do_contrato(
                nome_cliente           = self.cliente_atual,
                num_contrato           = raiz,
                arquivo_cliente        = Path(self.arquivo_cliente),
                arquivo_clientes_geral = ARQUIVO_CLIENTES,
                arquivo_fornecedores   = BASE_PATH / 'base_fornecedores.xlsx',
            )
            if paths:
                nomes = "\n".join(Path(p).name for p in paths)
                custom_messagebox("info", "Sucesso",
                    f"{len(paths)} contrato(s) gerado(s):\n{nomes}")
                for p in paths:
                    try:
                        os.startfile(p)
                    except Exception:
                        subprocess.run(['xdg-open', p], check=False)
            else:
                custom_messagebox("warning", "Aviso",
                    "Nenhum contrato gerado. Verifique os dados da planilha.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            custom_messagebox("error", "Erro", f"Erro ao gerar contrato:\n{str(e)}")
