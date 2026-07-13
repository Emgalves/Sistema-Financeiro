"""
Motor de cálculo das taxas de administração: base de cálculo, percentual
contratado, recálculo quando a base muda, e verificação de consistência
entre o valor esperado e o valor efetivamente lançado na planilha.

Extraído de Sistema_Entrada_Dados.py em [DATA_DA_EXTRACAO].
Nenhuma alteração de lógica foi feita nesta extração — apenas mudança
de localização e ajuste de imports.

Dependência cruzada não resolvida por esta extração:
    recalcular_taxas_afetadas() chama
    self.sistema.obter_administradores_cliente_CORRIGIDO(cliente),
    método que permanece em SistemaEntradaDados. Isso continua
    funcionando normalmente porque a instância de GestorTaxasAdministracao
    recebe `sistema_principal` (self.sistema) no __init__, exatamente
    como antes da extração.
"""
import logging
import os
from datetime import datetime

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

from src.config.config import PASTA_CLIENTES

# Mesmo logger usado no restante do sistema (Sistema_Entrada_Dados.py).
# Não definimos nível aqui: ele herda o nível efetivo do logger raiz,
# configurado uma única vez no arquivo principal (logging.basicConfig).
logger = logging.getLogger("sistema")


class GestorTaxasAdministracao:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal

    def recalcular_taxas_afetadas(self, data_referencia, cliente=None, mostrar_detalhes=True):
        """
        VERSÃO CORRIGIDA do recálculo de taxas usando a lógica validada do finalizacao_quinzena.py
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual

            if not cliente:
                return {"sucesso": False, "mensagem": "Nenhum cliente especificado"}

            logger.debug(f"DEBUG: Iniciando recálculo de taxas para {cliente} em {data_referencia}")

            # CORREÇÃO 1: Chamar o método corretamente (sem parâmetro self extra)
            novo_valor_base = self.calcular_base_calculo_taxa(data_referencia)
            logger.debug(f"DEBUG: Nova base calculada: R$ {novo_valor_base:.2f}")

            if novo_valor_base == 0:
                return {"sucesso": True, "mensagem": "Sem lançamentos base para recálculo"}

            # 2. Obter percentual usando método corrigido
            percentual_taxa = self.obter_percentual_taxa_cliente(cliente)
            logger.debug(f"DEBUG: Percentual encontrado: {percentual_taxa}%")

            if percentual_taxa == 0:
                return {"sucesso": True, "mensagem": "Sem taxa percentual configurada"}

            # 3. Calcular novo valor da taxa
            novo_valor_taxa = novo_valor_base * (percentual_taxa / 100)
            logger.debug(f"DEBUG: Novo valor da taxa: R$ {novo_valor_taxa:.2f}")

            # 4. Verificar valor atual das taxas na planilha
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_dados = wb["Dados"]

            # Converter data
            if isinstance(data_referencia, str):
                data_ref = datetime.strptime(data_referencia, '%d/%m/%Y')
            else:
                data_ref = data_referencia

            # Buscar taxas existentes (tipo 7) na data
            valor_atual_total = 0
            linhas_taxa = []

            for idx, row in enumerate(ws_dados.iter_rows(min_row=2, values_only=True), start=2):
                data_lancamento = row[0]
                if isinstance(data_lancamento, datetime):
                    if (data_lancamento.day == data_ref.day and
                        data_lancamento.month == data_ref.month and
                        data_lancamento.year == data_ref.year):

                        tipo_desp = row[1]
                        if tipo_desp == 7:  # Taxa ADM
                            # CORREÇÃO 2: Verificar status antes de incluir no valor atual
                            status = row[13] if len(row) > 13 else "ATIVO"  # Coluna N (STATUS)

                            if status == "ATIVO":  # Só considerar taxas ativas
                                valor = row[8]  # Coluna I
                                if valor:
                                    valor_numeric = float(str(valor).replace(',', '.'))
                                    valor_atual_total += valor_numeric
                                    linhas_taxa.append(idx)

            logger.debug(f"DEBUG: Valor atual total das taxas ATIVAS: R$ {valor_atual_total:.2f}")

            # 5. Verificar se precisa recalcular
            diferenca = abs(novo_valor_taxa - valor_atual_total)

            if diferenca < 0.01:  # Diferença menor que 1 centavo
                wb.close()
                return {"sucesso": True, "mensagem": f"Taxas já estão corretas (R$ {valor_atual_total:.2f})"}

            # 6. Se chegou aqui, precisa recalcular
            logger.debug(f"DEBUG: Diferença detectada: R$ {diferenca:.2f}")

            # CORREÇÃO 3: Marcar como excluído ao invés de deletar fisicamente
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

            for linha in linhas_taxa:
                # Marcar como EXCLUIDO
                ws_dados.cell(row=linha, column=14, value='EXCLUIDO')  # STATUS

                # Atualizar histórico
                historico_atual = ws_dados.cell(row=linha, column=16).value or ""
                novo_historico = f"{historico_atual} | EXCLUÍDA P/ RECÁLCULO EM: {timestamp}" if historico_atual else f"EXCLUÍDA P/ RECÁLCULO EM: {timestamp}"
                ws_dados.cell(row=linha, column=16, value=novo_historico)

            logger.debug(f"DEBUG: {len(linhas_taxa)} linhas de taxa marcadas como excluídas")

            # 7. Obter administradores e lançar novas taxas
            administradores = self.sistema.obter_administradores_cliente_CORRIGIDO(cliente)

            if not administradores:
                wb.close()
                return {"sucesso": False, "mensagem": "Nenhum administrador encontrado"}

            # 8. Lançar novas taxas (usar a mesma lógica do finalizacao_quinzena.py)
            taxa_total_percentual = sum(adm['percentual'] for adm in administradores)

            for adm in administradores:
                valor_adm = (novo_valor_taxa * adm['percentual']) / taxa_total_percentual

                # Determinar data de vencimento e quinzena
                if data_ref.day == 5:
                    dt_vencto = data_ref
                    while dt_vencto.weekday() >= 5:  # Ajustar fim de semana
                        dt_vencto += relativedelta(days=1)
                    quinzena = "1ª"
                else:
                    dt_vencto = data_ref
                    quinzena = "2ª"

                # Inserir nova linha
                proxima_linha = ws_dados.max_row + 1

                # CORREÇÃO 4: Gerar ID sequencial consistente
                id_lancamento = self._obter_proximo_id_sequencial(ws_dados)

                # Preencher dados (mesma estrutura do finalizacao_quinzena.py)
                ws_dados.cell(row=proxima_linha, column=1, value=data_ref)
                ws_dados.cell(row=proxima_linha, column=1).number_format = 'DD/MM/YYYY'
                ws_dados.cell(row=proxima_linha, column=2, value=7)  # Tipo taxa ADM
                ws_dados.cell(row=proxima_linha, column=3, value=adm['cnpj_cpf'])
                ws_dados.cell(row=proxima_linha, column=4, value=adm['nome'])

                referencia = f"ADM. OBRA REF. {quinzena} QUINZ. {data_ref.strftime('%m/%Y')}"
                ws_dados.cell(row=proxima_linha, column=5, value=referencia)
                ws_dados.cell(row=proxima_linha, column=6, value='')  # NF

                ws_dados.cell(row=proxima_linha, column=7, value=valor_adm)
                ws_dados.cell(row=proxima_linha, column=7).number_format = '#,##0.00'
                ws_dados.cell(row=proxima_linha, column=8, value=1)  # Dias
                ws_dados.cell(row=proxima_linha, column=9, value=valor_adm)
                ws_dados.cell(row=proxima_linha, column=9).number_format = '#,##0.00'

                ws_dados.cell(row=proxima_linha, column=10, value=dt_vencto)
                ws_dados.cell(row=proxima_linha, column=10).number_format = 'DD/MM/YYYY'
                ws_dados.cell(row=proxima_linha, column=11, value='ADM')

                # Buscar dados bancários
                from src.config.utils import buscar_dados_bancarios_fornecedor
                dados_bancarios = buscar_dados_bancarios_fornecedor(adm['cnpj_cpf'])
                ws_dados.cell(row=proxima_linha, column=12, value=dados_bancarios)

                # CORREÇÃO 5: Observação mais detalhada
                obs = f"RECÁLCULO AUTO - BASE: R$ {novo_valor_base:.2f} - {timestamp}"
                ws_dados.cell(row=proxima_linha, column=13, value=obs)

                # CORREÇÃO 6: Status e ID
                ws_dados.cell(row=proxima_linha, column=14, value='ATIVO')  # STATUS
                ws_dados.cell(row=proxima_linha, column=15, value=id_lancamento)  # ID_LANCAMENTO

                # Histórico inicial
                historico_inicial = f"CRIADO POR RECÁLCULO EM: {timestamp}"
                ws_dados.cell(row=proxima_linha, column=16, value=historico_inicial)

                logger.debug(f"DEBUG: Taxa lançada para {adm['nome']}: R$ {valor_adm:.2f} (ID: {id_lancamento})")

            # Salvar arquivo
            wb.save(arquivo_cliente)

            mensagem = f"Taxas recalculadas com sucesso! "
            mensagem += f"Base: R$ {novo_valor_base:.2f} | "
            mensagem += f"Taxa: {percentual_taxa}% | "
            mensagem += f"Valor total: R$ {novo_valor_taxa:.2f}"

            return {"sucesso": True, "mensagem": mensagem}

        except Exception as e:
            if 'wb' in locals():
                wb.close()
            logger.debug(f"DEBUG: Erro no recálculo: {str(e)}")
            import traceback
            logger.debug(f"DEBUG: Traceback completo: {traceback.format_exc()}")
            return {"sucesso": False, "mensagem": f"Erro no recálculo: {str(e)}"}

    def _obter_proximo_id_sequencial(self, worksheet):
        """
        Obtém o próximo ID sequencial disponível (compatível com sistema principal)
        """
        try:
            max_id = 0

            # Percorrer coluna 15 (ID_LANCAMENTO) para encontrar o maior ID
            for row in range(2, worksheet.max_row + 1):
                id_valor = worksheet.cell(row=row, column=15).value
                if id_valor is not None:
                    try:
                        id_int = int(float(id_valor))
                        if id_int > max_id:
                            max_id = id_int
                    except (ValueError, TypeError):
                        continue

            return max_id + 1

        except Exception as e:
            logger.debug(f"DEBUG: Erro ao obter próximo ID: {str(e)}")
            # Fallback: usar número da linha como ID
            return worksheet.max_row

    def identificar_lancamentos_taxa_admin(self, df):
        """
        Identifica lançamentos de taxa de administração com padrões mais amplos
        """
        if df.empty:
            return pd.DataFrame()

        mask_taxa = df['TP_DESP'] == 7

        taxas = df[mask_taxa].copy()
        logger.debug(f"DEBUG: Taxas encontradas (tp_desp=7): {len(taxas)} registros")

        if not taxas.empty:
            logger.debug(f"DEBUG: Referências das taxas: {taxas['REFERÊNCIA'].tolist()}")

        return taxas

    def calcular_base_calculo_taxa(self, data_referencia, df=None):
        """
        VERSÃO UNIFICADA - Calcula valor base seguindo a lógica corrigida do finalizacao_quinzena.py

        Parâmetros:
        - data_referencia: Data para cálculo (str ou datetime)
        - df: DataFrame opcional (para compatibilidade com código existente)
            Se não fornecido, lê diretamente da planilha
        """
        try:
            logger.debug(f"DEBUG: Calculando valor base para {data_referencia}")

            # Se DataFrame foi fornecido, usar lógica compatível
            if df is not None:
                return self._calcular_base_por_dataframe(df, data_referencia)

            # Caso contrário, usar lógica corrigida da planilha
            return self._calcular_base_por_planilha(data_referencia)

        except Exception as e:
            logger.debug(f"DEBUG: Erro ao calcular valor base: {str(e)}")
            return 0

    def _calcular_base_por_planilha(self, data_referencia):
        """
        Método interno - Calcula base lendo diretamente da planilha
        (Lógica corrigida do finalizacao_quinzena.py)
        """
        try:
            cliente_atual = self.sistema.cliente_atual

            arquivo_cliente = PASTA_CLIENTES / f"{cliente_atual}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_dados = wb["Dados"]

            # Converter data de referência se necessário
            if isinstance(data_referencia, str):
                data_ref = datetime.strptime(data_referencia, '%d/%m/%Y')
            else:
                data_ref = data_referencia

            logger.debug(f"DEBUG: Data de referência (planilha): {data_ref.strftime('%d/%m/%Y')}")

            valor_base = 0
            lancamentos_encontrados = 0

            # Usar a mesma lógica do finalizacao_quinzena.py
            for row in ws_dados.iter_rows(min_row=2, values_only=True):
                data_lancamento = row[0]  # Coluna A

                if isinstance(data_lancamento, datetime):
                    # Verificar se é da mesma data (dia, mês, ano)
                    if (data_lancamento.day == data_ref.day and
                        data_lancamento.month == data_ref.month and
                        data_lancamento.year == data_ref.year):

                        tipo_desp = row[1]  # Coluna B (TP_DESP)
                        status = row[13] if len(row) > 13 else "ATIVO"  # Coluna N (STATUS)

                        # Incluir apenas tipos 1 a 6 e status ATIVO
                        if (isinstance(tipo_desp, (int, float)) and 1 <= tipo_desp <= 6 and
                            status == "ATIVO"):
                            valor = row[8]  # Coluna I (VALOR)

                            if valor:
                                try:
                                    valor_numeric = float(str(valor).replace(',', '.'))
                                    valor_base += valor_numeric
                                    lancamentos_encontrados += 1

                                    logger.debug(f"DEBUG: Lançamento incluído - Tipo: {tipo_desp}, Valor: R$ {valor_numeric:.2f}")

                                except (ValueError, TypeError) as e:
                                    logger.debug(f"DEBUG: Erro ao processar valor '{valor}': {e}")
                                    continue

            logger.debug(f"DEBUG: Valor base total (planilha): R$ {valor_base:.2f}")
            logger.debug(f"DEBUG: Total de lançamentos incluídos: {lancamentos_encontrados}")

            wb.close()
            return valor_base

        except Exception as e:
            logger.debug(f"DEBUG: Erro ao calcular valor base por planilha: {str(e)}")
            if 'wb' in locals():
                wb.close()
            return 0

    def _calcular_base_por_dataframe(self, df, data_referencia):
        """
        Método interno - Calcula base usando DataFrame fornecido
        (Para compatibilidade com verificações existentes)
        """
        try:
            # Converter data de referência
            if isinstance(data_referencia, str):
                data_ref = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref = pd.to_datetime(data_referencia)

            logger.debug(f"DEBUG: Data de referência (DataFrame): {data_ref.strftime('%d/%m/%Y')}")

            # Garantir que DATA_REL existe e está em formato datetime
            if 'DATA_REL' not in df.columns:
                logger.debug("DEBUG: Coluna DATA_REL não encontrada no DataFrame")
                return 0

            df = df.copy()
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')

            # Filtrar dados para a data específica
            df_data = df[df['DATA_REL'].dt.date == data_ref.date()].copy()

            if df_data.empty:
                logger.debug(f"DEBUG: Nenhum lançamento encontrado para {data_ref.strftime('%d/%m/%Y')}")
                return 0

            # Garantir que STATUS existe
            if 'STATUS' not in df_data.columns:
                df_data['STATUS'] = 'ATIVO'

            # Filtrar apenas lançamentos ativos e tipos 1-6
            df_base = df_data[
                (df_data['STATUS'] == 'ATIVO') &
                (df_data['TP_DESP'].isin([1, 2, 3, 4, 5, 6]))
            ].copy()

            if df_base.empty:
                logger.debug("DEBUG: Nenhum lançamento ativo dos tipos 1-6 encontrado")
                return 0

            # Converter valores para numérico
            df_base['VALOR_NUM'] = pd.to_numeric(
                df_base['VALOR'].astype(str).str.replace('R$', '').str.replace(',', '.'),
                errors='coerce'
            ).fillna(0)

            valor_base = df_base['VALOR_NUM'].sum()

            logger.debug(f"DEBUG: Valor base total (DataFrame): R$ {valor_base:.2f}")
            logger.debug(f"DEBUG: Lançamentos incluídos: {len(df_base)}")

            return valor_base

        except Exception as e:
            logger.debug(f"DEBUG: Erro ao calcular valor base por DataFrame: {str(e)}")
            return 0

    def excluir_taxas_base_zerada(self, arquivo_cliente, taxas_existentes):
        """
        Marca taxas como excluídas quando a base for zerada
        """
        try:
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']

            taxas_excluidas = []
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

            for _, taxa in taxas_existentes.iterrows():
                id_taxa = taxa.get('ID_LANCAMENTO')
                if pd.isna(id_taxa):
                    continue

                # Encontrar linha na planilha
                for row_num in range(2, ws.max_row + 1):
                    if ws.cell(row=row_num, column=15).value == id_taxa:
                        # Marcar como excluído
                        ws.cell(row=row_num, column=14, value='EXCLUIDO')  # STATUS

                        # Atualizar histórico
                        historico_atual = ws.cell(row=row_num, column=16).value or ""
                        novo_historico = f"{historico_atual} | EXCLUÍDA (BASE ZERADA) EM: {timestamp}" if historico_atual else f"EXCLUÍDA (BASE ZERADA) EM: {timestamp}"
                        ws.cell(row=row_num, column=16, value=novo_historico)

                        taxas_excluidas.append({
                            'id': id_taxa,
                            'referencia': taxa.get('REFERÊNCIA', ''),
                            'valor': taxa.get('VALOR', 0)
                        })
                        break

            wb.save(arquivo_cliente)

            return {
                "sucesso": True,
                "mensagem": f"Taxas excluídas por base zerada: {len(taxas_excluidas)} itens",
                "detalhes": taxas_excluidas,
                "nova_base": 0,
                "novo_valor_total": 0
            }

        except Exception as e:
            return {"sucesso": False, "mensagem": f"Erro ao excluir taxas: {str(e)}"}

    def atualizar_taxas_na_planilha(self, arquivo_cliente, taxas_existentes, novo_valor, nova_base, percentual):
        """
        Atualiza os valores das taxas EXISTENTES na planilha com histórico detalhado

        IMPORTANTE: Este método ATUALIZA taxas já lançadas, não cria novas!
        Quando uma taxa já foi lançada e a base muda, ajustamos o valor da taxa existente.
        """
        try:
            wb = load_workbook(arquivo_cliente)
            ws = wb['Dados']

            taxas_atualizadas = []
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

            logger.debug(f"DEBUG: Atualizando {len(taxas_existentes)} taxas já lançadas")
            logger.debug(f"DEBUG: Novo valor total a distribuir: R$ {novo_valor:,.2f}")

            # Se há múltiplas taxas, distribuir proporcionalmente
            if len(taxas_existentes) > 1:
                logger.debug(f"DEBUG: Distribuindo entre {len(taxas_existentes)} taxas existentes")

                # Calcular total atual das taxas ATIVAS para proporção
                total_atual = 0
                taxas_ativas = []

                for _, taxa in taxas_existentes.iterrows():
                    status = taxa.get('STATUS', 'ATIVO')
                    if status != 'EXCLUIDO':
                        try:
                            valor_atual = float(str(taxa.get('VALOR', 0)).replace(',', '.'))
                            total_atual += valor_atual
                            taxas_ativas.append((taxa, valor_atual))
                        except:
                            taxas_ativas.append((taxa, 0))

                if total_atual == 0:
                    # Se total atual é zero, dividir igualmente entre taxas ativas
                    valor_por_taxa = novo_valor / len(taxas_ativas) if taxas_ativas else 0
                    proporcoes = [valor_por_taxa] * len(taxas_ativas)
                    logger.debug(f"DEBUG: Divisão igual: R$ {valor_por_taxa:,.2f} por taxa")
                else:
                    # Calcular proporcionalmente ao valor atual
                    proporcoes = []
                    for taxa, valor_atual in taxas_ativas:
                        proporcao = (valor_atual / total_atual) * novo_valor
                        proporcoes.append(proporcao)
                        logger.debug(f"DEBUG: Taxa {taxa.get('ID_LANCAMENTO')}: R$ {valor_atual:,.2f} → R$ {proporcao:,.2f}")

                # Usar apenas taxas ativas para atualização
                taxas_para_processar = [(taxa, prop) for (taxa, _), prop in zip(taxas_ativas, proporcoes)]
            else:
                # Apenas uma taxa - usar valor total
                taxa_unica = taxas_existentes.iloc[0]
                if taxa_unica.get('STATUS', 'ATIVO') != 'EXCLUIDO':
                    taxas_para_processar = [(taxa_unica, novo_valor)]
                    logger.debug(f"DEBUG: Taxa única: R$ {novo_valor:,.2f}")
                else:
                    taxas_para_processar = []
                    logger.debug(f"DEBUG: Taxa única está excluída, não atualizando")

            # Atualizar cada taxa EXISTENTE na planilha
            for taxa, valor_novo in taxas_para_processar:
                id_taxa = taxa.get('ID_LANCAMENTO')
                if pd.isna(id_taxa):
                    logger.debug(f"DEBUG: Taxa sem ID, pulando")
                    continue

                logger.debug(f"DEBUG: Procurando taxa ID {id_taxa} na planilha")

                # Encontrar linha na planilha pelo ID
                linha_encontrada = False
                for row_num in range(2, ws.max_row + 1):
                    id_na_planilha = ws.cell(row=row_num, column=15).value  # ID_LANCAMENTO

                    if id_na_planilha == id_taxa:
                        linha_encontrada = True
                        valor_antigo = ws.cell(row=row_num, column=9).value or 0  # VALOR

                        logger.debug(f"DEBUG: Encontrada linha {row_num}, atualizando valor: R$ {valor_antigo:,.2f} → R$ {valor_novo:,.2f}")

                        # ATUALIZAR O VALOR DA TAXA EXISTENTE
                        ws.cell(row=row_num, column=9, value=round(valor_novo, 2))

                        # Se for tipo 1 (com dias), atualizar valor unitário também
                        tp_desp = ws.cell(row=row_num, column=2).value
                        if tp_desp == 1:
                            dias = ws.cell(row=row_num, column=8).value or 1
                            if dias > 0:
                                vr_unit_novo = round(valor_novo / dias, 2)
                                ws.cell(row=row_num, column=7, value=vr_unit_novo)
                                logger.debug(f"DEBUG: Valor unitário atualizado: R$ {vr_unit_novo:,.2f}")

                        # Garantir que status seja ATIVO (caso tenha sido excluído por engano)
                        status_atual = ws.cell(row=row_num, column=14).value
                        if status_atual != 'ATIVO':
                            ws.cell(row=row_num, column=14, value='ATIVO')
                            logger.debug(f"DEBUG: Status corrigido de {status_atual} para ATIVO")

                        # Atualizar observação com informações detalhadas do recálculo
                        obs_atual = ws.cell(row=row_num, column=13).value or ""
                        # Limpar observações de recálculos anteriores para evitar texto muito longo
                        if "RECALC:" in obs_atual:
                            obs_base = obs_atual.split(" - RECALC:")[0]
                        else:
                            obs_base = obs_atual

                        nova_obs = f"{obs_base} - TAXA ADM {percentual}% - BASE: R$ {nova_base:,.2f} - RECALC: {timestamp}".strip()
                        ws.cell(row=row_num, column=13, value=nova_obs)

                        # Atualizar histórico de alterações
                        historico_atual = ws.cell(row=row_num, column=16).value or ""
                        acao = f"RECALC AUTO: R$ {valor_antigo:,.2f} → R$ {valor_novo:,.2f} (Base: R$ {nova_base:,.2f}) - {timestamp}"

                        if historico_atual:
                            # Limitar histórico para não ficar muito longo (manter últimas 5 ações)
                            historico_partes = historico_atual.split(' | ')
                            if len(historico_partes) >= 5:
                                historico_partes = historico_partes[-4:]  # Manter últimas 4
                            novo_historico = ' | '.join(historico_partes) + ' | ' + acao
                        else:
                            novo_historico = acao

                        ws.cell(row=row_num, column=16, value=novo_historico)

                        taxas_atualizadas.append({
                            'id': id_taxa,
                            'linha': row_num,
                            'referencia': taxa.get('REFERÊNCIA', ''),
                            'valor_antigo': valor_antigo,
                            'valor_novo': valor_novo,
                            'diferenca': valor_novo - float(valor_antigo),
                        })

                        logger.debug(f"✅ Taxa ID {id_taxa} atualizada com sucesso na linha {row_num}")
                        break

                if not linha_encontrada:
                    logger.debug(f"❌ ERRO: Taxa ID {id_taxa} não encontrada na planilha!")
                    # Isso é um problema - taxa existe no DataFrame mas não na planilha
                    # Pode indicar inconsistência nos dados

            # Salvar alterações na planilha
            wb.save(arquivo_cliente)
            logger.debug(f"✅ Planilha salva com {len(taxas_atualizadas)} taxas atualizadas")

            return {
                "sucesso": True,
                "mensagem": f"Taxas EXISTENTES recalculadas: {len(taxas_atualizadas)} itens atualizados",
                "detalhes": taxas_atualizadas,
                "nova_base": nova_base,
                "novo_valor_total": novo_valor,
                "percentual": percentual,
                "observacao": "ATUALIZAÇÃO de taxas já lançadas, não criação de novas taxas"
            }

        except Exception as e:
            import traceback
            logger.debug(f"DEBUG: Erro ao atualizar taxas existentes: {traceback.format_exc()}")
            return {"sucesso": False, "mensagem": f"Erro ao atualizar taxas na planilha: {str(e)}"}

    def criar_nova_taxa_se_necessario(self, data_referencia, cliente=None):
        """
        MÉTODO SEPARADO: Cria nova taxa apenas quando não existe nenhuma para a data

        Este método deve ser usado apenas quando:
        1. Não existe nenhuma taxa para a data/quinzena
        2. O usuário está finalizando a quinzena pela primeira vez

        NÃO usar este método quando já existem taxas lançadas!
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual

            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"

            # Verificar se já existem taxas para esta data
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')

            if isinstance(data_referencia, str):
                data_ref_dt = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref_dt = pd.to_datetime(data_referencia)

            df_data = df[df['DATA_REL'].dt.date == data_ref_dt.date()]
            taxas_existentes = self.identificar_lancamentos_taxa_admin(df_data)

            if not taxas_existentes.empty:
                return {
                    "sucesso": False,
                    "mensagem": f"Já existem {len(taxas_existentes)} taxa(s) para esta data. Use o recálculo ao invés de criar nova."
                }

            # Calcular base e valor da nova taxa
            base_calculo = self.calcular_base_calculo_taxa(df, data_ref_dt.date())

            if base_calculo <= 0:
                return {
                    "sucesso": False,
                    "mensagem": "Base de cálculo zerada. Não é possível criar taxa de administração."
                }

            percentual = self.obter_percentual_taxa_cliente(cliente)  # or self.percentual_padrao
            valor_taxa = base_calculo * (percentual / 100)

            # Aqui você implementaria a lógica para criar um novo lançamento de taxa
            # (Similar ao que já existe no sistema para criar lançamentos normais)

            return {
                "sucesso": True,
                "mensagem": f"Nova taxa criada: R$ {valor_taxa:,.2f} ({percentual}% de R$ {base_calculo:,.2f})",
                "valor_taxa": valor_taxa,
                "base_calculo": base_calculo,
                "percentual": percentual
            }

        except Exception as e:
            return {"sucesso": False, "mensagem": f"Erro ao criar nova taxa: {str(e)}"}

    def distinguir_cenarios_taxa(self, data_referencia, cliente=None):
        """
        MÉTODO UTILITÁRIO: Distingue entre diferentes cenários de taxa

        Retorna:
        - "sem_taxa": Não há taxa para esta data (primeira finalização)
        - "taxa_existente": Há taxa que pode ser recalculada
        - "taxa_excluida": Há taxa mas está excluída
        - "multiplas_taxas": Há múltiplas taxas (situação complexa)
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual

            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')

            if isinstance(data_referencia, str):
                data_ref_dt = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref_dt = pd.to_datetime(data_referencia)

            df_data = df[df['DATA_REL'].dt.date == data_ref_dt.date()]
            taxas_todas = self.identificar_lancamentos_taxa_admin(df_data)

            if taxas_todas.empty:
                return "sem_taxa", "Nenhuma taxa encontrada para esta data"

            # Adicionar coluna STATUS se não existir
            if 'STATUS' not in taxas_todas.columns:
                taxas_todas['STATUS'] = 'ATIVO'

            taxas_ativas = taxas_todas[taxas_todas['STATUS'] != 'EXCLUIDO']
            taxas_excluidas = taxas_todas[taxas_todas['STATUS'] == 'EXCLUIDO']

            if len(taxas_ativas) > 1:
                return "multiplas_taxas", f"{len(taxas_ativas)} taxas ativas encontradas"
            elif len(taxas_ativas) == 1:
                return "taxa_existente", f"1 taxa ativa encontrada (ID: {taxas_ativas.iloc[0].get('ID_LANCAMENTO', 'N/A')})"
            elif len(taxas_excluidas) > 0:
                return "taxa_excluida", f"{len(taxas_excluidas)} taxa(s) excluída(s) encontrada(s)"
            else:
                return "sem_taxa", "Nenhuma taxa ativa encontrada"

        except Exception as e:
            return "erro", f"Erro ao analisar cenário: {str(e)}"

    def obter_percentual_taxa_cliente(self, cliente):
        """
        VERSÃO CORRIGIDA - Busca percentual de taxa seguindo a mesma lógica do finalizacao_quinzena.py
        """
        try:
            logger.debug(f"DEBUG: Buscando percentual da taxa para cliente {cliente}")

            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)

            if 'Contratos_ADM' not in wb.sheetnames:
                logger.debug("DEBUG: Aba 'Contratos_ADM' não encontrada")
                wb.close()
                return 0

            ws_contratos = wb['Contratos_ADM']
            logger.debug(f"DEBUG: Aba 'Contratos_ADM' carregada")

            # CORREÇÃO 1: Usar a mesma lógica do finalizacao_quinzena.py
            # 1º Passo: Encontrar contratos ativos
            contratos_ativos = set()
            for row in ws_contratos.iter_rows(min_row=3, values_only=True):  # Começar da linha 3
                if row[0] and row[3] == 'ATIVO':  # Coluna A (Nº Contrato) e Coluna D (Status)
                    contratos_ativos.add(str(row[0]))
                    logger.debug(f"DEBUG: Contrato ativo encontrado: {row[0]}")

            logger.debug(f"DEBUG: Contratos ativos: {contratos_ativos}")

            if not contratos_ativos:
                logger.debug("DEBUG: Nenhum contrato ativo encontrado")
                wb.close()
                return 0

            # CORREÇÃO 2: Para cada contrato ativo, buscar administradores com taxa percentual
            taxa_total = 0
            administradores_encontrados = []

            for num_contrato in contratos_ativos:
                logger.debug(f"DEBUG: Verificando administradores do contrato {num_contrato}")

                for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                    # CORREÇÃO 3: Verificar se pertence ao contrato (coluna G) e é do tipo Percentual (coluna J)
                    if (str(row[6]) == num_contrato and      # Coluna G (Nº Contrato)
                        row[9] == 'Percentual'):            # Coluna J (Tipo)

                        # CORREÇÃO 4: Extrair percentual da coluna K
                        percentual_raw = row[10]  # Coluna K (Valor/Percentual)

                        logger.debug(f"DEBUG: Administrador encontrado:")
                        logger.debug(f"  - CNPJ/CPF: {row[7]}")     # Coluna H
                        logger.debug(f"  - Nome: {row[8]}")         # Coluna I
                        logger.debug(f"  - Tipo: {row[9]}")         # Coluna J
                        logger.debug(f"  - Percentual bruto: '{percentual_raw}'")  # Coluna K

                        try:
                            # CORREÇÃO 5: Processar o percentual corretamente
                            if percentual_raw:
                                # Converter para string e limpar
                                percentual_str = str(percentual_raw).strip()

                                # Remover % se existir e converter vírgula para ponto
                                percentual_limpo = percentual_str.replace('%', '').replace(',', '.')

                                percentual_float = float(percentual_limpo)
                                taxa_total += percentual_float

                                administradores_encontrados.append({
                                    'cnpj_cpf': row[7],
                                    'nome': row[8],
                                    'percentual': percentual_float
                                })

                                logger.debug(f"DEBUG: Percentual processado: {percentual_float}%")

                        except (ValueError, TypeError) as e:
                            logger.debug(f"DEBUG: Erro ao processar percentual '{percentual_raw}': {e}")
                            continue

            logger.debug(f"DEBUG: Taxa total encontrada: {taxa_total}%")
            logger.debug(f"DEBUG: Administradores encontrados: {len(administradores_encontrados)}")

            wb.close()
            return taxa_total

        except Exception as e:
            logger.debug(f"DEBUG: Erro ao obter percentual: {str(e)}")
            if 'wb' in locals():
                wb.close()
            return 0

    def verificar_necessidade_recalculo(self, data_referencia, cliente=None):
        """
        VERSÃO CORRIGIDA - Verifica se há necessidade de recálculo usando os métodos unificados
        """
        try:
            if not cliente:
                cliente = self.sistema.cliente_atual

            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"

            if not os.path.exists(arquivo_cliente):
                return False, "Arquivo do cliente não encontrado"

            logger.debug(f"DEBUG: Verificando necessidade de recálculo para {cliente} em {data_referencia}")

            # Ler dados da planilha
            df = pd.read_excel(arquivo_cliente, sheet_name='Dados')
            df = df.fillna("")

            # Converter data
            if isinstance(data_referencia, str):
                data_ref_dt = pd.to_datetime(data_referencia, format='%d/%m/%Y')
            else:
                data_ref_dt = pd.to_datetime(data_referencia)

            # Filtrar para a data específica
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            df_data = df[df['DATA_REL'].dt.date == data_ref_dt.date()].copy()

            # Verificar se há taxas existentes
            taxas_existentes = self.identificar_lancamentos_taxa_admin(df_data)

            if taxas_existentes.empty:
                return False, "Nenhuma taxa encontrada para esta data"

            logger.debug(f"DEBUG: {len(taxas_existentes)} taxa(s) encontrada(s)")

            # CORREÇÃO: Usar o método unificado para calcular base
            # Primeiro tentar com DataFrame (mais rápido)
            base_atual = self.calcular_base_calculo_taxa(data_referencia, df=df)

            logger.debug(f"DEBUG: Base atual calculada: R$ {base_atual:.2f}")

            # Obter percentual da taxa
            percentual = self.obter_percentual_taxa_cliente(cliente)

            if percentual == 0:
                return False, "Percentual de taxa não configurado"

            logger.debug(f"DEBUG: Percentual de taxa: {percentual}%")

            # Calcular valor esperado da taxa
            valor_esperado = base_atual * (percentual / 100)
            logger.debug(f"DEBUG: Valor esperado da taxa: R$ {valor_esperado:.2f}")

            # Somar valor atual das taxas ATIVAS
            valor_atual_taxas = 0
            taxas_ativas = 0

            for _, taxa in taxas_existentes.iterrows():
                status = taxa.get('STATUS', 'ATIVO')
                if status != 'EXCLUIDO':
                    try:
                        valor_taxa = float(str(taxa.get('VALOR', 0)).replace(',', '.'))
                        valor_atual_taxas += valor_taxa
                        taxas_ativas += 1
                        logger.debug(f"DEBUG: Taxa ativa ID {taxa.get('ID_LANCAMENTO', 'N/A')}: R$ {valor_taxa:.2f}")
                    except (ValueError, TypeError):
                        logger.debug(f"DEBUG: Erro ao processar valor da taxa: {taxa.get('VALOR', 'N/A')}")
                        pass

            logger.debug(f"DEBUG: Valor atual total das taxas ativas: R$ {valor_atual_taxas:.2f}")
            logger.debug(f"DEBUG: Taxas ativas encontradas: {taxas_ativas}")

            # Calcular diferença
            diferenca = abs(valor_esperado - valor_atual_taxas)
            tolerancia = 0.01  # R$ 0,01

            logger.debug(f"DEBUG: Diferença: R$ {diferenca:.2f} (tolerância: R$ {tolerancia:.2f})")

            if diferenca > tolerancia:
                mensagem = f"Recálculo necessário - Base: R$ {base_atual:.2f} ({percentual}%) = R$ {valor_esperado:.2f}, Atual: R$ {valor_atual_taxas:.2f}, Diferença: R$ {diferenca:.2f}"
                return True, mensagem

            mensagem = f"Taxas consistentes - Base: R$ {base_atual:.2f} ({percentual}%) = R$ {valor_esperado:.2f}"
            return False, mensagem

        except Exception as e:
            import traceback
            logger.debug(f"DEBUG: Erro na verificação: {traceback.format_exc()}")
            return False, f"Erro na verificação: {str(e)}"
