"""
Processamento de taxas de administração do tipo "Fixo" (valor fixo por
parcela, diferente do tipo "Percentual" tratado em GestorTaxasAdministracao):
gera os lançamentos periódicos na planilha do cliente para administradores
com contrato ativo do tipo fixo.

Extraído de Sistema_Entrada_Dados.py em [DATA_DA_EXTRACAO].
Nenhuma alteração de lógica foi feita nesta extração — apenas mudança
de localização e ajuste de imports.

ATENÇÃO — pendência conhecida (não corrigida nesta extração):
    Em processar_lancamentos_fixos, `row[10].replace(',', '.')` assume que
    o valor da parcela sempre vem como string do Excel. Se a célula for
    numérica (int/float), isso quebra com AttributeError — mesmo padrão de
    bug já visto e corrigido em outras extrações (NaN/float vindo do
    pandas/openpyxl). Aguardando confirmação de uso real antes de corrigir.
"""
import logging

from openpyxl import load_workbook

from src.config.config import PASTA_CLIENTES

# Mesmo logger usado no restante do sistema (Sistema_Entrada_Dados.py).
logger = logging.getLogger("sistema")


class GestaoTaxasFixas:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal

        # Import tardio (dentro do __init__, não no topo do arquivo) para
        # evitar import circular: GestorParcelas ainda está definida dentro
        # de Sistema_Entrada_Dados.py, que é justamente quem importa esta
        # classe. Quando GestorParcelas for extraída para
        # src/parcelamento/gestor_parcelas.py, atualizar esta linha para
        # `from src.parcelamento.gestor_parcelas import GestorParcelas`
        # e mover o import para o topo do arquivo, junto aos demais.
        from src.parcelamento.gestor_parcelas import GestorParcelas
        self.gestor_parcelas = GestorParcelas(self)

    def processar_lancamentos_fixos(self, cliente, data_ref):
        """Processa os lançamentos de taxas fixas para a data de referência"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb['Contratos_ADM']

            lancamentos_gerados = []

            # Buscar contratos ativos com taxa fixa
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verifica se é registro de administrador e tipo fixo
                if (row[6] and  # Tem nº contrato na coluna G
                    row[9] == 'Fixo' and  # É tipo fixo
                    self.contrato_ativo(ws, row[6])):  # Contrato está ativo

                    # Verificar se já tem lançamento para este período
                    if not self.tem_lancamento(ws, row[6], row[7], data_ref):
                        # Preparar dados para o lançamento
                        dados_lancamento = {
                            'data_rel': data_ref,
                            'cnpj_cpf': row[7],  # CNPJ/CPF
                            'nome': row[8],      # Nome/Razão Social
                            'referencia': f'ADM FIXA REF. {data_ref.strftime("%m/%Y")}',
                            'valor': float(row[10].replace(',', '.')),  # Valor/Parcela
                            'dt_vencto': self.calcular_vencimento(data_ref)
                        }

                        # Registrar lançamento no sistema
                        self.sistema.dados_para_incluir.append(dados_lancamento)
                        lancamentos_gerados.append(dados_lancamento)

                        # Registrar na aba de controle
                        self.registrar_lancamento(ws, dados_lancamento)

            wb.save(arquivo_cliente)
            return lancamentos_gerados

        except Exception as e:
            raise Exception(f"Erro ao processar lançamentos fixos: {str(e)}")

    def contrato_ativo(self, ws, num_contrato):
        """Verifica se o contrato está ativo"""
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == num_contrato:  # Coluna A (Nº Contrato)
                return row[3] == 'ATIVO'  # Coluna D (Status)
        return False

    def tem_lancamento(self, ws, num_contrato, cnpj_cpf, data_ref):
        """Verifica se já existe lançamento para o período"""
        data_str = data_ref.strftime("%d/%m/%Y")
        for row in ws.iter_rows(min_row=3, values_only=True):
            if (row[25] and  # Tem referência na coluna PARCELAS
                row[24] == num_contrato and  # Mesmo contrato
                row[26] == cnpj_cpf and  # Mesmo CNPJ/CPF
                row[28] == data_str):  # Mesma data
                return True
        return False

    def calcular_vencimento(self, data_ref):
        """Calcula data de vencimento (dia 5 do mês seguinte)"""
        if data_ref.day == 5:
            vencto = data_ref.replace(day=20)
        else:  # day == 20
            if data_ref.month == 12:
                vencto = data_ref.replace(year=data_ref.year + 1, month=1, day=5)
            else:
                vencto = data_ref.replace(month=data_ref.month + 1, day=5)
        return vencto

    def registrar_lancamento(self, ws, dados):
        """Registra o lançamento na aba de controle"""
        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=26, value=dados['cnpj_cpf'])
        ws.cell(row=proxima_linha, column=27, value=dados['nome'])
        ws.cell(row=proxima_linha, column=28, value=dados['data_rel'])
        ws.cell(row=proxima_linha, column=29, value=dados['valor'])
        ws.cell(row=proxima_linha, column=30, value='LANÇADO')
