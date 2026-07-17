import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from tkcalendar import DateEntry
import pandas as pd
from pathlib import Path
import xlwings as xw
from openpyxl import load_workbook
import openpyxl
import babel

# Adicionar diretório raiz ao path para importar módulos corretamente
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações e utilitários
try:
    from src.config.logger_config import system_logger, log_action
    logger = system_logger.get_logger()
    logger.info("Logger importado com sucesso em gestao_medicoes.py")
except Exception as e:
    print(f"Erro ao importar logger: {str(e)}")
    
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_FORNECEDORES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH
    )
    print("Configurações importadas com sucesso")
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    raise

# Importar o utils.py
from src.config.utils import atualizar_combobox_clientes, cliente_esta_ativo, obter_info_cliente

from src.modules.gerador_contrato import GeradorContrato
from src.config.window_config import configurar_janela
print("window_config importado pelo caminho alternativo")
    
# Importar funções auxiliares
from src.config.utils import (
    formatar_cnpj_cpf,
    buscar_dados_bancarios_fornecedor,
    validar_cnpj,
    validar_cpf,
    formatar_cnpj_cpf,
    validar_data,
    aplicar_formatacao_celula,
    formatar_moeda_br
)

class GestaoMedicoes:
    """Classe principal para gestão de medições"""
    
    _instancia_ativa = None

    @staticmethod
    def formatar_nome_cidade(cidade):
        """
        Formata o nome da cidade como nome próprio.
        Primeira letra maiúscula, resto minúscula, exceto elementos de ligação.
        
        Args:
            cidade (str): Nome da cidade a ser formatado
            
        Returns:
            str: Nome da cidade formatado
        """
        if not cidade:
            return ""
        
        # Elementos de ligação que devem permanecer em minúscula
        elementos_ligacao = ['de', 'da', 'do', 'das', 'dos', 'e']
        
        # Dividir em palavras
        palavras = cidade.lower().split()
        
        # Formatar cada palavra
        palavras_formatadas = []
        for palavra in palavras:
            if palavra in elementos_ligacao:
                palavras_formatadas.append(palavra)
            else:
                palavras_formatadas.append(palavra.capitalize())
        
        return ' '.join(palavras_formatadas)
    
    def __init__(self, parent=None):
        """Inicializa a interface de gestão de medições"""
        
        # ================================================================
        # PROTEÇÃO: Evitar múltiplas instâncias
        # ================================================================
        if GestaoMedicoes._instancia_ativa is not None:
            try:
                if GestaoMedicoes._instancia_ativa.root.winfo_exists():
                    GestaoMedicoes._instancia_ativa.root.lift()
                    GestaoMedicoes._instancia_ativa.root.focus_force()
                    logger.warning("⚠️ Janela de medições já está aberta - trazendo para frente")
                    return
            except:
                GestaoMedicoes._instancia_ativa = None
        
        GestaoMedicoes._instancia_ativa = self
        
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
            
            # ============================================================
            # TRIPLA PROTEÇÃO DE VISIBILIDADE
            # ============================================================
            
            # CAMADA 1: Topmost inicial forte (5 segundos)
            self.root.attributes('-topmost', True)
            self.root.after(5000, self._desativar_topmost)
            
            # CAMADA 2: Lift e foco inicial
            self.root.lift()
            self.root.focus_force()
            
            # CAMADA 3: Binds automáticos para trazer janela quando necessário
            # Quando janela pai ganhar foco, traz filha também
            parent.bind('<FocusIn>', self._trazer_para_frente, add='+')
            # Quando clicar na janela pai, traz filha também
            parent.bind('<Button-1>', self._trazer_para_frente, add='+')
            
            # CAMADA 4: Timer de verificação periódica (backup)
            self._verificar_visibilidade()
            
            # Protocolo de fechamento
            self.root.protocol("WM_DELETE_WINDOW", self.voltar_menu)
            
            logger.info("✅ Janela Gestão de Medições inicializada com tripla proteção de visibilidade")
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            logger.info("✅ Janela Gestão de Medições inicializada como janela principal")
            
        configurar_janela(self.root, "Gestão de Medições")
        
        # Configuração de variáveis
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.contrato_atual = None
        self.fornecedor_atual = None
        self.dados_para_incluir = []
        # Contexto de fornecedor (selecionado na aba Fornecedor)
        self.cnpj_fornecedor_selecionado = None
        self.nome_fornecedor_selecionado = None
        
        # Configurar interface
        self.setup_gui()

        self._gerador_contrato = None  # instanciado sob demanda
        self.servicos_selecionados = []
    
    @property
    def gerador_contrato(self):
        if self._gerador_contrato is None:
            self._gerador_contrato = GeradorContrato()
        return self._gerador_contrato

    def _desativar_topmost(self):
        """
        Desativa topmost após 5 segundos.
        Permite alternância normal de janelas depois.
        """
        try:
            if self.root and self.root.winfo_exists():
                self.root.attributes('-topmost', False)
                logger.debug("✓ Topmost desativado após 5s - janela pode ser alternada normalmente")
        except Exception as e:
            logger.debug(f"⚠️ Erro ao desativar topmost: {e}")
    
    def _trazer_para_frente(self, event=None):
        """
        Traz janela de medições para frente automaticamente.
        Chamado por binds quando janela pai recebe foco ou clique.
        
        SOLUÇÃO PRINCIPAL: Isto garante que a janela nunca fica oculta!
        """
        try:
            if self.root and self.root.winfo_exists():
                # Verificar se janela não está minimizada
                if self.root.state() != 'withdrawn':
                    self.root.lift()
                    logger.debug("✓ Janela trazida para frente via bind")
        except Exception as e:
            logger.debug(f"⚠️ Erro ao trazer para frente: {e}")
    
    def _verificar_visibilidade(self):
        """
        Timer de backup: verifica periodicamente se janela está visível.
        Se janela pai está ativa mas filha está oculta, traz para frente.
        
        Reagendado a cada 2 segundos enquanto janela existir.
        """
        try:
            if self.root and self.root.winfo_exists():
                # Se janela está mapeada (visível no sistema)
                if self.root.winfo_ismapped():
                    # Verificar se janela pai existe e está ativa
                    if self.menu_principal and self.menu_principal.winfo_exists():
                        try:
                            # Tentar pegar o foco atual
                            foco_atual = self.root.focus_get()
                            # Se pai tem foco mas filha não, trazer filha
                            if foco_atual and foco_atual != self.root:
                                parent_window = foco_atual.winfo_toplevel()
                                if parent_window == self.menu_principal:
                                    self.root.lift()
                                    logger.debug("✓ Janela trazida para frente via timer")
                        except:
                            pass
                
                # Reagendar próxima verificação
                self.root.after(2000, self._verificar_visibilidade)
        except:
            # Janela foi destruída, parar timer
            pass


    def converter_valor_brasileiro_para_float(self, valor_str):
        """
        Converte string em formato brasileiro para float.
        
        IMPORTANTE: Esta função é um MÉTODO da classe, então precisa de 'self' e 'valor_str'
        
        Exemplos de entrada aceitos:
        - "R$ 15.293,58" → 15293.58
        - "15.293,58" → 15293.58
        - "R$ 5.017,98" → 5017.98
        - "5017,98" → 5017.98
        - "15293.58" → 15293.58 (já em formato numérico)
        - 15293.58 → 15293.58 (já é número)
        
        Args:
            valor_str: String ou número em formato brasileiro
            
        Returns:
            float: Valor numérico convertido
            
        Raises:
            ValueError: Se não conseguir converter
        """
        # Se já for número, retorna direto
        if isinstance(valor_str, (int, float)):
            return float(valor_str)
        
        # Se for None ou vazio, retorna 0
        if not valor_str or valor_str == '':
            return 0.0
        
        # Converter para string e limpar
        valor_limpo = str(valor_str).strip()
        
        # Remover símbolo de moeda e espaços
        valor_limpo = valor_limpo.replace('R$', '').replace('$', '').strip()
        
        # DETECTAR FORMATO:
        # Formato brasileiro: ponto para milhar, vírgula para decimal
        # Formato americano: vírgula para milhar, ponto para decimal
        
        tem_virgula = ',' in valor_limpo
        tem_ponto = '.' in valor_limpo
        
        if tem_virgula and tem_ponto:
            # Tem ambos - identificar qual é decimal
            pos_virgula = valor_limpo.rfind(',')
            pos_ponto = valor_limpo.rfind('.')
            
            if pos_virgula > pos_ponto:
                # Formato brasileiro: vírgula vem depois (é o decimal)
                # "15.293,58" → remover ponto, trocar vírgula por ponto
                valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
            else:
                # Formato americano: ponto vem depois (é o decimal)
                # "15,293.58" → remover vírgula
                valor_limpo = valor_limpo.replace(',', '')
        
        elif tem_virgula and not tem_ponto:
            # Só vírgula - formato brasileiro
            # "5017,98" → trocar vírgula por ponto
            valor_limpo = valor_limpo.replace(',', '.')
        
        elif tem_ponto and not tem_virgula:
            # Só ponto - pode ser milhar OU decimal
            # Analisar contexto
            partes = valor_limpo.split('.')
            
            # Se tem apenas um ponto e:
            # - 2 casas depois E
            # - Parte antes tem <= 3 dígitos
            # Então provavelmente é decimal
            if len(partes) == 2 and len(partes[-1]) == 2 and len(partes[0]) <= 3:
                # "15.29" → é decimal, mantém
                pass
            else:
                # "15.293" ou "1.500.293" → é milhar, remove
                valor_limpo = valor_limpo.replace('.', '')
        
        # Tentar converter
        try:
            return float(valor_limpo)
        except ValueError as e:
            raise ValueError(f"Não foi possível converter '{valor_str}' para número: {e}")
            
    def setup_gui(self):
        """Configuração da interface gráfica"""
        # Notebook (abas)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar abas — ordem define os índices
        self.aba_selecao = ttk.Frame(self.notebook)            # 0
        self.aba_fornecedor = ttk.Frame(self.notebook)         # 1
        self.aba_contratos = ttk.Frame(self.notebook)          # 2
        self.aba_medicoes = ttk.Frame(self.notebook)           # 3
        self.aba_relatorios = ttk.Frame(self.notebook)         # 4 — NOVA
        self.aba_contratos_emissao = ttk.Frame(self.notebook)  # 5
        
        self.notebook.add(self.aba_selecao, text='Seleção de Cliente')
        self.notebook.add(self.aba_fornecedor, text='Fornecedor')
        self.notebook.add(self.aba_contratos, text='Contratos')
        self.notebook.add(self.aba_medicoes, text='Medições')
        self.notebook.add(self.aba_relatorios, text='Relatórios')
        self.notebook.add(self.aba_contratos_emissao, text='Contrato')

        # Configurar cada aba
        self.setup_aba_selecao()
        self.setup_aba_fornecedor()
        self.setup_aba_contratos()
        self.setup_aba_medicoes()
        self.setup_aba_relatorios()
        self.setup_aba_emissao_contrato()
        self.notebook.bind('<<NotebookTabChanged>>', self._ao_trocar_aba)

    def _ao_trocar_aba(self, event=None):
        """Carrega contratos automaticamente ao entrar na aba Contratos."""
        aba_atual = self.notebook.select()
        if aba_atual == str(self.aba_contratos):
            if self.cliente_atual and self.arquivo_cliente:
                self.carregar_contratos()

    def setup_aba_selecao(self):
        """Configura a aba de seleção de cliente"""
        frame_principal = ttk.Frame(self.aba_selecao)
        frame_principal.pack(expand=True, fill='both', padx=10, pady=5)

        # --- SELEÇÃO DO CLIENTE ---
        frame_selecao = ttk.LabelFrame(frame_principal, text="Seleção do Cliente")
        frame_selecao.pack(fill='x', pady=10)

        frame_cliente = ttk.Frame(frame_selecao)
        frame_cliente.pack(fill='x', padx=10, pady=10)

        ttk.Label(frame_cliente, text="Selecione o Cliente:", font=('Arial', 11)).pack(side='left', pady=5)
        self.cliente_combobox = ttk.Combobox(frame_cliente, width=60, font=('Arial', 11))
        self.cliente_combobox.pack(side='left', padx=5, fill='x', expand=True)

        # --- DADOS DO CLIENTE (preenchidos ao selecionar) ---
        frame_dados_cliente = ttk.LabelFrame(frame_principal, text="Dados do Cliente")
        frame_dados_cliente.pack(fill='x', pady=5, padx=5)

        dados_grid = ttk.Frame(frame_dados_cliente)
        dados_grid.pack(fill='x', padx=10, pady=8)

        ttk.Label(dados_grid, text="CNO:").grid(row=0, column=0, sticky='w', padx=5, pady=3)
        self.ent_sel_cno = ttk.Entry(dados_grid, state='readonly', width=22)
        self.ent_sel_cno.grid(row=0, column=1, sticky='w', padx=5, pady=3)

        ttk.Label(dados_grid, text="CPF:").grid(row=0, column=2, sticky='w', padx=5, pady=3)
        self.ent_sel_cpf = ttk.Entry(dados_grid, state='readonly', width=22)
        self.ent_sel_cpf.grid(row=0, column=3, sticky='w', padx=5, pady=3)

        ttk.Label(dados_grid, text="Estado Civil:").grid(row=1, column=0, sticky='w', padx=5, pady=3)
        self.ent_sel_estado_civil = ttk.Entry(dados_grid, state='readonly', width=22)
        self.ent_sel_estado_civil.grid(row=1, column=1, sticky='w', padx=5, pady=3)

        ttk.Label(dados_grid, text="Cidade:").grid(row=1, column=2, sticky='w', padx=5, pady=3)
        self.ent_sel_cidade = ttk.Entry(dados_grid, state='readonly', width=22)
        self.ent_sel_cidade.grid(row=1, column=3, sticky='w', padx=5, pady=3)

        ttk.Label(dados_grid, text="Endereço:").grid(row=2, column=0, sticky='w', padx=5, pady=3)
        self.ent_sel_endereco = ttk.Entry(dados_grid, state='readonly', width=70)
        self.ent_sel_endereco.grid(row=2, column=1, columnspan=3, sticky='ew', padx=5, pady=3)

        dados_grid.columnconfigure(1, weight=1)
        dados_grid.columnconfigure(3, weight=1)

        # --- BOTÕES ---
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 12, 'bold'), padding=(15, 10))

        frame_gerenciar = ttk.Frame(frame_principal)
        frame_gerenciar.pack(pady=15)

        ttk.Button(frame_gerenciar,
                text="Continuar →",
                command=self.continuar_para_fornecedor,
                style='Big.TButton').pack(side='right', padx=10)

        frame_botoes_selecao = ttk.Frame(frame_principal)
        frame_botoes_selecao.pack(fill='x', side='bottom', pady=10)

        ttk.Button(frame_botoes_selecao,
                text="Voltar ao Menu",
                command=self.voltar_menu,
                style='Big.TButton').pack(side='left', padx=10)

        self.atualizar_lista_clientes()
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)

        # Pré-selecionar o último cliente usado em outro módulo (ex.: Entrada de Dados)
        try:
            try:
                from src.estado_compartilhado import estado_sessao
            except ImportError:
                from estado_compartilhado import estado_sessao

            valores_disponiveis = self.cliente_combobox['values']
            if estado_sessao.ultimo_cliente and estado_sessao.ultimo_cliente in valores_disponiveis:
                self.cliente_combobox.set(estado_sessao.ultimo_cliente)
                self.selecionar_cliente()
        except Exception as e:
            logger.warning(f"Não foi possível pré-selecionar último cliente: {e}")
        
    def setup_aba_fornecedor(self):
        """Configura a aba de seleção de fornecedor/empreiteiro"""
        frame_principal = ttk.Frame(self.aba_fornecedor)
        frame_principal.pack(expand=True, fill='both', padx=10, pady=5)

        # Label do cliente em contexto
        self.lbl_cliente_fornecedor = ttk.Label(
            frame_principal,
            text="Cliente: Nenhum selecionado",
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_fornecedor.pack(anchor='w', padx=5, pady=5)

        # --- BUSCA DE FORNECEDOR ---
        frame_busca = ttk.LabelFrame(frame_principal, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', pady=5)

        busca_inner = ttk.Frame(frame_busca)
        busca_inner.pack(fill='x', padx=10, pady=8)

        ttk.Label(busca_inner, text="Nome / CNPJ:").pack(side='left', padx=5)
        self.ent_busca_forn_aba = ttk.Entry(busca_inner, width=40)
        self.ent_busca_forn_aba.pack(side='left', padx=5, fill='x', expand=True)
        self.ent_busca_forn_aba.bind('<KeyRelease>', self.buscar_fornecedor_aba)

        ttk.Button(busca_inner, text="🔍 Buscar",
                command=self.buscar_fornecedor_aba).pack(side='left', padx=2)
        ttk.Button(busca_inner, text="↻ Todos",
                command=self.listar_todos_fornecedores_aba).pack(side='left', padx=2)
        ttk.Button(busca_inner, text="➕ Novo Fornecedor",
                command=self.novo_fornecedor).pack(side='left', padx=2)
        ttk.Button(busca_inner, text="✏️ Editar Fornecedor",
                command=self.editar_fornecedor_aba).pack(side='left', padx=2)

        # Listbox de resultados
        listbox_frame = ttk.Frame(frame_busca)
        listbox_frame.pack(fill='x', padx=10, pady=5)

        self.lst_fornecedor_aba = tk.Listbox(listbox_frame, height=5, width=80,
                                             selectmode='single', exportselection=False)
        self.lst_fornecedor_aba.pack(side='left', fill='both', expand=True)

        scrollbar_forn = ttk.Scrollbar(listbox_frame, orient='vertical',
                                       command=self.lst_fornecedor_aba.yview)
        scrollbar_forn.pack(side='right', fill='y')
        self.lst_fornecedor_aba.config(yscrollcommand=scrollbar_forn.set)
        self.lst_fornecedor_aba.bind('<<ListboxSelect>>', self.selecionar_fornecedor_aba)

        # --- DADOS DO FORNECEDOR (preenchidos ao selecionar) ---
        frame_dados_forn = ttk.LabelFrame(frame_principal, text="Dados do Fornecedor")
        frame_dados_forn.pack(fill='x', pady=5, padx=5)

        dados_grid = ttk.Frame(frame_dados_forn)
        dados_grid.pack(fill='x', padx=10, pady=8)

        ttk.Label(dados_grid, text="CNPJ/CPF:").grid(row=0, column=0, sticky='w', padx=5, pady=3)
        self.ent_forn_cnpj = ttk.Entry(dados_grid, state='readonly', width=25)
        self.ent_forn_cnpj.grid(row=0, column=1, sticky='w', padx=5, pady=3)

        ttk.Label(dados_grid, text="Endereço:").grid(row=1, column=0, sticky='w', padx=5, pady=3)
        self.ent_forn_endereco = ttk.Entry(dados_grid, state='readonly', width=70)
        self.ent_forn_endereco.grid(row=1, column=1, columnspan=3, sticky='ew', padx=5, pady=3)

        ttk.Label(dados_grid, text="Dados Bancários:").grid(row=2, column=0, sticky='nw', padx=5, pady=3)
        self.txt_forn_dados_bancarios = tk.Text(dados_grid, height=2, width=70,
                                                state='disabled', wrap='word')
        self.txt_forn_dados_bancarios.grid(row=2, column=1, columnspan=3, sticky='ew', padx=5, pady=3)

        dados_grid.columnconfigure(1, weight=1)

        # --- BOTÕES ---
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=15)

        ttk.Button(frame_botoes,
                text="Continuar →",
                command=self.continuar_para_contratos_filtrado,
                style='Big.TButton').pack(side='right', padx=10)

        ttk.Button(frame_botoes,
                text="← Voltar",
                command=lambda: self.notebook.select(self.aba_selecao),
                style='Big.TButton').pack(side='left', padx=10)

    def setup_aba_contratos(self):
        """Configura a aba de contratos"""
        # Frame principal
        frame_principal = ttk.Frame(self.aba_contratos)
        frame_principal.pack(expand=True, fill='both', padx=10, pady=5)
        
        # Label para cliente atual
        self.lbl_cliente_contratos = ttk.Label(
            frame_principal, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_contratos.pack(anchor='w', padx=5, pady=5)

        # --- BARRA DE FILTROS ---
        frame_filtros = ttk.Frame(frame_principal)
        frame_filtros.pack(fill='x', padx=5, pady=(0, 4))

        ttk.Label(frame_filtros, text="Status:").pack(side='left', padx=(0, 4))
        self.var_filtro_status = tk.StringVar(value="Ativo")
        for texto, valor in [("Todos", "Todos"), ("Ativo", "Ativo"), ("Concluído", "Concluído")]:
            ttk.Radiobutton(
                frame_filtros, text=texto,
                variable=self.var_filtro_status, value=valor,
                command=self.carregar_contratos
            ).pack(side='left', padx=2)

        ttk.Label(frame_filtros, text="  Buscar:").pack(side='left', padx=(10, 4))
        self.var_busca_contratos = tk.StringVar()
        self.var_busca_contratos.trace_add('write', lambda *_: self.carregar_contratos())
        ttk.Entry(frame_filtros, textvariable=self.var_busca_contratos, width=22).pack(side='left')

        self.lbl_cont_contratos = ttk.Label(frame_filtros, text="", foreground='#555')
        self.lbl_cont_contratos.pack(side='right', padx=10)

        # Segunda linha: indicador do filtro de fornecedor ativo
        frame_forn_filtro = ttk.Frame(frame_principal)
        frame_forn_filtro.pack(fill='x', padx=5, pady=(0, 2))
        self.lbl_filtro_forn = ttk.Label(
            frame_forn_filtro, text="", foreground='#0056b3', font=('Arial', 9)
        )
        self.lbl_filtro_forn.pack(side='left')
        self.btn_limpar_filtro_forn = ttk.Button(
            frame_forn_filtro, text="✕ Todos os fornecedores",
            command=self._limpar_filtro_fornecedor, width=22
        )
        # O botão só aparece quando há filtro ativo (controlado por _atualizar_lbl_filtro_fornecedor)

        # --- LISTA DE CONTRATOS ---
        frame_contratos = ttk.LabelFrame(frame_principal, text="Contratos Registrados")
        frame_contratos.pack(fill='both', expand=True, pady=5)

        colunas = ('ID', 'Nº/Emp', 'Fornecedor', 'Descrição', 'Data Início', 'Data Final', 'Valor Global', 'Valor Pago', 'Saldo', 'Pend.')
        self.tree_contratos = ttk.Treeview(frame_contratos, columns=colunas, show='headings', height=10)
        self._sort_contratos = {'col': None, 'rev': False}

        # Cabeçalhos clicáveis para ordenação
        self.tree_contratos.heading('ID', text='ID')
        for col in ('Nº/Emp', 'Fornecedor', 'Descrição', 'Data Início', 'Data Final', 'Valor Global', 'Valor Pago', 'Saldo', 'Pend.'):
            self.tree_contratos.heading(
                col, text=col,
                command=lambda c=col: self._ordenar_contratos(c)
            )

        # Ajustar larguras das colunas (ID oculto — largura 0)
        self.tree_contratos.column('ID', width=0, minwidth=0, stretch=False)
        self.tree_contratos.column('Nº/Emp', width=55, anchor='center')
        self.tree_contratos.column('Fornecedor', width=150)
        self.tree_contratos.column('Descrição', width=200)
        self.tree_contratos.column('Data Início', width=100, anchor='center')
        self.tree_contratos.column('Data Final', width=100, anchor='center')
        self.tree_contratos.column('Valor Global', width=110, anchor='e')
        self.tree_contratos.column('Valor Pago', width=110, anchor='e')
        self.tree_contratos.column('Saldo', width=110, anchor='e')
        self.tree_contratos.column('Pend.', width=45, anchor='center')
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_contratos, orient='vertical', command=self.tree_contratos.yview)
        scrollx = ttk.Scrollbar(frame_contratos, orient='horizontal', command=self.tree_contratos.xview)
        self.tree_contratos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionamento
        self.tree_contratos.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Frame para botões
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Novo Contrato",
                  command=self.abrir_novo_contrato).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Reemitir Contrato",
                  command=self.reemitir_contrato).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Editar Contrato",
                  command=self.editar_contrato).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Excluir Contrato",
                  command=self.excluir_contrato).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Selecionar",
                  command=self.selecionar_contrato).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Voltar",
                  command=lambda: self.notebook.select(self.aba_fornecedor)).pack(side='right', padx=5)

        # Binding para duplo clique
        self.tree_contratos.bind('<Double-1>', lambda e: self.selecionar_contrato())

    def setup_aba_medicoes(self):
        """Configura a aba de medições"""
        # Frame principal
        frame_principal = ttk.Frame(self.aba_medicoes)
        frame_principal.pack(expand=True, fill='both', padx=10, pady=5)
        
        # Labels para cliente e contrato atuais
        frame_info = ttk.Frame(frame_principal)
        frame_info.pack(fill='x', pady=5)
        
        self.lbl_cliente_medicoes = ttk.Label(
            frame_info, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_medicoes.pack(side='left', padx=10)
        
        self.lbl_contrato_medicoes = ttk.Label(
            frame_info, 
            text="Contrato: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_contrato_medicoes.pack(side='left', padx=10)
        
        # Frame para lista de medições
        frame_medicoes = ttk.LabelFrame(frame_principal, text="Medições Registradas")
        frame_medicoes.pack(fill='both', expand=True, pady=5)
        
        # Treeview para medições
        colunas = ('ID', 'Data Medição', 'Data Pagamento', 'Referência', 'Valor', 'Status')
        self.tree_medicoes = ttk.Treeview(frame_medicoes, columns=colunas, show='headings', height=10)
        
        # Configurar colunas
        self.tree_medicoes.heading('ID', text='ID')
        self.tree_medicoes.heading('Data Medição', text='Data Medição')
        self.tree_medicoes.heading('Data Pagamento', text='Data Pagamento')
        self.tree_medicoes.heading('Referência', text='Referência')
        self.tree_medicoes.heading('Valor', text='Valor')
        self.tree_medicoes.heading('Status', text='Status')
        
        # Ajustar larguras das colunas
        self.tree_medicoes.column('ID', width=50, anchor='center')
        self.tree_medicoes.column('Data Medição', width=100, anchor='center')
        self.tree_medicoes.column('Data Pagamento', width=100, anchor='center')
        self.tree_medicoes.column('Referência', width=300)
        self.tree_medicoes.column('Valor', width=100, anchor='e')
        self.tree_medicoes.column('Status', width=100, anchor='center')
        
        # Scrollbars
        scrolly = ttk.Scrollbar(frame_medicoes, orient='vertical', command=self.tree_medicoes.yview)
        scrollx = ttk.Scrollbar(frame_medicoes, orient='horizontal', command=self.tree_medicoes.xview)
        self.tree_medicoes.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionamento
        self.tree_medicoes.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')
        
        # Frame para botões
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=10)

        ttk.Button(frame_botoes, text="Nova Medição", 
                command=self.nova_medicao).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Editar Medição", 
                command=self.editar_medicao).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Excluir Medição", 
           command=self.excluir_medicao).pack(side='left')
        ttk.Button(frame_botoes, text="Lançar no Cliente", 
                command=self.lancar_medicao).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Vincular a Lançamento", 
                command=self.vincular_medicao, 
                style='Accent.TButton').pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Voltar", 
                command=lambda: self.notebook.select(self.aba_contratos)).pack(side='right', padx=5)
        
        # Botão para voltar ao menu principal
        ttk.Button(frame_principal, text="Voltar ao Menu Principal", 
                 command=self.voltar_menu).pack(side='bottom', pady=10)

    def setup_aba_relatorios(self):
        """Configura a aba de relatórios e importação (escopo: cliente atual)"""
        frame_principal = ttk.Frame(self.aba_relatorios)
        frame_principal.pack(expand=True, fill='both', padx=15, pady=10)

        # Label de contexto
        self.lbl_cliente_relatorios = ttk.Label(
            frame_principal,
            text="Cliente: Nenhum selecionado",
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_relatorios.pack(anchor='w', padx=5, pady=(0, 10))

        # --- SEÇÃO 1: RELATÓRIO DE CONTRATOS (Excel) ---
        frame_excel = ttk.LabelFrame(frame_principal, text="Relatório de Contratos e Medições", padding=12)
        frame_excel.pack(fill='x', pady=8)

        ttk.Label(frame_excel,
            text="Exporta todos os contratos e medições do cliente selecionado para uma janela interativa,\n"
                 "com filtros por status e data. Permite exportar para Excel.",
            foreground='#444444', font=('Arial', 9)).pack(anchor='w', pady=(0, 8))

        ttk.Button(frame_excel,
            text="📊 Abrir Relatório de Contratos",
            command=self.abrir_relatorio_contratos,
            style='Big.TButton').pack(anchor='w')

        # --- SEÇÃO 2: IMPORTAR MEDIÇÕES ---
        frame_import = ttk.LabelFrame(frame_principal, text="Importar Medições", padding=12)
        frame_import.pack(fill='x', pady=8)

        ttk.Label(frame_import,
            text="Importa medições a partir de um arquivo .xlsx gerado pelo relatório de contratos.\n"
                 "As medições são associadas automaticamente aos contratos existentes do cliente.",
            foreground='#444444', font=('Arial', 9)).pack(anchor='w', pady=(0, 8))

        ttk.Button(frame_import,
            text="📥 Importar Medições (.xlsx)",
            command=self.importar_medicoes_relatorio,
            style='Big.TButton').pack(anchor='w')

        # --- SEÇÃO 3: RELATÓRIO QUINZENAL PDF ---
        frame_pdf = ttk.LabelFrame(frame_principal, text="Relatório Quinzenal PDF", padding=12)
        frame_pdf.pack(fill='x', pady=8)

        ttk.Label(frame_pdf,
            text="Gera o relatório quinzenal de medições em PDF para o cliente selecionado.\n"
                 "Se um fornecedor estiver selecionado, o relatório será filtrado por ele.",
            foreground='#444444', font=('Arial', 9)).pack(anchor='w', pady=(0, 8))

        frame_data_pdf = ttk.Frame(frame_pdf)
        frame_data_pdf.pack(anchor='w', pady=(0, 8))

        ttk.Label(frame_data_pdf, text="Data de referência:").pack(side='left', padx=(0, 8))
        self.ent_data_relatorio_pdf = DateEntry(frame_data_pdf, width=14,
                                                date_pattern='dd/mm/yyyy', locale='pt_BR')
        # Sugere a próxima quinzena (dia 5 ou 20), mesma regra usada na
        # Interface de Relatório (src.config_relatorio_quinzenal).
        try:
            from src.config_relatorio_quinzenal import calcular_data_rel_automatica
            data_sugerida = calcular_data_rel_automatica()
        except Exception:
            data_sugerida = datetime.now()
        self.ent_data_relatorio_pdf.set_date(data_sugerida)
        self.ent_data_relatorio_pdf.pack(side='left')

        # Indicador do filtro de fornecedor ativo
        self.lbl_filtro_pdf = ttk.Label(frame_data_pdf,
            text="", foreground='#0056b3', font=('Arial', 9, 'italic'))
        self.lbl_filtro_pdf.pack(side='left', padx=12)

        ttk.Button(frame_pdf,
            text="📄 Gerar PDF Quinzenal",
            command=self.gerar_relatorio_pdf_quinzenal,
            style='Big.TButton').pack(anchor='w')

        # --- BOTÃO VOLTAR ---
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', side='bottom', pady=10)

        ttk.Button(frame_botoes, text="Voltar ao Menu Principal",
                command=self.voltar_menu).pack(side='left', padx=5)

    # --- Métodos da aba Relatórios ---

    def abrir_relatorio_contratos(self):
        """Abre a janela de relatório de contratos (RelatorioContratos)"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
            return
        try:
            from src.relatorio_contratos_medicoes import RelatorioContratos
            janela = RelatorioContratos(parent=self.root)
            # Pré-selecionar o cliente e disparar selecionar_cliente para
            # que cliente_atual e arquivo_cliente sejam definidos na janela
            if hasattr(janela, 'cliente_combobox'):
                janela.cliente_combobox.set(self.cliente_atual)
            if hasattr(janela, 'selecionar_cliente'):
                janela.selecionar_cliente()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir relatório de contratos:\n{str(e)}", parent=self.root)

    def importar_medicoes_relatorio(self):
        """Chama o ImportadorMedicoes com o contexto do cliente atual"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
            return
        try:
            from src.importador_medicoes import ImportadorMedicoes
            importador = ImportadorMedicoes(sistema_principal=self)
            importador.importar_medicoes()
            ids_alterados = getattr(importador, 'ids_contratos_alterados', [])
            if ids_alterados:
                # Garantir que todos os contratos alterados fiquem visíveis,
                # incluindo os que ficaram CONCLUÍDO após a importação.
                filtro_status = getattr(self, 'var_filtro_status', None)
                if filtro_status:
                    filtro_status.set('Todos')
            self.carregar_contratos()
            if ids_alterados:
                self._destacar_contratos_alterados(ids_alterados)
                self.notebook.select(self.aba_contratos)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao iniciar importador:\n{str(e)}", parent=self.root)

    def _ordenar_contratos(self, coluna):
        """Ordena o Treeview de contratos pela coluna clicada (toggle asc/desc)."""
        estado = self._sort_contratos
        reverso = (estado['col'] == coluna) and not estado['rev']
        estado['col'] = coluna
        estado['rev'] = reverso

        # Coletar itens com o valor da coluna para ordenação
        idx_col = ('ID', 'Nº/Emp', 'Fornecedor', 'Descrição', 'Data Início', 'Data Final',
                   'Valor Global', 'Valor Pago', 'Saldo', 'Pend.').index(coluna)

        def _chave(item_id):
            val = self.tree_contratos.item(item_id)['values'][idx_col]
            val_str = str(val or '')
            # Tentar converter valor monetário para número
            if val_str.startswith('R$'):
                try:
                    return float(val_str.replace('R$', '').replace('.', '').replace(',', '.').strip())
                except ValueError:
                    pass
            # Tentar converter datas (dd/mm/aaaa → aaaa-mm-dd para ordenação correta)
            if len(val_str) == 10 and val_str[2] == '/' and val_str[5] == '/':
                try:
                    from datetime import datetime as _dt
                    return _dt.strptime(val_str, '%d/%m/%Y')
                except ValueError:
                    pass
            # Tentar número inteiro
            try:
                return int(val_str)
            except ValueError:
                pass
            return val_str.lower()

        itens = sorted(self.tree_contratos.get_children(''), key=_chave, reverse=reverso)
        for pos, item_id in enumerate(itens):
            self.tree_contratos.move(item_id, '', pos)

        # Atualizar texto dos cabeçalhos com indicador de direção
        for col in ('Nº/Emp', 'Fornecedor', 'Descrição', 'Data Início', 'Data Final',
                    'Valor Global', 'Valor Pago', 'Saldo', 'Pend.'):
            base = col.rstrip(' ▲▼')
            if col == coluna:
                self.tree_contratos.heading(col, text=f"{base} {'▼' if reverso else '▲'}")
            else:
                self.tree_contratos.heading(col, text=base)

    def _destacar_contratos_alterados(self, ids_alterados):
        """Marca com fundo amarelo os contratos cujas medições foram importadas."""
        ids_set = {str(i) for i in ids_alterados}
        # foreground padrão para garantir legibilidade sobre fundo amarelo
        self.tree_contratos.tag_configure('alterado', background='#FFF176', foreground='#000000')
        for item in self.tree_contratos.get_children():
            valores = self.tree_contratos.item(item)['values']
            if valores and str(valores[0]) in ids_set:
                # Substituir tags completamente para que 'alterado' tenha prioridade
                self.tree_contratos.item(item, tags=('alterado',))

    def _atualizar_lbl_filtro_fornecedor(self):
        """Atualiza o indicador visual do filtro de fornecedor na aba Contratos."""
        if not hasattr(self, 'lbl_filtro_forn'):
            return
        nome = getattr(self, 'nome_fornecedor_selecionado', None)
        cnpj = getattr(self, 'cnpj_fornecedor_selecionado', None)
        if cnpj:
            texto = f"Filtro: {nome or cnpj}"
            self.lbl_filtro_forn.config(text=texto)
            self.btn_limpar_filtro_forn.pack(side='left', padx=(6, 0))
        else:
            self.lbl_filtro_forn.config(text="")
            self.btn_limpar_filtro_forn.pack_forget()

    def _limpar_filtro_fornecedor(self):
        """Remove o filtro de fornecedor e recarrega a lista de contratos."""
        self.cnpj_fornecedor_selecionado = None
        self.nome_fornecedor_selecionado = None
        self._atualizar_lbl_filtro_fornecedor()
        self.carregar_contratos()

    def gerar_relatorio_pdf_quinzenal(self):
        """Gera o relatório quinzenal PDF usando o cliente e data selecionados"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
            return
        if not self.arquivo_cliente:
            messagebox.showwarning("Aviso", "Arquivo do cliente não encontrado!", parent=self.root)
            return
        try:
            from src.gerar_relatorio_quinzenal_pdf import RelatorioQuinzenalPDF

            data_ref = self.ent_data_relatorio_pdf.get_date()
            # Converter para datetime se necessário
            if not isinstance(data_ref, datetime):
                data_ref = datetime(data_ref.year, data_ref.month, data_ref.day)

            gerador = RelatorioQuinzenalPDF(
                arquivo_cliente=str(self.arquivo_cliente),
                arquivo_clientes=str(ARQUIVO_CLIENTES)
            )
            arquivo_pdf = gerador.gerar_pdf(data_referencia=data_ref)

            if arquivo_pdf:
                messagebox.showinfo(
                    "Sucesso",
                    f"Relatório gerado com sucesso:\n{arquivo_pdf}",
                    parent=self.root
                )
                # Abrir o PDF automaticamente
                try:
                    import subprocess, platform
                    if platform.system() == 'Windows':
                        os.startfile(str(arquivo_pdf))
                    elif platform.system() == 'Darwin':
                        subprocess.run(['open', str(arquivo_pdf)])
                    else:
                        subprocess.run(['xdg-open', str(arquivo_pdf)])
                except Exception:
                    pass  # Abertura automática é opcional
            else:
                messagebox.showwarning("Aviso", "O relatório não foi gerado. Não existem medições no período.", parent=self.root)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar PDF:\n{str(e)}", parent=self.root)

    def setup_aba_emissao_contrato(self):
        """Configura a aba de contrato (novo e reemissão)"""

        # Frame principal com scroll
        main_frame = ttk.Frame(self.aba_contratos_emissao)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # --- CONTEXTO: Cliente e Fornecedor (somente leitura, herdados do contexto) ---
        frame_contexto = ttk.LabelFrame(scrollable_frame, text="Contexto", padding=8)
        frame_contexto.pack(fill='x', pady=5)

        self.lbl_contrato_cliente = ttk.Label(frame_contexto,
            text="Cliente: Nenhum selecionado",
            font=('Arial', 10, 'bold'), foreground='#0056b3')
        self.lbl_contrato_cliente.pack(anchor='w')

        self.lbl_contrato_fornecedor = ttk.Label(frame_contexto,
            text="Fornecedor: Nenhum selecionado",
            font=('Arial', 10, 'bold'), foreground='#0056b3')
        self.lbl_contrato_fornecedor.pack(anchor='w')

        # --- DADOS DO CONTRATO ---
        frame_dados_contrato = ttk.LabelFrame(scrollable_frame, text="Dados do Contrato", padding=10)
        frame_dados_contrato.pack(fill='x', pady=5)

        dados_contrato_grid = ttk.Frame(frame_dados_contrato)
        dados_contrato_grid.pack(fill='x', pady=5)

        row = 0
        ttk.Label(dados_contrato_grid, text="Data do Contrato:").grid(row=row, column=0, sticky='w', padx=5, pady=2)
        self.ent_data_contrato = DateEntry(dados_contrato_grid, width=15,
                                date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.ent_data_contrato.grid(row=row, column=1, sticky='w', padx=5, pady=2)
        self.ent_data_contrato.set_date(datetime.now())
        ttk.Label(dados_contrato_grid, text="⚠️ Verifique a data ao reemitir",
                foreground='#b8860b', font=('Arial', 8, 'italic')).grid(
                row=row, column=2, columnspan=2, sticky='w', padx=5, pady=2)

        row += 1
        ttk.Label(dados_contrato_grid, text="Data Início:").grid(row=row, column=0, sticky='w', padx=5, pady=2)
        self.ent_data_inicio = DateEntry(dados_contrato_grid, width=15,
                                date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.ent_data_inicio.grid(row=row, column=1, sticky='w', padx=5, pady=2)
        self.ent_data_inicio.set_date(datetime.now() + timedelta(days=1))
        self.ent_data_inicio.bind('<<DateEntrySelected>>', self.ao_mudar_data_inicio)

        ttk.Label(dados_contrato_grid, text="Data Fim:").grid(row=row, column=2, sticky='w', padx=5, pady=2)
        self.ent_data_fim = DateEntry(dados_contrato_grid, width=15,
                                    date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.ent_data_fim.grid(row=row, column=3, sticky='w', padx=5, pady=2)
        self.ent_data_fim.bind('<<DateEntrySelected>>', self.calcular_prazo_contrato)

        row += 1
        ttk.Label(dados_contrato_grid, text="Prazo (dias):").grid(row=row, column=0, sticky='w', padx=5, pady=2)
        self.ent_prazo_dias = ttk.Entry(dados_contrato_grid, width=15)
        self.ent_prazo_dias.bind('<KeyRelease>', self.ao_mudar_dias)
        self.ent_prazo_dias.bind('<FocusOut>', self.ao_mudar_dias)
        self.ent_prazo_dias.grid(row=row, column=1, sticky='w', padx=5, pady=2)

        self.var_data_condicionada = tk.BooleanVar(value=False)
        self.chk_data_condicionada = ttk.Checkbutton(
            dados_contrato_grid,
            text="📦 Início condicionado ao recebimento de material",
            variable=self.var_data_condicionada,
            command=self.alternar_data_condicionada
        )
        self.chk_data_condicionada.grid(row=row, column=2, columnspan=2, sticky='w', padx=5, pady=2)

        row += 1
        ttk.Label(dados_contrato_grid, text="Valor Global:").grid(row=row, column=0, sticky='w', padx=5, pady=2)
        self.ent_valor_global = ttk.Entry(dados_contrato_grid, width=20)
        self.ent_valor_global.grid(row=row, column=1, sticky='w', padx=5, pady=2)
        self.ent_valor_global.insert(0, "R$ 0,00")
        self.ent_valor_global.bind('<FocusOut>', self.formatar_valor_global)

        ttk.Label(dados_contrato_grid, text="Multa:").grid(row=row, column=2, sticky='w', padx=5, pady=2)
        self.ent_multa = ttk.Entry(dados_contrato_grid, width=20)
        self.ent_multa.grid(row=row, column=3, sticky='w', padx=5, pady=2)
        self.ent_multa.insert(0, "R$ 4.000,00")

        row += 1
        ttk.Label(dados_contrato_grid, text="Endereço da Obra:").grid(row=row, column=0, sticky='w', padx=5, pady=2)
        self.ent_endereco_obra = ttk.Entry(dados_contrato_grid, width=60)
        self.ent_endereco_obra.grid(row=row, column=1, columnspan=3, sticky='ew', padx=5, pady=2)

        # --- DESCRIÇÃO DOS SERVIÇOS ---
        frame_servicos = ttk.LabelFrame(scrollable_frame, text="Descrição dos Serviços", padding=10)
        frame_servicos.pack(fill='both', expand=True, pady=5)

        ttk.Button(frame_servicos, text="📋 Selecionar Serviços",
                command=self.abrir_selecao_servicos).pack(pady=5)

        servicos_frame = ttk.Frame(frame_servicos)
        servicos_frame.pack(fill='both', expand=True, pady=5)

        ttk.Label(servicos_frame, text="Serviços selecionados:",
                font=('Arial', 9, 'bold')).pack(anchor='w')

        self.txt_servicos_selecionados = tk.Text(servicos_frame, height=5, width=70, wrap='word')
        self.txt_servicos_selecionados.pack(fill='both', expand=True, pady=5)

        scrollbar_servicos = ttk.Scrollbar(self.txt_servicos_selecionados)
        scrollbar_servicos.pack(side='right', fill='y')
        self.txt_servicos_selecionados.config(yscrollcommand=scrollbar_servicos.set)
        scrollbar_servicos.config(command=self.txt_servicos_selecionados.yview)

        # --- OBSERVAÇÕES ---
        frame_obs = ttk.LabelFrame(scrollable_frame, text="Observações", padding=10)
        frame_obs.pack(fill='x', pady=5)

        self.txt_observacoes_contrato = tk.Text(frame_obs, height=3, width=70, wrap='word')
        self.txt_observacoes_contrato.pack(fill='x', pady=5)

        # --- BOTÕES DE AÇÃO ---
        frame_botoes = ttk.Frame(scrollable_frame)
        frame_botoes.pack(fill='x', pady=10)

        style = ttk.Style()
        style.configure('Action.TButton', font=('Arial', 10, 'bold'))

        ttk.Button(frame_botoes, text="📄 Gerar Contrato",
                command=self.gerar_contrato_final,
                style='Action.TButton').pack(side='left', padx=5)

        ttk.Button(frame_botoes, text="💾 Salvar Contrato",
                command=self.incluir_contrato_na_aba,
                style='Action.TButton').pack(side='left', padx=5)

        ttk.Button(frame_botoes, text="🔄 Limpar Formulário",
                command=self.limpar_formulario_contrato).pack(side='left', padx=5)

        ttk.Button(frame_botoes, text="📂 Abrir Pasta de Contratos",
                command=self.abrir_pasta_contratos).pack(side='left', padx=5)

        ttk.Button(frame_botoes, text="Voltar",
                command=lambda: self.notebook.select(self.aba_contratos)).pack(side='right', padx=5)

        # Empacotar canvas e scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Bind de mudança de aba para carregar dados do contexto
        self.notebook.bind('<<NotebookTabChanged>>', self.on_tab_changed_contrato, add='+')

    def centralizar_janela(self, janela, largura=600, altura=400):
        """Centraliza a janela na tela e mantém visível - CORRIGIDO v3"""
        # Atualizar a geometria
        janela.update_idletasks()
        
        # Obter dimensões da TELA
        tela_largura = janela.winfo_screenwidth()
        tela_altura = janela.winfo_screenheight()
        
        # Calcular posição centralizada na TELA
        x = (tela_largura - largura) // 2
        y = (tela_altura - altura) // 2
        
        # Garantir que não fique fora da tela
        x = max(0, x)
        y = max(0, y)
        
        # Aplicar geometria
        janela.geometry(f"{largura}x{altura}+{x}+{y}")
        
        # CORREÇÃO CRÍTICA: Configuração que NÃO bloqueia campos
        try:
            # 1. Configurar hierarquia primeiro
            janela.transient(self.root)
            
            # 2. Forçar para frente COM topmost
            janela.attributes('-topmost', True)
            janela.lift()
            janela.focus_force()
            janela.update()
            
            # 3. IMPORTANTE: Desabilitar topmost DEPOIS para permitir interação
            # mas manter a janela modal
            janela.after(200, lambda: janela.attributes('-topmost', False))
            
            # 4. Aplicar grab_set para modalidade
            janela.grab_set()
            
            # 5. Garantir que volte ao topo se clicar na janela pai
            def trazer_de_volta(event):
                try:
                    janela.lift()
                    janela.focus_force()
                except:
                    pass
            
            # Bind apenas no clique da janela pai
            self.root.bind('<Button-1>', trazer_de_volta, add='+')
            
            # Garantir visibilidade inicial
            janela.lift()
            janela.focus_force()
            
        except Exception as e:
            # Fallback simples
            janela.lift()
            janela.focus_force()
            janela.grab_set()

    def criar_janela_modal(self, titulo, largura=600, altura=400):
        """Cria janela modal"""
        janela = tk.Toplevel(self.root)
        janela.title(titulo)
        self.centralizar_janela(janela, largura, altura)
        return janela

    def formatar_documento(self, documento):
        
        try:
            if not documento:
                return ''
            
            # Converter para string e limpar
            doc_str = str(documento).strip()
            
            # Remover .0 de floats
            if doc_str.endswith('.0'):
                doc_str = doc_str[:-2]
            
            # Extrair apenas dígitos
            digitos = ''.join(filter(str.isdigit, doc_str))
            
            if not digitos:
                return doc_str
            
            # Formatar baseado no tamanho
            if len(digitos) == 11:  # CPF
                return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
            elif len(digitos) == 14:  # CNPJ
                return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
            else:
                # Retornar sem formatação se tamanho inválido
                return digitos
                
        except Exception as e:
            logger.warning(f"Erro ao formatar documento: {e}")
            return str(documento)
    
    def adicionar_busca_fornecedores(self, frame_fornecedor):
                
        # Frame de busca
        busca_frame = ttk.Frame(frame_fornecedor)
        busca_frame.pack(fill='x', pady=5)
        
        ttk.Label(busca_frame, text="Buscar:").pack(side='left', padx=5)
        
        # Campo de busca
        self.ent_busca_fornecedor = ttk.Entry(busca_frame, width=40)
        self.ent_busca_fornecedor.pack(side='left', padx=5, fill='x', expand=True)
        
        # Bind para busca ao digitar
        self.ent_busca_fornecedor.bind('<KeyRelease>', self.buscar_fornecedor_contrato)
        
        ttk.Button(busca_frame, text="🔍 Buscar", 
                command=lambda: self.buscar_fornecedor_contrato()).pack(side='left', padx=5)
        
        ttk.Button(busca_frame, text="↻ Mostrar Todos",
                command=self.atualizar_lista_fornecedores_contrato).pack(side='left', padx=5)
    
    def buscar_fornecedor_contrato(self, event=None):
        
        try:
            termo = self.ent_busca_fornecedor.get().strip()
            
            if not termo:
                # Se vazio, mostrar todos
                self.atualizar_lista_fornecedores_contrato()
                return
            
            logger.info(f"Buscando fornecedores com termo: {termo}")
            
            from openpyxl import load_workbook
            
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Normalizar termo de busca
            termo_upper = termo.upper()
            termo_numerico = ''.join(filter(str.isdigit, termo))
            
            fornecedores_encontrados = []
            
            # Iterar pelas linhas
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:  # Pular sem CNPJ/CPF
                        continue
                    
                    # Dados da linha
                    cnpj_cpf_raw = row[0]
                    razao_social = row[2] if len(row) > 2 else ''
                    nome_fantasia = row[3] if len(row) > 3 else ''
                    
                    # Converter CNPJ/CPF para comparação
                    row_cnpj = ''.join(filter(str.isdigit, str(cnpj_cpf_raw)))
                    
                    # Verificar se termo está em:
                    # 1. CNPJ/CPF (dígitos)
                    # 2. Nome Fantasia
                    # 3. Razão Social
                    match = False
                    
                    if termo_numerico and termo_numerico in row_cnpj:
                        match = True
                    elif nome_fantasia and termo_upper in str(nome_fantasia).upper():
                        match = True
                    elif razao_social and termo_upper in str(razao_social).upper():
                        match = True
                    
                    if match:
                        # Formatar CNPJ/CPF
                        cnpj_formatado = self.formatar_documento(cnpj_cpf_raw)
                        
                        # Escolher nome a exibir
                        if nome_fantasia and not pd.isna(nome_fantasia):
                            nome = str(nome_fantasia).strip()
                        elif razao_social and not pd.isna(razao_social):
                            nome = str(razao_social).strip()
                        else:
                            nome = f'Fornecedor_Linha_{row_idx}'
                        
                        # Adicionar à lista: "NOME - CNPJ/CPF"
                        item_lista = f"{nome} - {cnpj_formatado}"
                        fornecedores_encontrados.append(item_lista)
                    
                except Exception as e:
                    logger.error(f"Erro ao processar linha {row_idx}: {e}")
                    continue
            
            wb.close()
            
            # Habilitar e limpar listbox (removendo placeholder se houver)
            self.lst_fornecedor_contrato.config(state='normal')
            self.lst_fornecedor_contrato.delete(0, tk.END)
            
            # Atualizar listbox com resultados
            if fornecedores_encontrados:
                for fornecedor in sorted(fornecedores_encontrados):
                    self.lst_fornecedor_contrato.insert(tk.END, fornecedor)
                logger.info(f"✅ {len(fornecedores_encontrados)} fornecedores encontrados")
            else:
                logger.warning(f"⚠️ Nenhum fornecedor encontrado com '{termo}'")
                messagebox.showinfo("Busca", f"Nenhum fornecedor encontrado com '{termo}'", parent=self.root)
            
        except Exception as e:
            logger.error(f"❌ Erro ao buscar fornecedores: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao buscar:\n{str(e)}", parent=self.root)

    def normalizar_cnpj_cpf(self, cnpj_cpf_input):
        """
        Normaliza CNPJ/CPF para busca consistente.
        Remove formatação e garante comparação confiável.
        
        Args:
            cnpj_cpf_input: String com CNPJ/CPF formatado ou não
            
        Returns:
            dict com 'limpo' (só números) e 'formatado' (com máscara)
        """
        try:
            # Garantir que é string
            cnpj_str = str(cnpj_cpf_input) if cnpj_cpf_input else ""
            
            # Remover todos os caracteres não numéricos
            apenas_numeros = ''.join(filter(str.isdigit, cnpj_str))
            
            # Se vazio, retornar vazio
            if not apenas_numeros:
                return {'limpo': '', 'formatado': '', 'tipo': 'INVÁLIDO'}
            
            # Determinar se é CPF ou CNPJ pelo tamanho
            tamanho = len(apenas_numeros)
            
            if tamanho <= 11:
                # É CPF - garantir 11 dígitos com zeros à esquerda
                cpf_completo = apenas_numeros.zfill(11)
                cpf_formatado = f"{cpf_completo[:3]}.{cpf_completo[3:6]}.{cpf_completo[6:9]}-{cpf_completo[9:]}"
                return {
                    'limpo': cpf_completo,
                    'formatado': cpf_formatado,
                    'tipo': 'CPF'
                }
            else:
                # É CNPJ - garantir 14 dígitos com zeros à esquerda
                cnpj_completo = apenas_numeros.zfill(14)
                cnpj_formatado = f"{cnpj_completo[:2]}.{cnpj_completo[2:5]}.{cnpj_completo[5:8]}/{cnpj_completo[8:12]}-{cnpj_completo[12:]}"
                return {
                    'limpo': cnpj_completo,
                    'formatado': cnpj_formatado,
                    'tipo': 'CNPJ'
                }
        except Exception as e:
            logger.error(f"Erro ao normalizar CNPJ/CPF '{cnpj_cpf_input}': {str(e)}")
            return {'limpo': '', 'formatado': '', 'tipo': 'INVÁLIDO'}

    def buscar_fornecedor_por_nome_agenda(self, nome_fornecedor):
        """
        Busca um fornecedor na base de dados pelo nome (razão social ou nome fantasia).
        Usado pelo ImportadorMedicoes para obter o CNPJ a partir do nome.

        Args:
            nome_fornecedor (str): Nome do fornecedor a ser buscado

        Returns:
            dict com 'cnpj_cpf' ou None se não encontrar
        """
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb.active  # Aba 'Fornecedores'

            nome_busca = nome_fornecedor.strip().upper()

            melhor_match = None
            melhor_score = 0
            melhor_tipo = ""

            for row in ws.iter_rows(min_row=2, values_only=True):
                cnpj      = str(row[0]).strip() if row[0] else ""
                razao     = str(row[2]).strip().upper() if row[2] else ""
                nome_fan  = str(row[3]).strip().upper() if row[3] else ""

                # Correspondência exata
                if nome_busca == razao or nome_busca == nome_fan:
                    wb.close()
                    tipo = str(row[1]).strip().upper() if row[1] else ""  # coluna B
                    return {'cnpj_cpf': cnpj, 'tipo_pessoa': tipo}

                # Correspondência parcial
                for candidato in [razao, nome_fan]:
                    if not candidato:
                        continue
                    if nome_busca in candidato or candidato in nome_busca:
                        score = len(set(nome_busca.split()) & set(candidato.split()))
                        if score > melhor_score:
                            melhor_score = score
                            melhor_match = cnpj
                            melhor_tipo  = tipo

            wb.close()

            if melhor_match and melhor_score > 0:
                logger.info(f"Fornecedor '{nome_fornecedor}' → CNPJ '{melhor_match}' (score={melhor_score})")
                return {'cnpj_cpf': melhor_match, 'tipo_pessoa': melhor_tipo}

            logger.warning(f"Fornecedor não encontrado na base: '{nome_fornecedor}'")
            return None

        except Exception as e:
            logger.error(f"Erro ao buscar fornecedor por nome '{nome_fornecedor}': {str(e)}")
            return None

    # Funções da aba Cliente
    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes no combobox usando a função centralizada"""
        try:
            # Usar a função centralizada (apenas clientes ativos)
            self.info_clientes = atualizar_combobox_clientes(self.cliente_combobox, mostrar_inativos=False)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}", parent=self.root)

    # E modifique o método selecionar_cliente:

    def selecionar_cliente(self, event=None):
        """Atualiza o cliente selecionado e preenche painel de dados na aba Seleção"""
        self.cliente_atual = self.cliente_combobox.get()

        if self.cliente_atual:
            # Lembrar este cliente para outros módulos (ex.: Entrada de Dados)
            try:
                from src.estado_compartilhado import estado_sessao
            except ImportError:
                from estado_compartilhado import estado_sessao
            estado_sessao.ultimo_cliente = self.cliente_atual

            if not cliente_esta_ativo(self.cliente_atual):
                messagebox.showwarning(
                    "Cliente Inativo",
                    f"O cliente '{self.cliente_atual}' está inativo (contrato finalizado). " +
                    "Os dados serão mostrados somente para consulta.",
                    parent=self.root
                )

            info_cliente = obter_info_cliente(self.cliente_atual)
            texto_cliente = f"Cliente: {self.cliente_atual}"
            if info_cliente and not info_cliente['ativo']:
                texto_cliente += " (INATIVO)"

            # Atualizar labels nas abas dependentes
            for attr in ('lbl_cliente_contratos', 'lbl_cliente_medicoes',
                         'lbl_cliente_fornecedor', 'lbl_cliente_relatorios', 'lbl_cliente_resumo'):
                if hasattr(self, attr):
                    getattr(self, attr).config(text=texto_cliente)

            # Definir caminho do arquivo
            if info_cliente and 'arquivo' in info_cliente:
                self.arquivo_cliente = info_cliente['arquivo']
            else:
                self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"

            # Preencher painel de dados do cliente na aba Seleção
            try:
                dados_cliente = self._obter_dados_cliente(self.cliente_atual)
                if dados_cliente:
                    def _set(entry, valor):
                        entry.config(state='normal')
                        entry.delete(0, tk.END)
                        entry.insert(0, valor or '')
                        entry.config(state='readonly')

                    _set(self.ent_sel_cno,         dados_cliente.get('cno', ''))
                    _set(self.ent_sel_cpf,         dados_cliente.get('cnpj_cpf', ''))
                    _set(self.ent_sel_estado_civil, dados_cliente.get('estado_civil', ''))
                    _set(self.ent_sel_cidade,
                         self.formatar_nome_cidade(dados_cliente.get('cidade', '')))
                    _set(self.ent_sel_endereco,    dados_cliente.get('endereco', ''))
            except Exception as e:
                logger.warning(f"Não foi possível carregar dados do cliente no painel: {e}")

            self.verificar_aba_medicoes()

    def _obter_dados_cliente(self, nome_cliente):
        """
        Lê dados do cliente diretamente do Clientes.xlsx.
        Independente do GeradorContrato — sem dependência do python-docx.
        """
        try:
            df = pd.read_excel(ARQUIVO_CLIENTES)
 
            # Localizar coluna de nome
            col_cliente = None
            for nome_col in ['Nome', 'Cliente', 'cliente', 'CLIENTE', 'nome', 'NOME']:
                if nome_col in df.columns:
                    col_cliente = nome_col
                    break
 
            if not col_cliente:
                logger.warning("_obter_dados_cliente: coluna de nome não encontrada")
                return None
 
            cliente_row = df[df[col_cliente] == nome_cliente]
            if cliente_row.empty:
                logger.warning(f"_obter_dados_cliente: '{nome_cliente}' não encontrado")
                return None
 
            cliente = cliente_row.iloc[0]
 
            def _safe(col, default=''):
                try:
                    val = cliente.get(col, default)
                    if pd.isna(val) or val is None:
                        return default
                    return str(val).strip()
                except:
                    return default
 
            # Formatar CPF/CNPJ
            cpf_raw = cliente['CPF'] if 'CPF' in cliente.index else None
            cpf_formatado = 'não informado'
            if cpf_raw is not None and not pd.isna(cpf_raw):
                if isinstance(cpf_raw, float):
                    cpf_str = str(int(cpf_raw))
                elif isinstance(cpf_raw, int):
                    cpf_str = str(cpf_raw)
                else:
                    cpf_str = str(cpf_raw).strip()
 
                apenas_numeros = ''.join(filter(str.isdigit, cpf_str))
                if len(apenas_numeros) == 11:
                    cpf_formatado = f"{apenas_numeros[:3]}.{apenas_numeros[3:6]}.{apenas_numeros[6:9]}-{apenas_numeros[9:11]}"
                elif len(apenas_numeros) == 14:
                    cpf_formatado = f"{apenas_numeros[:2]}.{apenas_numeros[2:5]}.{apenas_numeros[5:8]}/{apenas_numeros[8:12]}-{apenas_numeros[12:14]}"
                else:
                    cpf_formatado = cpf_str
 
            # Formatar CNO
            cno_raw = _safe('CNO', '')
            cno_formatado = ''
            if cno_raw:
                if '.' in cno_raw and '/' in cno_raw:
                    cno_formatado = cno_raw
                else:
                    cno_limpo = ''.join(filter(str.isdigit, cno_raw)).zfill(13)
                    if len(cno_limpo) >= 12:
                        cno_formatado = f"{cno_limpo[:2]}.{cno_limpo[2:5]}.{cno_limpo[5:10]}/{cno_limpo[10:12]}"
                    else:
                        cno_formatado = cno_raw
 
            return {
                'nome':         _safe(col_cliente, nome_cliente),
                'cnpj_cpf':     cpf_formatado,
                'cno':          cno_formatado,
                'estado_civil': _safe('Estado Civil', 'não informado'),
                'endereco':     _safe('Endereço', 'não informado'),
                'cidade':       _safe('Cidade', 'Belo Horizonte'),
                'estado':       _safe('Estado', 'MG'),
            }
 
        except Exception as e:
            logger.error(f"_obter_dados_cliente: erro ao ler dados: {e}")
            return None
        
    def continuar_para_fornecedor(self):
        """Avança para a aba de seleção de fornecedor.
        Também limpa o filtro de fornecedor anterior para que a aba
        Contratos exiba todos os contratos do cliente caso o usuário
        navegue diretamente para ela sem selecionar um fornecedor.
        """
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
            return
        # Limpar contexto de fornecedor anterior
        self.cnpj_fornecedor_selecionado = None
        self.nome_fornecedor_selecionado = None
        self.notebook.select(self.aba_fornecedor)

    def continuar_para_contratos(self):
        """Mantido por compatibilidade — use continuar_para_fornecedor no fluxo principal"""
        self.continuar_para_fornecedor()

    def continuar_para_contratos_filtrado(self):
        """Avança para a aba de Contratos filtrando pelo fornecedor selecionado"""
        selecionado = self.lst_fornecedor_aba.curselection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um fornecedor primeiro!", parent=self.root)
            return

        item = self.lst_fornecedor_aba.get(selecionado[0]).strip()
        if not item or item.startswith("👆"):
            messagebox.showwarning("Aviso", "Selecione um fornecedor válido!", parent=self.root)
            return

        # Extrair CNPJ do formato "NOME - CNPJ"
        if " - " in item:
            partes = item.rsplit(" - ", 1)
            self.nome_fornecedor_selecionado = partes[0].strip()
            self.cnpj_fornecedor_selecionado = partes[1].strip()
        else:
            self.nome_fornecedor_selecionado = item
            self.cnpj_fornecedor_selecionado = ""

        # Atualizar label da aba Contratos
        if hasattr(self, 'lbl_cliente_contratos'):
            self.lbl_cliente_contratos.config(
                text=f"Cliente: {self.cliente_atual}  |  Fornecedor: {self.nome_fornecedor_selecionado}"
            )

        self.notebook.select(self.aba_contratos)
        self.carregar_contratos()

        # Atualizar indicador de filtro na aba Relatórios
        if hasattr(self, 'lbl_filtro_pdf'):
            if self.nome_fornecedor_selecionado:
                self.lbl_filtro_pdf.config(
                    text=f"Filtro ativo: {self.nome_fornecedor_selecionado}"
                )
            else:
                self.lbl_filtro_pdf.config(text="")

    # --- Métodos da aba Fornecedor ---

    def buscar_fornecedor_aba(self, event=None):
        """Busca fornecedores na nova aba de seleção"""
        try:
            termo = self.ent_busca_forn_aba.get().strip()
            if not termo:
                self.listar_todos_fornecedores_aba()
                return

            from openpyxl import load_workbook as _lw
            wb = _lw(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']

            termo_upper = termo.upper()
            termo_num = ''.join(filter(str.isdigit, termo))
            encontrados = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                cnpj_raw = row[0]
                razao = row[2] if len(row) > 2 else ''
                fantasia = row[3] if len(row) > 3 else ''
                cnpj_num = ''.join(filter(str.isdigit, str(cnpj_raw)))

                match = (
                    (termo_num and termo_num in cnpj_num) or
                    (fantasia and termo_upper in str(fantasia).upper()) or
                    (razao and termo_upper in str(razao).upper())
                )
                if match:
                    nome = str(fantasia).strip() if fantasia and str(fantasia).strip() else str(razao).strip()
                    cnpj_fmt = self.formatar_documento(cnpj_raw)
                    encontrados.append(f"{nome} - {cnpj_fmt}")

            wb.close()
            self._preencher_listbox_forn_aba(sorted(encontrados))

            if not encontrados:
                messagebox.showinfo("Busca", f"Nenhum fornecedor encontrado com '{termo}'", parent=self.root)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar fornecedores:\n{str(e)}", parent=self.root)

    def listar_todos_fornecedores_aba(self):
        """Lista todos os fornecedores na aba de seleção"""
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']

            fornecedores = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                razao = row[2] if len(row) > 2 else ''
                fantasia = row[3] if len(row) > 3 else ''
                nome = str(fantasia).strip() if fantasia and str(fantasia).strip() else str(razao).strip()
                cnpj_fmt = self.formatar_documento(row[0])
                fornecedores.append(f"{nome} - {cnpj_fmt}")

            wb.close()
            self._preencher_listbox_forn_aba(sorted(fornecedores))

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar fornecedores:\n{str(e)}", parent=self.root)

    def _preencher_listbox_forn_aba(self, itens):
        """Preenche a listbox da aba fornecedor com os itens fornecidos"""
        self.lst_fornecedor_aba.delete(0, tk.END)
        for item in itens:
            self.lst_fornecedor_aba.insert(tk.END, item)

    def selecionar_fornecedor_aba(self, event=None):
        """Carrega dados do fornecedor selecionado na aba de seleção"""
        selecionado = self.lst_fornecedor_aba.curselection()
        if not selecionado:
            return

        item = self.lst_fornecedor_aba.get(selecionado[0]).strip()
        if not item:
            return

        try:
            if " - " in item:
                partes = item.rsplit(" - ", 1)
                nome_busca = partes[0].strip()
                cnpj_busca = partes[1].strip()
            else:
                nome_busca = item
                cnpj_busca = None

            from openpyxl import load_workbook as _lw
            wb = _lw(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']

            fornecedor = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                cnpj_fmt = self.formatar_documento(row[0])
                if cnpj_busca and cnpj_fmt == cnpj_busca:
                    fornecedor = row
                    break

            wb.close()

            if not fornecedor:
                messagebox.showwarning("Aviso", f"Fornecedor não encontrado.", parent=self.root)
                return

            cnpj_cpf = self.formatar_documento(fornecedor[0])
            endereco = str(fornecedor[15]).strip() if len(fornecedor) > 15 and fornecedor[15] else '[ENDEREÇO NÃO CADASTRADO]'

            def _set(entry, valor):
                entry.config(state='normal')
                entry.delete(0, tk.END)
                entry.insert(0, valor)
                entry.config(state='readonly')

            _set(self.ent_forn_cnpj, cnpj_cpf)
            _set(self.ent_forn_endereco, endereco)

            # Dados bancários
            try:
                dados_banc = buscar_dados_bancarios_fornecedor(cnpj_cpf, "PIX") or 'Dados bancários não cadastrados'
            except Exception:
                dados_banc = 'Erro ao buscar dados bancários'

            self.txt_forn_dados_bancarios.config(state='normal')
            self.txt_forn_dados_bancarios.delete('1.0', tk.END)
            self.txt_forn_dados_bancarios.insert('1.0', dados_banc)
            self.txt_forn_dados_bancarios.config(state='disabled')

            # Salvar contexto do fornecedor imediatamente (sem precisar clicar "Continuar →")
            self.nome_fornecedor_selecionado = nome_busca
            self.cnpj_fornecedor_selecionado = cnpj_cpf

            logger.info(f"Fornecedor selecionado na aba: {nome_busca} ({cnpj_cpf})")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do fornecedor:\n{str(e)}", parent=self.root)

    # --- Métodos da aba Contrato (novo / reemissão) ---

    def abrir_novo_contrato(self):
        """Navega para aba Contrato com formulário limpo para novo contrato"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
            return
        if not getattr(self, 'cnpj_fornecedor_selecionado', None):
            resposta = messagebox.askyesno(
                "Fornecedor não selecionado",
                "Nenhum fornecedor está selecionado.\n\nDeseja ir para a aba Fornecedor para selecionar um?",
                parent=self.root
            )
            if resposta:
                self.notebook.select(self.aba_fornecedor)
            return
        self.limpar_formulario_contrato()
        self.notebook.select(self.aba_contratos_emissao)
        self.root.after(50, self._preencher_contexto_aba_contrato)

    def reemitir_contrato(self):
        """Preenche a aba Contrato com dados de um contrato existente para reemissão"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um contrato para reemitir!", parent=self.root)
            return

        valores = self.tree_contratos.item(selecionado)['values']
        id_contrato = valores[0]

        try:
            contrato = self.obter_dados_contrato(id_contrato)
            if not contrato:
                messagebox.showerror("Erro", "Não foi possível carregar os dados do contrato!", parent=self.root)
                return

            self.limpar_formulario_contrato()

            # ── Restaurar contexto do fornecedor a partir do contrato ──
            if contrato.get('nome'):
                self.nome_fornecedor_selecionado = contrato['nome']
            if contrato.get('cnpj'):
                self.cnpj_fornecedor_selecionado = contrato['cnpj']

            # Preencher campos com dados do contrato existente
            if contrato['data_inicio']:
                try:
                    dt = contrato['data_inicio'] if isinstance(contrato['data_inicio'], datetime)                          else datetime.strptime(str(contrato['data_inicio']), '%d/%m/%Y')
                    self.ent_data_inicio.set_date(dt)
                except Exception:
                    pass

            if contrato['data_final']:
                try:
                    dt = contrato['data_final'] if isinstance(contrato['data_final'], datetime)                          else datetime.strptime(str(contrato['data_final']), '%d/%m/%Y')
                    self.ent_data_fim.set_date(dt)
                    self.calcular_prazo_contrato()
                except Exception:
                    pass

            if contrato['valor_global']:
                self.ent_valor_global.delete(0, tk.END)
                valor_fmt = f"R$ {float(contrato['valor_global']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                self.ent_valor_global.insert(0, valor_fmt)

            if contrato['descricao']:
                self.txt_servicos_selecionados.delete('1.0', tk.END)
                self.txt_servicos_selecionados.insert('1.0', contrato['descricao'])

            if contrato['observacao']:
                self.txt_observacoes_contrato.delete('1.0', tk.END)
                self.txt_observacoes_contrato.insert('1.0', contrato['observacao'])

            self.notebook.select(self.aba_contratos_emissao)
            # Forçar atualização dos labels de contexto imediatamente,
            # sem depender do evento <<NotebookTabChanged>> (assíncrono no tkinter)
            self.root.after(50, self._preencher_contexto_aba_contrato)
            logger.info(f"Contrato {id_contrato} carregado para reemissão")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar contrato para reemissão:\n{str(e)}", parent=self.root)
            
    def verificar_aba_medicoes(self):
        """Verifica se a aba de medições existe na planilha do cliente e cria se necessário"""
        try:
            if not os.path.exists(self.arquivo_cliente):
                messagebox.showerror("Erro", f"Arquivo do cliente '{self.cliente_atual}' não encontrado!", parent=self.root)
                return False
                
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de medições já existe
            if "Medicoes" not in wb.sheetnames:
                # Criar aba de medições
                ws = wb.create_sheet("Medicoes")
                
                # Definir cabeçalhos
                headers = [
                    "ID_Contrato", "ID_Medicao", "CNPJ_Fornecedor", "Nome_Fornecedor", 
                    "Data_Medicao", "Data_Pagamento", "Referencia", "Valor", 
                    "Status", "Data_Lancamento", "Observacao"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                wb.save(self.arquivo_cliente)
                return True
            
            return True
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao verificar aba de medições: {str(e)}", parent=self.root)
            return False
    
    # Funções da aba Contratos
    def carregar_contratos(self):
        """Carrega os contratos do cliente atual"""
        try:
            if not self.arquivo_cliente:
                messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
                return
            
            # CORREÇÃO: Atualizar o label do cliente na aba de contratos
            if self.cliente_atual:
                # Verificar se cliente está ativo
                info_cliente = obter_info_cliente(self.cliente_atual)
                texto_cliente = f"Cliente: {self.cliente_atual}"
                if info_cliente and not info_cliente['ativo']:
                    texto_cliente += " (INATIVO)"
                
                if hasattr(self, 'lbl_cliente_contratos'):
                    self.lbl_cliente_contratos.config(text=texto_cliente)
                
            # Limpar treeview
            for item in self.tree_contratos.get_children():
                self.tree_contratos.delete(item)
                
            # Verificar existência da aba de medições
            self.verificar_aba_medicoes()
                
            # Abrir arquivo do cliente
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de contratos existe
            if "Contratos_Medicao" not in wb.sheetnames:
                # Criar aba de contratos
                ws = wb.create_sheet("Contratos_Medicao")
                
                # Definir cabeçalhos
                headers = [
                    "ID_Contrato", "CNPJ_Fornecedor", "Nome_Fornecedor", "Descricao", 
                    "Data_Inicio", "Data_Final", "Valor_Global", "Valor_Pago", "Saldo", "Status", "Observacao"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                wb.save(self.arquivo_cliente)
                return
                
            # Carregar dados da aba Contratos_Medicao
            ws = wb["Contratos_Medicao"]

            # Filtros ativos
            cnpj_filtro = getattr(self, 'cnpj_fornecedor_selecionado', None)
            cnpj_filtro_num = ''.join(filter(str.isdigit, cnpj_filtro)) if cnpj_filtro else ''
            filtro_status = getattr(self, 'var_filtro_status', None)
            status_sel = filtro_status.get() if filtro_status else 'Todos'
            busca_var = getattr(self, 'var_busca_contratos', None)
            busca_txt = busca_var.get().strip().lower() if busca_var else ''

            # Tags de cor por status
            self.tree_contratos.tag_configure('concluido',    background='#E8E8E8', foreground='#666666')
            self.tree_contratos.tag_configure('ativo',        background='#FFFFFF')
            self.tree_contratos.tag_configure('alterado',     background='#FFF176')
            self.tree_contratos.tag_configure('pend',         background='#FFF3E0')  # laranja claro
            self.tree_contratos.tag_configure('pend_conc',    background='#F5E6CC', foreground='#666666')

            # ── RECÁLCULO SILENCIOSO ─────────────────────────────────────────────
            # Recalcula Valor_Pago e Saldo em Contratos_Medicao a partir das
            # medições reais (aba Medicoes), ignorando EXCLUÍDAS.
            # Corrige divergências causadas por exclusões manuais ou falhas
            # anteriores, sem exibir mensagem ao usuário.
            if 'Medicoes' in wb.sheetnames:
                ws_med_rc = wb['Medicoes']
                ws_cont_rc = wb['Contratos_Medicao']

                # Somar valores pagos por contrato (apenas medições não excluídas)
                totais_rc = {}
                for r in ws_med_rc.iter_rows(min_row=2, values_only=True):
                    if r[0] and r[7]:
                        st = str(r[8] or '').strip().upper()
                        if st in ('EXCLUÍDO', 'EXCLUIDO'):
                            continue
                        totais_rc[r[0]] = totais_rc.get(r[0], 0.0) + float(r[7])

                # Comparar com o gravado e corrigir se necessário
                houve_correcao = False
                for idx, r in enumerate(ws_cont_rc.iter_rows(min_row=2, values_only=False), 2):
                    id_cont = r[0].value
                    if not id_cont:
                        continue
                    valor_global = float(r[6].value or 0)
                    valor_pago_gravado = float(r[7].value or 0)
                    total_real = totais_rc.get(id_cont, 0.0)
                    saldo_real = valor_global - total_real

                    # Arredondar para evitar falsos positivos com float
                    if abs(total_real - valor_pago_gravado) >= 0.01:
                        r[7].value = total_real
                        r[7].number_format = '#,##0.00'
                        r[8].value = saldo_real
                        r[8].number_format = '#,##0.00'
                        # Atualizar status se saldo zerou ou voltou a ter saldo
                        status_atual = str(r[9].value or '').strip().upper()
                        if saldo_real <= 0 and status_atual == 'ATIVO':
                            r[9].value = 'CONCLUÍDO'
                        elif saldo_real > 0 and status_atual == 'CONCLUÍDO':
                            r[9].value = 'ATIVO'
                        houve_correcao = True

                if houve_correcao:
                    wb.save(self.arquivo_cliente)
                    # Reler a aba para garantir que a treeview use os dados corrigidos
                    ws = wb['Contratos_Medicao']
            # ── FIM DO RECÁLCULO SILENCIOSO ───────────────────────────────────────

            # Contar medições PENDENTE por contrato (antes de popular treeview)
            pendentes_por_contrato = {}
            if 'Medicoes' in wb.sheetnames:
                ws_med = wb['Medicoes']
                for r_med in ws_med.iter_rows(min_row=2, values_only=True):
                    id_cont = r_med[0]
                    st_med = str(r_med[8] or '').strip().upper()
                    if id_cont and st_med == 'PENDENTE':
                        pendentes_por_contrato[id_cont] = pendentes_por_contrato.get(id_cont, 0) + 1

            # 1ª passagem: índice sequencial por empreiteiro
            seq_por_cnpj = {}
            id_para_seq = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                cnpj_num = ''.join(filter(str.isdigit, str(row[1] or '')))
                seq_por_cnpj[cnpj_num] = seq_por_cnpj.get(cnpj_num, 0) + 1
                id_para_seq[row[0]] = seq_por_cnpj[cnpj_num]

            # 2ª passagem: popular treeview
            total_visiveis = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                # Filtro por fornecedor (CNPJ)
                if cnpj_filtro_num:
                    cnpj_linha = ''.join(filter(str.isdigit, str(row[1] or '')))
                    if cnpj_linha != cnpj_filtro_num:
                        continue

                # Filtro por status
                status_linha = str(row[9] or 'ATIVO').strip().upper() if len(row) > 9 else 'ATIVO'
                if status_sel == 'Ativo' and status_linha not in ('ATIVO', ''):
                    continue
                if status_sel == 'Concluído' and status_linha != 'CONCLUÍDO':
                    continue

                # Filtro por texto (fornecedor ou descrição)
                if busca_txt:
                    forn = str(row[2] or '').lower()
                    desc = str(row[3] or '').lower()
                    if busca_txt not in forn and busca_txt not in desc:
                        continue

                num_seq = id_para_seq.get(row[0], '?')

                data_inicio = row[4].strftime('%d/%m/%Y') if isinstance(row[4], datetime) else str(row[4] or '')
                data_final = ''
                if row[5]:
                    data_final = row[5].strftime('%d/%m/%Y') if isinstance(row[5], datetime) else str(row[5])

                valor_global = f"R$ {row[6]:,.2f}" if isinstance(row[6], (int, float)) else str(row[6] or '0,00')
                valor_pago   = f"R$ {row[7]:,.2f}" if isinstance(row[7], (int, float)) else str(row[7] or '0,00')
                saldo        = f"R$ {row[8]:,.2f}" if isinstance(row[8], (int, float)) else str(row[8] or '0,00')

                pend_count = pendentes_por_contrato.get(row[0], 0)
                pend_str = str(pend_count) if pend_count > 0 else ''

                if pend_count > 0:
                    tag = 'pend_conc' if status_linha == 'CONCLUÍDO' else 'pend'
                else:
                    tag = 'concluido' if status_linha == 'CONCLUÍDO' else 'ativo'

                self.tree_contratos.insert('', 'end', tags=(tag,), values=(
                    row[0], num_seq, row[2], row[3],
                    data_inicio, data_final,
                    valor_global, valor_pago, saldo, pend_str
                ))
                total_visiveis += 1

            # Atualizar contador e indicador de filtro de fornecedor
            if hasattr(self, 'lbl_cont_contratos'):
                self.lbl_cont_contratos.config(text=f"{total_visiveis} contrato(s)")
            self._atualizar_lbl_filtro_fornecedor()

            wb.close()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar contratos: {str(e)}", parent=self.root)
    
    def novo_contrato(self):
        """Abre janela para cadastro de novo contrato"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
            return
            
        # Criar janela
        janela = self.criar_janela_modal("Novo Contrato", largura=800, altura=600)
        
        # Frame principal com scroll
        canvas = tk.Canvas(janela)
        scrollbar = ttk.Scrollbar(janela, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Frame principal dentro do scroll
        frame = ttk.Frame(scrollable_frame, padding="10")
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Frame para busca de fornecedor
        frame_busca = ttk.LabelFrame(frame, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', pady=5)
        
        ttk.Label(frame_busca, text="Nome:").grid(row=0, column=0, padx=5, pady=5)
        busca_entry = ttk.Entry(frame_busca, width=30)
        busca_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Button(frame_busca, text="Buscar", 
                  command=lambda: self.buscar_fornecedor(tree_fornecedores, busca_entry.get())
                 ).grid(row=0, column=2, padx=5, pady=5)
                 
        ttk.Button(frame_busca, text="Novo Fornecedor", 
                  command=self.novo_fornecedor
                 ).grid(row=0, column=3, padx=5, pady=5)
        
        # Frame para lista de fornecedores
        frame_fornecedores = ttk.LabelFrame(frame, text="Fornecedores")
        frame_fornecedores.pack(fill='x', pady=5)
        
        # Treeview para fornecedores
        tree_fornecedores = ttk.Treeview(
            frame_fornecedores, 
            columns=('CNPJ/CPF', 'Nome', 'Categoria'),
            show='headings',
            height=4
        )
        tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        tree_fornecedores.heading('Nome', text='Nome')
        tree_fornecedores.heading('Categoria', text='Categoria')
        
        tree_fornecedores.column('CNPJ/CPF', width=150)
        tree_fornecedores.column('Nome', width=250)
        tree_fornecedores.column('Categoria', width=100)
        
        tree_fornecedores.pack(fill='x', padx=5, pady=5)
        
        # Binding para selecionar fornecedor
        tree_fornecedores.bind('<Double-1>', lambda e: self.selecionar_fornecedor_contrato(
            tree_fornecedores, cnpj_entry, nome_entry
        ))
        
        # Frame para dados do contrato
        frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
        frame_contrato.pack(fill='x', pady=10)
        
        # CNPJ Fornecedor
        ttk.Label(frame_contrato, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        cnpj_entry = ttk.Entry(frame_contrato, width=30)
        cnpj_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        cnpj_entry.config(state='readonly')
        
        # Nome Fornecedor
        ttk.Label(frame_contrato, text="Nome:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        nome_entry = ttk.Entry(frame_contrato, width=50)
        nome_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        nome_entry.config(state='readonly')
        
        # Descrição do Contrato
        ttk.Label(frame_contrato, text="Descrição:*").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        descricao_entry = ttk.Entry(frame_contrato, width=90)
        descricao_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        
        # Data de Início
        ttk.Label(frame_contrato, text="Data de Início:*").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        data_inicio = DateEntry(frame_contrato, width=12, background='darkblue',
                              foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        data_inicio.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        # Data Final
        ttk.Label(frame_contrato, text="Data Final:*").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        data_final = DateEntry(frame_contrato, width=12, background='darkblue',
                              foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        data_final.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        
        # Valor Global
        ttk.Label(frame_contrato, text="Valor Global (R$):*").grid(row=5, column=0, padx=5, pady=5, sticky='e')
        valor_global = ttk.Entry(frame_contrato, width=20)
        valor_global.grid(row=5, column=1, padx=5, pady=5, sticky='w')
        
        # Observações
        ttk.Label(frame_contrato, text="Observações:").grid(row=6, column=0, padx=5, pady=5, sticky='ne')
        observacoes = tk.Text(frame_contrato, width=40, height=4)
        observacoes.grid(row=6, column=1, padx=5, pady=5, sticky='w')
        
        # Frame para botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, 
                 text="Salvar", 
                 command=lambda: self.salvar_contrato(
                     janela,
                     cnpj_entry.get(),
                     nome_entry.get(),
                     descricao_entry.get(),
                     data_inicio.get(),
                     data_final.get(),
                     valor_global.get(),
                     observacoes.get("1.0", "end-1c")
                 )).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes, 
                 text="Cancelar", 
                 command=janela.destroy).pack(side='left', padx=5)
                 
    def selecionar_fornecedor_contrato(self, tree, cnpj_entry, nome_entry):
        """Seleciona um fornecedor da lista para o contrato"""
        selecionado = tree.selection()
        if not selecionado:
            return
            
        # Obter valores selecionados
        valores = tree.item(selecionado)['values']
        
        # Preencher os campos
        cnpj_entry.config(state='normal')
        cnpj_entry.delete(0, tk.END)
        cnpj_entry.insert(0, valores[0])
        cnpj_entry.config(state='readonly')
        
        nome_entry.config(state='normal')
        nome_entry.delete(0, tk.END)
        nome_entry.insert(0, valores[1])
        nome_entry.config(state='readonly')
    
    def buscar_fornecedor(self, tree, termo):
        """Busca fornecedores na base com tratamento melhorado para CNPJ/CPF"""
        try:
            # Limpar treeview
            for item in tree.get_children():
                tree.delete(item)
                    
            if not termo:
                return
                    
            # Carregar planilha de fornecedores
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Buscar fornecedores que contenham o termo
            termo = termo.upper()
            # Verificar se o termo pode ser um CNPJ/CPF (apenas dígitos)
            termo_numerico = ''.join(filter(str.isdigit, termo))
            
            encontrados = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular linhas sem CNPJ/CPF
                    continue
                    
                # Converter CNPJ/CPF da linha para comparação
                row_cnpj = ''.join(filter(str.isdigit, str(row[0])))
                
                # Verificar se termo está no CNPJ/CPF, nome ou razão social
                if (termo_numerico and termo_numerico in row_cnpj) or \
                (termo in str(row[3]).upper()) or \
                (termo in str(row[2]).upper()):
                    # Formatar o CNPJ/CPF corretamente
                    cnpj_formatado = self.formatar_documento(row[0])
                    encontrados.append((cnpj_formatado, row[3], row[11]))
                        
            # Adicionar à treeview
            for fornecedor in encontrados:
                tree.insert('', 'end', values=fornecedor)
                    
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar fornecedores: {str(e)}", parent=self.root)
    
    def novo_fornecedor(self):
        """Abre janela para cadastro de novo fornecedor"""
        try:
            from src.Sistema_Entrada_Dados import SistemaEntradaDados

            # Criar instância apontando para a janela atual como root
            sistema = SistemaEntradaDados(self.root)

            # Criar a janela do formulário diretamente no contexto do sistema
            sistema.janela_fornecedor = tk.Toplevel(self.root)
            sistema.janela_fornecedor.title("Novo Fornecedor")
            sistema.janela_fornecedor.geometry("800x750")
            sistema.janela_fornecedor.transient(self.root)
            sistema.janela_fornecedor.grab_set()

            # Centralizar
            sistema.janela_fornecedor.update_idletasks()
            largura, altura = 800, 750
            pos_x = (sistema.janela_fornecedor.winfo_screenwidth() // 2) - (largura // 2)
            pos_y = (sistema.janela_fornecedor.winfo_screenheight() // 2) - (altura // 2)
            sistema.janela_fornecedor.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")
            sistema.janela_fornecedor.lift()
            sistema.janela_fornecedor.focus_force()
            sistema.janela_fornecedor.attributes('-topmost', True)
            sistema.janela_fornecedor.after(
                500, lambda: sistema.janela_fornecedor.attributes('-topmost', False)
            )

            # Abrir o formulário diretamente
            sistema.setup_formulario_fornecedor()

            # Ao fechar, atualizar a lista de fornecedores na aba atual
            sistema.janela_fornecedor.protocol(
                "WM_DELETE_WINDOW",
                lambda: [sistema.janela_fornecedor.destroy(),
                        self.listar_todos_fornecedores_aba()]
            )

        except ImportError:
            messagebox.showerror(
                "Erro", "Módulo de cadastro de fornecedor não encontrado", parent=self.root
            )
        except Exception as e:
            messagebox.showerror(
                "Erro", f"Erro ao abrir cadastro de fornecedor:\n{str(e)}", parent=self.root
            )

    def editar_fornecedor_aba(self):
        """Abre janela de edição para o fornecedor selecionado na aba Fornecedor"""
        selecionado = self.lst_fornecedor_aba.curselection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um fornecedor para editar!", parent=self.root)
            return

        item = self.lst_fornecedor_aba.get(selecionado[0]).strip()
        if not item:
            return

        cnpj_busca = None
        if " - " in item:
            cnpj_busca = item.rsplit(" - ", 1)[1].strip()

        if not cnpj_busca:
            messagebox.showwarning("Aviso", "Não foi possível identificar o CNPJ/CPF do fornecedor!", parent=self.root)
            return

        try:
            from src.Sistema_Entrada_Dados import SistemaEntradaDados

            sistema = SistemaEntradaDados(self.root)
            fornecedor = sistema.buscar_fornecedor_completo(cnpj_busca)
            if not fornecedor:
                messagebox.showwarning("Aviso", f"Fornecedor não encontrado: {cnpj_busca}", parent=self.root)
                return

            sistema.janela_fornecedor = tk.Toplevel(self.root)
            titulo = fornecedor.get('nome') or fornecedor.get('razao_social') or cnpj_busca
            sistema.janela_fornecedor.title(f"Editar Fornecedor: {titulo}")
            sistema.janela_fornecedor.geometry("800x750")
            sistema.janela_fornecedor.transient(self.root)

            sistema.janela_fornecedor.update_idletasks()
            largura, altura = 800, 750
            pos_x = (sistema.janela_fornecedor.winfo_screenwidth() // 2) - (largura // 2)
            pos_y = (sistema.janela_fornecedor.winfo_screenheight() // 2) - (altura // 2)
            sistema.janela_fornecedor.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")

            sistema.setup_formulario_fornecedor(modo_edicao=True)

            # Pré-preencher campos com dados existentes
            cnpj_cpf = str(fornecedor['cnpj_cpf']).strip()
            tipo_pessoa = 'PJ' if len(''.join(filter(str.isdigit, cnpj_cpf))) > 11 else 'PF'

            sistema.campos_form['cnpj_cpf'].insert(0, cnpj_cpf)
            sistema.campos_form['cnpj_cpf'].config(state='readonly')
            sistema.campos_form['tipo_pessoa'].set(tipo_pessoa)
            sistema.campos_form['tipo_pessoa'].config(state='disabled')
            sistema.campos_form['razao_social'].insert(0, fornecedor.get('razao_social') or '')
            sistema.campos_form['razao_social'].config(state='readonly')
            sistema.campos_form['nome'].insert(0, fornecedor.get('nome') or '')
            sistema.campos_form['telefone'].insert(0, fornecedor.get('telefone') or '')
            sistema.campos_form['email'].insert(0, fornecedor.get('email') or '')
            sistema.campos_form['banco'].set(fornecedor.get('banco') or '')
            sistema.campos_form['op'].insert(0, fornecedor.get('op') or '')
            sistema.campos_form['agencia'].insert(0, fornecedor.get('agencia') or '')
            sistema.campos_form['conta'].insert(0, fornecedor.get('conta') or '')
            sistema.campos_form['chave_pix'].insert(0, fornecedor.get('chave_pix') or '')

            cat_widget = sistema.campos_form.get('categoria')
            if cat_widget:
                if isinstance(cat_widget, ttk.Combobox):
                    cat_widget.set(fornecedor.get('categoria') or '')
                else:
                    cat_widget.insert(0, fornecedor.get('categoria') or '')

            for campo in ('especificacao', 'vinculo', 'endereco', 'responsavel'):
                widget = sistema.campos_form.get(campo)
                if widget:
                    widget.insert(0, fornecedor.get(campo) or '')

            # Ao fechar, recarregar lista e re-selecionar o fornecedor
            def _ao_fechar():
                sistema.janela_fornecedor.destroy()
                self._recarregar_e_selecionar_fornecedor(cnpj_busca)

            sistema.janela_fornecedor.protocol("WM_DELETE_WINDOW", _ao_fechar)

            sistema.janela_fornecedor.lift()
            sistema.janela_fornecedor.focus_force()
            sistema.janela_fornecedor.grab_set()

        except ImportError:
            messagebox.showerror("Erro", "Módulo de cadastro de fornecedor não encontrado", parent=self.root)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir edição de fornecedor:\n{str(e)}", parent=self.root)

    def _recarregar_e_selecionar_fornecedor(self, cnpj_busca):
        """Após edição, recarrega a lista e re-seleciona o fornecedor pelo CNPJ."""
        self.listar_todos_fornecedores_aba()
        cnpj_num = ''.join(filter(str.isdigit, cnpj_busca))
        for i in range(self.lst_fornecedor_aba.size()):
            item = self.lst_fornecedor_aba.get(i)
            if cnpj_num in ''.join(filter(str.isdigit, item)):
                self.lst_fornecedor_aba.selection_clear(0, tk.END)
                self.lst_fornecedor_aba.selection_set(i)
                self.lst_fornecedor_aba.see(i)
                self.selecionar_fornecedor_aba()
                break

    def salvar_contrato(self, janela, cnpj, nome, descricao, data_inicio, data_final, valor_global, observacoes):
        """Salva um novo contrato"""
        try:
            # Validar campos obrigatórios
            if not cnpj or not nome or not descricao or not data_inicio or not valor_global:
                messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!", parent=self.root)
                return
            
            # Validar valor global
            try:
                valor = float(valor_global.replace(',', '.'))
                if valor <= 0:
                    messagebox.showerror("Erro", "Valor global deve ser maior que zero!", parent=self.root)
                    return
            except ValueError:
                messagebox.showerror("Erro", "Valor global inválido!", parent=self.root)
                return
            
            # *** NOVO: Validar data final ***
            if data_final:
                try:
                    data_fim_obj = datetime.strptime(data_final, '%d/%m/%Y')
                    data_inicio_obj = datetime.strptime(data_inicio, '%d/%m/%Y')
                    if data_fim_obj < data_inicio_obj:
                        messagebox.showwarning("Aviso", "Data final não pode ser anterior à data inicial!", parent=self.root)
                        return
                except ValueError:
                    messagebox.showerror("Erro", "Data final inválida! Use o formato dd/mm/aaaa", parent=self.root)
                    return
            
            # Abrir arquivo do cliente
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de contratos existe
            if "Contratos_Medicao" not in wb.sheetnames:
                messagebox.showerror("Erro", "Aba de contratos não encontrada!", parent=self.root)
                wb.close()
                return
            
            ws = wb["Contratos_Medicao"]
            
            # Gerar ID do contrato (próximo número sequencial)
            next_id = 1
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], int) and row[0] >= next_id:
                    next_id = row[0] + 1
            
            # Converter data
            try:
                data = datetime.strptime(data_inicio, '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Erro", "Data inválida! Use o formato dd/mm/aaaa", parent=self.root)
                wb.close()
                return
            
            # Adicionar novo contrato
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1, value=next_id)                # ID_Contrato
            ws.cell(row=proxima_linha, column=2, value=cnpj)                   # CNPJ_Fornecedor
            ws.cell(row=proxima_linha, column=3, value=nome)                   # Nome_Fornecedor
            ws.cell(row=proxima_linha, column=4, value=descricao.upper())      # Descricao
            
            # Data de Início
            data_inicio_cell = ws.cell(row=proxima_linha, column=5, value=data)
            data_inicio_cell.number_format = 'DD/MM/YYYY'
            
            # *** NOVO: Data Final ***
            if data_final:
                try:
                    data_fim_obj = datetime.strptime(data_final, '%d/%m/%Y')
                    data_fim_cell = ws.cell(row=proxima_linha, column=6, value=data_fim_obj)
                    data_fim_cell.number_format = 'DD/MM/YYYY'
                except:
                    ws.cell(row=proxima_linha, column=6, value=None)
            else:
                ws.cell(row=proxima_linha, column=6, value=None)
            
            # Formatação para células de valor (AGORA NA COLUNA 7+)
            valor_cell = ws.cell(row=proxima_linha, column=7, value=valor)     # Valor_Global
            valor_cell.number_format = '#.##0,00'
            
            zero_cell_1 = ws.cell(row=proxima_linha, column=8, value=0)        # Valor_Pago
            zero_cell_1.number_format = '#.##0,00'
            
            saldo_cell = ws.cell(row=proxima_linha, column=9, value=valor)     # Saldo
            saldo_cell.number_format = '#.##0,00'
            
            ws.cell(row=proxima_linha, column=10, value="ATIVO")               # Status
            ws.cell(row=proxima_linha, column=11, value=observacoes.upper())   # Observacao
            
            # Salvar arquivo
            wb.save(self.arquivo_cliente)
            wb.close()
            
            messagebox.showinfo("Sucesso", "Contrato cadastrado com sucesso!", parent=self.root)
            
            # CORREÇÃO: Só fechar janela se foi passada
            if janela:
                janela.destroy()
            
            # Atualizar lista de contratos
            self.carregar_contratos()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar contrato: {str(e)}", parent=self.root)
            try:
                wb.close()
            except:
                pass
    
    def editar_contrato(self):
        """Edita o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um contrato para editar", parent=self.root)
            return
            
        # Obter ID do contrato selecionado
        valores = self.tree_contratos.item(selecionado)['values']
        id_contrato = valores[0]
        
        # Obter dados completos do contrato
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb["Contratos_Medicao"]
            
            dados_contrato = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == id_contrato:
                    dados_contrato = {
                        'id': row[0],
                        'cnpj': row[1],
                        'nome': row[2],
                        'descricao': row[3],
                        'data_inicio': row[4],
                        'data_final': row[5],      # *** NOVO ***
                        'valor_global': row[6],    # Mudou de coluna 5 para 6
                        'valor_pago': row[7],      # Mudou de coluna 6 para 7
                        'saldo': row[8],           # Mudou de coluna 7 para 8
                        'status': row[9],          # Mudou de coluna 8 para 9
                        'observacao': row[10]      # Mudou de coluna 9 para 10
                    }
                    break
                    
            wb.close()
            
            if not dados_contrato:
                messagebox.showerror("Erro", "Contrato não encontrado!, parent=self.root")
                return
                
            # Criar janela de edição
            janela = self.criar_janela_modal("Editar Contrato", largura=800, altura=600)
                        
            # Frame principal
            frame = ttk.Frame(janela, padding="10")
            frame.pack(fill='both', expand=True)

            # Frame para dados do fornecedor
            frame_fornecedor = ttk.LabelFrame(frame, text="Dados do Fornecedor")
            frame_fornecedor.pack(fill='x', pady=5)
            
            # CNPJ/CPF
            ttk.Label(frame_fornecedor, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
            cnpj_entry = ttk.Entry(frame_fornecedor, width=30)
            cnpj_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            cnpj_entry.insert(0, dados_contrato['cnpj'])
            cnpj_entry.config(state='readonly')
            
            # Nome
            ttk.Label(frame_fornecedor, text="Nome:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
            nome_entry = ttk.Entry(frame_fornecedor, width=50)
            nome_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
            nome_entry.insert(0, dados_contrato['nome'])
            nome_entry.config(state='readonly')
            
            # Frame para dados do contrato
            frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
            frame_contrato.pack(fill='x', pady=10)
            
            # Descrição
            ttk.Label(frame_contrato, text="Descrição:*").grid(row=0, column=0, padx=5, pady=5, sticky='e')
            descricao_entry = ttk.Entry(frame_contrato, width=80)
            descricao_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            descricao_entry.insert(0, dados_contrato['descricao'])
            
            # Data de Início
            ttk.Label(frame_contrato, text="Data de Início:*").grid(row=1, column=0, padx=5, pady=5, sticky='e')
            data_inicio = DateEntry(frame_contrato, width=12, background='darkblue',
                                foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
            data_inicio.grid(row=1, column=1, padx=5, pady=5, sticky='w')
            if isinstance(dados_contrato['data_inicio'], datetime):
                data_inicio.set_date(dados_contrato['data_inicio'])

            # Data Final
            ttk.Label(frame_contrato, text="Data Final:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
            data_final = DateEntry(frame_contrato, width=12, background='darkblue',
                                foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
            data_final.grid(row=2, column=1, padx=5, pady=5, sticky='w')
            if dados_contrato.get('data_final') and isinstance(dados_contrato['data_final'], datetime):
                data_final.set_date(dados_contrato['data_final'])

            # Valor Global (ajuste o row para 3)
            ttk.Label(frame_contrato, text="Valor Global (R$):*").grid(row=3, column=0, padx=5, pady=5, sticky='e')
            valor_global = ttk.Entry(frame_contrato, width=20)
            valor_global.grid(row=3, column=1, padx=5, pady=5, sticky='w')
            valor_global.insert(0, str(dados_contrato['valor_global']))
            
            # Valor Pago (somente leitura)
            ttk.Label(frame_contrato, text="Valor Pago (R$):").grid(row=4, column=0, padx=5, pady=5, sticky='e')
            valor_pago = ttk.Entry(frame_contrato, width=20)
            valor_pago.grid(row=4, column=1, padx=5, pady=5, sticky='w')
            valor_pago.insert(0, str(dados_contrato['valor_pago']))
            valor_pago.config(state='readonly')
            
            # Saldo (somente leitura)
            ttk.Label(frame_contrato, text="Saldo (R$):").grid(row=5, column=0, padx=5, pady=5, sticky='e')
            saldo = ttk.Entry(frame_contrato, width=20)
            saldo.grid(row=5, column=1, padx=5, pady=5, sticky='w')
            saldo.insert(0, str(dados_contrato['saldo']))
            saldo.config(state='readonly')
            
            # Status
            ttk.Label(frame_contrato, text="Status:").grid(row=6, column=0, padx=5, pady=5, sticky='e')
            status = ttk.Combobox(frame_contrato, values=["ATIVO", "CONCLUÍDO", "CANCELADO"], width=15)
            status.grid(row=6, column=1, padx=5, pady=5, sticky='w')
            status.set(dados_contrato['status'] if dados_contrato['status'] else "ATIVO")
            
            # Observações
            ttk.Label(frame_contrato, text="Observações:").grid(row=7, column=0, padx=5, pady=5, sticky='ne')
            observacoes = tk.Text(frame_contrato, width=40, height=4)
            observacoes.grid(row=7, column=1, padx=5, pady=5, sticky='w')
            observacoes.insert("1.0", dados_contrato['observacao'] if dados_contrato['observacao'] else "")
            
            # Frame para botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)
            
            ttk.Button(frame_botoes, text="Salvar", 
                command=lambda: self.atualizar_contrato(
                    janela,
                    dados_contrato['id'],
                    descricao_entry.get(),
                    data_inicio.get_date().strftime('%d/%m/%Y'),
                    data_final.get_date().strftime('%d/%m/%Y'),  # ← NOVO PARÂMETRO!
                    valor_global.get(),
                    status.get(),
                    observacoes.get('1.0', tk.END).strip()
                )).pack(side='left', padx=5)
            
            ttk.Button(frame_botoes, 
                     text="Cancelar", 
                     command=janela.destroy).pack(side='left', padx=5)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao editar contrato: {str(e)}", parent=self.root)

    def verificar_medicoes_contrato(self, id_contrato):
        """Verifica se o contrato possui medições registradas"""
        try:
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de medições existe
            if "Medicoes" not in wb.sheetnames:
                wb.close()
                return False
            
            ws = wb["Medicoes"]
            
            # Verificar se há medições para este contrato
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == id_contrato:  # Coluna B contém ID_Contrato
                    wb.close()
                    return True
            
            wb.close()
            return False
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao verificar medições: {str(e)}", parent=self.root)
            return True  # Em caso de erro, assume que tem medições (segurança)

    def excluir_contrato(self):
        """Exclui o contrato selecionado (apenas se não tiver medições)"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um contrato para excluir", parent=self.root)
            return
        
        # Obter dados do contrato selecionado
        valores = self.tree_contratos.item(selecionado)['values']
        id_contrato = valores[0]
        fornecedor = valores[1]
        descricao = valores[2]
        
        try:
            # Verificar se o contrato possui medições
            if self.verificar_medicoes_contrato(id_contrato):
                messagebox.showerror(
                    "Erro", 
                    f"Não é possível excluir o contrato ID {id_contrato}!\n\n"
                    "Este contrato possui medições registradas.\n"
                    "Para excluí-lo, primeiro exclua todas as medições associadas.",
                    parent=self.root
                )
                return
            
            # Confirmação de exclusão
            confirmacao = messagebox.askyesno(
                "Confirmar Exclusão",
                f"Deseja realmente excluir o contrato?\n\n"
                f"ID: {id_contrato}\n"
                f"Fornecedor: {fornecedor}\n"
                f"Descrição: {descricao}\n\n"
                "Esta ação não pode ser desfeita!",
                parent=self.root
            )
            
            if not confirmacao:
                return
            
            # Abrir arquivo do cliente
            wb = load_workbook(self.arquivo_cliente)
            ws = wb["Contratos_Medicao"]
            
            # Buscar e excluir a linha do contrato
            linha_excluir = None
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), 2):
                if row[0] == id_contrato:
                    linha_excluir = idx
                    break
            
            if not linha_excluir:
                messagebox.showerror("Erro", "Contrato não encontrado na planilha!", parent=self.root)
                wb.close()
                return
            
            # Excluir a linha
            ws.delete_rows(linha_excluir, 1)
            
            # Salvar arquivo
            wb.save(self.arquivo_cliente)
            wb.close()
            
            messagebox.showinfo(
                "Sucesso", 
                f"Contrato ID {id_contrato} excluído com sucesso!",
                parent=self.root
            )
            
            # Limpar seleção atual se for o contrato excluído
            if hasattr(self, 'contrato_atual') and self.contrato_atual == id_contrato:
                self.contrato_atual = None
                self.fornecedor_atual = None
                if hasattr(self, 'lbl_contrato_medicoes'):
                    self.lbl_contrato_medicoes.config(text="Contrato: Nenhum selecionado")
            
            # Atualizar lista de contratos
            self.carregar_contratos()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir contrato: {str(e)}", parent=self.root)
            try:
                wb.close()
            except:
                pass

    def atualizar_contrato(self, janela, id_contrato, descricao, data_inicio, 
                       data_final, valor_global, status, observacoes):
        """Atualiza os dados de um contrato"""
        try:
            # Validar campos obrigatórios
            if not descricao or not data_inicio or not valor_global:
                messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!", parent=self.root)
                return
            
            # Validar valor global
            try:
                valor = float(valor_global.replace(',', '.'))
                if valor <= 0:
                    messagebox.showerror("Erro", "Valor global deve ser maior que zero!", parent=self.root)
                    return
            except ValueError:
                messagebox.showerror("Erro", "Valor global inválido!", parent=self.root)
                return
            
            # *** NOVO: Validar data final ***
            if data_final:
                try:
                    data_fim_obj = datetime.strptime(data_final, '%d/%m/%Y')
                    data_inicio_obj = datetime.strptime(data_inicio, '%d/%m/%Y')
                    if data_fim_obj < data_inicio_obj:
                        messagebox.showwarning("Aviso", "Data final não pode ser anterior à data inicial!", parent=self.root)
                        return
                except ValueError:
                    messagebox.showerror("Erro", "Data final inválida! Use o formato dd/mm/aaaa", parent=self.root)
                    return
            
            # Abrir arquivo do cliente
            wb = load_workbook(self.arquivo_cliente)
            ws = wb["Contratos_Medicao"]
            
            # Buscar contrato pelo ID
            valor_pago = 0
            row_index = None
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), 2):
                if row[0] == id_contrato:
                    row_index = idx
                    # ⚠️ ATENÇÃO: Valor_Pago agora está na coluna 8 (não 7)
                    valor_pago = ws.cell(row=row_index, column=8).value or 0
                    break
                    
            if not row_index:
                messagebox.showerror("Erro", "Contrato não encontrado!", parent=self.root)
                wb.close()
                return
                
            # Converter data início
            try:
                data = datetime.strptime(data_inicio, '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Erro", "Data inválida! Use o formato dd/mm/aaaa", parent=self.root)
                wb.close()
                return
                
            # Calcular novo saldo
            saldo = valor - valor_pago
            
            # Atualizar dados
            ws.cell(row=row_index, column=4, value=descricao.upper())     # Descricao
            
            # Data de Início
            data_inicio_cell = ws.cell(row=row_index, column=5, value=data)
            data_inicio_cell.number_format = 'DD/MM/YYYY'
            
            # *** NOVO: Data Final ***
            if data_final:
                try:
                    data_fim_obj = datetime.strptime(data_final, '%d/%m/%Y')
                    data_fim_cell = ws.cell(row=row_index, column=6, value=data_fim_obj)
                    data_fim_cell.number_format = 'DD/MM/YYYY'
                except:
                    ws.cell(row=row_index, column=6, value=None)
            else:
                ws.cell(row=row_index, column=6, value=None)
            
            # ⚠️ COLUNAS MUDARAM DE POSIÇÃO!
            # Valor Global (ERA coluna 6, AGORA é coluna 7)
            valor_cell = ws.cell(row=row_index, column=7, value=valor)
            valor_cell.number_format = '#.##0,00'
            
            # Valor Pago (ERA coluna 7, AGORA é coluna 8)
            valor_pago_cell = ws.cell(row=row_index, column=8, value=valor_pago)
            valor_pago_cell.number_format = '#.##0,00'

            # Saldo (ERA coluna 8, AGORA é coluna 9)
            saldo_cell = ws.cell(row=row_index, column=9, value=saldo)
            saldo_cell.number_format = '#.##0,00'
            
            # Status (ERA coluna 9, AGORA é coluna 10)
            ws.cell(row=row_index, column=10, value=status)
            
            # Observacao (ERA coluna 10, AGORA é coluna 11)
            ws.cell(row=row_index, column=11, value=observacoes.upper())
            
            # Salvar arquivo
            wb.save(self.arquivo_cliente)
            wb.close()
            
            messagebox.showinfo("Sucesso", "Contrato atualizado com sucesso!", parent=self.root)
            janela.destroy()
            
            # Atualizar lista de contratos
            self.carregar_contratos()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar contrato: {str(e)}", parent=self.root)
            try:
                wb.close()
            except:
                pass
    
    def selecionar_contrato(self):
        """Seleciona um contrato para visualizar/editar medições"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um contrato", parent=self.root)
            return
            
        # Obter ID e nome do fornecedor do contrato selecionado
        valores = self.tree_contratos.item(selecionado)['values']
        id_contrato = valores[0]
        nome_fornecedor = valores[1]
        
        # Armazenar contrato atual
        self.contrato_atual = id_contrato
        self.fornecedor_atual = nome_fornecedor
        
        # Atualizar label na aba de medições
        self.lbl_contrato_medicoes.config(text=f"Contrato: {id_contrato} - {nome_fornecedor}")
        
        # Carregar medições do contrato
        self.carregar_medicoes()
        
        # Mudar para a aba de medições
        self.notebook.select(self.aba_medicoes)  # Índice da aba de medições
    
    def on_tab_changed_contrato(self, event):
        """Chamado quando a aba é alterada"""
        if self.notebook.select() == str(self.aba_contratos_emissao):
            self._preencher_contexto_aba_contrato()

    def _preencher_contexto_aba_contrato(self):
        """Preenche os labels de contexto (cliente e fornecedor) na aba Contrato"""
        nome_cliente = self.cliente_atual or "Nenhum selecionado"
        nome_forn = getattr(self, 'nome_fornecedor_selecionado', None) or "Nenhum selecionado"
        cnpj_forn = getattr(self, 'cnpj_fornecedor_selecionado', None) or ""

        if hasattr(self, 'lbl_contrato_cliente'):
            self.lbl_contrato_cliente.config(text=f"Cliente: {nome_cliente}")
        if hasattr(self, 'lbl_contrato_fornecedor'):
            txt = f"Fornecedor: {nome_forn}"
            if cnpj_forn:
                txt += f"  ({cnpj_forn})"
            self.lbl_contrato_fornecedor.config(text=txt)

        # Pré-preencher endereço da obra com endereço do cliente (se campo vazio)
        try:
            if hasattr(self, 'ent_endereco_obra') and not self.ent_endereco_obra.get().strip():
                dados_cliente = self.gerador_contrato.obter_dados_cliente(self.cliente_atual)
                if dados_cliente:
                    self.ent_endereco_obra.delete(0, tk.END)
                    self.ent_endereco_obra.insert(0, dados_cliente.get('endereco', ''))
        except Exception:
            pass

    def carregar_dados_cliente_contrato(self):
        """Carrega dados do cliente selecionado na aba de contrato"""
        if not self.cliente_atual:
            self.lbl_cliente_contrato.config(text="Cliente: Nenhum selecionado")
            return
        
        try:
            dados_cliente = self.gerador_contrato.obter_dados_cliente(self.cliente_atual)
            
            if dados_cliente:
                self.lbl_cliente_contrato.config(text=f"Cliente: {dados_cliente['nome']}")
                
                # Preencher campos
                self.ent_cno.config(state='normal')
                self.ent_cno.delete(0, tk.END)
                self.ent_cno.insert(0, dados_cliente.get('cno', ''))
                self.ent_cno.config(state='readonly')
                
                self.ent_cpf_cliente.config(state='normal')
                self.ent_cpf_cliente.delete(0, tk.END)
                self.ent_cpf_cliente.insert(0, dados_cliente.get('cnpj_cpf', ''))
                self.ent_cpf_cliente.config(state='readonly')
                
                self.ent_estado_civil.config(state='normal')
                self.ent_estado_civil.delete(0, tk.END)
                self.ent_estado_civil.insert(0, dados_cliente.get('estado_civil', ''))
                self.ent_estado_civil.config(state='readonly')
                
                self.ent_cidade.config(state='normal')
                self.ent_cidade.delete(0, tk.END)
                cidade_formatada = self.formatar_nome_cidade(dados_cliente.get('cidade', ''))
                self.ent_cidade.insert(0, cidade_formatada)
                self.ent_cidade.config(state='readonly')
                
                self.ent_endereco_cliente.config(state='normal')
                self.ent_endereco_cliente.delete(0, tk.END)
                self.ent_endereco_cliente.insert(0, dados_cliente.get('endereco', ''))
                self.ent_endereco_cliente.config(state='readonly')
                
                # Pré-preencher endereço da obra com endereço do cliente
                self.ent_endereco_obra.delete(0, tk.END)
                self.ent_endereco_obra.insert(0, dados_cliente.get('endereco', ''))
                
                logger.info(f"Dados do cliente carregados na aba de contrato: {self.cliente_atual}")
            else:
                messagebox.showwarning("Aviso", "Não foi possível carregar os dados do cliente.", parent=self.root)
        except Exception as e:
            logger.error(f"Erro ao carregar dados do cliente para contrato: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar dados do cliente: {str(e)}", parent=self.root)

    def atualizar_lista_fornecedores_contrato(self):
        
        try:
            logger.info("Carregando lista de fornecedores...")
            
            # Usar openpyxl igual ao método buscar_fornecedor
            from openpyxl import load_workbook
            
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            fornecedores = []
            fornecedores_com_problema = []
            
            # Iterar pelas linhas (começando da linha 2, pulando cabeçalho)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # row[0] = CNPJ/CPF
                    # row[2] = Razão Social  
                    # row[3] = Nome Fantasia
                    
                    if not row[0]:  # Pular linhas sem CNPJ/CPF
                        continue
                    
                    # === OBTER CNPJ/CPF ===
                    cnpj_cpf_raw = row[0]
                    
                    # Converter e limpar
                    if pd.isna(cnpj_cpf_raw) or cnpj_cpf_raw == '':
                        cnpj_cpf_formatado = 'Sem CPF/CNPJ'
                        fornecedores_com_problema.append(f"Linha {row_idx}")
                    else:
                        # Usar o mesmo método de formatação
                        cnpj_cpf_formatado = self.formatar_documento(cnpj_cpf_raw)
                    
                    # === OBTER NOME (priorizar Nome Fantasia, depois Razão Social) ===
                    nome_fantasia = row[3] if len(row) > 3 else None
                    razao_social = row[2] if len(row) > 2 else None
                    
                    # Escolher qual nome usar
                    if nome_fantasia and not pd.isna(nome_fantasia) and str(nome_fantasia).strip():
                        nome = str(nome_fantasia).strip()
                    elif razao_social and not pd.isna(razao_social) and str(razao_social).strip():
                        nome = str(razao_social).strip()
                    else:
                        nome = f'Fornecedor_Linha_{row_idx}'
                        fornecedores_com_problema.append(nome)
                    
                    # Adicionar à lista: "NOME - CNPJ/CPF"
                    item_lista = f"{nome} - {cnpj_cpf_formatado}"
                    fornecedores.append(item_lista)
                    
                except Exception as e:
                    logger.error(f"Erro ao processar fornecedor linha {row_idx}: {e}")
                    continue
            
            wb.close()
            
            # Habilitar e limpar listbox (removendo placeholder se houver)
            self.lst_fornecedor_contrato.config(state='normal')
            self.lst_fornecedor_contrato.delete(0, tk.END)
            
            # Atualizar listbox (ordenado)
            for fornecedor in sorted(fornecedores):
                self.lst_fornecedor_contrato.insert(tk.END, fornecedor)
            
            # Logs informativos
            logger.info(f"✅ {len(fornecedores)} fornecedores carregados")
            
            if fornecedores_com_problema:
                logger.warning(
                    f"⚠️ {len(fornecedores_com_problema)} fornecedores com dados incompletos"
                )
            
        except Exception as e:
            logger.error(f"❌ Erro ao carregar fornecedores: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao carregar fornecedores:\n{str(e)}", parent=self.root)


    def carregar_dados_fornecedor_contrato(self, event=None):
        
        # Obter item selecionado do Listbox
        selection = self.lst_fornecedor_contrato.curselection()
        if not selection:
            return
        
        # Pegar o item selecionado no formato "NOME - CNPJ"
        item_selecionado = self.lst_fornecedor_contrato.get(selection[0]).strip()
        
        if not item_selecionado or item_selecionado.startswith("👆"):
            return
        
        try:
            # Extrair CNPJ do formato "NOME - CNPJ"
            if " - " in item_selecionado:
                partes = item_selecionado.rsplit(" - ", 1)  # Split da direita para evitar problemas com " - " no nome
                fornecedor_nome = partes[0].strip()
                cnpj_busca = partes[1].strip()
            else:
                # Fallback: se não tiver o formato esperado, usar o item todo como nome
                fornecedor_nome = item_selecionado
                cnpj_busca = None
            
            logger.info(f"Carregando dados do fornecedor: {fornecedor_nome} (CNPJ: {cnpj_busca})")
            
            from openpyxl import load_workbook
            
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Buscar fornecedor pelo CNPJ (mais preciso) ou pelo NOME
            fornecedor_encontrado = None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pular sem CNPJ/CPF
                    continue
                
                # Obter CNPJ formatado da linha
                cnpj_cpf_raw = row[0]
                cnpj_linha = self.formatar_documento(cnpj_cpf_raw)
                
                # Buscar por CNPJ (mais confiável) ou por nome
                if cnpj_busca and cnpj_linha == cnpj_busca:
                    # Encontrou pelo CNPJ - método preferencial
                    fornecedor_encontrado = row
                    break
                
                # Fallback: buscar por nome se CNPJ não disponível
                if not cnpj_busca:
                    razao_social = row[2] if len(row) > 2 else ''
                    nome_fantasia = row[3] if len(row) > 3 else ''
                    
                    # Verificar se é o fornecedor procurado
                    if (nome_fantasia and str(nome_fantasia).strip() == fornecedor_nome) or \
                       (razao_social and str(razao_social).strip() == fornecedor_nome):
                        fornecedor_encontrado = row
                        break
            
            wb.close()
            
            if not fornecedor_encontrado:
                messagebox.showwarning(
                    "Aviso",
                    f"Fornecedor '{fornecedor_nome}' não encontrado na planilha.", 
                    parent=self.root
                )
                return
            
            # Extrair dados do fornecedor
            cnpj_cpf_raw = fornecedor_encontrado[0]
            endereco_raw = fornecedor_encontrado[15] if len(fornecedor_encontrado) > 15 else ''  
            dados_banc_raw = fornecedor_encontrado[14] if len(fornecedor_encontrado) > 14 else ''
            
            # Formatar CNPJ/CPF
            cnpj_cpf_formatado = self.formatar_documento(cnpj_cpf_raw)
            
            # === PREENCHER CNPJ/CPF ===
            self.ent_cnpj_fornecedor.config(state='normal')
            self.ent_cnpj_fornecedor.delete(0, tk.END)
            self.ent_cnpj_fornecedor.insert(0, cnpj_cpf_formatado)
            self.ent_cnpj_fornecedor.config(state='readonly')
            
            # === PREENCHER ENDEREÇO ===
            if pd.isna(endereco_raw) or not str(endereco_raw).strip():
                endereco = '[ENDEREÇO NÃO CADASTRADO]'
                logger.warning(f"Fornecedor '{fornecedor_nome}' sem endereço cadastrado")
            else:
                endereco = str(endereco_raw).strip()
            
            self.ent_endereco_fornecedor.config(state='normal')
            self.ent_endereco_fornecedor.delete(0, tk.END)
            self.ent_endereco_fornecedor.insert(0, endereco)
            self.ent_endereco_fornecedor.config(state='readonly')
            
            # === BUSCAR DADOS BANCÁRIOS ===
            try:
                dados_bancarios = buscar_dados_bancarios_fornecedor(cnpj_cpf_formatado, "PIX")
                
                if not dados_bancarios or dados_bancarios.strip() == '':
                    dados_bancarios = (
                        f"⚠️ Dados bancários não encontrados para:\n"
                        f"{fornecedor_nome}\n"
                        f"CPF/CNPJ: {cnpj_cpf_formatado}\n\n"
                        f"Por favor, cadastre os dados bancários."
                    )
                    logger.warning(f"Dados bancários não encontrados para {fornecedor_nome}")
                
            except Exception as e:
                logger.error(f"Erro ao buscar dados bancários: {e}")
                dados_bancarios = f"❌ Erro ao buscar dados bancários: {str(e)}"
            
            self.txt_dados_bancarios.config(state='normal')
            self.txt_dados_bancarios.delete('1.0', tk.END)
            self.txt_dados_bancarios.insert('1.0', dados_bancarios)
            self.txt_dados_bancarios.config(state='disabled')
            
            logger.info(f"✅ Dados do fornecedor '{fornecedor_nome}' carregados")
            
        except Exception as e:
            logger.error(f"❌ Erro ao carregar dados do fornecedor: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao carregar dados do fornecedor:\n{str(e)}", parent=self.root)

    def ao_mudar_data_contrato(self, event=None):
        """Quando Data do Contrato mudar, apenas registra"""
        try:
            data_contrato = self.ent_data_contrato.get_date()
            logger.info(f"Data do Contrato alterada: {data_contrato}")
            # NÃO sincronizar com Data Início - são independentes
        except Exception as e:
            logger.error(f"Erro ao alterar data contrato: {e}")
    
    def ao_mudar_data_inicio(self, event=None):
        """Quando Data Início mudar, apenas recalcula prazo"""
        try:
            data_inicio = self.ent_data_inicio.get_date()
            logger.info(f"Data Início alterada: {data_inicio}")
            # NÃO sincronizar com Data do Contrato - são independentes
            
            # Calcular prazo se não estiver em modo condicionado
            if not self.var_data_condicionada.get():
                self.calcular_prazo_contrato(event)
        except Exception as e:
            logger.error(f"Erro ao alterar data início: {e}")
    
    def ao_mudar_dias(self, event=None):
        """Quando usuário digitar dias, recalcula Data Fim"""
        try:
            dias_str = self.ent_prazo_dias.get().strip()
            if not dias_str or not dias_str.isdigit():
                return
            
            dias_uteis = int(dias_str)
            data_inicio = self.ent_data_inicio.get_date()
            
            # CORREÇÃO: Data início é DIA 0 (não conta)
            # Calcular data_fim somando dias úteis A PARTIR do dia seguinte
            data_atual = data_inicio
            dias_contados = 0
            
            # Começar contando do dia SEGUINTE à data início
            while dias_contados < dias_uteis:
                data_atual += pd.Timedelta(days=1)
                if data_atual.weekday() < 5:  # Segunda a Sexta
                    dias_contados += 1
            
            # Atualizar Data Fim
            self.ent_data_fim.set_date(data_atual)
            
            logger.info(f"Data Fim calculada: {data_atual.strftime('%d/%m/%Y')} ({dias_uteis} dias úteis, excluindo data início)")
        except Exception as e:
            logger.error(f"Erro ao calcular data fim: {e}")
     
    def calcular_prazo_contrato(self, event=None):
        """Calcula o prazo em DIAS ÚTEIS entre Data Início e Data Fim"""
        try:
            data_inicio = self.ent_data_inicio.get_date()
            data_fim = self.ent_data_fim.get_date()
            
            if data_inicio and data_fim:
                # Verificar se data fim é fim de semana
                if data_fim.weekday() >= 5:  # Sábado ou Domingo
                    dia_semana = 'sábado' if data_fim.weekday() == 5 else 'domingo'
                    logger.warning(f"⚠️ Data final cai em {dia_semana}: {data_fim.strftime('%d/%m/%Y')}")
                    
                # CORREÇÃO: Calcular dias úteis EXCLUINDO a data de início
                # Data início = DIA 0 (não conta)
                dias_uteis = 0
                data_atual = data_inicio + pd.Timedelta(days=1)  # Começa no dia seguinte
                
                while data_atual <= data_fim:
                    # 0 = Segunda, 6 = Domingo
                    if data_atual.weekday() < 5:  # Segunda a Sexta
                        dias_uteis += 1
                    data_atual += pd.Timedelta(days=1)
                
                self.ent_prazo_dias.delete(0, tk.END)
                self.ent_prazo_dias.insert(0, str(dias_uteis))
                
                logger.info(f"Prazo calculado: {dias_uteis} dias úteis (excluindo data início)")
        except Exception as e:
            logger.error(f"Erro ao calcular prazo: {e}")

    def alternar_data_condicionada(self):
        """
        Alterna entre modo de data condicionada e modo normal.
        
        IMPORTANTE: A Data do Contrato NUNCA é afetada!
        Apenas Data Início e Data Fim são condicionadas.
        """
        try:
            condicionado = self.var_data_condicionada.get()
            
            if condicionado:
                logger.info("✓ Modo data condicionada ATIVADO")
                
                # Desabilitar APENAS Data Início e Data Fim
                self.ent_data_inicio.configure(state='disabled')
                self.ent_data_fim.configure(state='disabled')
                
                # Data do Contrato permanece HABILITADA
                
                messagebox.showinfo(
                    "Data Condicionada Ativada",
                    "✓ As datas de INÍCIO e FIM foram desabilitadas.\n\n"
                    "📋 O contrato incluirá a seguinte cláusula:\n"
                    "\"O início dos serviços está condicionado ao\n"
                    "recebimento do material necessário na obra.\"\n\n"
                    "⏱️ O prazo em dias permanece como referência contratual.\n\n"
                    "📅 A DATA DO CONTRATO não é afetada (sempre conhecida).",
                    parent=self.root
                )
                
            else:
                logger.info("✓ Modo data condicionada DESATIVADO")
                
                self.ent_data_inicio.configure(state='normal')
                self.ent_data_fim.configure(state='normal')
                
                # Recalcular prazo automaticamente
                self.calcular_prazo_contrato()
                
        except Exception as e:
            logger.error(f"❌ Erro ao alternar data condicionada: {e}")
            messagebox.showerror(
                "Erro",
                f"Erro ao alternar modo de data:\n{str(e)}",
                parent=self.root
            )

    def abrir_selecao_servicos(self):
        """Abre janela para seleção de serviços - VERSÃO COM COMBOBOX"""
        janela = tk.Toplevel(self.root)
        janela.title("Seleção de Serviços")
        janela.geometry("700x650")
        
        # Frame principal
        main_frame = ttk.Frame(janela, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        # ========================================
        # NOVA SEÇÃO: ADICIONAR SERVIÇO RÁPIDO
        # ========================================
        add_frame = ttk.LabelFrame(main_frame, text="➕ Adicionar Novo Serviço", padding=10)
        add_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(add_frame, text="Serviço:").pack(side='left', padx=5)
        
        # Combobox com autocompletar
        from src.configuracoes_sistema import GerenciadorConfiguracoes
        
        combo_novo = ttk.Combobox(add_frame, width=40)
        combo_novo.pack(side='left', padx=5, fill='x', expand=True)
        
        def atualizar_combo():
            servicos = GerenciadorConfiguracoes.listar_todos_servicos()
            combo_novo['values'] = servicos
        
        atualizar_combo()
        
        # Autocompletar
        def autocompletar(event):
            valor = combo_novo.get()
            if not valor:
                atualizar_combo()
                return
            
            todos = GerenciadorConfiguracoes.listar_todos_servicos()
            filtrados = [s for s in todos if valor.lower() in s.lower()]
            combo_novo['values'] = filtrados
        
        combo_novo.bind('<KeyRelease>', autocompletar)
        
        def adicionar_novo_servico():
            nome = combo_novo.get().strip()
            if not nome:
                messagebox.showwarning("Aviso", "Digite o nome do serviço!", parent=self.root)
                return
            
            existentes = GerenciadorConfiguracoes.listar_todos_servicos()
            
            if nome not in existentes:
                if messagebox.askyesno("Novo Serviço", 
                                    f"O serviço '{nome}' não existe.\n\nDeseja adicioná-lo?"):
                    if GerenciadorConfiguracoes.adicionar_servico_rapido(nome):
                        messagebox.showinfo("Sucesso", f"Serviço '{nome}' adicionado!", parent=self.root)
                        atualizar_combo()
                        # Recarregar lista abaixo
                        if categoria_nomes:
                            cmb_categoria.current(0)
                            on_categoria_changed(None)
                    else:
                        messagebox.showerror("Erro", "Não foi possível adicionar!", parent=self.root)
                        return
            
            # Adicionar à lista de selecionados
            if nome not in self.servicos_selecionados:
                self.servicos_selecionados.append(nome)
                atualizar_texto_servicos()
                combo_novo.set('')
        
        ttk.Button(add_frame, text="➕ Adicionar", 
                command=adicionar_novo_servico).pack(side='left', padx=5)
        
        # ========================================
        # SEÇÃO ORIGINAL: SELEÇÃO POR CATEGORIA
        # ========================================
        
        ttk.Label(main_frame, text="Ou selecione da lista por categoria:", 
                font=('Arial', 10, 'bold')).pack(pady=5)
        
        # Frame para categoria
        cat_frame = ttk.Frame(main_frame)
        cat_frame.pack(fill='x', pady=5)
        
        ttk.Label(cat_frame, text="Categoria:").pack(side='left', padx=5)
        cmb_categoria = ttk.Combobox(cat_frame, width=40, state='readonly')
        cmb_categoria.pack(side='left', padx=5, fill='x', expand=True)
        
        # Frame para lista de serviços com scroll
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill='both', expand=True, pady=10)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox_servicos = tk.Listbox(list_frame, selectmode='multiple', 
                                    yscrollcommand=scrollbar.set, height=15)
        listbox_servicos.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox_servicos.yview)
        
        # Dicionário para armazenar serviços por categoria
        servicos_por_categoria = {}
        
        # Carregar categorias
        categorias = self.gerador_contrato.listar_categorias_servicos()
        categoria_nomes = [f"{cat['nome']} ({cat['qtd_servicos']} serviços)" for cat in categorias]
        cmb_categoria['values'] = categoria_nomes
        
        # Armazenar mapping de categorias
        for i, cat in enumerate(categorias):
            servicos_por_categoria[categoria_nomes[i]] = {
                'id': cat['id'],
                'servicos': self.gerador_contrato.listar_servicos_categoria(cat['id'])
            }
        
        def on_categoria_changed(event):
            """Atualiza lista de serviços ao mudar categoria"""
            cat_selecionada = cmb_categoria.get()
            if cat_selecionada:
                servicos = servicos_por_categoria[cat_selecionada]['servicos']
                listbox_servicos.delete(0, tk.END)
                for servico in servicos:
                    listbox_servicos.insert(tk.END, servico)
        
        cmb_categoria.bind('<<ComboboxSelected>>', on_categoria_changed)
        
        # Selecionar primeira categoria por padrão
        if categoria_nomes:
            cmb_categoria.current(0)
            on_categoria_changed(None)
        
        # Frame para botões de seleção
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=5)
        
        def selecionar_todos():
            listbox_servicos.select_set(0, tk.END)
        
        def limpar_selecao():
            listbox_servicos.selection_clear(0, tk.END)
        
        ttk.Button(btn_frame, text="Selecionar Todos", 
                command=selecionar_todos).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Limpar Seleção", 
                command=limpar_selecao).pack(side='left', padx=5)
        
        # Label contador
        lbl_contador = ttk.Label(btn_frame, text="0 serviços selecionados")
        lbl_contador.pack(side='right', padx=5)
        
        def atualizar_contador(event=None):
            qtd = len(listbox_servicos.curselection())
            lbl_contador.config(text=f"{qtd} serviços selecionados")
        
        listbox_servicos.bind('<<ListboxSelect>>', atualizar_contador)
        
        # Frame para botões finais
        final_frame = ttk.Frame(main_frame)
        final_frame.pack(fill='x', pady=10)
        
        def atualizar_texto_servicos():
            """Atualiza o texto de serviços no campo principal"""
            descricao = self.gerador_contrato.concatenar_servicos(self.servicos_selecionados)
            self.txt_servicos_selecionados.delete('1.0', tk.END)
            self.txt_servicos_selecionados.insert('1.0', descricao)
        
        def confirmar_selecao():
            """Confirma seleção de serviços"""
            indices = listbox_servicos.curselection()
            servicos_da_lista = [listbox_servicos.get(i) for i in indices]
            
            # Adicionar aos já selecionados
            for servico in servicos_da_lista:
                if servico not in self.servicos_selecionados:
                    self.servicos_selecionados.append(servico)
            
            if not self.servicos_selecionados:
                messagebox.showwarning("Aviso", "Selecione ao menos um serviço!", parent=self.root)
                return
            
            # Atualizar texto
            atualizar_texto_servicos()
            
            logger.info(f"{len(self.servicos_selecionados)} serviços selecionados")
            janela.destroy()
        
        ttk.Button(final_frame, text="✓ Confirmar", 
                command=confirmar_selecao, 
                style='Action.TButton').pack(side='left', padx=5)
        
        ttk.Button(final_frame, text="✗ Cancelar", 
                command=janela.destroy).pack(side='left', padx=5)

    def formatar_valor_global(self, event=None):
        """Formata o valor global para padrão brasileiro"""
        try:
            valor_str = self.ent_valor_global.get().strip()
            if not valor_str or valor_str == "R$ 0,00":
                return
            
            # Remover tudo exceto dígitos e vírgula/ponto
            valor_limpo = valor_str.replace('R$', '').replace('.', '').replace(',', '.').strip()
            
            # Converter para float
            valor_float = float(valor_limpo)
            
            # Formatar como moeda brasileira
            valor_formatado = f"R$ {valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Atualizar campo
            self.ent_valor_global.delete(0, tk.END)
            self.ent_valor_global.insert(0, valor_formatado)
            
        except Exception as e:
            logger.error(f"Erro ao formatar valor: {e}")
    
    def ajustar_data_util(self, data):
        """Ajusta uma data para o próximo dia útil se cair em fim de semana"""
        # Se for sábado (5), avançar 2 dias para segunda
        if data.weekday() == 5:
            data = data + pd.Timedelta(days=2)
            logger.info(f"Data ajustada de sábado para segunda: {data.strftime('%d/%m/%Y')}")
        # Se for domingo (6), avançar 1 dia para segunda
        elif data.weekday() == 6:
            data = data + pd.Timedelta(days=1)
            logger.info(f"Data ajustada de domingo para segunda: {data.strftime('%d/%m/%Y')}")
        return data

    def gerar_contrato_final(self):
        """Gera o contrato final em DOCX"""
        
        # Validar cliente selecionado
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
            return
        
        # Validar fornecedor selecionado (a partir do contexto da aba Fornecedor)
        if not getattr(self, 'cnpj_fornecedor_selecionado', None):
            messagebox.showwarning(
                "Aviso",
                "Nenhum fornecedor selecionado.\n\nVolte à aba Fornecedor e selecione um fornecedor.",
                parent=self.root
            )
            return
        
        # CORREÇÃO MELHORIA 2: Validar serviços - aceita lista OU campo de texto
        servicos_texto = self.txt_servicos_selecionados.get('1.0', tk.END).strip()
        if not self.servicos_selecionados and not servicos_texto:
            messagebox.showwarning("Aviso", "Selecione ao menos um serviço ou digite a descrição!", parent=self.root)
            return
        
        # Validar campos obrigatórios
        if not self.ent_valor_global.get() or self.ent_valor_global.get() == "R$ 0,00":
            messagebox.showwarning("Aviso", "Informe o valor global do contrato!", parent=self.root)
            return
        
        if not self.ent_endereco_obra.get():
            messagebox.showwarning("Aviso", "Informe o endereço da obra!", parent=self.root)
            return
        
        # Validar prazo em dias (sempre obrigatório)
        if not self.ent_prazo_dias.get() or self.ent_prazo_dias.get() == '0':
            messagebox.showwarning("Aviso", "Informe o prazo em dias do contrato!", parent=self.root)
            return
        
        # Se não for condicionado, validar datas específicas
        data_condicionada = self.var_data_condicionada.get()

        if not data_condicionada:
            try:
                data_inicio = self.ent_data_inicio.get_date()
                data_fim = self.ent_data_fim.get_date()
                
                if data_fim <= data_inicio:
                    messagebox.showwarning(
                        "Aviso", 
                        "A data fim deve ser posterior à data de início!", 
                        parent=self.root
                    )
                    return
                    
                # Ajustar data fim se cair em fim de semana (somente modo normal)
                data_fim_ajustada = self.ajustar_data_util(data_fim)
                
                if data_fim_ajustada != data_fim:
                    self.ent_data_fim.set_date(data_fim_ajustada)
                    messagebox.showinfo(
                        "Data Ajustada",
                        f"A data final foi ajustada de {data_fim.strftime('%d/%m/%Y')} "
                        f"para {data_fim_ajustada.strftime('%d/%m/%Y')}\n\n"
                        "Contratos devem terminar em dias úteis.", 
                        parent=self.root
                    )
                    self.calcular_prazo_contrato()
                    
            except Exception as e:
                messagebox.showerror(
                    "Erro", 
                    f"Erro ao validar datas: {str(e)}", 
                    parent=self.root
                )
                return
        else:
            # Modo condicionado: não validar datas mas logar
            logger.info("⚠️ Contrato com data condicionada - datas não validadas")
        
        try:
            # Obter dados do cliente
            dados_cliente = self.gerador_contrato.obter_dados_cliente(self.cliente_atual)
            
            # Obter dados do fornecedor a partir do contexto (aba Fornecedor)
            nome_fornecedor = getattr(self, 'nome_fornecedor_selecionado', None)
            cnpj_forn = getattr(self, 'cnpj_fornecedor_selecionado', None)

            if not nome_fornecedor and not cnpj_forn:
                messagebox.showwarning(
                    "Aviso",
                    "Nenhum fornecedor selecionado.\n\nVolte à aba Fornecedor e selecione um fornecedor.",
                    parent=self.root
                )
                return

            dados_fornecedor = self.gerador_contrato.obter_dados_fornecedor_por_nome(nome_fornecedor)
            
            # Validar se fornecedor foi encontrado
            if not dados_fornecedor:
                messagebox.showerror(
                    "Erro",
                    f"Não foi possível carregar os dados do fornecedor '{nome_fornecedor}'."
                    f"Verifique se o fornecedor está cadastrado corretamente.", 
                    parent=self.root
                )
                return
            
            # Obter dados bancários a partir do contexto
            cnpj_forn_ctx = getattr(self, 'cnpj_fornecedor_selecionado', '') or ''
            dados_bancarios = self.obter_dados_bancarios(cnpj_forn_ctx)
            
            # Ajustar data fim se cair em fim de semana
            data_fim = self.ent_data_fim.get_date()
            data_fim_ajustada = self.ajustar_data_util(data_fim)
            
            # Se a data foi ajustada, atualizar o campo
            if data_fim_ajustada != data_fim:
                self.ent_data_fim.set_date(data_fim_ajustada)
                messagebox.showinfo(
                    "Data Ajustada",
                    f"A data final foi ajustada de {data_fim.strftime('%d/%m/%Y')} "
                    f"para {data_fim_ajustada.strftime('%d/%m/%Y')}\n\n"
                    "Contratos devem terminar em dias úteis.", 
                    parent=self.root
                )
                # Recalcular o número de dias úteis
                self.calcular_prazo_contrato()
            
            # Preparar dados do contrato
            # Preparar dados do contrato
            dados_contrato = {
                'data': self.ent_data_contrato.get_date().strftime('%d/%m/%Y'),  # Data do contrato (SEMPRE)
                'cidade': self.formatar_nome_cidade(dados_cliente['cidade']),
                'cliente_nome': dados_cliente['nome'],
                'cliente_cno': dados_cliente['cno'],
                'cliente_cpf': dados_cliente['cnpj_cpf'],
                'cliente_estado_civil': dados_cliente['estado_civil'],
                'cliente_endereco': dados_cliente['endereco'],
                'fornecedor_nome': dados_fornecedor['nome'],
                'fornecedor_razao_social': dados_fornecedor.get('razao_social') or dados_fornecedor['nome'],
                'fornecedor_cnpj_cpf': dados_fornecedor['cnpj_cpf'],
                'fornecedor_endereco': dados_fornecedor['endereco'],
                'endereco_obra': self.ent_endereco_obra.get(),
                'dias': self.ent_prazo_dias.get() or '0',
                'valor': self.ent_valor_global.get(),
                'multa': self.ent_multa.get(),
                'dados_bancarios': dados_bancarios,
                'data_condicionada': self.var_data_condicionada.get()
            }

            # Adicionar datas de início/fim conforme modo
            if not self.var_data_condicionada.get():
                # Modo normal: datas específicas
                dados_contrato['data_inicio'] = self.ent_data_inicio.get_date().strftime('%d/%m/%Y')
                dados_contrato['data_fim'] = self.ent_data_fim.get_date().strftime('%d/%m/%Y')
            else:
                # Modo condicionado: textos placeholder
                dados_contrato['data_inicio'] = 'A definir conforme recebimento de material'
                dados_contrato['data_fim'] = 'A definir conforme recebimento de material'
            
            # CORREÇÃO MELHORIA 2: Priorizar texto do campo sobre lista
            servicos_texto_campo = self.txt_servicos_selecionados.get('1.0', tk.END).strip()
            
            if servicos_texto_campo:
                # Se há texto no campo (digitado ou selecionado), usar esse
                dados_contrato['descricao'] = servicos_texto_campo
            elif hasattr(self, 'servicos_selecionados') and self.servicos_selecionados:
                # Se não há texto mas tem lista, concatenar lista
                servicos_texto = self.gerador_contrato.concatenar_servicos(self.servicos_selecionados)
                dados_contrato['descricao'] = servicos_texto
            else:
                # Fallback (não deveria chegar aqui por causa da validação)
                dados_contrato['descricao'] = '[SERVIÇOS A DEFINIR]'
            
            # Gerar contrato
            arquivo_gerado = self.gerador_contrato.gerar_contrato(dados_contrato)
            
            if arquivo_gerado:
                messagebox.showinfo(
                    "Sucesso",
                    f"Contrato gerado com sucesso!\n\n"
                    f"Arquivo: {Path(arquivo_gerado).name}\n"
                    f"Local: {self.gerador_contrato.PASTA_CONTRATOS}",
                    parent=self.root
                )
                
                # Perguntar se quer abrir a pasta
                if messagebox.askyesno("Abrir pasta?", "Deseja abrir a pasta de contratos?", parent=self.root):
                    self.abrir_pasta_contratos()
            else:
                messagebox.showerror("Erro", "Falha ao gerar o contrato. Verifique os logs.", parent=self.root)
                
        except Exception as e:
            logger.error(f"Erro ao gerar contrato: {e}")
            messagebox.showerror("Erro", f"Erro ao gerar contrato:\n{str(e)}", parent=self.root)

    def limpar_formulario_contrato(self):
        """Limpa todos os campos do formulário de contrato"""
        try:
            # Resetar datas
            self.ent_data_contrato.set_date(datetime.now())
            self.ent_data_inicio.set_date(datetime.now() + timedelta(days=1))
            self.ent_data_fim.set_date(datetime.now() + timedelta(days=31))

            # Limpar valores monetários
            self.ent_valor_global.delete(0, tk.END)
            self.ent_valor_global.insert(0, "R$ 0,00")
            self.ent_valor_global.bind('<FocusOut>', self.formatar_valor_global)

            self.ent_multa.delete(0, tk.END)
            self.ent_multa.insert(0, "R$ 0,00")

            # Limpar endereço da obra
            self.ent_endereco_obra.delete(0, tk.END)

            # Limpar checkbox de data condicionada
            if hasattr(self, 'var_data_condicionada'):
                self.var_data_condicionada.set(False)

            # Limpar serviços
            self.servicos_selecionados = []
            self.txt_servicos_selecionados.delete('1.0', tk.END)

            # Limpar observações
            if hasattr(self, 'txt_observacoes_contrato'):
                self.txt_observacoes_contrato.delete('1.0', tk.END)

            # Recalcular prazo
            self.calcular_prazo_contrato()

            logger.info("Formulário de contrato limpo")

        except Exception as e:
            logger.error(f"Erro ao limpar formulário: {e}")

    def abrir_pasta_contratos(self):
        """Abre a pasta de contratos no explorador de arquivos"""
        try:
            import subprocess
            import platform
            
            pasta = self.gerador_contrato.PASTA_CONTRATOS
            
            if not pasta.exists():
                pasta.mkdir(parents=True, exist_ok=True)
            
            if platform.system() == 'Windows':
                os.startfile(str(pasta))
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', str(pasta)])
            else:  # Linux
                subprocess.run(['xdg-open', str(pasta)])
                
            logger.info(f"Pasta de contratos aberta: {pasta}")
            
        except Exception as e:
            logger.error(f"Erro ao abrir pasta de contratos: {e}")
            messagebox.showerror("Erro", f"Erro ao abrir pasta: {str(e)}, parent=self.root")

    # Funções da aba Medições
    def carregar_medicoes(self):
        """
        Carrega as medições do contrato selecionado
        VERSÃO ATUALIZADA: Filtra medições excluídas por padrão
        """
        try:
            if not self.contrato_atual:
                messagebox.showwarning("Aviso", "Selecione um contrato primeiro!", parent=self.root)
                return
                
            # Limpar treeview
            for item in self.tree_medicoes.get_children():
                self.tree_medicoes.delete(item)
                
            # Abrir arquivo do cliente
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se a aba de medições existe
            if "Medicoes" not in wb.sheetnames:
                messagebox.showerror("Erro", "Aba de medições não encontrada!", parent=self.root)
                wb.close()
                return
                
            ws = wb["Medicoes"]
            
            # Percorrer as linhas (pulando o cabeçalho)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == self.contrato_atual:  # Filtrar pelo ID do contrato atual
                    # === FILTRO DE EXCLUSÃO ===
                    # Não mostrar medições com status "EXCLUÍDO"
                    status = row[8] if row[8] else ""
                    if status == "EXCLUÍDO":
                        continue  # Pula esta linha
                    
                    # Formatação de dados
                    try:
                        data_medicao = row[4].strftime('%d/%m/%Y') if isinstance(row[4], datetime) else row[4]
                        data_pagamento = row[5].strftime('%d/%m/%Y') if isinstance(row[5], datetime) else row[5]
                        valor = f"R$ {float(row[7]):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if row[7] else "R$ 0,00"
                    except (ValueError, TypeError, AttributeError):
                        data_medicao = str(row[4] or "")
                        data_pagamento = str(row[5] or "")
                        valor = str(row[7] or "R$ 0,00")
                    
                    # Adicionar à treeview
                    self.tree_medicoes.insert('', 'end', values=(
                        row[1],             # ID Medição
                        data_medicao,       # Data Medição
                        data_pagamento,     # Data Pagamento
                        row[6],             # Referência
                        valor,              # Valor
                        row[8]              # Status
                    ))
                
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar medições: {str(e)}", parent=self.root)
    
    def nova_medicao(self):
        """Abre janela para cadastro de nova medição"""
        if not self.contrato_atual:
            messagebox.showwarning("Aviso", "Selecione um contrato primeiro!", parent=self.root)
            return
        
        # Verificar se o contrato tem saldo disponível
        saldo = self.verificar_saldo_contrato()
        if saldo <= 0:
            messagebox.showwarning("Aviso", "Este contrato não possui saldo disponível para novas medições!", parent=self.root)
            return
            
        # Criar janela de cadastro
        janela = self.criar_janela_modal("Nova Medição", largura=700, altura=500)
        
        # Frame principal
        frame = ttk.Frame(janela, padding="10")
        frame.pack(fill='both', expand=True)
        
        # Informações do contrato
        frame_info = ttk.LabelFrame(frame, text="Informações do Contrato")
        frame_info.pack(fill='x', pady=5)
        
        # Obter informações do contrato
        try:
            contrato = self.obter_dados_contrato(self.contrato_atual)
            
            if contrato:
                ttk.Label(frame_info, text=f"ID: {contrato['id']}").grid(row=0, column=0, padx=5, pady=2, sticky='w')
                ttk.Label(frame_info, text=f"Fornecedor: {contrato['nome']}").grid(row=0, column=1, padx=5, pady=2, sticky='w')
                ttk.Label(frame_info, text=f"Valor Global: {formatar_moeda_br(contrato['valor_global'])}").grid(row=1, column=0, padx=5, pady=2, sticky='w')
                ttk.Label(frame_info, text=f"Saldo: {formatar_moeda_br(contrato['saldo'])}").grid(row=1, column=1, padx=5, pady=2, sticky='w')
        except Exception as e:
            ttk.Label(frame_info, text=f"Erro ao carregar dados: {str(e)}").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        
        # Frame para dados da medição
        frame_medicao = ttk.LabelFrame(frame, text="Dados da Medição")
        frame_medicao.pack(fill='x', pady=10)
        
        # Data da Medição
        ttk.Label(frame_medicao, text="Data da Medição:*").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        data_medicao = DateEntry(frame_medicao, width=12, background='darkblue',
                              foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        data_medicao.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        data_medicao.set_date(datetime.now())
        
        # Data de Pagamento
        ttk.Label(frame_medicao, text="Data de Pagamento:*").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        data_pagamento = DateEntry(frame_medicao, width=12, background='darkblue',
                                foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        data_pagamento.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        # Calcular data de pagamento padrão (data da medição)
        # data_pagamento = data_medicao
        
        # Referência
        ttk.Label(frame_medicao, text="Referência:*").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        referencia = ttk.Entry(frame_medicao, width=40)
        referencia.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        
        # Valor
        ttk.Label(frame_medicao, text="Valor (R$):*").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        valor = ttk.Entry(frame_medicao, width=20)
        valor.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        # Saldo Máximo (como referência)
        ttk.Label(frame_medicao, text=f"Saldo disponível: {formatar_moeda_br(contrato['saldo'])}").grid(row=3, column=2, padx=5, sticky='w')

        # Observações
        ttk.Label(frame_medicao, text="Observações:").grid(row=4, column=0, padx=5, pady=5, sticky='ne')
        observacoes = tk.Text(frame_medicao, width=40, height=4)
        observacoes.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky='w')
        
        # Frame para botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, 
                 text="Salvar", 
                 command=lambda: self.salvar_medicao(
                     janela,
                     self.contrato_atual,
                     data_medicao.get(),
                     data_pagamento.get(),
                     referencia.get(),
                     valor.get(),
                     observacoes.get("1.0", "end-1c")
                 )).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes, 
                 text="Cancelar", 
                 command=janela.destroy).pack(side='left', padx=5)
        
    def verificar_saldo_contrato(self):
        """Verifica o saldo disponível do contrato atual"""
        try:
            # Obter dados do contrato
            contrato = self.obter_dados_contrato(self.contrato_atual)
            if not contrato:
                return 0
                
            return float(contrato['saldo'])
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao verificar saldo: {str(e)}", parent=self.root)
            return 0
            
    def obter_dados_contrato(self, id_contrato):
        """Obtém os dados completos de um contrato pelo ID"""
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb["Contratos_Medicao"]
            
            dados_contrato = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == id_contrato:
                    dados_contrato = {
                        'id': row[0],
                        'cnpj': row[1],
                        'nome': row[2],
                        'descricao': row[3],
                        'data_inicio': row[4],
                        'data_final': row[5],
                        'valor_global': row[6],
                        'valor_pago': row[7],
                        'saldo': row[8],
                        'status': row[9],
                        'observacao': row[10]
                    }
                    break
                    
            wb.close()
            return dados_contrato
            
        except Exception as e:
            logger.error(f"Erro ao obter dados do contrato: {str(e)}")
            return None
    
    def salvar_medicao(self, janela, id_contrato, data_medicao, data_pagamento, referencia, valor, observacoes):
        """Salva uma nova medição"""
        try:
            # Validar campos obrigatórios
            if not data_medicao or not data_pagamento or not referencia or not valor:
                messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!", parent=self.root)
                return
                
            # Validar valor
            try:
                valor_float = float(valor.replace(',', '.'))
                if valor_float <= 0:
                    messagebox.showerror("Erro", "Valor deve ser maior que zero!", parent=self.root)
                    return
            except ValueError:
                messagebox.showerror("Erro", "Valor inválido!", parent=self.root)
                return
                
            # Verificar saldo disponível
            saldo = self.verificar_saldo_contrato()
            if valor_float > saldo:
                messagebox.showerror("Erro", f"Valor excede o saldo disponível de R$ {saldo:.2f}!", parent=self.root)
                return
                
            # Converter datas
            try:
                data_med = datetime.strptime(data_medicao, '%d/%m/%Y')
                data_pag = datetime.strptime(data_pagamento, '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Erro", "Data inválida! Use o formato dd/mm/aaaa", parent=self.root)
                return
                
            # Abrir arquivo do cliente
            wb = load_workbook(self.arquivo_cliente)
            ws_medicoes = wb["Medicoes"]
            
            # Gerar ID da medição (próximo número sequencial para o contrato)
            # Medições EXCLUÍDAS são ignoradas — o ID pode ser reutilizado
            next_id = 1
            for row in ws_medicoes.iter_rows(min_row=2, values_only=True):
                status_row = str(row[8] or '').strip().upper()
                if status_row in ('EXCLUÍDO', 'EXCLUIDO'):
                    continue
                if row[0] == id_contrato and row[1] and isinstance(row[1], int) and row[1] >= next_id:
                    next_id = row[1] + 1
            
            # Obter dados do fornecedor
            contrato = self.obter_dados_contrato(id_contrato)
            if not contrato:
                messagebox.showerror("Erro", "Não foi possível obter dados do contrato!", parent=self.root)
                wb.close()
                return
                
            # Adicionar nova medição
            proxima_linha = ws_medicoes.max_row + 1
            ws_medicoes.cell(row=proxima_linha, column=1, value=id_contrato)  # ID_Contrato
            ws_medicoes.cell(row=proxima_linha, column=2, value=next_id)      # ID_Medicao

            # Salvar CNPJ como texto para preservar zeros à esquerda
            cnpj_limpo = ''.join(filter(str.isdigit, str(contrato['cnpj'])))

            # Determinar se é CPF ou CNPJ e garantir formatação adequada
            if len(cnpj_limpo) <= 11:
                # É um CPF
                cnpj_formatado = cnpj_limpo.zfill(11)
                cnpj_formatado = f"{cnpj_formatado[:3]}.{cnpj_formatado[3:6]}.{cnpj_formatado[6:9]}-{cnpj_formatado[9:]}"
            else:
                # É um CNPJ
                cnpj_formatado = cnpj_limpo.zfill(14)
                cnpj_formatado = f"{cnpj_formatado[:2]}.{cnpj_formatado[2:5]}.{cnpj_formatado[5:8]}/{cnpj_formatado[8:12]}-{cnpj_formatado[12:]}"

            # Garantir que o CNPJ seja armazenado como texto na planilha
            cnpj_cell = ws_medicoes.cell(row=proxima_linha, column=3, value=cnpj_formatado)

            ws_medicoes.cell(row=proxima_linha, column=4, value=contrato['nome'])  # Nome_Fornecedor
            ws_medicoes.cell(row=proxima_linha, column=5, value=data_med)     # Data_Medicao
            ws_medicoes.cell(row=proxima_linha, column=6, value=data_pag)     # Data_Pagamento
            ws_medicoes.cell(row=proxima_linha, column=7, value=referencia.upper())  # Referencia
            
            # Formatação para valor
            valor_cell = ws_medicoes.cell(row=proxima_linha, column=8, value=valor_float)  # Valor
            valor_cell.number_format = '#.##0,00'
            
            ws_medicoes.cell(row=proxima_linha, column=9, value="PENDENTE")   # Status
            ws_medicoes.cell(row=proxima_linha, column=10, value=None)        # Data_Lancamento
            ws_medicoes.cell(row=proxima_linha, column=11, value=observacoes.upper())  # Observacao
            
            # Atualizar saldo do contrato na aba Contratos_Medicao
            ws_contratos = wb["Contratos_Medicao"]
            for idx, row in enumerate(ws_contratos.iter_rows(min_row=2, max_col=1, values_only=True), 2):
                if row[0] == id_contrato:
                    # Obter valores atuais
                    valor_global = ws_contratos.cell(row=idx, column=7).value or 0
                    valor_pago = ws_contratos.cell(row=idx, column=8).value or 0
                    
                    # Atualizar valor pago
                    novo_valor_pago = valor_pago + valor_float
                    valor_pago_cell = ws_contratos.cell(row=idx, column=8, value=novo_valor_pago)
                    valor_pago_cell.number_format = '#.##0,00'
                    
                    # Atualizar saldo
                    novo_saldo = valor_global - novo_valor_pago
                    saldo_cell = ws_contratos.cell(row=idx, column=9, value=novo_saldo)
                    saldo_cell.number_format = '#.##0,00'
                    
                    # Se saldo zerou, atualizar status para CONCLUÍDO
                    if novo_saldo <= 0:
                        ws_contratos.cell(row=idx, column=10, value="CONCLUÍDO")
                    
                    break
            
            # Salvar arquivo
            wb.save(self.arquivo_cliente)
            wb.close()
            
            messagebox.showinfo("Sucesso", "Medição cadastrada com sucesso!", parent=self.root)
            janela.destroy()
            
            # Atualizar lista de medições
            self.carregar_medicoes()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar medição: {str(e)}", parent=self.root)
            try:
                wb.close()
            except:
                pass
    
    def editar_medicao(self):
        """Edita a medição selecionada"""
        selecionado = self.tree_medicoes.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione uma medição para editar", parent=self.root)
            return
            
        # Obter ID da medição selecionada
        valores = self.tree_medicoes.item(selecionado)['values']
        id_medicao = valores[0]
        
        # Obter dados completos da medição
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb["Medicoes"]
            
            dados_medicao = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == self.contrato_atual and row[1] == id_medicao:
                    if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                        continue
                    dados_medicao = {
                        'id_contrato': row[0],
                        'id_medicao': row[1],
                        'cnpj': row[2],
                        'nome': row[3],
                        'data_medicao': row[4],
                        'data_pagamento': row[5],
                        'referencia': row[6],
                        'valor': row[7],
                        'status': row[8],
                        'data_lancamento': row[9],
                        'observacao': row[10]
                    }
                    break
                    
            wb.close()
            
            if not dados_medicao:
                messagebox.showerror("Erro", "Medição não encontrada!", parent=self.root)
                return
                
            # Verificar se já foi lançada
            if dados_medicao['status'] != 'PENDENTE':
                messagebox.showwarning("Aviso", "Esta medição já foi lançada e não pode ser editada!", parent=self.root)
                return
                
            # Criar janela de edição
            janela = self.criar_janela_modal("Editar Medição", largura=700, altura=500)
            
            # Frame principal
            frame = ttk.Frame(janela, padding="10")
            frame.pack(fill='both', expand=True)

            # Informações do contrato
            frame_info = ttk.LabelFrame(frame, text="Informações da Medição")
            frame_info.pack(fill='x', pady=5)
            
            ttk.Label(frame_info, text=f"Contrato: {dados_medicao['id_contrato']}").grid(row=0, column=0, padx=5, pady=2, sticky='w')
            ttk.Label(frame_info, text=f"Medição: {dados_medicao['id_medicao']}").grid(row=0, column=1, padx=5, pady=2, sticky='w')
            ttk.Label(frame_info, text=f"Fornecedor: {dados_medicao['nome']}").grid(row=1, column=0, columnspan=2, padx=5, pady=2, sticky='w')
            
            # Frame para dados da medição
            frame_medicao = ttk.LabelFrame(frame, text="Dados da Medição")
            frame_medicao.pack(fill='x', pady=10)
            
            # Data da Medição
            ttk.Label(frame_medicao, text="Data da Medição:*").grid(row=0, column=0, padx=5, pady=5, sticky='e')
            data_medicao = DateEntry(frame_medicao, width=12, background='darkblue',
                                  foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
            data_medicao.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            if isinstance(dados_medicao['data_medicao'], datetime):
                data_medicao.set_date(dados_medicao['data_medicao'])
            
            # Data de Pagamento
            ttk.Label(frame_medicao, text="Data de Pagamento:*").grid(row=1, column=0, padx=5, pady=5, sticky='e')
            data_pagamento = DateEntry(frame_medicao, width=12, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
            data_pagamento.grid(row=1, column=1, padx=5, pady=5, sticky='w')
            if isinstance(dados_medicao['data_pagamento'], datetime):
                data_pagamento.set_date(dados_medicao['data_pagamento'])
            
            # Referência
            ttk.Label(frame_medicao, text="Referência:*").grid(row=2, column=0, padx=5, pady=5, sticky='e')
            referencia = ttk.Entry(frame_medicao, width=80)
            referencia.grid(row=2, column=1, padx=5, pady=5, sticky='w')
            referencia.insert(0, dados_medicao['referencia'] if dados_medicao['referencia'] else "")
            
            # Valor original (somente leitura)
            ttk.Label(frame_medicao, text="Valor Original (R$):").grid(row=3, column=0, padx=5, pady=5, sticky='e')
            valor_original = ttk.Entry(frame_medicao, width=15)
            valor_original.grid(row=3, column=1, padx=5, pady=5, sticky='w')

            # Formatar o valor usando a função antes de inserir no campo
            valor_formatado = formatar_moeda_br(dados_medicao['valor']).replace('R$ ', '')  # Remove o "R$ " para ficar só o número
            valor_original.insert(0, valor_formatado)
            valor_original.config(state='readonly')
            
            # Valor novo
            ttk.Label(frame_medicao, text="Novo Valor (R$):*").grid(row=4, column=0, padx=5, pady=5, sticky='e')
            valor_novo = ttk.Entry(frame_medicao, width=15)
            valor_novo.grid(row=4, column=1, padx=5, pady=5, sticky='w')
            valor_novo.insert(0, str(dados_medicao['valor']))
            
            # Observações
            ttk.Label(frame_medicao, text="Observações:").grid(row=5, column=0, padx=5, pady=5, sticky='ne')
            observacoes = tk.Text(frame_medicao, width=40, height=4)
            observacoes.grid(row=5, column=1, padx=5, pady=5, sticky='w')
            observacoes.insert("1.0", dados_medicao['observacao'] if dados_medicao['observacao'] else "")
            
            # Frame para botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)
            
            ttk.Button(frame_botoes, 
                     text="Salvar", 
                     command=lambda: self.atualizar_medicao(
                         janela,
                         self.contrato_atual,
                         id_medicao,
                         data_medicao.get(),
                         data_pagamento.get(),
                         referencia.get(),
                         valor_original.get(),
                         valor_novo.get(),
                         observacoes.get("1.0", "end-1c")
                     )).pack(side='left', padx=5)
            
            ttk.Button(frame_botoes, 
                     text="Cancelar", 
                     command=janela.destroy).pack(side='left', padx=5)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao editar medição: {str(e)}", parent=self.root)
    
    def atualizar_medicao(self, janela, id_contrato, id_medicao, data_medicao, data_pagamento, 
                    referencia, valor_original, valor_novo, observacoes):
        """Atualiza os dados de uma medição"""
        try:
            # Validar campos obrigatórios
            if not data_medicao or not data_pagamento or not referencia or not valor_novo:
                messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!", parent=self.root)
                return
                
            # Validar e converter valores
            try:
                # Limpar e converter valor original
                valor_original_limpo = str(valor_original).replace('R$', '').strip()
                valor_original_limpo = valor_original_limpo.replace('.', '')
                valor_original_limpo = valor_original_limpo.replace(',', '.')
                valor_org = float(valor_original_limpo)
                
                # Limpar e converter valor novo
                valor_novo_limpo = str(valor_novo).replace('R$', '').strip()
                valor_novo_limpo = valor_novo_limpo.replace('.', '')
                valor_novo_limpo = valor_novo_limpo.replace(',', '.')
                valor_novo_float = float(valor_novo_limpo)
                
                if valor_novo_float <= 0:
                    messagebox.showerror("Erro", "Valor deve ser maior que zero!", parent=self.root)
                    return
            except (ValueError, AttributeError) as e:
                messagebox.showerror("Erro", f"Valor inválido! Detalhes: {str(e)}", parent=self.root)
                return
                
            # Se valor mudou, verificar saldo do contrato
            if valor_novo_float != valor_org:
                diferenca = valor_novo_float - valor_org
                
                if diferenca > 0:  # Apenas se aumentou
                    saldo = self.verificar_saldo_contrato()
                    if diferenca > saldo:
                        messagebox.showerror("Erro", 
                                        f"O aumento de R$ {diferenca:.2f} excede o saldo disponível de R$ {saldo:.2f}!", 
                                        parent=self.root)
                        return
                
            # Converter datas
            try:
                data_med = datetime.strptime(data_medicao, '%d/%m/%Y')
                data_pag = datetime.strptime(data_pagamento, '%d/%m/%Y')
            except ValueError:
                messagebox.showerror("Erro", "Data inválida! Use o formato dd/mm/aaaa", parent=self.root)
                return
                
            # Abrir arquivo do cliente
            wb = load_workbook(self.arquivo_cliente)
            ws_medicoes = wb["Medicoes"]
            
            # Buscar medição pelo ID
            medicao_row = None
            for idx, row in enumerate(ws_medicoes.iter_rows(min_row=2, values_only=True), 2):
                if row[0] == id_contrato and row[1] == id_medicao:
                    if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                        continue
                    medicao_row = idx
                    break
                    
            if not medicao_row:
                messagebox.showerror("Erro", "Medição não encontrada!", parent=self.root)
                wb.close()
                return
                
            # Atualizar dados da medição
            ws_medicoes.cell(row=medicao_row, column=5, value=data_med)  # Data_Medicao
            ws_medicoes.cell(row=medicao_row, column=6, value=data_pag)  # Data_Pagamento
            ws_medicoes.cell(row=medicao_row, column=7, value=referencia.upper())  # Referencia
            
            # Atualizar valor apenas se mudou
            if valor_novo_float != valor_org:
                # Atualizar valor na medição
                valor_cell = ws_medicoes.cell(row=medicao_row, column=8, value=valor_novo_float)
                valor_cell.number_format = '#.##0,00'
                
                # Calcular a diferença
                diferenca = valor_novo_float - valor_org
                
                # Atualizar valores do contrato
                ws_contratos = wb["Contratos_Medicao"]
                contrato_encontrado = False
                
                for idx, row in enumerate(ws_contratos.iter_rows(min_row=2, values_only=True), 2):
                    if row[0] == id_contrato:
                        contrato_encontrado = True
                        
                        valor_col_global = ws_contratos.cell(row=idx, column=7).value  # Coluna G
                        valor_col_pago = ws_contratos.cell(row=idx, column=8).value    # Coluna H
                        
                        # Verificar se os valores são numéricos
                        if isinstance(valor_col_global, datetime):
                            messagebox.showerror("Erro", 
                                            f"A coluna G (Valor Global) contém uma data em vez de um valor numérico!\n"
                                            f"Valor encontrado: {valor_col_global}", 
                                            parent=self.root)
                            wb.close()
                            return
                            
                        if isinstance(valor_col_pago, datetime):
                            messagebox.showerror("Erro", 
                                            f"A coluna H (Valor Pago) contém uma data em vez de um valor numérico!\n"
                                            f"Valor encontrado: {valor_col_pago}", 
                                            parent=self.root)
                            wb.close()
                            return
                        
                        # Converter para float com segurança
                        try:
                            valor_global = float(valor_col_global) if valor_col_global is not None else 0.0
                            valor_pago = float(valor_col_pago) if valor_col_pago is not None else 0.0
                        except (ValueError, TypeError) as e:
                            messagebox.showerror("Erro", 
                                            f"Erro ao converter valores do contrato:\n"
                                            f"Valor Global: {valor_col_global}\n"
                                            f"Valor Pago: {valor_col_pago}\n"
                                            f"Erro: {str(e)}", 
                                            parent=self.root)
                            wb.close()
                            return
                        
                        # Ajustar com a diferença
                        novo_valor_pago = valor_pago + diferenca
                        valor_pago_cell = ws_contratos.cell(row=idx, column=8, value=novo_valor_pago)  # Coluna H
                        valor_pago_cell.number_format = '#.##0,00'
                        
                        # Atualizar saldo
                        novo_saldo = valor_global - novo_valor_pago
                        saldo_cell = ws_contratos.cell(row=idx, column=9, value=novo_saldo)  # Coluna I
                        saldo_cell.number_format = '#.##0,00'
                        
                        # Se saldo zerou ou ficou negativo, atualizar status para CONCLUÍDO
                        if novo_saldo <= 0:
                            ws_contratos.cell(row=idx, column=10, value="CONCLUÍDO")  # Coluna J
                        
                        break

                if not contrato_encontrado:
                    messagebox.showerror("Erro", f"Contrato {id_contrato} não encontrado na planilha!", parent=self.root)
                    wb.close()
                    return
            
            # Atualizar observações
            ws_medicoes.cell(row=medicao_row, column=11, value=observacoes.upper())  # Observacao
            
            # Salvar arquivo
            wb.save(self.arquivo_cliente)
            wb.close()
            
            messagebox.showinfo("Sucesso", "Medição atualizada com sucesso!", parent=self.root)
            janela.destroy()
            
            # Atualizar lista de medições
            self.carregar_medicoes()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar medição: {str(e)}", parent=self.root)
            import traceback
            print(traceback.format_exc())
            try:
                wb.close()
            except:
                pass
    
    def excluir_medicao(self):
        """
        Exclusão lógica da medição selecionada - mantém registro mas marca como excluída
        
        CARACTERÍSTICAS:
        - Não permite exclusão de medições LANÇADAS ou VINCULADAS
        - Verifica vinculações na aba Vinculacoes antes de permitir exclusão
        - Devolve o valor ao saldo do contrato
        - Registra em log de auditoria
        - Mantém histórico completo para rastreabilidade
        """
        try:
            # Verificar se há seleção
            selecionado = self.tree_medicoes.selection()
            if not selecionado:
                messagebox.showwarning("Aviso", "Selecione uma medição para excluir!", parent=self.root)
                return
            
            # Obter dados da medição
            valores = self.tree_medicoes.item(selecionado)['values']
            id_medicao = valores[0]
            status_atual = valores[5]  # Coluna Status
            
            # Verificar se já está excluída
            if status_atual == "EXCLUÍDO":
                messagebox.showinfo("Informação", "Esta medição já está excluída!", parent=self.root)
                return
            
            # Buscar dados completos da medição
            wb = load_workbook(self.arquivo_cliente)
            ws_medicoes = wb["Medicoes"]
            
            dados_medicao = None
            linha_medicao = None
            
            for idx, row in enumerate(ws_medicoes.iter_rows(min_row=2, values_only=True), 2):
                if row[0] == self.contrato_atual and row[1] == id_medicao:
                    if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                        continue
                    dados_medicao = {
                        'id_medicao': row[1],
                        'cnpj': row[2],
                        'nome': row[3],
                        'data_medicao': row[4],
                        'data_pagamento': row[5],
                        'referencia': row[6],
                        'valor': row[7],
                        'status': row[8],
                        'data_lancamento': row[9],
                        'observacao': row[10]
                    }
                    linha_medicao = idx
                    break
            
            if not dados_medicao:
                wb.close()
                messagebox.showerror("Erro", "Medição não encontrada!", parent=self.root)
                return
            
            # === VERIFICAÇÕES CRÍTICAS ===
            
            # 1. Verificar se já foi lançada ou vinculada
            if dados_medicao['status'] in ['LANÇADO', 'VINCULADO']:
                wb.close()
                messagebox.showerror(
                    "Erro - Não Permitido",
                    f"❌ NÃO É POSSÍVEL EXCLUIR!\n\n"
                    f"Status atual: {dados_medicao['status']}\n\n"
                    f"Esta medição já foi {dados_medicao['status'].lower()} e possui "
                    f"impactos financeiros registrados.\n\n"
                    f"Para corrigi-la:\n"
                    f"• Se LANÇADO: localize e corrija na aba 'Dados' do cliente\n"
                    f"• Se VINCULADO: desvincule primeiro ou ajuste manualmente\n\n"
                    f"⚠️ A exclusão só é permitida para medições com status PENDENTE.",
                    parent=self.root
                )
                return
            
            # 2. Verificar se há vinculações (mesmo que status não indique)
            vinculacoes = []
            if "Vinculacoes" in wb.sheetnames:
                ws_vinc = wb["Vinculacoes"]
                for row in ws_vinc.iter_rows(min_row=2, values_only=True):
                    if row[0] == self.contrato_atual and row[1] == id_medicao:
                        vinculacoes.append(row[2])  # Linha do lançamento
            
            if vinculacoes:
                wb.close()
                linhas_str = ", ".join(map(str, vinculacoes))
                messagebox.showerror(
                    "Erro - Vinculação Detectada",
                    f"❌ NÃO É POSSÍVEL EXCLUIR!\n\n"
                    f"Esta medição possui {len(vinculacoes)} vinculação(ões) ativa(s):\n"
                    f"Linhas dos lançamentos: {linhas_str}\n\n"
                    f"É necessário desvincular antes de excluir.\n"
                    f"Contate o suporte se necessário.",
                    parent=self.root
                )
                return
            
            # === MONTAR MENSAGEM DE CONFIRMAÇÃO ===
            
            # Formatar data
            try:
                if isinstance(dados_medicao['data_medicao'], datetime):
                    data_str = dados_medicao['data_medicao'].strftime('%d/%m/%Y')
                else:
                    data_str = str(dados_medicao['data_medicao'])
            except:
                data_str = "N/D"
            
            # Formatar valor
            try:
                valor_float = float(dados_medicao['valor'])
                valor_str = f"R$ {valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except:
                valor_str = str(dados_medicao['valor'])
            
            mensagem_confirmacao = (
                f"⚠️ CONFIRMA A EXCLUSÃO DESTA MEDIÇÃO?\n\n"
                f"📋 DADOS DA MEDIÇÃO:\n"
                f"{'='*50}\n"
                f"Contrato: {self.contrato_atual}\n"
                f"Medição: #{id_medicao}\n"
                f"Fornecedor: {dados_medicao['nome']}\n"
                f"CNPJ/CPF: {dados_medicao['cnpj']}\n"
                f"Data: {data_str}\n"
                f"Valor: {valor_str}\n"
                f"Referência: {dados_medicao['referencia']}\n"
                f"Status atual: {dados_medicao['status']}\n"
                f"{'='*50}\n\n"
                f"📌 O QUE ACONTECERÁ:\n"
                f"• Status mudará para 'EXCLUÍDO'\n"
                f"• Valor será devolvido ao saldo do contrato\n"
                f"• Medição NÃO aparecerá mais nos relatórios\n"
                f"• Registro permanecerá no histórico para auditoria\n"
                f"• Data/hora da exclusão serão registradas\n\n"
                f"⚠️ Esta ação NÃO pode ser desfeita automaticamente!\n\n"
                f"Deseja continuar?"
            )
            
            resposta = messagebox.askyesno(
                "Confirmar Exclusão",
                mensagem_confirmacao,
                parent=self.root
            )
            
            if not resposta:
                wb.close()
                return
            
            # === EXECUTAR EXCLUSÃO LÓGICA ===
            
            try:
                hoje = datetime.now()
                
                # 1. Atualizar status da medição
                ws_medicoes.cell(row=linha_medicao, column=9, value="EXCLUÍDO")
                
                # 2. Registrar data da exclusão (reutilizar coluna Data_Lancamento)
                data_cell = ws_medicoes.cell(row=linha_medicao, column=10, value=hoje)
                data_cell.number_format = 'DD/MM/YYYY HH:MM:SS'
                
                # 3. Adicionar observação sobre exclusão
                obs_atual = ws_medicoes.cell(row=linha_medicao, column=11).value or ""
                nova_obs = f"{obs_atual} [EXCLUÍDO EM {hoje.strftime('%d/%m/%Y %H:%M:%S')}]"
                ws_medicoes.cell(row=linha_medicao, column=11, value=nova_obs)
                
                # 4. Devolver valor ao saldo do contrato
                ws_contratos = wb["Contratos_Medicao"]
                
                for idx, row in enumerate(ws_contratos.iter_rows(min_row=2, values_only=True), 2):
                    if row[0] == self.contrato_atual:
                        # Obter valores atuais
                        valor_global = float(row[6]) if row[6] else 0
                        valor_pago_atual = float(row[7]) if row[7] else 0
                        
                        # Calcular novo valor pago (subtraindo a medição excluída)
                        valor_medicao = float(dados_medicao['valor'])
                        novo_valor_pago = valor_pago_atual - valor_medicao
                        
                        # Atualizar valor pago
                        valor_pago_cell = ws_contratos.cell(row=idx, column=8, value=novo_valor_pago)
                        valor_pago_cell.number_format = '#.##0,00'
                        
                        # Atualizar saldo
                        novo_saldo = valor_global - novo_valor_pago
                        saldo_cell = ws_contratos.cell(row=idx, column=9, value=novo_saldo)
                        saldo_cell.number_format = '#.##0,00'
                        
                        # Se estava CONCLUÍDO e agora tem saldo, voltar para ATIVO
                        if row[9] == "CONCLUÍDO" and novo_saldo > 0:
                            ws_contratos.cell(row=idx, column=10, value="ATIVO")
                        
                        break
                
                # 5. Registrar exclusão em log (se existir aba de auditoria)
                if "Auditoria" not in wb.sheetnames:
                    ws_audit = wb.create_sheet("Auditoria")
                    # Criar cabeçalhos
                    headers = ['Data_Hora', 'Acao', 'ID_Contrato', 'ID_Medicao', 
                            'Valor', 'Usuario', 'Detalhes']
                    for col, header in enumerate(headers, start=1):
                        ws_audit.cell(row=1, column=col, value=header)
                else:
                    ws_audit = wb["Auditoria"]
                
                proxima_linha = ws_audit.max_row + 1
                
                ws_audit.cell(row=proxima_linha, column=1, value=hoje)
                ws_audit.cell(row=proxima_linha, column=2, value="EXCLUSÃO_MEDICAO")
                ws_audit.cell(row=proxima_linha, column=3, value=self.contrato_atual)
                ws_audit.cell(row=proxima_linha, column=4, value=id_medicao)
                
                valor_cell = ws_audit.cell(row=proxima_linha, column=5, value=float(dados_medicao['valor']))
                valor_cell.number_format = '#.##0,00'
                
                # Usuario (se houver sistema de login, senão deixa vazio)
                ws_audit.cell(row=proxima_linha, column=6, value="SISTEMA")
                
                detalhes = (
                    f"Fornecedor: {dados_medicao['nome']} | "
                    f"Referência: {dados_medicao['referencia']} | "
                    f"Status anterior: {dados_medicao['status']}"
                )
                ws_audit.cell(row=proxima_linha, column=7, value=detalhes)
                
                # Salvar tudo
                wb.save(self.arquivo_cliente)
                wb.close()
                
                # Mensagem de sucesso
                messagebox.showinfo(
                    "Exclusão Concluída",
                    f"✅ Medição #{id_medicao} excluída com sucesso!\n\n"
                    f"📊 RESUMO:\n"
                    f"• Status: EXCLUÍDO\n"
                    f"• Valor devolvido ao contrato: {valor_str}\n"
                    f"• Registrado em: {hoje.strftime('%d/%m/%Y %H:%M:%S')}\n"
                    f"• Log de auditoria atualizado\n\n"
                    f"Esta medição não aparecerá mais em relatórios,\n"
                    f"mas permanece no histórico para auditoria.",
                    parent=self.root
                )
                
                # Atualizar lista de medições
                self.carregar_medicoes()
                
                logger.info(f"Medição excluída: Contrato {self.contrato_atual}, Medição #{id_medicao}, Valor: {dados_medicao['valor']}")
                
            except Exception as e:
                wb.close()
                raise Exception(f"Erro ao executar exclusão: {str(e)}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir medição: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()
            try:
                wb.close()
            except:
                pass

    def lancar_medicao(self):
        """Lança a medição selecionada na planilha do cliente"""
        # Verificar se há seleção
        selecionado = self.tree_medicoes.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione uma medição para lançar", parent=self.root)
            return
            
        # Obter ID da medição selecionada
        valores = self.tree_medicoes.item(selecionado)['values']
        id_medicao = valores[0]
        
        # Verificar se já foi lançada
        if valores[5] != "PENDENTE":
            messagebox.showwarning("Aviso", "Esta medição já foi lançada!", parent=self.root)
            return

        # Validar data de pagamento — impedir lançamento acidental de medições com data passada
        data_pag_str = str(valores[2] or valores[1] or '').strip()
        if data_pag_str:
            try:
                data_pag_dt = datetime.strptime(data_pag_str, '%d/%m/%Y').date()
                hoje_dt = datetime.now().date()
                if data_pag_dt < hoje_dt:
                    messagebox.showwarning(
                        "Data anterior a hoje",
                        f"A data de pagamento desta medição ({data_pag_str}) é anterior à data atual "
                        f"({hoje_dt.strftime('%d/%m/%Y')}).\n\n"
                        f"Para lançar uma medição retroativa use 'Vincular a Lançamento', "
                        f"que a associa a um registro já existente na planilha do cliente.",
                        parent=self.root
                    )
                    return
            except ValueError:
                pass

        # Obter dados completos da medição
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb["Medicoes"]
            
            dados_medicao = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == self.contrato_atual and row[1] == id_medicao:
                    if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                        continue
                    dados_medicao = {
                        'id_contrato': row[0],
                        'id_medicao': row[1],
                        'cnpj': row[2],
                        'nome': row[3],
                        'data_medicao': row[4],
                        'data_pagamento': row[5],
                        'referencia': row[6],
                        'valor': row[7],
                        'status': row[8],
                        'data_lancamento': row[9],
                        'observacao': row[10]
                    }
                    break
                    
            if not dados_medicao:
                messagebox.showerror("Erro", "Medição não encontrada!", parent=self.root)
                wb.close()
                return
                
            # Calcular data de relatório
            data_pagamento = dados_medicao['data_medicao']
            dt_vencto = data_pagamento
            
            # Estimar data para relatório
            hoje = datetime.now()
            data_rel = dt_vencto
            
            # Ajustar para dia 5 ou 20 mais próximo
            # if hoje.day < 5:
            #     data_rel = hoje.replace(day=5)
            # elif hoje.day < 20:
            #     data_rel = hoje.replace(day=20)
            # else:
            #     # Próximo mês, dia 5
            #     if hoje.month == 12:
            #         data_rel = hoje.replace(year=hoje.year+1, month=1, day=5)
            #     else:
            #         data_rel = hoje.replace(month=hoje.month+1, day=5)
                    
            # Verificar se já existe dados_para_incluir e criar se não existir
            if not hasattr(self, 'dados_para_incluir'):
                self.dados_para_incluir = []
                
            # Preparar dados do lançamento
            dados_lancamento = {
                'data': data_rel.strftime('%d/%m/%Y'),  # Data do relatório
                'tp_desp': '2',  # Tipo de despesa (serviço/material)
                'cnpj_cpf': dados_medicao['cnpj'],
                'nome': dados_medicao['nome'],
                'categoria': 'SERV',  # Categoria padrão (ajustar conforme necessário)
                'referencia': dados_medicao['referencia'],
                'nf': '',  # NF em branco (pode ser ajustado)
                'vr_unit': str(dados_medicao['valor']),
                'dias': 1,
                'valor': str(dados_medicao['valor']),
                'dt_vencto': dt_vencto.strftime('%d/%m/%Y'),
                'dados_bancarios': self.obter_dados_bancarios(dados_medicao['cnpj']),
                'observacao': f"MEDIÇÃO {id_medicao} - {dados_medicao['observacao'] or ''}",
                'forma_pagamento': 'PIX',  # Forma padrão
                'status': 'ATIVO'  # Status do lançamento
            }
            
            # Adicionar à lista de dados para incluir
            self.dados_para_incluir.append(dados_lancamento)
            
            # Atualizar status da medição na planilha (ignora registros EXCLUÍDO)
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if row[0] == self.contrato_atual and row[1] == id_medicao:
                    if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                        continue
                    ws.cell(row=idx, column=9, value="LANÇADO")  # Status
                    ws.cell(row=idx, column=10, value=hoje)      # Data_Lancamento
                    break
                    
            # Salvar alterações
            wb.save(self.arquivo_cliente)
            wb.close()
            
            # Confirmar ao usuário
            messagebox.showinfo("Sucesso", 
                             f"Medição lançada com sucesso!\n\n"
                             f"Fornecedor: {dados_medicao['nome']}\n"
                             f"Valor: R$ {float(dados_medicao['valor']):.2f}\n"
                             f"Vencimento: {dt_vencto.strftime('%d/%m/%Y')}", 
                             parent=self.root)
            
            # Atualizar lista de medições
            self.carregar_medicoes()
            
            # Perguntar se deseja enviar os dados imediatamente
            if messagebox.askyesno("Enviar Dados", 
                                 "Deseja enviar os dados para a planilha do cliente agora?"):
                self.enviar_dados()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao lançar medição: {str(e)}", parent=self.root)
            try:
                wb.close()
            except:
                pass
    
    def vincular_medicao(self):
        """Vincula uma medição a um lançamento existente na aba Dados"""
        try:
            # Verificar se há medição selecionada
            selecao = self.tree_medicoes.selection()
            if not selecao:
                messagebox.showwarning("Aviso", "Selecione uma medição para vincular!", parent=self.root)
                return
            
            # Obter dados da medição selecionada
            item = self.tree_medicoes.item(selecao[0])
            valores = item['values']
            id_medicao = valores[0]
            
            # Verificar status atual
            try:
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Medicoes']
                
                # Buscar a medição (ignora registros EXCLUÍDO)
                medicao_encontrada = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == self.contrato_atual and row[1] == id_medicao:
                        if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                            continue
                        status_atual = row[8] if row[8] else ""
                        
                        if status_atual in ["LANÇADO", "VINCULADO"]:
                            wb.close()
                            messagebox.showwarning(
                                "Aviso", 
                                f"Esta medição já está {status_atual}!\\n\\n"
                                "Não é possível vincular novamente.", 
                                parent=self.root
                            )
                            return
                        medicao_encontrada = True
                        break
                
                wb.close()
                
                if not medicao_encontrada:
                    messagebox.showerror("Erro", "Medição não encontrada!", parent=self.root)
                    return
                    
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao verificar status: {str(e)}", parent=self.root)
                return
            
            # Buscar dados completos da medição
            dados_medicao = self.obter_dados_medicao(id_medicao)
            if not dados_medicao:
                messagebox.showerror("Erro", "Não foi possível obter dados da medição!", parent=self.root)
                return
            
            # Abrir janela de seleção de lançamento
            self.abrir_janela_selecao_lancamento(id_medicao, dados_medicao)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao vincular medição: {str(e)}", parent=self.root)

    def adicionar_rastreamento_vinculacao(self, id_contrato, id_medicao, linha_lancamento, valor_medicao, observacao=""):
        """
        Adiciona registro de vinculação na aba Vinculacoes para rastreamento.
        Permite múltiplas vinculações ao mesmo lançamento.
        
        Args:
            id_contrato: ID do contrato
            id_medicao: ID da medição
            linha_lancamento: Linha do lançamento na aba Dados
            valor_medicao: Valor da medição vinculada
            observacao: Observação adicional
        """
        try:
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se aba Vinculacoes existe, se não, criar
            if "Vinculacoes" not in wb.sheetnames:
                ws = wb.create_sheet("Vinculacoes")
                # Criar cabeçalhos
                headers = ['ID_Contrato', 'ID_Medicao', 'Linha_Lancamento', 'Data_Vinculacao', 
                        'Valor_Medicao', 'Observacao']
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)
            else:
                ws = wb["Vinculacoes"]
            
            # Adicionar novo registro
            proxima_linha = ws.max_row + 1
            hoje = datetime.now()
            
            ws.cell(row=proxima_linha, column=1, value=id_contrato)
            ws.cell(row=proxima_linha, column=2, value=id_medicao)
            ws.cell(row=proxima_linha, column=3, value=linha_lancamento)
            
            data_cell = ws.cell(row=proxima_linha, column=4, value=hoje)
            data_cell.number_format = 'DD/MM/YYYY HH:MM:SS'
            
            valor_cell = ws.cell(row=proxima_linha, column=5, value=float(valor_medicao))
            valor_cell.number_format = '#.##0,00'
            
            ws.cell(row=proxima_linha, column=6, value=observacao)
            
            wb.save(self.arquivo_cliente)
            wb.close()
            
            logger.info(f"Rastreamento adicionado: Contrato {id_contrato}, Medição {id_medicao}, Linha {linha_lancamento}")
            
        except Exception as e:
            logger.error(f"Erro ao adicionar rastreamento: {str(e)}")
            try:
                wb.close()
            except:
                pass

    def obter_vinculacoes_lancamento(self, linha_lancamento):
        """
        VERSÃO CORRIGIDA - Com conversão correta ao calcular valores.
        """
        try:
            wb = load_workbook(self.arquivo_cliente)
            
            # Verificar se aba existe
            if 'Vinculacoes' not in wb.sheetnames:
                wb.close()
                return []
            
            ws_vinc = wb['Vinculacoes']
            vinculacoes = []
            
            for row in ws_vinc.iter_rows(min_row=2, values_only=True):
                if row[2] == linha_lancamento:  # Coluna C: Linha_Lancamento
                    vinculacoes.append({
                        'id_contrato': row[0],
                        'id_medicao': row[1],
                        'linha_lancamento': row[2],
                        'data_vinculacao': row[3],
                        'valor_medicao': row[4],  # Já vem como número da planilha
                        'observacao': row[5] if len(row) > 5 else ""
                    })
            
            wb.close()
            return vinculacoes
            
        except Exception as e:
            print(f"Erro ao obter vinculações: {e}")
            return []

    def buscar_lancamentos_por_cnpj(self, cnpj_cpf_normalizado):
        """
        Busca lançamentos na aba Dados usando CNPJ/CPF normalizado.
        Mais robusto que busca por nome.
        
        Args:
            cnpj_cpf_normalizado: Dict com 'limpo' e 'formatado'
            
        Returns:
            Lista de tuplas (índice_linha, dados_linha)
        """
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Dados']
            
            lancamentos_encontrados = []
            cnpj_limpo = cnpj_cpf_normalizado['limpo']
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Coluna 3 (índice 2) contém CNPJ/CPF
                cnpj_celula = str(row[2]) if row[2] else ""
                
                # Normalizar CNPJ da célula
                cnpj_celula_norm = self.normalizar_cnpj_cpf(cnpj_celula)
                
                # Comparar apenas os números
                if cnpj_celula_norm['limpo'] == cnpj_limpo:
                    lancamentos_encontrados.append((idx, row))
            
            wb.close()
            return lancamentos_encontrados
            
        except Exception as e:
            logger.error(f"Erro ao buscar lançamentos por CNPJ: {str(e)}")
            try:
                wb.close()
            except:
                pass
            return []
    
    def abrir_janela_selecao_lancamento(self, id_medicao, dados_medicao):
        """
        VERSÃO CORRIGIDA - Interface melhorada com filtro de data e nova coluna de saldo.
        """
        try:
            # Criar janela modal
            janela = tk.Toplevel(self.root)
            janela.title("Vincular a Lançamento Existente")
            self.centralizar_janela(janela, largura=1200, altura=650)
            
            # Frame de informações da medição
            frame_info = ttk.LabelFrame(janela, text="Dados da Medição", padding=10)
            frame_info.pack(fill='x', padx=10, pady=5)
            
            # Normalizar CNPJ para exibição
            cnpj_norm = self.normalizar_cnpj_cpf(dados_medicao['cnpj'])
            
            # Extrair data da medição
            try:
                if isinstance(dados_medicao['data_medicao'], str):
                    data_med_str = dados_medicao['data_medicao']
                else:
                    data_med_str = dados_medicao['data_medicao'].strftime('%d/%m/%Y')
            except:
                data_med_str = "N/D"
            
            info_text = f"""Fornecedor: {dados_medicao['nome']}
    CNPJ/CPF: {cnpj_norm['formatado']} ({cnpj_norm['tipo']})
    Valor: R$ {float(dados_medicao['valor']):,.2f}
    Data Medição: {data_med_str}
    Referência: {dados_medicao['referencia']}"""
            
            ttk.Label(frame_info, text=info_text, justify='left', font=('Arial', 10)).pack()
            
            # Frame de filtros
            frame_filtros = ttk.LabelFrame(janela, text="Filtros de Busca", padding=10)
            frame_filtros.pack(fill='x', padx=10, pady=5)
            
            # Linha 1: Nome e checkbox de data
            ttk.Label(frame_filtros, text="Buscar por Nome:").grid(row=0, column=0, sticky='w', padx=5)
            var_filtro_nome = tk.StringVar(value=dados_medicao['nome'])
            entry_filtro = ttk.Entry(frame_filtros, textvariable=var_filtro_nome, width=40)
            entry_filtro.grid(row=0, column=1, sticky='ew', padx=5)
            
            # NOVO: Filtro por data
            var_filtrar_data = tk.BooleanVar(value=True)
            ttk.Checkbutton(
                frame_filtros, 
                text="Filtrar por data (mesmo mês/ano)", 
                variable=var_filtrar_data
            ).grid(row=0, column=2, sticky='w', padx=10)
            
            # Botões de busca
            btn_buscar = ttk.Button(
                frame_filtros, 
                text="🔍 Buscar",
                command=lambda: self.buscar_lancamentos_existentes(
                    tree_lancamentos, 
                    dados_medicao, 
                    var_filtro_nome.get(),
                    var_filtrar_data.get()
                )
            )
            btn_buscar.grid(row=0, column=3, padx=5)
            
            btn_buscar_cnpj = ttk.Button(
                frame_filtros,
                text="🔍 Só CNPJ",
                command=lambda: self.buscar_lancamentos_existentes(
                    tree_lancamentos,
                    dados_medicao,
                    "",  # Sem filtro de nome
                    var_filtrar_data.get()
                )
            )
            btn_buscar_cnpj.grid(row=0, column=4, padx=5)
            
            frame_filtros.columnconfigure(1, weight=1)
            
            # Frame de informações adicionais
            frame_info_busca = ttk.Frame(janela)
            frame_info_busca.pack(fill='x', padx=10, pady=2)
            
            label_info = ttk.Label(
                frame_info_busca,
                text="💡 Busca por: CNPJ + Data (mesmo mês) + Saldo suficiente. Desmarque 'Filtrar por data' se não encontrar.",
                foreground='#0066cc',
                font=('Arial', 9)
            )
            label_info.pack()
            
            # Frame para lista de lançamentos
            frame_lancamentos = ttk.LabelFrame(janela, text="Lançamentos Encontrados na Aba 'Dados'", padding=5)
            frame_lancamentos.pack(fill='both', expand=True, padx=10, pady=5)
            
            # NOVA ESTRUTURA: Treeview com coluna de Saldo Disponível
            colunas = ('Linha', 'Data', 'Nome', 'CNPJ/CPF', 'Valor Total', 'Saldo Disp.', 'Status', 'Vencimento', 'Referência', 'Obs')
            tree_lancamentos = ttk.Treeview(frame_lancamentos, columns=colunas, show='headings', height=12)
            
            # Configurar colunas
            larguras = {
                'Linha': 60, 
                'Data': 90, 
                'Nome': 220,
                'CNPJ/CPF': 130, 
                'Valor Total': 100,
                'Saldo Disp.': 100,  # NOVA COLUNA
                'Status': 100,       # NOVA COLUNA
                'Vencimento': 90, 
                'Referência': 120, 
                'Obs': 150
            }
            
            for col in colunas:
                tree_lancamentos.heading(col, text=col)
                tree_lancamentos.column(col, width=larguras.get(col, 100))
            
            # Scrollbars
            scrolly = ttk.Scrollbar(frame_lancamentos, orient='vertical', command=tree_lancamentos.yview)
            scrollx = ttk.Scrollbar(frame_lancamentos, orient='horizontal', command=tree_lancamentos.xview)
            tree_lancamentos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
            
            tree_lancamentos.pack(side='left', fill='both', expand=True)
            scrolly.pack(side='right', fill='y')
            scrollx.pack(side='bottom', fill='x')
            
            # Buscar lançamentos automaticamente ao abrir
            self.buscar_lancamentos_existentes(
                tree_lancamentos, 
                dados_medicao, 
                var_filtro_nome.get(), 
                var_filtrar_data.get()
            )
            
            # Frame para botões de ação
            frame_botoes = ttk.Frame(janela)
            frame_botoes.pack(fill='x', padx=10, pady=10)
            
            ttk.Button(
                frame_botoes, 
                text="✓ Vincular Selecionado",
                command=lambda: self.confirmar_vinculacao(
                    janela, id_medicao, tree_lancamentos, dados_medicao
                )
            ).pack(side='left', padx=5)
            
            ttk.Button(
                frame_botoes,
                text="📊 Ver Vinculações",
                command=lambda: self.mostrar_vinculacoes_lancamento(tree_lancamentos)
            ).pack(side='left', padx=5)
            
            ttk.Button(
                frame_botoes, 
                text="✕ Cancelar",
                command=janela.destroy
            ).pack(side='right', padx=5)
            
            # Label de instruções
            ttk.Label(
                janela, 
                text="💡 Selecione o lançamento com saldo suficiente e clique em 'Vincular Selecionado'",
                foreground='#666',
                font=('Arial', 9)
            ).pack(pady=5)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir janela de seleção: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()

    def buscar_lancamentos_existentes(self, tree, dados_medicao, filtro_nome, usar_filtro_data=True):
        """
        VERSÃO CORRIGIDA - Com conversão correta de valores brasileiros.
        """
        try:
            # Limpar treeview
            for item in tree.get_children():
                tree.delete(item)
            
            # Carregar planilha
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Dados']
            
            # Normalizar CNPJ/CPF da medição
            cnpj_medicao_norm = self.normalizar_cnpj_cpf(dados_medicao['cnpj'])
            
            # Valor da medição - CONVERSÃO CORRETA
            try:
                valor_medicao = self.converter_valor_brasileiro_para_float(dados_medicao['valor'])
            except ValueError as e:
                messagebox.showerror("Erro", f"Erro ao converter valor da medição: {str(e)}", parent=self.root)
                wb.close()
                return
            
            # Data da medição (para comparação)
            try:
                if isinstance(dados_medicao['data_medicao'], str):
                    data_medicao = datetime.strptime(dados_medicao['data_medicao'], '%d/%m/%Y')
                else:
                    data_medicao = dados_medicao['data_medicao']
            except:
                data_medicao = None
            
            # Buscar lançamentos
            encontrados = 0
            lancamentos_possiveis = []
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Extrair dados da linha
                data_rel = row[0]
                cnpj_cpf = str(row[2]) if row[2] else ""
                nome = str(row[3]) if row[3] else ""
                referencia = str(row[4]) if row[4] else ""
                valor_lancamento = row[8] if row[8] else 0
                dt_vencto = row[9]
                observacao = str(row[12]) if row[12] else ""
                status = str(row[13]).strip().upper() if row[13] else ""  # ← NOVA LINHA
            
                # === FILTRO DE STATUS - PULAR SE EXCLUÍDO ===
                if status == "EXCLUIDO":
                    continue
                
                # === CRITÉRIO 1: CNPJ/CPF (obrigatório) ===
                cnpj_linha_norm = self.normalizar_cnpj_cpf(cnpj_cpf)
                match_cnpj = (cnpj_linha_norm['limpo'] == cnpj_medicao_norm['limpo']) if cnpj_medicao_norm['limpo'] else False
                
                # === CRITÉRIO 2: Nome (opcional, fallback) ===
                match_nome = False
                if filtro_nome and not match_cnpj:
                    filtro_lower = filtro_nome.lower()
                    nome_lower = nome.lower()
                    palavras_filtro = filtro_lower.split()
                    palavras_nome = nome_lower.split()
                    match_nome = all(
                        any(palavra_filtro in palavra_nome for palavra_nome in palavras_nome)
                        for palavra_filtro in palavras_filtro
                    )
                
                # Se não deu match em CNPJ nem em nome, pular
                if not (match_cnpj or match_nome):
                    continue
                
                # === CRITÉRIO 3: Data (se habilitado) ===
                match_data = True  # Padrão: aceita qualquer data
                if usar_filtro_data and data_medicao and data_rel:
                    try:
                        if isinstance(data_rel, datetime):
                            data_lancamento = data_rel
                        else:
                            data_lancamento = datetime.strptime(str(data_rel), '%d/%m/%Y')
                        
                        # Aceita se for:
                        # 1. Mesma data exata, OU
                        # 2. Mesmo mês e ano
                        match_data = (
                            data_lancamento.date() == data_medicao.date() or
                            (data_lancamento.month == data_medicao.month and 
                            data_lancamento.year == data_medicao.year)
                        )
                    except:
                        match_data = True  # Se não conseguir comparar, aceita
                
                if not match_data:
                    continue
                
                # === CRITÉRIO 4: Valor e Saldo - COM CONVERSÃO CORRETA ===
                try:
                    valor_lancamento_float = self.converter_valor_brasileiro_para_float(valor_lancamento)
                except:
                    continue
                
                # Calcular saldo disponível do lançamento
                vinculacoes_existentes = self.obter_vinculacoes_lancamento(idx)
                valor_ja_vinculado = sum(
                    self.converter_valor_brasileiro_para_float(v['valor_medicao']) 
                    for v in vinculacoes_existentes
                )
                saldo_disponivel = valor_lancamento_float - valor_ja_vinculado
                
                # Debug
                print(f"DEBUG Linha {idx}:")
                print(f"  Valor lançamento: {valor_lancamento} → {valor_lancamento_float}")
                print(f"  Já vinculado: {valor_ja_vinculado}")
                print(f"  Saldo disponível: {saldo_disponivel}")
                print(f"  Valor medição: {valor_medicao}")
                
                # Aceita se:
                # 1. Valor exato: saldo_disponivel == valor_medicao (com margem de R$ 0.01)
                # 2. Saldo suficiente: saldo_disponivel >= valor_medicao
                valor_exato = abs(saldo_disponivel - valor_medicao) <= 0.01
                saldo_suficiente = saldo_disponivel >= valor_medicao
                
                if not (valor_exato or saldo_suficiente):
                    print(f"  REJEITADO: Saldo insuficiente")
                    continue
                
                print(f"  ACEITO!")
                
                # === Preparar dados para exibição ===
                
                # Formatar datas
                data_formatada = data_rel.strftime('%d/%m/%Y') if isinstance(data_rel, datetime) else str(data_rel)
                vencto_formatado = dt_vencto.strftime('%d/%m/%Y') if isinstance(dt_vencto, datetime) else str(dt_vencto)
                
                # Formatar valores - USAR FORMATO BRASILEIRO
                valor_formatado = f"R$ {valor_lancamento_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                saldo_formatado = f"R$ {saldo_disponivel:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                # Indicadores visuais
                tipo_match = "CNPJ" if match_cnpj else "NOME"
                info_vinculacoes = ""
                if vinculacoes_existentes:
                    qtd = len(vinculacoes_existentes)
                    info_vinculacoes = f" [{qtd}V]"
                
                # Indicador de adequação
                if valor_exato:
                    indicador_valor = "🎯 EXATO"
                elif saldo_disponivel >= valor_medicao * 2:
                    indicador_valor = "✅ SALDO OK"
                else:
                    indicador_valor = "⚠️ SALDO JUSTO"
                
                nome_display = f"{nome} ({tipo_match}){info_vinculacoes}"
                
                # Adicionar à lista com score para ordenação
                lancamentos_possiveis.append({
                    'idx': idx,
                    'data': data_formatada,
                    'nome': nome_display,
                    'cnpj': cnpj_cpf,
                    'valor_total': valor_formatado,
                    'saldo_disp': saldo_formatado,
                    'vencimento': vencto_formatado,
                    'referencia': referencia,
                    'observacao': observacao,
                    'score': (
                        100 if valor_exato else 0 +  # Valor exato = prioridade máxima
                        50 if match_cnpj else 0 +     # Match CNPJ > nome
                        20 if match_data else 0        # Match data é bônus
                    ),
                    'indicador': indicador_valor,
                    'saldo_real': saldo_disponivel
                })
            
            # Ordenar por score (melhores matches primeiro)
            lancamentos_possiveis.sort(key=lambda x: x['score'], reverse=True)
            
            # Adicionar ao treeview
            for lanc in lancamentos_possiveis:
                tree.insert('', 'end', values=(
                    lanc['idx'],
                    lanc['data'],
                    lanc['nome'],
                    lanc['cnpj'],
                    lanc['valor_total'],
                    lanc['saldo_disp'],
                    lanc['indicador'],
                    lanc['vencimento'],
                    lanc['referencia'],
                    lanc['observacao']
                ))
                encontrados += 1
            
            wb.close()
            
            # Mensagem de resultado
            if encontrados == 0:
                if usar_filtro_data:
                    # Retry automático sem filtro de data
                    messagebox.showinfo(
                        "Busca",
                        "Nenhum lançamento encontrado no mesmo mês/ano da medição.\n\n"
                        "Ampliando busca automaticamente (sem filtro de data)...",
                        parent=self.root
                    )
                    self.buscar_lancamentos_existentes(
                        tree, dados_medicao, filtro_nome, usar_filtro_data=False
                    )
                    return
                messagebox.showinfo(
                    "Busca",
                    "Nenhum lançamento encontrado com os critérios especificados.\n\n"
                    "Possíveis causas:\n"
                    "• Não há lançamento com saldo suficiente\n"
                    "• CNPJ/CPF não corresponde\n"
                    "• Lançamento já está totalmente vinculado\n\n"
                    "Dicas:\n"
                    "• Verifique se o lançamento existe na aba Dados\n"
                    "• Confirme CNPJ/CPF do fornecedor\n"
                    "• Tente buscar só por CNPJ (botão 'Só CNPJ')",
                    parent=self.root
                )
            else:
                # Mensagem de sucesso com informações
                msg_data = ""
                if usar_filtro_data and data_medicao:
                    msg_data = f"• Data: {data_medicao.strftime('%m/%Y')} (mesmo mês/ano)\n"
                
                messagebox.showinfo(
                    "Busca",
                    f"Encontrados {encontrados} lançamento(s) compatível(is).\n\n"
                    f"Critérios aplicados:\n"
                    f"• CNPJ/CPF: {cnpj_medicao_norm['formatado']}\n"
                    f"• Nome: {filtro_nome or 'Qualquer'}\n"
                    f"{msg_data}"
                    f"• Valor medição: R$ {valor_medicao:,.2f}\n"
                    f"• Lógica: Saldo disponível >= Valor medição\n\n"
                    f"🎯 = Valor exato\n"
                    f"✅ = Saldo OK (>= 2x medição)\n"
                    f"⚠️ = Saldo justo (>= medição)"
                )
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar lançamentos: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()

    def confirmar_vinculacao(self, janela, id_medicao, tree, dados_medicao):
        """
        VERSÃO CORRIGIDA - Com conversão correta de valores brasileiros.
        """
        try:
            # Verificar seleção
            selecao = tree.selection()
            if not selecao:
                messagebox.showwarning("Aviso", "Selecione um lançamento para vincular!", parent=self.root)
                return
            
            # Obter dados do lançamento selecionado
            item = tree.item(selecao[0])
            valores = item['values']
            linha_lancamento = valores[0]
            nome_lancamento_display = valores[2]
            nome_lancamento = nome_lancamento_display.split(' (')[0]
            valor_total_lanc_str = valores[4]  # String formatada
            saldo_disponivel_str = valores[5]  # String formatada
            indicador = valores[6]
            
            # === CONVERSÃO CORRETA DE VALORES ===
            try:
                saldo_disponivel = self.converter_valor_brasileiro_para_float(saldo_disponivel_str)
            except ValueError as e:
                messagebox.showerror(
                    "Erro de Conversão",
                    f"Erro ao converter saldo disponível:\n{saldo_disponivel_str}\n\n{str(e)}", 
                    parent=self.root
                )
                return
            
            try:
                valor_medicao = self.converter_valor_brasileiro_para_float(dados_medicao['valor'])
            except ValueError as e:
                messagebox.showerror(
                    "Erro de Conversão",
                    f"Erro ao converter valor da medição:\n{dados_medicao['valor']}\n\n{str(e)}", 
                    parent=self.root
                )
                return
            
            # Debug: mostrar valores convertidos
            print(f"DEBUG - Conversões:")
            print(f"  Saldo string: {saldo_disponivel_str}")
            print(f"  Saldo float: {saldo_disponivel}")
            print(f"  Medição string: {dados_medicao['valor']}")
            print(f"  Medição float: {valor_medicao}")
            
            # === VALIDAÇÃO CRÍTICA: Verificar saldo disponível ===
            if saldo_disponivel < valor_medicao:
                messagebox.showerror(
                    "Saldo Insuficiente",
                    f"❌ NÃO É POSSÍVEL VINCULAR!\n\n"
                    f"Valor da medição: R$ {valor_medicao:,.2f}\n"
                    f"Saldo disponível: R$ {saldo_disponivel:,.2f}\n"
                    f"Faltam: R$ {(valor_medicao - saldo_disponivel):,.2f}\n\n"
                    f"Este lançamento não tem saldo suficiente para esta medição.\n"
                    f"Procure outro lançamento ou divida a medição."
                )
                return
            
            # Verificar vinculações existentes
            vinculacoes_existentes = self.obter_vinculacoes_lancamento(linha_lancamento)
            
            # Calcular novo saldo após vinculação
            novo_saldo = saldo_disponivel - valor_medicao
            
            # Preparar mensagem
            mensagem_vinculacoes = ""
            if vinculacoes_existentes:
                qtd = len(vinculacoes_existentes)
                valor_total_vinc = sum(
                    self.converter_valor_brasileiro_para_float(v['valor_medicao']) 
                    for v in vinculacoes_existentes
                )
                mensagem_vinculacoes = (
                    f"\n\n⚠️ ATENÇÃO: Este lançamento já possui {qtd} vinculação(ões):\n"
                    f"Total já vinculado: R$ {valor_total_vinc:,.2f}\n"
                    f"Saldo atual: R$ {saldo_disponivel:,.2f}\n\n"
                    "Esta medição será ADICIONADA às vinculações existentes."
                )
            
            # Aviso sobre saldo final
            aviso_saldo = ""
            if novo_saldo <= 0.01:
                aviso_saldo = "\n\n✅ Este lançamento ficará TOTALMENTE VINCULADO (saldo zerado)."
            else:
                aviso_saldo = f"\n\n💰 Saldo restante após vinculação: R$ {novo_saldo:,.2f}"
            
            # Confirmar com usuário
            resposta = messagebox.askyesno(
                "Confirmar Vinculação",
                f"Confirma a vinculação?\n\n"
                f"MEDIÇÃO #{id_medicao}\n"
                f"Fornecedor: {dados_medicao['nome']}\n"
                f"CNPJ/CPF: {dados_medicao['cnpj']}\n"
                f"Valor: R$ {valor_medicao:,.2f}\n\n"
                f"SERÁ VINCULADA AO LANÇAMENTO:\n"
                f"Linha: {linha_lancamento}\n"
                f"Nome: {nome_lancamento}\n"
                f"Valor total: {valor_total_lanc_str}\n"
                f"Saldo disponível: R$ {saldo_disponivel:,.2f}\n"
                f"Status: {indicador}"
                f"{mensagem_vinculacoes}"
                f"{aviso_saldo}\n\n"
                f"Esta ação marcará a medição como 'VINCULADO'."
            )
            
            if not resposta:
                return
            
            # === EXECUTAR VINCULAÇÃO ===
            wb = load_workbook(self.arquivo_cliente)
            ws_medicoes = wb['Medicoes']
            
            # Atualizar status e dados da medição
            hoje = datetime.now()
            
            for idx, row in enumerate(ws_medicoes.iter_rows(min_row=2, values_only=True), 2):
                if row[0] == self.contrato_atual and row[1] == id_medicao:
                    if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                        continue
                    ws_medicoes.cell(row=idx, column=9, value="VINCULADO")  # Status
                    
                    data_cell = ws_medicoes.cell(row=idx, column=10, value=hoje)
                    data_cell.number_format = 'DD/MM/YYYY HH:MM:SS'
                    
                    # Observação com mais informações
                    obs_atual = ws_medicoes.cell(row=idx, column=11).value or ""
                    info_adicional = ""
                    if vinculacoes_existentes:
                        info_adicional = f" (LANÇAMENTO COM {len(vinculacoes_existentes)} VINC. ANTERIOR(ES))"
                    
                    info_saldo = f" SALDO PÓS-VINC: R$ {novo_saldo:.2f}"
                    
                    nova_obs = f"{obs_atual} [VINCULADO À LINHA {linha_lancamento}{info_adicional}{info_saldo}]"
                    ws_medicoes.cell(row=idx, column=11, value=nova_obs)
                    break
            
            # Salvar alterações
            wb.save(self.arquivo_cliente)
            wb.close()
            
            # Adicionar rastreamento da vinculação
            self.adicionar_rastreamento_vinculacao(
                id_contrato=self.contrato_atual,
                id_medicao=id_medicao,
                linha_lancamento=linha_lancamento,
                valor_medicao=valor_medicao,
                observacao=f"Vinculado a {nome_lancamento} - Saldo restante: R$ {novo_saldo:.2f}"
            )
            
            # Mensagem de sucesso
            total_vinculacoes = len(vinculacoes_existentes) + 1
            msg_final = f"Medição #{id_medicao} vinculada com sucesso!\n\n"
            msg_final += f"Status: VINCULADO\n"
            msg_final += f"Linha do lançamento: {linha_lancamento}\n"
            msg_final += f"Total de vinculações neste lançamento: {total_vinculacoes}\n"
            msg_final += f"Saldo restante no lançamento: R$ {novo_saldo:,.2f}\n\n"
            
            if novo_saldo <= 0.01:
                msg_final += "✅ Lançamento TOTALMENTE VINCULADO!\n"
            elif novo_saldo < valor_medicao:
                msg_final += "⚠️ Saldo restante insuficiente para nova medição deste valor.\n"
            else:
                msg_final += "💰 Saldo ainda disponível para mais vinculações.\n"
            
            msg_final += "\n📊 Use o relatório de vinculações para ver detalhes."
            
            messagebox.showinfo("Sucesso", msg_final, parent=self.root)
            
            # Fechar janela e atualizar lista
            janela.destroy()
            self.carregar_medicoes()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao confirmar vinculação: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()
            try:
                wb.close()
            except:
                pass

    def mostrar_vinculacoes_lancamento(self, tree):
        """
        MÉTODO NOVO - Mostra todas as medições vinculadas ao lançamento selecionado.
        """
        try:
            # Verificar seleção
            selecao = tree.selection()
            if not selecao:
                messagebox.showinfo(
                    "Informação",
                    "Selecione um lançamento para ver suas vinculações.", 
                    parent=self.root
                )
                return
            
            # Obter linha do lançamento
            item = tree.item(selecao[0])
            linha_lancamento = item['values'][0]
            
            # Buscar vinculações
            vinculacoes = self.obter_vinculacoes_lancamento(linha_lancamento)
            
            if not vinculacoes:
                messagebox.showinfo(
                    "Vinculações",
                    f"Lançamento da linha {linha_lancamento} não possui vinculações.", 
                    parent=self.root
                )
                return
            
            # Montar mensagem
            valor_total = sum(v['valor_medicao'] for v in vinculacoes)
            
            mensagem = f"📊 VINCULAÇÕES DO LANÇAMENTO (Linha {linha_lancamento})\n"
            mensagem += f"{'='*60}\n\n"
            mensagem += f"Total de medições vinculadas: {len(vinculacoes)}\n"
            mensagem += f"Valor total vinculado: R$ {valor_total:,.2f}\n\n"
            mensagem += "DETALHES:\n"
            mensagem += "-" * 60 + "\n"
            
            for i, v in enumerate(vinculacoes, 1):
                data_vinc = v['data_vinculacao']
                data_str = data_vinc.strftime('%d/%m/%Y %H:%M') if isinstance(data_vinc, datetime) else str(data_vinc)
                
                mensagem += f"\n{i}. Contrato {v['id_contrato']} - Medição #{v['id_medicao']}\n"
                mensagem += f"   Valor: R$ {float(v['valor_medicao']):,.2f}\n"
                mensagem += f"   Data: {data_str}\n"
                if v['observacao']:
                    mensagem += f"   Obs: {v['observacao']}\n"
            
            # Criar janela de detalhes
            janela_detalhes = tk.Toplevel()
            janela_detalhes.title(f"Vinculações - Lançamento Linha {linha_lancamento}")
            self.centralizar_janela(janela_detalhes, 700, 500)
            
            # Text widget com scroll
            frame = ttk.Frame(janela_detalhes, padding=10)
            frame.pack(fill='both', expand=True)
            
            text_widget = tk.Text(frame, wrap='word', font=('Courier', 10))
            scrollbar = ttk.Scrollbar(frame, command=text_widget.yview)
            text_widget.configure(yscrollcommand=scrollbar.set)
            
            text_widget.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            text_widget.insert('1.0', mensagem)
            text_widget.configure(state='disabled')
            
            # Botão fechar
            ttk.Button(
                janela_detalhes,
                text="Fechar",
                command=janela_detalhes.destroy
            ).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao mostrar vinculações: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()

    def obter_dados_medicao(self, id_medicao):
        """Obtém todos os dados de uma medição específica"""
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Medicoes']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == self.contrato_atual and row[1] == id_medicao:
                    if str(row[8] or '').strip().upper() in ('EXCLUÍDO', 'EXCLUIDO'):
                        continue
                    dados = {
                        'id_medicao': row[1],
                        'cnpj': row[2],
                        'nome': row[3],
                        'data_medicao': row[4].strftime('%d/%m/%Y') if isinstance(row[4], datetime) else str(row[4]),
                        'data_pagamento': row[5].strftime('%d/%m/%Y') if row[5] and isinstance(row[5], datetime) else "",
                        'referencia': row[6] or "",
                        'valor': row[7] or 0,
                        'status': row[8] or "",
                        'observacao': row[10] or ""
                    }
                    wb.close()
                    return dados
            
            wb.close()
            return None
            
        except Exception as e:
            logger.error(f"Erro ao obter dados da medição: {str(e)}")
            try:
                wb.close()
            except:
                pass
            return None

    def obter_dados_bancarios(self, cnpj):
        """
        VERSÃO CORRIGIDA - Obtém os dados bancários do fornecedor com tratamento robusto.
        Funciona tanto em VSCode quanto em executável.
        """
        try:
            # Normalizar CNPJ/CPF
            cnpj_normalizado = self.normalizar_cnpj_cpf(cnpj)
            
            if not cnpj_normalizado['limpo']:
                return "DADOS BANCÁRIOS NÃO CADASTRADOS"
            
            # Tentar primeiro com formatação
            dados_bancarios = buscar_dados_bancarios_fornecedor(
                cnpj_normalizado['formatado'], 
                "PIX"
            )
            
            # Se não encontrou, tentar sem formatação
            if not dados_bancarios or dados_bancarios == "DADOS BANCÁRIOS NÃO CADASTRADOS":
                dados_bancarios = buscar_dados_bancarios_fornecedor(
                    cnpj_normalizado['limpo'], 
                    "PIX"
                )
            
            # Se ainda não encontrou, tentar variações
            if not dados_bancarios or dados_bancarios == "DADOS BANCÁRIOS NÃO CADASTRADOS":
                # Tentar com o valor original também
                dados_bancarios = buscar_dados_bancarios_fornecedor(str(cnpj), "PIX")
            
            return dados_bancarios if dados_bancarios else "DADOS BANCÁRIOS NÃO CADASTRADOS"
            
        except Exception as e:
            logger.error(f"Erro ao obter dados bancários: {str(e)}")
            return "DADOS BANCÁRIOS NÃO CADASTRADOS"
    
    def enviar_dados(self):
        """Envia os dados para a planilha do cliente"""
        try:
            if not self.dados_para_incluir:
                messagebox.showwarning("Aviso", "Não há dados para enviar!", parent=self.root)
                return
                
            # Verificar arquivo do cliente
            arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
            
            try:
                workbook = load_workbook(arquivo_cliente)
            except PermissionError:
                messagebox.showerror(
                    "Erro", 
                    f"A planilha '{self.cliente_atual}.xlsx' está aberta!\n\n"
                    "Por favor:\n"
                    "1. Feche a planilha\n"
                    "2. Clique em OK\n"
                    "3. Tente enviar novamente"
                )
                return
            
            sheet = workbook["Dados"]
            
            # Processar registros
            for dados in self.dados_para_incluir:
                proxima_linha = sheet.max_row + 1
                
                # Converter e salvar data de referência
                data_rel = datetime.strptime(dados['data'], '%d/%m/%Y')
                data_cell = sheet.cell(row=proxima_linha, column=1, value=data_rel)
                data_cell.number_format = 'DD/MM/YYYY'

                # Converter tipo de despesa para número
                tp_desp_cell = sheet.cell(row=proxima_linha, column=2, value=int(dados['tp_desp']))
                tp_desp_cell.number_format = '0'

                sheet.cell(row=proxima_linha, column=3, value=dados['cnpj_cpf'])
                sheet.cell(row=proxima_linha, column=4, value=dados['nome'])
                sheet.cell(row=proxima_linha, column=5, value=dados['referencia'])
                sheet.cell(row=proxima_linha, column=6, value=dados['nf'])

                # Valores financeiros
                vr_unit = float(dados['vr_unit'].replace(',', '.'))
                vr_unit_cell = sheet.cell(row=proxima_linha, column=7, value=vr_unit)
                aplicar_formatacao_celula(vr_unit_cell)

                sheet.cell(row=proxima_linha, column=8, value=int(dados.get('dias', 1)))

                valor = float(dados['valor'].replace(',', '.'))
                valor_cell = sheet.cell(row=proxima_linha, column=9, value=valor)
                aplicar_formatacao_celula(valor_cell)

                dt_vencto = datetime.strptime(dados['dt_vencto'], '%d/%m/%Y')
                dt_vencto_cell = sheet.cell(row=proxima_linha, column=10, value=dt_vencto)
                dt_vencto_cell.number_format = 'DD/MM/YYYY'

                sheet.cell(row=proxima_linha, column=11, value=dados['categoria'])
                sheet.cell(row=proxima_linha, column=12, value=dados['dados_bancarios'])
                sheet.cell(row=proxima_linha, column=13, value=dados['observacao'])
                sheet.cell(row=proxima_linha, column=14, value=dados['status'])

            try:
                # Tentar salvar o arquivo
                workbook.save(arquivo_cliente)
                messagebox.showinfo("Sucesso", "Dados salvos com sucesso na planilha do cliente!", parent=self.root)
                    
                # Limpar após salvar
                self.dados_para_incluir.clear()
                
            except PermissionError:
                messagebox.showerror(
                    "Erro", 
                    f"Não foi possível salvar! A planilha '{self.cliente_atual}.xlsx' está aberta.\n\n"
                    "Por favor:\n"
                    "1. Feche a planilha\n"
                    "2. Clique em OK\n"
                    "3. Tente enviar novamente"
                )
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar arquivo: {str(e)}", parent=self.root)
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar dados: {str(e)}", parent=self.root)
    
    # MELHORIA 1: Método para incluir contrato na aba Contratos
    def incluir_contrato_na_aba(self):
        """Salva o contrato da aba Contrato na planilha do cliente"""
        try:
            if not self.cliente_atual:
                messagebox.showwarning("Aviso", "Selecione um cliente primeiro!", parent=self.root)
                return

            nome_fornecedor = getattr(self, 'nome_fornecedor_selecionado', None)
            cnpj_fornecedor = getattr(self, 'cnpj_fornecedor_selecionado', None)

            if not nome_fornecedor or not cnpj_fornecedor:
                messagebox.showwarning("Aviso", "Selecione um fornecedor na aba Fornecedor antes de salvar!", parent=self.root)
                return

            if not self.ent_valor_global.get() or self.ent_valor_global.get() == "R$ 0,00":
                messagebox.showwarning("Aviso", "Informe o valor global do contrato!", parent=self.root)
                return

            descricao = self.txt_servicos_selecionados.get('1.0', tk.END).strip()
            if not descricao:
                messagebox.showwarning("Aviso", "Informe a descrição dos serviços!", parent=self.root)
                return

            data_inicio = self.ent_data_inicio.get_date().strftime('%d/%m/%Y')
            data_final = self.ent_data_fim.get_date().strftime('%d/%m/%Y')
            valor_global = self.ent_valor_global.get()
            valor_limpo = valor_global.replace('R$', '').replace('.', '').replace(',', '.').strip()

            # Observações: campo dedicado, com fallback para descrição
            observacoes = self.txt_observacoes_contrato.get('1.0', tk.END).strip()
            if not observacoes:
                observacoes = f"Contrato de {descricao[:80]}"

            self.salvar_contrato(None, cnpj_fornecedor, nome_fornecedor, descricao,
                                 data_inicio, data_final, valor_limpo, observacoes)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar contrato: {str(e)}", parent=self.root)
            import traceback
            traceback.print_exc()
    
    def voltar_menu(self):
        """Volta ao menu principal"""
        try:
            # Verificar dados não enviados
            if hasattr(self, 'dados_para_incluir') and self.dados_para_incluir:
                if messagebox.askyesno("Aviso", "Existem dados não enviados. Deseja enviá-los antes de sair?", parent=self.root):
                    self.enviar_dados()
            
            # ============================================================
            # LIMPAR BINDS E TIMERS (IMPORTANTE!)
            # ============================================================
            if self.menu_principal:
                try:
                    # Remover binds criados
                    self.menu_principal.unbind('<FocusIn>')
                    self.menu_principal.unbind('<Button-1>')
                    logger.debug("✓ Binds removidos da janela pai")
                except Exception as e:
                    logger.debug(f"⚠️ Erro ao remover binds: {e}")
            
            # Limpar instância ativa
            GestaoMedicoes._instancia_ativa = None
            logger.debug("✓ Instância ativa limpa")
            
            # Fechar a janela atual (isso também para o timer)
            self.root.destroy()
            logger.info("✅ Janela de Gestão de Medições fechada")
            
            # Mostrar janela principal
            if self.menu_principal:
                self.menu_principal.deiconify()
                self.menu_principal.lift()
                self.menu_principal.focus_force()
                logger.info("✅ Menu principal restaurado")
                
        except Exception as e:
            logger.error(f"❌ Erro ao voltar para menu principal: {e}")
            # Mesmo com erro, limpar tudo
            GestaoMedicoes._instancia_ativa = None
            try:
                if self.menu_principal:
                    self.menu_principal.unbind('<FocusIn>')
                    self.menu_principal.unbind('<Button-1>')
            except:
                pass
            try:
                self.root.destroy()
            except:
                pass

class ComboboxServicosSimples(ttk.Combobox):
    """Combobox com autocompletar para serviços"""
    
    def __init__(self, parent, **kwargs):
        kwargs['state'] = 'normal'  # Permite digitação
        super().__init__(parent, **kwargs)
        
        self.atualizar_valores()
        self.bind('<KeyRelease>', self.autocompletar)
        self.bind('<FocusOut>', self.validar_novo)
    
    def atualizar_valores(self):
        """Atualiza lista de serviços"""
        try:
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            servicos = GerenciadorConfiguracoes.listar_todos_servicos()
            self['values'] = servicos
        except Exception as e:
            print(f"Erro ao carregar serviços: {e}")
            self['values'] = []
    
    def autocompletar(self, event):
        """Autocompletar enquanto digita"""
        valor = self.get()
        if not valor:
            self.atualizar_valores()
            return
        
        # Filtrar
        try:
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            todos = GerenciadorConfiguracoes.listar_todos_servicos()
            filtrados = [s for s in todos if valor.lower() in s.lower()]
            self['values'] = filtrados
        except:
            pass
    
    def validar_novo(self, event=None):
        """Valida e oferece adicionar novo serviço"""
        valor = self.get().strip()
        if not valor:
            return
        
        try:
            from src.configuracoes_sistema import GerenciadorConfiguracoes
            existentes = GerenciadorConfiguracoes.listar_todos_servicos()
            
            if valor not in existentes:
                resposta = messagebox.askyesno(
                    "Novo Serviço",
                    f"O serviço '{valor}' não existe.\n\nDeseja adicioná-lo?", 
                    parent=self.root
                )
                
                if resposta:
                    if GerenciadorConfiguracoes.adicionar_servico_rapido(valor):
                        messagebox.showinfo("Sucesso", f"Serviço '{valor}' adicionado!", parent=self.root)
                        self.atualizar_valores()
                    else:
                        messagebox.showerror("Erro", "Não foi possível adicionar!", parent=self.root)
        except Exception as e:
            print(f"Erro na validação: {e}")

def main():
    """Função principal para executar o módulo de forma independente"""
    app = GestaoMedicoes()
    app.root.mainloop()
    
if __name__ == "__main__":
    main()