"""
Configuração do Relatório Quinzenal de Medições para integração com relatorios_interface.py
Versão modificada com:
- Combobox para seleção de clientes (ao invés de seleção de arquivo)
- Data de referência automática (dias 5 e 20)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
import sys
import os
import logging

# Configurar logger
logger = logging.getLogger(__name__)

# Importar DateEntry se disponível
try:
    from tkcalendar import DateEntry
    tem_tkcalendar = True
except ImportError:
    tem_tkcalendar = False

# Importar configurações do sistema
try:
    from src.config.config import ARQUIVO_CLIENTES, PASTA_CLIENTES
    usa_config_sistema = True
except ImportError:
    ARQUIVO_CLIENTES = None
    PASTA_CLIENTES = None
    usa_config_sistema = False

# Importar gerador de PDF
try:
    from gerar_relatorio_quinzenal_pdf import RelatorioQuinzenalPDF
except ImportError:
    try:
        from src.gerar_relatorio_quinzenal_pdf import RelatorioQuinzenalPDF
    except ImportError:
        RelatorioQuinzenalPDF = None


def carregar_clientes():
    """Carrega a lista de clientes ativos do arquivo de clientes"""
    try:
        # Importar bibliotecas necessárias
        import pandas as pd
        from openpyxl import load_workbook
        
        # Caminho para o arquivo de clientes
        try:
            from src.config.config import ARQUIVO_CLIENTES
            logger.info(f"Carregando clientes de: {ARQUIVO_CLIENTES}")
        except ImportError:
            # Caminho padrão se não conseguir importar das configurações
            ARQUIVO_CLIENTES = "dados/clientes.xlsx"
            logger.warning(f"Usando caminho padrão para clientes: {ARQUIVO_CLIENTES}")
        
        # Verificar se o arquivo existe
        if not os.path.exists(ARQUIVO_CLIENTES):
            logger.warning(f"Arquivo de clientes não encontrado: {ARQUIVO_CLIENTES}")
            return ['Todos os Clientes']
        
        # Carregar o arquivo usando pandas
        try:
            # Ler o arquivo Excel
            df = pd.read_excel(ARQUIVO_CLIENTES, sheet_name='Clientes')
            
            # Debug: mostrar as colunas disponíveis
            logger.info(f"Colunas disponíveis: {df.columns.tolist()}")
            
            # Verificar se a coluna E existe (coluna 4 em índice baseado em 0)
            # Ou verificar pelo nome da coluna se existir
            if len(df.columns) >= 5:  # Verifica se tem pelo menos 5 colunas (A-E)
                # Filtrar clientes ativos (coluna E vazia)
                coluna_status = df.columns[4]  # Coluna E (índice 4)
                logger.info(f"Coluna de status: {coluna_status}")
                
                # Considera como vazio: None, NaN, '', etc.
                df_ativos = df[df[coluna_status].isna() | (df[coluna_status] == '')]
                
                # Verificar se a primeira coluna contém os nomes dos clientes
                coluna_nome = df.columns[0]  # Coluna A
                logger.info(f"Coluna de nome: {coluna_nome}")
                
                # Extrair nomes dos clientes ativos (assumindo que estão na primeira coluna)
                clientes_ativos = df_ativos[coluna_nome].dropna().tolist()
                
                logger.info(f"Total de clientes ativos encontrados: {len(clientes_ativos)}")
                
                # Ordenar alfabeticamente
                clientes_ativos.sort()
                
                # Adicionar "Todos os Clientes" no início
                clientes = ['Todos os Clientes'] + clientes_ativos
                
                return clientes
            else:
                logger.warning("Arquivo não tem colunas suficientes (precisa de pelo menos 5 colunas - A até E)")
                return ['Todos os Clientes']
            
        except Exception as e:
            logger.error(f"Erro ao ler arquivo Excel com pandas: {str(e)}")
            # Tentar com openpyxl como fallback
            try:
                workbook = load_workbook(ARQUIVO_CLIENTES)
                sheet = workbook['Clientes']
                
                clientes = ['Todos os Clientes']
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Verifica se a coluna E (índice 4) está vazia
                    if row[0] and (len(row) < 5 or not row[4]):
                        clientes.append(row[0])
                
                workbook.close()
                clientes.sort()  # Ordenar alfabeticamente (mantendo "Todos os Clientes" primeiro)
                return clientes
                
            except Exception as inner_e:
                logger.error(f"Erro ao ler arquivo Excel com openpyxl: {str(inner_e)}")
                return ['Todos os Clientes']
            
    except Exception as e:
        logger.error(f"Erro ao carregar clientes: {str(e)}", exc_info=True)
        return ['Todos os Clientes']


def calcular_data_rel_automatica():
    """Calcula automaticamente a data do relatório baseado na regra dos dias 5 e 20"""
    try:
        hoje = datetime.now()
        
        if 6 <= hoje.day <= 20:
            # Entre dia 6 e 20: relatório do dia 20 do mês atual
            data_rel = hoje.replace(day=20)
        else:
            if hoje.day > 20:
                # Após dia 20: relatório do dia 5 do próximo mês
                data_rel = (hoje + relativedelta(months=1)).replace(day=5)
            else:
                # Antes do dia 6: relatório do dia 5 do mês atual
                data_rel = hoje.replace(day=5)
        
        logger.info(f"Data calculada automaticamente: {data_rel.strftime('%d/%m/%Y')}")
        return data_rel
        
    except Exception as e:
        logger.error(f"Erro ao calcular data automática: {str(e)}")
        # Fallback: retorna data atual
        return datetime.now()


def obter_caminho_arquivo_cliente(nome_cliente):
    """Obtém o caminho do arquivo Excel do cliente baseado no nome"""
    try:
        if not usa_config_sistema or not PASTA_CLIENTES:
            return None
        
        # Normalizar o nome do cliente para o formato do arquivo
        # Geralmente: nome_cliente.xlsx ou NOME_CLIENTE.xlsx
        nome_normalizado = nome_cliente.replace(' ', '_')
        
        # Tentar diferentes variações do nome
        variações = [
            f"{nome_normalizado}.xlsx",
            f"{nome_normalizado.upper()}.xlsx",
            f"{nome_normalizado.lower()}.xlsx",
        ]
        
        for variacao in variações:
            caminho = Path(PASTA_CLIENTES) / variacao
            if caminho.exists():
                logger.info(f"Arquivo encontrado: {caminho}")
                return str(caminho)
        
        # Se não encontrou, procurar na pasta
        pasta = Path(PASTA_CLIENTES)
        if pasta.exists():
            for arquivo in pasta.glob("*.xlsx"):
                # Comparar sem considerar case e underscores
                nome_arquivo = arquivo.stem.replace('_', ' ').lower()
                nome_busca = nome_cliente.replace('_', ' ').lower()
                if nome_arquivo == nome_busca:
                    logger.info(f"Arquivo encontrado por busca: {arquivo}")
                    return str(arquivo)
        
        logger.warning(f"Arquivo não encontrado para cliente: {nome_cliente}")
        return None
        
    except Exception as e:
        logger.error(f"Erro ao obter caminho do arquivo do cliente: {str(e)}")
        return None


def configurar_relatorio_quinzenal(parent_frame, sistema_relatorios):
    """
    Cria a interface de configuração do relatório quinzenal no painel direito
    
    Args:
        parent_frame: Frame onde as configurações serão adicionadas
        sistema_relatorios: Instância do SistemaRelatorios para acesso a métodos
    """
    
    # Limpar frame anterior
    for widget in parent_frame.winfo_children():
        widget.destroy()
    
    # Variáveis de controle
    cliente_selecionado_var = tk.StringVar()
    arquivo_clientes_var = tk.StringVar()
    
    # Pré-carregar Clientes.xlsx se disponível
    if usa_config_sistema and ARQUIVO_CLIENTES and ARQUIVO_CLIENTES.exists():
        arquivo_clientes_var.set(str(ARQUIVO_CLIENTES))
    
    # ========== TÍTULO ==========
    titulo_frame = ttk.Frame(parent_frame)
    titulo_frame.pack(fill='x', pady=(0, 15))
    
    ttk.Label(
        titulo_frame,
        text="Relatório Quinzenal de Medições (PDF)",
        font=('Arial', 14, 'bold')
    ).pack(anchor='w')
    
    ttk.Label(
        titulo_frame,
        text="Relatório PDF de medições da quinzena (dias 5 e 20)",
        font=('Arial', 9),
        foreground='gray'
    ).pack(anchor='w')
    
    # ========== DESCRIÇÃO ==========
    desc_frame = ttk.LabelFrame(parent_frame, text="ℹ️ Sobre este Relatório", padding=10)
    desc_frame.pack(fill='x', pady=(0, 15))
    
    desc_text = """✓ Contratos com medições na quinzena (dias 5 e 20)
✓ Histórico completo de medições por contrato
✓ Medições da quinzena destacadas em amarelo
✓ Resumo financeiro completo"""
    
    ttk.Label(desc_frame, text=desc_text, justify='left', font=('Arial', 9)).pack(anchor='w')
    
    # ========== SELEÇÃO DE CLIENTE ==========
    cliente_frame = ttk.LabelFrame(parent_frame, text="👤 Seleção de Cliente", padding=10)
    cliente_frame.pack(fill='x', pady=(0, 15))
    
    ttk.Label(
        cliente_frame,
        text="Selecione o cliente:",
        font=('Arial', 10, 'bold')
    ).pack(anchor='w', pady=(0, 5))
    
    # Carregar lista de clientes
    lista_clientes = carregar_clientes()
    
    # Combobox de clientes
    cliente_combobox = ttk.Combobox(
        cliente_frame,
        textvariable=cliente_selecionado_var,
        values=lista_clientes,
        state='readonly',
        width=40,
        font=('Arial', 10)
    )
    cliente_combobox.pack(anchor='w', pady=(0, 5))
    
    # Selecionar "Todos os Clientes" por padrão
    if lista_clientes:
        cliente_combobox.current(0)
    
    # Informação sobre a pasta
    if usa_config_sistema and PASTA_CLIENTES:
        ttk.Label(
            cliente_frame,
            text=f"📂 Pasta de clientes: {PASTA_CLIENTES}",
            font=('Arial', 8),
            foreground='blue'
        ).pack(anchor='w', pady=(5, 0))
    
    # Botão para atualizar lista de clientes
    def atualizar_clientes():
        nova_lista = carregar_clientes()
        cliente_combobox['values'] = nova_lista
        if nova_lista:
            cliente_combobox.current(0)
        messagebox.showinfo("Info", f"Lista atualizada com {len(nova_lista)} clientes")
    
    ttk.Button(
        cliente_frame,
        text="🔄 Atualizar Lista",
        command=atualizar_clientes
    ).pack(anchor='w', pady=(8, 0))
    
    # ========== ARQUIVO CLIENTES.XLSX (Oculto, mas mantido) ==========
    # Mantido para compatibilidade com o código de geração
    
    # ========== DATA DE REFERÊNCIA AUTOMÁTICA ==========
    data_frame = ttk.LabelFrame(parent_frame, text="📅 Data de Referência", padding=10)
    data_frame.pack(fill='x', pady=(0, 15))
    
    # Calcular data automática
    data_automatica = calcular_data_rel_automatica()
    
    ttk.Label(
        data_frame,
        text="Data do relatório (calculada automaticamente):",
        font=('Arial', 10)
    ).pack(anchor='w', pady=(0, 5))
    
    # Frame para exibir a data
    frame_data_display = ttk.Frame(data_frame)
    frame_data_display.pack(anchor='w', pady=(0, 10))
    
    # Label com a data calculada
    data_label = ttk.Label(
        frame_data_display,
        text=data_automatica.strftime('%d/%m/%Y'),
        font=('Arial', 12, 'bold'),
        foreground='#006600'
    )
    data_label.pack(side='left', padx=(0, 10))
    
    # Ícone de calendário
    ttk.Label(
        frame_data_display,
        text="📅",
        font=('Arial', 14)
    ).pack(side='left')
    
    # Explicação da lógica
#     info_logica = ttk.Frame(data_frame)
#     info_logica.pack(anchor='w', fill='x')
    
#     ttk.Label(
#         info_logica,
#         text="ℹ️ Lógica de cálculo:",
#         font=('Arial', 9, 'bold'),
#         foreground='#0066CC'
#     ).pack(anchor='w')
    
#     texto_logica = """• Dias 6 a 20: Relatório do dia 20 do mês atual
# • Dias 21 a 31: Relatório do dia 5 do próximo mês
# • Dias 1 a 5: Relatório do dia 5 do mês atual"""
    
#     ttk.Label(
#         info_logica,
#         text=texto_logica,
#         font=('Arial', 8),
#         foreground='#666666',
#         justify='left'
#     ).pack(anchor='w', pady=(3, 8))
    
    # Opção para alterar data manualmente (se necessário)
    var_alterar_data = tk.BooleanVar(value=False)
    
    def toggle_alteracao_data():
        if var_alterar_data.get():
            frame_data_manual.pack(anchor='w', pady=(5, 0))
        else:
            frame_data_manual.pack_forget()
            data_label.config(text=data_automatica.strftime('%d/%m/%Y'))
    
    check_alterar = ttk.Checkbutton(
        data_frame,
        text="Alterar data manualmente",
        variable=var_alterar_data,
        command=toggle_alteracao_data
    )
    check_alterar.pack(anchor='w')
    
    # Frame para alteração manual (inicialmente oculto)
    frame_data_manual = ttk.Frame(data_frame)
    
    ttk.Label(
        frame_data_manual,
        text="Nova data:",
        font=('Arial', 9)
    ).pack(side='left', padx=(0, 5))
    
    if tem_tkcalendar:
        data_entry = DateEntry(
            frame_data_manual,
            width=12,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            locale='pt_BR',
            font=('Arial', 10)
        )
        data_entry.pack(side='left', padx=(0, 5))
        data_entry.set_date(data_automatica)
    else:
        data_entry = ttk.Entry(frame_data_manual, width=12, font=('Arial', 10))
        data_entry.pack(side='left', padx=(0, 5))
        data_entry.insert(0, data_automatica.strftime('%d/%m/%Y'))
    
    def aplicar_data_manual():
        try:
            if tem_tkcalendar:
                nova_data = data_entry.get_date()
            else:
                data_str = data_entry.get()
                nova_data = datetime.strptime(data_str, '%d/%m/%Y')
            
            data_label.config(text=nova_data.strftime('%d/%m/%Y'))
            messagebox.showinfo("Sucesso", f"Data alterada para {nova_data.strftime('%d/%m/%Y')}")
        except Exception as e:
            messagebox.showerror("Erro", f"Data inválida: {str(e)}")
    
    ttk.Button(
        frame_data_manual,
        text="✓ Aplicar",
        command=aplicar_data_manual
    ).pack(side='left')
    
    # Informação sobre quinzenas
    ttk.Separator(data_frame, orient='horizontal').pack(fill='x', pady=(10, 8))
    
    info_quinzena = ttk.Frame(data_frame)
    info_quinzena.pack(anchor='w', fill='x')
    
    ttk.Label(
        info_quinzena,
        text="📊 Quinzenas:",
        font=('Arial', 9, 'bold')
    ).pack(anchor='w')
    
    ttk.Label(
        info_quinzena,
        text="1ª Quinzena: dia 21 ao dia 5  |  2ª Quinzena: dia 6 ao dia 20",
        font=('Arial', 8),
        foreground='#666666'
    ).pack(anchor='w', pady=(2, 0))
    
    # ========== SEPARADOR ==========
    ttk.Separator(parent_frame, orient='horizontal').pack(fill='x', pady=15)
    
    # ========== BOTÃO GERAR (DESTACADO) ==========
    btn_frame = ttk.Frame(parent_frame)
    btn_frame.pack(fill='x', pady=(5, 15))
    
    # Instrução
    ttk.Label(
        btn_frame,
        text="📌 Pronto para gerar! Clique no botão abaixo:",
        font=('Arial', 9, 'bold'),
        foreground='#006600'
    ).pack(anchor='w', pady=(0, 8))
    
    def gerar_relatorio():
        """Gera o relatório quinzenal"""
        # Validar seleção de cliente
        cliente = cliente_selecionado_var.get()
        
        if not cliente or cliente == "Todos os Clientes":
            messagebox.showwarning(
                "Aviso", 
                "Por favor, selecione um cliente específico.\n\n"
                "A opção 'Todos os Clientes' não está disponível para este relatório."
            )
            return
        
        # Validar arquivo Clientes.xlsx
        if not arquivo_clientes_var.get():
            if usa_config_sistema and ARQUIVO_CLIENTES and ARQUIVO_CLIENTES.exists():
                arquivo_clientes_var.set(str(ARQUIVO_CLIENTES))
            else:
                messagebox.showwarning(
                    "Aviso", 
                    "Arquivo Clientes.xlsx não encontrado.\n\n"
                    "Verifique a configuração do sistema."
                )
                return
        
        # Obter caminho do arquivo do cliente
        arquivo_cliente = obter_caminho_arquivo_cliente(cliente)
        
        if not arquivo_cliente:
            # Se não encontrou automaticamente, perguntar ao usuário
            resposta = messagebox.askyesno(
                "Arquivo não encontrado",
                f"O arquivo para o cliente '{cliente}' não foi encontrado automaticamente.\n\n"
                f"Deseja selecionar o arquivo manualmente?"
            )
            
            if resposta:
                pasta_inicial = str(PASTA_CLIENTES) if usa_config_sistema and PASTA_CLIENTES and PASTA_CLIENTES.exists() else None
                
                arquivo_cliente = filedialog.askopenfilename(
                    title=f"Selecionar Arquivo do Cliente: {cliente}",
                    initialdir=pasta_inicial,
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                )
                
                if not arquivo_cliente:
                    return
            else:
                return
        
        try:
            # Obter data de referência
            if var_alterar_data.get():
                # Data foi alterada manualmente
                if tem_tkcalendar:
                    data_ref = data_entry.get_date()
                else:
                    try:
                        data_str = data_entry.get()
                        data_ref = datetime.strptime(data_str, '%d/%m/%Y')
                    except ValueError:
                        messagebox.showerror("Erro", "Data inválida! Use o formato DD/MM/AAAA")
                        return
            else:
                # Usar data automática
                data_ref = data_automatica
            
            # === PASSO 1: VERIFICAR SE HÁ DADOS ANTES DE PEDIR ARQUIVO ===
            # Criar janela de verificação
            verificacao = tk.Toplevel(sistema_relatorios.root)
            verificacao.title("Verificando dados")
            verificacao.geometry("350x100")
            verificacao.transient(sistema_relatorios.root)
            verificacao.grab_set()
            
            # Centralizar
            verificacao.update_idletasks()
            x = (verificacao.winfo_screenwidth() // 2) - 175
            y = (verificacao.winfo_screenheight() // 2) - 50
            verificacao.geometry(f"+{x}+{y}")
            
            frame_verif = ttk.Frame(verificacao, padding=20)
            frame_verif.pack(fill='both', expand=True)
            
            ttk.Label(frame_verif, text="Verificando medições...", font=('Arial', 11)).pack(pady=10)
            
            progress_bar_verif = ttk.Progressbar(frame_verif, mode='indeterminate', length=250)
            progress_bar_verif.pack()
            progress_bar_verif.start()
            
            verificacao.update()
            
            # Criar gerador e verificar dados
            if RelatorioQuinzenalPDF is None:
                raise ImportError("Módulo gerar_relatorio_quinzenal_pdf não encontrado")
            
            gerador = RelatorioQuinzenalPDF(
                arquivo_cliente,
                arquivo_clientes_var.get()
            )
            
            # Carregar dados do cliente
            gerador.carregar_dados_cliente()
            
            # Verificar se há medições na quinzena
            contratos_encontrados = gerador.filtrar_medicoes_quinzena(data_ref)
            
            progress_bar_verif.stop()
            verificacao.destroy()
            
            # Se não encontrou medições, avisar e parar
            if not contratos_encontrados:
                data_inicio, data_fim = gerador.identificar_quinzena(data_ref)
                messagebox.showinfo(
                    "Aviso",
                    f"Nenhuma medição foi encontrada na quinzena especificada.\n\n"
                    f"Cliente: {cliente}\n"
                    f"Período verificado:\n"
                    f"  • De: {data_inicio.strftime('%d/%m/%Y')}\n"
                    f"  • Até: {data_fim.strftime('%d/%m/%Y')}\n\n"
                    f"Verifique a data de referência."
                )
                return
            
            # === PASSO 2: HÁ DADOS! AGORA PODE PEDIR ARQUIVO DE SAÍDA ===
            # Sugerir nome do arquivo
            nome_cliente_arquivo = cliente.replace(' ', '_').upper()
            nome_sugerido = f"REL_MEDICOES_{nome_cliente_arquivo}_{data_ref.strftime('%d-%m-%Y')}.pdf"
            
            # Solicitar local de salvamento
            arquivo_saida = filedialog.asksaveasfilename(
                title="Salvar Relatório PDF",
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=nome_sugerido
            )
            
            if not arquivo_saida:
                return
            
            # === PASSO 3: GERAR O PDF ===
            # Criar janela de progresso
            progress = tk.Toplevel(sistema_relatorios.root)
            progress.title("Gerando PDF")
            progress.geometry("400x120")
            progress.transient(sistema_relatorios.root)
            progress.grab_set()
            
            # Centralizar
            progress.update_idletasks()
            x = (progress.winfo_screenwidth() // 2) - 200
            y = (progress.winfo_screenheight() // 2) - 60
            progress.geometry(f"+{x}+{y}")
            
            frame_prog = ttk.Frame(progress, padding=20)
            frame_prog.pack(fill='both', expand=True)
            
            ttk.Label(frame_prog, text="Gerando relatório...", font=('Arial', 11)).pack(pady=10)
            
            progress_bar = ttk.Progressbar(frame_prog, mode='indeterminate', length=300)
            progress_bar.pack()
            progress_bar.start()
            
            progress.update()
            
            # Gerar PDF (já temos os dados carregados no gerador)
            resultado = gerador.gerar_pdf(data_ref, arquivo_saida)
            
            progress_bar.stop()
            progress.destroy()
            
            # === PASSO 4: MOSTRAR RESULTADO ===
            if resultado:
                # Perguntar se quer abrir
                total_contratos = len(gerador.contratos_quinzena) if hasattr(gerador, 'contratos_quinzena') else 0
                
                resposta = messagebox.askyesno(
                    "Sucesso!",
                    f"✅ Relatório gerado com sucesso!\n\n"
                    f"Cliente: {cliente}\n"
                    f"Arquivo: {Path(resultado).name}\n"
                    f"Total de contratos: {total_contratos}\n\n"
                    f"Deseja abrir o PDF?"
                )
                
                if resposta:
                    sistema_relatorios.abrir_arquivo(resultado)
            else:
                # Não deveria chegar aqui, pois já verificamos antes
                messagebox.showwarning(
                    "Aviso",
                    "Não foi possível gerar o relatório."
                )
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    # Botão grande e destacado
    btn_gerar = ttk.Button(
        btn_frame,
        text="🚀 Gerar Relatório PDF",
        command=gerar_relatorio,
        style='Accentuated.TButton'
    )
    btn_gerar.pack(fill='x', ipady=8)
    
    # ========== DICA FINAL ==========
    dica_frame = ttk.Frame(parent_frame, relief='solid', borderwidth=1)
    dica_frame.pack(fill='x', pady=(10, 0))
    
    # Fundo cinza claro
    dica_frame.configure(style='Info.TFrame')
    
    dica_content = ttk.Frame(dica_frame, padding=10)
    dica_content.pack(fill='x')
    
    ttk.Label(
        dica_content,
        text="💡 Dica",
        font=('Arial', 9, 'bold'),
        foreground='#666666'
    ).pack(anchor='w')
    
    ttk.Label(
        dica_content,
        text="O relatório mostra o histórico completo de cada contrato,\ncom destaque visual para as medições da quinzena atual.",
        font=('Arial', 8),
        foreground='#666666',
        justify='left'
    ).pack(anchor='w', pady=(3, 0))


# Para compatibilidade com o sistema de relatórios
class ConfiguracaoRelatorioQuinzenal:
    """Classe wrapper para manter compatibilidade"""
    
    def __init__(self, parent_frame, sistema_relatorios):
        configurar_relatorio_quinzenal(parent_frame, sistema_relatorios)