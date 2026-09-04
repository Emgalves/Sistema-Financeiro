import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.ticker as mticker

# Adicionar diretório raiz ao path para importar módulos corretamente
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        PASTA_CLIENTES,
        BASE_PATH
    )
    print("Configurações importadas com sucesso")
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Definir valores padrão em caso de falha
    BASE_PATH = Path(".")
    ARQUIVO_CLIENTES = BASE_PATH / "dados" / "clientes.xlsx"
    PASTA_CLIENTES = BASE_PATH / "dados" / "clientes"

# Caminho do logo (deve estar na mesma pasta do script ou configurado)
LOGO_PATH = Path(__file__).parent / "logo3.png"
if not LOGO_PATH.exists():
    # Tentar na pasta de saída
    LOGO_PATH = BASE_PATH / "outputs" / "logo3.png"

# Dados da empresa exibidos no cabeçalho dos relatórios (mesmos dados usados
# no Relatório Quinzenal de Medições, para manter uniformidade entre os relatórios)
EMPRESA_INFO = {
    'endereco': 'Rua Zodiaco, 87 Sala 07 – Santa Lúcia - Belo Horizonte - MG',
    'fones': '(31) 3654-6616 / (31) 99974-1241 / (31) 98711-1139',
    'email': 'rvr.engenharia@gmail.com',
}

# Importar o utils.py
from src.config.utils import atualizar_combobox_clientes, cliente_esta_ativo, obter_info_cliente

try:
    from src.config.window_config import configurar_janela
    print("window_config importado com sucesso")
except ImportError as e:
    print(f"Erro ao importar window_config: {str(e)}")
    # Implementação simples de configurar_janela como fallback
    def configurar_janela(janela, titulo="Janela", largura=800, altura=600):
        janela.title(titulo)
        janela.geometry(f"{largura}x{altura}")
        janela.resizable(True, True)
        
        # Centralizar na tela
        janela.update_idletasks()
        width = janela.winfo_width()
        height = janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (width // 2)
        y = (janela.winfo_screenheight() // 2) - (height // 2)
        janela.geometry(f'{width}x{height}+{x}+{y}')

# Funções auxiliares
def formatar_moeda_br(valor):
    """Formata um valor numérico como moeda brasileira"""
    try:
        valor_float = float(valor)
        return f"R$ {valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return f"R$ 0,00"

def formatar_valor_sem_simbolo(valor):
    """Formata um valor numérico sem o símbolo R$"""
    try:
        valor_float = float(valor)
        return f"{valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return "0,00"

def cor_texto_contraste(cor_rgb):
    """
    Recebe uma cor (tupla RGB ou RGBA, valores 0-1) e retorna 'black' ou 'white',
    a que tiver melhor contraste sobre essa cor de fundo (fórmula de luminância
    relativa). Usado para rótulos de valor sobre barras de cores claras (ex.: os
    tons pastel das paletas Set3/Paired), onde texto branco fixo fica ilegível.
    """
    try:
        r, g, b = cor_rgb[0], cor_rgb[1], cor_rgb[2]
        luminancia = 0.299 * r + 0.587 * g + 0.114 * b
        return 'black' if luminancia > 0.6 else 'white'
    except Exception:
        return 'black'

def obter_imagem_logo_reportlab(largura_max_pt, altura_max_pt):
    """
    Retorna um objeto Image do ReportLab com o logo redimensionado
    mantendo a proporção original, ou None se o arquivo não existir
    ou não puder ser lido.
    """
    try:
        if not LOGO_PATH.exists():
            print(f"Aviso: logo não encontrado em {LOGO_PATH}")
            return None

        from reportlab.platypus import Image as RLImage

        largura_final = largura_max_pt
        altura_final = altura_max_pt
        try:
            from PIL import Image as PILImage
            with PILImage.open(LOGO_PATH) as img:
                largura_orig, altura_orig = img.size
            proporcao = min(largura_max_pt / largura_orig, altura_max_pt / altura_orig)
            largura_final = largura_orig * proporcao
            altura_final = altura_orig * proporcao
        except ImportError:
            # Sem Pillow disponível: usa tamanho máximo fixo (pode distorcer)
            pass

        return RLImage(str(LOGO_PATH), width=largura_final, height=altura_final)
    except Exception as e:
        print(f"Aviso: não foi possível carregar o logo: {e}")
        return None

class RelatorioCategoria:
    """Classe para geração de relatórios por categoria de despesa agrupado por mês de vencimento"""
    
    def __init__(self, parent=None):
        """Inicializa a interface do relatório"""
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Relatório por Categoria - Agrupado por Mês de Vencimento", 1200, 1000)
        
        # Definição das categorias de despesa
        self.categorias_despesas = {
            'ADM': "ADMINISTRATIVO",
            'DIV': "DIVERSOS",
            'LOC': "LOCAÇÃO", 
            'MAT': "MATERIAL", 
            'MO': "MÃO-DE-OBRA", 
            'SERV': "SERVIÇOS",
            'TAX': "TAXA ADMINISTRAÇÃO",
            'TP': "TARIFAS/TRIBUTOS PÚBLICOS"
        }
        
        # Configuração de variáveis
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.data_referencia = datetime.now()
        self.df_despesas = None
        self.df_por_mes = None  # Mudança: agora agrupamos por mês
        self.data_selecionada = None  # Mudança: em vez de data_selecionada
        self.dados_grafico = {}
        
        # Configurar interface
        self.setup_gui()
    
    def setup_gui(self):
        """Configuração da interface gráfica principal"""
        # Frame principal
        self.frame_principal = ttk.Frame(self.root, padding=10)
        self.frame_principal.pack(fill='both', expand=True)
        
        # Frame para seleção (TOPO - FIXO)
        self.frame_selecao = ttk.LabelFrame(self.frame_principal, text="Seleção de Cliente")
        self.frame_selecao.pack(fill='x', side='top', pady=(0, 10))
        
        # Container para cliente
        frame_cliente = ttk.Frame(self.frame_selecao)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Selecione o Cliente:", font=('Arial', 11)).pack(side='left', pady=5)
        self.cliente_combobox = ttk.Combobox(frame_cliente, width=40, font=('Arial', 11))
        self.cliente_combobox.pack(side='left', padx=5)
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)
        
        # Container para botão de gerar relatório
        frame_data = ttk.Frame(self.frame_selecao)
        frame_data.pack(fill='x', padx=10, pady=10)
        
        # Botão de gerar relatório
        ttk.Button(
            frame_data,
            text="Gerar Relatório",
            command=self.gerar_relatorio,
            style='Big.TButton'
        ).pack(side='left', padx=20)
        
        # BOTÕES NA PARTE INFERIOR (FIXO)
        frame_botoes = ttk.Frame(self.frame_principal)
        frame_botoes.pack(fill='x', side='bottom', pady=(10, 0))
        
        ttk.Button(
            frame_botoes,
            text="Exportar para Excel",
            command=self.exportar_excel
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes,
            text="Exportar para PDF",
            command=self.exportar_pdf
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes,
            text="Voltar ao Menu",
            command=self.voltar_menu
        ).pack(side='right', padx=5)
        
        # Frame para resultados - MEIO (EXPANSÍVEL)
        self.frame_resultados = ttk.LabelFrame(self.frame_principal, text="Resultados")
        self.frame_resultados.pack(fill='both', expand=True, pady=(0, 10))
        
        # Notebook (abas)
        self.notebook = ttk.Notebook(self.frame_resultados)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Abas
        self.aba_resumo = ttk.Frame(self.notebook)
        self.aba_detalhes = ttk.Frame(self.notebook)
        self.aba_grafico = ttk.Frame(self.notebook)
        
        self.notebook.add(self.aba_resumo, text='Resumo')
        self.notebook.add(self.aba_detalhes, text='Detalhes')
        self.notebook.add(self.aba_grafico, text='Gráfico')
        
        # Configurar cada aba
        self.setup_aba_resumo()
        self.setup_aba_detalhes()
        self.setup_aba_grafico()
        
        # Estilo para botões grandes
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 11, 'bold'), padding=(10, 5))
        
        # Carregar lista de clientes
        self.atualizar_lista_clientes()
    
    def setup_aba_resumo(self):
        """Configura a aba de resumo do relatório por mês de vencimento e categoria de despesa"""
        # Frame para informações do cliente
        frame_info = ttk.Frame(self.aba_resumo, padding=5)
        frame_info.pack(fill='x', pady=5)
        
        self.lbl_cliente_resumo = ttk.Label(
            frame_info, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_resumo.pack(side='left', padx=10)
        
        self.lbl_data_resumo = ttk.Label(
            frame_info, 
            text=f"Agrupamento: Por Mês de Vencimento", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_data_resumo.pack(side='left', padx=10)
        
        # Frame para o TreeView com os dados por mês
        frame_resumo = ttk.Frame(self.aba_resumo, padding=5)
        frame_resumo.pack(fill='both', expand=True, pady=5)
        
        # Criar TreeView para os dados por mês
        # Colunas: 'mes_ano', categorias (ADM, DIV, LOC, MAT, MO, SERV, TAX, TP), 'total'
        colunas = ['mes_ano']
        for categoria in self.categorias_despesas.keys():
            colunas.append(f'cat_{categoria}')
        colunas.append('total')
        
        self.tv_resumo = ttk.Treeview(frame_resumo, columns=colunas, show='headings', height=15)
        
        # Configurar cabeçalhos
        self.tv_resumo.heading('mes_ano', text='Mês/Ano')
        for categoria in self.categorias_despesas.keys():
            # Usar a sigla da categoria para o cabeçalho
            self.tv_resumo.heading(f'cat_{categoria}', text=categoria)
            
        self.tv_resumo.heading('total', text='Total (R$)')
        
        # Configurar colunas
        self.tv_resumo.column('mes_ano', width=120, anchor='center')
        for categoria in self.categorias_despesas.keys():
            self.tv_resumo.column(f'cat_{categoria}', width=100, anchor='e', stretch=True)
        self.tv_resumo.column('total', width=120, anchor='e')
        
        # Configurar scrollbars
        scrollbar_y = ttk.Scrollbar(frame_resumo, orient='vertical', command=self.tv_resumo.yview)
        scrollbar_x = ttk.Scrollbar(frame_resumo, orient='horizontal', command=self.tv_resumo.xview)
        self.tv_resumo.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Adicionar à tela
        self.tv_resumo.pack(side='top', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        
        # Adicionar evento de seleção
        self.tv_resumo.bind('<<TreeviewSelect>>', self.selecionar_data)
        
        # Frame para resumo de totais
        frame_totais = ttk.LabelFrame(self.aba_resumo, text="Resumo Financeiro", padding=10)
        frame_totais.pack(fill='x', pady=10, padx=10)
        
        # Total de Despesas (primeira linha, primeira coluna)
        ttk.Label(frame_totais, text="Total de Despesas:", font=('Arial', 11, 'bold')).grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.lbl_total_geral = ttk.Label(frame_totais, text="0,00", font=('Arial', 11))
        self.lbl_total_geral.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        # Criar labels dinâmicos para cada categoria
        self.labels_categorias = {}
        row = 1
        col = 0
        
        for categoria, nome_completo in self.categorias_despesas.items():
            # Label do nome da categoria
            ttk.Label(frame_totais, text=f"{categoria}:", font=('Arial', 10, 'bold')).grid(row=row, column=col, sticky='e', padx=5, pady=2)
            
            # Label do valor da categoria
            self.labels_categorias[categoria] = ttk.Label(frame_totais, text="0,00", font=('Arial', 10))
            self.labels_categorias[categoria].grid(row=row, column=col+1, sticky='w', padx=5, pady=2)
            
            # Avançar para próxima posição
            col += 2
            if col >= 12:  # Máximo de 6 colunas (6 considerando label + valor)
                col = 0
                row += 1
    
    def setup_aba_detalhes(self):
        """Configura a aba de detalhes do relatório para o mês selecionado"""
        # Frame para informações do mês selecionado
        frame_info_mes = ttk.Frame(self.aba_detalhes, padding=5)
        frame_info_mes.pack(fill='x', pady=5)
        
        self.lbl_mes_detalhe = ttk.Label(
            frame_info_mes, 
            text="Mês Selecionado: Nenhum", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_mes_detalhe.pack(side='left', padx=10)
        
        self.lbl_total_mes_detalhe = ttk.Label(
            frame_info_mes, 
            text="Total: R$ 0,00", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_total_mes_detalhe.pack(side='left', padx=10)
        
        # Frame para a tabela de detalhes
        frame_tabela = ttk.Frame(self.aba_detalhes, padding=5)
        frame_tabela.pack(fill='both', expand=True, pady=5)
        
        # Criar TreeView para os lançamentos do mês selecionado
        colunas = ('dt_vencto', 'categoria', 'nome', 'referencia', 'data_rel', 'valor', 'observacao')
        self.tv_detalhes = ttk.Treeview(frame_tabela, columns=colunas, show='headings', height=15)
        
        # Configurar cabeçalhos
        self.tv_detalhes.heading('dt_vencto', text='Dt. Vencimento')
        self.tv_detalhes.heading('categoria', text='Cat')  # MUDOU
        self.tv_detalhes.heading('nome', text='Nome')
        self.tv_detalhes.heading('referencia', text='Referência')
        self.tv_detalhes.heading('data_rel', text='Data Relatório')
        self.tv_detalhes.heading('valor', text='Valor')  # MUDOU - sem (R$)
        self.tv_detalhes.heading('observacao', text='Observação')
        
        # Configurar colunas
        self.tv_detalhes.column('dt_vencto', width=100, anchor='center')
        self.tv_detalhes.column('categoria', width=50, anchor='center')  # MUDOU largura
        self.tv_detalhes.column('nome', width=180, anchor='w')
        self.tv_detalhes.column('referencia', width=220, anchor='w')
        self.tv_detalhes.column('data_rel', width=100, anchor='center')
        self.tv_detalhes.column('valor', width=100, anchor='e')  # MUDOU largura
        self.tv_detalhes.column('observacao', width=180, anchor='w')

        
        # Configurar scrollbars
        scrollbar_y = ttk.Scrollbar(frame_tabela, orient='vertical', command=self.tv_detalhes.yview)
        scrollbar_x = ttk.Scrollbar(frame_tabela, orient='horizontal', command=self.tv_detalhes.xview)
        self.tv_detalhes.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Adicionar à tela
        self.tv_detalhes.pack(side='top', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        
        # Frame para resumo de totais (igual ao da aba resumo)
        frame_totais_detalhes = ttk.LabelFrame(self.aba_detalhes, text="Resumo Financeiro", padding=10)
        frame_totais_detalhes.pack(fill='x', pady=10, padx=10)
        
        # Total de Despesas (primeira linha, primeira coluna)
        ttk.Label(frame_totais_detalhes, text="Total de Despesas:", font=('Arial', 11, 'bold')).grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.lbl_total_geral_detalhes = ttk.Label(frame_totais_detalhes, text="0,00", font=('Arial', 11))
        self.lbl_total_geral_detalhes.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        # Criar labels dinâmicos para cada categoria na aba detalhes
        self.labels_categorias_detalhes = {}
        row = 1
        col = 0
        
        for categoria, nome_completo in self.categorias_despesas.items():
            # Label do nome da categoria
            ttk.Label(frame_totais_detalhes, text=f"{categoria}:", font=('Arial', 10, 'bold')).grid(row=row, column=col, sticky='e', padx=5, pady=2)
            
            # Label do valor da categoria
            self.labels_categorias_detalhes[categoria] = ttk.Label(frame_totais_detalhes, text="0,00", font=('Arial', 10))
            self.labels_categorias_detalhes[categoria].grid(row=row, column=col+1, sticky='w', padx=5, pady=2)
            
            # Avançar para próxima posição
            col += 2
            if col >= 12:  # Máximo de 6 colunas (considerando label + valor)
                col = 0
                row += 1
    
    def setup_aba_grafico(self):
        """Configura a aba de gráficos"""
        # Frame para controles do gráfico
        frame_controles = ttk.Frame(self.aba_grafico, padding=5)
        frame_controles.pack(fill='x', pady=5)
        
        ttk.Label(frame_controles, text="Tipo de Gráfico:").pack(side='left', padx=5)
        self.combo_tipo_grafico = ttk.Combobox(frame_controles, values=[
            "Gráfico de Pizza - Totais",
            "Gráfico de Barras - Totais", 
            "Gráfico de Linha do Tempo",
            "Gráfico de Pizza - Data Selecionada",
            "Gráfico de Barras - Data Selecionada"
        ], state='readonly', width=35)
        self.combo_tipo_grafico.pack(side='left', padx=5)
        self.combo_tipo_grafico.current(0)
        
        ttk.Button(frame_controles, text="Atualizar Gráfico", command=self.atualizar_grafico).pack(side='left', padx=20)
        
        # Frame para informações do mês no gráfico
        frame_info_grafico = ttk.Frame(self.aba_grafico, padding=5)
        frame_info_grafico.pack(fill='x', pady=5)
        
        self.lbl_data_grafico = ttk.Label(
            frame_info_grafico, 
            text="Visualização: Dados Gerais", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_data_grafico.pack(side='left', padx=10)
        
        # Frame para o gráfico
        self.frame_grafico = ttk.Frame(self.aba_grafico)
        self.frame_grafico.pack(fill='both', expand=True, pady=5)
    
    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes no combobox usando a função centralizada"""
        try:
            # Usar a função centralizada (apenas clientes ativos)
            self.info_clientes = atualizar_combobox_clientes(self.cliente_combobox, mostrar_inativos=False)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")

    def selecionar_cliente(self, event=None):
        """Atualiza o cliente selecionado"""
        self.cliente_atual = self.cliente_combobox.get()
        
        if self.cliente_atual:
            # Verificar se o cliente está ativo (extra proteção)
            if not cliente_esta_ativo(self.cliente_atual):
                messagebox.showwarning(
                    "Cliente Inativo", 
                    f"O cliente '{self.cliente_atual}' está inativo (contrato finalizado). " +
                    "Os dados serão mostrados somente para consulta."
                )
            
            # Obter informações do cliente
            info_cliente = obter_info_cliente(self.cliente_atual)
            
            # Atualizar label
            if hasattr(self, 'lbl_cliente_resumo'):
                texto_cliente = f"Cliente: {self.cliente_atual}"
                if info_cliente and not info_cliente['ativo']:
                    texto_cliente += " (INATIVO)"
                self.lbl_cliente_resumo.config(text=texto_cliente)
            
            # Definir o caminho do arquivo
            if info_cliente and 'arquivo' in info_cliente:
                self.arquivo_cliente = info_cliente['arquivo']
            else:
                self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
    
    def gerar_relatorio(self):
        """Gera o relatório com base nos dados selecionados"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
            
        # Definir data de referência como data atual (apenas para o relatório exportado)
        self.data_referencia = datetime.now()
            
        # Carregar dados
        if not self.carregar_dados():
            return
        
        # Preencher resumo
        self.preencher_resumo()
        
        # Resetar resumo financeiro da aba detalhes (ainda não há data selecionada)
        if hasattr(self, 'lbl_total_geral_detalhes'):
            self.lbl_total_geral_detalhes.config(text="0,00")
            
            # PARA CATEGORIA (usar este bloco no relatorio_categoria.py):
            if hasattr(self, 'labels_categorias_detalhes'):
                for categoria in self.categorias_despesas.keys():
                    if categoria in self.labels_categorias_detalhes:
                        self.labels_categorias_detalhes[categoria].config(text="0,00")

                        
        # Limpar detalhes (pois ainda não há data selecionada)
        if hasattr(self, 'tv_detalhes'):
            for item in self.tv_detalhes.get_children():
                self.tv_detalhes.delete(item)
        
        # Resetar data selecionada
        self.data_selecionada = None
        
        # Atualizar label do gráfico
        if hasattr(self, 'lbl_data_grafico'):
            self.lbl_data_grafico.config(text="Visualização: Dados Gerais")
        
        # Gerar gráfico inicial de totais
        self.atualizar_grafico()
        
        # Selecionar aba de resumo
        self.notebook.select(0)  # Índice 0 corresponde à primeira aba (resumo)
    
    def carregar_dados(self):
        """Carrega os dados para o relatório a partir da aba 'Dados'"""
        try:
            if not os.path.exists(self.arquivo_cliente):
                messagebox.showerror("Erro", f"Arquivo do cliente '{self.cliente_atual}' não encontrado!")
                return False
            
            # Carregar dados do Excel - da aba 'Dados'
            try:
                # Usar pandas para ler a aba Dados
                self.df_despesas = pd.read_excel(self.arquivo_cliente, sheet_name='Dados')
                print(f"Colunas do DataFrame: {self.df_despesas.columns.tolist()}")
                
                # Verificar se as colunas necessárias existem
                colunas_necessarias = ['VALOR', 'DATA_REL']
                for coluna in colunas_necessarias:
                    if coluna not in self.df_despesas.columns:
                        messagebox.showerror("Erro", f"A coluna '{coluna}' não foi encontrada na aba Dados!")
                        return False
                
                # Verificar se existe a coluna K (categoria) - índice 10 (base 0)
                if len(self.df_despesas.columns) < 11:
                    messagebox.showerror("Erro", "A coluna K (categoria) não foi encontrada na aba Dados!")
                    return False
                
                # Nomear a coluna K como 'CATEGORIA' se ainda não tiver nome
                nome_coluna_k = self.df_despesas.columns[10]  # Coluna K (índice 10)
                if nome_coluna_k != 'CATEGORIA':
                    # Renomear a coluna K para CATEGORIA
                    self.df_despesas = self.df_despesas.rename(columns={nome_coluna_k: 'CATEGORIA'})
                
                print(f"Coluna de categoria identificada: {nome_coluna_k} -> CATEGORIA")
                
                # NOVO: Filtrar apenas lançamentos ativos
                if 'STATUS' in self.df_despesas.columns:
                    # Filtrar apenas registros com STATUS = 'ATIVO'
                    df_original_len = len(self.df_despesas)
                    self.df_despesas = self.df_despesas[
                        self.df_despesas['STATUS'].str.upper().str.strip() == 'ATIVO'
                    ].copy()
                    df_filtrado_len = len(self.df_despesas)
                    
                    print(f"Cliente {self.cliente_atual}: {df_original_len} registros totais, {df_filtrado_len} ativos processados")
                    
                    # Se não há registros ativos, mostrar aviso
                    if self.df_despesas.empty:
                        messagebox.showinfo("Aviso", f"Nenhum lançamento ativo encontrado para o cliente {self.cliente_atual}")
                        return False
                else:
                    # Se não existe a coluna STATUS, processar todos (compatibilidade)
                    print(f"Cliente {self.cliente_atual}: Coluna STATUS não encontrada, processando todos os registros")
                
                # Converter DATA_REL para datetime
                self.df_despesas['DATA_REL'] = pd.to_datetime(self.df_despesas['DATA_REL'], errors='coerce')
                
                # Converter DT_VENCTO para datetime (se existir)
                if 'DT_VENCTO' in self.df_despesas.columns:
                    self.df_despesas['DT_VENCTO'] = pd.to_datetime(self.df_despesas['DT_VENCTO'], format='%d/%m/%Y', errors='coerce')
                else:
                    # Se não existir, criar coluna vazia
                    self.df_despesas['DT_VENCTO'] = pd.NaT

                # Garantir valores numéricos para a coluna valor
                self.df_despesas['VALOR'] = pd.to_numeric(self.df_despesas['VALOR'], errors='coerce').fillna(0)
                
                # Processar categorias - converter para string e limpar
                self.df_despesas['CATEGORIA'] = self.df_despesas['CATEGORIA'].astype(str).str.upper().str.strip()
                
                # Substituir valores vazios ou 'nan' por 'DIV' (DIVERSOS)
                self.df_despesas['CATEGORIA'] = self.df_despesas['CATEGORIA'].replace(['', 'NAN', 'NONE'], 'DIV')
                
                # Verificar se existem categorias não mapeadas
                categorias_encontradas = set(self.df_despesas['CATEGORIA'].unique())
                categorias_validas = set(self.categorias_despesas.keys())
                categorias_invalidas = categorias_encontradas - categorias_validas
                
                if categorias_invalidas:
                    print(f"Categorias não mapeadas encontradas: {categorias_invalidas}")
                    # Substituir categorias inválidas por 'DIV'
                    self.df_despesas.loc[self.df_despesas['CATEGORIA'].isin(categorias_invalidas), 'CATEGORIA'] = 'DIV'
                
                # Ordenar dados por data
                self.df_despesas = self.df_despesas.sort_values(by='DATA_REL')
                
                # Agrupar dados por data e categoria
                self.preparar_dados_por_data()
                
                return True
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao carregar dados do Excel: {str(e)}")
                import traceback
                traceback.print_exc()
                return False
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
            return False
    
    def preparar_dados_por_data(self):
        """Prepara os dados agrupados por mês de vencimento"""
        try:
            # Criar coluna de mês/ano baseada em DT_VENCTO
            self.df_despesas['mes_ano_vencto'] = self.df_despesas['DT_VENCTO'].dt.to_period('M')
            
            # Agrupar por mês/ano de vencimento e categoria de despesa
            df_pivot = self.df_despesas.pivot_table(
                index='mes_ano_vencto', 
                columns='CATEGORIA', 
                values='VALOR', 
                aggfunc='sum'
            ).fillna(0)
            
            # Resetar o índice
            df_pivot = df_pivot.reset_index()
            
            # Criar colunas para cada categoria se não existirem
            for categoria in self.categorias_despesas.keys():
                if categoria not in df_pivot.columns:
                    df_pivot[categoria] = 0.0
            
            # Calcular total por mês
            df_pivot['total'] = df_pivot[[cat for cat in self.categorias_despesas.keys() if cat in df_pivot.columns]].sum(axis=1)
            
            # Ordenar por mês/ano (ascendente)
            df_pivot = df_pivot.sort_values(by='mes_ano_vencto')
            
            # Armazenar o DataFrame
            self.df_por_data = df_pivot
            
            # Preparar dados para gráficos
            self.preparar_dados_grafico()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao preparar dados por mês: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preparar_dados_grafico(self):
        """Prepara os dados para os gráficos"""
        try:
            # Inicializar dicionário de dados para gráficos
            self.dados_grafico = {}
            
            # Se não temos dados ou data selecionada, não há o que fazer
            if not hasattr(self, 'df_por_data') or self.df_por_data.empty:
                return
                
            # Dados para gráfico de pizza e barras por data selecionada
            # Serão preenchidos quando uma data for selecionada
            self.dados_grafico['pizza'] = None
            self.dados_grafico['barras'] = None
            
        except Exception as e:
            print(f"Erro ao preparar dados para gráfico: {str(e)}")
            import traceback
            traceback.print_exc()

    
    def preencher_resumo(self):
        """Preenche os dados da aba de resumo"""
        try:
            # Limpar TreeView
            for item in self.tv_resumo.get_children():
                self.tv_resumo.delete(item)
            
            # Adicionar dados à TreeView
            for _, row in self.df_por_data.iterrows():
                # Formatar mês/ano
                mes_ano_str = row['mes_ano_vencto'].strftime('%m/%Y')
                
                # Preparar valores para cada categoria (SEM R$)
                valores = []
                for categoria in self.categorias_despesas.keys():
                    valor_formatado = formatar_valor_sem_simbolo(row[categoria]) if categoria in row else "0,00"
                    valores.append(valor_formatado)
                
                # Adicionar total (SEM R$)
                total_formatado = formatar_valor_sem_simbolo(row['total'])
                
                # Inserir na treeview
                self.tv_resumo.insert(
                    '', 'end', 
                    values=[mes_ano_str] + valores + [total_formatado]
                )
            
            # Atualizar labels de totais (SEM R$)
            total_geral = self.df_por_data['total'].sum()
            self.lbl_total_geral.config(text=formatar_valor_sem_simbolo(total_geral))
            
            for categoria in self.categorias_despesas.keys():
                if categoria in self.df_por_data.columns:
                    total_tipo = self.df_por_data[categoria].sum()
                else:
                    total_tipo = 0
                
                if categoria in self.labels_categorias:
                    self.labels_categorias[categoria].config(text=formatar_valor_sem_simbolo(total_tipo))
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao preencher resumo: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def selecionar_data(self, event=None):
        """Atualiza o mês selecionado e preenche as abas de detalhes e gráfico"""
        try:
            # Obter seleção atual
            selecao = self.tv_resumo.selection()
            if not selecao:
                return
                
            # Obter mês/ano selecionado
            item = self.tv_resumo.item(selecao[0])
            mes_ano_str = item['values'][0]  # Primeira coluna é o mês/ano
            
            # Converter string mês/ano para Period
            try:
                mes, ano = mes_ano_str.split('/')
                self.mes_ano_selecionado = pd.Period(year=int(ano), month=int(mes), freq='M')
                self.data_selecionada = self.mes_ano_selecionado
            except ValueError:
                messagebox.showerror("Erro", f"Formato de mês/ano inválido: {mes_ano_str}")
                return
            
            # Atualizar labels
            self.lbl_mes_detalhe.config(text=f"Mês Selecionado: {mes_ano_str}")
            self.lbl_data_grafico.config(text=f"Mês Selecionado: {mes_ano_str}")
            
            # Encontrar o total do mês no DataFrame
            df_mes = self.df_por_data[self.df_por_data['mes_ano_vencto'] == self.mes_ano_selecionado]
            if not df_mes.empty:
                total_mes = df_mes.iloc[0]['total']
                self.lbl_total_mes_detalhe.config(text=f"Total: {formatar_moeda_br(total_mes)}")
            
            # Filtrar dados para o mês selecionado
            self.df_despesas['mes_ano_vencto_temp'] = self.df_despesas['DT_VENCTO'].dt.to_period('M')
            df_filtrado = self.df_despesas[self.df_despesas['mes_ano_vencto_temp'] == self.mes_ano_selecionado].copy()
            
            # Atualizar resumo financeiro e detalhes
            self.atualizar_resumo_financeiro_detalhes(df_filtrado)
            self.preencher_detalhes(df_filtrado)
            self.preparar_grafico_data_selecionada(df_filtrado)
            
            # Atualizar tipo de gráfico
            if "Data Selecionada" not in self.combo_tipo_grafico.get():
                self.combo_tipo_grafico.set("Gráfico de Pizza - Data Selecionada")
            
            self.atualizar_grafico()
            
            # Alternar para aba de detalhes
            self.notebook.select(1)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao selecionar mês: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def atualizar_resumo_financeiro_detalhes(self, df_filtrado):
        """Atualiza o resumo financeiro da aba detalhes com dados da data selecionada"""
        try:
            if df_filtrado.empty:
                self.lbl_total_geral_detalhes.config(text="0,00")
                for categoria in self.categorias_despesas.keys():
                    if categoria in self.labels_categorias_detalhes:
                        self.labels_categorias_detalhes[categoria].config(text="0,00")
                return
            
            # Calcular total (SEM R$)
            total_data = df_filtrado['VALOR'].sum()
            if hasattr(self, 'lbl_total_geral_detalhes'):
                self.lbl_total_geral_detalhes.config(text=formatar_valor_sem_simbolo(total_data))
            
            # Calcular totais por categoria
            totais_por_categoria = df_filtrado.groupby('CATEGORIA')['VALOR'].sum()
            
            # Atualizar labels (SEM R$)
            for categoria in self.categorias_despesas.keys():
                if categoria in self.labels_categorias_detalhes:
                    valor_categoria = totais_por_categoria.get(categoria, 0)
                    self.labels_categorias_detalhes[categoria].config(text=formatar_valor_sem_simbolo(valor_categoria))
                    
        except Exception as e:
            print(f"Erro ao atualizar resumo financeiro detalhes: {str(e)}")

    def preparar_grafico_data_selecionada(self, df_filtrado):
        """Prepara dados para os gráficos da data selecionada"""
        try:
            if df_filtrado.empty:
                self.dados_grafico['pizza'] = None
                self.dados_grafico['barras'] = None
                return
                
            # Agrupar por categoria
            df_agrupado = df_filtrado.groupby('CATEGORIA')['VALOR'].sum().reset_index()
            
            # Adicionar nome completo da categoria
            df_agrupado['categoria_nome'] = df_agrupado['CATEGORIA'].apply(
                lambda x: f"{x} - {self.categorias_despesas.get(x, 'Não classificado')}" if pd.notna(x) else "Não classificado"
            )
            
            # Ordenar por categoria
            df_agrupado = df_agrupado.sort_values(by='CATEGORIA')
            
            # Armazenar para gráficos
            self.dados_grafico['pizza'] = df_agrupado.copy()
            self.dados_grafico['barras'] = df_agrupado.copy()
            
        except Exception as e:
            print(f"Erro ao preparar gráfico para data selecionada: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def preencher_detalhes(self, df_filtrado):
        """Preenche os detalhes para a data selecionada"""
        try:
            # Limpar tabela
            for item in self.tv_detalhes.get_children():
                self.tv_detalhes.delete(item)
            
            # Verificar se o DataFrame está vazio
            if df_filtrado.empty:
                return
            
            # Ordenar por Categoria, Nome e Valor
            df_ordenado = df_filtrado.copy()
            
            if 'NOME' not in df_ordenado.columns:
                df_ordenado['NOME'] = ''
                
            df_ordenado = df_ordenado.sort_values(
                by=['CATEGORIA', 'NOME', 'VALOR'], 
                ascending=[True, True, False]
            )
            
            # Adicionar dados à tabela
            for _, row in df_ordenado.iterrows():
                # Data de vencimento
                dt_vencto_str = ''
                if 'DT_VENCTO' in row and pd.notna(row['DT_VENCTO']):
                    dt_vencto_str = row['DT_VENCTO'].strftime('%d/%m/%Y')
                
                # Categoria (APENAS O CÓDIGO)
                categoria = row['CATEGORIA'] if pd.notna(row['CATEGORIA']) else 'DIV'
                
                # Nome
                nome = row.get('NOME', '') if pd.notna(row.get('NOME', '')) else ''
                
                # Referência e NF
                referencia = row.get('REFERÊNCIA', '') if pd.notna(row.get('REFERÊNCIA', '')) else ''
                nf = row.get('NF', '') if pd.notna(row.get('NF', '')) else ''
                if referencia and nf:
                    referencia = f"{referencia} - NF: {nf}"
                elif nf:
                    referencia = f"NF: {nf}"
                
                # Data relatório
                data_str = row['DATA_REL'].strftime('%d/%m/%Y') if pd.notna(row['DATA_REL']) else ''
                
                # Valor (SEM R$)
                valor = formatar_valor_sem_simbolo(row['VALOR'])
                
                # Observação
                observacao = row.get('OBSERVAÇÃO', '') if pd.notna(row.get('OBSERVAÇÃO', '')) else ''
                
                # Inserir na tabela (ordem: dt_vencto, categoria, nome, referencia, data_rel, valor, observacao)
                self.tv_detalhes.insert(
                    '', 'end', 
                    values=(
                        dt_vencto_str,
                        categoria,  # APENAS CÓDIGO
                        nome,
                        referencia,
                        data_str,
                        valor,
                        observacao
                    )
                )
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao preencher detalhes: {str(e)}")
            import traceback
            traceback.print_exc()
 
    def determinar_agrupamento_temporal(self):
        """Determina o melhor agrupamento temporal baseado no período dos dados"""
        if not hasattr(self, 'df_por_data') or self.df_por_data.empty:
            return 'dia'
        
        # Calcular diferença baseado nos períodos de mês/ano
        data_inicio = self.df_por_data['mes_ano_vencto'].min().to_timestamp()
        data_fim = self.df_por_data['mes_ano_vencto'].max().to_timestamp()
        dias_total = (data_fim - data_inicio).days
        
        # Definir agrupamento baseado no período
        if dias_total <= 31:  # Até 1 mês
            return 'dia'
        elif dias_total <= 93:  # Até 3 meses
            return 'semana'
        elif dias_total <= 365:  # Até 1 ano
            return 'mes'
        elif dias_total <= 1095:  # Até 3 anos
            return 'trimestre'
        else:  # Mais de 3 anos
            return 'ano'

    def preparar_dados_timeline(self):
        """Prepara dados para gráfico de linha do tempo baseado em mês de vencimento"""
        try:
            if not hasattr(self, 'df_despesas') or self.df_despesas.empty:
                return None
            
            agrupamento = self.determinar_agrupamento_temporal()
            df_timeline = self.df_despesas.copy()
            
            # **MODIFICAR: Criar coluna de agrupamento temporal baseado em DT_VENCTO**
            if agrupamento == 'dia':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.date
                formato_periodo = lambda x: x.strftime('%d/%m/%Y')
            elif agrupamento == 'semana':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('W')
                formato_periodo = lambda x: f"Sem {x.week}/{x.year}"
            elif agrupamento == 'mes':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('M')
                formato_periodo = lambda x: f"{x.month:02d}/{x.year}"
            elif agrupamento == 'trimestre':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('Q')
                formato_periodo = lambda x: f"Q{x.quarter}/{x.year}"
            else:  # ano
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('Y')
                formato_periodo = lambda x: str(x.year)
            
            # Agrupar por período e categoria
            df_agrupado = df_timeline.groupby(['periodo', 'CATEGORIA'])['VALOR'].sum().reset_index()
            
            # Criar pivot para ter categorias como colunas
            df_pivot = df_agrupado.pivot(index='periodo', columns='CATEGORIA', values='VALOR').fillna(0)
            
            # Garantir que todas as categorias existam
            for categoria in self.categorias_despesas.keys():
                if categoria not in df_pivot.columns:
                    df_pivot[categoria] = 0
            
            # Resetar índice e formatar período
            df_pivot = df_pivot.reset_index()
            df_pivot['periodo_str'] = df_pivot['periodo'].apply(formato_periodo)
            
            # Calcular total por período
            df_pivot['total'] = df_pivot[[cat for cat in self.categorias_despesas.keys() if cat in df_pivot.columns]].sum(axis=1)
            
            return {
                'dados': df_pivot,
                'agrupamento': agrupamento,
                'categorias': list(self.categorias_despesas.keys())
            }
            
        except Exception as e:
            print(f"Erro ao preparar dados de timeline: {str(e)}")
            import traceback
            traceback.print_exc()
            return None  

    def atualizar_grafico(self, event=None):
        """Atualiza o gráfico com base na seleção"""
        try:
            tipo_grafico = self.combo_tipo_grafico.get()
            
            # Limpar frame do gráfico
            self.limpar_grafico()
            
            # Verificar se há dados
            if not hasattr(self, 'df_por_data') or self.df_por_data.empty:
                return
            
            # Criar figura
            fig, ax = plt.subplots(figsize=(12, 7))
            
            if "Linha do Tempo" in tipo_grafico:
                self.criar_grafico_timeline(fig, ax)
            elif "Data Selecionada" in tipo_grafico:
                # Gráficos para data específica
                if not self.data_selecionada:
                    ax.text(0.5, 0.5, "Selecione uma data na aba Resumo\npara ver os gráficos da data específica", 
                        horizontalalignment='center', verticalalignment='center',
                        transform=ax.transAxes, fontsize=14)
                else:
                    if "Pizza" in tipo_grafico:
                        self.criar_grafico_pizza(fig, ax)
                    elif "Barras" in tipo_grafico:
                        self.criar_grafico_barras(fig, ax)
            else:
                # Gráficos de totais gerais
                if "Pizza" in tipo_grafico:
                    self.criar_grafico_pizza_totais(fig, ax)
                elif "Barras" in tipo_grafico:
                    self.criar_grafico_barras_totais(fig, ax)
            
            # Exibir o gráfico
            canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar gráfico: {str(e)}")
            import traceback
            traceback.print_exc()

    def limpar_grafico(self):
        """Limpa o gráfico atual"""
        for widget in self.frame_grafico.winfo_children():
            widget.destroy()
    
    
    def determinar_agrupamento_temporal(self):
        """Determina o melhor agrupamento temporal baseado no período dos dados"""
        if not hasattr(self, 'df_por_data') or self.df_por_data.empty:
            return 'dia'
        
        # Calcular diferença em dias entre primeira e última data
        data_inicio = self.df_por_data['mes_ano_vencto'].min().to_timestamp()
        data_fim = self.df_por_data['mes_ano_vencto'].max().to_timestamp()
        dias_total = (data_fim - data_inicio).days
        
        # Definir agrupamento baseado no período
        if dias_total <= 31:  # Até 1 mês
            return 'dia'
        elif dias_total <= 93:  # Até 3 meses
            return 'semana'
        elif dias_total <= 365:  # Até 1 ano
            return 'mes'
        elif dias_total <= 1095:  # Até 3 anos
            return 'trimestre'
        else:  # Mais de 3 anos
            return 'ano'

    def preparar_dados_timeline(self):
        """Prepara dados para gráfico de linha do tempo baseado em mês de vencimento"""
        try:
            if not hasattr(self, 'df_despesas') or self.df_despesas.empty:
                return None
            
            agrupamento = self.determinar_agrupamento_temporal()
            df_timeline = self.df_despesas.copy()
            
            # **MODIFICAR: Criar coluna de agrupamento temporal baseado em DT_VENCTO**
            if agrupamento == 'dia':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.date
                formato_periodo = lambda x: x.strftime('%d/%m/%Y')
            elif agrupamento == 'semana':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('W')
                formato_periodo = lambda x: f"Sem {x.week}/{x.year}"
            elif agrupamento == 'mes':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('M')
                formato_periodo = lambda x: f"{x.month:02d}/{x.year}"
            elif agrupamento == 'trimestre':
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('Q')
                formato_periodo = lambda x: f"Q{x.quarter}/{x.year}"
            else:  # ano
                df_timeline['periodo'] = df_timeline['DT_VENCTO'].dt.to_period('Y')
                formato_periodo = lambda x: str(x.year)
            
            # Agrupar por período e categoria
            df_agrupado = df_timeline.groupby(['periodo', 'CATEGORIA'])['VALOR'].sum().reset_index()
            
            # Criar pivot para ter categorias como colunas
            df_pivot = df_agrupado.pivot(index='periodo', columns='CATEGORIA', values='VALOR').fillna(0)
            
            # Garantir que todas as categorias existam
            for categoria in self.categorias_despesas.keys():
                if categoria not in df_pivot.columns:
                    df_pivot[categoria] = 0
            
            # Resetar índice e formatar período
            df_pivot = df_pivot.reset_index()
            df_pivot['periodo_str'] = df_pivot['periodo'].apply(formato_periodo)
            
            # Calcular total por período
            df_pivot['total'] = df_pivot[[cat for cat in self.categorias_despesas.keys() if cat in df_pivot.columns]].sum(axis=1)
            
            return {
                'dados': df_pivot,
                'agrupamento': agrupamento,
                'categorias': list(self.categorias_despesas.keys())
            }
            
        except Exception as e:
            print(f"Erro ao preparar dados de timeline: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def criar_grafico_pizza(self, fig, ax):
        """Cria um gráfico de pizza com as categorias da data selecionada"""
        try:
            # Usar os dados para gráfico de pizza
            df = self.dados_grafico.get('pizza')
            
            if df is None or df.empty:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Filtrar apenas valores maiores que zero
            df = df[df['VALOR'] > 0].copy()
            
            if df.empty:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Ordenar por valor (maior para menor) para melhor visualização
            df = df.sort_values('VALOR', ascending=False)
            
            # Cores para o gráfico
            colors = plt.cm.Set3.colors  # Paleta mais suave
            
            # Calcular percentuais
            total = df['VALOR'].sum()
            df['percentual'] = (df['VALOR'] / total) * 100
            
            # Função para formatar labels com percentual
            def autopct_format(pct):
                return f'{pct:.1f}%' if pct > 2 else ''  # Só mostra % se for maior que 2%
            
            # Criar o gráfico de pizza
            wedges, texts, autotexts = ax.pie(
                df['VALOR'], 
                labels=None,  # Não colocar labels direto nas fatias
                autopct=autopct_format,
                startangle=90,
                colors=colors,
                wedgeprops={'edgecolor': 'w', 'linewidth': 2},
                pctdistance=0.85
            )
            
            # Melhorar aparência dos percentuais
            for autotext in autotexts:
                autotext.set_fontsize(10)
                autotext.set_fontweight('bold')
                autotext.set_color('white')
            
            # Criar legenda ao lado do gráfico com valores
            legend_labels = []
            for _, row in df.iterrows():
                categoria = row['CATEGORIA']
                categoria_desc = self.categorias_despesas.get(categoria, 'Não classificado')
                valor = row['VALOR']
                pct = row['percentual']
                
                # Formato: "MAT - MATERIAL: R$ 22.639,76 (24.1%)"
                valor_formatado = formatar_valor_sem_simbolo(valor)
                legend_labels.append(f"{categoria} - {categoria_desc}: {valor_formatado} ({pct:.1f}%)")
            
            # Adicionar legenda
            ax.legend(
                wedges,
                legend_labels,
                title="Categorias de Despesa",
                loc="center left",
                bbox_to_anchor=(1, 0, 0.5, 1),
                fontsize=9,
                title_fontsize=10
            )
            
            # Adicionar título
            mes_ano_str = self.data_selecionada.strftime('%m/%Y')
            data_final = self.data_selecionada.to_timestamp('M')
            data_final_str = data_final.strftime('%d/%m/%Y')
            
            ax.set_title(
                f'Distribuição por Categoria de Despesa - {data_final_str}', 
                fontsize=13, 
                pad=20,
                fontweight='bold'
            )
            
            # Ajustar layout para não cortar a legenda
            fig.tight_layout()
            
        except Exception as e:
            print(f"Erro ao criar gráfico de pizza: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Mostrar erro no gráfico
            ax.text(0.5, 0.5, f"Erro ao gerar gráfico: {str(e)}", 
                horizontalalignment='center', verticalalignment='center',
                transform=ax.transAxes, fontsize=12, color='red')

    def criar_grafico_barras(self, fig, ax, mostrar_titulo=True):
        """Cria um gráfico de barras com as categorias da data selecionada"""
        try:
            # Usar os dados para gráfico de barras
            df = self.dados_grafico.get('barras')
            
            if df is None or df.empty:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Filtrar apenas valores maiores que zero
            df = df[df['VALOR'] > 0].copy()
            
            if df.empty:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Ordenar por valor (maior para menor)
            df = df.sort_values(by='VALOR', ascending=True)  # ascending=True para barras horizontais
            
            # Criar labels curtos (apenas código da categoria)
            labels_curtos = [row['CATEGORIA'] for _, row in df.iterrows()]
            
            # Cores para o gráfico (mesma paleta do pizza)
            colors = plt.cm.Set3.colors[:len(df)]
            
            # Criar o gráfico de barras horizontais
            bars = ax.barh(labels_curtos, df['VALOR'], color=colors, edgecolor='white', linewidth=1.5)
            
            # Adicionar valores nas barras com formatação
            for bar, cor_barra in zip(bars, colors):
                width = bar.get_width()
                # Posicionar o texto dentro da barra se ela for grande, fora se for pequena
                max_value = df['VALOR'].max()
                if width > max_value * 0.1:  # Se a barra tem mais de 10% do máximo
                    label_x_pos = width / 2
                    ha = 'center'
                    # Cor do texto com contraste adequado à cor real da barra
                    # (paletas pastel como Set3 têm tons claros onde texto branco
                    # fixo fica ilegível — ex.: amarelo claro)
                    color = cor_texto_contraste(cor_barra)
                    weight = 'bold'
                else:  # Barra pequena, colocar valor fora
                    label_x_pos = width + width * 0.02
                    ha = 'left'
                    color = 'black'
                    weight = 'normal'
                
                valor_formatado = formatar_valor_sem_simbolo(width)
                ax.text(label_x_pos, bar.get_y() + bar.get_height()/2, 
                    valor_formatado,
                    va='center', ha=ha, fontsize=9, color=color, fontweight=weight)
            
            # Ajustar formatação do eixo x (valores)
            def format_real(x, pos):
                if x >= 1000:
                    return f'{x/1000:.0f}k'
                return f'{x:.0f}'
            
            ax.xaxis.set_major_formatter(mticker.FuncFormatter(format_real))
            
            # Criar legenda com descrições completas
            legend_labels = []
            for _, row in df.sort_values('VALOR', ascending=False).iterrows():
                categoria = row['CATEGORIA']
                categoria_desc = self.categorias_despesas.get(categoria, 'Não classificado')
                legend_labels.append(f"{categoria} - {categoria_desc}")
            
            # Pegar as cores na ordem correta (maior para menor)
            df_sorted = df.sort_values('VALOR', ascending=False)
            # Cores na mesma ordem da legenda (maior para menor). A lista `colors`
            # está alinhada com `df` (ordem ascendente, mesma das barras), então a
            # ordem descendente usada na legenda é exatamente o inverso dela.
            cores_ordenadas = list(colors)[::-1]
            
            # Criar patches para a legenda
            from matplotlib.patches import Patch
            legend_patches = [Patch(facecolor=cor, edgecolor='white', linewidth=1.5) 
                            for cor in cores_ordenadas]
            
            # Adicionar legenda
            ax.legend(
                legend_patches,
                legend_labels,
                title="Categorias de Despesa",
                loc="lower right",
                fontsize=8,
                title_fontsize=9,
                framealpha=0.95
            )
            
            # Adicionar títulos e labels
            mes_ano_str = self.data_selecionada.strftime('%m/%Y')
            data_final = self.data_selecionada.to_timestamp('M')
            data_final_str = data_final.strftime('%d/%m/%Y')
            
            if mostrar_titulo:
                ax.set_title(
                    f'Valores por Categoria de Despesa - {data_final_str}', 
                    fontsize=13,
                    pad=20,
                    fontweight='bold'
                )
            ax.set_xlabel('Valor', fontsize=11, fontweight='bold')
            ax.set_ylabel('Categoria', fontsize=11, fontweight='bold')
            
            # Adicionar grid leve
            ax.grid(axis='x', linestyle='--', alpha=0.3, zorder=0)
            ax.set_axisbelow(True)  # Grid atrás das barras
            
            # Ajustar margens
            ax.margins(y=0.02)
            
            # Ajustar layout
            fig.tight_layout()
            
        except Exception as e:
            print(f"Erro ao criar gráfico de barras: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Mostrar erro no gráfico
            ax.text(0.5, 0.5, f"Erro ao gerar gráfico: {str(e)}", 
                horizontalalignment='center', verticalalignment='center',
                transform=ax.transAxes, fontsize=12, color='red')

    def criar_grafico_pizza_totais(self, fig, ax):
        """Cria um gráfico de pizza com os totais por categoria"""
        try:
            # Calcular totais por categoria
            totais = {}
            for categoria in self.categorias_despesas.keys():
                if categoria in self.df_por_data.columns:
                    totais[categoria] = self.df_por_data[categoria].sum()
                else:
                    totais[categoria] = 0
            
            # Filtrar categorias com valor > 0
            totais_filtrados = {k: v for k, v in totais.items() if v > 0}
            
            if not totais_filtrados:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Preparar dados e ordenar por valor (maior para menor)
            categorias = list(totais_filtrados.keys())
            valores = list(totais_filtrados.values())
            
            # Ordenar
            dados_ordenados = sorted(zip(categorias, valores), key=lambda x: x[1], reverse=True)
            categorias = [x[0] for x in dados_ordenados]
            valores = [x[1] for x in dados_ordenados]
            
            # Paleta intermediária (Paired - mais balanceada)
            colors = plt.cm.Paired.colors
            
            # Calcular percentuais
            total = sum(valores)
            
            # Função para formatar labels com percentual
            def autopct_format(pct):
                return f'{pct:.1f}%' if pct > 2 else ''
            
            # Criar o gráfico de pizza
            wedges, texts, autotexts = ax.pie(
                valores, 
                labels=None,
                autopct=autopct_format,
                startangle=90,
                colors=colors,
                wedgeprops={'edgecolor': 'w', 'linewidth': 2},
                pctdistance=0.85
            )
            
            # Melhorar aparência dos percentuais
            for autotext in autotexts:
                autotext.set_fontsize(10)
                autotext.set_fontweight('bold')
                autotext.set_color('white')
            
            # Criar legenda ao lado do gráfico
            legend_labels = []
            for cat, val in zip(categorias, valores):
                categoria_desc = self.categorias_despesas.get(cat, 'Não classificado')
                valor_formatado = formatar_valor_sem_simbolo(val)
                pct = (val / total) * 100
                legend_labels.append(f"{cat} - {categoria_desc}: {valor_formatado} ({pct:.1f}%)")
            
            # Adicionar legenda
            ax.legend(
                wedges,
                legend_labels,
                title="Categorias de Despesa",
                loc="center left",
                bbox_to_anchor=(1, 0, 0.5, 1),
                fontsize=9,
                title_fontsize=10
            )
            
            # Adicionar título
            total_geral = sum(valores)
            ax.set_title(
                f'Distribuição Total por Categoria - {formatar_moeda_br(total_geral)}', 
                fontsize=13, 
                pad=20,
                fontweight='bold'
            )
            
            # Ajustar layout
            fig.tight_layout()
            
        except Exception as e:
            print(f"Erro ao criar gráfico de pizza totais: {str(e)}")
            ax.text(0.5, 0.5, f"Erro ao gerar gráfico: {str(e)}", 
                horizontalalignment='center', verticalalignment='center',
                transform=ax.transAxes, fontsize=12, color='red')

    def criar_grafico_barras_totais(self, fig, ax, mostrar_titulo=True):
        """Cria um gráfico de barras com os totais por categoria"""
        try:
            # Calcular totais por categoria
            totais = {}
            for categoria in self.categorias_despesas.keys():
                if categoria in self.df_por_data.columns:
                    totais[categoria] = self.df_por_data[categoria].sum()
                else:
                    totais[categoria] = 0
            
            # Filtrar categorias com valor > 0 e ordenar por valor
            totais_filtrados = {k: v for k, v in totais.items() if v > 0}
            totais_ordenados = dict(sorted(totais_filtrados.items(), key=lambda x: x[1], reverse=False))
            
            if not totais_ordenados:
                ax.text(0.5, 0.5, "Não há dados para exibir", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            # Preparar dados
            categorias = list(totais_ordenados.keys())
            valores = list(totais_ordenados.values())
            
            # Labels curtos (apenas código)
            labels_curtos = categorias
            
            # Paleta intermediária (Paired)
            colors = plt.cm.Paired.colors[:len(categorias)]
            
            # Criar o gráfico de barras horizontais
            bars = ax.barh(labels_curtos, valores, color=colors, edgecolor='white', linewidth=1.5)
            
            # Adicionar valores nas barras
            for bar, cor_barra in zip(bars, colors):
                width = bar.get_width()
                max_value = max(valores)
                
                # Posicionar texto
                if width > max_value * 0.1:
                    label_x_pos = width / 2
                    ha = 'center'
                    # Cor do texto com contraste adequado à cor real da barra
                    color = cor_texto_contraste(cor_barra)
                    weight = 'bold'
                else:
                    label_x_pos = width + width * 0.02
                    ha = 'left'
                    color = 'black'
                    weight = 'normal'
                
                valor_formatado = formatar_valor_sem_simbolo(width)
                ax.text(label_x_pos, bar.get_y() + bar.get_height()/2, 
                    valor_formatado,
                    va='center', ha=ha, fontsize=9, color=color, fontweight=weight)
            
            # Formatação do eixo x (sem símbolo de moeda, alinhado com o restante do relatório)
            def format_real(x, pos):
                if x >= 1000000:
                    return f'{x/1000000:.1f}M'
                elif x >= 1000:
                    return f'{x/1000:.0f}k'
                return f'{x:.0f}'
            
            ax.xaxis.set_major_formatter(mticker.FuncFormatter(format_real))
            
            # Criar legenda com descrições completas
            legend_labels = []
            categorias_ordenadas_desc = sorted(categorias, key=lambda x: totais_ordenados[x], reverse=True)
            
            for cat in categorias_ordenadas_desc:
                categoria_desc = self.categorias_despesas.get(cat, 'Não classificado')
                legend_labels.append(f"{cat} - {categoria_desc}")
            
            # Cores na ordem da legenda (maior para menor)
            indices_ordenados = [categorias.index(cat) for cat in categorias_ordenadas_desc]
            cores_ordenadas = [colors[i] for i in indices_ordenados]
            
            # Criar patches para legenda
            from matplotlib.patches import Patch
            legend_patches = [Patch(facecolor=cor, edgecolor='white', linewidth=1.5) 
                            for cor in cores_ordenadas]
            
            # Adicionar legenda
            ax.legend(
                legend_patches,
                legend_labels,
                title="Categorias de Despesa",
                loc="lower right",
                fontsize=8,
                title_fontsize=9,
                framealpha=0.95
            )
            
            # Adicionar títulos
            total_geral = sum(valores)
            if mostrar_titulo:
                ax.set_title(
                    f'Totais por Categoria - {formatar_moeda_br(total_geral)}', 
                    fontsize=13,
                    pad=20,
                    fontweight='bold'
                )
            ax.set_xlabel('Valor', fontsize=11, fontweight='bold')
            ax.set_ylabel('Categoria', fontsize=11, fontweight='bold')
            
            # Grid leve
            ax.grid(axis='x', linestyle='--', alpha=0.3, zorder=0)
            ax.set_axisbelow(True)
            
            # Ajustar margens
            ax.margins(y=0.02)
            
            # Ajustar layout
            fig.tight_layout()
            
        except Exception as e:
            print(f"Erro ao criar gráfico de barras totais: {str(e)}")
            ax.text(0.5, 0.5, f"Erro ao gerar gráfico: {str(e)}", 
                horizontalalignment='center', verticalalignment='center',
                transform=ax.transAxes, fontsize=12, color='red')

    def criar_grafico_timeline(self, fig, ax):
        """Cria um gráfico de linha do tempo"""
        try:
            dados_timeline = self.preparar_dados_timeline()
            
            if not dados_timeline or dados_timeline['dados'].empty:
                ax.text(0.5, 0.5, "Não há dados suficientes para linha do tempo", 
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=14)
                return
            
            df = dados_timeline['dados']
            agrupamento = dados_timeline['agrupamento']
            categorias = dados_timeline['categorias']
            
            # Cores para cada categoria
            colors = plt.cm.tab10.colors
            
            # Plotar linha para cada categoria
            for i, categoria in enumerate(categorias):
                if categoria in df.columns and df[categoria].sum() > 0:
                    ax.plot(df['periodo_str'], df[categoria], 
                        marker='o', linewidth=2, markersize=4,
                        color=colors[i % len(colors)],
                        label=f"{categoria} - {self.categorias_despesas[categoria]}")
            
            # Plotar linha do total
            ax.plot(df['periodo_str'], df['total'], 
                marker='s', linewidth=3, markersize=5,
                color='black', linestyle='--',
                label='Total Geral')
            
            # Configurar eixos
            ax.set_xlabel(f'Período ({agrupamento.title()})', fontsize=11)
            ax.set_ylabel('Valor (R$)', fontsize=11)
            
            # Formatar eixo Y
            def format_real(x, pos):
                return f'R$ {x:,.0f}'.replace(',', '.')
            ax.yaxis.set_major_formatter(mticker.FuncFormatter(format_real))
            
            # Rotacionar labels do eixo X se necessário
            if len(df) > 10:
                ax.tick_params(axis='x', rotation=45)
            
            # Adicionar grid
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # Adicionar legenda
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
            
            # **MODIFICAR O TÍTULO:**
            # Usar mes_ano_vencto do df_por_data
            mes_inicio = self.df_por_data['mes_ano_vencto'].min().strftime('%m/%Y')
            mes_fim = self.df_por_data['mes_ano_vencto'].max().strftime('%m/%Y')
            ax.set_title(f'Evolução das Despesas por Categoria (Mês de Vencimento)\n{mes_inicio} a {mes_fim}', 
                        fontsize=14, pad=20)
            
            # Ajustar layout
            fig.tight_layout()
            
        except Exception as e:
            print(f"Erro ao criar gráfico de timeline: {str(e)}")
            ax.text(0.5, 0.5, f"Erro ao gerar gráfico: {str(e)}", 
                horizontalalignment='center', verticalalignment='center',
                transform=ax.transAxes, fontsize=12, color='red')
    
    def exportar_excel(self):
        """Exporta o relatório para um arquivo Excel"""
        if not hasattr(self, 'cliente_atual') or not self.cliente_atual:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return
            
        # Solicitar nome do arquivo ao usuário
        data_str = self.data_referencia.strftime('%d-%m-%Y')
        nome_padrao = f"Relatorio_Categoria_{self.cliente_atual}_{data_str}.xlsx"
        
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            initialfile=nome_padrao
        )
        
        if not arquivo:
            return
            
        try:
            # Criar workbook
            wb = Workbook()
            
            # Remover a aba padrão
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Criar aba de resumo
            ws_resumo = wb.create_sheet("Resumo")

            # Logomarca (se disponível) - inserida à direita do cabeçalho, sem sobrepor o texto
            try:
                if LOGO_PATH.exists():
                    from openpyxl.drawing.image import Image as XLImage
                    xl_logo = XLImage(str(LOGO_PATH))
                    largura_alvo_px = 110
                    proporcao = largura_alvo_px / xl_logo.width
                    xl_logo.width = largura_alvo_px
                    xl_logo.height = xl_logo.height * proporcao
                    ws_resumo.add_image(xl_logo, "K1")
            except Exception as e:
                print(f"Aviso: não foi possível inserir o logo no Excel: {e}")

            # Adicionar cabeçalho
            ws_resumo['A1'] = "Relatório por Categoria de Despesa"
            ws_resumo['A1'].font = Font(size=14, bold=True)
            ws_resumo.merge_cells('A1:I1')
            ws_resumo['A1'].alignment = Alignment(horizontal='center')
            
            ws_resumo['A2'] = f"Cliente: {self.cliente_atual}"
            ws_resumo['A2'].font = Font(size=12, bold=True)
            ws_resumo.merge_cells('A2:I2')
            
            ws_resumo['A3'] = f"Data do relatório: {data_str}"
            ws_resumo['A3'].font = Font(size=12)
            ws_resumo.merge_cells('A3:I3')
            
            # Adicionar cabeçalhos da tabela
            headers = ["Data"] + list(self.categorias_despesas.keys()) + ["Total"]
            for col, header in enumerate(headers, start=1):
                cell = ws_resumo.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(fgColor="DDDDDD", fill_type="solid")
            
            # Adicionar dados
            for i, (_, row) in enumerate(self.df_por_data.iterrows(), start=6):
                # Mês/Ano formatado
                mes_ano_str = row['mes_ano_vencto'].strftime('%m/%Y')
                ws_resumo.cell(row=i, column=1, value=mes_ano_str)
                
                # Valores por categoria
                for j, categoria in enumerate(self.categorias_despesas.keys(), start=2):
                    ws_resumo.cell(row=i, column=j, value=row[categoria] if categoria in row else 0)
                    ws_resumo.cell(row=i, column=j).number_format = "#.##0,00"
                
                # Total
                ws_resumo.cell(row=i, column=len(headers), value=row['total'])
                ws_resumo.cell(row=i, column=len(headers)).number_format = "#.##0,00"
            
            # Ajustar largura das colunas
            ws_resumo.column_dimensions['A'].width = 15
            for col in range(2, len(headers) + 1):
                ws_resumo.column_dimensions[get_column_letter(col)].width = 15
            
            # Adicionar totais
            total_row = 6 + len(self.df_por_data)
            
            ws_resumo.cell(row=total_row, column=1, value="TOTAL")
            ws_resumo.cell(row=total_row, column=1).font = Font(bold=True)
            
            # Totais por categoria
            for j, categoria in enumerate(self.categorias_despesas.keys(), start=2):
                formula = f"=SUM({get_column_letter(j)}6:{get_column_letter(j)}{total_row-1})"
                ws_resumo.cell(row=total_row, column=j, value=formula)
                ws_resumo.cell(row=total_row, column=j).font = Font(bold=True)
                ws_resumo.cell(row=total_row, column=j).number_format = "#.##0,00"
            
            # Total geral
            total_formula = f"=SUM({get_column_letter(len(headers))}6:{get_column_letter(len(headers))}{total_row-1})"
            ws_resumo.cell(row=total_row, column=len(headers), value=total_formula)
            ws_resumo.cell(row=total_row, column=len(headers)).font = Font(bold=True)
            ws_resumo.cell(row=total_row, column=len(headers)).number_format = "#.##0,00"
            
            # Criar aba de detalhes se tivermos um mês/ano selecionado
            if hasattr(self, 'mes_ano_selecionado') and self.mes_ano_selecionado:
                ws_detalhes = wb.create_sheet("Detalhes")

                mes_ano_str_detalhe = self.mes_ano_selecionado.strftime('%m/%Y')
                ws_detalhes['A1'] = f"Detalhes do Mês: {mes_ano_str_detalhe}"
                ws_detalhes['A1'].font = Font(size=14, bold=True)
                ws_detalhes.merge_cells('A1:F1')
                ws_detalhes['A1'].alignment = Alignment(horizontal='center')

                # Filtrar dados para o mês selecionado
                self.df_despesas['mes_ano_vencto_temp'] = self.df_despesas['DT_VENCTO'].dt.to_period('M')
                df_filtrado = self.df_despesas[self.df_despesas['mes_ano_vencto_temp'] == self.mes_ano_selecionado].copy()
                df_filtrado['CATEGORIA'] = df_filtrado['CATEGORIA'].fillna('DIV')

                # CABEÇALHOS (coluna "Cat" removida: cada categoria agora forma seu próprio bloco)
                headers = ["Dt. Vencimento", "Nome", "Referência", "Data Relatório", "Valor", "Observação"]
                num_cols = len(headers)

                fill_grupo = PatternFill(fgColor="34495E", fill_type="solid")
                fonte_grupo = Font(bold=True, color="FFFFFF", size=11)
                fill_cabecalho = PatternFill(fgColor="DDDDDD", fill_type="solid")
                fill_subtotal = PatternFill(fgColor="EAECEE", fill_type="solid")
                fill_total_geral = PatternFill(fgColor="2C3E50", fill_type="solid")
                fonte_total_geral = Font(bold=True, color="FFFFFF")

                # Ordem das categorias conforme a legenda
                ordem_categorias = list(self.categorias_despesas.keys())

                linha_atual = 3
                total_geral_mes = 0.0

                for categoria_cod in ordem_categorias:
                    df_cat = df_filtrado[df_filtrado['CATEGORIA'] == categoria_cod].sort_values('DT_VENCTO')
                    if df_cat.empty:
                        continue

                    categoria_nome = self.categorias_despesas.get(categoria_cod, 'NÃO CLASSIFICADO')

                    # Cabeçalho do grupo (linha mesclada, destacada)
                    ws_detalhes.cell(row=linha_atual, column=1, value=f"{categoria_cod} - {categoria_nome}")
                    ws_detalhes.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=num_cols)
                    cell_grupo = ws_detalhes.cell(row=linha_atual, column=1)
                    cell_grupo.font = fonte_grupo
                    cell_grupo.fill = fill_grupo
                    cell_grupo.alignment = Alignment(horizontal='left', vertical='center')
                    linha_atual += 1

                    # Cabeçalho das colunas
                    for col, header in enumerate(headers, start=1):
                        cell = ws_detalhes.cell(row=linha_atual, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = fill_cabecalho
                    linha_atual += 1
                    linha_dados_inicio = linha_atual

                    # Linhas de dados da categoria
                    for _, row in df_cat.iterrows():
                        if pd.notna(row['DT_VENCTO']):
                            ws_detalhes.cell(row=linha_atual, column=1, value=row['DT_VENCTO'])
                            ws_detalhes.cell(row=linha_atual, column=1).number_format = "dd/mm/yyyy"

                        if 'NOME' in row and pd.notna(row['NOME']):
                            ws_detalhes.cell(row=linha_atual, column=2, value=row['NOME'])

                        if 'REFERÊNCIA' in row and pd.notna(row['REFERÊNCIA']):
                            ws_detalhes.cell(row=linha_atual, column=3, value=row['REFERÊNCIA'])

                        if pd.notna(row['DATA_REL']):
                            ws_detalhes.cell(row=linha_atual, column=4, value=row['DATA_REL'])
                            ws_detalhes.cell(row=linha_atual, column=4).number_format = "dd/mm/yyyy"

                        ws_detalhes.cell(row=linha_atual, column=5, value=row['VALOR'])
                        ws_detalhes.cell(row=linha_atual, column=5).number_format = "#.##0,00"

                        if 'OBSERVAÇÃO' in row and pd.notna(row['OBSERVAÇÃO']):
                            ws_detalhes.cell(row=linha_atual, column=6, value=row['OBSERVAÇÃO'])

                        linha_atual += 1

                    linha_dados_fim = linha_atual - 1

                    # Subtotal da categoria
                    ws_detalhes.cell(row=linha_atual, column=4, value=f"Subtotal {categoria_cod}")
                    ws_detalhes.cell(row=linha_atual, column=4).font = Font(bold=True)
                    formula_subtotal = f"=SUM(E{linha_dados_inicio}:E{linha_dados_fim})"
                    ws_detalhes.cell(row=linha_atual, column=5, value=formula_subtotal)
                    ws_detalhes.cell(row=linha_atual, column=5).font = Font(bold=True)
                    ws_detalhes.cell(row=linha_atual, column=5).number_format = "#.##0,00"
                    for col in range(1, num_cols + 1):
                        ws_detalhes.cell(row=linha_atual, column=col).fill = fill_subtotal

                    total_geral_mes += df_cat['VALOR'].sum()
                    linha_atual += 2  # linha em branco entre categorias

                # Total geral do mês (soma de todas as categorias)
                ws_detalhes.cell(row=linha_atual, column=4, value="TOTAL GERAL DO MÊS")
                ws_detalhes.cell(row=linha_atual, column=4).font = fonte_total_geral
                ws_detalhes.cell(row=linha_atual, column=5, value=total_geral_mes)
                ws_detalhes.cell(row=linha_atual, column=5).font = fonte_total_geral
                ws_detalhes.cell(row=linha_atual, column=5).number_format = "#.##0,00"
                for col in range(1, num_cols + 1):
                    ws_detalhes.cell(row=linha_atual, column=col).fill = fill_total_geral
                linha_atual += 3

                # Ajustar largura das colunas
                ws_detalhes.column_dimensions['A'].width = 14
                ws_detalhes.column_dimensions['B'].width = 28
                ws_detalhes.column_dimensions['C'].width = 38
                ws_detalhes.column_dimensions['D'].width = 16
                ws_detalhes.column_dimensions['E'].width = 14
                ws_detalhes.column_dimensions['F'].width = 40

                # LEGENDA DAS CATEGORIAS
                legenda_row = linha_atual
                ws_detalhes.cell(row=legenda_row, column=1, value="Legenda de Categorias:")
                ws_detalhes.cell(row=legenda_row, column=1).font = Font(bold=True)

                for i, (codigo, descricao) in enumerate(self.categorias_despesas.items(), start=1):
                    ws_detalhes.cell(row=legenda_row + i, column=1, value=f"{codigo}:")
                    ws_detalhes.cell(row=legenda_row + i, column=2, value=descricao)
                    ws_detalhes.merge_cells(f'B{legenda_row + i}:E{legenda_row + i}')
            
            # Salvar o arquivo
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso para:\n{arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def exportar_pdf(self):
        """Exporta o relatório para um arquivo PDF"""
        if not hasattr(self, 'cliente_atual') or not self.cliente_atual:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return
            
        # Solicitar nome do arquivo ao usuário
        data_str = self.data_referencia.strftime('%d-%m-%Y')
        nome_padrao = f"Relatorio_Categoria_{self.cliente_atual}_{data_str}.pdf"
        
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Arquivos PDF", "*.pdf")],
            initialfile=nome_padrao
        )
        
        if not arquivo:
            return
            
        try:
            # Verificar se temos a biblioteca reportlab disponível
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable, KeepTogether, PageBreak
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.lib.enums import TA_RIGHT
                import matplotlib.pyplot as plt
                import io
            except ImportError:
                messagebox.showerror("Erro", "Biblioteca ReportLab não encontrada. Por favor instale usando 'pip install reportlab'.")
                return
            
            # Criar documento PDF
            doc = SimpleDocTemplate(arquivo, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            heading1_style = styles['Heading1']
            heading2_style = styles['Heading2']
            normal_style = styles['Normal']
            
            # Cabeçalho: logo à esquerda, dados da empresa à direita
            # (mesmo padrão visual do Relatório Quinzenal de Medições)
            logo_img = obter_imagem_logo_reportlab(1.6*inch, 0.85*inch)

            empresa_style = ParagraphStyle(
                'EmpresaInfo',
                parent=normal_style,
                fontSize=7.5,
                leading=10,
                alignment=TA_RIGHT,
                textColor=colors.HexColor('#444444')
            )
            empresa_text = (
                f"{EMPRESA_INFO['endereco']}<br/>"
                f"{EMPRESA_INFO['fones']}<br/>"
                f"{EMPRESA_INFO['email']}"
            )
            empresa_paragraph = Paragraph(empresa_text, empresa_style)

            celula_logo = logo_img if logo_img is not None else ""
            header_table = Table(
                [[celula_logo, empresa_paragraph]],
                colWidths=[2.2*inch, 4.3*inch]
            )
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
                ('RIGHTPADDING', (1, 0), (1, 0), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 0.1*inch))
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#34495E')))
            story.append(Spacer(1, 0.15*inch))

            # Título
            story.append(Paragraph(f"Relatório por Categoria de Despesa", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Informações do cliente
            story.append(Paragraph(f"Cliente: {self.cliente_atual}", heading1_style))
            story.append(Paragraph(f"Data do relatório: {data_str}", normal_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Resumo por data
            story.append(Paragraph("Resumo por Data", heading2_style))
            story.append(Spacer(1, 0.1*inch))
            
            # Cabeçalhos
            headers = ["Mês/Ano"] + list(self.categorias_despesas.keys()) + ["Total (R$)"] 
            
            # Dados da tabela de resumo
            table_data = [headers]
            
            for _, row in self.df_por_data.iterrows():
                # Formatar data
                mes_ano_str = row['mes_ano_vencto'].strftime('%m/%Y')
                
                # Preparar valores para cada categoria
                valores = [mes_ano_str]  # <-- Usar mes_ano_str em vez de data_str
                for categoria in self.categorias_despesas.keys():
                    valor_formatado = f"{row[categoria]:,.2f}".replace(',', '.').replace('.', ',') if categoria in row else "0,00"
                    valores.append(valor_formatado)
                
                # Adicionar total
                total_formatado = f"{row['total']:,.2f}".replace(',', '.').replace('.', ',')
                valores.append(total_formatado)
                
                table_data.append(valores)
            
            # Adicionar linha de total
            if not self.df_por_data.empty:
                total_row = ["TOTAL"]
                for categoria in self.categorias_despesas.keys():
                    total_cat = self.df_por_data[categoria].sum() if categoria in self.df_por_data.columns else 0
                    total_formatado = f"{total_cat:,.2f}".replace(',', '.').replace('.', ',')
                    total_row.append(total_formatado)
                
                # Total geral
                total_geral = self.df_por_data['total'].sum()
                total_geral_formatado = f"{total_geral:,.2f}".replace(',', '.').replace('.', ',')
                total_row.append(total_geral_formatado)
                
                table_data.append(total_row)
            
            # Criar tabela de resumo
            # Calcular larguras de colunas
            num_categorias = len(self.categorias_despesas)
            col_width_data = 1.0*inch  # Data
            col_width_cat = (6.0*inch) / num_categorias  # Dividir espaço restante entre categorias
            col_width_total = 1.0*inch  # Total
            
            col_widths = [col_width_data] + [col_width_cat] * num_categorias + [col_width_total]
            
            resumo_table = Table(table_data, colWidths=col_widths)
            
            # Estilo da tabela
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),  # Fonte menor para caber
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ])
            
            # Destacar a linha de total
            if not self.df_por_data.empty:
                table_style.add('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey)
                table_style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
            
            resumo_table.setStyle(table_style)
            story.append(resumo_table)
            story.append(Spacer(1, 0.2*inch))
            
            # Se tiver um mês/ano selecionado, adicionar detalhes
            if hasattr(self, 'mes_ano_selecionado') and self.mes_ano_selecionado:
                # Detalhes do mês selecionado
                mes_ano_str_detalhe = self.mes_ano_selecionado.strftime('%m/%Y')
                story.append(Paragraph(f"Detalhes - Mês: {mes_ano_str_detalhe}", heading2_style))
                story.append(Spacer(1, 0.1*inch))
                
                # Filtrar dados
                self.df_despesas['mes_ano_vencto_temp'] = self.df_despesas['DT_VENCTO'].dt.to_period('M')
                df_filtrado = self.df_despesas[self.df_despesas['mes_ano_vencto_temp'] == self.mes_ano_selecionado].copy()
                
                if not df_filtrado.empty:
                    # Criar estilo com fonte pequena e consistente
                    style_pequeno = ParagraphStyle('Pequeno', parent=normal_style, fontSize=7, leading=8)

                    # Garantir categoria preenchida
                    df_filtrado['CATEGORIA'] = df_filtrado['CATEGORIA'].fillna('DIV')

                    # Ordenar as categorias conforme a ordem definida em self.categorias_despesas
                    ordem_categorias = list(self.categorias_despesas.keys())

                    # Largura das colunas do bloco de detalhes (sem a coluna "Cat", pois cada
                    # categoria agora vira um bloco/tabela própria com cabeçalho de grupo)
                    col_widths = [0.9*inch, 2.1*inch, 2.6*inch, 0.9*inch]

                    total_geral_mes = 0.0

                    for categoria_cod in ordem_categorias:
                        df_cat = df_filtrado[df_filtrado['CATEGORIA'] == categoria_cod]
                        if df_cat.empty:
                            continue

                        # Ordenar lançamentos da categoria por data de vencimento
                        df_cat = df_cat.sort_values('DT_VENCTO')

                        categoria_nome = self.categorias_despesas.get(categoria_cod, 'NÃO CLASSIFICADO')

                        # Linha 0: cabeçalho do grupo (mesclada) | Linha 1: cabeçalho das colunas
                        headers = ["Dt Venc", "Nome", "Referência", "Valor"]
                        table_data = [
                            [f"{categoria_cod} - {categoria_nome}", "", "", ""],
                            headers,
                        ]

                        for _, row in df_cat.iterrows():
                            # Data vencimento
                            dt_vencto = row['DT_VENCTO'].strftime('%d/%m/%y') if pd.notna(row['DT_VENCTO']) else ''

                            # Nome e Referência
                            nome = row.get('NOME', '') if pd.notna(row.get('NOME', '')) else ''
                            referencia = row.get('REFERÊNCIA', '') if pd.notna(row.get('REFERÊNCIA', '')) else ''

                            # QUEBRAR TEXTO LONGO com estilo consistente
                            if len(nome) > 30:
                                nome = Paragraph(nome, style_pequeno)
                            if len(referencia) > 35:
                                referencia = Paragraph(referencia, style_pequeno)

                            # Valor SEM R$
                            valor = f"{row['VALOR']:,.2f}".replace(',', '.').replace('.', ',', 1)

                            table_data.append([dt_vencto, nome, referencia, valor])

                        # Subtotal da categoria
                        subtotal_cat = df_cat['VALOR'].sum()
                        total_geral_mes += subtotal_cat
                        subtotal_formatado = f"{subtotal_cat:,.2f}".replace(',', '.').replace('.', ',', 1)
                        table_data.append(["", "", f"Subtotal {categoria_cod}", subtotal_formatado])

                        detalhes_table = Table(table_data, colWidths=col_widths, repeatRows=2)

                        table_style = TableStyle([
                            # Cabeçalho de grupo (linha 0) - mesclado e destacado
                            ('SPAN', (0, 0), (-1, 0)),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 9),
                            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                            ('LEFTPADDING', (0, 0), (0, 0), 6),
                            ('TOPPADDING', (0, 0), (-1, 0), 5),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),

                            # Cabeçalho das colunas (linha 1)
                            ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),
                            ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),
                            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),

                            # Corpo da tabela
                            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                            ('ALIGN', (1, 1), (2, -1), 'LEFT'),
                            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('FONTSIZE', (0, 1), (-1, -1), 7),
                            ('BOTTOMPADDING', (0, 1), (-1, 1), 6),
                            ('GRID', (0, 1), (-1, -1), 0.5, colors.grey),
                            ('WORDWRAP', (0, 0), (-1, -1), True),

                            # Linha de subtotal (última linha)
                            ('SPAN', (0, -1), (1, -1)),
                            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EAECEE')),
                            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                        ])

                        detalhes_table.setStyle(table_style)
                        story.append(detalhes_table)
                        story.append(Spacer(1, 0.15*inch))

                    # Total geral do mês (soma de todas as categorias)
                    total_geral_formatado = f"{total_geral_mes:,.2f}".replace(',', '.').replace('.', ',', 1)
                    total_geral_table = Table(
                        [["TOTAL GERAL DO MÊS", total_geral_formatado]],
                        colWidths=[col_widths[0] + col_widths[1] + col_widths[2], col_widths[3]]
                    )
                    total_geral_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2C3E50')),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                        ('LEFTPADDING', (0, 0), (0, 0), 6),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    story.append(total_geral_table)
                    story.append(Spacer(1, 0.2*inch))
                    
                    # LEGENDA DAS CATEGORIAS
                    legenda_style = ParagraphStyle(
                        'Legenda',
                        parent=normal_style,
                        fontSize=8,
                        textColor=colors.grey
                    )
                    
                    story.append(Spacer(1, 0.15*inch))
                    story.append(Paragraph("Legenda de Categorias:", legenda_style))
                    story.append(Spacer(1, 0.03*inch))
                    
                    legenda_text = " | ".join([f"{cod}: {desc}" for cod, desc in self.categorias_despesas.items()])
                    story.append(Paragraph(legenda_text, legenda_style))
                    story.append(Spacer(1, 0.2*inch))
                    
                    # Adicionar gráficos: um do mês selecionado e outro do total acumulado
                    # do período, ambos em barras horizontais (mais legível que pizza
                    # quando há muitas categorias com valores pequenos).

                    # Garantir que os dados de gráfico de barras estejam atualizados
                    # para o mês selecionado (independente de o usuário ter passado
                    # pela tela de seleção antes de exportar)
                    self.preparar_grafico_data_selecionada(df_filtrado)
                    if not hasattr(self, 'data_selecionada') or not self.data_selecionada:
                        self.data_selecionada = self.mes_ano_selecionado

                    # Iniciar os gráficos em página nova, com tamanho reduzido o
                    # suficiente para os dois caberem juntos numa única página
                    story.append(PageBreak())

                    # --- Gráfico 1: barras do mês selecionado ---
                    # (mostrar_titulo=False porque o valor total do mês já está no
                    # título da seção, logo abaixo — evita repetir a mesma informação)
                    fig_mes, ax_mes = plt.subplots(figsize=(7, 4))
                    self.criar_grafico_barras(fig_mes, ax_mes, mostrar_titulo=False)

                    img_buffer_mes = io.BytesIO()
                    fig_mes.savefig(img_buffer_mes, format='png', dpi=100, bbox_inches='tight')
                    img_buffer_mes.seek(0)
                    plt.close(fig_mes)

                    titulo_grafico_mes = (
                        f"Gráfico - Categorias do Mês {mes_ano_str_detalhe} "
                        f"— Total: {formatar_moeda_br(total_geral_mes)}"
                    )

                    # --- Gráfico 2: barras do total acumulado no período (todos os meses) ---
                    mes_inicio_periodo = self.df_por_data['mes_ano_vencto'].min().strftime('%m/%Y')
                    mes_fim_periodo = self.df_por_data['mes_ano_vencto'].max().strftime('%m/%Y')
                    total_periodo = self.df_por_data['total'].sum()

                    fig_total, ax_total = plt.subplots(figsize=(7, 4))
                    self.criar_grafico_barras_totais(fig_total, ax_total, mostrar_titulo=False)

                    img_buffer_total = io.BytesIO()
                    fig_total.savefig(img_buffer_total, format='png', dpi=100, bbox_inches='tight')
                    img_buffer_total.seek(0)
                    plt.close(fig_total)

                    titulo_grafico_total = (
                        f"Gráfico - Total Acumulado ({mes_inicio_periodo} a {mes_fim_periodo}) "
                        f"— Total: {formatar_moeda_br(total_periodo)}"
                    )

                    # Ambos os gráficos num único bloco, para o ReportLab tentar
                    # manter os dois juntos na mesma página
                    story.append(KeepTogether([
                        Paragraph(titulo_grafico_mes, heading2_style),
                        Spacer(1, 0.1*inch),
                        Image(img_buffer_mes, width=6.3*inch, height=3.6*inch, kind='proportional'),
                        Spacer(1, 0.25*inch),
                        Paragraph(titulo_grafico_total, heading2_style),
                        Spacer(1, 0.1*inch),
                        Image(img_buffer_total, width=6.3*inch, height=3.6*inch, kind='proportional'),
                    ]))
                else:
                    story.append(Paragraph("Não há lançamentos para esta data.", normal_style))
            
            # Construir o PDF
            doc.build(story)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso para:\n{arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para PDF: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def voltar_menu(self):
        """Volta ao menu principal"""
        self.root.destroy()
        
        # Mostrar janela principal
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()

def main():
    """Função principal para executar o módulo de forma independente"""
    app = RelatorioCategoria()
    app.root.mainloop()
    
if __name__ == "__main__":
    main()