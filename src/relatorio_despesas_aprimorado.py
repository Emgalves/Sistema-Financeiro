import sys
import os
import pandas as pd
import xlwings as xw
import openpyxl
import warnings
import platform
import subprocess
import tkinter as tk
import numpy as np
import tempfile
from tkinter import Tk
from openpyxl import load_workbook
from tkinter import ttk, messagebox, filedialog, StringVar, Toplevel, BooleanVar, Scrollbar, Text
from tkcalendar import Calendar
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    PageTemplate, Frame, Spacer, PageBreak, Image
)
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from reportlab.platypus import KeepTogether
import logging

# Criar diretório para logs se não existir
log_dir = 'logs'
os.makedirs(log_dir, exist_ok=True)

# Nome do arquivo de log com data
log_file = os.path.join(log_dir, f'relatorio_{datetime.now().strftime("%Y%m%d")}.log')

# Configuração do logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()  # Para exibir também no console
    ]
)

# Criar logger específico para o módulo
logger = logging.getLogger(__name__)

# Configuração inicial
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# VARIÁVEL GLOBAL PARA MENU PRINCIPAL
_MENU_PRINCIPAL_GLOBAL = None

def definir_menu_principal(menu):
    """Define o menu principal globalmente"""
    global _MENU_PRINCIPAL_GLOBAL
    _MENU_PRINCIPAL_GLOBAL = menu
    print(f"✅ Menu principal definido: {menu}")

def obter_menu_principal():
    """Obtém o menu principal global"""
    global _MENU_PRINCIPAL_GLOBAL
    return _MENU_PRINCIPAL_GLOBAL

# Variáveis globais
arquivo_path = None
arquivo_selecionado = None
data_selecionada = None
incluir_futuros = None
status_label = None
root = None
handler = None

def aplicar_configuracoes_externas():
    """Aplica configurações passadas por arquivo externo"""
    try:
        import sys
        import json
        
        # Verificar se foi passado arquivo de configuração
        if '--config' in sys.argv:
            config_index = sys.argv.index('--config')
            if config_index + 1 < len(sys.argv):
                config_file = sys.argv[config_index + 1]
                
                # Carregar configurações
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                return config
                
    except Exception as e:
        print(f"Erro ao carregar configurações: {str(e)}")
    
    return None

class RelatorioUI:
    def __init__(self, parent):
        logger.info("Iniciando RelatorioUI")
        if parent is None:
            self.root = tk.Tk()
        else:
            self.root = parent
            
        logger.debug(f"Parent configurado: {parent}")
        
        self.arquivo_selecionado = StringVar(self.root, value="Nenhum arquivo selecionado")
        self.data_selecionada = StringVar(self.root, value=datetime.now().strftime('%d/%m/%Y'))
        logger.debug(f"Data inicial configurada: {self.data_selecionada.get()}")
        
        self.incluir_futuros = BooleanVar(value=True)
        self.status_label = None
        self.handler = RelatorioHandler()
        self.arquivos_lote = []
        self.menu_principal = None  # Adicionado aqui, antes do setup_ui
        self.setup_ui()

    def setup_ui(self):
        logger.info("Iniciando configuração da interface")

        self.root.title("Gerador de Relatório de Despesas")
        self.root.geometry("850x1000")
        self.root.update_idletasks()


        # Container principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Data
        logger.debug("Configurando campos de data")
        frame_data = ttk.Frame(main_frame)
        frame_data.pack(pady=10, padx=20, fill='x')
        
        self.data_selecionada.set(datetime.now().strftime('%d/%m/%Y'))
        self.arquivo_selecionado.set("Nenhum arquivo selecionado")
        ttk.Label(frame_data, text="Data do relatório:").pack(side='left', padx=(0, 10))
        ttk.Label(frame_data, textvariable=self.data_selecionada, width=10).pack(side='left')
        ttk.Button(frame_data, text="Escolher Data", command=self.escolher_data).pack(side='left', padx=5)

        # Relatório Individual
        frame_arquivo = ttk.LabelFrame(main_frame, text="Relatório Individual")
        frame_arquivo.pack(pady=10, padx=20, fill='x')

        self.arquivo_selecionado.set("Nenhum arquivo selecionado")
        ttk.Button(frame_arquivo, text="Escolher arquivo", 
                  command=self.selecionar_arquivo_local).pack(pady=5, fill='x')
        ttk.Label(frame_arquivo, textvariable=self.arquivo_selecionado).pack(pady=5)
        button_frame = ttk.Frame(frame_arquivo)
        button_frame.pack(pady=5, fill='x')

        ttk.Button(button_frame, text="Gerar com Preview",
                command=self.gerar_relatorio).pack(side='left', padx=(0, 5), fill='x', expand=True)
        ttk.Button(button_frame, text="Gerar Direto",
                command=self.gerar_relatorio_sem_preview).pack(side='left', padx=(5, 0), fill='x', expand=True)

        # Relatório em Lote
        frame_lote = ttk.LabelFrame(main_frame, text="Relatório em Lote")
        frame_lote.pack(pady=10, padx=20, fill='x')
        ttk.Button(frame_lote, text="Selecionar Arquivos para Lote", 
                  command=self.selecionar_arquivos_lote).pack(pady=5, fill='x')

        # Checkbox para lançamentos futuros
        ttk.Checkbutton(main_frame, text="Incluir lançamentos futuros",
                       variable=self.incluir_futuros).pack(pady=10, anchor='w')

        self.incluir_excluidos = BooleanVar(value=False)
        ttk.Checkbutton(main_frame, text="Incluir lançamentos excluídos no relatório",
                    variable=self.incluir_excluidos).pack(pady=5, anchor='w')


        # Status label
        self.status_label = ttk.Label(main_frame, text="", wraplength=350)
        self.status_label.pack(pady=10)

        # Adicione esta linha ao final do método:
        self.adicionar_botao_pendentes()

        logger.info("Interface configurada com sucesso")

        

    def escolher_data(self):
        top = Toplevel(self.root)
        top.title("Selecione a Data")
        
        x = self.root.winfo_x() + 50
        y = self.root.winfo_y() + 50
        top.geometry(f"+{x}+{y}")
        
        cal = Calendar(top,
                      selectmode='day',
                      year=datetime.now().year,
                      month=datetime.now().month,
                      day=datetime.now().day,
                      locale='pt_BR',
                      date_pattern='dd/mm/yyyy')
        cal.pack(padx=10, pady=10)
        
        def definir_data():
            data = cal.get_date()
            self.data_selecionada.set(data)
            top.destroy()
            
        ttk.Button(top, text="Confirmar", command=definir_data).pack(pady=5)
        top.transient(self.root)
        top.grab_set()
        self.root.wait_window(top)

        

    def selecionar_arquivo_local(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.arquivo_path = arquivo
            nome_arquivo = os.path.basename(arquivo)
            self.arquivo_selecionado.set(nome_arquivo)
            self.root.update_idletasks()

    def selecionar_arquivos_lote(self):
        files = filedialog.askopenfilenames(
            title="Selecione os arquivos Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if files:
            self.arquivos_lote = files
            self.processar_lote(files)

    def adicionar_filtro_status_relatorio(self):
        """Adiciona filtro para incluir/excluir lançamentos excluídos no relatório"""
        # Adicionar ao frame de filtros existente
        frame_status = ttk.Frame(self.root)
        frame_status.pack(pady=5, padx=20, fill='x')
        
        self.incluir_excluidos = BooleanVar(value=False)
        ttk.Checkbutton(
            frame_status, 
            text="Incluir lançamentos excluídos no relatório",
            variable=self.incluir_excluidos
        ).pack(anchor='w')

    def gerar_relatorio(self):
        """Versão corrigida do método gerar_relatorio com preview"""
        try:
            if not self.arquivo_path:
                logger.warning("Tentativa de gerar relatório sem arquivo selecionado")
                self.status_label.config(text="Selecione um arquivo Excel!")
                return

            logger.info(f"Iniciando geração de relatório para arquivo: {self.arquivo_path}")
            data_rel = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
            
            # CORREÇÃO: Verificar se deve incluir excluídos
            incluir_excluidos = hasattr(self, 'incluir_excluidos') and self.incluir_excluidos.get()
            logger.info(f"Incluir excluídos: {incluir_excluidos}")
                
            # CORREÇÃO: Carregar dados passando o parâmetro incluir_excluidos
            df = self.handler.carregar_dados_excel(self.arquivo_path, incluir_excluidos)
            
            # CORREÇÃO: Processar dados passando o parâmetro incluir_excluidos
            df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2 = self.handler.processar_dados(
                df, data_rel, incluir_excluidos
            )
                
            # CORREÇÃO: Processar lançamentos futuros - USAR O MÉTODO CORRETO
            df_futuro = None
            if self.incluir_futuros.get():
                # Verificar se o método existe antes de chamar
                if hasattr(self.handler, 'processar_lancamentos_futuros'):
                    df_futuro = self.handler.processar_lancamentos_futuros(df, data_rel, incluir_excluidos)
                else:
                    logger.warning("Método processar_lancamentos_futuros não encontrado, pulando lançamentos futuros")
                    df_futuro = None
                    
            # Processar workbook
            workbook = load_workbook(self.arquivo_path, data_only=True)
            ws_resumo = workbook['RESUMO']
            nome_cliente = ws_resumo['A3'].value
                
            # CORREÇÃO: Obter número do relatório e valor acumulado passando o parâmetro incluir_excluidos
            numero_relatorio = self.handler.obter_numero_relatorio(ws_resumo, data_rel)
            valor_acumulado = self.handler.calcular_acumulado_dados(df, data_rel, incluir_excluidos)
            
            logger.info(f"Número do relatório: {numero_relatorio}")
            logger.info(f"Valor acumulado calculado: {valor_acumulado:,.2f}")
                
            dados_completos = {
                'df_filtrado': df_filtrado,
                'df_diaria': df_diaria,
                'df_tp_desp_1': df_tp_desp_1,
                'df_tp_desp_2': df_tp_desp_2,
                'df_futuro': df_futuro,
                'df_original': df,
                'incluir_futuros': self.incluir_futuros.get(),
                'incluir_excluidos': incluir_excluidos,
                'data_relatorio': data_rel,
                'nome_cliente': nome_cliente,
                'endereco_cliente': ws_resumo['A4'].value,
                'numero_relatorio': numero_relatorio,
                'acumulado': valor_acumulado
            }
            
            logger.debug("Verificando dados antes de mostrar preview:")
            logger.debug(f"dados_completos['acumulado']: {dados_completos['acumulado']}")
            logger.debug(f"Tipo do acumulado: {type(dados_completos['acumulado'])}")
            
            # NOVO: Mostrar preview antes de gerar o PDF final
            visualizador = VisualizadorRelatorio(self.root)
            
            # Passar referência ao arquivo_path para o visualizador
            visualizador.arquivo_path = self.arquivo_path
            
            preview_window = visualizador.mostrar_preview(dados_completos)
            
            self.status_label.config(text=f"Preview do relatório exibido para {nome_cliente}")
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório: {str(e)}", exc_info=True)
            self.status_label.config(text=f"Erro: {str(e)}")

    def gerar_relatorio_sem_preview(self):
        """Método corrigido sem preview - para casos especiais"""
        try:
            if not self.arquivo_path:
                logger.warning("Tentativa de gerar relatório sem arquivo selecionado")
                self.status_label.config(text="Selecione um arquivo Excel!")
                return

            logger.info(f"Iniciando geração de relatório para arquivo: {self.arquivo_path}")
            data_rel = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
            
            # CORREÇÃO: Verificar se deve incluir excluídos
            incluir_excluidos = hasattr(self, 'incluir_excluidos') and self.incluir_excluidos.get()
            logger.info(f"Incluir excluídos: {incluir_excluidos}")
                
            # CORREÇÃO: Carregar dados passando o parâmetro incluir_excluidos
            df = self.handler.carregar_dados_excel(self.arquivo_path, incluir_excluidos)
            
            # CORREÇÃO: Processar dados passando o parâmetro incluir_excluidos
            df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2 = self.handler.processar_dados(
                df, data_rel, incluir_excluidos
            )
                
            # CORREÇÃO: Processar lançamentos futuros - USAR O MÉTODO CORRETO
            df_futuro = None
            if self.incluir_futuros.get():
                # Verificar se o método existe antes de chamar
                if hasattr(self.handler, 'processar_lancamentos_futuros'):
                    df_futuro = self.handler.processar_lancamentos_futuros(df, data_rel, incluir_excluidos)
                else:
                    logger.warning("Método processar_lancamentos_futuros não encontrado, pulando lançamentos futuros")
                    df_futuro = None
                    
            # Processar workbook
            workbook = load_workbook(self.arquivo_path, data_only=True)
            ws_resumo = workbook['RESUMO']
            nome_cliente = ws_resumo['A3'].value
                
            # CORREÇÃO: Obter número do relatório e valor acumulado passando o parâmetro incluir_excluidos
            numero_relatorio = self.handler.obter_numero_relatorio(ws_resumo, data_rel)
            valor_acumulado = self.handler.calcular_acumulado_dados(df, data_rel, incluir_excluidos)
            
            logger.info(f"Número do relatório: {numero_relatorio}")
            logger.info(f"Valor acumulado calculado: {valor_acumulado:,.2f}")
                
            dados_completos = {
                'df_filtrado': df_filtrado,
                'df_diaria': df_diaria,
                'df_tp_desp_1': df_tp_desp_1,
                'df_tp_desp_2': df_tp_desp_2,
                'df_futuro': df_futuro,
                'df_original': df,
                'incluir_futuros': self.incluir_futuros.get(),
                'incluir_excluidos': incluir_excluidos,
                'data_relatorio': data_rel,
                'nome_cliente': nome_cliente,
                'endereco_cliente': ws_resumo['A4'].value,
                'numero_relatorio': numero_relatorio,
                'acumulado': valor_acumulado
            }
            
            logger.debug("Verificando dados antes de gerar PDF:")
            logger.debug(f"dados_completos['acumulado']: {dados_completos['acumulado']}")
            logger.debug(f"Tipo do acumulado: {type(dados_completos['acumulado'])}")
            
            # Gerar nome do arquivo
            data_formatada = data_rel.strftime('%d-%m-%Y')
            nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
            
            # CORREÇÃO: Adicionar sufixo se incluir excluídos
            if incluir_excluidos:
                nome_arquivo = nome_arquivo.replace('.pdf', ' (com excluídos).pdf')
                
            caminho_output = os.path.join(os.path.dirname(self.arquivo_path), nome_arquivo)
            
            # Gerar o PDF com os dados completos
            self.handler.gerar_relatorio_pdf(dados_completos, caminho_output, self.arquivo_path)
            
            self.status_label.config(text=f"Relatório gerado com sucesso para {nome_cliente}")
            self.criar_dialog_relatorio_gerado(nome_cliente, data_formatada)
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório: {str(e)}", exc_info=True)
            self.status_label.config(text=f"Erro: {str(e)}")

    def processar_lote(self, arquivos):
        """Processa arquivos em lote - VERSÃO CORRIGIDA"""
        try:
            logger.info(f"Iniciando processamento em lote de {len(arquivos)} arquivos")
            
            # CORREÇÃO: Verificar se deve incluir excluídos
            incluir_excluidos = hasattr(self, 'incluir_excluidos') and self.incluir_excluidos.get()
            logger.info(f"Lote - Incluir excluídos: {incluir_excluidos}")

            progress_window = Toplevel(self.root)
            progress_window.title("Gerando Relatórios em Lote")
            progress_window.geometry("600x400")
            progress_window.transient(self.root)

            # Label para mostrar progresso
            progress_label = ttk.Label(progress_window, text="Processando...", font=('Helvetica', 10))
            progress_label.pack(pady=10)

            # Barra de progresso
            progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
            progress_bar.pack(pady=10)

            # Lista para mostrar arquivos processados
            lista_processados = tk.Listbox(progress_window, width=50, height=10)
            lista_processados.pack(pady=10, padx=10)

            # Configurar barra de progresso
            total_arquivos = len(arquivos)
            progress_bar['maximum'] = total_arquivos
            
            # Processar cada arquivo
            for i, arquivo in enumerate(arquivos, 1):
                try:
                    arquivo_nome = os.path.basename(arquivo)
                    progress_label.config(text=f"Processando: {arquivo_nome}")
                    progress_bar['value'] = i
                    
                    wb = load_workbook(arquivo, data_only=True)
                    try:
                        ws_resumo = wb['RESUMO']
                        nome_cliente = ws_resumo['A3'].value
                        logger.debug(f"Cliente: {nome_cliente}")
                        
                        data_rel = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
                        
                        # CORREÇÃO: Carregar dados usando parâmetro incluir_excluidos
                        df = self.handler.carregar_dados_excel(arquivo, incluir_excluidos)
                        df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2 = self.handler.processar_dados(
                            df, data_rel, incluir_excluidos
                        )
                        
                        # CORREÇÃO: Processar lançamentos futuros
                        df_futuro = None
                        if self.incluir_futuros.get():
                            # Verificar se o método existe antes de chamar
                            if hasattr(self.handler, 'processar_lancamentos_futuros'):
                                df_futuro = self.handler.processar_lancamentos_futuros(df, data_rel, incluir_excluidos)
                            else:
                                logger.warning("Método processar_lancamentos_futuros não encontrado")
                                df_futuro = None
                        
                        # CORREÇÃO: Obter valor acumulado
                        numero_relatorio = self.handler.obter_numero_relatorio(ws_resumo, data_rel)
                        valor_acumulado = self.handler.calcular_acumulado_dados(df, data_rel, incluir_excluidos)
                        
                        logger.info(f"Arquivo: {arquivo_nome}")
                        logger.info(f"Número do relatório: {numero_relatorio}")
                        logger.info(f"Valor acumulado calculado: {valor_acumulado:,.2f}")
                        
                        dados_completos = {
                            'df_filtrado': df_filtrado,
                            'df_diaria': df_diaria,
                            'df_tp_desp_1': df_tp_desp_1,
                            'df_tp_desp_2': df_tp_desp_2,
                            'df_futuro': df_futuro,
                            'df_original': df,
                            'incluir_futuros': self.incluir_futuros.get(),
                            'incluir_excluidos': incluir_excluidos,
                            'data_relatorio': data_rel,
                            'nome_cliente': nome_cliente,
                            'endereco_cliente': ws_resumo['A4'].value,
                            'numero_relatorio': numero_relatorio,
                            'acumulado': valor_acumulado
                        }
                        
                        # Gerar nome do arquivo
                        data_formatada = data_rel.strftime('%d-%m-%Y')
                        nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
                        
                        # CORREÇÃO: Adicionar sufixo se incluir excluídos
                        if incluir_excluidos:
                            nome_arquivo = nome_arquivo.replace('.pdf', ' (com excluídos).pdf')
                            
                        caminho_output = os.path.join(os.path.dirname(arquivo), nome_arquivo)
                        
                        # Gerar relatório
                        self.handler.gerar_relatorio_pdf(dados_completos, caminho_output, arquivo)
                        
                        lista_processados.insert(tk.END, f"✓ {arquivo_nome} - Concluído")
                        lista_processados.see(tk.END)
                        
                    finally:
                        wb.close()

                except Exception as e:
                    logger.error(f"Erro ao processar arquivo {arquivo_nome}: {str(e)}", exc_info=True)
                    lista_processados.insert(tk.END, f"✗ {arquivo_nome} - Erro: {str(e)}")

                # Atualizar interface
                progress_window.update()

            # Finalização
            progress_label.config(text="Processamento concluído!")
            ttk.Button(
                progress_window, 
                text="Fechar", 
                command=lambda: self.criar_dialog_relatorio_gerado(None, None) or progress_window.destroy()
            ).pack(pady=10)
            
        except Exception as e:
            logger.error(f"Erro no processamento em lote: {str(e)}", exc_info=True)
            raise
        
    def gerar_relatorio_lote():
        try:
            # Verificar se há arquivos selecionados
            if not self.arquivo_path:  # Usar self em vez de variável global
                self.status_label.config(text="Selecione um arquivo Excel!")
                return
            
            processar_lote(arquivos_selecionados)


            status_label.config(text="Relatórios em lote gerados com sucesso!")

            # Criar diálogo após gerar os relatórios em lote
            # criar_dialog_relatorio_gerado(None, None)

        except Exception as e:
            erro = str(e)
            print(f"Erro ao gerar relatórios em lote: {erro}")
            status_label.config(text=f"Erro: {erro}")


    def criar_dialog_relatorio_gerado(self, nome_cliente, data_formatada):
        """VERSÃO CORRIGIDA do diálogo pós-geração"""
        dialog = Toplevel(self.root)
        dialog.title("Relatório Gerado")
        dialog.geometry("350x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        msg = f"Relatório individual gerado com sucesso para:\n{nome_cliente}\nData: {data_formatada}" if nome_cliente else "Relatórios em lote gerados com sucesso!"
        
        ttk.Label(dialog, text=msg, font=('Helvetica', 10, 'bold')).pack(pady=10)
        
        def continuar():
            """Continua na interface atual"""
            dialog.destroy()
            
        def voltar_menu():
            """Volta ao menu principal de forma robusta"""
            try:
                dialog.destroy()
                self.root.destroy()
                
                # Buscar menu principal usando as múltiplas estratégias
                menu_encontrado = False
                
                # Estratégia 1: Referência direta
                if hasattr(self, 'menu_principal') and self.menu_principal:
                    if hasattr(self.menu_principal, 'winfo_exists') and self.menu_principal.winfo_exists():
                        self.menu_principal.deiconify()
                        self.menu_principal.lift()
                        self.menu_principal.focus_force()
                        menu_encontrado = True
                
                # Estratégia 2: Variável global
                if not menu_encontrado:
                    menu_global = obter_menu_principal()
                    if menu_global and hasattr(menu_global, 'winfo_exists') and menu_global.winfo_exists():
                        menu_global.deiconify()
                        menu_global.lift()
                        menu_global.focus_force()
                        menu_encontrado = True
                
                # Estratégia 3: Executar sistema principal
                if not menu_encontrado:
                    import subprocess
                    import sys
                    
                    possible_paths = [
                        "sistema_principal.py",
                        "src/sistema_principal.py"
                    ]
                    
                    for path in possible_paths:
                        if os.path.exists(path):
                            subprocess.Popen([sys.executable, path])
                            break
                            
            except Exception as e:
                logger.error(f"Erro ao voltar ao menu: {str(e)}")
                try:
                    dialog.destroy()
                    self.root.destroy()
                except:
                    pass
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill='x', pady=20)
        
        ttk.Button(btn_frame, text="Gerar Outro Relatório", 
                command=continuar).pack(pady=5, padx=10, fill='x')
        ttk.Button(btn_frame, text="Voltar ao Menu Principal", 
                command=voltar_menu).pack(pady=5, padx=10, fill='x')     

    def processar_lancamentos_futuros(self, df, data_relatorio, incluir_excluidos=False):
        """Versão corrigida que considera status de exclusão"""
        try:
            # Converter a data do relatório para datetime usando formato explícito
            try:
                self.data_ref = pd.to_datetime(data_relatorio)
            except:
                self.data_ref = pd.to_datetime(data_relatorio, format='%d/%m/%Y')

            # Converter a coluna DATA_REL para datetime
            df = df.copy()
            
            # CORREÇÃO: Só filtrar excluídos se incluir_excluidos for False
            if not incluir_excluidos and 'STATUS' in df.columns:
                df = df[df['STATUS'] != 'EXCLUIDO'].copy()
                print(f"Lançamentos futuros - registros após filtrar excluídos: {len(df)}")
            else:
                print(f"Lançamentos futuros - incluindo todos os registros: {len(df)}")
            
            # Verificar se as colunas necessárias existem
            if 'DATA_REL' not in df.columns:
                logger.error("Coluna DATA_REL não encontrada no DataFrame")
                return pd.DataFrame()
            
            if 'DT_VENCTO' not in df.columns:
                logger.warning("Coluna DT_VENCTO não encontrada, usando DATA_REL como substituto")
                df['DT_VENCTO'] = df['DATA_REL']
            
            # Converter colunas para datetime
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            df['DT_VENCTO'] = pd.to_datetime(df['DT_VENCTO'], format='%d/%m/%Y', errors='coerce')
            
            # Remover registros com datas inválidas
            df = df.dropna(subset=['DATA_REL'])
            
            # Formatar a data de vencimento para DD/MM/AAAA
            df['DT_VENCTO'] = df['DT_VENCTO'].dt.strftime('%d/%m/%Y').fillna('')

            # Filtrar apenas lançamentos futuros baseado em DATA_REL
            df_futuro = df[(df['DATA_REL'] > self.data_ref) & (df['TP_DESP'] != 1)].copy()

            if df_futuro.empty:
                logger.info("Nenhum lançamento futuro encontrado")
                return df_futuro

            # Ordenar por data de vencimento
            df_futuro = df_futuro.sort_values('DATA_REL')

            # Agrupar por período baseado na DATA_REL
            def classificar_periodo(data_rel):
                """Classifica o período baseado na diferença de dias"""
                try:
                    diff_days = (data_rel - self.data_ref).days
                    if diff_days <= 30:
                        return "Próximos 30 dias"
                    elif diff_days <= 60:
                        return "31 a 60 dias"
                    else:
                        return "Após 60 dias"
                except:
                    return "Após 60 dias"

            df_futuro['periodo'] = df_futuro['DATA_REL'].apply(classificar_periodo)

            logger.info(f"Processados {len(df_futuro)} lançamentos futuros")
            return df_futuro
            
        except Exception as e:
            logger.error(f"Erro ao processar lançamentos futuros: {str(e)}", exc_info=True)
            return pd.DataFrame()

    def adicionar_botao_pendentes(self):
        """
        Adiciona botão para gerar relatório de lançamentos pendentes
        """
        frame_pendentes = ttk.LabelFrame(self.root, text="Relatório de Lançamentos Pendentes")
        frame_pendentes.pack(pady=10, padx=20, fill='x')
        
        def selecionar_pasta():
            try:
                # Obter a data selecionada
                data_ref = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
                print(f"\nData de referência selecionada: {data_ref}")
                
                # Selecionar pasta
                pasta = filedialog.askdirectory(
                    title="Selecione a pasta com os arquivos dos clientes"
                )
                
                if pasta:
                    print(f"Pasta selecionada: {pasta}")
                    arquivo_saida = os.path.join(pasta, "relatorio_lancamentos_pendentes.html")
                    
                    # Criar instância do relatório
                    relatorio = RelatorioLancamentosPendentes()
                    
                    # Gerar relatório passando a data de referência
                    if relatorio.gerar_relatorio_pendentes(pasta, arquivo_saida, data_ref):
                        messagebox.showinfo(
                            "Sucesso",
                            f"Relatório gerado com sucesso!\nSalvo em: {arquivo_saida}"
                        )
                    else:
                        messagebox.showwarning(
                            "Aviso",
                            "Nenhum lançamento pendente encontrado para o período especificado."
                        )
                        
            except Exception as e:
                print(f"Erro ao gerar relatório: {str(e)}")
                messagebox.showerror(
                    "Erro",
                    "Erro ao gerar relatório. Verifique o console para mais detalhes."
                )
        
        # Adicionar botão
        ttk.Button(
            frame_pendentes,
            text="Gerar Relatório de Lançamentos Pendentes",
            command=selecionar_pasta
        ).pack(pady=5, fill='x')  

        

class RelatorioConfig:
    """Classe para gerenciar configurações e estilos do relatório"""
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
        
    def setup_custom_styles(self):
        """Configura os estilos personalizados para o relatório"""
        self.style_heading = ParagraphStyle(
            'HeadingStyle',
            parent=self.styles['Heading1'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            leftIndent=0,
            textColor=colors.black,
            spaceBefore=20,
            spaceAfter=12
        )
        
        self.style_normal = ParagraphStyle(
            'NormalStyle',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=12,
            textColor=colors.black,
            spaceBefore=12,
            spaceAfter=6
        )
        
        self.style_despesa = ParagraphStyle(
            name='TipoDespesa',
            parent=self.styles['Normal'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            leftIndent=0,
            firstLineIndent=0,
            rightIndent=0,
            spaceBefore=12,
            spaceAfter=6,
            keepWithNext=True
        )




def resource_path(relative_path):
    """Obtém o caminho absoluto para recursos empacotados"""
    try:
        # PyInstaller cria um temp folder e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



class IndentedFlowable:
    """Classe para ajudar na indentação de elementos como tabelas"""
    def __init__(self, flowable, leftIndent=0, rightIndent=0):
        self.flowable = flowable
        self.leftIndent = leftIndent
        self.rightIndent = rightIndent
        self.width = 0
        self.height = 0
        self.canv = None

    def wrap(self, availWidth, availHeight):
        """Define o tamanho do elemento"""
        self.width, self.height = self.flowable.wrap(
            availWidth - self.leftIndent - self.rightIndent, 
            availHeight
        )
        return (self.width + self.leftIndent + self.rightIndent, 
                self.height)

    def draw(self):
        """Desenha o elemento na posição correta"""
        self.flowable.drawOn(
            self.canv,
            self.canv._x + self.leftIndent,
            self.canv._y
        )

    def split(self, availWidth, availHeight):
        """Divide o elemento se necessário"""
        # Ajusta a largura disponível para a indentação
        availWidth = availWidth - self.leftIndent - self.rightIndent
        flowables = self.flowable.split(availWidth, availHeight)
        return [IndentedFlowable(f, self.leftIndent, self.rightIndent) for f in flowables]

    # Métodos de espaçamento
    def getSpaceBefore(self):
        return getattr(self.flowable, 'spaceBefore', 0)

    def getSpaceAfter(self):
        return getattr(self.flowable, 'spaceAfter', 0)

    def setSpaceBefore(self, space):
        self.flowable.spaceBefore = space

    def setSpaceAfter(self, space):
        self.flowable.spaceAfter = space

    # Propriedades de espaçamento
    spaceBefore = property(getSpaceBefore, setSpaceBefore)
    spaceAfter = property(getSpaceAfter, setSpaceAfter)

    # Métodos de controle de quebra de página
    def getKeepWithNext(self):
        return getattr(self.flowable, 'keepWithNext', 0)

    def setKeepWithNext(self, value):
        self.flowable.keepWithNext = value

    keepWithNext = property(getKeepWithNext, setKeepWithNext)

    # Métodos adicionais que podem ser necessários
    def identity(self, maxLen=None):
        return "IndentedFlowable: " + self.flowable.identity(maxLen)

    def drawOn(self, canvas, x, y, _sW=0):
        self.canv = canvas
        canvas.saveState()
        self.flowable.drawOn(canvas, x + self.leftIndent, y, _sW)
        canvas.restoreState()

    # Delegação de outros atributos ao flowable interno
    def __getattr__(self, name):
        return getattr(self.flowable, name)




class RelatorioHandler:
    def __init__(self):
        self.config = RelatorioConfig()
        self.tipos_despesas = {
            1: "1) DESPESAS COM COLABORADORES",
            2: "2) TRANSF. PROGR. - MATERIAIS, LOCAÇÕES E PREST.SERVIÇOS",
            3: "3) BOLETOS - MATERIAIS, PREST. SERVIÇOS, IMPOSTOS, ETC.",
            4: "4) RESSARCIMENTOS E RESTITUIÇÕES",
            5: "5) DESPESAS PAGAS PELO CLIENTE",
            6: "6) PAGAMENTOS CAIXA DE OBRA",
            7: "7) ADMINISTRAÇÃO DA OBRA - % SOBRE DESPESAS",
        }

        # Verificar se a logomarca existe na mesma pasta do script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.logo_path = os.path.join(script_dir, "logo1.png")
        if not os.path.exists(self.logo_path):
            self.logo_path = None
            print("Aviso: Logomarca não encontrada na pasta do script.")
        
        self.tipos_despesas_futuras = {
            "Próximos 30 dias": lambda x: x <= self.data_ref + pd.Timedelta(days=30),
            "31 a 60 dias": lambda x: (x > self.data_ref + pd.Timedelta(days=30)) & 
                                     (x <= self.data_ref + pd.Timedelta(days=60)),
            "Após 60 dias": lambda x: x > self.data_ref + pd.Timedelta(days=60)
        }
        self.data_ref = None

    def gerar_relatorio_direto(self, arquivo_path, data_relatorio, incluir_futuros=True, output_callback=None):
        """
        Novo método para geração direta do relatório sem interface própria
        
        Args:
            arquivo_path: Caminho do arquivo Excel
            data_relatorio: Data do relatório (datetime)
            incluir_futuros: Se deve incluir lançamentos futuros
            output_callback: Função para retornar mensagens de status
        """
        try:
            # Carregar e processar dados
            df = self.carregar_dados_excel(arquivo_path)
            df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2 = self.processar_dados(df, data_relatorio)
            
            # Processar lançamentos futuros
            df_futuro = None
            if incluir_futuros:
                df_futuro = self.processar_lancamentos_futuros(df, data_relatorio)
            
            # Processar workbook
            workbook = load_workbook(arquivo_path, data_only=True)
            ws_resumo = workbook['RESUMO']
            nome_cliente = ws_resumo['A3'].value
            
            # Obter número do relatório e valor acumulado
            numero_relatorio = self.obter_numero_relatorio(ws_resumo, data_relatorio)
            valor_acumulado = self.calcular_acumulado_dados(df, data_relatorio)
            
            dados_completos = {
                'df_filtrado': df_filtrado,
                'df_diaria': df_diaria,
                'df_tp_desp_1': df_tp_desp_1,
                'df_tp_desp_2': df_tp_desp_2,
                'df_futuro': df_futuro,
                'df_original': df,
                'incluir_futuros': incluir_futuros,
                'data_relatorio': data_relatorio,
                'nome_cliente': nome_cliente,
                'endereco_cliente': ws_resumo['A4'].value,
                'numero_relatorio': numero_relatorio,
                'acumulado': valor_acumulado
            }
            
            # Gerar nome do arquivo
            data_formatada = data_relatorio.strftime('%d-%m-%Y')
            nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
            caminho_output = os.path.join(os.path.dirname(arquivo_path), nome_arquivo)
            
            # Gerar o PDF
            self.gerar_relatorio_pdf(dados_completos, caminho_output, arquivo_path)
            
            if output_callback:
                output_callback(f"Relatório gerado com sucesso: {nome_arquivo}")
            
            return caminho_output
            
        except Exception as e:
            erro_msg = f"Erro ao gerar relatório: {str(e)}"
            if output_callback:
                output_callback(erro_msg)
            raise Exception(erro_msg)       
        
    def selecionar_arquivo(self):
        """Interface para seleção do arquivo Excel"""
        root = Tk()
        root.withdraw()
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        return arquivo

    def obter_numero_relatorio(self, ws_resumo, data_relatorio):
        """
        Método para obter o número do relatório baseado na data.
        """
        try:
            logger.info(f"\nObtendo número do relatório para data {data_relatorio}")
            
            # Converter data para datetime
            data_ref = pd.to_datetime(data_relatorio).date()
            logger.debug(f"Data de referência processada: {data_ref}")
            
            # Encontrar primeira data na planilha
            primeira_data = None
            primeira_linha = None
            
            for row in range(9, 150):  # Buscar nas primeiras 150 linhas
                cell_value = ws_resumo.cell(row=row, column=1).value
                if isinstance(cell_value, (datetime, date)):
                    primeira_data = cell_value.date() if isinstance(cell_value, datetime) else cell_value
                    primeira_linha = row
                    break
                    
            if not primeira_data:
                logger.warning("Nenhuma data encontrada na planilha")
                return 1
                
            logger.debug(f"Primeira data encontrada: {primeira_data} na linha {primeira_linha}")
            
            # Usar a primeira data encontrada como data inicial
            data_inicial = primeira_data
            logger.debug(f"Data inicial: {data_inicial}")
            
            # Calcular número do relatório
            numero = 1
            data_atual = data_inicial
            
            while data_atual <= data_ref:
                logger.debug(f"Verificando data: {data_atual}")
                
                if data_atual == data_ref:
                    logger.info(f"Número do relatório calculado: {numero}")
                    return numero
                    
                # Avançar para próxima data (5 ou 20 do mês)
                if data_atual.day == 5:
                    data_atual = data_atual.replace(day=20)
                else:  # day == 20
                    if data_atual.month == 12:
                        data_atual = data_atual.replace(year=data_atual.year + 1, month=1, day=5)
                    else:
                        data_atual = data_atual.replace(month=data_atual.month + 1, day=5)
                numero += 1
                
            logger.warning(f"Data {data_ref} não encontrada na sequência. Último número calculado: {numero}")
            return numero
            
        except Exception as e:
            logger.error(f"Erro ao obter número do relatório: {str(e)}", exc_info=True)
            return 1

    def calcular_acumulado_dados(self, df, data_relatorio, incluir_excluidos=False):
        """
        Versão modificada que considera apenas registros ativos ou todos dependendo da opção
        """
        try:
            logger.info(f"Calculando acumulado para data {data_relatorio}")
            
            # Criar cópia do DataFrame
            df = df.copy()
            
            # CORREÇÃO: Só filtrar excluídos se incluir_excluidos for False
            if not incluir_excluidos and 'STATUS' in df.columns:
                df = df[df['STATUS'] != 'EXCLUIDO'].copy()
                logger.info(f"Acumulado - registros após filtrar excluídos: {len(df)}")
            else:
                logger.info(f"Acumulado - incluindo todos os registros: {len(df)}")
            
            if 'VALOR' not in df.columns:
                logger.error("Coluna 'VALOR' não encontrada no DataFrame")
                return 0.0
                
            # Garantir que data_relatorio seja datetime
            if isinstance(data_relatorio, str):
                data_relatorio = pd.to_datetime(data_relatorio)
            elif isinstance(data_relatorio, date):
                data_relatorio = pd.to_datetime(data_relatorio)
                
            logger.debug(f"Data de referência processada: {data_relatorio}")
            
            # Converter DATA_REL para datetime
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            df = df.dropna(subset=['DATA_REL'])
            
            # Filtrar registros anteriores à data do relatório
            df_anterior = df[df['DATA_REL'] < data_relatorio].copy()
            
            if df_anterior.empty:
                logger.warning("Nenhum registro anterior encontrado")
                return 0.0
                
            # Converter valores para numérico
            logger.debug("Processando valores...")
            df_anterior['VALOR_NUMERICO'] = df_anterior['VALOR'].replace({',': '.', 'R$': '', ' ': ''}, regex=True)
            df_anterior['VALOR_NUMERICO'] = pd.to_numeric(df_anterior['VALOR_NUMERICO'], errors='coerce').fillna(0)
            
            # Calcular soma
            valor_acumulado = float(df_anterior['VALOR_NUMERICO'].sum())
            
            status_info = "incluindo excluídos" if incluir_excluidos else "apenas ativos"
            logger.info(f"Valor acumulado calculado ({status_info}): {valor_acumulado:,.2f}")
            logger.debug(f"Total de registros considerados: {len(df_anterior)}")
            
            return valor_acumulado
                
        except Exception as e:
            logger.error(f"Erro ao calcular acumulado: {str(e)}", exc_info=True)
            return 0.0

    def carregar_dados_excel(self, arquivo_excel, incluir_excluidos=False):
        """Versão modificada que considera status de exclusão"""
        try:
            df = pd.read_excel(arquivo_excel, sheet_name='Dados')
            df = df.fillna("")
            
            # Verificar colunas necessárias
            colunas_necessarias = {'DATA_REL', 'TP_DESP', 'REFERÊNCIA', 'DT_VENCTO', 'VALOR', 'NF'}
            if not colunas_necessarias.issubset(df.columns):
                raise ValueError(f"Colunas necessárias ausentes: {colunas_necessarias - set(df.columns)}")
            
            # Adicionar coluna STATUS se não existir
            if 'STATUS' not in df.columns:
                df['STATUS'] = 'ATIVO'
            
            # CORREÇÃO: Só filtrar excluídos se incluir_excluidos for False
            if not incluir_excluidos:
                df = df[df['STATUS'] != 'EXCLUIDO'].copy()
                print(f"Registros após filtrar excluídos: {len(df)}")
            else:
                print(f"Incluindo todos os registros (incluindo excluídos): {len(df)}")
            
            # Converter NF para string antes de processar
            df['NF'] = df['NF'].astype(str)
            
            # Concatenar NF com REFERÊNCIA apenas para TP_DESP != 1
            mascara = (df['TP_DESP'] != 1) & (df['NF'].notna()) & (df['NF'].str.strip() != '') & (df['NF'] != 'nan')
            df.loc[mascara, 'REFERÊNCIA'] = df[mascara].apply(
                lambda row: f"{row['REFERÊNCIA']} (NF: {row['NF'].strip()})", 
                axis=1
            )
            
            return df
            
        except Exception as e:
            raise Exception(f"Erro ao carregar arquivo Excel: {str(e)}")

    def processar_dados(self, df, data_relatorio, incluir_excluidos=False):
        """Versão corrigida que preserva todas as colunas essenciais"""
        # Converter data para datetime usando formato explícito
        try:
            data_rel = pd.to_datetime(data_relatorio)
        except:
            # Se falhar, tenta converter assumindo formato brasileiro
            data_rel = pd.to_datetime(data_relatorio, format='%d/%m/%Y')
        
        # Criar cópia do DataFrame para não modificar o original
        df = df.copy()
        
        # Log para debug
        logger.debug(f"Colunas do DataFrame original: {df.columns.tolist()}")
        
        # CORREÇÃO: Só filtrar excluídos se incluir_excluidos for False
        if not incluir_excluidos and 'STATUS' in df.columns:
            df = df[df['STATUS'] != 'EXCLUIDO'].copy()
            print(f"Processando dados - registros após filtrar excluídos: {len(df)}")
        else:
            print(f"Processando dados - incluindo todos os registros: {len(df)}")
        
        # Adicionar coluna de índice original para manter ordem de entrada
        df = df.reset_index(drop=True)
        df['ordem_original'] = df.index
        
        # Converter corretamente a coluna DT_VENCTO para datetime para ordenação
        if 'DT_VENCTO' in df.columns:
            try:
                df['DT_VENCTO_SORT'] = pd.to_datetime(df['DT_VENCTO'], 
                                                    format='mixed', 
                                                    errors='coerce', 
                                                    dayfirst=True)
                
                df['DT_VENCTO_DISPLAY'] = df['DT_VENCTO_SORT'].dt.strftime('%d/%m/%Y')
                
            except Exception as e:
                print(f"Erro ao converter DT_VENCTO: {str(e)}")
                df['DT_VENCTO_SORT'] = pd.to_datetime('2000-01-01')
                df['DT_VENCTO_DISPLAY'] = df['DT_VENCTO']
        
        # Aplicar a restrição de dados bancários para tp_desp 3 e 5
        if 'DADOS_BANCARIOS' in df.columns:
            df['DADOS_BANCARIOS_ORIGINAL'] = df['DADOS_BANCARIOS']
            df.loc[df['TP_DESP'].isin([3, 5]), 'DADOS_BANCARIOS'] = ''

        # VERIFICAÇÃO CRÍTICA: Garantir que TP_DESP existe
        if 'TP_DESP' not in df.columns:
            logger.error("ERRO CRÍTICO: Coluna TP_DESP não encontrada no DataFrame!")
            logger.error(f"Colunas disponíveis: {df.columns.tolist()}")
            raise ValueError("Coluna TP_DESP não encontrada no DataFrame")

        # Filtrar dados (considerando a opção de incluir excluídos)
        df_filtrado = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] != 1)
        ].copy()  # IMPORTANTE: usar .copy() para evitar warnings
        
        # Log para debug
        logger.debug(f"df_filtrado criado com {len(df_filtrado)} registros")
        logger.debug(f"Colunas do df_filtrado: {df_filtrado.columns.tolist()}")
        
        # NOVA LÓGICA: Separar TP_DESP == 5 dos demais para ordenação diferente
        df_tp5 = df_filtrado[df_filtrado['TP_DESP'] == 5].copy()
        df_outros = df_filtrado[df_filtrado['TP_DESP'] != 5].copy()
        
        # Ordenar apenas os outros tipos (não o tipo 5)
        if not df_outros.empty:
            df_outros = df_outros.sort_values(
                by=['TP_DESP', 'DT_VENCTO_SORT', 'VALOR'], 
                ascending=[True, True, False]
            )
        
        # Para TP_DESP == 5, manter ordem original de entrada
        if not df_tp5.empty:
            df_tp5 = df_tp5.sort_values('ordem_original')
        
        # Combinar os DataFrames: primeiro os outros tipos ordenados, depois o tipo 5 na ordem original
        if not df_outros.empty and not df_tp5.empty:
            df_filtrado = pd.concat([df_outros, df_tp5], ignore_index=True)
        elif not df_outros.empty:
            df_filtrado = df_outros
        elif not df_tp5.empty:
            df_filtrado = df_tp5
        else:
            df_filtrado = pd.DataFrame()  # Vazio se nenhum dos dois tiver dados
        
        # Processar os outros DataFrames normalmente
        df_diaria = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'] == 'DIÁRIA')
        ].copy()
        
        df_tp_desp_1 = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'].isin(['SALÁRIO', 'TRANSPORTE', 'CAFÉ']))
        ].copy()

        df_tp_desp_2 = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'].isin(['FÉRIAS', 'RESCISÃO', '13º SALÁRIO']))
        ].copy()
        
        # Substituir DT_VENCTO pela versão formatada uniformemente antes de retornar
        if 'DT_VENCTO_DISPLAY' in df_filtrado.columns:
            df_filtrado['DT_VENCTO'] = df_filtrado['DT_VENCTO_DISPLAY']
            
        # CORREÇÃO CRÍTICA: Preservar colunas essenciais
        colunas_essenciais = [
            'TP_DESP', 'NOME', 'REFERÊNCIA', 'VALOR', 'DATA_REL', 'DT_VENCTO',
            'DADOS_BANCARIOS', 'DIAS', 'VR_UNIT', 'NF', 'STATUS'
        ]
        
        # Remover apenas as colunas temporárias, preservando as essenciais
        colunas_temporarias = ['DT_VENCTO_SORT', 'DT_VENCTO_DISPLAY', 'ordem_original', 'DADOS_BANCARIOS_ORIGINAL']
        
        for df_temp in [df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2]:
            # Verificar se o DataFrame não está vazio
            if df_temp.empty:
                continue
                
            # Log das colunas antes da limpeza
            logger.debug(f"DataFrame com {len(df_temp)} registros - Colunas antes da limpeza: {df_temp.columns.tolist()}")
            
            # Remover apenas colunas temporárias que existem
            colunas_para_remover = [col for col in colunas_temporarias if col in df_temp.columns]
            
            if colunas_para_remover:
                df_temp.drop(columns=colunas_para_remover, inplace=True)
                logger.debug(f"Colunas removidas: {colunas_para_remover}")
            
            # Verificar se TP_DESP ainda existe após limpeza
            if 'TP_DESP' not in df_temp.columns and not df_temp.empty:
                logger.error(f"ERRO: TP_DESP foi removida inadvertidamente do DataFrame com {len(df_temp)} registros!")
                logger.error(f"Colunas após limpeza: {df_temp.columns.tolist()}")
            else:
                logger.debug(f"TP_DESP preservada - Colunas após limpeza: {df_temp.columns.tolist()}")
        
        # Verificação final
        logger.info(f"df_filtrado final: {len(df_filtrado)} registros")
        if not df_filtrado.empty:
            logger.info(f"Colunas finais do df_filtrado: {df_filtrado.columns.tolist()}")
            if 'TP_DESP' in df_filtrado.columns:
                logger.info(f"Tipos de despesa únicos: {df_filtrado['TP_DESP'].unique()}")
            else:
                logger.error("ERRO FINAL: TP_DESP não está presente no df_filtrado!")
        
        return df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2
    
    # def processar_dados(self, df, data_relatorio, incluir_excluidos=False):
    #     """Versão corrigida do processar_dados com debug melhorado"""
        
    #     # CORREÇÃO: Garantir formato correto da data
    #     from datetime import datetime, date
    #     import pandas as pd
        
    #     if isinstance(data_relatorio, str):
    #         data_rel = pd.to_datetime(data_relatorio)
    #     elif isinstance(data_relatorio, date) and not isinstance(data_relatorio, datetime):
    #         data_rel = pd.to_datetime(datetime.combine(data_relatorio, datetime.min.time()))
    #     else:
    #         data_rel = pd.to_datetime(data_relatorio)
        
    #     logger.info(f"Data de referência para processamento: {data_rel}")
    #     logger.info(f"Tipo da data de referência: {type(data_rel)}")
        
    #     # Criar cópia do DataFrame
    #     df = df.copy()
        
    #     # Log para debug
    #     logger.debug(f"Colunas do DataFrame original: {df.columns.tolist()}")
    #     logger.info(f"Total de registros originais: {len(df)}")
        
    #     # Filtrar excluídos se necessário
    #     if not incluir_excluidos and 'STATUS' in df.columns:
    #         df = df[df['STATUS'] != 'EXCLUIDO'].copy()
    #         logger.info(f"Processando dados - registros após filtrar excluídos: {len(df)}")
    #     else:
    #         logger.info(f"Processando dados - incluindo todos os registros: {len(df)}")
        
    #     # CORREÇÃO: Converter DATA_REL para datetime ANTES de comparar
    #     df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
        
    #     # Remover registros com datas inválidas
    #     registros_antes = len(df)
    #     df = df.dropna(subset=['DATA_REL'])
    #     registros_depois = len(df)
        
    #     if registros_antes != registros_depois:
    #         logger.warning(f"Removidos {registros_antes - registros_depois} registros com DATA_REL inválida")
        
    #     # DEBUG: Mostrar datas únicas disponíveis
    #     datas_unicas = sorted(df['DATA_REL'].dt.date.unique())
    #     logger.info(f"Datas únicas disponíveis no arquivo: {datas_unicas}")
        
    #     # DEBUG: Verificar se a data procurada existe
    #     data_procurada = data_rel.date()
    #     logger.info(f"Data procurada: {data_procurada}")
        
    #     if data_procurada not in datas_unicas:
    #         logger.warning(f"Data {data_procurada} não encontrada no arquivo!")
    #         logger.info("Datas próximas disponíveis:")
    #         for data_disp in datas_unicas:
    #             diff = abs((data_disp - data_procurada).days)
    #             logger.info(f"  {data_disp} (diferença: {diff} dias)")
        
    #     # Aplicar filtros
    #     df_filtrado = df[
    #         (df['DATA_REL'].dt.date == data_rel.date()) & 
    #         (df['TP_DESP'] != 1)
    #     ].copy()
        
    #     df_diaria = df[
    #         (df['DATA_REL'].dt.date == data_rel.date()) & 
    #         (df['TP_DESP'] == 1) & 
    #         (df['REFERÊNCIA'] == 'DIÁRIA')
    #     ].copy()
        
    #     df_tp_desp_1 = df[
    #         (df['DATA_REL'].dt.date == data_rel.date()) & 
    #         (df['TP_DESP'] == 1) & 
    #         (df['REFERÊNCIA'].isin(['SALÁRIO', 'TRANSPORTE', 'CAFÉ']))
    #     ].copy()

    #     df_tp_desp_2 = df[
    #         (df['DATA_REL'].dt.date == data_rel.date()) & 
    #         (df['TP_DESP'] == 1) & 
    #         (df['REFERÊNCIA'].isin(['FÉRIAS', 'RESCISÃO', '13º SALÁRIO']))
    #     ].copy()
        
    #     # Log dos resultados
    #     logger.info(f"Resultados do processamento:")
    #     logger.info(f"  - df_filtrado: {len(df_filtrado)} registros")
    #     logger.info(f"  - df_diaria: {len(df_diaria)} registros")
    #     logger.info(f"  - df_tp_desp_1: {len(df_tp_desp_1)} registros")
    #     logger.info(f"  - df_tp_desp_2: {len(df_tp_desp_2)} registros")
        
    #     # DEBUG: Se não encontrou nada, mostrar amostra dos dados
    #     if len(df_filtrado) == 0 and len(df_diaria) == 0 and len(df_tp_desp_1) == 0 and len(df_tp_desp_2) == 0:
    #         logger.warning("NENHUM DADO ENCONTRADO! Mostrando amostra dos dados:")
    #         amostra = df.head(10)[['DATA_REL', 'TP_DESP', 'REFERÊNCIA', 'VALOR']].copy()
    #         amostra['DATA_REL_DATE'] = amostra['DATA_REL'].dt.date
    #         logger.info(f"Amostra dos dados:\n{amostra}")
        
    #     return df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2
    
    def adicionar_lancamentos_futuros(self, elementos, dados):
        """Adiciona a seção de lançamentos futuros ao relatório"""
        if not dados['df_futuro'].empty:
            elementos.append(PageBreak())
            elementos.append(Paragraph("LANÇAMENTOS FUTUROS", self.config.style_heading))
            
            total_geral_futuro = 0
            
            # Agrupar por período e tipo de despesa
            for periodo in ["Próximos 30 dias", "31 a 60 dias", "Após 60 dias"]:
                df_periodo = dados['df_futuro'][dados['df_futuro']['periodo'] == periodo]
                
                if not df_periodo.empty:
                    # Adicionar título do período com estilo destacado
                    elementos.append(Paragraph(
                        f"\n{periodo}",
                        ParagraphStyle(
                            'PeriodoStyle',
                            parent=self.config.style_heading,
                            fontSize=14,
                            leading=16,
                            spaceBefore=12,
                            spaceAfter=6,
                            textColor=colors.HexColor('#2F4F4F')  # Cor mais escura para destaque
                        )
                    ))
                    
                    total_periodo = 0
                    
                    # Agrupar por tipo de despesa dentro do período
                    for tipo in sorted(df_periodo['TP_DESP'].unique()):
                        df_tipo = df_periodo[df_periodo['TP_DESP'] == tipo]
                        if not df_tipo.empty:
                            elementos.append(Paragraph(
                                self.tipos_despesas.get(tipo, f"Tipo {tipo}"),
                                self.config.style_normal
                            ))
                            
                            # Renomear colunas para corresponder ao formato esperado
                            df_tipo = df_tipo.rename(columns={
                                'DT_VENCTO': 'VENCIMENTO',
                                'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
                            })
                            
                            tabela = self.criar_tabela_despesas(
                                df_tipo,
                                ['NOME', 'VENCIMENTO', 'REFERÊNCIA', 'VALOR', 'DADOS BANCÁRIOS'],
                                [240, 70, 220, 80, 170]
                            )
                            elementos.append(tabela)
                            elementos.append(Spacer(1, 12))
                            
                            total_periodo += df_tipo['VALOR'].sum()
                    
                    # Adicionar subtotal do período
                    elementos.append(Paragraph(
                        f"Subtotal {periodo}: {self.formatar_numero(total_periodo)}",
                        ParagraphStyle(
                            'SubtotalStyle',
                            parent=self.config.style_normal,
                            fontSize=10,
                            leading=12,
                            spaceBefore=6,
                            spaceAfter=12,
                            textColor=colors.HexColor('#4A4A4A')
                        )
                    ))
                    
                    total_geral_futuro += total_periodo
            
            # Adicionar total geral dos lançamentos futuros
            elementos.append(Paragraph(
                f"\nTotal Geral de Lançamentos Futuros: {self.formatar_numero(total_geral_futuro)}",
                self.config.style_heading
            ))
    
    def formatar_numero(self, valor):
        """Formata valor numérico, tratando possíveis strings e NaN"""
        if pd.isna(valor) or valor == "":
            return "0,00"
        try:
            if isinstance(valor, str):
                # Remover caracteres não numéricos, mantendo ponto decimal
                valor = valor.replace('R$', '').replace(' ', '')
                # Se tiver vírgula como separador decimal, converter para ponto
                if ',' in valor and '.' not in valor:
                    valor = valor.replace(',', '.')
                # Se tiver tanto vírgula quanto ponto, assume que a vírgula é separador de milhar
                elif ',' in valor and '.' in valor:
                    valor = valor.replace('.', '').replace(',', '.')
                
                valor = float(valor)
            
            # Formatar com separador de milhar e vírgula como separador decimal
            return f"{float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except Exception as e:
            logger.error(f"Erro ao formatar número: {str(e)} - Valor: {valor}")
            return "0,00"  # Retorna zero formatado em caso de erro

    def formatar_data(self, data):
        """Formata data para o padrão brasileiro"""
        if pd.isna(data):
            return ''
        try:
            return pd.to_datetime(data).strftime('%d/%m/%Y')
        except:
            return str(data)

    def consolidar_despesas_colaboradores(self, df):
        """Consolida as despesas dos colaboradores"""
        try:
            # Criar cópia e tratar valores nulos
            df = df.copy()
            df = df.fillna("")
            df = df.infer_objects(copy=False)
                
            agregacoes = {
                'SALÁRIO': ['SALÁRIO'],
                'TRANSPORTE': ['TRANSPORTE'],
                'CAFÉ': ['CAFÉ']
            }
                
            if 'DADOS_BANCARIOS' in df.columns:
                df = df.rename(columns={'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'})
                
            resultados = []
            total_colunas = {
                'SALÁRIO': 0.0,
                'TRANSPORTE': 0.0,
                'CAFÉ': 0.0
            }
                
            for nome, grupo in df.groupby('NOME'):
                linha = {'NOME': nome}
                    
                for coluna, referencias in agregacoes.items():
                    valores_grupo = grupo[grupo['REFERÊNCIA'].isin(referencias)]['VALOR']
                    valor = pd.to_numeric(valores_grupo, errors='coerce').sum()
                    linha[coluna] = valor if not pd.isna(valor) else 0.0
                    total_colunas[coluna] += linha[coluna]
                        
                # Pegar DIAS do lançamento de TRANSPORTE ou CAFÉ (o que for maior)
                dias_transporte = grupo[grupo['REFERÊNCIA'] == 'TRANSPORTE']['DIAS'].iloc[0] if not grupo[grupo['REFERÊNCIA'] == 'TRANSPORTE'].empty else 0
                dias_cafe = grupo[grupo['REFERÊNCIA'] == 'CAFÉ']['DIAS'].iloc[0] if not grupo[grupo['REFERÊNCIA'] == 'CAFÉ'].empty else 0
                    
                # Converter para inteiro e pegar o maior valor
                dias_transporte = int(dias_transporte) if pd.notnull(dias_transporte) else 0
                dias_cafe = int(dias_cafe) if pd.notnull(dias_cafe) else 0
                linha['DIAS'] = max(dias_transporte, dias_cafe)
                    
                linha['DADOS BANCÁRIOS'] = grupo['DADOS BANCÁRIOS'].iloc[0] if not grupo['DADOS BANCÁRIOS'].empty else ''
                linha['TOTAL'] = sum(linha.get(col, 0) for col in total_colunas.keys())
                    
                resultados.append(linha)
                
            # Criar DataFrame com os resultados
            df_result = pd.DataFrame(resultados)
                
            # Definir ordem das colunas
            colunas_ordem = ['NOME', 'SALÁRIO', 'DIAS', 
                            'TRANSPORTE', 'CAFÉ', 'TOTAL', 'DADOS BANCÁRIOS']
                
            # Reordenar colunas
            df_result = df_result.reindex(columns=colunas_ordem)
            
            # ADICIONAR ESTA LINHA:
            df_result = df_result.sort_values('TOTAL', ascending=False)
                
            return df_result
                
        except Exception as e:
            print(f"Erro ao consolidar despesas: {str(e)}")
            raise

    def consolidar_despesas_colaboradores1(self, df):
        """Consolida as despesas  13º, férias e rescisão dos colaboradores"""
        try:
            # Criar cópia e tratar valores nulos
            df = df.copy()
            df = df.fillna("")
            df = df.infer_objects(copy=False)
                
            agregacoes1 = {
                '13º SALÁRIO': ['13º SALÁRIO'],
                'FÉRIAS': ['FÉRIAS'],
                'RESCISÃO': ['RESCISÃO']
            }
                
            if 'DADOS_BANCARIOS' in df.columns:
                df = df.rename(columns={'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'})
                
            resultados1 = []
            total_colunas = {
                '13º SALÁRIO': 0.0,
                'FÉRIAS': 0.0,
                'RESCISÃO': 0.0
            }
                
            for nome, grupo in df.groupby('NOME'):
                linha = {'NOME': nome}
                    
                for coluna, referencias in agregacoes1.items():
                    valores_grupo = grupo[grupo['REFERÊNCIA'].isin(referencias)]['VALOR']
                    valor = pd.to_numeric(valores_grupo, errors='coerce').sum()
                    linha[coluna] = valor if not pd.isna(valor) else 0.0
                    total_colunas[coluna] += linha[coluna]
                        
                    
                linha['DADOS BANCÁRIOS'] = grupo['DADOS BANCÁRIOS'].iloc[0] if not grupo['DADOS BANCÁRIOS'].empty else ''
                linha['TOTAL'] = sum(linha.get(col, 0) for col in total_colunas.keys())
                    
                resultados1.append(linha)
                
            # Criar DataFrame com os resultados
            df_result1 = pd.DataFrame(resultados1)
                
            # Definir ordem das colunas
            colunas_ordem = ['NOME', '13º SALÁRIO', 'FÉRIAS', 
                            'RESCISÃO', 'TOTAL', 'DADOS BANCÁRIOS']
                
            # Reordenar colunas
            df_result1 = df_result1.reindex(columns=colunas_ordem)
            
            # ADICIONAR ESTA LINHA:
            df_result1 = df_result1.sort_values('TOTAL', ascending=False)
                
            return df_result1
                
        except Exception as e:
            print(f"Erro ao consolidar despesas: {str(e)}")
            raise

    def criar_tabela_despesas(self, dados, colunas, larguras, incluir_total=True):
        """Versão modificada que pode mostrar status dos lançamentos"""
        dados_formatados = dados.copy()
        dados_formatados = dados_formatados.fillna("")
        dados_formatados = dados_formatados.infer_objects()

        # Verificar se deve mostrar coluna de status
        mostrar_status = 'STATUS' in dados_formatados.columns and \
                        dados_formatados['STATUS'].nunique() > 1

        if mostrar_status and 'STATUS' not in colunas:
            colunas = list(colunas) + ['STATUS']
            larguras = list(larguras) + [60]

        # Estilo para o cabeçalho com quebra de linha
        estilo_cabecalho = ParagraphStyle(
            'CabecalhoTabela',
            parent=self.config.style_normal,
            fontSize=8,
            leading=10,
            alignment=1,
            textColor=colors.whitesmoke
        )

        # Estilo para células com quebra de texto
        estilo_celula = ParagraphStyle(
            'CelulaTabela',
            parent=self.config.style_normal,
            fontSize=8,
            leading=10,
            alignment=0
        )

        # Converter cabeçalhos simples em Paragraphs com quebras de linha
        cabecalhos_formatados = []
        for coluna in colunas:
            if '/' in coluna:
                texto_formatado = Paragraph(coluna.replace('/', '<br/>'), estilo_cabecalho)
            elif ' - ' in coluna:
                texto_formatado = Paragraph(coluna.replace(' - ', '<br/>'), estilo_cabecalho)
            else:
                texto_formatado = Paragraph(coluna, estilo_cabecalho)
            cabecalhos_formatados.append(texto_formatado)

        colunas_numericas = ['VALOR', 'TOTAL', 'SALÁRIO', 'RESCISÃO', '13º SALÁRIO', 
                            'TRANSPORTE', 'CAFÉ', 'FÉRIAS', 'DIÁRIA', 'DIAS']

        # Colunas para centralizar
        colunas_centralizadas = ['DT_VENCTO', 'VENCIMENTO', 'STATUS']

        # Processar dados linha por linha
        dados_tabela = [cabecalhos_formatados]
        for _, linha in dados_formatados.iterrows():
            linha_formatada = []
            for i, coluna in enumerate(colunas):
                valor = linha[coluna] if coluna in linha.index else ""
                
                # Formatar números
                if coluna in colunas_numericas:
                    valor = pd.to_numeric(valor, errors='coerce')
                    valor = 0 if pd.isna(valor) else valor
                    if coluna == 'DIAS':
                        valor = str(int(valor))
                    else:
                        valor = self.formatar_numero(valor)
                    linha_formatada.append(valor)
                
                # Formatar datas
                elif coluna in ['DT_VENCTO', 'VENCIMENTO']:
                    try:
                        valor = pd.to_datetime(valor, dayfirst=True).strftime('%d/%m/%Y')
                    except:
                        valor = str(valor)
                    linha_formatada.append(valor)
                
                # Adicionar quebra de texto para a coluna Referência
                elif coluna == 'REFERÊNCIA':
                    valor = str(valor)
                    linha_formatada.append(Paragraph(valor, estilo_celula))
                
                # Tratar coluna STATUS
                elif coluna == 'STATUS':
                    valor = str(valor) if valor else "ATIVO"
                    linha_formatada.append(valor)
                
                # Outras colunas
                else:
                    linha_formatada.append(str(valor))
                    
            dados_tabela.append(linha_formatada)

        # Adicionar linha de total se necessário (mesmo código anterior)
        if incluir_total:
            coluna_valor = next((i for i, col in enumerate(colunas) 
                            if col in ['VALOR', 'TOTAL']), -1)
            if coluna_valor >= 0:
                if 'SALÁRIO' in colunas and 'TRANSPORTE' in colunas or '13º SALÁRIO' in colunas and 'FÉRIAS' in colunas:
                    linha_total = [''] * len(colunas)
                    linha_total[0] = 'Subtotal'
                    
                    for i, col in enumerate(colunas):
                        if col in ['SALÁRIO', 'FÉRIAS', 'RESCISÃO', '13º SALÁRIO', 'TRANSPORTE', 'CAFÉ', 'TOTAL']:
                            total = dados[col].sum()
                            linha_total[i] = self.formatar_numero(total)
                        elif col == 'DIAS':
                            linha_total[i] = ''
                            
                    dados_tabela.append(linha_total)
                
                else:
                    total = dados[colunas[coluna_valor]].sum()
                    linha_total = [''] * len(colunas)
                    linha_total[coluna_valor-1] = 'Subtotal'
                    linha_total[coluna_valor] = self.formatar_numero(total)
                    dados_tabela.append(linha_total)

        # Criar tabela com os dados formatados
        tabela = Table(dados_tabela, colWidths=larguras, repeatRows=1)
        
        # Definir estilos da tabela
        estilo_tabela = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        # Alinhar colunas numéricas à direita
        for i, col in enumerate(colunas):
            if col in colunas_numericas:
                estilo_tabela.append(('ALIGN', (i, 1), (i, -1), 'RIGHT'))

        # Alinhar colunas de data e status ao centro
        for i, col in enumerate(colunas):
            if col in colunas_centralizadas:
                estilo_tabela.append(('ALIGN', (i, 0), (i, -1), 'CENTER'))

        # Destacar linhas com status EXCLUIDO
        if mostrar_status:
            status_col_idx = colunas.index('STATUS')
            for row_idx, linha in enumerate(dados_tabela[1:], 1):  # Pular cabeçalho
                if row_idx < len(dados_tabela) - (1 if incluir_total else 0):  # Não aplicar à linha de total
                    if len(linha) > status_col_idx and linha[status_col_idx] == 'EXCLUIDO':
                        estilo_tabela.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightgrey))
                        estilo_tabela.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.red))

        if incluir_total:
            estilo_tabela.extend([
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ])

        tabela.setStyle(TableStyle(estilo_tabela))
        return tabela

    def criar_resumo_despesas(self, dados):
        """Cria o resumo das despesas para o relatório - VERSÃO CORRIGIDA"""
        logger.debug("\nIniciando criar_resumo_despesas")
        logger.debug(f"Dados recebidos - acumulado: {dados.get('acumulado')}")
        
        # VERIFICAÇÃO CRÍTICA DOS DADOS
        df_filtrado = dados.get('df_filtrado', pd.DataFrame())
        df_tp_desp_1 = dados.get('df_tp_desp_1', pd.DataFrame())
        df_tp_desp_2 = dados.get('df_tp_desp_2', pd.DataFrame())
        df_diaria = dados.get('df_diaria', pd.DataFrame())
        
        logger.debug(f"df_filtrado: {len(df_filtrado)} registros")
        logger.debug(f"df_tp_desp_1: {len(df_tp_desp_1)} registros")
        logger.debug(f"df_tp_desp_2: {len(df_tp_desp_2)} registros")
        logger.debug(f"df_diaria: {len(df_diaria)} registros")
        
        # Verificar se df_filtrado tem a coluna TP_DESP
        if not df_filtrado.empty:
            logger.debug(f"Colunas do df_filtrado: {df_filtrado.columns.tolist()}")
            if 'TP_DESP' not in df_filtrado.columns:
                logger.error("ERRO CRÍTICO: df_filtrado não contém a coluna TP_DESP!")
                # Tentar recuperar dos dados originais se possível
                df_original = dados.get('df_original', pd.DataFrame())
                if not df_original.empty and 'TP_DESP' in df_original.columns:
                    logger.warning("Tentando recriar df_filtrado dos dados originais...")
                    data_relatorio = dados.get('data_relatorio')
                    if data_relatorio:
                        data_rel = pd.to_datetime(data_relatorio)
                        df_filtrado = df_original[
                            (df_original['DATA_REL'] == data_rel) & 
                            (df_original['TP_DESP'] != 1)
                        ].copy()
                        logger.info(f"df_filtrado recriado com {len(df_filtrado)} registros")
                    else:
                        logger.error("Não foi possível recriar df_filtrado - data_relatorio não disponível")
                        # Criar DataFrame vazio com colunas mínimas necessárias
                        df_filtrado = pd.DataFrame(columns=['TP_DESP', 'VALOR'])
                else:
                    logger.error("Não foi possível recuperar df_filtrado")
                    # Criar DataFrame vazio com colunas mínimas necessárias
                    df_filtrado = pd.DataFrame(columns=['TP_DESP', 'VALOR'])
        
        subtotais = {}
        
        # Calcular subtotais por tipo de despesa
        for tipo, descricao in self.tipos_despesas.items():
            valor = 0
            
            try:
                if tipo == 1:
                    # Somar todas as despesas de colaboradores (incluindo diárias, férias, rescisão, 13º)
                    valor1 = 0
                    valor2 = 0
                    valor3 = 0
                    
                    # TP_DESP_1 (Salário, Transporte, Café)
                    if not df_tp_desp_1.empty and 'VALOR' in df_tp_desp_1.columns:
                        try:
                            valores_numericos = pd.to_numeric(df_tp_desp_1['VALOR'], errors='coerce').fillna(0)
                            valor1 = valores_numericos.sum()
                        except Exception as e:
                            logger.warning(f"Erro ao somar df_tp_desp_1: {str(e)}")
                            valor1 = 0
                    
                    # TP_DESP_2 (13º, Férias, Rescisão)
                    if not df_tp_desp_2.empty and 'VALOR' in df_tp_desp_2.columns:
                        try:
                            valores_numericos = pd.to_numeric(df_tp_desp_2['VALOR'], errors='coerce').fillna(0)
                            valor2 = valores_numericos.sum()
                        except Exception as e:
                            logger.warning(f"Erro ao somar df_tp_desp_2: {str(e)}")
                            valor2 = 0
                    
                    # Diárias
                    if not df_diaria.empty and 'VALOR' in df_diaria.columns:
                        try:
                            valores_numericos = pd.to_numeric(df_diaria['VALOR'], errors='coerce').fillna(0)
                            valor3 = valores_numericos.sum()
                        except Exception as e:
                            logger.warning(f"Erro ao somar df_diaria: {str(e)}")
                            valor3 = 0
                    
                    valor = valor1 + valor2 + valor3
                    logger.debug(f"Tipo {tipo}: valor1={valor1}, valor2={valor2}, valor3={valor3}, total={valor}")
                    
                else:
                    # Somar outras despesas usando df_filtrado
                    if not df_filtrado.empty and 'TP_DESP' in df_filtrado.columns and 'VALOR' in df_filtrado.columns:
                        try:
                            df_tipo = df_filtrado[df_filtrado['TP_DESP'] == tipo]
                            if not df_tipo.empty:
                                valores_numericos = pd.to_numeric(df_tipo['VALOR'], errors='coerce').fillna(0)
                                valor = valores_numericos.sum()
                            logger.debug(f"Tipo {tipo}: {len(df_tipo)} registros, valor={valor}")
                        except Exception as e:
                            logger.error(f"Erro ao processar tipo {tipo}: {str(e)}")
                            valor = 0
                    else:
                        logger.warning(f"df_filtrado vazio ou sem colunas necessárias para tipo {tipo}")
                        valor = 0
                        
            except Exception as e:
                logger.error(f"Erro geral ao calcular subtotal para tipo {tipo}: {str(e)}")
                valor = 0
                
            subtotais[tipo] = valor

        # Calcular despesas agrupadas
        despesas_a_pagar = sum(subtotais.get(tp, 0) for tp in [1, 2, 3, 4, 7])
        despesas_pagas_cliente = sum(subtotais.get(tp, 0) for tp in [5])
        despesas_pagas_caixa = sum(subtotais.get(tp, 0) for tp in [6])

        total_quinzena = sum(subtotais.values())
        
        # Garantir que temos os valores corretos
        acumulado = dados.get('acumulado', 0)
        numero_relatorio = dados.get('numero_relatorio', 1)
        
        logger.debug(f"Valores para cálculo:")
        logger.debug(f"- Total quinzena: {total_quinzena}")
        logger.debug(f"- Acumulado: {acumulado}")
        
        total_obra = total_quinzena + acumulado

        # Criar tabelas de resumo com formatação consistente
        tabela_subtotais = []
        for tipo, descricao in self.tipos_despesas.items():
            if tipo in subtotais:
                valor_formatado = self.formatar_numero(subtotais[tipo])
                tabela_subtotais.append([descricao, valor_formatado])

        tabela_totais = [
            ['DESPESAS A PAGAR', self.formatar_numero(despesas_a_pagar)],
            ['DESPESAS PAGAS PELO CLIENTE', self.formatar_numero(despesas_pagas_cliente)],
            ['COMPLEMENTO DE CAIXA', self.formatar_numero(despesas_pagas_caixa)],
            [''],
            ['TOTAL DA QUINZENA', self.formatar_numero(total_quinzena)],
            [f'TOTAL ACUMULADO RELATÓRIO Nº {numero_relatorio - 1}',
            self.formatar_numero(acumulado)],
            ['TOTAL DA OBRA', self.formatar_numero(total_obra)]
        ]
        
        logger.debug("Tabela totais criada:")
        for linha in tabela_totais:
            logger.debug(f"Linha: {linha}")

        return tabela_subtotais, tabela_totais

    def adicionar_cabecalho(self, elementos, dados):
        """Adiciona cabeçalho ao relatório PDF - VERSÃO CORRIGIDA"""
        try:
            logger.info("=== INICIANDO ADIÇÃO DO CABEÇALHO ===")
            
            # Verificar se elementos é uma lista válida
            if not isinstance(elementos, list):
                logger.error("ERRO: elementos não é uma lista válida")
                return
            
            # Criar estilo para informações da empresa
            style_empresa = ParagraphStyle(
                'StyleEmpresa',
                parent=self.config.style_normal,
                fontSize=10,
                leading=12,
                alignment=2,  # Alinhamento à direita
                spaceBefore=0,
                spaceAfter=0
            )
            
            # Informações da empresa (texto completo)
            texto_empresa = [
                "Rua Zodiaco, 87 Sala 07 – Santa Lúcia - Belo Horizonte - MG",
                "(31) 3654-6616 / (31) 99974-1241 / (31) 98711-1139",
                "rvr.engenharia@gmail.com"
            ]
            
            # Tentar carregar a logo
            logo_carregada = False
            logo_element = None
            
            # Verificar múltiplos locais para a logo
            possíveis_logos = [
                self.logo_path,  # Caminho atual definido no __init__
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo1.png"),
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png"),
                "logo1.png",
                "logo.png"
            ]
            
            for logo_path in possíveis_logos:
                if logo_path and os.path.exists(logo_path):
                    try:
                        logger.info(f"Tentando carregar logo: {logo_path}")
                        logo_element = Image(logo_path, width=180, height=90)
                        logo_carregada = True
                        logger.info("Logo carregada com sucesso!")
                        break
                    except Exception as e:
                        logger.warning(f"Erro ao carregar logo {logo_path}: {str(e)}")
                        continue
            
            if not logo_carregada:
                logger.warning("Nenhuma logo encontrada, continuando sem logo")
            
            # Criar cabeçalho
            if logo_carregada and logo_element:
                logger.info("Criando cabeçalho COM logo")
                
                # Converter texto da empresa em Paragraphs
                paragraphs_empresa = []
                for linha in texto_empresa:
                    paragraphs_empresa.append(Paragraph(linha, style_empresa))
                
                # Criar tabela com logo + informações
                cabecalho_data = [[logo_element, paragraphs_empresa]]
                
                cabecalho_table = Table(
                    cabecalho_data,
                    colWidths=[200, 550],  # Logo: 200pt, Texto: 550pt
                    rowHeights=[100]
                )
                
                cabecalho_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),     # Logo à esquerda
                    ('ALIGN', (1, 0), (1, 0), 'RIGHT'),    # Texto à direita
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),   # Alinhamento no topo
                    ('LEFTPADDING', (0, 0), (0, 0), 0),
                    ('RIGHTPADDING', (1, 0), (1, 0), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                
                elementos.append(cabecalho_table)
                
            else:
                logger.info("Criando cabeçalho SEM logo")
                
                # Cabeçalho apenas com texto (alinhado à direita)
                for linha in texto_empresa:
                    elementos.append(Paragraph(linha, style_empresa))
                
                # Adicionar espaço extra para compensar a falta da logo
                elementos.append(Spacer(1, 60))
            
            # Espaço após cabeçalho da empresa
            elementos.append(Spacer(1, 25))
            
            # === INFORMAÇÕES DO CLIENTE E RELATÓRIO ===
            try:
                # Formatar data
                if isinstance(dados.get('data_relatorio'), str):
                    data_formatada = dados.get('data_relatorio')
                else:
                    data_formatada = pd.to_datetime(dados.get('data_relatorio')).strftime('%d/%m/%Y')
                
                # Estilo para informações do cliente
                style_cliente_nome = ParagraphStyle(
                    'ClienteNome',
                    parent=self.config.style_normal,
                    fontSize=14,
                    leading=16,
                    alignment=0,  # Esquerda
                    spaceBefore=0,
                    spaceAfter=0,
                    fontName='Helvetica-Bold'
                )
                
                style_cliente_info = ParagraphStyle(
                    'ClienteInfo',
                    parent=self.config.style_normal,
                    fontSize=11,
                    leading=13,
                    alignment=0,  # Esquerda
                    spaceBefore=0,
                    spaceAfter=0
                )
                
                style_relatorio_info = ParagraphStyle(
                    'RelatorioInfo',
                    parent=self.config.style_normal,
                    fontSize=11,
                    leading=13,
                    alignment=2,  # Direita
                    spaceBefore=0,
                    spaceAfter=0
                )
                
                # Dados do cliente e relatório
                info_cliente_data = [
                    [
                        Paragraph(dados.get('nome_cliente', 'CLIENTE NÃO INFORMADO'), style_cliente_nome),
                        Paragraph(f"Relatório nº: {dados.get('numero_relatorio', 'N/A')}", style_relatorio_info)
                    ],
                    [
                        Paragraph(dados.get('endereco_cliente', 'ENDEREÇO NÃO INFORMADO'), style_cliente_info),
                        Paragraph(f"Data: {data_formatada}", style_relatorio_info)
                    ]
                ]
                
                cliente_table = Table(
                    info_cliente_data,
                    colWidths=[550, 200],  # Cliente: 550pt, Relatório: 200pt
                    rowHeights=[25, 25]
                )
                
                cliente_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),     # Coluna cliente à esquerda
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),    # Coluna relatório à direita
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Centralizado verticalmente
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                
                elementos.append(cliente_table)
                logger.info("Informações do cliente adicionadas com sucesso")
                
            except Exception as e:
                logger.error(f"Erro ao processar informações do cliente: {str(e)}")
                # Fallback em caso de erro
                elementos.append(Paragraph("ERRO: Informações do cliente indisponíveis", self.config.style_normal))
            
            # Espaço final antes do conteúdo
            elementos.append(Spacer(1, 30))
            
            logger.info("=== CABEÇALHO CONCLUÍDO COM SUCESSO ===")
            
        except Exception as e:
            logger.error(f"ERRO CRÍTICO no cabeçalho: {str(e)}", exc_info=True)
            # Em caso de erro total, adicionar pelo menos um título
            try:
                elementos.append(Paragraph("SISTEMA DE RELATÓRIOS", self.config.style_heading))
                elementos.append(Spacer(1, 30))
            except:
                pass  # Se nem isso funcionar, continuar sem cabeçalho

    
    def adicionar_detalhes(self, elementos, dados):
        """Adiciona os detalhes das despesas ao relatório - VERSÃO CORRIGIDA"""
        logger.info("Iniciando adição de detalhes ao relatório")

        elementos.append(Paragraph("DETALHES DAS DESPESAS", self.config.style_heading))

        # VERIFICAÇÃO CRÍTICA DOS DADOS
        df_filtrado = dados.get('df_filtrado', pd.DataFrame())
        df_tp_desp_1 = dados.get('df_tp_desp_1', pd.DataFrame())
        df_tp_desp_2 = dados.get('df_tp_desp_2', pd.DataFrame())
        df_diaria = dados.get('df_diaria', pd.DataFrame())
        
        logger.debug(f"=== VERIFICAÇÃO DE DADOS ===")
        logger.debug(f"df_filtrado: {len(df_filtrado)} registros")
        logger.debug(f"df_tp_desp_1: {len(df_tp_desp_1)} registros")
        logger.debug(f"df_tp_desp_2: {len(df_tp_desp_2)} registros")
        logger.debug(f"df_diaria: {len(df_diaria)} registros")
        
        if not df_filtrado.empty:
            logger.debug(f"Colunas df_filtrado: {df_filtrado.columns.tolist()}")
        else:
            logger.warning("df_filtrado está vazio!")
            
        # RECUPERAR df_filtrado SE NECESSÁRIO
        if df_filtrado.empty or 'TP_DESP' not in df_filtrado.columns:
            logger.warning("df_filtrado vazio ou sem TP_DESP, tentando recuperar...")
            
            df_original = dados.get('df_original', pd.DataFrame())
            data_relatorio = dados.get('data_relatorio')
            incluir_excluidos = dados.get('incluir_excluidos', False)
            
            if not df_original.empty and data_relatorio:
                try:
                    data_rel = pd.to_datetime(data_relatorio)
                    
                    # Recriar df_filtrado dos dados originais
                    df_temp = df_original.copy()
                    
                    # Filtrar excluídos se necessário
                    if not incluir_excluidos and 'STATUS' in df_temp.columns:
                        df_temp = df_temp[df_temp['STATUS'] != 'EXCLUIDO'].copy()
                    
                    # Filtrar para obter df_filtrado
                    df_filtrado = df_temp[
                        (df_temp['DATA_REL'] == data_rel) & 
                        (df_temp['TP_DESP'] != 1)
                    ].copy()
                    
                    logger.info(f"df_filtrado recuperado com {len(df_filtrado)} registros")
                    
                    # Atualizar dados para usar o df_filtrado recuperado
                    dados['df_filtrado'] = df_filtrado
                    
                except Exception as e:
                    logger.error(f"Erro ao recuperar df_filtrado: {str(e)}")
                    # Criar DataFrame vazio com estrutura mínima
                    df_filtrado = pd.DataFrame(columns=['TP_DESP', 'NOME', 'REFERÊNCIA', 'VALOR'])

        # 1. Despesas com Colaboradores - Funcionários (Salários, VT e VR)
        if not df_tp_desp_1.empty:
            logger.debug("Processando despesas com colaboradores - funcionários")
            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - SALÁRIO/ADIANTAMENTO, TRANSPORTE E CAFÉ", 
                                self.config.style_despesa))
            
            try:
                df_consolidado = self.consolidar_despesas_colaboradores(df_tp_desp_1)
                logger.debug(f"Total de funcionários processados: {len(df_consolidado)}")

                tabela = self.criar_tabela_despesas(
                    df_consolidado,
                    ['NOME', 'SALÁRIO', 'DIAS', 
                    'TRANSPORTE', 'CAFÉ', 'TOTAL', 'DADOS BANCÁRIOS'],
                    [220, 80, 40, 70, 70, 80, 210]
                )
                elementos.append(tabela)
                elementos.append(Spacer(1, 12))
            except Exception as e:
                logger.error(f"Erro ao processar df_tp_desp_1: {str(e)}")

        # 2. Despesas com Colaboradores - Funcionários (13º, Férias e Rescisão)
        if not df_tp_desp_2.empty:
            logger.debug("Processando despesas com colaboradores - 13º, férias, rescisão")
            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - 13º SALÁRIO, FÉRIAS E RESCISÃO", 
                                self.config.style_despesa))
            
            try:
                df_consolidado1 = self.consolidar_despesas_colaboradores1(df_tp_desp_2)
                logger.debug(f"Total de funcionários processados: {len(df_consolidado1)}")

                tabela = self.criar_tabela_despesas(
                    df_consolidado1,
                    ['NOME', '13º SALÁRIO', 'FÉRIAS', 'RESCISÃO',  
                    'TOTAL', 'DADOS BANCÁRIOS'],
                    [240, 70, 70, 70, 70, 240]
                )
                elementos.append(tabela)
                elementos.append(Spacer(1, 12))
            except Exception as e:
                logger.error(f"Erro ao processar df_tp_desp_2: {str(e)}")
        
        # 3. Despesas com Colaboradores - Diaristas
        if not df_diaria.empty:
            logger.debug("Processando despesas com colaboradores - diaristas")

            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - DIÁRIAS", 
                                self.config.style_despesa))
            
            try:
                # Renomear colunas para corresponder ao formato esperado
                df_diaria_formatado = df_diaria.copy()
                df_diaria_formatado = df_diaria_formatado.rename(columns={
                    'VR_UNIT': 'DIÁRIA',
                    'VALOR': 'TOTAL',
                    'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
                })
                tabela = self.criar_tabela_despesas(
                    df_diaria_formatado,
                    ['NOME', 'DIÁRIA', 'DIAS', 'TOTAL', 'DADOS BANCÁRIOS'],
                    [284, 80, 50, 90, 280]
                )
                elementos.append(tabela)
                elementos.append(Spacer(1, 12))
            except Exception as e:
                logger.error(f"Erro ao processar df_diaria: {str(e)}")

        # 4. Outras despesas - COM VERIFICAÇÃO ROBUSTA
        for tipo in range(2, 8):
            try:
                # Verificar se df_filtrado tem dados e as colunas necessárias
                if df_filtrado.empty:
                    logger.debug(f"Pulando tipo {tipo} - df_filtrado está vazio")
                    continue
                    
                if 'TP_DESP' not in df_filtrado.columns:
                    logger.warning(f"Pulando tipo {tipo} - coluna TP_DESP não encontrada")
                    continue
                    
                df_tipo = df_filtrado[df_filtrado['TP_DESP'] == tipo]
                
                if not df_tipo.empty:
                    logger.debug(f"Processando despesas tipo {tipo} - {len(df_tipo)} registros")
                    elementos.append(Paragraph(self.tipos_despesas[tipo], 
                                        self.config.style_despesa))
                    
                    # Renomear colunas para corresponder ao formato esperado
                    df_tipo = df_tipo.rename(columns={
                        'DT_VENCTO': 'VENCIMENTO',
                        'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
                    })
                    
                    # Verificar se as colunas necessárias existem
                    colunas_necessarias = ['NOME', 'VENCIMENTO', 'REFERÊNCIA', 'VALOR', 'DADOS BANCÁRIOS']
                    colunas_faltantes = [col for col in colunas_necessarias if col not in df_tipo.columns]
                    
                    if colunas_faltantes:
                        logger.warning(f"Tipo {tipo}: Colunas faltantes: {colunas_faltantes}")
                        # Adicionar colunas faltantes com valores vazios
                        for col in colunas_faltantes:
                            df_tipo[col] = ''
                    
                    tabela = self.criar_tabela_despesas(
                        df_tipo,
                        ['NOME', 'VENCIMENTO', 'REFERÊNCIA', 'VALOR', 'DADOS BANCÁRIOS'],
                        [220, 70, 250, 80, 170]
                    )
                    elementos.append(tabela)
                    elementos.append(Spacer(1, 16))
                else:
                    logger.debug(f"Tipo {tipo}: Nenhum registro encontrado")
                    
            except Exception as e:
                logger.error(f"Erro ao processar tipo {tipo}: {str(e)}", exc_info=True)
                continue
                
        logger.info("Detalhes adicionados com sucesso")

    def gerar_relatorio_pdf(self, dados, caminho_output, arquivo_excel):
        """Gera o relatório PDF final"""
        try:
            logger.debug("\nIniciando geração do PDF")
            logger.debug(f"Dados recebidos - acumulado: {dados.get('acumulado')}")
            logger.debug(f"Tipo do acumulado: {type(dados.get('acumulado'))}")
            
            # Criar cópia dos dados para não modificar o original
            dados_pdf = dados.copy()
            
            doc = SimpleDocTemplate(
                caminho_output, 
                pagesize=landscape(A4),
                rightMargin=30,
                leftMargin=30,
                topMargin=40,
                bottomMargin=30
            )
                
            elementos = []
            
            # Adicionar cabeçalho
            self.adicionar_cabecalho(elementos, dados_pdf)
            
            # Adicionar resumo
            elementos.append(Paragraph("RESUMO DAS DESPESAS", self.config.style_heading))
            
            logger.debug("Antes de criar_resumo_despesas:")
            logger.debug(f"Acumulado: {dados_pdf.get('acumulado')}")
            
            tabela_subtotais, tabela_totais = self.criar_resumo_despesas(dados_pdf)
            
            # Log para debug
            logger.debug("Tabela totais gerada:")
            for linha in tabela_totais:
                logger.debug(f"Linha: {linha}")
            
            # Criar tabelas com estilos específicos
            estilo_subtotais = TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ])

            estilo_totais = TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('BOX', (0, 0), (-1, 0), 1, colors.grey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ])

            tabela_esquerda = Table(tabela_subtotais, colWidths=[300, 70])
            tabela_esquerda.setStyle(estilo_subtotais)

            tabela_direita = Table(tabela_totais, colWidths=[180, 70])
            tabela_direita.setStyle(estilo_totais)

            # Criar tabela que combina as duas anteriores
            tabela_resumo = Table(
                [[tabela_esquerda, Spacer(1, 12), tabela_direita]],
                colWidths=[400, 60, 280]
            )
        
            elementos.append(tabela_resumo)
            
            # Adicionar quebra de página
            elementos.append(PageBreak())
            
            # Adicionar detalhes
            self.adicionar_detalhes(elementos, dados)

            if dados.get('incluir_futuros', True) and dados.get('df_futuro') is not None:
                self.adicionar_lancamentos_futuros(elementos, dados)

            # Carregar e processar taxas de administração
            # df_taxas = self.carregar_taxas_administracao(arquivo_excel)
            # if not df_taxas.empty:
            #     df_taxas_processadas = self.processar_taxas_pendentes(df_taxas, dados['data_relatorio'])
            #     if not df_taxas_processadas.empty:
            #         self.adicionar_taxas_administracao(elementos, df_taxas_processadas, self.config)

            # Gerar PDF
            doc.build(elementos)

        except Exception as e:
            logger.error(f"Erro na geração do relatório: {str(e)}", exc_info=True)
            raise
 
    def validar_integridade_dados(self, df, local="DataFrame"):
        """
        Valida a integridade dos dados essenciais
        
        Args:
            df: DataFrame a ser validado
            local: String descritiva do local onde o DataFrame está sendo validado
        
        Returns:
            bool: True se os dados são válidos, False caso contrário
        """
        try:
            logger.debug(f"Validando integridade de dados em: {local}")
            
            # Verificar se o DataFrame não está vazio
            if df.empty:
                logger.warning(f"{local}: DataFrame está vazio")
                return True  # DataFrame vazio é tecnicamente válido
            
            # Verificar colunas essenciais
            colunas_essenciais = ['TP_DESP', 'NOME', 'REFERÊNCIA', 'VALOR', 'DATA_REL']
            colunas_faltantes = [col for col in colunas_essenciais if col not in df.columns]
            
            if colunas_faltantes:
                logger.error(f"{local}: Colunas essenciais ausentes: {colunas_faltantes}")
                logger.error(f"{local}: Colunas disponíveis: {df.columns.tolist()}")
                return False
            
            # Verificar tipos de dados
            if df['TP_DESP'].dtype not in ['int64', 'float64', 'object']:
                logger.warning(f"{local}: Tipo de dado inesperado para TP_DESP: {df['TP_DESP'].dtype}")
            
            # Verificar valores de TP_DESP
            tipos_validos = list(range(1, 8))  # 1 a 7
            tipos_encontrados = df['TP_DESP'].unique()
            tipos_invalidos = [t for t in tipos_encontrados if t not in tipos_validos]
            
            if tipos_invalidos:
                logger.warning(f"{local}: Tipos de despesa inválidos encontrados: {tipos_invalidos}")
            
            # Verificar se há valores nulos em colunas críticas
            for col in ['TP_DESP', 'DATA_REL']:
                nulos = df[col].isnull().sum()
                if nulos > 0:
                    logger.warning(f"{local}: {nulos} valores nulos encontrados na coluna {col}")
            
            logger.debug(f"{local}: Validação concluída - {len(df)} registros, {len(df.columns)} colunas")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao validar integridade de dados em {local}: {str(e)}")
            return False

    def carregar_dados_excel_com_validacao(self, arquivo_excel, incluir_excluidos=False):
        """Versão do carregar_dados_excel com validação robusta"""
        try:
            logger.info(f"Carregando dados de: {arquivo_excel}")
            
            # Carregar dados
            df = pd.read_excel(arquivo_excel, sheet_name='Dados')
            df = df.fillna("")
            
            logger.info(f"Dados carregados: {len(df)} registros, {len(df.columns)} colunas")
            logger.debug(f"Colunas carregadas: {df.columns.tolist()}")
            
            # Verificar colunas necessárias
            colunas_necessarias = {'DATA_REL', 'TP_DESP', 'REFERÊNCIA', 'DT_VENCTO', 'VALOR', 'NF'}
            colunas_faltantes = colunas_necessarias - set(df.columns)
            
            if colunas_faltantes:
                raise ValueError(f"Colunas necessárias ausentes: {colunas_faltantes}")
            
            # Adicionar coluna STATUS se não existir
            if 'STATUS' not in df.columns:
                df['STATUS'] = 'ATIVO'
                logger.info("Coluna STATUS adicionada com valor padrão 'ATIVO'")
            
            # Validar integridade inicial
            if not self.validar_integridade_dados(df, "Dados carregados"):
                raise ValueError("Falha na validação de integridade dos dados carregados")
            
            # Filtrar excluídos se necessário
            if not incluir_excluidos:
                df_original_size = len(df)
                df = df[df['STATUS'] != 'EXCLUIDO'].copy()
                registros_excluidos = df_original_size - len(df)
                if registros_excluidos > 0:
                    logger.info(f"Filtrados {registros_excluidos} registros excluídos")
                print(f"Registros após filtrar excluídos: {len(df)}")
            else:
                print(f"Incluindo todos os registros (incluindo excluídos): {len(df)}")
            
            # Validar após filtragem
            if not self.validar_integridade_dados(df, "Dados após filtragem"):
                raise ValueError("Falha na validação após filtragem de excluídos")
            
            # Converter NF para string antes de processar
            df['NF'] = df['NF'].astype(str)
            
            # Concatenar NF com REFERÊNCIA apenas para TP_DESP != 1
            mascara = (df['TP_DESP'] != 1) & (df['NF'].notna()) & (df['NF'].str.strip() != '') & (df['NF'] != 'nan')
            df.loc[mascara, 'REFERÊNCIA'] = df[mascara].apply(
                lambda row: f"{row['REFERÊNCIA']} (NF: {row['NF'].strip()})", 
                axis=1
            )
            
            # Validação final
            if not self.validar_integridade_dados(df, "Dados finais"):
                raise ValueError("Falha na validação final dos dados")
            
            logger.info(f"Dados carregados e validados com sucesso: {len(df)} registros")
            return df
            
        except Exception as e:
            logger.error(f"Erro ao carregar arquivo Excel: {str(e)}", exc_info=True)
            raise Exception(f"Erro ao carregar arquivo Excel: {str(e)}")

    def gerar_relatorio_pdf_com_validacao(self, dados, caminho_output, arquivo_excel):
        """Versão do gerar_relatorio_pdf com validação de dados"""
        try:
            logger.info("=== INICIANDO GERAÇÃO DE PDF COM VALIDAÇÃO ===")
            
            # Validar dados de entrada
            df_filtrado = dados.get('df_filtrado', pd.DataFrame())
            df_tp_desp_1 = dados.get('df_tp_desp_1', pd.DataFrame())
            df_tp_desp_2 = dados.get('df_tp_desp_2', pd.DataFrame())
            df_diaria = dados.get('df_diaria', pd.DataFrame())
            
            # Validar cada DataFrame
            dataframes_para_validar = [
                (df_filtrado, "df_filtrado"),
                (df_tp_desp_1, "df_tp_desp_1"),
                (df_tp_desp_2, "df_tp_desp_2"),
                (df_diaria, "df_diaria")
            ]
            
            for df, nome in dataframes_para_validar:
                if not df.empty:
                    if not self.validar_integridade_dados(df, nome):
                        logger.error(f"Falha na validação de {nome}")
                        raise ValueError(f"Dados inválidos em {nome}")
            
            # Verificar dados obrigatórios
            campos_obrigatorios = ['nome_cliente', 'data_relatorio', 'numero_relatorio', 'acumulado']
            campos_faltantes = [campo for campo in campos_obrigatorios if campo not in dados]
            
            if campos_faltantes:
                logger.error(f"Campos obrigatórios ausentes: {campos_faltantes}")
                raise ValueError(f"Campos obrigatórios ausentes: {campos_faltantes}")
            
            # Chamar o método original de geração de PDF
            self.gerar_relatorio_pdf_original(dados, caminho_output, arquivo_excel)
            
            logger.info("=== PDF GERADO COM SUCESSO ===")
            
        except Exception as e:
            logger.error(f"Erro na geração de PDF com validação: {str(e)}", exc_info=True)
            raise

    # Método para aplicar as correções na classe RelatorioHandler
    def aplicar_correcoes_relatorio_handler(handler_instance):
        """
        Aplica as correções na instância do RelatorioHandler
        """
        # Backup dos métodos originais
        handler_instance.carregar_dados_excel_original = handler_instance.carregar_dados_excel
        handler_instance.processar_dados_original = handler_instance.processar_dados
        handler_instance.criar_resumo_despesas_original = handler_instance.criar_resumo_despesas
        handler_instance.gerar_relatorio_pdf_original = handler_instance.gerar_relatorio_pdf
        
        # Adicionar novos métodos
        handler_instance.validar_integridade_dados = validar_integridade_dados.__get__(handler_instance)
        handler_instance.carregar_dados_excel = carregar_dados_excel_com_validacao.__get__(handler_instance)
        # handler_instance.processar_dados = processar_dados.__get__(handler_instance) 
        # handler_instance.criar_resumo_despesas = criar_resumo_despesas_corrigido.__get__(handler_instance)
        handler_instance.gerar_relatorio_pdf = gerar_relatorio_pdf_com_validacao.__get__(handler_instance)
        
        logger.info("Correções aplicadas ao RelatorioHandler")
        
class RelatorioLancamentosPendentes:
    def __init__(self):
        self.config = RelatorioConfig()

    def obter_ultima_data_fechamento(self, df):
        """
        Obtém a última data de fechamento (última DATA_REL usada)
        """
        if 'DATA_REL' not in df.columns or df.empty:
            return None
        return pd.to_datetime(df['DATA_REL']).max()

    def processar_arquivo_cliente(self, caminho_arquivo, data_referencia):
        """
        Processa um arquivo de cliente individual
        
        Parameters:
        -----------
        caminho_arquivo : str
            Caminho completo para o arquivo Excel
        data_referencia : datetime
            Data de referência para filtrar lançamentos
            
        Returns:
        --------
        dict ou None
            Dicionário com os dados processados ou None se houver erro
        """
        try:
            print(f"\nProcessando arquivo: {caminho_arquivo}")
            print(f"Data de referência: {data_referencia}")
            
            # Carregar dados do arquivo
            df = pd.read_excel(caminho_arquivo, sheet_name='Dados')
            df = df.fillna("")
            
            wb = load_workbook(caminho_arquivo, data_only=True)
            ws_resumo = wb['RESUMO']
            
            # Obter informações do cliente
            nome_cliente = ws_resumo['A3'].value
            print(f"Cliente: {nome_cliente}")
            
            # Converter DATA_REL para datetime
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], dayfirst=True)
            
            # Garantir que data_referencia seja datetime completo
            if isinstance(data_referencia, date) and not isinstance(data_referencia, datetime):
                # Se é apenas date, converter para datetime com hora zero
                data_referencia = datetime.combine(data_referencia, datetime.min.time())
            elif isinstance(data_referencia, str):
                # Se é string, converter para datetime
                data_referencia = pd.to_datetime(data_referencia)
            
            # Agora ambos são datetime, podemos comparar
            df_pendentes = df[df['DATA_REL'] > data_referencia].copy()
            
            # Remover duplicatas baseado em todas as colunas relevantes
            colunas_check = ['DATA_REL', 'TP_DESP', 'NOME', 'REFERÊNCIA', 'VALOR']
            df_pendentes = df_pendentes.drop_duplicates(subset=colunas_check)
            print(f"Lançamentos encontrados (após remover duplicatas): {len(df_pendentes)}")
            
            if df_pendentes.empty:
                print("Nenhum lançamento pendente encontrado")
                return None
            
            # Identificar parcelamentos
            df_pendentes['is_parcelamento'] = df_pendentes['REFERÊNCIA'].str.contains(
                'parcela|parcelamento', 
                case=False, 
                na=False
            )
            
            # Converter valores para float
            df_pendentes['VALOR'] = pd.to_numeric(
                df_pendentes['VALOR'].astype(str)
                .str.replace('R$', '')
                .str.replace(',', '.')
                .str.strip(), 
                errors='coerce'
            ).fillna(0.0)
            
            # Converter tipo de despesa para inteiro
            df_pendentes['TP_DESP'] = df_pendentes['TP_DESP'].astype(int)
            
            # Formatar datas
            if 'DT_VENCTO' in df_pendentes.columns:
                df_pendentes['DT_VENCTO'] = pd.to_datetime(
                    df_pendentes['DT_VENCTO'], 
                    format='%d/%m/%Y', 
                    errors='coerce'
                )
            
            # Ordenar por data
            df_pendentes = df_pendentes.sort_values(['DATA_REL', 'TP_DESP'])
            
            return {
                'nome_cliente': nome_cliente,
                'ultima_data': data_referencia,
                'lancamentos': df_pendentes,
                'arquivo': caminho_arquivo
            }
            
        except Exception as e:
            print(f"Erro ao processar arquivo {caminho_arquivo}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def gerar_relatorio_html(self, dados_clientes, caminho_saida):
        """
        Gera um relatório HTML com os lançamentos pendentes
        """
        def formatar_valor(valor):
            """Formata valor para o padrão brasileiro"""
            return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
        try:
            # Lista para armazenar as partes do HTML
            html_parts = []
            
            # Cabeçalho do documento
            html_parts.extend([
                '<!DOCTYPE html>',
                '<html>',
                '<head>',
                '<meta charset="utf-8">',
                '<title>Relatório de Lançamentos Pendentes</title>',
                '<style>',
                'body { font-family: Arial, sans-serif; margin: 20px; background-color: #f0f2f5; }',
                'h1 { color: #2c3e50; text-align: center; margin-bottom: 30px; }',
                '.cliente { background-color: white; margin: 20px 0; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }',
                '.cliente-header { background-color: #f8f9fa; padding: 15px; margin: -20px -20px 20px -20px; border-radius: 8px 8px 0 0; border-bottom: 1px solid #dee2e6; }',
                '.cliente-header h2 { margin: 0; color: #2c3e50; }',
                'table { width: 100%; border-collapse: collapse; margin-top: 15px; background-color: white; }',
                'th, td { padding: 12px; text-align: left; border: 1px solid #dee2e6; font-size: 14px; }',
                'th { background-color: #f8f9fa; font-weight: bold; color: #495057; }',
                'tr:nth-child(even) { background-color: #f8f9fa; }',
                '.parcelamento { background-color: #fff3e0; }',
                '.valor { text-align: right; }',
                '.resumo { margin-top: 20px; padding: 15px; background-color: #e8f5e9; border-radius: 5px; font-weight: bold; }',
                '.data-geracao { text-align: center; color: #6c757d; margin-bottom: 30px; }',
                '</style>',
                '</head>',
                '<body>',
                '<h1>Relatório de Lançamentos Pendentes</h1>',
                f'<p class="data-geracao">Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>'
            ])

            # Processar dados de cada cliente
            for dados in dados_clientes:
                if dados is None or dados['lancamentos'].empty:
                    continue

                df = dados['lancamentos']
                total_cliente = df['VALOR'].sum()

                # Cabeçalho do cliente
                html_parts.extend([
                    '<div class="cliente">',
                    '<div class="cliente-header">',
                    f'<h2>{dados["nome_cliente"]}</h2>',
                    f'<p>Última data de fechamento: {dados["ultima_data"].strftime("%d/%m/%Y")}</p>',
                    '</div>',
                    '<table>',
                    '<tr>',
                    '<th>Data</th>',
                    '<th>Tipo</th>',
                    '<th>Nome</th>',
                    '<th>Referência</th>',
                    '<th>Vencimento</th>',
                    '<th>Valor</th>',
                    '</tr>'
                ])

                # Ordenar por data e tipo
                df = df.sort_values(['DATA_REL', 'TP_DESP'])
                
                # Adicionar linhas de dados
                for _, row in df.iterrows():
                    classe = 'parcelamento' if row['is_parcelamento'] else ''
                    valor = float(row['VALOR']) if pd.notnull(row['VALOR']) else 0.0
                    
                    html_parts.extend([
                        f'<tr class="{classe}">',
                        f'<td>{row["DATA_REL"].strftime("%d/%m/%Y")}</td>',
                        f'<td>{int(row["TP_DESP"])}</td>',
                        f'<td>{row["NOME"]}</td>',
                        f'<td>{row["REFERÊNCIA"]}</td>',
                        f'<td>{row["DT_VENCTO"].strftime("%d/%m/%Y") if pd.notnull(row["DT_VENCTO"]) else ""}</td>',
                        f'<td class="valor">{formatar_valor(valor)}</td>',
                        '</tr>'
                    ])

                # Fechar tabela e adicionar resumo
                html_parts.extend([
                    '</table>',
                    '<div class="resumo">',
                    f'<p>Total de lançamentos: R$ {formatar_valor(total_cliente)}</p>',
                    '</div>',
                    '</div>'
                ])

            # Fechar documento HTML
            html_parts.extend([
                '</body>',
                '</html>'
            ])

            # Juntar todas as partes e salvar
            html_content = '\n'.join(html_parts)
            
            with open(caminho_saida, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
            print(f"Relatório HTML gerado com sucesso em: {caminho_saida}")
            
        except Exception as e:
            print(f"Erro ao gerar relatório HTML: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def processar_pasta(self, pasta, data_referencia=None):
        """
        Processa todos os arquivos Excel da pasta
        
        Parameters:
        -----------
        pasta : str
            Caminho da pasta contendo os arquivos Excel
        data_referencia : datetime, optional
            Data de referência para filtrar lançamentos
            
        Returns:
        --------
        list
            Lista com os dados processados de cada cliente
        """
        try:
            print(f"\nProcessando pasta: {pasta}")
            print(f"Data de referência: {data_referencia}")
            
            # Se data_referencia não foi fornecida, usar data atual
            if data_referencia is None:
                data_referencia = datetime.now()
                
            arquivos = [f for f in os.listdir(pasta) if f.endswith('.xlsx')]
            print(f"Encontrados {len(arquivos)} arquivos Excel")
            
            dados_clientes = []
            for arquivo in arquivos:
                caminho_completo = os.path.join(pasta, arquivo)
                dados = self.processar_arquivo_cliente(caminho_completo, data_referencia)
                if dados is not None:
                    dados_clientes.append(dados)
                    
            print(f"Total de clientes processados: {len(dados_clientes)}")
            return dados_clientes
            
        except Exception as e:
            print(f"Erro ao processar pasta: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def gerar_relatorio_pendentes(self, pasta_entrada, arquivo_saida, data_referencia):
        """
        Método principal para gerar o relatório de lançamentos pendentes
        
        Parameters:
        -----------
        self: RelatorioLancamentosPendentes
            Instância da classe
        pasta_entrada : str
            Caminho da pasta com os arquivos dos clientes
        arquivo_saida : str
            Caminho onde o relatório HTML será salvo
        data_referencia : datetime
            Data de referência para filtrar lançamentos
        """
        try:
            print("\nGerando relatório de lançamentos pendentes...")
            print(f"Pasta de entrada: {pasta_entrada}")
            print(f"Arquivo de saída: {arquivo_saida}")
            print(f"Data de referência: {data_referencia}")

            # Garantir que data_referencia é datetime
            if isinstance(data_referencia, date) and not isinstance(data_referencia, datetime):
                data_referencia = datetime.combine(data_referencia, datetime.min.time())
            elif isinstance(data_referencia, str):
                data_referencia = pd.to_datetime(data_referencia)
            
            # Processar todos os arquivos da pasta
            dados_clientes = self.processar_pasta(pasta_entrada, data_referencia)
            
            if not dados_clientes:
                print("Nenhum dado encontrado para processar")
                return False
                
            # Gerar relatório HTML
            self.gerar_relatorio_html(dados_clientes, arquivo_saida)
            
            # Abrir o relatório no navegador padrão
            if platform.system() == 'Darwin':       # macOS
                subprocess.run(['open', arquivo_saida])
            elif platform.system() == 'Windows':    # Windows
                os.startfile(arquivo_saida)
            else:                                   # Linux
                subprocess.run(['xdg-open', arquivo_saida])
                
            return True
            
        except Exception as e:
            print(f"Erro ao gerar relatório de lançamentos pendentes: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

class VisualizadorRelatorio:
    def __init__(self, parent):
        self.parent = parent
        self.dados_preview = None
        self.elementos_preview = None
        
    def formatar_valor(self, valor):
        """Formata valor para o padrão brasileiro"""
        try:
            if pd.isna(valor) or valor == "":
                return "0,00"
            if isinstance(valor, str):
                valor = valor.replace('R$', '').replace(' ', '')
                if ',' in valor and '.' not in valor:
                    valor = valor.replace(',', '.')
                elif ',' in valor and '.' in valor:
                    valor = valor.replace('.', '').replace(',', '.')
                valor = float(valor)
            return f"{float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return "0,00"
    
    def gerar_preview_textual(self, dados):
        """Gera um preview textual do relatório"""
        preview_text = []
        
        # Cabeçalho
        preview_text.append("=" * 80)
        preview_text.append("PREVIEW DO RELATÓRIO DE DESPESAS")
        preview_text.append("=" * 80)
        preview_text.append("")
        
        # Informações do cliente
        preview_text.append(f"CLIENTE: {dados.get('nome_cliente', 'N/A')}")
        preview_text.append(f"ENDEREÇO: {dados.get('endereco_cliente', 'N/A')}")
        
        # Informações do relatório
        data_formatada = dados.get('data_relatorio')
        if hasattr(data_formatada, 'strftime'):
            data_formatada = data_formatada.strftime('%d/%m/%Y')
        preview_text.append(f"RELATÓRIO Nº: {dados.get('numero_relatorio', 'N/A')}")
        preview_text.append(f"DATA: {data_formatada}")
        preview_text.append("")
        
        # Resumo das despesas
        preview_text.append("-" * 50)
        preview_text.append("RESUMO DAS DESPESAS")
        preview_text.append("-" * 50)
        
        # Calcular totais por tipo
        tipos_despesas = {
            1: "DESPESAS COM COLABORADORES",
            2: "TRANSF. PROGR. - MATERIAIS, LOCAÇÕES E PREST.SERVIÇOS", 
            3: "BOLETOS - MATERIAIS, PREST. SERVIÇOS, IMPOSTOS, ETC.",
            4: "RESSARCIMENTOS E RESTITUIÇÕES",
            5: "DESPESAS PAGAS PELO CLIENTE",
            6: "PAGAMENTOS CAIXA DE OBRA",
            7: "ADMINISTRAÇÃO DA OBRA"
        }
        
        subtotais = {}
        
        # Verificar se os DataFrames existem e têm dados
        df_tp_desp_1 = dados.get('df_tp_desp_1', pd.DataFrame())
        df_tp_desp_2 = dados.get('df_tp_desp_2', pd.DataFrame())
        df_diaria = dados.get('df_diaria', pd.DataFrame())
        df_filtrado = dados.get('df_filtrado', pd.DataFrame())
        
        for tipo, descricao in tipos_despesas.items():
            valor = 0
            if tipo == 1:
                # Somar todas as despesas de colaboradores
                valor1 = 0
                valor2 = 0
                valor3 = 0
                
                if not df_tp_desp_1.empty and 'VALOR' in df_tp_desp_1.columns:
                    try:
                        valor1 = pd.to_numeric(df_tp_desp_1['VALOR'], errors='coerce').fillna(0).sum()
                    except:
                        valor1 = 0
                        
                if not df_tp_desp_2.empty and 'VALOR' in df_tp_desp_2.columns:
                    try:
                        valor2 = pd.to_numeric(df_tp_desp_2['VALOR'], errors='coerce').fillna(0).sum()
                    except:
                        valor2 = 0
                        
                if not df_diaria.empty and 'VALOR' in df_diaria.columns:
                    try:
                        valor3 = pd.to_numeric(df_diaria['VALOR'], errors='coerce').fillna(0).sum()
                    except:
                        valor3 = 0
                        
                valor = valor1 + valor2 + valor3
            else:
                # Verificar se df_filtrado tem dados e a coluna TP_DESP existe
                if not df_filtrado.empty and 'TP_DESP' in df_filtrado.columns and 'VALOR' in df_filtrado.columns:
                    try:
                        df_tipo = df_filtrado[df_filtrado['TP_DESP'] == tipo]
                        if not df_tipo.empty:
                            valor = pd.to_numeric(df_tipo['VALOR'], errors='coerce').fillna(0).sum()
                    except Exception as e:
                        print(f"Erro ao processar tipo {tipo}: {str(e)}")
                        valor = 0
            
            subtotais[tipo] = valor
            if valor > 0:  # Só mostra se tiver valor
                preview_text.append(f"{tipo}) {descricao}: R$ {self.formatar_valor(valor)}")
        
        preview_text.append("")
        
        # Totais consolidados
        despesas_a_pagar = sum(subtotais.get(tp, 0) for tp in [1, 2, 3, 4, 7])
        despesas_pagas_cliente = sum(subtotais.get(tp, 0) for tp in [5])
        despesas_pagas_caixa = sum(subtotais.get(tp, 0) for tp in [6])
        total_quinzena = sum(subtotais.values())
        acumulado = dados.get('acumulado', 0)
        total_obra = total_quinzena + acumulado
        
        preview_text.append(f"DESPESAS A PAGAR: R$ {self.formatar_valor(despesas_a_pagar)}")
        preview_text.append(f"DESPESAS PAGAS PELO CLIENTE: R$ {self.formatar_valor(despesas_pagas_cliente)}")
        preview_text.append(f"COMPLEMENTO DE CAIXA: R$ {self.formatar_valor(despesas_pagas_caixa)}")
        preview_text.append("")
        preview_text.append(f"TOTAL DA QUINZENA: R$ {self.formatar_valor(total_quinzena)}")
        preview_text.append(f"TOTAL ACUMULADO RELATÓRIO Nº {dados.get('numero_relatorio', 1) - 1}: R$ {self.formatar_valor(acumulado)}")
        preview_text.append(f"TOTAL DA OBRA: R$ {self.formatar_valor(total_obra)}")
        preview_text.append("")
        
        # Detalhes das despesas
        preview_text.append("-" * 50)
        preview_text.append("DETALHES DAS DESPESAS")
        preview_text.append("-" * 50)
        
        # Colaboradores - Salários, Transporte, Café
        if not df_tp_desp_1.empty:
            preview_text.append("")
            preview_text.append("1) DESPESAS COM COLABORADORES - SALÁRIO/ADIANTAMENTO, TRANSPORTE E CAFÉ")
            preview_text.append("Nome".ljust(25) + "Referência".ljust(15) + "Valor".rjust(15))
            preview_text.append("-" * 55)
            
            for _, row in df_tp_desp_1.iterrows():
                nome = str(row.get('NOME', ''))[:24]
                valor_num = pd.to_numeric(row.get('VALOR', 0), errors='coerce')
                if pd.isna(valor_num):
                    valor_num = 0
                valor = f"R$ {self.formatar_valor(valor_num)}"
                referencia = str(row.get('REFERÊNCIA', ''))[:14]
                preview_text.append(f"{nome.ljust(25)} {referencia.ljust(15)} {valor.rjust(15)}")
        
        # Colaboradores - 13º, Férias, Rescisão
        if not df_tp_desp_2.empty:
            preview_text.append("")
            preview_text.append("1) DESPESAS COM COLABORADORES - 13º SALÁRIO, FÉRIAS E RESCISÃO")
            preview_text.append("Nome".ljust(25) + "Referência".ljust(15) + "Valor".rjust(15))
            preview_text.append("-" * 55)
            
            for _, row in df_tp_desp_2.iterrows():
                nome = str(row.get('NOME', ''))[:24]
                valor_num = pd.to_numeric(row.get('VALOR', 0), errors='coerce')
                if pd.isna(valor_num):
                    valor_num = 0
                valor = f"R$ {self.formatar_valor(valor_num)}"
                referencia = str(row.get('REFERÊNCIA', ''))[:14]
                preview_text.append(f"{nome.ljust(25)} {referencia.ljust(15)} {valor.rjust(15)}")
        
        # Diaristas
        if not df_diaria.empty:
            preview_text.append("")
            preview_text.append("1) DESPESAS COM COLABORADORES - DIÁRIAS")
            preview_text.append("Nome".ljust(25) + "Dias".ljust(8) + "Valor".rjust(15))
            preview_text.append("-" * 48)
            
            for _, row in df_diaria.iterrows():
                nome = str(row.get('NOME', ''))[:24]
                valor_num = pd.to_numeric(row.get('VALOR', 0), errors='coerce')
                if pd.isna(valor_num):
                    valor_num = 0
                valor = f"R$ {self.formatar_valor(valor_num)}"
                dias = str(row.get('DIAS', ''))[:7]
                preview_text.append(f"{nome.ljust(25)} {dias.ljust(8)} {valor.rjust(15)}")
        
        # Outras despesas (tipos 2-7)
        if not df_filtrado.empty and 'TP_DESP' in df_filtrado.columns:
            for tipo in range(2, 8):
                try:
                    df_tipo = df_filtrado[df_filtrado['TP_DESP'] == tipo]
                    if not df_tipo.empty:
                        preview_text.append("")
                        preview_text.append(f"{tipos_despesas[tipo]}")
                        
                        # Verificar se é tipo 5 (ordem especial)
                        if tipo == 5:
                            preview_text.append("(Mantida ordem de entrada dos dados)")
                            
                        preview_text.append("Nome".ljust(25) + "Referência".ljust(30) + "Valor".rjust(15))
                        preview_text.append("-" * 70)
                        
                        for _, row in df_tipo.iterrows():
                            nome = str(row.get('NOME', ''))[:24]
                            referencia = str(row.get('REFERÊNCIA', ''))[:29]
                            valor_num = pd.to_numeric(row.get('VALOR', 0), errors='coerce')
                            if pd.isna(valor_num):
                                valor_num = 0
                            valor = f"R$ {self.formatar_valor(valor_num)}"
                            preview_text.append(f"{nome.ljust(25)} {referencia.ljust(30)} {valor.rjust(15)}")
                except Exception as e:
                    print(f"Erro ao processar tipo {tipo} no preview: {str(e)}")
                    continue
        
        # Lançamentos futuros
        df_futuro = dados.get('df_futuro')
        if dados.get('incluir_futuros') and df_futuro is not None and not df_futuro.empty:
            preview_text.append("")
            preview_text.append("-" * 50)
            preview_text.append("LANÇAMENTOS FUTUROS")
            preview_text.append("-" * 50)
            
            try:
                for periodo in ["Próximos 30 dias", "31 a 60 dias", "Após 60 dias"]:
                    if 'periodo' in df_futuro.columns:
                        df_periodo = df_futuro[df_futuro['periodo'] == periodo]
                        if not df_periodo.empty:
                            preview_text.append(f"\n{periodo}:")
                            for _, row in df_periodo.iterrows():
                                nome = str(row.get('NOME', ''))[:24]
                                referencia = str(row.get('REFERÊNCIA', ''))[:29]
                                valor_num = pd.to_numeric(row.get('VALOR', 0), errors='coerce')
                                if pd.isna(valor_num):
                                    valor_num = 0
                                valor = f"R$ {self.formatar_valor(valor_num)}"
                                preview_text.append(f"  {nome.ljust(25)} {referencia.ljust(30)} {valor.rjust(15)}")
            except Exception as e:
                preview_text.append(f"Erro ao processar lançamentos futuros: {str(e)}")
        
        return "\n".join(preview_text)
    
    def mostrar_preview(self, dados):
        """Mostra a janela de preview do relatório - VERSÃO CORRIGIDA"""
        self.dados_preview = dados
        
        # Criar janela de preview
        preview_window = Toplevel(self.parent)
        preview_window.title("Preview do Relatório")
        preview_window.geometry("900x1000")
        preview_window.transient(self.parent)
        
        # Frame principal
        main_frame = ttk.Frame(preview_window)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Label de título
        title_label = ttk.Label(main_frame, text="Preview do Relatório", 
                            font=('Helvetica', 14, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # Frame para o texto com scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill='both', expand=True)
        
        # Área de texto com scrollbar
        text_widget = Text(text_frame, wrap='none', font=('Courier', 9))
        scrollbar_v = Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        scrollbar_h = Scrollbar(text_frame, orient='horizontal', command=text_widget.xview)
        
        text_widget.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        # Layout dos widgets
        text_widget.grid(row=0, column=0, sticky='nsew')
        scrollbar_v.grid(row=0, column=1, sticky='ns')
        scrollbar_h.grid(row=1, column=0, sticky='ew')
        
        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)
        
        # Gerar e inserir o preview textual
        try:
            preview_text = self.gerar_preview_textual(dados)
            text_widget.insert('1.0', preview_text)
        except Exception as e:
            error_text = f"Erro ao gerar preview: {str(e)}\n\nDados disponíveis:\n"
            for key, value in dados.items():
                if isinstance(value, pd.DataFrame):
                    error_text += f"{key}: DataFrame com {len(value)} linhas\n"
                    if not value.empty:
                        error_text += f"  Colunas: {list(value.columns)}\n"
                else:
                    error_text += f"{key}: {type(value).__name__}\n"
            text_widget.insert('1.0', error_text)
        
        text_widget.configure(state='disabled')  # Somente leitura
        
        # Frame para botões
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(10, 0))
        
        # CORREÇÃO: Função para cancelar e voltar ao menu
        def cancelar_e_voltar():
            """Cancela o preview e volta ao menu principal"""
            try:
                # Fechar janela de preview
                preview_window.destroy()
                
                # Fechar interface atual
                self.parent.destroy()
                
                # Buscar e mostrar menu principal
                menu_principal = self.obter_menu_principal()
                if menu_principal:
                    menu_principal.deiconify()
                    menu_principal.lift()
                    menu_principal.focus_force()
                    logger.info("Cancelado preview - retornando ao menu principal")
                else:
                    logger.warning("Menu principal não encontrado após cancelar preview")
                    
            except Exception as e:
                logger.error(f"Erro ao cancelar e voltar: {str(e)}")
                try:
                    preview_window.destroy()
                    self.parent.destroy()
                except:
                    pass
        
        # Função para continuar editando
        def continuar_editando():
            """Fecha apenas o preview mas mantém a interface atual"""
            preview_window.destroy()
            # Mantém a interface atual aberta para mais modificações
        
        # Botões CORRIGIDOS
        ttk.Button(button_frame, text="Gerar PDF Temporário", 
                command=lambda: self.gerar_pdf_temporario(dados)).pack(side='left', padx=(0, 10))
        
        ttk.Button(button_frame, text="Gerar e Salvar PDF", 
                command=lambda: self.confirmar_geracao(preview_window, dados)).pack(side='left', padx=(0, 10))
        
        # NOVO: Botão para continuar editando
        ttk.Button(button_frame, text="Continuar Editando", 
                command=continuar_editando).pack(side='left', padx=(0, 10))
        
        # CORRIGIDO: Botão cancelar agora volta ao menu
        ttk.Button(button_frame, text="Cancelar e Voltar ao Menu", 
                command=cancelar_e_voltar).pack(side='right')
        
        # CORREÇÃO: Configurar fechamento da janela (X) para também voltar ao menu
        preview_window.protocol("WM_DELETE_WINDOW", cancelar_e_voltar)
        
        # Centralizar janela
        preview_window.transient(self.parent)
        preview_window.grab_set()
        
        return preview_window

    def obter_menu_principal(self):
        """Obtém referência ao menu principal de forma robusta"""
        try:
            # Primeiro: tentar usar a referência direta da interface
            if hasattr(self.parent, 'menu_principal') and self.parent.menu_principal:
                return self.parent.menu_principal
            
            # Segundo: tentar usar a variável global
            menu_global = obter_menu_principal()
            if menu_global:
                return menu_global
            
            # Terceiro: procurar na hierarquia de janelas
            current = self.parent
            while current:
                if hasattr(current, 'menu_principal') and current.menu_principal:
                    return current.menu_principal
                current = getattr(current, 'master', None)
            
            logger.warning("Menu principal não encontrado")
            return None
            
        except Exception as e:
            logger.error(f"Erro ao obter menu principal: {str(e)}")
            return None
    
    def gerar_pdf_temporario(self, dados):
        """Gera um PDF temporário para visualização"""
        try:
            # Criar arquivo temporário
            temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            
            # Usar o handler existente para gerar o PDF
            from relatorio_despesas_aprimorado import RelatorioHandler  # Importar da forma correta
            handler = RelatorioHandler()
            handler.gerar_relatorio_pdf(dados, temp_path, "")
            
            # Abrir o PDF temporário
            self.abrir_arquivo(temp_path)
            
            # Agendar remoção do arquivo temporário após alguns segundos
            self.parent.after(10000, lambda: self.remover_arquivo_temporario(temp_path))
            
        except Exception as e:
            print(f"Erro ao gerar PDF temporário: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def abrir_arquivo(self, caminho):
        """Abre arquivo com o programa padrão do sistema"""
        try:
            if platform.system() == 'Darwin':       # macOS
                subprocess.run(['open', caminho])
            elif platform.system() == 'Windows':    # Windows
                os.startfile(caminho)
            else:                                   # Linux
                subprocess.run(['xdg-open', caminho])
        except Exception as e:
            print(f"Erro ao abrir arquivo: {str(e)}")
    
    def remover_arquivo_temporario(self, caminho):
        """Remove arquivo temporário"""
        try:
            if os.path.exists(caminho):
                os.unlink(caminho)
        except Exception as e:
            print(f"Erro ao remover arquivo temporário: {str(e)}")
    
    def confirmar_geracao(self, preview_window, dados):
        """Confirma a geração do PDF final - VERSÃO MELHORADA"""
        try:
            # Fechar janela de preview
            preview_window.destroy()
            
            # Gerar nome do arquivo
            data_formatada = dados['data_relatorio'].strftime('%d-%m-%Y')
            nome_cliente = dados['nome_cliente']
            nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
            
            if dados.get('incluir_excluidos', False):
                nome_arquivo = nome_arquivo.replace('.pdf', ' (com excluídos).pdf')
                
            # Obter caminho do arquivo
            arquivo_original = getattr(self, 'arquivo_path', '')
            if arquivo_original:
                caminho_output = os.path.join(os.path.dirname(arquivo_original), nome_arquivo)
            else:
                caminho_output = nome_arquivo
            
            # Mostrar mensagem de processamento
            progress_msg = messagebox.showinfo(
                "Processando", 
                "Gerando relatório PDF...\nPor favor, aguarde.",
                icon='info'
            )
            
            # Atualizar interface
            self.parent.update()
            
            # Gerar o PDF
            from relatorio_despesas_aprimorado import RelatorioHandler
            handler = RelatorioHandler()
            handler.gerar_relatorio_pdf(dados, caminho_output, arquivo_original)
            
            # Mostrar mensagem de sucesso COM OPÇÕES
            resposta = messagebox.askyesnocancel(
                "Relatório Gerado com Sucesso!", 
                f"Relatório gerado com sucesso!\n\n"
                f"Cliente: {nome_cliente}\n"
                f"Arquivo: {nome_arquivo}\n\n"
                f"Sim: Abrir o PDF e voltar ao menu\n"
                f"Não: Apenas voltar ao menu\n"
                f"Cancelar: Gerar outro relatório",
                icon='question'
            )
            
            # Processar resposta
            if resposta is True:  # Sim - Abrir PDF e voltar
                self.abrir_arquivo(caminho_output)
                self.voltar_ao_menu_principal()
                
            elif resposta is False:  # Não - Apenas voltar
                self.voltar_ao_menu_principal()
                
            else:  # Cancelar - Manter interface para gerar outro
                # Não faz nada, mantém interface aberta
                pass
            
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar PDF: {str(e)}")
            logger.error(f"Erro ao gerar PDF: {str(e)}")
            return False 

    def voltar_ao_menu_principal(self):
        """Método unificado para voltar ao menu principal"""
        try:
            # Fechar interface atual
            self.parent.destroy()
            
            # Buscar menu principal
            menu_principal = self.obter_menu_principal()
            
            if menu_principal and hasattr(menu_principal, 'winfo_exists'):
                if menu_principal.winfo_exists():
                    menu_principal.deiconify()
                    menu_principal.lift()
                    menu_principal.focus_force()
                    logger.info("Retornado ao menu principal com sucesso")
                    return
            
            # Se não encontrou menu, tentar executar sistema principal
            logger.warning("Menu principal não encontrado, tentando executar sistema principal")
            
            possible_paths = [
                "sistema_principal.py",
                "src/sistema_principal.py"
            ]
            
            for path in possible_paths:
                if os.path.exists(path):
                    subprocess.Popen([sys.executable, path])
                    logger.info(f"Sistema principal executado: {path}")
                    return
            
            logger.error("Sistema principal não encontrado")
            
        except Exception as e:
            logger.error(f"Erro ao voltar ao menu principal: {str(e)}")
            try:
                self.parent.destroy()
            except:
                pass 

    obter_menu_principal = obter_menu_principal
    voltar_ao_menu_principal = voltar_ao_menu_principal

def main():
    # Tentar carregar configurações externas
    config_externa = aplicar_configuracoes_externas()
    
    app = RelatorioUI(None)
    
    # Se há configurações externas, aplicar
    if config_externa:
        try:
            app.data_selecionada.set(config_externa['data'])
            app.incluir_futuros.set(config_externa['incluir_futuros'])
            app.incluir_excluidos.set(config_externa['incluir_excluidos'])
            
            if config_externa['arquivo']:
                app.arquivo_path = config_externa['arquivo']
                app.arquivo_selecionado.set(os.path.basename(config_externa['arquivo']))
            
            if config_externa['arquivos_lote']:
                app.arquivos_lote = config_externa['arquivos_lote']
                
            print("Configurações externas aplicadas com sucesso!")
            
        except Exception as e:
            print(f"Erro ao aplicar configurações externas: {str(e)}")
    
    app.root.mainloop()
