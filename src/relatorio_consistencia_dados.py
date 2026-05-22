import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date as _date
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def add_project_root():
    import sys
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

try:
    from src.config.config import ARQUIVO_CLIENTES, PASTA_CLIENTES
except ImportError:
    from config.config import ARQUIVO_CLIENTES, PASTA_CLIENTES

try:
    from src.config.window_config import configurar_janela
except ImportError:
    def configurar_janela(janela, titulo, largura=1150, altura=820):
        janela.title(titulo)
        sw, sh = janela.winfo_screenwidth(), janela.winfo_screenheight()
        largura = min(largura, sw)
        altura = min(altura, sh)
        janela.geometry(f"{largura}x{altura}+0+0")
        janela.resizable(True, True)
        janela.grid_rowconfigure(0, weight=1)
        janela.grid_columnconfigure(0, weight=1)
        janela.lift()
        janela.focus_force()


# ---------------------------------------------------------------------------
# Funções auxiliares
# ---------------------------------------------------------------------------

def _fmt_moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (TypeError, ValueError):
        return "R$ 0,00"


def _normalizar_cnpj(valor):
    if not valor:
        return ""
    return re.sub(r'\D', '', str(valor))


def _get(row, idx):
    """Lê posição de iter_rows de forma segura (evita IndexError)."""
    try:
        return row[idx]
    except IndexError:
        return None


def _extrair_id_medicao_da_obs(observacao):
    """Extrai o id_medicao da observação 'MEDIÇÃO {id} - ...'"""
    if not observacao:
        return None
    m = re.search(r'MEDI[ÇC][AÃ]O\s+(\S+)', str(observacao), re.IGNORECASE)
    return m.group(1) if m else None


def _extrair_parcelas_da_referencia(referencia):
    """Extrai números de parcela de referências ADM.

    Formatos suportados:
      'TAXA ADM - X/Y' ou 'TAXA ADM - X,Y/Z'  (formato atual)
      '... - PARC. X/Y' ou '... - PARCELA X/Y' (formato legado do Sistema de Entrada de Dados)
    """
    if not referencia:
        return []
    ref = str(referencia)
    # Formato atual: TAXA ADM - X/Y (pode ter múltiplas parcelas separadas por vírgula)
    m = re.search(r'TAXA\s+ADM\s*[-–]\s*([0-9,\s]+)\s*/\s*\d+', ref, re.IGNORECASE)
    if m:
        nums = []
        for p in m.group(1).split(','):
            try:
                nums.append(int(p.strip()))
            except ValueError:
                pass
        return nums
    # Formato legado: "... - PARC. X/Y" ou "... - PARC. X,Y,Z/N" ou "... - PARCELA X/Y"
    # Inclui variante "SINAL + PARC. X,Y,Z/N"
    m = re.search(r'PARC(?:ELA)?\.?\s+([0-9,\s]+)\s*/\s*\d+', ref, re.IGNORECASE)
    if m:
        nums = []
        for p in m.group(1).split(','):
            try:
                nums.append(int(p.strip()))
            except ValueError:
                pass
        return nums
    return []


# ---------------------------------------------------------------------------
# Classe principal
# ---------------------------------------------------------------------------

class RelatorioConsistenciaDados:
    """
    Verificacao de consistencia entre a aba 'Dados' e as abas
    'Medicoes' / 'Contratos_ADM' / 'Contratos_Medicao'.

    Sentido 1 - Dados -> Origem
        Registros ATIVO em 'Dados' cujo CNPJ possui contrato ativo em
        'Contratos_Medicao' (empreiteiros) ou 'Contratos_ADM' (administradores),
        mas sem correspondencia na aba de origem respectiva.

    Sentido 2 - Origem -> Dados
        Medicoes com status LANCADO em 'Medicoes' e parcelas com
        status PAGO/VINCULADO em 'Contratos_ADM' que nao possuem entrada
        correspondente em 'Dados'.
        (Status VINCULADO em Medicoes indica que a medicao foi registrada apos
        o pagamento e vinculada a ele — por definicao ja tem correspondencia em Dados.)
    """

    def __init__(self, parent=None, cliente_inicial=None):
        self.parent = parent
        self.root = tk.Toplevel(parent) if parent else tk.Tk()
        configurar_janela(
            self.root,
            "Verificacao de Consistencia — Dados / Medicoes / Contratos ADM",
            1150, 820
        )
        self.root.protocol("WM_DELETE_WINDOW", self._fechar)

        self.cliente_atual = None
        self.arquivo_cliente = None
        self._cliente_inicial = cliente_inicial

        self._dados_sem_origem_med: list = []
        self._dados_sem_origem_adm: list = []
        self._origem_sem_dados_med: list = []
        self._origem_sem_dados_adm: list = []

        self._setup_gui()

    # ------------------------------------------------------------------
    # GUI
    # ------------------------------------------------------------------

    def _setup_gui(self):
        fp = ttk.Frame(self.root, padding=10)
        fp.pack(fill='both', expand=True)
        fp.rowconfigure(1, weight=1)
        fp.columnconfigure(0, weight=1)

        # Selecao de cliente
        frame_sel = ttk.LabelFrame(fp, text="Selecao de Cliente")
        frame_sel.grid(row=0, column=0, sticky='ew', pady=(0, 6))

        inner = ttk.Frame(frame_sel)
        inner.pack(fill='x', padx=10, pady=8)

        ttk.Label(inner, text="Cliente:", font=('Arial', 11)).pack(side='left')
        self.cmb_cliente = ttk.Combobox(inner, width=50, font=('Arial', 11), state='readonly')
        self.cmb_cliente.pack(side='left', padx=8)
        self.cmb_cliente.bind('<<ComboboxSelected>>', self._on_cliente_selecionado)
        ttk.Button(inner, text="Gerar Relatorio", command=self._gerar_relatorio).pack(side='left', padx=8)

        # Grade 2 x 2 de resultados
        frame_grade = ttk.Frame(fp)
        frame_grade.grid(row=1, column=0, sticky='nsew', pady=(0, 4))
        frame_grade.rowconfigure(0, weight=1)
        frame_grade.rowconfigure(1, weight=1)
        frame_grade.columnconfigure(0, weight=1)
        frame_grade.columnconfigure(1, weight=1)

        # Linha superior — empreiteiros
        frm_dsm = ttk.LabelFrame(frame_grade,
                                  text="Dados s/ correspondencia em Medicoes  (ATIVO)",
                                  padding=4)
        frm_dsm.grid(row=0, column=0, sticky='nsew', padx=(0, 3), pady=(0, 3))

        frm_osm = ttk.LabelFrame(frame_grade,
                                  text="Medicoes LANCADAS s/ entrada em Dados",
                                  padding=4)
        frm_osm.grid(row=0, column=1, sticky='nsew', padx=(3, 0), pady=(0, 3))

        # Linha inferior — administradores
        frm_dsa = ttk.LabelFrame(frame_grade,
                                  text="Dados s/ correspondencia em Contratos ADM  (ATIVO)",
                                  padding=4)
        frm_dsa.grid(row=1, column=0, sticky='nsew', padx=(0, 3), pady=(3, 0))

        frm_osa = ttk.LabelFrame(frame_grade,
                                  text="Parcelas ADM PENDENTE (vencidas e aguardando evento)",
                                  padding=4)
        frm_osa.grid(row=1, column=1, sticky='nsew', padx=(3, 0), pady=(3, 0))

        cols_dados = ('Linha', 'Data', 'CNPJ', 'Nome', 'Referencia', 'Valor', 'Observacao')
        larg_dados = {
            'Linha': 45, 'Data': 80, 'CNPJ': 130, 'Nome': 155,
            'Referencia': 210, 'Valor': 85, 'Observacao': 180,
        }
        self._tree_dsm = self._criar_treeview(frm_dsm, cols_dados, larg_dados)
        self._tree_dsa = self._criar_treeview(frm_dsa, cols_dados, larg_dados)

        cols_med = ('Contrato', 'ID Med.', 'CNPJ', 'Nome', 'Referencia', 'Valor', 'Data Med.')
        larg_med = {
            'Contrato': 60, 'ID Med.': 55, 'CNPJ': 130, 'Nome': 155,
            'Referencia': 210, 'Valor': 85, 'Data Med.': 80,
        }
        self._tree_osm = self._criar_treeview(frm_osm, cols_med, larg_med)

        cols_adm = ('Tipo', 'Contrato', 'Parcela', 'CNPJ', 'Nome', 'Condição', 'Valor', 'Fase')
        larg_adm = {
            'Tipo': 100, 'Contrato': 65, 'Parcela': 50, 'CNPJ': 130,
            'Nome': 140, 'Condição': 120, 'Valor': 85, 'Fase': 180,
        }
        self._tree_osa = self._criar_treeview(frm_osa, cols_adm, larg_adm, selectmode='extended')

        # Barra de resumo
        self._setup_resumo_bar(fp)

        # Botoes inferiores
        frame_bot = ttk.Frame(fp)
        frame_bot.grid(row=3, column=0, sticky='ew', pady=(2, 0))
        ttk.Button(frame_bot, text="Exportar para Excel",
                   command=self._exportar_excel).pack(side='left', padx=4)
        ttk.Button(frame_bot, text="Fechar",
                   command=self._fechar).pack(side='right', padx=4)
        ttk.Button(frame_bot, text="Vincular Selecionados (Dados ↔ Parcelas ADM)",
                   command=self._vincular_selecionados).pack(side='left', padx=12)

        self._carregar_clientes()

    def _setup_resumo_bar(self, parent):
        frame = ttk.LabelFrame(parent, text="Resumo", padding=5)
        frame.grid(row=2, column=0, sticky='ew', pady=(0, 4))

        inner = ttk.Frame(frame)
        inner.pack(fill='x', padx=6, pady=3)

        self._lbl_cliente_res = ttk.Label(
            inner, text="Cliente: —", font=('Arial', 10, 'bold'), foreground='#0056b3')
        self._lbl_cliente_res.pack(side='left', padx=(0, 16))

        ttk.Separator(inner, orient='vertical').pack(side='left', fill='y', padx=6)

        for attr, rotulo in (
            ('_lbl_dsm', 'Dados s/ Med.:'),
            ('_lbl_osm', 'Med. s/ lancto.:'),
            ('_lbl_dsa', 'Dados s/ ADM:'),
            ('_lbl_osa', 'ADM PENDENTE:'),
        ):
            ttk.Label(inner, text=rotulo, font=('Arial', 9)).pack(side='left', padx=(12, 3))
            lbl = ttk.Label(inner, text="—", font=('Arial', 9, 'bold'))
            lbl.pack(side='left')
            setattr(self, attr, lbl)

    @staticmethod
    def _criar_treeview(parent, cols, larguras, selectmode='browse'):
        frame = ttk.Frame(parent)
        frame.pack(fill='both', expand=True, padx=5, pady=5)
        tree = ttk.Treeview(frame, columns=cols, show='headings', selectmode=selectmode)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=larguras.get(c, 100), anchor='w')
        sy = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        sx = ttk.Scrollbar(frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side='right', fill='y')
        sx.pack(side='bottom', fill='x')
        tree.pack(side='left', fill='both', expand=True)
        return tree

    # ------------------------------------------------------------------
    # Carregamento de clientes
    # ------------------------------------------------------------------

    def _carregar_clientes(self):
        try:
            import pandas as pd
            df = pd.read_excel(ARQUIVO_CLIENTES)
            col_nome = next(
                (c for c in ['Nome', 'nome', 'NOME', 'Cliente'] if c in df.columns), None)
            if not col_nome:
                return
            col_final = next(
                (c for c in ['Data Final', 'data_final'] if c in df.columns), None)
            if col_final:
                df = df[df[col_final].isna()]
            nomes = sorted(df[col_nome].dropna().tolist())
            self.cmb_cliente['values'] = nomes
            if self._cliente_inicial and self._cliente_inicial in nomes:
                self.cmb_cliente.set(self._cliente_inicial)
            elif nomes:
                self.cmb_cliente.current(0)
            self._on_cliente_selecionado()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes:\n{e}", parent=self.root)

    def _on_cliente_selecionado(self, event=None):
        nome = self.cmb_cliente.get()
        if not nome:
            return
        self.cliente_atual = nome
        self.arquivo_cliente = Path(PASTA_CLIENTES) / f"{nome}.xlsx"

    # ------------------------------------------------------------------
    # Geracao do relatorio
    # ------------------------------------------------------------------

    def _gerar_relatorio(self):
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente.", parent=self.root)
            return
        if not self.arquivo_cliente or not self.arquivo_cliente.exists():
            messagebox.showerror(
                "Erro", f"Arquivo nao encontrado:\n{self.arquivo_cliente}", parent=self.root)
            return

        try:
            wb = load_workbook(self.arquivo_cliente, data_only=True)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir planilha:\n{e}", parent=self.root)
            return

        abas_req = ('Dados', 'Medicoes', 'Contratos_ADM', 'Contratos_Medicao')
        faltam = [a for a in abas_req if a not in wb.sheetnames]
        if faltam:
            wb.close()
            messagebox.showerror(
                "Erro", f"Abas necessarias nao encontradas: {', '.join(faltam)}", parent=self.root)
            return

        try:
            dados = self._ler_dados(wb['Dados'])
            medicoes = self._ler_medicoes(wb['Medicoes'])
            adm = self._ler_contratos_adm(wb['Contratos_ADM'])
            empreiteiro_cnpjs = self._cnpjs_contratos_medicao(wb['Contratos_Medicao'])
        except Exception as e:
            wb.close()
            messagebox.showerror("Erro", f"Erro ao ler abas:\n{e}", parent=self.root)
            return
        finally:
            wb.close()

        # Sentido 1: Dados ATIVO -> Origem
        self._dados_sem_origem_med = []
        self._dados_sem_origem_adm = []

        for reg in dados:
            if reg['status'].upper() != 'ATIVO':
                continue
            cnpj = reg['cnpj']
            if cnpj in empreiteiro_cnpjs:
                if not self._tem_correspondencia_medicao(reg, medicoes):
                    self._dados_sem_origem_med.append(reg)
            elif cnpj in adm['cnpjs']:
                if not self._tem_correspondencia_adm(reg, adm):
                    self._dados_sem_origem_adm.append(reg)

        # Sentido 2: Origem -> Dados
        self._origem_sem_dados_med = [
            m for m in medicoes['lancadas']
            if not self._medicao_tem_entrada_em_dados(m, dados)
        ]
        # Parcelas PENDENTE vencidas: ja sao vencidas e nao pagas por definicao
        self._origem_sem_dados_adm = list(adm['processadas'])

        self._preencher_tree_dados(self._tree_dsm, self._dados_sem_origem_med)
        self._preencher_tree_dados(self._tree_dsa, self._dados_sem_origem_adm)
        self._preencher_tree_medicoes(self._tree_osm, self._origem_sem_dados_med)
        self._preencher_tree_adm(self._tree_osa, self._origem_sem_dados_adm)
        self._atualizar_resumo()

    # ------------------------------------------------------------------
    # Leitura das abas
    # ------------------------------------------------------------------

    def _ler_dados(self, ws):
        registros = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            cnpj = _normalizar_cnpj(_get(row, 2))
            if not cnpj:
                continue
            data_val = _get(row, 0)
            data_str = (data_val.strftime('%d/%m/%Y')
                        if isinstance(data_val, datetime) else str(data_val or ''))
            registros.append({
                'linha': idx,
                'data': data_str,
                'cnpj': cnpj,
                'cnpj_raw': str(_get(row, 2) or ''),
                'nome': str(_get(row, 3) or '').strip(),
                'referencia': str(_get(row, 4) or '').strip(),
                'valor': _get(row, 8) or 0,
                'observacao': str(_get(row, 12) or '').strip(),
                'status': str(_get(row, 13) or '').strip(),
            })
        return registros

    def _ler_medicoes(self, ws):
        # lancadas: apenas LANCADO — usado para Sentido 2
        # ids/refs/vals: LANCADO + VINCULADO — usados para Sentido 1 (lookup)
        lancadas = []
        ids_set = set()    # (id_medicao, cnpj)
        refs_set = set()   # (cnpj, referencia_lower)
        vals_set = set()   # (cnpj, data_str, valor_arredondado) — fallback para VINCULADO

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not _get(row, 0):
                continue
            status_raw = str(_get(row, 8) or '').strip().upper()
            # normaliza acento: LANÇADO -> LANCADO
            status = status_raw.replace('Ç', 'C').replace('ç', 'c').upper()
            if status not in ('LANCADO', 'VINCULADO'):
                continue

            id_med = str(_get(row, 1) or '').strip()
            cnpj = _normalizar_cnpj(_get(row, 2))
            ref = str(_get(row, 6) or '').strip().lower()

            ids_set.add((id_med, cnpj))
            if ref:
                refs_set.add((cnpj, ref))

            data_val = _get(row, 4)
            data_str = (data_val.strftime('%d/%m/%Y')
                        if isinstance(data_val, datetime) else str(data_val or ''))
            valor = _get(row, 7) or 0
            try:
                vals_set.add((cnpj, data_str, round(float(valor), 2)))
            except (TypeError, ValueError):
                pass

            # Sentido 2: somente LANCADO (VINCULADO ja tem correspondencia em Dados por definicao)
            if status == 'LANCADO':
                lancadas.append({
                    'id_contrato': str(_get(row, 0) or ''),
                    'id_medicao': id_med,
                    'cnpj': cnpj,
                    'cnpj_raw': str(_get(row, 2) or ''),
                    'nome': str(_get(row, 3) or '').strip(),
                    'referencia': str(_get(row, 6) or '').strip(),
                    'valor': valor,
                    'data_medicao': data_str,
                })

        return {'lancadas': lancadas, 'ids': ids_set, 'refs': refs_set, 'vals': vals_set}

    def _ler_contratos_adm(self, ws):
        # processadas:     todas as parcelas PENDENTE (Sentido 2)
        #   tipo='VENCIDA'           → tem Data Vencimento e já passou
        #   tipo='AGUARDANDO_EVENTO' → sem data (pagamento por evento/fase) ou data futura
        # pares:            parcelas PAGO/VINCULADO por (cnpj, num_parcela) (lookup Sentido 1)
        # linhas_vinculadas: linhas da aba Dados referenciadas em parcelas VINCULADO (fallback)
        # cnpjs:            todos os CNPJs presentes em Contratos_ADM
        processadas = []
        pares = set()
        linhas_vinculadas = set()
        cnpjs = set()

        hoje = datetime.now().date()

        for row in ws.iter_rows(min_row=3, values_only=True):
            cnpj_parc = _normalizar_cnpj(_get(row, 26))
            if cnpj_parc:
                cnpjs.add(cnpj_parc)

            if _get(row, 24) is None:
                continue

            status = str(_get(row, 30) or '').strip().upper()
            cnpj = _normalizar_cnpj(_get(row, 26))
            try:
                num_parc = int(_get(row, 25))
            except (TypeError, ValueError):
                continue

            if status in ('PAGO', 'VINCULADO'):
                pares.add((cnpj, num_parc))
                # Extrair linha da aba Dados gravada pela operação de vincular
                obs = str(_get(row, 35) or '')
                m_linha = re.search(
                    r'\[VINCULADO\s+[AÀ]\s+DESPESA\s+DA\s+LINHA\s+(\d+)\s+DE\s+DADOS\]',
                    obs, re.IGNORECASE
                )
                if m_linha:
                    linhas_vinculadas.add(int(m_linha.group(1)))

            elif status == 'PENDENTE':
                venc_raw = _get(row, 28)
                fase = str(_get(row, 32) or '').strip()

                # Normalizar data de vencimento
                venc_date = None
                if isinstance(venc_raw, datetime):
                    venc_date = venc_raw.date()
                elif isinstance(venc_raw, _date):
                    venc_date = venc_raw

                # Determinar tipo e label de condição
                if venc_date is not None and venc_date <= hoje:
                    tipo = 'VENCIDA'
                    condicao = venc_date.strftime('%d/%m/%Y')
                elif venc_date is not None:
                    tipo = 'AGUARDANDO_EVENTO'
                    condicao = f"Vence {venc_date.strftime('%d/%m/%Y')}"
                else:
                    tipo = 'AGUARDANDO_EVENTO'
                    condicao = fase if fase else 'Aguardando evento/fase'

                processadas.append({
                    'num_contrato': str(_get(row, 24) or ''),
                    'num_parcela': num_parc,
                    'cnpj': cnpj,
                    'cnpj_raw': str(_get(row, 26) or ''),
                    'nome': str(_get(row, 27) or '').strip(),
                    'valor': _get(row, 29) or 0,
                    'condicao': condicao,
                    'fase': fase,
                    'tipo': tipo,
                    # mantido para compatibilidade com exportação legada
                    'vencimento': condicao,
                })

        return {
            'processadas': processadas,
            'pares': pares,
            'linhas_vinculadas': linhas_vinculadas,
            'cnpjs': cnpjs,
        }

    def _cnpjs_contratos_medicao(self, ws):
        """CNPJs com contrato ATIVO em Contratos_Medicao (empreiteiros)."""
        cnpjs = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not _get(row, 0):
                continue
            status = str(_get(row, 9) or 'ATIVO').strip().upper()
            if status == 'ATIVO':
                cnpj = _normalizar_cnpj(_get(row, 1))
                if cnpj:
                    cnpjs.add(cnpj)
        return cnpjs

    # ------------------------------------------------------------------
    # Correspondencia
    # ------------------------------------------------------------------

    def _tem_correspondencia_medicao(self, reg, medicoes):
        cnpj = reg['cnpj']
        ref = reg['referencia'].lower()
        id_med = _extrair_id_medicao_da_obs(reg['observacao'])
        if id_med and (id_med, cnpj) in medicoes['ids']:
            return True
        if cnpj and ref and (cnpj, ref) in medicoes['refs']:
            return True
        # Fallback: CNPJ + data + valor (cobre medicoes VINCULADO com ref divergente)
        try:
            val = round(float(reg['valor'] or 0), 2)
            if (cnpj, reg['data'], val) in medicoes['vals']:
                return True
        except (TypeError, ValueError):
            pass
        return False

    def _tem_correspondencia_adm(self, reg, adm):
        # Verificação direta: linha desta entrada foi gravada na observação de parcela VINCULADO
        if reg['linha'] in adm.get('linhas_vinculadas', set()):
            return True
        # Verificação por referência: extrai número(s) de parcela e confere em pares
        cnpj = reg['cnpj']
        parcelas = _extrair_parcelas_da_referencia(reg['referencia'])
        return any((cnpj, n) in adm['pares'] for n in parcelas)

    def _medicao_tem_entrada_em_dados(self, med, dados):
        cnpj = med['cnpj']
        ref = med['referencia'].lower()
        id_med = med['id_medicao']
        try:
            val_med = round(float(med['valor'] or 0), 2)
        except (TypeError, ValueError):
            val_med = None
        for reg in dados:
            if reg['cnpj'] != cnpj:
                continue
            if _extrair_id_medicao_da_obs(reg['observacao']) == id_med:
                return True
            if ref and reg['referencia'].lower() == ref:
                return True
            # Fallback: data + valor (cobre medicoes VINCULADO com ref divergente)
            if val_med is not None:
                try:
                    if (reg['data'] == med['data_medicao']
                            and round(float(reg['valor'] or 0), 2) == val_med):
                        return True
                except (TypeError, ValueError):
                    pass
        return False

    def _parcela_tem_entrada_em_dados(self, parc, dados):
        cnpj = parc['cnpj']
        num = parc['num_parcela']
        for reg in dados:
            if reg['cnpj'] != cnpj:
                continue
            if num in _extrair_parcelas_da_referencia(reg['referencia']):
                return True
        return False

    # ------------------------------------------------------------------
    # Preenchimento de treeviews
    # ------------------------------------------------------------------

    def _preencher_tree_dados(self, tree, lista):
        for item in tree.get_children():
            tree.delete(item)
        for reg in lista:
            tree.insert('', 'end', values=(
                reg['linha'], reg['data'], reg['cnpj_raw'], reg['nome'],
                reg['referencia'], _fmt_moeda(reg['valor']),
                reg['observacao'][:80],
            ))

    def _preencher_tree_medicoes(self, tree, lista):
        for item in tree.get_children():
            tree.delete(item)
        for m in lista:
            tree.insert('', 'end', values=(
                m['id_contrato'], m['id_medicao'], m['cnpj_raw'],
                m['nome'], m['referencia'][:60],
                _fmt_moeda(m['valor']), m['data_medicao'],
            ))

    def _preencher_tree_adm(self, tree, lista):
        tree.tag_configure('vencida', foreground='#cc0000')
        tree.tag_configure('evento', foreground='#b35a00')
        for item in tree.get_children():
            tree.delete(item)
        for p in lista:
            tipo = p.get('tipo', 'VENCIDA')
            rotulo = 'Vencida' if tipo == 'VENCIDA' else 'Ag. evento/fase'
            tag = 'vencida' if tipo == 'VENCIDA' else 'evento'
            tree.insert('', 'end', tags=(tag,), values=(
                rotulo, p['num_contrato'], p['num_parcela'], p['cnpj_raw'],
                p['nome'], p['condicao'], _fmt_moeda(p['valor']),
                p['fase'][:80],
            ))

    def _atualizar_resumo(self):
        self._lbl_cliente_res.config(text=f"Cliente: {self.cliente_atual}")
        for lbl, n in [
            (self._lbl_dsm, len(self._dados_sem_origem_med)),
            (self._lbl_dsa, len(self._dados_sem_origem_adm)),
            (self._lbl_osm, len(self._origem_sem_dados_med)),
            (self._lbl_osa, len(self._origem_sem_dados_adm)),
        ]:
            lbl.config(text=str(n), foreground='red' if n else 'green')

    # ------------------------------------------------------------------
    # Exportacao
    # ------------------------------------------------------------------

    def _exportar_excel(self):
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Gere o relatorio antes de exportar.", parent=self.root)
            return
        tem_dados = any([self._dados_sem_origem_med, self._dados_sem_origem_adm,
                         self._origem_sem_dados_med, self._origem_sem_dados_adm])
        if not tem_dados:
            messagebox.showinfo("Info", "Nenhuma inconsistencia encontrada.", parent=self.root)
            return

        nome_padrao = (
            f"Consistencia_{self.cliente_atual}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        caminho = filedialog.asksaveasfilename(
            defaultextension='.xlsx', filetypes=[('Excel', '*.xlsx')],
            initialfile=nome_padrao, parent=self.root)
        if not caminho:
            return

        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            titulo_font = Font(size=13, bold=True)
            cab_font = Font(size=10, bold=True)
            cab_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

            def _cab(ws, titulo, colunas):
                ws['A1'] = f"Verificacao de Consistencia — {self.cliente_atual}"
                ws['A1'].font = titulo_font
                ws.merge_cells(f'A1:{get_column_letter(len(colunas))}1')
                ws['A2'] = titulo
                ws.merge_cells(f'A2:{get_column_letter(len(colunas))}2')
                ws['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                ws.merge_cells(f'A3:{get_column_letter(len(colunas))}3')
                for ci, col in enumerate(colunas, 1):
                    cel = ws.cell(row=5, column=ci, value=col)
                    cel.font = cab_font
                    cel.fill = cab_fill
                    cel.border = borda
                    cel.alignment = Alignment(horizontal='center')

            cols_d = ['Linha', 'Data', 'CNPJ', 'Nome', 'Referencia', 'Valor (R$)', 'Observacao']

            # Aba 1 — Dados s/ origem (Med.)
            ws1 = wb.create_sheet('Dados s-origem (Med)')
            _cab(ws1, 'Registros ATIVO em Dados sem correspondencia em Medicoes', cols_d)
            for li, r in enumerate(self._dados_sem_origem_med, 6):
                ws1.cell(li, 1, r['linha']); ws1.cell(li, 2, r['data'])
                ws1.cell(li, 3, r['cnpj_raw']); ws1.cell(li, 4, r['nome'])
                ws1.cell(li, 5, r['referencia'])
                ws1.cell(li, 6, float(r['valor'] or 0)).number_format = '#,##0.00'
                ws1.cell(li, 7, r['observacao'])

            # Aba 2 — Dados s/ origem (ADM)
            ws2 = wb.create_sheet('Dados s-origem (ADM)')
            _cab(ws2, 'Registros ATIVO em Dados sem correspondencia em Contratos_ADM', cols_d)
            for li, r in enumerate(self._dados_sem_origem_adm, 6):
                ws2.cell(li, 1, r['linha']); ws2.cell(li, 2, r['data'])
                ws2.cell(li, 3, r['cnpj_raw']); ws2.cell(li, 4, r['nome'])
                ws2.cell(li, 5, r['referencia'])
                ws2.cell(li, 6, float(r['valor'] or 0)).number_format = '#,##0.00'
                ws2.cell(li, 7, r['observacao'])

            # Aba 3 — Medicoes s/ lancamento
            ws3 = wb.create_sheet('Medicoes s-lancamento')
            cols_m = ['Contrato', 'ID Med.', 'CNPJ', 'Nome', 'Referencia',
                      'Valor (R$)', 'Data Med.']
            _cab(ws3, 'Medicoes LANCADAS sem entrada em Dados', cols_m)
            for li, m in enumerate(self._origem_sem_dados_med, 6):
                ws3.cell(li, 1, m['id_contrato']); ws3.cell(li, 2, m['id_medicao'])
                ws3.cell(li, 3, m['cnpj_raw']); ws3.cell(li, 4, m['nome'])
                ws3.cell(li, 5, m['referencia'])
                ws3.cell(li, 6, float(m['valor'] or 0)).number_format = '#,##0.00'
                ws3.cell(li, 7, m['data_medicao'])

            # Aba 4 — Parcelas ADM PENDENTE (vencidas e aguardando evento)
            ws4 = wb.create_sheet('Parcelas ADM PENDENTE')
            cols_a = ['Tipo', 'Contrato', 'Parcela', 'CNPJ', 'Nome', 'Condição', 'Valor (R$)', 'Fase']
            _cab(ws4, 'Parcelas ADM PENDENTE: vencidas (data expirada) e aguardando evento/fase (sem data)', cols_a)
            for li, p in enumerate(self._origem_sem_dados_adm, 6):
                tipo_label = 'Vencida' if p.get('tipo') == 'VENCIDA' else 'Ag. evento/fase'
                ws4.cell(li, 1, tipo_label); ws4.cell(li, 2, p['num_contrato'])
                ws4.cell(li, 3, p['num_parcela']); ws4.cell(li, 4, p['cnpj_raw'])
                ws4.cell(li, 5, p['nome']); ws4.cell(li, 6, p['condicao'])
                ws4.cell(li, 7, float(p['valor'] or 0)).number_format = '#,##0.00'
                ws4.cell(li, 8, p['fase'])

            wb.save(caminho)
            messagebox.showinfo("Sucesso", f"Arquivo exportado:\n{caminho}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar:\n{e}", parent=self.root)

    # ------------------------------------------------------------------
    # Vincular Selecionados
    # ------------------------------------------------------------------

    def _vincular_selecionados(self):
        """Vincula a entrada selecionada em Dados às parcelas ADM PENDENTE selecionadas."""
        sel_dados = self._tree_dsa.selection()
        sel_parc  = self._tree_osa.selection()

        if not sel_dados:
            messagebox.showwarning(
                "Seleção incompleta",
                "Selecione uma linha em 'Dados s/ correspondência em Contratos ADM'.",
                parent=self.root)
            return
        if len(sel_dados) > 1:
            messagebox.showwarning(
                "Seleção inválida",
                "Selecione apenas UMA linha de Dados para vincular.",
                parent=self.root)
            return
        if not sel_parc:
            messagebox.showwarning(
                "Seleção incompleta",
                "Selecione ao menos uma parcela em 'Parcelas ADM PENDENTE'.",
                parent=self.root)
            return

        # Dados da entrada selecionada
        vals_dados = self._tree_dsa.item(sel_dados[0])['values']
        # cols: ('Linha', 'Data', 'CNPJ', 'Nome', 'Referencia', 'Valor', 'Observacao')
        linha_dados = int(vals_dados[0])
        data_dados  = str(vals_dados[1])
        cnpj_dados  = _normalizar_cnpj(str(vals_dados[2]))

        # Dados das parcelas selecionadas
        # cols: ('Tipo', 'Contrato', 'Parcela', 'CNPJ', 'Nome', 'Condição', 'Valor', 'Fase')
        parcelas_sel = []
        for iid in sel_parc:
            v = self._tree_osa.item(iid)['values']
            parcelas_sel.append({
                'num_contrato': str(v[1]),
                'num_parcela' : int(v[2]),
                'cnpj'        : _normalizar_cnpj(str(v[3])),
                'nome'        : str(v[4]),
            })

        # Validar que todos os CNPJs das parcelas coincidem com o da entrada
        cnpjs_parc = {p['cnpj'] for p in parcelas_sel}
        if len(cnpjs_parc) > 1:
            messagebox.showwarning(
                "Seleção inválida",
                "As parcelas selecionadas pertencem a administradores diferentes.\n"
                "Selecione parcelas do mesmo administrador.",
                parent=self.root)
            return
        if cnpj_dados not in cnpjs_parc:
            resp = messagebox.askyesno(
                "CNPJ diferente",
                f"O CNPJ da entrada em Dados ({vals_dados[2]}) é diferente do CNPJ "
                f"das parcelas selecionadas ({parcelas_sel[0]['nome']}).\n\n"
                "Deseja continuar mesmo assim?",
                parent=self.root)
            if not resp:
                return

        # Confirmação final
        lista_parc = ', '.join(str(p['num_parcela']) for p in parcelas_sel)
        resp = messagebox.askyesno(
            "Confirmar vinculação",
            f"Vincular a entrada da linha {linha_dados} de Dados\n"
            f"({vals_dados[3]} — {vals_dados[5]})\n\n"
            f"às parcelas: {lista_parc}\n"
            f"Administrador: {parcelas_sel[0]['nome']}\n\n"
            "Confirma?",
            parent=self.root)
        if not resp:
            return

        # Converter data da entrada para objeto date (usado em Data Pagamento)
        data_pag = None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                from datetime import datetime as _dt
                data_pag = _dt.strptime(data_dados, fmt).date()
                break
            except ValueError:
                pass
        if data_pag is None:
            from datetime import date as _today
            data_pag = _today.today()

        # Executar vinculação na planilha
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']

            vinculadas = 0
            nao_encontradas = []

            for parc in parcelas_sel:
                encontrou = False
                for row_idx, row in enumerate(
                        ws.iter_rows(min_row=3, values_only=False), start=3):
                    num_cont_cell = row[24].value  # col Y
                    num_parc_cell = row[25].value  # col Z
                    cnpj_cell     = row[26].value  # col AA
                    status_cell   = str(row[30].value or '').strip().upper()  # col AE

                    if (str(num_cont_cell or '').strip() == parc['num_contrato'] and
                            status_cell == 'PENDENTE'):
                        try:
                            num_parc_val = int(num_parc_cell)
                        except (TypeError, ValueError):
                            continue
                        if (num_parc_val == parc['num_parcela'] and
                                _normalizar_cnpj(str(cnpj_cell or '')) == parc['cnpj']):
                            # Atualizar status e data pagamento
                            ws.cell(row=row_idx, column=31, value='VINCULADO')
                            ws.cell(row=row_idx, column=32, value=data_pag)
                            ws.cell(row=row_idx, column=32).number_format = 'DD/MM/YYYY'
                            # Registrar linha de Dados na observação
                            obs_atual = str(ws.cell(row=row_idx, column=36).value or '')
                            nova_obs = (obs_atual + f' [VINCULADO À DESPESA DA LINHA '
                                        f'{linha_dados} DE DADOS]').strip()
                            ws.cell(row=row_idx, column=36, value=nova_obs)
                            vinculadas += 1
                            encontrou = True
                            break

                if not encontrou:
                    nao_encontradas.append(parc['num_parcela'])

            wb.save(self.arquivo_cliente)
            wb.close()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar vinculação:\n{e}", parent=self.root)
            return

        # Resultado
        if nao_encontradas:
            messagebox.showwarning(
                "Vinculação parcial",
                f"{vinculadas} parcela(s) vinculada(s).\n"
                f"Não encontradas (já pagas ou contrato divergente): "
                f"{', '.join(str(n) for n in nao_encontradas)}",
                parent=self.root)
        else:
            messagebox.showinfo(
                "Sucesso",
                f"{vinculadas} parcela(s) vinculada(s) à linha {linha_dados} de Dados.",
                parent=self.root)

        # Regenerar relatório
        self._gerar_relatorio()

    # ------------------------------------------------------------------
    # Fechar
    # ------------------------------------------------------------------

    def _fechar(self):
        self.root.destroy()
        if self.parent:
            try:
                self.parent.deiconify()
                self.parent.lift()
                self.parent.focus_force()
            except Exception:
                pass


def main():
    app = RelatorioConsistenciaDados()
    app.root.mainloop()


if __name__ == '__main__':
    main()
