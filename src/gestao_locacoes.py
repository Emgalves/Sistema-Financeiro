# gestao_locacoes.py
"""
Módulo de Gestão de Locações Recorrentes - VERSÃO COMPLETA
===========================================================

ESTRUTURA: Usa ABAS no arquivo Excel existente do cliente

Abas criadas automaticamente:
- LOC_Contratos (contratos de locação)
- LOC_Equipamentos (equipamentos locados)
- LOC_Pagamentos (pagamentos mensais)

Autor: Sistema de Gestão de Medições
Data: Novembro 2025
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from pathlib import Path
import calendar
from openpyxl import load_workbook

# Imports do sistema principal
from src.config.utils import custom_messagebox, PASTA_CLIENTES
from tkcalendar import DateEntry


class GerenciadorLocacoes:
    """Gerenciador principal de locações recorrentes"""
    
    # Nomes das abas
    ABA_CONTRATOS = 'LOC_Contratos'
    ABA_EQUIPAMENTOS = 'LOC_Equipamentos'
    ABA_PAGAMENTOS = 'LOC_Pagamentos'
    
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.janela = None
        self.arquivo_cliente = None
        
        # DataFrames
        self.df_contratos = pd.DataFrame()
        self.df_equipamentos = pd.DataFrame()
        self.df_pagamentos = pd.DataFrame()
        
        # Widgets
        self.tree_contratos = None
        self.tree_equipamentos = None
        self.tree_pagamentos = None
        self.tree_alertas = None
    
    def inicializar_abas_locacoes(self):
        """Inicializa as abas de locações no arquivo Excel do cliente"""
        if not self.sistema.cliente_atual:
            return False
        
        self.arquivo_cliente = PASTA_CLIENTES / f"{self.sistema.cliente_atual}.xlsx"
        
        if not self.arquivo_cliente.exists():
            custom_messagebox("error", "Erro", 
                f"Arquivo do cliente {self.sistema.cliente_atual} não encontrado!")
            return False
        
        self._verificar_criar_abas()
        return True
    
    def _verificar_criar_abas(self):
        """Verifica e cria abas se necessário"""
        try:
            wb = load_workbook(self.arquivo_cliente)
            abas_existentes = wb.sheetnames
            
            # Criar abas faltantes
            if self.ABA_CONTRATOS not in abas_existentes:
                df = self._criar_estrutura_contratos()
                with pd.ExcelWriter(self.arquivo_cliente, engine='openpyxl', 
                                  mode='a') as writer:
                    df.to_excel(writer, sheet_name=self.ABA_CONTRATOS, index=False)
            
            if self.ABA_EQUIPAMENTOS not in abas_existentes:
                df = self._criar_estrutura_equipamentos()
                with pd.ExcelWriter(self.arquivo_cliente, engine='openpyxl', 
                                  mode='a') as writer:
                    df.to_excel(writer, sheet_name=self.ABA_EQUIPAMENTOS, index=False)
            
            if self.ABA_PAGAMENTOS not in abas_existentes:
                df = self._criar_estrutura_pagamentos()
                with pd.ExcelWriter(self.arquivo_cliente, engine='openpyxl', 
                                  mode='a') as writer:
                    df.to_excel(writer, sheet_name=self.ABA_PAGAMENTOS, index=False)
            
        except Exception as e:
            print(f"Erro ao verificar abas: {e}")
            raise
    
    def _criar_estrutura_contratos(self):
        """Estrutura da aba de contratos"""
        return pd.DataFrame(columns=[
            'ID_CONTRATO', 'NUMERO_CONTRATO', 'FORNECEDOR_CNPJ', 'FORNECEDOR_NOME',
            'DATA_INICIO', 'DATA_FIM_ATUAL', 'NUMERO_DIAS', 'CONTRATO_ORIGEM',
            'NUMERO_RENOVACAO', 'RENOVAR', 'FATURA_UNICA', 'STATUS', 
            'TIPO_EQUIPAMENTO', 'DESCRICAO_EQUIPAMENTO', 'IDENTIFICACAO', 
            'QUANTIDADE', 'VALOR_UNITARIO', 'OBSERVACAO', 'CATEGORIA',
            'DATA_CADASTRO', 'ULTIMA_ATUALIZACAO'
        ])
    
    def _criar_estrutura_equipamentos(self):
        """Estrutura da aba de equipamentos"""
        return pd.DataFrame(columns=[
            'ID_EQUIPAMENTO', 'ID_CONTRATO', 'TIPO_EQUIPAMENTO', 'DESCRICAO',
            'IDENTIFICACAO', 'QUANTIDADE', 'VALOR_UNITARIO', 'VALOR_TOTAL',
            'DATA_INICIO_USO', 'DATA_FIM_USO', 'STATUS', 'LOCALIZACAO',
            'OBSERVACAO', 'DATA_CADASTRO', 'ULTIMA_ATUALIZACAO'
        ])
    
    def _criar_estrutura_pagamentos(self):
        """Estrutura da aba de pagamentos"""
        return pd.DataFrame(columns=[
            'ID_PAGAMENTO', 'ID_CONTRATO', 'ID_LANCAMENTO', 'MES_REFERENCIA',
            'DATA_VENCIMENTO', 'DATA_PAGAMENTO', 'VALOR_FATURA', 'NUMERO_NF',
            'STATUS', 'OBSERVACAO', 'DATA_CADASTRO', 'ULTIMA_ATUALIZACAO'
        ])
    
    def carregar_dados(self):
        """Carrega dados das abas"""
        try:
            # Contratos
            try:
                self.df_contratos = pd.read_excel(self.arquivo_cliente, 
                                                  sheet_name=self.ABA_CONTRATOS)
                self.df_contratos = self.df_contratos.fillna("")
            except:
                self.df_contratos = self._criar_estrutura_contratos()
            
            # Equipamentos
            try:
                self.df_equipamentos = pd.read_excel(self.arquivo_cliente, 
                                                     sheet_name=self.ABA_EQUIPAMENTOS)
                self.df_equipamentos = self.df_equipamentos.fillna("")
            except:
                self.df_equipamentos = self._criar_estrutura_equipamentos()
            
            # Pagamentos
            try:
                self.df_pagamentos = pd.read_excel(self.arquivo_cliente, 
                                                   sheet_name=self.ABA_PAGAMENTOS)
                self.df_pagamentos = self.df_pagamentos.fillna("")
            except:
                self.df_pagamentos = self._criar_estrutura_pagamentos()
            
            self.atualizar_interface()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao carregar dados: {e}")
    
    def salvar_dados(self, tipo='todos'):
        """Salva dados nas abas"""
        try:
            with pd.ExcelWriter(self.arquivo_cliente, engine='openpyxl', 
                              mode='a', if_sheet_exists='overlay') as writer:
                
                book = writer.book
                
                if tipo in ['contratos', 'todos']:
                    if self.ABA_CONTRATOS in book.sheetnames:
                        book.remove(book[self.ABA_CONTRATOS])
                    self.df_contratos.to_excel(writer, sheet_name=self.ABA_CONTRATOS, index=False)
                
                if tipo in ['equipamentos', 'todos']:
                    if self.ABA_EQUIPAMENTOS in book.sheetnames:
                        book.remove(book[self.ABA_EQUIPAMENTOS])
                    self.df_equipamentos.to_excel(writer, sheet_name=self.ABA_EQUIPAMENTOS, index=False)
                
                if tipo in ['pagamentos', 'todos']:
                    if self.ABA_PAGAMENTOS in book.sheetnames:
                        book.remove(book[self.ABA_PAGAMENTOS])
                    self.df_pagamentos.to_excel(writer, sheet_name=self.ABA_PAGAMENTOS, index=False)
            
            return True
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar: {e}")
            return False
    
    def abrir_gestao_locacoes(self):
        """Abre janela principal"""
        if not self.sistema.cliente_atual:
            custom_messagebox("error", "Erro", "Selecione um cliente primeiro!")
            return
        
        if not self.inicializar_abas_locacoes():
            return
        
        self.janela = tk.Toplevel(self.sistema.root)
        self.janela.title(f"Gestão de Locações - {self.sistema.cliente_atual}")
        self.janela.geometry("1400x900")
        self.janela.transient(self.sistema.root)
        self.janela.grab_set()
        
        self.criar_interface()
        self.carregar_dados()
    
    def criar_interface(self):
        """Cria interface principal"""
        main_frame = ttk.Frame(self.janela, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Título
        frame_titulo = ttk.Frame(main_frame)
        frame_titulo.pack(fill='x', pady=(0, 10))
        
        ttk.Label(frame_titulo, text="Gestão de Locações Recorrentes", 
                 font=('TkDefaultFont', 14, 'bold')).pack(side='left')
        
        ttk.Label(frame_titulo, text=f"📁 {self.sistema.cliente_atual}.xlsx", 
                 font=('TkDefaultFont', 9), foreground='gray').pack(side='left', padx=(20, 0))
        
        # Resumo
        self.frame_resumo = ttk.LabelFrame(frame_titulo, text="Resumo")
        self.frame_resumo.pack(side='right')
        
        self.label_contratos_ativos = ttk.Label(self.frame_resumo, text="Ativos: 0")
        self.label_contratos_ativos.pack(side='left', padx=10)
        
        self.label_valor_mensal = ttk.Label(self.frame_resumo, text="Total: R$ 0,00")
        self.label_valor_mensal.pack(side='left', padx=10)
        
        self.label_alertas = ttk.Label(self.frame_resumo, text="Alertas: 0", foreground='red')
        self.label_alertas.pack(side='left', padx=10)
        
        # Notebook
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(0, 10))
        
        self.criar_aba_contratos()
        self.criar_aba_equipamentos()
        self.criar_aba_pagamentos()
        self.criar_aba_alertas()
        self.criar_aba_relatorios()
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x')
        
        ttk.Button(frame_botoes, text="🔄 Atualizar", 
                  command=self.carregar_dados).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="💾 Salvar", 
                  command=lambda: self.salvar_dados('todos')).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="📂 Abrir Excel", 
                  command=self.abrir_arquivo_excel).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Fechar", 
                  command=self.janela.destroy).pack(side='right', padx=5)
    
    def criar_aba_contratos(self):
        """Cria aba de contratos"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="📋 Contratos")
        
        # Controles
        frame_ctrl = ttk.Frame(frame)
        frame_ctrl.pack(fill='x', pady=(0, 10))
        
        # Filtros
        frame_filtros = ttk.LabelFrame(frame_ctrl, text="Filtros")
        frame_filtros.pack(side='left', padx=(0, 10))
        
        self.var_filtro_status = tk.StringVar(value="TODOS")
        ttk.Radiobutton(frame_filtros, text="Todos", variable=self.var_filtro_status, 
                       value="TODOS", command=self.atualizar_lista_contratos).pack(side='left', padx=5)
        ttk.Radiobutton(frame_filtros, text="Ativos", variable=self.var_filtro_status, 
                       value="ATIVO", command=self.atualizar_lista_contratos).pack(side='left', padx=5)
        
        # Ações
        frame_acoes = ttk.Frame(frame_ctrl)
        frame_acoes.pack(side='right')
        
        ttk.Button(frame_acoes, text="➕ Novo", 
                  command=self.novo_contrato).pack(side='left', padx=5)
        ttk.Button(frame_acoes, text="✏️ Editar", 
                  command=self.editar_contrato).pack(side='left', padx=5)
        ttk.Button(frame_acoes, text="🔄 Renovar Selecionados", 
                  command=self.renovar_contratos_marcados).pack(side='left', padx=5)
        ttk.Button(frame_acoes, text="☑️ Marcar/Desmarcar", 
                  command=self.toggle_checkbox).pack(side='left', padx=5)
        
        # Lista
        frame_lista = ttk.Frame(frame)
        frame_lista.pack(fill='both', expand=True)
        
        colunas = ('Renovar', 'ID', 'Número', 'Fornecedor', 'Item',
                  'Início', 'Vencimento', 'Dias', 'Qtd', 'Valor Unit.', 'Valor Total', 'Status')
        
        self.tree_contratos = ttk.Treeview(frame_lista, columns=colunas, 
                                          show='headings', height=20)
        
        # Configurar colunas
        self.tree_contratos.heading('Renovar', text='☑')
        self.tree_contratos.heading('ID', text='ID')
        self.tree_contratos.heading('Número', text='Número')
        self.tree_contratos.heading('Fornecedor', text='Fornecedor')
        self.tree_contratos.heading('Item', text='Item')
        self.tree_contratos.heading('Início', text='Início')
        self.tree_contratos.heading('Vencimento', text='Vencimento')
        self.tree_contratos.heading('Dias', text='Dias')
        self.tree_contratos.heading('Qtd', text='Qtd')
        self.tree_contratos.heading('Valor Unit.', text='Valor Unit.')
        self.tree_contratos.heading('Valor Total', text='Valor Total')
        self.tree_contratos.heading('Status', text='Status')
        
        self.tree_contratos.column('Renovar', width=40, anchor='center')
        self.tree_contratos.column('ID', width=40, anchor='center')
        self.tree_contratos.column('Número', width=100)
        self.tree_contratos.column('Fornecedor', width=200)
        self.tree_contratos.column('Item', width=200)  # Aumentado para descrição
        self.tree_contratos.column('Início', width=85, anchor='center')
        self.tree_contratos.column('Vencimento', width=85, anchor='center')
        self.tree_contratos.column('Dias', width=50, anchor='center')
        self.tree_contratos.column('Qtd', width=50, anchor='center')
        self.tree_contratos.column('Valor Unit.', width=100, anchor='e')
        self.tree_contratos.column('Valor Total', width=110, anchor='e')
        self.tree_contratos.column('Status', width=80, anchor='center')
        
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', 
                               command=self.tree_contratos.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient='horizontal',
                               command=self.tree_contratos.xview)
        self.tree_contratos.configure(yscrollcommand=scrolly.set, 
                                     xscrollcommand=scrollx.set)
        
        self.tree_contratos.grid(row=0, column=0, sticky='nsew')
        scrolly.grid(row=0, column=1, sticky='ns')
        scrollx.grid(row=1, column=0, sticky='ew')
        
        frame_lista.grid_rowconfigure(0, weight=1)
        frame_lista.grid_columnconfigure(0, weight=1)
        
        # Tags
        self.tree_contratos.tag_configure('ativo', background='#e8f5e8')
        self.tree_contratos.tag_configure('vencido', background='#ffe4e1')
        self.tree_contratos.tag_configure('alerta', background='#fff8dc')
        self.tree_contratos.tag_configure('renovacao', background='#e6f3ff')
        
        self.tree_contratos.bind('<<TreeviewSelect>>', self.on_select_contrato)
        self.tree_contratos.bind('<Double-1>', self.on_double_click_checkbox)
    
    def criar_aba_equipamentos(self):
        """Cria aba de equipamentos"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="🔧 Equipamentos")
        
        self.label_contrato_sel = ttk.Label(frame, text="Selecione um contrato", 
                                            font=('TkDefaultFont', 10, 'bold'))
        self.label_contrato_sel.pack(pady=10)
        
        # Botões
        frame_btns = ttk.Frame(frame)
        frame_btns.pack(fill='x', pady=5)
        
        ttk.Button(frame_btns, text="➕ Adicionar", 
                  command=self.adicionar_equipamento).pack(side='left', padx=5)
        ttk.Button(frame_btns, text="📦 Devolver", 
                  command=self.devolver_equipamento).pack(side='left', padx=5)
        ttk.Button(frame_btns, text="⚠️ Reportar Perda", 
                  command=self.reportar_perda).pack(side='left', padx=5)
        
        # Lista
        frame_lista = ttk.Frame(frame)
        frame_lista.pack(fill='both', expand=True)
        
        colunas = ('ID', 'Tipo', 'Descrição', 'Identificação', 
                  'Qtd', 'Valor Unit.', 'Valor Total', 'Status')
        
        self.tree_equipamentos = ttk.Treeview(frame_lista, columns=colunas, 
                                             show='headings', height=15)
        
        for col in colunas:
            self.tree_equipamentos.heading(col, text=col)
        
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', 
                               command=self.tree_equipamentos.yview)
        self.tree_equipamentos.configure(yscrollcommand=scrolly.set)
        
        self.tree_equipamentos.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        
        self.tree_equipamentos.tag_configure('em_uso', background='#e8f5e8')
        self.tree_equipamentos.tag_configure('devolvido', background='#e6f3ff')
        self.tree_equipamentos.tag_configure('perdido', background='#ffe4e1')
    
    def criar_aba_pagamentos(self):
        """Cria aba de pagamentos"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="💰 Pagamentos")
        
        # Botões
        frame_btns = ttk.Frame(frame)
        frame_btns.pack(fill='x', pady=5)
        
        ttk.Button(frame_btns, text="➕ Registrar Pagamento", 
                  command=self.registrar_pagamento).pack(side='left', padx=5)
        ttk.Button(frame_btns, text="🔗 Vincular", 
                  command=self.vincular_pagamento).pack(side='left', padx=5)
        
        # Lista
        frame_lista = ttk.Frame(frame)
        frame_lista.pack(fill='both', expand=True)
        
        colunas = ('ID', 'Contrato', 'Fornecedor', 'Mês', 
                  'Vencimento', 'Valor', 'NF', 'Status')
        
        self.tree_pagamentos = ttk.Treeview(frame_lista, columns=colunas, 
                                           show='headings', height=15)
        
        for col in colunas:
            self.tree_pagamentos.heading(col, text=col)
        
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', 
                               command=self.tree_pagamentos.yview)
        self.tree_pagamentos.configure(yscrollcommand=scrolly.set)
        
        self.tree_pagamentos.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        
        self.tree_pagamentos.tag_configure('pago', background='#e8f5e8')
        self.tree_pagamentos.tag_configure('pendente', background='#fff8dc')
        self.tree_pagamentos.tag_configure('atrasado', background='#ffe4e1')
    
    def criar_aba_alertas(self):
        """Cria aba de alertas"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="⚠️ Alertas")
        
        # Resumo
        frame_resumo = ttk.LabelFrame(frame, text="Resumo de Alertas")
        frame_resumo.pack(fill='x', pady=10, padx=10)
        
        self.label_vencendo = ttk.Label(frame_resumo, text="Vencendo: 0", foreground='orange')
        self.label_vencendo.grid(row=0, column=0, padx=20, pady=5)
        
        self.label_vencidos = ttk.Label(frame_resumo, text="Vencidos: 0", foreground='red')
        self.label_vencidos.grid(row=0, column=1, padx=20, pady=5)
        
        self.label_atrasados = ttk.Label(frame_resumo, text="Atrasados: 0", foreground='red')
        self.label_atrasados.grid(row=1, column=0, padx=20, pady=5)
        
        # Lista
        frame_lista = ttk.LabelFrame(frame, text="Alertas Detalhados")
        frame_lista.pack(fill='both', expand=True, padx=10, pady=10)
        
        colunas = ('Tipo', 'Contrato', 'Fornecedor', 'Descrição', 'Data', 'Criticidade')
        
        self.tree_alertas = ttk.Treeview(frame_lista, columns=colunas, 
                                        show='headings', height=15)
        
        for col in colunas:
            self.tree_alertas.heading(col, text=col)
        
        scrolly = ttk.Scrollbar(frame_lista, orient='vertical', 
                               command=self.tree_alertas.yview)
        self.tree_alertas.configure(yscrollcommand=scrolly.set)
        
        self.tree_alertas.pack(side='left', fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        
        self.tree_alertas.tag_configure('critica', background='#ffe4e1')
        self.tree_alertas.tag_configure('alta', background='#fff8dc')
        self.tree_alertas.tag_configure('media', background='#fffacd')
    
    def criar_aba_relatorios(self):
        """Cria aba de relatórios"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="📊 Relatórios")
        
        frame_sel = ttk.LabelFrame(frame, text="Selecionar Relatório")
        frame_sel.pack(fill='x', pady=10, padx=10)
        
        ttk.Button(frame_sel, text="📄 Contratos por Fornecedor", 
                  command=lambda: self.gerar_relatorio('fornecedor')).pack(fill='x', padx=10, pady=5)
        ttk.Button(frame_sel, text="📊 Evolução de Custos", 
                  command=lambda: self.gerar_relatorio('custos')).pack(fill='x', padx=10, pady=5)
        ttk.Button(frame_sel, text="🔧 Equipamentos por Tipo", 
                  command=lambda: self.gerar_relatorio('equipamentos')).pack(fill='x', padx=10, pady=5)
        ttk.Button(frame_sel, text="🎯 Relatório Completo", 
                  command=lambda: self.gerar_relatorio('completo')).pack(fill='x', padx=10, pady=5)
    
    # ============================================================================
    # MÉTODOS DE DADOS
    # ============================================================================
    
    def atualizar_interface(self):
        """Atualiza toda interface"""
        self.atualizar_lista_contratos()
        self.atualizar_resumo()
        self.atualizar_alertas()
    
    def atualizar_lista_contratos(self):
        """Atualiza lista de contratos"""
        for item in self.tree_contratos.get_children():
            self.tree_contratos.delete(item)
        
        if self.df_contratos.empty:
            return
        
        hoje = datetime.now().date()
        filtro = self.var_filtro_status.get()
        
        for idx, row in self.df_contratos.iterrows():
            try:
                status = row.get('STATUS', 'ATIVO')
                
                if filtro != "TODOS" and status != filtro:
                    continue
                
                # Checkbox visual
                renovar = row.get('RENOVAR', '')
                checkbox = '☑' if renovar == 'S' else '☐'
                
                # Determinar tag
                tag = 'ativo'
                num_renovacao = row.get('NUMERO_RENOVACAO', 0)
                if num_renovacao and num_renovacao > 0:
                    tag = 'renovacao'
                elif status == 'VENCIDO':
                    tag = 'vencido'
                elif row.get('DATA_FIM_ATUAL'):
                    try:
                        data_fim = pd.to_datetime(row['DATA_FIM_ATUAL']).date()
                        if 0 < (data_fim - hoje).days <= 7:
                            tag = 'alerta'
                    except:
                        pass
                
                # Calcular valor total
                qtd = float(row.get('QUANTIDADE', 1))
                valor_unit = float(row.get('VALOR_UNITARIO', 0))
                num_dias = int(row.get('NUMERO_DIAS', 30))
                valor_total = qtd * valor_unit * num_dias
                
                # Formatar valores
                valor_unit_fmt = f"R$ {valor_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                valor_total_fmt = f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                # Formatar datas
                data_inicio = row.get('DATA_INICIO', '')
                if data_inicio:
                    try:
                        data_inicio = pd.to_datetime(data_inicio).strftime('%d/%m/%Y')
                    except:
                        pass
                
                data_fim = row.get('DATA_FIM_ATUAL', '')
                if data_fim:
                    try:
                        data_fim = pd.to_datetime(data_fim).strftime('%d/%m/%Y')
                    except:
                        pass
                
                # Número do contrato com renovação
                numero_base = row.get('NUMERO_CONTRATO', '')
                if num_renovacao and num_renovacao > 0:
                    numero_display = f"{numero_base}.{int(num_renovacao)}"
                else:
                    numero_display = numero_base
                
                # Descrição do equipamento (analítico)
                item_desc = row.get('DESCRICAO_EQUIPAMENTO', '')
                
                valores = (
                    checkbox,
                    row.get('ID_CONTRATO', ''),
                    numero_display,
                    row.get('FORNECEDOR_NOME', ''),
                    item_desc,  # Mudado de TIPO_EQUIPAMENTO para DESCRICAO_EQUIPAMENTO
                    data_inicio,
                    data_fim,
                    num_dias,
                    int(qtd),
                    valor_unit_fmt,
                    valor_total_fmt,
                    status
                )
                
                self.tree_contratos.insert('', 'end', values=valores, tags=(tag,))
                
            except Exception as e:
                print(f"Erro ao processar contrato {idx}: {e}")
                continue
    
    def atualizar_lista_equipamentos(self, id_contrato):
        """Atualiza lista de equipamentos do contrato"""
        for item in self.tree_equipamentos.get_children():
            self.tree_equipamentos.delete(item)
        
        if self.df_equipamentos.empty:
            return
        
        equipamentos = self.df_equipamentos[
            self.df_equipamentos['ID_CONTRATO'] == id_contrato
        ]
        
        for idx, row in equipamentos.iterrows():
            try:
                status = row.get('STATUS', 'EM_USO')
                tag = 'em_uso'
                if status == 'DEVOLVIDO':
                    tag = 'devolvido'
                elif status == 'PERDIDO':
                    tag = 'perdido'
                
                valor_unit = float(row.get('VALOR_UNITARIO', 0))
                valor_total = float(row.get('VALOR_TOTAL', 0))
                
                valores = (
                    row.get('ID_EQUIPAMENTO', ''),
                    row.get('TIPO_EQUIPAMENTO', ''),
                    row.get('DESCRICAO', ''),
                    row.get('IDENTIFICACAO', ''),
                    int(row.get('QUANTIDADE', 0)),
                    f"R$ {valor_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    status
                )
                
                self.tree_equipamentos.insert('', 'end', values=valores, tags=(tag,))
                
            except Exception as e:
                print(f"Erro ao processar equipamento: {e}")
                continue
    
    def atualizar_resumo(self):
        """Atualiza resumo no cabeçalho"""
        try:
            ativos = len(self.df_contratos[self.df_contratos['STATUS'] == 'ATIVO'])
            
            # Calcular valor total usando a fórmula: Qtd × Valor Unit. × Dias
            valor_total = 0
            for idx, row in self.df_contratos[self.df_contratos['STATUS'] == 'ATIVO'].iterrows():
                try:
                    qtd = float(row.get('QUANTIDADE', 1))
                    valor_unit = float(row.get('VALOR_UNITARIO', 0))
                    dias = int(row.get('NUMERO_DIAS', 30))
                    valor_total += qtd * valor_unit * dias
                except:
                    pass
            
            self.label_contratos_ativos.config(text=f"Ativos: {ativos}")
            self.label_valor_mensal.config(
                text=f"Total: R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            )
            
        except Exception as e:
            print(f"Erro ao atualizar resumo: {e}")
    
    def atualizar_alertas(self):
        """Atualiza aba de alertas"""
        for item in self.tree_alertas.get_children():
            self.tree_alertas.delete(item)
        
        hoje = datetime.now().date()
        vencendo = 0
        vencidos = 0
        atrasados = 0
        
        # Contratos vencendo/vencidos
        for idx, row in self.df_contratos.iterrows():
            if row.get('STATUS') != 'ATIVO':
                continue
            
            try:
                data_fim = pd.to_datetime(row['DATA_FIM_ATUAL']).date()
                dias = (data_fim - hoje).days
                
                if dias < 0:
                    vencidos += 1
                    self.tree_alertas.insert('', 'end', values=(
                        'Contrato Vencido',
                        row.get('NUMERO_CONTRATO', ''),
                        row.get('FORNECEDOR_NOME', ''),
                        f'Vencido há {abs(dias)} dias',
                        data_fim.strftime('%d/%m/%Y'),
                        'CRÍTICA'
                    ), tags=('critica',))
                elif 0 < dias <= 30:
                    vencendo += 1
                    criticidade = 'ALTA' if dias <= 15 else 'MÉDIA'
                    tag = 'alta' if dias <= 15 else 'media'
                    self.tree_alertas.insert('', 'end', values=(
                        'Contrato Vencendo',
                        row.get('NUMERO_CONTRATO', ''),
                        row.get('FORNECEDOR_NOME', ''),
                        f'Vence em {dias} dias',
                        data_fim.strftime('%d/%m/%Y'),
                        criticidade
                    ), tags=(tag,))
            except:
                continue
        
        # Pagamentos atrasados
        for idx, row in self.df_pagamentos.iterrows():
            if row.get('STATUS') != 'PENDENTE':
                continue
            
            try:
                vencimento = pd.to_datetime(row['DATA_VENCIMENTO']).date()
                if vencimento < hoje:
                    atrasados += 1
                    dias = (hoje - vencimento).days
                    
                    # Buscar contrato
                    contrato = self.df_contratos[
                        self.df_contratos['ID_CONTRATO'] == row.get('ID_CONTRATO')
                    ]
                    
                    if not contrato.empty:
                        self.tree_alertas.insert('', 'end', values=(
                            'Pagamento Atrasado',
                            contrato.iloc[0]['NUMERO_CONTRATO'],
                            contrato.iloc[0]['FORNECEDOR_NOME'],
                            f'Atrasado há {dias} dias',
                            vencimento.strftime('%d/%m/%Y'),
                            'CRÍTICA'
                        ), tags=('critica',))
            except:
                continue
        
        # Atualizar labels
        self.label_vencendo.config(text=f"Vencendo: {vencendo}")
        self.label_vencidos.config(text=f"Vencidos: {vencidos}")
        self.label_atrasados.config(text=f"Atrasados: {atrasados}")
        self.label_alertas.config(text=f"Alertas: {vencendo + vencidos + atrasados}")
    
    # ============================================================================
    # MÉTODOS DE AÇÃO
    # ============================================================================
    
    def on_select_contrato(self, event=None):
        """Ao selecionar contrato"""
        selected = self.tree_contratos.selection()
        if not selected:
            return
        
        valores = self.tree_contratos.item(selected[0], 'values')
        id_contrato = int(valores[1])  # ID está na coluna 1 agora
        
        self.label_contrato_sel.config(
            text=f"Contrato: {valores[2]} - {valores[3]}"
        )
        
        self.atualizar_lista_equipamentos(id_contrato)
    
    def on_double_click_checkbox(self, event):
        """Duplo clique na linha para marcar/desmarcar"""
        # Verificar se clicou na coluna de checkbox
        region = self.tree_contratos.identify_region(event.x, event.y)
        if region == "cell":
            column = self.tree_contratos.identify_column(event.x)
            if column == '#1':  # Primeira coluna (Renovar)
                self.toggle_checkbox()
    
    def toggle_checkbox(self):
        """Marca/desmarca checkbox do contrato selecionado"""
        selected = self.tree_contratos.selection()
        if not selected:
            custom_messagebox("warning", "Aviso", "Selecione um contrato!")
            return
        
        try:
            valores = self.tree_contratos.item(selected[0], 'values')
            id_contrato = int(valores[1])
            
            # Encontrar no DataFrame
            idx = self.df_contratos[self.df_contratos['ID_CONTRATO'] == id_contrato].index[0]
            
            # Toggle
            atual = self.df_contratos.at[idx, 'RENOVAR']
            novo = 'N' if atual == 'S' else 'S'
            self.df_contratos.at[idx, 'RENOVAR'] = novo
            self.df_contratos.at[idx, 'ULTIMA_ATUALIZACAO'] = datetime.now()
            
            # Salvar
            if self.salvar_dados('contratos'):
                self.atualizar_lista_contratos()
                # Manter seleção
                for item in self.tree_contratos.get_children():
                    if self.tree_contratos.item(item, 'values')[1] == str(id_contrato):
                        self.tree_contratos.selection_set(item)
                        self.tree_contratos.see(item)
                        break
        
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao alternar checkbox: {e}")
    
    def renovar_contratos_marcados(self):
        """Renova todos os contratos marcados com checkbox"""
        try:
            # Filtrar contratos marcados e ativos
            marcados = self.df_contratos[
                (self.df_contratos['RENOVAR'] == 'S') & 
                (self.df_contratos['STATUS'] == 'ATIVO')
            ]
            
            if marcados.empty:
                custom_messagebox("warning", "Aviso", 
                    "Nenhum contrato marcado para renovação!\n\n"
                    "Marque os contratos desejados clicando no checkbox ☐")
                return
            
            # Confirmar
            qtd = len(marcados)
            msg = f"Renovar {qtd} contrato(s) marcado(s)?\n\n"
            msg += "Serão criadas novas renovações com:\n"
            msg += "- Número incrementado (.1, .2, etc)\n"
            msg += "- Data inicial = vencimento anterior + 1 dia\n"
            msg += "- Mesmo número de dias do contrato"
            
            if not custom_messagebox("yesno", "Confirmar Renovação", msg):
                return
            
            renovados = []
            erros = []
            
            for idx, contrato in marcados.iterrows():
                try:
                    resultado = self._criar_renovacao_incremental(contrato)
                    if resultado:
                        renovados.append(contrato['NUMERO_CONTRATO'])
                        # Desmarcar checkbox original
                        self.df_contratos.at[idx, 'RENOVAR'] = 'N'
                    else:
                        erros.append(contrato['NUMERO_CONTRATO'])
                except Exception as e:
                    erros.append(f"{contrato['NUMERO_CONTRATO']}: {str(e)}")
            
            # Salvar tudo
            if renovados:
                self.salvar_dados('contratos')
                self.salvar_dados('equipamentos')
            
            # Mensagem de resultado
            msg_result = f"✅ {len(renovados)} contrato(s) renovado(s):\n"
            msg_result += "\n".join([f"  • {num}" for num in renovados[:5]])
            if len(renovados) > 5:
                msg_result += f"\n  ... e mais {len(renovados)-5}"
            
            if erros:
                msg_result += f"\n\n❌ {len(erros)} erro(s):\n"
                msg_result += "\n".join([f"  • {err}" for err in erros[:3]])
            
            custom_messagebox("info", "Renovação Concluída", msg_result)
            self.carregar_dados()
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao renovar: {e}")
            import traceback
            traceback.print_exc()
    
    def _criar_renovacao_incremental(self, contrato_original):
        """
        Cria renovação incremental de um contrato
        
        Retorna: True se sucesso, False se erro
        """
        try:
            # Gerar novo ID
            if self.df_contratos.empty:
                novo_id = 1
            else:
                novo_id = int(self.df_contratos['ID_CONTRATO'].max()) + 1
            
            # Determinar número de renovação
            numero_base = contrato_original['NUMERO_CONTRATO']
            contrato_origem = contrato_original.get('CONTRATO_ORIGEM', '')
            
            if not contrato_origem:
                # É o contrato original, usar ele como origem
                contrato_origem = numero_base
                numero_renovacao = 1
            else:
                # Já é renovação, incrementar
                numero_renovacao = int(contrato_original.get('NUMERO_RENOVACAO', 0)) + 1
            
            # Calcular novas datas
            data_fim_anterior = pd.to_datetime(contrato_original['DATA_FIM_ATUAL']).date()
            nova_data_inicio = data_fim_anterior + timedelta(days=1)
            
            num_dias = int(contrato_original.get('NUMERO_DIAS', 30))
            nova_data_fim = nova_data_inicio + timedelta(days=num_dias - 1)
            
            # Criar novo contrato (renovação)
            nova_renovacao = {
                'ID_CONTRATO': novo_id,
                'NUMERO_CONTRATO': numero_base,  # Mantém número base
                'CONTRATO_ORIGEM': contrato_origem,
                'NUMERO_RENOVACAO': numero_renovacao,
                'FORNECEDOR_CNPJ': contrato_original['FORNECEDOR_CNPJ'],
                'FORNECEDOR_NOME': contrato_original['FORNECEDOR_NOME'],
                'DATA_INICIO': nova_data_inicio,
                'DATA_FIM_ATUAL': nova_data_fim,
                'NUMERO_DIAS': num_dias,
                'RENOVAR': 'N',
                'FATURA_UNICA': contrato_original.get('FATURA_UNICA', ''),
                'STATUS': 'ATIVO',
                'TIPO_EQUIPAMENTO': contrato_original.get('TIPO_EQUIPAMENTO', ''),
                'DESCRICAO_EQUIPAMENTO': contrato_original.get('DESCRICAO_EQUIPAMENTO', ''),
                'IDENTIFICACAO': contrato_original.get('IDENTIFICACAO', ''),
                'QUANTIDADE': contrato_original.get('QUANTIDADE', 1),
                'VALOR_UNITARIO': contrato_original.get('VALOR_UNITARIO', 0),
                'OBSERVACAO': f"Renovação {numero_renovacao} do contrato {contrato_origem}",
                'CATEGORIA': contrato_original.get('CATEGORIA', ''),
                'DATA_CADASTRO': datetime.now(),
                'ULTIMA_ATUALIZACAO': datetime.now()
            }
            
            # Adicionar ao DataFrame
            self.df_contratos = pd.concat([
                self.df_contratos,
                pd.DataFrame([nova_renovacao])
            ], ignore_index=True)
            
            # Copiar equipamento se existir
            if contrato_original.get('TIPO_EQUIPAMENTO'):
                self._criar_equipamento_da_renovacao(novo_id, contrato_original, nova_data_inicio)
            
            return True
            
        except Exception as e:
            print(f"Erro ao criar renovação: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _criar_equipamento_da_renovacao(self, id_contrato_novo, contrato_original, data_inicio):
        """Cria equipamento para renovação"""
        try:
            # Gerar ID
            if self.df_equipamentos.empty:
                novo_id_eq = 1
            else:
                novo_id_eq = int(self.df_equipamentos['ID_EQUIPAMENTO'].max()) + 1
            
            qtd = float(contrato_original.get('QUANTIDADE', 1))
            valor_unit = float(contrato_original.get('VALOR_UNITARIO', 0))
            num_dias = int(contrato_original.get('NUMERO_DIAS', 30))
            valor_total = qtd * valor_unit * num_dias
            
            novo_equipamento = {
                'ID_EQUIPAMENTO': novo_id_eq,
                'ID_CONTRATO': id_contrato_novo,
                'TIPO_EQUIPAMENTO': contrato_original.get('TIPO_EQUIPAMENTO', ''),
                'DESCRICAO': contrato_original.get('DESCRICAO_EQUIPAMENTO', ''),
                'IDENTIFICACAO': contrato_original.get('IDENTIFICACAO', ''),
                'QUANTIDADE': qtd,
                'VALOR_UNITARIO': valor_unit,
                'VALOR_TOTAL': valor_total,
                'DATA_INICIO_USO': data_inicio,
                'DATA_FIM_USO': None,
                'STATUS': 'EM_USO',
                'LOCALIZACAO': '',
                'OBSERVACAO': 'Criado automaticamente na renovação',
                'DATA_CADASTRO': datetime.now(),
                'ULTIMA_ATUALIZACAO': datetime.now()
            }
            
            self.df_equipamentos = pd.concat([
                self.df_equipamentos,
                pd.DataFrame([novo_equipamento])
            ], ignore_index=True)
            
        except Exception as e:
            print(f"Erro ao criar equipamento da renovação: {e}")
    
    def novo_contrato(self):
        """Cria novo contrato com equipamento integrado"""
        janela = tk.Toplevel(self.janela)
        janela.title("Novo Contrato")
        janela.geometry("750x1000")
        janela.transient(self.janela)
        janela.grab_set()
        
        # Frame com scrollbar
        canvas = tk.Canvas(janela)
        scrollbar = ttk.Scrollbar(janela, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas, padding="15")
        
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # ============ DADOS DO CONTRATO ============
        ttk.Label(frame, text="═══ DADOS DO CONTRATO ═══", 
                 font=('TkDefaultFont', 10, 'bold')).pack(anchor='w', pady=(10,5))
        
        ttk.Label(frame, text="Número do Contrato:").pack(anchor='w', pady=2)
        entry_numero = ttk.Entry(frame, width=30)
        entry_numero.pack(anchor='w', pady=2)
        
        # ============ BUSCA DE FORNECEDOR ============
        ttk.Label(frame, text="Fornecedor:", 
                 font=('TkDefaultFont', 10, 'bold')).pack(anchor='w', pady=(10,2))
        
        # Frame de busca
        frame_busca_forn = ttk.Frame(frame)
        frame_busca_forn.pack(anchor='w', pady=2)
        
        ttk.Label(frame_busca_forn, text="Buscar por nome:").pack(side='left', padx=(0,5))
        entry_busca_forn = ttk.Entry(frame_busca_forn, width=30)
        entry_busca_forn.pack(side='left', padx=5)
        
        # Variável para armazenar fornecedor selecionado
        fornecedor_selecionado = {'cnpj': '', 'nome': ''}
        
        # Listbox para resultados
        frame_resultados_forn = ttk.Frame(frame)
        frame_resultados_forn.pack(anchor='w', pady=2)
        
        listbox_forn = tk.Listbox(frame_resultados_forn, width=50, height=5)
        listbox_forn.pack(side='left')
        
        scroll_forn = ttk.Scrollbar(frame_resultados_forn, orient='vertical', 
                                    command=listbox_forn.yview)
        scroll_forn.pack(side='left', fill='y')
        listbox_forn.config(yscrollcommand=scroll_forn.set)
        
        # Esconder listbox inicialmente
        frame_resultados_forn.pack_forget()
        
        # Label para mostrar fornecedor selecionado
        label_forn_selecionado = ttk.Label(frame, text="Nenhum fornecedor selecionado", 
                                          foreground='gray')
        label_forn_selecionado.pack(anchor='w', pady=2)
        
        # Campos ocultos para CNPJ e Nome
        entry_cnpj = ttk.Entry(frame, width=20)
        entry_nome = ttk.Entry(frame, width=50)
        
        def buscar_fornecedor_nome(*args):
            """Busca fornecedor ao digitar"""
            termo = entry_busca_forn.get().strip()
            
            if len(termo) < 3:
                frame_resultados_forn.pack_forget()
                return
            
            try:
                # Usar método do sistema principal
                resultados = self.sistema.buscar_fornecedores_por_nome_parcial(termo)
                
                print(f"DEBUG BUSCA: Termo '{termo}' retornou {len(resultados)} resultados")
                
                # Limpar listbox
                listbox_forn.delete(0, tk.END)
                
                if resultados:
                    # Mostrar listbox SEMPRE que houver resultados
                    frame_resultados_forn.pack(anchor='w', pady=2, after=entry_busca_forn.master)
                    
                    # Adicionar resultados
                    count = 0
                    for forn in resultados[:10]:  # Limitar a 10 resultados
                        try:
                            # Debug: Mostrar estrutura do resultado (só primeira vez)
                            if count == 0:
                                print(f"DEBUG: Estrutura do fornecedor: {forn.keys()}")
                            
                            # CORRIGIDO: Cache usa 'cnpj_cpf' não 'cnpj'
                            cnpj = forn.get('cnpj_cpf', forn.get('cnpj', forn.get('CNPJ', '')))
                            cnpj_fmt = forn.get('cnpj_formatado', forn.get('CNPJ_FORMATADO', ''))
                            nome = forn.get('nome', forn.get('NOME', ''))
                            
                            # Se não tiver formatado, formatar
                            if not cnpj_fmt and cnpj:
                                try:
                                    from src.config.utils import formatar_cnpj_cpf
                                    cnpj_fmt = formatar_cnpj_cpf(cnpj)
                                    print(f"DEBUG: CNPJ formatado: {cnpj} → {cnpj_fmt}")
                                except Exception as e:
                                    print(f"DEBUG: Erro ao formatar CNPJ '{cnpj}': {e}")
                                    cnpj_fmt = cnpj
                            
                            if nome:  # Só precisa do nome
                                # Se não tiver CNPJ, usa "SEM CNPJ"
                                if not cnpj_fmt or not cnpj:
                                    cnpj_fmt = "SEM CNPJ/CPF"
                                
                                linha = f"{cnpj_fmt} - {nome}"
                                listbox_forn.insert(tk.END, linha)
                                count += 1
                                print(f"DEBUG: Linha {count}: {linha}")
                        except Exception as e:
                            print(f"DEBUG: Erro ao processar fornecedor: {e}")
                            continue
                    
                    print(f"DEBUG: Total de {count} itens adicionados à listbox")
                    
                    # Forçar atualização visual
                    listbox_forn.update_idletasks()
                    frame_resultados_forn.update_idletasks()
                    canvas.update_idletasks()
                    
                    # Verificar se realmente tem itens
                    total_items = listbox_forn.size()
                    print(f"DEBUG: Listbox agora tem {total_items} itens")
                    
                else:
                    print("DEBUG: Lista de resultados está vazia")
                    frame_resultados_forn.pack_forget()
                    
            except Exception as e:
                print(f"DEBUG: ERRO CRÍTICO ao buscar fornecedor: {e}")
                import traceback
                traceback.print_exc()
        
        def selecionar_fornecedor_da_lista(event=None):
            """Seleciona fornecedor da lista"""
            try:
                if not listbox_forn.curselection():
                    return
                
                idx = listbox_forn.curselection()[0]
                selecionado = listbox_forn.get(idx)
                
                # Extrair CNPJ e nome
                partes = selecionado.split(' - ', 1)
                if len(partes) == 2:
                    cnpj_fmt = partes[0].strip()
                    nome = partes[1].strip()
                    
                    # Remover formatação do CNPJ
                    cnpj_limpo = ''.join(filter(str.isdigit, cnpj_fmt))
                    
                    # Armazenar
                    fornecedor_selecionado['cnpj'] = cnpj_limpo
                    fornecedor_selecionado['nome'] = nome
                    
                    # Preencher campos ocultos
                    entry_cnpj.delete(0, tk.END)
                    entry_cnpj.insert(0, cnpj_limpo)
                    entry_nome.delete(0, tk.END)
                    entry_nome.insert(0, nome)
                    
                    # Atualizar label
                    label_forn_selecionado.config(
                        text=f"✓ {cnpj_fmt} - {nome}",
                        foreground='green'
                    )
                    
                    # Esconder listbox
                    frame_resultados_forn.pack_forget()
                    entry_busca_forn.delete(0, tk.END)
                    
            except Exception as e:
                print(f"Erro ao selecionar fornecedor: {e}")
        
        # Binds
        entry_busca_forn.bind('<KeyRelease>', buscar_fornecedor_nome)
        listbox_forn.bind('<Double-Button-1>', selecionar_fornecedor_da_lista)
        listbox_forn.bind('<Return>', selecionar_fornecedor_da_lista)
        
        ttk.Label(frame, text="Categoria:").pack(anchor='w', pady=2)
        combo_cat = ttk.Combobox(frame, 
                                values=['EQUIPAMENTOS', 'VEICULOS', 'IMOVEIS', 'FERRAMENTAS'],
                                state='readonly', width=28)
        combo_cat.set('EQUIPAMENTOS')
        combo_cat.pack(anchor='w', pady=2)
        
        ttk.Label(frame, text="Data de Início:").pack(anchor='w', pady=2)
        data_inicio = DateEntry(frame, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_inicio.pack(anchor='w', pady=2)
        
        ttk.Label(frame, text="Número de Dias (1, 7, 14, 30, etc):").pack(anchor='w', pady=2)
        frame_dias = ttk.Frame(frame)
        frame_dias.pack(anchor='w', pady=2)
        
        var_dias = tk.IntVar(value=30)
        ttk.Radiobutton(frame_dias, text="1 dia", variable=var_dias, value=1).pack(side='left', padx=5)
        ttk.Radiobutton(frame_dias, text="7 dias", variable=var_dias, value=7).pack(side='left', padx=5)
        ttk.Radiobutton(frame_dias, text="14 dias", variable=var_dias, value=14).pack(side='left', padx=5)
        ttk.Radiobutton(frame_dias, text="30 dias", variable=var_dias, value=30).pack(side='left', padx=5)
        
        ttk.Label(frame, text="Ou digite manualmente:").pack(anchor='w', pady=2)
        spin_dias = ttk.Spinbox(frame, from_=1, to=999, width=10, textvariable=var_dias)
        spin_dias.pack(anchor='w', pady=2)
        
        # ============ DADOS DO EQUIPAMENTO ============
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=15)
        ttk.Label(frame, text="═══ EQUIPAMENTO ═══", 
                 font=('TkDefaultFont', 10, 'bold')).pack(anchor='w', pady=(5,5))
        
        ttk.Label(frame, text="Tipo de Equipamento:").pack(anchor='w', pady=2)
        combo_tipo = ttk.Combobox(frame, values=[
            'BETONEIRA', 'GUINDASTE', 'ANDAIME', 'ESCORA', 
            'VIBRADOR', 'SERRA', 'GERADOR', 'COMPRESSOR', 
            'EMPILHADEIRA', 'SERRA MARMORE', 'OUTROS'
        ], width=30)
        combo_tipo.pack(anchor='w', pady=2)
        
        ttk.Label(frame, text="Descrição do Equipamento:").pack(anchor='w', pady=2)
        entry_desc = ttk.Entry(frame, width=50)
        entry_desc.pack(anchor='w', pady=2)
        
        ttk.Label(frame, text="Identificação (Nº Série/Placa/Tag):").pack(anchor='w', pady=2)
        entry_id = ttk.Entry(frame, width=30)
        entry_id.pack(anchor='w', pady=2)
        
        ttk.Label(frame, text="Quantidade:").pack(anchor='w', pady=2)
        spin_qtd = ttk.Spinbox(frame, from_=1, to=999, width=10)
        spin_qtd.set('1')
        spin_qtd.pack(anchor='w', pady=2)
        
        ttk.Label(frame, text="Valor Unitário por Dia (R$):").pack(anchor='w', pady=2)
        entry_valor = ttk.Entry(frame, width=15)
        entry_valor.insert(0, "0,00")
        entry_valor.pack(anchor='w', pady=2)
        
        # Cálculo automático
        frame_calculo = ttk.LabelFrame(frame, text="Valor Total")
        frame_calculo.pack(fill='x', pady=10)
        
        label_calculo = ttk.Label(frame_calculo, 
                                 text="R$ 0,00",
                                 font=('TkDefaultFont', 14, 'bold'),
                                 foreground='blue')
        label_calculo.pack(padx=10, pady=5)  # Reduzido de pady=10 para pady=5
        
        def atualizar_calculo(*args):
            try:
                qtd = int(spin_qtd.get())
                valor_str = entry_valor.get().replace('.', '').replace(',', '.')
                valor = float(valor_str)
                dias = var_dias.get()
                
                total = qtd * valor * dias
                total_fmt = f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                label_calculo.config(text=total_fmt)
            except:
                label_calculo.config(text="R$ 0,00")
        
        # Bind para atualizar cálculo
        spin_qtd.bind('<KeyRelease>', atualizar_calculo)
        spin_qtd.bind('<<Increment>>', atualizar_calculo)
        spin_qtd.bind('<<Decrement>>', atualizar_calculo)
        entry_valor.bind('<KeyRelease>', atualizar_calculo)
        spin_dias.bind('<KeyRelease>', atualizar_calculo)
        spin_dias.bind('<<Increment>>', atualizar_calculo)
        spin_dias.bind('<<Decrement>>', atualizar_calculo)
        
        # ============ CHECKBOXES ============
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=15)
        
        var_renovar = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame, text="☑ Marcar para renovação automática", 
                       variable=var_renovar).pack(anchor='w', pady=5)
        
        var_fatura = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame, text="☑ Agrupar em fatura única", 
                       variable=var_fatura).pack(anchor='w', pady=5)
        
        ttk.Label(frame, text="Observações:").pack(anchor='w', pady=5)
        text_obs = tk.Text(frame, width=50, height=2)  # Reduzido de 3 para 2
        text_obs.pack(anchor='w', pady=2)
        
        def salvar():
            try:
                # Validações
                if not entry_numero.get().strip():
                    custom_messagebox("error", "Erro", "Número do contrato é obrigatório!")
                    return
                
                if not fornecedor_selecionado['cnpj']:
                    custom_messagebox("error", "Erro", 
                        "Selecione um fornecedor!\n\n"
                        "Digite o nome e selecione da lista.")
                    return
                
                if not combo_tipo.get():
                    custom_messagebox("error", "Erro", "Tipo de equipamento é obrigatório!")
                    return
                
                # Gerar ID
                if self.df_contratos.empty:
                    novo_id = 1
                else:
                    novo_id = int(self.df_contratos['ID_CONTRATO'].max()) + 1
                
                # Calcular data fim
                num_dias = var_dias.get()
                data_ini = data_inicio.get_date()
                data_fim = data_ini + timedelta(days=num_dias - 1)
                
                # Valores
                qtd = float(spin_qtd.get())
                valor_unit = float(entry_valor.get().replace('.', '').replace(',', '.'))
                
                # Criar contrato
                novo = {
                    'ID_CONTRATO': novo_id,
                    'NUMERO_CONTRATO': entry_numero.get().strip().upper(),
                    'CONTRATO_ORIGEM': '',
                    'NUMERO_RENOVACAO': 0,
                    'FORNECEDOR_CNPJ': fornecedor_selecionado['cnpj'],  # Usar fornecedor selecionado
                    'FORNECEDOR_NOME': fornecedor_selecionado['nome'],  # Usar fornecedor selecionado
                    'DATA_INICIO': data_ini,
                    'DATA_FIM_ATUAL': data_fim,
                    'NUMERO_DIAS': num_dias,
                    'RENOVAR': 'S' if var_renovar.get() else 'N',
                    'FATURA_UNICA': 'S' if var_fatura.get() else 'N',
                    'STATUS': 'ATIVO',
                    'TIPO_EQUIPAMENTO': combo_tipo.get().upper(),
                    'DESCRICAO_EQUIPAMENTO': entry_desc.get().strip().upper(),
                    'IDENTIFICACAO': entry_id.get().strip().upper(),
                    'QUANTIDADE': qtd,
                    'VALOR_UNITARIO': valor_unit,
                    'OBSERVACAO': text_obs.get('1.0', 'end-1c').strip(),
                    'CATEGORIA': combo_cat.get(),
                    'DATA_CADASTRO': datetime.now(),
                    'ULTIMA_ATUALIZACAO': datetime.now()
                }
                
                # Adicionar contrato
                self.df_contratos = pd.concat([
                    self.df_contratos,
                    pd.DataFrame([novo])
                ], ignore_index=True)
                
                # Criar equipamento
                if self.df_equipamentos.empty:
                    novo_id_eq = 1
                else:
                    novo_id_eq = int(self.df_equipamentos['ID_EQUIPAMENTO'].max()) + 1
                
                valor_total = qtd * valor_unit * num_dias
                
                novo_eq = {
                    'ID_EQUIPAMENTO': novo_id_eq,
                    'ID_CONTRATO': novo_id,
                    'TIPO_EQUIPAMENTO': combo_tipo.get().upper(),
                    'DESCRICAO': entry_desc.get().strip().upper(),
                    'IDENTIFICACAO': entry_id.get().strip().upper(),
                    'QUANTIDADE': qtd,
                    'VALOR_UNITARIO': valor_unit,
                    'VALOR_TOTAL': valor_total,
                    'DATA_INICIO_USO': data_ini,
                    'DATA_FIM_USO': None,
                    'STATUS': 'EM_USO',
                    'LOCALIZACAO': '',
                    'OBSERVACAO': 'Criado com o contrato',
                    'DATA_CADASTRO': datetime.now(),
                    'ULTIMA_ATUALIZACAO': datetime.now()
                }
                
                self.df_equipamentos = pd.concat([
                    self.df_equipamentos,
                    pd.DataFrame([novo_eq])
                ], ignore_index=True)
                
                # Salvar
                if self.salvar_dados('todos'):
                    custom_messagebox("info", "Sucesso", 
                        f"Contrato {entry_numero.get()} cadastrado!\n"
                        f"Valor total: R$ {valor_total:,.2f}")
                    janela.destroy()
                    self.carregar_dados()
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro ao salvar: {e}")
                import traceback
                traceback.print_exc()
        
        # Botões
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=15)
        frame_btns = ttk.Frame(frame)
        frame_btns.pack(fill='x', pady=(10, 0))
        
        ttk.Button(frame_btns, text="💾 Salvar Contrato", 
                  command=salvar, width=20).pack(side='left', padx=5)
        ttk.Button(frame_btns, text="❌ Cancelar", 
                  command=janela.destroy, width=15).pack(side='left', padx=5)
    
    def editar_contrato(self):
        """Edita contrato selecionado"""
        selected = self.tree_contratos.selection()
        if not selected:
            custom_messagebox("warning", "Aviso", "Selecione um contrato!")
            return
        
        custom_messagebox("info", "Info", 
            "Use o Excel para editar contratos.\n"
            "Clique em '📂 Abrir Excel' e edite a aba LOC_Contratos.")
    
    def adicionar_equipamento(self):
        """Adiciona equipamento"""
        selected = self.tree_contratos.selection()
        if not selected:
            custom_messagebox("warning", "Aviso", "Selecione um contrato primeiro!")
            return
        
        valores = self.tree_contratos.item(selected[0], 'values')
        id_contrato = int(valores[1])  # Coluna 1 agora (coluna 0 é checkbox)
        numero = valores[2]  # Coluna 2 agora
        
        janela = tk.Toplevel(self.janela)
        janela.title(f"Adicionar Equipamento - {numero}")
        janela.geometry("600x500")
        janela.transient(self.janela)
        janela.grab_set()
        
        frame = ttk.Frame(janela, padding="15")
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text="Adicionar Equipamento", 
                 font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 15))
        
        # Campos
        ttk.Label(frame, text="Tipo:").pack(anchor='w', pady=5)
        combo_tipo = ttk.Combobox(frame, values=[
            'BETONEIRA', 'GUINDASTE', 'ANDAIME', 'ESCORA', 
            'VIBRADOR', 'SERRA', 'GERADOR', 'COMPRESSOR', 'OUTROS'
        ], width=30)
        combo_tipo.pack(anchor='w', pady=5)
        
        ttk.Label(frame, text="Descrição:").pack(anchor='w', pady=5)
        entry_desc = ttk.Entry(frame, width=50)
        entry_desc.pack(anchor='w', pady=5)
        
        ttk.Label(frame, text="Identificação (Nº Série/Placa/Tag):").pack(anchor='w', pady=5)
        entry_id = ttk.Entry(frame, width=30)
        entry_id.pack(anchor='w', pady=5)
        
        ttk.Label(frame, text="Quantidade:").pack(anchor='w', pady=5)
        spin_qtd = ttk.Spinbox(frame, from_=1, to=999, width=10)
        spin_qtd.set('1')
        spin_qtd.pack(anchor='w', pady=5)
        
        ttk.Label(frame, text="Valor Unitário por Dia (R$):").pack(anchor='w', pady=5)
        entry_valor = ttk.Entry(frame, width=15)
        entry_valor.insert(0, "0,00")
        entry_valor.pack(anchor='w', pady=5)
        
        ttk.Label(frame, text="Localização:").pack(anchor='w', pady=5)
        entry_local = ttk.Entry(frame, width=40)
        entry_local.pack(anchor='w', pady=5)
        
        # Buscar número de dias do contrato para mostrar no cálculo
        contrato_info = self.df_contratos[self.df_contratos['ID_CONTRATO'] == id_contrato]
        dias_contrato = 30  # Padrão
        if not contrato_info.empty:
            dias_contrato = int(contrato_info.iloc[0].get('NUMERO_DIAS', 30))
        
        # Frame de cálculo
        frame_calc = ttk.LabelFrame(frame, text="Valor Total")
        frame_calc.pack(fill='x', pady=10)
        
        label_calc = ttk.Label(frame_calc, 
                              text=f"R$ 0,00 ({dias_contrato} dias)",
                              font=('TkDefaultFont', 12, 'bold'),
                              foreground='blue')
        label_calc.pack(padx=10, pady=5)
        
        def atualizar_calc(*args):
            try:
                qtd = int(spin_qtd.get())
                valor_str = entry_valor.get().replace('.', '').replace(',', '.')
                valor = float(valor_str)
                
                total = qtd * valor * dias_contrato
                total_fmt = f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                label_calc.config(text=f"{total_fmt} ({dias_contrato} dias)")
            except:
                label_calc.config(text=f"R$ 0,00 ({dias_contrato} dias)")
        
        # Bind para atualizar
        spin_qtd.bind('<KeyRelease>', atualizar_calc)
        spin_qtd.bind('<<Increment>>', atualizar_calc)
        spin_qtd.bind('<<Decrement>>', atualizar_calc)
        entry_valor.bind('<KeyRelease>', atualizar_calc)
        
        def salvar():
            try:
                if not combo_tipo.get():
                    custom_messagebox("error", "Erro", "Selecione o tipo!")
                    return
                
                # Gerar ID
                if self.df_equipamentos.empty:
                    novo_id = 1
                else:
                    novo_id = int(self.df_equipamentos['ID_EQUIPAMENTO'].max()) + 1
                
                # Buscar número de dias do contrato
                contrato = self.df_contratos[self.df_contratos['ID_CONTRATO'] == id_contrato]
                if contrato.empty:
                    custom_messagebox("error", "Erro", "Contrato não encontrado!")
                    return
                
                num_dias = int(contrato.iloc[0].get('NUMERO_DIAS', 30))
                
                qtd = int(spin_qtd.get())
                valor_unit = float(entry_valor.get().replace(',', '.'))
                valor_total = qtd * valor_unit * num_dias  # CORRIGIDO: Incluir número de dias
                
                # Criar registro
                novo = {
                    'ID_EQUIPAMENTO': novo_id,
                    'ID_CONTRATO': id_contrato,
                    'TIPO_EQUIPAMENTO': combo_tipo.get().upper(),
                    'DESCRICAO': entry_desc.get().strip().upper(),
                    'IDENTIFICACAO': entry_id.get().strip().upper(),
                    'QUANTIDADE': qtd,
                    'VALOR_UNITARIO': valor_unit,
                    'VALOR_TOTAL': valor_total,
                    'DATA_INICIO_USO': datetime.now().date(),
                    'DATA_FIM_USO': None,
                    'STATUS': 'EM_USO',
                    'LOCALIZACAO': entry_local.get().strip().upper(),
                    'OBSERVACAO': '',
                    'DATA_CADASTRO': datetime.now(),
                    'ULTIMA_ATUALIZACAO': datetime.now()
                }
                
                # Adicionar
                self.df_equipamentos = pd.concat([
                    self.df_equipamentos,
                    pd.DataFrame([novo])
                ], ignore_index=True)
                
                # Salvar
                if self.salvar_dados('equipamentos'):
                    custom_messagebox("info", "Sucesso", 
                        f"Equipamento adicionado!\n"
                        f"Valor total: R$ {valor_total:,.2f}")
                    janela.destroy()
                    self.atualizar_lista_equipamentos(id_contrato)
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro: {e}")
                import traceback
                traceback.print_exc()
        
        frame_btns = ttk.Frame(frame)
        frame_btns.pack(fill='x', pady=(20, 0))
        
        ttk.Button(frame_btns, text="Salvar", command=salvar).pack(side='left', padx=5)
        ttk.Button(frame_btns, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)
    
    def devolver_equipamento(self):
        """Marca equipamento como devolvido"""
        selected = self.tree_equipamentos.selection()
        if not selected:
            custom_messagebox("warning", "Aviso", "Selecione um equipamento!")
            return
        
        valores = self.tree_equipamentos.item(selected[0], 'values')
        id_eq = int(valores[0])
        tipo = valores[1]
        
        if not custom_messagebox("yesno", "Confirmar", 
            f"Marcar como DEVOLVIDO?\n\n{tipo}\n{valores[3]}"):
            return
        
        try:
            idx = self.df_equipamentos[
                self.df_equipamentos['ID_EQUIPAMENTO'] == id_eq
            ].index[0]
            
            self.df_equipamentos.at[idx, 'STATUS'] = 'DEVOLVIDO'
            self.df_equipamentos.at[idx, 'DATA_FIM_USO'] = datetime.now().date()
            self.df_equipamentos.at[idx, 'ULTIMA_ATUALIZACAO'] = datetime.now()
            
            if self.salvar_dados('equipamentos'):
                custom_messagebox("info", "Sucesso", "Equipamento devolvido!")
                
                # Recarregar lista
                selected_contrato = self.tree_contratos.selection()
                if selected_contrato:
                    valores_c = self.tree_contratos.item(selected_contrato[0], 'values')
                    self.atualizar_lista_equipamentos(int(valores_c[1]))  # Coluna 1 é ID agora
            
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro: {e}")
    
    def reportar_perda(self):
        """Reporta perda de equipamento"""
        selected = self.tree_equipamentos.selection()
        if not selected:
            custom_messagebox("warning", "Aviso", "Selecione um equipamento!")
            return
        
        valores = self.tree_equipamentos.item(selected[0], 'values')
        id_eq = int(valores[0])
        tipo = valores[1]
        
        janela = tk.Toplevel(self.janela)
        janela.title("Reportar Perda/Dano")
        janela.geometry("500x400")
        janela.transient(self.janela)
        janela.grab_set()
        
        frame = ttk.Frame(janela, padding="15")
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text=f"Reportar Perda/Dano", 
                 font=('TkDefaultFont', 12, 'bold')).pack(pady=(0, 10))
        
        ttk.Label(frame, text=f"Tipo: {tipo}").pack(anchor='w', pady=2)
        ttk.Label(frame, text=f"Identificação: {valores[3]}").pack(anchor='w', pady=2)
        
        ttk.Label(frame, text="").pack(pady=5)
        
        ttk.Label(frame, text="Tipo de Ocorrência:").pack(anchor='w', pady=5)
        var_tipo = tk.StringVar(value="PERDIDO")
        ttk.Radiobutton(frame, text="Perdido", variable=var_tipo, 
                       value="PERDIDO").pack(anchor='w', padx=20)
        ttk.Radiobutton(frame, text="Danificado", variable=var_tipo, 
                       value="DANIFICADO").pack(anchor='w', padx=20)
        
        ttk.Label(frame, text="Data da Ocorrência:").pack(anchor='w', pady=5)
        data_ocor = DateEntry(frame, width=15, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_ocor.pack(anchor='w', pady=5)
        
        ttk.Label(frame, text="Descrição Detalhada:").pack(anchor='w', pady=5)
        text_desc = tk.Text(frame, width=50, height=8)
        text_desc.pack(anchor='w', pady=5)
        
        def confirmar():
            try:
                idx = self.df_equipamentos[
                    self.df_equipamentos['ID_EQUIPAMENTO'] == id_eq
                ].index[0]
                
                self.df_equipamentos.at[idx, 'STATUS'] = var_tipo.get()
                self.df_equipamentos.at[idx, 'DATA_FIM_USO'] = data_ocor.get_date()
                
                obs_atual = self.df_equipamentos.at[idx, 'OBSERVACAO']
                nova_obs = f"{obs_atual}\n[{var_tipo.get()} - {data_ocor.get_date()}]: {text_desc.get('1.0', 'end-1c')}"
                self.df_equipamentos.at[idx, 'OBSERVACAO'] = nova_obs
                self.df_equipamentos.at[idx, 'ULTIMA_ATUALIZACAO'] = datetime.now()
                
                if self.salvar_dados('equipamentos'):
                    custom_messagebox("info", "Sucesso", 
                        f"Equipamento marcado como {var_tipo.get()}!")
                    janela.destroy()
                    
                    # Recarregar
                    selected_contrato = self.tree_contratos.selection()
                    if selected_contrato:
                        valores_c = self.tree_contratos.item(selected_contrato[0], 'values')
                        self.atualizar_lista_equipamentos(int(valores_c[1]))  # Coluna 1 é ID agora
                
            except Exception as e:
                custom_messagebox("error", "Erro", f"Erro: {e}")
        
        frame_btns = ttk.Frame(frame)
        frame_btns.pack(fill='x', pady=(10, 0))
        
        ttk.Button(frame_btns, text="Confirmar", command=confirmar).pack(side='left', padx=5)
        ttk.Button(frame_btns, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)
    
    def registrar_pagamento(self):
        """Registra pagamento"""
        selected = self.tree_contratos.selection()
        if not selected:
            custom_messagebox("warning", "Aviso", "Selecione um contrato!")
            return
        
        custom_messagebox("info", "Info", 
            "Use o Excel para registrar pagamentos.\n"
            "Clique em '📂 Abrir Excel' e adicione na aba LOC_Pagamentos.")
    
    def vincular_pagamento(self):
        """Vincula pagamento a lançamento"""
        custom_messagebox("info", "Em Desenvolvimento", 
            "Funcionalidade de vinculação será implementada em breve.\n\n"
            "Por enquanto, edite manualmente no Excel:\n"
            "- Aba LOC_Pagamentos: coluna ID_LANCAMENTO\n"
            "- Aba Dados: coluna ID_LANCAMENTO")
    
    def gerar_relatorio(self, tipo):
        """Gera relatório"""
        try:
            if tipo == 'fornecedor':
                self.relatorio_por_fornecedor()
            elif tipo == 'custos':
                self.relatorio_custos()
            elif tipo == 'equipamentos':
                self.relatorio_equipamentos()
            elif tipo == 'completo':
                self.relatorio_completo()
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao gerar relatório: {e}")
    
    def relatorio_por_fornecedor(self):
        """Relatório por fornecedor"""
        if self.df_contratos.empty:
            custom_messagebox("warning", "Aviso", "Nenhum contrato cadastrado!")
            return
        
        ativos = self.df_contratos[self.df_contratos['STATUS'] == 'ATIVO']
        
        rel = ativos.groupby('FORNECEDOR_NOME').agg({
            'ID_CONTRATO': 'count',
            'VALOR_MENSAL': 'sum',
            'NUMERO_CONTRATO': lambda x: ', '.join(x)
        }).reset_index()
        
        rel.columns = ['Fornecedor', 'Qtd Contratos', 'Valor Mensal Total', 'Contratos']
        
        self.salvar_relatorio(rel, 'Contratos_Por_Fornecedor')
    
    def relatorio_custos(self):
        """Relatório de custos"""
        if self.df_pagamentos.empty:
            custom_messagebox("warning", "Aviso", "Nenhum pagamento registrado!")
            return
        
        rel = self.df_pagamentos.groupby('MES_REFERENCIA').agg({
            'VALOR_FATURA': 'sum',
            'ID_PAGAMENTO': 'count'
        }).reset_index()
        
        rel.columns = ['Mês', 'Valor Total', 'Qtd Pagamentos']
        rel = rel.sort_values('Mês')
        
        self.salvar_relatorio(rel, 'Evolucao_Custos')
    
    def relatorio_equipamentos(self):
        """Relatório de equipamentos"""
        if self.df_equipamentos.empty:
            custom_messagebox("warning", "Aviso", "Nenhum equipamento cadastrado!")
            return
        
        em_uso = self.df_equipamentos[self.df_equipamentos['STATUS'] == 'EM_USO']
        
        rel = em_uso.groupby('TIPO_EQUIPAMENTO').agg({
            'QUANTIDADE': 'sum',
            'VALOR_TOTAL': 'sum',
            'ID_EQUIPAMENTO': 'count'
        }).reset_index()
        
        rel.columns = ['Tipo', 'Qtd Total', 'Valor Mensal', 'Nº Itens']
        
        self.salvar_relatorio(rel, 'Equipamentos_Por_Tipo')
    
    def relatorio_completo(self):
        """Relatório completo"""
        try:
            data = datetime.now().strftime('%Y%m%d_%H%M')
            nome = f"Locacoes_Completo_{self.sistema.cliente_atual}_{data}.xlsx"
            
            caminho = filedialog.asksaveasfilename(
                title="Salvar Relatório",
                defaultextension=".xlsx",
                initialfile=nome,
                filetypes=[("Excel", "*.xlsx")]
            )
            
            if caminho:
                with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
                    self.df_contratos.to_excel(writer, sheet_name='Contratos', index=False)
                    self.df_equipamentos.to_excel(writer, sheet_name='Equipamentos', index=False)
                    self.df_pagamentos.to_excel(writer, sheet_name='Pagamentos', index=False)
                    
                    # Resumo
                    resumo = pd.DataFrame([
                        {'Métrica': 'Total Contratos', 'Valor': len(self.df_contratos)},
                        {'Métrica': 'Contratos Ativos', 
                         'Valor': len(self.df_contratos[self.df_contratos['STATUS'] == 'ATIVO'])},
                        {'Métrica': 'Total Equipamentos', 'Valor': len(self.df_equipamentos)},
                        {'Métrica': 'Equipamentos em Uso', 
                         'Valor': len(self.df_equipamentos[self.df_equipamentos['STATUS'] == 'EM_USO'])},
                    ])
                    resumo.to_excel(writer, sheet_name='Resumo', index=False)
                
                custom_messagebox("info", "Sucesso", f"Relatório salvo em:\n{caminho}")
        
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro ao salvar: {e}")
    
    def salvar_relatorio(self, df, nome_base):
        """Salva relatório"""
        try:
            data = datetime.now().strftime('%Y%m%d_%H%M')
            nome = f"{nome_base}_{self.sistema.cliente_atual}_{data}.xlsx"
            
            caminho = filedialog.asksaveasfilename(
                title="Salvar Relatório",
                defaultextension=".xlsx",
                initialfile=nome,
                filetypes=[("Excel", "*.xlsx")]
            )
            
            if caminho:
                df.to_excel(caminho, index=False)
                custom_messagebox("info", "Sucesso", f"Relatório salvo!")
        
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro: {e}")
    
    def abrir_arquivo_excel(self):
        """Abre arquivo Excel"""
        try:
            import os
            import platform
            
            if platform.system() == 'Windows':
                os.startfile(self.arquivo_cliente)
            elif platform.system() == 'Darwin':
                os.system(f'open "{self.arquivo_cliente}"')
            else:
                os.system(f'xdg-open "{self.arquivo_cliente}"')
            
            custom_messagebox("info", "Arquivo Aberto", 
                "Feche o arquivo antes de salvar no sistema!")
        
        except Exception as e:
            custom_messagebox("error", "Erro", f"Erro: {e}")


# ====================================================================================
# INTEGRAÇÃO
# ====================================================================================

def integrar_gestao_locacoes(sistema_principal):
    """
    Integra ao sistema principal
    
    Usage:
        from gestao_locacoes_completo import integrar_gestao_locacoes
        
        # Adicionar botão
        ttk.Button(frame, text="🔧 Gestão de Locações", 
                  command=lambda: integrar_gestao_locacoes(sistema)).pack()
    """
    gerenciador = GerenciadorLocacoes(sistema_principal)
    gerenciador.abrir_gestao_locacoes()
    return gerenciador


if __name__ == "__main__":
    print("=" * 80)
    print("Módulo de Gestão de Locações - VERSÃO COMPLETA")
    print("=" * 80)
    print()
    print("Usa ABAS no arquivo Excel do cliente:")
    print("  - LOC_Contratos")
    print("  - LOC_Equipamentos")
    print("  - LOC_Pagamentos")
    print()
    print("Para integrar:")
    print("  from gestao_locacoes_completo import integrar_gestao_locacoes")
    print("=" * 80)