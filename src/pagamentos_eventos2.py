import os
import sys
from pathlib import Path
import re
from datetime import datetime
from decimal import Decimal

import tkinter as tk
from tkinter import ttk, messagebox, StringVar
from tkcalendar import DateEntry

from dateutil.relativedelta import relativedelta

import pandas as pd
from openpyxl import load_workbook

# Adicionar diretório raiz ao path
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações do sistema
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH,
        ARQUIVO_FORNECEDORES
    )
    from src.config.utils import formatar_cnpj_cpf, aplicar_formatacao_celula
    from src.config.window_config import configurar_janela
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Tentar caminho alternativo
    try:
        from config.config import (
            ARQUIVO_CLIENTES,
            ARQUIVO_MODELO,
            PASTA_CLIENTES,
            BASE_PATH,
            ARQUIVO_FORNECEDORES
        )
        from config.utils import formatar_cnpj_cpf, aplicar_formatacao_celula
        from config.window_config import configurar_janela
    except ImportError as e2:
        print(f"Erro ao importar configurações (caminho alternativo): {str(e2)}")
        raise

class GestaoEventos:
    def __init__(self, parent=None):
        self.parent = parent
        self.janela = None
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.contratos = []
        self.eventos = []
        self.administradores_contratos = {}
        self.cliente_selecionado = False  # Variável para controlar resultado da seleção
        self.notebook = None  # Para armazenar o notebook de abas
        self.aba_contratos = None
        self.aba_eventos = None
        self.aba_pagamentos = None
        self.tree_contratos = None
        self.tree_eventos = None
        self.tree_pagamentos = None
        self.tree_adm = None
        
        # Labels para detalhes do contrato
        self.lbl_contrato = None
        self.lbl_periodo = None
        self.lbl_status = None
        self.lbl_valor_total = None
        self.lbl_administradores = None
        self.lbl_eventos = None
        
        # Campos para eventos
        self.contrato_selecionado = None
        self.evento_descricao = None
        self.evento_percentual = None
        self.evento_status = None
        self.evento_data = None
        
        # Campos para pagamentos
        self.pagto_contrato = None
        self.pagto_status = None
        self.pagto_cnpj = None
        self.pagto_nome = None
        self.pagto_valor = None
        self.pagto_vencimento = None
        self.pagto_status_atual = None
        self.pagto_data = None

    def fechar_janela(self):
        """Fecha a janela corretamente"""
        print("Fechando janela")
        if self.janela:
            self.janela.destroy()
        
    def carregar_lista_clientes(self):
        """Carrega a lista de clientes disponíveis"""
        try:
            print("Carregando lista de clientes...")
            # Verificar se o arquivo existe
            if not os.path.exists(ARQUIVO_CLIENTES):
                print(f"Arquivo não encontrado: {ARQUIVO_CLIENTES}")
                messagebox.showwarning("Aviso", "Arquivo de clientes não encontrado!")
                return []
                
            workbook = load_workbook(ARQUIVO_CLIENTES)
            sheet = workbook['Clientes']
            print(f"Planilha aberta, abas disponíveis: {workbook.sheetnames}")
            print(f"Número de linhas na planilha: {sheet.max_row}")
            
            clientes = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome do cliente
                    print(f"Cliente encontrado: {row[0]}")
                    clientes.append(row[0])
                    
            workbook.close()
            print(f"Total de clientes carregados: {len(clientes)}")
            return sorted(clientes)
            
        except Exception as e:
            print(f"Erro ao carregar clientes: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
            return []
    
    def abrir_janela_eventos(self, cliente=None):
        """Abre a janela principal de gestão de eventos para o cliente selecionado"""
        print("Entrando em abrir_janela_eventos")
        
        # Se a janela já existir, apenas traz para frente
        if self.janela and self.janela.winfo_exists():
            print("Janela já existe, trazendo para frente")
            self.janela.lift()
            self.janela.focus_force()
            return

        print("Criando nova janela")
        # Cria nova janela
        self.janela = tk.Toplevel(self.parent)
        configurar_janela(self.janela, "Gestão de Pagamentos por Eventos", 900, 700)
        
        # Garantir visibilidade
        self.janela.attributes('-topmost', True)
        self.janela.update()
        self.janela.attributes('-topmost', False)

        # Define o cliente atual
        if cliente:
            print(f"Cliente fornecido: {cliente}")
            self.cliente_atual = cliente
            self.arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        else:
            print("Cliente não fornecido, solicitando seleção")
            # Se não tiver cliente selecionado, solicita seleção
            if not self.selecionar_cliente():
                print("Cliente não selecionado, fechando janela")
                self.janela.destroy()
                return

        # Configurar fechamento adequado
        def fechar_janela_local():
            print("Fechando janela")
            self.janela.destroy()
        
        # Usar a função local para fechamento
        self.janela.protocol("WM_DELETE_WINDOW", fechar_janela_local)
        
        # Frame principal
        frame_principal = ttk.Frame(self.janela, padding=10)
        frame_principal.pack(fill='both', expand=True)

        # Cabeçalho com informações do cliente
        frame_cabecalho = ttk.Frame(frame_principal)
        frame_cabecalho.pack(fill='x', pady=5)

        ttk.Label(
            frame_cabecalho, 
            text=f"Cliente: {self.cliente_atual}", 
            font=('Arial', 12, 'bold')
        ).pack(side='left')

        # Notebook para abas
        self.notebook = ttk.Notebook(frame_principal)
        self.notebook.pack(fill='both', expand=True, pady=10)

        # Aba de contratos
        self.aba_contratos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_contratos, text="Contratos")

        # Aba de eventos
        self.aba_eventos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_eventos, text="Eventos")

        # Aba de pagamentos
        self.aba_pagamentos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_pagamentos, text="Pagamentos")

        # Configurar cada aba
        self.configurar_aba_contratos()
        self.configurar_aba_eventos()
        self.configurar_aba_pagamentos()

         # Botões principais
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=10)

        ttk.Button(
            frame_botoes,
            text="Fechar",
            command=fechar_janela_local,
            width=15
        ).pack(side='right', padx=5)

        # Carregar dados iniciais
        self.carregar_contratos()
    
    def selecionar_cliente(self):
        """Seleciona um cliente usando uma caixa de diálogo simplificada"""
        # Carregar a lista de clientes
        clientes = self.carregar_lista_clientes()
        if not clientes:
            messagebox.showwarning("Aviso", "Nenhum cliente encontrado!")
            return False
        
        # Criar janela de diálogo
        dialogo = tk.Toplevel(self.janela)
        dialogo.title("Selecionar Cliente")
        dialogo.geometry("400x200")
        dialogo.resizable(False, False)
        dialogo.transient(self.janela)
        dialogo.grab_set()
        
        # Variável para resultado
        resultado = [False]  # Usamos uma lista para poder modificar de dentro da função
        
        # Frame principal
        frame = ttk.Frame(dialogo, padding=20)
        frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            frame, 
            text="Selecione um Cliente",
            font=("Arial", 12, "bold")
        ).pack(pady=10)
        
        # Combobox para seleção
        cliente_var = tk.StringVar()
        combo = ttk.Combobox(
            frame, 
            textvariable=cliente_var,
            values=clientes,
            state="readonly",
            width=40
        )
        combo.pack(pady=10)
        if clientes:
            combo.set(clientes[0])  # Selecionar o primeiro cliente por padrão
        
        # Função para confirmar
        def confirmar():
            if not cliente_var.get():
                messagebox.showwarning("Aviso", "Selecione um cliente!")
                return
            
            self.cliente_atual = cliente_var.get()
            self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
            resultado[0] = True
            dialogo.destroy()
        
        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(pady=20)
        
        ttk.Button(
            frame_botoes,
            text="Confirmar",
            command=confirmar,
            width=15
        ).pack(side='left', padx=10)
        
        ttk.Button(
            frame_botoes,
            text="Cancelar",
            command=dialogo.destroy,
            width=15
        ).pack(side='left', padx=10)
        
        # Primeiro é necessário atualizar a geometria para obter o tamanho real
        dialogo.update_idletasks()
        
        # Obter o tamanho e posição da janela pai
        width = dialogo.winfo_width()
        height = dialogo.winfo_height()
        
        # Centralizar em relação à janela pai
        parent_x = self.janela.winfo_x()
        parent_y = self.janela.winfo_y()
        parent_width = self.janela.winfo_width()
        parent_height = self.janela.winfo_height()
        
        # Calcular a posição para centralizar
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        # Garantir que a janela não fique fora da tela
        screen_width = self.janela.winfo_screenwidth()
        screen_height = self.janela.winfo_screenheight()
        
        # Ajustar se necessário
        if x < 0:
            x = 0
        elif x + width > screen_width:
            x = screen_width - width
            
        if y < 0:
            y = 0
        elif y + height > screen_height:
            y = screen_height - height
        
        # Definir a posição
        dialogo.geometry(f"{width}x{height}+{x}+{y}")
        
        # Adicionar tecla Enter para confirmar
        dialogo.bind('<Return>', lambda event: confirmar())
        
        # Trazer para frente e focar no combobox
        dialogo.lift()
        combo.focus_set()
        
        # Aguardar seleção
        self.janela.wait_window(dialogo)
        
        # Retornar resultado
        return resultado[0]
    
    def confirmar_selecao_cliente(self):
        """Confirma a seleção do cliente e inicializa a interface"""
        cliente = self.var_cliente.get()
        if not cliente:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro")
            return
            
        self.cliente_atual = cliente
        self.arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        print(f"Cliente selecionado: {cliente}")
        print(f"Arquivo: {self.arquivo_cliente}")
        
        # Verificar existência do arquivo
        if not os.path.exists(self.arquivo_cliente):
            messagebox.showerror("Erro", f"Arquivo do cliente não encontrado: {self.arquivo_cliente}")
            return
            
        # Inicializar interface principal
        self.inicializar_interface()
    
    def inicializar_interface(self):
        """Inicializa a interface principal com abas"""
        print("Inicializando interface principal")
        
        # Limpar janela atual
        for widget in self.janela.winfo_children():
            widget.destroy()
            
        # Frame principal
        frame_principal = ttk.Frame(self.janela, padding=10)
        frame_principal.pack(fill='both', expand=True)
        
        # Cabeçalho com informações do cliente
        frame_cabecalho = ttk.Frame(frame_principal)
        frame_cabecalho.pack(fill='x', pady=5)
        
        ttk.Label(
            frame_cabecalho, 
            text=f"Cliente: {self.cliente_atual}", 
            font=('Arial', 12, 'bold')
        ).pack(side='left')
        
        # Notebook para abas
        self.notebook = ttk.Notebook(frame_principal)
        self.notebook.pack(fill='both', expand=True, pady=10)
        
        # Aba de contratos
        self.aba_contratos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_contratos, text="Contratos")
        
        # Aba de eventos
        self.aba_eventos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_eventos, text="Eventos")
        
        # Aba de pagamentos
        self.aba_pagamentos = ttk.Frame(self.notebook)
        self.notebook.add(self.aba_pagamentos, text="Pagamentos")
        
        # Configurar cada aba
        self.configurar_aba_contratos()
        self.configurar_aba_eventos()
        self.configurar_aba_pagamentos()
        
        # Botões principais
        frame_botoes = ttk.Frame(frame_principal)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(
            frame_botoes,
            text="Fechar",
            command=self.fechar_janela,
            width=15
        ).pack(side='right', padx=5)
        
        # Carregar dados iniciais
        self.carregar_contratos()
        
    def configurar_aba_contratos(self):
        """Configura a aba de gestão de contratos"""
        print("Configurando aba de contratos")
        frame = ttk.Frame(self.aba_contratos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para lista de contratos
        frame_lista = ttk.LabelFrame(frame, text="Contratos Disponíveis")
        frame_lista.pack(fill='both', expand=True, pady=5)

        # Treeview para contratos
        colunas = ('Nº Contrato', 'Data Início', 'Data Fim', 'Status', 'Valor Total')
        self.tree_contratos = ttk.Treeview(frame_lista, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_contratos.heading(col, text=col)
            if col == 'Valor Total':
                self.tree_contratos.column(col, width=120, anchor='e')
            else:
                self.tree_contratos.column(col, width=120)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree_contratos.yview)
        scroll_x = ttk.Scrollbar(frame_lista, orient='horizontal', command=self.tree_contratos.xview)
        self.tree_contratos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_contratos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do contrato selecionado
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Contrato")
        frame_detalhes.pack(fill='x', pady=10)

        frame_detalhes_grid = ttk.Frame(frame_detalhes)
        frame_detalhes_grid.pack(fill='x', padx=5, pady=5)

        # Criar labels e campos para detalhes
        ttk.Label(frame_detalhes_grid, text="Contrato:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.lbl_contrato = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_contrato.grid(row=0, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Período:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.lbl_periodo = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_periodo.grid(row=1, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Status:").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.lbl_status = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_status.grid(row=2, column=1, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Valor Total:").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.lbl_valor_total = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_valor_total.grid(row=0, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Administradores:").grid(row=1, column=2, sticky='w', padx=5, pady=2)
        self.lbl_administradores = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_administradores.grid(row=1, column=3, sticky='w', padx=5, pady=2)

        ttk.Label(frame_detalhes_grid, text="Eventos:").grid(row=2, column=2, sticky='w', padx=5, pady=2)
        self.lbl_eventos = ttk.Label(frame_detalhes_grid, text="-")
        self.lbl_eventos.grid(row=2, column=3, sticky='w', padx=5, pady=2)

        # Frame para administradores do contrato
        frame_admins = ttk.LabelFrame(frame, text="Administradores do Contrato")
        frame_admins.pack(fill='both', expand=True, pady=5)

        # Treeview para administradores
        colunas_adm = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas')
        self.tree_adm = ttk.Treeview(frame_admins, columns=colunas_adm, show='headings', height=4)
        
        for col in colunas_adm:
            self.tree_adm.heading(col, text=col)
            if col in ['Valor/Percentual', 'Valor Total']:
                self.tree_adm.column(col, width=120, anchor='e')
            elif col == 'Nº Parcelas':
                self.tree_adm.column(col, width=80, anchor='center')
            else:
                self.tree_adm.column(col, width=120)

        # Scrollbars
        scroll_y_adm = ttk.Scrollbar(frame_admins, orient='vertical', command=self.tree_adm.yview)
        scroll_x_adm = ttk.Scrollbar(frame_admins, orient='horizontal', command=self.tree_adm.xview)
        self.tree_adm.configure(yscrollcommand=scroll_y_adm.set, xscrollcommand=scroll_x_adm.set)
        
        self.tree_adm.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y_adm.pack(side='right', fill='y')
        scroll_x_adm.pack(side='bottom', fill='x')

        # Botões de ação
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes,
            text="Definir Eventos",
            command=self.definir_eventos,
            width=20
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Atualizar Eventos",
            command=self.atualizar_eventos_contrato,
            width=20
        ).pack(side='left', padx=5)

        # Binding para mostrar detalhes ao selecionar contrato
        self.tree_contratos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_contrato)

    def configurar_aba_eventos(self):
        """Configura a aba de gestão de eventos"""
        print("Configurando aba de eventos")
        frame = ttk.Frame(self.aba_eventos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para contrato selecionado
        frame_contrato = ttk.LabelFrame(frame, text="Contrato")
        frame_contrato.pack(fill='x', pady=5)

        ttk.Label(frame_contrato, text="Contrato Selecionado:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.contrato_selecionado = ttk.Combobox(frame_contrato, state='readonly', width=40)
        self.contrato_selecionado.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.contrato_selecionado.bind('<<ComboboxSelected>>', self.carregar_eventos_contrato)

        # Frame para botão de adicionar evento
        frame_adicionar = ttk.Frame(frame)
        frame_adicionar.pack(fill='x', pady=5)

        ttk.Button(
            frame_adicionar,
            text="Adicionar Evento",
            command=self.adicionar_evento,
            width=20
        ).pack(side='left', padx=5)

        # Frame para lista de eventos
        frame_eventos = ttk.LabelFrame(frame, text="Eventos do Contrato")
        frame_eventos.pack(fill='both', expand=True, pady=5)

        # Treeview para eventos
        colunas = ('ID', 'Descrição', 'Percentual', 'Valor', 'Status', 'Data Conclusão')
        self.tree_eventos = ttk.Treeview(frame_eventos, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_eventos.heading(col, text=col)
            if col == 'ID':
                self.tree_eventos.column(col, width=50, anchor='center')
            elif col in ['Percentual', 'Valor']:
                self.tree_eventos.column(col, width=100, anchor='e')
            elif col == 'Status':
                self.tree_eventos.column(col, width=100, anchor='center')
            else:
                self.tree_eventos.column(col, width=150)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_eventos, orient='vertical', command=self.tree_eventos.yview)
        scroll_x = ttk.Scrollbar(frame_eventos, orient='horizontal', command=self.tree_eventos.xview)
        self.tree_eventos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_eventos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do evento selecionado
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Evento")
        frame_detalhes.pack(fill='x', pady=10)

        # Grid para formulário
        frame_form = ttk.Frame(frame_detalhes)
        frame_form.pack(fill='x', padx=5, pady=5)

        ttk.Label(frame_form, text="Descrição:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.evento_descricao = ttk.Entry(frame_form, width=50)
        self.evento_descricao.grid(row=0, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Percentual (%):").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.evento_percentual = ttk.Entry(frame_form, width=15)
        self.evento_percentual.grid(row=1, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Status:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.evento_status = ttk.Combobox(frame_form, values=['pendente', 'concluido'], state='readonly', width=15)
        self.evento_status.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        self.evento_status.set('pendente')

        ttk.Label(frame_form, text="Data Conclusão:").grid(row=3, column=0, padx=5, pady=2, sticky='w')
        self.evento_data = DateEntry(frame_form, width=15, state='normal', date_pattern='dd/mm/yyyy')
        self.evento_data.grid(row=3, column=1, padx=5, pady=2, sticky='w')

        # Botões de ação para eventos
        frame_botoes = ttk.Frame(frame_detalhes)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes,
            text="Salvar Evento",
            command=self.salvar_evento,
            width=15
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Marcar como Concluído",
            command=self.concluir_evento,
            width=20
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Remover Evento",
            command=self.remover_evento,
            width=15
        ).pack(side='left', padx=5)

        # Binding para seleção de evento
        self.tree_eventos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_evento)

    def configurar_aba_pagamentos(self):
        """Configura a aba de pagamentos de eventos"""
        print("Configurando aba de pagamentos")
        frame = ttk.Frame(self.aba_pagamentos, padding=10)
        frame.pack(fill='both', expand=True)

        # Frame para controle de filtros
        frame_filtros = ttk.LabelFrame(frame, text="Filtros")
        frame_filtros.pack(fill='x', pady=5)

        ttk.Label(frame_filtros, text="Contrato:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.pagto_contrato = ttk.Combobox(frame_filtros, state='readonly', width=40)
        self.pagto_contrato.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.pagto_contrato.bind('<<ComboboxSelected>>', self.filtrar_pagamentos)

        ttk.Label(frame_filtros, text="Status:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        self.pagto_status = ttk.Combobox(frame_filtros, values=['Todos', 'Pendente', 'Pago'], state='readonly', width=15)
        self.pagto_status.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        self.pagto_status.set('Todos')
        self.pagto_status.bind('<<ComboboxSelected>>', self.filtrar_pagamentos)

        # Frame para lista de pagamentos
        frame_pagamentos = ttk.LabelFrame(frame, text="Pagamentos Previstos")
        frame_pagamentos.pack(fill='both', expand=True, pady=5)

        # Treeview para pagamentos
        colunas = ('Contrato', 'Evento', 'CNPJ/CPF', 'Nome', 'Data Vencimento', 'Valor', 'Status', 'Data Pagamento')
        self.tree_pagamentos = ttk.Treeview(frame_pagamentos, columns=colunas, show='headings')
        
        for col in colunas:
            self.tree_pagamentos.heading(col, text=col)
            if col in ['Data Vencimento', 'Data Pagamento']:
                self.tree_pagamentos.column(col, width=120, anchor='center')
            elif col == 'Valor':
                self.tree_pagamentos.column(col, width=100, anchor='e')
            elif col == 'Status':
                self.tree_pagamentos.column(col, width=100, anchor='center')
            elif col in ['Contrato', 'Evento']:
                self.tree_pagamentos.column(col, width=80, anchor='center')
            else:
                self.tree_pagamentos.column(col, width=150)

        # Scrollbars
        scroll_y = ttk.Scrollbar(frame_pagamentos, orient='vertical', command=self.tree_pagamentos.yview)
        scroll_x = ttk.Scrollbar(frame_pagamentos, orient='horizontal', command=self.tree_pagamentos.xview)
        self.tree_pagamentos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_pagamentos.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Frame para detalhes do pagamento
        frame_detalhes = ttk.LabelFrame(frame, text="Detalhes do Pagamento")
        frame_detalhes.pack(fill='x', pady=10)

        # Grid para formulário
        frame_form = ttk.Frame(frame_detalhes)
        frame_form.pack(fill='x', padx=5, pady=5)

        ttk.Label(frame_form, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        self.pagto_cnpj = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_cnpj.grid(row=0, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Nome:").grid(row=0, column=2, padx=5, pady=2, sticky='w')
        self.pagto_nome = ttk.Entry(frame_form, width=40, state='readonly')
        self.pagto_nome.grid(row=0, column=3, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Valor:").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        self.pagto_valor = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_valor.grid(row=1, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Vencimento:").grid(row=1, column=2, padx=5, pady=2, sticky='w')
        self.pagto_vencimento = ttk.Entry(frame_form, width=20, state='readonly')
        self.pagto_vencimento.grid(row=1, column=3, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Status:").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        self.pagto_status_atual = ttk.Combobox(frame_form, values=['Pendente', 'Pago'], state='readonly', width=15)
        self.pagto_status_atual.grid(row=2, column=1, padx=5, pady=2, sticky='w')

        ttk.Label(frame_form, text="Data Pagamento:").grid(row=2, column=2, padx=5, pady=2, sticky='w')
        self.pagto_data = DateEntry(frame_form, width=15, state='normal', date_pattern='dd/mm/yyyy')
        self.pagto_data.grid(row=2, column=3, padx=5, pady=2, sticky='w')

        # Botões de ação para pagamentos
        frame_botoes = ttk.Frame(frame_detalhes)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(
            frame_botoes,
            text="Registrar Pagamento",
            command=self.registrar_pagamento,
            width=20
        ).pack(side='left', padx=5)

        ttk.Button(
            frame_botoes,
            text="Gerar Lançamento",
            command=self.gerar_lancamento_pagamento,
            width=20
        ).pack(side='left', padx=5)

        # Binding para seleção de pagamento
        self.tree_pagamentos.bind('<<TreeviewSelect>>', self.mostrar_detalhes_pagamento)