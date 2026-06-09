"""
ImportadorMedicoes - Importa medições de contratos a partir de relatórios Excel

Este módulo permite que os usuários:
1. Exportem o relatório de contratos do sistema
2. Atualizem as medições no Excel
3. Importem as medições atualizadas de volta para o sistema

Autor: Sistema de Gestão
Data: Dezembro 2025
"""

import os
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta
from tkinter import filedialog
import logging
import shutil

# Configurar logging
logger = logging.getLogger(__name__)

# Importar dependências do sistema
try:
    from src.config.utils import custom_messagebox
    from src.config import BASE_PATH, PASTA_CLIENTES
    
    # Verificar se os imports foram bem-sucedidos
    if BASE_PATH == Path(".") or PASTA_CLIENTES == Path("."):
        raise ImportError("Caminhos do config não foram carregados corretamente")
    
    logger.info(f"✓ Config importado com sucesso")
    logger.info(f"  BASE_PATH: {BASE_PATH}")
    logger.info(f"  PASTA_CLIENTES: {PASTA_CLIENTES}")
    
except ImportError as e:
    logger.error(f"❌ Erro ao importar config: {e}")
    logger.error("Usando fallback - ISSO NÃO DEVERIA ACONTECER EM PRODUÇÃO!")
    
    # Para testes standalone - NUNCA DEVE SER USADO EM PRODUÇÃO
    BASE_PATH = Path("C:/Users/Obras/sistema_gestao_testes/testes/Financeiro/Planilhas_Base")
    PASTA_CLIENTES = Path("C:/Users/Obras/sistema_gestao_testes/testes/Financeiro/Clientes")
    
    def custom_messagebox(tipo, titulo, mensagem):
        """Fallback para custom_messagebox"""
        import tkinter as tk
        from tkinter import messagebox
        
        root = tk.Tk()
        root.withdraw()
        
        if tipo == "yesno":
            result = messagebox.askyesno(titulo, mensagem)
        elif tipo == "error":
            messagebox.showerror(titulo, mensagem)
            result = False
        elif tipo == "warning":
            messagebox.showwarning(titulo, mensagem)
            result = False
        else:  # info
            messagebox.showinfo(titulo, mensagem)
            result = False
        
        root.destroy()
        return result


class ImportadorMedicoes:
    """
    Classe responsável por importar medições de contratos a partir de planilhas Excel.
    
    Fluxo:
    1. Sistema exporta relatório → Relatorio_[CLIENTE]_[DATA].xlsx
    2. Usuário atualiza medições nas abas de cada contrato
    3. Sistema importa medições → Valida → Salva em [CLIENTE].xlsx
    """
    
    def __init__(self, sistema_principal):
        """
        Inicializa o importador de medições.
        
        Args:
            sistema_principal: Referência ao sistema principal
        """
        self.sistema = sistema_principal
        self.pasta_medicoes = Path(BASE_PATH) / "Relatorios_Medicoes"
        
        # Criar pasta se não existir
        os.makedirs(self.pasta_medicoes, exist_ok=True)

        # Pasta para arquivos importados
        self.pasta_importados = self.pasta_medicoes / "Importados"
        os.makedirs(self.pasta_importados, exist_ok=True)
        
        # Configurações de validação
        self.validacoes_ativadas = {
            'saldo': True,
            'sequencia_id': True,
            'data_medicao': True,
            'status_pendente': True
        }
        
        # Cache para fornecedores
        self.cache_fornecedores = {}
        
        # Log das configurações de caminho com verificação de tipos
        logger.info("=" * 60)
        logger.info("ImportadorMedicoes inicializado (VERSÃO CORRIGIDA)")
        logger.info(f"BASE_PATH: {BASE_PATH}")
        logger.info(f"PASTA_CLIENTES: {PASTA_CLIENTES}")
        logger.info(f"Pasta de medições: {self.pasta_medicoes}")
        logger.info(f"Pasta importados: {self.pasta_importados}")
        logger.info("=" * 60)
    
    def selecionar_arquivo_relatorio(self):
        """
        Permite ao usuário selecionar um arquivo de relatório de contratos.
        
        Returns:
            str: Caminho do arquivo selecionado ou None se cancelar
        """
        arquivo = filedialog.askopenfilename(
            title="Selecione o Relatório de Contratos (Medições Atualizadas)",
            filetypes=[
                ("Arquivos Excel", "*.xlsx *.xls"),
                ("Todos os arquivos", "*.*")
            ],
            initialdir=str(self.pasta_medicoes)
        )
        
        if not arquivo:
            logger.info("Seleção de arquivo cancelada pelo usuário")
            return None
        
        # Verificar se é formato antigo (.xls)
        extensao = os.path.splitext(arquivo)[1].lower()
        if extensao == '.xls':
            custom_messagebox("warning", 
                "Formato Antigo", 
                "O arquivo está no formato Excel 97-2003 (.xls).\n\n"
                "Por favor, abra o arquivo no Excel e salve como 'Pasta de Trabalho do Excel' (.xlsx)."
            )
            return None
        
        logger.info(f"Arquivo selecionado: {arquivo}")
        return arquivo
    
    def validar_estrutura_relatorio(self, arquivo):
        """
        Valida se o arquivo possui a estrutura esperada de relatório de contratos.
        
        Args:
            arquivo (str): Caminho do arquivo Excel
            
        Returns:
            tuple: (bool, str) - (válido, mensagem de erro)
        """
        try:
            wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
            
            # Verificar se tem abas
            if len(wb.sheetnames) == 0:
                return False, "O arquivo não contém nenhuma aba."
            
            # Verificar se tem aba de resumo
            if "Resumo Contratos" not in wb.sheetnames:
                return False, "Aba 'Resumo Contratos' não encontrada. Este não parece ser um relatório de contratos válido."
            
            # Verificar se tem pelo menos uma aba de contrato
            abas_contratos = [aba for aba in wb.sheetnames if aba.startswith("Contrato ")]
            if len(abas_contratos) == 0:
                # Aceitar planilha modelo: sem contratos, mas com Servicos_Adicionais
                if 'Servicos_Adicionais' in wb.sheetnames:
                    wb.close()
                    logger.info("Planilha modelo detectada (0 contratos, Servicos_Adicionais presente)")
                    return True, "Planilha modelo (0 contratos) — apenas Servicos_Adicionais"
                return False, "Nenhuma aba de contrato encontrada (ex: 'Contrato 1', 'Contrato 2', etc.)"
            
            # Validar estrutura de uma aba de contrato
            primeira_aba = abas_contratos[0]
            sheet = wb[primeira_aba]
            
            # Verificar células críticas
            if sheet['B3'].value is None:
                return False, f"Estrutura inválida na aba '{primeira_aba}': Célula B3 (ID do Contrato) está vazia."
            
            if sheet['B4'].value is None:
                return False, f"Estrutura inválida na aba '{primeira_aba}': Célula B4 (Fornecedor) está vazia."
            
            wb.close()
            
            logger.info(f"Estrutura do relatório validada: {len(abas_contratos)} contratos encontrados")
            return True, f"{len(abas_contratos)} contrato(s) encontrado(s)"
            
        except Exception as e:
            logger.error(f"Erro ao validar estrutura do relatório: {str(e)}")
            return False, f"Erro ao validar arquivo: {str(e)}"
    
    def extrair_medicoes_contrato(self, sheet, id_contrato, fornecedor, saldo_disponivel):
        """
        Extrai as medições de uma aba de contrato específica.
        
        REGRA DE STATUS:
        - Medições com status PENDENTE ou vazio (None/string vazia) são consideradas NOVAS
        - Medições com status LANÇADO ou VINCULADO são IGNORADAS (já foram processadas)
        - Usuários podem deixar o campo status vazio por esquecimento ou desconhecimento
        - O sistema assume que campos vazios são novas medições pendentes
        
        Args:
            sheet: Planilha do openpyxl
            id_contrato: ID do contrato
            fornecedor: Nome do fornecedor
            saldo_disponivel: Saldo disponível no contrato
            
        Returns:
            list: Lista de dicionários com as medições encontradas
        """
        medicoes = []
        erros = []

        # Linha 12 = cabeçalho, dados começam na linha 13
        # Colunas: A=ID (auto se vazio), B=Data Medição, C=Referência, D=Valor, E=Status
        linha_atual = 13
        medicoes_ids = []

        while True:
            id_medicao   = sheet.cell(row=linha_atual, column=1).value  # A
            data_medicao = sheet.cell(row=linha_atual, column=2).value  # B
            referencia   = sheet.cell(row=linha_atual, column=3).value  # C
            valor        = sheet.cell(row=linha_atual, column=4).value  # D
            status       = sheet.cell(row=linha_atual, column=5).value  # E

            # Fim dos dados quando todas as colunas relevantes estão vazias
            if data_medicao is None and referencia is None and valor is None:
                break

            # Auto-atribuir ID se célula estiver vazia
            id_auto = id_medicao is None or str(id_medicao).strip() == ''
            if id_auto:
                id_medicao = (max(medicoes_ids) + 1) if medicoes_ids else 1
            else:
                try:
                    id_medicao = int(id_medicao)
                except (ValueError, TypeError):
                    erros.append(f"Linha {linha_atual}: ID inválido '{id_medicao}'")
                    linha_atual += 1
                    continue

            # Aceita PENDENTE ou vazio; pula LANÇADO/VINCULADO (já processados)
            status_str = str(status).upper().strip() if status is not None else ''

            if status_str and status_str != 'PENDENTE':
                medicoes_ids.append(id_medicao)
                linha_atual += 1
                continue

            # ===== VALIDAÇÕES =====

            # 1. Validar sequência de IDs (só para IDs preenchidos manualmente)
            if self.validacoes_ativadas['sequencia_id'] and not id_auto:
                if medicoes_ids:
                    ultimo_id = max(medicoes_ids)
                    if id_medicao != ultimo_id + 1:
                        erros.append(
                            f"Linha {linha_atual}: ID {id_medicao} não é sequencial. "
                            f"Esperado: {ultimo_id + 1}"
                        )
                        linha_atual += 1
                        continue
            
            # 2. Validar Data de Medição
            if self.validacoes_ativadas['data_medicao']:
                if data_medicao is None:
                    erros.append(f"Linha {linha_atual}: Data de Medição obrigatória")
                    linha_atual += 1
                    continue
                
                # Validar regra de data
                hoje = datetime.now()
                dia_hoje = hoje.day
                
                # Converter data_medicao para datetime se necessário
                if isinstance(data_medicao, str):
                    try:
                        data_medicao = datetime.strptime(data_medicao, '%Y-%m-%d')
                    except ValueError:
                        try:
                            data_medicao = datetime.strptime(data_medicao, '%d/%m/%Y')
                        except ValueError:
                            erros.append(f"Linha {linha_atual}: Data de Medição inválida '{data_medicao}'")
                            linha_atual += 1
                            continue
                
                dia_medicao = data_medicao.day
                
                # Regra: Data Medição (coluna B) não pode ser < 5 se hoje é > 5 e <= 20, 
                # ou < 20 se hoje é > 20 e <= 5
                # if 5 < dia_hoje <= 20:
                #     # Período 1: hoje entre 6 e 20
                #     # Data de medição não pode ser < 5
                #     if dia_medicao < 5:
                #         erros.append(
                #             f"Linha {linha_atual}: Data de Medição (dia {dia_medicao}) não pode ser "
                #             f"anterior ao dia 5 quando estamos entre os dias 6 e 20"
                #         )
                #         linha_atual += 1
                #         continue
                # elif dia_hoje > 20 or dia_hoje <= 5:
                #     # Período 2: hoje entre 21 e 5 do mês seguinte
                #     # Data de medição não pode ser < 20
                #     if dia_medicao < 20:
                #         erros.append(
                #             f"Linha {linha_atual}: Data de Medição (dia {dia_medicao}) não pode ser "
                #             f"anterior ao dia 20 quando estamos entre os dias 21 e 5"
                #         )
                #         linha_atual += 1
                #         continue
            
            # 3. Validar Referência
            if referencia is None or str(referencia).strip().upper() == '':
                erros.append(f"Linha {linha_atual}: Referência obrigatória")
                linha_atual += 1
                continue
            
            # 4. Validar Valor
            if valor is None:
                erros.append(f"Linha {linha_atual}: Valor obrigatório")
                linha_atual += 1
                continue
            
            try:
                valor_float = float(valor)
                if valor_float <= 0:
                    erros.append(f"Linha {linha_atual}: Valor deve ser maior que zero")
                    linha_atual += 1
                    continue
            except (ValueError, TypeError):
                erros.append(f"Linha {linha_atual}: Valor inválido '{valor}'")
                linha_atual += 1
                continue
            
            # 5. Validar se valor não ultrapassa saldo
            if self.validacoes_ativadas['saldo']:
                if saldo_disponivel <= 0:
                    erros.append(
                        f"Linha {linha_atual}: Contrato {id_contrato} não possui saldo disponível "
                        f"(Saldo: R$ {saldo_disponivel:.2f})"
                    )
                    linha_atual += 1
                    continue
                
                if valor_float > saldo_disponivel:
                    erros.append(
                        f"Contrato {id_contrato} — linha {linha_atual}: as medições ultrapassam o saldo "
                        f"disponível (R$ {saldo_disponivel:,.2f}). "
                        f"Se for aditivo, utilize a aba 'Servicos_Adicionais' para criar um novo contrato."
                    )
                    linha_atual += 1
                    continue
            
            # ===== MEDIÇÃO VÁLIDA - ADICIONAR À LISTA =====

            medicao = {
                'id_contrato': id_contrato,
                'id_medicao': id_medicao,
                'fornecedor': fornecedor,
                'data_medicao': data_medicao,
                'data_pagamento': data_medicao,  # sem coluna própria; usa data de medição
                'referencia': str(referencia).strip().upper(),
                'valor': valor_float,
                'status': 'PENDENTE',
                'linha_origem': linha_atual
            }
            
            medicoes.append(medicao)
            medicoes_ids.append(id_medicao)
            
            # Atualizar saldo para próximas validações
            saldo_disponivel -= valor_float
            
            linha_atual += 1
        
        return medicoes, erros
    
    def processar_servicos_adicionais(self, arquivo, arquivo_cliente):
        medicoes_adic = []
        erros = []

        try:
            # ── ETAPA 1: Ler a aba do relatório e fechar imediatamente ──────
            linhas_adicionais = []

            wb_rel = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

            if 'Servicos_Adicionais' not in wb_rel.sheetnames:
                wb_rel.close()
                return medicoes_adic, erros

            sheet = wb_rel['Servicos_Adicionais']

            # Detectar formato: novo (10 colunas) ou legado (7 colunas)
            # Cabeçalho está na linha 4; col C do novo formato = "Data Início"
            cabecalho_c = str(sheet.cell(row=4, column=3).value or '').strip().lower()
            formato_novo = 'início' in cabecalho_c or 'inicio' in cabecalho_c

            for linha in range(5, sheet.max_row + 1):
                fornecedor = sheet.cell(row=linha, column=1).value
                descricao  = sheet.cell(row=linha, column=2).value

                if formato_novo:
                    # A=Fornecedor, B=Descrição, C=Data Início, D=Data Fim,
                    # E=Valor Contrato, F=Data Medição, G=Data Pagamento,
                    # H=Referência, I=Valor Medição, J=Observação
                    dt_inicio      = sheet.cell(row=linha, column=3).value
                    dt_fim         = sheet.cell(row=linha, column=4).value
                    valor_contrato = sheet.cell(row=linha, column=5).value
                    dt_med         = sheet.cell(row=linha, column=6).value
                    dt_pag         = sheet.cell(row=linha, column=7).value
                    referencia     = sheet.cell(row=linha, column=8).value
                    valor          = sheet.cell(row=linha, column=9).value
                    observacao     = sheet.cell(row=linha, column=10).value
                else:
                    # Formato legado: A=Fornecedor, B=Descrição, C=Data Medição,
                    # D=Data Pagamento, E=Referência, F=Valor, G=Observação
                    dt_inicio      = None
                    dt_fim         = None
                    valor_contrato = None
                    dt_med         = sheet.cell(row=linha, column=3).value
                    dt_pag         = sheet.cell(row=linha, column=4).value
                    referencia     = sheet.cell(row=linha, column=5).value
                    valor          = sheet.cell(row=linha, column=6).value
                    observacao     = sheet.cell(row=linha, column=7).value

                # Linha vazia = fim
                if not fornecedor and not descricao and not valor:
                    break

                linhas_adicionais.append({
                    'linha': linha,
                    'fornecedor': fornecedor,
                    'descricao': descricao,
                    'dt_inicio': dt_inicio,
                    'dt_fim': dt_fim,
                    'valor_contrato': valor_contrato,
                    'dt_med': dt_med,
                    'dt_pag': dt_pag,
                    'referencia': referencia,
                    'valor': valor,
                    'observacao': observacao
                })

            wb_rel.close()  # Arquivo do relatório fechado AQUI, de vez

            # ── ETAPA 2: Validar e gravar no arquivo do cliente ─────────────
            if not linhas_adicionais:
                return medicoes_adic, erros

            wb_cliente = openpyxl.load_workbook(arquivo_cliente)

            # Criar abas com cabeçalhos se ainda não existirem (cliente sem contratos)
            if 'Contratos_Medicao' not in wb_cliente.sheetnames:
                ws_new = wb_cliente.create_sheet('Contratos_Medicao')
                for col, h in enumerate(
                    ["ID_Contrato", "CNPJ_Fornecedor", "Nome_Fornecedor", "Descricao",
                     "Data_Inicio", "Data_Final", "Valor_Global", "Valor_Pago",
                     "Saldo", "Status", "Observacao"], 1
                ):
                    ws_new.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)
                logger.info("Aba Contratos_Medicao criada (cliente novo)")

            if 'Medicoes' not in wb_cliente.sheetnames:
                ws_new = wb_cliente.create_sheet('Medicoes')
                for col, h in enumerate(
                    ["ID_Contrato", "ID_Medicao", "CNPJ_Fornecedor", "Nome_Fornecedor",
                     "Data_Medicao", "Data_Pagamento", "Referencia", "Valor",
                     "Status", "Data_Lancamento", "Observacao"], 1
                ):
                    ws_new.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)
                logger.info("Aba Medicoes criada (cliente novo)")

            ws_contratos = wb_cliente['Contratos_Medicao']
            ws_medicoes  = wb_cliente['Medicoes']

            def _normalizar_data(val, label, erros, linha):
                if val is None:
                    return None, False
                if isinstance(val, str):
                    try:
                        return datetime.strptime(val.strip(), '%d/%m/%Y'), False
                    except ValueError:
                        erros.append(f"Linha {linha}: {label} inválida"); return None, True
                return val, False

            for item in linhas_adicionais:
                linha          = item['linha']
                fornecedor     = item['fornecedor']
                descricao      = item['descricao']
                dt_inicio      = item['dt_inicio']
                dt_fim         = item['dt_fim']
                valor_contrato = item['valor_contrato']
                dt_med         = item['dt_med']
                dt_pag         = item['dt_pag']
                referencia     = item['referencia']
                valor          = item['valor']
                observacao     = item['observacao']

                # Validações obrigatórias
                if not fornecedor:
                    erros.append(f"Linha {linha}: Fornecedor obrigatório"); continue
                if not descricao:
                    erros.append(f"Linha {linha}: Descrição obrigatória"); continue
                if not dt_med:
                    erros.append(f"Linha {linha}: Data de Medição obrigatória"); continue
                if not valor:
                    erros.append(f"Linha {linha}: Valor obrigatório"); continue

                try:
                    valor_float = float(valor)
                    if valor_float <= 0:
                        erros.append(f"Linha {linha}: Valor deve ser maior que zero"); continue
                except (ValueError, TypeError):
                    erros.append(f"Linha {linha}: Valor inválido '{valor}'"); continue

                # Normalizar datas
                dt_med,    err = _normalizar_data(dt_med, "Data de Medição", erros, linha)
                if err: continue
                dt_pag,    _   = _normalizar_data(dt_pag, "Data de Pagamento", erros, linha)
                if dt_pag is None: dt_pag = dt_med
                dt_inicio, _   = _normalizar_data(dt_inicio, "Data Início", erros, linha)
                dt_fim,    _   = _normalizar_data(dt_fim, "Data Fim", erros, linha)

                # Auto-calcular datas do contrato quando não informadas
                if dt_med:
                    import calendar
                    if dt_inicio is None:
                        dt_inicio = dt_med.replace(day=1)
                    if dt_fim is None:
                        data_2m = dt_med + relativedelta(months=2)
                        dt_fim = data_2m.replace(
                            day=calendar.monthrange(data_2m.year, data_2m.month)[1]
                        )

                # Valor do contrato: usa o informado, senão usa o valor da medição
                try:
                    valor_contrato_float = float(valor_contrato) if valor_contrato else valor_float
                except (ValueError, TypeError):
                    valor_contrato_float = valor_float

                saldo_contrato = max(0.0, valor_contrato_float - valor_float)
                status_contrato = "CONCLUÍDO" if saldo_contrato == 0 else "ATIVO"

                # Buscar CNPJ do fornecedor
                cnpj = self.buscar_cnpj_fornecedor(str(fornecedor).strip())
                if not cnpj:
                    erros.append(
                        f"Linha {linha}: Fornecedor '{fornecedor}' não encontrado no cadastro."
                    )
                    continue

                # Próximo ID de contrato
                next_id_contrato = 1
                for row in ws_contratos.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0] and isinstance(row[0], int) and row[0] >= next_id_contrato:
                        next_id_contrato = row[0] + 1

                # Criar contrato adicional
                pl = ws_contratos.max_row + 1
                ws_contratos.cell(row=pl, column=1,  value=next_id_contrato)
                ws_contratos.cell(row=pl, column=2,  value=cnpj)
                ws_contratos.cell(row=pl, column=3,  value=str(fornecedor).strip().upper())
                ws_contratos.cell(row=pl, column=4,  value=str(descricao).strip().upper())
                c = ws_contratos.cell(row=pl, column=5, value=dt_inicio or dt_med)
                c.number_format = 'DD/MM/YYYY'
                c = ws_contratos.cell(row=pl, column=6, value=dt_fim)
                c.number_format = 'DD/MM/YYYY'
                c = ws_contratos.cell(row=pl, column=7, value=valor_contrato_float)
                c.number_format = '#.##0,00'
                c = ws_contratos.cell(row=pl, column=8, value=valor_float)
                c.number_format = '#.##0,00'
                c = ws_contratos.cell(row=pl, column=9, value=saldo_contrato)
                c.number_format = '#.##0,00'
                ws_contratos.cell(row=pl, column=10, value=status_contrato)
                obs = f"SERVIÇO ADICIONAL — {str(observacao).strip().upper() if observacao else ''}"
                ws_contratos.cell(row=pl, column=11, value=obs)

                # Criar medição vinculada
                pm = ws_medicoes.max_row + 1
                ws_medicoes.cell(row=pm, column=1,  value=next_id_contrato)
                ws_medicoes.cell(row=pm, column=2,  value=1)
                ws_medicoes.cell(row=pm, column=3,  value=cnpj)
                ws_medicoes.cell(row=pm, column=4,  value=str(fornecedor).strip().upper())
                c = ws_medicoes.cell(row=pm, column=5, value=dt_med)
                c.number_format = 'DD/MM/YYYY'
                c = ws_medicoes.cell(row=pm, column=6, value=dt_pag)
                c.number_format = 'DD/MM/YYYY'
                ref = str(referencia).strip().upper() if referencia else str(descricao).strip().upper()
                ws_medicoes.cell(row=pm, column=7,  value=ref)
                c = ws_medicoes.cell(row=pm, column=8, value=valor_float)
                c.number_format = '#.##0,00'
                ws_medicoes.cell(row=pm, column=9,  value="PENDENTE")
                ws_medicoes.cell(row=pm, column=10, value=None)
                ws_medicoes.cell(row=pm, column=11, value=obs)

                logger.info(
                    f"✓ Serviço adicional: Contrato {next_id_contrato} | "
                    f"{fornecedor} | R$ {valor_float:.2f}"
                )

            wb_cliente.save(arquivo_cliente)
            wb_cliente.close()

        except Exception as e:
            erros.append(f"Erro ao processar serviços adicionais: {str(e)}")
            logger.error(f"Erro em processar_servicos_adicionais: {e}")

        return medicoes_adic, erros

    def buscar_cnpj_fornecedor(self, nome_fornecedor):
        if nome_fornecedor in self.cache_fornecedores:
            return self.cache_fornecedores[nome_fornecedor]

        try:
            fornecedor = None

            if hasattr(self.sistema, 'buscar_fornecedor_por_nome_agenda'):
                fornecedor = self.sistema.buscar_fornecedor_por_nome_agenda(nome_fornecedor)

            if fornecedor and 'cnpj_cpf' in fornecedor:
                cnpj = fornecedor['cnpj_cpf']
                self.cache_fornecedores[nome_fornecedor] = cnpj
                return cnpj

            # ── FALLBACK: leitura direta do arquivo de fornecedores ──────────
            try:
                from src.config.config import ARQUIVO_FORNECEDORES
                import openpyxl as _ox
                wb = _ox.load_workbook(ARQUIVO_FORNECEDORES, read_only=True, data_only=True)
                ws = wb.active
                nome_busca = nome_fornecedor.strip().upper()
                melhor_cnpj, melhor_score = None, 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    cnpj  = str(row[0]).strip() if row[0] else ""
                    razao = str(row[2]).strip().upper() if row[2] else ""
                    nome_ = str(row[3]).strip().upper() if row[3] else ""
                    for candidato in [razao, nome_]:
                        if nome_busca == candidato:
                            wb.close()
                            self.cache_fornecedores[nome_fornecedor] = cnpj
                            return cnpj
                        if nome_busca in candidato or candidato in nome_busca:
                            score = len(set(nome_busca.split()) & set(candidato.split()))
                            if score > melhor_score:
                                melhor_score, melhor_cnpj = score, cnpj
                wb.close()
                if melhor_cnpj:
                    self.cache_fornecedores[nome_fornecedor] = melhor_cnpj
                    return melhor_cnpj
            except Exception as fe:
                logger.warning(f"Fallback de busca direta falhou: {fe}")
            # ─────────────────────────────────────────────────────────────────

            self.cache_fornecedores[nome_fornecedor] = None
            return None

        except Exception as e:
            logger.error(f"Erro ao buscar CNPJ do fornecedor '{nome_fornecedor}': {str(e)}")
            return None
    
    def processar_relatorio(self, arquivo):
        """
        Processa o relatório completo e extrai todas as medições novas.
        
        Args:
            arquivo (str): Caminho do arquivo Excel
            
        Returns:
            tuple: (medicoes, erros, resumo)
                - medicoes: lista de dicionários com as medições
                - erros: lista de mensagens de erro
                - resumo: dicionário com estatísticas
        """
        medicoes_total = []
        erros_total = []
        
        resumo = {
            'contratos_processados': 0,
            'medicoes_encontradas': 0,
            'valor_total': 0,
            'contratos_com_erro': [],
            'aditivos': {},  # {id_contrato: novo_valor_global}
        }
        
        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            
            # Processar cada aba de contrato
            abas_contratos = [aba for aba in wb.sheetnames if aba.startswith("Contrato ")]
            
            for aba_nome in abas_contratos:
                try:
                    sheet = wb[aba_nome]
                    
                    # Extrair informações do contrato
                    id_contrato = sheet['B3'].value
                    fornecedor = sheet['B4'].value
                    saldo = sheet['B8'].value
                    
                    # Validar dados do contrato
                    if id_contrato is None:
                        erros_total.append(f"{aba_nome}: ID do contrato não encontrado")
                        resumo['contratos_com_erro'].append(aba_nome)
                        continue
                    
                    if fornecedor is None:
                        erros_total.append(f"{aba_nome}: Fornecedor não encontrado")
                        resumo['contratos_com_erro'].append(aba_nome)
                        continue
                    
                    if saldo is None:
                        saldo = 0
                    
                    try:
                        saldo_float = float(saldo)
                    except (ValueError, TypeError):
                        erros_total.append(f"{aba_nome}: Saldo inválido '{saldo}'")
                        resumo['contratos_com_erro'].append(aba_nome)
                        continue
                    
                    # Validar se contrato tem saldo (apenas se validação ativada)
                    if self.validacoes_ativadas['saldo'] and saldo_float <= 0:
                        logger.info(f"{aba_nome}: Sem saldo disponível, pulando")
                        resumo['contratos_processados'] += 1
                        continue
                    
                    # Buscar CNPJ do fornecedor
                    cnpj = self.buscar_cnpj_fornecedor(str(fornecedor).strip())
                    
                    # Ler aditivo (B9): novo Valor Global informado pelo usuário
                    aditivo_raw = sheet['B9'].value
                    if aditivo_raw is not None:
                        try:
                            aditivo_float = float(aditivo_raw)
                            if aditivo_float > 0:
                                resumo['aditivos'][id_contrato] = aditivo_float
                                valor_pago_aba = sheet['E7'].value
                                pago = float(valor_pago_aba) if valor_pago_aba else 0
                                saldo_float = aditivo_float - pago
                        except (ValueError, TypeError):
                            erros_total.append(f"{aba_nome}: Valor de aditivo inválido em B9 '{aditivo_raw}'")

                    # Extrair medições desta aba
                    medicoes, erros = self.extrair_medicoes_contrato(
                        sheet,
                        id_contrato,
                        fornecedor,
                        saldo_float
                    )
                    
                    # Adicionar CNPJ às medições
                    for medicao in medicoes:
                        medicao['cnpj_fornecedor'] = cnpj if cnpj else 'NÃO ENCONTRADO'
                    
                    medicoes_total.extend(medicoes)
                    erros_total.extend([f"{aba_nome}: {erro}" for erro in erros])
                    
                    resumo['contratos_processados'] += 1
                    resumo['medicoes_encontradas'] += len(medicoes)
                    resumo['valor_total'] += sum(m['valor'] for m in medicoes)
                    
                    if erros:
                        resumo['contratos_com_erro'].append(aba_nome)
                    
                except Exception as e:
                    erro_msg = f"{aba_nome}: Erro ao processar - {str(e)}"
                    erros_total.append(erro_msg)
                    resumo['contratos_com_erro'].append(aba_nome)
                    logger.error(erro_msg)
            
            wb.close()
            
        except Exception as e:
            erro_msg = f"Erro ao processar relatório: {str(e)}"
            erros_total.append(erro_msg)
            logger.error(erro_msg)
        
        return medicoes_total, erros_total, resumo
    
    def preparar_dados_para_sistema(self, medicoes):
        """
        Prepara os dados das medições para serem salvos no sistema.
        
        IMPORTANTE: Mantém datas como datetime para salvar corretamente no Excel!
        """
        medicoes_preparadas = []
        
        for medicao in medicoes:
            data_medicao = medicao.get('data_medicao')
            data_pagamento = medicao.get('data_pagamento')
            
            # ✅ CORREÇÃO: Manter como datetime, não converter para string!
            # Se data_pagamento não existe, usar data_medicao
            if data_pagamento is None:
                data_pagamento = data_medicao
            
            medicoes_preparadas.append({
                'ID_Contrato': medicao['id_contrato'],
                'ID_Medicao': medicao['id_medicao'],
                'CNPJ_Fornecedor': medicao['cnpj_fornecedor'],
                'Nome_Fornecedor': medicao['fornecedor'],
                'Data_Medicao': data_medicao,  # ✅ DATETIME!
                'Data_Pagamento': data_pagamento,  # ✅ DATETIME!
                'Referencia': medicao['referencia'],
                'Valor': medicao['valor'],
                'Status': medicao['status']
            })
        
        return medicoes_preparadas
    
    def salvar_medicoes_cliente(self, cliente, medicoes_preparadas):
        """
        Salva as medições no arquivo do cliente.
        
        Args:
            cliente (str): Nome do cliente
            medicoes_preparadas (list): Lista com as medições preparadas
            
        Returns:
            tuple: (bool, str) - (sucesso, mensagem)
        """
        try:
            # Garantir que PASTA_CLIENTES seja um Path
            pasta_clientes_path = Path(PASTA_CLIENTES) if not isinstance(PASTA_CLIENTES, Path) else PASTA_CLIENTES
            base_path = Path(BASE_PATH) if not isinstance(BASE_PATH, Path) else BASE_PATH
            
            # Tentar diferentes caminhos possíveis para o arquivo do cliente
            # A ordem importa - começar pelos mais prováveis
            possiveis_caminhos = [
                pasta_clientes_path / f"{cliente}.xlsx",  # Caminho configurado (prioritário)
                base_path / f"{cliente}.xlsx",  # BASE_PATH
                base_path / "Clientes" / f"{cliente}.xlsx",  # Subpasta Clientes em BASE_PATH
            ]
            
            # Se o sistema tiver um atributo com o caminho do cliente atual
            if hasattr(self.sistema, 'arquivo_cliente_atual'):
                possiveis_caminhos.insert(0, Path(self.sistema.arquivo_cliente_atual))
            
            # Log dos caminhos que serão verificados
            logger.info("=" * 60)
            logger.info("Buscando arquivo do cliente:")
            logger.info(f"Cliente: {cliente}")
            logger.info(f"PASTA_CLIENTES (config): {PASTA_CLIENTES}")
            logger.info(f"BASE_PATH (config): {BASE_PATH}")
            logger.info("Caminhos a verificar:")
            for idx, caminho in enumerate(possiveis_caminhos, 1):
                logger.info(f"  {idx}. {caminho}")
            logger.info("=" * 60)
            
            # Procurar o arquivo
            arquivo_cliente = None
            for caminho in possiveis_caminhos:
                logger.debug(f"Verificando caminho: {caminho}")
                if caminho.exists():
                    arquivo_cliente = caminho
                    logger.info(f"✓ Arquivo do cliente encontrado em: {arquivo_cliente}")
                    break
                else:
                    logger.debug(f"✗ Caminho não existe: {caminho}")
            
            # Verificar se arquivo foi encontrado
            if not arquivo_cliente:
                caminhos_tentados = '\n'.join([f"  • {str(p)}" for p in possiveis_caminhos])
                
                # Perguntar se deseja selecionar manualmente
                mensagem_erro = (
                    f"📁 Arquivo do cliente não encontrado automaticamente.\n\n"
                    f"Cliente: {cliente}\n\n"
                    f"Caminhos verificados:\n{caminhos_tentados}\n\n"
                    f"Deseja selecionar o arquivo manualmente?"
                )
                
                if custom_messagebox("yesno", "Arquivo Não Encontrado", mensagem_erro):
                    # Permitir seleção manual
                    arquivo_selecionado = filedialog.askopenfilename(
                        title=f"Selecione o arquivo do cliente {cliente}",
                        filetypes=[
                            ("Arquivos Excel", "*.xlsx"),
                            ("Todos os arquivos", "*.*")
                        ],
                        initialdir=str(PASTA_CLIENTES)
                    )
                    
                    if arquivo_selecionado:
                        arquivo_cliente = Path(arquivo_selecionado)
                        logger.info(f"Arquivo selecionado manualmente: {arquivo_cliente}")
                    else:
                        return False, "Seleção de arquivo cancelada pelo usuário"
                else:
                    return False, "Importação cancelada - arquivo não encontrado"

            
            # Carregar arquivo
            wb = openpyxl.load_workbook(arquivo_cliente)
            
            # Verificar se aba Medicoes existe
            if 'Medicoes' not in wb.sheetnames:
                wb.close()
                return False, "Aba 'Medicoes' não encontrada no arquivo do cliente"
            
            sheet = wb['Medicoes']
            
            # ===== PROTEÇÃO CONTRA DUPLICAÇÃO =====
            # Verificar se alguma das medições já existe
            medicoes_existentes = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # ID_Contrato e ID_Medicao
                    chave = f"{row[0]}_{row[1]}"  # Exemplo: "3_12"
                    medicoes_existentes.add(chave)
            
            # Verificar se alguma medição a ser importada já existe
            medicoes_duplicadas = []
            for medicao in medicoes_preparadas:
                chave = f"{medicao['ID_Contrato']}_{medicao['ID_Medicao']}"
                if chave in medicoes_existentes:
                    medicoes_duplicadas.append(
                        f"Contrato {medicao['ID_Contrato']} - Medição {medicao['ID_Medicao']}"
                    )
            
            if medicoes_duplicadas:
                wb.close()
                mensagem_erro = (
                    f"⚠️ DUPLICAÇÃO DETECTADA!\n\n"
                    f"As seguintes medições JÁ EXISTEM no arquivo:\n\n"
                )
                for dup in medicoes_duplicadas[:5]:
                    mensagem_erro += f"  • {dup}\n"
                
                if len(medicoes_duplicadas) > 5:
                    mensagem_erro += f"  • ... e mais {len(medicoes_duplicadas) - 5}\n"
                
                mensagem_erro += (
                    f"\n❌ Importação CANCELADA para evitar duplicação.\n\n"
                    f"💡 Dica: Verifique se este relatório já foi importado antes."
                )
                
                logger.warning(f"Importação cancelada - {len(medicoes_duplicadas)} medições duplicadas detectadas")
                
                # Mostrar alerta ao usuário
                custom_messagebox("warning", "Duplicação Detectada", mensagem_erro)
                
                return False, f"Importação cancelada - {len(medicoes_duplicadas)} medições já existem"
            # ===== FIM DA PROTEÇÃO =====
            
            # Encontrar próxima linha vazia
            proxima_linha = sheet.max_row + 1
            
            # Adicionar cada medição
            for medicao in medicoes_preparadas:
                sheet.cell(row=proxima_linha, column=1, value=medicao['ID_Contrato'])  # A
                sheet.cell(row=proxima_linha, column=2, value=medicao['ID_Medicao'])  # B
                sheet.cell(row=proxima_linha, column=3, value=medicao['CNPJ_Fornecedor'])  # C
                cnpj_val = str(medicao['CNPJ_Fornecedor']).strip()
                apenas_numeros = ''.join(filter(str.isdigit, cnpj_val))
                tipo_pessoa = medicao.get('tipo_pessoa', '').upper()

                if tipo_pessoa == 'PJ':
                    cnpj_completo = apenas_numeros.zfill(14)
                    cnpj_fmt = f"{cnpj_completo[:2]}.{cnpj_completo[2:5]}.{cnpj_completo[5:8]}/{cnpj_completo[8:12]}-{cnpj_completo[12:]}"
                elif tipo_pessoa == 'PF':
                    cnpj_completo = apenas_numeros.zfill(11)
                    cnpj_fmt = f"{cnpj_completo[:3]}.{cnpj_completo[3:6]}.{cnpj_completo[6:9]}-{cnpj_completo[9:]}"
                else:
                    cnpj_fmt = cnpj_val  # tipo desconhecido, mantém original
                sheet.cell(row=proxima_linha, column=3, value=cnpj_fmt)
                sheet.cell(row=proxima_linha, column=4, value=medicao['Nome_Fornecedor'])  # D
                sheet.cell(row=proxima_linha, column=5, value=medicao['Data_Medicao'])  # E
                sheet.cell(row=proxima_linha, column=6, value=medicao['Data_Pagamento'])  # F
                sheet.cell(row=proxima_linha, column=7, value=medicao['Referencia'])  # G
                sheet.cell(row=proxima_linha, column=8, value=medicao['Valor'])  # H
                sheet.cell(row=proxima_linha, column=9, value=medicao['Status'])  # I
                
                # Formatar células de data (colunas E e F)
                for col_date in [5, 6]:  # Colunas E (Data_Medicao) e F (Data_Pagamento)
                    cell = sheet.cell(row=proxima_linha, column=col_date)
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'DD/MM/YYYY'
                
                # Formatar célula de valor (coluna H)
                sheet.cell(row=proxima_linha, column=8).number_format = '#,##0.00'

                proxima_linha += 1
            
            # Salvar arquivo
            wb.save(arquivo_cliente)
            wb.close()
            
            logger.info(f"{len(medicoes_preparadas)} medições salvas no arquivo {arquivo_cliente}")
            return True, f"{len(medicoes_preparadas)} medições salvas com sucesso"
            
        except Exception as e:
            erro_msg = f"Erro ao salvar medições: {str(e)}"
            logger.error(erro_msg)
            return False, erro_msg
        
    def atualizar_saldos_contratos(self, arquivo_cliente, medicoes_importadas):
        """
        Atualiza os valores pagos e saldos na aba Contratos_Medicao
        
        NOVA FUNCIONALIDADE:
        - Quando Saldo = 0, atualiza Status para "CONCLUÍDO"
        
        Args:
            arquivo_cliente (Path): Caminho do arquivo do cliente
            medicoes_importadas (list): Lista de medições que foram importadas
            
        Returns:
            tuple: (sucesso, mensagem)
        """
        try:
            logger.info("=" * 60)
            logger.info("ATUALIZANDO SALDOS DOS CONTRATOS")
            logger.info("=" * 60)
            
            # Carregar arquivo
            wb = openpyxl.load_workbook(arquivo_cliente)
            
            # Verificar se aba Contratos_Medicao existe
            if 'Contratos_Medicao' not in wb.sheetnames:
                wb.close()
                return False, "Aba 'Contratos_Medicao' não encontrada no arquivo do cliente"
            
            # Verificar se aba Medicoes existe
            if 'Medicoes' not in wb.sheetnames:
                wb.close()
                return False, "Aba 'Medicoes' não encontrada no arquivo do cliente"
            
            sheet_contratos = wb['Contratos_Medicao']
            sheet_medicoes = wb['Medicoes']
            
            # ESTRUTURA DA ABA Contratos_Medicao:
            # A: ID_Contrato
            # B: CNPJ_Fornecedor
            # C: Nome_Fornecedor
            # D: Descricao
            # E: Data_Inicio
            # F: Data_Final
            # G: Valor_Global
            # H: Valor_Pago      <<< SERÁ ATUALIZADO
            # I: Saldo           <<< SERÁ ATUALIZADO
            # J: Status          <<< SERÁ ATUALIZADO SE SALDO = 0
            # K: Observacao
            
            # Mapear contratos por ID
            contratos_map = {}
            for row in sheet_contratos.iter_rows(min_row=2, values_only=False):
                if row[0].value:  # Se tem ID_Contrato
                    id_contrato = row[0].value
                    contratos_map[id_contrato] = {
                        'row_idx': row[0].row,
                        'valor_global': row[6].value or 0,  # Coluna G
                        'valor_pago_atual': row[7].value or 0,  # Coluna H
                        'status_atual': row[9].value  # Coluna J
                    }
            
            logger.info(f"Encontrados {len(contratos_map)} contratos na aba Contratos_Medicao")
            
            # Calcular totais pagos por contrato a partir de TODAS as medições
            # (não apenas as recém-importadas)
            totais_por_contrato = {}
            for row in sheet_medicoes.iter_rows(min_row=2, values_only=True):
                if row[0] and row[7]:  # Se tem ID_Contrato e Valor
                    id_contrato = row[0]  # Coluna A
                    valor = float(row[7])  # Coluna H (Valor)
                    
                    if id_contrato in totais_por_contrato:
                        totais_por_contrato[id_contrato] += valor
                    else:
                        totais_por_contrato[id_contrato] = valor
            
            logger.info(f"Calculados totais pagos para {len(totais_por_contrato)} contratos")
            
            # Atualizar cada contrato
            contratos_atualizados = 0
            contratos_concluidos = 0  # Contador de contratos que foram concluídos
            ids_alterados = []        # IDs dos contratos que tiveram saldo atualizado

            for id_contrato, total_pago in totais_por_contrato.items():
                if id_contrato in contratos_map:
                    contrato_info = contratos_map[id_contrato]
                    row_idx = contrato_info['row_idx']
                    valor_global = float(contrato_info['valor_global'])
                    status_atual = contrato_info['status_atual']

                    # Calcular novo saldo
                    novo_saldo = valor_global - total_pago

                    # Arredondar para evitar problemas com float
                    # (ex: 0.0000001 deve ser considerado 0)
                    if abs(novo_saldo) < 0.01:  # Menos de 1 centavo = 0
                        novo_saldo = 0

                    # Determinar novo status
                    novo_status = status_atual
                    if novo_saldo == 0 and status_atual and status_atual.upper() == "ATIVO":
                        novo_status = "CONCLUÍDO"
                        contratos_concluidos += 1
                        logger.info(f"⭐ Contrato {id_contrato} será marcado como CONCLUÍDO (saldo zerado)")

                    # Atualizar células
                    # Coluna H: Valor_Pago
                    sheet_contratos.cell(row=row_idx, column=8).value = total_pago

                    # Coluna I: Saldo
                    sheet_contratos.cell(row=row_idx, column=9).value = novo_saldo

                    # Coluna J: Status (atualiza apenas se mudou para CONCLUÍDO)
                    if novo_status != status_atual:
                        sheet_contratos.cell(row=row_idx, column=10).value = novo_status

                    contratos_atualizados += 1
                    ids_alterados.append(id_contrato)

                    # Log detalhado
                    log_msg = f"Contrato {id_contrato}: Valor Pago = R$ {total_pago:,.2f} | Saldo = R$ {novo_saldo:,.2f}"
                    if novo_status != status_atual:
                        log_msg += f" | Status: {status_atual} → {novo_status}"
                    logger.info(log_msg)
            
            # Salvar arquivo
            wb.save(arquivo_cliente)
            wb.close()
            
            logger.info("=" * 60)
            logger.info(f"✓ SALDOS ATUALIZADOS COM SUCESSO!")
            logger.info(f"✓ {contratos_atualizados} contratos atualizados")
            if contratos_concluidos > 0:
                logger.info(f"⭐ {contratos_concluidos} contrato(s) marcado(s) como CONCLUÍDO")
            logger.info("=" * 60)
            
            # Expor lista de IDs alterados para que a UI possa destacar as linhas
            self.ids_contratos_alterados = ids_alterados

            # Mensagem de retorno
            mensagem = f"{contratos_atualizados} contrato(s) atualizado(s)"
            if contratos_concluidos > 0:
                mensagem += f" ({contratos_concluidos} concluído(s))"

            return True, mensagem
            
        except Exception as e:
            erro_msg = f"Erro ao atualizar saldos dos contratos: {str(e)}"
            logger.error(erro_msg)
            import traceback
            logger.error(traceback.format_exc())
            return False, erro_msg
    
    def mover_para_importados(self, arquivo_relatorio):
        """
        MÉTODO NOVO - Move o arquivo importado para a pasta Importados
        
        Args:
            arquivo_relatorio (str): Caminho do arquivo original
            
        Returns:
            tuple: (sucesso, novo_caminho)
        """
        try:
            arquivo_path = Path(arquivo_relatorio)
            
            # Verificar se arquivo existe
            if not arquivo_path.exists():
                logger.warning(f"Arquivo não encontrado para mover: {arquivo_relatorio}")
                return False, None
            
            # Gerar nome de destino
            nome_arquivo = arquivo_path.name
            destino = self.pasta_importados / nome_arquivo
            
            # Se já existe arquivo com mesmo nome, adicionar timestamp
            if destino.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                nome_base = arquivo_path.stem
                extensao = arquivo_path.suffix
                nome_arquivo = f"{nome_base}_{timestamp}{extensao}"
                destino = self.pasta_importados / nome_arquivo
            
            # Mover arquivo
            shutil.move(str(arquivo_path), str(destino))
            
            logger.info(f"Arquivo movido com sucesso:")
            logger.info(f"  De: {arquivo_path}")
            logger.info(f"  Para: {destino}")
            
            return True, str(destino)
            
        except Exception as e:
            erro_msg = f"Erro ao mover arquivo: {str(e)}"
            logger.error(erro_msg)
            return False, None
        
    # def renomear_arquivo_importado(self, arquivo_relatorio):
    #     """
    #     Renomeia o arquivo de relatório após importação bem-sucedida.
    #     Adiciona sufixo _IMPORTADO_[DATA_HORA] para evitar duplicidade.
        
    #     Args:
    #         arquivo_relatorio (str): Caminho do arquivo original
            
    #     Returns:
    #         tuple: (bool, str) - (sucesso, novo_nome)
    #     """
    #     try:
    #         arquivo_path = Path(arquivo_relatorio)
            
    #         # Verificar se arquivo existe
    #         if not arquivo_path.exists():
    #             logger.warning(f"Arquivo não encontrado para renomear: {arquivo_relatorio}")
    #             return False, None
            
    #         # Gerar timestamp
    #         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    #         nome_base = arquivo_path.stem  # Nome sem extensão
    #         extensao = arquivo_path.suffix  # .xlsx
            
    #         # Se o arquivo já tem _IMPORTADO no nome, não adicionar novamente
    #         if "_IMPORTADO_" in nome_base:
    #             logger.info("Arquivo já foi renomeado anteriormente, não renomeando novamente")
    #             return False, None
            
    #         # Gerar novo nome
    #         novo_nome = f"{nome_base}_IMPORTADO_{timestamp}{extensao}"
    #         novo_caminho = arquivo_path.parent / novo_nome
            
    #         # Renomear arquivo
    #         arquivo_path.rename(novo_caminho)
            
    #         logger.info(f"Arquivo renomeado com sucesso:")
    #         logger.info(f"  De: {arquivo_path.name}")
    #         logger.info(f"  Para: {novo_nome}")
            
    #         return True, novo_nome
            
    #     except Exception as e:
    #         erro_msg = f"Erro ao renomear arquivo: {str(e)}"
    #         logger.error(erro_msg)
    #         return False, None

    def _tem_servicos_adicionais(self, arquivo):
        """Retorna True se Servicos_Adicionais contiver pelo menos uma linha com dados."""
        try:
            wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
            if 'Servicos_Adicionais' not in wb.sheetnames:
                wb.close()
                return False
            sheet = wb['Servicos_Adicionais']
            for row in sheet.iter_rows(min_row=5, values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row[:3]):
                    wb.close()
                    return True
            wb.close()
            return False
        except Exception:
            return False

    def importar_medicoes(self):
        """
        Método principal para importar medições de contratos.
        
        Este é o método que deve ser chamado pelo botão na interface.
        """
        # Verificar se há um cliente selecionado
        if not hasattr(self.sistema, 'cliente_atual') or not self.sistema.cliente_atual:
            custom_messagebox("error", 
                "Erro", 
                "Nenhum cliente selecionado.\n\n"
                "Por favor, selecione um cliente antes de importar medições."
            )
            return
        
        cliente_atual = self.sistema.cliente_atual.upper()
        
        # Passo 1: Selecionar arquivo
        arquivo = self.selecionar_arquivo_relatorio()
        if not arquivo:
            return
        
        # Passo 2: Validar estrutura
        valido, mensagem = self.validar_estrutura_relatorio(arquivo)
        if not valido:
            custom_messagebox("error", 
                "Estrutura Inválida", 
                f"O arquivo selecionado não possui a estrutura esperada:\n\n{mensagem}"
            )
            return
        
        # Confirmar com usuário
        if not custom_messagebox("yesno", 
            "Confirmar Importação", 
            f"📊 Relatório válido encontrado!\n\n"
            f"• {mensagem}\n"
            f"• Cliente: {cliente_atual}\n\n"
            f"Deseja processar as medições?"
        ):
            return
        
        # Passo 3: Processar relatório
        medicoes, erros, resumo = self.processar_relatorio(arquivo)
        tem_adicionais = self._tem_servicos_adicionais(arquivo)

        # Verificar se encontrou medições
        if len(medicoes) == 0:
            if not tem_adicionais:
                mensagem_resultado = (
                    f"✅ Processamento concluído!\n\n"
                    f"📈 Resumo:\n"
                    f"• Contratos verificados: {resumo['contratos_processados']}\n"
                    f"• Medições novas encontradas: 0\n\n"
                )

                if len(erros) > 0:
                    mensagem_resultado += f"⚠️ Foram encontrados {len(erros)} avisos:\n\n"
                    for erro in erros[:5]:
                        mensagem_resultado += f"• {erro}\n"
                    if len(erros) > 5:
                        mensagem_resultado += f"• ... e mais {len(erros) - 5} avisos\n"
                else:
                    mensagem_resultado += "ℹ️ Nenhuma medição nova (status PENDENTE) foi encontrada."

                custom_messagebox("info", "Resultado", mensagem_resultado)
                return

            # Nenhuma medição regular, mas há serviços adicionais para processar
            if not custom_messagebox("yesno",
                "Serviços Adicionais Encontrados",
                f"Nenhuma medição regular (status PENDENTE) foi encontrada nos contratos existentes.\n\n"
                f"⚠️ Foram encontrados serviços adicionais na aba 'Servicos_Adicionais'.\n\n"
                f"• Cliente: {cliente_atual}\n\n"
                f"Deseja importar os serviços adicionais?"
            ):
                return

            arquivo_cliente = self._obter_arquivo_cliente(cliente_atual)
            if arquivo_cliente:
                medicoes_adic, erros_adic = self.processar_servicos_adicionais(arquivo, arquivo_cliente)
                if erros_adic:
                    erros_resumo = '\n'.join(erros_adic[:5])
                    custom_messagebox(
                        "warning",
                        "Avisos — Serviços Adicionais",
                        f"Alguns serviços adicionais não puderam ser importados:\n\n{erros_resumo}"
                    )

            if custom_messagebox("yesno",
                "Mover Arquivo?",
                "Deseja mover o arquivo importado para a pasta 'Importados'?\n\n"
                "Isso organiza os arquivos e evita reimportação acidental.\n\n"
                "Recomendado: SIM"
            ):
                self.mover_para_importados(arquivo)

            custom_messagebox("info", "Concluído", "Serviços adicionais importados com sucesso!")
            return
        
        # Passo 4: Mostrar resumo e pedir confirmação
        mensagem_confirmacao = (
            f"📊 Medições Encontradas\n\n"
            f"• Total de medições: {resumo['medicoes_encontradas']}\n"
            f"• Valor total: R$ {resumo['valor_total']:,.2f}\n"
            f"• Contratos processados: {resumo['contratos_processados']}\n"
        )
        
        if len(erros) > 0:
            mensagem_confirmacao += f"\n⚠️ Avisos: {len(erros)}\n"
        
        if len(resumo['contratos_com_erro']) > 0:
            mensagem_confirmacao += f"\n⚠️ Contratos com problemas: {len(resumo['contratos_com_erro'])}\n"
        
        mensagem_confirmacao += "\nDeseja importar estas medições?"
        
        if not custom_messagebox("yesno", "Confirmar Importação", mensagem_confirmacao):
            custom_messagebox("info", "Cancelado", "Importação cancelada pelo usuário.")
            return
        
        # Passo 5: Preparar dados
        medicoes_preparadas = self.preparar_dados_para_sistema(medicoes)
        
        # Passo 6: Salvar no arquivo do cliente
        sucesso, mensagem_salvamento = self.salvar_medicoes_cliente(cliente_atual, medicoes_preparadas)
        
        # Passo 7: Atualizar saldos dos contratos
        arquivo_cliente = self._obter_arquivo_cliente(cliente_atual)
        sucesso_saldos = False
        mensagem_saldos = ""

        if arquivo_cliente and resumo.get('aditivos'):
            self._aplicar_aditivos(arquivo_cliente, resumo['aditivos'])

        if arquivo_cliente:
            sucesso_saldos, mensagem_saldos = self.atualizar_saldos_contratos(
                arquivo_cliente,
                medicoes_preparadas
            )
            
            if not sucesso_saldos:
                logger.warning(f"Aviso: {mensagem_saldos}")
        
        
        # Passo 6b: Processar serviços adicionais (se existirem na planilha)
        if arquivo_cliente:
            medicoes_adic, erros_adic = self.processar_servicos_adicionais(
                arquivo, arquivo_cliente
            )
            if erros_adic:
                erros_resumo = '\n'.join(erros_adic[:5])
                custom_messagebox(
                    "warning",
                    "Avisos — Serviços Adicionais",
                    f"Alguns serviços adicionais não puderam ser importados:\n\n{erros_resumo}"
                )
            elif medicoes_adic is not None:
                # Opcional: acumular no resumo final
                pass

        # Passo 8: Mover arquivo para pasta Importados (SUBSTITUI renomear)
        arquivo_movido = False
        novo_caminho = None
        
        if custom_messagebox("yesno", 
            "Mover Arquivo?", 
            f"Deseja mover o arquivo importado para a pasta 'Importados'?\n\n"
            f"Isso organiza os arquivos e evita reimportação acidental.\n\n"
            f"Recomendado: SIM"
        ):
            sucesso_mover, novo_caminho = self.mover_para_importados(arquivo)
            if sucesso_mover:
                arquivo_movido = True
                logger.info(f"✓ Arquivo movido: {novo_caminho}")
            else:
                logger.warning("⚠ Não foi possível mover o arquivo")
        
        # Passo 9: Mostrar resultado final
        mensagem_final = (
            f"✅ Importação Concluída!\n\n"
            f"📊 Resumo:\n"
            f"• Medições: {len(medicoes_preparadas)}\n"
            f"• Valor: R$ {resumo['valor_total']:,.2f}\n"
            f"• Cliente: {cliente_atual}\n"
        )
        
        if sucesso_saldos:
            mensagem_final += f"• Saldos atualizados: {mensagem_saldos}\n"
        
        if arquivo_movido:
            mensagem_final += f"\n📁 Arquivo movido para:\n{Path(novo_caminho).name}\n"
        
        custom_messagebox("info", "Sucesso", mensagem_final)
    def _aplicar_aditivos(self, arquivo_cliente, aditivos):
        """Atualiza Valor Global de contratos que receberam aditivo na planilha (B9)."""
        try:
            wb = openpyxl.load_workbook(arquivo_cliente)
            if 'Contratos_Medicao' not in wb.sheetnames:
                wb.close()
                return
            ws = wb['Contratos_Medicao']
            for row in ws.iter_rows(min_row=2, values_only=False):
                id_cont = row[0].value
                if id_cont in aditivos:
                    row[6].value = aditivos[id_cont]  # coluna G = Valor_Global
                    row[6].number_format = '#,##0.00'
                    logger.info(f"Aditivo aplicado: Contrato {id_cont} → Valor Global R$ {aditivos[id_cont]:,.2f}")
            wb.save(arquivo_cliente)
            wb.close()
        except Exception as e:
            logger.error(f"Erro ao aplicar aditivos: {e}")

    def _obter_arquivo_cliente(self, cliente):
        """
        Método auxiliar para obter o caminho do arquivo do cliente
        
        Args:
            cliente (str): Nome do cliente
            
        Returns:
            Path: Caminho do arquivo ou None se não encontrar
        """
        try:
            # Possíveis localizações do arquivo
            possiveis_caminhos = [
                Path(PASTA_CLIENTES) / f"{cliente}.xlsx",
                Path(PASTA_CLIENTES) / cliente / f"{cliente}.xlsx",
                Path(BASE_PATH) / "Clientes" / f"{cliente}.xlsx"
            ]
            
            # Procurar o arquivo
            for caminho in possiveis_caminhos:
                if caminho.exists():
                    return caminho
            
            return None
            
        except Exception as e:
            logger.error(f"Erro ao obter arquivo do cliente: {e}")
            return None