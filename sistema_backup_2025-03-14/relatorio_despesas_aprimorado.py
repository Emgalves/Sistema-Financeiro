import sys
import os
import pandas as pd
import xlwings as xw
import openpyxl
import warnings
import platform
import subprocess
import tkinter as tk
import numpy as np
from tkinter import Tk
from openpyxl import load_workbook
from tkinter import ttk, messagebox, filedialog, StringVar, Toplevel, BooleanVar
from tkcalendar import Calendar
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    PageTemplate, Frame, Spacer, PageBreak, Image
)
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from reportlab.platypus import KeepTogether
import logging

# Criar diretório para logs se não existir
log_dir = 'logs'
os.makedirs(log_dir, exist_ok=True)

# Nome do arquivo de log com data
log_file = os.path.join(log_dir, f'relatorio_{datetime.now().strftime("%Y%m%d")}.log')

# Configuração do logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()  # Para exibir também no console
    ]
)

# Criar logger específico para o módulo
logger = logging.getLogger(__name__)

# Configuração inicial
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# Variáveis globais
arquivo_path = None
arquivo_selecionado = None
data_selecionada = None
incluir_futuros = None
status_label = None
root = None
handler = None



class RelatorioUI:
    def __init__(self, parent):
        logger.info("Iniciando RelatorioUI")
        if parent is None:
            self.root = tk.Tk()
        else:
            self.root = parent
            
        logger.debug(f"Parent configurado: {parent}")
        
        self.arquivo_selecionado = StringVar(self.root, value="Nenhum arquivo selecionado")
        self.data_selecionada = StringVar(self.root, value=datetime.now().strftime('%d/%m/%Y'))
        logger.debug(f"Data inicial configurada: {self.data_selecionada.get()}")
        
        self.incluir_futuros = BooleanVar(value=True)
        self.status_label = None
        self.handler = RelatorioHandler()
        self.arquivos_lote = []
        self.menu_principal = None  # Adicionado aqui, antes do setup_ui
        self.setup_ui()

    def setup_ui(self):
        logger.info("Iniciando configuração da interface")

        self.root.title("Gerador de Relatório de Despesas")
        self.root.geometry("500x500")
        self.root.update_idletasks()


        # Container principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Data
        logger.debug("Configurando campos de data")
        frame_data = ttk.Frame(main_frame)
        frame_data.pack(pady=10, padx=20, fill='x')
        
        self.data_selecionada.set(datetime.now().strftime('%d/%m/%Y'))
        self.arquivo_selecionado.set("Nenhum arquivo selecionado")
        ttk.Label(frame_data, text="Data do relatório:").pack(side='left', padx=(0, 10))
        ttk.Label(frame_data, textvariable=self.data_selecionada, width=10).pack(side='left')
        ttk.Button(frame_data, text="Escolher Data", command=self.escolher_data).pack(side='left', padx=5)

        # Relatório Individual
        frame_arquivo = ttk.LabelFrame(main_frame, text="Relatório Individual")
        frame_arquivo.pack(pady=10, padx=20, fill='x')

        self.arquivo_selecionado.set("Nenhum arquivo selecionado")
        ttk.Button(frame_arquivo, text="Escolher arquivo", 
                  command=self.selecionar_arquivo_local).pack(pady=5, fill='x')
        ttk.Label(frame_arquivo, textvariable=self.arquivo_selecionado).pack(pady=5)
        ttk.Button(frame_arquivo, text="Gerar Relatório Individual",
                  command=self.gerar_relatorio).pack(pady=5, fill='x')

        # Relatório em Lote
        frame_lote = ttk.LabelFrame(main_frame, text="Relatório em Lote")
        frame_lote.pack(pady=10, padx=20, fill='x')
        ttk.Button(frame_lote, text="Selecionar Arquivos para Lote", 
                  command=self.selecionar_arquivos_lote).pack(pady=5, fill='x')

        # Checkbox para lançamentos futuros
        ttk.Checkbutton(main_frame, text="Incluir lançamentos futuros",
                       variable=self.incluir_futuros).pack(pady=10, anchor='w')

        # Status label
        self.status_label = ttk.Label(main_frame, text="", wraplength=350)
        self.status_label.pack(pady=10)

        # Adicione esta linha ao final do método:
        self.adicionar_botao_pendentes()

        logger.info("Interface configurada com sucesso")

        

    def escolher_data(self):
        top = Toplevel(self.root)
        top.title("Selecione a Data")
        
        x = self.root.winfo_x() + 50
        y = self.root.winfo_y() + 50
        top.geometry(f"+{x}+{y}")
        
        cal = Calendar(top,
                      selectmode='day',
                      year=datetime.now().year,
                      month=datetime.now().month,
                      day=datetime.now().day,
                      locale='pt_BR',
                      date_pattern='dd/mm/yyyy')
        cal.pack(padx=10, pady=10)
        
        def definir_data():
            data = cal.get_date()
            self.data_selecionada.set(data)
            top.destroy()
            
        ttk.Button(top, text="Confirmar", command=definir_data).pack(pady=5)
        top.transient(self.root)
        top.grab_set()
        self.root.wait_window(top)

        

    def selecionar_arquivo_local(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.arquivo_path = arquivo
            nome_arquivo = os.path.basename(arquivo)
            self.arquivo_selecionado.set(nome_arquivo)
            self.root.update_idletasks()

    def selecionar_arquivos_lote(self):
        files = filedialog.askopenfilenames(
            title="Selecione os arquivos Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if files:
            self.arquivos_lote = files
            self.processar_lote(files)

    def gerar_relatorio(self):
        try:
            if not self.arquivo_path:
                logger.warning("Tentativa de gerar relatório sem arquivo selecionado")
                self.status_label.config(text="Selecione um arquivo Excel!")
                return

            logger.info(f"Iniciando geração de relatório para arquivo: {self.arquivo_path}")
            data_rel = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
                
            # Carregar e processar dados
            df = self.handler.carregar_dados_excel(self.arquivo_path)  # Usar self.handler
            df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2 = self.handler.processar_dados(df, data_rel)
                
            # Processar lançamentos futuros
            df_futuro = None
            if self.incluir_futuros.get():
                df_futuro = self.handler.processar_lancamentos_futuros(df, data_rel)
                    
            # Processar workbook
            workbook = load_workbook(self.arquivo_path, data_only=True)
            ws_resumo = workbook['RESUMO']
            nome_cliente = ws_resumo['A3'].value
                
            # Obter número do relatório e valor acumulado
            numero_relatorio = self.handler.obter_numero_relatorio(ws_resumo, data_rel)
            valor_acumulado = self.handler.calcular_acumulado_dados(df, data_rel)
            
            logger.info(f"Número do relatório: {numero_relatorio}")
            logger.info(f"Valor acumulado calculado: {valor_acumulado:,.2f}")
                
            dados_completos = {
                'df_filtrado': df_filtrado,
                'df_diaria': df_diaria,
                'df_tp_desp_1': df_tp_desp_1,
                'df_tp_desp_2': df_tp_desp_2,
                'df_futuro': df_futuro,
                'df_original': df,
                'incluir_futuros': self.incluir_futuros.get(),
                'data_relatorio': data_rel,
                'nome_cliente': nome_cliente,
                'endereco_cliente': ws_resumo['A4'].value,
                'numero_relatorio': numero_relatorio,
                'acumulado': valor_acumulado  # Valor direto, sem conversão
            }
            
            logger.debug("Verificando dados antes de gerar PDF:")
            logger.debug(f"dados_completos['acumulado']: {dados_completos['acumulado']}")
            logger.debug(f"Tipo do acumulado: {type(dados_completos['acumulado'])}")
            
            # Gerar nome do arquivo
            data_formatada = data_rel.strftime('%d-%m-%Y')
            nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
            caminho_output = os.path.join(os.path.dirname(self.arquivo_path), nome_arquivo)
            
            # Gerar o PDF com os dados completos
            self.handler.gerar_relatorio_pdf(dados_completos, caminho_output, self.arquivo_path)
            
            self.status_label.config(text=f"Relatório gerado com sucesso para {nome_cliente}")
            self.criar_dialog_relatorio_gerado(nome_cliente, data_formatada)
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório: {str(e)}", exc_info=True)
            self.status_label.config(text=f"Erro: {str(e)}")


    def processar_lote(self, arquivos):
        """Processa arquivos em lote"""
        try:
            logger.info(f"Iniciando processamento em lote de {len(arquivos)} arquivos")

            progress_window = Toplevel(self.root)
            progress_window.title("Gerando Relatórios em Lote")
            progress_window.geometry("600x400")
            progress_window.transient(self.root)

            # Label para mostrar progresso
            progress_label = ttk.Label(progress_window, text="Processando...", font=('Helvetica', 10))
            progress_label.pack(pady=10)

            # Barra de progresso
            progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
            progress_bar.pack(pady=10)

            # Lista para mostrar arquivos processados
            lista_processados = tk.Listbox(progress_window, width=50, height=10)
            lista_processados.pack(pady=10, padx=10)

            # Configurar barra de progresso
            total_arquivos = len(arquivos)
            progress_bar['maximum'] = total_arquivos
            
            # Processar cada arquivo
            for i, arquivo in enumerate(arquivos, 1):
                try:
                    arquivo_nome = os.path.basename(arquivo)
                    progress_label.config(text=f"Processando: {arquivo_nome}")
                    progress_bar['value'] = i
                    
                    wb = load_workbook(arquivo, data_only=True)
                    try:
                        ws_resumo = wb['RESUMO']
                        nome_cliente = ws_resumo['A3'].value
                        logger.debug(f"Cliente: {nome_cliente}")
                        
                        data_rel = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
                        
                        # Carregar dados usando handler
                        df = self.handler.carregar_dados_excel(arquivo)
                        df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2 = self.handler.processar_dados(df, data_rel)
                        
                        # Processar lançamentos futuros
                        df_futuro = None
                        if self.incluir_futuros.get():
                            df_futuro = self.handler.processar_lancamentos_futuros(df, data_rel)
                        
                        # Obter número do relatório e valor acumulado
                        numero_relatorio = self.handler.obter_numero_relatorio(ws_resumo, data_rel)
                        valor_acumulado = self.handler.calcular_acumulado_dados(df, data_rel)
                        
                        logger.info(f"Arquivo: {arquivo_nome}")
                        logger.info(f"Número do relatório: {numero_relatorio}")
                        logger.info(f"Valor acumulado calculado: {valor_acumulado:,.2f}")
                        
                        dados_completos = {
                            'df_filtrado': df_filtrado,
                            'df_diaria': df_diaria,
                            'df_tp_desp_1': df_tp_desp_1,
                            'df_tp_desp_2': df_tp_desp_2,
                            'df_futuro': df_futuro,
                            'df_original': df,
                            'incluir_futuros': self.incluir_futuros.get(),
                            'data_relatorio': data_rel,
                            'nome_cliente': nome_cliente,
                            'endereco_cliente': ws_resumo['A4'].value,
                            'numero_relatorio': numero_relatorio,
                            'acumulado': valor_acumulado
                        }
                        
                        # Gerar nome do arquivo
                        data_formatada = data_rel.strftime('%d-%m-%Y')
                        nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
                        caminho_output = os.path.join(os.path.dirname(arquivo), nome_arquivo)
                        
                        # Gerar relatório
                        self.handler.gerar_relatorio_pdf(dados_completos, caminho_output, arquivo)
                        
                        lista_processados.insert(tk.END, f"✓ {arquivo_nome} - Concluído")
                        lista_processados.see(tk.END)
                        
                    finally:
                        wb.close()

                except Exception as e:
                    logger.error(f"Erro ao processar arquivo {arquivo_nome}: {str(e)}", exc_info=True)
                    lista_processados.insert(tk.END, f"✗ {arquivo_nome} - Erro: {str(e)}")

                # Atualizar interface
                progress_window.update()

            # Finalização
            progress_label.config(text="Processamento concluído!")
            ttk.Button(
                progress_window, 
                text="Fechar", 
                command=lambda: self.criar_dialog_relatorio_gerado(None, None) or progress_window.destroy()
            ).pack(pady=10)
            
        except Exception as e:
            logger.error(f"Erro no processamento em lote: {str(e)}", exc_info=True)
            raise


        
    def gerar_relatorio_lote():
        try:
            # Verificar se há arquivos selecionados
            if not self.arquivo_path:  # Usar self em vez de variável global
                self.status_label.config(text="Selecione um arquivo Excel!")
                return
            
            processar_lote(arquivos_selecionados)


            status_label.config(text="Relatórios em lote gerados com sucesso!")

            # Criar diálogo após gerar os relatórios em lote
            # criar_dialog_relatorio_gerado(None, None)

        except Exception as e:
            erro = str(e)
            print(f"Erro ao gerar relatórios em lote: {erro}")
            status_label.config(text=f"Erro: {erro}")


    def criar_dialog_relatorio_gerado(self, nome_cliente, data_formatada):
        dialog = Toplevel(self.root)
        dialog.title("Relatório Gerado")
        dialog.geometry("300x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        msg = f"Relatório individual gerado com sucesso para:\n{nome_cliente}\nData: {data_formatada}" if nome_cliente else "Relatórios em lote gerados com sucesso!"
        
        ttk.Label(dialog, text=msg, font=('Helvetica', 10, 'bold')).pack(pady=10)
        
        def continuar():
            dialog.destroy()
            
        def voltar_menu():
            dialog.destroy()
            self.root.destroy()
            if self.menu_principal:
                self.menu_principal.deiconify()
                self.menu_principal.lift()
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill='x', pady=10)
        
        ttk.Button(btn_frame, text="Gerar Outro Relatório", 
                  command=continuar).pack(pady=5, padx=10, fill='x')
        ttk.Button(btn_frame, text="Voltar ao Menu Principal", 
                  command=voltar_menu).pack(pady=5, padx=10, fill='x')          

    def adicionar_botao_pendentes(self):
        """
        Adiciona botão para gerar relatório de lançamentos pendentes
        """
        frame_pendentes = ttk.LabelFrame(self.root, text="Relatório de Lançamentos Pendentes")
        frame_pendentes.pack(pady=10, padx=20, fill='x')
        
        def selecionar_pasta():
            try:
                # Obter a data selecionada
                data_ref = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
                print(f"\nData de referência selecionada: {data_ref}")
                
                # Selecionar pasta
                pasta = filedialog.askdirectory(
                    title="Selecione a pasta com os arquivos dos clientes"
                )
                
                if pasta:
                    print(f"Pasta selecionada: {pasta}")
                    arquivo_saida = os.path.join(pasta, "relatorio_lancamentos_pendentes.html")
                    
                    # Criar instância do relatório
                    relatorio = RelatorioLancamentosPendentes()
                    
                    # Gerar relatório passando a data de referência
                    if relatorio.gerar_relatorio_pendentes(pasta, arquivo_saida, data_ref):
                        messagebox.showinfo(
                            "Sucesso",
                            f"Relatório gerado com sucesso!\nSalvo em: {arquivo_saida}"
                        )
                    else:
                        messagebox.showwarning(
                            "Aviso",
                            "Nenhum lançamento pendente encontrado para o período especificado."
                        )
                        
            except Exception as e:
                print(f"Erro ao gerar relatório: {str(e)}")
                messagebox.showerror(
                    "Erro",
                    "Erro ao gerar relatório. Verifique o console para mais detalhes."
                )
        
        # Adicionar botão
        ttk.Button(
            frame_pendentes,
            text="Gerar Relatório de Lançamentos Pendentes",
            command=selecionar_pasta
        ).pack(pady=5, fill='x')  

        

class RelatorioConfig:
    """Classe para gerenciar configurações e estilos do relatório"""
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
        
    def setup_custom_styles(self):
        """Configura os estilos personalizados para o relatório"""
        self.style_heading = ParagraphStyle(
            'HeadingStyle',
            parent=self.styles['Heading1'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            leftIndent=0,
            textColor=colors.black,
            spaceBefore=20,
            spaceAfter=12
        )
        
        self.style_normal = ParagraphStyle(
            'NormalStyle',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=12,
            textColor=colors.black,
            spaceBefore=12,
            spaceAfter=6
        )
        
        self.style_despesa = ParagraphStyle(
            name='TipoDespesa',
            parent=self.styles['Normal'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            leftIndent=0,
            firstLineIndent=0,
            rightIndent=0,
            spaceBefore=12,
            spaceAfter=6,
            keepWithNext=True
        )




def resource_path(relative_path):
    """Obtém o caminho absoluto para recursos empacotados"""
    try:
        # PyInstaller cria um temp folder e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



class IndentedFlowable:
    """Classe para ajudar na indentação de elementos como tabelas"""
    def __init__(self, flowable, leftIndent=0, rightIndent=0):
        self.flowable = flowable
        self.leftIndent = leftIndent
        self.rightIndent = rightIndent
        self.width = 0
        self.height = 0
        self.canv = None

    def wrap(self, availWidth, availHeight):
        """Define o tamanho do elemento"""
        self.width, self.height = self.flowable.wrap(
            availWidth - self.leftIndent - self.rightIndent, 
            availHeight
        )
        return (self.width + self.leftIndent + self.rightIndent, 
                self.height)

    def draw(self):
        """Desenha o elemento na posição correta"""
        self.flowable.drawOn(
            self.canv,
            self.canv._x + self.leftIndent,
            self.canv._y
        )

    def split(self, availWidth, availHeight):
        """Divide o elemento se necessário"""
        # Ajusta a largura disponível para a indentação
        availWidth = availWidth - self.leftIndent - self.rightIndent
        flowables = self.flowable.split(availWidth, availHeight)
        return [IndentedFlowable(f, self.leftIndent, self.rightIndent) for f in flowables]

    # Métodos de espaçamento
    def getSpaceBefore(self):
        return getattr(self.flowable, 'spaceBefore', 0)

    def getSpaceAfter(self):
        return getattr(self.flowable, 'spaceAfter', 0)

    def setSpaceBefore(self, space):
        self.flowable.spaceBefore = space

    def setSpaceAfter(self, space):
        self.flowable.spaceAfter = space

    # Propriedades de espaçamento
    spaceBefore = property(getSpaceBefore, setSpaceBefore)
    spaceAfter = property(getSpaceAfter, setSpaceAfter)

    # Métodos de controle de quebra de página
    def getKeepWithNext(self):
        return getattr(self.flowable, 'keepWithNext', 0)

    def setKeepWithNext(self, value):
        self.flowable.keepWithNext = value

    keepWithNext = property(getKeepWithNext, setKeepWithNext)

    # Métodos adicionais que podem ser necessários
    def identity(self, maxLen=None):
        return "IndentedFlowable: " + self.flowable.identity(maxLen)

    def drawOn(self, canvas, x, y, _sW=0):
        self.canv = canvas
        canvas.saveState()
        self.flowable.drawOn(canvas, x + self.leftIndent, y, _sW)
        canvas.restoreState()

    # Delegação de outros atributos ao flowable interno
    def __getattr__(self, name):
        return getattr(self.flowable, name)




class RelatorioHandler:
    def __init__(self):
        self.config = RelatorioConfig()
        self.tipos_despesas = {
            1: "1) DESPESAS COM COLABORADORES",
            2: "2) TRANSF. PROGR. - MATERIAIS, LOCAÇÕES E PREST.SERVIÇOS",
            3: "3) BOLETOS - MATERIAIS, PREST. SERVIÇOS, IMPOSTOS, ETC.",
            4: "4) RESSARCIMENTOS E RESTITUIÇÕES",
            5: "5) DESPESAS PAGAS PELO CLIENTE",
            6: "6) PAGAMENTOS CAIXA DE OBRA",
            7: "7) ADMINISTRAÇÃO DA OBRA"
        }

        # Verificar se a logomarca existe na mesma pasta do script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.logo_path = os.path.join(script_dir, "logo1.png")
        if not os.path.exists(self.logo_path):
            self.logo_path = None
            print("Aviso: Logomarca não encontrada na pasta do script.")
        
        self.tipos_despesas_futuras = {
            "Próximos 30 dias": lambda x: x <= self.data_ref + pd.Timedelta(days=30),
            "31 a 60 dias": lambda x: (x > self.data_ref + pd.Timedelta(days=30)) & 
                                     (x <= self.data_ref + pd.Timedelta(days=60)),
            "Após 60 dias": lambda x: x > self.data_ref + pd.Timedelta(days=60)
        }
        self.data_ref = None


       
        
    def selecionar_arquivo(self):
        """Interface para seleção do arquivo Excel"""
        root = Tk()
        root.withdraw()
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        return arquivo

    def obter_numero_relatorio(self, ws_resumo, data_relatorio):
        """
        Método para obter o número do relatório baseado na data.
        """
        try:
            logger.info(f"\nObtendo número do relatório para data {data_relatorio}")
            
            # Converter data para datetime
            data_ref = pd.to_datetime(data_relatorio).date()
            logger.debug(f"Data de referência processada: {data_ref}")
            
            # Encontrar primeira data na planilha
            primeira_data = None
            primeira_linha = None
            
            for row in range(9, 150):  # Buscar nas primeiras 150 linhas
                cell_value = ws_resumo.cell(row=row, column=1).value
                if isinstance(cell_value, (datetime, date)):
                    primeira_data = cell_value.date() if isinstance(cell_value, datetime) else cell_value
                    primeira_linha = row
                    break
                    
            if not primeira_data:
                logger.warning("Nenhuma data encontrada na planilha")
                return 1
                
            logger.debug(f"Primeira data encontrada: {primeira_data} na linha {primeira_linha}")
            
            # Usar a primeira data encontrada como data inicial
            data_inicial = primeira_data
            logger.debug(f"Data inicial: {data_inicial}")
            
            # Calcular número do relatório
            numero = 1
            data_atual = data_inicial
            
            while data_atual <= data_ref:
                logger.debug(f"Verificando data: {data_atual}")
                
                if data_atual == data_ref:
                    logger.info(f"Número do relatório calculado: {numero}")
                    return numero
                    
                # Avançar para próxima data (5 ou 20 do mês)
                if data_atual.day == 5:
                    data_atual = data_atual.replace(day=20)
                else:  # day == 20
                    if data_atual.month == 12:
                        data_atual = data_atual.replace(year=data_atual.year + 1, month=1, day=5)
                    else:
                        data_atual = data_atual.replace(month=data_atual.month + 1, day=5)
                numero += 1
                
            logger.warning(f"Data {data_ref} não encontrada na sequência. Último número calculado: {numero}")
            return numero
            
        except Exception as e:
            logger.error(f"Erro ao obter número do relatório: {str(e)}", exc_info=True)
            return 1

    def calcular_acumulado_dados(self, df, data_relatorio):
        """
        Calcula o valor acumulado somando todos os valores da aba 'Dados' 
        com DATA_REL anterior à data do relatório.
        """
        try:
            logger.info(f"Calculando acumulado para data {data_relatorio}")
            
            # Criar cópia do DataFrame e garantir que estamos usando a coluna correta
            df = df.copy()
            if 'VALOR' not in df.columns:
                logger.error("Coluna 'VALOR' não encontrada no DataFrame")
                return 0.0
                
            # Garantir que data_relatorio seja datetime
            if isinstance(data_relatorio, str):
                data_relatorio = pd.to_datetime(data_relatorio)
            elif isinstance(data_relatorio, date):
                data_relatorio = pd.to_datetime(data_relatorio)
                
            logger.debug(f"Data de referência processada: {data_relatorio}")
            
            # Converter DATA_REL para datetime
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'], errors='coerce')
            df = df.dropna(subset=['DATA_REL'])
            
            # Filtrar registros anteriores à data do relatório
            df_anterior = df[df['DATA_REL'] < data_relatorio].copy()
            
            if df_anterior.empty:
                logger.warning("Nenhum registro anterior encontrado")
                return 0.0
                
            # Converter valores para numérico
            logger.debug("Processando valores...")
            df_anterior['VALOR_NUMERICO'] = df_anterior['VALOR'].replace({',': '.', 'R$': '', ' ': ''}, regex=True)
            df_anterior['VALOR_NUMERICO'] = pd.to_numeric(df_anterior['VALOR_NUMERICO'], errors='coerce').fillna(0)
            
            # Log para debug
            logger.debug("\nPrimeiros 5 valores processados:")
            for idx, row in df_anterior.head().iterrows():
                logger.debug(f"Linha {idx}:")
                logger.debug(f"  Data: {row['DATA_REL']}")
                logger.debug(f"  Valor original: {row['VALOR']}")
                logger.debug(f"  Valor convertido: {row['VALOR_NUMERICO']}")
                
            # Calcular soma
            valor_acumulado = float(df_anterior['VALOR_NUMERICO'].sum())
            
            logger.info(f"Valor acumulado calculado: {valor_acumulado:,.2f}")
            logger.debug(f"Total de registros considerados: {len(df_anterior)}")
            
            return valor_acumulado
                
        except Exception as e:
            logger.error(f"Erro ao calcular acumulado: {str(e)}", exc_info=True)
            return 0.0


##  Esse método foi substituído por CALCULAR ACUMULADO DADOS EM 21/02/2025 ""
    # def calcular_acumulado_dados(self, df, data_relatorio):
    #     """
    #     Calcula o valor acumulado somando todos os valores da aba 'Dados' 
    #     com DATA_REL anterior à data do relatório.
        
    #     Parameters:
    #     -----------
    #     df : pandas.DataFrame
    #         DataFrame com os dados da aba 'Dados'
    #     data_relatorio : datetime
    #         Data do relatório sendo gerado
            
    #     Returns:
    #     --------
    #     float
    #         Valor acumulado total
    #     """
    #     try:
    #         # Garantir que data_relatorio seja datetime
    #         logger.info("Iniciando cálculo de valor acumulado")
    #         data_rel = pd.to_datetime(data_relatorio)
            
    #         # Converter DATA_REL para datetime se necessário
    #         df['DATA_REL'] = pd.to_datetime(df['DATA_REL'])
                    
    #         # Filtrar registros anteriores à data do relatório
    #         df_anterior = df[df['DATA_REL'] < data_rel].copy()

    #         logger.debug(f"Registros encontrados para cálculo: {len(df_anterior)}")
            
    #         # Converter VALOR para numérico de forma mais eficiente
    #         df_anterior['VALOR'] = pd.to_numeric(
    #             df_anterior['VALOR'].replace(r'[R$, ]', '', regex=True), 
    #             errors='coerce'
    #         ).fillna(0)
            
    #         # Calcular soma total
    #         valor_acumulado = float(df_anterior['VALOR'].sum())
    #         logger.info(f"Valor acumulado calculado: {valor_acumulado:,.2f}")
    #         return valor_acumulado
            
    #     except Exception as e:
    #         logger.error(f"Erro ao calcular acumulado: {str(e)}", exc_info=True)
    #         return 0.0


##  Esse método foi substituído por CALCULAR ACUMULADO DADOS EM 11/02/2025 ""

##    def obter_acumulado(self, ws_resumo, data_relatorio):
##        
##        Obtém o valor da coluna ACUMULADO do relatório anterior
##        
##        print("\n=== INÍCIO OBTER ACUMULADO ===")
##        
##        try:
##            # Processar data_relatorio
##            data_ref = pd.to_datetime(data_relatorio).date()
##            print(f"Data de referência processada: {data_ref}")
##            
##            # Encontrar a linha do relatório atual
##            linha_atual = None
##            for row in range(9, ws_resumo.max_row + 1):
##                data_cell = ws_resumo.cell(row=row, column=1).value
##                if isinstance(data_cell, datetime):
##                    if data_cell.date() == data_ref:
##                        linha_atual = row
##                        break
##            
##            if linha_atual and linha_atual > 9:
##                # Pegar valor da coluna L (ACUMULADO) da linha anterior
##                valor_anterior = ws_resumo.cell(row=linha_atual-1, column=12).value
##                
##                # Converter para float se necessário
##                if isinstance(valor_anterior, str):
##                    valor_anterior = float(valor_anterior.replace('.', '').replace(',', '.'))
##                
##                print(f"Encontrado relatório anterior, acumulado: {valor_anterior}")
##                return float(valor_anterior or 0)
##                
##            print("Nenhum relatório anterior encontrado")
##            return 0.0
##            
##        except Exception as e:
##            print(f"Erro ao obter acumulado: {str(e)}")
##            return 0.0
##        finally:
##            print("=== FIM OBTER ACUMULADO ===\n")
 
    
    def carregar_dados_excel(self, arquivo_excel):
        try:
            df = pd.read_excel(arquivo_excel, sheet_name='Dados')
            df = df.fillna("")
            
            # Verificar colunas necessárias
            colunas_necessarias = {'DATA_REL', 'TP_DESP', 'REFERÊNCIA', 'DT_VENCTO', 'VALOR', 'NF'}
            if not colunas_necessarias.issubset(df.columns):
                raise ValueError(f"Colunas necessárias ausentes: {colunas_necessarias - set(df.columns)}")
            
            # Converter NF para string antes de processar
            df['NF'] = df['NF'].astype(str)
            
            # Concatenar NF com REFERÊNCIA apenas para TP_DESP != 1
            mascara = (df['TP_DESP'] != 1) & (df['NF'].notna()) & (df['NF'].str.strip() != '') & (df['NF'] != 'nan')
            df.loc[mascara, 'REFERÊNCIA'] = df[mascara].apply(
                lambda row: f"{row['REFERÊNCIA']} (NF: {row['NF'].strip()})", 
                axis=1
            )
            
            
            return df
            
        except Exception as e:
            raise Exception(f"Erro ao carregar arquivo Excel: {str(e)}")
        

        
            
    def processar_dados(self, df, data_relatorio):
        """Processa os dados conforme os critérios especificados"""
        # Converter data para datetime usando formato explícito
        try:
            data_rel = pd.to_datetime(data_relatorio)
        except:
            # Se falhar, tenta converter assumindo formato brasileiro
            data_rel = pd.to_datetime(data_relatorio, format='%d/%m/%Y')
        
        # Criar cópia do DataFrame para não modificar o original
        df = df.copy()
        
        # Formatar as datas no DataFrame usando formato brasileiro
        if 'DT_VENCTO' in df.columns:
            df['DT_VENCTO'] = pd.to_datetime(df['DT_VENCTO'], dayfirst=True)  # Forçar interpretação dia/mês
            df['DT_VENCTO'] = df['DT_VENCTO'].dt.strftime('%d/%m/%Y')
        
        # Filtrar dados
        df_filtrado = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] != 1)
        ].sort_values(
            by=['TP_DESP', 'DT_VENCTO', 'VALOR'], 
            ascending=[True, True, False]  # True para ordenar vencimento do mais antigo
        )
        
        df_diaria = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'] == 'DIÁRIA')
        ].sort_values(
            by=['TP_DESP', 'DT_VENCTO', 'VALOR'], 
            ascending=[True, False, False]
        )
        
        df_tp_desp_1 = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'].isin(['SALÁRIO', 'TRANSPORTE', 'CAFÉ']))
        ]

        df_tp_desp_2 = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'].isin(['FÉRIAS', 'RESCISÃO', '13º SALÁRIO']))
        ]
        
        
        return df_filtrado, df_diaria, df_tp_desp_1, df_tp_desp_2

    def processar_lancamentos_futuros(self, df, data_relatorio):
        """Processa os lançamentos futuros do DataFrame usando DATA_REL"""
        # Converter a data do relatório para datetime usando formato explícito
        try:
            self.data_ref = pd.to_datetime(data_relatorio)
        except:
            # Se falhar, tenta converter assumindo formato brasileiro
            self.data_ref = pd.to_datetime(data_relatorio, format='%d/%m/%Y')

        # Converter a coluna DATA_REL para datetime
        df = df.copy()
        df['DATA_REL'] = pd.to_datetime(df['DATA_REL'])
        df['DT_VENCTO'] = pd.to_datetime(df['DT_VENCTO'], format='%d/%m/%Y', errors='coerce')
        
        # Formatar a data de vencimento para DD/MM/AAAA
        df['DT_VENCTO'] = df['DT_VENCTO'].dt.strftime('%d/%m/%Y')

        # Filtrar apenas lançamentos futuros baseado em DATA_REL
        df_futuro = df[(df['DATA_REL'] > self.data_ref) & (df['TP_DESP'] != 1)].copy()

        # Ordenar por data de vencimento
        df_futuro = df_futuro.sort_values('DT_VENCTO')

        # Agrupar por período baseado na DATA_REL
        df_futuro['periodo'] = df_futuro['DATA_REL'].apply(
            lambda x: next(
                (nome for nome, func in self.tipos_despesas_futuras.items() 
                 if func(x)),
                "Após 60 dias"
            )
        )

        return df_futuro
    
    def adicionar_lancamentos_futuros(self, elementos, dados):
        """Adiciona a seção de lançamentos futuros ao relatório"""
        if not dados['df_futuro'].empty:
            elementos.append(PageBreak())
            elementos.append(Paragraph("LANÇAMENTOS FUTUROS", self.config.style_heading))
            
            total_geral_futuro = 0
            
            # Agrupar por período e tipo de despesa
            for periodo in ["Próximos 30 dias", "31 a 60 dias", "Após 60 dias"]:
                df_periodo = dados['df_futuro'][dados['df_futuro']['periodo'] == periodo]
                
                if not df_periodo.empty:
                    # Adicionar título do período com estilo destacado
                    elementos.append(Paragraph(
                        f"\n{periodo}",
                        ParagraphStyle(
                            'PeriodoStyle',
                            parent=self.config.style_heading,
                            fontSize=14,
                            leading=16,
                            spaceBefore=12,
                            spaceAfter=6,
                            textColor=colors.HexColor('#2F4F4F')  # Cor mais escura para destaque
                        )
                    ))
                    
                    total_periodo = 0
                    
                    # Agrupar por tipo de despesa dentro do período
                    for tipo in sorted(df_periodo['TP_DESP'].unique()):
                        df_tipo = df_periodo[df_periodo['TP_DESP'] == tipo]
                        if not df_tipo.empty:
                            elementos.append(Paragraph(
                                self.tipos_despesas.get(tipo, f"Tipo {tipo}"),
                                self.config.style_normal
                            ))
                            
                            # Renomear colunas para corresponder ao formato esperado
                            df_tipo = df_tipo.rename(columns={
                                'DT_VENCTO': 'VENCIMENTO',
                                'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
                            })
                            
                            tabela = self.criar_tabela_despesas(
                                df_tipo,
                                ['NOME', 'VENCIMENTO', 'REFERÊNCIA', 'VALOR', 'DADOS BANCÁRIOS'],
                                [240, 70, 220, 80, 170]
                            )
                            elementos.append(tabela)
                            elementos.append(Spacer(1, 12))
                            
                            total_periodo += df_tipo['VALOR'].sum()
                    
                    # Adicionar subtotal do período
                    elementos.append(Paragraph(
                        f"Subtotal {periodo}: {self.formatar_numero(total_periodo)}",
                        ParagraphStyle(
                            'SubtotalStyle',
                            parent=self.config.style_normal,
                            fontSize=10,
                            leading=12,
                            spaceBefore=6,
                            spaceAfter=12,
                            textColor=colors.HexColor('#4A4A4A')
                        )
                    ))
                    
                    total_geral_futuro += total_periodo
            
            # Adicionar total geral dos lançamentos futuros
            elementos.append(Paragraph(
                f"\nTotal Geral de Lançamentos Futuros: {self.formatar_numero(total_geral_futuro)}",
                self.config.style_heading
            ))
    
    def formatar_numero(self, valor):
        """Formata valor numérico, tratando possíveis strings e NaN"""
        if pd.isna(valor) or valor == "":
            return "0,00"
        try:
            if isinstance(valor, str):
                # Remover caracteres não numéricos, mantendo ponto decimal
                valor = valor.replace('R$', '').replace(' ', '')
                # Se tiver vírgula como separador decimal, converter para ponto
                if ',' in valor and '.' not in valor:
                    valor = valor.replace(',', '.')
                # Se tiver tanto vírgula quanto ponto, assume que a vírgula é separador de milhar
                elif ',' in valor and '.' in valor:
                    valor = valor.replace('.', '').replace(',', '.')
                
                valor = float(valor)
            
            # Formatar com separador de milhar e vírgula como separador decimal
            return f"{float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except Exception as e:
            logger.error(f"Erro ao formatar número: {str(e)} - Valor: {valor}")
            return "0,00"  # Retorna zero formatado em caso de erro

    def formatar_data(self, data):
        """Formata data para o padrão brasileiro"""
        if pd.isna(data):
            return ''
        try:
            return pd.to_datetime(data).strftime('%d/%m/%Y')
        except:
            return str(data)

    def consolidar_despesas_colaboradores(self, df):
        """Consolida as despesas dos colaboradores"""
        try:
            # Criar cópia e tratar valores nulos
            df = df.copy()
            df = df.fillna("")
            df = df.infer_objects(copy=False)
                
            agregacoes = {
                'SALÁRIO': ['SALÁRIO'],
                'TRANSPORTE': ['TRANSPORTE'],
                'CAFÉ': ['CAFÉ']
            }
                
            if 'DADOS_BANCARIOS' in df.columns:
                df = df.rename(columns={'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'})
                
            resultados = []
            total_colunas = {
                'SALÁRIO': 0.0,
                'TRANSPORTE': 0.0,
                'CAFÉ': 0.0
            }
                
            for nome, grupo in df.groupby('NOME'):
                linha = {'NOME': nome}
                    
                for coluna, referencias in agregacoes.items():
                    valores_grupo = grupo[grupo['REFERÊNCIA'].isin(referencias)]['VALOR']
                    valor = pd.to_numeric(valores_grupo, errors='coerce').sum()
                    linha[coluna] = valor if not pd.isna(valor) else 0.0
                    total_colunas[coluna] += linha[coluna]
                        
                # Pegar DIAS do lançamento de TRANSPORTE ou CAFÉ (o que for maior)
                dias_transporte = grupo[grupo['REFERÊNCIA'] == 'TRANSPORTE']['DIAS'].iloc[0] if not grupo[grupo['REFERÊNCIA'] == 'TRANSPORTE'].empty else 0
                dias_cafe = grupo[grupo['REFERÊNCIA'] == 'CAFÉ']['DIAS'].iloc[0] if not grupo[grupo['REFERÊNCIA'] == 'CAFÉ'].empty else 0
                    
                # Converter para inteiro e pegar o maior valor
                dias_transporte = int(dias_transporte) if pd.notnull(dias_transporte) else 0
                dias_cafe = int(dias_cafe) if pd.notnull(dias_cafe) else 0
                linha['DIAS'] = max(dias_transporte, dias_cafe)
                    
                linha['DADOS BANCÁRIOS'] = grupo['DADOS BANCÁRIOS'].iloc[0] if not grupo['DADOS BANCÁRIOS'].empty else ''
                linha['TOTAL'] = sum(linha.get(col, 0) for col in total_colunas.keys())
                    
                resultados.append(linha)
                
            # Criar DataFrame com os resultados
            df_result = pd.DataFrame(resultados)
                
            # Definir ordem das colunas
            colunas_ordem = ['NOME', 'SALÁRIO', 'DIAS', 
                            'TRANSPORTE', 'CAFÉ', 'TOTAL', 'DADOS BANCÁRIOS']
                
            return df_result.reindex(columns=colunas_ordem)
                
        except Exception as e:
            print(f"Erro ao consolidar despesas: {str(e)}")
            raise

    def consolidar_despesas_colaboradores1(self, df):
        """Consolida as despesas  13º, férias e rescisão dos colaboradores"""
        try:
            # Criar cópia e tratar valores nulos
            df = df.copy()
            df = df.fillna("")
            df = df.infer_objects(copy=False)
                
            agregacoes1 = {
                '13º SALÁRIO': ['13º SALÁRIO'],
                'FÉRIAS': ['FÉRIAS'],
                'RESCISÃO': ['RESCISÃO']
            }
                
            if 'DADOS_BANCARIOS' in df.columns:
                df = df.rename(columns={'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'})
                
            resultados1 = []
            total_colunas = {
                '13º SALÁRIO': 0.0,
                'FÉRIAS': 0.0,
                'RESCISÃO': 0.0
            }
                
            for nome, grupo in df.groupby('NOME'):
                linha = {'NOME': nome}
                    
                for coluna, referencias in agregacoes1.items():
                    valores_grupo = grupo[grupo['REFERÊNCIA'].isin(referencias)]['VALOR']
                    valor = pd.to_numeric(valores_grupo, errors='coerce').sum()
                    linha[coluna] = valor if not pd.isna(valor) else 0.0
                    total_colunas[coluna] += linha[coluna]
                        
                    
                linha['DADOS BANCÁRIOS'] = grupo['DADOS BANCÁRIOS'].iloc[0] if not grupo['DADOS BANCÁRIOS'].empty else ''
                linha['TOTAL'] = sum(linha.get(col, 0) for col in total_colunas.keys())
                    
                resultados1.append(linha)
                
            # Criar DataFrame com os resultados
            df_result1 = pd.DataFrame(resultados1)
                
            # Definir ordem das colunas
            colunas_ordem = ['NOME', '13º SALÁRIO', 'FÉRIAS', 
                            'RESCISÃO', 'TOTAL', 'DADOS BANCÁRIOS']
                
            return df_result1.reindex(columns=colunas_ordem)
                
        except Exception as e:
            print(f"Erro ao consolidar despesas: {str(e)}")
            raise

    def criar_tabela_despesas(self, dados, colunas, larguras, incluir_total=True):
        """Cria uma tabela formatada para o relatório"""
        dados_formatados = dados.copy()
        dados_formatados = dados_formatados.fillna("")
        dados_formatados = dados_formatados.infer_objects()

        # Estilo para o cabeçalho com quebra de linha
        estilo_cabecalho = ParagraphStyle(
            'CabecalhoTabela',
            parent=self.config.style_normal,
            fontSize=8,
            leading=10,
            alignment=1,
            textColor=colors.whitesmoke
        )

        # Estilo para células com quebra de texto
        estilo_celula = ParagraphStyle(
            'CelulaTabela',
            parent=self.config.style_normal,
            fontSize=8,
            leading=10,
            alignment=0  # Alinhamento à esquerda
        )

        # Converter cabeçalhos simples em Paragraphs com quebras de linha
        cabecalhos_formatados = []
        for coluna in colunas:
            if '/' in coluna:
                texto_formatado = Paragraph(coluna.replace('/', '<br/>'), estilo_cabecalho)
            elif ' - ' in coluna:
                texto_formatado = Paragraph(coluna.replace(' - ', '<br/>'), estilo_cabecalho)
            else:
                texto_formatado = Paragraph(coluna, estilo_cabecalho)
            cabecalhos_formatados.append(texto_formatado)

        colunas_numericas = ['VALOR', 'TOTAL', 'SALÁRIO', 'RESCISÃO', '13º SALÁRIO', 
                            'TRANSPORTE', 'CAFÉ', 'FÉRIAS', 'DIÁRIA', 'DIAS']

        # Processar dados linha por linha
        dados_tabela = [cabecalhos_formatados]
        for _, linha in dados_formatados.iterrows():
            linha_formatada = []
            for i, coluna in enumerate(colunas):
                valor = linha[coluna]
                
                # Formatar números
                if coluna in colunas_numericas:
                    valor = pd.to_numeric(valor, errors='coerce')
                    valor = 0 if pd.isna(valor) else valor
                    if coluna == 'DIAS':
                        valor = str(int(valor))  # Converter para inteiro e depois string
                    else:
                        valor = self.formatar_numero(valor)
                    linha_formatada.append(valor)
                
                # Formatar datas
                elif coluna in ['DT_VENCTO', 'VENCIMENTO']:
                    try:
                        valor = pd.to_datetime(valor, dayfirst=True).strftime('%d/%m/%Y')
                    except:
                        valor = str(valor)
                    linha_formatada.append(valor)
                
                # Adicionar quebra de texto para a coluna Referência
                elif coluna == 'REFERÊNCIA':
                    valor = str(valor)
                    linha_formatada.append(Paragraph(valor, estilo_celula))
                
                # Tratar coluna NF
                elif coluna == 'NF':
                    valor = str(valor) if valor else ""
                    linha_formatada.append(valor)
                

                # Outras colunas
                else:
                    linha_formatada.append(str(valor))
                    
            dados_tabela.append(linha_formatada)

        # Adicionar linha de total se necessário
        if incluir_total:
            coluna_valor = next((i for i, col in enumerate(colunas) 
                            if col in ['VALOR', 'TOTAL']), -1)
            if coluna_valor >= 0:
                # Se for a tabela de colaboradores (verificar pelas colunas características)
                if 'SALÁRIO' in colunas and 'TRANSPORTE' in colunas or '13º SALÁRIO' in colunas and 'FÉRIAS' in colunas:
                    linha_total = [''] * len(colunas)
                    linha_total[0] = 'Subtotal'
                    
                    # Calcular total para cada coluna numérica
                    for i, col in enumerate(colunas):
                        if col in ['SALÁRIO', 'FÉRIAS', 'RESCISÃO', '13º SALÁRIO', 'TRANSPORTE', 'CAFÉ', 'TOTAL']:
                            total = dados[col].sum()
                            linha_total[i] = self.formatar_numero(total)
                        elif col == 'DIAS':
                            linha_total[i] = ''  # Deixar DIAS vazio
                            
                    dados_tabela.append(linha_total)
                
                else:
                    # Para outras tabelas, manter o comportamento original
                    total = dados[colunas[coluna_valor]].sum()
                    linha_total = [''] * len(colunas)
                    linha_total[coluna_valor-1] = 'Subtotal'
                    linha_total[coluna_valor] = self.formatar_numero(total)
                    dados_tabela.append(linha_total)

        # Criar tabela com os dados formatados
        tabela = Table(dados_tabela, colWidths=larguras, repeatRows=1)
        
        # Definir estilos da tabela
        estilo_tabela = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        # Alinhar colunas numéricas à direita
        for i, col in enumerate(colunas):
            if col in colunas_numericas:
                estilo_tabela.append(('ALIGN', (i, 1), (i, -1), 'RIGHT'))

        if incluir_total:
            estilo_tabela.extend([
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ])

        tabela.setStyle(TableStyle(estilo_tabela))
        return tabela

    def criar_resumo_despesas(self, dados):
        """Cria o resumo das despesas para o relatório"""
        logger.debug("\nIniciando criar_resumo_despesas")
        logger.debug(f"Dados recebidos - acumulado: {dados.get('acumulado')}")
        
        subtotais = {}
        # Calcular subtotais por tipo de despesa
        for tipo, descricao in self.tipos_despesas.items():
            valor = 0
            if tipo == 1:
                # Somar todas as despesas de colaboradores (incluindo diárias, férias, rescisão, 13º)
                # O cálculo permanece o mesmo, pois continuamos somando todos os valores tp_desp=1,
                # apenas a apresentação visual é que mudou.
                valor = (dados['df_tp_desp_1']['VALOR'].sum() +
                    dados['df_tp_desp_2']['VALOR'].sum() +
                    dados['df_diaria']['VALOR'].sum())
            else:
                # Somar outras despesas
                df_tipo = dados['df_filtrado'][dados['df_filtrado']['TP_DESP'] == tipo]
                valor = df_tipo['VALOR'].sum()
                
            subtotais[tipo] = valor

        # Calcular despesas agrupadas
        despesas_a_pagar = sum(subtotais.get(tp, 0) for tp in [1, 2, 3, 4, 7])
        despesas_pagas_cliente = sum(subtotais.get(tp, 0) for tp in [5])
        despesas_pagas_caixa = sum(subtotais.get(tp, 0) for tp in [6])

        total_quinzena = sum(subtotais.values())
        
        # Garantir que temos os valores corretos
        acumulado = dados.get('acumulado', 0)
        numero_relatorio = dados.get('numero_relatorio', 1)
        
        logger.debug(f"Valores para cálculo:")
        logger.debug(f"- Total quinzena: {total_quinzena}")
        logger.debug(f"- Acumulado: {acumulado}")
        
        total_obra = total_quinzena + acumulado

        # Criar tabelas de resumo com formatação consistente
        tabela_subtotais = []
        for tipo, descricao in self.tipos_despesas.items():
            if tipo in subtotais:
                valor_formatado = self.formatar_numero(subtotais[tipo])
                tabela_subtotais.append([descricao, valor_formatado])

        tabela_totais = [
            ['DESPESAS A PAGAR', self.formatar_numero(despesas_a_pagar)],
            ['DESPESAS PAGAS PELO CLIENTE', self.formatar_numero(despesas_pagas_cliente)],
            ['COMPLEMENTO DE CAIXA', self.formatar_numero(despesas_pagas_caixa)],
            [''],
            ['TOTAL DA QUINZENA', self.formatar_numero(total_quinzena)],
            [f'TOTAL ACUMULADO RELATÓRIO Nº {numero_relatorio - 1}',
            self.formatar_numero(acumulado)],
            ['TOTAL DA OBRA', self.formatar_numero(total_obra)]
        ]
        
        logger.debug("Tabela totais criada:")
        for linha in tabela_totais:
            logger.debug(f"Linha: {linha}")

        return tabela_subtotais, tabela_totais

    def adicionar_cabecalho(self, elementos, dados):
        
        try:
            if not isinstance(elementos, list):
                print("ERRO: elementos não é uma lista!")
                elementos = []
                
            # Criar estilo customizado com espaçamento de 0
            style_cabecalho = ParagraphStyle(
                'CabecalhoStyle',
                parent=self.config.style_normal,
                alignment=2,
                spaceBefore=0,
                spaceAfter=0,
                leading=12
            )

            try:
##                print(f"Antes de verificar logo - self.logo_path: {self.logo_path}")
##                print(f"Caminho da logo existe? {os.path.exists(self.logo_path)}")
                
                if self.logo_path and os.path.exists(self.logo_path):
##                    print("Tentando criar Image")
                    logo = Image(self.logo_path, width=200, height=100)
##                    print("Image criada com sucesso")
                    
                    info_empresa = [
                        Paragraph("Rua Zodiaco, 87 Sala 07 – Santa Lúcia - Belo Horizonte - MG", style_cabecalho),
                        Paragraph("(31) 3654-6616 / (31) 99974-1241 / (31) 98711-1139", style_cabecalho),
                        Paragraph("rvr.engenharia@gmail.com", style_cabecalho)
                    ]
                    
##                    print("Criando tabela do cabeçalho")
                    cabecalho_table = Table(
                        [[logo, info_empresa]], 
                        colWidths=[80, 650],
                        rowHeights=[60]
                    )
                    
                    cabecalho_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                        ('VALIGN', (0, 0), (1, 0), 'TOP'),
                        ('RIGHTPADDING', (1, 0), (1, 0), 0),
                    ]))
                    
##                    print("Adicionando tabela aos elementos")
                    elementos.append(cabecalho_table)
##                    print("Tabela adicionada com sucesso")
                    
            except Exception as e:
                print(f"Erro ao processar logo: {str(e)}")

                
        except Exception as e:
            print(f"Aviso: Não foi possível adicionar a logo ao cabeçalho: {e}")
            # Continua sem a logo, apenas com as informações
            info_empresa = [
                Paragraph("Rua Zodiaco, 87 Sala 07 – Santa Lúcia - Belo Horizonte - MG", style_cabecalho),
                Paragraph("(31) 3654-6616 / (31) 99974-1241 / (31) 98711-1139", style_cabecalho),
                Paragraph("rvr.engenharia@gmail.com", style_cabecalho)
            ]
            elementos.extend(info_empresa)

        # Espaço após o cabeçalho
        elementos.append(Spacer(1, 40))
        
        # Criar tabela para nome/endereço do cliente e número/data do relatório
        data_formatada = pd.to_datetime(dados.get('data_relatorio')).strftime('%d/%m/%Y')
        info_cliente = [
            [Paragraph(dados.get('nome_cliente', ''), self.config.style_heading),
             Paragraph(f"Relatório nº: {dados.get('numero_relatorio', '')}", self.config.style_normal)],
            [Paragraph(dados.get('endereco_cliente', ''), self.config.style_normal),
             Paragraph(f"Data: {data_formatada}", self.config.style_normal)]
        ]

        cliente_table = Table(
            info_cliente,
            colWidths=[680, 100],  # Ajuste as larguras conforme necessário
            rowHeights=[20, 20]   
        )
        cliente_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Alinhar informações do cliente à esquerda
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),   # Alinhar número e data à direita
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elementos.append(cliente_table)

    
    def adicionar_detalhes(self, elementos, dados):
        """Adiciona os detalhes das despesas ao relatório"""
        logger.info("Iniciando adição de detalhes ao relatório")

        elementos.append(Paragraph("DETALHES DAS DESPESAS", self.config.style_heading))
    
        # 1. Despesas com Colaboradores - Funcionários (Salários, VT e VR)
        if not dados['df_tp_desp_1'].empty:
            logger.debug("Processando despesas com colaboradores - funcionários")
            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - SALÁRIO/ADIANTAMENTO, TRANSPORTE E CAFÉ", 
                                self.config.style_despesa))
            df_consolidado = self.consolidar_despesas_colaboradores(dados['df_tp_desp_1'])
            logger.debug(f"Total de funcionários processados: {len(df_consolidado)}")

            tabela = self.criar_tabela_despesas(
                df_consolidado,
                ['NOME', 'SALÁRIO', 'DIAS', 
                 'TRANSPORTE', 'CAFÉ', 'TOTAL', 'DADOS BANCÁRIOS'],
                [220, 80, 40, 70, 70, 80, 210]
            )
            elementos.append(tabela)
            elementos.append(Spacer(1, 12))

        # 2. Despesas com Colaboradores - Funcionários (13º, Férias e Rescisão)
        if not dados['df_tp_desp_2'].empty:
            logger.debug("Processando despesas com colaboradores - funcionários")
            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - 13º SALÁRIO, FÉRIAS E RESCISÃO", 
                                self.config.style_despesa))
            df_consolidado1 = self.consolidar_despesas_colaboradores1(dados['df_tp_desp_2'])
            logger.debug(f"Total de funcionários processados: {len(df_consolidado1)}")

            tabela = self.criar_tabela_despesas(
                df_consolidado1,
                ['NOME', '13º SALÁRIO', 'FÉRIAS', 'RESCISÃO',  
                 'TOTAL', 'DADOS BANCÁRIOS'],
                [240, 70, 70, 70, 70, 240]
            )
            elementos.append(tabela)
            elementos.append(Spacer(1, 12))
        
        # 3. Despesas com Colaboradores - Diaristas
        if not dados['df_diaria'].empty:
            logger.debug("Processando despesas com colaboradores - diaristas")

            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - DIÁRIAS", 
                                self.config.style_despesa))
            # Renomear colunas para corresponder ao formato esperado
            df_diaria_formatado = dados['df_diaria'].copy()
            df_diaria_formatado = df_diaria_formatado.rename(columns={
                'VR_UNIT': 'DIÁRIA',
                'VALOR': 'TOTAL',
                'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
            })
            tabela = self.criar_tabela_despesas(
                df_diaria_formatado,
                ['NOME', 'DIÁRIA', 'DIAS', 'TOTAL', 'DADOS BANCÁRIOS'],
                [284, 80, 50, 90, 280]
            )
            elementos.append(tabela)
            elementos.append(Spacer(1, 12))
    
        # 4. Outras despesas
        for tipo in range(2, 8):
            df_tipo = dados['df_filtrado'][dados['df_filtrado']['TP_DESP'] == tipo]
            if not df_tipo.empty:
                logger.debug(f"Processando despesas tipo {tipo}")
                elementos.append(Paragraph(self.tipos_despesas[tipo], 
                                    self.config.style_despesa))
                # Renomear colunas para corresponder ao formato esperado
                df_tipo = df_tipo.rename(columns={
                    'DT_VENCTO': 'VENCIMENTO',
                    'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
                })
                tabela = self.criar_tabela_despesas(
                    df_tipo,
                    ['NOME', 'VENCIMENTO', 'REFERÊNCIA', 'VALOR', 'DADOS BANCÁRIOS'],
                    [220, 70, 250, 80, 170]
                )
                elementos.append(tabela)
                elementos.append(Spacer(1, 16))
        logger.info("Detalhes adicionados com sucesso")


    def carregar_taxas_administracao(self, arquivo_excel):
        """
        Carrega e processa os dados de taxas de administração da aba Contratos_ADM,
        considerando a estrutura específica da planilha:
        - Linha 1: Títulos dos blocos
        - Linha 2: Subtítulos
        - Linha 3: Dados do contrato
        - Linha 4: Dados dos administradores
        - Linha 5: Início dos dados das parcelas
        """
        logger.info(f"Iniciando carregamento de taxas de administração: {arquivo_excel}")

        try:
            workbook = load_workbook(arquivo_excel, data_only=True)
            if 'Contratos_ADM' not in workbook.sheetnames:
                logger.warning("Aba 'Contratos_ADM' não encontrada no arquivo")
                return pd.DataFrame()

            ws_contratos = workbook['Contratos_ADM']
            logger.debug(f"Total de linhas na planilha: {ws_contratos.max_row}")
            
            # Colunas para dados das parcelas com mapeamento correto
            colunas_parcelas = {
                'Y': 'referencia',      # Número do contrato
                'Z': 'numero_parcela',  # Número da parcela
                'AA': 'cpf_cnpj',       # CNPJ/CPF do Administrador
                'AB': 'administrador',   # Nome do Administrador
                'AC': 'data_vencimento', # Data Vencimento
                'AD': 'valor_parcela',   # Valor da parcela
                'AE': 'status',         # Status (PENDENTE/PAGO)
                'AF': 'data_pagamento'  # Data Pagamento
            }
            
            dados = []
            linha_atual = 5  # Começar da linha 5
            
            while linha_atual <= ws_contratos.max_row:
                valor_coluna_y = ws_contratos[f'Y{linha_atual}'].value
                logger.debug(f"Processando linha {linha_atual} - Valor Y: {valor_coluna_y}")
                
                if valor_coluna_y:
                    try:
                        linha = {}
                        
                        # Processar cada coluna
                        for col, nome in colunas_parcelas.items():
                            valor = ws_contratos[f'{col}{linha_atual}'].value
                            print(f"Coluna {col} ({nome}): {valor}")
                            
                            # Tratamento específico para cada tipo de campo
                            if nome == 'valor_parcela':
                                try:
                                    if isinstance(valor, str):
                                        valor = float(valor.replace('R$', '').replace('.', '').replace(',', '.').strip())
                                    elif isinstance(valor, (int, float)):
                                        valor = float(valor)
                                    else:
                                        valor = 0.0
                                    print(f"Valor parcela convertido: {valor}")
                                except (ValueError, TypeError) as e:
                                    print(f"Erro ao converter valor_parcela: {e}")
                                    valor = 0.0
                                    
                            elif nome == 'data_vencimento' or nome == 'data_pagamento':
                                if isinstance(valor, datetime):
                                    valor = valor.date()
                                elif valor:
                                    try:
                                        valor = pd.to_datetime(valor).date()
                                    except:
                                        valor = None
                                print(f"Data convertida: {valor}")
                                        
                            elif nome == 'status':
                                valor = str(valor).upper() if valor else ''
                                print(f"Status convertido: {valor}")
                                
                            else:
                                # Outros campos mantêm o valor original
                                valor = str(valor) if valor is not None else ''
                                
                            linha[nome] = valor
                        
                        # Verificar apenas as validações necessárias
                        print("\nValidando dados:")
                        print(f"Tem referência: {bool(linha['referencia'])}")
                        print(f"Tem número da parcela: {bool(linha['numero_parcela'])}")
                        print(f"Valor parcela > 0: {linha['valor_parcela'] > 0}")
                        print(f"Data vencimento existe: {linha['data_vencimento'] is not None}")
                        print(f"Status é PENDENTE: {linha['status'] == 'PENDENTE'}")
                        
                        # Verificações simplificadas
                        if (linha['referencia'] and 
                            linha['numero_parcela'] and 
                            linha['valor_parcela'] > 0 and 
                            linha['data_vencimento'] is not None and
                            linha['status'] == 'PENDENTE'):
                            dados.append(linha)
                            print("Linha adicionada aos dados!")
                        else:
                            print("Linha não atendeu aos critérios de validação")
                            
                        linha_atual += 1
                            
                    except Exception as e:
                        print(f"Erro ao processar linha {linha_atual}: {str(e)}")
                        linha_atual += 1
                        
                else:
                    linha_atual += 1
            
            # Criar DataFrame apenas com dados válidos
            df = pd.DataFrame(dados) if dados else pd.DataFrame()
            
            logger.info(f"Total de parcelas encontradas: {len(dados)}")

            if dados:
                print("\nPrimeira parcela:")
                for k, v in dados[0].items():
                    print(f"{k}: {v}")
            
            return df
            
        except Exception as e:
            logger.error(f"Erro ao carregar taxas de administração: {str(e)}", exc_info=True)
            return pd.DataFrame()

    def processar_taxas_pendentes(self, df_contratos, data_relatorio):
        """
        Processa as taxas pendentes, agrupando por administrador e selecionando as próximas parcelas
        """
        if df_contratos.empty:
            return pd.DataFrame()
            
        try:
            # Converter data_relatorio para datetime
            data_ref = pd.to_datetime(data_relatorio).date()
            
            # Filtrar apenas parcelas futuras em relação à data do relatório
            df_futuro = df_contratos[
                pd.to_datetime(df_contratos['data_vencimento']).dt.date > data_ref
            ].copy()
            
            # Ordenar por data de vencimento e limitar a 3 parcelas por contrato/administrador
            df_futuro = df_futuro.sort_values(['referencia', 'administrador', 'data_vencimento'])
            
            # Agrupar por contrato e administrador e pegar as 3 primeiras parcelas de cada grupo
            df_final = df_futuro.groupby(['referencia', 'administrador']).head(3)
            
            # Ordenar resultado final
            df_final = df_final.sort_values(['administrador', 'data_vencimento'])
            
            print("\nParcelas processadas:")
            for _, row in df_final.iterrows():
                print(f"Administrador: {row['administrador']}")
                print(f"Contrato: {row['referencia']}")
                print(f"Parcela: {row['numero_parcela']}")
                print(f"Vencimento: {row['data_vencimento']}")
                print(f"Valor: {row['valor_parcela']}\n")
            
            return df_final
            
        except Exception as e:
            print(f"Erro ao processar taxas pendentes: {str(e)}")
            return pd.DataFrame()

        

    def adicionar_taxas_administracao(self, elementos, dados_taxas, config):
        """
        Adiciona a seção de taxas de administração pendentes ao relatório
        """
        if dados_taxas.empty:
            return
            
        try:
            # Adicionar quebra de página antes da seção de taxas
            elementos.append(PageBreak())
            
            # Adicionar título e descrição
            elementos.append(Paragraph(
                "TAXAS DE ADMINISTRAÇÃO VINCENDAS",
                config.style_heading
            ))
            elementos.append(Paragraph(
                "(Próximas 3 parcelas a vencer por contrato)",
                ParagraphStyle(
                    'SubtitleStyle',
                    parent=config.style_normal,
                    fontSize=9,
                    leading=12,
                    textColor=colors.HexColor('#4A4A4A'),
                    spaceBefore=2,
                    spaceAfter=12
                )
            ))
            
            total_geral = 0.0
            
            # Agrupar por administrador
            for administrador, grupo in dados_taxas.groupby('administrador'):
                # Adicionar nome do administrador e CNPJ/CPF com estilo melhorado
                cpf_cnpj = grupo['cpf_cnpj'].iloc[0]
                elementos.append(Paragraph(
                    f"{administrador}",
                    ParagraphStyle(
                        'AdminStyle',
                        parent=config.style_despesa,
                        fontSize=11,
                        leading=13,
                        textColor=colors.HexColor('#2F4F4F'),
                        spaceBefore=12,
                        spaceAfter=2
                    )
                ))
                elementos.append(Paragraph(
                    f"CNPJ/CPF: {cpf_cnpj}",
                    ParagraphStyle(
                        'CpfCnpjStyle',
                        parent=config.style_normal,
                        fontSize=8,
                        leading=10,
                        leftIndent=10,
                        textColor=colors.HexColor('#666666'),
                        spaceBefore=0,
                        spaceAfter=6
                    )
                ))
                
                subtotal_admin = 0.0
                
                # Criar lista de parcelas por contrato para este administrador
                for contrato, parcelas in grupo.groupby('referencia'):
                    # Cabeçalho da tabela
                    cabecalhos = ['Nº Parcela', 'Data Vencimento', 'Valor']
                    dados_tabela = [cabecalhos]
                    
                    # Adicionar linhas de dados
                    for _, parcela in parcelas.iterrows():
                        data = pd.to_datetime(parcela['data_vencimento']).strftime('%d/%m/%Y')
                        valor = self.formatar_numero(parcela['valor_parcela'])
                        dados_tabela.append([
                            f"Parcela {parcela['numero_parcela']}",
                            data,
                            f"R$ {valor}"
                        ])
                        subtotal_admin += float(parcela['valor_parcela'])
                        total_geral += float(parcela['valor_parcela'])
                    
                    # Criar e adicionar título do contrato
                    elementos.append(Paragraph(
                        f"Contrato {contrato}:",
                        ParagraphStyle(
                            'ContratoStyle',
                            parent=config.style_normal,
                            fontSize=9,
                            leading=11,
                            leftIndent=20,
                            textColor=colors.HexColor('#2F4F4F'),
                            spaceBefore=6,
                            spaceAfter=3
                        )
                    ))
                    
                    # Criar tabela com estilo melhorado
                    tabela = Table(
                        dados_tabela,
                        colWidths=[100, 100, 100],
                        style=TableStyle([
                            # Estilo do cabeçalho
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E6E6E6')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2F4F4F')),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 8),
                            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                            ('TOPPADDING', (0, 0), (-1, 0), 6),
                            
                            # Estilo das células de dados
                            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                            ('ALIGN', (0, 1), (1, -1), 'LEFT'),
                            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                            ('TOPPADDING', (0, 1), (-1, -1), 4),
                            ('LEFTPADDING', (0, 0), (-1, -1), 10),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                            
                            # Grades e bordas
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#999999')),
                        ])
                    )
                    
                    # Adicionar indentação na tabela
                    elementos.append(IndentedFlowable(tabela, leftIndent=30))
                    elementos.append(Spacer(1, 6))
                
                # Adicionar subtotal do administrador
                elementos.append(Paragraph(
                    f"Subtotal {administrador}: R$ {self.formatar_numero(subtotal_admin)}",
                    ParagraphStyle(
                        'SubtotalStyle',
                        parent=config.style_normal,
                        fontSize=9,
                        leading=11,
                        leftIndent=30,
                        textColor=colors.HexColor('#2F4F4F'),
                        spaceBefore=3,
                        spaceAfter=12,
                        alignment=TA_CENTER  # Centraliza o texto
                    )
                ))
            
            # Adicionar total geral
            elementos.append(Paragraph(
                f"Total de Taxas Vincendas: R$ {self.formatar_numero(total_geral)}",
                ParagraphStyle(
                    'TotalStyle',
                    parent=config.style_heading,
                    fontSize=10,
                    leading=12,
                    textColor=colors.HexColor('#2F4F4F'),
                    spaceBefore=12,
                    spaceAfter=6
                )
            ))
            
        except Exception as e:
            print(f"Erro ao adicionar taxas de administração: {str(e)}")
            raise # Para ajudar no debug
    

    def gerar_relatorio_pdf(self, dados, caminho_output, arquivo_excel):
        """Gera o relatório PDF final"""
        try:
            logger.debug("\nIniciando geração do PDF")
            logger.debug(f"Dados recebidos - acumulado: {dados.get('acumulado')}")
            logger.debug(f"Tipo do acumulado: {type(dados.get('acumulado'))}")
            
            # Criar cópia dos dados para não modificar o original
            dados_pdf = dados.copy()
            
            doc = SimpleDocTemplate(
                caminho_output, 
                pagesize=landscape(A4),
                rightMargin=30,
                leftMargin=30,
                topMargin=40,
                bottomMargin=30
            )
                
            elementos = []
            
            # Adicionar cabeçalho
            self.adicionar_cabecalho(elementos, dados_pdf)
            
            # Adicionar resumo
            elementos.append(Paragraph("RESUMO DAS DESPESAS", self.config.style_heading))
            
            logger.debug("Antes de criar_resumo_despesas:")
            logger.debug(f"Acumulado: {dados_pdf.get('acumulado')}")
            
            tabela_subtotais, tabela_totais = self.criar_resumo_despesas(dados_pdf)
            
            # Log para debug
            logger.debug("Tabela totais gerada:")
            for linha in tabela_totais:
                logger.debug(f"Linha: {linha}")
            
            # Criar tabelas com estilos específicos
            estilo_subtotais = TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ])

            estilo_totais = TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('BOX', (0, 0), (-1, 0), 1, colors.grey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ])

            tabela_esquerda = Table(tabela_subtotais, colWidths=[300, 70])
            tabela_esquerda.setStyle(estilo_subtotais)

            tabela_direita = Table(tabela_totais, colWidths=[180, 70])
            tabela_direita.setStyle(estilo_totais)

            # Criar tabela que combina as duas anteriores
            tabela_resumo = Table(
                [[tabela_esquerda, Spacer(1, 12), tabela_direita]],
                colWidths=[400, 60, 280]
            )
        
            elementos.append(tabela_resumo)
            
            # Adicionar quebra de página
            elementos.append(PageBreak())
            
            # Adicionar detalhes
            self.adicionar_detalhes(elementos, dados)

            if dados.get('incluir_futuros', True) and dados.get('df_futuro') is not None:
                self.adicionar_lancamentos_futuros(elementos, dados)

            # Carregar e processar taxas de administração
            df_taxas = self.carregar_taxas_administracao(arquivo_excel)
            if not df_taxas.empty:
                df_taxas_processadas = self.processar_taxas_pendentes(df_taxas, dados['data_relatorio'])
                if not df_taxas_processadas.empty:
                    self.adicionar_taxas_administracao(elementos, df_taxas_processadas, self.config)

            # Gerar PDF
            doc.build(elementos)

        except Exception as e:
            logger.error(f"Erro na geração do relatório: {str(e)}", exc_info=True)
            raise


##  Esse método foi substituído porque a atualização dos dados que estava sobrescrevendo os valores calculados em 21/02/2025 ""            
    # def gerar_relatorio_pdf(self, dados, caminho_output, arquivo_excel):
    #     """Gera o relatório PDF final"""
    #     try:
    #         # Carregar workbook com data_only=True para pegar valores calculados
    #         workbook = load_workbook(arquivo_excel, data_only=True)
    #         ws_resumo = workbook['RESUMO']
            
    #         data_rel = pd.to_datetime(dados['data_relatorio'])
            
    #         # Obter número do relatório
    #         relatorio_num = self.obter_numero_relatorio(ws_resumo, data_rel)
            
    #         # Calcular acumulado usando a nova função
    #         df = dados.get('df_original')  # Pegamos o DataFrame original
    #         if df is None:
    #             df = self.carregar_dados_excel(arquivo_excel)
    #         acumulado = self.calcular_acumulado_dados(df, data_rel)
            
    #         # Atualizar dados
    #         dados.update({
    #             'numero_relatorio': relatorio_num or 1,
    #             'acumulado': acumulado
    #         })
            

    #         # Get report number directly from worksheet
    #         relatorio_num = None
    #         acumulado = 0.0
            
    #         for row in range(9, 150):  # Scan reasonable range of rows
    #             data_cell = ws_resumo.cell(row=row, column=1).value
    #             if isinstance(data_cell, datetime):
    #                 if data_cell.date() == data_rel.date():
    #                     relatorio_num = ws_resumo.cell(row=row, column=2).value
    #                     # Get previous report's accumulated value
    #                     if row > 9:
    #                         acumulado = ws_resumo.cell(row=row-1, column=12).value or 0.0
    #                     break

    #         # Update dados with correct values
    #         dados.update({
    #             'numero_relatorio': relatorio_num or 1,
    #             'acumulado': float(acumulado)
    #         })

    #         # Continue with PDF generation
    #         doc = SimpleDocTemplate(
    #                 caminho_output, 
    #                 pagesize=landscape(A4),
    #                 rightMargin=30,
    #                 leftMargin=30,
    #                 topMargin=40,
    #                 bottomMargin=30
    #         )
                
    #         elementos = []
            
    #         # Adicionar cabeçalho
    #         self.adicionar_cabecalho(elementos, dados)
            
    #         # Adicionar resumo
    #         elementos.append(Paragraph("RESUMO DAS DESPESAS", self.config.style_heading))
    #         tabela_subtotais, tabela_totais = self.criar_resumo_despesas(dados)
            
    #          # Criar tabelas com estilos específicos
    #         estilo_subtotais = TableStyle([
    #             ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Texto à esquerda
    #             ('ALIGN', (1, 0), (1, -1), 'RIGHT'),   # Valores à direita
    #             ('FONTSIZE', (0, 0), (-1, -1), 9),     # Tamanho da fonte
    #         ])

    #         estilo_totais = TableStyle([
    #             ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Texto à esquerda
    #             ('ALIGN', (1, 0), (1, -1), 'RIGHT'),   # Valores à direita
    #             ('FONTSIZE', (0, 0), (-1, -1), 9),     # Tamanho da fonte
    #             # Destacar "DESPESAS A PAGAR"
    #             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Negrito
    #             ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fundo cinza claro
    #             ('BOX', (0, 0), (-1, 0), 1, colors.grey),  # Borda ao redor
    #             # Negrito para "TOTAL DA OBRA" (última linha)
    #             ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    #             ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
    #         ])

    #         tabela_esquerda = Table(tabela_subtotais, colWidths=[300, 70])
    #         tabela_esquerda.setStyle(estilo_subtotais)

    #         tabela_direita = Table(tabela_totais, colWidths=[180, 70])
    #         tabela_direita.setStyle(estilo_totais)

    #         # Criar tabela que combina as duas anteriores
    #         tabela_resumo = Table(
    #             [[tabela_esquerda, Spacer(1, 12), tabela_direita]],
    #             colWidths=[400, 60, 280]
    #         )
        
    #         elementos.append(tabela_resumo)
            
    #         # Adicionar quebra de página
    #         elementos.append(PageBreak())
            
    #         # Adicionar detalhes
    #         self.adicionar_detalhes(elementos, dados)

    #         if dados.get('incluir_futuros', True) and dados.get('df_futuro') is not None:
    #             self.adicionar_lancamentos_futuros(elementos, dados)

    #         # Carregar e processar taxas de administração
    #         df_taxas = self.carregar_taxas_administracao(arquivo_excel)
    #         if not df_taxas.empty:
    #             df_taxas_processadas = self.processar_taxas_pendentes(df_taxas, data_rel)
    #             if not df_taxas_processadas.empty:
    #                 self.adicionar_taxas_administracao(elementos, df_taxas_processadas, self.config)

    #         # Gerar PDF
    #         doc.build(elementos)

    #     except Exception as e:
    #         print(f"Erro na geração do relatório: {e}")
    #         raise       
        
class RelatorioLancamentosPendentes:
    def __init__(self):
        self.config = RelatorioConfig()

    def obter_ultima_data_fechamento(self, df):
        """
        Obtém a última data de fechamento (última DATA_REL usada)
        """
        if 'DATA_REL' not in df.columns or df.empty:
            return None
        return pd.to_datetime(df['DATA_REL']).max()

    def processar_arquivo_cliente(self, caminho_arquivo, data_referencia):
        """
        Processa um arquivo de cliente individual
        
        Parameters:
        -----------
        caminho_arquivo : str
            Caminho completo para o arquivo Excel
        data_referencia : datetime
            Data de referência para filtrar lançamentos
            
        Returns:
        --------
        dict ou None
            Dicionário com os dados processados ou None se houver erro
        """
        try:
            print(f"\nProcessando arquivo: {caminho_arquivo}")
            print(f"Data de referência: {data_referencia}")
            
            # Carregar dados do arquivo
            df = pd.read_excel(caminho_arquivo, sheet_name='Dados')
            df = df.fillna("")
            
            wb = load_workbook(caminho_arquivo, data_only=True)
            ws_resumo = wb['RESUMO']
            
            # Obter informações do cliente
            nome_cliente = ws_resumo['A3'].value
            print(f"Cliente: {nome_cliente}")
            
            # Converter DATA_REL para datetime
            df['DATA_REL'] = pd.to_datetime(df['DATA_REL'])
            
            # Filtrar lançamentos posteriores à data de referência
            df_pendentes = df[df['DATA_REL'] > data_referencia].copy()
            
            # Remover duplicatas baseado em todas as colunas relevantes
            colunas_check = ['DATA_REL', 'TP_DESP', 'NOME', 'REFERÊNCIA', 'VALOR']
            df_pendentes = df_pendentes.drop_duplicates(subset=colunas_check)
            print(f"Lançamentos encontrados (após remover duplicatas): {len(df_pendentes)}")
            
            if df_pendentes.empty:
                print("Nenhum lançamento pendente encontrado")
                return None
            
            # Identificar parcelamentos
            df_pendentes['is_parcelamento'] = df_pendentes['REFERÊNCIA'].str.contains(
                'parcela|parcelamento', 
                case=False, 
                na=False
            )
            
            # Converter valores para float
            df_pendentes['VALOR'] = pd.to_numeric(
                df_pendentes['VALOR'].astype(str)
                .str.replace('R$', '')
                .str.replace(',', '.')
                .str.strip(), 
                errors='coerce'
            ).fillna(0.0)
            
            # Converter tipo de despesa para inteiro
            df_pendentes['TP_DESP'] = df_pendentes['TP_DESP'].astype(int)
            
            # Formatar datas
            if 'DT_VENCTO' in df_pendentes.columns:
                df_pendentes['DT_VENCTO'] = pd.to_datetime(
                    df_pendentes['DT_VENCTO'], 
                    format='%d/%m/%Y', 
                    errors='coerce'
                )
            
            # Ordenar por data
            df_pendentes = df_pendentes.sort_values(['DATA_REL', 'TP_DESP'])
            
            return {
                'nome_cliente': nome_cliente,
                'ultima_data': data_referencia,
                'lancamentos': df_pendentes,
                'arquivo': caminho_arquivo
            }
            
        except Exception as e:
            print(f"Erro ao processar arquivo {caminho_arquivo}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def gerar_relatorio_html(self, dados_clientes, caminho_saida):
        """
        Gera um relatório HTML com os lançamentos pendentes
        """
        def formatar_valor(valor):
            """Formata valor para o padrão brasileiro"""
            return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
        try:
            # Lista para armazenar as partes do HTML
            html_parts = []
            
            # Cabeçalho do documento
            html_parts.extend([
                '<!DOCTYPE html>',
                '<html>',
                '<head>',
                '<meta charset="utf-8">',
                '<title>Relatório de Lançamentos Pendentes</title>',
                '<style>',
                'body { font-family: Arial, sans-serif; margin: 20px; background-color: #f0f2f5; }',
                'h1 { color: #2c3e50; text-align: center; margin-bottom: 30px; }',
                '.cliente { background-color: white; margin: 20px 0; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }',
                '.cliente-header { background-color: #f8f9fa; padding: 15px; margin: -20px -20px 20px -20px; border-radius: 8px 8px 0 0; border-bottom: 1px solid #dee2e6; }',
                '.cliente-header h2 { margin: 0; color: #2c3e50; }',
                'table { width: 100%; border-collapse: collapse; margin-top: 15px; background-color: white; }',
                'th, td { padding: 12px; text-align: left; border: 1px solid #dee2e6; font-size: 14px; }',
                'th { background-color: #f8f9fa; font-weight: bold; color: #495057; }',
                'tr:nth-child(even) { background-color: #f8f9fa; }',
                '.parcelamento { background-color: #fff3e0; }',
                '.valor { text-align: right; }',
                '.resumo { margin-top: 20px; padding: 15px; background-color: #e8f5e9; border-radius: 5px; font-weight: bold; }',
                '.data-geracao { text-align: center; color: #6c757d; margin-bottom: 30px; }',
                '</style>',
                '</head>',
                '<body>',
                '<h1>Relatório de Lançamentos Pendentes</h1>',
                f'<p class="data-geracao">Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>'
            ])

            # Processar dados de cada cliente
            for dados in dados_clientes:
                if dados is None or dados['lancamentos'].empty:
                    continue

                df = dados['lancamentos']
                total_cliente = df['VALOR'].sum()

                # Cabeçalho do cliente
                html_parts.extend([
                    '<div class="cliente">',
                    '<div class="cliente-header">',
                    f'<h2>{dados["nome_cliente"]}</h2>',
                    f'<p>Última data de fechamento: {dados["ultima_data"].strftime("%d/%m/%Y")}</p>',
                    '</div>',
                    '<table>',
                    '<tr>',
                    '<th>Data</th>',
                    '<th>Tipo</th>',
                    '<th>Nome</th>',
                    '<th>Referência</th>',
                    '<th>Vencimento</th>',
                    '<th>Valor</th>',
                    '</tr>'
                ])

                # Ordenar por data e tipo
                df = df.sort_values(['DATA_REL', 'TP_DESP'])
                
                # Adicionar linhas de dados
                for _, row in df.iterrows():
                    classe = 'parcelamento' if row['is_parcelamento'] else ''
                    valor = float(row['VALOR']) if pd.notnull(row['VALOR']) else 0.0
                    
                    html_parts.extend([
                        f'<tr class="{classe}">',
                        f'<td>{row["DATA_REL"].strftime("%d/%m/%Y")}</td>',
                        f'<td>{int(row["TP_DESP"])}</td>',
                        f'<td>{row["NOME"]}</td>',
                        f'<td>{row["REFERÊNCIA"]}</td>',
                        f'<td>{row["DT_VENCTO"].strftime("%d/%m/%Y") if pd.notnull(row["DT_VENCTO"]) else ""}</td>',
                        f'<td class="valor">{formatar_valor(valor)}</td>',
                        '</tr>'
                    ])

                # Fechar tabela e adicionar resumo
                html_parts.extend([
                    '</table>',
                    '<div class="resumo">',
                    f'<p>Total de lançamentos: R$ {formatar_valor(total_cliente)}</p>',
                    '</div>',
                    '</div>'
                ])

            # Fechar documento HTML
            html_parts.extend([
                '</body>',
                '</html>'
            ])

            # Juntar todas as partes e salvar
            html_content = '\n'.join(html_parts)
            
            with open(caminho_saida, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
            print(f"Relatório HTML gerado com sucesso em: {caminho_saida}")
            
        except Exception as e:
            print(f"Erro ao gerar relatório HTML: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def processar_pasta(self, pasta, data_referencia=None):
        """
        Processa todos os arquivos Excel da pasta
        
        Parameters:
        -----------
        pasta : str
            Caminho da pasta contendo os arquivos Excel
        data_referencia : datetime, optional
            Data de referência para filtrar lançamentos
            
        Returns:
        --------
        list
            Lista com os dados processados de cada cliente
        """
        try:
            print(f"\nProcessando pasta: {pasta}")
            print(f"Data de referência: {data_referencia}")
            
            # Se data_referencia não foi fornecida, usar data atual
            if data_referencia is None:
                data_referencia = datetime.now()
                
            arquivos = [f for f in os.listdir(pasta) if f.endswith('.xlsx')]
            print(f"Encontrados {len(arquivos)} arquivos Excel")
            
            dados_clientes = []
            for arquivo in arquivos:
                caminho_completo = os.path.join(pasta, arquivo)
                dados = self.processar_arquivo_cliente(caminho_completo, data_referencia)
                if dados is not None:
                    dados_clientes.append(dados)
                    
            print(f"Total de clientes processados: {len(dados_clientes)}")
            return dados_clientes
            
        except Exception as e:
            print(f"Erro ao processar pasta: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def gerar_relatorio_pendentes(self, pasta_entrada, arquivo_saida, data_referencia):
        """
        Método principal para gerar o relatório de lançamentos pendentes
        
        Parameters:
        -----------
        self: RelatorioLancamentosPendentes
            Instância da classe
        pasta_entrada : str
            Caminho da pasta com os arquivos dos clientes
        arquivo_saida : str
            Caminho onde o relatório HTML será salvo
        data_referencia : datetime
            Data de referência para filtrar lançamentos
        """
        try:
            print("\nGerando relatório de lançamentos pendentes...")
            print(f"Pasta de entrada: {pasta_entrada}")
            print(f"Arquivo de saída: {arquivo_saida}")
            print(f"Data de referência: {data_referencia}")
            
            # Processar todos os arquivos da pasta
            dados_clientes = self.processar_pasta(pasta_entrada, data_referencia)
            
            if not dados_clientes:
                print("Nenhum dado encontrado para processar")
                return False
                
            # Gerar relatório HTML
            self.gerar_relatorio_html(dados_clientes, arquivo_saida)
            
            # Abrir o relatório no navegador padrão
            if platform.system() == 'Darwin':       # macOS
                subprocess.run(['open', arquivo_saida])
            elif platform.system() == 'Windows':    # Windows
                os.startfile(arquivo_saida)
            else:                                   # Linux
                subprocess.run(['xdg-open', arquivo_saida])
                
            return True
            
        except Exception as e:
            print(f"Erro ao gerar relatório de lançamentos pendentes: {str(e)}")
            import traceback
            traceback.print_exc()
            return False


        

def main():
    try:
        app = RelatorioUI(None)
        app.root.mainloop()
    except Exception as e:
        print(f"Erro durante a execução: {str(e)}")

if __name__ == "__main__":
    main()
