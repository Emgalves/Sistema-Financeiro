"""
SCRIPT ALTERNATIVO: Atualizar Estrutura do Excel Clientes.xlsx
Versão simplificada que aceita o caminho do arquivo como argumento

Uso:
    python atualizar_estrutura_excel_manual.py [caminho_do_arquivo]
    
Exemplos:
    python atualizar_estrutura_excel_manual.py Clientes.xlsx
    python atualizar_estrutura_excel_manual.py data/Clientes.xlsx
    python atualizar_estrutura_excel_manual.py C:\Projeto\data\Clientes.xlsx
"""

import openpyxl
from pathlib import Path
import shutil
from datetime import datetime
import sys

def criar_backup(arquivo):
    """Cria backup do arquivo antes de modificar"""
    try:
        # Criar pasta de backup no mesmo diretório do arquivo
        backup_folder = arquivo.parent / "backups"
        backup_folder.mkdir(exist_ok=True)
        
        # Nome do backup com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = backup_folder / f"{arquivo.stem}_backup_{timestamp}{arquivo.suffix}"
        
        # Copiar arquivo
        shutil.copy2(arquivo, backup_file)
        print(f"✓ Backup criado: {backup_file}")
        return True
    except Exception as e:
        print(f"✗ Erro ao criar backup: {e}")
        return False

def atualizar_estrutura(arquivo):
    """Adiciona as novas colunas ao Excel"""
    try:
        # Carregar workbook
        wb = openpyxl.load_workbook(arquivo)
        ws = wb['Clientes']
        
        print("\n📊 Estrutura atual:")
        print(f"   Total de colunas: {ws.max_column}")
        print(f"   Total de linhas: {ws.max_row}")
        
        # Verificar se já tem as novas colunas
        if ws.max_column >= 23:
            print("\n⚠️  Aviso: Arquivo parece já ter sido atualizado")
            resposta = input("Deseja continuar mesmo assim? (s/n): ")
            if resposta.lower() != 's':
                print("Operação cancelada")
                return False
        
        # Definir novos cabeçalhos (colunas L-W)
        novos_headers = {
            12: 'Logradouro',      # L
            13: 'Numero',          # M
            14: 'Complemento',     # N
            15: 'Localidade',      # O
            16: 'CEP',             # P
            17: 'Estado',          # Q
            18: 'Metragem_Alvara', # R
            19: 'Banco',           # S
            20: 'Agencia',         # T
            21: 'Conta',           # U
            22: 'Tipo_Conta',      # V
            23: 'PIX'              # W
        }
        
        # Adicionar novos cabeçalhos na linha 1
        print("\n📝 Adicionando novos cabeçalhos...")
        for col, header in novos_headers.items():
            cell = ws.cell(row=1, column=col)
            cell.value = header
            # Formatação (opcional)
            cell.font = openpyxl.styles.Font(bold=True)
            print(f"   Coluna {openpyxl.utils.get_column_letter(col)}: {header}")
        
        # Salvar arquivo
        wb.save(arquivo)
        print("\n✓ Estrutura atualizada com sucesso!")
        
        print("\n📊 Nova estrutura:")
        print(f"   Total de colunas: {ws.max_column}")
        print(f"   Total de linhas: {ws.max_row}")
        
        return True
        
    except Exception as e:
        print(f"\n✗ Erro ao atualizar estrutura: {e}")
        import traceback
        traceback.print_exc()
        return False

def verificar_estrutura(arquivo):
    """Verifica e exibe a estrutura atual do arquivo"""
    try:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb['Clientes']
        
        print("\n" + "="*60)
        print("ESTRUTURA ATUAL DO ARQUIVO")
        print("="*60)
        
        # Ler cabeçalhos
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            letra = openpyxl.utils.get_column_letter(col)
            headers.append((letra, cell_value))
        
        # Exibir cabeçalhos
        print("\nCabeçalhos encontrados:")
        for letra, valor in headers:
            print(f"   {letra:3} : {valor}")
        
        print(f"\nTotal de colunas: {len(headers)}")
        print(f"Total de registros: {ws.max_row - 1}")
        print("="*60 + "\n")
        
        return True
        
    except Exception as e:
        print(f"✗ Erro ao verificar estrutura: {e}")
        return False

def main():
    """Função principal"""
    print("\n" + "="*60)
    print("ATUALIZAÇÃO DA ESTRUTURA - Clientes.xlsx")
    print("Versão Manual (com caminho especificado)")
    print("="*60)
    
    # Obter caminho do arquivo
    if len(sys.argv) > 1:
        caminho_arquivo = sys.argv[1]
    else:
        print("\n📂 Nenhum caminho especificado.")
        print("\nUso:")
        print("  python atualizar_estrutura_excel_manual.py [caminho]")
        print("\nExemplos:")
        print("  python atualizar_estrutura_excel_manual.py Clientes.xlsx")
        print("  python atualizar_estrutura_excel_manual.py data/Clientes.xlsx")
        print()
        caminho_arquivo = input("Digite o caminho do arquivo Clientes.xlsx: ").strip()
        
        if not caminho_arquivo:
            print("✗ Operação cancelada")
            return
    
    arquivo = Path(caminho_arquivo)
    
    # Informações sobre o arquivo
    print(f"\n📋 Informações do arquivo:")
    print(f"   Caminho fornecido: {caminho_arquivo}")
    print(f"   Caminho absoluto:  {arquivo.absolute()}")
    print(f"   Nome do arquivo:   {arquivo.name}")
    print(f"   Diretório pai:     {arquivo.parent}")
    
    # Verificar se arquivo existe
    if not arquivo.exists():
        print(f"\n✗ Arquivo não encontrado!")
        print(f"\n📁 Tentando listar arquivos no diretório {arquivo.parent}:")
        
        if arquivo.parent.exists():
            try:
                arquivos_encontrados = list(arquivo.parent.glob("*.xlsx"))
                if arquivos_encontrados:
                    print("\n   Arquivos .xlsx encontrados:")
                    for arq in arquivos_encontrados:
                        print(f"   - {arq.name}")
                else:
                    print("   Nenhum arquivo .xlsx encontrado")
            except Exception as e:
                print(f"   Erro ao listar: {e}")
        else:
            print(f"   Diretório não existe: {arquivo.parent}")
        
        return
    
    print(f"✓ Arquivo encontrado!")
    
    # Mostrar estrutura atual
    if not verificar_estrutura(arquivo):
        return
    
    # Confirmar operação
    print("\n⚠️  ATENÇÃO:")
    print("   - Este script irá modificar o arquivo Clientes.xlsx")
    print("   - Um backup será criado automaticamente")
    print("   - As colunas L-W serão adicionadas")
    print()
    
    resposta = input("Deseja continuar? (s/n): ")
    
    if resposta.lower() != 's':
        print("\n✗ Operação cancelada pelo usuário")
        return
    
    # Criar backup
    print("\n📦 Criando backup...")
    if not criar_backup(arquivo):
        print("✗ Não foi possível criar backup. Operação cancelada por segurança.")
        return
    
    # Atualizar estrutura
    print("\n🔧 Atualizando estrutura...")
    if atualizar_estrutura(arquivo):
        print("\n✓ Operação concluída com sucesso!")
        
        # Verificar nova estrutura
        print("\n🔍 Verificando nova estrutura...")
        verificar_estrutura(arquivo)
        
        print("\n" + "="*60)
        print("PRÓXIMOS PASSOS:")
        print("="*60)
        print("1. Verifique se os cabeçalhos estão corretos")
        print("2. Implemente os novos métodos criar_novo_cliente() e editar_cliente()")
        print("3. Teste com cadastro de um novo cliente")
        print("4. Migre clientes antigos quando necessário")
        print("="*60 + "\n")
    else:
        print("\n✗ Falha na atualização")
        print("   Restaure o backup se necessário")

if __name__ == "__main__":
    main()
