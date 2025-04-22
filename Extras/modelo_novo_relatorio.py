import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Adicionar diretório raiz ao path para importar módulos corretamente
def add_project_root():
    import sys
    from pathlib import Path
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent
    if str(project_root) not in sys.path:
        sys.path.append(str(project_root))

add_project_root()

# Importar configurações
try:
    from config.config import (
        ARQUIVO_CLIENTES,
        PASTA_CLIENTES,
        BASE_PATH
    )
    print("Configurações importadas com sucesso")
except ImportError as e:
    print(f"Erro ao importar configurações: {str(e)}")
    # Definir valores padrão em caso de falha
    BASE_PATH = Path(".")
    ARQUIVO_CLIENTES = BASE_PATH / "dados" / "clientes.xlsx"
    PASTA_CLIENTES = BASE_PATH / "dados" / "clientes"

try:
    from config.window_config import configurar_janela
    print("window_config importado com sucesso")
except ImportError as e:
    print(f"Erro ao importar window_config: {str(e)}")
    # Implementação simples de configurar_janela como fallback
    def configurar_janela(janela, titulo="Janela", largura=800, altura=600):
        janela.title(titulo)
        janela.geometry(f"{largura}x{altura}")
        janela.resizable(True, True)
        
        # Centralizar na tela
        janela.update_idletasks()
        width = janela.winfo_width()
        height = janela.winfo_height()
        x = (janela.winfo_screenwidth() // 2) - (width // 2)
        y = (janela.winfo_screenheight() // 2) - (height // 2)
        janela.geometry(f'{width}x{height}+{x}+{y}')

# Importar funções auxiliares ou definir aqui
def formatar_moeda_br(valor):
    """Formata um valor numérico como moeda brasileira"""
    try:
        valor_float = float(valor)
        return f"R$ {valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except (ValueError, TypeError):
        return f"R$ 0,00"

class RelatorioModelo:
    """Classe base para implementação de relatórios"""
    def __init__(self, parent=None):
        """Inicializa a interface do relatório"""
        self.parent = parent
        
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent
        else:
            self.root = tk.Tk()
            self.menu_principal = None
            
        configurar_janela(self.root, "Relatório [TIPO DO RELATÓRIO]", 1000, 700)
        
        # Configuração de variáveis
        self.cliente_atual = None
        self.arquivo_cliente = None
        self.data_referencia = datetime.now()
        
        # Configurar interface
        self.setup_gui()
        
    def setup_gui(self):
        """Configuração da interface gráfica principal"""
        # Frame principal
        self.frame_principal = ttk.Frame(self.root, padding=10)
        self.frame_principal.pack(fill='both', expand=True)
        
        # Frame para seleção
        self.frame_selecao = ttk.LabelFrame(self.frame_principal, text="Seleção de Cliente e Data")
        self.frame_selecao.pack(fill='x', pady=10)
        
        # Container para cliente
        frame_cliente = ttk.Frame(self.frame_selecao)
        frame_cliente.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_cliente, text="Selecione o Cliente:", font=('Arial', 11)).pack(side='left', pady=5)
        self.cliente_combobox = ttk.Combobox(frame_cliente, width=40, font=('Arial', 11))
        self.cliente_combobox.pack(side='left', padx=5)
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)
        
        # Container para data
        frame_data = ttk.Frame(self.frame_selecao)
        frame_data.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_data, text="Data de Referência:", font=('Arial', 11)).pack(side='left', pady=5)
        
        # Usar DateEntry se disponível, caso contrário usar Entry simples
        try:
            from tkcalendar import DateEntry
            self.data_entry = DateEntry(
                frame_data, 
                width=12,
                background='darkblue',
                foreground='white',
                borderwidth=2,
                date_pattern='dd/mm/yyyy',
                locale='pt_BR',
                font=('Arial', 11)
            )
            self.data_entry.pack(side='left', padx=5)
            self.data_entry.set_date(datetime.now())
        except ImportError:
            self.data_var = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
            ttk.Entry(
                frame_data,
                textvariable=self.data_var,
                width=12,
                font=('Arial', 11)
            ).pack(side='left', padx=5)
        
        # Botão de gerar relatório
        ttk.Button(
            frame_data,
            text="Gerar Relatório",
            command=self.gerar_relatorio,
            style='Big.TButton'
        ).pack(side='left', padx=20)
        
        # Frame para resultados - com notebook para separar visões
        self.frame_resultados = ttk.LabelFrame(self.frame_principal, text="Resultados")
        self.frame_resultados.pack(fill='both', expand=True, pady=10)
        
        # Notebook (abas)
        self.notebook = ttk.Notebook(self.frame_resultados)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Abas
        self.aba_resumo = ttk.Frame(self.notebook)
        self.aba_detalhes = ttk.Frame(self.notebook)
        self.aba_grafico = ttk.Frame(self.notebook)
        
        self.notebook.add(self.aba_resumo, text='Resumo')
        self.notebook.add(self.aba_detalhes, text='Detalhes')
        self.notebook.add(self.aba_grafico, text='Gráfico')
        
        # Configurar cada aba
        self.setup_aba_resumo()
        self.setup_aba_detalhes()
        self.setup_aba_grafico()
        
        # Botões na parte inferior
        frame_botoes = ttk.Frame(self.frame_principal)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(
            frame_botoes,
            text="Exportar para Excel",
            command=self.exportar_excel
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes,
            text="Exportar para PDF",
            command=self.exportar_pdf
        ).pack(side='left', padx=5)
        
        ttk.Button(
            frame_botoes,
            text="Voltar ao Menu",
            command=self.voltar_menu
        ).pack(side='right', padx=5)
        
        # Estilo para botões grandes
        style = ttk.Style()
        style.configure('Big.TButton', font=('Arial', 11, 'bold'), padding=(10, 5))
        
        # Carregar lista de clientes
        self.atualizar_lista_clientes()
        
    def setup_aba_resumo(self):
        """Configura a aba de resumo do relatório"""
        # Frame para informações do cliente
        frame_info = ttk.Frame(self.aba_resumo, padding=5)
        frame_info.pack(fill='x', pady=5)
        
        self.lbl_cliente_resumo = ttk.Label(
            frame_info, 
            text="Cliente: Nenhum selecionado", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_cliente_resumo.pack(side='left', padx=10)
        
        self.lbl_data_resumo = ttk.Label(
            frame_info, 
            text=f"Data: {datetime.now().strftime('%d/%m/%Y')}", 
            font=('Arial', 12, 'bold'),
            foreground='#0056b3'
        )
        self.lbl_data_resumo.pack(side='left', padx=10)
        
        # Adicionar elementos específicos para resumo
        # IMPLEMENTE AQUI OS ELEMENTOS ESPECÍFICOS
        pass
    
    def setup_aba_detalhes(self):
        """Configura a aba de detalhes do relatório"""
        # Adicionar elementos específicos para detalhes
        # IMPLEMENTE AQUI OS ELEMENTOS ESPECÍFICOS
        pass
    
    def setup_aba_grafico(self):
        """Configura a aba de gráficos"""
        # Frame para controles do gráfico
        frame_controles = ttk.Frame(self.aba_grafico, padding=5)
        frame_controles.pack(fill='x', pady=5)
        
        ttk.Label(frame_controles, text="Tipo de Gráfico:").pack(side='left', padx=5)
        self.combo_tipo_grafico = ttk.Combobox(frame_controles, values=[
            "Gráfico de Pizza",
            "Gráfico de Barras",
            "Gráfico de Linha"
        ], state='readonly', width=30)
        self.combo_tipo_grafico.pack(side='left', padx=5)
        self.combo_tipo_grafico.current(0)
        
        ttk.Button(frame_controles, text="Atualizar Gráfico", command=self.atualizar_grafico).pack(side='left', padx=20)
        
        # Frame para o gráfico
        self.frame_grafico = ttk.Frame(self.aba_grafico)
        self.frame_grafico.pack(fill='both', expand=True, pady=5)
        
    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes no combobox"""
        try:
            # Carregar arquivo de clientes
            workbook = load_workbook(ARQUIVO_CLIENTES)
            sheet = workbook['Clientes']  # Assumindo que existe uma aba chamada 'Clientes'
            
            # Limpar lista atual
            self.cliente_combobox['values'] = []
            
            # Pegar todos os clientes (pulando o cabeçalho)
            clientes = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome do cliente está na primeira coluna
                    clientes.append(row[0])
            
            # Atualizar combobox
            self.cliente_combobox['values'] = sorted(clientes)
            workbook.close()
            
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo de clientes não encontrado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
    
    def selecionar_cliente(self, event=None):
        """Atualiza o cliente selecionado"""
        self.cliente_atual = self.cliente_combobox.get()
        
        if self.cliente_atual:
            # Atualizar label
            self.lbl_cliente_resumo.config(text=f"Cliente: {self.cliente_atual}")
            
            # Definir o caminho do arquivo
            self.arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
    
    def gerar_relatorio(self):
        """Gera o relatório com base nos dados selecionados"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
            
        # Obter data de referência
        try:
            # Verificar se estamos usando DateEntry ou Entry
            if hasattr(self, 'data_entry'):
                self.data_referencia = datetime.strptime(self.data_entry.get(), '%d/%m/%Y')
            else:
                self.data_referencia = datetime.strptime(self.data_var.get(), '%d/%m/%Y')
                
            self.lbl_data_resumo.config(text=f"Data: {self.data_referencia.strftime('%d/%m/%Y')}")
        except ValueError:
            messagebox.showerror("Erro", "Data inválida!")
            return
            
        # Carregar dados
        if not self.carregar_dados():
            return
        
        # Preencher resumo
        self.preencher_resumo()
        
        # Preencher detalhes
        self.preencher_detalhes()
        
        # Gerar gráfico inicial
        self.atualizar_grafico()
        
        # Selecionar aba de resumo
        self.notebook.select(0)
    
    def carregar_dados(self):
        """Carrega os dados para o relatório"""
        try:
            if not os.path.exists(self.arquivo_cliente):
                messagebox.showerror("Erro", f"Arquivo do cliente '{self.cliente_atual}' não encontrado!")
                return False
                
            # Carregar dados específicos do relatório
            # IMPLEMENTE AQUI A LÓGICA DE CARREGAMENTO
            
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
            return False
    
    def preencher_resumo(self):
        """Preenche os dados da aba de resumo"""
        # IMPLEMENTE AQUI A LÓGICA DO RESUMO
        pass
    
    def preencher_detalhes(self):
        """Preenche os dados da aba de detalhes"""
        # IMPLEMENTE AQUI A LÓGICA DOS DETALHES
        pass
    
    def atualizar_grafico(self):
        """Atualiza o gráfico com base no tipo selecionado"""
        tipo_grafico = self.combo_tipo_grafico.get()
        
        # Limpar frame do gráfico
        for widget in self.frame_grafico.winfo_children():
            widget.destroy()
            
        # Verificar se há dados para gerar o gráfico
        if not hasattr(self, 'dados_grafico') or not self.dados_grafico:
            return
            
        # Criar figura
        fig, ax = plt.subplots(figsize=(8, 6))
        
        if tipo_grafico == "Gráfico de Pizza":
            self.criar_grafico_pizza(fig, ax)
        elif tipo_grafico == "Gráfico de Barras":
            self.criar_grafico_barras(fig, ax)
        elif tipo_grafico == "Gráfico de Linha":
            self.criar_grafico_linha(fig, ax)
            
        # Exibir o gráfico
        canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    
    def criar_grafico_pizza(self, fig, ax):
        """Cria um gráfico de pizza"""
        # IMPLEMENTE AQUI A LÓGICA DO GRÁFICO DE PIZZA
        pass
    
    def criar_grafico_barras(self, fig, ax):
        """Cria um gráfico de barras"""
        # IMPLEMENTE AQUI A LÓGICA DO GRÁFICO DE BARRAS
        pass
    
    def criar_grafico_linha(self, fig, ax):
        """Cria um gráfico de linha"""
        # IMPLEMENTE AQUI A LÓGICA DO GRÁFICO DE LINHA
        pass
    
    def exportar_excel(self):
        """Exporta o relatório para um arquivo Excel"""
        if not hasattr(self, 'cliente_atual') or not self.cliente_atual:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return
            
        # Solicitar nome do arquivo ao usuário
        data_str = self.data_referencia.strftime('%d-%m-%Y')
        nome_padrao = f"Relatorio_{self.__class__.__name__}_{self.cliente_atual}_{data_str}.xlsx"
        
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            initialfile=nome_padrao
        )
        
        if not arquivo:
            return
            
        try:
            # Criar workbook
            wb = Workbook()
            
            # IMPLEMENTE AQUI A LÓGICA DE EXPORTAÇÃO PARA EXCEL
            
            # Salvar o arquivo
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso para:\n{arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel: {str(e)}")
    
    def exportar_pdf(self):
        """Exporta o relatório para um arquivo PDF"""
        if not hasattr(self, 'cliente_atual') or not self.cliente_atual:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return
            
        # Solicitar nome do arquivo ao usuário
        data_str = self.data_referencia.strftime('%d-%m-%Y')
        nome_padrao = f"Relatorio_{self.__class__.__name__}_{self.cliente_atual}_{data_str}.pdf"
        
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Arquivos PDF", "*.pdf")],
            initialfile=nome_padrao
        )
        
        if not arquivo:
            return
            
        try:
            # IMPLEMENTE AQUI A LÓGICA DE EXPORTAÇÃO PARA PDF
            # Esta função geralmente requer a biblioteca reportlab
            
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso para:\n{arquivo}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para PDF: {str(e)}")
    
    def voltar_menu(self):
        """Volta ao menu principal"""
        self.root.destroy()
        
        # Mostrar janela principal
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()

def main():
    """Função principal para executar o módulo de forma independente"""
    app = RelatorioModelo()
    app.root.mainloop()
    
if __name__ == "__main__":
    main()
