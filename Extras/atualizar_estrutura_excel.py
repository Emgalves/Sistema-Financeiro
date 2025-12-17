"""
SCRIPT: Atualizar Estrutura do Excel Clientes.xlsx
Adiciona as novas colunas L-W mantendo compatibilidade com dados existentes

ATENÇÃO: Este script deve ser executado ANTES de implementar os novos métodos
         Faça backup do arquivo Clientes.xlsx antes de executar!

Uso:
    python atualizar_estrutura_excel.py
"""

import openpyxl
from pathlib import Path
import shutil
from datetime import datetime
import sys
import os

# Adicionar diretório src ao path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Importar configurações do sistema
try:
    from src.config.config import (
        ARQUIVO_CLIENTES,
        ARQUIVO_MODELO,
        PASTA_CLIENTES,
        BASE_PATH
    )
    print("✓ Configurações do sistema importadas com sucesso")
except Exception as e:
    print(f"✗ Erro ao importar configurações do sistema: {e}")
    print("\nATENÇÃO: Execute este script a partir do diretório raiz do projeto:")
    print("  python atualizar_estrutura_excel.py")
    sys.exit(1)

# Configuração de backup
BACKUP_FOLDER = PASTA_CLIENTES / "backups"

def criar_backup():
    """Cria backup do arquivo antes de modificar"""
    try:
        # Criar pasta de backup se não existir
        BACKUP_FOLDER.mkdir(exist_ok=True)
        
        # Nome do backup com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = BACKUP_FOLDER / f"Clientes_backup_{timestamp}.xlsx"
        
        # Copiar arquivo
        shutil.copy2(ARQUIVO_CLIENTES, backup_file)
        print(f"✓ Backup criado: {backup_file}")
        return True
    except Exception as e:
        print(f"✗ Erro ao criar backup: {e}")
        return False

def atualizar_estrutura():
    """Adiciona as novas colunas ao Excel"""
    try:
        # Carregar workbook
        wb = openpyxl.load_workbook(ARQUIVO_CLIENTES)
        ws = wb['Clientes']
        
        print("\n📊 Estrutura atual:")
        print(f"   Total de colunas: {ws.max_column}")
        print(f"   Total de linhas: {ws.max_row}")
        
        # Verificar se já tem as novas colunas
        if ws.max_column >= 23:
            print("\n⚠️  Aviso: Arquivo parece já ter sido atualizado")
            resposta = input("Deseja continuar mesmo assim? (s/n): ")
            if resposta.lower() != 's':
                print("Operação cancelada")
                return False
        
        # Definir novos cabeçalhos (colunas L-W)
        novos_headers = {
            12: 'Logradouro',      # L
            13: 'Numero',          # M
            14: 'Complemento',     # N
            15: 'Localidade',      # O
            16: 'CEP',             # P
            17: 'Estado',          # Q
            18: 'Metragem_Alvara', # R
            19: 'Banco',           # S
            20: 'Agencia',         # T
            21: 'Conta',           # U
            22: 'Tipo_Conta',      # V
            23: 'PIX'              # W
        }
        
        # Adicionar novos cabeçalhos na linha 1
        print("\n📝 Adicionando novos cabeçalhos...")
        for col, header in novos_headers.items():
            cell = ws.cell(row=1, column=col)
            cell.value = header
            # Formatação (opcional)
            cell.font = openpyxl.styles.Font(bold=True)
            print(f"   Coluna {openpyxl.utils.get_column_letter(col)}: {header}")
        
        # Salvar arquivo
        wb.save(ARQUIVO_CLIENTES)
        print("\n✓ Estrutura atualizada com sucesso!")
        
        print("\n📊 Nova estrutura:")
        print(f"   Total de colunas: {ws.max_column}")
        print(f"   Total de linhas: {ws.max_row}")
        
        return True
        
    except FileNotFoundError:
        print(f"\n✗ Erro: Arquivo '{ARQUIVO_CLIENTES}' não encontrado!")
        print("   Certifique-se de que o arquivo existe no diretório atual")
        return False
    except Exception as e:
        print(f"\n✗ Erro ao atualizar estrutura: {e}")
        return False

def verificar_estrutura():
    """Verifica e exibe a estrutura atual do arquivo"""
    try:
        wb = openpyxl.load_workbook(ARQUIVO_CLIENTES)
        ws = wb['Clientes']
        
        print("\n" + "="*60)
        print("ESTRUTURA ATUAL DO ARQUIVO")
        print("="*60)
        
        # Ler cabeçalhos
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            letra = openpyxl.utils.get_column_letter(col)
            headers.append((letra, cell_value))
        
        # Exibir cabeçalhos
        print("\nCabeçalhos encontrados:")
        for letra, valor in headers:
            print(f"   {letra:3} : {valor}")
        
        print(f"\nTotal de colunas: {len(headers)}")
        print(f"Total de registros: {ws.max_row - 1}")
        print("="*60 + "\n")
        
        return True
        
    except Exception as e:
        print(f"✗ Erro ao verificar estrutura: {e}")
        return False

def main():
    """Função principal"""
    print("\n" + "="*60)
    print("ATUALIZAÇÃO DA ESTRUTURA - Clientes.xlsx")
    print("="*60)
    
    # Mostrar caminhos configurados
    print("\n📂 Caminhos do Sistema:")
    print(f"   BASE_PATH:        {BASE_PATH}")
    print(f"   PASTA_CLIENTES:   {PASTA_CLIENTES}")
    print(f"   ARQUIVO_CLIENTES: {ARQUIVO_CLIENTES}")
    print(f"   BACKUP_FOLDER:    {BACKUP_FOLDER}")
    
    # Verificar se arquivo existe
    if not ARQUIVO_CLIENTES.exists():
        print(f"\n✗ Arquivo não encontrado: {ARQUIVO_CLIENTES}")
        print("\n📋 Verificações realizadas:")
        print(f"   - Caminho absoluto: {ARQUIVO_CLIENTES.absolute()}")
        print(f"   - Existe? {ARQUIVO_CLIENTES.exists()}")
        print(f"   - Diretório pai existe? {ARQUIVO_CLIENTES.parent.exists()}")
        
        if ARQUIVO_CLIENTES.parent.exists():
            print(f"\n📁 Arquivos no diretório:")
            try:
                for arquivo in ARQUIVO_CLIENTES.parent.iterdir():
                    if arquivo.suffix in ['.xlsx', '.xls']:
                        print(f"   - {arquivo.name}")
            except Exception as e:
                print(f"   Erro ao listar: {e}")
        
        return
    
    # Mostrar estrutura atual
    verificar_estrutura()
    
    # Confirmar operação
    print("\n⚠️  ATENÇÃO:")
    print("   - Este script irá modificar o arquivo Clientes.xlsx")
    print("   - Um backup será criado automaticamente")
    print("   - As colunas L-W serão adicionadas")
    print()
    
    resposta = input("Deseja continuar? (s/n): ")
    
    if resposta.lower() != 's':
        print("\n✗ Operação cancelada pelo usuário")
        return
    
    # Criar backup
    print("\n📦 Criando backup...")
    if not criar_backup():
        print("✗ Não foi possível criar backup. Operação cancelada por segurança.")
        return
    
    # Atualizar estrutura
    print("\n🔧 Atualizando estrutura...")
    if atualizar_estrutura():
        print("\n✓ Operação concluída com sucesso!")
        
        # Verificar nova estrutura
        print("\n🔍 Verificando nova estrutura...")
        verificar_estrutura()
        
        print("\n" + "="*60)
        print("PRÓXIMOS PASSOS:")
        print("="*60)
        print("1. Verifique se os cabeçalhos estão corretos")
        print("2. Implemente os novos métodos criar_novo_cliente() e editar_cliente()")
        print("3. Teste com cadastro de um novo cliente")
        print("4. Migre clientes antigos quando necessário")
        print("="*60 + "\n")
    else:
        print("\n✗ Falha na atualização")
        print("   Restaure o backup se necessário")

if __name__ == "__main__":
    main()